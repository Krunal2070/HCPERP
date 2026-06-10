"""
HCP Stock — Flask Blueprint  (new FG / PM / BOM model)
======================================================
All routes mounted under /hcp_stock/*

Access:
    Restricted to admin & planning users (login_required + role gate).

Tabs:
    FG       — finished goods (with BOM editor)
    PM       — packing materials (Inward + Wastage feed off these)
    Inward   — PM inward register
    Dispatch — FG dispatch register (auto-consumes PM via BOM)
    Wastage  — PM actual wastage register
"""

import io
from datetime import datetime
from functools import wraps

from flask import (Blueprint, render_template, request, jsonify,
                   send_file, session, redirect, url_for)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import hcp_stock_db as db

# Common, data-driven sidebar (menu defined once in core/menus.py)
try:
    from menus import get_menu
except Exception:
    def get_menu(*a, **k): return None


hcp_stock_bp = Blueprint('hcp_stock', __name__, url_prefix='/hcp_stock')


# ─── auth + role gate (matches the rest of the portal) ──────────────────────
def login_required(view):
    @wraps(view)
    def wrapped(*a, **kw):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return view(*a, **kw)
    return wrapped


_ALLOWED_ROLES = {'admin', 'planning'}


def hcp_stock_role_required(view):
    @wraps(view)
    def wrapped(*a, **kw):
        role = (session.get('User_Type') or '').strip().lower()
        if role not in _ALLOWED_ROLES:
            api_paths = ('/hcp_stock/api', '/hcp_stock/import',
                         '/hcp_stock/export', '/hcp_stock/template')
            if any(request.path.startswith(p) for p in api_paths):
                return jsonify({'ok': False,
                                'error': 'Access denied — admin / planning only'}), 403
            return ("<div style='font-family:sans-serif;padding:60px;text-align:center;'>"
                    "<h2>403 — Access denied</h2>"
                    "<p>HCP Stock is restricted to <b>Admin</b> and "
                    "<b>Planning</b> users.</p>"
                    "<p><a href='/'>← Back to portal</a></p></div>"), 403
        return view(*a, **kw)
    return wrapped


def admin_only(view):
    """Stricter than role_required — only admin can access. Used for Recycle Bin
    + Audit Log endpoints, which expose privileged data."""
    @wraps(view)
    def wrapped(*a, **kw):
        role = (session.get('User_Type') or '').strip().lower()
        if role != 'admin':
            return jsonify({'ok': False,
                            'error': 'Access denied — admin only'}), 403
        return view(*a, **kw)
    return wrapped


def feature_required(slug):
    """Per-user feature gate. Admin always passes; other users must have the
    slug in their hcp_stock_permissions row (or in DEFAULT_NON_ADMIN_FEATURES
    when no row exists)."""
    def deco(view):
        @wraps(view)
        def wrapped(*a, **kw):
            role = (session.get('User_Type') or '').strip().lower()
            if role == 'admin':
                return view(*a, **kw)
            user_name = session.get('User_Name') or session.get('UID') or ''
            features = db.get_user_features(user_name, role=role)
            if slug not in features:
                return jsonify({
                    'ok': False,
                    'error': f"Access denied — feature '{slug}' is not granted "
                             f"to your user. Ask an admin to grant it via Sidebar → Permissions.",
                    'missing_feature': slug,
                }), 403
            return view(*a, **kw)
        return wrapped
    return deco


def _user_features_for_session():
    """Convenience: resolved features for the current session user."""
    role = (session.get('User_Type') or '').strip().lower()
    user_name = session.get('User_Name') or session.get('UID') or ''
    return db.get_user_features(user_name, role=role)


def _audit_user():
    """Extract user info from session for audit logging."""
    return {
        'user_name': session.get('User_Name') or session.get('UID') or 'unknown',
        'user_role': session.get('User_Type') or 'unknown',
    }


def _audit(action, entity, entity_id=None, summary='', details=None):
    """Wrapper around db.write_audit that injects session user info.
       Never raises — audit must not break the underlying CRUD."""
    try:
        u = _audit_user()
        db.write_audit(
            user_name=u['user_name'], user_role=u['user_role'],
            action=action, entity=entity, entity_id=entity_id,
            summary=summary, details=details,
        )
    except Exception as e:
        # Defensive — should never bubble up
        print(f"[HCP-Stock] audit logging failed: {e}")


def _f(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _dmy(v):
    """Display a stored date/datetime as DD/MM/YYYY (time kept if present).
    Storage and the wire format stay ISO (YYYY-MM-DD); this is for human-facing
    strings only — audit summaries and Excel/CSV export cells."""
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        try:
            return v.strftime('%d/%m/%Y')
        except Exception:
            return str(v)
    s = str(v).strip()
    if not s:
        return s
    import re as _re
    m = _re.match(r'^(\d{4})-(\d{2})-(\d{2})(?:[ T](\d{2}:\d{2}(?::\d{2})?))?', s)
    if not m:
        return s
    out = f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    return f"{out} {m.group(4)}" if m.group(4) else out


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/')
@login_required
@hcp_stock_role_required
def hcp_stock_page():
    db.ensure_tables()
    role = session.get('User_Type', '') or ''
    user_name = session.get('User_Name', 'User')
    is_admin = role.strip().lower() == 'admin'
    user_features = sorted(db.get_user_features(user_name, role=role))
    return render_template('hcp_stock.html',
                           user_name=user_name,
                           role=role,
                           is_admin=is_admin,
                           user_features=user_features,
                           feature_catalog=db.FEATURE_CATALOG,
                           sidebar_menu=get_menu('hcp_stock', is_admin=is_admin),
                           active_item=None)


# ═══════════════════════════════════════════════════════════════════════════════
# BRANDS  (read-only — sourced from procurement_brands)
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/api/brands')
@login_required
@hcp_stock_role_required
def api_brands():
    return jsonify({'ok': True, 'brands': db.list_brands()})


# ═══════════════════════════════════════════════════════════════════════════════
# PM master
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/api/pm', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_pm_list():
    brand_id = request.args.get('brand_id', type=int)
    rows = db.get_pm_rows(brand_id=brand_id)
    return jsonify({'ok': True, 'rows': rows,
                    'kpi': db.kpi_summary(brand_id=brand_id)})


@hcp_stock_bp.route('/api/pm', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('add_pm')
def api_pm_create():
    d = request.json or {}
    if not d.get('pm_name'): return jsonify({'ok': False, 'error': 'PM Name is required'}), 400
    if not d.get('brand_id'): return jsonify({'ok': False, 'error': 'Brand is required'}), 400
    payload = {
        'brand_id':            int(d['brand_id']),
        'pm_code':             (d.get('pm_code') or '').strip(),
        'pm_name':             d['pm_name'].strip(),
        'pm_type':             (d.get('pm_type') or '').strip(),
        'sku_size':            (d.get('sku_size') or '').strip(),
        'rate':                _f(d.get('rate')),
        'opening_stock':       _f(d.get('opening_stock')),
        'provisional_wastage': _f(d.get('provisional_wastage')),
        'low_stock_threshold': _f(d.get('low_stock_threshold')),
    }
    pm_id = db.upsert_pm(payload)
    _audit('create', 'pm', pm_id,
           summary=f"Created PM '{payload['pm_name']}'"
                   + (f" [{payload['pm_code']}]" if payload['pm_code'] else ''),
           details={'after': payload})
    return jsonify({'ok': True, 'id': pm_id})


@hcp_stock_bp.route('/api/pm/<int:pm_id>', methods=['PUT'])
@login_required
@hcp_stock_role_required
@feature_required('edit_pm')
def api_pm_update(pm_id):
    d = request.json or {}
    if not d.get('pm_name'): return jsonify({'ok': False, 'error': 'PM Name is required'}), 400
    if not d.get('brand_id'): return jsonify({'ok': False, 'error': 'Brand is required'}), 400
    before = db.get_pm(pm_id) or {}
    payload = {
        'brand_id':            int(d['brand_id']),
        'pm_code':             (d.get('pm_code') or '').strip(),
        'pm_name':             d['pm_name'].strip(),
        'pm_type':             (d.get('pm_type') or '').strip(),
        'sku_size':            (d.get('sku_size') or '').strip(),
        'rate':                _f(d.get('rate')),
        'opening_stock':       _f(d.get('opening_stock')),
        'provisional_wastage': _f(d.get('provisional_wastage')),
        'low_stock_threshold': _f(d.get('low_stock_threshold')),
    }
    db.upsert_pm(payload, pm_id=pm_id)
    _audit('update', 'pm', pm_id,
           summary=f"Updated PM '{payload['pm_name']}'"
                   + (f" [{payload['pm_code']}]" if payload['pm_code'] else ''),
           details={'before': {k: before.get(k) for k in payload.keys()},
                    'after':  payload})
    return jsonify({'ok': True})


@hcp_stock_bp.route('/api/pm/<int:pm_id>', methods=['DELETE'])
@login_required
@hcp_stock_role_required
@feature_required('delete_pm')
def api_pm_delete(pm_id):
    u = _audit_user()
    before = db.get_pm(pm_id) or {}
    db.delete_pm(pm_id, user_name=u['user_name'])
    _audit('delete', 'pm', pm_id,
           summary=f"Soft-deleted PM '{before.get('pm_name','')}'"
                   + (f" [{before.get('pm_code','')}]" if before.get('pm_code') else ''),
           details={'before': before})
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════════════════════════════════
# FG master + BOM
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/api/fg', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_fg_list():
    brand_id = request.args.get('brand_id', type=int)
    rows = db.get_fg_rows(brand_id=brand_id)
    return jsonify({'ok': True, 'rows': rows})


@hcp_stock_bp.route('/api/fg', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('add_fg')
def api_fg_create():
    d = request.json or {}
    if not d.get('product_name'): return jsonify({'ok': False, 'error': 'Product Name is required'}), 400
    if not d.get('brand_id'): return jsonify({'ok': False, 'error': 'Brand is required'}), 400
    payload = {
        'brand_id':     int(d['brand_id']),
        'product_code': (d.get('product_code') or '').strip(),
        'product_name': d['product_name'].strip(),
        'category':     (d.get('category') or '').strip(),
        'sku_size':     (d.get('sku_size') or '').strip(),
        'rate':         _f(d.get('rate')),
    }
    fg_id = db.upsert_fg(payload)
    bom_lines = d.get('bom') if isinstance(d.get('bom'), list) else None
    if bom_lines is not None:
        db.set_bom_for_fg(fg_id, bom_lines)
    _audit('create', 'fg', fg_id,
           summary=f"Created FG '{payload['product_name']}'"
                   + (f" [{payload['product_code']}]" if payload['product_code'] else ''),
           details={'after': payload, 'bom_count': len(bom_lines or [])})
    return jsonify({'ok': True, 'id': fg_id})


@hcp_stock_bp.route('/api/fg/<int:fg_id>', methods=['PUT'])
@login_required
@hcp_stock_role_required
@feature_required('edit_fg')
def api_fg_update(fg_id):
    d = request.json or {}
    if not d.get('product_name'): return jsonify({'ok': False, 'error': 'Product Name is required'}), 400
    if not d.get('brand_id'): return jsonify({'ok': False, 'error': 'Brand is required'}), 400
    before = db.get_fg(fg_id) or {}
    bom_before = db.get_bom_for_fg(fg_id)
    payload = {
        'brand_id':     int(d['brand_id']),
        'product_code': (d.get('product_code') or '').strip(),
        'product_name': d['product_name'].strip(),
        'category':     (d.get('category') or '').strip(),
        'sku_size':     (d.get('sku_size') or '').strip(),
        'rate':         _f(d.get('rate')),
    }
    db.upsert_fg(payload, fg_id=fg_id)
    bom_lines = d.get('bom') if isinstance(d.get('bom'), list) else None
    if bom_lines is not None:
        db.set_bom_for_fg(fg_id, bom_lines)
    _audit('update', 'fg', fg_id,
           summary=f"Updated FG '{payload['product_name']}'"
                   + (f" [{payload['product_code']}]" if payload['product_code'] else ''),
           details={'before': {k: before.get(k) for k in payload.keys()},
                    'after':  payload,
                    'bom_before_count': len(bom_before),
                    'bom_after_count':  len(bom_lines) if bom_lines is not None else len(bom_before)})
    return jsonify({'ok': True})


@hcp_stock_bp.route('/api/fg/<int:fg_id>', methods=['DELETE'])
@login_required
@hcp_stock_role_required
@feature_required('delete_fg')
def api_fg_delete(fg_id):
    u = _audit_user()
    before = db.get_fg(fg_id) or {}
    db.delete_fg(fg_id, user_name=u['user_name'])
    _audit('delete', 'fg', fg_id,
           summary=f"Soft-deleted FG '{before.get('product_name','')}'"
                   + (f" [{before.get('product_code','')}]" if before.get('product_code') else ''),
           details={'before': before})
    return jsonify({'ok': True})


@hcp_stock_bp.route('/api/fg/<int:fg_id>/bom', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_fg_bom(fg_id):
    return jsonify({'ok': True, 'lines': db.get_bom_for_fg(fg_id)})


@hcp_stock_bp.route('/api/fg/<int:fg_id>/bom', methods=['PUT'])
@login_required
@hcp_stock_role_required
@feature_required('edit_fg')
def api_fg_bom_save(fg_id):
    d = request.json or {}
    lines = d.get('lines') or []
    before = db.get_bom_for_fg(fg_id)
    db.set_bom_for_fg(fg_id, lines)
    fg = db.get_fg(fg_id) or {}
    _audit('update_bom', 'fg', fg_id,
           summary=f"Updated BOM for FG '{fg.get('product_name','')}' ({len(before)} → {len(lines)} lines)",
           details={'lines_before': len(before), 'lines_after': len(lines)})
    return jsonify({'ok': True})


@hcp_stock_bp.route('/api/fg/<int:fg_id>/requirement', methods=['PUT'])
@login_required
@hcp_stock_role_required
@feature_required('set_requirement')
def api_fg_set_requirement(fg_id):
    """Inline-edit endpoint for the Requirement column on the FG grid.
    Body: {"quantity": <number>} — non-negative; 0 clears the requirement.
    Audited with before/after values."""
    d = request.json or {}
    qty = _f(d.get('quantity'))
    if qty < 0:
        return jsonify({'ok': False, 'error': 'Requirement cannot be negative'}), 400
    fg = db.get_fg(fg_id)
    if not fg:
        return jsonify({'ok': False, 'error': 'FG not found'}), 404

    before, after = db.set_fg_requirement(fg_id, qty)
    # Only audit when value actually changed (avoids noise on no-op blur events)
    try:
        before_f = float(before) if before is not None else 0.0
    except Exception:
        before_f = 0.0
    if before_f != after:
        _audit('update_requirement', 'fg', fg_id,
               summary=f"Set requirement for '{fg.get('product_name','')}' "
                       f"= {after}{(' (was '+str(before_f)+')') if before_f else ''}",
               details={'before': before_f, 'after': after})
    return jsonify({'ok': True, 'requirement_qty': after})


@hcp_stock_bp.route('/api/fg/<int:fg_id>/pending_requirement', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_fg_pending_requirement(fg_id):
    """Return the FG's PENDING requirement quantity for the planner / can-build.
    pending = (sum of non-preclosed requirement entries) − (all dispatches ever),
    floored at 0. Used to default the Check-PM planner qty and the 'Pending
    Requirement' chip."""
    fg = db.get_fg(fg_id)
    if not fg:
        return jsonify({'ok': False, 'error': 'FG not found'}), 404
    pending = db.fg_pending_requirement(fg_id)
    return jsonify({'ok': True, 'fg_id': fg_id, 'pending': pending})


# ── Dated FG requirement entries (separate from the cached requirement_qty) ──
@hcp_stock_bp.route('/api/fg_requirement', methods=['GET'])
@login_required
@hcp_stock_role_required
@feature_required('report_req_history')
def api_fg_requirement_list():
    """List requirement entries with optional filters.
    Query params: brand_id, fg_id, from=YYYY-MM-DD, to=YYYY-MM-DD."""
    brand_id = request.args.get('brand_id', type=int)
    fg_id    = request.args.get('fg_id',    type=int)
    start    = (request.args.get('from') or '').strip() or None
    end      = (request.args.get('to')   or '').strip() or None
    rows = db.list_fg_requirements(brand_id=brand_id, fg_id=fg_id,
                                    start_date=start, end_date=end)
    return jsonify({'ok': True, 'rows': rows, 'count': len(rows)})


@hcp_stock_bp.route('/api/fg_requirement/batch', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('set_requirement')
def api_fg_requirement_batch():
    """Atomic multi-line FG requirement entry.
    Body: {"items": [{fg_id, entry_date, quantity, note}, ...]}
    """
    d = request.json or {}
    items = d.get('items') or []
    if not isinstance(items, list) or not items:
        return jsonify({'ok': False, 'error': 'items must be a non-empty list'}), 400

    cleaned = []
    today = datetime.now().strftime('%Y-%m-%d')
    for i, it in enumerate(items, start=1):
        if not it.get('fg_id'):
            return jsonify({'ok': False, 'error': f'Line {i}: FG (Product) is required',
                            'line_index': i}), 400
        qty = _f(it.get('quantity'))
        if qty <= 0:
            return jsonify({'ok': False, 'error': f'Line {i}: quantity must be > 0',
                            'line_index': i}), 400
        cleaned.append({
            'fg_id':      int(it['fg_id']),
            'entry_date': it.get('entry_date') or today,
            'quantity':   qty,
            'note':       (it.get('note') or '').strip(),
        })

    user_name = session.get('User_Name') or session.get('UID') or ''
    try:
        ids = db.add_fg_requirement_batch(cleaned, created_by=user_name)
        for idx, (new_id, payload) in enumerate(zip(ids, cleaned), start=1):
            fg = db.get_fg(payload['fg_id']) or {}
            _audit('create', 'fg_requirement', new_id,
                   summary=f"[Batch line {idx}/{len(ids)}] Requirement {payload['quantity']} "
                           f"of '{fg.get('product_name','')}' on {_dmy(payload['entry_date'])}",
                   details={'after': payload, 'batch_size': len(ids)})
        return jsonify({'ok': True, 'ids': ids, 'count': len(ids)})
    except ValueError as e:
        return jsonify({
            'ok': False,
            'error': str(e),
            'line_index': getattr(e, 'line_index', None),
        }), 400


@hcp_stock_bp.route('/api/fg_requirement/<int:req_id>', methods=['DELETE'])
@login_required
@hcp_stock_role_required
@feature_required('delete_requirement')
def api_fg_requirement_delete(req_id):
    """Soft-delete a requirement entry. Refreshes the FG's cached requirement_qty."""
    user_name = session.get('User_Name') or session.get('UID') or ''
    ok = db.soft_delete_fg_requirement(req_id, user_name=user_name)
    if not ok:
        return jsonify({'ok': False, 'error': 'Requirement entry not found'}), 404
    _audit('delete', 'fg_requirement', req_id,
           summary=f'Deleted requirement entry #{req_id}')
    return jsonify({'ok': True})


# ───────────────────────────────────────────────────────────────────────────
# ADMIN-ONLY: reset FG requirement
# ───────────────────────────────────────────────────────────────────────────
# Two distinct actions, both gated by admin role:
#
#   1. POST /api/fg/<fg_id>/requirement/reset_clear_all
#      Soft-deletes EVERY non-deleted requirement entry for this FG.
#      The cached value drops to 0. History is preserved (visible in Recycle
#      Bin and Requirement History) and individual entries can still be
#      restored. This is the "clean reset" — preserves data integrity.
#
#   2. POST /api/fg/<fg_id>/requirement/reset_override
#      Body: {"quantity": <number>}
#      Directly sets hcp_stock_fg.requirement_qty WITHOUT touching the
#      dated history. Breaks the "cached = sum of history" invariant. The
#      response includes the divergence so the UI can warn the admin. The
#      cached value will be re-synced on the next batch insert / soft-delete
#      against the history table.
# ───────────────────────────────────────────────────────────────────────────

@hcp_stock_bp.route('/api/fg/<int:fg_id>/requirement/reset_clear_all', methods=['POST'])
@login_required
@hcp_stock_role_required
@admin_only
def api_admin_fg_requirement_reset_clear(fg_id):
    """ADMIN: soft-delete all non-deleted requirement entries for this FG."""
    try:
        fg = db.get_fg(fg_id)
        if not fg:
            return jsonify({'ok': False, 'error': 'FG not found'}), 404

        user_name = session.get('User_Name') or session.get('UID') or ''
        n_deleted = db.admin_reset_fg_requirement_clear_all(fg_id, user_name=user_name)
        _audit('admin_reset_clear', 'fg_requirement', fg_id,
               summary=(f"Admin cleared {n_deleted} requirement entr"
                        f"{'y' if n_deleted == 1 else 'ies'} for "
                        f"'{fg.get('product_name','')}'"),
               details={'fg_id': fg_id, 'entries_soft_deleted': n_deleted,
                        'product_name': fg.get('product_name','')})
        return jsonify({'ok': True, 'entries_soft_deleted': n_deleted,
                        'fg_id': fg_id, 'product_name': fg.get('product_name','')})
    except Exception as e:
        # Log full traceback for the server admin, return a clean message to the UI
        import traceback
        print(f"[reset_clear_all] FAILED for fg_id={fg_id}: {e}")
        traceback.print_exc()
        return jsonify({'ok': False,
                        'error': f"Reset failed: {type(e).__name__}: {e}"}), 500


@hcp_stock_bp.route('/api/fg/<int:fg_id>/requirement/reset_override', methods=['POST'])
@login_required
@hcp_stock_role_required
@admin_only
def api_admin_fg_requirement_reset_override(fg_id):
    """ADMIN: directly set requirement_qty without touching history."""
    try:
        fg = db.get_fg(fg_id)
        if not fg:
            return jsonify({'ok': False, 'error': 'FG not found'}), 404

        d = request.json or {}
        qty_raw = d.get('quantity', 0)
        try:
            qty = float(qty_raw)
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Invalid quantity — must be a number'}), 400
        if qty < 0:
            return jsonify({'ok': False, 'error': 'Quantity cannot be negative'}), 400

        user_name = session.get('User_Name') or session.get('UID') or ''
        result = db.admin_reset_fg_requirement_override(fg_id, qty, user_name=user_name)
        if result is None:
            return jsonify({'ok': False, 'error': 'FG not found'}), 404

        _audit('admin_reset_override', 'fg_requirement', fg_id,
               summary=(f"Admin override: requirement_qty for '{fg.get('product_name','')}' "
                        f"set from {result['previous_qty']} → {result['new_qty']} "
                        f"(history sum: {result['history_sum']}, "
                        f"divergence: {result['divergence']:+g})"),
               details={
                   'fg_id':         fg_id,
                   'product_name':  fg.get('product_name',''),
                   'previous_qty':  result['previous_qty'],
                   'new_qty':       result['new_qty'],
                   'history_sum':   result['history_sum'],
                   'divergence':    result['divergence'],
               })
        return jsonify({'ok': True, **result, 'fg_id': fg_id,
                        'product_name': fg.get('product_name','')})
    except Exception as e:
        import traceback
        print(f"[reset_override] FAILED for fg_id={fg_id}: {e}")
        traceback.print_exc()
        return jsonify({'ok': False,
                        'error': f"Override failed: {type(e).__name__}: {e}"}), 500


@hcp_stock_bp.route('/api/reports/requirement_vs_dispatch', methods=['GET'])
@login_required
@hcp_stock_role_required
@feature_required('report_req_dispatch')
def api_requirement_vs_dispatch():
    """Variance report — per-FG totals of requirement vs dispatch.
    Query params: from=YYYY-MM-DD, to=YYYY-MM-DD, brand_id."""
    start = (request.args.get('from') or '').strip() or None
    end   = (request.args.get('to')   or '').strip() or None
    bid   = request.args.get('brand_id', type=int)
    rows  = db.requirement_vs_dispatch_report(start_date=start, end_date=end, brand_id=bid)

    # Aggregate totals across all rows
    tot_req  = sum(r['required_qty']   for r in rows)
    tot_disp = sum(r['dispatched_qty'] for r in rows)
    tot_short = sum(r['shortfall']     for r in rows)
    tot_surp  = sum(r['surplus']       for r in rows)

    return jsonify({
        'ok':    True,
        'date_range': {'from': start, 'to': end},
        'brand_id': bid,
        'rows':  rows,
        'count': len(rows),
        'totals': {
            'required_qty':    tot_req,
            'dispatched_qty':  tot_disp,
            'shortfall':       tot_short,
            'surplus':         tot_surp,
            'achievement_pct': round((tot_disp / tot_req * 100), 2) if tot_req > 0 else None,
        },
    })


# ───────────────────────────────────────────────────────────────────────────
# REPORTS HUB  —  PM Shortage, FG Buildability, PM Ledger, Dispatch Summary
# ───────────────────────────────────────────────────────────────────────────
@hcp_stock_bp.route('/api/reports/pm_shortage', methods=['GET'])
@login_required
@hcp_stock_role_required
@feature_required('report_pm_shortage')
def api_report_pm_shortage():
    """PM vs pending requirement: shortfall and suggested order quantity."""
    bid = request.args.get('brand_id', type=int)
    return jsonify({'ok': True, **db.report_pm_shortage(brand_id=bid)})


@hcp_stock_bp.route('/api/reports/buildability', methods=['GET'])
@login_required
@hcp_stock_role_required
@feature_required('report_buildability')
def api_report_buildability():
    """Per-FG: pending requirement vs buildable units, and the bottleneck PM."""
    bid = request.args.get('brand_id', type=int)
    return jsonify({'ok': True, **db.report_fg_buildability(brand_id=bid)})


@hcp_stock_bp.route('/api/reports/pm_ledger', methods=['GET'])
@login_required
@hcp_stock_role_required
@feature_required('report_pm_ledger')
def api_report_pm_ledger():
    """Running stock ledger for one PM. Requires ?pm_id=; optional from/to."""
    pm_id = request.args.get('pm_id', type=int)
    if not pm_id:
        return jsonify({'ok': False, 'error': 'pm_id is required'}), 400
    start = (request.args.get('from') or '').strip() or None
    end   = (request.args.get('to')   or '').strip() or None
    res = db.report_pm_ledger(pm_id, start_date=start, end_date=end)
    if not res.get('pm'):
        return jsonify({'ok': False, 'error': 'PM not found'}), 404
    return jsonify({'ok': True, **res})


@hcp_stock_bp.route('/api/reports/dispatch_summary', methods=['GET'])
@login_required
@hcp_stock_role_required
@feature_required('report_dispatch_summary')
def api_report_dispatch_summary():
    """FG dispatch totals grouped by brand / category / product / month."""
    bid   = request.args.get('brand_id', type=int)
    start = (request.args.get('from') or '').strip() or None
    end   = (request.args.get('to')   or '').strip() or None
    group = (request.args.get('group_by') or 'brand').strip().lower()
    return jsonify({'ok': True, **db.report_dispatch_summary(
        brand_id=bid, start_date=start, end_date=end, group_by=group)})


# ───────────────────────────────────────────────────────────────────────────
# REQUIREMENTS GRID  —  per-requirement ledger with FIFO dispatch tracking
# ───────────────────────────────────────────────────────────────────────────
@hcp_stock_bp.route('/api/requirement_ledger', methods=['GET'])
@login_required
@hcp_stock_role_required
@feature_required('view_requirements')
def api_requirement_ledger():
    """Feeds the standalone Requirements grid.

    Each row = one requirement entry with:
      entered, dispatched (FIFO-allocated FG dispatches on/after entry_date),
      preclosed, pending (= entered − dispatched − preclosed), and preclosure
      metadata. Pending is the figure reports consume.

    Query params: brand_id, fg_id, from, to (filter requirement entry_date),
                  include_preclosed=0|1 (default 1), as_of=YYYY-MM-DD.
    """
    brand_id = request.args.get('brand_id', type=int)
    fg_id    = request.args.get('fg_id',    type=int)
    start    = (request.args.get('from') or '').strip() or None
    end      = (request.args.get('to')   or '').strip() or None
    as_of    = (request.args.get('as_of') or '').strip() or None
    inc_pre  = (request.args.get('include_preclosed', '1').strip().lower()
                not in ('0', 'false', 'no'))

    rows = db.list_requirement_ledger(
        brand_id=brand_id, fg_id=fg_id,
        start_date=start, end_date=end,
        include_preclosed=inc_pre, as_of_date=as_of,
    )
    totals = {
        'entered':    round(sum(r['entered']    for r in rows), 4),
        'dispatched': round(sum(r['dispatched'] for r in rows), 4),
        'preclosed':  round(sum(r['preclosed']  for r in rows), 4),
        'pending':    round(sum(r['pending']    for r in rows), 4),
        'open_count':     sum(1 for r in rows if not r['is_preclosed']),
        'preclosed_count': sum(1 for r in rows if r['is_preclosed']),
    }
    return jsonify({'ok': True, 'rows': rows, 'count': len(rows),
                    'totals': totals,
                    'filters': {'brand_id': brand_id, 'fg_id': fg_id,
                                'from': start, 'to': end,
                                'include_preclosed': inc_pre, 'as_of': as_of}})


@hcp_stock_bp.route('/api/fg_requirement/<int:req_id>/preclose', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('preclose_requirement')
def api_requirement_preclose(req_id):
    """Force-close the remaining balance of one requirement (pending → 0).
    Body (optional): {"note": "<reason>"}. Audited."""
    d = request.json or {}
    note = (d.get('note') or '').strip()
    user_name = session.get('User_Name') or session.get('UID') or ''
    try:
        res = db.preclose_requirement(req_id, user_name=user_name, note=note)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'{type(e).__name__}: {e}'}), 500
    if not res:
        return jsonify({'ok': False,
                        'error': 'Requirement not found or already preclosed'}), 404
    _audit('preclose', 'fg_requirement', req_id,
           summary=(f"Preclosed requirement #{req_id} — closed balance "
                    f"{res['preclosed_qty']} (entered {res['entered']}, "
                    f"dispatched {res['dispatched']})"),
           details={**res, 'note': note})
    return jsonify({'ok': True, **res})


@hcp_stock_bp.route('/api/fg_requirement/<int:req_id>/revoke_preclosure', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('revoke_preclosure')
def api_requirement_revoke_preclosure(req_id):
    """Undo a preclosure — the requirement re-enters the pending pool."""
    user_name = session.get('User_Name') or session.get('UID') or ''
    try:
        ok = db.revoke_preclosure(req_id, user_name=user_name)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'{type(e).__name__}: {e}'}), 500
    if not ok:
        return jsonify({'ok': False,
                        'error': 'Requirement not found or was not preclosed'}), 404
    _audit('revoke_preclosure', 'fg_requirement', req_id,
           summary=f"Revoked preclosure on requirement #{req_id}")
    return jsonify({'ok': True})


@hcp_stock_bp.route('/api/fg/<int:fg_id>/availability', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_fg_availability(fg_id):
    """Hover-panel data: BOM lines enriched with PM stock for the requested
    date window, required qty, short qty, and possible-units.

    Query params:
      qty=N         hypothetical requirement (override stored)
      from=YYYY-MM-DD  inclusive lower bound on transaction dates
      to=YYYY-MM-DD    inclusive upper bound
    """
    fg = db.get_fg(fg_id)
    if not fg:
        return jsonify({'ok': False, 'error': 'FG not found'}), 404

    qty_override = request.args.get('qty')
    # The PM "can build" / required calc ALWAYS uses the FG's PENDING requirement:
    #   pending = (sum of all requirement entries) − dispatched − preclosed.
    # An explicit ?qty= override (used by the manual planner) still wins when given.
    basis = 'pending'
    start_date = (request.args.get('from') or '').strip() or None
    end_date   = (request.args.get('to')   or '').strip() or None

    if qty_override not in (None, ''):
        # Manual override: compute against an explicit requirement number.
        requirement = _f(qty_override)
        lines = db.fg_bom_with_availability(fg_id, requirement_qty=requirement,
                                            start_date=start_date, end_date=end_date,
                                            basis='entered')
        basis = 'override'
    else:
        # Default: pending requirement from the ledger.
        lines = db.fg_bom_with_availability(fg_id, basis='pending',
                                            start_date=start_date, end_date=end_date)
        requirement = lines[0]['basis_requirement_qty'] if lines \
                      else db.fg_pending_requirement(fg_id, as_of_date=end_date)

    # FG-level "possible" = MIN of possible_units across all PMs
    possible_total = None
    if lines:
        possible_total = min((ln['possible_units'] for ln in lines), default=0)

    has_short = any(ln['short_qty'] > 0 for ln in lines)
    total_short = sum(ln['short_qty'] for ln in lines)

    return jsonify({
        'ok': True,
        'fg': {
            'id':           fg['id'],
            'product_code': fg.get('product_code') or '',
            'product_name': fg.get('product_name') or '',
            'requirement_qty': requirement,
        },
        'date_range': {'from': start_date, 'to': end_date},
        'basis':       basis,
        'bom_lines':   lines,
        'possible':    possible_total or 0,
        'has_short':   has_short,
        'total_short': total_short,
        'feasible':    (not has_short) if requirement > 0 else True,
    })


@hcp_stock_bp.route('/api/pm/<int:pm_id>/fgs', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_pm_used_in_fgs(pm_id):
    """Reverse BOM: which FG products consume this PM. Used by the PM hover panel."""
    return jsonify({'ok': True, 'fgs': db.get_fgs_using_pm(pm_id)})


@hcp_stock_bp.route('/api/pm/<int:pm_id>/stats', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_pm_stats(pm_id):
    """Date-aware PM stats for the hover panel.
    Query params: from=YYYY-MM-DD, to=YYYY-MM-DD (both optional)."""
    pm = db.get_pm(pm_id)
    if not pm:
        return jsonify({'ok': False, 'error': 'PM not found'}), 404
    start = (request.args.get('from') or '').strip() or None
    end   = (request.args.get('to')   or '').strip() or None
    stats = db.pm_stats(pm_id, start_date=start, end_date=end) or {}
    return jsonify({
        'ok': True,
        'pm': {'id': pm['id'], 'pm_code': pm.get('pm_code',''),
               'pm_name': pm.get('pm_name',''), 'pm_type': pm.get('pm_type',''),
               'sku_size': pm.get('sku_size',''),
               'rate': float(pm.get('rate') or 0),
               'low_stock_threshold': float(pm.get('low_stock_threshold') or 0)},
        'date_range': {'from': start, 'to': end},
        'stats': stats,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# INWARD (PM)
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/api/inward', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_inward_list():
    brand_id = request.args.get('brand_id', type=int)
    pm_id    = request.args.get('pm_id', type=int)
    return jsonify({'ok': True, 'rows': db.list_inward(brand_id=brand_id, pm_id=pm_id)})


@hcp_stock_bp.route('/api/inward', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('add_inward')
def api_inward_create():
    d = request.json or {}
    if not d.get('pm_id'):     return jsonify({'ok': False, 'error': 'PM is required'}), 400
    if _f(d.get('quantity')) <= 0: return jsonify({'ok': False, 'error': 'Quantity must be > 0'}), 400
    payload = {
        'entry_date': d.get('entry_date') or datetime.now().strftime('%Y-%m-%d'),
        'pm_id':      int(d['pm_id']),
        'quantity':   _f(d['quantity']),
        'ref_no':     (d.get('ref_no') or '').strip(),
        'remarks':    (d.get('remarks') or '').strip(),
    }
    new_id = db.add_inward(payload)
    pm = db.get_pm(payload['pm_id']) or {}
    _audit('create', 'inward', new_id,
           summary=f"Inward {payload['quantity']} of '{pm.get('pm_name','')}' on {_dmy(payload['entry_date'])}",
           details={'after': payload})
    return jsonify({'ok': True, 'id': new_id})


@hcp_stock_bp.route('/api/inward/batch', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('add_inward')
def api_inward_batch_create():
    """Atomic multi-line inward.
    Body: {"items": [{entry_date, pm_id, quantity, ref_no, remarks}, ...]}
    Either all lines commit, or none do (rollback on first bad row).
    """
    d = request.json or {}
    items = d.get('items') or []
    if not isinstance(items, list) or not items:
        return jsonify({'ok': False, 'error': 'items must be a non-empty list'}), 400

    # Build clean payloads + light pre-validation (deeper validation is in db layer)
    cleaned = []
    today = datetime.now().strftime('%Y-%m-%d')
    for i, it in enumerate(items, start=1):
        if not it.get('pm_id'):
            return jsonify({'ok': False, 'error': f'Line {i}: PM is required',
                            'line_index': i}), 400
        qty = _f(it.get('quantity'))
        if qty <= 0:
            return jsonify({'ok': False, 'error': f'Line {i}: quantity must be > 0',
                            'line_index': i}), 400
        cleaned.append({
            'entry_date': it.get('entry_date') or today,
            'pm_id':      int(it['pm_id']),
            'quantity':   qty,
            'ref_no':     (it.get('ref_no') or '').strip(),
            'remarks':    (it.get('remarks') or '').strip(),
        })

    try:
        ids = db.add_inward_batch(cleaned)
        # Audit each line individually for traceability
        for idx, (new_id, payload) in enumerate(zip(ids, cleaned), start=1):
            pm = db.get_pm(payload['pm_id']) or {}
            _audit('create', 'inward', new_id,
                   summary=f"[Batch line {idx}/{len(ids)}] Inward {payload['quantity']} of '{pm.get('pm_name','')}' on {_dmy(payload['entry_date'])}",
                   details={'after': payload, 'batch_size': len(ids)})
        return jsonify({'ok': True, 'ids': ids, 'count': len(ids)})
    except ValueError as e:
        return jsonify({
            'ok': False,
            'error': str(e),
            'line_index': getattr(e, 'line_index', None),
        }), 400


@hcp_stock_bp.route('/api/inward/<int:txn_id>', methods=['PUT'])
@login_required
@hcp_stock_role_required
@feature_required('edit_inward')
def api_inward_update(txn_id):
    d = request.json or {}
    payload = {
        'entry_date': d.get('entry_date'),
        'pm_id':      int(d.get('pm_id') or 0),
        'quantity':   _f(d.get('quantity')),
        'ref_no':     (d.get('ref_no') or '').strip(),
        'remarks':    (d.get('remarks') or '').strip(),
    }
    db.update_inward(txn_id, payload)
    pm = db.get_pm(payload['pm_id']) or {}
    _audit('update', 'inward', txn_id,
           summary=f"Updated inward → {payload['quantity']} of '{pm.get('pm_name','')}' on {_dmy(payload['entry_date'])}",
           details={'after': payload})
    return jsonify({'ok': True})


@hcp_stock_bp.route('/api/inward/<int:txn_id>', methods=['DELETE'])
@login_required
@hcp_stock_role_required
@feature_required('delete_inward')
def api_inward_delete(txn_id):
    u = _audit_user()
    db.delete_inward(txn_id, user_name=u['user_name'])
    _audit('delete', 'inward', txn_id,
           summary=f"Soft-deleted inward txn id={txn_id}",
           details=None)
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════════════════════════════════
# WASTAGE (PM)
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/api/wastage', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_wastage_list():
    brand_id = request.args.get('brand_id', type=int)
    pm_id    = request.args.get('pm_id', type=int)
    return jsonify({'ok': True, 'rows': db.list_wastage(brand_id=brand_id, pm_id=pm_id)})


@hcp_stock_bp.route('/api/wastage', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('add_wastage')
def api_wastage_create():
    d = request.json or {}
    if not d.get('pm_id'):     return jsonify({'ok': False, 'error': 'PM is required'}), 400
    if _f(d.get('quantity')) <= 0: return jsonify({'ok': False, 'error': 'Quantity must be > 0'}), 400
    payload = {
        'entry_date': d.get('entry_date') or datetime.now().strftime('%Y-%m-%d'),
        'pm_id':      int(d['pm_id']),
        'quantity':   _f(d['quantity']),
        'reason':     (d.get('reason') or '').strip(),
        'remarks':    (d.get('remarks') or '').strip(),
    }
    new_id = db.add_wastage(payload)
    pm = db.get_pm(payload['pm_id']) or {}
    _audit('create', 'wastage', new_id,
           summary=f"Wastage {payload['quantity']} of '{pm.get('pm_name','')}' on {_dmy(payload['entry_date'])}",
           details={'after': payload})
    return jsonify({'ok': True, 'id': new_id})


@hcp_stock_bp.route('/api/wastage/batch', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('add_wastage')
def api_wastage_batch_create():
    """Atomic multi-line wastage.
    Body: {"items": [{entry_date, pm_id, quantity, reason, remarks}, ...]}
    Validates cumulatively — each line's wastage decreases available stock
    BEFORE the next line is checked. Rolls back the entire batch on first
    short / bad row.
    """
    d = request.json or {}
    items = d.get('items') or []
    if not isinstance(items, list) or not items:
        return jsonify({'ok': False, 'error': 'items must be a non-empty list'}), 400

    cleaned = []
    today = datetime.now().strftime('%Y-%m-%d')
    for i, it in enumerate(items, start=1):
        if not it.get('pm_id'):
            return jsonify({'ok': False, 'error': f'Line {i}: PM is required',
                            'line_index': i}), 400
        qty = _f(it.get('quantity'))
        if qty <= 0:
            return jsonify({'ok': False, 'error': f'Line {i}: quantity must be > 0',
                            'line_index': i}), 400
        if not (it.get('reason') or '').strip():
            return jsonify({'ok': False, 'error': f'Line {i}: reason is required',
                            'line_index': i}), 400
        cleaned.append({
            'entry_date': it.get('entry_date') or today,
            'pm_id':      int(it['pm_id']),
            'quantity':   qty,
            'reason':     (it.get('reason') or '').strip(),
            'remarks':    (it.get('remarks') or '').strip(),
        })

    try:
        ids = db.add_wastage_batch(cleaned)
        for idx, (new_id, payload) in enumerate(zip(ids, cleaned), start=1):
            pm = db.get_pm(payload['pm_id']) or {}
            _audit('create', 'wastage', new_id,
                   summary=f"[Batch line {idx}/{len(ids)}] Wastage {payload['quantity']} of '{pm.get('pm_name','')}' "
                           f"on {_dmy(payload['entry_date'])} (reason: {payload['reason']})",
                   details={'after': payload, 'batch_size': len(ids)})
        return jsonify({'ok': True, 'ids': ids, 'count': len(ids)})
    except db.StockShortError as e:
        return jsonify({
            'ok': False,
            'error': f"Wastage exceeds available stock at line {getattr(e, 'line_index', '?')}",
            'line_index': getattr(e, 'line_index', None),
            'shortages': e.shortages,
        }), 409
    except ValueError as e:
        return jsonify({
            'ok': False,
            'error': str(e),
            'line_index': getattr(e, 'line_index', None),
        }), 400


@hcp_stock_bp.route('/api/wastage/<int:txn_id>', methods=['PUT'])
@login_required
@hcp_stock_role_required
@feature_required('edit_wastage')
def api_wastage_update(txn_id):
    d = request.json or {}
    payload = {
        'entry_date': d.get('entry_date'),
        'pm_id':      int(d.get('pm_id') or 0),
        'quantity':   _f(d.get('quantity')),
        'reason':     (d.get('reason') or '').strip(),
        'remarks':    (d.get('remarks') or '').strip(),
    }
    db.update_wastage(txn_id, payload)
    pm = db.get_pm(payload['pm_id']) or {}
    _audit('update', 'wastage', txn_id,
           summary=f"Updated wastage → {payload['quantity']} of '{pm.get('pm_name','')}' on {_dmy(payload['entry_date'])}",
           details={'after': payload})
    return jsonify({'ok': True})


@hcp_stock_bp.route('/api/wastage/<int:txn_id>', methods=['DELETE'])
@login_required
@hcp_stock_role_required
@feature_required('delete_wastage')
def api_wastage_delete(txn_id):
    u = _audit_user()
    db.delete_wastage(txn_id, user_name=u['user_name'])
    _audit('delete', 'wastage', txn_id,
           summary=f"Soft-deleted wastage txn id={txn_id}",
           details=None)
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════════════════════════════════
# DISPATCH (FG) — auto-consumes PM via BOM
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/api/dispatch', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_dispatch_list():
    brand_id = request.args.get('brand_id', type=int)
    fg_id    = request.args.get('fg_id', type=int)
    return jsonify({'ok': True, 'rows': db.list_dispatch(brand_id=brand_id, fg_id=fg_id)})


@hcp_stock_bp.route('/api/dispatch', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('add_dispatch')
def api_dispatch_create():
    d = request.json or {}
    if not d.get('fg_id'):     return jsonify({'ok': False, 'error': 'Product is required'}), 400
    if _f(d.get('quantity')) <= 0: return jsonify({'ok': False, 'error': 'Quantity must be > 0'}), 400
    payload = {
        'entry_date': d.get('entry_date') or datetime.now().strftime('%Y-%m-%d'),
        'fg_id':      int(d['fg_id']),
        'quantity':   _f(d['quantity']),
        'ref_no':     (d.get('ref_no') or '').strip(),
        'remarks':    (d.get('remarks') or '').strip(),
        'deduct_rejection': bool(d.get('deduct_rejection')),
        'rejection_pct':    _f(d.get('rejection_pct')),
    }
    try:
        new_id = db.add_dispatch(payload)
        fg = db.get_fg(payload['fg_id']) or {}
        rej_note = (f" + {payload['rejection_pct'] or 2:g}% rejection wastage"
                    if payload['deduct_rejection'] else "")
        _audit('create', 'dispatch', new_id,
               summary=f"Dispatched {payload['quantity']} of '{fg.get('product_name','')}' on {_dmy(payload['entry_date'])}{rej_note}",
               details={'after': payload})
        return jsonify({'ok': True, 'id': new_id})
    except db.StockShortError as e:
        return jsonify({'ok': False, 'error': 'Insufficient PM stock for BOM',
                        'shortages': e.shortages}), 409


@hcp_stock_bp.route('/api/dispatch/batch', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('add_dispatch')
def api_dispatch_batch_create():
    """Atomic multi-line dispatch.
    Body: {"items": [{entry_date, fg_id, quantity, ref_no, remarks}, ...]}
    Either all lines commit, or none do (rollback on first shortage / bad row).
    """
    d = request.json or {}
    items = d.get('items') or []
    if not isinstance(items, list) or not items:
        return jsonify({'ok': False, 'error': 'items must be a non-empty list'}), 400

    # Build clean payloads + remember user-facing line numbers
    cleaned = []
    today = datetime.now().strftime('%Y-%m-%d')
    for i, it in enumerate(items, start=1):
        if not it.get('fg_id'):
            return jsonify({'ok': False, 'error': f'Line {i}: product is required'}), 400
        qty = _f(it.get('quantity'))
        if qty <= 0:
            return jsonify({'ok': False, 'error': f'Line {i}: quantity must be > 0'}), 400
        cleaned.append({
            'entry_date': it.get('entry_date') or today,
            'fg_id':      int(it['fg_id']),
            'quantity':   qty,
            'ref_no':     (it.get('ref_no') or '').strip(),
            'remarks':    (it.get('remarks') or '').strip(),
            'deduct_rejection': bool(it.get('deduct_rejection')),
            'rejection_pct':    _f(it.get('rejection_pct')),
        })

    try:
        ids = db.batch_add_dispatch(cleaned)
        # Audit each line, plus a summary entry
        for idx, (new_id, payload) in enumerate(zip(ids, cleaned), start=1):
            fg = db.get_fg(payload['fg_id']) or {}
            _audit('create', 'dispatch', new_id,
                   summary=f"[Batch line {idx}/{len(ids)}] Dispatched {payload['quantity']} of '{fg.get('product_name','')}' on {_dmy(payload['entry_date'])}",
                   details={'after': payload, 'batch_size': len(ids)})
        return jsonify({'ok': True, 'ids': ids, 'count': len(ids)})
    except db.StockShortError as e:
        return jsonify({
            'ok': False,
            'error': f"Insufficient PM stock at line {getattr(e, 'line_index', '?')}",
            'line_index': getattr(e, 'line_index', None),
            'shortages': e.shortages,
        }), 409
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


@hcp_stock_bp.route('/api/dispatch/<int:txn_id>', methods=['PUT'])
@login_required
@hcp_stock_role_required
@feature_required('edit_dispatch')
def api_dispatch_update(txn_id):
    d = request.json or {}
    payload = {
        'entry_date': d.get('entry_date'),
        'fg_id':      int(d.get('fg_id') or 0),
        'quantity':   _f(d.get('quantity')),
        'ref_no':     (d.get('ref_no') or '').strip(),
        'remarks':    (d.get('remarks') or '').strip(),
    }
    try:
        db.update_dispatch(txn_id, payload)
        fg = db.get_fg(payload['fg_id']) or {}
        _audit('update', 'dispatch', txn_id,
               summary=f"Updated dispatch → {payload['quantity']} of '{fg.get('product_name','')}' on {_dmy(payload['entry_date'])}",
               details={'after': payload})
        return jsonify({'ok': True})
    except db.StockShortError as e:
        return jsonify({'ok': False, 'error': 'Insufficient PM stock for BOM',
                        'shortages': e.shortages}), 409


@hcp_stock_bp.route('/api/dispatch/<int:txn_id>', methods=['DELETE'])
@login_required
@hcp_stock_role_required
@feature_required('delete_dispatch')
def api_dispatch_delete(txn_id):
    u = _audit_user()
    db.delete_dispatch(txn_id, user_name=u['user_name'])
    _audit('delete', 'dispatch', txn_id,
           summary=f"Soft-deleted dispatch txn id={txn_id}",
           details=None)
    return jsonify({'ok': True})


@hcp_stock_bp.route('/api/dispatch/<int:txn_id>/consumption', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_dispatch_consume(txn_id):
    return jsonify({'ok': True, 'lines': db.get_dispatch_consumption(txn_id)})


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL — TEMPLATE / IMPORT / EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
def _style_header(ws, last_col_letter, color='1E3A8A'):
    fill = PatternFill('solid', start_color=color)
    font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in ws[1]:
        cell.fill = fill; cell.font = font; cell.alignment = align
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{last_col_letter}1'


_PM_HEADERS = ['Brand Name', 'PM Code', 'PM Name', 'PM Type', 'SKU Size',
               'Rate', 'Opening Stock', 'Provisional Wastage', 'Low Stock Threshold']
_FG_HEADERS = ['Brand Name', 'Product Code', 'Product Name', 'Category',
               'SKU Size', 'Rate']
_BOM_HEADERS     = ['FG Product Code', 'FG Product Name', 'PM Code',
                    'PM Name', 'Qty per Unit']
_INWARD_HEADERS  = ['Date (YYYY-MM-DD)', 'PM Code', 'PM Name',
                    'Quantity', 'Ref No.', 'Remarks']
_DISPATCH_HEADERS= ['Date (YYYY-MM-DD)', 'Product Code', 'Product Name',
                    'Quantity', 'Ref No.', 'Remarks']
_WASTAGE_HEADERS = ['Date (YYYY-MM-DD)', 'PM Code', 'PM Name',
                    'Quantity', 'Reason', 'Remarks']


@hcp_stock_bp.route('/template/download')
@login_required
@hcp_stock_role_required
@feature_required('data_template')
def template_download():
    """Multi-sheet template covering ALL importable entities:
    PM master, FG master, BOM, Inward, Dispatch, Wastage."""
    wb = Workbook()

    # ── PM master ──
    ws = wb.active; ws.title = 'PM'
    ws.append(_PM_HEADERS)
    ws.append(['PLIX', 'BTL-30', '30ml Bottle', 'Bottle', '30ml',
               2.0, 10000, 50, 500])
    for i, w in enumerate([16, 22, 30, 16, 12, 10, 14, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _style_header(ws, get_column_letter(len(_PM_HEADERS)))

    # ── FG master ──
    ws2 = wb.create_sheet('FG')
    ws2.append(_FG_HEADERS)
    ws2.append(['PLIX', 'PLIX-RHS-30', 'Rosemary Hair Serum 30ml',
                'Hair Serum', '30ml', 199.0])
    for i, w in enumerate([16, 22, 38, 18, 12, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    _style_header(ws2, get_column_letter(len(_FG_HEADERS)), color='0d9488')

    # ── BOM ──
    wb_b = wb.create_sheet('BOM')
    wb_b.append(_BOM_HEADERS)
    wb_b.append(['PLIX-RHS-30', 'Rosemary Hair Serum 30ml',
                 'BTL-30', '30ml Bottle', 1])
    wb_b.append(['PLIX-RHS-30', 'Rosemary Hair Serum 30ml',
                 'CAP-DRP', 'Dropper Cap', 1])
    for i, w in enumerate([22, 38, 22, 32, 14], 1):
        wb_b.column_dimensions[get_column_letter(i)].width = w
    _style_header(wb_b, get_column_letter(len(_BOM_HEADERS)), color='a855f7')

    # ── Inward ──
    ws_in = wb.create_sheet('Inward')
    ws_in.append(_INWARD_HEADERS)
    ws_in.append(['2026-04-01', 'BTL-30', '30ml Bottle', 5000,
                  'GRN-001', 'Initial inward'])
    for i, w in enumerate([18, 22, 32, 12, 18, 24], 1):
        ws_in.column_dimensions[get_column_letter(i)].width = w
    _style_header(ws_in, get_column_letter(len(_INWARD_HEADERS)), color='10b981')

    # ── Dispatch ──
    ws_dp = wb.create_sheet('Dispatch')
    ws_dp.append(_DISPATCH_HEADERS)
    ws_dp.append(['2026-04-15', 'PLIX-RHS-30', 'Rosemary Hair Serum 30ml',
                  100, 'DC-001', 'First dispatch'])
    for i, w in enumerate([18, 22, 38, 12, 18, 24], 1):
        ws_dp.column_dimensions[get_column_letter(i)].width = w
    _style_header(ws_dp, get_column_letter(len(_DISPATCH_HEADERS)), color='f59e0b')

    # ── Wastage ──
    ws_wt = wb.create_sheet('Wastage')
    ws_wt.append(_WASTAGE_HEADERS)
    ws_wt.append(['2026-04-20', 'BTL-30', '30ml Bottle', 50,
                  'Damaged in transit', ''])
    for i, w in enumerate([18, 22, 32, 12, 22, 24], 1):
        ws_wt.column_dimensions[get_column_letter(i)].width = w
    _style_header(ws_wt, get_column_letter(len(_WASTAGE_HEADERS)), color='ef4444')

    # ── Brands reference ──
    ws3 = wb.create_sheet('Brands')
    ws3.append(['Brand Name'])
    for b in db.list_brands():
        ws3.append([b['brand_name']])
    _style_header(ws3, 'A', color='6366f1')
    ws3.column_dimensions['A'].width = 24

    # ── Instructions ──
    ws4 = wb.create_sheet('Instructions')
    ws4['A1'] = 'HCP Stock — Import Template'
    ws4['A1'].font = Font(bold=True, size=14, color='1E3A8A')
    notes = [
        '',
        'Sheets in this workbook:',
        '   • PM        — Packing Material master (brand + opening stock + low-stock threshold)',
        '   • FG        — Finished Good (Product) master',
        '   • BOM       — Links each FG to its PMs with qty-per-unit',
        '   • Inward    — PM inward register (receipts)',
        '   • Dispatch  — FG dispatch register (auto-consumes PM via BOM)',
        '   • Wastage   — PM actual wastage register',
        '',
        'Rules:',
        '1. Brand Name MUST already exist in procurement_brands (see "Brands" sheet).',
        '2. PM Name and Product Name are mandatory in the master sheets.',
        '3. For BOM / Inward / Dispatch / Wastage rows, the referenced PM or FG must',
        '   already exist (either imported in the same workbook or already in the DB).',
        '   Match is by PM Code / Product Code (preferred) — if blank we fall back to name.',
        '4. Date columns must be in YYYY-MM-DD format.',
        '5. Dispatches will auto-deduct PM via BOM. If any PM goes negative the row is',
        '   skipped and reported in the import errors.',
        '6. Existing PMs/FGs with the same Code are UPDATED, not duplicated.',
        '7. BOM rows are upserted (replace by FG+PM unique key).',
        '8. Inward/Dispatch/Wastage rows are always INSERTED (history is append-only).',
        '',
        'Import processing order: PM → FG → BOM → Inward → Dispatch → Wastage.',
        'This way the masters are always present before transactions reference them.',
    ]
    for i, n in enumerate(notes, 2):
        ws4.cell(row=i, column=1, value=n)
    ws4.column_dimensions['A'].width = 100

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True,
                     download_name=f'HCP_Stock_Template_{datetime.now().strftime("%Y%m%d")}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _brand_id_by_name(name):
    if not name: return None
    for b in db.list_brands():
        if b['brand_name'].strip().lower() == str(name).strip().lower():
            return b['id']
    return None


def _norm_match(s):
    """Normalize a code/name for matching: drop non-breaking spaces, collapse
    internal whitespace runs to one space, strip, lowercase. Mirrors the header
    normalization in _make_col so BOM cells match master rows even when the user
    typed extra spaces, trailing spaces, or pasted from a styled cell."""
    import re as _re
    s = str(s or '').replace('\xa0', ' ')
    return _re.sub(r'\s+', ' ', s).strip().lower()


def _pm_lookup_by_code_or_name(code, name):
    """Find a PM by code first (preferred), then by name. Returns id or None."""
    code = _norm_match(code)
    name = _norm_match(name)
    rows = db.get_pm_rows()
    if code:
        for r in rows:
            if _norm_match(r.get('pm_code')) == code:
                return r['id']
    if name:
        for r in rows:
            if _norm_match(r.get('pm_name')) == name:
                return r['id']
    return None


def _fg_lookup_by_code_or_name(code, name):
    code = _norm_match(code)
    name = _norm_match(name)
    rows = db.get_fg_rows()
    if code:
        for r in rows:
            if _norm_match(r.get('product_code')) == code:
                return r['id']
    if name:
        for r in rows:
            if _norm_match(r.get('product_name')) == name:
                return r['id']
    return None


def _format_date(v):
    """Coerce a cell value to YYYY-MM-DD string for storage.
    Accepts real datetimes, ISO 'YYYY-MM-DD', and human 'DD/MM/YYYY' or
    'DD-MM-YYYY' (the format this module displays/exports), so a file exported
    by this module round-trips cleanly back through import."""
    if v is None or v == '':
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s:
        return None
    import re as _re
    # Already ISO (optionally with time) → keep the date part.
    m = _re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # DD/MM/YYYY or DD-MM-YYYY (also accepts 2-digit year).
    m = _re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        # Guard against MM/DD slips: if first field > 12 it must be the day.
        if d > 31 or mo > 12:
            return s  # ambiguous/invalid — leave as-is rather than corrupt
        return f"{y:04d}-{mo:02d}-{d:02d}"
    return s


def _make_col(headers_row):
    """Build a {header_normalized → index} lookup for resolving columns by header name.
    Normalization:
      • strip leading/trailing whitespace (incl. non-breaking space \\xa0)
      • collapse multiple internal whitespace runs to a single space
      • lowercase
    This handles common Excel artefacts where headers were typed with extra
    spaces, copy-pasted from styled documents, or auto-formatted by Excel."""
    import re as _re
    def _norm(s):
        s = str(s or '').replace('\xa0', ' ')   # non-breaking space → regular
        return _re.sub(r'\s+', ' ', s).strip().lower()
    headers = [_norm(c.value) for c in headers_row]
    idx = {h: i for i, h in enumerate(headers) if h}
    def col(row, name):
        i = idx.get(_norm(name))
        return row[i] if i is not None and i < len(row) else None
    col._headers = headers       # exposed for debugging
    col._index   = idx
    return col


@hcp_stock_bp.route('/import', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('data_import')
def import_excel():
    f = request.files.get('file')
    if not f:
        return jsonify({'ok': False, 'error': 'No file uploaded'}), 400

    # ── Two-phase import controls ────────────────────────────────────────────
    #  dry_run=1   → validate only, write NOTHING. Returns errors + a list of
    #                PM/FG duplicates already present in the DB, plus the counts
    #                that *would* be imported. Lets the UI gently ask the user
    #                to resolve issues first (not forced) and warn about dupes.
    #  dup_mode    → how to treat a PM/FG that already exists in the DB:
    #                  'add'  (default) → ADD the imported quantity onto the
    #                                     existing item's stock (no new row)
    #                  'skip'           → leave existing item untouched
    def _truthy(v): return str(v).strip().lower() in ('1', 'true', 'yes', 'on')
    dry_run  = _truthy(request.form.get('dry_run') or request.args.get('dry_run'))
    dup_mode = (request.form.get('dup_mode') or request.args.get('dup_mode')
                or 'add').strip().lower()
    if dup_mode not in ('add', 'skip'):
        dup_mode = 'add'

    try:
        wb = load_workbook(f, data_only=True)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Cannot read Excel: {e}'}), 400

    # ─────────────────────────────────────────────────────────────────────────
    # Auto-detect file format.
    #
    #   "template" format  → user filled in our generated template
    #                        Sheets: PM, FG, BOM, Inward, Dispatch, Wastage
    #                        Headers in row 1, "Brand Name" column required.
    #
    #   "working" format   → the team's actual day-to-day file (HCP_WORKING__*.xlsx)
    #                        Sheets: 'Pm Data', 'Inward', 'Dispatch', 'Requirement'
    #                        Headers in row 3 (rows 1-2 are title + totals).
    #                        Brand inferred from the file (one workbook = one brand).
    #                        Inward & Dispatch sheets are wide-format pivots.
    # ─────────────────────────────────────────────────────────────────────────
    def _looks_like_template():
        # Standard template has 'PM' sheet with 'Brand Name' in row 1
        if 'PM' not in wb.sheetnames: return False
        ws = wb['PM']
        if ws.max_row < 1: return False
        h = [str(c.value or '').strip().lower() for c in ws[1]]
        return 'brand name' in h and 'pm name' in h

    def _looks_like_working():
        return 'Pm Data' in wb.sheetnames or 'PM Data' in wb.sheetnames

    if _looks_like_template():
        return _import_template_format(wb, dry_run=dry_run, dup_mode=dup_mode)
    elif _looks_like_working():
        return _import_working_format(wb)
    else:
        return jsonify({
            'ok': False,
            'error': "File format not recognized. Expected either:\n"
                     "  • Standard template with 'PM' sheet + 'Brand Name' column, OR\n"
                     "  • Working format with 'Pm Data' sheet (HCP team layout)\n"
                     f"Sheets found in your file: {wb.sheetnames}",
        }), 400


def _import_template_format(wb, dry_run=False, dup_mode='add'):
    """Original template-style importer (sheets: PM, FG, BOM, Inward, Dispatch, Wastage).

    Two-phase aware:
      • dry_run=True  → validate only, write nothing; collect `duplicates`.
      • dup_mode      → 'add' merges quantity onto an existing PM/FG; 'skip'
                        leaves the existing item untouched (no new row either way).
    """
    counts = {'pm': 0, 'fg': 0, 'bom': 0, 'inward': 0, 'dispatch': 0, 'wastage': 0}
    scanned = {'pm': 0, 'fg': 0, 'bom': 0, 'inward': 0, 'dispatch': 0, 'wastage': 0}
    merged = {'pm': 0, 'fg': 0}          # how many existing items got quantity added
    errors = []
    duplicates = []   # [{kind:'pm'|'fg', name, code, existing_stock, add_qty, row}]
    sheet_diag = {}  # per-sheet: header list + missing required columns

    # ── 1. PM master (must come first — everything else can reference PMs) ──
    if 'PM' in wb.sheetnames:
        ws = wb['PM']
        col = _make_col(ws[1])
        sheet_diag['PM'] = {'headers': list(col._headers),
                            'missing': [h for h in ('Brand Name','PM Name')
                                        if h.lower() not in col._index]}
        if sheet_diag['PM']['missing']:
            errors.append(f"PM sheet: missing required column(s): "
                          f"{', '.join(sheet_diag['PM']['missing'])}. "
                          f"Headers found: {sheet_diag['PM']['headers']}")
        else:
            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Skip rows where every cell is empty/None.  Note: a row whose
                # only filled value is "0" is still a valid candidate (could be
                # a PM with opening_stock=0), so we test ALL trimmed values.
                if all((v is None or str(v).strip() == '') for v in row): continue
                scanned['pm'] += 1
                brand = col(row, 'Brand Name'); name = col(row, 'PM Name')
                if not brand or not str(brand).strip():
                    errors.append(f'PM row {r_idx}: missing brand'); continue
                if not name or not str(name).strip():
                    errors.append(f'PM row {r_idx}: missing PM name'); continue
                bid = _brand_id_by_name(str(brand))
                if not bid:
                    errors.append(f'PM row {r_idx}: brand "{brand}" not found in '
                                  f'procurement_brands master — add the brand first'); continue
                pm_code = str(col(row, 'PM Code') or '').strip()
                add_qty = _f(col(row, 'Opening Stock'))

                # ── Duplicate detection: does this PM already exist? ──
                existing = db.find_pm_by_code_or_name(bid, pm_code, str(name).strip())
                if existing:
                    duplicates.append({
                        'kind': 'pm', 'row': r_idx,
                        'name': str(name).strip(), 'code': pm_code,
                        'brand': str(brand).strip(),
                        'existing_stock': _f(existing.get('opening_stock')),
                        'add_qty': add_qty,
                    })
                    if dry_run:
                        continue
                    if dup_mode == 'skip':
                        continue
                    # dup_mode == 'add' → add the imported quantity onto stock
                    try:
                        db.add_pm_stock(existing['id'], add_qty)
                        merged['pm'] += 1
                    except Exception as e:
                        errors.append(f'PM row {r_idx} ("{name}"): {e}')
                    continue

                if dry_run:
                    counts['pm'] += 1   # would be inserted
                    continue
                try:
                    db.upsert_pm({
                        'brand_id':            bid,
                        'pm_code':             pm_code,
                        'pm_name':             str(name).strip(),
                        'pm_type':             str(col(row, 'PM Type') or '').strip(),
                        'sku_size':            str(col(row, 'SKU Size') or '').strip(),
                        'rate':                _f(col(row, 'Rate')),
                        'opening_stock':       add_qty,
                        'provisional_wastage': _f(col(row, 'Provisional Wastage')),
                        'low_stock_threshold': _f(col(row, 'Low Stock Threshold')),
                    })
                    counts['pm'] += 1
                except Exception as e:
                    errors.append(f'PM row {r_idx} ("{name}"): {e}')

    # ── 2. FG master ──
    if 'FG' in wb.sheetnames:
        ws = wb['FG']
        col = _make_col(ws[1])
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all((v is None or str(v).strip() == '') for v in row): continue
            scanned['fg'] += 1
            brand = col(row, 'Brand Name'); name = col(row, 'Product Name')
            if not brand or not str(brand).strip():
                errors.append(f'FG row {r_idx}: missing brand'); continue
            if not name or not str(name).strip():
                errors.append(f'FG row {r_idx}: missing product name'); continue
            bid = _brand_id_by_name(str(brand))
            if not bid:
                errors.append(f'FG row {r_idx}: brand "{brand}" not in procurement_brands'); continue
            fg_code = str(col(row, 'Product Code') or '').strip()

            # ── Duplicate detection: does this FG already exist? ──
            existing = db.find_fg_by_code_or_name(bid, fg_code, str(name).strip())
            if existing:
                duplicates.append({
                    'kind': 'fg', 'row': r_idx,
                    'name': str(name).strip(), 'code': fg_code,
                    'brand': str(brand).strip(),
                    'existing_stock': None,   # FG stock is derived, not stored
                    'add_qty': None,
                })
                # An FG has no opening-stock column to accumulate, so a duplicate
                # is simply left as-is (master details unchanged). Dispatch rows
                # in the same file still post against the existing FG.
                continue

            if dry_run:
                counts['fg'] += 1   # would be inserted
                continue
            try:
                db.upsert_fg({
                    'brand_id':     bid,
                    'product_code': fg_code,
                    'product_name': str(name).strip(),
                    'category':     str(col(row, 'Category') or '').strip(),
                    'sku_size':     str(col(row, 'SKU Size') or '').strip(),
                    'rate':         _f(col(row, 'Rate')),
                })
                counts['fg'] += 1
            except Exception as e:
                errors.append(f'FG row {r_idx}: {e}')

    # ── 3. BOM (replace strategy: per FG, accumulate then write once) ──
    #     Transactional sheets (BOM/Inward/Dispatch/Wastage) are skipped during a
    #     dry-run: their lookups depend on PM/FG rows that haven't been written
    #     yet, so validating them now would produce misleading "not found" noise.
    #     The real (non-dry-run) pass processes them normally.
    if not dry_run and 'BOM' in wb.sheetnames:
        ws = wb['BOM']
        col = _make_col(ws[1])
        # Header diagnostics: if the key columns are missing/renamed, every row
        # would otherwise fail with a misleading "not found". Surface it clearly.
        _bom_required = ('FG Product Code', 'FG Product Name', 'PM Code', 'PM Name')
        sheet_diag['BOM'] = {
            'headers': list(col._headers),
            'missing': [h for h in _bom_required if h.lower() not in col._index],
        }
        # An FG can be identified by code OR name (same for PM), so only flag a
        # real problem when BOTH identifier columns for a side are absent.
        _fg_cols_missing = ('fg product code' not in col._index
                            and 'fg product name' not in col._index)
        _pm_cols_missing = ('pm code' not in col._index
                            and 'pm name' not in col._index)
        if _fg_cols_missing or _pm_cols_missing:
            need = []
            if _fg_cols_missing: need.append("'FG Product Code' or 'FG Product Name'")
            if _pm_cols_missing: need.append("'PM Code' or 'PM Name'")
            errors.append("BOM sheet: required column(s) not found — need "
                          + " and ".join(need)
                          + f". Headers found: {sheet_diag['BOM']['headers']}")
        # group rows by FG (skip if the identifier columns are missing — the
        # single header error above is clearer than one error per row)
        bom_by_fg = {}
        if not (_fg_cols_missing or _pm_cols_missing):
          for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all((v is None or str(v).strip() == '') for v in row): continue
            scanned['bom'] += 1
            fg_code  = col(row, 'FG Product Code')
            fg_name  = col(row, 'FG Product Name')
            pm_code  = col(row, 'PM Code')
            pm_name  = col(row, 'PM Name')
            qty      = _f(col(row, 'Qty per Unit'))
            if qty <= 0: qty = 1.0

            fg_id = _fg_lookup_by_code_or_name(fg_code, fg_name)
            if not fg_id:
                errors.append(f'BOM row {r_idx}: FG not found — code="{str(fg_code or "").strip()}", '
                              f'name="{str(fg_name or "").strip()}". '
                              f'Make sure this product exists in the FG sheet/master first.'); continue
            pm_id = _pm_lookup_by_code_or_name(pm_code, pm_name)
            if not pm_id:
                errors.append(f'BOM row {r_idx}: PM not found — code="{str(pm_code or "").strip()}", '
                              f'name="{str(pm_name or "").strip()}". '
                              f'Make sure this material exists in the PM sheet/master first.'); continue
            bom_by_fg.setdefault(fg_id, []).append({'pm_id': pm_id,
                                                    'qty_per_unit': qty})
        # write each FG's BOM
        for fg_id, lines in bom_by_fg.items():
            try:
                # merge with existing BOM (don't wipe lines outside of import)
                existing = {ln['pm_id']: float(ln['qty_per_unit'])
                            for ln in db.get_bom_for_fg(fg_id)}
                for ln in lines:
                    existing[ln['pm_id']] = ln['qty_per_unit']
                db.set_bom_for_fg(fg_id, [{'pm_id': k, 'qty_per_unit': v}
                                          for k, v in existing.items()])
                counts['bom'] += len(lines)
            except Exception as e:
                errors.append(f'BOM for fg_id={fg_id}: {e}')

    # ── 4. Inward (PM transactions) ──
    if not dry_run and 'Inward' in wb.sheetnames:
        ws = wb['Inward']
        col = _make_col(ws[1])
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all((v is None or str(v).strip() == '') for v in row): continue
            scanned['inward'] += 1
            date    = _format_date(col(row, 'Date (YYYY-MM-DD)') or col(row, 'Date'))
            pm_code = col(row, 'PM Code'); pm_name = col(row, 'PM Name')
            qty     = _f(col(row, 'Quantity'))
            if not date:
                errors.append(f'Inward row {r_idx}: invalid date'); continue
            if qty <= 0:
                errors.append(f'Inward row {r_idx}: quantity must be > 0'); continue
            pm_id = _pm_lookup_by_code_or_name(pm_code, pm_name)
            if not pm_id:
                errors.append(f'Inward row {r_idx}: PM "{pm_code or pm_name}" not found'); continue
            try:
                db.add_inward({
                    'entry_date': date, 'pm_id': pm_id, 'quantity': qty,
                    'ref_no':     str(col(row, 'Ref No.') or '').strip(),
                    'remarks':    str(col(row, 'Remarks') or '').strip(),
                })
                counts['inward'] += 1
            except Exception as e:
                errors.append(f'Inward row {r_idx}: {e}')

    # ── 5. Dispatch (FG transactions — auto-consume PM via BOM) ──
    if not dry_run and 'Dispatch' in wb.sheetnames:
        ws = wb['Dispatch']
        col = _make_col(ws[1])
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all((v is None or str(v).strip() == '') for v in row): continue
            scanned['inward'] += 1
            date    = _format_date(col(row, 'Date (YYYY-MM-DD)') or col(row, 'Date'))
            fg_code = col(row, 'Product Code'); fg_name = col(row, 'Product Name')
            qty     = _f(col(row, 'Quantity'))
            if not date:
                errors.append(f'Dispatch row {r_idx}: invalid date'); continue
            if qty <= 0:
                errors.append(f'Dispatch row {r_idx}: quantity must be > 0'); continue
            fg_id = _fg_lookup_by_code_or_name(fg_code, fg_name)
            if not fg_id:
                errors.append(f'Dispatch row {r_idx}: FG "{fg_code or fg_name}" not found'); continue
            try:
                db.add_dispatch({
                    'entry_date': date, 'fg_id': fg_id, 'quantity': qty,
                    'ref_no':     str(col(row, 'Ref No.') or '').strip(),
                    'remarks':    str(col(row, 'Remarks') or '').strip(),
                })
                counts['dispatch'] += 1
            except db.StockShortError as e:
                shorts = ', '.join(f"{s['pm_name']}: short {s['short_by']}" for s in e.shortages)
                errors.append(f'Dispatch row {r_idx}: skipped (PM shortage — {shorts})')
            except Exception as e:
                errors.append(f'Dispatch row {r_idx}: {e}')

    # ── 6. Wastage (PM transactions) ──
    if not dry_run and 'Wastage' in wb.sheetnames:
        ws = wb['Wastage']
        col = _make_col(ws[1])
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all((v is None or str(v).strip() == '') for v in row): continue
            scanned['inward'] += 1
            date    = _format_date(col(row, 'Date (YYYY-MM-DD)') or col(row, 'Date'))
            pm_code = col(row, 'PM Code'); pm_name = col(row, 'PM Name')
            qty     = _f(col(row, 'Quantity'))
            if not date:
                errors.append(f'Wastage row {r_idx}: invalid date'); continue
            if qty <= 0:
                errors.append(f'Wastage row {r_idx}: quantity must be > 0'); continue
            pm_id = _pm_lookup_by_code_or_name(pm_code, pm_name)
            if not pm_id:
                errors.append(f'Wastage row {r_idx}: PM "{pm_code or pm_name}" not found'); continue
            try:
                db.add_wastage({
                    'entry_date': date, 'pm_id': pm_id, 'quantity': qty,
                    'reason':     str(col(row, 'Reason') or '').strip(),
                    'remarks':    str(col(row, 'Remarks') or '').strip(),
                })
                counts['wastage'] += 1
            except Exception as e:
                errors.append(f'Wastage row {r_idx}: {e}')

    # ── Dry-run: validate only, write nothing. Hand the UI everything it needs
    #     to (a) gently flag blocking errors and (b) warn about duplicates. ──
    if dry_run:
        return jsonify({
            'ok': True,
            'dry_run': True,
            'counts': counts,        # rows that WOULD be newly inserted
            'scanned': scanned,
            'errors': errors,
            'duplicates': duplicates,
            'sheet_diag': sheet_diag,
        })

    # Audit log the import as a single summary entry
    summary_parts = []
    for k, v in counts.items():
        if v: summary_parts.append(f"{v} {k}")
    if merged['pm']: summary_parts.append(f"{merged['pm']} pm stock-merged")
    if merged['fg']: summary_parts.append(f"{merged['fg']} fg matched")
    _audit('import', 'workbook', None,
           summary='Imported: ' + (', '.join(summary_parts) or 'nothing'),
           details={'counts': counts, 'merged': merged,
                    'dup_mode': dup_mode, 'errors': errors[:50]})

    return jsonify({
        'ok': True,
        'dry_run': False,
        'dup_mode': dup_mode,
        # legacy keys for the old UI message
        'pm_inserted': counts['pm'],
        'fg_inserted': counts['fg'],
        # full breakdown
        'counts':      counts,
        'merged':      merged,
        'duplicates':  duplicates,
        'scanned':     scanned,
        'errors':      errors,
        'sheet_diag':  sheet_diag,
    })


def _import_working_format(wb):
    """Importer for the team's day-to-day Excel layout.

    Sheets recognised:
      • 'Pm Data'  — PM master (headers in row 3)
      • 'Inward'   — wide-format inward (date columns in row 2)
      • 'Dispatch' — wide-format FG dispatch (date columns in row 2) — needs FG master
      • 'Requirement' — FG products + per-PM requirement (also seeds the FG master)

    Brand: derived from the workbook title in row 1 of 'Pm Data'.
    For example "Plix Closing Stock From Date 01/04/2026" → brand "PLIX" (matched
    case-insensitively against procurement_brands).
    """
    counts  = {'pm': 0, 'fg': 0, 'bom': 0, 'inward': 0, 'dispatch': 0, 'wastage': 0}
    scanned = {'pm': 0, 'fg': 0, 'bom': 0, 'inward': 0, 'dispatch': 0, 'wastage': 0}
    errors = []
    sheet_diag = {'_format': 'working'}

    # ── A. Identify brand from workbook title ──
    brand_id = None
    brand_label = None
    if 'Pm Data' in wb.sheetnames or 'PM Data' in wb.sheetnames:
        ws = wb['Pm Data' if 'Pm Data' in wb.sheetnames else 'PM Data']
        title = str(ws.cell(1, 1).value or '').strip()
        # Try every brand name against the title (case-insensitive substring match)
        for b in db.list_brands():
            bn = str(b.get('brand_name') or b.get('name') or '').strip()
            if bn and bn.lower() in title.lower():
                brand_id = b['id']; brand_label = bn
                break
    if not brand_id:
        # fall back: check if there's exactly ONE brand defined → use it.
        brands = db.list_brands()
        if len(brands) == 1:
            brand_id = brands[0]['id']
            brand_label = brands[0].get('brand_name') or brands[0].get('name')
        else:
            errors.append("Could not infer brand from workbook title. "
                          "Please ensure your brand name appears in cell A1 of 'Pm Data' sheet, "
                          "and that the brand exists in the procurement_brands master.")
            return jsonify({'ok': True, 'counts': counts, 'scanned': scanned,
                            'errors': errors, 'sheet_diag': sheet_diag,
                            'pm_inserted': 0, 'fg_inserted': 0})

    sheet_diag['_brand'] = brand_label

    # ── B. PM master from 'Pm Data' sheet ──
    pm_sheet_name = 'Pm Data' if 'Pm Data' in wb.sheetnames else 'PM Data'
    ws = wb[pm_sheet_name]
    # Headers are on row 3
    headers_raw = [str(ws.cell(3, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    sheet_diag[pm_sheet_name] = {'headers': headers_raw, 'header_row': 3, 'missing': []}

    def _norm(s):
        import re as _re
        return _re.sub(r'\s+', ' ', str(s or '').replace('\xa0', ' ')).strip().lower()

    # Map normalized header → column index (1-based for openpyxl)
    h2c = {}
    for c, h in enumerate(headers_raw, start=1):
        nh = _norm(h)
        if nh:
            h2c[nh] = c

    # Required header check
    if not any(k.startswith('sku') for k in h2c) and 'pm name' not in h2c:
        sheet_diag[pm_sheet_name]['missing'] = ['SKU or PM Name']
        errors.append(f"'{pm_sheet_name}' sheet: cannot find SKU/PM Name column on row 3.")
    else:
        # find the OS column — it might be named "OS 01/04/2026" or similar
        os_col = h2c.get('opening stock')
        if not os_col:
            for k in h2c:
                if k.startswith('os ') or k == 'os':
                    os_col = h2c[k]; break

        col_sku    = h2c.get('sku') or h2c.get('pm code')
        col_cat    = h2c.get('category')
        col_size   = h2c.get('size') or h2c.get('sku size')
        col_pmname = h2c.get('pm name')
        col_rate   = h2c.get('rate')
        col_wast   = h2c.get('provisional wastage')

        for r in range(4, ws.max_row + 1):
            sku  = ws.cell(r, col_sku).value if col_sku else None
            pmn  = ws.cell(r, col_pmname).value if col_pmname else None
            cat  = ws.cell(r, col_cat).value if col_cat else None
            size = ws.cell(r, col_size).value if col_size else None
            rate = ws.cell(r, col_rate).value if col_rate else None
            os_v = ws.cell(r, os_col).value if os_col else None
            waste = ws.cell(r, col_wast).value if col_wast else None

            sku_s = (str(sku).strip() if sku is not None else '')
            pmn_s = (str(pmn).strip() if pmn is not None else '')
            cat_s = (str(cat).strip() if cat is not None else '')
            size_s = (str(size).strip() if size is not None else '')

            # Skip totally-empty rows
            if not sku_s and not pmn_s and not cat_s and not size_s and rate in (None, '') and os_v in (None, ''):
                continue

            # Skip section divider rows (only column A populated, no other meaningful data)
            if sku_s and not pmn_s and not cat_s and not size_s and rate in (None, '') and os_v in (None, ''):
                continue

            scanned['pm'] += 1

            # Build PM record. Use SKU as code; if PM Name is empty, fall back to SKU.
            pm_name_final = pmn_s or sku_s
            if not pm_name_final:
                errors.append(f"'{pm_sheet_name}' row {r}: no SKU or PM Name — cannot identify the row")
                continue

            try:
                # Concatenate Category + Size into PM Type if both present (richer detail)
                pm_type_final = cat_s
                db.upsert_pm({
                    'brand_id':            brand_id,
                    'pm_code':             sku_s,
                    'pm_name':             pm_name_final,
                    'pm_type':             pm_type_final,
                    'sku_size':            size_s,
                    'rate':                _f(rate),
                    'opening_stock':       _f(os_v),
                    'provisional_wastage': _f(waste),
                    'low_stock_threshold': 0.0,
                })
                counts['pm'] += 1
            except Exception as e:
                errors.append(f"'{pm_sheet_name}' row {r} ('{pm_name_final}'): {e}")

    # ── C. Inward (wide-format → unpivot) ──
    if 'Inward' in wb.sheetnames:
        ws_in = wb['Inward']
        # Headers on row 2
        in_headers = [ws_in.cell(2, c).value for c in range(1, ws_in.max_column + 1)]
        # Find the SKU + PM Name columns
        h2c_in = {}
        for c, h in enumerate(in_headers, start=1):
            nh = _norm(h)
            if nh:
                h2c_in[nh] = c
        col_sku_i  = h2c_in.get('sku')
        col_pmn_i  = h2c_in.get('pm name')

        # Date columns: any col in row 2 whose value is a datetime
        from datetime import datetime as _dt
        date_cols = []
        for c, h in enumerate(in_headers, start=1):
            if hasattr(h, 'strftime'):
                date_cols.append((c, h.strftime('%Y-%m-%d')))

        sheet_diag['Inward'] = {'headers': [str(h) for h in in_headers],
                                'header_row': 2,
                                'date_cols': len(date_cols),
                                'missing': []}

        if not col_sku_i and not col_pmn_i:
            errors.append("'Inward' sheet: cannot find SKU or PM Name column on row 2.")
        elif not date_cols:
            errors.append("'Inward' sheet: no date columns detected in row 2 (expected datetime headers).")
        else:
            # Build a {sku→pm_id} lookup once for performance
            pm_rows = db.get_pm_rows()
            sku2id  = {(r.get('pm_code') or '').strip().lower(): r['id']
                       for r in pm_rows if r.get('pm_code')}
            name2id = {(r.get('pm_name') or '').strip().lower(): r['id']
                       for r in pm_rows if r.get('pm_name')}

            for r in range(3, ws_in.max_row + 1):
                sku = (str(ws_in.cell(r, col_sku_i).value or '').strip()
                       if col_sku_i else '')
                pmn = (str(ws_in.cell(r, col_pmn_i).value or '').strip()
                       if col_pmn_i else '')

                # Skip section divider / empty rows (no data in any date col)
                date_qty_pairs = [(d, ws_in.cell(r, c).value)
                                  for c, d in date_cols]
                has_any_date_qty = any(v not in (None, '', 0) for _, v in date_qty_pairs)
                if not has_any_date_qty:
                    continue

                pm_id = (sku2id.get(sku.lower()) if sku else None) \
                     or (name2id.get(pmn.lower()) if pmn else None)
                if not pm_id:
                    errors.append(f"'Inward' row {r}: PM '{sku or pmn}' not found in master")
                    continue

                # Unpivot: each non-zero date cell becomes one inward transaction
                row_in = 0
                for date_str, qty in date_qty_pairs:
                    qf = _f(qty)
                    if qf <= 0:
                        continue
                    scanned['inward'] += 1
                    try:
                        db.add_inward({
                            'entry_date': date_str,
                            'pm_id':      pm_id,
                            'quantity':   qf,
                            'ref_no':     '',
                            'remarks':    'Imported from working-format Excel',
                        })
                        counts['inward'] += 1
                        row_in += 1
                    except Exception as e:
                        errors.append(f"'Inward' row {r} date {date_str}: {e}")

    # ── D. Audit summary ──
    summary_parts = []
    for k, v in counts.items():
        if v: summary_parts.append(f"{v} {k}")
    _audit('import_working', 'workbook', None,
           summary='Imported (working format): ' + (', '.join(summary_parts) or 'nothing'),
           details={'counts': counts, 'errors': errors[:50],
                    'brand': brand_label})

    return jsonify({
        'ok': True,
        'pm_inserted': counts['pm'],
        'fg_inserted': counts['fg'],
        'counts':      counts,
        'scanned':     scanned,
        'errors':      errors,
        'sheet_diag':  sheet_diag,
    })


@hcp_stock_bp.route('/export/pm')
@login_required
@hcp_stock_role_required
@feature_required('data_export')
def export_pm_excel():
    """Properly formatted single-sheet PM stock export.

    Layout matches the team's working format:
      • Row 1: workbook title with brand name + closing-stock-as-of date
      • Row 2: column totals (sum of numeric columns)
      • Row 3: bold header row with frozen pane below
      • Rows 4+: one row per active PM, grouped/sorted by PM Type then PM Name

    Visual polish:
      • Brand-coloured header row (or dark blue default)
      • Comma-formatted numbers, two-decimal rate
      • Closing-stock cells highlighted red when ≤ low_stock_threshold
      • Auto-sized columns
      • Filter applied to header row so users can sort/filter inside Excel
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    brand_id = request.args.get('brand_id', type=int)
    pm_rows  = db.get_pm_rows(brand_id=brand_id)

    # Resolve brand label for the title row
    brand_label = 'All Brands'
    if brand_id:
        for b in db.list_brands():
            if b['id'] == brand_id:
                brand_label = b.get('brand_name') or b.get('name') or 'Brand'
                break

    wb = Workbook()
    ws = wb.active
    ws.title = 'PM Stock'

    # ── Styles ──
    title_font  = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    title_fill  = PatternFill('solid', fgColor='1E3A8A')   # dark blue
    totals_font = Font(name='Calibri', size=11, bold=True, color='1E3A8A')
    totals_fill = PatternFill('solid', fgColor='E0E7FF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3A8A')
    body_font   = Font(name='Calibri', size=11)
    low_fill    = PatternFill('solid', fgColor='FEE2E2')   # light red
    low_font    = Font(name='Calibri', size=11, bold=True, color='B91C1C')
    thin = Side(style='thin', color='CBD5E1')
    body_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left        = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right       = Alignment(horizontal='right',  vertical='center')

    headers = ['Brand', 'PM Code', 'PM Name', 'PM Type', 'SKU Size',
               'Rate', 'Opening', 'Inward', 'BOM Consumed', 'Wastage',
               'Closing', 'Threshold', 'Status']
    n_cols = len(headers)

    # ── Row 1: title ──
    today_str = datetime.now().strftime('%d-%m-%Y')
    ws.cell(1, 1).value = f"{brand_label} — PM Closing Stock (as of {today_str})"
    ws.cell(1, 1).font = title_font
    ws.cell(1, 1).fill = title_fill
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 28

    # ── Row 2: totals ──
    totals = {
        'opening':  sum(_f(r.get('opening_stock'))   for r in pm_rows),
        'inward':   sum(_f(r.get('inward_qty'))      for r in pm_rows),
        'consumed': sum(_f(r.get('bom_consumed_qty'))for r in pm_rows),
        'wastage':  sum(_f(r.get('actual_wastage'))  for r in pm_rows),
        'closing':  sum(_f(r.get('closing_stock'))   for r in pm_rows),
    }
    ws.cell(2, 1).value = f"TOTAL ({len(pm_rows)} PMs)"
    ws.cell(2, 1).font = totals_font
    ws.cell(2, 1).alignment = left
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    for c, key in zip([7, 8, 9, 10, 11], ['opening','inward','consumed','wastage','closing']):
        cell = ws.cell(2, c)
        cell.value = totals[key]
        cell.font = totals_font
        cell.alignment = right
        cell.number_format = '#,##0.00'
    for c in range(1, n_cols + 1):
        ws.cell(2, c).fill = totals_fill
    ws.row_dimensions[2].height = 22

    # ── Row 3: headers ──
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(3, c)
        cell.value = h
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center
        cell.border = body_border
    ws.row_dimensions[3].height = 30

    # ── Body rows (sorted by PM Type, then PM Name) ──
    sorted_rows = sorted(
        pm_rows,
        key=lambda r: ((r.get('pm_type') or '').strip().lower(),
                       (r.get('pm_name') or '').strip().lower())
    )
    r_idx = 4
    for r in sorted_rows:
        closing   = _f(r.get('closing_stock'))
        threshold = _f(r.get('low_stock_threshold'))
        is_low    = (threshold > 0 and closing <= threshold)

        values = [
            r.get('brand_name') or '',
            r.get('pm_code') or '',
            r.get('pm_name') or '',
            r.get('pm_type') or '',
            r.get('sku_size') or '',
            _f(r.get('rate')),
            _f(r.get('opening_stock')),
            _f(r.get('inward_qty')),
            _f(r.get('bom_consumed_qty')),
            _f(r.get('actual_wastage')),
            closing,
            threshold,
            'LOW' if is_low else 'OK',
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(r_idx, c)
            cell.value = v
            cell.font  = low_font if (is_low and c == 11) else body_font
            cell.border = body_border
            if c in (1, 2, 4, 5, 13):
                cell.alignment = center
            elif c == 3:
                cell.alignment = left
            else:
                cell.alignment = right
                cell.number_format = '#,##0.00' if c != 6 else '#,##0.00'
            if is_low and c == 11:
                cell.fill = low_fill
            if c == 13 and is_low:
                cell.fill = low_fill
                cell.font = low_font
        r_idx += 1

    # ── Column widths ──
    widths = {1: 14, 2: 16, 3: 38, 4: 16, 5: 12, 6: 9, 7: 12, 8: 12, 9: 14,
              10: 12, 11: 13, 12: 12, 13: 10}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Freeze panes + filter ──
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}{r_idx - 1}"

    # ── Audit ──
    _audit('export_pm', 'workbook', None,
           summary=f'Exported {len(pm_rows)} PM rows for "{brand_label}"',
           details={'brand_id': brand_id, 'rows': len(pm_rows)})

    # ── Stream ──
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    import re as _re
    safe_brand = _re.sub(r'[^A-Za-z0-9]+', '_', brand_label).strip('_')
    fname = f"PM_Stock_{safe_brand}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(out,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=fname)


@hcp_stock_bp.route('/export')
@login_required
@hcp_stock_role_required
@feature_required('data_export')
def export_excel():
    """Multi-sheet workbook: PM, FG, Inward, Dispatch (with consumption), Wastage, BOM."""
    brand_id = request.args.get('brand_id', type=int)
    wb = Workbook()

    # ── PM Stock summary ──
    ws = wb.active; ws.title = 'PM Stock'
    headers = ['Brand', 'PM Code', 'PM Name', 'PM Type', 'SKU Size', 'Rate',
               'Opening', 'Inward', 'BOM Consumed', 'Actual Wastage',
               'Closing', 'Wastage %', 'Low Stock Threshold', 'Low?']
    ws.append(headers)
    for r in db.get_pm_rows(brand_id=brand_id):
        wpct = (_f(r.get('actual_wastage')) / _f(r.get('inward_qty')) * 100
                if _f(r.get('inward_qty')) else 0)
        ws.append([
            r.get('brand_name') or '', r.get('pm_code') or '', r.get('pm_name') or '',
            r.get('pm_type') or '', r.get('sku_size') or '', r.get('rate') or 0,
            r.get('opening_stock') or 0, r.get('inward_qty') or 0,
            r.get('bom_consumed_qty') or 0, r.get('actual_wastage') or 0,
            r.get('closing_stock') or 0, round(wpct, 2),
            r.get('low_stock_threshold') or 0,
            'YES' if r.get('is_low_stock') else '',
        ])
    _style_header(ws, get_column_letter(len(headers)))
    for i, w in enumerate([16,22,32,18,12,10,12,12,14,14,12,10,18,8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── FG ──
    ws = wb.create_sheet('FG')
    headers = ['Brand', 'Product Code', 'Product Name', 'Category',
               'SKU Size', 'Rate', 'Total Dispatched', 'BOM Lines']
    ws.append(headers)
    for r in db.get_fg_rows(brand_id=brand_id):
        ws.append([
            r.get('brand_name') or '', r.get('product_code') or '',
            r.get('product_name') or '', r.get('category') or '',
            r.get('sku_size') or '', r.get('rate') or 0,
            r.get('dispatch_qty') or 0, r.get('bom_lines') or 0,
        ])
    _style_header(ws, get_column_letter(len(headers)), color='0d9488')
    for i, w in enumerate([16,22,38,18,12,10,16,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Inward (PM) ──
    ws = wb.create_sheet('Inward')
    headers = ['Date', 'Brand', 'PM Code', 'PM Name', 'PM Type',
               'SKU Size', 'Quantity', 'Ref No.', 'Remarks']
    ws.append(headers)
    for t in db.list_inward(brand_id=brand_id):
        ws.append([_dmy(t.get('entry_date')), t.get('brand_name') or '',
                   t.get('pm_code') or '', t.get('pm_name') or '',
                   t.get('pm_type') or '', t.get('sku_size') or '',
                   t.get('quantity') or 0, t.get('ref_no') or '',
                   t.get('remarks') or ''])
    _style_header(ws, get_column_letter(len(headers)), color='10b981')
    for i, w in enumerate([12,16,22,32,18,12,12,18,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Dispatch (FG) ──
    ws = wb.create_sheet('Dispatch')
    headers = ['Date', 'Brand', 'Product Code', 'Product Name', 'Category',
               'SKU Size', 'Quantity', 'Ref No.', 'Remarks']
    ws.append(headers)
    for t in db.list_dispatch(brand_id=brand_id):
        ws.append([_dmy(t.get('entry_date')), t.get('brand_name') or '',
                   t.get('product_code') or '', t.get('product_name') or '',
                   t.get('category') or '', t.get('sku_size') or '',
                   t.get('quantity') or 0, t.get('ref_no') or '',
                   t.get('remarks') or ''])
    _style_header(ws, get_column_letter(len(headers)), color='f59e0b')
    for i, w in enumerate([12,16,22,38,18,12,12,18,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Wastage (PM) ──
    ws = wb.create_sheet('Wastage')
    headers = ['Date', 'Brand', 'PM Code', 'PM Name', 'PM Type',
               'SKU Size', 'Quantity', 'Reason', 'Remarks']
    ws.append(headers)
    for t in db.list_wastage(brand_id=brand_id):
        ws.append([_dmy(t.get('entry_date')), t.get('brand_name') or '',
                   t.get('pm_code') or '', t.get('pm_name') or '',
                   t.get('pm_type') or '', t.get('sku_size') or '',
                   t.get('quantity') or 0, t.get('reason') or '',
                   t.get('remarks') or ''])
    _style_header(ws, get_column_letter(len(headers)), color='ef4444')
    for i, w in enumerate([12,16,22,32,18,12,12,18,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── BOM master ──
    ws = wb.create_sheet('BOM')
    headers = ['FG Brand', 'FG Code', 'FG Name', 'PM Code', 'PM Name',
               'PM Type', 'Qty per Unit']
    ws.append(headers)
    for fg in db.get_fg_rows(brand_id=brand_id):
        for ln in db.get_bom_for_fg(fg['id']):
            ws.append([fg.get('brand_name') or '', fg.get('product_code') or '',
                       fg.get('product_name') or '',
                       ln.get('pm_code') or '', ln.get('pm_name') or '',
                       ln.get('pm_type') or '', ln.get('qty_per_unit') or 0])
    _style_header(ws, get_column_letter(len(headers)), color='a855f7')
    for i, w in enumerate([16,22,32,22,32,18,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True,
                     download_name=f'HCP_Stock_Export_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═══════════════════════════════════════════════════════════════════════════════
# SHARE — export a SUBSET of rows (selected via checkboxes) as Excel
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/share/export', methods=['POST'])
@login_required
@hcp_stock_role_required
def share_export():
    """
    POST { "kind": "pm" | "fg" | "inward" | "dispatch" | "wastage",
           "ids":  [1, 2, 3, ...] }
    Returns an Excel attachment containing ONLY the selected rows for the
    chosen grid. Used by the WhatsApp / Email share buttons.
    """
    d = request.json or {}
    kind = (d.get('kind') or '').strip().lower()
    ids  = [int(x) for x in (d.get('ids') or []) if str(x).isdigit()]
    if not kind or not ids:
        return jsonify({'ok': False, 'error': 'kind and ids are required'}), 400

    wb = Workbook(); ws = wb.active

    if kind == 'pm':
        ws.title = 'PM Stock'
        headers = ['Brand', 'PM Code', 'PM Name', 'PM Type', 'SKU Size', 'Rate',
                   'Opening', 'Inward', 'BOM Consumed', 'Actual Wastage',
                   'Closing', 'Wastage %']
        ws.append(headers)
        rows = [r for r in db.get_pm_rows() if r['id'] in ids]
        for r in rows:
            wpct = (_f(r.get('actual_wastage')) / _f(r.get('inward_qty')) * 100
                    if _f(r.get('inward_qty')) else 0)
            ws.append([r.get('brand_name') or '', r.get('pm_code') or '',
                       r.get('pm_name') or '', r.get('pm_type') or '',
                       r.get('sku_size') or '', r.get('rate') or 0,
                       r.get('opening_stock') or 0, r.get('inward_qty') or 0,
                       r.get('bom_consumed_qty') or 0, r.get('actual_wastage') or 0,
                       r.get('closing_stock') or 0, round(wpct, 2)])
        _style_header(ws, get_column_letter(len(headers)))
        for i, w in enumerate([16,22,32,18,12,10,12,12,14,14,12,10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    elif kind == 'fg':
        ws.title = 'FG'
        headers = ['Brand', 'Product Code', 'Product Name', 'Category',
                   'SKU Size', 'Rate', 'Total Dispatched', 'BOM Lines']
        ws.append(headers)
        rows = [r for r in db.get_fg_rows() if r['id'] in ids]
        for r in rows:
            ws.append([r.get('brand_name') or '', r.get('product_code') or '',
                       r.get('product_name') or '', r.get('category') or '',
                       r.get('sku_size') or '', r.get('rate') or 0,
                       r.get('dispatch_qty') or 0, r.get('bom_lines') or 0])
        _style_header(ws, get_column_letter(len(headers)), color='0d9488')
        for i, w in enumerate([16,22,38,18,12,10,16,12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    elif kind == 'inward':
        ws.title = 'Inward'
        headers = ['Date', 'Brand', 'PM Code', 'PM Name', 'PM Type',
                   'SKU Size', 'Quantity', 'Ref No.', 'Remarks']
        ws.append(headers)
        rows = [r for r in db.list_inward() if r['id'] in ids]
        for r in rows:
            ws.append([r.get('entry_date') or '', r.get('brand_name') or '',
                       r.get('pm_code') or '', r.get('pm_name') or '',
                       r.get('pm_type') or '', r.get('sku_size') or '',
                       r.get('quantity') or 0, r.get('ref_no') or '',
                       r.get('remarks') or ''])
        _style_header(ws, get_column_letter(len(headers)), color='10b981')
        for i, w in enumerate([12,16,22,32,18,12,12,18,22], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    elif kind == 'dispatch':
        ws.title = 'Dispatch'
        headers = ['Date', 'Brand', 'Product Code', 'Product Name', 'Category',
                   'SKU Size', 'Quantity', 'Ref No.', 'Remarks']
        ws.append(headers)
        rows = [r for r in db.list_dispatch() if r['id'] in ids]
        for r in rows:
            ws.append([r.get('entry_date') or '', r.get('brand_name') or '',
                       r.get('product_code') or '', r.get('product_name') or '',
                       r.get('category') or '', r.get('sku_size') or '',
                       r.get('quantity') or 0, r.get('ref_no') or '',
                       r.get('remarks') or ''])
        _style_header(ws, get_column_letter(len(headers)), color='f59e0b')
        for i, w in enumerate([12,16,22,38,18,12,12,18,22], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    elif kind == 'wastage':
        ws.title = 'Wastage'
        headers = ['Date', 'Brand', 'PM Code', 'PM Name', 'PM Type',
                   'SKU Size', 'Quantity', 'Reason', 'Remarks']
        ws.append(headers)
        rows = [r for r in db.list_wastage() if r['id'] in ids]
        for r in rows:
            ws.append([r.get('entry_date') or '', r.get('brand_name') or '',
                       r.get('pm_code') or '', r.get('pm_name') or '',
                       r.get('pm_type') or '', r.get('sku_size') or '',
                       r.get('quantity') or 0, r.get('reason') or '',
                       r.get('remarks') or ''])
        _style_header(ws, get_column_letter(len(headers)), color='ef4444')
        for i, w in enumerate([12,16,22,32,18,12,12,18,22], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        return jsonify({'ok': False, 'error': f'unknown kind: {kind}'}), 400

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    fname = f'HCP_Stock_{kind}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═══════════════════════════════════════════════════════════════════════════════
# CHECK AVAILABILITY  (multi-FG planning workbench)
# ═══════════════════════════════════════════════════════════════════════════════
def _build_availability_report(plan):
    """
    plan: list of {'fg_id': int, 'quantity': float}

    Returns a dict with:
      - lines:            per-FG breakdown (FG details + per-line BOM consumption)
      - pm_requirements:  per-PM aggregated requirement vs current closing
      - shortages:        list of PMs that go negative
      - feasible:         bool — True iff every PM has enough
      - totals:           {fg_total_units, pm_total_qty, shortage_count}
      - errors:           list of bad input rows (unknown FG, FG with no BOM, …)
    """
    # Index PM rows once for O(1) lookup
    pm_rows  = db.get_pm_rows()
    pm_by_id = {p['id']: p for p in pm_rows}

    fg_rows  = db.get_fg_rows()
    fg_by_id = {f['id']: f for f in fg_rows}

    lines    = []          # per-FG result objects
    pm_need  = {}          # pm_id → total qty required across all lines
    pm_meta  = {}          # pm_id → snapshot (name, code, current closing, …)
    errors   = []

    for idx, item in enumerate(plan or [], start=1):
        try:
            fg_id = int(item.get('fg_id') or 0)
            qty   = _f(item.get('quantity'))
        except (TypeError, ValueError):
            errors.append(f'Line {idx}: invalid fg_id or quantity')
            continue
        if fg_id <= 0 or qty <= 0:
            errors.append(f'Line {idx}: fg_id and positive quantity required')
            continue
        fg = fg_by_id.get(fg_id)
        if not fg:
            errors.append(f'Line {idx}: FG id={fg_id} not found')
            continue

        bom = db.get_bom_for_fg(fg_id)
        line_consumption = []
        for ln in bom:
            pm_id = ln['pm_id']
            need_for_line = _f(ln['qty_per_unit']) * qty
            line_consumption.append({
                'pm_id':         pm_id,
                'pm_code':       ln.get('pm_code') or '',
                'pm_name':       ln.get('pm_name') or '',
                'pm_type':       ln.get('pm_type') or '',
                'sku_size':      ln.get('sku_size') or '',
                'qty_per_unit':  _f(ln['qty_per_unit']),
                'qty_consumed':  round(need_for_line, 4),
            })
            # Aggregate across all FGs in the plan
            pm_need[pm_id] = pm_need.get(pm_id, 0.0) + need_for_line
            if pm_id not in pm_meta:
                pm = pm_by_id.get(pm_id) or {}
                pm_meta[pm_id] = {
                    'pm_id':         pm_id,
                    'pm_code':       pm.get('pm_code') or '',
                    'pm_name':       pm.get('pm_name') or ln.get('pm_name') or '',
                    'pm_type':       pm.get('pm_type') or '',
                    'sku_size':      pm.get('sku_size') or '',
                    'brand_name':    pm.get('brand_name') or '',
                    'closing_stock': _f(pm.get('closing_stock')),
                    'low_stock_threshold': _f(pm.get('low_stock_threshold')),
                }

        lines.append({
            'fg_id':         fg_id,
            'product_code':  fg.get('product_code') or '',
            'product_name':  fg.get('product_name') or '',
            'category':      fg.get('category') or '',
            'sku_size':      fg.get('sku_size') or '',
            'brand_name':    fg.get('brand_name') or '',
            'quantity':      qty,
            'bom_lines':     int(fg.get('bom_lines') or 0),
            'consumption':   line_consumption,
            'has_bom':       bool(line_consumption),
        })

    # Build the PM requirements list
    pm_requirements = []
    shortages       = []
    for pm_id, need in pm_need.items():
        meta = pm_meta[pm_id]
        avail = meta['closing_stock']
        short = avail - need
        is_short = short < -1e-6
        row = {
            'pm_id':                 pm_id,
            'pm_code':               meta['pm_code'],
            'pm_name':                meta['pm_name'],
            'pm_type':               meta['pm_type'],
            'sku_size':              meta['sku_size'],
            'brand_name':            meta['brand_name'],
            'available':             round(avail, 2),
            'required':              round(need, 2),
            'after_dispatch':        round(short, 2),
            'is_short':              is_short,
            'low_stock_threshold':   meta['low_stock_threshold'],
            'will_be_low':           (
                meta['low_stock_threshold'] > 0
                and not is_short
                and (avail - need) < meta['low_stock_threshold']
            ),
        }
        pm_requirements.append(row)
        if is_short:
            shortages.append({
                'pm_id':     pm_id,
                'pm_name':   meta['pm_name'],
                'pm_code':   meta['pm_code'],
                'available': round(avail, 2),
                'required':  round(need, 2),
                'short_by':  round(need - avail, 2),
            })
    # Sort: shortages first, then by name
    pm_requirements.sort(key=lambda r: (not r['is_short'], r['pm_name'].lower()))

    return {
        'feasible':        not shortages,
        'lines':           lines,
        'pm_requirements': pm_requirements,
        'shortages':       shortages,
        'totals': {
            'fg_lines':       len(lines),
            'fg_total_units': round(sum(_f(l['quantity']) for l in lines), 2),
            'pm_total_qty':   round(sum(r['required'] for r in pm_requirements), 2),
            'shortage_count': len(shortages),
        },
        'errors':          errors,
    }


@hcp_stock_bp.route('/api/share/fg_bom', methods=['POST'])
@login_required
@hcp_stock_role_required
def api_share_fg_bom():
    """Enriched share data for FG / dispatch WhatsApp messages.

    Body: { "fg_ids": [int, ...] }
    Returns per-FG: product name/code/brand + BOM lines, each line carrying the
    PM name, qty/unit and the PM's current available (closing) stock.
    """
    d = request.json or {}
    fg_ids = d.get('fg_ids') or []
    if not isinstance(fg_ids, list) or not fg_ids:
        return jsonify({'ok': False, 'error': 'fg_ids must be a non-empty list'}), 400

    # PM closing-stock lookup, computed once
    pm_by_id = {p['id']: p for p in db.get_pm_rows()}
    fg_by_id = {f['id']: f for f in db.get_fg_rows()}

    out = []
    for fid in fg_ids:
        try:
            fid = int(fid)
        except (TypeError, ValueError):
            continue
        fg = fg_by_id.get(fid)
        if not fg:
            continue
        bom = db.get_bom_for_fg(fid)
        lines = []
        for ln in bom:
            pm = pm_by_id.get(ln['pm_id']) or {}
            lines.append({
                'pm_id':        ln['pm_id'],
                'pm_code':      ln.get('pm_code') or '',
                'pm_name':      ln.get('pm_name') or '',
                'sku_size':     ln.get('sku_size') or '',
                'qty_per_unit': _f(ln.get('qty_per_unit')),
                'available':    round(_f(pm.get('closing_stock')), 2),
            })
        out.append({
            'fg_id':        fid,
            'product_code': fg.get('product_code') or '',
            'product_name': fg.get('product_name') or '',
            'brand_name':   fg.get('brand_name') or '',
            'sku_size':     fg.get('sku_size') or '',
            'rate':         _f(fg.get('rate')),
            'dispatch_qty': _f(fg.get('dispatch_qty')),
            'bom':          lines,
        })
    return jsonify({'ok': True, 'fgs': out})


@hcp_stock_bp.route('/check_availability', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('data_check_pm')
def api_check_availability():
    """Read-only PM availability check for a multi-FG dispatch plan.
       Does NOT write anything to the DB."""
    d = request.json or {}
    plan = d.get('plan') or []
    if not isinstance(plan, list) or not plan:
        return jsonify({'ok': False, 'error': 'plan must be a non-empty list'}), 400
    return jsonify({'ok': True, 'report': _build_availability_report(plan)})


@hcp_stock_bp.route('/check_availability/export', methods=['POST'])
@login_required
@hcp_stock_role_required
@feature_required('data_check_pm')
def api_check_availability_export():
    """Excel export of an availability check.
       Body: same as /check_availability — list of {fg_id, quantity}."""
    d = request.json or {}
    plan = d.get('plan') or []
    if not plan:
        return jsonify({'ok': False, 'error': 'plan must be a non-empty list'}), 400

    report = _build_availability_report(plan)
    wb = Workbook()

    # ─── Sheet 1: Summary ───
    ws = wb.active; ws.title = 'Summary'
    ws['A1'] = 'HCP Stock — Availability Check'
    ws['A1'].font = Font(bold=True, size=15, color='1E3A8A')
    ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A3'].font = Font(italic=True, color='64748B', size=11)

    feasible_label = 'YES — sufficient PM stock' if report['feasible'] else 'NO — PM shortage(s)'
    feasible_color = '047857' if report['feasible'] else 'B91C1C'
    ws['A5'] = 'Feasible:'
    ws['B5'] = feasible_label
    ws['A5'].font = Font(bold=True, size=12)
    ws['B5'].font = Font(bold=True, size=12, color=feasible_color)

    ws['A7']  = 'FG lines:';                ws['B7']  = report['totals']['fg_lines']
    ws['A8']  = 'Total FG units to ship:';  ws['B8']  = report['totals']['fg_total_units']
    ws['A9']  = 'Total PM units needed:';   ws['B9']  = report['totals']['pm_total_qty']
    ws['A10'] = 'PMs short:';               ws['B10'] = report['totals']['shortage_count']
    for r in range(7, 11):
        ws.cell(row=r, column=1).font = Font(bold=True, color='64748B', size=11)
        ws.cell(row=r, column=2).font = Font(bold=True, size=12, color='1E3A8A')
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 36

    if report['errors']:
        ws['A12'] = 'Input errors:'
        ws['A12'].font = Font(bold=True, color='B91C1C')
        for i, e in enumerate(report['errors'], start=13):
            ws.cell(row=i, column=1, value=f'  • {e}').font = Font(color='B91C1C', size=10)

    # ─── Sheet 2: PM Requirements ───
    ws2 = wb.create_sheet('PM Requirements')
    headers = ['Brand', 'PM Code', 'PM Name', 'PM Type', 'SKU Size',
               'Available (Closing)', 'Required (Cumulative)',
               'After Dispatch', 'Status']
    ws2.append(headers)
    for r in report['pm_requirements']:
        status = 'SHORT' if r['is_short'] else ('LOW after' if r['will_be_low'] else 'OK')
        ws2.append([
            r['brand_name'], r['pm_code'], r['pm_name'], r['pm_type'], r['sku_size'],
            r['available'], r['required'], r['after_dispatch'], status,
        ])
    _style_header(ws2, get_column_letter(len(headers)), color='0d9488')
    # Highlight shortage rows in red
    short_fill = PatternFill('solid', start_color='FEE2E2')
    short_font = Font(color='B91C1C', bold=True)
    low_fill   = PatternFill('solid', start_color='FEF3C7')
    for row_idx, r in enumerate(report['pm_requirements'], start=2):
        if r['is_short']:
            for col in range(1, len(headers)+1):
                ws2.cell(row=row_idx, column=col).fill = short_fill
            ws2.cell(row=row_idx, column=len(headers)).font = short_font
        elif r['will_be_low']:
            for col in range(1, len(headers)+1):
                ws2.cell(row=row_idx, column=col).fill = low_fill
    for i, w in enumerate([16,22,32,18,12,18,20,16,12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ─── Sheet 3: Per-FG Breakdown ───
    ws3 = wb.create_sheet('Per-FG Breakdown')
    headers = ['Brand', 'Product Code', 'Product Name', 'Category', 'SKU Size',
               'FG Qty', 'BOM PM Code', 'BOM PM Name', 'Qty/Unit', 'PM Consumption']
    ws3.append(headers)
    for line in report['lines']:
        if not line['consumption']:
            ws3.append([
                line['brand_name'], line['product_code'], line['product_name'],
                line['category'], line['sku_size'], line['quantity'],
                '— no BOM —', '', '', 0,
            ])
        else:
            for con in line['consumption']:
                ws3.append([
                    line['brand_name'], line['product_code'], line['product_name'],
                    line['category'], line['sku_size'], line['quantity'],
                    con['pm_code'], con['pm_name'], con['qty_per_unit'], con['qty_consumed'],
                ])
    _style_header(ws3, get_column_letter(len(headers)), color='1E3A8A')
    for i, w in enumerate([16,20,32,16,12,10,18,28,12,16], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    fname = f'HCP_Stock_Availability_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═══════════════════════════════════════════════════════════════════════════════
# RECYCLE BIN  (admin only)
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/admin/recycle_bin', methods=['GET'])
@login_required
@admin_only
def api_recycle_bin_list():
    """Return all soft-deleted rows grouped by entity."""
    return jsonify({'ok': True, 'data': db.list_recycle_bin()})


@hcp_stock_bp.route('/admin/recycle_bin/restore', methods=['POST'])
@login_required
@admin_only
def api_recycle_bin_restore():
    d = request.json or {}
    entity = (d.get('entity') or '').strip().lower()
    row_id = int(d.get('id') or 0)
    if entity not in db.SOFT_DELETE_TABLES:
        return jsonify({'ok': False, 'error': f'unknown entity: {entity}'}), 400
    if row_id <= 0:
        return jsonify({'ok': False, 'error': 'id required'}), 400
    db.restore_row(entity, row_id)
    _audit('restore', entity, row_id,
           summary=f"Restored {entity} id={row_id} from Recycle Bin",
           details=None)
    return jsonify({'ok': True})


@hcp_stock_bp.route('/admin/recycle_bin/purge', methods=['POST'])
@login_required
@admin_only
def api_recycle_bin_purge():
    """Permanently delete a row (hard delete). Cascades via FK as before."""
    d = request.json or {}
    entity = (d.get('entity') or '').strip().lower()
    row_id = int(d.get('id') or 0)
    if entity not in db.SOFT_DELETE_TABLES:
        return jsonify({'ok': False, 'error': f'unknown entity: {entity}'}), 400
    if row_id <= 0:
        return jsonify({'ok': False, 'error': 'id required'}), 400
    db.hard_delete(entity, row_id)
    _audit('hard_delete', entity, row_id,
           summary=f"Permanently deleted {entity} id={row_id}",
           details=None)
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════════════════════════════════
# AUDIT LOG  (admin only)
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/admin/audit', methods=['GET'])
@login_required
@admin_only
def api_audit_list():
    """Return audit entries — latest first. Optional filters: entity, action, user, q (search), limit."""
    rows = db.list_audit(
        limit    = request.args.get('limit', 500, type=int),
        entity   = request.args.get('entity') or None,
        action   = request.args.get('action') or None,
        user_name= request.args.get('user')   or None,
        q        = request.args.get('q')      or None,
    )
    return jsonify({'ok': True, 'rows': rows})


# ═══════════════════════════════════════════════════════════════════════════════
# CLEAR / REFRESH TABLES  (admin only — hard delete, no undo)
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/admin/clear/list', methods=['GET'])
@login_required
@admin_only
def api_clear_list():
    """List all clearable HCP Stock tables with current row counts."""
    out = []
    for key, spec in db.CLEARABLE_TABLES.items():
        rc = db.count_table_rows(spec['table'])
        deps = db.table_dependents(key)
        dep_total = sum(d['row_count'] for d in deps)
        out.append({
            'key':        key,
            'label':      spec['label'],
            'table':      spec['table'],
            'icon':       spec.get('icon',''),
            'row_count':  rc,
            'dep_count':  len(deps),
            'dep_total':  dep_total,
            'is_leaf':    not deps,
        })
    return jsonify({'ok': True, 'tables': out})


@hcp_stock_bp.route('/admin/clear/preview', methods=['GET'])
@login_required
@admin_only
def api_clear_preview():
    """Preview impact of clearing a single table — its row count + every
    dependent table that would be wiped on cascade."""
    key = (request.args.get('table') or '').strip().lower()
    spec = db.CLEARABLE_TABLES.get(key)
    if not spec:
        return jsonify({'ok': False, 'error': f'unknown table: {key}'}), 400
    main_rc = db.count_table_rows(spec['table'])
    deps = db.table_dependents(key)
    return jsonify({
        'ok':           True,
        'key':          key,
        'label':        spec['label'],
        'table':        spec['table'],
        'icon':         spec.get('icon',''),
        'row_count':    main_rc,
        'dependents':   deps,
        'dep_total':    sum(d['row_count'] for d in deps),
        'has_deps':     bool([d for d in deps if d['row_count'] > 0]),
        'confirm_text': f'CLEAR {key}',
    })


@hcp_stock_bp.route('/admin/clear', methods=['POST'])
@login_required
@admin_only
def api_clear_table():
    """Hard-DELETE every row in a table. Optional cascade=true to also wipe
    dependent tables. Requires confirm_text == "CLEAR <key>" (case-insensitive
    on the keyword, key part is exact)."""
    d = request.json or {}
    key = (d.get('table') or '').strip().lower()
    cascade = bool(d.get('cascade'))
    confirm = (d.get('confirm_text') or '').strip()

    spec = db.CLEARABLE_TABLES.get(key)
    if not spec:
        return jsonify({'ok': False, 'error': f'unknown table: {key}'}), 400

    expected = f'CLEAR {key}'
    if confirm.upper() != expected.upper():
        return jsonify({
            'ok': False,
            'error': f'Confirmation text mismatch — please type exactly: {expected}',
        }), 400

    try:
        cleared = db.clear_table(key, cascade=cascade)
    except RuntimeError as e:
        # dependents present but cascade not requested
        return jsonify({
            'ok': False,
            'error': str(e),
            'needs_cascade': True,
            'dependents':    db.table_dependents(key),
        }), 409
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    # Audit — single summary entry. NOTE: this audit row may be wiped on its
    # own next clear, but for the immediate operation it provides a trail.
    total_cleared = sum(cleared.values())
    summary_parts = []
    for k, n in cleared.items():
        summary_parts.append(f"{k}={n}")
    _audit('clear_table', key, None,
           summary=f"CLEARED {spec['label']}{' + cascade' if cascade else ''} "
                   f"— {total_cleared} rows total",
           details={'cleared': cleared, 'cascade': cascade})

    return jsonify({
        'ok':            True,
        'cleared':       cleared,
        'total_cleared': total_cleared,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# PER-USER PERMISSIONS  (admin only)
# ═══════════════════════════════════════════════════════════════════════════════
@hcp_stock_bp.route('/admin/permissions', methods=['GET'])
@login_required
@admin_only
def api_permissions_list():
    """Return: feature catalog + every permissions row + the resolved features
    for the calling admin (so the modal can also show their own state)."""
    return jsonify({
        'ok':            True,
        'catalog':       db.FEATURE_CATALOG,
        'permissions':   db.list_permissions(),
        'all_slugs':     sorted(db.ALL_FEATURE_SLUGS),
        'default_features': sorted(db.DEFAULT_NON_ADMIN_FEATURES),
    })


@hcp_stock_bp.route('/admin/permissions', methods=['POST'])
@login_required
@admin_only
def api_permissions_upsert():
    d = request.json or {}
    user_name = (d.get('user_name') or '').strip()
    features  = d.get('features') or []
    note      = (d.get('note') or '').strip()
    if not user_name:
        return jsonify({'ok': False, 'error': 'user_name is required'}), 400
    try:
        granted = db.upsert_permission(user_name, features, note=note)
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    _audit('upsert_permission', 'permission', None,
           summary=f"Granted {len(granted)} feature(s) to '{user_name}'",
           details={'user_name': user_name, 'features': granted, 'note': note})
    return jsonify({'ok': True, 'user_name': user_name, 'features': granted})


@hcp_stock_bp.route('/admin/permissions/<user_name>', methods=['DELETE'])
@login_required
@admin_only
def api_permissions_delete(user_name):
    user_name = (user_name or '').strip()
    if not user_name:
        return jsonify({'ok': False, 'error': 'user_name is required'}), 400
    db.delete_permission(user_name)
    _audit('delete_permission', 'permission', None,
           summary=f"Removed all custom permissions for '{user_name}' "
                   f"(reverts to defaults)",
           details={'user_name': user_name})
    return jsonify({'ok': True})


@hcp_stock_bp.route('/api/me/features', methods=['GET'])
@login_required
@hcp_stock_role_required
def api_me_features():
    """Anyone logged in can ask 'what can I do?' — used by the frontend on
    boot/refresh to refresh local feature flags after admin grants."""
    return jsonify({
        'ok':       True,
        'features': sorted(_user_features_for_session()),
        'is_admin': (session.get('User_Type') or '').strip().lower() == 'admin',
    })
