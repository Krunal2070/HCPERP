"""
modules/npd/npd_routes.py
────────────────────────────────────────────────────────────────────────────
NPD / EPD PROJECTS  (CRM lead -> NPD project flow)
OLD-SYSTEM-COMPATIBLE SCHEMA — tables/columns old HCP-ERP (models/npd.py)
jaise hi hain taaki old data seedha migrate ho jaye:
    npd_projects, milestone_masters (per-project), milestone_logs,
    npd_milestone_templates (admin master), npd_statuses, npd_activity_logs

  • CRM ka `lead_create_npd` gate yahan redirect karta hai:
        url_for('npd.create', lead_id=..., client_id=...)
  • project_type: 'npd' / 'existing'  (existing = EPD)
  • code: NPD-0001 (old gen_npd_code jaisa — last id + 1)
  • Visibility: NPD Manager / Management / Administrator dept + admin /
    manager role = sab; baaki = created_by / assigned_sc / assigned_rd /
    npd_poc / assigned_members me ho to.

Routes:
    GET  /npd                      -> projects list (?status, ?type, ?q)
    GET  /npd/new                  -> create form (lead_id/client_id prefill)
    POST /npd/new                  -> create + template milestones copy + logs
    GET  /npd/<id>                 -> project view (milestones + activity)
    POST /npd/<id>/status          -> project status update
    POST /npd/<id>/milestone/<mid> -> milestone status update (+ milestone_logs)
"""

from flask import (Blueprint, render_template, request, redirect, url_for,
                   session, flash)

from crm.crm_leads_routes import (_db, _is_admin, _has_full_visibility,
                                  _role, _uid, _uname, _user_map,
                                  login_required)

try:
    from menus import get_menu
except Exception:
    def get_menu(*a, **k):
        return None

from models.npd import (MS_STATUSES, PRIORITIES, ensure_npd_tables,  # noqa: F401
                        get_npd_statuses)

npd_bp = Blueprint('npd', __name__, url_prefix='/npd')


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
@npd_bp.before_request
def _npd_login():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return None


def _proj_visibility_sql(alias='p'):
    """Project-level visibility (old system ke assignment columns pe)."""
    if _has_full_visibility():
        return "", []
    u = _uid()
    return (f" AND ({alias}.created_by = %s OR {alias}.assigned_sc = %s "
            f"OR {alias}.assigned_rd = %s OR {alias}.npd_poc = %s "
            f"OR FIND_IN_SET(%s, COALESCE({alias}.assigned_members,'')) "
            f"OR FIND_IN_SET(%s, COALESCE({alias}.assigned_rd_members,'')))",
            [u, u, u, u, str(u), str(u)])


def _log(conn, pid, action):
    conn.execute(
        "INSERT INTO `npd_activity_logs` (project_id, user_id, action) "
        "VALUES (%s,%s,%s)", (pid, _uid(), action))


def _render(tpl, **kw):
    kw.setdefault('sidebar_menu',
                  get_menu('npd', role=_role(), is_admin=_is_admin()))
    kw.setdefault('user_name', _uname())
    kw.setdefault('role', session.get('User_Type'))
    kw.setdefault('ms_statuses', MS_STATUSES)
    return render_template(tpl, **kw)


# ─────────────────────────────────────────────────────────────────────────────
# LIST
# ─────────────────────────────────────────────────────────────────────────────
SORT_COLS = {
    'created':  'p.created_at',
    'code':     'p.code',
    'category': 'p.product_category',
    'client':   'COALESCE(p.client_name, p.client_company)',
    'product':  'p.product_name',
    'brand':    'p.reference_brand',
    'priority': "FIELD(p.priority,'Urgent','High','Normal','Low')",
    'last_connected': 'p.last_connected',
    'tat':      'p.created_at',
    'status':   'p.status',
    'start':    'p.project_start_date',
}


def _list_filters():
    """Listing ke saare query-params -> (where, params, state-dict)."""
    show_deleted = request.args.get('show') == 'deleted'
    st = request.args.get('status', '')
    pt = request.args.get('type', '')
    pr = request.args.get('priority', '')
    cat = (request.args.get('category') or '').strip()
    q = (request.args.get('q') or '').strip()
    sort = request.args.get('sort', 'created')
    if sort not in SORT_COLS:
        sort = 'created'
    sdir = 'asc' if request.args.get('dir') == 'asc' else 'desc'

    vis_frag, vis_params = _proj_visibility_sql('p')
    where = ("p.is_deleted=%s" + vis_frag)
    params = [1 if show_deleted else 0] + list(vis_params)
    if st:
        where += " AND p.status=%s"; params.append(st)
    if pt in ('npd', 'existing'):
        where += " AND p.project_type=%s"; params.append(pt)
    if pr in PRIORITIES:
        where += " AND p.priority=%s"; params.append(pr)
    if cat:
        where += " AND p.product_category=%s"; params.append(cat)
    if q:
        where += (" AND (p.code LIKE %s OR p.product_name LIKE %s "
                  "OR p.client_name LIKE %s OR p.client_company LIKE %s "
                  "OR p.reference_brand LIKE %s OR c.company_name LIKE %s)")
        params += [f"%{q}%"] * 6
    state = dict(show_deleted=show_deleted, cur_status=st, cur_type=pt,
                 cur_priority=pr, cur_category=cat, q=q,
                 sort=sort, sdir=sdir)
    return where, params, state


_LIST_SQL = (
    "SELECT p.*, "
    "c.company_name AS cm_company, "
    "DATEDIFF(NOW(), p.created_at) AS tat_days, "
    "(SELECT COUNT(*) FROM `milestone_masters` m "
    " WHERE m.project_id=p.id AND m.is_selected=1) AS ms_total, "
    "(SELECT COUNT(*) FROM `milestone_masters` m "
    " WHERE m.project_id=p.id AND m.is_selected=1 "
    " AND m.status='approved') AS ms_done "
    "FROM `npd_projects` p "
    "LEFT JOIN `client_masters` c ON c.id = p.client_id ")


def _members_count(r):
    ids = set()
    for k in ('assigned_sc', 'assigned_rd', 'npd_poc'):
        if r.get(k):
            ids.add(str(r[k]))
    for k in ('assigned_members', 'assigned_rd_members'):
        for x in (r.get(k) or '').split(','):
            if x.strip():
                ids.add(x.strip())
    return len(ids)


# Listing columns — npd_projects ke SAARE DB fields (Columns dropdown me
# show/hide). default=True wale shuru se dikhte hain (old listing jaisa).
COLUMN_DEFS = [
    # (key, label, default_visible)
    ('created',          'Create Date',        True),
    ('code',             'Project No',         True),
    ('project_type',     'Type',               False),
    ('category',         'Category',           True),
    ('client',           'Client Name',        True),
    ('client_company',   'Client Company',     False),
    ('client_email',     'Client Email',       False),
    ('client_phone',     'Client Phone',       False),
    ('client_coordinator','Client Coordinator',False),
    ('product',          'Product Name',       True),
    ('product_range',    'Product Range',      False),
    ('members',          'Members',            True),
    ('brand',            'Reference Brand',    True),
    ('reference_product_name', 'Reference Product', False),
    ('variant_type',     'Variant / Type',     False),
    ('area_of_application','Area of Application', False),
    ('market_level',     'Market Level',       False),
    ('appearance',       'Appearance',         False),
    ('fragrance',        'Fragrance',          False),
    ('viscosity',        'Viscosity',          False),
    ('ph_value',         'pH Value',           False),
    ('packaging_type',   'Packaging Type',     False),
    ('product_size',     'Product Size',       False),
    ('no_of_samples',    'No. of Samples',     False),
    ('moq',              'MOQ',                False),
    ('order_quantity',   'Order Quantity',     False),
    ('costing_range',    'Costing Range',      False),
    ('requirement_spec', 'Requirement Spec',   False),
    ('description',      'Description',        False),
    ('ingredients',      'Ingredients',        False),
    ('active_ingredients','Active Ingredients',False),
    ('product_claim',    'Product Claim',      False),
    ('label_claim',      'Label Claim',        False),
    ('video_link',       'Video Link',         False),
    ('priority',         'Priority',           True),
    ('npd_fee',          'NPD Fee',            False),
    ('custom_formulation','Custom Formulation',False),
    ('assigned_sc_nm',   'Sales Coordinator',  False),
    ('assigned_rd_nm',   'R&D Person',         False),
    ('npd_poc_nm',       'NPD POC',            False),
    ('last_connected',   'Last Connected',     True),
    ('milestones',       'Milestones',         True),
    ('tat',              'TAT',                True),
    ('status',           'Status',             True),
    ('start',            'Start Date',         True),
    ('project_lead_days','Lead Days',          False),
    ('project_end_date', 'End Date',           False),
    ('target_sample_date','Target Sample Date',False),
    ('delay_reason',     'Delay Reason',       False),
    ('created_by_nm',    'Created By',         False),
    ('updated',          'Updated At',         False),
]


@npd_bp.route('/')
@login_required
def projects():
    conn = _db()
    if not conn:
        flash('Database connection failed.', 'error')
        return redirect('/')
    try:
        statuses = get_npd_statuses(conn)
        where, params, state = _list_filters()
        order = f"ORDER BY {SORT_COLS[state['sort']]} {state['sdir'].upper()}, p.id DESC"
        rows = conn.execute(
            f"{_LIST_SQL} WHERE {where} {order} LIMIT 500",
            params).fetchall() or []
        for r in rows:
            r['members'] = _members_count(r)

        umap = _user_map(conn)
        vis_frag, vis_params = _proj_visibility_sql('p')
        deleted_count = (conn.execute(
            "SELECT COUNT(*) AS c FROM `npd_projects` p "
            "WHERE p.is_deleted=1" + vis_frag, vis_params)
            .fetchone() or {}).get('c', 0)
        cats = _npd_cats(conn)

        return _render('npd/projects.html', rows=rows, umap=umap,
                       statuses=statuses, cats=cats,
                       deleted_count=deleted_count, coldefs=COLUMN_DEFS,
                       sortable=list(SORT_COLS.keys()),
                       priorities=PRIORITIES, market_levels=MARKET_LEVELS,
                       active_item='npd-list', **state)
    finally:
        conn.close()


# ── EXPORT (current filters ke saath xlsx) ───────────────────────────────────
@npd_bp.route('/export')
@login_required
def export_projects():
    conn = _db()
    try:
        where, params, state = _list_filters()
        order = f"ORDER BY {SORT_COLS[state['sort']]} {state['sdir'].upper()}, p.id DESC"
        rows = conn.execute(f"{_LIST_SQL} WHERE {where} {order}",
                            params).fetchall() or []
        smap = {s['slug']: s['name'] for s in get_npd_statuses(conn)}
    finally:
        conn.close()

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "NPD Projects"
    headers = ['Create Date', 'Project No', 'Type', 'Category', 'Client Name',
               'Client Company', 'Client Email', 'Client Phone',
               'Client Coordinator', 'Product Name', 'Members',
               'Reference Brand', 'Reference Product', 'Variant/Type',
               'Area of Application', 'Market Level', 'Appearance',
               'Product Claim', 'Label Claim', 'Costing Range',
               'Odour/Fragrance', 'pH', 'No of Samples', 'MOQ',
               'Product Size', 'Viscosity', 'Packaging Option',
               'Order Quantity', 'Requirement Spec', 'Active Ingredients',
               'Ingredients', 'Video Link', 'NPD Fee Paid', 'Fee Amount',
               'Priority', 'Last Connected', 'Milestones', 'TAT (days)',
               'Status', 'Start Date', 'Lead Days', 'End Date']
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='7C3AED')
    for r in rows:
        ws.append([
            str(r['created_at'] or '')[:10], r['code'] or '',
            'NPD' if r['project_type'] == 'npd' else 'EPD',
            r['product_category'] or '',
            r['client_name'] or '', r['client_company'] or r['cm_company'] or '',
            r['client_email'] or '', r['client_phone'] or '',
            r['client_coordinator'] or '',
            r['product_name'] or '', _members_count(r),
            r['reference_brand'] or '',
            r['reference_product_name'] or r['reference_product'] or '',
            r['variant_type'] or '', r['area_of_application'] or '',
            r['market_level'] or '', r['appearance'] or '',
            r['product_claim'] or '', r['label_claim'] or '',
            r['costing_range'] or '', r['fragrance'] or '',
            r['ph_value'] or '', r['no_of_samples'] or 0, r['moq'] or '',
            r['product_size'] or '', r['viscosity'] or '',
            r['packaging_type'] or '', r['order_quantity'] or '',
            r['requirement_spec'] or '', r['active_ingredients'] or '',
            r['ingredients'] or '', r['video_link'] or '',
            'Yes' if r['npd_fee_paid'] else 'No',
            float(r['npd_fee_amount']) if r['npd_fee_amount'] else '',
            r['priority'] or 'Normal',
            str(r['last_connected'] or '')[:10],
            f"{r['ms_done']}/{r['ms_total']}", r['tat_days'] or 0,
            smap.get(r['status'], (r['status'] or '').replace('_', ' ').title()),
            str(r['project_start_date'] or '')[:10],
            r['project_lead_days'] or '',
            str(r['project_end_date'] or '')[:10]])
    widths = [12, 11, 7, 14, 16, 18, 20, 13, 18, 28, 9, 18, 22, 18, 16, 12,
              20, 24, 24, 16, 18, 10, 11, 9, 11, 18, 22, 13, 30, 24, 26, 22,
              11, 11, 9, 13, 11, 10, 14, 11, 10, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    from flask import send_file
    from datetime import datetime as _dt
    return send_file(buf, as_attachment=True,
                     download_name=f"npd_projects_{_dt.now():%Y%m%d_%H%M}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument'
                              '.spreadsheetml.sheet')


# ── IMPORT (dedicated page — CRM jaisa) ──────────────────────────────────────
IMP_ALIAS = {
    'project_type': ['project_type', 'project type', 'type'],
    'product_name': ['product_name', 'product name', 'product'],
    'product_category': ['product_category', 'product category', 'category'],
    'product_range': ['product_range', 'product range'],
    'client_name': ['client_name', 'client name', 'contact name', 'contact_name'],
    'client_company': ['client_company', 'client company', 'company',
                       'company name', 'company_name'],
    'client_email': ['client_email', 'client email', 'email'],
    'client_phone': ['client_phone', 'client phone', 'phone', 'mobile'],
    'reference_brand': ['reference_brand', 'reference brand', 'brand'],
    'priority': ['priority'],
    'order_quantity': ['order_quantity', 'order quantity', 'order qty', 'qty'],
    'moq': ['moq'],
    'product_size': ['product_size', 'product size', 'size'],
    'requirement_spec': ['requirement_spec', 'requirement spec',
                         'requirement', 'requirement specification', 'specs'],
    'description': ['description', 'desc'],
    'client_coordinator': ['client_coordinator', 'client coordinator',
                           'client co-ordinator', 'coordinator'],
    'area_of_application': ['area_of_application', 'area of application',
                            'area'],
    'market_level': ['market_level', 'market level'],
    'variant_type': ['variant_type', 'variant / type', 'variant/type',
                     'variant', 'variety'],
    'appearance': ['appearance'],
    'product_claim': ['product_claim', 'product claim'],
    'label_claim': ['label_claim', 'label claim'],
    'costing_range': ['costing_range', 'costing range', 'costing'],
    'fragrance': ['fragrance', 'odour / fragrance', 'odour/fragrance',
                  'odour', 'flavours'],
    'ph_value': ['ph_value', 'ph'],
    'viscosity': ['viscosity'],
    'packaging_type': ['packaging_type', 'packaging option', 'packaging'],
    'active_ingredients': ['active_ingredients', 'active ingredients',
                           'active ing. req', 'active ing req'],
    'ingredients': ['ingredients'],
    'video_link': ['video_link', 'video link', 'video'],
    'reference_product_name': ['reference_product_name',
                               'reference product name',
                               'reference product'],
    'reference_product': ['reference / existing hcp product',
                          'existing hcp product'],
    'no_of_samples': ['no_of_samples', 'no of samples', 'samples'],
}
# reverse lookup: header -> canonical field
_IMP_LOOKUP = {a: k for k, aliases in IMP_ALIAS.items() for a in aliases}


def _imp_norm(h):
    return str(h or '').strip().lower().replace('*', '').strip()


def _imp_parse(f):
    """Upload file -> (records, detected_headers). Header row auto-detect
    karta hai (pehli row jisme product/product_name jaisa column ho)."""
    raw_rows = []
    if f.filename.lower().endswith('.csv'):
        import csv, io
        txt = f.read().decode('utf-8-sig', errors='replace')
        for row in csv.reader(io.StringIO(txt)):
            raw_rows.append(list(row))
    else:
        from openpyxl import load_workbook
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            raw_rows.append(list(row))

    # header row dhundo: jis row me product-name jaisa koi alias ho
    hdr_idx, hdr = None, []
    for i, row in enumerate(raw_rows[:10]):
        cells = [_imp_norm(c) for c in row]
        if any(c in _IMP_LOOKUP and _IMP_LOOKUP[c] == 'product_name'
               for c in cells):
            hdr_idx, hdr = i, cells
            break
    if hdr_idx is None and raw_rows:           # fallback: pehli row
        hdr_idx, hdr = 0, [_imp_norm(c) for c in raw_rows[0]]

    recs = []
    for row in raw_rows[hdr_idx + 1:]:
        if not any(v not in (None, '') for v in row):
            continue
        rec = {}
        for i, v in enumerate(row):
            if i >= len(hdr):
                break
            field = _IMP_LOOKUP.get(hdr[i])
            if field and v is not None and str(v).strip() != '':
                rec[field] = str(v).strip()
        recs.append(rec)
    return recs, hdr


@npd_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_projects():
    if request.method == 'GET':
        return _render('npd/import_projects.html', active_item='npd-list')

    f = request.files.get('file')
    if not f or not f.filename:
        flash('No file selected (.xlsx or .csv).', 'error')
        return redirect(url_for('npd.import_projects'))
    if not f.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
        flash('Only .xlsx or .csv files are allowed.', 'error')
        return redirect(url_for('npd.import_projects'))

    try:
        recs, hdr = _imp_parse(f)
    except Exception as e:
        flash(f'Could not read the file: {e}', 'error')
        return redirect(url_for('npd.import_projects'))

    conn = _db()
    ok = skipped = 0
    try:
        for rec in recs:
            pname = rec.get('product_name')
            if not pname:
                skipped += 1
                continue
            ptype = (rec.get('project_type') or 'npd').lower()
            ptype = 'existing' if ptype in ('existing', 'epd') else 'npd'
            pr = (rec.get('priority') or 'Normal').capitalize()
            extra_cols = ['client_coordinator', 'area_of_application',
                          'market_level', 'variant_type', 'appearance',
                          'product_claim', 'label_claim', 'costing_range',
                          'fragrance', 'ph_value', 'viscosity',
                          'packaging_type', 'active_ingredients',
                          'ingredients', 'video_link',
                          'reference_product_name', 'reference_product',
                          'no_of_samples']
            cur = conn.execute(
                "INSERT INTO `npd_projects` (project_type, status, "
                "product_name, product_category, product_range, client_name,"
                " client_company, client_email, client_phone, "
                "reference_brand, priority, order_quantity, moq, "
                "product_size, requirement_spec, description, assigned_sc, "
                "milestone_master_created, created_by, "
                + ",".join(f"`{c}`" for c in extra_cols) + ") "
                "VALUES (%s,'not_started',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,"
                "%s,%s,%s,%s,%s,1,%s,"
                + ",".join(["%s"] * len(extra_cols)) + ")",
                [ptype, pname, rec.get('product_category'),
                 rec.get('product_range'), rec.get('client_name'),
                 rec.get('client_company'), rec.get('client_email'),
                 rec.get('client_phone'), rec.get('reference_brand'),
                 pr if pr in PRIORITIES else 'Normal',
                 rec.get('order_quantity'), rec.get('moq'),
                 rec.get('product_size'), rec.get('requirement_spec'),
                 rec.get('description'), _uid(), _uid()]
                + [(rec.get(c) or 0 if c == 'no_of_samples'
                    else rec.get(c)) for c in extra_cols])
            pid = cur.lastrowid
            conn.execute("UPDATE `npd_projects` SET code=%s WHERE id=%s",
                         ("NPD-%04d" % pid, pid))
            conn.execute(
                "INSERT INTO `milestone_masters` (project_id, milestone_type,"
                " title, description, is_selected, status, sort_order, "
                "created_by) SELECT %s, milestone_type, title, description, "
                "default_selected, 'pending', sort_order, %s "
                "FROM `npd_milestone_templates` WHERE is_active=1 "
                "AND applies_to IN ('both', %s) ORDER BY sort_order",
                (pid, _uid(), ptype))
            _log(conn, pid, "Project imported (NPD-%04d)" % pid)
            ok += 1
        conn.commit()
    except Exception as e:
        flash(f'Import error: {e}', 'error')
        conn.close()
        return redirect(url_for('npd.import_projects'))
    finally:
        try:
            conn.close()
        except Exception:
            pass

    if ok:
        flash(f'Import complete — {ok} project(s) created'
              + (f', {skipped} row(s) skipped (product name missing).'
                 if skipped else '.'), 'success')
        return redirect(url_for('npd.projects'))
    # 0 import -> user ko batao headers kya mile, taaki turant fix kar sake
    known = [h for h in hdr if h in _IMP_LOOKUP]
    flash('0 projects imported. The file must have a "product_name" (or "Product Name"/'
          '"Product") column. Headers found in the file: '
          + (', '.join(h for h in hdr if h) or '(none)')
          + (f' — recognized among them: {", ".join(known)}.' if known
             else ' — none of them were recognized. Please download '
                  'and use the template.'), 'error')
    return redirect(url_for('npd.import_projects'))


@npd_bp.route('/import/template')
@login_required
def import_template():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = Workbook(); ws = wb.active; ws.title = 'NPD Projects'
    headers = ['product_name', 'project_type', 'category', 'product_range',
               'client_name', 'client_company', 'client_email',
               'client_phone', 'client_coordinator', 'area_of_application',
               'market_level', 'reference_brand', 'reference_product_name',
               'variant_type', 'appearance', 'product_claim', 'label_claim',
               'costing_range', 'fragrance', 'ph_value', 'viscosity',
               'packaging_type', 'priority', 'order_quantity', 'moq',
               'product_size', 'no_of_samples', 'active_ingredients',
               'ingredients', 'video_link', 'requirement_spec',
               'description']
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='7C3AED')
    ws.append(['Vitamin C Face Wash', 'npd', 'Skin Care', 'Premium',
               'John Doe', 'ABC Corp', 'john@abc.com', '9999999999',
               'Neha Soni', 'Face', 'Premium', 'Foxtale',
               'Foxtale Oil Face Wash', 'Gel based', 'Clear gel',
               'Brightens skin', 'Vitamin C 10%', 'MRP ₹499-699',
               'Citrus', '5.5', 'Medium', 'Fliptop bottle', 'Normal',
               '500 units', '1000', '100ml', 5,
               'Vitamin C, Niacinamide', 'Aqua, Vitamin C, Glycerin',
               'https://example.com/ref', 'Clear gel face wash',
               'Urgent requirement'])
    dv_t = DataValidation(type='list', formula1='"npd,epd"', allow_blank=True)
    dv_p = DataValidation(type='list', formula1='"Urgent,High,Normal,Low"',
                          allow_blank=True)
    ws.add_data_validation(dv_t); ws.add_data_validation(dv_p)
    dv_t.add('B2:B1000'); dv_p.add('W2:W1000')
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
            max(13, min(26, len(headers[i - 1]) + 4))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    from flask import send_file
    return send_file(buf, as_attachment=True,
                     download_name='npd_import_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument'
                              '.spreadsheetml.sheet')


# ── FORM FIELDS (old NPD form ke saare fields) ──────────────────────────────
MARKET_LEVELS = ['Premium', 'Masstige', 'Mass', 'Luxury', 'Economy']

# text/select fields jo seedha column me jaate hain
FORM_FIELDS = ['client_id', 'client_name', 'client_company', 'client_email',
               'client_phone', 'client_coordinator', 'product_name',
               'product_category', 'area_of_application',
               'market_level', 'priority', 'status', 'description',
               'ingredients', 'video_link', 'active_ingredients',
               'reference_brand', 'reference_product_name', 'variant_type',
               'appearance', 'product_claim', 'label_claim', 'costing_range',
               'fragrance', 'ph_value', 'no_of_samples', 'moq',
               'product_size', 'viscosity', 'packaging_type',
               'order_quantity', 'requirement_spec', 'reference_product',
               'project_start_date', 'project_lead_days', 'project_end_date']


def _save_receipt(files):
    """npd_fee_receipt file -> static/uploads/npd me save, filename return."""
    f = files.get('npd_fee_receipt')
    if not f or not f.filename:
        return None
    import os
    from datetime import datetime as _dt
    from werkzeug.utils import secure_filename
    fn = f"receipt_{_dt.now():%Y%m%d%H%M%S}_{secure_filename(f.filename)}"
    f.save(os.path.join(_npd_upload_dir(), fn))
    return fn


def _form_values(f, conn):
    """POST form -> {column: value} dict (old form ke saare fields)."""
    vals = {}
    for k in FORM_FIELDS:
        v = (f.get(k) or '').strip()
        if k == 'priority':
            v = v if v in PRIORITIES else 'Normal'
        elif k == 'status':
            slugs = [s['slug'] for s in get_npd_statuses(conn)]
            v = v if v in slugs else 'not_started'
        elif k in ('no_of_samples', 'project_lead_days'):
            v = v or (0 if k == 'no_of_samples' else None)
        else:
            v = v or None
        vals[k] = v
    # multi-selects -> CSV (old format: comma-sep user ids)
    vals['assigned_members'] = ','.join(f.getlist('assigned_members')) or None
    rd = f.getlist('assigned_rd_members')
    vals['assigned_rd_members'] = ','.join(rd) or None
    vals['assigned_rd'] = int(rd[0]) if rd else None
    # NPD fee
    vals['npd_fee_paid'] = 1 if f.get('npd_fee_paid') else 0
    vals['npd_fee_amount'] = f.get('npd_fee_amount') or 10000
    return vals


def _form_context(conn, **extra):
    """Form (add/edit) render ke liye common data."""
    clients = conn.execute(
        "SELECT id, code, company_name, contact_name, email, mobile "
        "FROM `client_masters` WHERE is_deleted=0 "
        "ORDER BY COALESCE(NULLIF(contact_name,''), company_name)"
    ).fetchall() or []
    users = conn.execute(
        "SELECT id, COALESCE(NULLIF(full_name,''), username) AS nm "
        "FROM `User_Tbl` WHERE COALESCE(is_active,1)=1 "
        "ORDER BY nm").fetchall() or []
    cats = _npd_cats(conn)
    templates = conn.execute(
        "SELECT milestone_type, title, icon, applies_to, default_selected "
        "FROM `npd_milestone_templates` WHERE is_active=1 "
        "ORDER BY sort_order").fetchall() or []
    ctx = dict(clients=clients, users=users, cats=cats,
               ms_templates=templates, statuses=get_npd_statuses(conn),
               priorities=PRIORITIES, market_levels=MARKET_LEVELS)
    ctx.update(extra)
    return ctx


def _sync_milestones(conn, pid, ptype, selected):
    """Selected milestone types ko project ke milestone_masters me sync karo
    (missing template rows insert, is_selected update)."""
    tpls = conn.execute(
        "SELECT milestone_type, title, description, sort_order "
        "FROM `npd_milestone_templates` WHERE is_active=1 "
        "AND applies_to IN ('both', %s) ORDER BY sort_order",
        (ptype,)).fetchall() or []
    existing = {m['milestone_type']: m for m in (conn.execute(
        "SELECT id, milestone_type FROM `milestone_masters` "
        "WHERE project_id=%s", (pid,)).fetchall() or [])}
    for t in tpls:
        sel = 1 if t['milestone_type'] in selected else 0
        if t['milestone_type'] in existing:
            conn.execute(
                "UPDATE `milestone_masters` SET is_selected=%s "
                "WHERE id=%s", (sel, existing[t['milestone_type']]['id']))
        else:
            conn.execute(
                "INSERT INTO `milestone_masters` (project_id, "
                "milestone_type, title, description, is_selected, status, "
                "sort_order, created_by) VALUES (%s,%s,%s,%s,%s,'pending',"
                "%s,%s)", (pid, t['milestone_type'], t['title'],
                           t['description'], sel, t['sort_order'], _uid()))


# ── EDIT ─────────────────────────────────────────────────────────────────────
@npd_bp.route('/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
def edit(pid):
    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        proj = conn.execute(
            "SELECT p.* FROM `npd_projects` p WHERE p.id=%s "
            "AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not proj:
            flash('Project not found or you do not have access.', 'error')
            return redirect(url_for('npd.projects'))

        if request.method == 'POST':
            f = request.form
            if not (f.get('product_name') or '').strip():
                flash('Product name is required.', 'error')
                return redirect(request.url)
            vals = _form_values(f, conn)
            # fee receipt: nayi file aayi to replace, warna purani rehne do
            rc = _save_receipt(request.files)
            if rc:
                vals['npd_fee_receipt'] = rc
            # fee abhi paid hua to timestamp (old system jaisa)
            if vals['npd_fee_paid'] and not proj.get('npd_fee_paid'):
                vals['npd_fee_paid_at'] = None  # set via NOW() below
            sets = [f"{k}=%s" for k in vals] + ["updated_by=%s"]
            params = list(vals.values()) + [_uid()]
            sql = ("UPDATE `npd_projects` SET " + ", ".join(sets)
                   + (", npd_fee_paid_at=NOW()"
                      if (vals['npd_fee_paid']
                          and not proj.get('npd_fee_paid')) else "")
                   + " WHERE id=%s")
            # vals me None-marker hata do (NOW() handle alag se)
            if 'npd_fee_paid_at' in vals:
                del vals['npd_fee_paid_at']
                sets = [f"{k}=%s" for k in vals] + ["updated_by=%s"]
                params = list(vals.values()) + [_uid()]
                sql = ("UPDATE `npd_projects` SET " + ", ".join(sets)
                       + ", npd_fee_paid_at=NOW() WHERE id=%s")
            conn.execute(sql, params + [pid])
            _sync_milestones(conn, pid, proj['project_type'],
                             set(f.getlist('milestones')))
            _log(conn, pid, f"Project {proj['code']} edited")
            conn.commit()
            flash('Project updated.', 'success')
            return redirect(url_for('npd.view', pid=pid))

        selected_ms = {m['milestone_type'] for m in (conn.execute(
            "SELECT milestone_type FROM `milestone_masters` "
            "WHERE project_id=%s AND is_selected=1", (pid,)).fetchall()
            or [])}
        return _render('npd/project_form.html',
                       **_form_context(conn, v=proj, edit=proj,
                                       lead=None,
                                       sel_client=proj.get('client_id'),
                                       sel_type=proj['project_type'],
                                       selected_ms=selected_ms),
                       active_item='npd-list')
    finally:
        conn.close()


# ── INLINE EDIT (listing se cell-level quick edit — leads jaisa) ─────────────
IE_ALLOWED = {
    'product_name', 'product_category', 'client_name', 'client_company',
    'client_email', 'client_phone', 'client_coordinator', 'reference_brand',
    'reference_product_name', 'variant_type', 'area_of_application',
    'market_level', 'appearance', 'fragrance', 'viscosity', 'ph_value',
    'packaging_type', 'product_size', 'no_of_samples', 'moq',
    'order_quantity', 'costing_range', 'priority', 'status',
    'product_claim', 'label_claim', 'requirement_spec', 'description',
    'video_link', 'delay_reason',
}


@npd_bp.route('/<int:pid>/inline-edit', methods=['POST'])
@login_required
def inline_edit(pid):
    from flask import jsonify
    data = request.get_json(silent=True) or {}
    field = (data.get('field') or '').strip()
    value = data.get('value', '')
    if field not in IE_ALLOWED:
        return jsonify(success=False, error='Field not allowed'), 400
    if isinstance(value, str):
        value = value.strip()
    conn = _db()
    try:
        if field == 'priority' and value not in PRIORITIES:
            return jsonify(success=False, error='Invalid priority'), 400
        if field == 'status':
            if value not in [s['slug'] for s in get_npd_statuses(conn)]:
                return jsonify(success=False, error='Invalid status'), 400
        if field == 'no_of_samples':
            try:
                value = int(value) if value not in ('', None) else 0
            except Exception:
                return jsonify(success=False, error='Invalid number'), 400
        if value == '':
            value = None

        vis_frag, vis_params = _proj_visibility_sql('p')
        proj = conn.execute(
            "SELECT p.id, p.code FROM `npd_projects` p "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not proj:
            return jsonify(success=False, error='Project not found'), 404

        extra = ""
        if field == 'status':
            extra = (", completed_at=CASE WHEN %s='finish' THEN NOW() "
                     "ELSE completed_at END, "
                     "cancelled_at=CASE WHEN %s='cancelled' THEN NOW() "
                     "ELSE cancelled_at END")
        sql = (f"UPDATE `npd_projects` SET `{field}`=%s, updated_by=%s"
               + extra + " WHERE id=%s")
        params = [value, _uid()] + ([value, value] if extra else []) + [pid]
        conn.execute(sql, params)
        _log(conn, pid, f"Inline edit: {field} updated")
        conn.commit()
        return jsonify(success=True, field=field, value=value)
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify(success=False, error=str(e)), 500
    finally:
        conn.close()


# ── DELETE / RESTORE (soft) ──────────────────────────────────────────────────
@npd_bp.route('/<int:pid>/delete', methods=['POST'])
@login_required
def delete(pid):
    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('npd_projects')
        cur = conn.execute(
            "UPDATE `npd_projects` SET is_deleted=1, deleted_at=NOW(), "
            "deleted_by=%s WHERE id=%s AND is_deleted=0" + vis_frag,
            [_uid(), pid] + vis_params)
        if cur.rowcount:
            _log(conn, pid, "Project deleted")
            conn.commit()
            flash('Project deleted (you can find it in the Deleted tab).', 'success')
        else:
            flash('You do not have access.', 'error')
    finally:
        conn.close()
    return redirect(url_for('npd.projects'))


@npd_bp.route('/<int:pid>/restore', methods=['POST'])
@login_required
def restore(pid):
    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('npd_projects')
        cur = conn.execute(
            "UPDATE `npd_projects` SET is_deleted=0, deleted_at=NULL, "
            "deleted_by=NULL WHERE id=%s AND is_deleted=1" + vis_frag,
            [pid] + vis_params)
        if cur.rowcount:
            _log(conn, pid, "Project restored")
            conn.commit()
            flash('Project restored.', 'success')
        else:
            flash('You do not have access.', 'error')
    finally:
        conn.close()
    return redirect(url_for('npd.projects', show='deleted'))


def _purge_one(conn, pid):
    """Project + saare child records permanently delete (commit caller pe)."""
    conn.execute("DELETE FROM `milestone_logs` WHERE milestone_id IN "
                 "(SELECT id FROM `milestone_masters` WHERE project_id=%s)",
                 (pid,))
    for tbl in ('milestone_masters', 'npd_activity_logs',
                'npd_comments', 'npd_notes', 'client_dispatch'):
        conn.execute(f"DELETE FROM `{tbl}` WHERE project_id=%s", (pid,))
    # dispatch items project_id carry karte hain; tokens batch-level
    # hote hain (kai projects share karte hain) — unhe mat chhedo
    conn.execute("DELETE FROM `office_dispatch_items` WHERE project_id=%s",
                 (pid,))
    conn.execute("DELETE FROM `npd_projects` WHERE id=%s", (pid,))


@npd_bp.route('/<int:pid>/purge', methods=['POST'])
@login_required
def purge(pid):
    """Permanent delete — sirf deleted projects, sirf full-visibility users."""
    if not _has_full_visibility():
        flash('You do not have permission to permanently delete.', 'error')
        return redirect(url_for('npd.projects', show='deleted'))
    conn = _db()
    try:
        proj = conn.execute(
            "SELECT id, code FROM `npd_projects` "
            "WHERE id=%s AND is_deleted=1", (pid,)).fetchone()
        if not proj:
            flash('Project not found in the Deleted tab — delete it first.',
                  'error')
            return redirect(url_for('npd.projects', show='deleted'))
        _purge_one(conn, pid)
        conn.commit()
        flash(f"{proj['code']} permanently deleted.", 'success')
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f'Permanent delete failed: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('npd.projects', show='deleted'))


@npd_bp.route('/bulk-action', methods=['POST'])
@login_required
def bulk_action():
    """Selected projects pe bulk delete / restore / purge (AJAX JSON)."""
    from flask import jsonify
    data = request.get_json(silent=True) or {}
    action = data.get('action')
    try:
        ids = [int(i) for i in (data.get('ids') or [])][:200]
    except Exception:
        ids = []
    if action not in ('delete', 'restore', 'purge') or not ids:
        return jsonify(success=False, error='Invalid request'), 400
    conn = _db()
    try:
        ph = ",".join(["%s"] * len(ids))
        vis_frag, vis_params = _proj_visibility_sql('npd_projects')
        if action == 'delete':
            cur = conn.execute(
                f"UPDATE `npd_projects` SET is_deleted=1, deleted_at=NOW(), "
                f"deleted_by=%s WHERE id IN ({ph}) AND is_deleted=0"
                + vis_frag, [_uid()] + ids + vis_params)
            n = cur.rowcount
        elif action == 'restore':
            cur = conn.execute(
                f"UPDATE `npd_projects` SET is_deleted=0, deleted_at=NULL, "
                f"deleted_by=NULL WHERE id IN ({ph}) AND is_deleted=1"
                + vis_frag, ids + vis_params)
            n = cur.rowcount
        else:  # purge — sirf full-visibility
            if not _has_full_visibility():
                return jsonify(success=False,
                               error='You do not have permission'), 403
            rows = conn.execute(
                f"SELECT id FROM `npd_projects` "
                f"WHERE id IN ({ph}) AND is_deleted=1", ids).fetchall() or []
            for r in rows:
                _purge_one(conn, r['id'])
            n = len(rows)
        conn.commit()
        return jsonify(success=True, count=n)
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify(success=False, error=str(e)), 500
    finally:
        conn.close()


# ── NPD FORM (printable / download) ──────────────────────────────────────────
@npd_bp.route('/<int:pid>/form')
@login_required
def form_print(pid):
    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        proj = conn.execute(
            "SELECT p.*, c.company_name AS cm_company, c.code AS cm_code "
            "FROM `npd_projects` p "
            "LEFT JOIN `client_masters` c ON c.id = p.client_id "
            "WHERE p.id=%s" + vis_frag, [pid] + vis_params).fetchone()
        if not proj:
            flash('Project not found or you do not have access.', 'error')
            return redirect(url_for('npd.projects'))
        umap = _user_map(conn)
        return render_template('npd/npd_form_print.html', p=proj, umap=umap)
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# CREATE  (endpoint name 'create' — CRM gate url_for('npd.create') use karta hai)
# ─────────────────────────────────────────────────────────────────────────────
@npd_bp.route('/new', methods=['GET', 'POST'])
@login_required
def create():
    conn = _db()
    if not conn:
        flash('Database connection failed.', 'error')
        return redirect('/')
    try:
        if request.method == 'POST':
            f = request.form
            ptype = f.get('project_type') \
                if f.get('project_type') in ('npd', 'existing') else 'npd'
            if not (f.get('product_name') or '').strip():
                flash('Product name is required.', 'error')
                return redirect(request.url)

            vals = _form_values(f, conn)
            vals['project_type'] = ptype
            vals['lead_id'] = f.get('lead_id') or None
            vals['assigned_sc'] = f.get('assigned_sc') or _uid()
            vals['milestone_master_created'] = 1
            vals['created_by'] = _uid()
            rc = _save_receipt(request.files)
            if rc:
                vals['npd_fee_receipt'] = rc

            cols = list(vals.keys())
            sql = ("INSERT INTO `npd_projects` ("
                   + ",".join(f"`{c}`" for c in cols)
                   + (", npd_fee_paid_at" if vals['npd_fee_paid'] else "")
                   + ") VALUES (" + ",".join(["%s"] * len(cols))
                   + (", NOW()" if vals['npd_fee_paid'] else "") + ")")
            cur = conn.execute(sql, list(vals.values()))
            pid = cur.lastrowid

            code = "NPD-%04d" % pid
            conn.execute("UPDATE `npd_projects` SET code=%s WHERE id=%s",
                         (code, pid))

            # milestones: form me jo select kiye wahi follow honge
            selected = set(f.getlist('milestones'))
            if not selected:        # kuch select nahi -> defaults
                selected = {t['milestone_type'] for t in (conn.execute(
                    "SELECT milestone_type FROM `npd_milestone_templates` "
                    "WHERE is_active=1 AND default_selected=1").fetchall()
                    or [])}
            _sync_milestones(conn, pid, ptype, selected)

            _log(conn, pid, f"Project {code} created")
            if f.get('lead_id'):
                try:
                    conn.execute(
                        "INSERT INTO `lead_activity_logs` "
                        "(lead_id, user_id, action) VALUES (%s,%s,%s)",
                        (f.get('lead_id'), _uid(),
                         f"{'NPD' if ptype == 'npd' else 'EPD'} Project "
                         f"{code} created"))
                except Exception:
                    pass
            conn.commit()
            flash(f'Project {code} created.', 'success')
            return redirect(url_for('npd.view', pid=pid))

        # ── GET: prefill from lead/client ────────────────────────────────
        lead = None
        lead_id = request.args.get('lead_id')
        client_id = request.args.get('client_id')
        if lead_id:
            lead = conn.execute(
                "SELECT id, contact_name, company_name, email, "
                "phone AS mobile, client_id, category, product_range, "
                "requirement_spec, product_name, "
                "order_quantity AS order_qty "
                "FROM `leads` WHERE id=%s AND is_deleted=0",
                (lead_id,)).fetchone()
            if lead and not client_id:
                client_id = lead.get('client_id')

        # lead -> form prefill dict (v)
        v = {}
        if lead:
            # NOTE: client_name/company prefill NAHI — client dropdown se
            # select hoga, company auto-fill hogi (warna validation bypass
            # ho jaata tha aur company me lead ki default value aati thi)
            v = {'client_email': lead.get('email'),
                 'client_phone': lead.get('mobile'),
                 'product_name': lead.get('product_name'),
                 'product_category': lead.get('category'),
                 'product_range': lead.get('product_range'),
                 'requirement_spec': lead.get('requirement_spec'),
                 'order_quantity': lead.get('order_qty')}

        # Client CO-Ordinator default = jo login hai (sirf naye form pe)
        if not v.get('client_coordinator'):
            try:
                u = conn.execute(
                    "SELECT COALESCE(NULLIF(full_name,''), username) AS nm "
                    "FROM `User_Tbl` WHERE id=%s", (_uid(),)).fetchone()
                v['client_coordinator'] = (u or {}).get('nm') or _uname()
            except Exception:
                v['client_coordinator'] = _uname()

        sel_type = request.args.get('type') \
            if request.args.get('type') in ('npd', 'existing') else 'npd'
        return _render('npd/project_form.html',
                       **_form_context(conn, v=v, edit=None, lead=lead,
                                       sel_client=int(client_id)
                                       if client_id else None,
                                       sel_type=sel_type, selected_ms=None),
                       active_item='npd-new')
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# VIEW
# ─────────────────────────────────────────────────────────────────────────────
@npd_bp.route('/<int:pid>')
@login_required
def view(pid):
    conn = _db()
    if not conn:
        flash('Database connection failed.', 'error')
        return redirect('/')
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        proj = conn.execute(
            "SELECT p.*, c.company_name AS cm_company, "
            "c.code AS cm_code, l.contact_name AS lead_contact, "
            "l.company_name AS lead_company "
            "FROM `npd_projects` p "
            "LEFT JOIN `client_masters` c ON c.id = p.client_id "
            "LEFT JOIN `leads` l ON l.id = p.lead_id "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not proj:
            flash('Project not found or you do not have access.', 'error')
            return redirect(url_for('npd.projects'))

        miles = conn.execute(
            "SELECT * FROM `milestone_masters` "
            "WHERE project_id=%s AND is_selected=1 "
            "ORDER BY sort_order, id", (pid,)).fetchall() or []
        acts = conn.execute(
            "SELECT * FROM `npd_activity_logs` WHERE project_id=%s "
            "ORDER BY created_at DESC LIMIT 50", (pid,)).fetchall() or []
        comments = conn.execute(
            "SELECT * FROM `npd_comments` WHERE project_id=%s "
            "AND is_internal=0 AND (board IS NULL OR board='') "
            "ORDER BY created_at DESC", (pid,)).fetchall() or []
        internal = conn.execute(
            "SELECT * FROM `npd_comments` WHERE project_id=%s "
            "AND is_internal=1 AND (board IS NULL OR board='') "
            "ORDER BY created_at DESC", (pid,)).fetchall() or []
        acomments = conn.execute(
            "SELECT * FROM `npd_comments` WHERE project_id=%s "
            "AND board='artwork' ORDER BY created_at DESC",
            (pid,)).fetchall() or []
        qcomments = conn.execute(
            "SELECT * FROM `npd_comments` WHERE project_id=%s "
            "AND board='artwork_qc' ORDER BY created_at DESC",
            (pid,)).fetchall() or []
        note = conn.execute(
            "SELECT * FROM `npd_notes` WHERE project_id=%s",
            (pid,)).fetchone() or {}
        umap = _user_map(conn)
        # Quotations (project se ya project ke lead se bani hui)
        quots = []
        try:
            from crm.crm_leads_routes import _ensure_quotations
            _ensure_quotations(conn)
            q_where = "q.is_deleted=0 AND (q.project_id=%s"
            q_params = [pid]
            if proj.get('lead_id'):
                q_where += " OR q.lead_id=%s"
                q_params.append(proj['lead_id'])
            q_where += ")"
            import json as _qjson
            for q in (conn.execute(
                    "SELECT q.* FROM `quotations` q WHERE " + q_where +
                    " ORDER BY q.id DESC", q_params).fetchall() or []):
                try:
                    _items = _qjson.loads(q.get('items_json') or '[]')
                except Exception:
                    _items = []
                quots.append({'id': q['id'], 'number': q['quot_number'],
                              'date': str(q['quot_date'] or '')[:10],
                              'valid_until': str(q['valid_until'] or '')[:10],
                              'company': q['bill_company'] or '',
                              'address': q['bill_address'] or '',
                              'phone': q['bill_phone'] or '',
                              'email': q['bill_email'] or '',
                              'gst_no': q['bill_gst'] or '',
                              'gst_pct': float(q['gst_pct'] or 18),
                              'subject': q['subject'] or '',
                              'terms': q['terms'] or '',
                              'notes': q['notes'] or '',
                              'items': _items,
                              'total': float(q['total_amount'] or 0),
                              'status': q['status'] or 'draft',
                              'by': umap.get(q['created_by'], '')})
        except Exception:
            quots = []
        packs = [dict(r) for r in (conn.execute(
            "SELECT * FROM `npd_packing_rows` WHERE project_id=%s "
            "ORDER BY id", (pid,)).fetchall() or [])]
        for pk in packs:
            pk['cost'] = float(pk['cost']) if pk.get('cost') is not None \
                else None
            pk['created_at'] = str(pk.get('created_at') or '')
            pk['updated_at'] = str(pk.get('updated_at') or '')
        bom_rows_db = conn.execute(
            "SELECT b.*, r.id AS rid, r.sr_no, r.inci_name, r.qty_pct "
            "FROM `npd_boms` b LEFT JOIN `npd_bom_rows` r ON r.bom_id=b.id "
            "WHERE b.project_id=%s ORDER BY b.id DESC, r.sr_no, r.id",
            (pid,)).fetchall() or []
        boms, _bm = [], {}
        for r in bom_rows_db:
            b = _bm.get(r['id'])
            if not b:
                b = {'id': r['id'], 'product_name': r['product_name'] or '',
                     'code': r['code'] or '', 'variant': r['variant'] or '',
                     'by': umap.get(r['created_by'], ''),
                     'created_at': str(r['created_at'] or '')[:16],
                     'rows': []}
                _bm[r['id']] = b
                boms.append(b)
            if r['rid']:
                b['rows'].append({'inci': r['inci_name'] or '',
                                  'qty': float(r['qty_pct'] or 0)})
        mrows = conn.execute(
            "SELECT l.* FROM `milestone_logs` l "
            "JOIN `milestone_masters` mm ON mm.id = l.milestone_id "
            "WHERE mm.project_id=%s ORDER BY l.created_at DESC, l.id DESC",
            (pid,)).fetchall() or []
        mlogs = {}
        for r in mrows:
            mlogs.setdefault(r['milestone_id'], []).append(dict(r))
        # Har milestone ke log me uske related project activities bhi merge
        # karo (legacy data + saari activities).
        def _ms_matcher(mt, title):
            mp = {
                'bom': lambda a: a.startswith("BOM '")
                                  or a.startswith('BOM '),
                'ingredients':
                    lambda a: 'Ingredients List' in a,
                'quotation': lambda a: a.startswith('Quotation '),
                'packing_material':
                    lambda a: a.startswith('Packing material'),
                'artwork': lambda a: a.startswith('Artwork / Design'),
                'artwork_qc': lambda a: a.startswith('Artwork QC'),
            }
            base = mp.get(mt, lambda a: False)
            tprefix = (f"Milestone '{title}'" if title else None)
            sprefix = (f"'{title}'" if title else None)
            return lambda a: base(a) or (tprefix and tprefix in a) \
                or (sprefix and a.startswith(sprefix))

        for _m in miles:
            mat = _ms_matcher(_m.get('milestone_type'), _m.get('title'))
            have = {(lg.get('action'), str(lg.get('created_at') or ''))
                    for lg in mlogs.get(_m['id'], [])}
            extra = [{'milestone_id': _m['id'], 'action': a['action'],
                      'created_by': a['user_id'],
                      'created_at': a['created_at']}
                     for a in acts
                     if mat(str(a.get('action') or ''))
                     and (a['action'], str(a['created_at'] or ''))
                     not in have]
            if extra:
                merged = mlogs.get(_m['id'], []) + extra
                merged.sort(key=lambda x: str(x.get('created_at') or ''),
                            reverse=True)
                mlogs[_m['id']] = merged
        pnotes = conn.execute(
            "SELECT * FROM `npd_personal_notes` "
            "WHERE project_id=%s AND user_id=%s "
            "ORDER BY created_at DESC, id DESC", (pid, _uid())).fetchall() or []
        # comment files (multi-attachment table) + legacy single attachment
        frows = conn.execute(
            "SELECT * FROM `npd_comment_files` WHERE project_id=%s "
            "ORDER BY id", (pid,)).fetchall() or []
        fmap = {}
        for fr in frows:
            fmap.setdefault(fr['comment_id'], []).append(
                {'id': fr['id'], 'name': fr['file_name'],
                 'path': fr['file_path']})
        comments = [dict(c) for c in comments]
        internal = [dict(c) for c in internal]
        acomments = [dict(c) for c in acomments]
        qcomments = [dict(c) for c in qcomments]
        for c in comments + internal + acomments + qcomments:
            c['_files'] = list(fmap.get(c['id'], []))
            if c.get('attachment'):                      # legacy single file
                c['_files'].append({'id': '', 'name':
                                    c['attachment'].split('_', 2)[-1],
                                    'path': c['attachment']})
        # mentions ke liye users
        all_users = [{'id': r['id'], 'name': r['nm']} for r in (conn.execute(
            "SELECT id, COALESCE(NULLIF(full_name,''), username) AS nm "
            "FROM `User_Tbl` WHERE COALESCE(is_active,1)=1 "
            "ORDER BY nm").fetchall() or [])]
        fda_users = [{'id': r['id'], 'name': r['nm'], 'email': r['email']}
                      for r in (conn.execute(
            "SELECT id, COALESCE(NULLIF(full_name,''), username) AS nm, "
            "email FROM `User_Tbl` WHERE COALESCE(is_active,1)=1 "
            "AND COALESCE(email,'') != '' ORDER BY nm").fetchall() or [])]
        fda_ms = next((m for m in miles if m.get('milestone_type') == 'fda'), None)
        fda_request = None
        fda_entries = []
        if fda_ms:
            fda_request = conn.execute(
                "SELECT * FROM `npd_fda_requests` WHERE project_id=%s "
                "ORDER BY id DESC LIMIT 1", (pid,)).fetchone()
            fda_entries = conn.execute(
                "SELECT * FROM `npd_fda_entries` WHERE project_id=%s "
                "ORDER BY id", (pid,)).fetchall() or []

        bc_ms = next((m for m in miles if m.get('milestone_type') == 'barcode'), None)
        bc_designs = []
        if bc_ms:
            bc_designs = conn.execute(
                "SELECT * FROM `npd_barcode_designs` WHERE project_id=%s "
                "ORDER BY sr_no, id", (pid,)).fetchall() or []
        # Attachments list: SIRF Discussion board ki files (internal nahi)
        # + milestone attachments
        attachments = [{'name': f['name'], 'path': f['path'],
                        'src': 'Discussion',
                        'when': c['created_at'], 'by': c['user_id']}
                       for c in comments for f in c['_files']]
        attachments += [{'name': f['name'], 'path': f['path'],
                         'src': 'Artwork / Design',
                         'when': c['created_at'], 'by': c['user_id']}
                        for c in acomments for f in c['_files']]
        attachments += [{'name': f['name'], 'path': f['path'],
                         'src': 'Artwork QC',
                         'when': c['created_at'], 'by': c['user_id']}
                        for c in qcomments for f in c['_files']]
        for m in miles:
            for fn in (m.get('attachments') or '').split(','):
                if fn.strip():
                    attachments.append({'name': fn.strip(),
                                        'path': fn.strip(),
                                        'src': f"Milestone: {m['title']}",
                                        'when': m.get('updated_at'),
                                        'by': m.get('created_by')})
        done = sum(1 for m in miles if m['status'] == 'approved')
        pct = round(done / len(miles) * 100) if miles else 0

        # Duration + START permission
        from datetime import datetime as _dt
        def _pdt(v):
            try:
                return _dt.strptime(str(v)[:19], '%Y-%m-%d %H:%M:%S')
            except Exception:
                return None
        st_at, fin_at = _pdt(proj.get('started_at')), _pdt(proj.get('finished_at'))
        if st_at:
            dur_days = ((fin_at or _dt.now()) - st_at).days
        else:
            dur_days = 0
        # elapsed seconds for the live duration timer (visible to all viewers)
        if st_at and fin_at:
            elapsed_seconds = int((fin_at - st_at).total_seconds())
        elif st_at:
            elapsed_seconds = int((_dt.now() - st_at).total_seconds())
        else:
            elapsed_seconds = 0
        # START/FINISH only after assignment, and only for the assigned R&D person
        can_start = _is_rd_assignee(conn, pid, _uid())

        return _render('npd/project_view.html', p=proj, miles=miles,
                       acts=acts, comments=comments, internal=internal,
                       note=note, pnotes=pnotes, mlogs=mlogs, boms=boms,
                       quots=quots, packs=packs, acomments=acomments, qcomments=qcomments,
                       attachments=attachments,
                       dur_days=dur_days, can_start=can_start,
                       elapsed_seconds=elapsed_seconds,
                       started=bool(st_at), finished=bool(fin_at),
                       umap=umap, pct=pct, all_users=all_users,
                       fda_users=fda_users, fda_request=fda_request,
                       fda_entries=fda_entries, bc_designs=bc_designs,
                       my_uid=_uid(),
                       statuses=get_npd_statuses(conn), active_item='npd-list')
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# DISCUSSION / NOTE / TIMER
# ─────────────────────────────────────────────────────────────────────────────
def _npd_upload_dir():
    import os
    from flask import current_app
    d = os.path.join(current_app.root_path, 'static', 'uploads', 'npd')
    os.makedirs(d, exist_ok=True)
    return d


@npd_bp.route('/<int:pid>/comment', methods=['POST'])
@login_required
def add_comment(pid):
    import os
    from datetime import datetime as _dt
    from flask import jsonify
    from werkzeug.utils import secure_filename
    is_internal = request.form.get('is_internal', '0') == '1'
    board = (request.form.get('board') or '').strip().lower()
    if board not in ('', 'artwork', 'artwork_qc'):
        board = ''
    comment = (request.form.get('comment') or '').strip()
    files = request.files.getlist('files')
    legacy = request.files.get('file')
    if legacy and legacy.filename:
        files = files + [legacy]
    ajax = request.is_json or request.headers.get('X-Requested-With')
    if not comment and not any(f and f.filename for f in files):
        if ajax:
            return jsonify(ok=False, error='A comment or file is required.'), 400
        flash('A comment or file is required.', 'error')
        return redirect(url_for('npd.view', pid=pid)
                        + ('#internal' if is_internal else '#discussion'))
    conn = _db()
    try:
        aw_ms_id = None
        if board in _BOARD_MS:
            aw_ms_id, aw_locked = _board_lock(conn, pid, board)
            if aw_locked:
                if ajax:
                    return jsonify(ok=False,
                                   error='Milestone is marked as done — '
                                         'the board is locked.'), 403
                flash('Milestone is marked as done — the board is locked.',
                      'error')
                return redirect(url_for('npd.view', pid=pid) + '#milestones')
        cur = conn.execute(
            "INSERT INTO `npd_comments` (project_id, user_id, comment, "
            "is_internal, board) VALUES (%s,%s,%s,%s,%s)",
            (pid, _uid(), comment or '(file)', 1 if is_internal else 0,
             board or None))
        cid = cur.lastrowid
        saved = []
        for fs in files:
            if not fs or not fs.filename:
                continue
            fname = secure_filename(fs.filename)
            stored = f"{pid}_{cid}_{_dt.now():%Y%m%d%H%M%S}_{fname}"
            path = os.path.join(_npd_upload_dir(), stored)
            fs.save(path)
            c2 = conn.execute(
                "INSERT INTO `npd_comment_files` (project_id, comment_id, "
                "file_name, file_path, file_size, created_by) "
                "VALUES (%s,%s,%s,%s,%s,%s)",
                (pid, cid, fname, stored, os.path.getsize(path), _uid()))
            saved.append({'id': c2.lastrowid, 'name': fname, 'path': stored})
        action = ('Artwork / Design comment added' if board == 'artwork'
                  else 'Artwork QC comment added' if board == 'artwork_qc'
                  else ('Internal comment' if is_internal else 'Comment')
                  + ' added')
        _log(conn, pid, action)
        if board in _BOARD_MS and aw_ms_id:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (aw_ms_id, action, _uid()))
        conn.commit()
        if ajax:
            return jsonify(ok=True, id=cid, user=_uname(), comment=comment,
                           created_at=_dt.now().strftime('%d-%m-%Y %I:%M %p'),
                           files=saved)
    finally:
        conn.close()
    return redirect(url_for('npd.view', pid=pid)
                    + ('#internal' if is_internal else '#discussion'))


@npd_bp.route('/comment/<int:cid>/delete', methods=['POST'])
@login_required
def comment_delete(cid):
    import os
    from flask import jsonify
    conn = _db()
    try:
        d = conn.execute("SELECT * FROM `npd_comments` WHERE id=%s",
                         (cid,)).fetchone()
        if not d:
            return jsonify(ok=False, error='Not found'), 404
        if d['user_id'] != _uid():
            return jsonify(ok=False,
                           error='You can only delete your own comments'), 403
        if (d.get('board') or '') in _BOARD_MS:
            _aw, _lk = _board_lock(conn, d['project_id'], d['board'])
            if _lk:
                return jsonify(ok=False,
                               error='Milestone is marked as done — '
                                     'the board is locked.'), 403
        for fr in (conn.execute("SELECT file_path FROM `npd_comment_files` "
                                "WHERE comment_id=%s", (cid,)).fetchall()
                   or []):
            try:
                os.remove(os.path.join(_npd_upload_dir(), fr['file_path']))
            except Exception:
                pass
        conn.execute("DELETE FROM `npd_comment_files` WHERE comment_id=%s",
                     (cid,))
        conn.execute("DELETE FROM `npd_comments` WHERE id=%s", (cid,))
        brd = d.get('board') or ''
        if brd == 'artwork':
            act_del = 'Artwork / Design comment removed'
        elif brd == 'artwork_qc':
            act_del = 'Artwork QC comment removed'
        elif d.get('is_internal'):
            act_del = 'Internal comment removed'
        else:
            act_del = 'Comment removed'
        _log(conn, d['project_id'], act_del)
        if brd in _BOARD_MS:
            _bm, _ = _board_lock(conn, d['project_id'], brd)
            if _bm:
                conn.execute(
                    "INSERT INTO `milestone_logs` (milestone_id, action, "
                    "created_by) VALUES (%s,%s,%s)",
                    (_bm, act_del, _uid()))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@npd_bp.route('/comment/<int:cid>/edit', methods=['POST'])
@login_required
def comment_edit(cid):
    """Owner-only edit — text + attachments add/remove (lead board jaisa)."""
    import os
    from datetime import datetime as _dt
    from flask import jsonify
    from werkzeug.utils import secure_filename
    conn = _db()
    try:
        d = conn.execute("SELECT * FROM `npd_comments` WHERE id=%s",
                         (cid,)).fetchone()
        if not d:
            return jsonify(ok=False, error='Not found'), 404
        if d['user_id'] != _uid():
            return jsonify(ok=False,
                           error='You can only edit your own comments.'), 403
        if (d.get('board') or '') in _BOARD_MS:
            _aw, _lk = _board_lock(conn, d['project_id'], d['board'])
            if _lk:
                return jsonify(ok=False,
                               error='Milestone is marked as done — '
                                     'the board is locked.'), 403
        comment = (request.form.get('comment') or '').strip()
        removed = [x for x in (request.form.get('removed') or '').split(',')
                   if x.strip().isdigit()]
        conn.execute("UPDATE `npd_comments` SET comment=%s, edited_at=NOW() "
                     "WHERE id=%s", (comment, cid))
        for aid in removed:
            ar = conn.execute(
                "SELECT file_path FROM `npd_comment_files` "
                "WHERE id=%s AND comment_id=%s", (aid, cid)).fetchone()
            if ar:
                try:
                    os.remove(os.path.join(_npd_upload_dir(),
                                           ar['file_path']))
                except Exception:
                    pass
                conn.execute("DELETE FROM `npd_comment_files` WHERE id=%s",
                             (aid,))
        for fs in request.files.getlist('files'):
            if not fs or not fs.filename:
                continue
            fname = secure_filename(fs.filename)
            stored = (f"{d['project_id']}_{cid}_"
                      f"{_dt.now():%Y%m%d%H%M%S}_{fname}")
            path = os.path.join(_npd_upload_dir(), stored)
            fs.save(path)
            conn.execute(
                "INSERT INTO `npd_comment_files` (project_id, comment_id, "
                "file_name, file_path, file_size, created_by) "
                "VALUES (%s,%s,%s,%s,%s,%s)",
                (d['project_id'], cid, fname, stored,
                 os.path.getsize(path), _uid()))
        brd = d.get('board') or ''
        if brd == 'artwork':
            act_ed = 'Artwork / Design comment edited'
        elif brd == 'artwork_qc':
            act_ed = 'Artwork QC comment edited'
        elif d.get('is_internal'):
            act_ed = 'Internal comment edited'
        else:
            act_ed = 'Comment edited'
        _log(conn, d['project_id'], act_ed)
        if brd in _BOARD_MS:
            _bm, _ = _board_lock(conn, d['project_id'], brd)
            if _bm:
                conn.execute(
                    "INSERT INTO `milestone_logs` (milestone_id, action, "
                    "created_by) VALUES (%s,%s,%s)",
                    (_bm, act_ed, _uid()))
        conn.commit()
        rem = [{'id': r['id'], 'name': r['file_name'], 'path': r['file_path']}
               for r in (conn.execute(
                   "SELECT * FROM `npd_comment_files` WHERE comment_id=%s "
                   "ORDER BY id", (cid,)).fetchall() or [])]
        return jsonify(ok=True, files=rem)
    finally:
        conn.close()


@npd_bp.route('/<int:pid>/note', methods=['POST'])
@login_required
def save_note(pid):
    content = request.form.get('content') or ''
    conn = _db()
    try:
        conn.execute(
            "INSERT INTO `npd_notes` (project_id, content, updated_by) "
            "VALUES (%s,%s,%s) ON DUPLICATE KEY UPDATE content=VALUES(content),"
            " updated_by=VALUES(updated_by)", (pid, content, _uid()))
        _log(conn, pid, 'Note updated')
        conn.commit()
        flash('Note saved.', 'success')
    finally:
        conn.close()
    return redirect(url_for('npd.view', pid=pid) + '#note')


@npd_bp.route('/<int:pid>/note/add', methods=['POST'])
@login_required
def personal_note_add(pid):
    """Lead-style personal note (sirf apne ko dikhta hai) — AJAX."""
    from datetime import datetime as _dt
    from flask import jsonify
    note = (request.form.get('note') or '').strip()
    if not note:
        return jsonify(ok=False, error='Note cannot be empty.'), 400
    conn = _db()
    try:
        cur = conn.execute(
            "INSERT INTO `npd_personal_notes` (project_id, user_id, note) "
            "VALUES (%s,%s,%s)", (pid, _uid(), note))
        nid = cur.lastrowid
        conn.commit()
        return jsonify(ok=True, id=nid, note=note,
                       created_at=_dt.now().strftime('%d-%m-%Y %I:%M %p'))
    finally:
        conn.close()


@npd_bp.route('/note/<int:nid>/delete', methods=['POST'])
@login_required
def personal_note_delete(nid):
    from flask import jsonify
    conn = _db()
    try:
        n = conn.execute("SELECT * FROM `npd_personal_notes` WHERE id=%s",
                         (nid,)).fetchone()
        if n and (n['user_id'] == _uid() or _is_admin()):
            conn.execute("DELETE FROM `npd_personal_notes` WHERE id=%s",
                         (nid,))
            conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@npd_bp.route('/<int:pid>/start', methods=['POST'])
@login_required
def start_project(pid):
    conn = _db()
    try:
        proj = conn.execute(
            "SELECT assigned_rd, assigned_sc, npd_poc, created_by, started_at "
            "FROM `npd_projects` WHERE id=%s AND is_deleted=0",
            (pid,)).fetchone()
        if not proj:
            flash('Project not found.', 'error')
            return redirect(url_for('npd.projects'))
        if not _is_rd_assignee(conn, pid, _uid()):
            flash('This project is not assigned to you — only the assigned '
                  'R&D person can start it.', 'error')
            return redirect(url_for('npd.view', pid=pid))
        if not proj['started_at']:
            conn.execute(
                "UPDATE `npd_projects` SET started_at=NOW(), "
                "status=CASE WHEN status='not_started' "
                "THEN 'sample_inprocess' ELSE status END WHERE id=%s", (pid,))
            _log(conn, pid, 'Project STARTED')
            # ── R&D assignment timer + log (safe if R&D tables absent) ──
            try:
                conn.execute(
                    "UPDATE `rd_sub_assignments` SET "
                    "started_at=COALESCE(started_at, NOW()), "
                    "status=CASE WHEN status='not_started' THEN 'in_progress' "
                    "ELSE status END "
                    "WHERE project_id=%s AND user_id=%s AND is_active=1",
                    (pid, _uid()))
                conn.execute(
                    "INSERT INTO `rd_project_logs` (project_id, user_id, event, "
                    "detail) VALUES (%s,%s,'started',%s)",
                    (pid, _uid(), f"Started by {_uname()}"))
            except Exception:
                pass
            conn.commit()
            flash('Project started.', 'success')
    finally:
        conn.close()
    return redirect(url_for('npd.view', pid=pid))


@npd_bp.route('/<int:pid>/finish', methods=['POST'])
@login_required
def finish_project(pid):
    conn = _db()
    try:
        if not _is_rd_assignee(conn, pid, _uid()):
            flash('Only the assigned R&D person can finish this project.', 'error')
            return redirect(url_for('npd.view', pid=pid))
        cur = conn.execute(
            "UPDATE `npd_projects` SET finished_at=NOW(), "
            "total_duration_seconds=TIMESTAMPDIFF(SECOND, started_at, NOW()), "
            "status=CASE WHEN status IN ('sample_inprocess','not_started','in_process') "
            "            THEN 'sample_ready' ELSE status END "
            "WHERE id=%s AND started_at IS NOT NULL "
            "AND finished_at IS NULL", (pid,))
        if cur.rowcount:
            _log(conn, pid, 'Project FINISHED — status: Sample Ready')
            # ── R&D assignment timer + log (safe if R&D tables absent) ──
            try:
                conn.execute(
                    "UPDATE `rd_sub_assignments` SET finished_at=NOW(), "
                    "status='finished', "
                    "total_seconds=TIMESTAMPDIFF(SECOND, COALESCE(started_at, NOW()), NOW()) "
                    "WHERE project_id=%s AND user_id=%s AND is_active=1",
                    (pid, _uid()))
                conn.execute(
                    "INSERT INTO `rd_project_logs` (project_id, user_id, event, "
                    "detail) VALUES (%s,%s,'finished',%s)",
                    (pid, _uid(), f"Finished by {_uname()} — Sample Ready"))
            except Exception:
                pass
            conn.commit()
            flash('Sample marked Ready.', 'success')
    finally:
        conn.close()
    return redirect(url_for('npd.view', pid=pid))


# ─────────────────────────────────────────────────────────────────────────────
# STATUS UPDATES
# ─────────────────────────────────────────────────────────────────────────────
@npd_bp.route('/<int:pid>/status', methods=['POST'])
@login_required
def set_status(pid):
    st = request.form.get('status')
    conn = _db()
    try:
        if st not in [s['slug'] for s in get_npd_statuses(conn)]:
            flash('Invalid status.', 'error')
            return redirect(url_for('npd.view', pid=pid))
        vis_frag, vis_params = _proj_visibility_sql('npd_projects')
        cur = conn.execute(
            "UPDATE `npd_projects` SET status=%s, updated_by=%s, "
            "completed_at=CASE WHEN %s='finish' THEN NOW() "
            "             ELSE completed_at END, "
            "cancelled_at=CASE WHEN %s='cancelled' THEN NOW() "
            "             ELSE cancelled_at END "
            "WHERE id=%s AND is_deleted=0" + vis_frag,
            [st, _uid(), st, st, pid] + vis_params)
        if cur.rowcount:
            _log(conn, pid, f"Status changed to {st}")
            conn.commit()
            flash('Status updated.', 'success')
        else:
            flash('You do not have access.', 'error')
    finally:
        conn.close()
    return redirect(url_for('npd.view', pid=pid))


@npd_bp.route('/<int:pid>/bom/save', methods=['POST'])
@login_required
def bom_save(pid):
    """BOM formulation sheet save (naya ya edit) — AJAX JSON."""
    from flask import jsonify
    data = request.get_json(silent=True) or {}
    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403
        bid = data.get('id')
        pn = (data.get('product_name') or '').strip()
        code = (data.get('code') or '').strip()
        var = (data.get('variant') or '').strip()
        rows = [r for r in (data.get('rows') or [])
                if (r.get('inci') or '').strip() or r.get('qty')]
        if not pn and not rows:
            return jsonify(ok=False,
                           error='Product name or at least one row is '
                                 'required.'), 400
        total = 0.0
        for r in rows:
            try:
                total += float(r.get('qty') or 0)
            except Exception:
                pass
        if abs(round(total, 3) - 100.0) > 0.0005:
            return jsonify(ok=False,
                           error='BOM total must be exactly 100%% — '
                                 'current total is %.3f%%.'
                                 % round(total, 3)), 400
        ms = conn.execute(
            "SELECT id, status FROM `milestone_masters` WHERE project_id=%s "
            "AND (milestone_type='bom' OR title='BOM') LIMIT 1",
            (pid,)).fetchone()
        ms_id = (ms or {}).get('id')
        ms_status = (ms or {}).get('status') if ms else None
        if ms_status == 'approved':
            return jsonify(ok=False,
                           error='Milestone is marked as done — '
                                 'BOM is locked.'), 403
        if bid:
            b = conn.execute(
                "SELECT id FROM `npd_boms` WHERE id=%s AND project_id=%s",
                (bid, pid)).fetchone()
            if not b:
                return jsonify(ok=False, error='BOM not found'), 404
            conn.execute(
                "UPDATE `npd_boms` SET product_name=%s, code=%s, variant=%s "
                "WHERE id=%s", (pn, code, var, bid))
            conn.execute("DELETE FROM `npd_bom_rows` WHERE bom_id=%s",
                         (bid,))
            action = f"BOM '{pn or bid}' updated"
        else:
            cur = conn.execute(
                "INSERT INTO `npd_boms` (project_id, milestone_id, "
                "product_name, code, variant, created_by) "
                "VALUES (%s,%s,%s,%s,%s,%s)",
                (pid, ms_id, pn, code, var, _uid()))
            bid = cur.lastrowid
            action = f"BOM '{pn or bid}' created"
        for i, r in enumerate(rows, start=1):
            try:
                qty = round(float(r.get('qty') or 0), 3)
            except Exception:
                qty = 0
            conn.execute(
                "INSERT INTO `npd_bom_rows` (bom_id, sr_no, inci_name, "
                "qty_pct) VALUES (%s,%s,%s,%s)",
                (bid, i, (r.get('inci') or '').strip(), qty))
        _log(conn, pid, action)
        if ms_id:  # BOM milestone ke apne log me bhi
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms_id, action, _uid()))
        conn.commit()
        out_rows = [{'inci': (r.get('inci') or '').strip(),
                     'qty': round(float(r.get('qty') or 0), 3)
                     if str(r.get('qty') or '').strip() else 0}
                    for r in rows]
        from datetime import datetime as _dt
        return jsonify(ok=True, bom={'id': bid, 'product_name': pn,
                                     'code': code, 'variant': var,
                                     'by': _uname(),
                                     'created_at':
                                     _dt.now().strftime('%Y-%m-%d %H:%M'),
                                     'rows': out_rows})
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify(ok=False, error=str(e)), 500
    finally:
        conn.close()


@npd_bp.route('/bom/<int:bid>/delete', methods=['POST'])
@login_required
def bom_delete(bid):
    from flask import jsonify
    conn = _db()
    try:
        b = conn.execute("SELECT * FROM `npd_boms` WHERE id=%s",
                         (bid,)).fetchone()
        if not b:
            return jsonify(ok=False, error='Not found'), 404
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p WHERE p.id=%s" + vis_frag,
            [b['project_id']] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403
        ms_st = conn.execute(
            "SELECT status FROM `milestone_masters` WHERE project_id=%s "
            "AND (milestone_type='bom' OR title='BOM') LIMIT 1",
            (b['project_id'],)).fetchone()
        if ms_st and ms_st.get('status') == 'approved':
            return jsonify(ok=False,
                           error='Milestone is marked as done — '
                                 'BOM is locked.'), 403
        conn.execute("DELETE FROM `npd_bom_rows` WHERE bom_id=%s", (bid,))
        conn.execute("DELETE FROM `npd_boms` WHERE id=%s", (bid,))
        action = f"BOM '{b['product_name'] or bid}' deleted"
        _log(conn, b['project_id'], action)
        ms_id = b.get('milestone_id') or ((conn.execute(
            "SELECT id FROM `milestone_masters` WHERE project_id=%s "
            "AND (milestone_type='bom' OR title='BOM') LIMIT 1",
            (b['project_id'],)).fetchone() or {}).get('id'))
        if ms_id:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms_id, action, _uid()))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@npd_bp.route('/<int:pid>/quotation', methods=['POST'])
@login_required
def npd_create_quotation(pid):
    """Lead wale Create Quotation jaisa — NPD project se. PDF return karta hai."""
    import json as _json
    from datetime import datetime as _dt
    from flask import send_file
    from crm.crm_leads_routes import (_ensure_quotations, _next_quot_number,
                                      _build_quot_pdf, log_activity)  # noqa
    conn = _db()
    try:
        _ensure_quotations(conn)
        vis_frag, vis_params = _proj_visibility_sql('p')
        proj = conn.execute(
            "SELECT p.* FROM `npd_projects` p "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not proj:
            flash('Project not found or you do not have access.', 'error')
            return redirect(url_for('npd.projects'))
        qms = conn.execute(
            "SELECT id, status FROM `milestone_masters` WHERE project_id=%s "
            "AND (milestone_type='quotation' OR title='Quotation') LIMIT 1",
            (pid,)).fetchone()
        if qms and qms['status'] == 'approved':
            flash('Milestone is marked as done — quotations are locked.',
                  'error')
            return redirect(url_for('npd.view', pid=pid) + '#milestones')

        quot_number = ((request.form.get('quot_number') or '').strip()
                       or _next_quot_number(conn))
        quot_date = (request.form.get('quot_date')
                     or _dt.now().strftime('%Y-%m-%d'))
        valid_until = request.form.get('valid_until') or ''
        subject = request.form.get('quot_subject', '')
        bill_company = (request.form.get('bill_company')
                        or proj.get('client_company') or '').strip()
        bill_address = (request.form.get('bill_address') or '').strip()
        bill_phone = (request.form.get('bill_phone')
                      or proj.get('client_phone') or '').strip()
        bill_email = (request.form.get('bill_email')
                      or proj.get('client_email') or '').strip()
        bill_gst = (request.form.get('bill_gst') or '').strip()
        try:
            gst_pct = float(request.form.get('quot_gst_pct') or 18)
        except Exception:
            gst_pct = 18.0
        terms = request.form.get('terms', '')
        notes = request.form.get('notes', '')

        names = request.form.getlist('item_name[]')
        sizes = request.form.getlist('item_size[]')
        codes = request.form.getlist('item_code[]')
        uoms = request.form.getlist('item_uom[]')
        costs = request.form.getlist('item_cost[]')
        moqs = request.form.getlist('item_moq[]')
        pm_specs = request.form.getlist('item_pm_spec[]')
        pm_costs = request.form.getlist('item_pm_cost[]')
        cats = request.form.getlist('item_category[]')
        finals = request.form.getlist('item_final_cost[]')

        def _f(lst, i):
            try:
                return float(lst[i]) if i < len(lst) and lst[i] else 0.0
            except Exception:
                return 0.0

        def _s(lst, i):
            return lst[i] if i < len(lst) else ''

        if not bill_company:
            flash('Company Name is required.', 'error')
            return redirect(url_for('npd.view', pid=pid) + '#milestones')

        items, sub_total = [], 0.0
        for i, nm in enumerate(names):
            if not nm.strip():
                continue
            moq = _f(moqs, i)
            fc = _f(finals, i)
            amt = moq * fc
            sub_total += amt
            items.append({'name': nm.strip(), 'size': _s(sizes, i),
                          'uom': _s(uoms, i), 'code': _s(codes, i),
                          'moq': moq, 'final_cost': fc, 'amount': amt,
                          'cost': _f(costs, i), 'pm_spec': _s(pm_specs, i),
                          'pm_cost': _f(pm_costs, i),
                          'category': _s(cats, i)})
        if not items:
            flash('Please add at least one product.', 'error')
            return redirect(url_for('npd.view', pid=pid) + '#milestones')

        gst_amount = sub_total * gst_pct / 100.0
        total_amount = sub_total + gst_amount
        try:
            qd = _dt.strptime(quot_date, '%Y-%m-%d').strftime('%Y-%m-%d')
        except Exception:
            qd = _dt.now().strftime('%Y-%m-%d')
        vu = None
        if valid_until:
            try:
                vu = _dt.strptime(valid_until,
                                  '%Y-%m-%d').strftime('%Y-%m-%d')
            except Exception:
                vu = None
        edit_id = (request.form.get('quotation_id') or '').strip()
        try:
            if edit_id.isdigit():
                exists = conn.execute(
                    "SELECT id, quot_number FROM `quotations` "
                    "WHERE id=%s AND project_id=%s AND is_deleted=0",
                    (int(edit_id), pid)).fetchone()
                if not exists:
                    flash('Quotation not found.', 'error')
                    return redirect(url_for('npd.view', pid=pid)
                                    + '#milestones')
                quot_number = exists['quot_number']
                conn.execute(
                    "UPDATE `quotations` SET quot_date=%s, valid_until=%s, "
                    "subject=%s, bill_company=%s, bill_address=%s, "
                    "bill_phone=%s, bill_email=%s, bill_gst=%s, gst_pct=%s, "
                    "sub_total=%s, gst_amount=%s, total_amount=%s, "
                    "items_json=%s, terms=%s, notes=%s WHERE id=%s",
                    (qd, vu, subject, bill_company, bill_address, bill_phone,
                     bill_email, bill_gst, gst_pct, sub_total, gst_amount,
                     total_amount, _json.dumps(items), terms, notes,
                     int(edit_id)))
                action = f'Quotation updated: {quot_number}'
                _log(conn, pid, action)
                if qms:
                    conn.execute(
                        "INSERT INTO `milestone_logs` (milestone_id, action, "
                        "created_by) VALUES (%s,%s,%s)",
                        (qms['id'], action, _uid()))
                conn.commit()
                pdf = _build_quot_pdf({
                    'quot_number': quot_number,
                    'quot_date': _dt.strptime(qd, '%Y-%m-%d')
                    .strftime('%d-%m-%Y'),
                    'valid_until': (_dt.strptime(vu, '%Y-%m-%d')
                                    .strftime('%d-%m-%Y') if vu else '-'),
                    'subject': subject, 'bill_company': bill_company,
                    'bill_address': bill_address, 'bill_phone': bill_phone,
                    'bill_email': bill_email, 'bill_gst': bill_gst,
                    'gst_pct': gst_pct, 'items': items,
                    'sub_total': sub_total, 'gst_amount': gst_amount,
                    'total_amount': total_amount, 'terms': terms,
                    'notes': notes, 'by_name': _uname()})
                return send_file(pdf, mimetype='application/pdf',
                                 as_attachment=False,
                                 download_name='%s.pdf'
                                 % quot_number.replace('/', '_'))
            conn.execute(
                "INSERT INTO `quotations` (quot_number, lead_id, project_id, "
                "quot_date, valid_until, subject, bill_company, bill_address,"
                " bill_phone, bill_email, bill_gst, gst_pct, sub_total, "
                "gst_amount, total_amount, items_json, terms, notes, status, "
                "created_by, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,"
                "%s,%s,%s,%s,%s,%s,%s,%s,%s,'draft',%s,NOW())",
                (quot_number, proj.get('lead_id'), pid, qd, vu, subject,
                 bill_company, bill_address, bill_phone, bill_email, bill_gst,
                 gst_pct, sub_total, gst_amount, total_amount,
                 _json.dumps(items), terms, notes, _uid()))
            action = f'Quotation generated: {quot_number}'
            _log(conn, pid, action)
            if qms:
                conn.execute(
                    "INSERT INTO `milestone_logs` (milestone_id, action, "
                    "created_by) VALUES (%s,%s,%s)",
                    (qms['id'], action, _uid()))
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
            flash('Quotation number "%s" already exists, please retry.'
                  % quot_number, 'error')
            return redirect(url_for('npd.view', pid=pid) + '#milestones')

        pdf = _build_quot_pdf({
            'quot_number': quot_number,
            'quot_date': _dt.strptime(qd, '%Y-%m-%d').strftime('%d-%m-%Y'),
            'valid_until': (_dt.strptime(vu, '%Y-%m-%d').strftime('%d-%m-%Y')
                            if vu else '-'),
            'subject': subject, 'bill_company': bill_company,
            'bill_address': bill_address, 'bill_phone': bill_phone,
            'bill_email': bill_email, 'bill_gst': bill_gst,
            'gst_pct': gst_pct, 'items': items, 'sub_total': sub_total,
            'gst_amount': gst_amount, 'total_amount': total_amount,
            'terms': terms, 'notes': notes, 'by_name': _uname()})
        return send_file(pdf, mimetype='application/pdf',
                         as_attachment=False,
                         download_name='%s.pdf'
                         % quot_number.replace('/', '_'))
    finally:
        conn.close()


_BOARD_MS = {
    'artwork': ('artwork', 'Artwork / Design'),
    'artwork_qc': ('artwork_qc', 'Artwork QC Approval'),
}


def _board_lock(conn, pid, board):
    """Board wale milestone (artwork / artwork_qc) ka (id, locked)."""
    mt, title = _BOARD_MS.get(board, (None, None))
    if not mt:
        return (None, False)
    ms = conn.execute(
        "SELECT id, status FROM `milestone_masters` WHERE project_id=%s "
        "AND (milestone_type=%s OR title=%s) LIMIT 1",
        (pid, mt, title)).fetchone()
    return ((ms or {}).get('id'),
            bool(ms and ms.get('status') == 'approved'))


def _artwork_lock(conn, pid):
    return _board_lock(conn, pid, 'artwork')


def _packing_lock(conn, pid):
    """Packing milestone done ho to (id, locked) return karo."""
    ms = conn.execute(
        "SELECT id, status FROM `milestone_masters` WHERE project_id=%s "
        "AND (milestone_type='packing_material' OR title='Packing Material') "
        "LIMIT 1", (pid,)).fetchone()
    return ((ms or {}).get('id'),
            bool(ms and ms.get('status') == 'approved'))


@npd_bp.route('/<int:pid>/inline-upload', methods=['POST'])
@login_required
def board_inline_upload(pid):
    """Gmail-style inline image upload — artwork / artwork_qc boards only."""
    import os
    from datetime import datetime as _dt
    from flask import jsonify
    from werkzeug.utils import secure_filename
    board = (request.form.get('board') or '').strip().lower()
    if board not in ('artwork', 'artwork_qc'):
        return jsonify(ok=False, error='Inline upload not allowed.'), 400
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify(ok=False, error='File is required.'), 400
    if (f.mimetype or '').split('/')[0] != 'image':
        return jsonify(ok=False, error='Only images allowed inline.'), 400
    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403
        _ms_id, locked = _board_lock(conn, pid, board)
        if locked:
            return jsonify(ok=False,
                           error='Milestone is marked as done — '
                                 'the board is locked.'), 403
        fname = secure_filename(f.filename) or 'image'
        stored = f"{pid}_inline_{_dt.now():%Y%m%d%H%M%S}_{fname}"
        f.save(os.path.join(_npd_upload_dir(), stored))
        url = '/static/uploads/npd/' + stored
        ms_id, _lk = _board_lock(conn, pid, board)
        label = ('Artwork / Design' if board == 'artwork'
                 else 'Artwork QC')
        action = f'{label} inline image uploaded'
        _log(conn, pid, action)
        if ms_id:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms_id, action, _uid()))
        conn.commit()
        return jsonify(ok=True, url=url, path=stored)
    finally:
        conn.close()


@npd_bp.route('/<int:pid>/packing/save', methods=['POST'])
@login_required
def packing_save(pid):
    """Packing Material row add/update (AJAX, auto-save)."""
    from flask import jsonify
    data = request.get_json(silent=True) or {}
    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403
        ms_id, locked = _packing_lock(conn, pid)
        if locked:
            return jsonify(ok=False,
                           error='Milestone is marked as done — '
                                 'packing material is locked.'), 403
        cat = (data.get('category') or 'Primary').strip()[:50]
        vendor = (data.get('vendor_name') or '').strip()[:200]
        st = (data.get('filling_status') or 'pending').strip().lower()
        if st not in ('pending', 'in_process', 'done'):
            st = 'pending'
        cost = None
        try:
            if str(data.get('cost') or '').strip() != '':
                cost = float(data.get('cost'))
        except Exception:
            cost = None
        rid = data.get('id')
        if rid:
            r = conn.execute(
                "SELECT id FROM `npd_packing_rows` "
                "WHERE id=%s AND project_id=%s", (rid, pid)).fetchone()
            if not r:
                return jsonify(ok=False, error='Row not found'), 404
            conn.execute(
                "UPDATE `npd_packing_rows` SET category=%s, vendor_name=%s, "
                "cost=%s, filling_status=%s WHERE id=%s",
                (cat, vendor, cost, st, rid))
            conn.commit()
            return jsonify(ok=True, id=int(rid))
        cur = conn.execute(
            "INSERT INTO `npd_packing_rows` (project_id, milestone_id, "
            "category, vendor_name, cost, filling_status, created_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (pid, ms_id, cat, vendor, cost, st, _uid()))
        rid = cur.lastrowid
        action = 'Packing material row added'
        _log(conn, pid, action)
        if ms_id:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms_id, action, _uid()))
        conn.commit()
        return jsonify(ok=True, id=rid)
    finally:
        conn.close()


@npd_bp.route('/packing/<int:rid>/upload', methods=['POST'])
@login_required
def packing_upload(rid):
    """Packing row file upload (image / filling / coa)."""
    from flask import jsonify
    from werkzeug.utils import secure_filename
    import os
    field = request.form.get('field')
    col = {'image': 'image_path', 'filling': 'filling_image_path',
           'coa': 'coa_path'}.get(field)
    f = request.files.get('file')
    if not col or not f or not f.filename:
        return jsonify(ok=False, error='File is required.'), 400
    conn = _db()
    try:
        r = conn.execute("SELECT * FROM `npd_packing_rows` WHERE id=%s",
                         (rid,)).fetchone()
        if not r:
            return jsonify(ok=False, error='Row not found'), 404
        pid = r['project_id']
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p WHERE p.id=%s" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403
        ms_id, locked = _packing_lock(conn, pid)
        if locked:
            return jsonify(ok=False,
                           error='Milestone is marked as done — '
                                 'packing material is locked.'), 403
        sub = os.path.join('packing', str(pid))
        folder = os.path.join(_npd_upload_dir(), sub)
        os.makedirs(folder, exist_ok=True)
        fn = ('%d_%s_%s' % (rid, field,
                            secure_filename(f.filename) or 'file'))[:120]
        f.save(os.path.join(folder, fn))
        rel = (sub + '/' + fn).replace('\\', '/')
        conn.execute("UPDATE `npd_packing_rows` SET " + col + "=%s "
                     "WHERE id=%s", (rel, rid))
        labels = {'image': 'image', 'filling': 'filling image',
                  'coa': 'COA'}
        action = f"Packing material {labels[field]} uploaded"
        _log(conn, pid, action)
        if ms_id:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms_id, action, _uid()))
        conn.commit()
        return jsonify(ok=True, path=rel)
    finally:
        conn.close()


@npd_bp.route('/packing/<int:rid>/delete', methods=['POST'])
@login_required
def packing_delete(rid):
    from flask import jsonify
    conn = _db()
    try:
        r = conn.execute("SELECT * FROM `npd_packing_rows` WHERE id=%s",
                         (rid,)).fetchone()
        if not r:
            return jsonify(ok=False, error='Row not found'), 404
        pid = r['project_id']
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p WHERE p.id=%s" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403
        ms_id, locked = _packing_lock(conn, pid)
        if locked:
            return jsonify(ok=False,
                           error='Milestone is marked as done — '
                                 'packing material is locked.'), 403
        import os
        for col in ('image_path', 'filling_image_path', 'coa_path'):
            if r.get(col):
                try:
                    os.remove(os.path.join(_npd_upload_dir(), r[col]))
                except Exception:
                    pass
        conn.execute("DELETE FROM `npd_packing_rows` WHERE id=%s", (rid,))
        action = 'Packing material row deleted'
        _log(conn, pid, action)
        if ms_id:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms_id, action, _uid()))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@npd_bp.route('/quotation/<int:qid>/delete', methods=['POST'])
@login_required
def npd_quotation_delete(qid):
    """Quotation soft-delete (NPD panel se) — milestone done ho to locked."""
    from flask import jsonify
    conn = _db()
    try:
        q = conn.execute(
            "SELECT * FROM `quotations` WHERE id=%s AND is_deleted=0",
            (qid,)).fetchone()
        if not q or not q.get('project_id'):
            return jsonify(ok=False, error='Quotation not found.'), 404
        pid = q['project_id']
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p WHERE p.id=%s" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403
        qms = conn.execute(
            "SELECT id, status FROM `milestone_masters` WHERE project_id=%s "
            "AND (milestone_type='quotation' OR title='Quotation') LIMIT 1",
            (pid,)).fetchone()
        if qms and qms['status'] == 'approved':
            return jsonify(ok=False,
                           error='Milestone is marked as done — '
                                 'quotations are locked.'), 403
        conn.execute(
            "UPDATE `quotations` SET is_deleted=1, deleted_at=NOW() "
            "WHERE id=%s", (qid,))
        action = f"Quotation deleted: {q['quot_number']}"
        _log(conn, pid, action)
        if qms:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (qms['id'], action, _uid()))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@npd_bp.route('/<int:pid>/milestone/<int:mid>/sheet', methods=['POST'])
@login_required
def milestone_sheet(pid, mid):
    """Milestone sheet content (rich HTML) save — AJAX.
    Ingredients List & Marketing Sheet isi se save hoti hai (notes col)."""
    from flask import jsonify
    from datetime import datetime as _dt
    data = request.get_json(silent=True) or {}
    content = (data.get('content') or '').strip()
    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403
        m = conn.execute(
            "SELECT id, title, status FROM `milestone_masters` "
            "WHERE id=%s AND project_id=%s", (mid, pid)).fetchone()
        if not m:
            return jsonify(ok=False, error='Milestone not found.'), 404
        if m['status'] == 'approved':
            return jsonify(ok=False,
                           error='Milestone is marked as done — '
                                 'sheet is locked.'), 403
        conn.execute("UPDATE `milestone_masters` SET notes=%s WHERE id=%s",
                     (content, mid))
        action = f"'{m['title']}' sheet updated"
        conn.execute(
            "INSERT INTO `milestone_logs` (milestone_id, action, "
            "created_by) VALUES (%s,%s,%s)", (mid, action, _uid()))
        _log(conn, pid, action)
        conn.commit()
        return jsonify(ok=True, action=action, user=_uname(),
                       at=_dt.now().strftime('%Y-%m-%d %H:%M'))
    finally:
        conn.close()


@npd_bp.route('/<int:pid>/milestone/<int:mid>', methods=['POST'])
@login_required
def set_milestone(pid, mid):
    from flask import jsonify
    ajax = request.is_json or request.headers.get('X-Requested-With')
    st = request.form.get('status')
    if st not in [s[0] for s in MS_STATUSES]:
        if ajax:
            return jsonify(ok=False, error='Invalid milestone status.'), 400
        flash('Invalid milestone status.', 'error')
        return redirect(url_for('npd.view', pid=pid))
    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            if ajax:
                return jsonify(ok=False,
                               error='You do not have access.'), 403
            flash('You do not have access.', 'error')
            return redirect(url_for('npd.projects'))

        row = conn.execute(
            "SELECT title, status, milestone_type FROM `milestone_masters` "
            "WHERE id=%s AND project_id=%s", (mid, pid)).fetchone()
        if row:
            # barcode milestone: har design ka barcode upload hona zaruri hai
            if row.get('milestone_type') == 'barcode' and st == 'approved':
                tot = conn.execute(
                    "SELECT COUNT(*) AS n FROM `npd_barcode_designs` "
                    "WHERE project_id=%s", (pid,)).fetchone()
                pend = conn.execute(
                    "SELECT COUNT(*) AS n FROM `npd_barcode_designs` "
                    "WHERE project_id=%s AND (barcode_path IS NULL OR barcode_path='')",
                    (pid,)).fetchone()
                total = (tot or {}).get('n', 0)
                pending = (pend or {}).get('n', 0)
                if total == 0:
                    msg = ('Upload at least one design and its barcode '
                           'before marking this milestone as done.')
                    if ajax:
                        return jsonify(ok=False, error=msg), 400
                    flash(msg, 'error')
                    return redirect(url_for('npd.view', pid=pid))
                if pending > 0:
                    msg = (f'{pending} design(s) still need a barcode upload. '
                           f'Upload all barcodes before marking this milestone as done.')
                    if ajax:
                        return jsonify(ok=False, error=msg), 400
                    flash(msg, 'error')
                    return redirect(url_for('npd.view', pid=pid))
            conn.execute(
                "UPDATE `milestone_masters` SET status=%s, "
                "completed_at=CASE WHEN %s='approved' THEN NOW() "
                "             ELSE completed_at END, "
                "approved_by=CASE WHEN %s='approved' THEN %s "
                "             ELSE approved_by END, "
                "approved_at=CASE WHEN %s='approved' THEN NOW() "
                "             ELSE approved_at END "
                "WHERE id=%s AND project_id=%s",
                (st, st, st, _uid(), st, mid, pid))
            # old system jaisa milestone_logs entry
            conn.execute(
                "INSERT INTO `milestone_logs` "
                "(milestone_id, action, old_status, new_status, created_by) "
                "VALUES (%s,%s,%s,%s,%s)",
                (mid, f"Status: {row['status']} -> {st}",
                 row['status'], st, _uid()))
            _log(conn, pid, f"Milestone '{row['title']}' -> {st}")
            conn.commit()
            if ajax:
                from datetime import datetime as _dt
                return jsonify(
                    ok=True, status=st,
                    action=f"Status: {row['status']} -> {st}",
                    user=_uname(),
                    at=_dt.now().strftime('%Y-%m-%d %H:%M'))
            flash('Milestone updated.', 'success')
        elif ajax:
            return jsonify(ok=False, error='Milestone not found.'), 404
    finally:
        conn.close()
    return redirect(url_for('npd.view', pid=pid))


# ─────────────────────────────────────────────────────────────────────────────
# FDA MILESTONE  (Step 1: send request mail, Step 2: FDA entries + docs)
# ─────────────────────────────────────────────────────────────────────────────
def _fda_milestone(conn, pid):
    return conn.execute(
        "SELECT * FROM `milestone_masters` WHERE project_id=%s "
        "AND milestone_type='fda' LIMIT 1", (pid,)).fetchone()


@npd_bp.route('/<int:pid>/fda/send-mail', methods=['POST'])
@login_required
def fda_send_mail(pid):
    """Step 1 — send the FDA request mail (To/CC = selected employees)."""
    from flask import jsonify
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.utils import formataddr
    from datetime import datetime as _dt

    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403

        ms = _fda_milestone(conn, pid)

        to_ids = request.form.getlist('to_ids') or request.form.getlist('to_ids[]')
        cc_ids = request.form.getlist('cc_ids') or request.form.getlist('cc_ids[]')
        subject = (request.form.get('subject') or '').strip()
        body = request.form.get('body') or ''

        if not to_ids:
            return jsonify(ok=False, error='Please select at least one recipient.'), 400
        if not subject:
            return jsonify(ok=False, error='Subject is required.'), 400
        if not body.strip():
            return jsonify(ok=False, error='Message body is required.'), 400

        def _emails_for(ids):
            if not ids:
                return []
            placeholders = ','.join(['%s'] * len(ids))
            rows = conn.execute(
                "SELECT id, COALESCE(NULLIF(full_name,''), username) AS nm, "
                "email FROM `User_Tbl` WHERE id IN (" + placeholders + ")",
                ids).fetchall() or []
            return [r for r in rows if (r.get('email') or '').strip()]

        to_rows = _emails_for(to_ids)
        cc_rows = _emails_for(cc_ids)
        if not to_rows:
            return jsonify(ok=False, error='Selected recipients have no email address on file.'), 400

        to_emails = [r['email'].strip() for r in to_rows]
        cc_emails = [r['email'].strip() for r in cc_rows]
        to_names = ', '.join(r['nm'] for r in to_rows)

        # ── send via shared SMTP config ──
        try:
            import core.config as _cfg
        except Exception:
            _cfg = None
        server = getattr(_cfg, 'MAIL_SERVER', 'smtp.gmail.com')
        port = int(getattr(_cfg, 'MAIL_PORT', 587) or 587)
        use_tls = getattr(_cfg, 'MAIL_USE_TLS', True)
        username = getattr(_cfg, 'MAIL_USERNAME', 'no-reply@hcpwellness.in')
        password = getattr(_cfg, 'MAIL_PASSWORD', '')
        sender = username or 'no-reply@hcpwellness.in'

        if not password:
            return jsonify(ok=False, error='SMTP password not configured on server.'), 500

        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = subject
            msg['From'] = formataddr(('HCP Wellness Pvt. Ltd.', sender))
            msg['To'] = ', '.join(to_emails)
            if cc_emails:
                msg['Cc'] = ', '.join(cc_emails)
            msg['Reply-To'] = sender
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            html_body = body.replace('\n', '<br>')
            msg.attach(MIMEText(html_body, 'html', 'utf-8'))

            s = smtplib.SMTP(server, port, timeout=25)
            if use_tls:
                s.starttls()
            s.login(username, password)
            s.sendmail(sender, to_emails + cc_emails, msg.as_string())
            s.quit()
        except Exception as e:
            return jsonify(ok=False, error='Email bhejne me dikkat: ' + str(e)), 500

        conn.execute(
            "INSERT INTO `npd_fda_requests` "
            "(project_id, milestone_id, to_emails, cc_emails, subject, body, sent_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (pid, ms['id'] if ms else None, ','.join(to_emails),
             ','.join(cc_emails), subject, body, _uid()))
        action = f"FDA request mail sent to {to_names}"
        _log(conn, pid, action)
        if ms:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms['id'], action, _uid()))
        conn.commit()

        return jsonify(ok=True, sent_by=_uname(),
                       sent_at=_dt.now().strftime('%d-%m-%Y %H:%M'),
                       to=', '.join(to_emails), cc=', '.join(cc_emails),
                       subject=subject, body=body)
    finally:
        conn.close()


@npd_bp.route('/<int:pid>/fda/entry/add', methods=['POST'])
@login_required
def fda_entry_add(pid):
    """Step 2 — add a new FDA entry (product + 4 document uploads)."""
    from flask import jsonify
    from werkzeug.utils import secure_filename
    import os

    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403

        product_name = (request.form.get('product_name') or '').strip()
        if not product_name:
            return jsonify(ok=False, error='Product Name is required.'), 400

        ms = _fda_milestone(conn, pid)

        field_map = {
            'free_sale_certificate': 'free_sale_certificate',
            'product_permission': 'product_permission',
            'msds': 'msds',
            'dossier': 'dossier',
        }
        sub = os.path.join('fda', str(pid))
        folder = os.path.join(_npd_upload_dir(), sub)
        os.makedirs(folder, exist_ok=True)

        saved = {}
        for field, col in field_map.items():
            f = request.files.get(field)
            if f and f.filename:
                fn = ('%s_%s_%s' % (field, _dt_stamp(), secure_filename(f.filename) or 'file'))[:150]
                f.save(os.path.join(folder, fn))
                saved[col] = (sub + '/' + fn).replace('\\', '/')

        cols = ['project_id', 'milestone_id', 'product_name'] + list(saved.keys()) + ['created_by']
        vals = [pid, ms['id'] if ms else None, product_name] + list(saved.values()) + [_uid()]
        placeholders = ','.join(['%s'] * len(vals))
        cur = conn.execute(
            "INSERT INTO `npd_fda_entries` (`" + "`,`".join(cols) + "`) "
            "VALUES (" + placeholders + ")", vals)
        eid = cur.lastrowid

        action = f"FDA entry added for '{product_name}'"
        _log(conn, pid, action)
        if ms:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms['id'], action, _uid()))
        conn.commit()

        row = conn.execute(
            "SELECT * FROM `npd_fda_entries` WHERE id=%s", (eid,)).fetchone()
        return jsonify(ok=True, entry={
            'id': row['id'], 'product_name': row['product_name'],
            'free_sale_certificate': row.get('free_sale_certificate'),
            'product_permission': row.get('product_permission'),
            'msds': row.get('msds'), 'dossier': row.get('dossier')})
    finally:
        conn.close()


def _dt_stamp():
    from datetime import datetime as _dt
    return _dt.now().strftime('%Y%m%d%H%M%S')


@npd_bp.route('/<int:pid>/fda/entry/<int:eid>/delete', methods=['POST'])
@login_required
def fda_entry_delete(pid, eid):
    from flask import jsonify
    import os

    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403

        row = conn.execute(
            "SELECT * FROM `npd_fda_entries` WHERE id=%s AND project_id=%s",
            (eid, pid)).fetchone()
        if not row:
            return jsonify(ok=False, error='Entry not found.'), 404

        for col in ('free_sale_certificate', 'product_permission', 'msds', 'dossier'):
            if row.get(col):
                try:
                    os.remove(os.path.join(_npd_upload_dir(), row[col]))
                except Exception:
                    pass

        conn.execute("DELETE FROM `npd_fda_entries` WHERE id=%s", (eid,))
        action = f"FDA entry deleted for '{row['product_name']}'"
        _log(conn, pid, action)
        ms = _fda_milestone(conn, pid)
        if ms:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms['id'], action, _uid()))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# BARCODE MILESTONE  (Designer upload 1024x1024 + Barcode Team upload)
# ─────────────────────────────────────────────────────────────────────────────
def _barcode_milestone(conn, pid):
    return conn.execute(
        "SELECT * FROM `milestone_masters` WHERE project_id=%s "
        "AND milestone_type='barcode' LIMIT 1", (pid,)).fetchone()


def _bc_design_json(conn, row):
    """Build the JSON payload for one barcode design row (used by JS re-render)."""
    umap = _user_map(conn)
    return {
        'id': row['id'], 'sr_no': row['sr_no'],
        'design_path': row.get('design_path'),
        'design_width': row.get('design_width'),
        'design_height': row.get('design_height'),
        'barcode_path': row.get('barcode_path'),
        'by': umap.get(row.get('created_by'), ''),
        'created_at': str(row.get('created_at') or '')[:16],
    }


@npd_bp.route('/<int:pid>/barcode/upload-design', methods=['POST'])
@login_required
def barcode_upload_design(pid):
    """Designer uploads a new barcode design (must be exactly 1024x1024 px)."""
    from flask import jsonify
    from werkzeug.utils import secure_filename
    from PIL import Image
    import os

    f = request.files.get('design')
    if not f or not f.filename:
        return jsonify(ok=False, error='Please choose an image to upload.'), 400

    conn = _db()
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p "
            "WHERE p.id=%s AND p.is_deleted=0" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403

        ms = _barcode_milestone(conn, pid)
        if ms and ms.get('status') == 'approved':
            return jsonify(ok=False,
                           error='Milestone is marked as done — barcode is locked.'), 403

        # validate dimensions (must be exactly 1024 x 1024)
        try:
            img = Image.open(f.stream)
            w, h = img.size
        except Exception:
            return jsonify(ok=False, error='Could not read the image file.'), 400
        if w != 1024 or h != 1024:
            return jsonify(ok=False,
                           error=f'Image must be exactly 1024 x 1024 px '
                                 f'(uploaded image is {w} x {h} px).'), 400

        f.stream.seek(0)
        sub = os.path.join('barcode', str(pid))
        folder = os.path.join(_npd_upload_dir(), sub)
        os.makedirs(folder, exist_ok=True)
        fn = ('design_%s_%s' % (_dt_stamp(), secure_filename(f.filename) or 'file'))[:150]
        f.save(os.path.join(folder, fn))
        rel = (sub + '/' + fn).replace('\\', '/')

        nxt = conn.execute(
            "SELECT COALESCE(MAX(sr_no),0)+1 AS n FROM `npd_barcode_designs` "
            "WHERE project_id=%s", (pid,)).fetchone()
        sr_no = nxt['n'] if nxt else 1

        cur = conn.execute(
            "INSERT INTO `npd_barcode_designs` "
            "(project_id, milestone_id, sr_no, design_path, design_width, "
            "design_height, created_by) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (pid, ms['id'] if ms else None, sr_no, rel, w, h, _uid()))
        did = cur.lastrowid

        action = f"Barcode design #{sr_no} uploaded"
        _log(conn, pid, action)
        if ms:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms['id'], action, _uid()))
        conn.commit()

        row = conn.execute(
            "SELECT * FROM `npd_barcode_designs` WHERE id=%s", (did,)).fetchone()
        return jsonify(ok=True, design=_bc_design_json(conn, row))
    finally:
        conn.close()


@npd_bp.route('/barcode/<int:did>/upload-barcode', methods=['POST'])
@login_required
def barcode_upload_barcode(did):
    """Barcode Team uploads the barcode file against an existing design row."""
    from flask import jsonify
    from werkzeug.utils import secure_filename
    import os

    f = request.files.get('barcode')
    if not f or not f.filename:
        return jsonify(ok=False, error='Please choose a file to upload.'), 400

    conn = _db()
    try:
        row = conn.execute(
            "SELECT * FROM `npd_barcode_designs` WHERE id=%s", (did,)).fetchone()
        if not row:
            return jsonify(ok=False, error='Design not found.'), 404
        pid = row['project_id']
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p WHERE p.id=%s" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403

        ms = _barcode_milestone(conn, pid)
        if ms and ms.get('status') == 'approved':
            return jsonify(ok=False,
                           error='Milestone is marked as done — barcode is locked.'), 403

        sub = os.path.join('barcode', str(pid))
        folder = os.path.join(_npd_upload_dir(), sub)
        os.makedirs(folder, exist_ok=True)
        fn = ('barcode_%d_%s_%s' % (did, _dt_stamp(), secure_filename(f.filename) or 'file'))[:150]
        f.save(os.path.join(folder, fn))
        rel = (sub + '/' + fn).replace('\\', '/')

        conn.execute(
            "UPDATE `npd_barcode_designs` SET barcode_path=%s WHERE id=%s",
            (rel, did))
        action = f"Barcode file uploaded for design #{row['sr_no']}"
        _log(conn, pid, action)
        if ms:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms['id'], action, _uid()))
        conn.commit()
        out = conn.execute(
            "SELECT * FROM `npd_barcode_designs` WHERE id=%s", (did,)).fetchone()
        return jsonify(ok=True, path=rel, design=_bc_design_json(conn, out))
    finally:
        conn.close()


@npd_bp.route('/barcode/<int:did>/delete', methods=['POST'])
@login_required
def barcode_delete(did):
    from flask import jsonify
    import os

    conn = _db()
    try:
        row = conn.execute(
            "SELECT * FROM `npd_barcode_designs` WHERE id=%s", (did,)).fetchone()
        if not row:
            return jsonify(ok=False, error='Design not found.'), 404
        pid = row['project_id']
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p WHERE p.id=%s" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403

        ms = _barcode_milestone(conn, pid)
        if ms and ms.get('status') == 'approved':
            return jsonify(ok=False,
                           error='Milestone is marked as done — barcode is locked.'), 403

        if row.get('barcode_path'):
            return jsonify(ok=False,
                           error='Barcode already uploaded — designer image is locked. '
                                 'Delete the barcode first to remove this design.'), 403

        for col in ('design_path', 'barcode_path'):
            if row.get(col):
                try:
                    os.remove(os.path.join(_npd_upload_dir(), row[col]))
                except Exception:
                    pass

        conn.execute("DELETE FROM `npd_barcode_designs` WHERE id=%s", (did,))
        action = f"Barcode design #{row['sr_no']} deleted"
        _log(conn, pid, action)
        if ms:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms['id'], action, _uid()))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@npd_bp.route('/barcode/<int:did>/change-design', methods=['POST'])
@login_required
def barcode_change_design(did):
    """Designer re-uploads (changes) the design image (must be 1024x1024 px)."""
    from flask import jsonify
    from werkzeug.utils import secure_filename
    from PIL import Image
    import os

    f = request.files.get('design')
    if not f or not f.filename:
        return jsonify(ok=False, error='Please choose an image to upload.'), 400

    conn = _db()
    try:
        row = conn.execute(
            "SELECT * FROM `npd_barcode_designs` WHERE id=%s", (did,)).fetchone()
        if not row:
            return jsonify(ok=False, error='Design not found.'), 404
        pid = row['project_id']
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p WHERE p.id=%s" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403

        ms = _barcode_milestone(conn, pid)
        if ms and ms.get('status') == 'approved':
            return jsonify(ok=False,
                           error='Milestone is marked as done — barcode is locked.'), 403

        if row.get('barcode_path'):
            return jsonify(ok=False,
                           error='Barcode already uploaded — designer image is locked. '
                                 'Delete the barcode first to change the design.'), 403

        # validate dimensions (must be exactly 1024 x 1024)
        try:
            img = Image.open(f.stream)
            w, h = img.size
        except Exception:
            return jsonify(ok=False, error='Could not read the image file.'), 400
        if w != 1024 or h != 1024:
            return jsonify(ok=False,
                           error=f'Image must be exactly 1024 x 1024 px '
                                 f'(uploaded image is {w} x {h} px).'), 400

        f.stream.seek(0)
        sub = os.path.join('barcode', str(pid))
        folder = os.path.join(_npd_upload_dir(), sub)
        os.makedirs(folder, exist_ok=True)
        fn = ('design_%s_%s' % (_dt_stamp(), secure_filename(f.filename) or 'file'))[:150]
        f.save(os.path.join(folder, fn))
        rel = (sub + '/' + fn).replace('\\', '/')

        old = row.get('design_path')
        conn.execute(
            "UPDATE `npd_barcode_designs` SET design_path=%s, design_width=%s, "
            "design_height=%s WHERE id=%s", (rel, w, h, did))
        if old:
            try:
                os.remove(os.path.join(_npd_upload_dir(), old))
            except Exception:
                pass

        action = f"Barcode design #{row['sr_no']} image changed"
        _log(conn, pid, action)
        if ms:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms['id'], action, _uid()))
        conn.commit()

        out = conn.execute(
            "SELECT * FROM `npd_barcode_designs` WHERE id=%s", (did,)).fetchone()
        return jsonify(ok=True, design=_bc_design_json(conn, out))
    finally:
        conn.close()


@npd_bp.route('/barcode/<int:did>/change-barcode', methods=['POST'])
@login_required
def barcode_change_barcode(did):
    """Barcode Team re-uploads (changes) the barcode file against a design."""
    from flask import jsonify
    from werkzeug.utils import secure_filename
    import os

    f = request.files.get('barcode')
    if not f or not f.filename:
        return jsonify(ok=False, error='Please choose a file to upload.'), 400

    conn = _db()
    try:
        row = conn.execute(
            "SELECT * FROM `npd_barcode_designs` WHERE id=%s", (did,)).fetchone()
        if not row:
            return jsonify(ok=False, error='Design not found.'), 404
        pid = row['project_id']
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p WHERE p.id=%s" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403

        ms = _barcode_milestone(conn, pid)
        if ms and ms.get('status') == 'approved':
            return jsonify(ok=False,
                           error='Milestone is marked as done — barcode is locked.'), 403

        sub = os.path.join('barcode', str(pid))
        folder = os.path.join(_npd_upload_dir(), sub)
        os.makedirs(folder, exist_ok=True)
        fn = ('barcode_%d_%s_%s' % (did, _dt_stamp(), secure_filename(f.filename) or 'file'))[:150]
        f.save(os.path.join(folder, fn))
        rel = (sub + '/' + fn).replace('\\', '/')

        old = row.get('barcode_path')
        conn.execute(
            "UPDATE `npd_barcode_designs` SET barcode_path=%s WHERE id=%s",
            (rel, did))
        if old:
            try:
                os.remove(os.path.join(_npd_upload_dir(), old))
            except Exception:
                pass

        action = f"Barcode file changed for design #{row['sr_no']}"
        _log(conn, pid, action)
        if ms:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms['id'], action, _uid()))
        conn.commit()
        return jsonify(ok=True, path=rel)
    finally:
        conn.close()


@npd_bp.route('/barcode/<int:did>/delete-barcode', methods=['POST'])
@login_required
def barcode_delete_barcode(did):
    """Barcode Team deletes ONLY the barcode file (design row stays)."""
    from flask import jsonify
    import os

    conn = _db()
    try:
        row = conn.execute(
            "SELECT * FROM `npd_barcode_designs` WHERE id=%s", (did,)).fetchone()
        if not row:
            return jsonify(ok=False, error='Design not found.'), 404
        pid = row['project_id']
        vis_frag, vis_params = _proj_visibility_sql('p')
        ok = conn.execute(
            "SELECT p.id FROM `npd_projects` p WHERE p.id=%s" + vis_frag,
            [pid] + vis_params).fetchone()
        if not ok:
            return jsonify(ok=False, error='You do not have access.'), 403

        ms = _barcode_milestone(conn, pid)
        if ms and ms.get('status') == 'approved':
            return jsonify(ok=False,
                           error='Milestone is marked as done — barcode is locked.'), 403

        if row.get('barcode_path'):
            try:
                os.remove(os.path.join(_npd_upload_dir(), row['barcode_path']))
            except Exception:
                pass

        conn.execute(
            "UPDATE `npd_barcode_designs` SET barcode_path=NULL WHERE id=%s",
            (did,))
        action = f"Barcode file deleted for design #{row['sr_no']}"
        _log(conn, pid, action)
        if ms:
            conn.execute(
                "INSERT INTO `milestone_logs` (milestone_id, action, "
                "created_by) VALUES (%s,%s,%s)", (ms['id'], action, _uid()))
        conn.commit()

        out = conn.execute(
            "SELECT * FROM `npd_barcode_designs` WHERE id=%s", (did,)).fetchone()
        return jsonify(ok=True, design=_bc_design_json(conn, out))
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# R&D PROJECTS  (Unallotted / Allotted / Closed tabs + member chips + Log)
# ─────────────────────────────────────────────────────────────────────────────
# Closed statuses (project considered finished/cancelled)
_RD_CLOSED = {'completed', 'complete', 'closed', 'done', 'project_closed',
              'cancelled', 'finish', 'finished'}

_RD_DEPT_OK = {'rd', 'randd', 'rnd', 'researchanddevelopment',
               'researchdevelopment', 'researchndevelopment'}


def _is_rd_dept(dept):
    import re as _re
    if not dept:
        return False
    n = _re.sub(r'[^a-z]', '', str(dept).lower())
    return (n in _RD_DEPT_OK or n.startswith('researchdevelop')
            or n.startswith('researchanddevelop'))


def _rd_people(conn):
    """Active R&D-department users from User_Tbl (for the Assign modal)."""
    rows = conn.execute(
        "SELECT id, full_name, username, designation, department "
        "FROM `User_Tbl` WHERE is_active=1 ORDER BY full_name") .fetchall() or []
    out = []
    for r in rows:
        if _is_rd_dept(r.get('department')):
            out.append({
                'id': r['id'],
                'name': r.get('full_name') or r.get('username') or f"#{r['id']}",
                'designation': r.get('designation') or 'R&D',
            })
    return out


def _is_rd_assignee(conn, pid, uid):
    """True only if `uid` is an active R&D assignee of project `pid`.
    Used to gate START/FINISH — project must be assigned AND only the
    assigned R&D person can start/finish. Safe if R&D tables absent."""
    try:
        r = conn.execute(
            "SELECT 1 FROM `rd_sub_assignments` "
            "WHERE project_id=%s AND user_id=%s AND is_active=1 LIMIT 1",
            (pid, uid)).fetchone()
        return bool(r)
    except Exception:
        return False


@npd_bp.route('/rd-projects')
@login_required
def rd_projects():
    """R&D Manager view — assign R&D team to projects (3 tabs)."""
    conn = _db()
    try:
        q = (request.args.get('q') or '').strip()
        umap = _user_map(conn)

        # status badges (name/slug/color) — same as NPD projects list
        statuses = conn.execute(
            "SELECT name, slug, color FROM `npd_statuses` ORDER BY sort_order"
        ).fetchall()

        where = "WHERE p.is_deleted=0"
        params = []
        if q:
            where += (" AND (p.product_name LIKE %s OR p.code LIKE %s "
                      "OR p.client_company LIKE %s OR p.client_name LIKE %s)")
            like = f"%{q}%"
            params += [like, like, like, like]
        rows = conn.execute(
            "SELECT * FROM `npd_projects` p " + where + " ORDER BY p.id DESC",
            params).fetchall()

        # active sub-assignments → member chips + which projects are allotted
        subs = conn.execute(
            "SELECT project_id, user_id, variant_code, status "
            "FROM `rd_sub_assignments` WHERE is_active=1").fetchall()
        assigned_pids = set()
        members = {}          # pid -> [ {name, variant_code, status} ]
        assign_map = {}       # pid -> [ {user_id, variant_code} ]  (for prefill)
        my_pids = set()
        uid = _uid()
        for s in subs:
            pid = s['project_id']
            assigned_pids.add(pid)
            members.setdefault(pid, []).append({
                'name': umap.get(s['user_id'], '—'),
                'variant_code': s.get('variant_code') or '',
                'status': s.get('status') or 'not_started',
            })
            assign_map.setdefault(pid, []).append({
                'user_id': s['user_id'],
                'variant_code': s.get('variant_code') or '',
            })
            if s['user_id'] == uid:
                my_pids.add(pid)
        for pid in members:
            members[pid].sort(key=lambda m: (m['name'] or '').lower())

        full = _has_full_visibility()

        def _visible(p):
            # Manager / admin → all. Else only projects assigned to me.
            if full:
                return True
            if p['id'] in my_pids:
                return True
            if (p.get('assigned_rd') == uid or p.get('assigned_sc') == uid
                    or p.get('npd_poc') == uid):
                return True
            return False

        unallotted, allotted, closed = [], [], []
        for p in rows:
            stt = (p.get('status') or '').lower()
            if stt in _RD_CLOSED:
                if _visible(p):
                    closed.append(p)
            elif p['id'] in assigned_pids:
                if _visible(p):
                    allotted.append(p)
            else:
                # Unallotted is the assign pool — only managers act on it.
                if full:
                    unallotted.append(p)

        return _render('npd/rd_projects.html',
                       sidebar_menu=get_menu('rd', role=_role(),
                                             is_admin=_is_admin()),
                       active_item='rd-list',
                       q=q, statuses=statuses,
                       unallotted=unallotted, allotted=allotted, closed=closed,
                       members=members, assign_map=assign_map,
                       rd_people=_rd_people(conn),
                       umap=umap, is_manager=full)
    finally:
        conn.close()


@npd_bp.route('/rd-projects/<int:pid>/log')
@login_required
def rd_project_log(pid):
    """JSON for the R&D Project Log modal — team summary + activity timeline."""
    from flask import jsonify
    conn = _db()
    try:
        umap = _user_map(conn)
        subs = conn.execute(
            "SELECT * FROM `rd_sub_assignments` WHERE project_id=%s "
            "ORDER BY id", (pid,)).fetchall()
        logs = conn.execute(
            "SELECT * FROM `rd_project_logs` WHERE project_id=%s "
            "ORDER BY id", (pid,)).fetchall()

        def _dur(s):
            tot = s.get('total_seconds') or 0
            if not tot and s.get('started_at') and s.get('finished_at'):
                try:
                    tot = int((s['finished_at'] - s['started_at']).total_seconds())
                except Exception:
                    tot = 0
            h, rem = divmod(int(tot), 3600)
            m, sec = divmod(rem, 60)
            return '%02d:%02d:%02d' % (h, m, sec)

        return jsonify(
            ok=True,
            members=[{
                'name': umap.get(s['user_id'], '—'),
                'variant_code': s.get('variant_code') or '',
                'status': s.get('status') or 'not_started',
                'started': str(s.get('started_at') or '')[:19],
                'finished': str(s.get('finished_at') or '')[:19],
                'duration': _dur(s),
            } for s in subs],
            logs=[{
                'event': (l.get('event') or '').replace('_', ' ').title(),
                'detail': l.get('detail') or '',
                'user': umap.get(l.get('user_id'), 'System'),
                'at': str(l.get('created_at') or '')[:19],
            } for l in logs])
    finally:
        conn.close()


@npd_bp.route('/rd-projects/<int:pid>/assign', methods=['POST'])
@login_required
def rd_assign(pid):
    """Assign / reassign R&D team (manager only). Body: {members:[{user_id,variant_code}]}"""
    from flask import jsonify
    if not _has_full_visibility():
        return jsonify(ok=False,
                       error='Only R&D Manager can assign or reassign projects.'), 403

    data = request.get_json(silent=True) or {}
    items = data.get('members') or []

    conn = _db()
    try:
        proj = conn.execute(
            "SELECT id FROM `npd_projects` WHERE id=%s AND is_deleted=0",
            (pid,)).fetchone()
        if not proj:
            return jsonify(ok=False, error='Project not found.'), 404

        umap = _user_map(conn)
        allowed = {p['id'] for p in _rd_people(conn)}

        keep_ids, names = [], []
        for it in items:
            try:
                uid = int(it.get('user_id'))
            except (TypeError, ValueError):
                continue
            if uid not in allowed:
                continue
            variant = (str(it.get('variant_code') or '').strip())[:100]
            keep_ids.append(uid)
            names.append(umap.get(uid, f'#{uid}'))
            ex = conn.execute(
                "SELECT id FROM `rd_sub_assignments` "
                "WHERE project_id=%s AND user_id=%s", (pid, uid)).fetchone()
            if ex:
                conn.execute(
                    "UPDATE `rd_sub_assignments` SET variant_code=%s, "
                    "assigned_by=%s, assigned_at=NOW(), is_active=1 WHERE id=%s",
                    (variant, _uid(), ex['id']))
            else:
                conn.execute(
                    "INSERT INTO `rd_sub_assignments` "
                    "(project_id, user_id, variant_code, assigned_by, "
                    "assigned_at, status, is_active) "
                    "VALUES (%s,%s,%s,%s,NOW(),'not_started',1)",
                    (pid, uid, variant, _uid()))

        if not keep_ids:
            return jsonify(ok=False,
                           error='Select at least one R&D person.'), 400

        # deactivate stale (un-checked) assignments
        removed = []
        stale = conn.execute(
            "SELECT id, user_id FROM `rd_sub_assignments` "
            "WHERE project_id=%s AND is_active=1", (pid,)).fetchall()
        for s in stale:
            if s['user_id'] not in keep_ids:
                conn.execute(
                    "UPDATE `rd_sub_assignments` SET is_active=0 WHERE id=%s",
                    (s['id'],))
                removed.append(umap.get(s['user_id'], f"#{s['user_id']}"))

        # backward-compat project fields
        members_str = ','.join('u_%d' % u for u in keep_ids)
        conn.execute(
            "UPDATE `npd_projects` SET assigned_rd=%s, assigned_rd_members=%s "
            "WHERE id=%s", (keep_ids[0], members_str, pid))

        # logs
        detail = ("Members: " + ', '.join(names)
                  + (" — Removed: " + ', '.join(removed) if removed else '')
                  + " — by " + _uname())
        _log(conn, pid, "R&D team assigned: " + ', '.join(names))
        conn.execute(
            "INSERT INTO `rd_project_logs` (project_id, user_id, event, detail) "
            "VALUES (%s,%s,'assigned',%s)", (pid, _uid(), detail))
        conn.commit()

        subs = conn.execute(
            "SELECT user_id, variant_code, status FROM `rd_sub_assignments` "
            "WHERE project_id=%s AND is_active=1 ORDER BY id", (pid,)).fetchall()
        out = [{'name': umap.get(s['user_id'], '—'),
                'variant_code': s.get('variant_code') or '',
                'status': s.get('status') or 'not_started',
                'user_id': s['user_id']} for s in subs]
        out.sort(key=lambda m: (m['name'] or '').lower())
        return jsonify(ok=True, members=out, removed=removed)
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# NPD MASTERS  (Milestone Master · NPD Status Master · NPD Category Master)
# ─────────────────────────────────────────────────────────────────────────────
def _masters_guard():
    """Only admin / manager can manage masters."""
    return _has_full_visibility()


def _slugify(s):
    import re as _re
    return _re.sub(r'_+', '_', _re.sub(r'[^a-z0-9]+', '_',
                   (s or '').strip().lower())).strip('_')


def _npd_cats(conn):
    """Category options for the NPD form — strictly from NPD Category Master.
    No fallback: whatever is in npd_categories (active) is what the form shows."""
    try:
        rows = conn.execute(
            "SELECT name FROM `npd_categories` WHERE is_active=1 "
            "ORDER BY sort_order, name").fetchall()
        return [r['name'] for r in (rows or [])]
    except Exception:
        return []


# ── Milestone Master ─────────────────────────────────────────────────────────
@npd_bp.route('/milestone-master')
@login_required
def milestone_master():
    conn = _db()
    try:
        rows = conn.execute(
            "SELECT * FROM `npd_milestone_templates` "
            "ORDER BY sort_order, id").fetchall()
        return _render('npd/milestone_master.html', rows=rows,
                       can_edit=_masters_guard(), active_item='npd-ms-master')
    finally:
        conn.close()


@npd_bp.route('/milestone-master/save', methods=['POST'])
@login_required
def milestone_master_save():
    if not _masters_guard():
        flash('You do not have access.', 'error')
        return redirect(url_for('npd.milestone_master'))
    f = request.form
    mid = (f.get('id') or '').strip()
    title = (f.get('title') or '').strip()
    mtype = (f.get('milestone_type') or '').strip().lower().replace(' ', '_') \
        or _slugify(title)
    if not title or not mtype:
        flash('Type and Title are required.', 'error')
        return redirect(url_for('npd.milestone_master'))
    desc = (f.get('description') or '').strip()
    icon = (f.get('icon') or '📌').strip() or '📌'
    applies = f.get('applies_to') or 'both'
    dflt = 1 if f.get('default_selected') else 0
    mand = 1 if f.get('is_mandatory') else 0
    sort = int(f.get('sort_order') or 0)
    active = 1 if f.get('is_active') else 0
    conn = _db()
    try:
        if mid:
            conn.execute(
                "UPDATE `npd_milestone_templates` SET title=%s, description=%s, "
                "icon=%s, applies_to=%s, default_selected=%s, is_mandatory=%s, "
                "sort_order=%s, is_active=%s, modified_by=%s, modified_at=NOW() "
                "WHERE id=%s",
                (title, desc, icon, applies, dflt, mand, sort, active,
                 _uid(), mid))
            flash('Milestone updated.', 'success')
        else:
            dup = conn.execute(
                "SELECT id FROM `npd_milestone_templates` WHERE milestone_type=%s",
                (mtype,)).fetchone()
            if dup:
                flash(f'Milestone type "{mtype}" already exists.', 'error')
                return redirect(url_for('npd.milestone_master'))
            conn.execute(
                "INSERT INTO `npd_milestone_templates` (milestone_type, title, "
                "description, icon, applies_to, default_selected, is_mandatory, "
                "sort_order, is_active, created_by) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (mtype, title, desc, icon, applies, dflt, mand, sort, active,
                 _uid()))
            flash('Milestone added.', 'success')
        conn.commit()
    finally:
        conn.close()
    return redirect(url_for('npd.milestone_master'))


@npd_bp.route('/milestone-master/<int:mid>/delete', methods=['POST'])
@login_required
def milestone_master_delete(mid):
    if not _masters_guard():
        flash('You do not have access.', 'error')
        return redirect(url_for('npd.milestone_master'))
    conn = _db()
    try:
        conn.execute("DELETE FROM `npd_milestone_templates` WHERE id=%s", (mid,))
        conn.commit()
        flash('Milestone deleted.', 'success')
    finally:
        conn.close()
    return redirect(url_for('npd.milestone_master'))


# ── NPD Status Master ────────────────────────────────────────────────────────
@npd_bp.route('/status-master')
@login_required
def status_master():
    conn = _db()
    try:
        rows = conn.execute(
            "SELECT * FROM `npd_statuses` ORDER BY sort_order, id").fetchall()
        return _render('npd/npd_status_master.html', rows=rows,
                       can_edit=_masters_guard(), active_item='npd-st-master')
    finally:
        conn.close()


@npd_bp.route('/status-master/save', methods=['POST'])
@login_required
def status_master_save():
    if not _masters_guard():
        flash('You do not have access.', 'error')
        return redirect(url_for('npd.status_master'))
    f = request.form
    sid = (f.get('id') or '').strip()
    name = (f.get('name') or '').strip()
    if not name:
        flash('Name is required.', 'error')
        return redirect(url_for('npd.status_master'))
    slug = (f.get('slug') or '').strip().lower() or _slugify(name)
    color = (f.get('color') or '#6b7280').strip() or '#6b7280'
    icon = (f.get('icon') or '🔵').strip() or '🔵'
    sort = int(f.get('sort_order') or 0)
    active = 1 if f.get('is_active') else 0
    conn = _db()
    try:
        if sid:
            conn.execute(
                "UPDATE `npd_statuses` SET name=%s, slug=%s, color=%s, icon=%s, "
                "sort_order=%s, is_active=%s, modified_by=%s, modified_at=NOW() "
                "WHERE id=%s",
                (name, slug, color, icon, sort, active, _uid(), sid))
            flash('Status updated.', 'success')
        else:
            dup = conn.execute(
                "SELECT id FROM `npd_statuses` WHERE name=%s OR slug=%s",
                (name, slug)).fetchone()
            if dup:
                flash('A status with this name/slug already exists.', 'error')
                return redirect(url_for('npd.status_master'))
            conn.execute(
                "INSERT INTO `npd_statuses` (name, slug, color, icon, "
                "sort_order, is_active, created_by) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (name, slug, color, icon, sort, active, _uid()))
            flash('Status added.', 'success')
        conn.commit()
    finally:
        conn.close()
    return redirect(url_for('npd.status_master'))


@npd_bp.route('/status-master/<int:sid>/delete', methods=['POST'])
@login_required
def status_master_delete(sid):
    if not _masters_guard():
        flash('You do not have access.', 'error')
        return redirect(url_for('npd.status_master'))
    conn = _db()
    try:
        conn.execute("DELETE FROM `npd_statuses` WHERE id=%s", (sid,))
        conn.commit()
        flash('Status deleted.', 'success')
    finally:
        conn.close()
    return redirect(url_for('npd.status_master'))


# ── NPD Category Master ──────────────────────────────────────────────────────
@npd_bp.route('/category-master')
@login_required
def category_master():
    conn = _db()
    try:
        try:
            rows = conn.execute(
                "SELECT * FROM `npd_categories` ORDER BY sort_order, name"
            ).fetchall()
        except Exception:
            rows = []
        return _render('npd/npd_category_master.html', rows=rows,
                       can_edit=_masters_guard(), active_item='npd-cat-master')
    finally:
        conn.close()


@npd_bp.route('/category-master/save', methods=['POST'])
@login_required
def category_master_save():
    if not _masters_guard():
        flash('You do not have access.', 'error')
        return redirect(url_for('npd.category_master'))
    f = request.form
    cid = (f.get('id') or '').strip()
    name = (f.get('name') or '').strip()
    if not name:
        flash('Name is required.', 'error')
        return redirect(url_for('npd.category_master'))
    sort = int(f.get('sort_order') or 0)
    active = 1 if f.get('is_active') else 0
    conn = _db()
    try:
        if cid:
            conn.execute(
                "UPDATE `npd_categories` SET name=%s, sort_order=%s, is_active=%s "
                "WHERE id=%s", (name, sort, active, cid))
            flash('Category updated.', 'success')
        else:
            dup = conn.execute(
                "SELECT id FROM `npd_categories` WHERE name=%s", (name,)).fetchone()
            if dup:
                flash('This category already exists.', 'error')
                return redirect(url_for('npd.category_master'))
            conn.execute(
                "INSERT INTO `npd_categories` (name, sort_order, is_active, "
                "created_by) VALUES (%s,%s,%s,%s)", (name, sort, active, _uid()))
            flash('Category added.', 'success')
        conn.commit()
    finally:
        conn.close()
    return redirect(url_for('npd.category_master'))


@npd_bp.route('/category-master/<int:cid>/delete', methods=['POST'])
@login_required
def category_master_delete(cid):
    if not _masters_guard():
        flash('You do not have access.', 'error')
        return redirect(url_for('npd.category_master'))
    conn = _db()
    try:
        conn.execute("DELETE FROM `npd_categories` WHERE id=%s", (cid,))
        conn.commit()
        flash('Category deleted.', 'success')
    finally:
        conn.close()
    return redirect(url_for('npd.category_master'))


# ─────────────────────────────────────────────────────────────────────────────
PERIODS = {'all': 'All Time', 'today': 'Today', 'yesterday': 'Yesterday',
           'last_7_days': 'Last 7 Days', 'last_30_days': 'Last 30 Days'}
_PAL = ['#8b5cf6', '#3b82f6', '#10b981', '#f59e0b', '#ef4444',
        '#06b6d4', '#ec4899', '#64748b']


@npd_bp.route('/dashboard')
@login_required
def dashboard():
    from datetime import datetime as _dt, timedelta

    period = (request.args.get('period') or 'all').lower()
    if period not in PERIODS:
        period = 'all'
    now = _dt.now()
    pf = pt = None
    if period == 'today':
        pf = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == 'yesterday':
        y = now - timedelta(days=1)
        pf = y.replace(hour=0, minute=0, second=0, microsecond=0)
        pt = y.replace(hour=23, minute=59, second=59)
    elif period == 'last_7_days':
        pf = now - timedelta(days=7)
    elif period == 'last_30_days':
        pf = now - timedelta(days=30)

    conn = _db()
    if not conn:
        flash('Database connection failed.', 'error')
        return redirect('/')
    try:
        vis_frag, vis_params = _proj_visibility_sql('p')

        def _between(col):
            frag, prm = "", []
            if pf is not None:
                frag += f" AND {col} >= %s"; prm.append(pf)
            if pt is not None:
                frag += f" AND {col} <= %s"; prm.append(pt)
            return frag, prm

        proj_rng, proj_prm = _between('p.created_at')
        samp_rng, samp_prm = _between('t.dispatched_at')

        P_BASE = ("FROM `npd_projects` p WHERE p.is_deleted=0 "
                  "AND p.project_type='npd'" + vis_frag + proj_rng)
        P_PRM = vis_params + proj_prm

        def _scalar(sql, prm):
            r = conn.execute(sql, prm).fetchone()
            return (list(r.values())[0] if r else 0) or 0

        a = {}
        # ── Project KPIs ─────────────────────────────────────────────────
        total_proj = _scalar(f"SELECT COUNT(*) {P_BASE}", P_PRM)
        a['leads_converted'] = _scalar(
            f"SELECT COUNT(*) {P_BASE} AND p.lead_id IS NOT NULL", P_PRM)
        a['active_projects'] = _scalar(
            f"SELECT COUNT(*) {P_BASE} "
            f"AND p.status NOT IN ('finish','complete','completed','cancelled')", P_PRM)
        a['new_this_month'] = _scalar(
            "SELECT COUNT(*) FROM `npd_projects` p WHERE p.is_deleted=0 "
            "AND p.project_type='npd'" + vis_frag + " AND p.created_at >= %s",
            vis_params + [now.replace(day=1, hour=0, minute=0, second=0,
                                      microsecond=0)])
        a['completed_projects'] = _scalar(
            f"SELECT COUNT(*) {P_BASE} AND p.status='finish'", P_PRM)

        # ── Sample KPIs (office_dispatch_items + tokens) ─────────────────
        S_BASE = ("FROM `office_dispatch_items` i "
                  "JOIN `office_dispatch_tokens` t ON t.id = i.token_id "
                  "JOIN `npd_projects` p ON p.id = i.project_id "
                  "WHERE p.is_deleted=0 AND p.project_type='npd'"
                  + vis_frag + samp_rng)
        S_PRM = vis_params + samp_prm

        def _scount(extra="", prm=None):
            try:
                return _scalar(f"SELECT COUNT(*) {S_BASE}{extra}",
                               S_PRM + (prm or []))
            except Exception:
                return 0

        s_total = _scount()
        s_pending = _scount(" AND i.approval_status='pending'")
        s_approved = _scount(" AND i.approval_status='approved'")
        s_rejected = _scount(" AND i.approval_status='rejected'")
        s_sent = _scount(" AND i.sent_to_client_at IS NOT NULL")
        a.update(total_samples=s_total, in_development=s_pending,
                 internal_approved=s_approved, samples_dispatched=s_sent,
                 internal_rejected=s_rejected,
                 # client-side metrics: old system me bhi data field nahi tha
                 client_approved=0, client_rejected=0, rework_samples=0,
                 pending_feedback=0, in_transit=0, delivered=0,
                 overdue=0, on_time=0)

        # ── Dispatches + avg cycle ───────────────────────────────────────
        cd_rng, cd_prm = _between('d.dispatched_at')
        try:
            a['total_dispatches'] = _scalar(
                "SELECT COUNT(*) FROM `client_dispatch` d "
                "JOIN `npd_projects` p ON p.id = d.project_id "
                "WHERE 1=1" + vis_frag + cd_rng, vis_params + cd_prm)
        except Exception:
            a['total_dispatches'] = 0
        try:
            a['avg_cycle'] = float(_scalar(
                f"SELECT ROUND(AVG(DATEDIFF(i.actioned_at, t.dispatched_at)),1) "
                f"{S_BASE} AND i.actioned_at IS NOT NULL "
                f"AND i.actioned_at >= t.dispatched_at", S_PRM) or 0)
        except Exception:
            a['avg_cycle'] = 0

        # ── Lifecycle funnel ─────────────────────────────────────────────
        a['funnel'] = [
            {'label': 'Leads Converted', 'value': a['leads_converted'], 'color': '#8b5cf6'},
            {'label': 'NPD Projects Created', 'value': total_proj, 'color': '#6366f1'},
            {'label': 'Samples Created', 'value': s_total, 'color': '#3b82f6'},
            {'label': 'Internal Approved', 'value': s_approved, 'color': '#06b6d4'},
            {'label': 'Client Sent', 'value': s_sent, 'color': '#10b981'},
            {'label': 'Client Approved', 'value': a['client_approved'], 'color': '#f59e0b'},
            {'label': 'Projects Completed', 'value': a['completed_projects'], 'color': '#ef4444'},
        ]

        a['sample_status'] = [x for x in [
            {'label': 'In Development', 'value': s_pending, 'color': '#3b82f6'},
            {'label': 'Internal Approval', 'value': s_approved, 'color': '#06b6d4'},
            {'label': 'Client Sent', 'value': s_sent, 'color': '#10b981'},
            {'label': 'Client Rejected', 'value': s_rejected, 'color': '#ef4444'},
        ] if x['value']]

        # ── Project stage overview ───────────────────────────────────────
        a['project_stage'] = []
        try:
            smap = {s['slug']: s['name'] for s in get_npd_statuses(conn)}
            rows = conn.execute(
                f"SELECT p.status AS s, COUNT(*) AS c {P_BASE} "
                f"GROUP BY p.status ORDER BY c DESC", P_PRM).fetchall() or []
            st = [{'label': smap.get(r['s'],
                                     (r['s'] or 'Unknown').replace('_', ' ').title()),
                   'value': r['c']} for r in rows]
            for i, x in enumerate(st):
                x['color'] = _PAL[i % len(_PAL)]
            a['project_stage'] = st[:8]
        except Exception:
            pass

        a['client_trend'] = []      # client approval data nahi hai abhi

        # ── Rejection analysis (internal) ────────────────────────────────
        a['rejection_internal'] = []
        try:
            rr = conn.execute(
                f"SELECT COALESCE(NULLIF(TRIM(i.reject_reason),''),"
                f"'Not specified') AS r, COUNT(*) AS c {S_BASE} "
                f"AND i.approval_status='rejected' GROUP BY r "
                f"ORDER BY c DESC LIMIT 6", S_PRM).fetchall() or []
            a['rejection_internal'] = [
                {'label': r['r'], 'value': r['c'],
                 'color': _PAL[i % len(_PAL)]} for i, r in enumerate(rr)]
        except Exception:
            pass
        a['rejection_client'] = []
        a['top_rework'] = []
        a['pending_feedback_rows'] = []

        # ── Recent dispatches ────────────────────────────────────────────
        a['recent_dispatches'] = []
        try:
            for d in (conn.execute(
                    "SELECT d.token_no, d.dispatched_at, p.code, "
                    "p.client_name, p.client_company "
                    "FROM `client_dispatch` d "
                    "JOIN `npd_projects` p ON p.id = d.project_id "
                    "WHERE 1=1" + vis_frag + cd_rng +
                    " ORDER BY d.dispatched_at DESC LIMIT 6",
                    vis_params + cd_prm).fetchall() or []):
                a['recent_dispatches'].append({
                    'no': d['token_no'] or '—', 'project': d['code'] or '—',
                    'client': d['client_name'] or d['client_company'] or '—',
                    'date': str(d['dispatched_at'] or '')[:10]})
        except Exception:
            pass

        # ── Recently rejected samples ────────────────────────────────────
        a['rejected_recent'] = []
        try:
            for it in (conn.execute(
                    f"SELECT i.sample_code, i.reject_reason, p.code "
                    f"{S_BASE} AND i.approval_status='rejected' "
                    f"ORDER BY t.dispatched_at DESC LIMIT 6",
                    S_PRM).fetchall() or []):
                a['rejected_recent'].append({
                    'sample_code': it['sample_code'] or '—',
                    'project': it['code'] or '—',
                    'reason': it['reject_reason'] or '—'})
        except Exception:
            pass

        # ── Team workload (assigned R&D / creator) ──────────────────────
        a['team_workload'] = []
        try:
            rows = conn.execute(
                "SELECT COALESCE(NULLIF(u.full_name,''), u.username) AS nm, "
                "i.approval_status AS st, COUNT(*) AS c "
                "FROM `office_dispatch_items` i "
                "JOIN `office_dispatch_tokens` t ON t.id = i.token_id "
                "JOIN `npd_projects` p ON p.id = i.project_id "
                "JOIN `User_Tbl` u ON u.id = COALESCE(p.assigned_rd, "
                "p.created_by) WHERE p.is_deleted=0" + vis_frag + samp_rng +
                " GROUP BY nm, st", vis_params + samp_prm).fetchall() or []
            wm = {}
            for r in rows:
                d = wm.setdefault(r['nm'], {'name': r['nm'], 'in_process': 0,
                                            'completed': 0, 'rejected': 0})
                if r['st'] == 'pending':
                    d['in_process'] += r['c']
                elif r['st'] == 'approved':
                    d['completed'] += r['c']
                elif r['st'] == 'rejected':
                    d['rejected'] += r['c']
            a['team_workload'] = sorted(
                wm.values(), key=lambda x: (x['in_process'] + x['completed']
                                            + x['rejected']),
                reverse=True)[:6]
        except Exception:
            pass

        # When reached from the R&D menu (?nav=rd) show the R&D sidebar
        if (request.args.get('nav') or '') == 'rd':
            return _render('npd/npd_dashboard.html', a=a, period=period,
                           periods=PERIODS,
                           sidebar_menu=get_menu('rd', role=_role(),
                                                 is_admin=_is_admin()),
                           active_item='rd-dash')
        return _render('npd/npd_dashboard.html', a=a, period=period,
                       periods=PERIODS, active_item='npd-dash')
    finally:
        conn.close()
