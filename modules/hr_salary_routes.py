"""
hr_salary_routes.py  —  HR Salary Calculation (Blueprint)
=========================================================
Self-contained module for the HR payroll workflow used by HCP Wellness Pvt. Ltd.

Mirrors the SALARY_FACTOR workbook (sheet "March-26 Plant") formula set:

    BH  Gross (input)                          BI = ROUND(BH*50%)   Basic
    BJ  = ROUND(BI*40%)    HRA                 BK = ROUND(BI*30%)   Conveyance
    BL  Medical (fixed, editable, default 1200)
    BM  = BH - (BI+BJ+BK+BL)                   Other Allowance

    AO  Present = COUNTIF("P") + HD*0.5 + (PL-HD/SL-HD/CL-HD)*0.5
    AP  Paid Holiday = COUNTIF("PH") + COUNTIF("PHP")
    AR/AS/AT  PL/CL/SL used = COUNTIF(full) + COUNTIF(*-HD)*0.5
    AU  Total WO = COUNTIF("WO")+COUNTIF("WOP")+COUNTIF("WO-HD")
    AV  Total WOP = COUNTIF("WOP")+COUNTIF("WO-HD")*0.5
    AW  Total PHP = COUNTIF("PHP")+COUNTIF("PH-HD")*0.5
    BA/BB/BC  Closing PL/CL/SL  = Opening - Used

    BD  OT hours = SUM of per-day OT
    BE  OT days  = BD / 8
    BF  Actual days in month
    BG  Paid days = AO + AP + AR + AU + AS + AT

    BN..BR  Earned head = ROUND(head / BF * BG)        (pro-rated)
    BS  Production incentive
          = ROUND(BH/BF*BE) + ROUND(BH/BF*AV) + ROUND(BH/BF*AW)
    BT  Gross earning = BN+BO+BP+BQ+BR
    BU  Arrear
    BV  Total earned = BT + BS + BU
    BW  Loan outstanding    BY = loan deducted this month    BX = BW - BY

    CA  PT  = 200 if BV >= 12000 else 0
    CB  PF  = 1800 if BN >= 15000 else ROUND(BN*12%)
    CC  TDS
    CD  Total deduction = BY + BZ + CA + CB + CC
    CE  PAID SALARY = BV - CD

Features
--------
* Employee master (add / edit / soft-delete / search)
* Per-employee per-month attendance grid (31 days × status + OT hours)
* Live on-screen recalculation and authoritative server-side save
* Bulk Excel import of employees (with template download)
* Month-level xlsx export
* WhatsApp salary-slip facility (single + bulk) via web.whatsapp.com:
  the UI opens a queue of wa.me/<phone>?text=<slip> links the user
  clicks through in sequence. No phone number ever leaves the server,
  no third-party WhatsApp API.

Access control
--------------
All routes require an authenticated session AND User_Type in ('admin', 'HR').
"""

from flask import (Blueprint, render_template, request, jsonify, session,
                   redirect, url_for, send_file)
from functools import wraps
from datetime import datetime, timedelta
from io import BytesIO
import tempfile
import re
import base64
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.utils import formataddr, make_msgid, formatdate

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import sampling_portal  # shared DB helper (MySQL connection pool)


# ═══════════════════════════════════════════════════════════════════════════════
# Blueprint
# ═══════════════════════════════════════════════════════════════════════════════

hr_salary_bp = Blueprint('hr_salary', __name__)


# ═══════════════════════════════════════════════════════════════════════════════
# DB schema — idempotent
# ═══════════════════════════════════════════════════════════════════════════════

def ensure_hr_salary_tables():
    """Create all HR payroll tables if they don't exist. Safe to run on every start."""
    conn = sampling_portal.get_db_connection()
    if not conn:
        return
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS hr_employees (
                id                INT AUTO_INCREMENT PRIMARY KEY,
                emp_id            VARCHAR(32)  NOT NULL UNIQUE,
                emp_name          VARCHAR(200) NOT NULL,
                gender            VARCHAR(16)  DEFAULT '',
                doj               DATE         NULL,
                designation       VARCHAR(200) DEFAULT '',
                department        VARCHAR(200) DEFAULT '',
                emp_status        VARCHAR(64)  DEFAULT 'PERMANENT',
                mobile            VARCHAR(32)  DEFAULT '',
                email             VARCHAR(200) DEFAULT '',
                gross_salary      DECIMAL(12,2) DEFAULT 0,
                medical_fixed     DECIMAL(12,2) DEFAULT 1200,
                loan_outstanding  DECIMAL(12,2) DEFAULT 0,
                opening_pl        DECIMAL(6,2)  DEFAULT 0,
                opening_cl        DECIMAL(6,2)  DEFAULT 0,
                opening_sl        DECIMAL(6,2)  DEFAULT 0,
                pf_applicable     TINYINT(1)    DEFAULT 1,
                is_active         TINYINT(1)   DEFAULT 1,
                created_at        DATETIME     DEFAULT CURRENT_TIMESTAMP,
                updated_at        DATETIME     DEFAULT CURRENT_TIMESTAMP
                                               ON UPDATE CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS hr_salary_periods (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                emp_row_id      INT NOT NULL,
                period_month    VARCHAR(7)  NOT NULL,
                actual_days     INT         DEFAULT 31,
                gross_salary    DECIMAL(12,2) DEFAULT 0,
                medical_fixed   DECIMAL(12,2) DEFAULT 1200,
                ot_hours        DECIMAL(8,2)  DEFAULT 0,
                arrear          DECIMAL(12,2) DEFAULT 0,
                loan_amount     DECIMAL(12,2) DEFAULT 0,
                loan_deduction  DECIMAL(12,2) DEFAULT 0,
                other_deduction DECIMAL(12,2) DEFAULT 0,
                tds             DECIMAL(12,2) DEFAULT 0,
                opening_pl      DECIMAL(6,2)  DEFAULT 0,
                opening_cl      DECIMAL(6,2)  DEFAULT 0,
                opening_sl      DECIMAL(6,2)  DEFAULT 0,
                pf_applicable   TINYINT(1)    DEFAULT 1,
                remarks         VARCHAR(500) DEFAULT '',
                updated_at      DATETIME     DEFAULT CURRENT_TIMESTAMP
                                             ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uniq_emp_period (emp_row_id, period_month)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS hr_salary_days (
                id           INT AUTO_INCREMENT PRIMARY KEY,
                period_id    INT NOT NULL,
                day_num      INT NOT NULL,
                status_code  VARCHAR(16) DEFAULT '',
                ot_hours     DECIMAL(6,2) DEFAULT 0,
                UNIQUE KEY uniq_period_day (period_id, day_num)
            )
        """)
        # SMTP config — singleton row (id=1) set once by admin
        conn.execute("""
            CREATE TABLE IF NOT EXISTS hr_smtp_config (
                id           INT PRIMARY KEY,
                smtp_host    VARCHAR(200) DEFAULT '',
                smtp_port    INT          DEFAULT 587,
                use_tls      TINYINT(1)   DEFAULT 1,
                use_ssl      TINYINT(1)   DEFAULT 0,
                username     VARCHAR(200) DEFAULT '',
                password_enc VARCHAR(500) DEFAULT '',
                from_email   VARCHAR(200) DEFAULT '',
                from_name    VARCHAR(200) DEFAULT '',
                reply_to     VARCHAR(200) DEFAULT '',
                updated_at   DATETIME     DEFAULT CURRENT_TIMESTAMP
                                          ON UPDATE CURRENT_TIMESTAMP
            )
        """)
        # Loan register — one row per loan agreement
        conn.execute("""
            CREATE TABLE IF NOT EXISTS hr_loans (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                emp_row_id      INT NOT NULL,
                loan_date       DATE NOT NULL,
                principal       DECIMAL(12,2) NOT NULL DEFAULT 0,
                schedule_mode   VARCHAR(16) DEFAULT 'fixed_emi',
                                          -- 'fixed_emi' or 'fixed_tenure'
                emi_amount      DECIMAL(12,2) DEFAULT 0,
                tenure_months   INT           DEFAULT 0,
                start_month     VARCHAR(7)    NOT NULL,  -- YYYY-MM first deduction month
                auto_deduct     TINYINT(1)    DEFAULT 1,
                status          VARCHAR(16)   DEFAULT 'active',
                                          -- 'active', 'closed', 'cancelled'
                purpose         VARCHAR(200)  DEFAULT '',
                remarks         VARCHAR(500)  DEFAULT '',
                created_at      DATETIME      DEFAULT CURRENT_TIMESTAMP,
                updated_at      DATETIME      DEFAULT CURRENT_TIMESTAMP
                                              ON UPDATE CURRENT_TIMESTAMP
            )
        """)
        # Loan EMI schedule — one row per installment
        conn.execute("""
            CREATE TABLE IF NOT EXISTS hr_loan_schedule (
                id                INT AUTO_INCREMENT PRIMARY KEY,
                loan_id           INT NOT NULL,
                installment_no    INT NOT NULL,
                period_month      VARCHAR(7)  NOT NULL,
                scheduled_amount  DECIMAL(12,2) NOT NULL DEFAULT 0,
                paid_amount       DECIMAL(12,2) DEFAULT 0,
                paid_in_period_id INT           DEFAULT NULL,
                status            VARCHAR(16)   DEFAULT 'pending',
                                                -- 'pending', 'paid', 'partial', 'skipped'
                paid_at           DATETIME      DEFAULT NULL,
                UNIQUE KEY uniq_loan_inst (loan_id, installment_no)
            )
        """)
        # Add mobile column for older installs (ignore if already exists)
        try:
            conn.execute("ALTER TABLE hr_employees ADD COLUMN mobile VARCHAR(32) DEFAULT ''")
        except Exception:
            pass
        # Add email column for older installs
        try:
            conn.execute("ALTER TABLE hr_employees ADD COLUMN email VARCHAR(200) DEFAULT ''")
        except Exception:
            pass
        # Add pf_applicable to employee master — default 1 (covered)
        try:
            conn.execute("ALTER TABLE hr_employees ADD COLUMN pf_applicable TINYINT(1) DEFAULT 1")
        except Exception:
            pass
        # Add pf_applicable to period (month snapshot so it's editable per-month)
        try:
            conn.execute("ALTER TABLE hr_salary_periods ADD COLUMN pf_applicable TINYINT(1) DEFAULT 1")
        except Exception:
            pass
        conn.commit()
    except Exception:
        pass
    finally:
        try: conn.close()
        except Exception: pass


# ═══════════════════════════════════════════════════════════════════════════════
# Auth decorator (admin / HR only)
# ═══════════════════════════════════════════════════════════════════════════════

def hr_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        role = (session.get('User_Type') or '').strip()
        if role not in ('admin', 'HR'):
            return ("<div style='font-family:sans-serif;padding:40px;text-align:center'>"
                    "<h2>Access denied</h2>"
                    "<p>HR Salary is restricted to <b>admin</b> and <b>HR</b> users.</p>"
                    "<p><a href='/'>← Back to portal</a></p></div>"), 403
        return f(*args, **kwargs)
    return wrapper


# ═══════════════════════════════════════════════════════════════════════════════
# Status codes & calculation engine
# ═══════════════════════════════════════════════════════════════════════════════

HR_STATUS_CODES = [
    'P', 'WO', 'WOP', 'PH', 'PHP', 'PL', 'CL', 'SL', 'HD', 'L', 'A',
    'PL-HD', 'CL-HD', 'SL-HD', 'WO-HD', 'PH-HD',
]


def _count(codes, *targets):
    """Count occurrences of any target code in codes list (case-insensitive)."""
    s = {t.upper() for t in targets}
    return sum(1 for c in codes if (c or '').upper() in s)


def calc_salary(period, day_codes, ot_hours_list):
    """
    Pure function: given the period header + daily codes + per-day OT,
    return every BI..CE column exactly as the workbook computes them.

    Fields returned (mirrors Excel columns AO through CE):

        Attendance:
          AO_total_present        Present (P) + half-days
          AP_paid_holiday         Paid holiday (PH + PHP)
          AQ_total_leave          Total leave (L + A, unpaid)
          AR_pl_used              PL taken this month
          AS_cl_used              CL taken this month
          AT_sl_used              SL taken this month
          AU_total_wo             Total weekly-offs (WO + WOP + WO-HD)
          AV_total_wop            Weekly-offs worked (WOP + WO-HD/2)
          AW_total_php            Holiday worked (PHP + PH-HD/2)
          AX_opening_pl/cl/sl     Leave balances at start of month
          BA_closing_pl/cl/sl     Leave balances at end of month

        Overtime:
          BD_ot_hours             Total OT hours
          BE_ot_days              OT in days (BD/8)

        Days:
          BF_actual_days          Days in month
          BG_total_days           Paid days (AO + AP + AR + AU + AS + AT)

        Salary "rate" (monthly full value):
          BH_gross                Gross salary per month (CTC)
          BI_basic_rate           = round(BH × 50%)
          BJ_hra_rate             = round(BI × 40%)
          BK_conveyance_rate      = round(BI × 30%)
          BL_medical_rate         = fixed (usually 1200)
          BM_other_allowance_rate = BH - (BI+BJ+BK+BL)

        Salary "earned" (pro-rated for paid days):
          BN_basic_earned         = round(BI / BF × BG)
          BO_hra_earned           = round(BJ / BF × BG)
          BP_conveyance_earned    = round(BK / BF × BG)
          BQ_medical_earned       = round(BL / BF × BG)
          BR_other_allowance_earned = round(BM / BF × BG)

        Earnings total:
          BS_prod_incentive       = round(BH/BF × BE) + round(BH/BF × AV) + round(BH/BF × AW)
          BT_gross_earning        = sum of BN..BR
          BU_arrear               Additional arrear (manual input)
          BV_total_earned         = BT + BS + BU

        Loans:
          BW_loan_amount          Current outstanding loan (opening balance for month)
          BX_loan_balance         = BW - BY (closing balance)
          BY_loan_deduction       EMI deducted this month

        Deductions:
          BZ_other_deduction      Other deductions (manual input)
          CA_pt                   Professional Tax (200 if BV >= 12000, else 0)
          CB_pf                   Provident Fund (1800 if BN >= 15000, else BN × 12%).
                                  Only applied if employee's pf_applicable=1.
          CC_tds                  TDS (manual input)
          CD_total_deduction      = BY + BZ + CA + CB + CC
          CE_paid_salary          = BV - CD (net pay)
    """
    BH = float(period.get('gross_salary')  or 0)
    BL = float(period.get('medical_fixed') or 0)
    BF = int(period.get('actual_days')     or 31)
    if BF <= 0:
        BF = 31
    BU = float(period.get('arrear')          or 0)
    BW = float(period.get('loan_amount')     or 0)
    BY = float(period.get('loan_deduction')  or 0)
    BZ = float(period.get('other_deduction') or 0)
    CC = float(period.get('tds')             or 0)

    # PF applicable flag — per employee. Some workers are exempt.
    # Defaults to True if not specified (backward compatibility with periods
    # created before this flag existed).
    pf_applicable = period.get('pf_applicable')
    if pf_applicable is None:
        pf_applicable = True
    pf_applicable = bool(int(pf_applicable)) if str(pf_applicable).strip() != '' else True

    # ── Attendance counts ──
    AO = (_count(day_codes, 'P')
          + _count(day_codes, 'HD')    * 0.5
          + _count(day_codes, 'PL-HD') * 0.5
          + _count(day_codes, 'SL-HD') * 0.5
          + _count(day_codes, 'CL-HD') * 0.5)
    AP = _count(day_codes, 'PH') + _count(day_codes, 'PHP')
    AQ = _count(day_codes, 'L')  + _count(day_codes, 'A')
    AR = _count(day_codes, 'PL') + _count(day_codes, 'PL-HD') * 0.5
    AS_ = _count(day_codes, 'CL') + _count(day_codes, 'CL-HD') * 0.5
    AT = _count(day_codes, 'SL') + _count(day_codes, 'SL-HD') * 0.5
    AU = _count(day_codes, 'WO') + _count(day_codes, 'WOP') + _count(day_codes, 'WO-HD')
    AV = _count(day_codes, 'WOP') + _count(day_codes, 'WO-HD') * 0.5
    AW = _count(day_codes, 'PHP') + _count(day_codes, 'PH-HD') * 0.5

    # ── Leave balances ──
    AX = float(period.get('opening_pl') or 0)
    AY = float(period.get('opening_cl') or 0)
    AZ = float(period.get('opening_sl') or 0)
    BA = AX - AR
    BB = AY - AS_
    BC = AZ - AT

    # ── Overtime ──
    day_ot_sum = sum(float(x or 0) for x in ot_hours_list)
    BD = day_ot_sum if day_ot_sum > 0 else float(period.get('ot_hours') or 0)
    BE = BD / 8.0

    # ── Paid days ──
    BG = AO + AP + AR + AU + AS_ + AT

    # ── Salary "rate" components (monthly full value) ──
    BI = round(BH * 0.50)
    BJ = round(BI * 0.40)
    BK = round(BI * 0.30)
    BM = BH - (BI + BJ + BK + BL)

    # ── Salary "earned" components (pro-rated for paid days) ──
    def prorate(head):
        return round(head / BF * BG) if BF else 0

    BN = prorate(BI)
    BO = prorate(BJ)
    BP = prorate(BK)
    BQ = prorate(BL)
    BR = prorate(BM)

    BS = ((round(BH / BF * BE) if BF else 0)
          + (round(BH / BF * AV) if BF else 0)
          + (round(BH / BF * AW) if BF else 0))

    BT = BN + BO + BP + BQ + BR
    BV = BT + BS + BU
    BX = BW - BY

    # ── Deductions ──
    CA = 200 if BV >= 12000 else 0
    if pf_applicable:
        CB = 1800 if BN >= 15000 else round(BN * 0.12)
    else:
        CB = 0
    CD = BY + BZ + CA + CB + CC
    CE = BV - CD

    return {
        # Attendance
        'AO_total_present': AO, 'AP_paid_holiday': AP, 'AQ_total_leave': AQ,
        'AR_pl_used': AR, 'AS_cl_used': AS_, 'AT_sl_used': AT,
        'AU_total_wo': AU, 'AV_total_wop': AV, 'AW_total_php': AW,

        # Leave balances
        'AX_opening_pl': AX, 'AY_opening_cl': AY, 'AZ_opening_sl': AZ,
        'BA_closing_pl': BA, 'BB_closing_cl': BB, 'BC_closing_sl': BC,

        # OT
        'BD_ot_hours': BD, 'BE_ot_days': BE,

        # Days
        'BF_actual_days': BF, 'BG_total_days': BG,
        'BH_gross': BH,

        # Rate components (monthly full)
        'BI_basic_rate':           BI,
        'BJ_hra_rate':             BJ,
        'BK_conveyance_rate':      BK,
        'BL_medical_rate':         BL,
        'BM_other_allowance_rate': BM,
        # Aliases for backward compat with any legacy callers
        'BI_basic': BI, 'BJ_hra': BJ, 'BK_conveyance': BK,
        'BL_medical': BL, 'BM_other_allow': BM,

        # Earned components (pro-rated)
        'BN_basic_earned':           BN,
        'BO_hra_earned':             BO,
        'BP_conveyance_earned':      BP,
        'BQ_medical_earned':         BQ,
        'BR_other_allowance_earned': BR,
        # Aliases
        'BN_earned_basic': BN, 'BO_earned_hra': BO, 'BP_earned_conv': BP,
        'BQ_earned_medical': BQ, 'BR_earned_other': BR,

        # Totals
        'BS_prod_incentive':   BS,
        'BT_gross_earning':    BT,
        'BU_arrear':           BU,
        'BV_total_earned':     BV,

        # Loans
        'BW_loan_amount':      BW,
        'BX_loan_balance':     BX,
        'BX_balance_loan':     BX,  # alias

        # Deductions
        'BY_loan_deduction':   BY,
        'BZ_other_deduction':  BZ,
        'CA_pt':               CA,
        'CB_pf':               CB,
        'CC_tds':              CC,
        'CD_total_deduction':  CD,
        'CE_paid_salary':      CE,

        # Flag
        'pf_applicable':       1 if pf_applicable else 0,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Small helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _days_in_month(period_month):
    """Return the number of days in YYYY-MM (fallback 31 on parse error)."""
    try:
        yr, mo = period_month.split('-')
        yr, mo = int(yr), int(mo)
        first_next = datetime(yr + 1, 1, 1) if mo == 12 else datetime(yr, mo + 1, 1)
        return (first_next - timedelta(days=1)).day
    except Exception:
        return 31


def _clean_mobile(raw):
    """
    Normalise a mobile number for wa.me links.
    Returns (digits-only string without +, e164-ish) or '' if unusable.
    Assumes Indian default if 10 digits and no country code.
    """
    if not raw:
        return ''
    s = re.sub(r'\D', '', str(raw))
    if not s:
        return ''
    # 10 digits → assume India (+91)
    if len(s) == 10:
        s = '91' + s
    # Strip leading zeros (e.g. 091XXXXXXXXXX)
    s = s.lstrip('0')
    if len(s) < 10 or len(s) > 15:
        return ''
    return s


def _fmt_inr(n):
    try:
        n = float(n)
    except Exception:
        n = 0
    # Indian number grouping (1,23,456.00)
    sign = '-' if n < 0 else ''
    n = abs(n)
    whole = int(n)
    frac = round(n - whole, 2)
    s = str(whole)
    if len(s) > 3:
        last3 = s[-3:]
        rest  = s[:-3]
        groups = []
        while len(rest) > 2:
            groups.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            groups.insert(0, rest)
        s = ','.join(groups) + ',' + last3
    if frac > 0:
        s += '.' + str(round(frac * 100)).zfill(2).rstrip('0').rstrip('.')
    return sign + '₹' + s


def _load_period_with_days(conn, emp_row_id, period_month, auto_create=True, emp_master=None):
    """
    Fetch period header + daily rows; optionally create from master on first visit.
    Returns (period_dict_or_None, day_rows_list).
    """
    p = conn.execute("""
        SELECT * FROM hr_salary_periods
        WHERE emp_row_id=%s AND period_month=%s
    """, (emp_row_id, period_month)).fetchone()

    if not p and auto_create and emp_master:
        last_day = _days_in_month(period_month)
        # emp_master may be a sqlite Row or dict — try dict conversion
        try:
            em_dict = dict(emp_master)
        except Exception:
            em_dict = emp_master
        pf_val = em_dict.get('pf_applicable', 1) if hasattr(em_dict, 'get') else 1
        pf_flag = 1 if pf_val in (1, True, '1') or pf_val is None else 0
        conn.execute("""
            INSERT INTO hr_salary_periods
                (emp_row_id, period_month, actual_days, gross_salary,
                 medical_fixed, ot_hours, arrear, loan_amount,
                 loan_deduction, other_deduction, tds,
                 opening_pl, opening_cl, opening_sl, pf_applicable, remarks)
            VALUES (%s,%s,%s,%s,%s,0,0,%s,0,0,0,%s,%s,%s,%s,'')
        """, (emp_row_id, period_month, last_day,
              em_dict['gross_salary']    or 0,
              em_dict['medical_fixed']   or 1200,
              em_dict['loan_outstanding'] or 0,
              em_dict['opening_pl']      or 0,
              em_dict['opening_cl']      or 0,
              em_dict['opening_sl']      or 0,
              pf_flag))
        conn.commit()
        p = conn.execute("""
            SELECT * FROM hr_salary_periods
            WHERE emp_row_id=%s AND period_month=%s
        """, (emp_row_id, period_month)).fetchone()

    if not p:
        return None, []

    p = dict(p)
    rows = conn.execute("""
        SELECT day_num, status_code, ot_hours
        FROM hr_salary_days
        WHERE period_id=%s
        ORDER BY day_num ASC
    """, (p['id'],)).fetchall()
    return p, [dict(r) for r in rows]


def _calc_from_period_and_days(period, day_rows):
    """Pad day rows up to actual_days and run calc."""
    day_map = {int(r['day_num']): r for r in day_rows}
    codes, ots = [], []
    for d in range(1, int(period.get('actual_days') or 31) + 1):
        r = day_map.get(d, {})
        codes.append((r.get('status_code') or '').strip())
        ots.append(float(r.get('ot_hours') or 0))
    return calc_salary(period, codes, ots)


# ═══════════════════════════════════════════════════════════════════════════════
# Page route
# ═══════════════════════════════════════════════════════════════════════════════

@hr_salary_bp.route('/hr_salary')
@hr_required
def hr_salary_page():
    return render_template(
        'hr_salary.html',
        role=session.get('User_Type'),
        user_name=session.get('User_Name'),
    )


# ═══════════════════════════════════════════════════════════════════════════════
# API: employee list + CRUD
# ═══════════════════════════════════════════════════════════════════════════════

@hr_salary_bp.route('/api/hr_salary/employees')
@hr_required
def api_employees():
    try:
        conn = sampling_portal.get_db_connection()
        rows = conn.execute("""
            SELECT id, emp_id, emp_name, gender, doj, designation, department,
                   emp_status, mobile, email, gross_salary, medical_fixed,
                   loan_outstanding,
                   opening_pl, opening_cl, opening_sl, pf_applicable, is_active
            FROM hr_employees
            WHERE is_active = 1
            ORDER BY CAST(emp_id AS UNSIGNED) ASC, emp_id ASC
        """).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            if d.get('doj'):
                try: d['doj'] = str(d['doj'])
                except Exception: pass
            out.append(d)
        return jsonify({'status': 'ok', 'employees': out})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/employee/save', methods=['POST'])
@hr_required
def api_employee_save():
    try:
        d = request.get_json(force=True) or {}
        emp_id   = (d.get('emp_id')   or '').strip()
        emp_name = (d.get('emp_name') or '').strip()
        if not emp_id or not emp_name:
            return jsonify({'status': 'error', 'message': 'emp_id and emp_name required'}), 400

        row = (emp_id, emp_name,
               (d.get('gender')      or '').strip(),
               (d.get('doj')         or None) or None,
               (d.get('designation') or '').strip(),
               (d.get('department')  or '').strip(),
               (d.get('emp_status')  or 'PERMANENT').strip(),
               (d.get('mobile')      or '').strip(),
               (d.get('email')       or '').strip(),
               float(d.get('gross_salary')     or 0),
               float(d.get('medical_fixed')    or 1200),
               float(d.get('loan_outstanding') or 0),
               float(d.get('opening_pl')       or 0),
               float(d.get('opening_cl')       or 0),
               float(d.get('opening_sl')       or 0),
               1 if d.get('pf_applicable', 1) else 0)

        conn = sampling_portal.get_db_connection()
        existing = conn.execute(
            "SELECT id FROM hr_employees WHERE emp_id=%s", (emp_id,)
        ).fetchone()
        if existing:
            conn.execute("""
                UPDATE hr_employees
                   SET emp_name=%s, gender=%s, doj=%s, designation=%s, department=%s,
                       emp_status=%s, mobile=%s, email=%s,
                       gross_salary=%s, medical_fixed=%s, loan_outstanding=%s,
                       opening_pl=%s, opening_cl=%s, opening_sl=%s,
                       pf_applicable=%s,
                       is_active=1
                 WHERE emp_id=%s
            """, row[1:] + (emp_id,))
        else:
            conn.execute("""
                INSERT INTO hr_employees
                    (emp_id, emp_name, gender, doj, designation, department,
                     emp_status, mobile, email, gross_salary, medical_fixed,
                     loan_outstanding, opening_pl, opening_cl, opening_sl,
                     pf_applicable)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, row)
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/employee/delete', methods=['POST'])
@hr_required
def api_employee_delete():
    try:
        d = request.get_json(force=True) or {}
        emp_id = (d.get('emp_id') or '').strip()
        if not emp_id:
            return jsonify({'status': 'error', 'message': 'emp_id required'}), 400
        conn = sampling_portal.get_db_connection()
        conn.execute("UPDATE hr_employees SET is_active=0 WHERE emp_id=%s", (emp_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# API: period / days CRUD + live calc
# ═══════════════════════════════════════════════════════════════════════════════

@hr_salary_bp.route('/api/hr_salary/period')
@hr_required
def api_period_get():
    try:
        emp_id = (request.args.get('emp_id') or '').strip()
        month  = (request.args.get('month')  or datetime.now().strftime('%Y-%m')).strip()
        if not emp_id:
            return jsonify({'status': 'error', 'message': 'emp_id required'}), 400

        conn = sampling_portal.get_db_connection()
        emp = conn.execute(
            "SELECT * FROM hr_employees WHERE emp_id=%s AND is_active=1",
            (emp_id,)
        ).fetchone()
        if not emp:
            conn.close()
            return jsonify({'status': 'error', 'message': 'employee not found'}), 404
        emp = dict(emp)
        if emp.get('doj'):
            try: emp['doj'] = str(emp['doj'])
            except Exception: pass

        period, day_rows = _load_period_with_days(conn, emp['id'], month,
                                                  auto_create=True, emp_master=emp)
        conn.close()

        # Pad days up to actual_days for the UI
        last_day = int(period.get('actual_days') or _days_in_month(month))
        day_map = {int(r['day_num']): r for r in day_rows}
        days = []
        codes, ots = [], []
        for d in range(1, last_day + 1):
            r = day_map.get(d, {})
            code = (r.get('status_code') or '').strip()
            ot   = float(r.get('ot_hours') or 0)
            days.append({'day_num': d, 'status_code': code, 'ot_hours': ot})
            codes.append(code); ots.append(ot)

        return jsonify({
            'status': 'ok',
            'employee': emp,
            'period': period,
            'days': days,
            'calc': calc_salary(period, codes, ots),
            'status_codes': HR_STATUS_CODES,
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/period/save', methods=['POST'])
@hr_required
def api_period_save():
    try:
        d = request.get_json(force=True) or {}
        emp_id = (d.get('emp_id')      or '').strip()
        month  = (d.get('period_month') or '').strip()
        if not emp_id or not month:
            return jsonify({'status': 'error', 'message': 'emp_id and period_month required'}), 400
        header = d.get('period') or {}
        days   = d.get('days')   or []

        conn = sampling_portal.get_db_connection()
        emp = conn.execute(
            "SELECT id FROM hr_employees WHERE emp_id=%s AND is_active=1",
            (emp_id,)
        ).fetchone()
        if not emp:
            conn.close()
            return jsonify({'status': 'error', 'message': 'employee not found'}), 404
        emp_row_id = dict(emp)['id']

        vals = (
            int(header.get('actual_days')    or 31),
            float(header.get('gross_salary')    or 0),
            float(header.get('medical_fixed')   or 1200),
            float(header.get('ot_hours')        or 0),
            float(header.get('arrear')          or 0),
            float(header.get('loan_amount')     or 0),
            float(header.get('loan_deduction')  or 0),
            float(header.get('other_deduction') or 0),
            float(header.get('tds')             or 0),
            float(header.get('opening_pl')      or 0),
            float(header.get('opening_cl')      or 0),
            float(header.get('opening_sl')      or 0),
            1 if header.get('pf_applicable', 1) else 0,
            (header.get('remarks') or '').strip()[:500],
        )
        existing = conn.execute("""
            SELECT id FROM hr_salary_periods
            WHERE emp_row_id=%s AND period_month=%s
        """, (emp_row_id, month)).fetchone()
        if existing:
            period_id = dict(existing)['id']
            conn.execute("""
                UPDATE hr_salary_periods
                   SET actual_days=%s, gross_salary=%s, medical_fixed=%s, ot_hours=%s,
                       arrear=%s, loan_amount=%s, loan_deduction=%s, other_deduction=%s,
                       tds=%s, opening_pl=%s, opening_cl=%s, opening_sl=%s,
                       pf_applicable=%s, remarks=%s
                 WHERE id=%s
            """, vals + (period_id,))
        else:
            conn.execute("""
                INSERT INTO hr_salary_periods
                    (emp_row_id, period_month, actual_days, gross_salary, medical_fixed,
                     ot_hours, arrear, loan_amount, loan_deduction, other_deduction,
                     tds, opening_pl, opening_cl, opening_sl, pf_applicable, remarks)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (emp_row_id, month) + vals)
            period_id = dict(conn.execute(
                "SELECT id FROM hr_salary_periods WHERE emp_row_id=%s AND period_month=%s",
                (emp_row_id, month)
            ).fetchone())['id']

        # Replace all day rows for this period
        conn.execute("DELETE FROM hr_salary_days WHERE period_id=%s", (period_id,))
        for d_row in days:
            try:
                dn   = int(d_row.get('day_num'))
                code = (d_row.get('status_code') or '').strip()
                ot   = float(d_row.get('ot_hours') or 0)
            except Exception:
                continue
            if dn < 1 or dn > 31:
                continue
            if code or ot:
                conn.execute("""
                    INSERT INTO hr_salary_days (period_id, day_num, status_code, ot_hours)
                    VALUES (%s,%s,%s,%s)
                """, (period_id, dn, code, ot))

        # ── Auto-deduct loan EMIs (Feature 2e) ───────────────────────────────
        # If the payload explicitly opts out, skip. Otherwise add unpaid auto-EMIs
        # for this month to the existing loan_deduction (never overwrite).
        apply_auto = d.get('apply_auto_loans', True)
        auto_emi_total = 0
        auto_emi_items = []
        if apply_auto:
            auto_emi_total, auto_emi_items = _collect_auto_loan_emis(
                conn, emp_row_id, month
            )
            if auto_emi_total > 0:
                # Add to existing deduction — HR might have typed an additional
                # manual amount, preserve it.
                new_deduction = float(vals[6]) + auto_emi_total  # vals[6] = loan_deduction
                conn.execute(
                    "UPDATE hr_salary_periods SET loan_deduction=%s WHERE id=%s",
                    (new_deduction, period_id)
                )
                _apply_auto_loan_emis(conn, period_id, auto_emi_items)
        conn.commit()

        # Re-read + re-calc for authoritative return
        period = dict(conn.execute(
            "SELECT * FROM hr_salary_periods WHERE id=%s", (period_id,)
        ).fetchone())
        day_rows = [dict(r) for r in conn.execute(
            "SELECT day_num, status_code, ot_hours FROM hr_salary_days WHERE period_id=%s",
            (period_id,)
        ).fetchall()]
        conn.close()
        return jsonify({
            'status': 'ok',
            'calc':   _calc_from_period_and_days(period, day_rows),
            'auto_loan_emi_applied': auto_emi_total,
            'auto_loan_emi_count':   len(auto_emi_items),
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/calc', methods=['POST'])
@hr_required
def api_calc():
    """Stateless live-preview calc — never writes to DB."""
    try:
        d = request.get_json(force=True) or {}
        period = d.get('period') or {}
        days   = d.get('days')   or []
        codes = [(r.get('status_code') or '').strip() for r in days]
        ots   = [float(r.get('ot_hours') or 0)         for r in days]
        return jsonify({'status': 'ok', 'calc': calc_salary(period, codes, ots)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# API: month-level summary & xlsx export
# ═══════════════════════════════════════════════════════════════════════════════

@hr_salary_bp.route('/api/hr_salary/month_summary')
@hr_required
def api_month_summary():
    try:
        month = (request.args.get('month') or datetime.now().strftime('%Y-%m')).strip()
        conn = sampling_portal.get_db_connection()
        emps = conn.execute("""
            SELECT id, emp_id, emp_name, department, designation, mobile, email
            FROM hr_employees
            WHERE is_active = 1
            ORDER BY CAST(emp_id AS UNSIGNED) ASC, emp_id ASC
        """).fetchall()
        out = []
        for e in emps:
            e = dict(e)
            p_row = conn.execute("""
                SELECT * FROM hr_salary_periods
                WHERE emp_row_id=%s AND period_month=%s
            """, (e['id'], month)).fetchone()
            if not p_row:
                out.append({**e, 'has_data': False})
                continue
            p = dict(p_row)
            day_rows = [dict(r) for r in conn.execute(
                "SELECT day_num, status_code, ot_hours FROM hr_salary_days WHERE period_id=%s",
                (p['id'],)
            ).fetchall()]
            c = _calc_from_period_and_days(p, day_rows)
            out.append({
                **e, 'has_data': True,
                'gross':            c['BH_gross'],
                'total_days':       c['BG_total_days'],
                'total_earned':     c['BV_total_earned'],
                'total_deduction':  c['CD_total_deduction'],
                'paid_salary':      c['CE_paid_salary'],
                'pf':               c['CB_pf'],
                'pt':               c['CA_pt'],
            })
        conn.close()
        return jsonify({'status': 'ok', 'month': month, 'rows': out})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/export')
@hr_required
def api_export():
    try:
        month = (request.args.get('month') or datetime.now().strftime('%Y-%m')).strip()
        conn = sampling_portal.get_db_connection()
        emps = conn.execute("""
            SELECT id, emp_id, emp_name, gender, doj, designation, department, emp_status
            FROM hr_employees WHERE is_active=1
            ORDER BY CAST(emp_id AS UNSIGNED) ASC, emp_id ASC
        """).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = f'Payroll {month}'
        headers = [
            'SR.NO', 'EMP ID', 'EMPLOYEE NAME', 'DEPARTMENT', 'DESIGNATION',
            'ACTUAL DAYS', 'TOTAL PRESENT', 'PAID HOLIDAY',
            'PL USED', 'CL USED', 'SL USED', 'WO', 'WOP', 'PHP',
            'OT HOURS', 'OT DAYS', 'PAID DAYS',
            'GROSS (BH)', 'BASIC (BI)', 'HRA (BJ)', 'CONV (BK)', 'MEDICAL (BL)', 'OTHER (BM)',
            'EARNED BASIC (BN)', 'EARNED HRA (BO)', 'EARNED CONV (BP)',
            'EARNED MED (BQ)', 'EARNED OTHER (BR)',
            'PROD INCENTIVE (BS)', 'GROSS EARNING (BT)', 'ARREAR (BU)', 'TOTAL EARNED (BV)',
            'LOAN AMT (BW)', 'LOAN DED (BY)', 'BALANCE (BX)', 'OTHER DED (BZ)',
            'PT (CA)', 'PF (CB)', 'TDS (CC)', 'TOTAL DED (CD)', 'PAID SALARY (CE)',
        ]
        ws.append(headers)
        sr = 0
        for e in emps:
            e = dict(e)
            p_row = conn.execute("""
                SELECT * FROM hr_salary_periods WHERE emp_row_id=%s AND period_month=%s
            """, (e['id'], month)).fetchone()
            if not p_row:
                continue
            p = dict(p_row)
            day_rows = [dict(r) for r in conn.execute(
                "SELECT day_num, status_code, ot_hours FROM hr_salary_days WHERE period_id=%s",
                (p['id'],)
            ).fetchall()]
            c = _calc_from_period_and_days(p, day_rows)
            sr += 1
            ws.append([
                sr, e['emp_id'], e['emp_name'], e.get('department') or '',
                e.get('designation') or '',
                c['BF_actual_days'], c['AO_total_present'], c['AP_paid_holiday'],
                c['AR_pl_used'], c['AS_cl_used'], c['AT_sl_used'],
                c['AU_total_wo'], c['AV_total_wop'], c['AW_total_php'],
                c['BD_ot_hours'], c['BE_ot_days'], c['BG_total_days'],
                c['BH_gross'], c['BI_basic'], c['BJ_hra'], c['BK_conveyance'],
                c['BL_medical'], c['BM_other_allow'],
                c['BN_earned_basic'], c['BO_earned_hra'], c['BP_earned_conv'],
                c['BQ_earned_medical'], c['BR_earned_other'],
                c['BS_prod_incentive'], c['BT_gross_earning'], c['BU_arrear'],
                c['BV_total_earned'],
                c['BW_loan_amount'], c['BY_loan_deduction'], c['BX_balance_loan'],
                c['BZ_other_deduction'],
                c['CA_pt'], c['CB_pf'], c['CC_tds'],
                c['CD_total_deduction'], c['CE_paid_salary'],
            ])
        conn.close()

        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', start_color='1E3A8A')
        for col_letter in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col_letter].width = 18
        ws.freeze_panes = 'C2'

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(tmp.name); tmp.close()
        return send_file(tmp.name,
                         as_attachment=True,
                         download_name=f'HR_Salary_{month}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 2 — Bulk Excel import of employees
# ═══════════════════════════════════════════════════════════════════════════════

IMPORT_COLUMNS = [
    ('emp_id',           'EMP ID',           'str',   True),
    ('emp_name',         'EMPLOYEE NAME',    'str',   True),
    ('gender',           'GENDER',           'str',   False),
    ('doj',              'DOJ',              'date',  False),
    ('designation',      'DESIGNATION',      'str',   False),
    ('department',       'DEPARTMENT',       'str',   False),
    ('emp_status',       'STATUS',           'str',   False),
    ('mobile',           'MOBILE',           'str',   False),
    ('email',            'EMAIL',            'str',   False),
    ('gross_salary',     'GROSS SALARY',     'num',   False),
    ('medical_fixed',    'MEDICAL (FIXED)',  'num',   False),
    ('loan_outstanding', 'LOAN OUTSTANDING', 'num',   False),
    ('opening_pl',       'OPENING PL',       'num',   False),
    ('opening_cl',       'OPENING CL',       'num',   False),
    ('opening_sl',       'OPENING SL',       'num',   False),
    ('pf_applicable',    'PF APPLICABLE',    'bool',  False),
]


@hr_salary_bp.route('/api/hr_salary/import_template')
@hr_required
def api_import_template():
    """Download a blank import template with headers + legend."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Employees'
    headers = [h for _, h, _, _ in IMPORT_COLUMNS]
    ws.append(headers)

    # Styling
    thin = Side(style='thin', color='CBD5E1')
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', start_color='1E3A8A')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ws.column_dimensions[cell.column_letter].width = max(16, len(cell.value or '') + 2)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'

    # Example row
    ws.append([
        '1001', 'RAMESH KUMAR SHARMA', 'MALE', '2021-07-01',
        'FOREMAN', 'PRODUCTION', 'PERMANENT', '9876543210', 'ramesh@example.com',
        25000, 1200, 0, 5, 1, 0,
    ])
    for cell in ws[2]:
        cell.font = Font(color='94A3B8', italic=True)
        cell.alignment = Alignment(horizontal='center')

    # Legend / instructions sheet
    ws2 = wb.create_sheet('Instructions')
    legend = [
        ['HR Salary — Employee Import Template'],
        [''],
        ['REQUIRED COLUMNS', 'EMP ID and EMPLOYEE NAME are mandatory.'],
        [''],
        ['DOJ format', 'YYYY-MM-DD (e.g. 2021-07-01). Leave blank if unknown.'],
        ['STATUS', 'PERMANENT / CONTRACT / TRAINEE / PROBATION'],
        ['GENDER', 'MALE / FEMALE / OTHER'],
        ['MOBILE', '10 digits (India default) or full international (e.g. 919876543210).'],
        ['EMAIL', 'Valid email for payslip delivery. Optional but needed for Email Slips feature.'],
        ['PF APPLICABLE', '1 / Y / YES / TRUE if employee is covered by EPF (default). 0 / N / NO if exempt.'],
        ['Numeric fields', 'GROSS SALARY, MEDICAL, LOAN OUTSTANDING, OPENING PL/CL/SL.'],
        [''],
        ['BEHAVIOUR', 'Existing emp_ids are UPDATED; new ones are INSERTED.'],
        ['', 'Soft-deleted employees (is_active=0) are re-activated on re-import.'],
        [''],
        ['STATUS CODES USED IN DAILY ATTENDANCE (not here)',
         'P, WO, WOP, PH, PHP, PL, CL, SL, HD, L, A, PL-HD, CL-HD, SL-HD, WO-HD, PH-HD'],
    ]
    for row in legend:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 80
    ws2['A1'].font = Font(bold=True, size=14, color='1E3A8A')

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name); tmp.close()
    return send_file(tmp.name,
                     as_attachment=True,
                     download_name='HR_Employee_Import_Template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@hr_salary_bp.route('/api/hr_salary/import', methods=['POST'])
@hr_required
def api_import():
    """
    Parse uploaded xlsx and upsert employees.
    Returns: { inserted, updated, skipped, errors[] }
    """
    try:
        f = request.files.get('file')
        if not f:
            return jsonify({'status': 'error', 'message': 'no file uploaded'}), 400

        wb = load_workbook(BytesIO(f.read()), data_only=True, read_only=True)
        # Use first sheet whose name starts with 'Employees', else first sheet
        ws = None
        for name in wb.sheetnames:
            if name.strip().lower().startswith('employee'):
                ws = wb[name]; break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return jsonify({'status': 'error', 'message': 'file is empty'}), 400

        # Map header text -> column index (case/space-insensitive)
        norm = lambda s: re.sub(r'\s+', ' ', str(s or '').strip().upper())
        header_map = {norm(h): i for i, h in enumerate(header_row)}

        # Resolve each logical column index
        col_idx = {}
        missing_required = []
        for field, label, _kind, required in IMPORT_COLUMNS:
            idx = header_map.get(norm(label))
            if idx is None and required:
                missing_required.append(label)
            col_idx[field] = idx
        if missing_required:
            return jsonify({
                'status': 'error',
                'message': 'missing required columns: ' + ', '.join(missing_required),
            }), 400

        def _get(row, field, kind):
            idx = col_idx.get(field)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            if v is None or (isinstance(v, str) and not v.strip()):
                return None
            if kind == 'str':
                return str(v).strip()
            if kind == 'num':
                try:
                    return float(v)
                except Exception:
                    return None
            if kind == 'date':
                if isinstance(v, datetime):
                    return v.strftime('%Y-%m-%d')
                s = str(v).strip()
                # try common formats
                for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
                    try:
                        return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
                    except Exception:
                        continue
                return None
            if kind == 'bool':
                # Accept 1/0, Y/N, YES/NO, TRUE/FALSE (case-insensitive)
                s = str(v).strip().upper()
                if s in ('1', 'Y', 'YES', 'TRUE', 'T', 'YEP', 'APPLICABLE'):
                    return 1
                if s in ('0', 'N', 'NO', 'FALSE', 'F', 'EXEMPT', 'NOT APPLICABLE'):
                    return 0
                return None
            return v

        conn = sampling_portal.get_db_connection()
        inserted = updated = skipped = 0
        errors = []
        row_num = 1  # header was row 1

        for row in rows_iter:
            row_num += 1
            if row is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue

            emp_id   = _get(row, 'emp_id',   'str')
            emp_name = _get(row, 'emp_name', 'str')
            if not emp_id or not emp_name:
                skipped += 1
                errors.append(f'Row {row_num}: emp_id and emp_name required — skipped')
                continue

            pf_raw = _get(row, 'pf_applicable', 'bool')
            pf_val = 1 if pf_raw is None else int(pf_raw)  # default ON if blank/missing

            payload = (
                emp_id, emp_name,
                _get(row, 'gender',           'str')  or '',
                _get(row, 'doj',              'date') or None,
                _get(row, 'designation',      'str')  or '',
                _get(row, 'department',       'str')  or '',
                _get(row, 'emp_status',       'str')  or 'PERMANENT',
                _get(row, 'mobile',           'str')  or '',
                _get(row, 'email',            'str')  or '',
                _get(row, 'gross_salary',     'num')  or 0,
                _get(row, 'medical_fixed',    'num')  or 1200,
                _get(row, 'loan_outstanding', 'num')  or 0,
                _get(row, 'opening_pl',       'num')  or 0,
                _get(row, 'opening_cl',       'num')  or 0,
                _get(row, 'opening_sl',       'num')  or 0,
                pf_val,
            )

            try:
                existing = conn.execute(
                    "SELECT id FROM hr_employees WHERE emp_id=%s", (emp_id,)
                ).fetchone()
                if existing:
                    conn.execute("""
                        UPDATE hr_employees
                           SET emp_name=%s, gender=%s, doj=%s, designation=%s,
                               department=%s, emp_status=%s, mobile=%s, email=%s,
                               gross_salary=%s, medical_fixed=%s, loan_outstanding=%s,
                               opening_pl=%s, opening_cl=%s, opening_sl=%s,
                               pf_applicable=%s,
                               is_active=1
                         WHERE emp_id=%s
                    """, payload[1:] + (emp_id,))
                    updated += 1
                else:
                    conn.execute("""
                        INSERT INTO hr_employees
                            (emp_id, emp_name, gender, doj, designation, department,
                             emp_status, mobile, email, gross_salary, medical_fixed,
                             loan_outstanding, opening_pl, opening_cl, opening_sl,
                             pf_applicable)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, payload)
                    inserted += 1
            except Exception as row_err:
                skipped += 1
                errors.append(f'Row {row_num} ({emp_id}): {row_err}')

        conn.commit()
        conn.close()
        return jsonify({
            'status': 'ok',
            'inserted': inserted,
            'updated': updated,
            'skipped': skipped,
            'errors': errors[:50],  # cap to avoid huge payloads
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 2b — Bulk Excel import of daily attendance
#
# Supports TWO layouts (auto-detected from header row):
#
#   WIDE (one row per employee, columns 1..31 are days):
#       EMP ID | EMP NAME | 1 | 2 | 3 | ... | 31
#          123 | X Y      | P |WO | P | ... | P
#     Month must be supplied (form field `month=YYYY-MM`) or inferred
#     from sheet name if sheet name is YYYY-MM / Mon-YY / "March 2026".
#     OT row is OPTIONAL: a second row per employee with EMP NAME == "OT"
#     or with status like "OT_HRS" — gives per-day OT hours.
#
#   LONG (one row per emp-day):
#       EMP ID | DATE       | STATUS | OT_HOURS | IN_TIME | OUT_TIME | HOURS
#          123 | 2026-03-01 | P      |        2 |   08:40 |    18:00 |
#     Month is derived per-row from DATE. Can span multiple months in one file.
#
# Biometric-only fallback (if STATUS column is blank / absent):
#     - IN_TIME + OUT_TIME both present  → status = P, OT = max(0, hrs-8) rounded
#     - Either blank                      → status = A (absent)
#
# `overwrite` form flag:
#     - True  → replace existing day codes
#     - False → keep existing, skip incoming for days that already have a code
# ═══════════════════════════════════════════════════════════════════════════════

# Status codes valid in the attendance table (kept in sync with workbook COUNTIFs)
VALID_ATTENDANCE_CODES = {
    'P', 'WO', 'WOP', 'PH', 'PHP', 'PL', 'CL', 'SL', 'HD', 'L', 'A',
    'PL-HD', 'CL-HD', 'SL-HD', 'WO-HD', 'PH-HD'
}


def _derive_month_from_sheet_name(name):
    """Attempt to extract YYYY-MM from a sheet name like 'March 2026', '2026-03', 'Mar-26'."""
    if not name:
        return None
    s = str(name).strip()
    # YYYY-MM directly
    m = re.match(r'^(20\d{2})[-_\s]?(\d{1,2})$', s)
    if m:
        yr = int(m.group(1)); mo = int(m.group(2))
        if 1 <= mo <= 12:
            return f'{yr:04d}-{mo:02d}'
    # Month-YY / Month YYYY / Month-YYYY
    months = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,
              'JUL':7,'AUG':8,'SEP':9,'SEPT':9,'OCT':10,'NOV':11,'DEC':12}
    su = s.upper()
    for name_, mo in months.items():
        m = re.match(rf'^{name_}[A-Z]*[-_\s]*(\d{{2,4}})', su)
        if m:
            yr_raw = int(m.group(1))
            yr = yr_raw if yr_raw >= 100 else 2000 + yr_raw
            return f'{yr:04d}-{mo:02d}'
    return None


def _parse_biometric_time(v):
    """
    Accept '08:40', '8:40', datetime.time, datetime.datetime, 0.36 (Excel fraction of day),
    '08:40:00'. Returns hours-as-float or None.
    """
    if v is None or v == '':
        return None
    # datetime.time
    try:
        from datetime import time as _dt_time
        if isinstance(v, _dt_time):
            return v.hour + v.minute/60 + v.second/3600
    except Exception:
        pass
    # datetime
    if isinstance(v, datetime):
        return v.hour + v.minute/60 + v.second/3600
    # Excel serial fraction (0..1)
    if isinstance(v, (int, float)):
        n = float(v)
        if 0 <= n < 2:   # treat fractions as day fraction
            return n * 24
        return None
    # String 'HH:MM' or 'HH:MM:SS'
    s = str(v).strip()
    if not s or s == '0' or s.upper() == 'NA':
        return None
    m = re.match(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?$', s)
    if m:
        return int(m.group(1)) + int(m.group(2))/60 + int(m.group(3) or 0)/3600
    return None


def _derive_status_from_bio(in_hrs, out_hrs):
    """Return (status, ot_hours) from IN/OUT biometric clock times (hours-of-day floats)."""
    if in_hrs is None or out_hrs is None:
        return 'A', 0
    worked = out_hrs - in_hrs
    if worked <= 0:
        return 'A', 0
    if worked < 4:
        return 'HD', 0           # under 4 hours → half day
    ot = max(0, int(round(worked - 8)))
    return 'P', ot


def _normalize_status_code(raw):
    """Uppercase + strip; return code if valid, else None."""
    if raw is None:
        return None
    s = str(raw).strip().upper()
    if not s:
        return None
    s = s.replace(' ', '')   # 'PL HD' → 'PLHD' … nope, keep hyphen
    # Allow both 'PL-HD' and 'PLHD'
    variants = {s, s.replace('HD', '-HD') if 'HD' in s and '-' not in s else s}
    for v in variants:
        if v in VALID_ATTENDANCE_CODES:
            return v
    # Common aliases
    aliases = {'PRESENT':'P', 'ABSENT':'A', 'LEAVE':'L',
               'HOLIDAY':'PH', 'HALFDAY':'HD', 'WEEKOFF':'WO',
               'OFF':'WO', 'WEEK-OFF':'WO'}
    return aliases.get(s)


def _ensure_period(conn, emp_row_id, period_month, emp_master=None):
    """Return (period_id, last_day). Creates period row if missing, seeded from emp master."""
    p = conn.execute(
        "SELECT id, actual_days FROM hr_salary_periods WHERE emp_row_id=%s AND period_month=%s",
        (emp_row_id, period_month)
    ).fetchone()
    if p:
        p = dict(p)
        return p['id'], int(p.get('actual_days') or _days_in_month(period_month))
    last_day = _days_in_month(period_month)
    seed = emp_master or {}
    conn.execute("""
        INSERT INTO hr_salary_periods
            (emp_row_id, period_month, actual_days, gross_salary,
             medical_fixed, ot_hours, arrear, loan_amount,
             loan_deduction, other_deduction, tds,
             opening_pl, opening_cl, opening_sl, remarks)
        VALUES (%s,%s,%s,%s,%s,0,0,%s,0,0,0,%s,%s,%s,'')
    """, (emp_row_id, period_month, last_day,
          seed.get('gross_salary') or 0,
          seed.get('medical_fixed') or 1200,
          seed.get('loan_outstanding') or 0,
          seed.get('opening_pl') or 0,
          seed.get('opening_cl') or 0,
          seed.get('opening_sl') or 0))
    pid = conn.execute(
        "SELECT id FROM hr_salary_periods WHERE emp_row_id=%s AND period_month=%s",
        (emp_row_id, period_month)
    ).fetchone()
    return dict(pid)['id'], last_day


@hr_salary_bp.route('/api/hr_salary/attendance/import_template')
@hr_required
def api_attendance_import_template():
    """Download a template workbook with WIDE, LONG, and Instructions sheets."""
    wb = Workbook()

    # Sheet 1: WIDE layout
    ws = wb.active
    ws.title = 'Wide (days 1-31)'
    wide_hdr = ['EMP ID', 'EMP NAME'] + [str(d) for d in range(1, 32)]
    ws.append(wide_hdr)
    # Example rows
    sample_codes_1 = ['P','P','WO','PH','P','P','P','P','P','WO',
                      'P','P','P','P','P','P','WO','P','P','P',
                      'P','P','P','WO','P','P','P','CL','P','P','WO']
    sample_codes_2 = ['P','P','WO','PH','P','SL-HD','P','P','P','WO',
                      'P','P','P','P','P','P','WO','P','P','P',
                      'P','P','P','WO','P','P','P','P','P','P','WO']
    ws.append(['1001', 'RAMESH KUMAR SHARMA'] + sample_codes_1)
    ws.append(['1002', 'MEENA DEVI JOSHI']    + sample_codes_2)

    # Sheet 2: LONG layout
    ws2 = wb.create_sheet('Long (per-day rows)')
    ws2.append(['EMP ID', 'DATE', 'STATUS', 'OT_HOURS', 'IN_TIME', 'OUT_TIME'])
    # A handful of example rows showing all three patterns
    ws2.append(['1001', '2026-03-01', 'P',  2, '08:40', '18:00'])
    ws2.append(['1001', '2026-03-02', 'P',  0, '08:38', '17:35'])
    ws2.append(['1001', '2026-03-03', 'WO', 0, '',      ''     ])
    ws2.append(['1002', '2026-03-01', '',   0, '08:45', '17:30'])   # status blank → derive from IN/OUT
    ws2.append(['1002', '2026-03-02', '',   0, '',      ''     ])   # both blank → A

    # Sheet 3: Instructions
    ws3 = wb.create_sheet('Instructions')
    rows = [
        ['HR Attendance — Bulk Import Template'],
        [''],
        ['LAYOUTS', 'You can use EITHER Wide OR Long — pick whichever matches your source.'],
        [''],
        ['WIDE', 'One row per employee; columns 1, 2, 3 ... 31 are day codes.'],
        ['',     'Month must be selected in the upload dialog (or sheet name may be "2026-03" / "March 2026").'],
        [''],
        ['LONG', 'One row per (employee, date). Month is derived per row.'],
        ['',     'Columns: EMP ID, DATE (YYYY-MM-DD), STATUS, OT_HOURS, IN_TIME, OUT_TIME.'],
        [''],
        ['STATUS CODES',
         'P · WO · WOP · PH · PHP · PL · CL · SL · HD · L · A · PL-HD · CL-HD · SL-HD · WO-HD · PH-HD'],
        [''],
        ['BIOMETRIC FALLBACK',
         'If STATUS is blank and IN_TIME + OUT_TIME are both filled → status = P with OT = max(0, hours-8).'],
        ['', 'If both IN_TIME and OUT_TIME are blank → status = A.'],
        ['', 'If worked hours < 4 → status = HD (half day).'],
        [''],
        ['OVERWRITE',
         'The upload dialog has an "Overwrite existing" checkbox.'],
        ['', 'Unchecked (safe mode): existing day codes are kept, only blank days are filled.'],
        ['', 'Checked: incoming value replaces whatever was there.'],
        [''],
        ['AUTO-CREATION',
         'If an employee has no salary period row for the imported month, one is created '
         'automatically (seeded from employee master: gross, medical, leave openings).'],
    ]
    for row in rows:
        ws3.append(row)

    # Styling — headers on every sheet
    for _ws in (ws, ws2):
        for cell in _ws[1]:
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill('solid', start_color='1E3A8A')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        _ws.row_dimensions[1].height = 28
        _ws.freeze_panes = 'C2' if _ws is ws else 'B2'
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 32
    for col_letter in [c for c in 'CDEFGHIJKLMNOPQRSTUVWXYZ'] + ['AA','AB','AC','AD','AE','AF','AG']:
        ws.column_dimensions[col_letter].width = 6
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 10
    ws2.column_dimensions['F'].width = 10
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 90
    ws3['A1'].font = Font(bold=True, size=14, color='1E3A8A')

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name); tmp.close()
    return send_file(tmp.name, as_attachment=True,
                     download_name='HR_Attendance_Import_Template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _detect_layout(header_row):
    """
    Return 'wide', 'long', or None based on header contents.
    WIDE: has EMP ID + at least one numeric day column (1..31).
    LONG: has EMP ID + DATE (any case).
    """
    norm = [re.sub(r'\s+', ' ', str(h or '').strip().upper()) for h in header_row]
    hdr_set = set(norm)
    has_emp = any(h in hdr_set for h in ('EMP ID', 'EMP_ID', 'EMPLOYEE ID', 'EMPID'))
    if not has_emp:
        return None
    has_date   = any(h in hdr_set for h in ('DATE', 'ATTENDANCE DATE', 'ATT DATE'))
    day_cols   = [h for h in norm if re.match(r'^(3[01]|[12]?\d)$', h)]
    # If both present, LONG wins (more specific)
    if has_date:
        return 'long'
    if len(day_cols) >= 7:   # at least a week of day columns is a clear signal
        return 'wide'
    return None


def _header_index(norm_headers, *candidates):
    """Return the first index whose header (upper-normalised) matches any candidate, else -1."""
    for i, h in enumerate(norm_headers):
        for c in candidates:
            if h == c:
                return i
    return -1


@hr_salary_bp.route('/api/hr_salary/attendance/import', methods=['POST'])
@hr_required
def api_attendance_import():
    """
    Parse attendance xlsx and write to hr_salary_days (+ auto-create period rows).

    Form-data:
        file       : the .xlsx
        month      : fallback YYYY-MM (used by WIDE layout if sheet name doesn't encode it)
        overwrite  : "1" or "0" — whether to replace existing codes

    Response JSON:
        { status, inserted, updated, skipped_existing, skipped_unknown_emp,
          employees_touched, months_touched, errors[] }
    """
    try:
        f = request.files.get('file')
        if not f:
            return jsonify({'status': 'error', 'message': 'no file uploaded'}), 400

        month_fallback = (request.form.get('month') or '').strip()
        overwrite = (request.form.get('overwrite') or '0').strip() in ('1', 'true', 'True', 'on', 'yes')

        wb = load_workbook(BytesIO(f.read()), data_only=True, read_only=True)

        # Build emp_id -> (emp_row_id, master_dict) lookup once
        conn = sampling_portal.get_db_connection()
        emp_rows = conn.execute(
            "SELECT * FROM hr_employees WHERE is_active=1"
        ).fetchall()
        emp_lookup = {}
        for r in emp_rows:
            d = dict(r)
            emp_lookup[str(d['emp_id']).strip()] = d

        # Accumulators
        inserted = 0
        updated = 0
        skipped_existing = 0
        skipped_unknown = 0
        errors = []
        employees_touched = set()
        months_touched = set()
        # Cache of (emp_row_id, period_month) -> (period_id, {day_num: row})
        existing_map_cache = {}
        # Reverse lookup: emp_row_id -> master dict (built once)
        emp_by_row_id = {m['id']: m for m in emp_lookup.values()}

        def _load_existing(emp_row_id, period_month):
            key = (emp_row_id, period_month)
            if key in existing_map_cache:
                return existing_map_cache[key]
            pid, _ = _ensure_period(conn, emp_row_id, period_month,
                                    emp_by_row_id.get(emp_row_id))
            rows = conn.execute(
                "SELECT id, day_num, status_code, ot_hours FROM hr_salary_days WHERE period_id=%s",
                (pid,)
            ).fetchall()
            dmap = {int(dict(r)['day_num']): dict(r) for r in rows}
            existing_map_cache[key] = (pid, dmap)
            return pid, dmap

        def _write_day(emp_row_id, period_month, day_num, status_code, ot_hours):
            """Insert or update one day. Respects the overwrite flag. Returns 'ins'|'upd'|'skip'."""
            nonlocal inserted, updated, skipped_existing
            pid, dmap = _load_existing(emp_row_id, period_month)
            existing = dmap.get(int(day_num))
            if existing and existing.get('id', -1) > 0:
                # Real DB row — decide whether to update or skip
                existing_code = (existing.get('status_code') or '').strip()
                if existing_code and not overwrite:
                    skipped_existing += 1
                    return 'skip'
                conn.execute("""
                    UPDATE hr_salary_days
                       SET status_code=%s, ot_hours=%s
                     WHERE id=%s
                """, (status_code, ot_hours, existing['id']))
                dmap[int(day_num)] = {**existing, 'status_code': status_code, 'ot_hours': ot_hours}
                updated += 1
                return 'upd'
            # Either no row yet, or only an in-memory placeholder from a previous iteration
            if existing and existing.get('id', -1) <= 0 and not overwrite:
                # Same run already wrote this day; keep the first write (safe mode)
                skipped_existing += 1
                return 'skip'
            if existing and existing.get('id', -1) <= 0 and overwrite:
                # Overwrite the in-memory placeholder: still need a real DB write,
                # but we should DELETE the earlier INSERT to avoid duplicate-key.
                # Simplest correct behaviour: UPDATE via unique index (period_id, day_num).
                conn.execute("""
                    UPDATE hr_salary_days
                       SET status_code=%s, ot_hours=%s
                     WHERE period_id=%s AND day_num=%s
                """, (status_code, ot_hours, pid, int(day_num)))
                dmap[int(day_num)] = {'id': 0, 'day_num': int(day_num),
                                      'status_code': status_code, 'ot_hours': ot_hours}
                updated += 1
                return 'upd'
            # Genuine new insert
            conn.execute("""
                INSERT INTO hr_salary_days (period_id, day_num, status_code, ot_hours)
                VALUES (%s,%s,%s,%s)
            """, (pid, int(day_num), status_code, ot_hours))
            dmap[int(day_num)] = {'id': 0, 'day_num': int(day_num),
                                  'status_code': status_code, 'ot_hours': ot_hours}
            inserted += 1
            return 'ins'

        # ── Iterate every sheet; skip Instructions / empty sheets ──
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() == 'instructions':
                continue
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                continue
            if not header or not any(h not in (None, '') for h in header):
                continue

            layout = _detect_layout(header)
            if layout is None:
                errors.append(f"Sheet '{sheet_name}': unrecognised header — skipped")
                continue

            norm = [re.sub(r'\s+', ' ', str(h or '').strip().upper()) for h in header]

            if layout == 'wide':
                # Determine the month: form field > sheet name hint
                month = month_fallback or _derive_month_from_sheet_name(sheet_name)
                if not month:
                    errors.append(
                        f"Sheet '{sheet_name}': WIDE layout needs month "
                        "(supply via form field or name the sheet like '2026-03')"
                    )
                    continue
                if not re.match(r'^20\d{2}-(0[1-9]|1[0-2])$', month):
                    errors.append(f"Sheet '{sheet_name}': invalid month '{month}'")
                    continue

                emp_idx = _header_index(norm, 'EMP ID', 'EMP_ID', 'EMPLOYEE ID', 'EMPID')
                # Build map: column-index -> day_num (1..31)
                day_col_map = {}
                for i, h in enumerate(norm):
                    m = re.match(r'^(3[01]|[12]?\d)$', h)
                    if m:
                        dn = int(m.group(1))
                        if 1 <= dn <= 31:
                            day_col_map[i] = dn

                # Track per-employee "OT row" pairings: we look for a row where the cell in
                # EMP NAME column says "OT" / "OT_HRS" with the same EMP ID
                row_num = 1
                # Accumulate all rows first; group by emp_id
                ws_rows = list(rows_iter)
                # Tag each row with (row_num, raw_row)
                tagged = [(row_num + 1 + i, r) for i, r in enumerate(ws_rows)]
                # Group consecutively by emp_id when possible
                for rnum, r in tagged:
                    if r is None or all(c in (None, '') for c in r):
                        continue
                    emp_id_val = r[emp_idx] if emp_idx >= 0 and emp_idx < len(r) else None
                    emp_id = str(emp_id_val).strip() if emp_id_val not in (None, '') else None
                    if not emp_id:
                        # Silent skip — blank separator row
                        continue
                    if emp_id not in emp_lookup:
                        skipped_unknown += 1
                        errors.append(f"Sheet '{sheet_name}' row {rnum}: emp_id '{emp_id}' not found")
                        continue
                    emp_master = emp_lookup[emp_id]
                    employees_touched.add(emp_id)
                    months_touched.add(month)

                    # Write each day column
                    for col_idx, day_num in day_col_map.items():
                        raw = r[col_idx] if col_idx < len(r) else None
                        code = _normalize_status_code(raw)
                        if not code:
                            # In WIDE layout we only write cells that have a recognised code
                            if raw not in (None, ''):
                                errors.append(
                                    f"Sheet '{sheet_name}' row {rnum} day {day_num}: "
                                    f"unknown status '{raw}' — skipped"
                                )
                            continue
                        try:
                            _write_day(emp_master['id'], month, day_num, code, 0)
                        except Exception as ex:
                            errors.append(f"Sheet '{sheet_name}' row {rnum} day {day_num}: {ex}")

            else:  # LONG
                emp_idx    = _header_index(norm, 'EMP ID', 'EMP_ID', 'EMPLOYEE ID', 'EMPID')
                date_idx   = _header_index(norm, 'DATE', 'ATTENDANCE DATE', 'ATT DATE')
                status_idx = _header_index(norm, 'STATUS', 'STATUS CODE')
                ot_idx     = _header_index(norm, 'OT_HOURS', 'OT HOURS', 'OT', 'OT(HRS)', 'OVERTIME HOURS')
                in_idx     = _header_index(norm, 'IN_TIME', 'IN TIME', 'IN', 'CHECK IN', 'CHECKIN')
                out_idx    = _header_index(norm, 'OUT_TIME', 'OUT TIME', 'OUT', 'CHECK OUT', 'CHECKOUT')

                row_num = 1
                for r in rows_iter:
                    row_num += 1
                    if r is None or all(c in (None, '') for c in r):
                        continue
                    emp_id_val = r[emp_idx] if emp_idx >= 0 and emp_idx < len(r) else None
                    emp_id = str(emp_id_val).strip() if emp_id_val not in (None, '') else None
                    if not emp_id:
                        continue
                    if emp_id not in emp_lookup:
                        skipped_unknown += 1
                        errors.append(f"Sheet '{sheet_name}' row {row_num}: emp_id '{emp_id}' not found")
                        continue

                    # Date parse
                    date_val = r[date_idx] if date_idx >= 0 and date_idx < len(r) else None
                    parsed_date = None
                    if isinstance(date_val, datetime):
                        parsed_date = date_val
                    elif date_val not in (None, ''):
                        s = str(date_val).strip()
                        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d',
                                    '%d-%b-%Y', '%d %b %Y', '%d-%b-%y'):
                            try:
                                parsed_date = datetime.strptime(s, fmt); break
                            except Exception:
                                continue
                    if not parsed_date:
                        errors.append(f"Sheet '{sheet_name}' row {row_num}: unparseable date '{date_val}'")
                        continue
                    period_month = parsed_date.strftime('%Y-%m')
                    day_num = parsed_date.day

                    # Status
                    raw_status = r[status_idx] if status_idx >= 0 and status_idx < len(r) else None
                    code = _normalize_status_code(raw_status)

                    # OT
                    ot_raw = r[ot_idx] if ot_idx >= 0 and ot_idx < len(r) else None
                    try:
                        ot_hours = float(ot_raw) if ot_raw not in (None, '') else 0
                    except Exception:
                        ot_hours = 0

                    # Biometric fallback if status absent
                    if code is None:
                        in_raw  = r[in_idx]  if in_idx  >= 0 and in_idx  < len(r) else None
                        out_raw = r[out_idx] if out_idx >= 0 and out_idx < len(r) else None
                        in_hrs  = _parse_biometric_time(in_raw)
                        out_hrs = _parse_biometric_time(out_raw)
                        derived_code, derived_ot = _derive_status_from_bio(in_hrs, out_hrs)
                        code = derived_code
                        # If user put an explicit OT use that; else use derived
                        if not ot_hours and derived_ot:
                            ot_hours = derived_ot

                    if not code:
                        continue  # unrecognised and no bio data → skip silently

                    emp_master = emp_lookup[emp_id]
                    employees_touched.add(emp_id)
                    months_touched.add(period_month)

                    try:
                        _write_day(emp_master['id'], period_month, day_num, code, ot_hours)
                    except Exception as ex:
                        errors.append(f"Sheet '{sheet_name}' row {row_num}: {ex}")

        conn.commit()
        conn.close()
        return jsonify({
            'status': 'ok',
            'inserted': inserted,
            'updated': updated,
            'skipped_existing': skipped_existing,
            'skipped_unknown_emp': skipped_unknown,
            'employees_touched': len(employees_touched),
            'months_touched':    sorted(months_touched),
            'overwrite_mode':    overwrite,
            'errors':            errors[:100],
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 2c — Email salary slips via global SMTP
#
# Credential model
# ----------------
# A single row in `hr_smtp_config` (id=1) is set once by an admin.
# Password is XOR-obfuscated + base64-encoded before saving — same pattern
# already used by the Tally credential store in app.py.
#
# Routes
# ------
#   GET  /api/hr_salary/smtp/config       load current settings (password masked)
#   POST /api/hr_salary/smtp/config       save / update
#   POST /api/hr_salary/smtp/test         send a one-off test message to any address
#   POST /api/hr_salary/email/send        send payslips (single or bulk)
#
# The send route reuses `_build_slip_text()` + `_build_slip_pdf_bytes()`
# so the body / attachment are byte-identical to the WhatsApp slip + ZIP.
# ═══════════════════════════════════════════════════════════════════════════════

_HR_SMTP_KEY = b'hcp_hr_smtp_key_2025'   # XOR key — mirrors the Tally cred pattern


def _hr_obfuscate(text):
    """XOR + base64 encode. Empty-in, empty-out."""
    if not text:
        return ''
    data = text.encode('utf-8')
    xored = bytes(b ^ _HR_SMTP_KEY[i % len(_HR_SMTP_KEY)] for i, b in enumerate(data))
    return base64.b64encode(xored).decode('ascii')


def _hr_deobfuscate(token):
    """Reverse of _hr_obfuscate. Returns '' on any decode error."""
    if not token:
        return ''
    try:
        xored = base64.b64decode(token.encode('ascii'))
        return bytes(
            b ^ _HR_SMTP_KEY[i % len(_HR_SMTP_KEY)] for i, b in enumerate(xored)
        ).decode('utf-8')
    except Exception:
        return ''


def _load_smtp_config(unmask=False):
    """
    Load SMTP config row. If `unmask=True`, the real password is returned
    (used only internally by the send functions, never returned to the browser).
    """
    conn = sampling_portal.get_db_connection()
    if not conn:
        return None
    try:
        row = conn.execute(
            "SELECT * FROM hr_smtp_config WHERE id=1"
        ).fetchone()
    finally:
        try: conn.close()
        except Exception: pass
    if not row:
        return None
    d = dict(row)
    if unmask:
        d['password'] = _hr_deobfuscate(d.pop('password_enc', '') or '')
    else:
        d['has_password'] = bool(d.pop('password_enc', '') or '')
    return d


def _smtp_send(cfg, to_addrs, subject, body_text, attachments=None, cc=None,
               bcc=None, reply_to=None):
    """
    Low-level sender. Expects cfg dict from _load_smtp_config(unmask=True).
    attachments: list of (filename, bytes, mimetype='application/pdf')
    Returns (ok, error_message_or_None).
    """
    if not cfg:
        return False, 'SMTP is not configured'
    host = (cfg.get('smtp_host') or '').strip()
    port = int(cfg.get('smtp_port') or 0)
    if not host or not port:
        return False, 'SMTP host/port missing'

    user = (cfg.get('username') or '').strip()
    pwd  = cfg.get('password') or ''
    from_email = (cfg.get('from_email') or user or '').strip()
    from_name  = (cfg.get('from_name')  or '').strip()
    if not from_email:
        return False, 'From-email not configured'

    if isinstance(to_addrs, str):
        to_addrs = [to_addrs]
    to_addrs = [a.strip() for a in to_addrs if a and a.strip()]
    if not to_addrs:
        return False, 'No recipient'

    msg = MIMEMultipart()
    msg['From'] = formataddr((from_name or 'HR Payroll', from_email))
    msg['To']   = ', '.join(to_addrs)
    if cc:
        msg['Cc'] = ', '.join(cc)
    rep_to = (reply_to or cfg.get('reply_to') or '').strip()
    if rep_to:
        msg['Reply-To'] = rep_to
    msg['Subject'] = subject
    msg['Date']    = formatdate(localtime=True)
    msg['Message-ID'] = make_msgid()
    msg.attach(MIMEText(body_text or '', 'plain', 'utf-8'))

    for att in (attachments or []):
        try:
            fname, payload, mimetype = att[0], att[1], (att[2] if len(att) > 2 else 'application/pdf')
        except Exception:
            continue
        if not payload:
            continue
        sub = mimetype.split('/', 1)[1] if '/' in mimetype else 'octet-stream'
        part = MIMEApplication(payload, _subtype=sub)
        part.add_header('Content-Disposition', 'attachment', filename=fname)
        msg.attach(part)

    all_rcpts = to_addrs + list(cc or []) + list(bcc or [])

    use_ssl = bool(cfg.get('use_ssl'))
    use_tls = bool(cfg.get('use_tls'))

    try:
        if use_ssl:
            ctx = ssl.create_default_context()
            server = smtplib.SMTP_SSL(host, port, timeout=30, context=ctx)
        else:
            server = smtplib.SMTP(host, port, timeout=30)
        try:
            server.ehlo()
            if use_tls and not use_ssl:
                server.starttls(context=ssl.create_default_context())
                server.ehlo()
            if user and pwd:
                server.login(user, pwd)
            server.sendmail(from_email, all_rcpts, msg.as_string())
        finally:
            try: server.quit()
            except Exception: pass
        return True, None
    except smtplib.SMTPAuthenticationError as e:
        return False, f'Auth failed: {e.smtp_error.decode() if isinstance(e.smtp_error, bytes) else e.smtp_error}'
    except smtplib.SMTPException as e:
        return False, f'SMTP error: {e}'
    except Exception as e:
        return False, f'Connection error: {e}'


def _clean_email(v):
    """Return a trimmed lower-cased email if it looks valid, else ''."""
    if not v:
        return ''
    s = str(v).strip().lower()
    # Very lenient — just check for '@' + '.' to the right
    m = re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', s)
    return s if m else ''


# ── CONFIG LOAD / SAVE ────────────────────────────────────────────────────────

@hr_salary_bp.route('/api/hr_salary/smtp/config')
@hr_required
def api_smtp_config_get():
    """Return current SMTP config (password masked)."""
    try:
        cfg = _load_smtp_config(unmask=False)
        if not cfg:
            cfg = {
                'smtp_host': '', 'smtp_port': 587, 'use_tls': 1, 'use_ssl': 0,
                'username': '', 'from_email': '', 'from_name': 'HR Payroll',
                'reply_to': '', 'has_password': False,
            }
        # Don't leak the encrypted string
        cfg.pop('password_enc', None)
        cfg.pop('password', None)
        return jsonify({'status': 'ok', 'config': cfg})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/smtp/config', methods=['POST'])
@hr_required
def api_smtp_config_save():
    """
    Upsert the singleton SMTP config row.
    Both admin and HR can save (full peers on this module).
    """
    try:
        d = request.get_json(force=True) or {}
        host = (d.get('smtp_host') or '').strip()
        port = int(d.get('smtp_port') or 0)
        if not host or port <= 0 or port > 65535:
            return jsonify({'status': 'error',
                            'message': 'smtp_host and valid smtp_port required'}), 400

        use_tls  = 1 if d.get('use_tls') else 0
        use_ssl  = 1 if d.get('use_ssl') else 0
        username = (d.get('username') or '').strip()
        from_email = (d.get('from_email') or '').strip()
        from_name  = (d.get('from_name')  or 'HR Payroll').strip()
        reply_to   = (d.get('reply_to')   or '').strip()

        if not _clean_email(from_email):
            return jsonify({'status': 'error',
                            'message': 'from_email must be a valid email address'}), 400

        raw_pwd = d.get('password')
        # If the client sends empty/None and a password already exists, keep it.
        conn = sampling_portal.get_db_connection()
        existing = conn.execute(
            "SELECT id, password_enc FROM hr_smtp_config WHERE id=1"
        ).fetchone()

        if raw_pwd is None or raw_pwd == '':
            password_enc = dict(existing).get('password_enc', '') if existing else ''
        else:
            password_enc = _hr_obfuscate(str(raw_pwd))

        if existing:
            conn.execute("""
                UPDATE hr_smtp_config SET
                    smtp_host=%s, smtp_port=%s, use_tls=%s, use_ssl=%s,
                    username=%s, password_enc=%s,
                    from_email=%s, from_name=%s, reply_to=%s
                 WHERE id=1
            """, (host, port, use_tls, use_ssl, username, password_enc,
                  from_email, from_name, reply_to))
        else:
            conn.execute("""
                INSERT INTO hr_smtp_config
                    (id, smtp_host, smtp_port, use_tls, use_ssl,
                     username, password_enc, from_email, from_name, reply_to)
                VALUES (1,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (host, port, use_tls, use_ssl, username, password_enc,
                  from_email, from_name, reply_to))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ── TEST SEND ─────────────────────────────────────────────────────────────────

@hr_salary_bp.route('/api/hr_salary/smtp/test', methods=['POST'])
@hr_required
def api_smtp_test():
    """Send a small test message to the 'to' address in the body."""
    try:
        d = request.get_json(force=True) or {}
        to = _clean_email(d.get('to'))
        if not to:
            return jsonify({'status': 'error',
                            'message': 'Provide a valid test recipient'}), 400
        cfg = _load_smtp_config(unmask=True)
        if not cfg:
            return jsonify({'status': 'error',
                            'message': 'SMTP is not configured yet'}), 400
        body = (
            "This is a test message from the HCP Wellness HR Payroll module.\n\n"
            f"Sent at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"Host: {cfg.get('smtp_host')}:{cfg.get('smtp_port')}\n"
            "If you received this, SMTP is working correctly."
        )
        ok, err = _smtp_send(cfg, [to], 'HR Payroll — SMTP Test', body)
        if ok:
            return jsonify({'status': 'ok'})
        return jsonify({'status': 'error', 'message': err or 'Send failed'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ── BULK SEND ─────────────────────────────────────────────────────────────────

@hr_salary_bp.route('/api/hr_salary/email/send', methods=['POST'])
@hr_required
def api_email_send():
    """
    Send salary-slip emails.

    Body:
        {
          month:   "YYYY-MM",
          emp_ids: ["1","2",...]      # omit or empty → ALL active employees
          cc:      "hr@company.com"   # optional
          attach_pdf: true            # default true
        }

    Returns:
        {
          status: "ok",
          sent: [{ emp_id, emp_name, to }, ...],
          skipped: [{ emp_id, emp_name, reason }, ...]
        }
    """
    try:
        d = request.get_json(force=True) or {}
        month = (d.get('month') or datetime.now().strftime('%Y-%m')).strip()
        requested = {str(x).strip() for x in (d.get('emp_ids') or []) if str(x).strip()}
        cc_raw = (d.get('cc') or '').strip()
        cc_list = [_clean_email(x) for x in re.split(r'[,;\s]+', cc_raw) if x.strip()]
        cc_list = [x for x in cc_list if x]
        attach_pdf = d.get('attach_pdf', True)

        cfg = _load_smtp_config(unmask=True)
        if not cfg:
            return jsonify({'status': 'error',
                            'message': 'SMTP is not configured yet'}), 400

        conn = sampling_portal.get_db_connection()
        emps = conn.execute("""
            SELECT * FROM hr_employees WHERE is_active=1
            ORDER BY CAST(emp_id AS UNSIGNED) ASC, emp_id ASC
        """).fetchall()
        conn.close()

        sent, skipped = [], []
        for r in emps:
            e = dict(r)
            if requested and e['emp_id'] not in requested:
                continue

            to = _clean_email(e.get('email'))
            if not to:
                skipped.append({'emp_id': e['emp_id'], 'emp_name': e['emp_name'],
                                'reason': 'No email on file'})
                continue

            # Load period + calc
            emp, calc = _load_calc_for_emp(e['emp_id'], month)
            if not emp or not calc:
                skipped.append({'emp_id': e['emp_id'], 'emp_name': e['emp_name'],
                                'reason': f'No salary data for {month}'})
                continue

            # Body = the same formatted slip used for WhatsApp
            body_text = _build_slip_text(emp, month, calc)
            # Strip WhatsApp-specific asterisks for email readability
            body_text_clean = body_text.replace('*', '')
            subject = f'Salary Slip — {month} — {emp.get("emp_name") or emp.get("emp_id")}'

            attachments = []
            if attach_pdf:
                try:
                    pdf = _build_slip_pdf_bytes(emp, month, calc)
                    safe = re.sub(r'[^A-Za-z0-9_\-]+', '_',
                                  str(emp.get('emp_name') or 'slip'))[:40]
                    fname = f'Salary_{emp["emp_id"]}_{safe}_{month}.pdf'
                    attachments.append((fname, pdf, 'application/pdf'))
                except Exception as pdf_err:
                    skipped.append({'emp_id': e['emp_id'], 'emp_name': e['emp_name'],
                                    'reason': f'PDF build failed: {pdf_err}'})
                    continue

            ok, err = _smtp_send(cfg, [to], subject, body_text_clean,
                                 attachments=attachments, cc=cc_list)
            if ok:
                sent.append({'emp_id': e['emp_id'], 'emp_name': e['emp_name'],
                             'to': to, 'paid_salary': calc.get('CE_paid_salary')})
            else:
                skipped.append({'emp_id': e['emp_id'], 'emp_name': e['emp_name'],
                                'reason': err or 'Send failed'})

        return jsonify({'status': 'ok', 'sent': sent, 'skipped': skipped})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 2d — Year-to-date view (Indian fiscal year: Apr → Mar)
#
# Two endpoints:
#   GET /api/hr_salary/ytd?emp_id=<id>&fy=<YYYY>        single employee, 12 months
#   GET /api/hr_salary/ytd_company?fy=<YYYY>            all active employees by month
#
# `fy` = starting calendar year of the fiscal. Example:
#        fy=2025   →   2025-04 … 2026-03  (FY 2025-26)
# ═══════════════════════════════════════════════════════════════════════════════

def _fy_months(fy_start_year):
    """Return the 12 YYYY-MM strings for a fiscal year starting April `fy_start_year`."""
    months = []
    for m in range(4, 13):   # Apr..Dec of fy_start_year
        months.append(f'{fy_start_year:04d}-{m:02d}')
    for m in range(1, 4):    # Jan..Mar of fy_start_year + 1
        months.append(f'{fy_start_year+1:04d}-{m:02d}')
    return months


def _fy_for_date(dt):
    """Given a datetime/date, return the fiscal-year starting calendar year (Apr→Mar)."""
    if dt.month >= 4:
        return dt.year
    return dt.year - 1


def _current_fy():
    return _fy_for_date(datetime.now())


def _ytd_rows_for_emp(conn, emp_id_str, fy_start_year):
    """
    Return (emp_master, months_data[], totals{}) for one employee across the FY.
    months_data is a list of 12 dicts, one per month, with calc fields or has_data=False.
    totals sums the key numeric fields.
    """
    emp_row = conn.execute(
        "SELECT * FROM hr_employees WHERE emp_id=%s AND is_active=1",
        (emp_id_str,)
    ).fetchone()
    if not emp_row:
        return None, [], {}
    emp = dict(emp_row)

    months = _fy_months(fy_start_year)
    out = []
    sums = {
        'BH_gross': 0, 'BT_gross_earning': 0, 'BS_prod_incentive': 0,
        'BU_arrear': 0, 'BV_total_earned': 0,
        'BY_loan_deduction': 0, 'BZ_other_deduction': 0,
        'CA_pt': 0, 'CB_pf': 0, 'CC_tds': 0,
        'CD_total_deduction': 0, 'CE_paid_salary': 0,
        'AO_total_present': 0, 'AP_paid_holiday': 0,
        'AR_pl_used': 0, 'AS_cl_used': 0, 'AT_sl_used': 0,
        'AU_total_wo': 0, 'BG_total_days': 0,
        'BD_ot_hours': 0, 'BE_ot_days': 0,
    }
    months_with_data = 0

    for period_month in months:
        p_row = conn.execute("""
            SELECT * FROM hr_salary_periods
            WHERE emp_row_id=%s AND period_month=%s
        """, (emp['id'], period_month)).fetchone()
        if not p_row:
            out.append({'period_month': period_month, 'has_data': False})
            continue
        p = dict(p_row)
        day_rows = [dict(r) for r in conn.execute(
            "SELECT day_num, status_code, ot_hours FROM hr_salary_days WHERE period_id=%s",
            (p['id'],)
        ).fetchall()]
        calc = _calc_from_period_and_days(p, day_rows)
        months_with_data += 1

        # Accumulate
        for k in sums:
            sums[k] += float(calc.get(k) or 0)

        out.append({
            'period_month': period_month,
            'has_data': True,
            **{k: calc.get(k) for k in sums},  # cherry-pick only fields we sum
            # Additional per-month context that callers may want
            'BF_actual_days': calc.get('BF_actual_days'),
            'BA_closing_pl': calc.get('BA_closing_pl'),
            'BB_closing_cl': calc.get('BB_closing_cl'),
            'BC_closing_sl': calc.get('BC_closing_sl'),
            'AX_opening_pl': calc.get('AX_opening_pl'),
            'AY_opening_cl': calc.get('AY_opening_cl'),
            'AZ_opening_sl': calc.get('AZ_opening_sl'),
        })

    totals = {**sums, 'months_with_data': months_with_data}
    return emp, out, totals


@hr_salary_bp.route('/api/hr_salary/ytd')
@hr_required
def api_ytd_employee():
    """Per-employee YTD across 12 fiscal-year months."""
    try:
        emp_id = (request.args.get('emp_id') or '').strip()
        fy_raw = (request.args.get('fy') or '').strip()
        fy = int(fy_raw) if fy_raw.isdigit() else _current_fy()
        if not emp_id:
            return jsonify({'status': 'error', 'message': 'emp_id required'}), 400

        conn = sampling_portal.get_db_connection()
        emp, months_data, totals = _ytd_rows_for_emp(conn, emp_id, fy)
        conn.close()
        if not emp:
            return jsonify({'status': 'error', 'message': 'employee not found'}), 404

        # Strip emp fields that aren't safe to ship (none here, just remove internal pk)
        safe_emp = {k: v for k, v in emp.items() if k != 'id'}
        if safe_emp.get('doj'):
            try: safe_emp['doj'] = str(safe_emp['doj'])
            except Exception: pass

        return jsonify({
            'status':       'ok',
            'fy':           fy,
            'fy_label':     f'FY {fy}-{str(fy+1)[-2:]}',
            'months':       _fy_months(fy),
            'employee':     safe_emp,
            'months_data':  months_data,
            'totals':       totals,
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/ytd_company')
@hr_required
def api_ytd_company():
    """
    Company-wide YTD. For each active employee return the 12-month totals.
    Also returns per-month company aggregate (for the chart on the YTD tab).
    """
    try:
        fy_raw = (request.args.get('fy') or '').strip()
        fy = int(fy_raw) if fy_raw.isdigit() else _current_fy()

        conn = sampling_portal.get_db_connection()
        emp_rows = conn.execute("""
            SELECT id, emp_id, emp_name, department, designation
            FROM hr_employees WHERE is_active=1
            ORDER BY CAST(emp_id AS UNSIGNED) ASC, emp_id ASC
        """).fetchall()

        months = _fy_months(fy)
        employees_out = []
        # Per-month company aggregates
        company_totals = {
            m: {
                'period_month': m,
                'gross_earning': 0,
                'prod_incentive': 0,
                'arrear': 0,
                'total_earned': 0,
                'total_deduction': 0,
                'paid_salary': 0,
                'pf': 0, 'pt': 0, 'tds': 0,
                'employees_paid': 0,
            } for m in months
        }
        grand = {
            'gross_earning': 0, 'total_earned': 0, 'total_deduction': 0,
            'paid_salary': 0, 'pf': 0, 'pt': 0, 'tds': 0,
            'employees_with_any_data': 0,
        }

        for row in emp_rows:
            emp = dict(row)
            emp_totals = {
                'total_earned': 0, 'total_deduction': 0, 'paid_salary': 0,
                'pf': 0, 'pt': 0, 'tds': 0, 'months_paid': 0,
            }
            emp_monthly = []
            had_any = False

            for m in months:
                p = conn.execute("""
                    SELECT * FROM hr_salary_periods WHERE emp_row_id=%s AND period_month=%s
                """, (emp['id'], m)).fetchone()
                if not p:
                    emp_monthly.append({'period_month': m, 'has_data': False})
                    continue
                p = dict(p)
                dr = [dict(r) for r in conn.execute(
                    "SELECT day_num, status_code, ot_hours FROM hr_salary_days WHERE period_id=%s",
                    (p['id'],)
                ).fetchall()]
                calc = _calc_from_period_and_days(p, dr)
                had_any = True

                emp_monthly.append({
                    'period_month':    m,
                    'has_data':        True,
                    'total_earned':    calc['BV_total_earned'],
                    'total_deduction': calc['CD_total_deduction'],
                    'paid_salary':     calc['CE_paid_salary'],
                })
                emp_totals['total_earned']    += calc['BV_total_earned']
                emp_totals['total_deduction'] += calc['CD_total_deduction']
                emp_totals['paid_salary']     += calc['CE_paid_salary']
                emp_totals['pf']              += calc['CB_pf']
                emp_totals['pt']              += calc['CA_pt']
                emp_totals['tds']             += calc['CC_tds']
                emp_totals['months_paid']     += 1

                # Company per-month
                ct = company_totals[m]
                ct['gross_earning']   += calc['BT_gross_earning']
                ct['prod_incentive']  += calc['BS_prod_incentive']
                ct['arrear']          += calc['BU_arrear']
                ct['total_earned']    += calc['BV_total_earned']
                ct['total_deduction'] += calc['CD_total_deduction']
                ct['paid_salary']     += calc['CE_paid_salary']
                ct['pf']              += calc['CB_pf']
                ct['pt']              += calc['CA_pt']
                ct['tds']             += calc['CC_tds']
                ct['employees_paid']  += 1

            if had_any:
                grand['employees_with_any_data'] += 1
                for k in ('total_earned','total_deduction','paid_salary','pf','pt','tds'):
                    grand[k] = grand.get(k, 0) + emp_totals[k]

            employees_out.append({
                **emp, 'totals': emp_totals, 'monthly': emp_monthly,
            })
        conn.close()

        return jsonify({
            'status':       'ok',
            'fy':           fy,
            'fy_label':     f'FY {fy}-{str(fy+1)[-2:]}',
            'months':       months,
            'employees':    employees_out,
            'company_by_month': [company_totals[m] for m in months],
            'grand':        grand,
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 2e — Loan register with EMI schedule
#
# Data model
# ----------
#   hr_loans          — one row per loan (principal, schedule, auto-deduct flag)
#   hr_loan_schedule  — one row per EMI installment (period_month, status)
#
# Two scheduling modes:
#   'fixed_emi'     — user sets principal + emi_amount; tenure_months derived
#                     = ceil(principal / emi_amount), last installment smaller
#   'fixed_tenure'  — user sets principal + tenure_months; emi derived
#                     = round(principal / tenure, 0), residual in last installment
#
# Zero interest: the sum of scheduled_amount rows always equals principal.
#
# Auto-deduct: when a period is saved, for each active loan with auto_deduct=1
#              and a pending schedule row for that YYYY-MM, we mark the row paid
#              and include its amount in the period's loan_deduction (BY).
#
# Routes
#   GET  /api/hr_salary/loans?emp_id=X          list loans + schedules for one employee
#   GET  /api/hr_salary/loans/all               company-wide loan register
#   POST /api/hr_salary/loans/save              create or update a loan (regens schedule)
#   POST /api/hr_salary/loans/close             close a loan (mark status)
#   POST /api/hr_salary/loans/mark_paid         manually mark one installment paid/skipped
# ═══════════════════════════════════════════════════════════════════════════════

def _add_months_ym(ym, delta):
    """Advance 'YYYY-MM' by `delta` months. Returns 'YYYY-MM'."""
    y, m = [int(x) for x in ym.split('-')]
    total = (y * 12 + (m - 1)) + int(delta)
    ny, nm = divmod(total, 12)
    return f'{ny:04d}-{nm+1:02d}'


def _generate_loan_schedule(principal, mode, emi_amount, tenure_months, start_month):
    """
    Return a list of dicts: [{installment_no, period_month, scheduled_amount}, ...]
    Sum of scheduled_amount == principal (zero interest, no residual lost).
    """
    p = float(principal or 0)
    if p <= 0:
        return []

    mode = (mode or 'fixed_emi').strip()
    out = []

    if mode == 'fixed_tenure':
        tenure = int(tenure_months or 0)
        if tenure <= 0:
            return []
        base = round(p / tenure)          # round EMI to whole rupees
        running = 0
        for i in range(1, tenure + 1):
            if i < tenure:
                amt = base
            else:
                amt = round(p - running, 2)   # residual goes in last installment
            running += amt
            out.append({
                'installment_no':   i,
                'period_month':     _add_months_ym(start_month, i - 1),
                'scheduled_amount': amt,
            })
        return out

    # default: fixed_emi
    emi = float(emi_amount or 0)
    if emi <= 0:
        return []
    n_full = int(p // emi)
    residual = round(p - n_full * emi, 2)
    total = n_full + (1 if residual > 0 else 0)
    for i in range(1, total + 1):
        if i <= n_full and not (i == total and residual == 0):
            amt = emi
        else:
            amt = residual if residual > 0 else emi
        out.append({
            'installment_no':   i,
            'period_month':     _add_months_ym(start_month, i - 1),
            'scheduled_amount': round(amt, 2),
        })
    return out


def _loan_totals(loan_id, conn):
    """Return {scheduled, paid, pending, pending_count, paid_count, next_due} for one loan."""
    rows = [dict(r) for r in conn.execute(
        "SELECT scheduled_amount, paid_amount, status, period_month "
        "FROM hr_loan_schedule WHERE loan_id=%s ORDER BY installment_no ASC",
        (loan_id,)
    ).fetchall()]
    scheduled = sum(float(r['scheduled_amount'] or 0) for r in rows)
    paid      = sum(float(r['paid_amount'] or 0) for r in rows)
    pending_rows = [r for r in rows if (r['status'] or 'pending') == 'pending']
    next_due = pending_rows[0]['period_month'] if pending_rows else None
    return {
        'scheduled':     round(scheduled, 2),
        'paid':          round(paid, 2),
        'outstanding':   round(scheduled - paid, 2),
        'total_count':   len(rows),
        'paid_count':    sum(1 for r in rows if (r['status'] or '') == 'paid'),
        'pending_count': len(pending_rows),
        'next_due':      next_due,
    }


@hr_salary_bp.route('/api/hr_salary/loans')
@hr_required
def api_loans_for_emp():
    """List loans (with schedule) for one employee."""
    try:
        emp_id = (request.args.get('emp_id') or '').strip()
        if not emp_id:
            return jsonify({'status': 'error', 'message': 'emp_id required'}), 400
        conn = sampling_portal.get_db_connection()
        emp = conn.execute(
            "SELECT * FROM hr_employees WHERE emp_id=%s AND is_active=1", (emp_id,)
        ).fetchone()
        if not emp:
            conn.close()
            return jsonify({'status': 'error', 'message': 'employee not found'}), 404
        emp = dict(emp)

        loans_rows = [dict(r) for r in conn.execute(
            "SELECT * FROM hr_loans WHERE emp_row_id=%s "
            "ORDER BY (status='active') DESC, loan_date DESC, id DESC",
            (emp['id'],)
        ).fetchall()]
        out = []
        for lo in loans_rows:
            sched = [dict(r) for r in conn.execute(
                "SELECT * FROM hr_loan_schedule WHERE loan_id=%s ORDER BY installment_no ASC",
                (lo['id'],)
            ).fetchall()]
            # Stringify dates for JSON
            if lo.get('loan_date'):
                try: lo['loan_date'] = str(lo['loan_date'])
                except Exception: pass
            for s in sched:
                if s.get('paid_at'):
                    try: s['paid_at'] = str(s['paid_at'])
                    except Exception: pass
            lo['schedule'] = sched
            lo['totals']   = _loan_totals(lo['id'], conn)
            out.append(lo)
        conn.close()
        return jsonify({
            'status': 'ok',
            'employee': {'emp_id': emp['emp_id'], 'emp_name': emp['emp_name']},
            'loans': out,
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/loans/all')
@hr_required
def api_loans_all():
    """Company-wide loan register with per-loan totals."""
    try:
        status_filter = (request.args.get('status') or '').strip()
        conn = sampling_portal.get_db_connection()
        if status_filter in ('active', 'closed', 'cancelled'):
            rows = conn.execute(
                "SELECT l.*, e.emp_id, e.emp_name, e.department "
                "FROM hr_loans l JOIN hr_employees e ON e.id=l.emp_row_id "
                "WHERE l.status=%s "
                "ORDER BY (l.status='active') DESC, l.loan_date DESC, l.id DESC",
                (status_filter,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT l.*, e.emp_id, e.emp_name, e.department "
                "FROM hr_loans l JOIN hr_employees e ON e.id=l.emp_row_id "
                "ORDER BY (l.status='active') DESC, l.loan_date DESC, l.id DESC"
            ).fetchall()

        out = []
        grand = {'principal_total': 0, 'outstanding_total': 0, 'active_count': 0}
        for r in rows:
            d = dict(r)
            if d.get('loan_date'):
                try: d['loan_date'] = str(d['loan_date'])
                except Exception: pass
            d['totals'] = _loan_totals(d['id'], conn)
            out.append(d)
            grand['principal_total']  += float(d.get('principal') or 0)
            grand['outstanding_total'] += d['totals']['outstanding']
            if (d.get('status') or '') == 'active':
                grand['active_count'] += 1
        conn.close()
        grand['principal_total']  = round(grand['principal_total'], 2)
        grand['outstanding_total'] = round(grand['outstanding_total'], 2)
        return jsonify({'status': 'ok', 'loans': out, 'grand': grand})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/loans/save', methods=['POST'])
@hr_required
def api_loan_save():
    """
    Create or update a loan agreement.

    Body:
        {
          id:              null | <int>   # null = create
          emp_id:          "1001"
          loan_date:       "2026-03-15"
          principal:       50000
          schedule_mode:   "fixed_emi" | "fixed_tenure"
          emi_amount:      5000       (used when mode=fixed_emi)
          tenure_months:   10         (used when mode=fixed_tenure)
          start_month:     "2026-04"
          auto_deduct:     true
          purpose:         "Medical"
          remarks:         ""
        }

    Behaviour:
        - Regenerates the schedule rows from scratch every save (DELETE + INSERT).
          Existing paid_amount / paid_in_period_id values are preserved when
          the new schedule has the same (installment_no) — otherwise the old
          payments are orphaned and the operation is refused if any payments
          exist, unless `force=true` is passed.
    """
    try:
        d = request.get_json(force=True) or {}
        emp_id = (d.get('emp_id') or '').strip()
        if not emp_id:
            return jsonify({'status': 'error', 'message': 'emp_id required'}), 400

        loan_id       = d.get('id')
        principal     = float(d.get('principal') or 0)
        mode          = (d.get('schedule_mode') or 'fixed_emi').strip()
        emi_amount    = float(d.get('emi_amount') or 0)
        tenure_months = int(d.get('tenure_months') or 0)
        start_month   = (d.get('start_month') or '').strip()
        loan_date     = (d.get('loan_date') or '').strip() or None
        auto_deduct   = 1 if d.get('auto_deduct') else 0
        purpose       = (d.get('purpose') or '').strip()[:200]
        remarks       = (d.get('remarks') or '').strip()[:500]
        force         = bool(d.get('force'))

        if principal <= 0:
            return jsonify({'status': 'error', 'message': 'principal must be > 0'}), 400
        if mode not in ('fixed_emi', 'fixed_tenure'):
            return jsonify({'status': 'error', 'message': 'schedule_mode invalid'}), 400
        if mode == 'fixed_emi' and emi_amount <= 0:
            return jsonify({'status': 'error', 'message': 'emi_amount must be > 0'}), 400
        if mode == 'fixed_tenure' and tenure_months <= 0:
            return jsonify({'status': 'error', 'message': 'tenure_months must be > 0'}), 400
        if not re.match(r'^20\d{2}-(0[1-9]|1[0-2])$', start_month):
            return jsonify({'status': 'error',
                            'message': 'start_month must be YYYY-MM'}), 400

        # Generate schedule
        sched = _generate_loan_schedule(principal, mode, emi_amount, tenure_months, start_month)
        if not sched:
            return jsonify({'status': 'error',
                            'message': 'Could not generate schedule — check inputs'}), 400

        # Derived: store final tenure (count of installments) for display
        final_tenure = len(sched)
        # Derive an EMI to store (for display) — use most-common amount
        amounts = [s['scheduled_amount'] for s in sched]
        display_emi = amounts[0] if amounts else 0

        conn = sampling_portal.get_db_connection()
        emp = conn.execute(
            "SELECT id FROM hr_employees WHERE emp_id=%s AND is_active=1", (emp_id,)
        ).fetchone()
        if not emp:
            conn.close()
            return jsonify({'status': 'error', 'message': 'employee not found'}), 404
        emp_row_id = dict(emp)['id']

        if loan_id:
            # Update existing
            existing = conn.execute(
                "SELECT id FROM hr_loans WHERE id=%s", (loan_id,)
            ).fetchone()
            if not existing:
                conn.close()
                return jsonify({'status': 'error', 'message': 'loan not found'}), 404

            # If payments already exist, refuse unless force
            paid_rows = conn.execute(
                "SELECT COUNT(*) AS cnt FROM hr_loan_schedule "
                "WHERE loan_id=%s AND (status='paid' OR status='partial')",
                (loan_id,)
            ).fetchone()
            paid_cnt = dict(paid_rows).get('cnt', 0) if paid_rows else 0
            if paid_cnt > 0 and not force:
                conn.close()
                return jsonify({'status': 'error',
                                'message': f'{paid_cnt} installment(s) already paid. '
                                           'Pass force=true to regenerate the schedule.'}), 409

            conn.execute("""
                UPDATE hr_loans
                   SET loan_date=%s, principal=%s, schedule_mode=%s,
                       emi_amount=%s, tenure_months=%s, start_month=%s,
                       auto_deduct=%s, purpose=%s, remarks=%s
                 WHERE id=%s
            """, (loan_date, principal, mode, display_emi, final_tenure,
                  start_month, auto_deduct, purpose, remarks, loan_id))
            # Wipe and re-insert schedule
            conn.execute("DELETE FROM hr_loan_schedule WHERE loan_id=%s", (loan_id,))
            the_loan_id = loan_id
        else:
            conn.execute("""
                INSERT INTO hr_loans
                    (emp_row_id, loan_date, principal, schedule_mode,
                     emi_amount, tenure_months, start_month,
                     auto_deduct, status, purpose, remarks)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'active',%s,%s)
            """, (emp_row_id, loan_date, principal, mode,
                  display_emi, final_tenure, start_month,
                  auto_deduct, purpose, remarks))
            row = conn.execute(
                "SELECT id FROM hr_loans WHERE emp_row_id=%s "
                "ORDER BY id DESC LIMIT 1", (emp_row_id,)
            ).fetchone()
            the_loan_id = dict(row)['id']

        for s in sched:
            conn.execute("""
                INSERT INTO hr_loan_schedule
                    (loan_id, installment_no, period_month, scheduled_amount, status)
                VALUES (%s,%s,%s,%s,'pending')
            """, (the_loan_id, s['installment_no'], s['period_month'], s['scheduled_amount']))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'loan_id': the_loan_id, 'tenure': final_tenure})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/loans/close', methods=['POST'])
@hr_required
def api_loan_close():
    """Set a loan's status to 'closed' (or 'cancelled' if requested)."""
    try:
        d = request.get_json(force=True) or {}
        loan_id = d.get('id')
        target  = (d.get('status') or 'closed').strip()
        if not loan_id:
            return jsonify({'status': 'error', 'message': 'id required'}), 400
        if target not in ('closed', 'cancelled', 'active'):
            return jsonify({'status': 'error', 'message': 'invalid status'}), 400
        conn = sampling_portal.get_db_connection()
        conn.execute("UPDATE hr_loans SET status=%s WHERE id=%s", (target, loan_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/loans/mark_paid', methods=['POST'])
@hr_required
def api_loan_mark_paid():
    """
    Manually mark a single installment row as paid / partial / skipped / pending.
    Body: { schedule_id: <int>, status: 'paid'|'partial'|'skipped'|'pending', paid_amount: <num> }
    """
    try:
        d = request.get_json(force=True) or {}
        sid = d.get('schedule_id')
        st  = (d.get('status') or 'paid').strip()
        pa  = float(d.get('paid_amount') or 0)
        if not sid:
            return jsonify({'status': 'error', 'message': 'schedule_id required'}), 400
        if st not in ('paid', 'partial', 'skipped', 'pending'):
            return jsonify({'status': 'error', 'message': 'invalid status'}), 400

        conn = sampling_portal.get_db_connection()
        row = conn.execute(
            "SELECT scheduled_amount FROM hr_loan_schedule WHERE id=%s", (sid,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'status': 'error', 'message': 'schedule row not found'}), 404
        sched = float(dict(row)['scheduled_amount'] or 0)

        if st == 'pending':
            new_paid = 0
            paid_at = None
        elif st == 'skipped':
            new_paid = 0
            paid_at = datetime.now()
        elif st == 'paid':
            new_paid = pa if pa > 0 else sched
            paid_at = datetime.now()
        else:  # partial
            new_paid = pa
            paid_at = datetime.now()

        conn.execute(
            "UPDATE hr_loan_schedule SET status=%s, paid_amount=%s, paid_at=%s "
            " WHERE id=%s",
            (st, new_paid, paid_at, sid)
        )
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# Auto-deduction helper: called from api_period_save()
# ─────────────────────────────────────────────────────────────────────────────

def _collect_auto_loan_emis(conn, emp_row_id, period_month):
    """
    Return (total_emi_amount, list of {schedule_id, loan_id, scheduled_amount}).
    Only considers ACTIVE loans with auto_deduct=1 and a PENDING schedule row
    whose period_month == period_month.
    """
    rows = conn.execute("""
        SELECT s.id AS schedule_id, s.loan_id, s.scheduled_amount
          FROM hr_loan_schedule s
          JOIN hr_loans l ON l.id = s.loan_id
         WHERE l.emp_row_id = %s
           AND l.status = 'active'
           AND l.auto_deduct = 1
           AND s.period_month = %s
           AND s.status = 'pending'
    """, (emp_row_id, period_month)).fetchall()
    items = [dict(r) for r in rows]
    total = sum(float(it['scheduled_amount'] or 0) for it in items)
    return round(total, 2), items


def _apply_auto_loan_emis(conn, period_id, items):
    """Mark the given schedule rows as paid, linking to the given period_id."""
    if not items:
        return
    for it in items:
        conn.execute("""
            UPDATE hr_loan_schedule
               SET status='paid',
                   paid_amount=scheduled_amount,
                   paid_in_period_id=%s,
                   paid_at=%s
             WHERE id=%s
        """, (period_id, datetime.now(), it['schedule_id']))
    # If the loan now has zero pending rows, flip its status to 'closed'
    loan_ids = {it['loan_id'] for it in items}
    for lid in loan_ids:
        pending = conn.execute(
            "SELECT COUNT(*) AS cnt FROM hr_loan_schedule "
            " WHERE loan_id=%s AND status='pending'", (lid,)
        ).fetchone()
        if dict(pending).get('cnt', 0) == 0:
            conn.execute("UPDATE hr_loans SET status='closed' WHERE id=%s", (lid,))


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 3 — WhatsApp salary-slip facility (single + bulk)
#
# Strategy: server builds the slip TEXT + a canonical wa.me deep link
#           ("https://wa.me/<phone>?text=<url-encoded-slip>").
#           The browser (authenticated in web.whatsapp.com) opens each
#           link in a new tab — the user just presses "Send" on each.
#           No WhatsApp Business API, no credentials stored server-side.
# ═══════════════════════════════════════════════════════════════════════════════

def _build_slip_text(emp, period_month, calc):
    """Plain-text salary slip — optimised for WhatsApp (wa.me supports ~64k chars)."""
    lines = [
        '*HCP WELLNESS PVT. LTD.*',
        '*SALARY SLIP — ' + period_month + '*',
        '----------------------------------------',
        f"Emp ID  : {emp.get('emp_id') or ''}",
        f"Name    : {emp.get('emp_name') or ''}",
    ]
    if emp.get('department'):  lines.append(f"Dept    : {emp['department']}")
    if emp.get('designation'): lines.append(f"Role    : {emp['designation']}")

    lines += [
        '----------------------------------------',
        '*ATTENDANCE*',
        f"Actual Days       : {calc['BF_actual_days']}",
        f"Present           : {calc['AO_total_present']}",
        f"Paid Holiday      : {calc['AP_paid_holiday']}",
        f"PL / CL / SL used : {calc['AR_pl_used']} / {calc['AS_cl_used']} / {calc['AT_sl_used']}",
        f"WO / WOP / PHP    : {calc['AU_total_wo']} / {calc['AV_total_wop']} / {calc['AW_total_php']}",
        f"OT Hours (days)   : {calc['BD_ot_hours']} ({calc['BE_ot_days']})",
        f"*Paid Days*         : *{calc['BG_total_days']}*",
        '----------------------------------------',
        '*EARNINGS*',
        f"Basic             : {_fmt_inr(calc['BN_earned_basic'])}",
        f"HRA               : {_fmt_inr(calc['BO_earned_hra'])}",
        f"Conveyance        : {_fmt_inr(calc['BP_earned_conv'])}",
        f"Medical           : {_fmt_inr(calc['BQ_earned_medical'])}",
        f"Other Allowance   : {_fmt_inr(calc['BR_earned_other'])}",
        f"Prod. Incentive   : {_fmt_inr(calc['BS_prod_incentive'])}",
        f"Arrear            : {_fmt_inr(calc['BU_arrear'])}",
        f"*Total Earned*      : *{_fmt_inr(calc['BV_total_earned'])}*",
        '----------------------------------------',
        '*DEDUCTIONS*',
        f"Loan              : {_fmt_inr(calc['BY_loan_deduction'])}  (Bal: {_fmt_inr(calc['BX_balance_loan'])})",
        f"Other             : {_fmt_inr(calc['BZ_other_deduction'])}",
        f"PT                : {_fmt_inr(calc['CA_pt'])}",
        f"PF                : {_fmt_inr(calc['CB_pf'])}",
        f"TDS               : {_fmt_inr(calc['CC_tds'])}",
        f"*Total Deduction*   : *{_fmt_inr(calc['CD_total_deduction'])}*",
        '----------------------------------------',
        f"*PAID SALARY*       : *{_fmt_inr(calc['CE_paid_salary'])}*",
        '----------------------------------------',
        '_This is a system-generated slip._',
    ]
    return '\n'.join(lines)


@hr_salary_bp.route('/api/hr_salary/whatsapp/prepare', methods=['POST'])
@hr_required
def api_whatsapp_prepare():
    """
    Build wa.me links for the selected employees.

    Request  : { month: "YYYY-MM", emp_ids: ["1","2",...] }  (omit emp_ids for ALL active)
    Response : {
        status: "ok",
        items: [
            { emp_id, emp_name, mobile, has_mobile, has_period,
              wa_link, slip_text, warning? },
            ...
        ]
    }

    The frontend iterates `items` and window.open()s each `wa_link` one-by-one,
    waiting for user confirmation between sends.
    """
    try:
        d = request.get_json(force=True) or {}
        month = (d.get('month') or datetime.now().strftime('%Y-%m')).strip()
        requested = d.get('emp_ids') or []
        requested_set = {str(x).strip() for x in requested if str(x).strip()}

        conn = sampling_portal.get_db_connection()
        emps = conn.execute("""
            SELECT * FROM hr_employees WHERE is_active=1
            ORDER BY CAST(emp_id AS UNSIGNED) ASC, emp_id ASC
        """).fetchall()

        from urllib.parse import quote
        items = []
        for e in emps:
            e = dict(e)
            if requested_set and e['emp_id'] not in requested_set:
                continue

            mobile = _clean_mobile(e.get('mobile'))
            item = {
                'emp_id':     e['emp_id'],
                'emp_name':   e['emp_name'],
                'department': e.get('department') or '',
                'mobile':     mobile,
                'has_mobile': bool(mobile),
                'has_period': False,
                'wa_link':    '',
                'slip_text':  '',
            }

            p_row = conn.execute("""
                SELECT * FROM hr_salary_periods WHERE emp_row_id=%s AND period_month=%s
            """, (e['id'], month)).fetchone()
            if not p_row:
                item['warning'] = 'No salary data for this month'
                items.append(item); continue

            p = dict(p_row)
            day_rows = [dict(r) for r in conn.execute(
                "SELECT day_num, status_code, ot_hours FROM hr_salary_days WHERE period_id=%s",
                (p['id'],)
            ).fetchall()]
            calc = _calc_from_period_and_days(p, day_rows)

            slip = _build_slip_text(e, month, calc)
            item['has_period'] = True
            item['slip_text']  = slip
            item['paid_salary'] = calc['CE_paid_salary']
            if mobile:
                item['wa_link'] = f'https://wa.me/{mobile}?text={quote(slip)}'
            else:
                item['warning'] = 'No mobile number on file'
            items.append(item)

        conn.close()
        return jsonify({'status': 'ok', 'month': month, 'items': items})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/whatsapp/preview')
@hr_required
def api_whatsapp_preview():
    """Return a single slip's plain text — used by the 'Preview' button in the UI."""
    try:
        emp_id = (request.args.get('emp_id') or '').strip()
        month  = (request.args.get('month')  or datetime.now().strftime('%Y-%m')).strip()
        if not emp_id:
            return jsonify({'status': 'error', 'message': 'emp_id required'}), 400
        conn = sampling_portal.get_db_connection()
        e = conn.execute(
            "SELECT * FROM hr_employees WHERE emp_id=%s AND is_active=1", (emp_id,)
        ).fetchone()
        if not e:
            conn.close()
            return jsonify({'status': 'error', 'message': 'employee not found'}), 404
        e = dict(e)
        p_row = conn.execute("""
            SELECT * FROM hr_salary_periods WHERE emp_row_id=%s AND period_month=%s
        """, (e['id'], month)).fetchone()
        if not p_row:
            conn.close()
            return jsonify({'status': 'error', 'message': 'no salary data for this month'}), 404
        p = dict(p_row)
        day_rows = [dict(r) for r in conn.execute(
            "SELECT day_num, status_code, ot_hours FROM hr_salary_days WHERE period_id=%s",
            (p['id'],)
        ).fetchall()]
        conn.close()
        calc = _calc_from_period_and_days(p, day_rows)
        return jsonify({
            'status':    'ok',
            'slip_text': _build_slip_text(e, month, calc),
            'mobile':    _clean_mobile(e.get('mobile')),
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# PDF salary slip generation (reportlab)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_slip_pdf_bytes(emp, period_month, calc):
    """
    Build a one-page A4 salary slip PDF.
    Returns the PDF bytes — caller decides how to deliver (send_file / zip / etc.).
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=12*mm,  bottomMargin=12*mm,
    )

    styles = getSampleStyleSheet()
    h_company = ParagraphStyle('h_company', parent=styles['Heading1'],
                               fontSize=16, leading=20, alignment=TA_CENTER,
                               textColor=colors.HexColor('#1E3A8A'), spaceAfter=2)
    h_slip    = ParagraphStyle('h_slip', parent=styles['Heading2'],
                               fontSize=12, leading=14, alignment=TA_CENTER,
                               textColor=colors.HexColor('#475569'), spaceAfter=6)
    lbl       = ParagraphStyle('lbl', parent=styles['Normal'], fontSize=9,
                               textColor=colors.HexColor('#64748B'))
    val       = ParagraphStyle('val', parent=styles['Normal'], fontSize=10,
                               textColor=colors.HexColor('#1E293B'))

    story = []
    story.append(Paragraph('HCP WELLNESS PVT. LTD.', h_company))
    story.append(Paragraph(f'Salary Slip &mdash; {period_month}', h_slip))

    # Employee header block
    emp_tbl = Table([
        [Paragraph('Emp ID', lbl),       Paragraph(str(emp.get('emp_id') or ''), val),
         Paragraph('Department', lbl),   Paragraph(str(emp.get('department') or '—'), val)],
        [Paragraph('Name', lbl),         Paragraph(str(emp.get('emp_name') or ''), val),
         Paragraph('Designation', lbl),  Paragraph(str(emp.get('designation') or '—'), val)],
        [Paragraph('Gender', lbl),       Paragraph(str(emp.get('gender') or '—'), val),
         Paragraph('Status', lbl),       Paragraph(str(emp.get('emp_status') or '—'), val)],
    ], colWidths=[22*mm, 58*mm, 22*mm, 58*mm])
    emp_tbl.setStyle(TableStyle([
        ('GRID',   (0,0), (-1,-1), 0.4, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F8FAFC')),
        ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#F8FAFC')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING',  (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
    ]))
    story.append(emp_tbl)
    story.append(Spacer(1, 5*mm))

    # Attendance block
    att_tbl = Table([
        ['Actual Days', 'Present', 'Paid Hol.', 'PL Used', 'CL Used', 'SL Used',
         'WO', 'WOP', 'PHP', 'OT Hrs', 'OT Days', 'Paid Days'],
        [calc['BF_actual_days'], calc['AO_total_present'], calc['AP_paid_holiday'],
         calc['AR_pl_used'], calc['AS_cl_used'], calc['AT_sl_used'],
         calc['AU_total_wo'], calc['AV_total_wop'], calc['AW_total_php'],
         calc['BD_ot_hours'], calc['BE_ot_days'], calc['BG_total_days']],
    ])
    att_tbl.setStyle(TableStyle([
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN',    (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',   (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white]),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
    ]))
    story.append(att_tbl)
    story.append(Spacer(1, 5*mm))

    # Earnings & Deductions side-by-side
    def _row(label, amount, bold=False):
        style = val
        if bold:
            style = ParagraphStyle('b', parent=val, fontName='Helvetica-Bold')
        return [Paragraph(label, style), Paragraph(_fmt_inr(amount), style)]

    earn_rows = [
        [Paragraph('<b>EARNINGS</b>', val), Paragraph('<b>Amount</b>', val)],
        _row('Basic',             calc['BN_earned_basic']),
        _row('HRA',               calc['BO_earned_hra']),
        _row('Conveyance',        calc['BP_earned_conv']),
        _row('Medical',           calc['BQ_earned_medical']),
        _row('Other Allowance',   calc['BR_earned_other']),
        _row('Production Incentive', calc['BS_prod_incentive']),
        _row('Arrear',            calc['BU_arrear']),
        _row('TOTAL EARNED',      calc['BV_total_earned'], bold=True),
    ]
    ded_rows = [
        [Paragraph('<b>DEDUCTIONS</b>', val), Paragraph('<b>Amount</b>', val)],
        _row('Loan Repayment',    calc['BY_loan_deduction']),
        _row('Loan Balance',      calc['BX_balance_loan']),
        _row('Other',             calc['BZ_other_deduction']),
        _row('Professional Tax',  calc['CA_pt']),
        _row('PF',                calc['CB_pf']),
        _row('TDS',               calc['CC_tds']),
        [Paragraph('', val), Paragraph('', val)],
        _row('TOTAL DEDUCTION',   calc['CD_total_deduction'], bold=True),
    ]
    # Pad the shorter so both tables align
    while len(earn_rows) < len(ded_rows):
        earn_rows.insert(-1, [Paragraph('', val), Paragraph('', val)])
    while len(ded_rows) < len(earn_rows):
        ded_rows.insert(-1, [Paragraph('', val), Paragraph('', val)])

    def _make_side_table(rows):
        t = Table(rows, colWidths=[50*mm, 30*mm])
        t.setStyle(TableStyle([
            ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#E2E8F0')),
            ('BACKGROUND', (0,0), (-1,0),  colors.HexColor('#F1F5F9')),
            ('BACKGROUND', (0,-1),(-1,-1), colors.HexColor('#DCFCE7')),
            ('ALIGN',      (1,0), (1,-1), 'RIGHT'),
            ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING',  (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING',   (0,0), (-1,-1), 3),
            ('BOTTOMPADDING',(0,0), (-1,-1), 3),
        ]))
        return t

    side = Table(
        [[_make_side_table(earn_rows), _make_side_table(ded_rows)]],
        colWidths=[82*mm, 82*mm]
    )
    side.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(side)
    story.append(Spacer(1, 6*mm))

    # Final net
    net_tbl = Table([
        [Paragraph('<b>NET PAID SALARY</b>',
                   ParagraphStyle('nf', parent=val, fontSize=12,
                                  textColor=colors.white)),
         Paragraph('<b>' + _fmt_inr(calc['CE_paid_salary']) + '</b>',
                   ParagraphStyle('nv', parent=val, fontSize=12,
                                  textColor=colors.white, alignment=TA_RIGHT))],
    ], colWidths=[100*mm, 64*mm])
    net_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#1E3A8A')),
        ('LEFTPADDING',  (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ('TOPPADDING',   (0,0), (-1,-1), 8),
        ('BOTTOMPADDING',(0,0), (-1,-1), 8),
    ]))
    story.append(net_tbl)

    story.append(Spacer(1, 8*mm))
    story.append(Paragraph(
        '<i>This is a system-generated salary slip and does not require a signature.</i>',
        ParagraphStyle('foot', parent=styles['Normal'], fontSize=8,
                       textColor=colors.HexColor('#94A3B8'), alignment=TA_CENTER),
    ))

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    return pdf


def _load_calc_for_emp(emp_id, period_month):
    """Helper: fetch employee + computed calc for a month. Returns (emp, calc) or (None, None)."""
    conn = sampling_portal.get_db_connection()
    e = conn.execute(
        "SELECT * FROM hr_employees WHERE emp_id=%s AND is_active=1", (emp_id,)
    ).fetchone()
    if not e:
        conn.close()
        return None, None
    e = dict(e)
    p_row = conn.execute("""
        SELECT * FROM hr_salary_periods WHERE emp_row_id=%s AND period_month=%s
    """, (e['id'], period_month)).fetchone()
    if not p_row:
        conn.close()
        return e, None
    p = dict(p_row)
    day_rows = [dict(r) for r in conn.execute(
        "SELECT day_num, status_code, ot_hours FROM hr_salary_days WHERE period_id=%s",
        (p['id'],)
    ).fetchall()]
    conn.close()
    calc = _calc_from_period_and_days(p, day_rows)
    return e, calc


@hr_salary_bp.route('/api/hr_salary/slip_pdf')
@hr_required
def api_slip_pdf():
    """
    Single-employee salary slip as PDF download.
    Query:  ?emp_id=<id>&month=YYYY-MM
    """
    try:
        emp_id = (request.args.get('emp_id') or '').strip()
        month  = (request.args.get('month')  or datetime.now().strftime('%Y-%m')).strip()
        if not emp_id:
            return jsonify({'status': 'error', 'message': 'emp_id required'}), 400

        emp, calc = _load_calc_for_emp(emp_id, month)
        if not emp:
            return jsonify({'status': 'error', 'message': 'employee not found'}), 404
        if not calc:
            return jsonify({'status': 'error',
                            'message': 'No salary data for this month'}), 404

        pdf = _build_slip_pdf_bytes(emp, month, calc)
        safe_name = re.sub(r'[^A-Za-z0-9_\-]+', '_', str(emp.get('emp_name') or 'slip'))[:40]
        fname = f'Salary_{emp["emp_id"]}_{safe_name}_{month}.pdf'
        return send_file(
            BytesIO(pdf),
            as_attachment=True,
            download_name=fname,
            mimetype='application/pdf',
        )
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@hr_salary_bp.route('/api/hr_salary/slips_zip', methods=['POST'])
@hr_required
def api_slips_zip():
    """
    Bulk download salary slips as a single .zip of PDFs.
    Body:  { month: "YYYY-MM", emp_ids: ["1","3","7"] }  (omit emp_ids → all active)
    """
    try:
        d = request.get_json(force=True) or {}
        month = (d.get('month') or datetime.now().strftime('%Y-%m')).strip()
        requested = {str(x).strip() for x in (d.get('emp_ids') or []) if str(x).strip()}

        conn = sampling_portal.get_db_connection()
        emps = conn.execute("""
            SELECT emp_id FROM hr_employees WHERE is_active=1
            ORDER BY CAST(emp_id AS UNSIGNED) ASC, emp_id ASC
        """).fetchall()
        conn.close()

        import zipfile
        zip_buf = BytesIO()
        included, skipped = 0, 0
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for row in emps:
                eid = dict(row)['emp_id']
                if requested and eid not in requested:
                    continue
                emp, calc = _load_calc_for_emp(eid, month)
                if not emp or not calc:
                    skipped += 1
                    continue
                pdf = _build_slip_pdf_bytes(emp, month, calc)
                safe_name = re.sub(r'[^A-Za-z0-9_\-]+', '_',
                                   str(emp.get('emp_name') or 'slip'))[:40]
                zf.writestr(f'Salary_{eid}_{safe_name}_{month}.pdf', pdf)
                included += 1

        if included == 0:
            return jsonify({
                'status': 'error',
                'message': f'No slips could be generated. {skipped} employee(s) had no data for {month}.',
            }), 404

        zip_buf.seek(0)
        return send_file(
            zip_buf,
            as_attachment=True,
            download_name=f'Salary_Slips_{month}.zip',
            mimetype='application/zip',
        )
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
