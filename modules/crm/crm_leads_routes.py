"""
modules/crm/crm_leads_routes.py
────────────────────────────────────────────────────────────────────────────
CRM · LEAD MODULE  (ported into HCPERP.zip conventions)

  • Raw pymysql via sampling_portal.get_db_connection()  (NO SQLAlchemy)
  • Session-based auth (session['user_id'], session['role'] / User_Type)
  • Common, data-driven sidebar via core/menus.get_menu('crm', ...)
  • Templates: templates/crm/leads/*.html  (extend HCPERP.zip design — header
    + sidebar partials + hcptheme.css + common-sidebar.css)

Blueprint  : crm_bp  (url_prefix='/crm')
Tables     : leads, lead_discussions, lead_attachments, lead_reminders,
             lead_notes, lead_activity_logs, lead_contributions,
             contribution_config, lead_statuses, lead_sources,
             lead_categories, product_ranges
             (FK soft-ref User_Tbl.id — main project ka users table)

Registration (app.py):
    from crm import crm_bp, ensure_lead_tables
    app.register_blueprint(crm_bp)
    ensure_lead_tables()
"""

import os
import io
from datetime import datetime, timedelta
from functools import wraps

from flask import (Blueprint, render_template, redirect, url_for, request,
                   flash, jsonify, session, send_file, current_app, abort)
from werkzeug.utils import secure_filename

import sampling_portal  # main project ka pymysql bridge (core/ on sys.path)

# CRM schema model — single source of truth for the list/columns dropdown.
# (routes raw pymysql use karte hain, par column-metadata model se aata hai.)
try:
    from models.crm_lead import LEAD_LIST_COLUMNS
except Exception:
    try:
        from crm_lead import LEAD_LIST_COLUMNS
    except Exception:
        LEAD_LIST_COLUMNS = [
            ('created', 'Created', True), ('name', 'Name', True),
            ('company', 'Company', True), ('email', 'Email', True),
            ('mobile', 'Mobile', True), ('product', 'Product', True),
            ('team', 'Team', True), ('status', 'Status', True),
            ('lead_type', 'Lead Type', True), ('last_contact', 'Last Contact', True),
            ('age', 'Days (Age)', True), ('code', 'Code', False),
            ('position', 'Position', False), ('website', 'Website', False),
            ('alt_mobile', 'Alt. Mobile', False), ('source', 'Source', False),
            ('category', 'Category', False), ('product_range', 'Product Range', False),
            ('order_quantity', 'Order Qty', False), ('city', 'City', False),
            ('state', 'State', False), ('country', 'Country', False),
            ('zip_code', 'ZIP', False), ('avg_cost', 'Avg Cost', False),
            ('tags', 'Tags', False),
        ]

# Common, data-driven sidebar (menu defined once in core/menus.py)
try:
    from menus import get_menu
except Exception:
    def get_menu(*a, **k):
        return None

crm_bp = Blueprint('crm', __name__, url_prefix='/crm')

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx',
                      'xls', 'xlsx', 'txt', 'csv', 'ppt', 'pptx'}
BLOCKED_EXTENSIONS = {'php', 'phtml', 'php3', 'php4', 'php5', 'phar',
                      'jsp', 'asp', 'aspx', 'cgi', 'exe', 'bat', 'cmd',
                      'com', 'msi', 'dll', 'so', 'dylib'}
MAX_FILE_SIZE = 25 * 1024 * 1024  # 25 MB

LEAD_COLS_DEFAULT = ['created_at', 'name', 'company', 'email', 'mobile',
                     'product', 'team', 'status', 'lead_type', 'last_contact']
LEAD_COLS_ALL = {
    'created_at': 'Created Date', 'name': 'Name', 'company': 'Company',
    'email': 'Email', 'mobile': 'Mobile', 'product': 'Product',
    'category': 'Category', 'source': 'Source', 'city': 'City',
    'assigned_to': 'Assigned To', 'team': 'Team', 'follow_up': 'Follow Up Date',
    'status': 'Status', 'lead_type': 'Lead Type', 'last_contact': 'Last Contact',
    'priority': 'Priority', 'expected_value': 'Expected Value',
    'lead_age': 'Days (Age)',
}

DEFAULT_CONTRIB_POINTS = {
    'comment': 1, 'status_change': 2, 'close_fast': 8, 'close_slow': 0,
    'cancel': 0, 'follow_up': 1, 'reminder': 1, 'edit': 1,
}


# ─────────────────────────────────────────────────────────────────────────────
# DB / AUTH HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _db():
    return sampling_portal.get_db_connection()


def _uid():
    return session.get('user_id')


def _uname():
    return session.get('User_Name') or session.get('UID') or 'User'


def _role():
    return (session.get('User_Type') or session.get('role') or '').lower()


def _is_admin():
    return _role() == 'admin' or (session.get('UID', '') or '').lower() == 'admin'


def _is_admin_mgr():
    return _is_admin() or _role() == 'manager'


# ── Department-based full visibility ─────────────────────────────────────────
# In departments ke users ko SARI leads dikhengi (chahe kisi ko bhi assign ho):
#   NPD Manager, Management, Administrator (+ common variants)
FULL_VISIBILITY_DEPTS = {'npd manager', 'management', 'administrator',
                         'administration', 'admin'}


def _dept():
    """Current user ka department (login pe session['department'] me aata hai;
    na ho to User_Tbl se ek baar fetch karke session me cache)."""
    d = session.get('department')
    if d is None and _uid():
        conn = _db()
        if conn:
            try:
                row = conn.execute(
                    "SELECT department FROM `User_Tbl` WHERE id=%s",
                    (_uid(),)).fetchone()
                d = (row or {}).get('department') or ''
                session['department'] = d
            finally:
                conn.close()
    return (d or '').strip().lower()


def _has_full_visibility():
    """Admin/Manager role YA NPD Manager / Management / Administrator dept."""
    return _is_admin_mgr() or _dept() in FULL_VISIBILITY_DEPTS


def _unassigned_sql():
    """'Unassigned' ki definition: lead kisi NORMAL user ko assign nahi —
    yaani assigned_to khali ho YA NPD Manager / Management / Administrator
    department ke kisi user ko assigned ho (in logon ko assignment real
    assignment nahi ginte). Returns (sql_fragment, params)."""
    depts = sorted(FULL_VISIBILITY_DEPTS)
    ph = ','.join(['%s'] * len(depts))
    frag = ("(assigned_to IS NULL OR assigned_to = 0 OR assigned_to IN "
            "(SELECT id FROM `User_Tbl` "
            f"WHERE LOWER(TRIM(COALESCE(department,''))) IN ({ph})))")
    return frag, depts


def login_required(f):
    @wraps(f)
    def wrapper(*a, **k):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*a, **k)
    return wrapper


@crm_bp.before_request
def _crm_admin_only():
    """CRM ab sab logged-in users ke liye khula hai — visibility
    _visibility_sql() handle karti hai (normal user = sirf apni
    assigned/created/team leads; NPD Manager / Management / Administrator
    dept + admin/manager = sab).
    Wapas SIRF-admin lock chahiye to neeche wala block uncomment kar do:
        if not _is_admin():
            if request.is_json or request.headers.get('X-Requested-With'):
                return jsonify(ok=False, success=False,
                               error='Access denied.'), 403
            flash('The CRM module is currently available to admins only.', 'error')
            return redirect('/')
    """
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return None


def _upload_dir():
    d = os.path.join(current_app.root_path, 'static', 'uploads', 'leads')
    os.makedirs(d, exist_ok=True)
    return d


def _remove_upload(rel_path):
    """static/uploads/<rel_path> file ko disk se hatao (slash normalize karke)."""
    if not rel_path:
        return
    rel = str(rel_path).replace('\\', '/').lstrip('/')
    full = os.path.join(current_app.root_path, 'static', 'uploads',
                        *rel.split('/'))
    try:
        if os.path.exists(full):
            os.remove(full)
    except Exception:
        pass


def _parse_dt(s):
    """Kisi bhi common format ki datetime string ko MySQL 'Y-m-d H:M:S' me convert."""
    if not s:
        return None
    s = str(s).strip().replace('T', ' ')
    fmts = ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M',
            '%d-%m-%Y %H:%M:%S', '%d-%m-%Y %H:%M',
            '%d-%m-%Y %I:%M %p', '%d/%m/%Y %H:%M',
            '%m/%d/%Y %H:%M', '%Y-%m-%d')
    for f in fmts:
        try:
            return datetime.strptime(s, f).strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue
    return None


def _allowed_file(filename):
    if '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    return ext not in BLOCKED_EXTENSIONS


# ── User lookups (User_Tbl) ──────────────────────────────────────────────────
def get_team_users():
    """Active users from User_Tbl for assigned-to / team dropdowns."""
    conn = _db()
    if not conn:
        return []
    try:
        rows = conn.execute(
            "SELECT id, full_name, username, role, department "
            "FROM `User_Tbl` WHERE is_active=1 ORDER BY full_name, username"
        ).fetchall()
        return rows or []
    finally:
        conn.close()


def _user_map(conn):
    """id -> display name map for the User_Tbl."""
    rows = conn.execute(
        "SELECT id, full_name, username FROM `User_Tbl`").fetchall() or []
    return {r['id']: (r['full_name'] or r['username'] or f"#{r['id']}")
            for r in rows}


# ── Code generation ──────────────────────────────────────────────────────────
def gen_lead_code(conn):
    row = conn.execute("SELECT MAX(id) AS mx FROM `leads`").fetchone()
    nxt = ((row['mx'] or 0) + 1) if row else 1
    return f"LD-{nxt:04d}"


# ── Activity log + contributions ─────────────────────────────────────────────
def log_activity(conn, lead_id, action, user_id=None):
    uid = user_id or _uid()
    conn.execute(
        "INSERT INTO `lead_activity_logs` (lead_id, user_id, action, created_at) "
        "VALUES (%s, %s, %s, NOW())", (lead_id, uid, action))


def _contrib_points(conn, action_type):
    try:
        row = conn.execute(
            "SELECT points FROM `contribution_config` WHERE action_type=%s",
            (action_type,)).fetchone()
        if row:
            return row['points']
    except Exception:
        pass
    return DEFAULT_CONTRIB_POINTS.get(action_type, 1)


def add_contribution(conn, lead_id, action_type, user_id=None, note=''):
    uid = user_id or _uid()
    if not uid:
        return
    pts = _contrib_points(conn, action_type)
    conn.execute(
        "INSERT INTO `lead_contributions` "
        "(lead_id, user_id, action_type, points, note, created_at) "
        "VALUES (%s, %s, %s, %s, %s, NOW())",
        (lead_id, uid, action_type, pts, note))


def _handle_close_contribution(conn, lead):
    """Smart close points split among active contributors, by speed slab."""
    created = lead.get('created_at')
    if isinstance(created, str):
        try:
            created = datetime.strptime(created[:19], '%Y-%m-%d %H:%M:%S')
        except Exception:
            created = datetime.now()
    days = max(0, (datetime.now().date() - created.date()).days) if created else 0
    if days <= 7:
        slab_pts = _contrib_points(conn, 'close_fast')
    elif days <= 14:
        slab_pts = max(0, _contrib_points(conn, 'close_fast') - 2)
    elif days <= 21:
        slab_pts = max(0, _contrib_points(conn, 'close_fast') - 4)
    elif days <= 28:
        slab_pts = max(0, _contrib_points(conn, 'close_fast') - 6)
    else:
        slab_pts = _contrib_points(conn, 'close_slow')

    actives = conn.execute(
        "SELECT DISTINCT user_id FROM `lead_contributions` WHERE lead_id=%s",
        (lead['id'],)).fetchall() or []
    active_ids = [r['user_id'] for r in actives if r['user_id']]
    if not active_ids and _uid():
        active_ids = [_uid()]
    if not active_ids:
        return
    per = max(0, slab_pts // len(active_ids))
    for uid in active_ids:
        conn.execute(
            "INSERT INTO `lead_contributions` "
            "(lead_id, user_id, action_type, points, note, created_at) "
            "VALUES (%s, %s, 'close', %s, %s, NOW())",
            (lead['id'], uid, per, f'Closed in {days} day(s)'))


# ── Visibility WHERE clause (role-based) ─────────────────────────────────────
def _visibility_sql():
    """Return (sql_fragment, params) limiting normal users.
    Full visibility: admin/manager role YA NPD Manager / Management /
    Administrator department. Baaki users: sirf assigned / created / team."""
    if _has_full_visibility():
        return "", []
    uid = _uid()
    us = str(uid)
    frag = (" AND (assigned_to=%s OR created_by=%s "
            "OR team_members=%s "
            "OR team_members LIKE %s OR team_members LIKE %s OR team_members LIKE %s)")
    params = [uid, uid, us, f"{us},%", f"%,{us},%", f"%,{us}"]
    return frag, params


# ─────────────────────────────────────────────────────────────────────────────
# LIST
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads')
@login_required
def leads():
    status = request.args.get('status', '')
    # Dashboard mini-card filters:
    #   pending_followup | unassigned | client_linked | client_unlinked
    flt = request.args.get('filter', '')
    search = request.args.get('search', '')
    source = request.args.get('source', '')
    category = request.args.get('category', '')
    p_range = request.args.get('product_range', '')
    city = request.args.get('city', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    sort_by = request.args.get('sort_by', 'created_at')
    sort_dir = request.args.get('sort_dir', 'desc')
    show_trash = request.args.get('trash', '') == '1'

    conn = _db()
    if not conn:
        flash('Database connection failed.', 'error')
        return redirect('/')
    try:
        vis_frag, vis_params = _visibility_sql()

        where = ["is_deleted = %s"]
        params = [1 if show_trash else 0]

        # ── Dashboard mini-card filters ──────────────────────────────────
        # (visibility apne aap lagti hai — normal user ko sirf apni
        #  assigned/created/team leads hi dikhengi in filters me bhi)
        if flt == 'pending_followup':
            where.append("status IN ('open','in_process')")
            where.append("follow_up_date IS NOT NULL")
            where.append("follow_up_date <= CURDATE()")
        elif flt == 'open':
            # Dashboard ka 'Open Leads' KPI = open + in_process dono
            where.append("status IN ('open','in_process')")
        elif flt == 'unassigned':
            # Unassigned list SIRF full-visibility users ke liye
            # (Management / Administrator / NPD Manager / admin / manager)
            if not _has_full_visibility():
                flash('The unassigned leads list is only visible to Management, Administrator '
                      'aur NPD Manager ke liye hai.', 'error')
                return redirect(url_for('crm.leads'))
            # 'Unassigned' = kisi normal user ko assign nahi (privileged
            # dept walon ko assignment nahi ginti)
            ua_frag, ua_params = _unassigned_sql()
            where.append(ua_frag)
            params.extend(ua_params)
        elif flt == 'client_linked':
            where.append("client_id IS NOT NULL")
        elif flt == 'client_unlinked':
            where.append("client_id IS NULL")

        if not show_trash and status:
            where.append("status = %s")
            params.append(status)
        if source:
            where.append("source = %s")
            params.append(source)
        if category:
            where.append("category = %s")
            params.append(category)
        if p_range:
            where.append("product_range = %s")
            params.append(p_range)
        if city:
            where.append("city LIKE %s")
            params.append(f"%{city}%")
        if date_from:
            where.append("created_at >= %s")
            params.append(date_from + " 00:00:00")
        if date_to:
            where.append("created_at <= %s")
            params.append(date_to + " 23:59:59")
        if search:
            s = f"%{search}%"
            cols = ['contact_name', 'company_name', 'phone', 'alternate_mobile',
                    'email', 'product_name', 'category', 'product_range',
                    'source', 'city', 'state', 'country', 'zip_code', 'address',
                    'position', 'title', 'tags', 'remark', 'notes',
                    'lost_reason', 'requirement_spec', 'order_quantity', 'code']
            where.append("(" + " OR ".join(f"{c} LIKE %s" for c in cols) + ")")
            params.extend([s] * len(cols))

        sort_map = {'name': 'contact_name', 'mobile': 'phone',
                    'company': 'company_name'}
        col = sort_map.get(sort_by, sort_by)
        valid_sort = {'created_at', 'contact_name', 'company_name', 'phone',
                      'email', 'status', 'follow_up_date', 'priority',
                      'expected_value', 'city', 'category', 'source'}
        if col not in valid_sort:
            col = 'created_at'
        direction = 'ASC' if sort_dir == 'asc' else 'DESC'

        sql = ("SELECT * FROM `leads` WHERE " + " AND ".join(where)
               + vis_frag + f" ORDER BY {col} {direction}")
        all_leads = conn.execute(sql, params + vis_params).fetchall() or []

        umap = _user_map(conn)
        for ld in all_leads:
            ld['_age'] = _lead_age(ld)
            ld['_assigned_name'] = umap.get(ld.get('assigned_to'), '')
            ld['_team_names'] = [umap.get(int(x)) for x in
                                 (ld.get('team_members') or '').split(',')
                                 if x.strip().isdigit()]
            # avatar list = assigned + team (unique, names only)
            _av = []
            if ld['_assigned_name']:
                _av.append(ld['_assigned_name'])
            for nm in ld['_team_names']:
                if nm and nm not in _av:
                    _av.append(nm)
            ld['_avatars'] = _av

        # ── counts ──
        def _count(extra_where, extra_params):
            q = ("SELECT COUNT(*) AS c FROM `leads` WHERE is_deleted=0 "
                 + extra_where + vis_frag)
            r = conn.execute(q, extra_params + vis_params).fetchone()
            return r['c'] if r else 0

        total_count = _count("", [])
        statuses = conn.execute(
            "SELECT name, label, color FROM `lead_statuses` "
            "WHERE is_active=1 ORDER BY sort_order").fetchall() or []
        counts = {}
        for st in statuses:
            counts[st['name']] = _count(" AND status=%s", [st['name']])
        for s in ('open', 'in_process', 'close', 'cancel'):
            counts.setdefault(s, _count(" AND status=%s", [s]))

        # NPD / Existing-project counts — leads jo project me convert hue.
        # NPDProject integration is module ke scope me nahi, isliye 0 (UI parity).
        npd_count = 0
        epd_count = 0

        # trash count (visibility-aware)
        tq = "SELECT COUNT(*) AS c FROM `leads` WHERE is_deleted=1" + vis_frag
        deleted_count = (conn.execute(tq, vis_params).fetchone() or {}).get('c', 0)

        # filter option lists
        def _distinct(c):
            r = conn.execute(
                f"SELECT DISTINCT {c} AS v FROM `leads` "
                f"WHERE {c} IS NOT NULL AND {c}<>'' ORDER BY {c}").fetchall()
            return [x['v'] for x in (r or [])]

        all_sources = _distinct('source')
        all_categories = _distinct('category')
        all_ranges = _distinct('product_range')
        all_cities = _distinct('city')
        lead_sources = conn.execute(
            "SELECT name FROM `lead_sources` WHERE is_active=1 ORDER BY sort_order"
        ).fetchall() or []
        lead_categories = conn.execute(
            "SELECT name FROM `lead_categories` WHERE is_active=1 ORDER BY sort_order"
        ).fetchall() or []
        product_ranges = conn.execute(
            "SELECT name FROM `product_ranges` WHERE is_active=1 ORDER BY sort_order"
        ).fetchall() or []
        all_users = get_team_users()

        return render_template(
            'crm/leads/leads.html',
            leads=all_leads, counts=counts, total_count=total_count,
            list_columns=LEAD_LIST_COLUMNS,
            deleted_count=deleted_count, show_trash=show_trash,
            npd_count=npd_count, epd_count=epd_count,
            all_users=all_users, status=status, search=search,
            source=source, category=category, p_range=p_range, city=city,
            date_from=date_from, date_to=date_to,
            sort_by=sort_by, sort_dir=sort_dir,
            all_sources=all_sources, all_categories=all_categories,
            all_ranges=all_ranges, all_cities=all_cities,
            lead_statuses=statuses, lead_sources=lead_sources,
            lead_categories=lead_categories, product_ranges=product_ranges,
            is_admin=_is_admin(), is_admin_mgr=_is_admin_mgr(),
            sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
            active_item='leads', user_name=_uname(), role=session.get('User_Type'),
        )
    finally:
        conn.close()


def _lead_age(ld):
    created = ld.get('created_at')
    if not created:
        return 0
    if isinstance(created, str):
        try:
            created = datetime.strptime(created[:19], '%Y-%m-%d %H:%M:%S')
        except Exception:
            return 0
    cdate = created.date()
    if ld.get('status') in ('close', 'cancel') and ld.get('closed_at'):
        end = ld['closed_at']
        if isinstance(end, str):
            try:
                end = datetime.strptime(end[:19], '%Y-%m-%d %H:%M:%S')
            except Exception:
                end = datetime.now()
        edate = end.date()
    else:
        edate = datetime.now().date()
    return max(0, (edate - cdate).days)


@crm_bp.app_template_filter('fmtdt')
def _fmtdt(val, fmt='%d %b %Y'):
    """Safe date formatter — works whether the DB returns datetime or str."""
    if val is None or val == '':
        return ''
    if isinstance(val, str):
        s = val.strip()
        parsed = None
        for f in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M',
                  '%Y-%m-%d'):
            try:
                parsed = datetime.strptime(s[:19], f)
                break
            except Exception:
                continue
        if parsed is None:
            return s  # unknown format → show raw
        val = parsed
    try:
        return val.strftime(fmt)
    except Exception:
        return str(val)


# ─────────────────────────────────────────────────────────────────────────────
# ADD
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/add', methods=['GET', 'POST'])
@login_required
def lead_add():
    conn = _db()
    try:
        if request.method == 'POST':
            f = request.form
            code = gen_lead_code(conn)
            team = ','.join(request.form.getlist('team_members'))
            cur = conn.execute(
                "INSERT INTO `leads` "
                "(code, title, contact_name, company_name, email, website, phone, "
                " alternate_mobile, source, status, lead_type, priority, "
                " expected_value, average_cost, assigned_to, follow_up_date, "
                " notes, position, address, city, state, country, zip_code, "
                " product_name, category, product_range, order_quantity, "
                " requirement_spec, tags, remark, team_members, created_by, "
                " created_at) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,"
                "%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())",
                (code, f.get('title'), f.get('contact_name'), f.get('company_name'),
                 f.get('email'), f.get('website'), f.get('phone'),
                 f.get('alternate_mobile'), f.get('source'),
                 f.get('status', 'open'), f.get('lead_type', 'Quality'),
                 f.get('priority', 'medium'),
                 f.get('expected_value') or None,
                 f.get('average_cost') or None,
                 # AUTO-ASSIGN: form me kisi ko assign nahi kiya to
                 # creator ko hi assign kar do.
                 f.get('assigned_to') or _uid(),
                 f.get('follow_up_date') or None,
                 f.get('notes'), f.get('position'), f.get('address'),
                 f.get('city'), f.get('state'), f.get('country', 'India'),
                 f.get('zip_code'), f.get('product_name'), f.get('category'),
                 f.get('product_range'), f.get('order_quantity'),
                 f.get('requirement_spec'), f.get('tags'), f.get('remark'),
                 team, _uid()))
            new_id = cur.lastrowid
            if f.get('client_id'):
                conn.execute("UPDATE `leads` SET client_id=%s WHERE id=%s",
                             (f.get('client_id'), new_id))
            log_activity(conn, new_id, f"Lead created ({code})")
            conn.commit()
            flash(f'Lead {code} created.', 'success')
            return redirect(url_for('crm.lead_view', id=new_id))

        return render_template(
            'crm/leads/lead_form.html', lead=None, mode='add',
            all_users=get_team_users(),
            clients=conn.execute("SELECT id, code, company_name, contact_name FROM `client_masters` WHERE is_deleted=0 ORDER BY company_name").fetchall() or [],
            lead_statuses=_active(conn, 'lead_statuses'),
            lead_sources=_active(conn, 'lead_sources'),
            lead_categories=_active(conn, 'lead_categories'),
            product_ranges=_active(conn, 'product_ranges'),
            sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
            active_item='leads', user_name=_uname(),
            role=session.get('User_Type'))
    finally:
        conn.close()


def _active(conn, table):
    return conn.execute(
        f"SELECT * FROM `{table}` WHERE is_active=1 ORDER BY sort_order"
    ).fetchall() or []


# ─────────────────────────────────────────────────────────────────────────────
# EDIT
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def lead_edit(id):
    conn = _db()
    try:
        lead = conn.execute("SELECT * FROM `leads` WHERE id=%s", (id,)).fetchone()
        if not lead:
            abort(404)
        if request.method == 'POST':
            f = request.form
            team = ','.join(request.form.getlist('team_members'))
            conn.execute(
                "UPDATE `leads` SET title=%s, contact_name=%s, company_name=%s, "
                "email=%s, website=%s, phone=%s, alternate_mobile=%s, source=%s, "
                "lead_type=%s, priority=%s, expected_value=%s, average_cost=%s, "
                "assigned_to=%s, follow_up_date=%s, notes=%s, position=%s, "
                "address=%s, city=%s, state=%s, country=%s, zip_code=%s, "
                "product_name=%s, category=%s, product_range=%s, order_quantity=%s, "
                "requirement_spec=%s, tags=%s, remark=%s, team_members=%s, "
                "modified_by=%s WHERE id=%s",
                (f.get('title'), f.get('contact_name'), f.get('company_name'),
                 f.get('email'), f.get('website'), f.get('phone'),
                 f.get('alternate_mobile'), f.get('source'),
                 f.get('lead_type', 'Quality'), f.get('priority', 'medium'),
                 f.get('expected_value') or None, f.get('average_cost') or None,
                 f.get('assigned_to') or None,
                 f.get('follow_up_date') or None, f.get('notes'),
                 f.get('position'), f.get('address'), f.get('city'),
                 f.get('state'), f.get('country', 'India'), f.get('zip_code'),
                 f.get('product_name'), f.get('category'), f.get('product_range'),
                 f.get('order_quantity'), f.get('requirement_spec'),
                 f.get('tags'), f.get('remark'), team, _uid(), id))
            conn.execute("UPDATE `leads` SET client_id=%s WHERE id=%s",
                         (f.get('client_id') or None, id))
            log_activity(conn, id, "Lead details edited")
            add_contribution(conn, id, 'edit')
            conn.commit()
            flash('Lead updated.', 'success')
            return redirect(url_for('crm.lead_view', id=id))

        lead['_team_ids'] = [int(x) for x in (lead.get('team_members') or '').split(',')
                             if x.strip().isdigit()]
        return render_template(
            'crm/leads/lead_form.html', lead=lead, mode='edit',
            all_users=get_team_users(),
            clients=conn.execute("SELECT id, code, company_name, contact_name FROM `client_masters` WHERE is_deleted=0 ORDER BY company_name").fetchall() or [],
            lead_statuses=_active(conn, 'lead_statuses'),
            lead_sources=_active(conn, 'lead_sources'),
            lead_categories=_active(conn, 'lead_categories'),
            product_ranges=_active(conn, 'product_ranges'),
            sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
            active_item='leads', user_name=_uname(),
            role=session.get('User_Type'))
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# VIEW (detail)
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/<int:id>')
@login_required
def lead_view(id):
    from werkzeug.exceptions import HTTPException
    conn = _db()
    try:
        lead = conn.execute("SELECT * FROM `leads` WHERE id=%s", (id,)).fetchone()
        if not lead:
            abort(404)
        umap = {}
        try:
            umap = _user_map(conn) or {}
        except Exception as e:
            print(f"[lead_view] _user_map failed: {e}")
        try:
            lead['_age'] = _lead_age(lead)
        except Exception:
            lead['_age'] = 0
        lead['_assigned_name'] = umap.get(lead.get('assigned_to'), '')
        try:
            lead['_team'] = [{'id': int(x), 'name': umap.get(int(x), f'#{x}')}
                             for x in (str(lead.get('team_members') or '')).split(',')
                             if str(x).strip().isdigit()]
        except Exception:
            lead['_team'] = []
        lead['_created_name'] = umap.get(lead.get('created_by'), '')

        def q(sql, params=()):
            try:
                return conn.execute(sql, params).fetchall() or []
            except Exception as e:
                print(f"[lead_view] sub-query failed: {e}")
                return []

        discussions = q("SELECT * FROM `lead_discussions` WHERE lead_id=%s "
                        "ORDER BY created_at DESC", (id,))
        for d in discussions:
            d['_user'] = umap.get(d.get('user_id'), 'User')
            d['_files'] = q("SELECT * FROM `lead_attachments` WHERE discussion_id=%s",
                            (d['id'],))

        reminders = q("SELECT * FROM `lead_reminders` WHERE lead_id=%s ORDER BY remind_at",
                      (id,))
        for r in reminders:
            r['_user'] = umap.get(r.get('user_id'), 'User')

        notes = q("SELECT * FROM `lead_notes` WHERE lead_id=%s AND user_id=%s "
                  "ORDER BY created_at DESC", (id, _uid()))

        logs = q("SELECT * FROM `lead_activity_logs` WHERE lead_id=%s "
                 "ORDER BY created_at DESC LIMIT 100", (id,))
        for lg in logs:
            lg['_user'] = umap.get(lg.get('user_id'), 'System')

        attachments = q("SELECT * FROM `lead_attachments` WHERE lead_id=%s "
                        "ORDER BY created_at DESC", (id,))
        for a in attachments:
            a['_user'] = umap.get(a.get('uploaded_by'), 'User')

        contrib = q("SELECT user_id, SUM(points) AS pts, COUNT(*) AS cnt "
                    "FROM `lead_contributions` WHERE lead_id=%s GROUP BY user_id "
                    "ORDER BY pts DESC", (id,))
        for c in contrib:
            c['_user'] = umap.get(c.get('user_id'), 'User')

        try:
            statuses = _active(conn, 'lead_statuses')
        except Exception:
            statuses = []

        return render_template(
            'crm/leads/lead_view.html', lead=lead, discussions=discussions,
            reminders=reminders, notes=notes, logs=logs, attachments=attachments,
            contributions=contrib, lead_statuses=statuses,
            all_users=[{'id': k, 'name': v} for k, v in (umap or {}).items()],
            is_admin=_is_admin(), is_admin_mgr=_is_admin_mgr(), my_uid=_uid(),
            sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
            active_item='leads', user_name=_uname(),
            role=session.get('User_Type'))
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(tb)
        return ("<pre style='padding:24px;font:13px/1.5 monospace;white-space:pre-wrap;"
                "color:#b91c1c'>LEAD VIEW ERROR:\n\n" + str(e) + "\n\n" + tb + "</pre>"), 500
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# DELETE / RESTORE / PERMANENT
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/<int:id>/delete', methods=['POST'])
@login_required
def lead_delete(id):
    conn = _db()
    try:
        conn.execute("UPDATE `leads` SET is_deleted=1, deleted_at=NOW() "
                     "WHERE id=%s", (id,))
        log_activity(conn, id, "Lead moved to trash")
        conn.commit()
        if request.is_json or request.headers.get('X-Requested-With'):
            return jsonify(ok=True)
        flash('Lead moved to trash.', 'success')
        return redirect(url_for('crm.leads'))
    finally:
        conn.close()


@crm_bp.route('/leads/<int:id>/restore', methods=['POST'])
@login_required
def lead_restore(id):
    conn = _db()
    try:
        conn.execute("UPDATE `leads` SET is_deleted=0, deleted_at=NULL "
                     "WHERE id=%s", (id,))
        log_activity(conn, id, "Lead restored from trash")
        conn.commit()
        if request.headers.get('X-Requested-With'):
            return jsonify(ok=True)
        flash('Lead restored.', 'success')
        return redirect(url_for('crm.leads', trash='1'))
    finally:
        conn.close()


@crm_bp.route('/leads/<int:id>/permanent-delete', methods=['POST'])
@login_required
def lead_permanent_delete(id):
    if not _is_admin():
        return jsonify(ok=False, error='Admin only'), 403
    conn = _db()
    try:
        for t in ('lead_discussions', 'lead_attachments', 'lead_reminders',
                  'lead_notes', 'lead_activity_logs', 'lead_contributions'):
            conn.execute(f"DELETE FROM `{t}` WHERE lead_id=%s", (id,))
        conn.execute("DELETE FROM `leads` WHERE id=%s", (id,))
        conn.commit()
        if request.headers.get('X-Requested-With'):
            return jsonify(ok=True)
        flash('Lead permanently deleted.', 'success')
        return redirect(url_for('crm.leads', trash='1'))
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# STATUS CHANGE  (workflow)
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/<int:id>/status', methods=['POST'])
@crm_bp.route('/leads/<int:id>/update-status', methods=['POST'])
@login_required
def lead_update_status(id):
    new_status = (request.form.get('status')
                  or (request.json or {}).get('status') if request.is_json
                  else request.form.get('status'))
    lost_reason = request.form.get('lost_reason', '')
    if not new_status:
        return jsonify(ok=False, error='status required'), 400
    conn = _db()
    try:
        lead = conn.execute("SELECT * FROM `leads` WHERE id=%s", (id,)).fetchone()
        if not lead:
            return jsonify(ok=False, error='not found'), 404
        old = lead.get('status')
        closed = "NOW()" if new_status in ('close', 'cancel') else "NULL"
        conn.execute(
            f"UPDATE `leads` SET status=%s, lost_reason=%s, "
            f"closed_at={closed}, modified_by=%s WHERE id=%s",
            (new_status, lost_reason or lead.get('lost_reason'), _uid(), id))
        log_activity(conn, id, f"Status: {old} → {new_status}")
        add_contribution(conn, id, 'status_change',
                         note=f"{old} → {new_status}")
        if new_status == 'close':
            lead['status'] = new_status
            _handle_close_contribution(conn, lead)
        elif new_status == 'cancel':
            add_contribution(conn, id, 'cancel')
        conn.commit()
        if request.headers.get('X-Requested-With') or request.is_json:
            return jsonify(ok=True, status=new_status)
        flash('Status updated.', 'success')
        return redirect(url_for('crm.lead_view', id=id))
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# INLINE EDIT (ajax single field)
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/<int:id>/discussion/add', methods=['POST'])
@login_required
def lead_discussion_add(id):
    comment = (request.form.get('comment') or '').strip()
    files = request.files.getlist('files')
    if not comment and not any(f.filename for f in files):
        if request.is_json or request.headers.get('X-Requested-With'):
            return jsonify(ok=False, error='A comment or file is required.'), 400
        flash('A comment or file is required.', 'error')
        return redirect(url_for('crm.lead_view', id=id) + '#discussion')
    conn = _db()
    try:
        cur = conn.execute(
            "INSERT INTO `lead_discussions` (lead_id, user_id, comment, created_at) "
            "VALUES (%s, %s, %s, NOW())", (id, _uid(), comment or '(file)'))
        did = cur.lastrowid
        saved_files = []
        for fs in files:
            if not fs or not fs.filename:
                continue
            if not _allowed_file(fs.filename):
                continue
            fname = secure_filename(fs.filename)
            stamp = datetime.now().strftime('%Y%m%d%H%M%S')
            stored = f"{id}_{did}_{stamp}_{fname}"
            path = os.path.join(_upload_dir(), stored)
            fs.save(path)
            size = os.path.getsize(path)
            conn.execute(
                "INSERT INTO `lead_attachments` "
                "(lead_id, discussion_id, file_name, file_path, file_size, "
                " file_type, uploaded_by, created_at) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,NOW())",
                (id, did, fname, f"leads/{stored}", size,
                 fname.rsplit('.', 1)[-1].lower(), _uid()))
            _aid = conn.execute(
                "SELECT id FROM `lead_attachments` WHERE discussion_id=%s "
                "ORDER BY id DESC LIMIT 1", (did,)).fetchone()
            saved_files.append({'id': (_aid or {}).get('id'),
                                'name': fname, 'path': f"leads/{stored}"})
        log_activity(conn, id, "Added a discussion comment")
        add_contribution(conn, id, 'comment')
        conn.commit()
        if request.is_json or request.headers.get('X-Requested-With'):
            return jsonify(ok=True, id=did, user=_uname(), comment=comment,
                           created_at=datetime.now().strftime('%d-%m-%Y %I:%M %p'),
                           files=saved_files)
        flash('Comment added.', 'success')
    finally:
        conn.close()
    return redirect(url_for('crm.lead_view', id=id) + '#discussion')


@crm_bp.route('/leads/discussion/<int:did>/delete', methods=['POST'])
@login_required
def lead_discussion_delete(did):
    conn = _db()
    try:
        d = conn.execute("SELECT * FROM `lead_discussions` WHERE id=%s",
                         (did,)).fetchone()
        if not d:
            return jsonify(ok=False), 404
        if d['user_id'] != _uid():
            return jsonify(ok=False, error='You can only delete your own comments'), 403
        lead_id = d['lead_id']
        atts = conn.execute(
            "SELECT file_path FROM `lead_attachments` WHERE discussion_id=%s",
            (did,)).fetchall() or []
        for a in atts:
            _remove_upload(a['file_path'])
        conn.execute("DELETE FROM `lead_attachments` WHERE discussion_id=%s", (did,))
        conn.execute("DELETE FROM `lead_discussions` WHERE id=%s", (did,))
        conn.commit()
        if request.headers.get('X-Requested-With'):
            return jsonify(ok=True)
        flash('Comment deleted.', 'success')
        return redirect(url_for('crm.lead_view', id=lead_id) + '#discussion')
    finally:
        conn.close()


@crm_bp.route('/leads/discussion/<int:did>/edit', methods=['POST'])
@login_required
def lead_discussion_edit(did):
    """Sirf apni comment edit (owner-only). Text + attachments change kar sakte ho."""
    conn = _db()
    try:
        d = conn.execute("SELECT * FROM `lead_discussions` WHERE id=%s",
                         (did,)).fetchone()
        if not d:
            return jsonify(ok=False, error='Not found'), 404
        if d['user_id'] != _uid():
            return jsonify(ok=False, error='You can only edit your own comments.'), 403
        comment = (request.form.get('comment') or '').strip()
        removed = [x for x in (request.form.get('removed') or '').split(',')
                   if x.strip().isdigit()]
        new_files = request.files.getlist('files')
        lead_id = d['lead_id']

        conn.execute("UPDATE `lead_discussions` SET comment=%s WHERE id=%s",
                     (comment, did))

        # selected attachments hatao (file + DB dono se)
        for aid in removed:
            arow = conn.execute(
                "SELECT file_path FROM `lead_attachments` WHERE id=%s AND discussion_id=%s",
                (aid, did)).fetchone()
            if arow:
                _remove_upload(arow['file_path'])
                conn.execute("DELETE FROM `lead_attachments` WHERE id=%s", (aid,))

        # naye attachments add
        for fs in new_files:
            if not fs or not fs.filename or not _allowed_file(fs.filename):
                continue
            fname = secure_filename(fs.filename)
            stamp = datetime.now().strftime('%Y%m%d%H%M%S')
            stored = f"{lead_id}_{did}_{stamp}_{fname}"
            path = os.path.join(_upload_dir(), stored)
            fs.save(path)
            size = os.path.getsize(path)
            conn.execute(
                "INSERT INTO `lead_attachments` "
                "(lead_id, discussion_id, file_name, file_path, file_size, "
                " file_type, uploaded_by, created_at) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,NOW())",
                (lead_id, did, fname, f"leads/{stored}", size,
                 fname.rsplit('.', 1)[-1].lower(), _uid()))

        try:
            log_activity(conn, lead_id, "Edited a discussion comment")
        except Exception:
            pass
        conn.commit()
        frows = conn.execute(
            "SELECT id, file_name, file_path FROM `lead_attachments` "
            "WHERE discussion_id=%s ORDER BY id", (did,)).fetchall() or []
        files_out = [{'id': fr['id'], 'name': fr['file_name'],
                      'path': fr['file_path']} for fr in frows]
        return jsonify(ok=True, comment=comment, files=files_out)
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# REMINDERS
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/<int:id>/reminder/add', methods=['POST'])
@login_required
def lead_reminder_add(id):
    title = (request.form.get('title') or '').strip()
    remind_at = request.form.get('remind_at')
    desc = request.form.get('description', '')
    _ajax = request.is_json or request.headers.get('X-Requested-With')
    rdt = _parse_dt(remind_at)
    if not title or not rdt:
        if _ajax:
            return jsonify(ok=False, error='Title and a valid Date & Time are required.'), 400
        flash('Title and date/time are required.', 'error')
        return redirect(url_for('crm.lead_view', id=id) + '#reminders')
    conn = _db()
    try:
        cur = conn.execute(
            "INSERT INTO `lead_reminders` "
            "(lead_id, user_id, title, description, remind_at, created_at) "
            "VALUES (%s,%s,%s,%s,%s,NOW())",
            (id, _uid(), title, desc, rdt))
        rid = cur.lastrowid
        log_activity(conn, id, f"Reminder set: {title}")
        add_contribution(conn, id, 'reminder')
        conn.commit()
        if _ajax:
            disp = datetime.strptime(rdt, '%Y-%m-%d %H:%M:%S').strftime('%d-%m-%Y %H:%M')
            return jsonify(ok=True, id=rid, title=title, description=desc,
                           remind_at=disp, user=_uname())
        flash('Reminder added.', 'success')
    finally:
        conn.close()
    return redirect(url_for('crm.lead_view', id=id) + '#reminders')


@crm_bp.route('/leads/reminder/<int:rid>/done', methods=['POST'])
@login_required
def lead_reminder_done(rid):
    conn = _db()
    try:
        conn.execute("UPDATE `lead_reminders` SET is_done=1 WHERE id=%s", (rid,))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@crm_bp.route('/leads/reminder/<int:rid>/delete', methods=['POST'])
@login_required
def lead_reminder_delete(rid):
    conn = _db()
    try:
        conn.execute("DELETE FROM `lead_reminders` WHERE id=%s", (rid,))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


def _ensure_reminder_cols(conn):
    """lead_reminders me 'notified' column ensure karo (once)."""
    try:
        conn.execute("ALTER TABLE `lead_reminders` ADD COLUMN `notified` "
                     "TINYINT DEFAULT 0")
        conn.commit()
    except Exception:
        pass


@crm_bp.route('/reminders/due', methods=['GET'])
@login_required
def reminders_due():
    """Current user ke jo reminders due ho gaye + abhi tak notify nahi hue."""
    conn = _db()
    try:
        _ensure_reminder_cols(conn)
        rows = conn.execute(
            "SELECT r.id, r.lead_id, r.title, r.description, r.remind_at, "
            "       l.code AS lead_code, l.contact_name AS contact "
            "FROM `lead_reminders` r "
            "LEFT JOIN `leads` l ON l.id = r.lead_id "
            "WHERE r.user_id=%s AND r.is_done=0 "
            "  AND (r.notified=0 OR r.notified IS NULL) "
            "  AND r.remind_at <= NOW() "
            "ORDER BY r.remind_at", (_uid(),)).fetchall() or []
        if rows:
            ids = [r['id'] for r in rows]
            conn.execute(
                "UPDATE `lead_reminders` SET notified=1 WHERE id IN (%s)"
                % ",".join(['%s'] * len(ids)), tuple(ids))
            conn.commit()
        out = []
        for r in rows:
            try:
                disp = datetime.strptime(str(r['remind_at']), '%Y-%m-%d %H:%M:%S').strftime('%d %b %Y %I:%M %p')
            except Exception:
                disp = str(r['remind_at'])
            out.append({'id': r['id'], 'lead_id': r['lead_id'],
                        'title': r['title'], 'description': r['description'] or '',
                        'lead_code': r['lead_code'] or '',
                        'contact': r['contact'] or '',
                        'remind_at': disp})
        return jsonify(ok=True, reminders=out)
    except Exception:
        return jsonify(ok=True, reminders=[])
    finally:
        conn.close()


@crm_bp.route('/leads/reminder/<int:rid>/snooze', methods=['POST'])
@login_required
def lead_reminder_snooze(rid):
    """Reminder ko +N minutes aage badha do (default 5), notified reset."""
    try:
        mins = int(request.form.get('minutes') or request.args.get('minutes') or 5)
    except Exception:
        mins = 5
    conn = _db()
    try:
        _ensure_reminder_cols(conn)
        conn.execute(
            "UPDATE `lead_reminders` "
            "SET remind_at = DATE_ADD(NOW(), INTERVAL %s MINUTE), notified=0 "
            "WHERE id=%s", (mins, rid))
        conn.commit()
        return jsonify(ok=True, minutes=mins)
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# PERSONAL NOTES (private per user)
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/<int:id>/note/add', methods=['POST'])
@login_required
def lead_note_add(id):
    note = (request.form.get('note') or '').strip()
    _ajax = request.is_json or request.headers.get('X-Requested-With')
    if not note:
        if _ajax:
            return jsonify(ok=False, error='Note cannot be empty.'), 400
        return redirect(url_for('crm.lead_view', id=id) + '#notes')
    conn = _db()
    try:
        cur = conn.execute(
            "INSERT INTO `lead_notes` (lead_id, user_id, note, created_at) "
            "VALUES (%s,%s,%s,NOW())", (id, _uid(), note))
        nid = cur.lastrowid
        conn.commit()
        if _ajax:
            return jsonify(ok=True, id=nid, note=note,
                           created_at=datetime.now().strftime('%d-%m-%Y %I:%M %p'))
        flash('Note saved.', 'success')
    finally:
        conn.close()
    return redirect(url_for('crm.lead_view', id=id) + '#notes')


@crm_bp.route('/leads/note/<int:nid>/delete', methods=['POST'])
@login_required
def lead_note_delete(nid):
    conn = _db()
    try:
        n = conn.execute("SELECT * FROM `lead_notes` WHERE id=%s", (nid,)).fetchone()
        if n and (n['user_id'] == _uid() or _is_admin()):
            conn.execute("DELETE FROM `lead_notes` WHERE id=%s", (nid,))
            conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# EMAIL — SMTP direct send (no-reply@hcpwellness.in)
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/<int:id>/email/send', methods=['POST'])
@login_required
def lead_email_send(id):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.utils import formataddr

    def _mailcfg(key, default=''):
        """MAIL_* value app.config -> env -> config.py module se dhundo."""
        v = current_app.config.get(key)
        if v not in (None, ''):
            return v
        v = os.environ.get(key)
        if v not in (None, ''):
            return v
        for modname in ('config', 'app_config', 'settings', 'configuration'):
            try:
                mod = __import__(modname)
                v = getattr(mod, key, None)
                if v not in (None, ''):
                    return v
            except Exception:
                continue
        return default

    to_addr = (request.form.get('to') or '').strip()
    from_name = (request.form.get('from_name') or 'HCP Wellness Pvt. Ltd.').strip()
    from_email = (request.form.get('from_email') or '').strip()
    subject = (request.form.get('subject') or '').strip()
    body_html = request.form.get('body') or ''
    if not to_addr or not subject:
        return jsonify(ok=False, error='To and Subject are required.'), 400

    server = _mailcfg('MAIL_SERVER', 'smtp.gmail.com')
    port = int(_mailcfg('MAIL_PORT', 587) or 587)
    use_tls = _mailcfg('MAIL_USE_TLS', True)
    username = _mailcfg('MAIL_USERNAME', 'no-reply@hcpwellness.in')
    password = _mailcfg('MAIL_PASSWORD', '')
    sender = username or from_email or 'no-reply@hcpwellness.in'
    if not from_email:
        from_email = sender

    if not password:
        return jsonify(ok=False, error='SMTP password not found. Set '
                       'MAIL_PASSWORD in config.py (or load it in app.config).'), 500

    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = formataddr((from_name, from_email))
        msg['To'] = to_addr
        msg['Reply-To'] = from_email
        import re as _re
        plain = _re.sub(r'<br\s*/?>', '\n', body_html)
        plain = _re.sub(r'</li>', '\n', plain)
        plain = _re.sub(r'<[^>]+>', '', plain)
        msg.attach(MIMEText(plain, 'plain', 'utf-8'))
        msg.attach(MIMEText(body_html, 'html', 'utf-8'))

        s = smtplib.SMTP(server, port, timeout=25)
        if use_tls:
            s.starttls()
        s.login(username, password)
        s.sendmail(sender, [to_addr], msg.as_string())
        s.quit()

        try:
            conn = _db()
            log_activity(conn, id, f"Email sent to {to_addr}")
            conn.commit()
            conn.close()
        except Exception:
            pass
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error='Email bhejne me dikkat: ' + str(e)), 500


# ─────────────────────────────────────────────────────────────────────────────
# ATTACHMENT serve / download
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/attachment/<int:aid>')
@login_required
def lead_attachment_serve(aid):
    conn = _db()
    try:
        a = conn.execute("SELECT * FROM `lead_attachments` WHERE id=%s",
                         (aid,)).fetchone()
        if not a:
            abort(404)
        full = os.path.join(current_app.root_path, 'static', 'uploads',
                            a['file_path'])
        if not os.path.exists(full):
            abort(404)
        return send_file(full, as_attachment=False,
                         download_name=a['file_name'])
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# MASTER QUICK-ADD ( + buttons: status / source / category / range )
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/master/<kind>/add', methods=['POST'])
@login_required
def master_add(kind):
    name = (request.form.get('name') or '').strip()
    if not name:
        return jsonify(ok=False, error='name required'), 400
    tbl = {'status': 'lead_statuses', 'source': 'lead_sources',
           'category': 'lead_categories', 'range': 'product_ranges'}.get(kind)
    if not tbl:
        return jsonify(ok=False, error='bad kind'), 400
    key = name.lower().replace(' ', '_')
    conn = _db()
    try:
        if tbl == 'lead_statuses':
            label = request.form.get('label') or name
            color = request.form.get('color') or '#64748b'
            conn.execute(
                "INSERT INTO `lead_statuses` (name,label,color,sort_order,is_active)"
                " VALUES (%s,%s,%s,99,1) ON DUPLICATE KEY UPDATE is_active=1",
                (key, label, color))
            conn.commit()
            return jsonify(ok=True, name=key, label=label)
        conn.execute(
            f"INSERT INTO `{tbl}` (name,sort_order,is_active) "
            f"SELECT %s,99,1 FROM DUAL WHERE NOT EXISTS "
            f"(SELECT 1 FROM `{tbl}` WHERE name=%s)", (name, name))
        conn.execute(f"UPDATE `{tbl}` SET is_active=1 WHERE name=%s", (name,))
        conn.commit()
        return jsonify(ok=True, name=name, label=name)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# IMPORT (CSV / Excel)
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/import', methods=['GET', 'POST'])
@login_required
def leads_import():
    if request.method == 'GET':
        return render_template(
            'crm/leads/import_leads.html',
            sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
            active_item='imp-leads', user_name=_uname(),
            role=session.get('User_Type'))
    f = request.files.get('file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('crm.leads_import'))
    name = f.filename.lower()
    rows = []
    try:
        if name.endswith('.csv'):
            import csv
            data = f.read().decode('utf-8-sig', errors='ignore').splitlines()
            for r in csv.DictReader(data):
                rows.append({(k or '').strip().lower(): (str(v).strip() if v is not None else '')
                             for k, v in r.items()})
        elif name.endswith(('.xlsx', '.xls')):
            from openpyxl import load_workbook
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            hdr = [str(h or '').strip().lower() for h in next(it, [])]
            for r in it:
                rows.append({hdr[i]: (str(c).strip() if c is not None else '')
                             for i, c in enumerate(r) if i < len(hdr)})
        else:
            flash('Only .csv or .xlsx files are allowed.', 'error')
            return redirect(url_for('crm.leads_import'))
    except Exception as e:
        flash(f'Could not read the file: {e}', 'error')
        return redirect(url_for('crm.leads_import'))

    alias = {
        'name': 'contact_name', 'contact name': 'contact_name', 'contact_name': 'contact_name',
        'company': 'company_name', 'company name': 'company_name', 'company_name': 'company_name',
        'email': 'email', 'mobile': 'phone', 'phone': 'phone',
        'alt mobile': 'alternate_mobile', 'alternate_mobile': 'alternate_mobile',
        'product': 'product_name', 'product name': 'product_name', 'product_name': 'product_name',
        'category': 'category', 'product range': 'product_range', 'product_range': 'product_range',
        'source': 'source', 'city': 'city', 'state': 'state',
        'country': 'country', 'zip': 'zip_code', 'zip_code': 'zip_code', 'pincode': 'zip_code',
        'status': 'status', 'lead type': 'lead_type', 'lead_type': 'lead_type',
        'priority': 'priority', 'website': 'website', 'position': 'position', 'address': 'address',
        'tags': 'tags', 'order quantity': 'order_quantity', 'order_quantity': 'order_quantity',
        'requirement_spec': 'requirement_spec', 'requirement spec': 'requirement_spec',
        'remark': 'remark', 'remarks': 'remark',
    }
    LEAD_COLS = ['contact_name', 'company_name', 'email', 'phone', 'alternate_mobile',
                 'website', 'position', 'address', 'city', 'state', 'country', 'zip_code',
                 'product_name', 'category', 'product_range', 'order_quantity',
                 'requirement_spec', 'remark', 'source', 'tags', 'lead_type', 'priority']
    valid_status = {'open', 'in_process', 'close', 'cancel'}
    conn = _db()
    inserted, skipped = 0, 0
    try:
        for r in rows:
            rec = {}
            for k, v in r.items():
                if k in alias and v:
                    rec[alias[k]] = v
            cname = rec.get('contact_name')
            if not cname:
                skipped += 1
                continue
            st = (rec.get('status') or 'open').lower().replace(' ', '_')
            if st not in valid_status:
                st = 'open'
            code = gen_lead_code(conn)
            cols, ph, vals = ['code'], ['%s'], [code]
            for c in LEAD_COLS:
                if rec.get(c):
                    cols.append(c); ph.append('%s'); vals.append(rec[c])
            cols += ['status', 'created_by', 'created_at']
            ph += ['%s', '%s', 'NOW()']
            vals += [st, _uid()]
            # AUTO-ASSIGN: import file me assigned_to nahi tha to
            # import karne wale (creator) ko hi assign.
            if 'assigned_to' not in cols:
                cols.append('assigned_to'); ph.append('%s'); vals.append(_uid())
            sql = ("INSERT INTO `leads` (" + ",".join("`" + c + "`" for c in cols)
                   + ") VALUES (" + ",".join(ph) + ")")
            conn.execute(sql, vals)
            inserted += 1
        conn.commit()
        flash(f'{inserted} leads imported.'
              + (f' {skipped} skipped (name missing).' if skipped else ''), 'success')
    except Exception as e:
        flash(f'Import error: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('crm.leads'))


@crm_bp.route('/leads/import/template')
@login_required
def leads_import_template():
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = 'Leads'
    headers = ['name', 'position', 'email', 'mobile', 'alternate_mobile', 'company',
               'website', 'city', 'state', 'country', 'zip_code', 'product_name',
               'category', 'product_range', 'order_quantity', 'source', 'status',
               'lead_type', 'requirement_spec', 'remark', 'tags']
    ws.append(headers)
    ws.append(['John Doe', 'CEO', 'john@abc.com', '9999999999', '8888888888', 'ABC Corp',
               'www.abc.com', 'Mumbai', 'Maharashtra', 'India', '400001', 'Face Wash',
               'Skin Care', 'Premium', '500 units', 'HCP Website', 'open', 'Quality',
               'Vitamin C face wash', 'Urgent requirement', 'skincare, premium'])
    dv_st = DataValidation(type='list', formula1='"open,in_process,close,cancel"', allow_blank=True)
    dv_lt = DataValidation(type='list', formula1='"Quality,Non-Quality"', allow_blank=True)
    ws.add_data_validation(dv_st); ws.add_data_validation(dv_lt)
    dv_st.add('Q2:Q1000'); dv_lt.add('R2:R1000')
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    from flask import send_file
    return send_file(bio, as_attachment=True, download_name='leads_import_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT (Excel)
# ─────────────────────────────────────────────────────────────────────────────
@crm_bp.route('/leads/<int:id>/inline-edit', methods=['POST'])
@login_required
def lead_inline_edit(id):
    """List page se cell-level quick edit (AJAX JSON: {field, value})."""
    data = request.get_json(silent=True) or {}
    field = (data.get('field') or '').strip()
    value = data.get('value', '')
    allowed = {
        'contact_name', 'company_name', 'email', 'phone', 'alternate_mobile',
        'website', 'position', 'city', 'state', 'country', 'zip_code', 'source',
        'product_name', 'category', 'product_range', 'order_quantity', 'tags',
        'status', 'lead_type', 'average_cost',
    }
    if field not in allowed:
        return jsonify(success=False, error='Field not allowed'), 400
    if isinstance(value, str):
        value = value.strip()
    if field == 'lead_type' and value not in ('Quality', 'Non-Quality'):
        return jsonify(success=False, error='Invalid lead type'), 400
    if field == 'average_cost':
        try:
            value = float(value) if value not in ('', None) else None
        except Exception:
            return jsonify(success=False, error='Invalid number'), 400
    conn = _db()
    try:
        lead = conn.execute(
            "SELECT id FROM `leads` WHERE id=%s AND is_deleted=0", (id,)).fetchone()
        if not lead:
            return jsonify(success=False, error='Lead not found'), 404
        conn.execute(
            f"UPDATE `leads` SET `{field}`=%s, modified_by=%s, updated_at=NOW() "
            f"WHERE id=%s", (value, _uid(), id))
        try:
            log_activity(conn, id, f'Inline edit: {field} updated')
        except Exception:
            pass
        conn.commit()
        return jsonify(success=True, field=field, value=value)
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify(success=False, error=str(e)), 500
    finally:
        conn.close()


@crm_bp.route('/leads/export')
@login_required
def leads_export():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        flash('openpyxl not installed.', 'error')
        return redirect(url_for('crm.leads'))
    conn = _db()
    try:
        vis_frag, vis_params = _visibility_sql()
        rows = conn.execute(
            "SELECT * FROM `leads` WHERE is_deleted=0" + vis_frag
            + " ORDER BY created_at DESC", vis_params).fetchall() or []

        def d(val):
            """date → dd-mm-yyyy (string ya datetime dono)."""
            if not val:
                return ''
            if isinstance(val, str):
                v = val.strip()
                for f in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d'):
                    try:
                        from datetime import datetime as _dt
                        return _dt.strptime(v[:19], f).strftime('%d-%m-%Y')
                    except Exception:
                        continue
                return v
            try:
                return val.strftime('%d-%m-%Y')
            except Exception:
                return str(val)

        ST = {'open': 'Open', 'in_process': 'In Process',
              'close': 'Close', 'cancel': 'Cancel'}

        wb = Workbook()
        ws = wb.active
        ws.title = 'Leads'
        headers = ['Created Date', 'Code', 'Company', 'Contact Name', 'Position',
                   'Email', 'Mobile', 'Alt Mobile', 'Website', 'City', 'State',
                   'Country', 'Source', 'Product Name', 'Category', 'Product Range',
                   'Order Quantity', 'Expected Value', 'Priority', 'Status',
                   'Lead Type', 'Follow Up Date', 'Last Contact', 'Tags',
                   'Lead Age (Days)']
        ws.append(headers)
        for r in rows:
            ws.append([
                d(r.get('created_at')), r.get('code') or '',
                r.get('company_name') or '', r.get('contact_name') or '',
                r.get('position') or '', r.get('email') or '',
                r.get('phone') or '', r.get('alternate_mobile') or '',
                r.get('website') or '', r.get('city') or '', r.get('state') or '',
                r.get('country') or '', r.get('source') or '',
                r.get('product_name') or '', r.get('category') or '',
                r.get('product_range') or '', r.get('order_quantity') or '',
                (float(r['expected_value']) if r.get('expected_value') is not None
                 else ''),
                r.get('priority') or '', ST.get(r.get('status'), r.get('status') or ''),
                r.get('lead_type') or '', d(r.get('follow_up_date')),
                d(r.get('last_contact')), r.get('tags') or '',
                str(_lead_age(r)),
            ])

        # header styling
        hf = Font(bold=True, color='FFFFFF')
        fill = PatternFill('solid', fgColor='2563EB')
        for c in ws[1]:
            c.font = hf
            c.fill = fill
            c.alignment = Alignment(vertical='center')
        ws.freeze_panes = 'A2'
        widths = [13, 11, 24, 20, 14, 24, 13, 13, 18, 14, 14, 12, 16, 22, 14,
                  14, 14, 14, 10, 12, 11, 14, 14, 18, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        # Summary sheet
        s = wb.create_sheet('Summary')
        s['A1'] = 'HCP ERP — Leads Export'
        s['A1'].font = Font(bold=True, size=13)
        s['A3'] = 'Date'
        s['B3'] = datetime.now().strftime('%d-%m-%Y')
        s['A4'] = 'Total Leads'
        s['B4'] = len(rows)
        for cell in ('A3', 'A4'):
            s[cell].font = Font(bold=True)
        s.column_dimensions['A'].width = 18
        s.column_dimensions['B'].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"HCP_Leads_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return send_file(
            buf, as_attachment=True, download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.'
                     'spreadsheetml.sheet')
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# TABLE BOOTSTRAP  (idempotent — call once at app startup)
# ─────────────────────────────────────────────────────────────────────────────
def ensure_lead_tables():
    """Create all CRM Lead tables + seed masters/config if missing.
    Mirrors migrations/crm_leads_mysql.sql so first run never needs the CLI."""
    conn = _db()
    if not conn:
        print("[crm_leads] DB connection failed; tables not created.")
        return
    try:
        ddl = [
            """CREATE TABLE IF NOT EXISTS `lead_statuses` (
                id INT AUTO_INCREMENT PRIMARY KEY, name VARCHAR(30) NOT NULL,
                label VARCHAR(60) NOT NULL, color VARCHAR(20) DEFAULT '#64748b',
                icon VARCHAR(40), sort_order INT DEFAULT 0,
                is_active TINYINT(1) DEFAULT 1,
                UNIQUE KEY uq_lead_status_name (name)) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `lead_sources` (
                id INT AUTO_INCREMENT PRIMARY KEY, name VARCHAR(100) NOT NULL,
                sort_order INT DEFAULT 0, is_active TINYINT(1) DEFAULT 1,
                UNIQUE KEY uq_lead_source_name (name)) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `lead_categories` (
                id INT AUTO_INCREMENT PRIMARY KEY, name VARCHAR(100) NOT NULL,
                sort_order INT DEFAULT 0, is_active TINYINT(1) DEFAULT 1,
                UNIQUE KEY uq_lead_category_name (name)) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `product_ranges` (
                id INT AUTO_INCREMENT PRIMARY KEY, name VARCHAR(100) NOT NULL,
                sort_order INT DEFAULT 0, is_active TINYINT(1) DEFAULT 1,
                UNIQUE KEY uq_product_range_name (name)) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `leads` (
                id INT AUTO_INCREMENT PRIMARY KEY, code VARCHAR(30),
                title VARCHAR(200), contact_name VARCHAR(150) NOT NULL,
                company_name VARCHAR(200), email VARCHAR(150), website VARCHAR(200),
                phone VARCHAR(20), alternate_mobile VARCHAR(20), source VARCHAR(100),
                status VARCHAR(30) DEFAULT 'open', lead_type VARCHAR(20) DEFAULT 'Quality',
                priority VARCHAR(20) DEFAULT 'medium', expected_value DECIMAL(12,2),
                assigned_to INT, follow_up_date DATE, notes TEXT,
                lost_reason VARCHAR(200), customer_id INT, position VARCHAR(100),
                address TEXT, city VARCHAR(100), state VARCHAR(100),
                country VARCHAR(100) DEFAULT 'India', zip_code VARCHAR(10),
                average_cost DECIMAL(12,2) DEFAULT 0, product_name VARCHAR(200),
                category VARCHAR(100), product_range VARCHAR(100),
                order_quantity VARCHAR(100), requirement_spec TEXT, tags VARCHAR(300),
                remark TEXT, last_contact DATETIME, team_members TEXT,
                client_id INT, client_attachment VARCHAR(300),
                is_deleted TINYINT(1) NOT NULL DEFAULT 0, deleted_at DATETIME,
                closed_at DATETIME, created_by INT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP, modified_by INT,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_lead_code (code), KEY idx_lead_status (status),
                KEY idx_lead_deleted (is_deleted), KEY idx_lead_assigned (assigned_to),
                KEY idx_lead_created (created_at)) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `lead_discussions` (
                id INT AUTO_INCREMENT PRIMARY KEY, lead_id INT NOT NULL,
                user_id INT NOT NULL, comment TEXT NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                KEY idx_disc_lead (lead_id)) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `lead_attachments` (
                id INT AUTO_INCREMENT PRIMARY KEY, lead_id INT NOT NULL,
                discussion_id INT, file_name VARCHAR(300) NOT NULL,
                file_path VARCHAR(500) NOT NULL, file_size INT, file_type VARCHAR(100),
                uploaded_by INT, created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                KEY idx_att_lead (lead_id), KEY idx_att_disc (discussion_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `lead_reminders` (
                id INT AUTO_INCREMENT PRIMARY KEY, lead_id INT NOT NULL,
                user_id INT NOT NULL, title VARCHAR(300) NOT NULL, description TEXT,
                remind_at DATETIME NOT NULL, is_done TINYINT(1) DEFAULT 0,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                KEY idx_rem_lead (lead_id), KEY idx_rem_at (remind_at)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `lead_notes` (
                id INT AUTO_INCREMENT PRIMARY KEY, lead_id INT NOT NULL,
                user_id INT NOT NULL, note TEXT NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                KEY idx_note_lead (lead_id), KEY idx_note_user (user_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `lead_activity_logs` (
                id INT AUTO_INCREMENT PRIMARY KEY, lead_id INT NOT NULL, user_id INT,
                action VARCHAR(500) NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                KEY idx_log_lead (lead_id)) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `lead_contributions` (
                id INT AUTO_INCREMENT PRIMARY KEY, lead_id INT NOT NULL,
                user_id INT NOT NULL, action_type VARCHAR(30) NOT NULL,
                points INT DEFAULT 0, note VARCHAR(200),
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                KEY idx_contrib_lead (lead_id), KEY idx_contrib_user (user_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `contribution_config` (
                id INT AUTO_INCREMENT PRIMARY KEY, action_type VARCHAR(30) NOT NULL,
                label VARCHAR(100) NOT NULL, points INT DEFAULT 0,
                description VARCHAR(200), updated_by INT,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_contrib_action (action_type)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `client_masters` (
                id INT AUTO_INCREMENT PRIMARY KEY, code VARCHAR(20),
                company_name VARCHAR(200), contact_name VARCHAR(150) NOT NULL,
                position VARCHAR(100), email VARCHAR(150), website VARCHAR(200),
                mobile VARCHAR(20), alternate_mobile VARCHAR(20), gstin VARCHAR(20),
                status VARCHAR(20) DEFAULT 'active', address TEXT, city VARCHAR(100),
                state VARCHAR(100), country VARCHAR(100) DEFAULT 'India',
                zip_code VARCHAR(10), notes TEXT,
                is_deleted TINYINT(1) NOT NULL DEFAULT 0, deleted_at DATETIME,
                created_by INT, created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                modified_by INT,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_client_code (code), KEY idx_client_deleted (is_deleted)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `client_brands` (
                id INT AUTO_INCREMENT PRIMARY KEY, client_id INT NOT NULL,
                brand_name VARCHAR(200) NOT NULL, category VARCHAR(100),
                description TEXT, is_active TINYINT(1) DEFAULT 1,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                KEY idx_brand_client (client_id)) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
            """CREATE TABLE IF NOT EXISTS `client_addresses` (
                id INT AUTO_INCREMENT PRIMARY KEY, client_id INT NOT NULL,
                brand_index INT DEFAULT 0, title VARCHAR(100) NOT NULL DEFAULT 'Address',
                addr_type VARCHAR(20) DEFAULT 'billing', address TEXT, city VARCHAR(100),
                state VARCHAR(100), country VARCHAR(100) DEFAULT 'India',
                zip_code VARCHAR(10), is_default TINYINT(1) DEFAULT 0,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                KEY idx_caddr_client (client_id)) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
        ]
        for q in ddl:
            conn.execute(q)
        conn.commit()

        # seed (INSERT IGNORE)
        seed = [
            ("INSERT IGNORE INTO `lead_statuses` (name,label,color,sort_order,is_active) VALUES "
             "('open','Open','#2563eb',1,1),('in_process','In Process','#d97706',2,1),"
             "('close','Close (Won)','#16a34a',3,1),('cancel','Cancel','#dc2626',4,1)"),
            ("INSERT IGNORE INTO `lead_sources` (name,sort_order,is_active) VALUES "
             "('Website',1,1),('Referral',2,1),('IndiaMART',3,1),('Exhibition',4,1),"
             "('Cold Call',5,1),('Social Media',6,1),('Walk-in',7,1),('Other',8,1)"),
            ("INSERT IGNORE INTO `lead_categories` (name,sort_order,is_active) VALUES "
             "('Skin Care',1,1),('Hair Care',2,1),('Personal Care',3,1),"
             "('Cosmetics',4,1),('Ayurvedic',5,1),('Other',6,1)"),
            ("INSERT IGNORE INTO `product_ranges` (name,sort_order,is_active) VALUES "
             "('Premium',1,1),('Standard',2,1),('Economy',3,1)"),
            ("INSERT IGNORE INTO `contribution_config` (action_type,label,points,description) VALUES "
             "('comment','Comment / Discussion',1,'Comment add'),"
             "('status_change','Status Change',2,'Status change'),"
             "('close_fast','Fast Close (1-7d)',8,'Close within 7 days'),"
             "('close_slow','Slow Close (29d+)',0,'Close after 29 days'),"
             "('cancel','Cancel',0,'Lead cancel'),"
             "('follow_up','Follow Up',1,'Follow-up set'),"
             "('reminder','Reminder',1,'Reminder add'),"
             "('edit','Edit',1,'Lead edit')"),
        ]
        for q in seed:
            conn.execute(q)
        conn.commit()
        # ── BACKFILL (one-time, idempotent) ─────────────────────────────
        # Auto-assign rule se PEHLE bani leads ka assigned_to NULL hai —
        # unhe creator ko assign kar do, warna wo "Unassigned" me ginti hain.
        cur = conn.execute(
            "UPDATE `leads` SET assigned_to = created_by "
            "WHERE (assigned_to IS NULL OR assigned_to = 0) "
            "AND created_by IS NOT NULL AND created_by <> 0")
        if cur.rowcount:
            print(f"[crm_leads] backfill: {cur.rowcount} purani lead(s) "
                  f"creator ko auto-assign hui.")
        conn.commit()

        print("[crm_leads] tables ready.")
    except Exception as e:
        print(f"[crm_leads] ensure_lead_tables error: {e}")
    finally:
        conn.close()


# ═════════════════════════════════════════════════════════════════════════════
# CLIENT MASTER  (clients + brands + addresses)
# ═════════════════════════════════════════════════════════════════════════════
import json as _json


def gen_client_code(conn):
    row = conn.execute(
        "SELECT code FROM `client_masters` WHERE code LIKE 'CLT-%%' "
        "ORDER BY id DESC LIMIT 1").fetchone()
    n = 1
    if row and row.get('code'):
        try:
            n = int(row['code'].split('-')[1]) + 1
        except Exception:
            n = 1
    return f"CLT-{n:04d}"


@crm_bp.route('/clients')
@login_required
def clients():
    show_trash = request.args.get('trash', '') == '1'
    search = request.args.get('search', '')
    conn = _db()
    if not conn:
        flash('Database connection failed.', 'error')
        return redirect('/')
    try:
        where = ["is_deleted=%s"]
        params = [1 if show_trash else 0]
        if search:
            s = f"%{search}%"
            cols = ['code', 'company_name', 'contact_name', 'mobile', 'email',
                    'city', 'gstin']
            where.append("(" + " OR ".join(f"{c} LIKE %s" for c in cols) + ")")
            params.extend([s] * len(cols))
        rows = conn.execute(
            "SELECT * FROM `client_masters` WHERE " + " AND ".join(where)
            + " ORDER BY created_at DESC", params).fetchall() or []
        # brands per client (for Brands column)
        for c in rows:
            brs = conn.execute(
                "SELECT brand_name FROM `client_brands` WHERE client_id=%s "
                "AND is_active=1", (c['id'],)).fetchall() or []
            c['_brands'] = [b['brand_name'] for b in brs]
        total_count = (conn.execute(
            "SELECT COUNT(*) AS c FROM `client_masters` WHERE is_deleted=0"
        ).fetchone() or {}).get('c', 0)
        deleted_count = (conn.execute(
            "SELECT COUNT(*) AS c FROM `client_masters` WHERE is_deleted=1"
        ).fetchone() or {}).get('c', 0)
        return render_template(
            'crm/clients/clients.html', clients=rows, total_count=total_count,
            deleted_count=deleted_count, show_trash=show_trash, search=search,
            sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
            active_item='clients', user_name=_uname(),
            role=session.get('User_Type'))
    finally:
        conn.close()


def _save_client_brands(conn, client_id, brands):
    """brands = list of {brand_name, category, description, billing{...}, shipping{...}}"""
    conn.execute("DELETE FROM `client_brands` WHERE client_id=%s", (client_id,))
    conn.execute("DELETE FROM `client_addresses` WHERE client_id=%s", (client_id,))
    for idx, b in enumerate(brands):
        bn = (b.get('brand_name') or '').strip()
        if not bn:
            continue
        conn.execute(
            "INSERT INTO `client_brands` (client_id,brand_name,category,description,"
            "is_active,created_at) VALUES (%s,%s,%s,%s,1,NOW())",
            (client_id, bn, b.get('category'), b.get('description')))
        for kind in ('billing', 'shipping'):
            a = b.get(kind) or {}
            if any((a.get('address'), a.get('city'), a.get('state'),
                    a.get('zip_code'))):
                conn.execute(
                    "INSERT INTO `client_addresses` (client_id,brand_index,title,"
                    "addr_type,address,city,state,country,zip_code,is_default,"
                    "created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())",
                    (client_id, idx, kind.title(), kind, a.get('address'),
                     a.get('city'), a.get('state'), a.get('country', 'India'),
                     a.get('zip_code'), 1 if kind == 'billing' else 0))


@crm_bp.route('/clients/add', methods=['GET', 'POST'])
@login_required
def client_add():
    conn = _db()
    try:
        if request.method == 'POST':
            f = request.form
            code = gen_client_code(conn)
            cur = conn.execute(
                "INSERT INTO `client_masters` (code,company_name,contact_name,"
                "position,email,website,mobile,alternate_mobile,gstin,status,"
                "address,city,state,country,zip_code,notes,created_by,created_at) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())",
                (code, f.get('company_name'), f.get('contact_name'),
                 f.get('position'), f.get('email'), f.get('website'),
                 f.get('mobile'), f.get('alternate_mobile'), f.get('gstin'),
                 f.get('status', 'active'), f.get('address'), f.get('city'),
                 f.get('state'), f.get('country', 'India'), f.get('zip_code'),
                 f.get('notes'), _uid()))
            cid = cur.lastrowid
            try:
                brands = _json.loads(f.get('brands_json') or '[]')
            except Exception:
                brands = []
            _save_client_brands(conn, cid, brands)
            conn.commit()
            flash(f'Client {code} created.', 'success')
            return redirect(url_for('crm.clients'))
        return render_template(
            'crm/clients/client_form.html', client=None, mode='add', brands=[],
            sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
            active_item='clients', user_name=_uname(),
            role=session.get('User_Type'))
    finally:
        conn.close()


@crm_bp.route('/clients/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def client_edit(id):
    conn = _db()
    try:
        client = conn.execute("SELECT * FROM `client_masters` WHERE id=%s",
                              (id,)).fetchone()
        if not client:
            abort(404)
        if request.method == 'POST':
            f = request.form
            conn.execute(
                "UPDATE `client_masters` SET company_name=%s,contact_name=%s,"
                "position=%s,email=%s,website=%s,mobile=%s,alternate_mobile=%s,"
                "gstin=%s,status=%s,address=%s,city=%s,state=%s,country=%s,"
                "zip_code=%s,notes=%s,modified_by=%s WHERE id=%s",
                (f.get('company_name'), f.get('contact_name'), f.get('position'),
                 f.get('email'), f.get('website'), f.get('mobile'),
                 f.get('alternate_mobile'), f.get('gstin'),
                 f.get('status', 'active'), f.get('address'), f.get('city'),
                 f.get('state'), f.get('country', 'India'), f.get('zip_code'),
                 f.get('notes'), _uid(), id))
            try:
                brands = _json.loads(f.get('brands_json') or '[]')
            except Exception:
                brands = []
            _save_client_brands(conn, id, brands)
            conn.commit()
            flash('Client updated.', 'success')
            return redirect(url_for('crm.clients'))
        # build brands payload for edit
        brs = conn.execute(
            "SELECT * FROM `client_brands` WHERE client_id=%s ORDER BY id",
            (id,)).fetchall() or []
        addrs = conn.execute(
            "SELECT * FROM `client_addresses` WHERE client_id=%s", (id,)
        ).fetchall() or []
        brands = []
        for idx, b in enumerate(brs):
            bill = next((a for a in addrs if a['brand_index'] == idx
                         and a['addr_type'] == 'billing'), {})
            ship = next((a for a in addrs if a['brand_index'] == idx
                         and a['addr_type'] == 'shipping'), {})
            brands.append({
                'brand_name': b['brand_name'], 'category': b.get('category') or '',
                'description': b.get('description') or '',
                'billing': {k: (bill.get(k) or '') for k in
                            ('address', 'city', 'state', 'country', 'zip_code')},
                'shipping': {k: (ship.get(k) or '') for k in
                             ('address', 'city', 'state', 'country', 'zip_code')},
            })
        return render_template(
            'crm/clients/client_form.html', client=client, mode='edit',
            brands=brands,
            sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
            active_item='clients', user_name=_uname(),
            role=session.get('User_Type'))
    finally:
        conn.close()


@crm_bp.route('/clients/<int:id>/inline-edit', methods=['POST'])
@login_required
def client_inline_edit(id):
    """Client master list se cell-level quick edit (AJAX JSON: {field, value})."""
    data = request.get_json(silent=True) or {}
    field = (data.get('field') or '').strip()
    value = data.get('value', '')
    allowed = {'contact_name', 'company_name', 'mobile', 'email', 'city',
               'status', 'position', 'website', 'gstin', 'state'}
    if field not in allowed:
        return jsonify(success=False, error='Field not allowed'), 400
    if isinstance(value, str):
        value = value.strip()
    if field == 'status' and value not in ('active', 'inactive'):
        return jsonify(success=False, error='Invalid status'), 400
    if field == 'contact_name' and not value:
        return jsonify(success=False, error='Contact name cannot be empty'), 400
    conn = _db()
    try:
        row = conn.execute("SELECT id FROM `client_masters` "
                           "WHERE id=%s AND is_deleted=0", (id,)).fetchone()
        if not row:
            return jsonify(success=False, error='Client not found'), 404
        conn.execute(
            f"UPDATE `client_masters` SET `{field}`=%s, modified_by=%s "
            f"WHERE id=%s", (value or None, _uid(), id))
        conn.commit()
        return jsonify(success=True, field=field, value=value)
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify(success=False, error=str(e)), 500
    finally:
        conn.close()


@crm_bp.route('/clients/bulk-action', methods=['POST'])
@login_required
def clients_bulk_action():
    """Selected clients pe bulk delete / restore (AJAX JSON)."""
    data = request.get_json(silent=True) or {}
    action = data.get('action')
    try:
        ids = [int(i) for i in (data.get('ids') or [])][:200]
    except Exception:
        ids = []
    if action not in ('delete', 'restore') or not ids:
        return jsonify(success=False, error='Invalid request'), 400
    conn = _db()
    try:
        ph = ",".join(["%s"] * len(ids))
        if action == 'delete':
            cur = conn.execute(
                f"UPDATE `client_masters` SET is_deleted=1, deleted_at=NOW() "
                f"WHERE id IN ({ph}) AND is_deleted=0", ids)
        else:
            cur = conn.execute(
                f"UPDATE `client_masters` SET is_deleted=0, deleted_at=NULL "
                f"WHERE id IN ({ph}) AND is_deleted=1", ids)
        conn.commit()
        return jsonify(success=True, count=cur.rowcount)
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return jsonify(success=False, error=str(e)), 500
    finally:
        conn.close()


@crm_bp.route('/clients/<int:id>/delete', methods=['POST'])
@login_required
def client_delete(id):
    conn = _db()
    try:
        conn.execute("UPDATE `client_masters` SET is_deleted=1, deleted_at=NOW() "
                     "WHERE id=%s", (id,))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@crm_bp.route('/clients/<int:id>/restore', methods=['POST'])
@login_required
def client_restore(id):
    conn = _db()
    try:
        conn.execute("UPDATE `client_masters` SET is_deleted=0, deleted_at=NULL "
                     "WHERE id=%s", (id,))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


# ═════════════════════════════════════════════════════════════════════════════
# CLIENT IMPORT  (full page + template)
# ═════════════════════════════════════════════════════════════════════════════
@crm_bp.route('/clients/import', methods=['GET', 'POST'])
@login_required
def clients_import():
    if request.method == 'GET':
        return render_template(
            'crm/clients/import_clients.html',
            sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
            active_item='imp-clients', user_name=_uname(),
            role=session.get('User_Type'))

    f = request.files.get('file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('crm.clients_import'))

    name = f.filename.lower()
    rows = []
    try:
        raw = f.read()                          # poori file memory me (robust)
        if name.endswith('.csv'):
            import csv
            text = raw.decode('utf-8-sig', errors='ignore')
            for r in csv.DictReader(text.splitlines()):
                rows.append({(k or '').strip().lower(): (str(v).strip() if v is not None else '')
                             for k, v in r.items()})
        elif name.endswith('.xlsx'):
            import io as _io
            from openpyxl import load_workbook
            wb = load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            hdr = [str(h or '').strip().lower() for h in next(it, [])]
            for r in it:
                if not any(c is not None and str(c).strip() for c in r):
                    continue                    # blank row skip
                rows.append({hdr[i]: (str(c).strip() if c is not None else '')
                             for i, c in enumerate(r) if i < len(hdr)})
        elif name.endswith('.xls'):
            flash('Old .xls files are not supported. Please open the file in Excel '
                  '"Save As .xlsx" karke (ya CSV) upload karein.', 'error')
            return redirect(url_for('crm.clients_import'))
        else:
            flash('Only .csv or .xlsx files are allowed.', 'error')
            return redirect(url_for('crm.clients_import'))
    except Exception as e:
        flash(f'Could not read the file: {e}', 'error')
        return redirect(url_for('crm.clients_import'))

    if not rows:
        flash('No data rows found in the file (only a header row?).', 'error')
        return redirect(url_for('crm.clients_import'))

    # Header alias — spaces/underscores dono, kaafi variants cover
    alias = {
        'name': 'contact_name', 'contact': 'contact_name', 'contact name': 'contact_name',
        'contact_name': 'contact_name', 'client name': 'contact_name', 'person': 'contact_name',
        'company': 'company_name', 'company name': 'company_name', 'company_name': 'company_name',
        'firm': 'company_name', 'organisation': 'company_name', 'organization': 'company_name',
        'position': 'position', 'designation': 'position', 'title': 'position',
        'email': 'email', 'email id': 'email', 'e-mail': 'email', 'mail': 'email',
        'website': 'website', 'web': 'website', 'url': 'website',
        'mobile': 'mobile', 'mobile no': 'mobile', 'mobile number': 'mobile',
        'phone': 'mobile', 'phone no': 'mobile', 'contact no': 'mobile', 'number': 'mobile',
        'alt mobile': 'alternate_mobile', 'alternate mobile': 'alternate_mobile',
        'alternate_mobile': 'alternate_mobile', 'alt phone': 'alternate_mobile',
        'secondary mobile': 'alternate_mobile',
        'gstin': 'gstin', 'gst': 'gstin', 'gst no': 'gstin', 'gst number': 'gstin',
        'status': 'status',
        'address': 'address', 'addr': 'address',
        'city': 'city', 'state': 'state', 'country': 'country',
        'zip': 'zip_code', 'zip_code': 'zip_code', 'zipcode': 'zip_code',
        'pincode': 'zip_code', 'pin': 'zip_code', 'pin code': 'zip_code', 'postal code': 'zip_code',
        'notes': 'notes', 'note': 'notes', 'remarks': 'notes', 'remark': 'notes',
        # brand + shipping (import me 1 brand auto-create ke liye)
        'brand': 'brand', 'brand name': 'brand', 'brand_name': 'brand',
        'category': 'category', 'brand category': 'brand_category', 'brand_category': 'brand_category',
        'description': 'description', 'desc': 'description',
        'brand description': 'brand_description', 'brand_description': 'brand_description',
        'shipping address': 'ship_address', 'ship address': 'ship_address', 'shipping_address': 'ship_address',
        'shipping city': 'ship_city', 'ship city': 'ship_city',
        'shipping state': 'ship_state', 'ship state': 'ship_state',
        'shipping country': 'ship_country', 'ship country': 'ship_country',
        'shipping zip': 'ship_zip', 'ship zip': 'ship_zip', 'shipping zip_code': 'ship_zip', 'ship_zip': 'ship_zip',
    }
    CLIENT_COLS = ['company_name', 'contact_name', 'position', 'email', 'website',
                   'mobile', 'alternate_mobile', 'gstin', 'address', 'city', 'state',
                   'country', 'zip_code', 'notes']

    conn = _db()
    inserted, skipped, errors = 0, 0, []
    try:
        # Starting CLT-#### EK baar nikaal lo, fir locally increment (batch-safe)
        last = conn.execute(
            "SELECT code FROM `client_masters` WHERE code LIKE 'CLT-%%' "
            "ORDER BY id DESC LIMIT 1").fetchone()
        seq = 1
        if last and last.get('code'):
            try:
                seq = int(str(last['code']).split('-')[1]) + 1
            except Exception:
                seq = 1

        for idx, r in enumerate(rows, start=2):    # row 2 = pehli data row
            try:
                rec = {}
                for k, v in r.items():
                    if not v:
                        continue
                    key = (k or '').strip().lower()
                    tgt = alias.get(key) or alias.get(key.replace('_', ' '))
                    if tgt:
                        rec[tgt] = v
                cname = rec.get('contact_name')
                if not cname:
                    skipped += 1
                    continue
                st = (rec.get('status') or 'active').lower()
                if st not in ('active', 'inactive'):
                    st = 'active'
                code = f"CLT-{seq:04d}"
                seq += 1
                cols, ph, vals = ['code'], ['%s'], [code]
                for c in CLIENT_COLS:
                    if rec.get(c):
                        cols.append(c); ph.append('%s'); vals.append(rec[c])
                cols += ['status', 'created_by', 'created_at']
                ph   += ['%s', '%s', 'NOW()']
                vals += [st, _uid()]
                sql = ("INSERT INTO `client_masters` (" + ",".join("`" + c + "`" for c in cols)
                       + ") VALUES (" + ",".join(ph) + ")")
                cur = conn.execute(sql, vals)
                cid = cur.lastrowid

                # ── kam se kam 1 brand + Billing & Shipping address ──
                billing = {
                    'address': rec.get('address'), 'city': rec.get('city'),
                    'state': rec.get('state'), 'country': rec.get('country') or 'India',
                    'zip_code': rec.get('zip_code'),
                }
                shipping = {
                    'address':  rec.get('ship_address') or billing['address'],
                    'city':     rec.get('ship_city')    or billing['city'],
                    'state':    rec.get('ship_state')   or billing['state'],
                    'country':  rec.get('ship_country') or billing['country'],
                    'zip_code': rec.get('ship_zip')     or billing['zip_code'],
                }
                _save_client_brands(conn, cid, [{
                    'brand_name':  rec.get('brand') or rec.get('company_name') or cname,
                    'category':    rec.get('brand_category') or rec.get('category'),
                    'description': rec.get('brand_description') or rec.get('description'),
                    'billing':  billing,
                    'shipping': shipping,
                }])
                inserted += 1
            except Exception as row_err:
                errors.append(f'Row {idx}: {row_err}')
                continue
        conn.commit()
        msg = f'{inserted} clients import hue.'
        if skipped:
            msg += f' {skipped} skipped (name missing).'
        if errors:
            msg += f' {len(errors)} row error.'
        flash(msg, 'success' if inserted else 'error')
        if errors:
            flash('Pehla error → ' + errors[0], 'error')
    except Exception as e:
        flash(f'Import error: {e}', 'error')
    finally:
        conn.close()
    return redirect(url_for('crm.clients'))


@crm_bp.route('/clients/import/template')
@login_required
def clients_import_template():
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    import io
    wb = Workbook(); ws = wb.active; ws.title = 'Clients'
    headers = ['name', 'company', 'position', 'email', 'website', 'mobile',
               'alternate_mobile', 'gstin', 'status', 'address', 'city', 'state',
               'country', 'zip_code', 'notes',
               'brand', 'category', 'description',
               'ship_address', 'ship_city', 'ship_state', 'ship_country', 'ship_zip']
    ws.append(headers)
    ws.append(['Ramesh Shah', 'ABC Cosmetics Pvt Ltd', 'Director', 'ramesh@abc.com',
               'www.abc.com', '9999999999', '8888888888', '24ABCDE1234F1Z5', 'active',
               'Plot 12, GIDC', 'Surat', 'Gujarat', 'India', '395003', 'Net 30 terms',
               'XYZ Cosmetics', 'Skincare', 'Premium brand',
               'Plot 12, GIDC', 'Surat', 'Gujarat', 'India', '395003'])
    dv = DataValidation(type='list', formula1='"active,inactive"', allow_blank=True)
    ws.add_data_validation(dv); dv.add('I2:I1000')
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    from flask import send_file
    return send_file(bio, as_attachment=True, download_name='clients_import_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═════════════════════════════════════════════════════════════════════════════
# SAMPLE ORDERS  (ported from legacy CRM — same schema for data merge)
# ═════════════════════════════════════════════════════════════════════════════
def _ensure_sample_orders(conn):
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS `sample_orders` (
            id INT AUTO_INCREMENT PRIMARY KEY,
            order_number VARCHAR(50) UNIQUE NOT NULL,
            lead_id INT NOT NULL,
            order_date DATE,
            category VARCHAR(50) DEFAULT 'Sample Order',
            bill_company VARCHAR(200), bill_address TEXT, bill_phone VARCHAR(20),
            bill_email VARCHAR(150), bill_gst VARCHAR(20),
            gst_pct DECIMAL(5,2) DEFAULT 18, sub_total DECIMAL(12,2) DEFAULT 0,
            gst_amount DECIMAL(12,2) DEFAULT 0, total_amount DECIMAL(12,2) DEFAULT 0,
            items_json TEXT, terms TEXT, invoice_file VARCHAR(300),
            created_by INT, created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            is_deleted TINYINT DEFAULT 0, deleted_at DATETIME NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""")
        conn.commit()
    except Exception:
        pass


def _next_so_number(conn):
    row = conn.execute("SELECT order_number FROM `sample_orders` "
                       "WHERE order_number LIKE 'HCPSMPL%%' ORDER BY id DESC LIMIT 1").fetchone()
    num = 1
    if row and row.get('order_number'):
        try:
            num = int(str(row['order_number']).replace('HCPSMPL', '')) + 1
        except ValueError:
            num = 1
    cand = 'HCPSMPL%03d' % num
    while conn.execute("SELECT id FROM `sample_orders` WHERE order_number=%s",
                       (cand,)).fetchone():
        num += 1
        cand = 'HCPSMPL%03d' % num
    return cand


@crm_bp.route('/sample-orders/next-number')
@login_required
def sample_order_next_number():
    conn = _db()
    try:
        _ensure_sample_orders(conn)
        return jsonify(order_number=_next_so_number(conn))
    finally:
        conn.close()


def _build_so_pdf(d):
    """d = dict(order_number, order_date, category, bill_*, gst_pct, items[list],
       sub_total, gst_amount, total_amount, terms, by_name). Returns BytesIO."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable, Image)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=12*mm, bottomMargin=15*mm)
    W = A4[0] - 30*mm
    styles = getSampleStyleSheet()

    def S(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    normal = S('N', fontSize=9, leading=13)
    right = S('R', fontSize=9, alignment=TA_RIGHT)
    center = S('C', fontSize=9, alignment=TA_CENTER)
    small = S('Sm', fontSize=8, textColor=colors.HexColor('#6b7280'), leading=11)
    story = []

    company_info = (
        '<b>HCP Wellness Pvt. Ltd.</b><br/>'
        '403, Maruti Vertex Elanza,<br/>'
        'Opp. Global Hospital, Sindhu Bhavan Road, Bodakdev,<br/>'
        'Ahmedabad-380054, Gujarat, India.<br/>'
        '<b>GST :</b> 24AAFCH7246H1ZK'
    )
    logo_path = os.path.join(current_app.root_path, 'static', 'images', 'hcp-logo.png')
    if os.path.exists(logo_path):
        hdr_left = Image(logo_path, width=40*mm, height=18*mm, kind='proportional')
    else:
        hdr_left = Paragraph('<b><font size="16" color="#1e3a5f">HCP Wellness</font></b>', normal)
    hdr = Table([[hdr_left, Paragraph(company_info,
                 S('CI', fontSize=8.5, alignment=TA_RIGHT, leading=13))]],
                colWidths=[W*0.35, W*0.65])
    hdr.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                             ('TOPPADDING', (0, 0), (-1, -1), 6),
                             ('BOTTOMPADDING', (0, 0), (-1, -1), 6)]))
    story.append(hdr)
    story.append(HRFlowable(width='100%', thickness=1.5,
                            color=colors.HexColor('#1e3a5f'), spaceAfter=5))

    bl = []
    if d.get('bill_company'):
        bl.append('<b>%s</b>' % d['bill_company'])
    if d.get('bill_address'):
        bl.append(str(d['bill_address']).replace('\n', '<br/>'))
    if d.get('bill_phone'):
        bl.append(d['bill_phone'])
    if d.get('bill_email'):
        bl.append(d['bill_email'])
    if d.get('bill_gst'):
        bl.append('GST: %s' % d['bill_gst'])
    bill_txt = '<font size="7.5" color="#6b7280"><b>BILLING ADDRESS</b></font><br/>' + '<br/>'.join(bl)
    info_txt = ('<font size="7.5" color="#6b7280">Date</font><br/><b>%s</b><br/><br/>'
                '<font size="7.5" color="#6b7280">Order ID</font><br/><b>%s</b><br/><br/>'
                '<font size="7.5" color="#6b7280">Category</font><br/><b>%s</b>'
                % (d.get('order_date', ''), d.get('order_number', ''), d.get('category', 'Sample Order')))
    at = Table([[Paragraph(bill_txt, S('BT', fontSize=9, leading=14)),
                 Paragraph(info_txt, S('OI', fontSize=9, leading=13, alignment=TA_RIGHT))]],
               colWidths=[W*0.55, W*0.45])
    at.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('TOPPADDING', (0, 0), (-1, -1), 8),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                            ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))]))
    story.append(at)
    story.append(Spacer(1, 5*mm))

    th = [Paragraph('<b>Product Name</b>', normal),
          Paragraph('<b>Rate</b>', right), Paragraph('<b>Qty</b>', center),
          Paragraph('<b>Amount</b>', right)]
    rows = [th]
    for it in d.get('items', []):
        rows.append([Paragraph(str(it.get('name', '')), normal),
                     Paragraph('%.2f' % float(it.get('rate', 0) or 0), right),
                     Paragraph(str(it.get('qty', 0)), center),
                     Paragraph('%.2f' % float(it.get('amount', 0) or 0), right)])
    itbl = Table(rows, colWidths=[W*0.45, W*0.18, W*0.15, W*0.22], repeatRows=1)
    itbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 7), ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'), ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT')]))
    story.append(itbl)
    story.append(Spacer(1, 4*mm))

    gp = float(d.get('gst_pct', 0) or 0)
    totals = [[Paragraph('Sub Total', small), Paragraph('%.2f' % float(d.get('sub_total', 0)), right)],
              [Paragraph('GST (%g%%)' % gp, small), Paragraph('%.2f' % float(d.get('gst_amount', 0)), right)],
              [Paragraph('<b>Total Amount</b>', S('TB', fontSize=10, textColor=colors.HexColor('#1e3a5f'))),
               Paragraph('<b>%.2f</b>' % float(d.get('total_amount', 0)),
                         S('TRB', fontSize=10, alignment=TA_RIGHT, textColor=colors.HexColor('#1e3a5f')))]]
    ttbl = Table(totals, colWidths=[W*0.6, W*0.4])
    ttbl.setStyle(TableStyle([('TOPPADDING', (0, 0), (-1, -1), 5),
                              ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                              ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                              ('LINEBELOW', (0, 0), (-1, 1), 0.5, colors.HexColor('#e5e7eb')),
                              ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#e8f0fe'))]))
    story.append(ttbl)
    story.append(Spacer(1, 5*mm))

    terms = (d.get('terms') or '').replace('\n', '<br/>')
    ts = Table([[Paragraph('<b>Terms &amp; Conditions:</b><br/>' + terms, small),
                 Paragraph('<br/><br/><br/>________________________<br/><b>Authorised Signature</b>',
                           S('Sig', fontSize=9, alignment=TA_CENTER))]],
               colWidths=[W*0.6, W*0.4])
    ts.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    story.append(ts)
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#d1d5db')))
    story.append(Paragraph('<i>Generated %s &middot; %s</i>'
                           % (datetime.now().strftime('%d-%m-%Y %H:%M'), d.get('by_name', '')),
                           S('F', fontSize=7.5, textColor=colors.HexColor('#9ca3af'),
                             alignment=TA_CENTER, spaceBefore=3)))
    doc.build(story)
    buf.seek(0)
    return buf


@crm_bp.route('/leads/<int:id>/sample-order', methods=['POST'])
@login_required
def lead_sample_order(id):
    import json as _json
    conn = _db()
    try:
        _ensure_sample_orders(conn)
        lead = conn.execute("SELECT * FROM `leads` WHERE id=%s", (id,)).fetchone()
        if not lead:
            abort(404)

        so_number = (request.form.get('so_number') or '').strip() or _next_so_number(conn)
        so_date = request.form.get('so_date') or datetime.now().strftime('%Y-%m-%d')
        category = request.form.get('so_category', 'Sample Order')
        gst_pct = float(request.form.get('so_gst_pct', '18') or 0)
        bill_company = (request.form.get('bill_company') or lead.get('company_name') or '').strip()
        bill_address = (request.form.get('bill_address') or '').strip()
        bill_phone = (request.form.get('bill_phone') or lead.get('phone') or '').strip()
        bill_email = (request.form.get('bill_email') or lead.get('email') or '').strip()
        bill_gst = (request.form.get('bill_gst') or '').strip()
        terms = request.form.get('terms', '')
        names = request.form.getlist('item_name[]')
        qtys = request.form.getlist('item_qty[]')
        rates = request.form.getlist('item_rate[]')

        if not bill_company or not bill_address or not bill_phone:
            flash('Company, Address and Mobile are required.', 'error')
            return redirect(url_for('crm.lead_view', id=id))

        items, sub_total = [], 0.0
        for i, nm in enumerate(names):
            if not nm.strip():
                continue
            try:
                qty = float(qtys[i]) if i < len(qtys) and qtys[i] else 0
            except ValueError:
                qty = 0
            try:
                rate = float(rates[i]) if i < len(rates) and rates[i] else 0
            except ValueError:
                rate = 0
            amt = qty * rate
            sub_total += amt
            items.append({'name': nm.strip(), 'qty': qty, 'rate': rate, 'amount': amt})
        if not items:
            flash('Please add at least one product.', 'error')
            return redirect(url_for('crm.lead_view', id=id))

        gst_amount = sub_total * gst_pct / 100.0
        total_amount = sub_total + gst_amount
        try:
            od = datetime.strptime(so_date, '%Y-%m-%d').strftime('%Y-%m-%d')
        except Exception:
            od = datetime.now().strftime('%Y-%m-%d')

        try:
            conn.execute(
                "INSERT INTO `sample_orders` (order_number, lead_id, order_date, category, "
                "bill_company, bill_address, bill_phone, bill_email, bill_gst, gst_pct, "
                "sub_total, gst_amount, total_amount, items_json, terms, created_by, created_at) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())",
                (so_number, id, od, category, bill_company, bill_address, bill_phone,
                 bill_email, bill_gst, gst_pct, sub_total, gst_amount, total_amount,
                 _json.dumps(items), terms, _uid()))
            log_activity(conn, id, 'Sample Order generated: %s' % so_number)
            conn.commit()
        except Exception:
            conn.rollback()
            flash('Order number "%s" already exists, please refresh and retry.' % so_number, 'error')
            return redirect(url_for('crm.lead_view', id=id))

        pdf = _build_so_pdf({
            'order_number': so_number,
            'order_date': datetime.strptime(od, '%Y-%m-%d').strftime('%d-%m-%Y'),
            'category': category, 'bill_company': bill_company, 'bill_address': bill_address,
            'bill_phone': bill_phone, 'bill_email': bill_email, 'bill_gst': bill_gst,
            'gst_pct': gst_pct, 'items': items, 'sub_total': sub_total,
            'gst_amount': gst_amount, 'total_amount': total_amount, 'terms': terms,
            'by_name': _uname()})
        return send_file(pdf, mimetype='application/pdf', as_attachment=False,
                         download_name='%s.pdf' % so_number)
    finally:
        conn.close()


@crm_bp.route('/sample-orders')
@login_required
def sample_orders_list():
    conn = _db()
    try:
        _ensure_sample_orders(conn)
        trash = request.args.get('trash') == '1'
        search = (request.args.get('search') or '').strip()
        where = "s.is_deleted=%s" % (1 if trash else 0)
        params = []
        if search:
            where += " AND (s.order_number LIKE %s OR s.bill_company LIKE %s OR l.contact_name LIKE %s)"
            like = '%' + search + '%'
            params = [like, like, like]
        rows = conn.execute(
            "SELECT s.*, l.contact_name AS lead_contact, l.code AS lead_code "
            "FROM `sample_orders` s LEFT JOIN `leads` l ON l.id=s.lead_id "
            "WHERE " + where + " ORDER BY s.id DESC", tuple(params)).fetchall() or []
        cnt = conn.execute("SELECT is_deleted, COUNT(*) c FROM `sample_orders` "
                           "GROUP BY is_deleted").fetchall() or []
        active_n = sum(c['c'] for c in cnt if not c['is_deleted'])
        del_n = sum(c['c'] for c in cnt if c['is_deleted'])
        umap = _user_map(conn)
        for r in rows:
            r['_by'] = (umap or {}).get(r.get('created_by'), 'Administrator')
        return render_template('crm/sample_orders/list.html', orders=rows,
                               trash=trash, search=search, active_n=active_n, del_n=del_n,
                               sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
                               active_item='samples', user_name=_uname(),
                               role=session.get('User_Type'), is_admin=_is_admin(),
                               is_admin_mgr=_is_admin_mgr())
    finally:
        conn.close()


@crm_bp.route('/sample-orders/<int:soid>/pdf')
@login_required
def sample_order_pdf(soid):
    import json as _json
    conn = _db()
    try:
        s = conn.execute("SELECT * FROM `sample_orders` WHERE id=%s", (soid,)).fetchone()
        if not s:
            abort(404)
        items = []
        try:
            items = _json.loads(s.get('items_json') or '[]')
        except Exception:
            items = []
        od = s.get('order_date')
        try:
            od = datetime.strptime(str(od), '%Y-%m-%d').strftime('%d-%m-%Y')
        except Exception:
            od = str(od)
        pdf = _build_so_pdf({
            'order_number': s['order_number'], 'order_date': od,
            'category': s.get('category'), 'bill_company': s.get('bill_company'),
            'bill_address': s.get('bill_address'), 'bill_phone': s.get('bill_phone'),
            'bill_email': s.get('bill_email'), 'bill_gst': s.get('bill_gst'),
            'gst_pct': s.get('gst_pct'), 'items': items, 'sub_total': s.get('sub_total'),
            'gst_amount': s.get('gst_amount'), 'total_amount': s.get('total_amount'),
            'terms': s.get('terms'), 'by_name': _uname()})
        return send_file(pdf, mimetype='application/pdf', as_attachment=False,
                         download_name='%s.pdf' % s['order_number'])
    finally:
        conn.close()


# ═════════════════════════════════════════════════════════════════════════════
# QUOTATIONS  (ported from legacy CRM — same schema for data merge)
# ═════════════════════════════════════════════════════════════════════════════
def _ensure_quotations(conn):
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS `quotations` (
            id INT AUTO_INCREMENT PRIMARY KEY,
            quot_number VARCHAR(50) UNIQUE NOT NULL,
            lead_id INT NOT NULL,
            quot_date DATE, valid_until DATE, subject VARCHAR(300),
            bill_company VARCHAR(200), bill_address TEXT, bill_phone VARCHAR(20),
            bill_email VARCHAR(150), bill_gst VARCHAR(20),
            gst_pct DECIMAL(5,2) DEFAULT 18, sub_total DECIMAL(12,2) DEFAULT 0,
            gst_amount DECIMAL(12,2) DEFAULT 0, total_amount DECIMAL(12,2) DEFAULT 0,
            items_json TEXT, terms TEXT, notes TEXT,
            status VARCHAR(20) DEFAULT 'draft',
            email_sent_at DATETIME NULL, email_sent_to VARCHAR(150),
            created_by INT, created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            is_deleted TINYINT DEFAULT 0, deleted_at DATETIME NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""")
        conn.commit()
    except Exception:
        pass
    # NPD se bhi quotation ban sake (project link; lead optional)
    for alter in ("ALTER TABLE `quotations` ADD COLUMN project_id INT NULL",
                  "ALTER TABLE `quotations` MODIFY lead_id INT NULL"):
        try:
            conn.execute(alter)
            conn.commit()
        except Exception:
            pass


def _fy_qt():
    n = datetime.now()
    y = n.year
    if n.month >= 4:
        return '%02d-%02d' % (y % 100, (y + 1) % 100)
    return '%02d-%02d' % ((y - 1) % 100, y % 100)


def _next_quot_number(conn):
    fy = _fy_qt()
    row = conn.execute("SELECT quot_number FROM `quotations` "
                       "WHERE quot_number LIKE %s ORDER BY id DESC LIMIT 1",
                       ('QT-%%/' + fy,)).fetchone()
    num = 1
    if row and row.get('quot_number'):
        try:
            num = int(str(row['quot_number']).split('-')[1].split('/')[0]) + 1
        except Exception:
            num = 1
    cand = 'QT-%03d/%s' % (num, fy)
    while conn.execute("SELECT id FROM `quotations` WHERE quot_number=%s",
                       (cand,)).fetchone():
        num += 1
        cand = 'QT-%03d/%s' % (num, fy)
    return cand


@crm_bp.route('/quotations/next-number')
@login_required
def quotation_next_number():
    conn = _db()
    try:
        _ensure_quotations(conn)
        return jsonify(quot_number=_next_quot_number(conn))
    finally:
        conn.close()


def _build_quot_pdf(d):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable, Image)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=12*mm, bottomMargin=14*mm)
    W = A4[0] - 28*mm
    styles = getSampleStyleSheet()

    def S(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    normal = S('N', fontSize=8.5, leading=12)
    right = S('R', fontSize=8.5, alignment=TA_RIGHT)
    center = S('C', fontSize=8.5, alignment=TA_CENTER)
    small = S('Sm', fontSize=8, textColor=colors.HexColor('#6b7280'), leading=11)
    story = []

    company_info = (
        '<b>HCP Wellness Pvt. Ltd.</b><br/>403, Maruti Vertex Elanza,<br/>'
        'Opp. Global Hospital, Sindhu Bhavan Road, Bodakdev,<br/>'
        'Ahmedabad-380054, Gujarat, India.<br/><b>GST :</b> 24AAFCH7246H1ZK')
    logo_path = os.path.join(current_app.root_path, 'static', 'images', 'hcp-logo.png')
    if os.path.exists(logo_path):
        hl = Image(logo_path, width=40*mm, height=18*mm, kind='proportional')
    else:
        hl = Paragraph('<b><font size="16" color="#1e3a5f">HCP Wellness</font></b>', normal)
    hdr = Table([[hl, Paragraph(company_info, S('CI', fontSize=8, alignment=TA_RIGHT, leading=12))]],
                colWidths=[W*0.35, W*0.65])
    hdr.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
    story.append(hdr)
    story.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#1e3a5f'), spaceAfter=4))
    story.append(Paragraph('<b><font size="13" color="#1e3a5f">QUOTATION</font></b>', S('Q', spaceAfter=4)))

    bl = []
    if d.get('bill_company'):
        bl.append('<b>%s</b>' % d['bill_company'])
    if d.get('bill_address'):
        bl.append(str(d['bill_address']).replace('\n', '<br/>'))
    if d.get('bill_phone'):
        bl.append(d['bill_phone'])
    if d.get('bill_email'):
        bl.append(d['bill_email'])
    if d.get('bill_gst'):
        bl.append('GST: %s' % d['bill_gst'])
    bill_txt = '<font size="7.5" color="#6b7280"><b>BILL TO</b></font><br/>' + '<br/>'.join(bl)
    info = ('<font size="7.5" color="#6b7280">Quotation No</font><br/><b>%s</b><br/><br/>'
            '<font size="7.5" color="#6b7280">Date</font><br/><b>%s</b><br/><br/>'
            '<font size="7.5" color="#6b7280">Valid Until</font><br/><b>%s</b>'
            % (d.get('quot_number', ''), d.get('quot_date', ''), d.get('valid_until', '-') or '-'))
    at = Table([[Paragraph(bill_txt, S('BT', fontSize=8.5, leading=13)),
                 Paragraph(info, S('OI', fontSize=8.5, leading=12, alignment=TA_RIGHT))]],
               colWidths=[W*0.55, W*0.45])
    at.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('TOPPADDING', (0, 0), (-1, -1), 7), ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
                            ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))]))
    story.append(at)
    if d.get('subject'):
        story.append(Paragraph('<b>Subject:</b> %s' % d['subject'], S('Sub', fontSize=9, spaceBefore=6, spaceAfter=4)))
    story.append(Spacer(1, 3*mm))

    th = [Paragraph('<b>Product</b>', normal), Paragraph('<b>Size</b>', center),
          Paragraph('<b>UOM</b>', center), Paragraph('<b>MOQ</b>', center),
          Paragraph('<b>Final Cost</b>', right), Paragraph('<b>Amount</b>', right)]
    rows = [th]
    for it in d.get('items', []):
        rows.append([Paragraph(str(it.get('name', '')), normal),
                     Paragraph(str(it.get('size', '') or '-'), center),
                     Paragraph(str(it.get('uom', '') or '-'), center),
                     Paragraph(str(it.get('moq', 0)), center),
                     Paragraph('%.2f' % float(it.get('final_cost', 0) or 0), right),
                     Paragraph('%.2f' % float(it.get('amount', 0) or 0), right)])
    itbl = Table(rows, colWidths=[W*0.30, W*0.12, W*0.12, W*0.12, W*0.17, W*0.17], repeatRows=1)
    itbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0'))]))
    story.append(itbl)
    story.append(Spacer(1, 4*mm))

    gp = float(d.get('gst_pct', 0) or 0)
    totals = [[Paragraph('Sub Total', small), Paragraph('%.2f' % float(d.get('sub_total', 0)), right)],
              [Paragraph('GST (%g%%)' % gp, small), Paragraph('%.2f' % float(d.get('gst_amount', 0)), right)],
              [Paragraph('<b>Total Amount</b>', S('TB', fontSize=10, textColor=colors.HexColor('#1e3a5f'))),
               Paragraph('<b>%.2f</b>' % float(d.get('total_amount', 0)),
                         S('TRB', fontSize=10, alignment=TA_RIGHT, textColor=colors.HexColor('#1e3a5f')))]]
    ttbl = Table(totals, colWidths=[W*0.62, W*0.38])
    ttbl.setStyle(TableStyle([('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                              ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                              ('LINEBELOW', (0, 0), (-1, 1), 0.5, colors.HexColor('#e5e7eb')),
                              ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#e8f0fe'))]))
    story.append(ttbl)
    story.append(Spacer(1, 5*mm))

    extras = ''
    if d.get('notes'):
        extras += '<b>Notes:</b><br/>%s<br/><br/>' % str(d['notes']).replace('\n', '<br/>')
    extras += '<b>Terms &amp; Conditions:</b><br/>%s' % (d.get('terms') or '').replace('\n', '<br/>')
    ts = Table([[Paragraph(extras, small),
                 Paragraph('<br/><br/><br/>________________________<br/><b>Authorised Signature</b>',
                           S('Sig', fontSize=9, alignment=TA_CENTER))]],
               colWidths=[W*0.6, W*0.4])
    ts.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    story.append(ts)
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#d1d5db')))
    story.append(Paragraph('<i>Generated %s &middot; %s</i>'
                           % (datetime.now().strftime('%d-%m-%Y %H:%M'), d.get('by_name', '')),
                           S('F', fontSize=7.5, textColor=colors.HexColor('#9ca3af'),
                             alignment=TA_CENTER, spaceBefore=3)))
    doc.build(story)
    buf.seek(0)
    return buf


@crm_bp.route('/leads/<int:id>/create-npd')
@login_required
def lead_create_npd(id):
    """
    Create NPD gate (lead se):
      - Lead ka client connected hai (leads.client_id)  -> seedha NPD form
      - Connected nahi hai                               -> client page
    """
    conn = _db()
    try:
        lead = conn.execute(
            "SELECT id, client_id FROM `leads` WHERE id=%s", (id,)).fetchone()
    finally:
        conn.close()

    if not lead:
        flash('Lead not found.', 'error')
        return redirect(url_for('crm.leads'))

    ptype = request.args.get('type') \
        if request.args.get('type') in ('npd', 'existing') else 'npd'

    if lead.get('client_id'):
        # CLIENT CONNECTED -> seedha NPD form
        # NPD route ka naam alag ho to neeche `except` wala path apne hisaab se badlo.
        try:
            target = url_for('npd.create', lead_id=id,
                             client_id=lead['client_id'], type=ptype)
        except Exception:
            target = "/npd/new?lead_id=%s&client_id=%s&type=%s" % (
                id, lead['client_id'], ptype)
        return redirect(target)

    # CLIENT NAHI CONNECTED -> client page (pehle connect/add karo)
    flash('No Client is connected to this lead. '
          'Please connect/add a client first, then create the NPD.', 'info')
    return redirect(url_for('crm.clients', connect_lead=id, next='create-npd'))


@crm_bp.route('/leads/<int:id>/quotation', methods=['POST'])
@login_required
def lead_create_quotation(id):
    import json as _json
    conn = _db()
    try:
        _ensure_quotations(conn)
        lead = conn.execute("SELECT * FROM `leads` WHERE id=%s", (id,)).fetchone()
        if not lead:
            abort(404)

        quot_number = (request.form.get('quot_number') or '').strip() or _next_quot_number(conn)
        quot_date = request.form.get('quot_date') or datetime.now().strftime('%Y-%m-%d')
        valid_until = request.form.get('valid_until') or ''
        subject = request.form.get('quot_subject', '')
        bill_company = (request.form.get('bill_company') or lead.get('company_name') or '').strip()
        bill_address = (request.form.get('bill_address') or '').strip()
        bill_phone = (request.form.get('bill_phone') or lead.get('phone') or '').strip()
        bill_email = (request.form.get('bill_email') or lead.get('email') or '').strip()
        bill_gst = (request.form.get('bill_gst') or '').strip()
        terms = request.form.get('terms', '')
        notes = request.form.get('notes', '')
        gst_pct = float(request.form.get('quot_gst_pct', '18') or 0)

        names = request.form.getlist('item_name[]')
        sizes = request.form.getlist('item_size[]')
        uoms = request.form.getlist('item_uom[]')
        codes = request.form.getlist('item_code[]')
        moqs = request.form.getlist('item_moq[]')
        finals = request.form.getlist('item_final_cost[]')
        costs = request.form.getlist('item_cost[]')
        pm_specs = request.form.getlist('item_pm_spec[]')
        pm_costs = request.form.getlist('item_pm_cost[]')
        cats = request.form.getlist('item_category[]')

        def _f(lst, i):
            try:
                return float(lst[i]) if i < len(lst) and lst[i] else 0.0
            except ValueError:
                return 0.0

        def _s(lst, i):
            return lst[i] if i < len(lst) else ''

        if not bill_company:
            flash('Company Name is required.', 'error')
            return redirect(url_for('crm.lead_view', id=id))

        items, sub_total = [], 0.0
        for i, nm in enumerate(names):
            if not nm.strip():
                continue
            moq = _f(moqs, i)
            fc = _f(finals, i)
            amt = moq * fc
            sub_total += amt
            items.append({'name': nm.strip(), 'size': _s(sizes, i), 'uom': _s(uoms, i),
                          'code': _s(codes, i), 'moq': moq, 'final_cost': fc, 'amount': amt,
                          'cost': _f(costs, i), 'pm_spec': _s(pm_specs, i),
                          'pm_cost': _f(pm_costs, i), 'category': _s(cats, i)})
        if not items:
            flash('Please add at least one product.', 'error')
            return redirect(url_for('crm.lead_view', id=id))

        gst_amount = sub_total * gst_pct / 100.0
        total_amount = sub_total + gst_amount
        try:
            qd = datetime.strptime(quot_date, '%Y-%m-%d').strftime('%Y-%m-%d')
        except Exception:
            qd = datetime.now().strftime('%Y-%m-%d')
        vu = None
        if valid_until:
            try:
                vu = datetime.strptime(valid_until, '%Y-%m-%d').strftime('%Y-%m-%d')
            except Exception:
                vu = None

        try:
            conn.execute(
                "INSERT INTO `quotations` (quot_number, lead_id, quot_date, valid_until, subject, "
                "bill_company, bill_address, bill_phone, bill_email, bill_gst, gst_pct, sub_total, "
                "gst_amount, total_amount, items_json, terms, notes, status, created_by, created_at) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'draft',%s,NOW())",
                (quot_number, id, qd, vu, subject, bill_company, bill_address, bill_phone,
                 bill_email, bill_gst, gst_pct, sub_total, gst_amount, total_amount,
                 _json.dumps(items), terms, notes, _uid()))
            log_activity(conn, id, 'Quotation generated: %s' % quot_number)
            conn.commit()
        except Exception:
            conn.rollback()
            flash('Quotation number "%s" already exists, please retry.' % quot_number, 'error')
            return redirect(url_for('crm.lead_view', id=id))

        pdf = _build_quot_pdf({
            'quot_number': quot_number,
            'quot_date': datetime.strptime(qd, '%Y-%m-%d').strftime('%d-%m-%Y'),
            'valid_until': datetime.strptime(vu, '%Y-%m-%d').strftime('%d-%m-%Y') if vu else '-',
            'subject': subject, 'bill_company': bill_company, 'bill_address': bill_address,
            'bill_phone': bill_phone, 'bill_email': bill_email, 'bill_gst': bill_gst,
            'gst_pct': gst_pct, 'items': items, 'sub_total': sub_total, 'gst_amount': gst_amount,
            'total_amount': total_amount, 'terms': terms, 'notes': notes, 'by_name': _uname()})
        return send_file(pdf, mimetype='application/pdf', as_attachment=False,
                         download_name='%s.pdf' % quot_number.replace('/', '_'))
    finally:
        conn.close()


@crm_bp.route('/quotations')
@login_required
def quotations_list():
    conn = _db()
    try:
        _ensure_quotations(conn)
        trash = request.args.get('trash') == '1'
        search = (request.args.get('search') or '').strip()
        status_f = (request.args.get('status') or '').strip().lower()
        where = "q.is_deleted=%s" % (1 if trash else 0)
        params = []
        if search:
            where += (" AND (q.quot_number LIKE %s OR q.bill_company LIKE %s"
                      " OR q.subject LIKE %s OR l.contact_name LIKE %s"
                      " OR p.code LIKE %s)")
            like = '%' + search + '%'
            params = [like, like, like, like, like]
        if status_f in ('draft', 'sent', 'accepted', 'rejected'):
            where += " AND q.status=%s"
            params.append(status_f)
        rows = conn.execute(
            "SELECT q.*, l.contact_name AS lead_contact, l.code AS lead_code,"
            " p.id AS proj_id, p.code AS proj_code, "
            "p.product_name AS proj_product "
            "FROM `quotations` q LEFT JOIN `leads` l ON l.id=q.lead_id "
            "LEFT JOIN `npd_projects` p ON p.id=q.project_id "
            "WHERE " + where + " ORDER BY q.id DESC", tuple(params)).fetchall() or []
        cnt = conn.execute("SELECT is_deleted, COUNT(*) c FROM `quotations` "
                           "GROUP BY is_deleted").fetchall() or []
        active_n = sum(c['c'] for c in cnt if not c['is_deleted'])
        del_n = sum(c['c'] for c in cnt if c['is_deleted'])
        umap = _user_map(conn)
        for r in rows:
            r['_by'] = (umap or {}).get(r.get('created_by'), 'Administrator')
        return render_template('crm/quotations/list.html', quotations=rows,
                               trash=trash, search=search, status_f=status_f,
                               active_n=active_n, del_n=del_n,
                               sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
                               active_item='quotations', user_name=_uname(),
                               role=session.get('User_Type'), is_admin=_is_admin(),
                               is_admin_mgr=_is_admin_mgr())
    finally:
        conn.close()


@crm_bp.route('/quotations/<int:qid>/pdf')
@login_required
def quotation_pdf(qid):
    import json as _json
    conn = _db()
    try:
        q = conn.execute("SELECT * FROM `quotations` WHERE id=%s", (qid,)).fetchone()
        if not q:
            abort(404)
        try:
            items = _json.loads(q.get('items_json') or '[]')
        except Exception:
            items = []

        def _fmt(v):
            try:
                return datetime.strptime(str(v), '%Y-%m-%d').strftime('%d-%m-%Y')
            except Exception:
                return str(v) if v else '-'
        pdf = _build_quot_pdf({
            'quot_number': q['quot_number'], 'quot_date': _fmt(q.get('quot_date')),
            'valid_until': _fmt(q.get('valid_until')), 'subject': q.get('subject'),
            'bill_company': q.get('bill_company'), 'bill_address': q.get('bill_address'),
            'bill_phone': q.get('bill_phone'), 'bill_email': q.get('bill_email'),
            'bill_gst': q.get('bill_gst'), 'gst_pct': q.get('gst_pct'), 'items': items,
            'sub_total': q.get('sub_total'), 'gst_amount': q.get('gst_amount'),
            'total_amount': q.get('total_amount'), 'terms': q.get('terms'),
            'notes': q.get('notes'), 'by_name': _uname()})
        return send_file(pdf, mimetype='application/pdf', as_attachment=False,
                         download_name='%s.pdf' % str(q['quot_number']).replace('/', '_'))
    finally:
        conn.close()


# ═════════════════════════════════════════════════════════════════════════════
# SAMPLE ORDER / QUOTATION — list actions (delete, email, invoice, status)
# ═════════════════════════════════════════════════════════════════════════════
def _mail_get(key, default=''):
    try:
        v = current_app.config.get(key)
        if v not in (None, ''):
            return v
    except Exception:
        pass
    v = os.environ.get(key)
    if v not in (None, ''):
        return v
    for m in ('config', 'app_config', 'settings', 'configuration'):
        try:
            mod = __import__(m)
            v = getattr(mod, key, None)
            if v not in (None, ''):
                return v
        except Exception:
            continue
    return default


def _smtp_send(to_addr, subject, body_html, attach_bytes=None, attach_name=None):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    from email.utils import formataddr
    server = _mail_get('MAIL_SERVER', 'smtp.gmail.com')
    port = int(_mail_get('MAIL_PORT', 587) or 587)
    use_tls = _mail_get('MAIL_USE_TLS', True)
    username = _mail_get('MAIL_USERNAME', 'no-reply@hcpwellness.in')
    password = _mail_get('MAIL_PASSWORD', '')
    from_name = _mail_get('MAIL_FROM_NAME', 'HCP Wellness Pvt. Ltd.')
    if not to_addr:
        return False, 'Koi email address nahi hai'
    if not password:
        return False, 'SMTP password config.py me set karein'
    try:
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From'] = formataddr((from_name, username))
        msg['To'] = to_addr
        alt = MIMEMultipart('alternative')
        import re as _re
        plain = _re.sub(r'<[^>]+>', '', body_html)
        alt.attach(MIMEText(plain, 'plain', 'utf-8'))
        alt.attach(MIMEText(body_html, 'html', 'utf-8'))
        msg.attach(alt)
        if attach_bytes:
            ap = MIMEApplication(attach_bytes, _subtype='pdf')
            ap.add_header('Content-Disposition', 'attachment',
                          filename=attach_name or 'document.pdf')
            msg.attach(ap)
        s = smtplib.SMTP(server, port, timeout=25)
        if use_tls:
            s.starttls()
        s.login(username, password)
        s.sendmail(username, [to_addr], msg.as_string())
        s.quit()
        return True, None
    except Exception as e:
        return False, str(e)


@crm_bp.route('/sample-orders/<int:soid>/delete', methods=['POST'])
@login_required
def sample_order_delete(soid):
    conn = _db()
    try:
        conn.execute("UPDATE `sample_orders` SET is_deleted=1, deleted_at=NOW() "
                     "WHERE id=%s", (soid,))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@crm_bp.route('/sample-orders/<int:soid>/restore', methods=['POST'])
@login_required
def sample_order_restore(soid):
    conn = _db()
    try:
        conn.execute("UPDATE `sample_orders` SET is_deleted=0, deleted_at=NULL "
                     "WHERE id=%s", (soid,))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@crm_bp.route('/sample-orders/<int:soid>/purge', methods=['POST'])
@login_required
def sample_order_purge(soid):
    if not _is_admin_mgr():
        return jsonify(ok=False, error='You do not have permission'), 403
    conn = _db()
    try:
        s = conn.execute("SELECT invoice_file FROM `sample_orders` WHERE id=%s",
                         (soid,)).fetchone()
        if s and s.get('invoice_file'):
            try:
                p = os.path.join(current_app.root_path, 'static', 'uploads', s['invoice_file'])
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass
        conn.execute("DELETE FROM `sample_orders` WHERE id=%s", (soid,))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@crm_bp.route('/sample-orders/<int:soid>/email', methods=['POST'])
@login_required
def sample_order_email(soid):
    import json as _json
    conn = _db()
    try:
        s = conn.execute("SELECT * FROM `sample_orders` WHERE id=%s", (soid,)).fetchone()
        if not s:
            return jsonify(ok=False, error='Not found'), 404
        to_addr = (request.form.get('to') or s.get('bill_email') or '').strip()
        subj = (request.form.get('subject') or '').strip() or ('Sample Order %s — HCP Wellness' % s['order_number'])
        body = request.form.get('body')
        try:
            items = _json.loads(s.get('items_json') or '[]')
        except Exception:
            items = []
        od = s.get('order_date')
        try:
            od = datetime.strptime(str(od), '%Y-%m-%d').strftime('%d-%m-%Y')
        except Exception:
            od = str(od)
        pdf = _build_so_pdf({
            'order_number': s['order_number'], 'order_date': od, 'category': s.get('category'),
            'bill_company': s.get('bill_company'), 'bill_address': s.get('bill_address'),
            'bill_phone': s.get('bill_phone'), 'bill_email': s.get('bill_email'),
            'bill_gst': s.get('bill_gst'), 'gst_pct': s.get('gst_pct'), 'items': items,
            'sub_total': s.get('sub_total'), 'gst_amount': s.get('gst_amount'),
            'total_amount': s.get('total_amount'), 'terms': s.get('terms'), 'by_name': _uname()})
        if not body:
            body = ('Dear %s,<br><br>Please find attached the Sample Order <b>%s</b>.<br><br>'
                    'Regards,<br>HCP Wellness Pvt. Ltd.'
                    % (s.get('bill_company') or 'Sir/Madam', s['order_number']))
        ok, err = _smtp_send(to_addr, subj, body, pdf.getvalue(), '%s.pdf' % s['order_number'])
        if ok:
            log_activity(conn, s['lead_id'], 'Sample Order %s emailed to %s' % (s['order_number'], to_addr))
            conn.commit()
            return jsonify(ok=True, to=to_addr)
        return jsonify(ok=False, error=err), 500
    finally:
        conn.close()


@crm_bp.route('/sample-orders/<int:soid>/invoice-upload', methods=['POST'])
@login_required
def sample_order_invoice_upload(soid):
    conn = _db()
    try:
        f = request.files.get('invoice')
        if not f or not f.filename:
            return jsonify(ok=False, error='File chunno'), 400
        d = os.path.join(current_app.root_path, 'static', 'uploads', 'invoices')
        os.makedirs(d, exist_ok=True)
        fname = secure_filename('%s_%s' % (soid, f.filename))
        f.save(os.path.join(d, fname))
        conn.execute("UPDATE `sample_orders` SET invoice_file=%s WHERE id=%s",
                     ('invoices/%s' % fname, soid))
        conn.commit()
        return jsonify(ok=True, file=fname)
    finally:
        conn.close()


@crm_bp.route('/sample-orders/<int:soid>/invoice-download')
@login_required
def sample_order_invoice_download(soid):
    conn = _db()
    try:
        s = conn.execute("SELECT invoice_file FROM `sample_orders` WHERE id=%s",
                         (soid,)).fetchone()
        if not s or not s.get('invoice_file'):
            abort(404)
        path = os.path.join(current_app.root_path, 'static', 'uploads', s['invoice_file'])
        if not os.path.exists(path):
            abort(404)
        return send_file(path, as_attachment=True)
    finally:
        conn.close()


@crm_bp.route('/quotations/<int:qid>/delete', methods=['POST'])
@login_required
def quotation_delete(qid):
    conn = _db()
    try:
        conn.execute("UPDATE `quotations` SET is_deleted=1, deleted_at=NOW() WHERE id=%s", (qid,))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@crm_bp.route('/quotations/<int:qid>/restore', methods=['POST'])
@login_required
def quotation_restore(qid):
    conn = _db()
    try:
        conn.execute("UPDATE `quotations` SET is_deleted=0, deleted_at=NULL WHERE id=%s", (qid,))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@crm_bp.route('/quotations/<int:qid>/purge', methods=['POST'])
@login_required
def quotation_purge(qid):
    if not _is_admin_mgr():
        return jsonify(ok=False, error='You do not have permission'), 403
    conn = _db()
    try:
        conn.execute("DELETE FROM `quotations` WHERE id=%s", (qid,))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()


@crm_bp.route('/quotations/<int:qid>/status', methods=['POST'])
@login_required
def quotation_status(qid):
    st = (request.form.get('status') or '').strip().lower()
    if st not in ('draft', 'sent', 'accepted', 'rejected'):
        return jsonify(ok=False, error='Invalid status'), 400
    conn = _db()
    try:
        conn.execute("UPDATE `quotations` SET status=%s WHERE id=%s", (st, qid))
        conn.commit()
        return jsonify(ok=True, status=st)
    finally:
        conn.close()


@crm_bp.route('/quotations/<int:qid>/email', methods=['POST'])
@login_required
def quotation_email(qid):
    import json as _json
    conn = _db()
    try:
        q = conn.execute("SELECT * FROM `quotations` WHERE id=%s", (qid,)).fetchone()
        if not q:
            return jsonify(ok=False, error='Not found'), 404
        to_addr = (request.form.get('to') or q.get('bill_email') or '').strip()
        subj = (request.form.get('subject') or '').strip() or ('Quotation %s — HCP Wellness' % q['quot_number'])
        cbody = request.form.get('body')
        try:
            items = _json.loads(q.get('items_json') or '[]')
        except Exception:
            items = []

        def _fmt(v):
            try:
                return datetime.strptime(str(v), '%Y-%m-%d').strftime('%d-%m-%Y')
            except Exception:
                return str(v) if v else '-'
        pdf = _build_quot_pdf({
            'quot_number': q['quot_number'], 'quot_date': _fmt(q.get('quot_date')),
            'valid_until': _fmt(q.get('valid_until')), 'subject': q.get('subject'),
            'bill_company': q.get('bill_company'), 'bill_address': q.get('bill_address'),
            'bill_phone': q.get('bill_phone'), 'bill_email': q.get('bill_email'),
            'bill_gst': q.get('bill_gst'), 'gst_pct': q.get('gst_pct'), 'items': items,
            'sub_total': q.get('sub_total'), 'gst_amount': q.get('gst_amount'),
            'total_amount': q.get('total_amount'), 'terms': q.get('terms'),
            'notes': q.get('notes'), 'by_name': _uname()})
        if not cbody:
            cbody = ('Dear %s,<br><br>Please find attached our Quotation <b>%s</b>.<br><br>'
                     'Regards,<br>HCP Wellness Pvt. Ltd.'
                     % (q.get('bill_company') or 'Sir/Madam', q['quot_number']))
        ok, err = _smtp_send(to_addr, subj, cbody, pdf.getvalue(),
                             '%s.pdf' % str(q['quot_number']).replace('/', '_'))
        if ok:
            conn.execute("UPDATE `quotations` SET status='sent', email_sent_at=NOW(), "
                         "email_sent_to=%s WHERE id=%s", (to_addr, qid))
            log_activity(conn, q['lead_id'], 'Quotation %s emailed to %s' % (q['quot_number'], to_addr))
            conn.commit()
            return jsonify(ok=True, to=to_addr)
        return jsonify(ok=False, error=err), 500
    finally:
        conn.close()


@crm_bp.route('/quotations/products')
@login_required
def quotation_products_list():
    import json as _json
    conn = _db()
    try:
        _ensure_quotations(conn)
        search = (request.args.get('search') or '').strip()
        qrows = conn.execute(
            "SELECT q.*, l.contact_name AS lead_contact, "
            "p.id AS proj_id, p.code AS proj_code FROM `quotations` q "
            "LEFT JOIN `leads` l ON l.id=q.lead_id "
            "LEFT JOIN `npd_projects` p ON p.id=q.project_id "
            "WHERE q.is_deleted=0 "
            "ORDER BY q.created_at DESC").fetchall() or []
        rows = []
        sr = 0
        for q in qrows:
            try:
                items = _json.loads(q.get('items_json') or '[]')
            except Exception:
                items = []
            qd = q.get('quot_date')
            try:
                qd = datetime.strptime(str(qd), '%Y-%m-%d').strftime('%d-%m-%Y')
            except Exception:
                qd = str(qd)
            company = q.get('bill_company') or q.get('lead_contact') or '—'
            for it in items:
                if search:
                    blob = ('%s %s %s' % (it.get('name', ''), it.get('category', ''), company)).lower()
                    if search.lower() not in blob:
                        continue
                sr += 1
                rows.append({'sr': sr, 'quot_number': q['quot_number'], 'quot_id': q['id'],
                             'quot_date': qd, 'company': company, 'lead_id': q['lead_id'],
                             'lead_contact': q.get('lead_contact'),
                             'proj_id': q.get('proj_id'), 'proj_code': q.get('proj_code'),
                             'name': it.get('name', '—'), 'size': it.get('size', '') or '—',
                             'uom': it.get('uom', ''), 'cost': it.get('cost', 0),
                             'moq': it.get('moq', '—'), 'pm_spec': it.get('pm_spec', '') or '—',
                             'pm_cost': it.get('pm_cost', 0), 'category': it.get('category', '') or '—',
                             'final_cost': it.get('final_cost', 0), 'status': q.get('status')})
        return render_template('crm/quotations/products.html', rows=rows, search=search,
                               sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
                               active_item='quot-prod', user_name=_uname(),
                               role=session.get('User_Type'), is_admin=_is_admin(),
                               is_admin_mgr=_is_admin_mgr())
    finally:
        conn.close()


# ═════════════════════════════════════════════════════════════════════════════
# LEAD MASTERS  (Status / Source / Category / Product Range)
# Operates on the SAME tables the Lead form uses -> fully unified.
# ═════════════════════════════════════════════════════════════════════════════
_LM_LABELS = {'status': 'Lead Status', 'source': 'Lead Source',
              'category': 'Lead Category', 'product_range': 'Product Range'}
_LM_MAP = {'status': 'lead_statuses', 'source': 'lead_sources',
           'category': 'lead_categories', 'product_range': 'product_ranges',
           'range': 'product_ranges'}


@crm_bp.route('/lead-masters')
@login_required
def lead_masters():
    conn = _db()
    try:
        data = {}
        for t in ('status', 'source', 'category', 'product_range'):
            tbl = _LM_MAP[t]
            if tbl == 'lead_statuses':
                rows = conn.execute(
                    "SELECT id, COALESCE(NULLIF(label,''), name) AS name FROM `lead_statuses` "
                    "WHERE is_active=1 ORDER BY sort_order, id").fetchall() or []
            else:
                rows = conn.execute(
                    "SELECT id, name FROM `%s` WHERE is_active=1 ORDER BY sort_order, id" % tbl
                ).fetchall() or []
            data[t] = rows
        return render_template('crm/lead_masters.html', data=data, labels=_LM_LABELS,
                               types=['status', 'source', 'category', 'product_range'],
                               sidebar_menu=get_menu('crm', role=_role(), is_admin=_is_admin()),
                               active_item='lead-mstr', user_name=_uname(),
                               role=session.get('User_Type'), is_admin=_is_admin(),
                               is_admin_mgr=_is_admin_mgr())
    finally:
        conn.close()


@crm_bp.route('/lead-masters/add', methods=['POST'])
@login_required
def lead_master_add():
    t = (request.form.get('type') or '').strip()
    name = (request.form.get('name') or '').strip()
    tbl = _LM_MAP.get(t)
    if not tbl or not name:
        return jsonify(ok=False, error='Invalid Type/Name'), 400
    conn = _db()
    try:
        if tbl == 'lead_statuses':
            key = name.lower().replace(' ', '_')
            dup = conn.execute("SELECT id FROM `lead_statuses` WHERE name=%s OR label=%s",
                               (key, name)).fetchone()
            if dup:
                return jsonify(ok=False, error='Already exists'), 409
            conn.execute("INSERT INTO `lead_statuses` (name,label,color,sort_order,is_active) "
                         "VALUES (%s,%s,'#64748b',99,1)", (key, name))
        else:
            dup = conn.execute("SELECT id FROM `%s` WHERE name=%%s" % tbl, (name,)).fetchone()
            if dup:
                conn.execute("UPDATE `%s` SET is_active=1 WHERE id=%%s" % tbl, (dup['id'],))
                conn.commit()
                return jsonify(ok=True, id=dup['id'], name=name)
            conn.execute("INSERT INTO `%s` (name,sort_order,is_active) VALUES (%%s,99,1)" % tbl,
                         (name,))
        conn.commit()
        nid = conn.execute("SELECT LAST_INSERT_ID() id").fetchone()['id']
        return jsonify(ok=True, id=nid, name=name)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500
    finally:
        conn.close()


@crm_bp.route('/lead-masters/edit', methods=['POST'])
@login_required
def lead_master_edit():
    t = (request.form.get('type') or '').strip()
    mid = request.form.get('id')
    name = (request.form.get('name') or '').strip()
    tbl = _LM_MAP.get(t)
    if not tbl or not mid or not name:
        return jsonify(ok=False, error='Galat data'), 400
    conn = _db()
    try:
        if tbl == 'lead_statuses':
            conn.execute("UPDATE `lead_statuses` SET label=%s WHERE id=%s", (name, mid))
        else:
            conn.execute("UPDATE `%s` SET name=%%s WHERE id=%%s" % tbl, (name, mid))
        conn.commit()
        return jsonify(ok=True, name=name)
    finally:
        conn.close()


@crm_bp.route('/lead-masters/delete', methods=['POST'])
@login_required
def lead_master_delete():
    t = (request.form.get('type') or '').strip()
    mid = request.form.get('id')
    tbl = _LM_MAP.get(t)
    if not tbl or not mid:
        return jsonify(ok=False, error='Galat data'), 400
    conn = _db()
    try:
        conn.execute("DELETE FROM `%s` WHERE id=%%s" % tbl, (mid,))
        conn.commit()
        return jsonify(ok=True)
    finally:
        conn.close()
