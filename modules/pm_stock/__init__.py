"""
pm_stock/__init__.py
=====================
Packaging Material (PM) Stock Management — Flask blueprint
HCP Wellness Pvt Ltd

This file holds the Blueprint object and all route handlers.
Pure helper code lives in pm_stock/helpers.py.

Blueprint prefix : /pm_stock
"""

from flask import Blueprint, render_template, request, jsonify, session, send_file, redirect
from functools import wraps
from datetime import datetime, date
import io
import json
import os
import re
import secrets
import shutil
import traceback
import uuid
import sampling_portal

# Common, data-driven sidebar (menu defined once in core/menus.py)
try:
    from menus import get_menu
except Exception:                       # fallback so the module still loads
    def get_menu(*a, **k): return None

# Page-specific bits injected INSIDE the sidebar <nav> (via sidebar_extra_html):
#  • Hidden anchor stubs that background JS polls for badge counts / shortcuts
#  • The drag handle used to resize the sidebar
_PM_SIDEBAR_EXTRA = """
<div style="display:none" id="pmHiddenHooks">
  <a id="btn-new-grn"   onclick="pmvOpen('grn')"></a>
  <a id="btn-new-dn"    onclick="openDnModal()"></a>
  <a id="btn-new-allot" onclick="openAllotmentPicker()"></a>
  <a id="my-reprint-btn" onclick="openMyReprintRequestsModal()"><span id="my-reprint-badge" class="nav-badge">0</span></a>
  <a id="my-label-reissue-btn" onclick="openMyLabelReissuesModal()"><span id="my-label-reissue-badge" class="nav-badge">0</span></a>
  <a id="reprint-approvals-btn" onclick="openReprintApprovalsModal()"><span id="reprint-approvals-badge" class="nav-badge">0</span></a>
  <a id="label-reissue-approvals-btn" onclick="openLabelReissueApprovalsModal()"><span id="label-reissue-approvals-badge" class="nav-badge">0</span></a>
  <a id="fifo-override-approvals-btn" onclick="openFifoOverrideApprovalsModal()"><span id="fifo-override-approvals-badge" class="nav-badge">0</span></a>
</div>
<div class="sb-resize-handle" id="sbResizeHandle"
     title="Drag to resize sidebar (double-click to reset)"></div>
"""

# Re-export everything from helpers so route bodies can call them as if they
# were defined here (matches the pre-split pm_stock_routes.py layout).
from .helpers import *               # noqa: F401,F403
from .helpers import (                # explicit list for IDE / linter clarity
    _login_required, _user, _is_admin, _is_out_creator, _is_manager,
    ensure_pm_tables, ensure_fifo_lots_table, ensure_audit_tables,
    ensure_pm_settings_table, ensure_fifo_override_requests_table,
    ensure_label_reissue_requests_table,
    ensure_pm_material_lock_table, _material_lock_check,
    ensure_pm_stock_adjustment_tables,
    ensure_pm_trs_tables,
    _fifo_override_find_approved, _fifo_override_mark_used,
    _assemble_box_label,
    _setting_get, _setting_set, _fifo_is_enabled, _fifo_start_date,
    AUDIT_REVERSAL_SAFE, AUDIT_REVERSAL_GATED, AUDIT_REVERSAL_FINAL,
    AUDIT_ACTIONS,
    _audit_record, _label_print_record,
    _fifo_seq_to_code, _next_fifo_seq, _get_or_assign_fifo, _fifo_check_oldest,
    _next_voucher_no, _clean_for_code, _generate_product_code,
    _normalize_product_name, _find_duplicate_product,
    _brand_name_by_id, _grn_seq_part, _make_box_code,
    # Short-code helpers — assign + look up the 8-char A0000001 codes that
    # newly-printed labels encode into their QRs. Without these explicitly
    # in the named-import list, Python's import resolution picks up the
    # `from .helpers import *` wildcard fine at module level, but anywhere
    # this file does `from .helpers import (...)` (e.g. tooling that
    # statically analyses) skips them — and at runtime certain code paths
    # see "name not defined" errors.
    _short_seq_to_code, _code_to_short_seq, _next_short_code_seq,
    _gen_short_code, _assign_box_short_code, _reissue_box_short_code, _find_box_by_any_code,
    _make_group_code, _next_group_seq,
    _refresh_group_status, _create_group_for_boxes,
    _next_op_seq, _make_op_box_code,
    _create_opening_boxes, _create_boxes_for_grn_item, _delete_boxes_for_grn,
    _assert_grn_box_invariant,
    _heal_box_location,
    _get_godowns, _godown_summary, _floor_summary,
    VOUCHER_TYPES, _voucher_type_enabled, _block_if_disabled,
    _block_if_requester,
    _json_safe, _row_to_dict,
    _bin_soft_delete, _bin_restore, _reinsert_row, _bin_purge,
    VALID_REPRINT_SCOPES, _gen_print_token,
    _user_home_godown, _enforce_home_godown,
    _user_is_requester,
    _user_access_dict, _user_has_access, _block_if_no_access,
    PM_USER_ACCESS_KEYS,
    _verify_reprint_auth,
    _log_transfer_edit, _is_floor_godown, _post_stock_movement,
    _refresh_transfer_totals, _check_discrepancy,
    _RESET_CATEGORIES, _do_clear_categories,
    _table_exists,
    _parse_date_range,
)

pm_stock_bp = Blueprint('pm_stock', __name__)


# ── Bootstrap calls — run once at package import ─────────────────────────────
try:
    ensure_audit_tables()
except Exception as _e:
    import sys as _sys
    print(f"[pm_stock] WARNING: ensure_audit_tables() failed: {_e}", file=_sys.stderr)

try:
    ensure_fifo_lots_table()
except Exception as _e:
    import sys as _sys
    print(f"[pm_stock] WARNING: ensure_fifo_lots_table() failed: {_e}", file=_sys.stderr)

try:
    ensure_fifo_override_requests_table()
except Exception as _e:
    import sys as _sys
    print(f"[pm_stock] WARNING: ensure_fifo_override_requests_table() failed: {_e}", file=_sys.stderr)

try:
    ensure_label_reissue_requests_table()
except Exception as _e:
    import sys as _sys
    print(f"[pm_stock] WARNING: ensure_label_reissue_requests_table() failed: {_e}", file=_sys.stderr)

try:
    ensure_pm_material_lock_table()
except Exception as _e:
    import sys as _sys
    print(f"[pm_stock] WARNING: ensure_pm_material_lock_table() failed: {_e}", file=_sys.stderr)

try:
    ensure_pm_settings_table()
except Exception as _e:
    import sys as _sys
    print(f"[pm_stock] WARNING: ensure_pm_settings_table() failed: {_e}", file=_sys.stderr)

try:
    ensure_pm_tables()
except Exception as _e:
    # Don't crash module import on bootstrap errors — log and continue, so the
    # rest of the app stays up. The first request that hits a missing table
    # will surface a clearer error in the logs.
    import sys as _sys
    print(f"[pm_stock] WARNING: ensure_pm_tables() failed at import: {_e}", file=_sys.stderr)

try:
    ensure_pm_stock_adjustment_tables()
except Exception as _e:
    import sys as _sys
    print(f"[pm_stock] WARNING: ensure_pm_stock_adjustment_tables() failed: {_e}", file=_sys.stderr)

try:
    ensure_pm_trs_tables()
except Exception as _e:
    import sys as _sys
    print(f"[pm_stock] WARNING: ensure_pm_trs_tables() failed: {_e}", file=_sys.stderr)


# ── Material Request feature (separate module) ───────────────────────────
# The Material Request workflow lives in material_request.py. We import its
# helper functions for the OUT integration hook, and call register_routes()
# to mount its endpoints onto the existing pm_stock_bp so they share the
# /api/pm_stock/... URL namespace.
from . import material_request as _mr_mod
_mr_mod.register_routes(pm_stock_bp)
from . import pm_stock_report_routes as _rpt_mod
_rpt_mod.register_report_routes(pm_stock_bp)
# BOM (Bill of Materials) — sister module, same pattern: lazy-loads
# helpers and registers its endpoints onto pm_stock_bp.
from . import pm_stock_bom as _bom_mod
_bom_mod.register_routes(pm_stock_bp)
# Dispatch Entry — FG-driven stock decrement vouchers. Same modular
# pattern: ensure_dispatch_tables() at boot, register_routes() to mount
# the endpoints. Tables auto-create on first import; safe to no-op if
# they already exist.
from . import pm_stock_dispatch as _dsp_mod
try:
    _dsp_mod.ensure_dispatch_tables()
except Exception as _e:
    import sys as _sys
    print(f"[pm_stock] WARNING: ensure_dispatch_tables() failed: {_e}", file=_sys.stderr)
_dsp_mod.register_routes(pm_stock_bp)
# Find Box — read-only drill-down explorer (Godown → Item → Packages).
# No tables to create; pure read-side feature over pm_boxes.
# Import is guarded so a syntax/import error in this module can't take
# down the entire pm_stock blueprint at boot.
try:
    from . import pm_stock_findbox as _fbx_mod
    _fbx_mod.register_routes(pm_stock_bp)
except Exception as _e:
    import sys as _sys
    import traceback as _tb
    print(f"[pm_stock] WARNING: pm_stock_findbox failed to load: {_e}", file=_sys.stderr)
    _tb.print_exc(file=_sys.stderr)
_mr_link_transfer_to_request   = _mr_mod._link_transfer_to_request
_mr_unlink_transfer_from_request = _mr_mod._unlink_transfer_from_request


# ════════════════════════════════════════════════════════════════════════════
# PM GRN FILE ATTACHMENTS  —  Invoice files (GRN-level, multiple)
# ════════════════════════════════════════════════════════════════════════════
#
# Files are stored on the local filesystem under:
#   <flask-root>/uploads/pm_stock/grn/<grn_id>/invoice/<uuid>__<original-name>
# Only metadata + relative_path live in pm_grn_files (table auto-created).
#
# Endpoints (all under /api/pm_stock/grn/file/...):
#   GET    /diag                  — config + on-disk health check
#   POST   /upload                — upload one file
#   GET    /list?grn_id=...       — list files for a GRN
#   GET    /<file_id>             — download / inline-view (streams from disk)
#   DELETE /<file_id>             — remove (DB row + disk file)
#
# Authorization: any logged-in user can list/view; only edit-capable users
# can upload/delete (matches the GRN module's own permission model).

# ── Bootstrap: create pm_grn_files table on first run ──
try:
    _pmf_conn = sampling_portal.get_db_connection()
    if _pmf_conn:
        try:
            _pmf_conn.execute("""
                CREATE TABLE IF NOT EXISTS pm_grn_files (
                  id              INT AUTO_INCREMENT PRIMARY KEY,
                  grn_id          INT NOT NULL,
                  kind            ENUM('invoice') NOT NULL DEFAULT 'invoice',
                  original_name   VARCHAR(255) NOT NULL,
                  mime_type       VARCHAR(80)  NOT NULL,
                  size_bytes      INT          NOT NULL,
                  relative_path   VARCHAR(500) NOT NULL,
                  uploaded_by     VARCHAR(80)  NULL,
                  uploaded_at     DATETIME DEFAULT CURRENT_TIMESTAMP,
                  INDEX idx_grn  (grn_id),
                  INDEX idx_kind (kind)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            _pmf_conn.commit()
            print("✅ [pm_stock] pm_grn_files table ready (disk-storage mode)")
        except Exception as _e:
            print(f"⚠️  [pm_stock] pm_grn_files setup error: {_e}")
        try: _pmf_conn.close()
        except Exception: pass
except Exception:
    pass

# ── Resolve filesystem upload root ──
# Default: <flask-root>/uploads/pm_stock
# flask-root = parent of the pm_stock package directory.
_PMF_MODULE_DIR = os.path.dirname(os.path.abspath(__file__))
_PMF_APP_ROOT   = os.path.dirname(_PMF_MODULE_DIR)
# Fallback for legacy flat layouts where the module IS the app root.
if not os.path.isdir(os.path.join(_PMF_APP_ROOT, "uploads")) and \
   os.path.isdir(os.path.join(_PMF_MODULE_DIR, "uploads")):
    _PMF_APP_ROOT = _PMF_MODULE_DIR
_PM_UPLOAD_ROOT = os.path.join(_PMF_APP_ROOT, "uploads", "pm_stock")
try:
    os.makedirs(_PM_UPLOAD_ROOT, exist_ok=True)
    print(f"✅ [pm_stock] Upload root: {_PM_UPLOAD_ROOT}")
except Exception as _e:
    print(f"⚠️  [pm_stock] Could not create upload root {_PM_UPLOAD_ROOT}: {_e}")

# ── Constants & helpers ──
_PMF_ALLOWED_MIME = {"application/pdf", "image/jpeg", "image/jpg", "image/png"}
_PMF_MAX_BYTES    = 10 * 1024 * 1024   # 10 MB

def _pmf_mime_ok(mime):
    return (mime or "").lower() in _PMF_ALLOWED_MIME

def _pmf_safe_segment(s):
    """Sanitize a filename component — keep alphanumerics, dot, dash, underscore."""
    s = str(s or "").strip()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^A-Za-z0-9._\-]", "", s)
    if not s or set(s) <= {"."}:
        s = "file"
    return s[:120]

def _pmf_disk_path(relative_path):
    """Resolve a stored relative_path to an absolute filesystem path,
    with traversal guard. Returns None if the resolved path escapes the root."""
    rp = (relative_path or "").lstrip("/\\").replace("\\", "/")
    abs_path = os.path.normpath(os.path.join(_PM_UPLOAD_ROOT, *rp.split("/")))
    root = os.path.normpath(_PM_UPLOAD_ROOT)
    if not abs_path.startswith(root + os.sep) and abs_path != root:
        return None
    return abs_path


# ── /diag — health check ──────────────────────────────────────────────
@pm_stock_bp.route('/api/pm_stock/grn/file/diag', methods=['GET'])
@_login_required
def api_pm_grn_file_diag():
    """Disk + DB health check for the PM GRN file-upload subsystem."""
    out = {"status": "ok",
           "upload_root": _PM_UPLOAD_ROOT,
           "upload_root_exists":   os.path.isdir(_PM_UPLOAD_ROOT),
           "upload_root_writable": os.access(_PM_UPLOAD_ROOT, os.W_OK) if os.path.isdir(_PM_UPLOAD_ROOT) else False}
    try:
        total = 0; count = 0
        for root, _dirs, files in os.walk(_PM_UPLOAD_ROOT):
            for fn in files:
                try:
                    total += os.path.getsize(os.path.join(root, fn))
                    count += 1
                except Exception:
                    pass
        out["disk_files"] = count
        out["disk_bytes"] = total
    except Exception as e:
        out["disk_walk_error"] = str(e)
    try:
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            vrow = conn.execute("SELECT VERSION() AS v").fetchone()
            out["mysql_version"] = vrow["v"] if vrow else None
        except Exception:
            pass
        try:
            rows = conn.execute("""
                SELECT id, grn_id, kind, original_name, mime_type,
                       size_bytes, relative_path, uploaded_at
                FROM pm_grn_files ORDER BY id DESC LIMIT 50
            """).fetchall()
            missing = 0; mismatch = 0; recent = []
            for r in rows:
                d = dict(r)
                if d.get("uploaded_at"):
                    d["uploaded_at"] = str(d["uploaded_at"])
                abs_path = _pmf_disk_path(d.get("relative_path"))
                if not abs_path or not os.path.isfile(abs_path):
                    d["disk_status"] = "missing-on-disk"; missing += 1
                else:
                    actual = os.path.getsize(abs_path)
                    d["actual_bytes"] = actual
                    if actual != (d.get("size_bytes") or 0):
                        d["disk_status"] = "size-mismatch"; mismatch += 1
                    else:
                        d["disk_status"] = "ok"
                recent.append(d)
            out["recent_files"]  = recent
            out["files_missing"] = missing
            out["size_mismatch"] = mismatch
        except Exception as e:
            out["recent_files_error"] = str(e)
        try: conn.close()
        except Exception: pass
        return jsonify(out)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


# ── /upload — POST one file ──────────────────────────────────────────
@pm_stock_bp.route('/api/pm_stock/grn/file/upload', methods=['POST'])
@_login_required
def api_pm_grn_file_upload():
    grn_id = request.form.get("grn_id")
    if not grn_id:
        return jsonify({"status": "error", "message": "grn_id is required"}), 400
    if "file" not in request.files:
        return jsonify({"status": "error", "message": "No file part in request"}), 400

    f = request.files["file"]
    if not f or not f.filename:
        return jsonify({"status": "error", "message": "Empty filename"}), 400

    mime = (f.mimetype or "").lower()
    if not _pmf_mime_ok(mime):
        return jsonify({"status": "error",
                        "message": f"Unsupported file type: {mime or 'unknown'}. Allowed: PDF, JPG, PNG"}), 400

    blob = f.read()
    size = len(blob)
    if size == 0:
        return jsonify({"status": "error", "message": "File is empty"}), 400
    if size > _PMF_MAX_BYTES:
        return jsonify({"status": "error",
                        "message": f"File too large ({size//1024} KB). Max {_PMF_MAX_BYTES//(1024*1024)} MB"}), 400

    abs_disk_path = None  # for rollback
    try:
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500

        grn_row = conn.execute(
            "SELECT id FROM pm_grn WHERE id=%s", (grn_id,)
        ).fetchone()
        if not grn_row:
            return jsonify({"status": "error", "message": "GRN not found"}), 404

        safe_name = _pmf_safe_segment(f.filename)
        disk_name = f"{uuid.uuid4().hex}__{safe_name}"
        rel_dir   = f"grn/{int(grn_id)}/invoice"
        rel_path  = f"{rel_dir}/{disk_name}"

        abs_dir = os.path.normpath(os.path.join(_PM_UPLOAD_ROOT, *rel_dir.split("/")))
        os.makedirs(abs_dir, exist_ok=True)
        abs_disk_path = os.path.join(abs_dir, disk_name)

        # Atomic write: temp file, then rename.
        tmp_path = abs_disk_path + ".tmp"
        with open(tmp_path, "wb") as fh:
            fh.write(blob); fh.flush()
            try: os.fsync(fh.fileno())
            except Exception: pass
        os.replace(tmp_path, abs_disk_path)

        user = session.get("User_Name") or session.get("user_name")
        conn.execute("""
            INSERT INTO pm_grn_files
                (grn_id, kind, original_name, mime_type, size_bytes,
                 relative_path, uploaded_by)
            VALUES (%s, 'invoice', %s, %s, %s, %s, %s)
        """, (grn_id, f.filename, mime, size, rel_path, user))
        file_id = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
        conn.commit()

        return jsonify({"status": "ok", "file": {
            "id":            file_id,
            "grn_id":        int(grn_id),
            "kind":          "invoice",
            "original_name": f.filename,
            "mime_type":     mime,
            "size_bytes":    size,
        }})
    except Exception as e:
        traceback.print_exc()
        if abs_disk_path and os.path.exists(abs_disk_path):
            try: os.remove(abs_disk_path)
            except Exception: pass
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        try: conn.close()
        except Exception: pass


# ── /list — files for one GRN ────────────────────────────────────────
@pm_stock_bp.route('/api/pm_stock/grn/file/list', methods=['GET'])
@_login_required
def api_pm_grn_file_list():
    grn_id = request.args.get("grn_id")
    if not grn_id:
        return jsonify({"status": "error", "message": "grn_id is required"}), 400
    try:
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        rows = conn.execute("""
            SELECT id, grn_id, kind, original_name, mime_type, size_bytes,
                   uploaded_by, uploaded_at
            FROM pm_grn_files WHERE grn_id=%s ORDER BY id
        """, (grn_id,)).fetchall()
        files = []
        for r in rows:
            d = dict(r)
            if d.get("uploaded_at"):
                d["uploaded_at"] = str(d["uploaded_at"])
            files.append(d)
        return jsonify({"status": "ok", "files": files})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        try: conn.close()
        except Exception: pass


# ── GET /<id> — download/inline-view ─────────────────────────────────
@pm_stock_bp.route('/api/pm_stock/grn/file/<int:file_id>', methods=['GET'])
@_login_required
def api_pm_grn_file_download(file_id):
    force_download = request.args.get("download") in ("1", "true", "yes")
    try:
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        row = conn.execute("""
            SELECT original_name, mime_type, relative_path
            FROM pm_grn_files WHERE id=%s
        """, (file_id,)).fetchone()
        if not row:
            return jsonify({"status": "error", "message": "File not found"}), 404
        d = dict(row)
        if not d.get("relative_path"):
            return jsonify({"status": "error", "message": "File has no disk path"}), 410
        abs_path = _pmf_disk_path(d["relative_path"])
        if not abs_path or not os.path.isfile(abs_path):
            return jsonify({"status": "error", "message": "File missing on disk"}), 404
        return send_file(
            abs_path,
            mimetype=d["mime_type"] or "application/octet-stream",
            as_attachment=force_download,
            download_name=d["original_name"],
            conditional=True,
            max_age=300,
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        try: conn.close()
        except Exception: pass


# ── DELETE /<id> — remove ────────────────────────────────────────────
@pm_stock_bp.route('/api/pm_stock/grn/file/<int:file_id>', methods=['DELETE'])
@_login_required
def api_pm_grn_file_delete(file_id):
    try:
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        row = conn.execute(
            "SELECT id, relative_path FROM pm_grn_files WHERE id=%s", (file_id,)
        ).fetchone()
        if not row:
            return jsonify({"status": "error", "message": "File not found"}), 404
        d = dict(row)
        # DB delete first — disk delete after (orphaned disk file is harmless,
        # phantom DB row is not).
        conn.execute("DELETE FROM pm_grn_files WHERE id=%s", (file_id,))
        conn.commit()
        try:
            abs_path = _pmf_disk_path(d.get("relative_path"))
            if abs_path and os.path.isfile(abs_path):
                os.remove(abs_path)
        except Exception:
            pass
        return jsonify({"status": "ok"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        try: conn.close()
        except Exception: pass


# ════════════════════════════════════════════════════════════════════════════
# END  PM GRN FILE ATTACHMENTS
# ════════════════════════════════════════════════════════════════════════════


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@pm_stock_bp.route('/pm_stock')
@_login_required
def pm_stock_page():
    # Template lives under templates/pm_stock/pm_stock.html so all
    # PM-Stock templates (this page, audit page, etc.) sit together.
    # is_requester restricts the UI to Request/Stock/IN-only for users
    # whose home godown is FACTORY. Admins always get the full UI.
    try:
        is_req = bool(_user_is_requester())
    except Exception:
        is_req = False  # fail-open: prefer full UI on any lookup error
    # `access` is a dict {category: bool} from pm_user_access; admins
    # always get True for every key. Templates use it via `access.X`
    # in {% if %} guards to hide sidebar items / tabs the user can't use.
    try:
        access = _user_access_dict()
    except Exception:
        access = {k: True for k in PM_USER_ACCESS_KEYS}
    return render_template(
        'pm_stock/pm_stock.html',
        user_name=session.get('User_Name'),
        role=session.get('User_Type'),
        is_requester=is_req,
        is_manager=_is_manager(),
        access=access,
        sidebar_menu=get_menu('pm_stock', access=access,
                              role=session.get('User_Type'), is_requester=is_req),
        active_item='stock',
        sidebar_extra_html=_PM_SIDEBAR_EXTRA,
    )


@pm_stock_bp.route('/pm_stock/logout')
def pm_stock_logout():
    """Log the user out: clear the session and send them to the portal's
    login page. The parent portal owns the actual login screen, so we land
    on the site root which redirects unauthenticated users to login.
    """
    try:
        session.clear()
    except Exception:
        # Defensive: even if clear() fails, drop the auth keys explicitly.
        for k in ('logged_in', 'User_Name', 'UID', 'User_Type'):
            session.pop(k, None)
    return redirect('/')


# ════════════════════════════════════════════════════════════════════════════
# ADMIN: BACKFILL SHORT CODES
# ────────────────────────────────────────────────────────────────────────────
# One-time (or rerun-safe) admin tool that assigns a sequential short_code
# to every pm_boxes row whose short_code is NULL or empty.
#
# Why this exists: older boxes (created before short_codes were rolled out,
# or before the assign-on-create hook fired) carry only the long box_code
# (e.g. BEARBOX570-G0067-B001). Newer label prints embed the short_code
# in the QR. When an old box's label is reprinted today, the QR ends up
# encoding the long code instead — but scanners and the rest of the UI
# now expect the compact form. Backfilling lets reprinting produce
# scan-clean QRs for every box in the warehouse.
#
# Safe to rerun:
#   * _assign_box_short_code is idempotent — boxes that already have a
#     short_code are skipped.
#   * Each assignment is wrapped in its own savepoint-style try/except so a
#     transient row failure doesn't abort the whole batch.
#
# Performance: paginates by box_id in batches so a 200k-box site doesn't
# blow MySQL packet limits or hold a long lock. The endpoint reports
# {scanned, assigned, skipped, errors} so the admin can run it once and
# verify completion.
# ════════════════════════════════════════════════════════════════════════════
@pm_stock_bp.route('/api/pm_stock/admin/backfill_short_codes', methods=['POST'])
@_login_required
def api_admin_backfill_short_codes():
    """Assign short_codes to every pm_boxes row that doesn't have one yet.

    Admin only. Returns counts so the admin can tell whether the run did
    anything. Idempotent — running it again on a fully-coded warehouse
    returns scanned=N, assigned=0.

    Optional body: {"batch_size": 500} — pagination chunk (default 500).
    """
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403

    d = request.get_json(silent=True) or {}
    try:
        batch_size = max(50, min(5000, int(d.get('batch_size') or 500)))
    except Exception:
        batch_size = 500

    scanned   = 0
    assigned  = 0
    skipped   = 0
    errors    = 0
    last_err  = None

    conn = sampling_portal.get_db_connection()
    try:
        # Pull just the IDs of boxes lacking a short_code — keep the
        # working set tiny. The SELECT runs once at start of run; new
        # boxes created during the loop will be caught on a rerun.
        rows = conn.execute(
            """SELECT box_id FROM pm_boxes
               WHERE short_code IS NULL OR short_code = ''
               ORDER BY box_id"""
        ).fetchall() or []
        ids = [int(r['box_id'] if hasattr(r, 'get') else r[0]) for r in rows]
        scanned = len(ids)

        # Process in chunks. Each chunk commits independently so a failure
        # in chunk N+1 doesn't roll back chunk N.
        for i in range(0, len(ids), batch_size):
            chunk = ids[i:i + batch_size]
            for bid in chunk:
                try:
                    code = _assign_box_short_code(conn, bid)
                    if code:
                        # _assign_box_short_code returns the existing code
                        # if the row was already coded — but our SELECT
                        # filtered those out, so any return value here
                        # represents a fresh assignment. We don't rely on
                        # distinguishing the two cases.
                        assigned += 1
                    else:
                        skipped += 1
                except Exception as _bid_err:
                    errors += 1
                    last_err = str(_bid_err)
                    # Continue with the next box — one bad row should not
                    # abort the whole backfill.
                    continue
            # Commit chunk so partial progress sticks if a later chunk
            # blows up.
            try: conn.commit()
            except Exception: pass

        return jsonify({
            'status':   'ok',
            'scanned':  scanned,
            'assigned': assigned,
            'skipped':  skipped,
            'errors':   errors,
            'last_error': last_err,
            'message':  (f'Backfill complete: {assigned} new short_codes assigned '
                         f'(scanned {scanned}, errors {errors})'),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e),
                        'scanned': scanned, 'assigned': assigned,
                        'errors': errors}), 500
    finally:
        try: conn.close()
        except Exception: pass


@pm_stock_bp.route('/api/pm_stock/recycle_bin', methods=['GET'])
@_login_required
def api_recycle_bin_list():
    """List bin entries. Admin-only.
    Query params: entity_type, deleted_by, from_date, to_date, status (active|restored|all)
    """
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403

    entity_type = (request.args.get('entity_type') or '').strip()
    deleted_by  = (request.args.get('deleted_by')  or '').strip()
    from_date   = (request.args.get('from_date')   or '').strip()
    to_date     = (request.args.get('to_date')     or '').strip()
    status      = (request.args.get('status') or 'active').strip().lower()
    if status not in ('active', 'restored', 'all'):
        status = 'active'

    where, params = [], []
    if entity_type:
        where.append("entity_type = %s"); params.append(entity_type)
    if deleted_by:
        where.append("deleted_by LIKE %s"); params.append(f'%{deleted_by}%')
    if from_date:
        where.append("DATE(deleted_at) >= %s"); params.append(from_date)
    if to_date:
        where.append("DATE(deleted_at) <= %s"); params.append(to_date)
    if status == 'active':
        where.append("restored_at IS NULL")
    elif status == 'restored':
        where.append("restored_at IS NOT NULL")

    wc = (' WHERE ' + ' AND '.join(where)) if where else ''
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(f"""
            SELECT bin_id, entity_type, entity_label, entity_id,
                   payload_summary, deleted_by, deleted_at,
                   restored_at, restored_by, reason,
                   CHAR_LENGTH(payload) AS payload_size
            FROM pm_recycle_bin
            {wc}
            ORDER BY deleted_at DESC
            LIMIT 500
        """, params).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            for k in ('deleted_at', 'restored_at'):
                if d.get(k) is not None:
                    d[k] = str(d[k])
            out.append(d)
        return jsonify({'status': 'ok', 'entries': out, 'count': len(out)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/recycle_bin/<int:bin_id>', methods=['GET'])
@_login_required
def api_recycle_bin_detail(bin_id):
    """Return the full payload for inspection. Admin-only."""
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM pm_recycle_bin WHERE bin_id=%s",
            (bin_id,)
        ).fetchone()
        conn.close()
        if not row:
            return jsonify({'status': 'error', 'message': 'Not found'}), 404
        d = dict(row)
        for k in ('deleted_at', 'restored_at'):
            if d.get(k) is not None: d[k] = str(d[k])
        try:
            d['payload_parsed'] = json.loads(d.get('payload') or '{}')
        except Exception:
            d['payload_parsed'] = None
        return jsonify({'status': 'ok', 'entry': d})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/recycle_bin/restore', methods=['POST'])
@_login_required
def api_recycle_bin_restore():
    """Restore a bin entry. Admin-only.
    Body: { bin_id }
    """
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    d = request.get_json() or {}
    bin_id = d.get('bin_id')
    if not bin_id:
        return jsonify({'status': 'error', 'message': 'bin_id required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        info = _bin_restore(conn, int(bin_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', **info})
    except ValueError as ve:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(ve)}), 400
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/recycle_bin/purge', methods=['POST'])
@_login_required
def api_recycle_bin_purge():
    """Permanently delete a bin entry. Admin-only. Irreversible.
    Body: { bin_id }
    """
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    d = request.get_json() or {}
    bin_id = d.get('bin_id')
    if not bin_id:
        return jsonify({'status': 'error', 'message': 'bin_id required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        _bin_purge(conn, int(bin_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/reprint/request', methods=['POST'])
@_login_required
def api_reprint_request():
    """Create a reprint request. Body: {
       scope_type, voucher_kind, voucher_id, voucher_label, box_codes (list), reason,
       (edit-mode optional, paired):
         product_id, godown_id, new_no_of_box, new_per_box_qty
       (selective-mode — for per-box selection requests):
         product_id, godown_id, selections (list), removed_box_ids (list)
    }

    Three reprint flavours:
      1. Pure reprint (no fields beyond scope/box_codes/reason):
         redemption re-renders the labels with no DB changes.
      2. Edit reprint (new_no_of_box + new_per_box_qty paired):
         redemption recreates boxes with new uniform dimensions.
      3. Selective reprint (selections list set):
         redemption applies the same per-box selection logic admins
         use when they hit the selective-reprint endpoint directly.
         Constraint: sum(selections.per_box_qty) MUST equal the
         existing total — the server enforces this on redemption.

    Edit-mode rules:
      - new_no_of_box and new_per_box_qty must both be provided OR both absent.
      - If both are absent: pure reprint (no data change). Box records and
        ledger remain untouched; only labels are re-rendered.
      - If both are present: 'edit reprint' — the redemption step recreates
        boxes with new dimensions AND updates the ledger to match. The
        request payload carries the new values so the admin sees what
        they're approving.
      - product_id is required for both scopes when in edit mode.
      - godown_id is required only for OP scope edit mode.

    Selective-mode rules:
      - selections must be a non-empty list and edit-mode fields must be absent.
      - product_id is required.
      - godown_id is required for OP scope.
      - The total qty constraint is checked on redemption, not on submission,
        so the admin can reject a request whose totals don't match.

    Admins should not need this — calling it as admin still works but is
    discouraged (admins print directly without approval). The frontend
    skips the request flow for admins.
    """
    d = request.get_json() or {}
    scope_type    = (d.get('scope_type') or '').strip()
    voucher_kind  = (d.get('voucher_kind') or '').strip().lower() or None
    voucher_id    = d.get('voucher_id')
    voucher_label = (d.get('voucher_label') or '').strip()[:120] or None
    box_codes     = d.get('box_codes') or []
    reason        = (d.get('reason') or '').strip()[:500]

    # Edit-mode fields (all optional individually but with paired constraint)
    edit_product_id    = d.get('product_id')
    edit_godown_id     = d.get('godown_id')
    edit_no_of_box_raw = d.get('new_no_of_box')
    edit_per_box_qty_raw = d.get('new_per_box_qty')

    # Selective-mode fields
    selections_in    = d.get('selections')
    removed_in       = d.get('removed_box_ids') or []

    if scope_type not in VALID_REPRINT_SCOPES:
        return jsonify({'status': 'error', 'message': f'scope_type must be one of {VALID_REPRINT_SCOPES}'}), 400
    if not reason:
        return jsonify({'status': 'error', 'message': 'A reason is required'}), 400
    if scope_type in ('voucher_grn', 'voucher_op'):
        if not voucher_id:
            return jsonify({'status': 'error', 'message': 'voucher_id required for voucher-scope reprint'}), 400
        if voucher_kind not in ('grn', 'op'):
            return jsonify({'status': 'error', 'message': "voucher_kind must be 'grn' or 'op'"}), 400
    if scope_type == 'boxes':
        if not isinstance(box_codes, list) or not box_codes:
            return jsonify({'status': 'error', 'message': 'box_codes must be a non-empty list'}), 400
    box_csv = ','.join(str(c).strip() for c in box_codes if str(c).strip())[:65000] or None

    # Edit-mode: parse + validate paired constraint
    new_nob = None
    new_pbq = None
    is_edit = False
    has_any_edit = (edit_no_of_box_raw is not None and edit_no_of_box_raw != '') or \
                   (edit_per_box_qty_raw is not None and edit_per_box_qty_raw != '')
    if has_any_edit:
        try:
            new_nob = int(edit_no_of_box_raw or 0)
            new_pbq = float(edit_per_box_qty_raw or 0)
        except Exception:
            return jsonify({'status': 'error', 'message': 'new_no_of_box and new_per_box_qty must be numbers'}), 400
        if new_nob <= 0 or new_pbq <= 0:
            return jsonify({'status': 'error', 'message': 'When editing, both new_no_of_box and new_per_box_qty must be > 0'}), 400
        if not edit_product_id:
            return jsonify({'status': 'error', 'message': 'product_id required when reprinting with edits'}), 400
        if scope_type == 'voucher_op' and not edit_godown_id:
            return jsonify({'status': 'error', 'message': 'godown_id required when reprinting OP labels with edits'}), 400
        is_edit = True

    # Selective-mode: parse + validate
    selections_json = None
    is_selective = False
    if isinstance(selections_in, list) and selections_in:
        if is_edit:
            return jsonify({'status': 'error',
                'message': 'Cannot mix edit-mode (new_no_of_box/new_per_box_qty) with selective-mode (selections). Pick one.'}), 400
        if not edit_product_id:
            return jsonify({'status': 'error',
                'message': 'product_id required for selective reprint requests'}), 400
        if scope_type == 'voucher_op' and not edit_godown_id:
            return jsonify({'status': 'error',
                'message': 'godown_id required for selective reprint of OP labels'}), 400
        # Sanitise + size-cap each selection
        clean_selections = []
        for s in selections_in:
            if not isinstance(s, dict): continue
            try:
                pbq = float(s.get('per_box_qty') or 0)
                seq = int(s.get('box_seq') or 0)
            except Exception:
                return jsonify({'status':'error','message':'Each selection needs numeric per_box_qty and box_seq'}), 400
            if pbq <= 0:
                return jsonify({'status':'error','message':f'Selection box_seq={seq} has non-positive per_box_qty'}), 400
            clean_selections.append({
                'box_id':      (int(s['box_id']) if s.get('box_id') is not None else None),
                'box_seq':     seq,
                'per_box_qty': pbq,
                'is_new':      bool(s.get('is_new')),
            })
        if not clean_selections:
            return jsonify({'status': 'error','message': 'At least one valid selection is required'}), 400
        # Removed-box ids — optional, ints only
        clean_removed = []
        if isinstance(removed_in, list):
            for rid in removed_in:
                try: clean_removed.append(int(rid))
                except Exception: pass
        selections_json = json.dumps({
            'selections':      clean_selections,
            'removed_box_ids': clean_removed,
        })
        # Selections column is MEDIUMTEXT (16 MB capacity). Cap to ~5 MB
        # as a sanity bound — a 5 MB JSON blob would mean ~70k boxes in a
        # single reprint request, far beyond any realistic batch. Reject
        # anything bigger so an accidentally-pathological payload can't
        # bloat the DB.
        if len(selections_json) > 5_000_000:
            return jsonify({'status': 'error',
                'message': 'Too many box selections — split into multiple requests.'}), 400
        is_selective = True

    conn = sampling_portal.get_db_connection()
    try:
        # Reject duplicate pending requests for the exact same scope from the
        # same user — they should wait for the existing one to be decided.
        if scope_type in ('voucher_grn', 'voucher_op'):
            dupe = conn.execute(
                """SELECT req_id FROM pm_label_reprint_requests
                   WHERE status='pending' AND requested_by=%s
                     AND scope_type=%s AND voucher_kind=%s AND voucher_id=%s
                   LIMIT 1""",
                (_user(), scope_type, voucher_kind, int(voucher_id))
            ).fetchone()
            if dupe:
                conn.close()
                return jsonify({
                    'status': 'error',
                    'message': f'You already have a pending reprint request (#{dupe["req_id"]}) for this voucher. Wait for it to be decided.'
                }), 409

        cur = conn.execute(
            """INSERT INTO pm_label_reprint_requests
                 (scope_type, voucher_kind, voucher_id, voucher_label,
                  box_codes_csv, requested_by, reason, status,
                  product_id, godown_id, new_no_of_box, new_per_box_qty,
                  selections_json)
               VALUES (%s,%s,%s,%s,%s,%s,%s,'pending',%s,%s,%s,%s,%s)""",
            (scope_type, voucher_kind,
             int(voucher_id) if voucher_id else None,
             voucher_label, box_csv, _user(), reason,
             int(edit_product_id) if edit_product_id else None,
             int(edit_godown_id)  if edit_godown_id  else None,
             new_nob, new_pbq,
             selections_json)
        )
        req_id = cur.lastrowid

        # Audit
        try:
            _audit_record(
                conn,
                action='op.reprint_request',
                entity='reprint_request',
                entity_id=req_id,
                summary=f"Reprint requested for {voucher_kind or 'unknown'} {voucher_label or '?'} ({'selective' if is_selective else 'full'})",
                before=None,
                after={
                    'scope_type': scope_type, 'voucher_kind': voucher_kind,
                    'voucher_id': voucher_id, 'voucher_label': voucher_label,
                    'is_selective': bool(is_selective),
                    'reason': reason,
                },
            )
        except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'req_id': req_id, 'is_selective': is_selective})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/reprint/requests', methods=['GET'])
@_login_required
def api_reprint_list():
    """List reprint requests. Filters: status, scope, requested_by.

    Admins see everything. Non-admins only see their own requests.
    """
    status_f = (request.args.get('status') or '').strip().lower()
    scope_f  = (request.args.get('scope')  or '').strip().lower()
    requester_f = (request.args.get('requested_by') or '').strip()
    where, params = [], []
    if status_f:
        where.append("status=%s"); params.append(status_f)
    if scope_f:
        where.append("scope_type=%s"); params.append(scope_f)

    is_admin = _is_admin()
    if not is_admin:
        # Non-admin: filter to own requests regardless of any requested_by query
        where.append("requested_by=%s"); params.append(_user())
    elif requester_f:
        where.append("requested_by LIKE %s"); params.append(f'%{requester_f}%')

    wc = (' WHERE ' + ' AND '.join(where)) if where else ''
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(f"""
            SELECT req_id, scope_type, voucher_kind, voucher_id, voucher_label,
                   box_codes_csv, requested_by, requested_at, reason, status,
                   approved_by, approved_at, decided_note,
                   print_token, printed_at, printed_by,
                   product_id, godown_id, new_no_of_box, new_per_box_qty,
                   selections_json
            FROM pm_label_reprint_requests
            {wc}
            ORDER BY requested_at DESC
            LIMIT 200
        """, params).fetchall()
        # Pull per-box print status for every approved request that has
        # child rows. Pre-existing approvals (legacy) won't have any —
        # those just get an empty 'box_status' list and the UI falls back
        # to the old "Print all" button.
        req_ids = [r['req_id'] for r in rows]
        box_status_map = {}
        if req_ids:
            try:
                bs_rows = conn.execute(
                    f"""SELECT req_id, box_code, printed_at, printed_by
                        FROM pm_label_reprint_box_status
                        WHERE req_id IN ({','.join(['%s']*len(req_ids))})
                        ORDER BY req_id, box_code""",
                    req_ids
                ).fetchall()
                for bs in bs_rows:
                    box_status_map.setdefault(bs['req_id'], []).append({
                        'box_code':   bs['box_code'],
                        'printed':    bs['printed_at'] is not None,
                        'printed_at': str(bs['printed_at']) if bs['printed_at'] is not None else None,
                        'printed_by': bs['printed_by'],
                    })
            except Exception:
                # Table may not exist yet on a stale schema — just skip,
                # frontend handles missing box_status gracefully.
                pass
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            for k in ('requested_at', 'approved_at', 'printed_at'):
                if d.get(k) is not None: d[k] = str(d[k])
            # Coerce Decimal → float so the JS gets clean numbers
            if d.get('new_per_box_qty') is not None:
                try: d['new_per_box_qty'] = float(d['new_per_box_qty'])
                except Exception: d['new_per_box_qty'] = None
            # Parse selections_json so the UI can render it directly
            if d.get('selections_json'):
                try:
                    d['selections'] = json.loads(d['selections_json'])
                except Exception:
                    d['selections'] = None
            else:
                d['selections'] = None
            # Drop the raw column — clients use `selections` instead
            d.pop('selections_json', None)
            # Attach per-box print status (empty list for legacy approvals)
            d['box_status'] = box_status_map.get(d['req_id'], [])
            # Hide token from non-admin if not the requester (defense in depth)
            if not is_admin and d.get('requested_by') != _user():
                d['print_token'] = None
                d['box_codes_csv'] = None
            out.append(d)
        return jsonify({'status': 'ok', 'requests': out, 'count': len(out)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/reprint/pending_count', methods=['GET'])
@_login_required
def api_reprint_pending_count():
    """Polled by the admin notification badge. Returns count of pending
    reprint requests (admins see all, non-admins see own pending only)."""
    conn = sampling_portal.get_db_connection()
    try:
        if _is_admin():
            row = conn.execute(
                "SELECT COUNT(*) AS c FROM pm_label_reprint_requests WHERE status='pending'"
            ).fetchone()
        else:
            row = conn.execute(
                """SELECT COUNT(*) AS c FROM pm_label_reprint_requests
                   WHERE requested_by=%s AND status IN ('pending','approved')""",
                (_user(),)
            ).fetchone()
        conn.close()
        return jsonify({'status': 'ok', 'count': int((row or {}).get('c') or 0)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/reprint/<int:req_id>/approve', methods=['POST'])
@_login_required
def _approve_one_request(conn, req_id, note):
    """
    Approve a single reprint request and create one pm_label_reprint_box_status
    row per box in its CSV. Returns (ok: bool, message: str).

    Caller is responsible for transaction management (commit/rollback) and
    audit logging. This helper is shared by api_reprint_approve (single)
    and api_reprint_approve_bulk (multi).
    """
    row = conn.execute(
        "SELECT status, box_codes_csv FROM pm_label_reprint_requests WHERE req_id=%s",
        (req_id,)
    ).fetchone()
    if not row:
        return False, f'Request #{req_id} not found'
    if row['status'] != 'pending':
        return False, f"Request #{req_id} is '{row['status']}', not pending"
    token = _gen_print_token()
    conn.execute(
        """UPDATE pm_label_reprint_requests
           SET status='approved', approved_by=%s, approved_at=NOW(),
               decided_note=%s, print_token=%s
           WHERE req_id=%s""",
        (_user(), note, token, req_id)
    )
    # ── Per-box child rows ─────────────────────────────────────────
    # Parse the CSV and create one tracking row per box. These let the
    # user print labels one at a time; each box becomes individually
    # consumable. Requests with no CSV (rare — full-voucher reprints
    # that resolve to boxes at print time) get no child rows, which
    # the print endpoint detects and falls back to legacy bulk-print.
    csv = (row['box_codes_csv'] or '').strip()
    if csv:
        seen = set()
        for raw in csv.split(','):
            code = raw.strip()
            if not code or code in seen:
                continue
            seen.add(code)
            try:
                conn.execute(
                    """INSERT INTO pm_label_reprint_box_status (req_id, box_code)
                       VALUES (%s, %s)""",
                    (req_id, code[:120])
                )
            except Exception:
                # PK collision (re-approval edge case) — already tracked, skip
                pass
    return True, 'ok'


def api_reprint_approve(req_id):
    """Admin approves a reprint request. Issues a single-use print token."""
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    d = request.get_json() or {}
    note = (d.get('note') or '').strip()[:500] or None
    conn = sampling_portal.get_db_connection()
    try:
        ok, msg = _approve_one_request(conn, req_id, note)
        if not ok:
            conn.close()
            return jsonify({'status':'error','message': msg}), 409

        try:
            _audit_record(
                conn,
                action='op.reprint_approve',
                entity='reprint_request',
                entity_id=req_id,
                summary=f"Reprint request #{req_id} approved",
                before={'status': 'pending'},
                after={'status': 'approved', 'note': note},
            )
        except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'req_id': req_id})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/reprint/approve_bulk', methods=['POST'])
@_login_required
def api_reprint_approve_bulk():
    """
    Admin approves multiple pending reprint requests in a single call.
    Body: {req_ids: [int, ...], note: optional shared note}

    Each req_id is processed independently via _approve_one_request; the
    response includes a per-req status so the UI can show which ones
    succeeded and which were skipped (already decided, not found, etc.).
    All successful approvals commit in one transaction; an unexpected
    error rolls all of them back.
    """
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    d = request.get_json() or {}
    raw_ids = d.get('req_ids') or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return jsonify({'status':'error','message':'req_ids must be a non-empty list'}), 400
    # Coerce + dedupe
    try:
        req_ids = sorted(set(int(x) for x in raw_ids))
    except (TypeError, ValueError):
        return jsonify({'status':'error','message':'req_ids must contain integers'}), 400
    note = (d.get('note') or '').strip()[:500] or None

    conn = sampling_portal.get_db_connection()
    try:
        results = []
        approved_ids = []
        for rid in req_ids:
            ok, msg = _approve_one_request(conn, rid, note)
            results.append({'req_id': rid, 'ok': ok, 'message': msg})
            if ok:
                approved_ids.append(rid)

        # Single audit row summarising the batch action — individual
        # before/after pairs aren't worth the noise for bulk admin actions.
        if approved_ids:
            try:
                _audit_record(
                    conn,
                    action='op.reprint_approve_bulk',
                    entity='reprint_request',
                    entity_id=None,
                    summary=f"Bulk approved {len(approved_ids)} reprint request(s): {approved_ids}",
                    before=None,
                    after={'approved_ids': approved_ids, 'note': note},
                )
            except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({
            'status':       'ok',
            'approved':     len(approved_ids),
            'skipped':      len(req_ids) - len(approved_ids),
            'total':        len(req_ids),
            'per_request':  results,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/reprint/<int:req_id>/reject', methods=['POST'])
@_login_required
def api_reprint_reject(req_id):
    """Admin rejects a pending request."""
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    d = request.get_json() or {}
    note = (d.get('note') or '').strip()[:500] or None
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT status FROM pm_label_reprint_requests WHERE req_id=%s", (req_id,)
        ).fetchone()
        if not row:
            conn.close(); return jsonify({'status':'error','message':'Request not found'}), 404
        if row['status'] != 'pending':
            conn.close(); return jsonify({
                'status': 'error',
                'message': f"Request is currently '{row['status']}', not pending — cannot reject."
            }), 409
        conn.execute(
            """UPDATE pm_label_reprint_requests
               SET status='rejected', approved_by=%s, approved_at=NOW(), decided_note=%s
               WHERE req_id=%s""",
            (_user(), note, req_id)
        )

        try:
            _audit_record(
                conn,
                action='op.reprint_reject',
                entity='reprint_request',
                entity_id=req_id,
                summary=f"Reprint request #{req_id} rejected",
                before={'status': 'pending'},
                after={'status': 'rejected', 'note': note},
            )
        except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ════════════════════════════════════════════════════════════════════════════
# FIFO OVERRIDE APPROVAL WORKFLOW
# ════════════════════════════════════════════════════════════════════════════
#
# Non-admins can't force a FIFO violation through during Material OUT scanning.
# Instead they raise an override *request* (carrying the violation snapshot +
# a reason). An admin approves or rejects it. An approved request grants the
# requester a single-use pass to re-scan that EXACT box; the OUT scan path
# (api_voucher_scan_box) consumes it via _fifo_override_find_approved /
# _fifo_override_mark_used.
#
# Endpoints:
#   POST /api/pm_stock/fifo_override/request                  — non-admin raises a request
#   GET  /api/pm_stock/fifo_override/requests                 — list (admin: all; user: own)
#   GET  /api/pm_stock/fifo_override/pending_count            — badge poll
#   POST /api/pm_stock/fifo_override/<req_id>/approve         — admin approves
#   POST /api/pm_stock/fifo_override/<req_id>/reject          — admin rejects

@pm_stock_bp.route('/api/pm_stock/fifo_override/request', methods=['POST'])
@_login_required
def api_fifo_override_request():
    """A non-admin requests permission to override a FIFO violation for a
    specific box on a specific transfer.

    Body: {
        transfer_id (int, required),
        box_id      (int, required),
        reason      (str, required),
        fifo        (dict, optional — the violation snapshot from the
                     scan_box 409 response; used to capture context)
    }

    Admins don't need this (they force directly), but calling it as admin
    still works and is harmless — the frontend simply doesn't surface the
    request button for admins.
    """
    d = request.get_json() or {}
    transfer_id = d.get('transfer_id')
    box_id      = d.get('box_id')
    reason      = (d.get('reason') or '').strip()[:500]
    fifo        = d.get('fifo') or {}

    if not transfer_id or not box_id:
        return jsonify({'status': 'error', 'message': 'transfer_id and box_id are required'}), 400
    if not reason:
        return jsonify({'status': 'error', 'message': 'A reason is required'}), 400

    try:
        transfer_id = int(transfer_id)
        box_id      = int(box_id)
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'transfer_id and box_id must be integers'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        ok, msg, req_id = _fifo_override_insert_one(conn, transfer_id, box_id, reason, fifo)
        if not ok:
            conn.close()
            return jsonify({'status': 'error', 'message': msg}), 409
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'req_id': req_id})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


def _fifo_override_insert_one(conn, transfer_id, box_id, reason, fifo=None):
    """Insert ONE pending FIFO-override request. Shared by the single-request
    and bulk-request endpoints. Caller manages the transaction (commit/rollback)
    and HTTP shaping. Returns (ok: bool, message: str, req_id: int|None).

    `fifo` is the optional client-supplied violation snapshot dict. When the
    bulk path doesn't carry one, we recompute the snapshot server-side from
    _fifo_check_oldest so the admin still sees real context.
    """
    fifo = fifo or {}
    try:
        transfer_id = int(transfer_id)
        box_id      = int(box_id)
    except (TypeError, ValueError):
        return False, 'transfer_id and box_id must be integers', None
    reason = (reason or '').strip()[:500]
    if not reason:
        return False, 'A reason is required for every box', None

    t = conn.execute(
        "SELECT transfer_id, transfer_no, from_godown_id, status FROM pm_transfers WHERE transfer_id=%s",
        (transfer_id,)
    ).fetchone()
    if not t:
        return False, f'Transfer {transfer_id} not found', None

    b = conn.execute(
        """SELECT b.box_id, b.box_code, b.product_id, b.current_godown_id,
                  b.current_status, p.product_name
           FROM pm_boxes b
           LEFT JOIN pm_products p ON p.id = b.product_id
           WHERE b.box_id=%s LIMIT 1""",
        (box_id,)
    ).fetchone()
    if not b:
        return False, f'Box {box_id} not found', None

    # Reject duplicate open requests for the same box on this transfer.
    dupe = conn.execute(
        """SELECT req_id FROM pm_fifo_override_requests
           WHERE transfer_id=%s AND box_id=%s AND requested_by=%s
             AND status IN ('pending','approved')
           ORDER BY requested_at DESC LIMIT 1""",
        (transfer_id, box_id, _user())
    ).fetchone()
    if dupe:
        return False, f'Box {b["box_code"]} already has an open override request (#{dupe["req_id"]})', None

    # If the caller didn't pass a snapshot, recompute it server-side so the
    # admin still sees the violation context for bulk-submitted boxes.
    if not fifo:
        try:
            v = _fifo_check_oldest(
                conn, product_id=b['product_id'],
                godown_id=t['from_godown_id'], scanned_box_id=box_id
            )
            if v:
                fifo = v
        except Exception:
            fifo = {}

    def _g(k, default=None):
        v = fifo.get(k)
        return v if v not in (None, '') else default

    try:
        oqty = float(_g('oldest_total_qty', 0) or 0)
    except Exception:
        oqty = 0
    try:
        obc = int(_g('oldest_box_count', 0) or 0)
    except Exception:
        obc = 0

    cur = conn.execute(
        """INSERT INTO pm_fifo_override_requests
             (transfer_id, transfer_no, box_id, box_code, product_id, product_name,
              from_godown_id, scanned_fifo_code, oldest_fifo_code, oldest_voucher,
              oldest_supplier, oldest_date, oldest_box_count, oldest_total_qty,
              requested_by, reason, status)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'pending')""",
        (
            transfer_id, t['transfer_no'], box_id, b['box_code'],
            b['product_id'], b['product_name'], t['from_godown_id'],
            (_g('scanned_fifo_code') or '')[:16] or None,
            (_g('oldest_fifo_code') or '')[:16] or None,
            (_g('oldest_voucher') or '')[:64] or None,
            (_g('oldest_supplier') or '')[:255] or None,
            (str(_g('oldest_date') or '')[:20]) or None,
            obc, oqty,
            _user(), reason,
        )
    )
    req_id = cur.lastrowid

    try:
        _audit_record(
            conn,
            action='fifo_override.request',
            entity='fifo_override_request',
            entity_id=req_id,
            summary=f"FIFO override requested for box {b['box_code']} on {t['transfer_no']}",
            before=None,
            after={
                'transfer_no': t['transfer_no'], 'box_code': b['box_code'],
                'scanned_fifo_code': _g('scanned_fifo_code'),
                'oldest_fifo_code': _g('oldest_fifo_code'),
                'reason': reason,
            },
        )
    except Exception:
        pass
    return True, 'ok', req_id


@pm_stock_bp.route('/api/pm_stock/fifo_override/request_bulk', methods=['POST'])
@_login_required
def api_fifo_override_request_bulk():
    """A non-admin submits override requests for MULTIPLE boxes at once.
    Each box carries its own reason.

    Body: {
        transfer_id (int, required — all boxes belong to one transfer/OUT),
        items: [ { box_id (int), reason (str) }, ... ]   (required, non-empty)
    }

    Each item is processed independently; the response includes a per-item
    result so the UI can show which were created and which were skipped
    (duplicate open request, box not found, missing reason, etc.). All
    successful inserts commit together.
    """
    d = request.get_json() or {}
    transfer_id = d.get('transfer_id')
    items       = d.get('items') or []

    if not transfer_id:
        return jsonify({'status': 'error', 'message': 'transfer_id is required'}), 400
    if not isinstance(items, list) or not items:
        return jsonify({'status': 'error', 'message': 'items must be a non-empty list'}), 400
    try:
        transfer_id = int(transfer_id)
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'transfer_id must be an integer'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        results = []
        created_ids = []
        for it in items:
            if not isinstance(it, dict):
                results.append({'box_id': None, 'ok': False, 'message': 'Invalid item'})
                continue
            bid    = it.get('box_id')
            reason = (it.get('reason') or '').strip()[:500]
            fifo   = it.get('fifo') or {}
            ok, msg, rid = _fifo_override_insert_one(conn, transfer_id, bid, reason, fifo)
            results.append({'box_id': bid, 'ok': ok, 'message': msg, 'req_id': rid})
            if ok:
                created_ids.append(rid)

        conn.commit()
        conn.close()
        return jsonify({
            'status':      'ok',
            'created':     len(created_ids),
            'skipped':     len(results) - len(created_ids),
            'total':       len(results),
            'created_ids': created_ids,
            'per_item':    results,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/fifo_override/requests', methods=['GET'])
@_login_required
def api_fifo_override_list():
    """List FIFO override requests. Filters: status, requested_by (admin only).
    Admins see everything; non-admins see only their own."""
    status_f    = (request.args.get('status') or '').strip().lower()
    requester_f = (request.args.get('requested_by') or '').strip()
    where, params = [], []
    if status_f:
        where.append("status=%s"); params.append(status_f)

    is_admin = _is_admin()
    if not is_admin:
        where.append("requested_by=%s"); params.append(_user())
    elif requester_f:
        where.append("requested_by LIKE %s"); params.append(f'%{requester_f}%')

    wc = (' WHERE ' + ' AND '.join(where)) if where else ''
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(f"""
            SELECT req_id, transfer_id, transfer_no, box_id, box_code,
                   product_id, product_name, from_godown_id,
                   scanned_fifo_code, oldest_fifo_code, oldest_voucher,
                   oldest_supplier, oldest_date, oldest_box_count, oldest_total_qty,
                   requested_by, requested_at, reason, status,
                   decided_by, decided_at, decided_note, used_at
            FROM pm_fifo_override_requests
            {wc}
            ORDER BY
              CASE status WHEN 'pending' THEN 0 WHEN 'approved' THEN 1 ELSE 2 END,
              requested_at DESC
            LIMIT 300
        """, params).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            for k in ('requested_at', 'decided_at', 'used_at'):
                if d.get(k) is not None: d[k] = str(d[k])
            if d.get('oldest_total_qty') is not None:
                try: d['oldest_total_qty'] = float(d['oldest_total_qty'])
                except Exception: d['oldest_total_qty'] = 0
            out.append(d)
        return jsonify({'status': 'ok', 'requests': out, 'count': len(out)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/fifo_override/pending_count', methods=['GET'])
@_login_required
def api_fifo_override_pending_count():
    """Badge poll. Admins: count of pending requests. Non-admins: count of
    their own requests that are pending or approved-but-unused (i.e. still
    actionable by them)."""
    conn = sampling_portal.get_db_connection()
    try:
        if _is_admin():
            row = conn.execute(
                "SELECT COUNT(*) AS c FROM pm_fifo_override_requests WHERE status='pending'"
            ).fetchone()
        else:
            row = conn.execute(
                """SELECT COUNT(*) AS c FROM pm_fifo_override_requests
                   WHERE requested_by=%s AND status IN ('pending','approved')""",
                (_user(),)
            ).fetchone()
        conn.close()
        return jsonify({'status': 'ok', 'count': int((row or {}).get('c') or 0)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/fifo_override/<int:req_id>/approve', methods=['POST'])
@_login_required
def api_fifo_override_approve(req_id):
    """Admin approves a FIFO override request. The requester can then
    re-scan the exact box to push the OUT through (single-use)."""
    if not _user_has_access('fifo_override'):
        return jsonify({'status': 'error', 'message': 'You do not have access to FIFO Override approvals'}), 403
    d = request.get_json() or {}
    note = (d.get('note') or '').strip()[:500] or None
    conn = sampling_portal.get_db_connection()
    try:
        ok, msg = _fifo_override_approve_one(conn, req_id, note)
        if not ok:
            conn.close()
            return jsonify({'status': 'error', 'message': msg}), 409
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'req_id': req_id})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


def _fifo_override_approve_one(conn, req_id, note):
    """Approve a single pending FIFO-override request. Shared by the single
    and bulk approve endpoints. Caller manages transaction + HTTP shaping.
    Returns (ok: bool, message: str)."""
    row = conn.execute(
        "SELECT status, box_code, transfer_no FROM pm_fifo_override_requests WHERE req_id=%s",
        (req_id,)
    ).fetchone()
    if not row:
        return False, f'Request #{req_id} not found'
    if row['status'] != 'pending':
        return False, f"Request #{req_id} is '{row['status']}', not pending"
    conn.execute(
        """UPDATE pm_fifo_override_requests
           SET status='approved', decided_by=%s, decided_at=NOW(), decided_note=%s
           WHERE req_id=%s""",
        (_user(), note, req_id)
    )
    try:
        _audit_record(
            conn,
            action='fifo_override.approve',
            entity='fifo_override_request',
            entity_id=req_id,
            summary=f"FIFO override #{req_id} approved (box {row['box_code']} on {row['transfer_no']})",
            before={'status': 'pending'},
            after={'status': 'approved', 'note': note},
        )
    except Exception:
        pass
    return True, 'ok'


@pm_stock_bp.route('/api/pm_stock/fifo_override/approve_bulk', methods=['POST'])
@_login_required
def api_fifo_override_approve_bulk():
    """Admin approves multiple pending FIFO override requests in one call.
    Body: {req_ids: [int, ...], note: optional shared note}

    Each req_id is processed independently; the response carries a per-req
    result. All successful approvals commit together; an unexpected error
    rolls them all back. One summary audit row covers the batch."""
    if not _user_has_access('fifo_override'):
        return jsonify({'status': 'error', 'message': 'You do not have access to FIFO Override approvals'}), 403
    d = request.get_json() or {}
    raw_ids = d.get('req_ids') or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return jsonify({'status': 'error', 'message': 'req_ids must be a non-empty list'}), 400
    try:
        req_ids = sorted(set(int(x) for x in raw_ids))
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'req_ids must contain integers'}), 400
    note = (d.get('note') or '').strip()[:500] or None

    conn = sampling_portal.get_db_connection()
    try:
        results = []
        approved_ids = []
        for rid in req_ids:
            ok, msg = _fifo_override_approve_one(conn, rid, note)
            results.append({'req_id': rid, 'ok': ok, 'message': msg})
            if ok:
                approved_ids.append(rid)

        if approved_ids:
            try:
                _audit_record(
                    conn,
                    action='fifo_override.approve',
                    entity='fifo_override_request',
                    entity_id=None,
                    summary=f"Bulk approved {len(approved_ids)} FIFO override request(s): {approved_ids}",
                    before=None,
                    after={'approved_ids': approved_ids, 'note': note},
                )
            except Exception:
                pass

        conn.commit()
        conn.close()
        return jsonify({
            'status':       'ok',
            'approved':     len(approved_ids),
            'skipped':      len(req_ids) - len(approved_ids),
            'total':        len(req_ids),
            'per_request':  results,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/fifo_override/reject_bulk', methods=['POST'])
@_login_required
def api_fifo_override_reject_bulk():
    """Admin rejects multiple pending FIFO override requests in one call.
    Body: {req_ids: [int, ...], note: optional shared note}"""
    if not _user_has_access('fifo_override'):
        return jsonify({'status': 'error', 'message': 'You do not have access to FIFO Override approvals'}), 403
    d = request.get_json() or {}
    raw_ids = d.get('req_ids') or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return jsonify({'status': 'error', 'message': 'req_ids must be a non-empty list'}), 400
    try:
        req_ids = sorted(set(int(x) for x in raw_ids))
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'req_ids must contain integers'}), 400
    note = (d.get('note') or '').strip()[:500] or None

    conn = sampling_portal.get_db_connection()
    try:
        results = []
        rejected_ids = []
        for rid in req_ids:
            row = conn.execute(
                "SELECT status FROM pm_fifo_override_requests WHERE req_id=%s", (rid,)
            ).fetchone()
            if not row:
                results.append({'req_id': rid, 'ok': False, 'message': 'not found'}); continue
            if row['status'] != 'pending':
                results.append({'req_id': rid, 'ok': False, 'message': f"is '{row['status']}'"}); continue
            conn.execute(
                """UPDATE pm_fifo_override_requests
                   SET status='rejected', decided_by=%s, decided_at=NOW(), decided_note=%s
                   WHERE req_id=%s""",
                (_user(), note, rid)
            )
            results.append({'req_id': rid, 'ok': True, 'message': 'ok'})
            rejected_ids.append(rid)

        if rejected_ids:
            try:
                _audit_record(
                    conn,
                    action='fifo_override.reject',
                    entity='fifo_override_request',
                    entity_id=None,
                    summary=f"Bulk rejected {len(rejected_ids)} FIFO override request(s): {rejected_ids}",
                    before=None,
                    after={'rejected_ids': rejected_ids, 'note': note},
                )
            except Exception:
                pass

        conn.commit()
        conn.close()
        return jsonify({
            'status':      'ok',
            'rejected':    len(rejected_ids),
            'skipped':     len(req_ids) - len(rejected_ids),
            'total':       len(req_ids),
            'per_request': results,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/fifo_override/<int:req_id>/reject', methods=['POST'])
@_login_required
def api_fifo_override_reject(req_id):
    """Admin rejects a pending FIFO override request."""
    if not _user_has_access('fifo_override'):
        return jsonify({'status': 'error', 'message': 'You do not have access to FIFO Override approvals'}), 403
    d = request.get_json() or {}
    note = (d.get('note') or '').strip()[:500] or None
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT status, box_code, transfer_no FROM pm_fifo_override_requests WHERE req_id=%s",
            (req_id,)
        ).fetchone()
        if not row:
            conn.close(); return jsonify({'status': 'error', 'message': 'Request not found'}), 404
        if row['status'] != 'pending':
            conn.close(); return jsonify({
                'status': 'error',
                'message': f"Request is currently '{row['status']}', not pending — cannot reject."
            }), 409
        conn.execute(
            """UPDATE pm_fifo_override_requests
               SET status='rejected', decided_by=%s, decided_at=NOW(), decided_note=%s
               WHERE req_id=%s""",
            (_user(), note, req_id)
        )
        try:
            _audit_record(
                conn,
                action='fifo_override.reject',
                entity='fifo_override_request',
                entity_id=req_id,
                summary=f"FIFO override #{req_id} rejected (box {row['box_code']} on {row['transfer_no']})",
                before={'status': 'pending'},
                after={'status': 'rejected', 'note': note},
            )
        except Exception:
            pass
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'req_id': req_id})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/reprint/<int:req_id>/redeem', methods=['POST'])
@_login_required
def api_reprint_redeem(req_id):
    """Called by the requester just before firing the print. Validates the
    token, marks the request (or one box within it) as printed, returns
    the voucher payload needed by the print function.

    Body: { token, box_code? }

    Per-box mode (preferred for new requests):
      Caller passes `box_code`. The matching row in
      pm_label_reprint_box_status is marked printed. The parent request
      stays 'approved' until ALL child rows are printed, then flips to
      'printed' automatically.

    Legacy mode (fallback for old approvals with no child rows):
      Caller omits `box_code`. The whole request flips to 'printed' in
      one shot — same behaviour as before this change.
    """
    d = request.get_json() or {}
    token = (d.get('token') or '').strip()
    box_code = (d.get('box_code') or '').strip() or None
    if not token:
        return jsonify({'status': 'error', 'message': 'token required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            """SELECT * FROM pm_label_reprint_requests
               WHERE req_id=%s AND print_token=%s""",
            (req_id, token)
        ).fetchone()
        if not row:
            conn.close(); return jsonify({'status':'error','message':'Invalid request or token'}), 404
        # Only the original requester (or admin) may redeem
        if not _is_admin() and row['requested_by'] != _user():
            conn.close(); return jsonify({'status':'error','message':'Only the original requester can redeem this token'}), 403
        if row['status'] != 'approved':
            conn.close(); return jsonify({
                'status':'error',
                'message': f"Request status is '{row['status']}'. Only approved requests can be redeemed."
            }), 409

        # Check whether this request has per-box tracking. Approvals issued
        # AFTER this code shipped will have child rows; legacy approvals
        # (approved before deploy) won't — they use the old bulk path.
        child_rows = conn.execute(
            "SELECT COUNT(*) AS n FROM pm_label_reprint_box_status WHERE req_id=%s",
            (req_id,)
        ).fetchone()
        has_child_tracking = bool(child_rows and child_rows['n'] > 0)

        if has_child_tracking and box_code:
            # ── Per-box consumption path ─────────────────────────────
            target = conn.execute(
                """SELECT printed_at FROM pm_label_reprint_box_status
                   WHERE req_id=%s AND box_code=%s""",
                (req_id, box_code)
            ).fetchone()
            if not target:
                conn.close()
                return jsonify({
                    'status':'error',
                    'message': f"Box '{box_code}' is not part of this reprint approval"
                }), 404
            if target['printed_at']:
                conn.close()
                return jsonify({
                    'status':'error',
                    'message': f"Box '{box_code}' has already been printed for this request"
                }), 409
            conn.execute(
                """UPDATE pm_label_reprint_box_status
                   SET printed_at=NOW(), printed_by=%s
                   WHERE req_id=%s AND box_code=%s""",
                (_user(), req_id, box_code)
            )
            # If this was the last unprinted box, flip the parent request
            # to 'printed' and invalidate the token. Otherwise leave the
            # request 'approved' so remaining boxes can still be printed.
            remaining = conn.execute(
                """SELECT COUNT(*) AS n FROM pm_label_reprint_box_status
                   WHERE req_id=%s AND printed_at IS NULL""",
                (req_id,)
            ).fetchone()
            all_printed = (remaining and remaining['n'] == 0)
            if all_printed:
                conn.execute(
                    """UPDATE pm_label_reprint_requests
                       SET status='printed', printed_at=NOW(), printed_by=%s, print_token=NULL
                       WHERE req_id=%s""",
                    (_user(), req_id)
                )
            conn.commit()
            conn.close()
            return jsonify({
                'status': 'ok',
                'req_id': req_id,
                'box_code': box_code,
                'all_printed': bool(all_printed),
                'scope_type':    row['scope_type'],
                'voucher_kind':  row['voucher_kind'],
                'voucher_id':    row['voucher_id'],
                'voucher_label': row['voucher_label'],
                # For per-box mode return just THIS box as the CSV so the
                # frontend's existing print function can render one label.
                'box_codes_csv': box_code,
                'product_id':      row['product_id'],
                'godown_id':       row['godown_id'],
                'new_no_of_box':   row['new_no_of_box'],
                'new_per_box_qty': float(row['new_per_box_qty']) if row['new_per_box_qty'] is not None else None
            })

        # ── Legacy bulk-print path ────────────────────────────────────
        # Either pre-existing approval (no child rows) or new approval
        # with explicit no-box-code request (caller wants the whole
        # voucher). Mark the whole request printed in one shot.
        conn.execute(
            """UPDATE pm_label_reprint_requests
               SET status='printed', printed_at=NOW(), printed_by=%s, print_token=NULL
               WHERE req_id=%s""",
            (_user(), req_id)
        )
        conn.commit()
        conn.close()
        return jsonify({
            'status': 'ok',
            'req_id': req_id,
            'all_printed': True,
            'scope_type':    row['scope_type'],
            'voucher_kind':  row['voucher_kind'],
            'voucher_id':    row['voucher_id'],
            'voucher_label': row['voucher_label'],
            'box_codes_csv': row['box_codes_csv'],
            'product_id':      row['product_id'],
            'godown_id':       row['godown_id'],
            'new_no_of_box':   row['new_no_of_box'],
            'new_per_box_qty': float(row['new_per_box_qty']) if row['new_per_box_qty'] is not None else None
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/voucher_permissions', methods=['GET'])
@_login_required
def api_voucher_permissions_get():
    """Return the current enabled/disabled state of each voucher type. Open to
    all logged-in users (the frontend uses this to decide which forms to show)."""
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(
            "SELECT voucher_type, enabled FROM pm_voucher_permissions"
        ).fetchall()
        conn.close()
        # Default any missing types to True
        out = {vt: True for vt in VOUCHER_TYPES}
        for r in rows:
            vt = (r['voucher_type'] or '').strip().lower()
            if vt in VOUCHER_TYPES:
                out[vt] = bool(r['enabled'])
        return jsonify({'status': 'ok', 'permissions': out})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/voucher_permissions', methods=['POST'])
@_login_required
def api_voucher_permissions_set():
    """Admin-only: toggle one voucher type's enabled flag.
    Body: { voucher_type: 'grn'|'mtv'|'dn'|'opening', enabled: true|false }"""
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    d = request.get_json() or {}
    vt = (d.get('voucher_type') or '').strip().lower()
    en = bool(d.get('enabled'))
    if vt not in VOUCHER_TYPES:
        return jsonify({
            'status': 'error',
            'message': f"voucher_type must be one of {', '.join(VOUCHER_TYPES)}"
        }), 400
    conn = sampling_portal.get_db_connection()
    try:
        conn.execute(
            """INSERT INTO pm_voucher_permissions (voucher_type, enabled, updated_by)
                 VALUES (%s, %s, %s)
               ON DUPLICATE KEY UPDATE
                 enabled    = VALUES(enabled),
                 updated_by = VALUES(updated_by)""",
            (vt, 1 if en else 0, _user())
        )
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'voucher_type': vt, 'enabled': en})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════
# Per-user access control endpoints
# ═══════════════════════════════════════════════════════════════════════
#
# Three endpoints:
#   GET  /api/pm_stock/user_access/list    — admin: every user's flags
#   GET  /api/pm_stock/user_access/<user>  — admin: one user's flags
#   POST /api/pm_stock/user_access/save    — admin: persist a user's flags
#
# Plus a "list candidate users" endpoint:
#   GET  /api/pm_stock/user_access/users   — pulls usernames from User_Tbl
#                                            so the admin picker isn't
#                                            limited to users already in
#                                            pm_user_access.
#
# Defaults: every category is TRUE for any user without a row (table is
# additive — admin only writes rows when restricting). Admins always pass
# regardless of their row.

@pm_stock_bp.route('/api/pm_stock/user_access/users', methods=['GET'])
@_login_required
def api_pm_user_access_users():
    """Admin-only: list candidate usernames from user_tbl so the picker
    can offer every system user, not just those already in pm_user_access.

    Schema reference (user_tbl):
        username       VARCHAR(100) — login name (NOT 'User_Name'; lowercase)
        full_name      VARCHAR(200) — display name
        role           VARCHAR(50)  — application role (e.g. admin)
        user_type      VARCHAR(50)  — broad bucket (PM / QC / RnD / etc.)
        department     VARCHAR(100)
        designation    VARCHAR(100)
        is_active      TINYINT(1)   — only active users shown

    Falls back to existing pm_user_access rows if the user_tbl read fails
    for any reason, so the admin can still edit previously-configured
    users even on a malformed install.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        users = []
        # Primary fetch from User_Tbl. Notes on robustness:
        #   * Try mixed-case "User_Tbl" first (HCP portal canonical name),
        #     then fall back to lowercase "user_tbl". MySQL may or may not
        #     be case-sensitive on table names depending on the value of
        #     `lower_case_table_names` system variable on this server, so
        #     try both before giving up.
        #   * Don't hide inactive users — return is_active so the admin
        #     UI can decide whether to grey them out. Previously this
        #     filter excluded NULL or 0 is_active rows entirely, which
        #     hid legitimate users whose status flag had never been set.
        #   * COALESCE on role / user_type / department keeps the API
        #     shape consistent even when those columns are NULL.
        last_err = None
        for table_name in ('User_Tbl', 'user_tbl'):
            try:
                rows = conn.execute(f"""
                    SELECT username                       AS user_name,
                           COALESCE(full_name, username)  AS display_name,
                           COALESCE(`role`, '')           AS role,
                           COALESCE(user_type, '')        AS user_type,
                           COALESCE(department, '')       AS department,
                           COALESCE(is_active, 1)         AS is_active
                    FROM {table_name}
                    WHERE username IS NOT NULL AND username <> ''
                    ORDER BY COALESCE(is_active, 1) DESC, username
                """).fetchall()
                users = [dict(r) for r in rows]
                # First successful read wins — stop trying alternatives.
                break
            except Exception as _user_tbl_err:
                last_err = _user_tbl_err
                continue
        if not users and last_err is not None:
            import sys as _sys
            print(f"[pm_stock] user_tbl read failed for all table-name variants: {last_err}",
                  file=_sys.stderr)

        # Always merge in users who already have a pm_user_access row so
        # the admin can edit them even if the user_tbl read above came
        # back empty for any reason.
        existing = []
        try:
            ex_rows = conn.execute(
                "SELECT user_name FROM pm_user_access ORDER BY user_name"
            ).fetchall()
            existing = [r['user_name'] for r in ex_rows]
        except Exception:
            pass

        seen = set()
        merged = []
        for u in users:
            if u['user_name'] not in seen:
                seen.add(u['user_name'])
                # Normalize is_active to bool for JSON cleanliness
                u['is_active'] = bool(int(u.get('is_active') or 0))
                merged.append(u)
        for un in existing:
            if un not in seen:
                seen.add(un)
                # Stub entry — no role/dept available from user_tbl, but
                # they have a pm_user_access row so the admin definitely
                # wants to manage them. Mark is_active=True optimistically
                # so they show up at the top of the list.
                merged.append({
                    'user_name':    un,
                    'display_name': un,
                    'role':         '',
                    'user_type':    '',
                    'department':   '',
                    'is_active':    True,
                })

        conn.close()
        return jsonify({
            'status': 'ok',
            'users':  merged,
            # Counts so the admin UI can show "14 from user_tbl + 0 stubs = 14 total"
            # — handy for debugging "I expected more users" scenarios.
            'meta': {
                'from_user_tbl': len(users),
                'stubs_added':   len(merged) - len(users),
                'total':         len(merged),
            },
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/user_access/list', methods=['GET'])
@_login_required
def api_pm_user_access_list():
    """Admin-only: return every row from pm_user_access plus the canonical
    list of category keys so the frontend can render a stable column set
    even if the DB row only has some columns filled in."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        cols = ", ".join(PM_USER_ACCESS_KEYS)
        rows = conn.execute(
            f"""SELECT user_name, {cols},
                       updated_at, COALESCE(updated_by,'') AS updated_by
                FROM pm_user_access
                ORDER BY user_name"""
        ).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            if d.get('updated_at') is not None: d['updated_at'] = str(d['updated_at'])
            for k in PM_USER_ACCESS_KEYS:
                # Cast TINYINT → bool for clean JSON
                v = d.get(k)
                d[k] = bool(int(v)) if v is not None else True
            out.append(d)
        return jsonify({'status':'ok','keys':list(PM_USER_ACCESS_KEYS),'rows':out})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/user_access/<path:user_name>', methods=['GET'])
@_login_required
def api_pm_user_access_get(user_name):
    """Admin-only: fetch one user's access dict (defaults all-True when
    no row exists). Used to populate the admin modal when a user is
    selected from the picker."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    user_name = (user_name or '').strip()
    if not user_name:
        return jsonify({'status':'error','message':'user_name required'}), 400
    try:
        access = _user_access_dict(user=user_name)
        return jsonify({'status':'ok','user_name':user_name,'access':access})
    except Exception as e:
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/user_access/save', methods=['POST'])
@_login_required
def api_pm_user_access_save():
    """Admin-only: persist a user's access flags. Body shape:
        { user_name: 'bhavesh', access: { voucher_log: true, ... } }
    Unknown keys are ignored. Missing keys are NOT touched in DB — only
    the keys present in the payload get written. This makes it safe for
    the frontend to send a partial update (e.g. one toggle changed)."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    d = request.get_json() or {}
    user_name = (d.get('user_name') or '').strip()
    access    = d.get('access') or {}
    if not user_name:
        return jsonify({'status':'error','message':'user_name required'}), 400
    if not isinstance(access, dict):
        return jsonify({'status':'error','message':'access must be a dict'}), 400

    # Filter to known keys only — silently drop unknown ones rather than
    # erroring, so the frontend can be loose with what it sends.
    clean = {k: bool(access[k]) for k in PM_USER_ACCESS_KEYS if k in access}
    if not clean:
        return jsonify({'status':'error','message':'No valid keys in access dict'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # UPSERT pattern. INSERT supplies defaults (TRUE for any key not
        # in `clean`); UPDATE only overwrites the keys present in `clean`.
        all_cols   = list(PM_USER_ACCESS_KEYS)
        insert_cols = ['user_name'] + all_cols + ['updated_by']
        insert_vals = [user_name] + [1 if clean.get(k, True) else 0 for k in all_cols] + [_user()]
        # ON DUPLICATE KEY: only overwrite the keys that were actually sent
        update_clauses = [f"{k}=VALUES({k})" for k in clean.keys()]
        update_clauses.append("updated_by=VALUES(updated_by)")
        sql = (
            f"INSERT INTO pm_user_access ({', '.join(insert_cols)}) "
            f"VALUES ({', '.join(['%s']*len(insert_vals))}) "
            f"ON DUPLICATE KEY UPDATE {', '.join(update_clauses)}"
        )
        conn.execute(sql, insert_vals)
        conn.commit()
        # Return the resulting dict so frontend can update its local cache
        new_access = _user_access_dict(user=user_name, conn=conn)
        conn.close()
        return jsonify({'status':'ok','user_name':user_name,'access':new_access})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/user_access/delete', methods=['POST'])
@_login_required
def api_pm_user_access_delete():
    """Admin-only: remove a user's pm_user_access row entirely, reverting
    them to the default ALL-TRUE state. Useful when admin wants to undo
    earlier restrictions wholesale rather than toggle each flag back on."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    d = request.get_json() or {}
    user_name = (d.get('user_name') or '').strip()
    if not user_name:
        return jsonify({'status':'error','message':'user_name required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        conn.execute("DELETE FROM pm_user_access WHERE user_name=%s", (user_name,))
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','user_name':user_name,'reset':True})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════
#  ACCESS GROUPS  —  admin defines a group, sets feature access, assigns
#  users. Members inherit the group's access (unless they have an explicit
#  per-user pm_user_access row, which still wins). All admin-only.
# ═══════════════════════════════════════════════════════════════════════
@pm_stock_bp.route('/api/pm_stock/access_groups', methods=['GET'])
@_login_required
def api_access_groups_list():
    """List all access groups with their feature flags + member counts."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        cols = ", ".join(PM_USER_ACCESS_KEYS)
        groups = conn.execute(f"""
            SELECT group_id, group_name, note, {cols},
                   (SELECT COUNT(*) FROM pm_access_group_members m
                      WHERE m.group_id = g.group_id) AS member_count
            FROM pm_access_groups g
            ORDER BY group_name
        """).fetchall()
        out = []
        for g in groups:
            d = dict(g)
            d['access'] = {k: bool(int(d.pop(k))) for k in PM_USER_ACCESS_KEYS}
            out.append(d)
        conn.close()
        return jsonify({'status':'ok','keys':list(PM_USER_ACCESS_KEYS),'groups':out})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/access_groups/save', methods=['POST'])
@_login_required
def api_access_groups_save():
    """Create or update a group. Body: { group_id?:int, group_name:str,
    note?:str, access:{feature:bool,...} }. Omitting group_id creates a new
    group; including it updates that group's name/note/flags."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    d = request.get_json() or {}
    name = (d.get('group_name') or '').strip()
    if not name:
        return jsonify({'status':'error','message':'group_name required'}), 400
    access = d.get('access') or {}
    clean = {k: (1 if bool(access.get(k)) else 0) for k in PM_USER_ACCESS_KEYS}
    note = (d.get('note') or '').strip()[:300] or None
    gid = d.get('group_id')
    conn = sampling_portal.get_db_connection()
    try:
        if gid:
            set_cols = ", ".join(f"{k}=%s" for k in PM_USER_ACCESS_KEYS)
            conn.execute(
                f"""UPDATE pm_access_groups
                    SET group_name=%s, note=%s, {set_cols},
                        updated_by=%s, updated_at=NOW()
                    WHERE group_id=%s""",
                tuple([name, note] + [clean[k] for k in PM_USER_ACCESS_KEYS] + [_user(), int(gid)])
            )
            new_id = int(gid)
            action = 'access_group.update'
        else:
            ins_cols = ", ".join(PM_USER_ACCESS_KEYS)
            ph = ", ".join(["%s"] * len(PM_USER_ACCESS_KEYS))
            cur = conn.execute(
                f"""INSERT INTO pm_access_groups
                      (group_name, note, {ins_cols}, created_by)
                    VALUES (%s, %s, {ph}, %s)""",
                tuple([name, note] + [clean[k] for k in PM_USER_ACCESS_KEYS] + [_user()])
            )
            new_id = cur.lastrowid
            action = 'access_group.create'
        try:
            _audit_record(conn, action=action, entity='access_group', entity_id=new_id,
                          summary=f"Access group '{name}' saved", before=None,
                          after={'group_name': name, 'access': {k: bool(clean[k]) for k in PM_USER_ACCESS_KEYS}})
        except Exception:
            pass
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','group_id':new_id})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        msg = str(e)
        if 'uq_pm_group_name' in msg or 'Duplicate' in msg:
            msg = f"A group named '{name}' already exists."
        return jsonify({'status':'error','message':msg}), 500


@pm_stock_bp.route('/api/pm_stock/access_groups/<int:group_id>', methods=['DELETE'])
@_login_required
def api_access_groups_delete(group_id):
    """Delete a group and unassign its members (they revert to defaults)."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute("SELECT group_name FROM pm_access_groups WHERE group_id=%s", (group_id,)).fetchone()
        if not row:
            conn.close(); return jsonify({'status':'error','message':'Group not found'}), 404
        conn.execute("DELETE FROM pm_access_group_members WHERE group_id=%s", (group_id,))
        conn.execute("DELETE FROM pm_access_groups WHERE group_id=%s", (group_id,))
        try:
            _audit_record(conn, action='access_group.delete', entity='access_group',
                          entity_id=group_id, summary=f"Access group '{row['group_name']}' deleted",
                          before={'group_name': row['group_name']}, after=None)
        except Exception:
            pass
        conn.commit()
        conn.close()
        return jsonify({'status':'ok'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/access_groups/<int:group_id>/members', methods=['GET'])
@_login_required
def api_access_groups_members(group_id):
    """List the usernames assigned to a group."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(
            "SELECT user_name, assigned_by, assigned_at FROM pm_access_group_members WHERE group_id=%s ORDER BY user_name",
            (group_id,)
        ).fetchall()
        conn.close()
        out = []
        for r in rows:
            dd = dict(r)
            if dd.get('assigned_at') is not None: dd['assigned_at'] = str(dd['assigned_at'])
            out.append(dd)
        return jsonify({'status':'ok','members':out})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/access_groups/assign', methods=['POST'])
@_login_required
def api_access_groups_assign():
    """Assign a user to a group (one group per user — upsert). Body:
    { user_name:str, group_id:int }. Passing group_id=0/null unassigns."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    d = request.get_json() or {}
    uname = (d.get('user_name') or '').strip()
    if not uname:
        return jsonify({'status':'error','message':'user_name required'}), 400
    gid = d.get('group_id')
    conn = sampling_portal.get_db_connection()
    try:
        if not gid:        # unassign
            conn.execute("DELETE FROM pm_access_group_members WHERE user_name=%s", (uname,))
            action, summary = 'access_group.unassign', f"{uname} removed from group"
        else:
            gid = int(gid)
            grp = conn.execute("SELECT group_name FROM pm_access_groups WHERE group_id=%s", (gid,)).fetchone()
            if not grp:
                conn.close(); return jsonify({'status':'error','message':'Group not found'}), 404
            conn.execute(
                """INSERT INTO pm_access_group_members (user_name, group_id, assigned_by)
                   VALUES (%s,%s,%s)
                   ON DUPLICATE KEY UPDATE group_id=VALUES(group_id),
                                           assigned_by=VALUES(assigned_by), assigned_at=NOW()""",
                (uname, gid, _user())
            )
            action, summary = 'access_group.assign', f"{uname} assigned to '{grp['group_name']}'"
        try:
            _audit_record(conn, action=action, entity='access_group_member', entity_id=uname,
                          summary=summary, before=None, after={'group_id': gid})
        except Exception:
            pass
        conn.commit()
        conn.close()
        return jsonify({'status':'ok'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════
# Command Palette (Ctrl+K) endpoints
# ═══════════════════════════════════════════════════════════════════════
# Three endpoints back the palette UI:
#   GET  /api/pm_stock/palette/recent  — top-N most-used actions for the
#                                         current user (default 6)
#   POST /api/pm_stock/palette/track   — UPSERT a hit on (user, action_id)
#   GET  /api/pm_stock/palette/search  — fuzzy data lookup across vouchers
#                                         and products (returns up to 8
#                                         results per category)
# All endpoints are per-user-scoped — recent lookups never leak between
# users. Tracking is fire-and-forget; failures are ignored on the client
# so a missing table doesn't break the palette.

# ── Dock pins (per-user) ───────────────────────────────────────────────
# The floating dock shows usage-ranked actions by default, but the user can
# pin/unpin/reorder. We persist the user's pinned action_id list (ordered)
# as a JSON array in pm_dock_pins. Empty/absent = pure auto mode.
@pm_stock_bp.route('/api/pm_stock/dock/pins', methods=['GET'])
@_login_required
def api_dock_pins_get():
    user = _user()
    if not user or user == 'Unknown':
        return jsonify({'status':'ok','pins':[]})
    conn = sampling_portal.get_db_connection()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS pm_dock_pins (
                user_name VARCHAR(100) PRIMARY KEY,
                pins      TEXT,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        row = conn.execute("SELECT pins FROM pm_dock_pins WHERE user_name=%s", (user,)).fetchone()
        conn.close()
        import json as _json
        pins = []
        if row and row['pins']:
            try: pins = _json.loads(row['pins'])
            except Exception: pins = []
        return jsonify({'status':'ok','pins':pins})
    except Exception:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'ok','pins':[]})


@pm_stock_bp.route('/api/pm_stock/dock/pins', methods=['POST'])
@_login_required
def api_dock_pins_save():
    user = _user()
    if not user or user == 'Unknown':
        return jsonify({'status':'ok'})
    d = request.get_json() or {}
    pins = d.get('pins') or []
    if not isinstance(pins, list):
        return jsonify({'status':'error','message':'pins must be a list'}), 400
    # sanitise: strings only, cap length + count
    pins = [str(p).strip()[:80] for p in pins if str(p).strip()][:30]
    import json as _json
    conn = sampling_portal.get_db_connection()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS pm_dock_pins (
                user_name VARCHAR(100) PRIMARY KEY,
                pins      TEXT,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.execute(
            """INSERT INTO pm_dock_pins (user_name, pins) VALUES (%s, %s)
               ON DUPLICATE KEY UPDATE pins=VALUES(pins), updated_at=NOW()""",
            (user, _json.dumps(pins))
        )
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','pins':pins})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/palette/recent', methods=['GET'])
@_login_required
def api_palette_recent():
    """Return the current user's most-recently-used palette action IDs.
    Ranking blends recency (last_used_at) and frequency (hit_count) —
    we order by a simple combined score so a heavy hitter from yesterday
    still beats a one-off click this morning, but only just."""
    n = max(1, min(20, int(request.args.get('n') or 6)))
    user = _user()
    if not user or user == 'Unknown':
        return jsonify({'status':'ok','actions':[]})
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(
            """SELECT action_id, hit_count, last_used_at
               FROM pm_palette_usage
               WHERE user_name=%s
               ORDER BY last_used_at DESC, hit_count DESC
               LIMIT %s""",
            (user, n)
        ).fetchall()
        out = []
        for r in rows:
            out.append({
                'action_id': r['action_id'],
                'hit_count': int(r['hit_count'] or 0),
                'last_used_at': str(r['last_used_at']) if r.get('last_used_at') else None,
            })
        conn.close()
        return jsonify({'status':'ok','actions':out})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        # Fail-open so a missing table or bad row doesn't break the palette.
        return jsonify({'status':'ok','actions':[]})


@pm_stock_bp.route('/api/pm_stock/palette/track', methods=['POST'])
@_login_required
def api_palette_track():
    """Record that the current user invoked a palette action. UPSERT
    increments hit_count and bumps last_used_at via ON DUPLICATE KEY.
    Returns ok regardless — the palette doesn't care about failures."""
    user = _user()
    if not user or user == 'Unknown':
        return jsonify({'status':'ok'})
    d = request.get_json() or {}
    action_id = (d.get('action_id') or '').strip()[:80]
    if not action_id:
        return jsonify({'status':'error','message':'action_id required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        conn.execute(
            """INSERT INTO pm_palette_usage (user_name, action_id, hit_count)
               VALUES (%s, %s, 1)
               ON DUPLICATE KEY UPDATE
                 hit_count = hit_count + 1,
                 last_used_at = CURRENT_TIMESTAMP""",
            (user, action_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'status':'ok'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        # Tracking failures are non-fatal.
        return jsonify({'status':'ok'})


@pm_stock_bp.route('/api/pm_stock/palette/search', methods=['GET'])
@_login_required
def api_palette_search():
    """Fuzzy data lookup for the palette. Returns up to 8 voucher hits
    and 8 product hits matching the query string. Voucher search hits
    grn, dn, mtv, and material-request numbers in one pass. Product
    search hits name and product_code.

    The frontend treats each hit as a "data action" with a click handler
    that routes to the relevant detail view."""
    q = (request.args.get('q') or '').strip()
    if not q or len(q) < 2:
        return jsonify({'status':'ok','vouchers':[],'products':[]})
    like = f"%{q}%"
    conn = sampling_portal.get_db_connection()
    try:
        # Vouchers — union across GRN, DN, MTV, Material Request.
        # Each row carries `kind` so the frontend can route correctly.
        vouchers = []
        try:
            grn_rows = conn.execute(
                """SELECT 'grn' AS kind, id, grn_no AS voucher_no,
                          supplier AS detail1, grn_date AS detail2
                   FROM pm_grn
                   WHERE grn_no LIKE %s OR supplier LIKE %s
                   ORDER BY id DESC LIMIT 5""",
                (like, like)
            ).fetchall()
            vouchers.extend(dict(r) for r in grn_rows)
        except Exception: pass
        try:
            dn_rows = conn.execute(
                """SELECT 'dn' AS kind, id, dn_no AS voucher_no,
                          supplier AS detail1, dn_date AS detail2
                   FROM pm_dn
                   WHERE dn_no LIKE %s OR supplier LIKE %s
                   ORDER BY id DESC LIMIT 5""",
                (like, like)
            ).fetchall()
            vouchers.extend(dict(r) for r in dn_rows)
        except Exception: pass
        try:
            mtv_rows = conn.execute(
                """SELECT 'mtv' AS kind, transfer_id AS id, transfer_no AS voucher_no,
                          status AS detail1, out_at AS detail2
                   FROM pm_transfers
                   WHERE transfer_no LIKE %s
                   ORDER BY transfer_id DESC LIMIT 5""",
                (like,)
            ).fetchall()
            vouchers.extend(dict(r) for r in mtv_rows)
        except Exception: pass
        try:
            mr_rows = conn.execute(
                """SELECT 'mr' AS kind, id, request_no AS voucher_no,
                          requested_by AS detail1, request_date AS detail2
                   FROM pm_material_requests
                   WHERE request_no LIKE %s OR requested_by LIKE %s
                   ORDER BY id DESC LIMIT 5""",
                (like, like)
            ).fetchall()
            vouchers.extend(dict(r) for r in mr_rows)
        except Exception: pass

        # Sort vouchers by simple length-of-match-prefix preference, then
        # voucher_no descending so newer ones come first when prefix-tied.
        ql = q.lower()
        def _vscore(v):
            vn = (v.get('voucher_no') or '').lower()
            if vn.startswith(ql):     return 0
            if ql in vn:              return 1
            return 2
        vouchers.sort(key=lambda v: (_vscore(v), -(v.get('id') or 0)))
        vouchers = vouchers[:8]
        # Stringify dates so JSON is happy
        for v in vouchers:
            for k in ('detail2',):
                if hasattr(v.get(k), 'isoformat'):
                    v[k] = v[k].isoformat() if k != 'detail2' else str(v[k])

        # Products — match name OR code
        products = []
        try:
            p_rows = conn.execute(
                """SELECT id, product_name, product_code, pm_type
                   FROM pm_products
                   WHERE product_name LIKE %s OR product_code LIKE %s
                   ORDER BY
                     CASE
                       WHEN product_code LIKE %s THEN 0
                       WHEN product_name LIKE %s THEN 1
                       ELSE 2
                     END,
                     product_name
                   LIMIT 8""",
                (like, like, f"{q}%", f"{q}%")
            ).fetchall()
            products = [dict(r) for r in p_rows]
        except Exception: pass

        conn.close()
        return jsonify({'status':'ok','vouchers':vouchers,'products':products})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/user_home', methods=['GET'])
@_login_required
def api_pm_user_home():
    """Returns the current user's home godown info + admin flag.
    Frontend uses this to decide whether to lock location fields."""
    conn = sampling_portal.get_db_connection()
    try:
        is_admin = _is_admin()
        user = _user()
        home_id = None
        home_name = None
        if not is_admin and user and user != 'Unknown':
            row = conn.execute(
                """SELECT h.godown_id, COALESCE(g.name,'') AS name
                   FROM pm_user_home_godown h
                   LEFT JOIN procurement_godowns g ON g.id=h.godown_id
                   WHERE h.user_name=%s LIMIT 1""",
                (user,)
            ).fetchone()
            if row:
                home_id   = int(row['godown_id'])
                home_name = row['name'] or ''
        conn.close()
        return jsonify({
            'status':       'ok',
            'is_admin':     is_admin,
            'user_name':    user,
            'home_godown_id':   home_id,
            'home_godown_name': home_name,
        })
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/user_home/list', methods=['GET'])
@_login_required
def api_pm_user_home_list():
    """Admin-only: list all user→godown mappings."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin access required'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute("""
            SELECT h.user_name, h.godown_id,
                   COALESCE(g.name,'') AS godown_name,
                   COALESCE(h.note,'') AS note,
                   h.updated_at, COALESCE(h.updated_by,'') AS updated_by
            FROM pm_user_home_godown h
            LEFT JOIN procurement_godowns g ON g.id=h.godown_id
            ORDER BY g.name, h.user_name
        """).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            if d.get('updated_at') is not None: d['updated_at'] = str(d['updated_at'])
            out.append(d)
        return jsonify({'status':'ok','mappings':out})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/user_home/set', methods=['POST'])
@_login_required
def api_pm_user_home_set():
    """Admin-only: assign or update a user's home godown.
    Body: {user_name, godown_id, note?}"""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin access required'}), 403
    d = request.get_json() or {}
    user_name = (d.get('user_name') or '').strip()
    godown_id = d.get('godown_id')
    note      = (d.get('note') or '').strip() or None
    if not user_name:
        return jsonify({'status':'error','message':'user_name required'}), 400
    if godown_id is None:
        return jsonify({'status':'error','message':'godown_id required'}), 400
    try:
        godown_id = int(godown_id)
    except Exception:
        return jsonify({'status':'error','message':'godown_id must be an integer'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # Validate godown exists
        gd = conn.execute("SELECT id, name FROM procurement_godowns WHERE id=%s", (godown_id,)).fetchone()
        if not gd:
            conn.close()
            return jsonify({'status':'error','message':'Godown not found'}), 404

        # Snapshot before
        prev = conn.execute(
            "SELECT godown_id, note FROM pm_user_home_godown WHERE user_name=%s", (user_name,)
        ).fetchone()
        before_state = dict(prev) if prev else None

        conn.execute("""
            INSERT INTO pm_user_home_godown (user_name, godown_id, note, updated_by)
            VALUES (%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
                godown_id=VALUES(godown_id),
                note=VALUES(note),
                updated_by=VALUES(updated_by)
        """, (user_name, godown_id, note, _user()))

        try:
            _audit_record(
                conn,
                action='user_home.set',
                entity='user_home',
                entity_id=user_name,
                summary=f"{'Created' if before_state is None else 'Updated'} user-home: {user_name} → godown #{godown_id} ({gd['name']})",
                before=before_state,
                after={'godown_id': godown_id, 'note': note},
            )
        except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status':'ok','user_name':user_name,'godown_id':godown_id,'godown_name':gd['name']})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/user_home/delete', methods=['POST'])
@_login_required
def api_pm_user_home_delete():
    """Admin-only: remove a user's home godown lock.
    Body: {user_name}"""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin access required'}), 403
    d = request.get_json() or {}
    user_name = (d.get('user_name') or '').strip()
    if not user_name:
        return jsonify({'status':'error','message':'user_name required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        # Snapshot before-state for reversal
        prev = conn.execute(
            "SELECT godown_id, note FROM pm_user_home_godown WHERE user_name=%s", (user_name,)
        ).fetchone()
        before_state = dict(prev) if prev else None

        conn.execute("DELETE FROM pm_user_home_godown WHERE user_name=%s", (user_name,))

        if before_state:
            try:
                _audit_record(
                    conn,
                    action='user_home.delete',
                    entity='user_home',
                    entity_id=user_name,
                    summary=f"Removed home-godown lock for {user_name} (was godown #{before_state.get('godown_id')})",
                    before=before_state,
                    after=None,
                )
            except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status':'ok','user_name':user_name})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/user_home/bulk_set', methods=['POST'])
@_login_required
def api_pm_user_home_bulk_set():
    """Admin-only: assign multiple user→godown mappings in one request.

    Body: {
        mappings: [
            {user_name: str, godown_id: int|null, note: str (optional)}, ...
        ]
    }

    Behavior:
        - godown_id is a positive int → upsert (assign or update)
        - godown_id is null/0 → delete the user's mapping (clear lock)
        - Empty user_name rows are skipped silently

    Returns counts of inserted/updated and deleted mappings.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin access required'}), 403
    d = request.get_json() or {}
    mappings = d.get('mappings') or []
    if not isinstance(mappings, list):
        return jsonify({'status':'error','message':'mappings must be a list'}), 400
    if not mappings:
        return jsonify({'status':'error','message':'At least one mapping required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # Validate all godown ids in one shot — pulled to a set so we
        # only need a single query regardless of how many rows
        ids = sorted({int(m['godown_id']) for m in mappings
                      if m.get('godown_id') not in (None,'',0,'0')})
        valid_ids = set()
        if ids:
            placeholders = ','.join(['%s'] * len(ids))
            rows = conn.execute(
                f"SELECT id FROM procurement_godowns WHERE id IN ({placeholders})",
                tuple(ids)
            ).fetchall()
            valid_ids = {int(r['id']) for r in rows}
            missing = [i for i in ids if i not in valid_ids]
            if missing:
                conn.close()
                return jsonify({'status':'error',
                                'message':f'Unknown godown id(s): {missing}'}), 400

        upserted, deleted, skipped = 0, 0, 0
        for m in mappings:
            user_name = (m.get('user_name') or '').strip()
            if not user_name:
                skipped += 1
                continue
            note      = (m.get('note') or '').strip() or None
            godown_id = m.get('godown_id')
            # Empty / 0 / null godown_id → delete the lock
            if godown_id in (None, '', 0, '0'):
                conn.execute("DELETE FROM pm_user_home_godown WHERE user_name=%s",
                             (user_name,))
                deleted += 1
                continue
            try:
                godown_id = int(godown_id)
            except Exception:
                skipped += 1; continue
            conn.execute("""
                INSERT INTO pm_user_home_godown (user_name, godown_id, note, updated_by)
                VALUES (%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE
                    godown_id=VALUES(godown_id),
                    note=VALUES(note),
                    updated_by=VALUES(updated_by)
            """, (user_name, godown_id, note, _user()))
            upserted += 1
        conn.commit()
        conn.close()
        return jsonify({'status':'ok',
                        'upserted':upserted,
                        'deleted':deleted,
                        'skipped':skipped})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/admin_passcheck', methods=['POST'])
@_login_required
def api_admin_passcheck():
    """Re-verify an ADMIN user's own login password for a sensitive action.

    Used to gate the "Auto-Verify by Excel" shortcut on the GRN Verify modal
    (and any other future admin-only quick-action that should require a
    fresh password confirmation regardless of the active session).

    Body:
        password: str — the password to check, plain text from a password input.
        admin_username: str (optional) — if provided, check against THAT user
                        instead of the session user. Lets a non-admin operator
                        ask a supervisor to type credentials on their machine
                        without logging out. The target user MUST be admin.

    Response on success: {'status':'ok'}
    Response on failure: {'status':'error','message': <reason>}, 401.

    Security notes:
      - The endpoint never returns the password, never logs it, never echoes it.
      - It looks the password up via the same User_Tbl the rest of the app
        uses; it supports plain-text storage (parent HCP portal currently uses
        plaintext per existing app patterns) and also tries common hash
        comparisons defensively in case storage changes later.
      - To prevent brute force, repeated failures from the same session are
        rate-limited in-process (5 attempts per 60 seconds per session-name).
    """
    import time as _time
    d = request.get_json(silent=True) or {}
    pwd = d.get('password') or ''
    target_user = (d.get('admin_username') or '').strip() or None

    # Whichever user we're checking, the request itself must come from a
    # logged-in session (the @_login_required decorator already guarantees
    # this). We use the session name purely as a rate-limit key.
    session_name = (_user() or 'Unknown').strip()

    # ── Rate limit (per-process, per-session) ────────────────────────────
    now = _time.time()
    bucket = _ADMIN_PASSCHECK_BUCKET.setdefault(session_name, [])
    # Drop attempts older than the window
    cutoff = now - 60.0
    bucket[:] = [t for t in bucket if t >= cutoff]
    if len(bucket) >= 5:
        return jsonify({
            'status':'error',
            'message':'Too many attempts. Wait a minute and try again.'
        }), 429

    if not pwd:
        bucket.append(now)
        return jsonify({'status':'error','message':'Password required'}), 400

    # ── Look up the canonical 'admin' user and verify password ───────────
    # Per Tarak: the system has a single canonical admin user whose
    # username is literally 'admin'. We look that row up directly and
    # compare passwords. The User_Tbl column is `password_hash` (varchar
    # 256) and on this install stores a SHA-256 hex digest of the password.
    # We still try plaintext + the other common hash schemes defensively
    # in case storage ever changes.
    conn = sampling_portal.get_db_connection()
    try:
        row = None
        last_err = None
        for table_name in ('User_Tbl', 'user_tbl'):
            try:
                row = conn.execute(f"""
                    SELECT username, COALESCE(password_hash,'') AS password_hash,
                           COALESCE(is_active, 1) AS is_active
                      FROM {table_name}
                     WHERE LOWER(username) = 'admin'
                     LIMIT 1
                """).fetchone()
                if row:
                    break
            except Exception as _e:
                last_err = _e
                continue
        conn.close()

        if not row:
            bucket.append(now)
            return jsonify({'status':'error','message':"User 'admin' not found"}), 401
        if int(row.get('is_active') or 1) == 0:
            bucket.append(now)
            return jsonify({'status':'error','message':'Admin account disabled'}), 401

        stored = (row.get('password_hash') or '').strip()
        if not stored:
            bucket.append(now)
            return jsonify({'status':'error','message':'Admin has no password set'}), 401

        # 1) SHA-256 hex (this install — confirmed by inspecting User_Tbl)
        # 2) Other hash schemes + plaintext as defensive fallbacks
        ok = False
        try:
            import hashlib as _hashlib
            stored_lower = stored.lower()
            uname = (row.get('username') or 'admin')
            # Plain hashes
            for algo in ('sha256','md5','sha1','sha512'):
                if _hashlib.new(algo, pwd.encode('utf-8')).hexdigest() == stored_lower:
                    ok = True
                    break
            # Common salted patterns: username+password, password+username
            if not ok:
                for salt_combo in (uname + pwd, pwd + uname, uname.lower() + pwd, pwd + uname.lower()):
                    if _hashlib.sha256(salt_combo.encode('utf-8')).hexdigest() == stored_lower:
                        ok = True
                        break
        except Exception:
            pass
        if not ok and pwd == stored:
            ok = True   # plaintext fallback

        if not ok:
            bucket.append(now)
            return jsonify({'status':'error','message':'Incorrect password'}), 401

        # Success — clear this session's bucket so a future legitimate use
        # isn't penalised by historical failed attempts.
        _ADMIN_PASSCHECK_BUCKET[session_name] = []
        return jsonify({'status':'ok','admin_username': row.get('username') or 'admin'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        bucket.append(now)
        return jsonify({'status':'error','message':str(e)}), 500


# In-process rate-limit bucket for admin passcheck. Process-local; resets on
# Flask restart. Not a substitute for proper auth lockout but it stops casual
# brute-forcing of a 4-digit-style password in the same session window.
_ADMIN_PASSCHECK_BUCKET = {}


@pm_stock_bp.route('/api/pm_stock/user_directory', methods=['GET'])
@_login_required
def api_pm_user_directory():
    """Admin-only: return the list of users that the modal can map to home godowns.

    Pulls strictly from a user master table in the database. We auto-detect
    which table holds the user accounts by probing INFORMATION_SCHEMA for
    common column-name signatures (tables containing both a username column
    and either a password or role column). Whichever candidate table is
    found first is treated as the canonical user list. Each row also gets
    its current home-godown mapping (if any) joined in.

    Falls back to mining historical actor names from PM tables ONLY if no
    user master can be located — that fallback is intentionally minimal so
    the modal always shows something actionable.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin access required'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        # ── Locate the user master table ─────────────────────────────
        # Scan INFORMATION_SCHEMA for a table that looks like a user list:
        # has a "user_name"-style column AND a password/role/email column.
        # We prefer tables with the most signal columns (password > role >
        # email) so that an audit-only table like pm_voucher_audit doesn't
        # accidentally win over the real user master.
        candidate_user_cols = ['user_name','username','login','login_id','user_id','uid','user']
        candidate_signal_cols = ['password','passwd','password_hash','role','user_type','user_role','email','emailid']
        # Skip our own bookkeeping tables — they aren't user masters
        skip_tables = {
            'pm_user_home_godown','pm_voucher_audit','pm_label_reprint_requests',
            'pm_grn','pm_grn_items','pm_transfers','pm_transfer_items',
            'pm_box_movements','pm_boxes','pm_recycle_bin','pm_voucher_permissions',
            'pm_godown_txn','pm_floor_txn','pm_dn','pm_dn_items','pm_voucher_sequences',
            'pm_products','pm_mtv','pm_mtv_items',
        }
        # Build a list of (table, user_col) candidates
        table_rows = conn.execute("""
            SELECT TABLE_NAME, COLUMN_NAME
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
        """).fetchall()
        cols_by_table = {}
        for r in table_rows:
            t = (r['TABLE_NAME'] or '').lower()
            c = (r['COLUMN_NAME'] or '').lower()
            if t in skip_tables: continue
            if t.startswith('pm_'): continue   # skip our own pm_* tables
            cols_by_table.setdefault(t, set()).add(c)

        best = None  # (score, table, user_col)
        for t, cols in cols_by_table.items():
            user_col = next((u for u in candidate_user_cols if u in cols), None)
            if not user_col: continue
            score = sum(1 for s in candidate_signal_cols if s in cols)
            if score == 0: continue   # no auth-like signal → not a user table
            # Boost obviously named user tables
            name_boost = 0
            for hint in ('user','login','employee','staff','account'):
                if hint in t: name_boost += 1
            score += name_boost
            if not best or score > best[0]:
                best = (score, t, user_col, cols)

        users_from_master = []
        master_table_name = None
        if best:
            _score, master_table_name, user_col, cols = best
            # Pull a friendly display name and (optionally) role/active flag
            name_col = next((c for c in ('full_name','name','display_name','first_name')
                             if c in cols), None)
            role_col = next((c for c in ('user_type','role','user_role') if c in cols), None)
            active_col = next((c for c in ('active','is_active','status','enabled') if c in cols), None)

            # Build SELECT safely with identifier whitelisting (only
            # columns we just verified to exist).
            select_parts = [f"`{user_col}` AS user_name"]
            select_parts.append(f"`{name_col}` AS display_name" if name_col else "'' AS display_name")
            select_parts.append(f"`{role_col}` AS user_role"   if role_col else "'' AS user_role")
            if active_col:
                # Try to filter inactive users — heuristic: active='Y'/1/active/true
                where_clause = (f"WHERE (`{active_col}` IS NULL "
                                f"OR LOWER(CAST(`{active_col}` AS CHAR)) IN ('y','yes','1','true','active','enabled'))")
            else:
                where_clause = ""
            sql = (f"SELECT {', '.join(select_parts)} FROM `{master_table_name}` "
                   f"{where_clause} "
                   f"GROUP BY `{user_col}` "
                   f"ORDER BY `{user_col}`")
            try:
                master_rows = conn.execute(sql).fetchall()
                for r in master_rows:
                    nm = (r.get('user_name') or '').strip()
                    if not nm: continue
                    users_from_master.append({
                        'user_name':    nm,
                        'display_name': (r.get('display_name') or '').strip(),
                        'user_role':    (r.get('user_role') or '').strip(),
                    })
            except Exception:
                # If the auto-detected table doesn't query cleanly, drop back
                users_from_master = []
                master_table_name = None

        # ── Pull home-godown mappings ────────────────────────────────
        home_rows = conn.execute("""
            SELECT h.user_name, h.godown_id, COALESCE(g.name,'') AS godown_name,
                   COALESCE(h.note,'') AS note
            FROM pm_user_home_godown h
            LEFT JOIN procurement_godowns g ON g.id=h.godown_id
        """).fetchall()
        home_by_user = {r['user_name']: dict(r) for r in home_rows}

        # ── Build the final list ─────────────────────────────────────
        # Normal path: master table exists → show those users, joined with
        # home mappings. Mappings for users not in the master table are
        # still included (admin may have mapped a name that no longer
        # exists in the master, so they should at least see/clear it).
        users = []
        if users_from_master:
            for u in users_from_master:
                h = home_by_user.get(u['user_name']) or {}
                users.append({
                    'user_name':       u['user_name'],
                    'display_name':    u.get('display_name') or '',
                    'user_role':       u.get('user_role') or '',
                    'home_godown_id':  int(h['godown_id']) if h.get('godown_id') else None,
                    'home_godown_name':h.get('godown_name') or '',
                    'note':            h.get('note') or '',
                    'orphan':          False,
                })
            seen = {u['user_name'] for u in users}
            for h in home_rows:
                if h['user_name'] not in seen:
                    # An old mapping pointing at a user that's no longer in
                    # the master table — keep it visible so admin can clear it
                    users.append({
                        'user_name':       h['user_name'],
                        'display_name':    '',
                        'user_role':       '',
                        'home_godown_id':  int(h['godown_id']) if h.get('godown_id') else None,
                        'home_godown_name':h.get('godown_name') or '',
                        'note':            h.get('note') or '',
                        'orphan':          True,    # not in user master
                    })
        else:
            # Fallback: no user master table found. Surface whatever we know
            # from PM history + existing mappings so the modal isn't useless.
            actor_rows = conn.execute("""
                SELECT user_name FROM (
                    SELECT created_by AS user_name FROM pm_grn
                      WHERE created_by IS NOT NULL AND created_by <> '' AND created_by <> 'Unknown'
                    UNION
                    SELECT out_by FROM pm_transfers
                      WHERE out_by IS NOT NULL AND out_by <> '' AND out_by <> 'Unknown'
                    UNION
                    SELECT in_by FROM pm_transfers
                      WHERE in_by IS NOT NULL AND in_by <> '' AND in_by <> 'Unknown'
                    UNION
                    SELECT user_name FROM pm_user_home_godown
                ) u GROUP BY user_name ORDER BY user_name
            """).fetchall()
            for r in actor_rows:
                nm = r['user_name']
                if not nm: continue
                h = home_by_user.get(nm) or {}
                users.append({
                    'user_name':       nm,
                    'display_name':    '',
                    'user_role':       '',
                    'home_godown_id':  int(h['godown_id']) if h.get('godown_id') else None,
                    'home_godown_name':h.get('godown_name') or '',
                    'note':            h.get('note') or '',
                    'orphan':          False,
                })

        users.sort(key=lambda u: (u.get('user_name') or '').lower())
        conn.close()
        return jsonify({
            'status':            'ok',
            'users':             users,
            'source_table':      master_table_name or '',  # empty string = fallback used
        })
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/godowns')
@_login_required
def api_pm_godowns():
    """Return all active godowns from procurement_godowns.
       Adds is_floor=True for locations that are used as factory/floor destinations
       in pm_mtv (to_type='floor') or have pm_floor_txn entries with that godown_id.
    """
    conn = sampling_portal.get_db_connection()
    gdwns = _get_godowns(conn)

    # Detect floor locations from actual transaction data — don't trust type column
    # A location is floor if pm_floor_txn has issue/dispatch entries against it.
    # (Legacy MTV-based detection removed — pm_mtv table no longer exists.)
    try:
        floor_ids = set(
            r['godown_id'] for r in conn.execute(
                "SELECT DISTINCT godown_id FROM pm_floor_txn WHERE txn_type IN ('issue','dispatch') AND godown_id IS NOT NULL"
            ).fetchall()
        )
    except Exception:
        floor_ids = set()

    for g in gdwns:
        if g['godown_type'] == 'floor' or g['id'] in floor_ids:
            g['is_floor'] = True
        else:
            g['is_floor'] = False

    conn.close()
    return jsonify(gdwns)


# ═══════════════════════════════════════════════════════════════════════════
#  MATERIAL LOCK  —  Manager/admin can lock material from Material OUT
#  (independent of FIFO). Two parameter types: before_date and grn.
# ═══════════════════════════════════════════════════════════════════════════
@pm_stock_bp.route('/api/pm_stock/material_locks', methods=['GET'])
@_login_required
def api_material_locks_list():
    """List material-lock rules (newest first). Visible to any logged-in user
    so the UI can show why a scan was blocked; only managers/admins can edit."""
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute("""
            SELECT l.lock_id, l.product_id, l.product_name, l.mode, l.param_type,
                   l.godown_id, l.cutoff_date, l.grn_id, l.grn_no, l.note,
                   l.is_active, l.created_by, l.created_at, l.updated_by, l.updated_at,
                   COALESCE(g.name,'') AS godown_name,
                   COALESCE(p.product_name, l.product_name) AS product_name_live
            FROM pm_material_locks l
            LEFT JOIN procurement_godowns g ON g.id = l.godown_id
            LEFT JOIN pm_products p ON p.id = l.product_id
            ORDER BY l.is_active DESC, l.created_at DESC
            LIMIT 500
        """).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            for k in ('created_at', 'updated_at', 'cutoff_date'):
                if d.get(k) is not None: d[k] = str(d[k])
            out.append(d)
        return jsonify({'status': 'ok', 'locks': out,
                        'can_manage': _user_has_access('material_lock'),
                        'count': len(out)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/material_locks', methods=['POST'])
@_login_required
def api_material_locks_create():
    """Create a lock rule for a SPECIFIC item (product). Manager/admin only.

    Body: { product_id:int (required), mode:'block'|'allow',
            param_type:'before_date'|'grn',
            godown_id?:int|null (null=all locations),
            cutoff_date?:'YYYY-MM-DD' (for before_date),
            grn_id?:int (for grn), note?:str }
    """
    if not _user_has_access('material_lock'):
        return jsonify({'status': 'error', 'message': 'You do not have access to Material Lock'}), 403
    d = request.get_json() or {}
    try:
        product_id = int(d.get('product_id'))
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'product_id is required (rules are per item)'}), 400
    mode = (d.get('mode') or 'block').strip().lower()
    if mode not in ('block', 'allow'):
        return jsonify({'status': 'error', 'message': "mode must be 'block' or 'allow'"}), 400
    ptype = (d.get('param_type') or '').strip().lower()
    if ptype not in ('before_date', 'grn'):
        return jsonify({'status': 'error', 'message': "param_type must be 'before_date' or 'grn'"}), 400

    godown_id = d.get('godown_id')
    if godown_id in ('', 'null', None):
        godown_id = None
    else:
        try: godown_id = int(godown_id)
        except (TypeError, ValueError):
            return jsonify({'status': 'error', 'message': 'godown_id must be an integer or null'}), 400

    cutoff_date = None
    grn_id = None
    grn_no = None
    conn = sampling_portal.get_db_connection()
    try:
        prow = conn.execute("SELECT product_name FROM pm_products WHERE id=%s LIMIT 1", (product_id,)).fetchone()
        if not prow:
            conn.close(); return jsonify({'status': 'error', 'message': f'Product #{product_id} not found'}), 404
        product_name = prow['product_name']

        if ptype == 'before_date':
            cutoff_date = (d.get('cutoff_date') or '').strip()
            if not cutoff_date:
                conn.close(); return jsonify({'status': 'error', 'message': 'cutoff_date is required'}), 400
        else:  # grn
            try: grn_id = int(d.get('grn_id'))
            except (TypeError, ValueError):
                conn.close(); return jsonify({'status': 'error', 'message': 'grn_id is required'}), 400
            grow = conn.execute("SELECT grn_no FROM pm_grn WHERE id=%s LIMIT 1", (grn_id,)).fetchone()
            if not grow:
                conn.close(); return jsonify({'status': 'error', 'message': f'GRN #{grn_id} not found'}), 404
            grn_no = grow['grn_no']

        cur = conn.execute(
            """INSERT INTO pm_material_locks
                 (product_id, product_name, mode, param_type, godown_id,
                  cutoff_date, grn_id, grn_no, note, is_active, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,1,%s)""",
            (product_id, product_name, mode, ptype, godown_id,
             (cutoff_date or None), grn_id, grn_no,
             (d.get('note') or '').strip()[:500] or None, _user())
        )
        lock_id = cur.lastrowid
        try:
            _audit_record(
                conn, action='material_lock.create', entity='material_lock',
                entity_id=lock_id,
                summary=f"Material {mode} rule created for {product_name} ({ptype})",
                before=None,
                after={'product_id': product_id, 'mode': mode, 'param_type': ptype,
                       'godown_id': godown_id, 'cutoff_date': cutoff_date, 'grn_no': grn_no},
            )
        except Exception:
            pass
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'lock_id': lock_id})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/material_locks/<int:lock_id>/toggle', methods=['POST'])
@_login_required
def api_material_locks_toggle(lock_id):
    """Activate / deactivate a lock rule without deleting it. Manager/admin only."""
    if not _user_has_access('material_lock'):
        return jsonify({'status': 'error', 'message': 'You do not have access to Material Lock'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute("SELECT is_active FROM pm_material_locks WHERE lock_id=%s", (lock_id,)).fetchone()
        if not row:
            conn.close(); return jsonify({'status': 'error', 'message': 'Rule not found'}), 404
        new_state = 0 if row['is_active'] else 1
        conn.execute(
            "UPDATE pm_material_locks SET is_active=%s, updated_by=%s, updated_at=NOW() WHERE lock_id=%s",
            (new_state, _user(), lock_id)
        )
        try:
            _audit_record(
                conn, action='material_lock.toggle', entity='material_lock',
                entity_id=lock_id,
                summary=f"Material lock rule {'activated' if new_state else 'deactivated'}",
                before={'is_active': row['is_active']}, after={'is_active': new_state},
            )
        except Exception:
            pass
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'is_active': new_state})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/material_locks/<int:lock_id>', methods=['DELETE'])
@_login_required
def api_material_locks_delete(lock_id):
    """Delete a lock rule. Manager/admin only."""
    if not _user_has_access('material_lock'):
        return jsonify({'status': 'error', 'message': 'You do not have access to Material Lock'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute("SELECT mode, param_type FROM pm_material_locks WHERE lock_id=%s", (lock_id,)).fetchone()
        if not row:
            conn.close(); return jsonify({'status': 'error', 'message': 'Rule not found'}), 404
        conn.execute("DELETE FROM pm_material_locks WHERE lock_id=%s", (lock_id,))
        try:
            _audit_record(
                conn, action='material_lock.delete', entity='material_lock',
                entity_id=lock_id, summary="Material lock rule deleted",
                before={'mode': row['mode'], 'param_type': row['param_type']}, after=None,
            )
        except Exception:
            pass
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/material_locks/item_grns/<int:product_id>', methods=['GET'])
@_login_required
def api_material_lock_item_grns(product_id):
    """GRNs that actually contain boxes of the given product — used by the
    Material Lock form so the GRN combobox only offers GRNs relevant to the
    selected item. Distinct GRNs (real receipts, not opening stock), newest
    first."""
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute("""
            SELECT DISTINCT b.grn_id AS id, b.grn_no,
                   g.grn_date, COALESCE(g.supplier,'') AS supplier
            FROM pm_boxes b
            JOIN pm_grn g ON g.id = b.grn_id
            WHERE b.product_id=%s
              AND b.grn_id IS NOT NULL
              AND UPPER(COALESCE(b.grn_no,'')) NOT LIKE 'PM-OP/%%'
            ORDER BY g.grn_date DESC, g.id DESC
            LIMIT 300
        """, (product_id,)).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            if d.get('grn_date') is not None: d['grn_date'] = str(d['grn_date'])
            out.append(d)
        return jsonify({'status': 'ok', 'grns': out})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/material_locks/item_locations/<int:product_id>', methods=['GET'])
@_login_required
def api_material_lock_item_locations(product_id):
    """Locations where the given product CURRENTLY resides as live stock
    (in_stock boxes with a real location). Used by the Material Lock form so
    the Location dropdown only offers places the item actually sits right now.
    Includes a per-location box count for context."""
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute("""
            SELECT b.current_godown_id AS id, COALESCE(g.name,'') AS name,
                   COUNT(*) AS box_count
            FROM pm_boxes b
            LEFT JOIN procurement_godowns g ON g.id = b.current_godown_id
            WHERE b.product_id=%s
              AND b.current_status='in_stock'
              AND b.current_godown_id IS NOT NULL
            GROUP BY b.current_godown_id, g.name
            ORDER BY g.name
        """, (product_id,)).fetchall()
        # Flag floor locations the same way /godowns does (data-driven).
        try:
            floor_ids = set(
                r['godown_id'] for r in conn.execute(
                    "SELECT DISTINCT godown_id FROM pm_floor_txn WHERE txn_type IN ('issue','dispatch') AND godown_id IS NOT NULL"
                ).fetchall()
            )
        except Exception:
            floor_ids = set()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            d['is_floor'] = (d['id'] in floor_ids)
            out.append(d)
        return jsonify({'status': 'ok', 'locations': out})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/material_locks/why_blocked', methods=['GET'])
@_login_required
def api_material_lock_why_blocked():
    """Diagnostic: explain whether a given box (by short_code/box_code) OR a
    given product+location would be blocked from Material OUT by a lock rule,
    and show every lock rule that applies to the product. Read-only.

    Query: ?code=<short_or_box_code>  OR  ?product_id=<id>&godown_id=<id>
    """
    code = (request.args.get('code') or '').strip()
    product_id = request.args.get('product_id')
    godown_id = request.args.get('godown_id')
    conn = sampling_portal.get_db_connection()
    try:
        box = None
        if code:
            box = conn.execute(
                """SELECT b.box_id, b.product_id, b.current_godown_id, b.current_status,
                          b.grn_id, b.grn_no, b.short_code, b.box_code, g.grn_date,
                          COALESCE(gd.name,'') AS godown_name, COALESCE(p.product_name,'') AS product_name
                   FROM pm_boxes b
                   LEFT JOIN pm_grn g ON g.id = b.grn_id
                   LEFT JOIN procurement_godowns gd ON gd.id = b.current_godown_id
                   LEFT JOIN pm_products p ON p.id = b.product_id
                   WHERE b.short_code=%s OR b.box_code=%s LIMIT 1""",
                (code, code)
            ).fetchone()
            if not box:
                conn.close(); return jsonify({'status':'error','message':f'No box found for code "{code}"'}), 404
            product_id = box['product_id']; godown_id = box['current_godown_id']
            grn_id = box['grn_id']; grn_date = box['grn_date']; grn_no = box['grn_no']
            is_op = (box['grn_id'] is None) or str(box['grn_no'] or '').upper().startswith('PM-OP/')
        else:
            try:
                product_id = int(product_id); godown_id = int(godown_id) if godown_id else None
            except (TypeError, ValueError):
                conn.close(); return jsonify({'status':'error','message':'Provide code, or product_id (+godown_id)'}), 400
            grn_id = grn_date = grn_no = None; is_op = False

        # All lock rules for this product (any location), for transparency.
        rules = conn.execute(
            """SELECT l.lock_id, l.mode, l.param_type, l.godown_id, l.cutoff_date,
                      l.grn_id, l.grn_no, l.is_active, COALESCE(g.name,'') AS godown_name
               FROM pm_material_locks l
               LEFT JOIN procurement_godowns g ON g.id=l.godown_id
               WHERE l.product_id=%s ORDER BY l.is_active DESC, l.lock_id DESC""",
            (product_id,)
        ).fetchall()
        rule_list = []
        for r in rules:
            d = dict(r)
            for k in ('cutoff_date',):
                if d.get(k) is not None: d[k] = str(d[k])[:10]
            rule_list.append(d)

        blocked, why = _material_lock_check(
            conn, product_id=product_id, godown_id=godown_id,
            grn_id=grn_id, grn_date=grn_date, is_opening=is_op
        )
        conn.close()
        resp = {
            'status': 'ok',
            'blocked': blocked,
            'reason': why,
            'product_id': product_id,
            'godown_id': godown_id,
            'rules_for_product': rule_list,
        }
        if box:
            resp['box'] = {
                'short_code': box['short_code'], 'box_code': box['box_code'],
                'product_name': box['product_name'], 'godown_name': box['godown_name'],
                'grn_no': box['grn_no'], 'grn_date': str(box['grn_date'])[:10] if box['grn_date'] else None,
                'current_status': box['current_status'],
                'is_opening': is_op,
            }
        return jsonify(resp)
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/box_location_diag', methods=['GET'])
@_login_required
def api_box_location_diag():
    """Diagnostic for the 'Box is at a different location' OUT-scan error.

    Shows where a box CURRENTLY sits (current_godown_id + name + status) and
    its full movement history, so you can tell whether the box genuinely is
    elsewhere (correct refusal) or whether current_godown_id is stale /
    inconsistent with its movements (a data bug to repair). Read-only.

    Query: ?code=<short_or_box_code>  [&source_godown_id=<id>]
    """
    code = (request.args.get('code') or '').strip()
    src = request.args.get('source_godown_id')
    if not code:
        return jsonify({'status':'error','message':'code required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        b = conn.execute(
            """SELECT b.box_id, b.short_code, b.box_code, b.product_id,
                      COALESCE(p.product_name,'') AS product_name,
                      b.current_godown_id, b.current_status, b.per_box_qty,
                      b.parent_box_id, b.grn_no,
                      COALESCE(gd.name,'') AS current_godown_name
               FROM pm_boxes b
               LEFT JOIN pm_products p ON p.id=b.product_id
               LEFT JOIN procurement_godowns gd ON gd.id=b.current_godown_id
               WHERE b.short_code=%s OR b.box_code=%s LIMIT 1""",
            (code, code)
        ).fetchone()
        if not b:
            conn.close(); return jsonify({'status':'error','message':f'No box for code "{code}"'}), 404

        moves = conn.execute(
            """SELECT m.movement_id, m.movement_type, m.from_godown_id, m.to_godown_id,
                      m.qty, m.movement_at, m.moved_by, m.transfer_id, m.remarks,
                      COALESCE(fg.name,'') AS from_name, COALESCE(tg.name,'') AS to_name
               FROM pm_box_movements m
               LEFT JOIN procurement_godowns fg ON fg.id=m.from_godown_id
               LEFT JOIN procurement_godowns tg ON tg.id=m.to_godown_id
               WHERE m.box_id=%s ORDER BY m.movement_at, m.movement_id""",
            (b['box_id'],)
        ).fetchall()
        mv = []
        last_to = None
        for m in moves:
            d = dict(m)
            if d.get('movement_at') is not None: d['movement_at'] = str(d['movement_at'])
            d['qty'] = float(d['qty'] or 0)
            mv.append(d)
            if m['to_godown_id'] is not None:
                last_to = {'id': m['to_godown_id'], 'name': m['to_name']}

        # Consistency check: does current_godown_id match the destination of
        # the most recent movement that had a to_godown_id?
        consistent = (last_to is None) or (b['current_godown_id'] == last_to['id'])

        srcmatch = None
        if src:
            try: srcmatch = (b['current_godown_id'] == int(src))
            except (TypeError, ValueError): srcmatch = None

        conn.close()
        return jsonify({
            'status': 'ok',
            'box': {
                'short_code': b['short_code'], 'box_code': b['box_code'],
                'product_name': b['product_name'],
                'current_godown_id': b['current_godown_id'],
                'current_godown_name': b['current_godown_name'],
                'current_status': b['current_status'],
                'per_box_qty': float(b['per_box_qty'] or 0),
                'parent_box_id': b['parent_box_id'],
                'grn_no': b['grn_no'],
            },
            'movements': mv,
            'last_movement_to': last_to,
            'current_matches_last_movement': consistent,
            'source_godown_id_checked': (int(src) if src else None),
            'current_matches_source': srcmatch,
            'hint': (
                'current_godown_id is INCONSISTENT with the box movement history — '
                'likely a stale location from a relocation/reissue that updated the '
                'ledger but not pm_boxes.current_godown_id. This box may need a repair.'
                if not consistent else
                ('Box genuinely sits at a different location than the OUT source you chose — '
                 'pick that location as the source, or relocate the box first.'
                 if srcmatch is False else
                 'Box location looks consistent with its history.')
            ),
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/location_truth_audit', methods=['GET'])
@_login_required
def api_location_truth_audit():
    """READ-ONLY truth table for the box↔ledger location question.

    For each product (or one product_id), reports side by side:
      • ledger_godown_qty  — net qty the GODOWN ledger (pm_godown_txn) shows
      • ledger_factory_qty — net qty the FACTORY ledger (pm_floor_txn) shows
      • box_godown_qty     — in_stock box per_box_qty currently AT godown(s)
      • box_factory_qty    — in_stock box per_box_qty currently AT factory
    plus the mismatches between them. Makes NO changes — purely so we can SEE
    where each product's truth diverges before deciding how to repair.

    Query: ?product_id=<id>  (omit for all in_stock products) &only_mismatch=1
    """
    product_id = request.args.get('product_id')
    only_mismatch = request.args.get('only_mismatch') in ('1','true','yes')
    conn = sampling_portal.get_db_connection()
    try:
        floor_rows = conn.execute("SELECT id, name FROM procurement_godowns").fetchall()
        floor_ids = [g['id'] for g in floor_rows if _is_floor_godown(conn, g['id'])]
        floor_set = set(floor_ids)

        if product_id:
            pids = [int(product_id)]
        else:
            pr = conn.execute("""SELECT DISTINCT product_id FROM pm_boxes
                                 WHERE current_status='in_stock'
                                 UNION SELECT DISTINCT product_id FROM pm_godown_txn
                                 UNION SELECT DISTINCT product_id FROM pm_floor_txn""").fetchall()
            pids = [r['product_id'] for r in pr if r['product_id']]

        def _floor_ph():
            return ','.join(['%s']*len(floor_ids)) if floor_ids else 'NULL'

        rows = []
        for pid in pids:
            gl = conn.execute("""
                SELECT COALESCE(SUM(CASE WHEN txn_type IN ('opening','inward') THEN qty
                                         WHEN txn_type='outward' THEN -qty ELSE 0 END),0) AS q
                FROM pm_godown_txn WHERE product_id=%s""", (pid,)).fetchone()
            fl = conn.execute("""
                SELECT COALESCE(SUM(CASE WHEN txn_type IN ('floor_opening','issue','pm_return') THEN qty
                                         WHEN txn_type IN ('dispatch','rejection') THEN -qty ELSE 0 END),0) AS q
                FROM pm_floor_txn WHERE product_id=%s""", (pid,)).fetchone()
            ledger_godown = float(gl['q'] or 0)
            ledger_factory = float(fl['q'] or 0)

            if floor_ids:
                bf = conn.execute("""SELECT COALESCE(SUM(per_box_qty),0) AS q, COUNT(*) AS n
                                     FROM pm_boxes WHERE product_id=%s AND current_status='in_stock'
                                       AND current_godown_id IN ({})""".format(_floor_ph()),
                                  tuple([pid]+floor_ids)).fetchone()
                bg = conn.execute("""SELECT COALESCE(SUM(per_box_qty),0) AS q, COUNT(*) AS n
                                     FROM pm_boxes WHERE product_id=%s AND current_status='in_stock'
                                       AND current_godown_id IS NOT NULL
                                       AND current_godown_id NOT IN ({})""".format(_floor_ph()),
                                  tuple([pid]+floor_ids)).fetchone()
            else:
                bf = {'q':0,'n':0}
                bg = conn.execute("""SELECT COALESCE(SUM(per_box_qty),0) AS q, COUNT(*) AS n
                                     FROM pm_boxes WHERE product_id=%s AND current_status='in_stock'
                                       AND current_godown_id IS NOT NULL""", (pid,)).fetchone()
            box_factory = float(bf['q'] or 0); box_godown = float(bg['q'] or 0)

            godown_gap = round(ledger_godown - box_godown, 3)
            factory_gap = round(ledger_factory - box_factory, 3)
            has_mismatch = (abs(godown_gap) > 0.001 or abs(factory_gap) > 0.001)
            if only_mismatch and not has_mismatch:
                continue
            prow = conn.execute("SELECT product_name FROM pm_products WHERE id=%s", (pid,)).fetchone()
            rows.append({
                'product_id': pid,
                'product_name': prow['product_name'] if prow else str(pid),
                'ledger_godown_qty': ledger_godown,
                'ledger_factory_qty': ledger_factory,
                'box_godown_qty': box_godown, 'box_godown_count': bg['n'],
                'box_factory_qty': box_factory, 'box_factory_count': bf['n'] if isinstance(bf, dict)==False else bf['n'],
                'godown_gap_ledger_minus_box': godown_gap,
                'factory_gap_ledger_minus_box': factory_gap,
                'has_mismatch': has_mismatch,
            })

        conn.close()
        return jsonify({
            'status':'ok',
            'factory_godown_ids': floor_ids,
            'product_count': len(rows),
            'rows': rows,
            'legend': {
                'godown_gap_ledger_minus_box': 'positive = ledger shows MORE at godown than boxes do',
                'factory_gap_ledger_minus_box': 'positive = ledger shows MORE at factory than boxes do (the issue-to-factory case)',
            },
            'note': 'READ-ONLY. No changes made.',
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/admin/box_locations_for_product/<int:product_id>', methods=['GET'])
@_login_required
def api_admin_box_locations_for_product(product_id):
    """Diagnostic: list every in_stock box for a product with its current_godown_id.

    Returns a per-godown breakdown so an admin can see exactly which godown
    each box for this product is currently parked at. Used when Material OUT
    scanning fails with "Box is at a different location" — admin runs this,
    sees that e.g. 245 boxes are at FACTORY when the operator expected them
    at NBG, and can decide whether to:
      * fix the location via /box_relocate, OR
      * tell the operator to scan a different physical box, OR
      * reconcile via /box_location_reconcile if the ledger says they
        should be at factory anyway.

    Admin only.
    """
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403

    conn = sampling_portal.get_db_connection()
    try:
        # Product header
        p = conn.execute(
            """SELECT id, product_name, product_code, pm_type
               FROM pm_products WHERE id=%s""",
            (product_id,)
        ).fetchone()
        if not p:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Product not found'}), 404

        # All in_stock boxes for this product, grouped by godown
        rows = conn.execute(
            """SELECT b.box_id, b.box_code, b.short_code, b.per_box_qty,
                      b.current_godown_id, b.current_status, b.grn_no,
                      COALESCE(g.name, '(no godown set)') AS godown_name,
                      g.type AS godown_type
                 FROM pm_boxes b
                 LEFT JOIN procurement_godowns g ON g.id = b.current_godown_id
                WHERE b.product_id = %s
                  AND b.current_status = 'in_stock'
                ORDER BY b.current_godown_id, b.grn_no, b.box_seq""",
            (product_id,)
        ).fetchall() or []

        # Roll up per-godown summary
        summary = {}
        for r in rows:
            rd = dict(r) if hasattr(r, 'keys') else r
            gid = rd.get('current_godown_id') or 0
            if gid not in summary:
                summary[gid] = {
                    'godown_id':   gid,
                    'godown_name': rd.get('godown_name'),
                    'godown_type': rd.get('godown_type'),
                    'box_count':   0,
                    'total_qty':   0.0,
                    'boxes':       [],
                }
            summary[gid]['box_count'] += 1
            summary[gid]['total_qty'] += float(rd.get('per_box_qty') or 0)
            summary[gid]['boxes'].append({
                'box_id':     rd['box_id'],
                'box_code':   rd['box_code'],
                'short_code': rd.get('short_code'),
                'qty':        float(rd.get('per_box_qty') or 0),
                'grn_no':     rd.get('grn_no'),
            })

        return jsonify({
            'status':  'ok',
            'product': dict(p),
            'per_godown': list(summary.values()),
            'total_in_stock_boxes': len(rows),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        try: conn.close()
        except Exception: pass


@pm_stock_bp.route('/api/pm_stock/admin/box_relocate', methods=['POST'])
@_login_required
def api_admin_box_relocate():
    """Move one or more boxes' current_godown_id to a different godown.

    Admin only. Records a 'manual_relocate' pm_box_movements row per box
    for traceability. Use case: after running
    /admin/box_locations_for_product, admin sees boxes at the wrong
    godown and wants to fix them so operators can scan during Material
    OUT without "Box is at a different location" errors.

    Body: { box_ids: [int,...], to_godown_id: int, remarks?: str }
    """
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    body = request.get_json(silent=True) or {}
    box_ids = body.get('box_ids') or []
    to_godown_id = body.get('to_godown_id')
    remarks = (body.get('remarks') or 'Admin manual relocate').strip()
    if not box_ids or not to_godown_id:
        return jsonify({'status': 'error', 'message': 'box_ids and to_godown_id required'}), 400

    conn = sampling_portal.get_db_connection()
    moved = 0
    errors = []
    try:
        # Verify target godown exists
        gd = conn.execute(
            "SELECT id, name FROM procurement_godowns WHERE id=%s", (to_godown_id,)
        ).fetchone()
        if not gd:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Target godown not found'}), 404
        for bid in box_ids:
            try:
                bid_int = int(bid)
                b = conn.execute(
                    "SELECT box_id, current_godown_id, per_box_qty FROM pm_boxes WHERE box_id=%s",
                    (bid_int,)
                ).fetchone()
                if not b:
                    errors.append(f'Box {bid} not found')
                    continue
                from_gid = b['current_godown_id'] if hasattr(b, 'get') else b[1]
                qty      = float(b['per_box_qty'] if hasattr(b, 'get') else b[2] or 0)
                conn.execute(
                    "UPDATE pm_boxes SET current_godown_id=%s WHERE box_id=%s",
                    (to_godown_id, bid_int)
                )
                conn.execute(
                    """INSERT INTO pm_box_movements
                         (box_id, transfer_id, movement_type, from_godown_id, to_godown_id, qty, moved_by, remarks)
                       VALUES (%s, NULL, 'manual_relocate', %s, %s, %s, %s, %s)""",
                    (bid_int, from_gid, to_godown_id, qty, _user(), remarks)
                )
                moved += 1
            except Exception as _be:
                errors.append(f'Box {bid}: {_be}')
                continue
        conn.commit()
        # Audit a single rollup record for the whole batch
        try:
            _audit_record(
                conn, action='box.manual_relocate', entity='box', entity_id=0,
                summary=f"Moved {moved} box(es) to {gd['name']}",
                before=None, after={'to_godown_id': to_godown_id, 'count': moved, 'box_ids': box_ids[:50]},
            )
            conn.commit()
        except Exception: pass
        return jsonify({'status': 'ok', 'moved': moved, 'errors': errors})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        try: conn.close()
        except Exception: pass


@pm_stock_bp.route('/api/pm_stock/box_location_reconcile_preview', methods=['GET'])
@_login_required
def api_box_location_reconcile_preview():
    """PREVIEW (read-only) of the box↔ledger location reconciliation.

    Background: quantity-based 'Issue → Factory' moves the LEDGER (godown
    outward + floor issue) but never updates pm_boxes.current_godown_id. So
    boxes can still point to the Godown while the ledger says the stock is at
    Factory — which makes OUT scans fail at BOTH locations ('Box is at a
    different location' from Factory; 'no stock' from Godown).

    This computes, per product, the ledger's Factory (floor) closing qty vs.
    how much in_stock box quantity currently sits at NON-floor (godown)
    locations, and proposes moving the oldest godown boxes to the Factory
    location until the box quantity at Factory matches the ledger — WITHOUT
    making any change.

    Query: ?product_id=<id>  (omit for ALL products with a mismatch)
    """
    product_id = request.args.get('product_id')
    conn = sampling_portal.get_db_connection()
    try:
        # Resolve the Factory (floor) godown id(s).
        floor_rows = conn.execute(
            "SELECT id, name FROM procurement_godowns"
        ).fetchall()
        floor_ids = [g['id'] for g in floor_rows if _is_floor_godown(conn, g['id'])]
        if not floor_ids:
            conn.close(); return jsonify({'status':'error','message':'No factory/floor godown found'}), 400
        factory_id = floor_ids[0]
        factory_name = next((g['name'] for g in floor_rows if g['id'] == factory_id), 'Factory')

        # Products to inspect.
        if product_id:
            pids = [int(product_id)]
        else:
            pr = conn.execute("""
                SELECT DISTINCT product_id FROM pm_boxes
                WHERE current_status='in_stock' AND current_godown_id IS NOT NULL
            """).fetchall()
            pids = [r['product_id'] for r in pr]

        proposals = []
        for pid in pids:
            # Ledger floor (Factory) closing qty for this product.
            frow = conn.execute("""
                SELECT COALESCE(SUM(CASE WHEN txn_type IN ('floor_opening','issue','pm_return') THEN qty
                                         WHEN txn_type IN ('dispatch','rejection') THEN -qty ELSE 0 END),0) AS floor_qty
                FROM pm_floor_txn WHERE product_id=%s
            """, (pid,)).fetchone()
            ledger_factory_qty = float(frow['floor_qty'] or 0)

            # In_stock box quantity currently AT the factory vs at godown(s).
            box_factory = conn.execute("""
                SELECT COALESCE(SUM(per_box_qty),0) AS q, COUNT(*) AS n
                FROM pm_boxes WHERE product_id=%s AND current_status='in_stock'
                  AND current_godown_id IN ({})
            """.format(','.join(['%s']*len(floor_ids))),
                tuple([pid]+floor_ids)).fetchone()
            box_factory_qty = float(box_factory['q'] or 0)

            shortfall = ledger_factory_qty - box_factory_qty
            if shortfall <= 0:
                continue   # boxes at Factory already cover (or exceed) ledger — no move needed

            # Candidate godown boxes (oldest first) to move to Factory.
            cands = conn.execute("""
                SELECT b.box_id, b.short_code, b.box_code, b.per_box_qty,
                       b.current_godown_id, COALESCE(g.name,'') AS godown_name, b.grn_no
                FROM pm_boxes b
                LEFT JOIN procurement_godowns g ON g.id=b.current_godown_id
                WHERE b.product_id=%s AND b.current_status='in_stock'
                  AND b.current_godown_id IS NOT NULL
                  AND b.current_godown_id NOT IN ({})
                ORDER BY b.box_id
            """.format(','.join(['%s']*len(floor_ids))),
                tuple([pid]+floor_ids)).fetchall()

            picked = []; acc = 0.0
            for c in cands:
                if acc >= shortfall: break
                q = float(c['per_box_qty'] or 0)
                picked.append({
                    'box_id': c['box_id'], 'short_code': c['short_code'],
                    'box_code': c['box_code'], 'per_box_qty': q,
                    'from_godown_id': c['current_godown_id'], 'from_godown_name': c['godown_name'],
                    'grn_no': c['grn_no'],
                })
                acc += q
            if not picked:
                continue
            prow = conn.execute("SELECT product_name FROM pm_products WHERE id=%s", (pid,)).fetchone()
            proposals.append({
                'product_id': pid,
                'product_name': prow['product_name'] if prow else str(pid),
                'ledger_factory_qty': ledger_factory_qty,
                'box_factory_qty_before': box_factory_qty,
                'shortfall': shortfall,
                'boxes_to_move': picked,
                'qty_to_move': acc,
                'to_godown_id': factory_id, 'to_godown_name': factory_name,
            })

        conn.close()
        return jsonify({
            'status': 'ok',
            'factory_godown_id': factory_id,
            'factory_godown_name': factory_name,
            'product_count_with_mismatch': len(proposals),
            'proposals': proposals,
            'note': 'PREVIEW ONLY — no changes made. POST to /box_location_reconcile to apply.',
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/box_location_reconcile', methods=['POST'])
@_login_required
def api_box_location_reconcile():
    """APPLY the box↔ledger location reconciliation previewed above.
    Admin only. Moves the proposed godown boxes' current_godown_id to the
    Factory location and logs a pm_box_movements 'adjust' row each.

    Body: { product_id?:int, confirm:true }  (product_id omitted = all)
    Requires confirm:true as a guard against accidental calls.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    d = request.get_json() or {}
    if not d.get('confirm'):
        return jsonify({'status':'error','message':'Pass confirm:true to apply. Use the preview endpoint first.'}), 400

    # Re-run the same computation as the preview to get the authoritative set,
    # then apply. (We recompute server-side rather than trust a client list.)
    import urllib.parse as _u
    product_id = d.get('product_id')
    conn = sampling_portal.get_db_connection()
    try:
        floor_rows = conn.execute("SELECT id, name FROM procurement_godowns").fetchall()
        floor_ids = [g['id'] for g in floor_rows if _is_floor_godown(conn, g['id'])]
        if not floor_ids:
            conn.close(); return jsonify({'status':'error','message':'No factory/floor godown found'}), 400
        factory_id = floor_ids[0]

        if product_id:
            pids = [int(product_id)]
        else:
            pr = conn.execute("""SELECT DISTINCT product_id FROM pm_boxes
                                 WHERE current_status='in_stock' AND current_godown_id IS NOT NULL""").fetchall()
            pids = [r['product_id'] for r in pr]

        moved_total = 0; products_fixed = 0
        for pid in pids:
            frow = conn.execute("""
                SELECT COALESCE(SUM(CASE WHEN txn_type IN ('floor_opening','issue','pm_return') THEN qty
                                         WHEN txn_type IN ('dispatch','rejection') THEN -qty ELSE 0 END),0) AS floor_qty
                FROM pm_floor_txn WHERE product_id=%s""", (pid,)).fetchone()
            ledger_factory_qty = float(frow['floor_qty'] or 0)
            box_factory = conn.execute("""
                SELECT COALESCE(SUM(per_box_qty),0) AS q FROM pm_boxes
                WHERE product_id=%s AND current_status='in_stock'
                  AND current_godown_id IN ({})""".format(','.join(['%s']*len(floor_ids))),
                tuple([pid]+floor_ids)).fetchone()
            shortfall = ledger_factory_qty - float(box_factory['q'] or 0)
            if shortfall <= 0:
                continue
            cands = conn.execute("""
                SELECT box_id, per_box_qty, current_godown_id FROM pm_boxes
                WHERE product_id=%s AND current_status='in_stock'
                  AND current_godown_id IS NOT NULL
                  AND current_godown_id NOT IN ({})
                ORDER BY box_id""".format(','.join(['%s']*len(floor_ids))),
                tuple([pid]+floor_ids)).fetchall()
            acc = 0.0; n = 0
            for c in cands:
                if acc >= shortfall: break
                q = float(c['per_box_qty'] or 0)
                conn.execute("UPDATE pm_boxes SET current_godown_id=%s WHERE box_id=%s",
                             (factory_id, c['box_id']))
                conn.execute("""INSERT INTO pm_box_movements
                                  (box_id, transfer_id, movement_type, from_godown_id, to_godown_id, qty, moved_by, remarks)
                                VALUES (%s,NULL,'adjust',%s,%s,%s,%s,%s)""",
                             (c['box_id'], c['current_godown_id'], factory_id, q, _user(),
                              'Box↔ledger location reconcile (issued-to-factory boxes never moved)'))
                acc += q; n += 1; moved_total += 1
            if n:
                products_fixed += 1
                try:
                    _audit_record(conn, action='box_location.reconcile', entity='product',
                                  entity_id=pid, summary=f"Reconciled {n} box(es) to Factory for product {pid}",
                                  before={'box_factory_qty': float(box_factory['q'] or 0)},
                                  after={'moved_boxes': n, 'qty_moved': acc, 'to_godown_id': factory_id})
                except Exception: pass
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','products_fixed':products_fixed,'boxes_moved':moved_total})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/backfill_opening_godown', methods=['POST'])
@_login_required
def api_backfill_opening_godown():
    """
    One-time utility: assign a godown_id to all existing opening / floor_opening
    rows that currently have godown_id = NULL.

    Body: { "godown_id": <int> }

    Updates:
      pm_godown_txn WHERE txn_type='opening'       AND godown_id IS NULL
      pm_floor_txn  WHERE txn_type='floor_opening' AND godown_id IS NULL
    Returns counts of rows updated.
    """
    d         = request.get_json() or {}
    godown_id = d.get('godown_id')
    if not godown_id:
        return jsonify({'status': 'error', 'message': 'godown_id required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        r1 = conn.execute(
            "UPDATE pm_godown_txn SET godown_id=%s WHERE txn_type='opening' AND godown_id IS NULL",
            (godown_id,)
        )
        r2 = conn.execute(
            "UPDATE pm_floor_txn SET godown_id=%s WHERE txn_type='floor_opening' AND godown_id IS NULL",
            (godown_id,)
        )
        conn.commit()
        conn.close()
        return jsonify({
            'status': 'ok',
            'godown_updated': r1.rowcount,
            'floor_updated':  r2.rowcount,
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/opening/save', methods=['POST'])
@_login_required
def api_opening_save():
    """
    Update or create the opening balance for a product at a specific location.
    Body: { product_id, godown_qty, floor_qty, txn_date, godown_id }
    Lookup is scoped by (product_id, godown_id) so each location's opening
    is set independently. Existing rows are UPDATEd (set, not added); missing
    rows are INSERTed. Stray duplicates are deleted.
    """
    blocked = _block_if_disabled('opening')
    if blocked is not None: return blocked
    # Requesters can't create vouchers — they only submit Material Requests.
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    # New-voucher-entries access check (per-user access control)
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked
    d          = request.get_json() or {}
    product_id = d.get('product_id')
    godown_qty = float(d.get('godown_qty') or 0)
    floor_qty  = float(d.get('floor_qty')  or 0)
    txn_date   = d.get('txn_date') or str(date.today())
    godown_id  = d.get('godown_id') or None

    if not product_id:
        return jsonify({'status': 'error', 'message': 'product_id required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # Capture before-state for both godown + floor opening (used by the
        # audit log and the eventual reversal flow)
        before_g = conn.execute(
            """SELECT COALESCE(SUM(qty),0) AS qty FROM pm_godown_txn
               WHERE product_id=%s AND txn_type='opening'
                 AND COALESCE(godown_id, 0) = COALESCE(%s, 0)""",
            (product_id, godown_id)
        ).fetchone()
        before_f = conn.execute(
            """SELECT COALESCE(SUM(qty),0) AS qty FROM pm_floor_txn
               WHERE product_id=%s AND txn_type='floor_opening'
                 AND COALESCE(godown_id, 0) = COALESCE(%s, 0)""",
            (product_id, godown_id)
        ).fetchone()
        before_state = {
            'godown_qty': float((before_g or {}).get('qty') or 0),
            'floor_qty':  float((before_f or {}).get('qty') or 0),
            'godown_id':  godown_id,
            'txn_date':   str(txn_date),
        }

        # ── Godown opening (scoped by godown_id) ──────────────────────────────
        ex_g = conn.execute(
            """SELECT id FROM pm_godown_txn
               WHERE product_id=%s AND txn_type='opening'
                 AND COALESCE(godown_id, 0) = COALESCE(%s, 0)
               ORDER BY id LIMIT 1""",
            (product_id, godown_id)
        ).fetchone()
        if ex_g:
            conn.execute(
                "UPDATE pm_godown_txn SET qty=%s, txn_date=%s, godown_id=%s WHERE id=%s",
                (godown_qty, txn_date, godown_id, ex_g['id'])
            )
            # Dedupe — remove any other opening rows for this (product, godown)
            conn.execute(
                """DELETE FROM pm_godown_txn
                   WHERE product_id=%s AND txn_type='opening'
                     AND COALESCE(godown_id, 0) = COALESCE(%s, 0)
                     AND id <> %s""",
                (product_id, godown_id, ex_g['id'])
            )
        else:
            conn.execute(
                """INSERT INTO pm_godown_txn (product_id, txn_date, txn_type, qty, remarks, created_by, godown_id)
                   VALUES (%s,%s,'opening',%s,'Opening Balance',%s,%s)""",
                (product_id, txn_date, godown_qty, _user(), godown_id)
            )

        # ── Floor opening (scoped by godown_id) ───────────────────────────────
        ex_f = conn.execute(
            """SELECT id FROM pm_floor_txn
               WHERE product_id=%s AND txn_type='floor_opening'
                 AND COALESCE(godown_id, 0) = COALESCE(%s, 0)
               ORDER BY id LIMIT 1""",
            (product_id, godown_id)
        ).fetchone()
        if ex_f:
            conn.execute(
                "UPDATE pm_floor_txn SET qty=%s, txn_date=%s, godown_id=%s WHERE id=%s",
                (floor_qty, txn_date, godown_id, ex_f['id'])
            )
            conn.execute(
                """DELETE FROM pm_floor_txn
                   WHERE product_id=%s AND txn_type='floor_opening'
                     AND COALESCE(godown_id, 0) = COALESCE(%s, 0)
                     AND id <> %s""",
                (product_id, godown_id, ex_f['id'])
            )
        else:
            conn.execute(
                """INSERT INTO pm_floor_txn (product_id, txn_date, txn_type, qty, remarks, created_by, godown_id)
                   VALUES (%s,%s,'floor_opening',%s,'Factory Opening Balance',%s,%s)""",
                (product_id, txn_date, floor_qty, _user(), godown_id)
            )

        try:
            after_state = {
                'godown_qty': float(godown_qty),
                'floor_qty':  float(floor_qty),
                'godown_id':  godown_id,
                'txn_date':   str(txn_date),
            }
            if before_state != after_state:
                _audit_record(
                    conn,
                    action='product.opening_edit',
                    entity='product',
                    entity_id=product_id,
                    summary=f"Opening edited (godown {before_state['godown_qty']:g}→{after_state['godown_qty']:g}, "
                            f"floor {before_state['floor_qty']:g}→{after_state['floor_qty']:g})",
                    before=before_state,
                    after=after_state,
                )
        except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/product_stock/<int:product_id>')
@_login_required
def api_pm_product_stock(product_id):
    """Total live stock for one product, split Godown vs Factory, for the
    product-page hover tooltip. Computed from the ledger tables (the source
    of truth), same as the stock views.

      godown_qty  = SUM(opening+inward) - SUM(outward)        [pm_godown_txn]
      factory_qty = SUM(floor_opening+issue+pm_return) - SUM(dispatch+rejection) [pm_floor_txn]
      total       = godown_qty + factory_qty
    """
    conn = sampling_portal.get_db_connection()
    try:
        g = conn.execute("""
            SELECT COALESCE(SUM(CASE WHEN txn_type IN ('opening','inward') THEN qty
                                     WHEN txn_type='outward' THEN -qty ELSE 0 END),0) AS q
            FROM pm_godown_txn WHERE product_id=%s""", (product_id,)).fetchone()
        f = conn.execute("""
            SELECT COALESCE(SUM(CASE WHEN txn_type IN ('floor_opening','issue','pm_return') THEN qty
                                     WHEN txn_type IN ('dispatch','rejection') THEN -qty ELSE 0 END),0) AS q
            FROM pm_floor_txn WHERE product_id=%s""", (product_id,)).fetchone()
        conn.close()
        godown = float(g['q'] or 0)
        factory = float(f['q'] or 0)
        return jsonify({
            'status': 'ok',
            'product_id': product_id,
            'godown_qty': godown,
            'factory_qty': factory,
            'total_qty': godown + factory,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/products')
@_login_required
def api_pm_products():
    conn = sampling_portal.get_db_connection()
    rows = conn.execute(
        """
        SELECT p.id, p.product_name, p.pm_type,
               COALESCE(p.product_code,'') as product_code,
               COALESCE(p.brand_id,0) as brand_id,
               COALESCE(b.name,'') as brand_name,
               COALESCE(b.color,'') as brand_color,
               COALESCE(p.min_stock,0) as min_stock,
               COALESCE(p.primary_uom,'Nos') as primary_uom,
               COALESCE(p.alt_uom,'')        as alt_uom,
               p.alt_to_primary_ratio        as alt_to_primary_ratio
        FROM pm_products p
        LEFT JOIN procurement_brands b ON b.id = p.brand_id
        WHERE p.is_active=1 ORDER BY p.product_name
    """
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@pm_stock_bp.route('/api/pm_stock/add_product', methods=['POST'])
@_login_required
def api_pm_add_product():
    blocked = _block_if_disabled('opening')
    if blocked is not None: return blocked
    # Requesters can't create vouchers — they only submit Material Requests.
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    # New-voucher-entries access check (per-user access control)
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked
    d        = request.get_json() or {}
    name     = (d.get('product_name') or '').strip()
    pm       = (d.get('pm_type') or '').strip()
    brand_id = d.get('brand_id') or None
    # UOM (Phase 1). Primary defaults to 'Nos' (Tally-style). Alt is optional;
    # if set, ratio must be a positive number (1 primary = ratio × alt).
    # Preserves whatever casing the user picked ("Kg", "Nos"); only the
    # equality check is case-insensitive so alt ≠ primary even when the
    # user types different cases.
    primary_uom = (d.get('primary_uom') or 'Nos').strip() or 'Nos'
    alt_uom_raw = (d.get('alt_uom') or '').strip()
    alt_uom     = alt_uom_raw or None
    ratio_raw   = d.get('alt_to_primary_ratio')
    alt_ratio   = None
    if alt_uom:
        try:
            alt_ratio = float(ratio_raw)
        except (TypeError, ValueError):
            return jsonify({'status':'error',
                            'message':'alt_to_primary_ratio must be a number when alt_uom is set'}), 400
        if alt_ratio <= 0:
            return jsonify({'status':'error',
                            'message':'alt_to_primary_ratio must be greater than zero'}), 400
        if alt_uom.lower() == primary_uom.lower():
            return jsonify({'status':'error',
                            'message':'Alternate UOM must differ from primary UOM'}), 400
    if not name or not pm:
        return jsonify({'status': 'error', 'message': 'product_name and pm_type required'}), 400
    if not brand_id:
        return jsonify({'status': 'error', 'message': 'Brand is required (used to generate product code)'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        brand_name = _brand_name_by_id(conn, brand_id)
        if not brand_name:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Selected brand not found'}), 400

        # Duplicate guard (whitespace/case-insensitive). Catches "visual
        # duplicates" that differ only by spacing or case — e.g.
        # 'Pack Of 4 Box' vs 'Pack of 4 Box' vs 'Pack Of  4 Box'.
        existing = _find_duplicate_product(conn, name, pm)

        if existing:
            conn.close()
            return jsonify({
                'status':       'duplicate',
                'id':           existing['id'],
                'product_code': existing['product_code'] or '',
                'existing_name': existing['product_name'],
                'message': (f"A product with this name already exists: "
                            f"\"{existing['product_name']}\" ({existing['product_code'] or 'no code'}). "
                            "Duplicate not created.")
            }), 409

        new_code = _generate_product_code(conn, brand_name, pm)
        conn.execute(
            """INSERT INTO pm_products
                 (product_name, pm_type, brand_id, product_code,
                  primary_uom, alt_uom, alt_to_primary_ratio)
               VALUES (%s, %s, %s, %s, %s, %s, %s)""",
            (name, pm, brand_id, new_code, primary_uom, alt_uom, alt_ratio)
        )
        conn.commit()
        row = conn.execute(
            "SELECT id FROM pm_products WHERE product_name=%s AND pm_type=%s", (name, pm)
        ).fetchone()
        pid = row['id']

        # Audit trail
        try:
            _audit_record(
                conn,
                action='product.create',
                entity='product',
                entity_id=pid,
                summary=f"Created product '{name}' [{pm}] · code {new_code} · brand {brand_name}",
                before=None,
                after={'id': pid, 'product_name': name, 'pm_type': pm,
                       'brand_id': int(brand_id), 'product_code': new_code},
            )
            conn.commit()
        except Exception: pass

        conn.close()
        return jsonify({'status': 'ok', 'id': pid, 'product_code': new_code})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/godown/save', methods=['POST'])
@_login_required
def api_godown_save():
    d          = request.get_json() or {}
    product_id = d.get('product_id')
    txn_type   = d.get('txn_type')
    qty        = float(d.get('qty') or 0)
    txn_date   = d.get('txn_date') or str(date.today())
    remarks    = (d.get('remarks') or '').strip()
    godown_id  = d.get('godown_id') or None

    if not product_id or not txn_type or qty <= 0:
        return jsonify({'status': 'error', 'message': 'product_id, txn_type and qty>0 required'}), 400
    if txn_type not in ('opening', 'inward', 'outward'):
        return jsonify({'status': 'error', 'message': 'Invalid txn_type for godown'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # ── Opening is special: SET, don't ADD. If an opening row already
        # exists for this (product_id, godown_id), UPDATE it instead of
        # inserting a new one. This prevents duplicate openings inflating
        # stock when a user re-edits a product's opening qty.
        if txn_type == 'opening':
            existing = conn.execute(
                """SELECT id FROM pm_godown_txn
                   WHERE product_id=%s AND txn_type='opening'
                     AND COALESCE(godown_id, 0) = COALESCE(%s, 0)
                   ORDER BY id LIMIT 1""",
                (product_id, godown_id)
            ).fetchone()
            if existing:
                conn.execute(
                    """UPDATE pm_godown_txn
                       SET qty=%s, txn_date=%s, remarks=%s, created_by=%s
                       WHERE id=%s""",
                    (qty, txn_date, remarks or 'Opening Balance', _user(), existing['id'])
                )
                # Also dedupe — if there are stray duplicate opening rows for
                # the same (product_id, godown_id), remove them so re-edits
                # don't accumulate. Only the row we just updated survives.
                conn.execute(
                    """DELETE FROM pm_godown_txn
                       WHERE product_id=%s AND txn_type='opening'
                         AND COALESCE(godown_id, 0) = COALESCE(%s, 0)
                         AND id <> %s""",
                    (product_id, godown_id, existing['id'])
                )
                conn.commit()
                conn.close()
                return jsonify({'status': 'ok', 'mode': 'updated', 'id': existing['id']})

        # Normal flow for inward/outward (and opening when no row exists yet)
        voucher_no = _next_voucher_no(conn, 'PM-GTXN', txn_date)
        conn.execute(
            """INSERT INTO pm_godown_txn (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
            (product_id, txn_date, txn_type, qty, remarks, _user(), voucher_no, godown_id)
        )
        if txn_type == 'outward':
            fvn = _next_voucher_no(conn, 'PM-FTXN', txn_date)
            conn.execute(
                """INSERT INTO pm_floor_txn (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
                   VALUES (%s, %s, 'issue', %s, %s, %s, %s, %s)""",
                (product_id, txn_date, qty,
                 f"[Auto: Issue from Godown] {remarks}".strip(), _user(), fvn, godown_id)
            )
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'voucher_no': voucher_no})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/floor/save', methods=['POST'])
@_login_required
def api_floor_save():
    d          = request.get_json() or {}
    product_id = d.get('product_id')
    txn_type   = d.get('txn_type')
    qty        = float(d.get('qty') or 0)
    txn_date   = d.get('txn_date') or str(date.today())
    remarks    = (d.get('remarks') or '').strip()
    godown_id  = d.get('godown_id') or None

    if not product_id or not txn_type or qty <= 0:
        return jsonify({'status': 'error', 'message': 'product_id, txn_type and qty>0 required'}), 400

    valid_floor = ('floor_opening', 'issue', 'dispatch', 'rejection', 'pm_return')
    if txn_type not in valid_floor:
        return jsonify({'status': 'error', 'message': 'Invalid txn_type for floor'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # ── floor_opening is special: SET, don't ADD. (Same logic as godown
        # 'opening' — re-editing a product's opening qty must replace the
        # previous value, not pile on top of it.)
        if txn_type == 'floor_opening':
            existing = conn.execute(
                """SELECT id FROM pm_floor_txn
                   WHERE product_id=%s AND txn_type='floor_opening'
                     AND COALESCE(godown_id, 0) = COALESCE(%s, 0)
                   ORDER BY id LIMIT 1""",
                (product_id, godown_id)
            ).fetchone()
            if existing:
                conn.execute(
                    """UPDATE pm_floor_txn
                       SET qty=%s, txn_date=%s, remarks=%s, created_by=%s
                       WHERE id=%s""",
                    (qty, txn_date, remarks or 'Factory Opening Balance', _user(), existing['id'])
                )
                # Dedupe stragglers
                conn.execute(
                    """DELETE FROM pm_floor_txn
                       WHERE product_id=%s AND txn_type='floor_opening'
                         AND COALESCE(godown_id, 0) = COALESCE(%s, 0)
                         AND id <> %s""",
                    (product_id, godown_id, existing['id'])
                )
                conn.commit()
                conn.close()
                return jsonify({'status': 'ok', 'mode': 'updated', 'id': existing['id']})

        voucher_no = _next_voucher_no(conn, 'PM-FTXN', txn_date)
        conn.execute(
            """INSERT INTO pm_floor_txn (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
            (product_id, txn_date, txn_type, qty, remarks, _user(), voucher_no, godown_id)
        )
        if txn_type == 'pm_return':
            gvn = _next_voucher_no(conn, 'PM-GTXN', txn_date)
            conn.execute(
                """INSERT INTO pm_godown_txn (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
                   VALUES (%s, %s, 'inward', %s, %s, %s, %s, %s)""",
                (product_id, txn_date, qty,
                 f"[Auto: Return from Floor] {remarks}".strip(), _user(), gvn, godown_id)
            )
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'voucher_no': voucher_no})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/summary')
@_login_required
def api_pm_summary():
    from_date  = request.args.get('from_date')
    to_date    = request.args.get('to_date')
    pm_type    = request.args.get('pm_type', '').strip()
    search     = request.args.get('search', '').strip().lower()
    brand_id   = request.args.get('brand_id', '').strip()
    godown_id  = request.args.get('godown_id', '').strip() or None

    conn   = sampling_portal.get_db_connection()
    godown = _godown_summary(conn, from_date=from_date, to_date=to_date,
                             godown_id=int(godown_id) if godown_id else None)
    # Floor summary is NEVER filtered by godown_id — the factory is one unified location.
    # godown_id only applies to the godown stock view.
    floor  = _floor_summary(conn, from_date=from_date, to_date=to_date)
    conn.close()

    floor_idx = {r['id']: r for r in floor}
    combined = []
    for g in godown:
        f = floor_idx.get(g['id'], {})
        combined.append({
            'id':             g['id'],
            'product_name':   g['product_name'],
            'product_code':   g.get('product_code', ''),
            'pm_type':        g['pm_type'],
            'brand_id':       g.get('brand_id', 0),
            'brand_name':     g.get('brand_name', ''),
            'brand_color':    g.get('brand_color', ''),
            'op':             g['op'],
            'inward':         g['inward'],
            'outward':        g['outward'],
            'godown_stock':   g['godown_stock'],
            'min_stock':      g.get('min_stock', 0),
            'godown_last_txn': g.get('last_txn_date', ''),
            'floor_op':       f.get('floor_op', 0),
            'issue':          f.get('issue', 0),
            'dispatch':       f.get('dispatch', 0),
            'rejection':      f.get('rejection', 0),
            'pm_return':      f.get('pm_return', 0),
            'remaining':      f.get('remaining', 0),
            'floor_last_txn': f.get('last_txn_date', ''),
        })

    if pm_type:
        combined = [r for r in combined if r['pm_type'].lower() == pm_type.lower()]
    if search:
        # Search matches product name, PM type, OR product code (case-insensitive).
        combined = [r for r in combined
                    if search in r['product_name'].lower()
                    or search in r['pm_type'].lower()
                    or search in (r.get('product_code') or '').lower()]
    if brand_id:
        try:
            bid = int(brand_id)
            brand_prod_ids = {g['id'] for g in godown if g.get('brand_id') == bid}
            combined = [r for r in combined if r['id'] in brand_prod_ids]
        except (ValueError, TypeError):
            pass

    return jsonify(combined)


# ════════════════════════════════════════════════════════════════════════════
# PM-GTXN VOUCHER DETAIL — GET / PUT / DELETE one godown_txn row
# ────────────────────────────────────────────────────────────────────────────
# The per-product transaction ledger renders voucher numbers like
# `PG/26-27/01787` for godown_txn rows (opening / inward / outward). Before
# this endpoint, those vouchers had no detail or edit UI — admins could only
# adjust opening rows via the dedicated "Edit Opening" button at the bottom
# of the ledger modal, and inward/outward rows had no edit path at all
# (they were created as side effects of GRN / DN saves but lived in
# pm_godown_txn as standalone ledger entries that could drift out of sync
# with their parent voucher).
#
# These endpoints unlock direct admin control over those rows so admin can:
#   * GET   /api/pm_stock/godown_txn/<id> → read full row + linked context
#   * PUT   /api/pm_stock/godown_txn/<id> → edit qty / txn_date / remarks
#   * DELETE /api/pm_stock/godown_txn/<id> → remove the row entirely
#
# Restrictions:
#   * Admin only — these are write paths that bypass the usual GRN/DN flow
#     and should not be available to ordinary users.
#   * DELETE is hard-gated: the caller must pass {"confirm": "DELETE"} so a
#     misclick on a "delete" button can't wipe a row.
#   * Editing qty rewrites the qty in place. Stock totals recompute from
#     pm_godown_txn on every page load, so this just updates the ledger
#     row without touching any other table. (pm_boxes are NOT touched —
#     they're a parallel layer.)
# ════════════════════════════════════════════════════════════════════════════
@pm_stock_bp.route('/api/pm_stock/godown_txn/<int:txn_id>', methods=['GET', 'PUT', 'DELETE'])
@_login_required
def api_godown_txn_detail(txn_id):
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403

    conn = sampling_portal.get_db_connection()
    try:
        # ── GET: fetch the row + lookups ─────────────────────────────
        if request.method == 'GET':
            row = conn.execute(
                """SELECT t.id, t.product_id, t.godown_id, t.txn_type, t.qty,
                          t.txn_date, t.voucher_no, t.remarks, t.created_by,
                          p.product_name, p.product_code, p.pm_type,
                          g.name AS godown_name
                     FROM pm_godown_txn t
                     LEFT JOIN pm_products       p ON p.id = t.product_id
                     LEFT JOIN procurement_godowns g ON g.id = t.godown_id
                    WHERE t.id = %s
                    LIMIT 1""",
                (txn_id,)
            ).fetchone()
            if not row:
                conn.close()
                return jsonify({'status': 'error', 'message': 'Voucher not found'}), 404
            d = dict(row)
            # Stringify date for JSON cleanliness
            if d.get('txn_date'):
                d['txn_date'] = str(d['txn_date'])
            return jsonify({'status': 'ok', 'voucher': d})

        # ── PUT: update qty / txn_date / remarks ─────────────────────
        if request.method == 'PUT':
            body = request.get_json(silent=True) or {}
            existing = conn.execute(
                "SELECT id, qty, txn_date, remarks FROM pm_godown_txn WHERE id=%s",
                (txn_id,)
            ).fetchone()
            if not existing:
                conn.close()
                return jsonify({'status': 'error', 'message': 'Voucher not found'}), 404

            # Build the SET clause from whichever fields the client sent.
            # Missing fields keep their existing values — partial PATCH-like
            # semantics on a PUT, which is what the modal needs.
            new_qty = body.get('qty', existing['qty'] if hasattr(existing, 'get') else existing[1])
            new_date = body.get('txn_date', existing['txn_date'] if hasattr(existing, 'get') else existing[2])
            new_remarks = body.get('remarks', existing['remarks'] if hasattr(existing, 'get') else existing[3])

            try:
                new_qty = float(new_qty)
            except Exception:
                conn.close()
                return jsonify({'status': 'error', 'message': 'qty must be a number'}), 400
            if new_qty == 0:
                conn.close()
                return jsonify({'status': 'error', 'message': 'qty cannot be zero (delete the voucher instead)'}), 400

            conn.execute(
                """UPDATE pm_godown_txn
                      SET qty=%s, txn_date=%s, remarks=%s
                    WHERE id=%s""",
                (new_qty, new_date, new_remarks, txn_id)
            )
            conn.commit()
            # Audit the edit so it's visible in the audit log feed
            try:
                _audit_record(
                    conn,
                    action='godown_txn.edit',
                    entity='godown_txn',
                    entity_id=txn_id,
                    summary=f"Edited PG voucher #{txn_id}",
                    before={'qty': float(existing['qty'] if hasattr(existing, 'get') else existing[1] or 0),
                            'txn_date': str(existing['txn_date'] if hasattr(existing, 'get') else existing[2]),
                            'remarks': existing['remarks'] if hasattr(existing, 'get') else existing[3]},
                    after={'qty': new_qty, 'txn_date': str(new_date), 'remarks': new_remarks},
                )
                conn.commit()
            except Exception: pass
            conn.close()
            return jsonify({'status': 'ok', 'message': 'Voucher updated'})

        # ── DELETE: remove the row (admin only, requires explicit confirm) ──
        if request.method == 'DELETE':
            body = request.get_json(silent=True) or {}
            if body.get('confirm') != 'DELETE':
                conn.close()
                return jsonify({
                    'status':  'error',
                    'message': 'DELETE requires confirm:"DELETE" in body',
                }), 400
            existing = conn.execute(
                """SELECT id, product_id, godown_id, txn_type, qty, voucher_no, remarks
                   FROM pm_godown_txn WHERE id=%s""",
                (txn_id,)
            ).fetchone()
            if not existing:
                conn.close()
                return jsonify({'status': 'error', 'message': 'Voucher not found'}), 404
            ex_d = dict(existing) if hasattr(existing, 'keys') else None
            conn.execute("DELETE FROM pm_godown_txn WHERE id=%s", (txn_id,))
            conn.commit()
            try:
                _audit_record(
                    conn,
                    action='godown_txn.delete',
                    entity='godown_txn',
                    entity_id=txn_id,
                    summary=f"Deleted PG voucher #{txn_id} ({ex_d.get('voucher_no') if ex_d else ''})",
                    before=ex_d,
                    after=None,
                )
                conn.commit()
            except Exception: pass
            conn.close()
            return jsonify({'status': 'ok', 'message': 'Voucher deleted'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/summary/per_godown')
@_login_required
def api_pm_summary_per_godown():
    """
    Returns stock for every product broken down per godown.
    Response: { godowns: [{id,name},...], rows: [{product_id, product_name, pm_type, brand_name, total_stock, by_godown:{godown_id: stock}}] }
    """
    to_date = request.args.get('to_date') or str(date.today())
    conn    = sampling_portal.get_db_connection()
    try:
        godowns = _get_godowns(conn)
        # Only actual storage godowns (not floor/billing/shipping)
        storage = [g for g in godowns if g.get('godown_type','godown') not in ('floor','billing','shipping')]

        # Per-product per-godown stock
        rows = conn.execute("""
            SELECT
                p.id, p.product_name, p.pm_type,
                COALESCE(b.name,'') AS brand_name,
                COALESCE(b.color,'') AS brand_color,
                t.godown_id,
                COALESCE(SUM(CASE WHEN t.txn_type='opening' THEN t.qty ELSE 0 END),0) AS op,
                COALESCE(SUM(CASE WHEN t.txn_type='inward'  THEN t.qty ELSE 0 END),0) AS inward,
                COALESCE(SUM(CASE WHEN t.txn_type='outward' THEN t.qty ELSE 0 END),0) AS outward
            FROM pm_products p
            LEFT JOIN procurement_brands b ON b.id = p.brand_id
            LEFT JOIN pm_godown_txn t ON t.product_id = p.id AND t.txn_date <= %s
            WHERE p.is_active = 1
            GROUP BY p.id, p.product_name, p.pm_type, b.name, b.color, t.godown_id
            ORDER BY p.product_name
        """, (to_date,)).fetchall()

        # Build product map
        prod_map = {}  # product_id -> {meta, by_godown}
        for r in rows:
            pid = r['id']
            if pid not in prod_map:
                prod_map[pid] = {
                    'product_id':   pid,
                    'product_name': r['product_name'],
                    'pm_type':      r['pm_type'],
                    'brand_name':   r['brand_name'],
                    'brand_color':  r['brand_color'],
                    'by_godown':    {}
                }
            gid = r['godown_id']
            stock = float(r['op']) + float(r['inward']) - float(r['outward'])
            if gid is not None:
                prod_map[pid]['by_godown'][str(gid)] = round(stock, 3)
            else:
                # opening entries with NULL godown_id → add to total but not per-godown
                prod_map[pid]['by_godown']['_null'] = (
                    prod_map[pid]['by_godown'].get('_null', 0) + round(stock, 3)
                )

        # Compute total stock per product (sum across all godowns)
        for p in prod_map.values():
            p['total_godown_stock'] = round(sum(
                v for k, v in p['by_godown'].items() if k != '_null'
            ) + p['by_godown'].get('_null', 0), 3)

        conn.close()
        return jsonify({
            'godowns': [{'id': g['id'], 'name': g['name'], 'city': g.get('city','')} for g in storage],
            'rows':    list(prod_map.values())
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/factory_report')
@_login_required
def api_factory_report():
    """
    Returns factory floor transactions filtered by type.
    type: dispatch | rejection | pm_return
    """
    txn_type  = request.args.get('type', 'dispatch')
    from_date = request.args.get('from_date') or '2000-01-01'
    to_date   = request.args.get('to_date')   or str(date.today())
    pm_type   = request.args.get('pm_type', '').strip()
    search    = request.args.get('search', '').strip()
    brand_id  = request.args.get('brand_id', '').strip()

    valid_types = ('dispatch', 'rejection', 'pm_return')
    if txn_type not in valid_types:
        return jsonify({'status': 'error', 'message': 'Invalid type'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute("""
            SELECT
                t.id, t.txn_date, t.qty, t.remarks, t.voucher_no, t.created_by, t.created_at,
                p.id AS product_id, p.product_name,
                COALESCE(p.product_code,'') AS product_code,
                p.pm_type,
                COALESCE(b.id,0)    AS brand_id,
                COALESCE(b.name,'') AS brand_name,
                COALESCE(b.color,'') AS brand_color,
                COALESCE(gd.name,'') AS location
            FROM pm_floor_txn t
            JOIN pm_products p ON p.id = t.product_id
            LEFT JOIN procurement_brands b ON b.id = p.brand_id
            LEFT JOIN procurement_godowns gd ON gd.id = t.godown_id
            WHERE t.txn_type = %s
              AND t.txn_date BETWEEN %s AND %s
            ORDER BY t.txn_date DESC, t.created_at DESC
        """, (txn_type, from_date, to_date)).fetchall()

        result = []
        for r in rows:
            d = dict(r)
            for col in ('txn_date', 'created_at'):
                if d.get(col) and hasattr(d[col], 'isoformat'):
                    d[col] = d[col].isoformat() if col == 'txn_date' else str(d[col])
            result.append(d)

        # Apply filters
        if pm_type:
            result = [r for r in result if r['pm_type'].lower() == pm_type.lower()]
        if search:
            sl = search.lower()
            result = [r for r in result
                      if sl in r['product_name'].lower()
                      or sl in r['pm_type'].lower()
                      or sl in (r.get('product_code') or '').lower()]
        if brand_id:
            try:
                bid = int(brand_id)
                result = [r for r in result if r['brand_id'] == bid]
            except: pass

        # Summary totals per product
        prod_totals = {}
        for r in result:
            pid = r['product_id']
            if pid not in prod_totals:
                prod_totals[pid] = {'product_name': r['product_name'], 'pm_type': r['pm_type'],
                                    'brand_name': r['brand_name'], 'total_qty': 0, 'count': 0}
            prod_totals[pid]['total_qty'] += float(r['qty'])
            prod_totals[pid]['count']     += 1

        conn.close()
        return jsonify({
            'rows':         result,
            'prod_totals':  list(prod_totals.values()),
            'grand_total':  sum(float(r['qty']) for r in result),
            'type':         txn_type
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════
# REPORTS HUB — printable reports opened from the Reports modal.
#   1) Stock Item Ledger     — one product, all movements + running balance
#   2) Detailed GRN Summary  — GRNs in a period with items/supplier
# Default period (client-side): current month → today.
# ════════════════════════════════════════════════════════════════════════

@pm_stock_bp.route('/api/pm_stock/reports/item_ledger')
@_login_required
def api_report_item_ledger():
    """Stock movement ledger for ONE product over a period, with a running
    balance. Combines godown + floor ledgers (the same tables stock totals
    are derived from), plus an opening balance carried from before the
    period start so the running balance is correct.

    Query: product_id (required), from_date, to_date.
    Sign convention: opening/inward/floor_opening/pm_return = +qty (into the
    system / returned), outward/issue/dispatch/rejection = -qty (out).
    """
    pid = request.args.get('product_id', '').strip()
    if not pid.isdigit():
        return jsonify({'status': 'error', 'message': 'product_id required'}), 400
    pid = int(pid)
    from_date = request.args.get('from_date') or str(date.today().replace(day=1))
    to_date   = request.args.get('to_date')   or str(date.today())

    # txn_types that ADD to stock vs REMOVE from stock.
    PLUS  = ('opening', 'inward', 'floor_opening', 'pm_return')
    MINUS = ('outward', 'issue', 'dispatch', 'rejection')

    conn = sampling_portal.get_db_connection()
    try:
        prod = conn.execute(
            "SELECT id, product_name, pm_type, COALESCE(product_code,'') AS product_code FROM pm_products WHERE id=%s",
            (pid,)
        ).fetchone()
        if not prod:
            conn.close(); return jsonify({'status': 'error', 'message': 'Product not found'}), 404

        # Opening balance = net of everything strictly BEFORE from_date.
        # NOTE: Reports exclude the FLOOR/factory ledger entirely — only the
        # godown ledger (pm_godown_txn) is counted, per requirement.
        def _net_before(table):
            row = conn.execute(
                f"""SELECT
                      COALESCE(SUM(CASE WHEN txn_type IN ('opening','inward','floor_opening','pm_return') THEN qty
                                        WHEN txn_type IN ('outward','issue','dispatch','rejection') THEN -qty
                                        ELSE 0 END), 0) AS net
                    FROM {table}
                    WHERE product_id=%s AND txn_date < %s""",
                (pid, from_date)
            ).fetchone()
            return float(row['net'] or 0)
        opening_balance = _net_before('pm_godown_txn')

        # All movements within the period — GODOWN ledger only (floor excluded).
        rows = conn.execute(
            """
            SELECT txn_date, txn_type, qty, voucher_no, godown_id, remarks, created_at, created_by, 'godown' AS ledger
              FROM pm_godown_txn WHERE product_id=%s AND txn_date BETWEEN %s AND %s
            ORDER BY txn_date ASC, created_at ASC
            """,
            (pid, from_date, to_date)
        ).fetchall()

        # Resolve godown names once.
        gids = sorted({r['godown_id'] for r in rows if r['godown_id']})
        gname = {}
        if gids:
            placeholders = ','.join(['%s'] * len(gids))
            for g in conn.execute(f"SELECT id, name FROM procurement_godowns WHERE id IN ({placeholders})", tuple(gids)).fetchall():
                gname[g['id']] = g['name']
        conn.close()

        TYPE_LABEL = {
            'opening':'Opening', 'inward':'Inward (GRN)', 'outward':'Outward (Transfer)',
            'floor_opening':'Floor Opening', 'issue':'Issue', 'dispatch':'Dispatch',
            'rejection':'Rejection', 'pm_return':'PM Return',
        }
        running = opening_balance
        out = []
        total_in = 0.0; total_out = 0.0
        for r in rows:
            q = float(r['qty'] or 0)
            is_in = r['txn_type'] in PLUS
            signed = q if is_in else -q
            running += signed
            if is_in: total_in += q
            else:     total_out += q
            out.append({
                'date':       str(r['txn_date']),
                'type':       r['txn_type'],
                'type_label': TYPE_LABEL.get(r['txn_type'], r['txn_type']),
                'in_qty':     q if is_in else 0,
                'out_qty':    q if not is_in else 0,
                'balance':    round(running, 2),
                'voucher_no': r['voucher_no'] or '',
                'location':   gname.get(r['godown_id'], '') if r['godown_id'] else ('Factory' if r['ledger']=='floor' else ''),
                'remarks':    r['remarks'] or '',
                'by':         r['created_by'] or '',
            })
        return jsonify({
            'status': 'ok',
            'product': {'id': prod['id'], 'name': prod['product_name'],
                        'pm_type': prod['pm_type'], 'code': prod['product_code']},
            'from_date': from_date, 'to_date': to_date,
            'opening_balance': round(opening_balance, 2),
            'closing_balance': round(running, 2),
            'total_in': round(total_in, 2), 'total_out': round(total_out, 2),
            'rows': out, 'count': len(out),
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/reports/grn_summary')
@_login_required
def api_report_grn_summary():
    """Detailed GRN summary for a period: each GRN with its supplier, date,
    and line items (product name + qty). Query: from_date, to_date, search
    (optional, matches grn_no/supplier).
    """
    from_date = request.args.get('from_date') or str(date.today().replace(day=1))
    to_date   = request.args.get('to_date')   or str(date.today())
    search    = (request.args.get('search') or '').strip()

    conn = sampling_portal.get_db_connection()
    try:
        where = ["g.grn_date BETWEEN %s AND %s"]
        params = [from_date, to_date]
        if search:
            where.append("(g.grn_no LIKE %s OR g.supplier LIKE %s)")
            params += [f'%{search}%', f'%{search}%']
        wc = "WHERE " + " AND ".join(where)
        grns = conn.execute(
            f"""SELECT g.id, g.grn_no, g.grn_date, g.supplier, g.po_number,
                       COALESCE(gd.name,'') AS godown_name,
                       COALESCE(g.verification_status,'verified') AS verification_status
                FROM pm_grn g
                LEFT JOIN procurement_godowns gd ON gd.id = g.godown_id
                {wc}
                ORDER BY g.grn_date DESC, g.id DESC""",
            tuple(params)
        ).fetchall()

        gids = [g['id'] for g in grns]
        items_by_grn = {}
        if gids:
            placeholders = ','.join(['%s'] * len(gids))
            irows = conn.execute(
                f"""SELECT gi.grn_id, gi.qty_received, gi.no_of_box,
                           COALESCE(p.product_name,'') AS product_name,
                           COALESCE(p.pm_type,'') AS pm_type
                    FROM pm_grn_items gi
                    LEFT JOIN pm_products p ON p.id = gi.product_id
                    WHERE gi.grn_id IN ({placeholders})
                    ORDER BY p.product_name""",
                tuple(gids)
            ).fetchall()
            for it in irows:
                items_by_grn.setdefault(it['grn_id'], []).append({
                    'product_name': it['product_name'], 'pm_type': it['pm_type'],
                    'qty_received': float(it['qty_received'] or 0),
                    'no_of_box': int(it['no_of_box'] or 0),
                })
        conn.close()

        out = []
        grand_qty = 0.0; grand_items = 0
        for g in grns:
            items = items_by_grn.get(g['id'], [])
            gqty = sum(i['qty_received'] for i in items)
            grand_qty += gqty; grand_items += len(items)
            out.append({
                'id': g['id'], 'grn_no': g['grn_no'], 'grn_date': str(g['grn_date']),
                'supplier': g['supplier'] or '', 'po_number': g['po_number'] or '',
                'godown_name': g['godown_name'], 'status': g['verification_status'],
                'items': items, 'item_count': len(items), 'total_qty': round(gqty, 2),
            })
        return jsonify({
            'status': 'ok', 'from_date': from_date, 'to_date': to_date,
            'grns': out, 'grn_count': len(out),
            'grand_qty': round(grand_qty, 2), 'grand_items': grand_items,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transactions')
@_login_required
def api_pm_transactions():
    from_date    = request.args.get('from_date') or '2000-01-01'
    to_date      = request.args.get('to_date')   or str(date.today())
    pm_type      = request.args.get('pm_type', '').strip()
    search       = request.args.get('search', '').strip()
    source       = request.args.get('source', 'all')   # all | grn | mtv | godown | floor
    # 'godown' and 'floor' are UI-side labels; the report needs all voucher types
    if source in ('godown', 'floor'):
        source = 'all'
    product_id   = request.args.get('product_id', '').strip()
    godown_id    = request.args.get('godown_id', '').strip()

    conn = sampling_portal.get_db_connection()
    rows = []

    # ── GRN vouchers ──────────────────────────────────────────────────────────
    if source in ('all', 'grn'):
        pid_filter = 'AND gi.product_id = %s' if product_id else ''
        gdn_filter = 'AND g.godown_id = %s'   if godown_id  else ''
        params_g = [from_date, to_date]
        if product_id: params_g.append(int(product_id))
        if godown_id:  params_g.append(int(godown_id))
        grn_rows = conn.execute(f"""
            SELECT gi.id, gi.product_id, p.product_name, p.pm_type,
                   COALESCE(p.product_code,'') AS product_code,
                   g.grn_date AS txn_date,
                   g.grn_no   AS voucher_no,
                   gi.qty_received AS qty,
                   COALESCE(g.supplier,'') AS supplier,
                   COALESCE(g.remarks,'')  AS remarks,
                   COALESCE(g.party_invoice_no,'') AS party_invoice_no,
                   g.created_by,
                   g.created_at,
                   g.godown_id,
                   COALESCE(gd.name,'') AS godown_name,
                   COALESCE(g.supplier,'') AS from_location,
                   COALESCE(gd.name,'')    AS to_location,
                   'inward' AS txn_type,
                   'grn' AS source,
                   g.id AS voucher_id
            FROM pm_grn_items gi
            JOIN pm_grn g ON g.id = gi.grn_id
            JOIN pm_products p ON p.id = gi.product_id
            LEFT JOIN procurement_godowns gd ON gd.id = g.godown_id
            WHERE g.grn_date BETWEEN %s AND %s {pid_filter} {gdn_filter}
            ORDER BY g.grn_date DESC, g.created_at DESC
        """, params_g).fetchall()
        rows += [dict(r) for r in grn_rows]

    # MTV vouchers removed — all material transfer is now in pm_transfers.
    # Existing godown txns posted by the new transfer flow are tagged
    # [PM-MT:...] in remarks and show up in the godown branch below.
    if source in ('all', 'mtv'):
        pid_filter = 'AND mi.product_id = %s' if product_id else ''
        if godown_id:
            gdn_filter = 'AND (m.from_godown = %s OR m.to_godown = %s)'
        else:
            gdn_filter = ''
        params_m = [from_date, to_date]
        if product_id: params_m.append(int(product_id))
        if godown_id:  params_m += [int(godown_id), int(godown_id)]
        mtv_rows = conn.execute(f"""
            SELECT mi.id, mi.product_id, p.product_name, p.pm_type,
                   COALESCE(p.product_code,'') AS product_code,
                   m.mtv_date AS txn_date,
                   m.mtv_no   AS voucher_no,
                   mi.qty,
                   m.remarks,
                   m.created_by,
                   m.created_at,
                   m.from_godown AS godown_id,
                   COALESCE(gf.name,'') AS from_location,
                   COALESCE(gt.name,'') AS to_location,
                   COALESCE(gf.name,'') AS godown_name,
                   'transfer' AS txn_type,
                   'mtv' AS source,
                   m.id AS voucher_id
            FROM pm_mtv_items mi
            JOIN pm_mtv m ON m.id = mi.mtv_id
            JOIN pm_products p ON p.id = mi.product_id
            LEFT JOIN procurement_godowns gf ON gf.id = m.from_godown
            LEFT JOIN procurement_godowns gt ON gt.id = m.to_godown
            WHERE m.mtv_date BETWEEN %s AND %s {pid_filter} {gdn_filter}
            ORDER BY m.mtv_date DESC, m.created_at DESC
        """, params_m).fetchall()
        rows += [dict(r) for r in mtv_rows]

    # ── Direct godown transactions (opening / inward / outward) ──
    # Exclude rows auto-created by MTV ([MTV:] or [Auto:]) AND
    # rows auto-created by GRN ([PM GRN:] or [GRN:]) — those are
    # already shown via the GRN source to avoid double-counting.
    if source in ('all', 'godown'):
        pid_filter = 'AND t.product_id = %s' if product_id else ''
        gdn_filter = 'AND t.godown_id = %s'  if godown_id  else ''
        params_gd = [from_date, to_date]
        if product_id: params_gd.append(int(product_id))
        if godown_id:  params_gd.append(int(godown_id))
        gd_rows = conn.execute(f"""
            SELECT t.id, t.product_id, p.product_name, p.pm_type,
                   COALESCE(p.product_code,'') AS product_code,
                   t.txn_date,
                   t.voucher_no,
                   t.qty,
                   t.remarks,
                   t.created_by,
                   t.created_at,
                   t.godown_id,
                   COALESCE(gd.name,'') AS godown_name,
                   '' AS from_location,
                   COALESCE(gd.name,'') AS to_location,
                   t.txn_type,
                   'godown' AS source,
                   t.id AS voucher_id
            FROM pm_godown_txn t
            JOIN pm_products p ON p.id = t.product_id
            LEFT JOIN procurement_godowns gd ON gd.id = t.godown_id
            WHERE t.txn_date BETWEEN %s AND %s {pid_filter} {gdn_filter}
              AND (t.remarks IS NULL OR (
                   t.remarks NOT LIKE '%%[MTV:%%'
               AND t.remarks NOT LIKE '%%[Auto:%%'
               AND t.remarks NOT LIKE '%%[PM GRN:%%'
               AND t.remarks NOT LIKE '%%[GRN:%%'
              ))
            ORDER BY t.txn_date DESC, t.created_at DESC
        """, params_gd).fetchall()
        rows += [dict(r) for r in gd_rows]

    # ── Direct floor/factory transactions (issue / dispatch / rejection / pm_return) ──
    # Exclude rows auto-created by MTV (remarks contain [MTV:] or [Auto:]) —
    # those are already shown via the MTV source.
    if source in ('all', 'floor'):
        pid_filter = 'AND t.product_id = %s' if product_id else ''
        gdn_filter = 'AND t.godown_id = %s'  if godown_id  else ''
        params_fl = [from_date, to_date]
        if product_id: params_fl.append(int(product_id))
        if godown_id:  params_fl.append(int(godown_id))
        fl_rows = conn.execute(f"""
            SELECT t.id, t.product_id, p.product_name, p.pm_type,
                   COALESCE(p.product_code,'') AS product_code,
                   t.txn_date,
                   t.voucher_no,
                   t.qty,
                   t.remarks,
                   t.created_by,
                   t.created_at,
                   t.godown_id,
                   COALESCE(gd.name,'') AS godown_name,
                   '' AS from_location,
                   COALESCE(gd.name,'') AS to_location,
                   t.txn_type,
                   'floor' AS source,
                   t.id AS voucher_id
            FROM pm_floor_txn t
            JOIN pm_products p ON p.id = t.product_id
            LEFT JOIN procurement_godowns gd ON gd.id = t.godown_id
            WHERE t.txn_date BETWEEN %s AND %s {pid_filter} {gdn_filter}
              AND (t.remarks IS NULL OR (t.remarks NOT LIKE '%%[MTV:%%' AND t.remarks NOT LIKE '%%[Auto:%%' AND t.remarks NOT LIKE '%%[Repaired]%%'))
            ORDER BY t.txn_date DESC, t.created_at DESC
        """, params_fl).fetchall()
        rows += [dict(r) for r in fl_rows]

    conn.close()

    # Normalise dates
    for r in rows:
        for col in ('txn_date', 'created_at'):
            v = r.get(col)
            if v and hasattr(v, 'isoformat'):
                r[col] = v.isoformat() if col == 'txn_date' else str(v)
        # qty field normalisation (grn uses qty_received alias, mtv uses qty)
        if 'qty_received' in r:
            r['qty'] = r.pop('qty_received')

    # Filters
    if pm_type:
        rows = [r for r in rows if r['pm_type'].lower() == pm_type.lower()]
    if search:
        s = search.lower()
        rows = [r for r in rows if
                s in (r.get('product_name')  or '').lower() or
                s in (r.get('pm_type')       or '').lower() or
                s in (r.get('product_code')  or '').lower() or
                s in (r.get('voucher_no')    or '').lower() or
                s in (r.get('godown_name')   or '').lower()]

    rows.sort(key=lambda r: (r.get('txn_date', ''), r.get('created_at', '')), reverse=True)
    return jsonify(rows)

# ── Version / Deploy Health-Check ─────────────────────────────────────────
# Used to verify that an updated __init__.py is actually live on the server
# AND that helpers.py migrations ran successfully. After every deploy,
# visit this URL in a browser to confirm:
#   1. PM_STOCK_VERSION below matches the version you expect (file uploaded)
#   2. Flask reloaded (response reflects current code, not cached process)
#   3. Migrations applied (checks return true for expected columns/tables)
#
# Bumping convention: YYYY-MM-DD-shortname. Bump every deploy that touches
# __init__.py so cache vs. live mismatches surface immediately.
PM_STOCK_VERSION = "2026-05-30-phase-a-purchase-orders-fix1"

@pm_stock_bp.route('/api/pm_stock/_version')
def api_pm_version():
    """Public — no login required, so it can be checked even on auth issues."""
    checks = {}
    try:
        conn = sampling_portal.get_db_connection()
        # Does pm_purchase_orders table exist?
        try:
            conn.execute("SELECT 1 FROM pm_purchase_orders LIMIT 1").fetchone()
            checks['table_pm_purchase_orders'] = True
        except Exception:
            checks['table_pm_purchase_orders'] = False
        # Does pm_po_items table exist?
        try:
            conn.execute("SELECT 1 FROM pm_po_items LIMIT 1").fetchone()
            checks['table_pm_po_items'] = True
        except Exception:
            checks['table_pm_po_items'] = False
        # Does pm_grn.po_id column exist?
        try:
            conn.execute("SELECT po_id FROM pm_grn LIMIT 1").fetchone()
            checks['column_pm_grn_po_id'] = True
        except Exception:
            checks['column_pm_grn_po_id'] = False
        # UOM Phase 3A columns on MR items
        try:
            conn.execute("SELECT entered_uom, entered_qty FROM pm_material_request_items LIMIT 1").fetchone()
            checks['column_pm_mr_items_entered_uom'] = True
        except Exception:
            checks['column_pm_mr_items_entered_uom'] = False
        # UOM Phase 2B column on GRN items
        try:
            conn.execute("SELECT entered_uom FROM pm_grn_items LIMIT 1").fetchone()
            checks['column_pm_grn_items_entered_uom'] = True
        except Exception:
            checks['column_pm_grn_items_entered_uom'] = False
        # UOM Phase 1 columns on products
        try:
            conn.execute("SELECT primary_uom, alt_uom, alt_to_primary_ratio FROM pm_products LIMIT 1").fetchone()
            checks['column_pm_products_uom'] = True
        except Exception:
            checks['column_pm_products_uom'] = False
        conn.close()
    except Exception as e:
        return jsonify({
            'status': 'error',
            'version': PM_STOCK_VERSION,
            'message': f'DB connection failed: {e}',
            'checks': checks,
        }), 500
    all_ok = all(checks.values())
    return jsonify({
        'status': 'ok' if all_ok else 'partial',
        'version': PM_STOCK_VERSION,
        'all_migrations_ok': all_ok,
        'checks': checks,
        'hint': (
            'All checks pass — code + migrations are live.' if all_ok
            else 'Some migrations have NOT run. Either Flask hasn\'t restarted, OR helpers.py wasn\'t deployed. Check the false items above — they tell you which migrations are missing.'
        ),
    })


@pm_stock_bp.route('/api/pm_stock/pm_types')
@_login_required
def api_pm_types():
    conn  = sampling_portal.get_db_connection()
    rows  = conn.execute(
        "SELECT DISTINCT pm_type FROM pm_products WHERE is_active=1 ORDER BY pm_type"
    ).fetchall()
    conn.close()
    return jsonify([r['pm_type'] for r in rows])


@pm_stock_bp.route('/api/pm_stock/grn/<int:grn_id>/heal_missing_boxes', methods=['POST'])
@_login_required
def api_grn_heal_missing_boxes(grn_id):
    """
    Heal a GRN that has missing pm_boxes rows due to the old box-code collision
    bug. Before the fix, when two GRN items shared the same product, the second
    item's boxes would restart box_seq at 1 and silently fail to insert because
    of the UNIQUE constraint on box_code. This endpoint scans the GRN and
    creates any missing pm_boxes + pm_box_movements rows so each physical box
    becomes scannable.

    Admin only — non-destructive. Re-runnable. Returns a per-product summary
    of what was healed.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403

    conn = sampling_portal.get_db_connection()
    try:
        grn = conn.execute(
            "SELECT id, grn_no, godown_id, grn_date FROM pm_grn WHERE id=%s",
            (grn_id,)
        ).fetchone()
        if not grn:
            conn.close(); return jsonify({'status':'error','message':'GRN not found'}), 404

        grn_no    = grn['grn_no']
        godown_id = grn['godown_id']
        # Items in entry order — defines which per_box_qty maps to which seq range
        items = conn.execute(
            """SELECT id AS grn_item_id, product_id, no_of_box, box_count
               FROM pm_grn_items
               WHERE grn_id=%s
               ORDER BY id ASC""",
            (grn_id,)
        ).fetchall()

        # Group items by product so we can compute expected vs actual per product
        by_prod = {}
        for it in items:
            pid = it['product_id']
            by_prod.setdefault(pid, []).append(dict(it))

        summary = []
        total_created = 0
        for pid, prod_items in by_prod.items():
            expected = sum(int(x['no_of_box'] or 0) for x in prod_items)
            existing = conn.execute(
                """SELECT box_id, box_seq, per_box_qty
                   FROM pm_boxes
                   WHERE grn_no=%s AND product_id=%s
                   ORDER BY box_seq""",
                (grn_no, pid)
            ).fetchall()
            actual = len(existing)
            if actual >= expected:
                # Nothing missing for this product
                summary.append({'product_id': pid, 'expected': expected, 'before': actual,
                                'created': 0, 'after': actual})
                continue

            # We need to create (expected - actual) more boxes. The first
            # `actual` "expected slots" are already filled (in entry order
            # across items). Slot indices > actual map back to (item, offset)
            # within that item.
            pcode_row = conn.execute(
                "SELECT COALESCE(product_code,'') AS pc FROM pm_products WHERE id=%s", (pid,)
            ).fetchone()
            product_code = (pcode_row['pc'] if pcode_row else '') or ''

            # Build slot map: ordered (per_box_qty, grn_item_id) tuples, one per expected box.
            slot_map = []
            for it in prod_items:
                for _ in range(int(it['no_of_box'] or 0)):
                    slot_map.append({'per_box_qty': float(it['box_count'] or 0),
                                     'grn_item_id': it['grn_item_id']})

            # Slots 0..actual-1 are already there. Create slots actual..expected-1.
            next_seq = (max((int(b['box_seq'] or 0) for b in existing), default=0)) + 1
            created_here = 0
            for slot_idx in range(actual, expected):
                slot = slot_map[slot_idx]
                seq  = next_seq + (slot_idx - actual)
                box_code = _make_box_code(product_code, grn_no, seq)
                # Defensive: skip if somehow already there
                exists = conn.execute(
                    "SELECT box_id FROM pm_boxes WHERE box_code=%s LIMIT 1", (box_code,)
                ).fetchone()
                if exists:
                    continue
                cur = conn.execute(
                    """INSERT INTO pm_boxes
                         (box_code, grn_id, grn_no, grn_item_id, product_id, product_code,
                          box_seq, total_boxes, per_box_qty, current_godown_id, current_status,
                          created_by)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'in_stock',%s)""",
                    (box_code, grn_id, grn_no, slot['grn_item_id'], pid, product_code,
                     seq, expected, float(slot['per_box_qty']), godown_id, _user())
                )
                box_id = cur.lastrowid
                # Compact 8-char sequential short_code for the QR payload
                # (e.g. A0000001). Idempotent — _assign_box_short_code
                # returns any existing code unchanged if somehow already set.
                _assign_box_short_code(conn, box_id)
                conn.execute(
                    """INSERT INTO pm_box_movements
                         (box_id, movement_type, to_godown_id, qty, moved_by, remarks)
                       VALUES (%s, 'grn_create', %s, %s, %s, %s)""",
                    (box_id, godown_id, float(slot['per_box_qty']), _user(),
                     f'GRN {grn_no} box {seq}/{expected} (healed)')
                )
                created_here += 1

            # Backfill total_boxes on ALL rows for this product so denominators line up
            conn.execute(
                "UPDATE pm_boxes SET total_boxes=%s WHERE grn_no=%s AND product_id=%s",
                (expected, grn_no, pid)
            )
            total_created += created_here
            summary.append({'product_id': pid, 'expected': expected, 'before': actual,
                            'created': created_here, 'after': actual + created_here})

        conn.commit()
        conn.close()
        return jsonify({
            'status':'ok',
            'grn_no':         grn_no,
            'total_created':  total_created,
            'per_product':    summary,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/labels/summary', methods=['GET'])
@_login_required
def api_labels_summary():
    """
    Summary of label-print activity within a date range. Used by the Daily
    Report (and the Daily Work Report WhatsApp share) to show how many
    labels were printed for the chosen day, broken down by kind
    (fresh GRN vs reprints vs OP vs DN vs manual).

    Query params:
      from_date, to_date — inclusive YYYY-MM-DD bounds (both required)

    `total_labels` is the SUM of label_count across rows in the range;
    `total_events` is the COUNT(*) of print events (each event may print
    1 to N labels). Both useful: labels = how much paper used;
    events = how many times the print button was hit.
    """
    from_date = (request.args.get('from_date') or '').strip()
    to_date   = (request.args.get('to_date')   or '').strip()
    if not from_date or not to_date:
        return jsonify({'status':'error','message':'from_date and to_date required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(
            """SELECT print_kind,
                      COUNT(*)                      AS event_count,
                      COALESCE(SUM(label_count), 0) AS label_sum
               FROM pm_label_print_log
               WHERE DATE(ts) >= %s AND DATE(ts) <= %s
               GROUP BY print_kind""",
            (from_date, to_date)
        ).fetchall()

        kinds = ['grn_fresh','grn_reprint','grn_selective',
                 'op_fresh','op_reprint','op_selective',
                 'dn_fresh','manual']
        by_kind    = {k: 0 for k in kinds}
        ev_by_kind = {k: 0 for k in kinds}
        total_labels = 0
        total_events = 0
        for r in rows:
            k = r['print_kind']
            n = int(r['label_sum'] or 0)
            e = int(r['event_count'] or 0)
            if k in by_kind:
                by_kind[k]    += n
                ev_by_kind[k] += e
            total_labels += n
            total_events += e
        conn.close()
        return jsonify({
            'status':'ok',
            'from_date': from_date,
            'to_date':   to_date,
            'total_labels': total_labels,
            'total_events': total_events,
            'by_kind':       by_kind,
            'events_by_kind': ev_by_kind,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/grn/<int:grn_id>/resync_box_qty', methods=['POST'])
@_login_required
def api_grn_resync_box_qty(grn_id):
    """
    Repair: re-sync pm_boxes.per_box_qty to the GRN line each box belongs to.

    Fixes the "120-box scans as 170" bug, which happens when a product has TWO
    GRN lines with different per-box quantities (e.g. 35×170 + 1×120) and boxes
    created by older code got the wrong per_box_qty.

    The correct per-box qty for a line is derived: qty_received / no_of_box
    (pm_grn_items has no per_box_qty column). Each box is matched to its line
    via pm_boxes.grn_item_id (NOT NULL by schema).

    SAFETY (this writes to live stock data):
      • Default is DRY-RUN — returns what WOULD change, changes nothing.
        Pass {"apply": true} to actually write.
      • Only boxes whose grn_item_id points to a valid line are touched.
      • A box whose line is AMBIGUOUS (line has no_of_box<=0 so per-box can't be
        derived) is reported as 'skipped', never guessed.
      • Only 'in_stock' boxes are corrected by default (history of
        consumed/dispatched boxes is left intact) unless {"include_moved": true}.
      • Every change is logged with before→after.

    Originally admin-only; opened up so any logged-in user can run the
    repair when they spot a wrong per-box scan reading. The dry-run
    default + audit-trail of every change keep it safe.
    """
    body = request.get_json(silent=True) or {}
    apply_changes = bool(body.get('apply'))
    include_moved = bool(body.get('include_moved'))
    conn = sampling_portal.get_db_connection()
    try:
        grn = conn.execute(
            "SELECT id, grn_no FROM pm_grn WHERE id=%s", (grn_id,)
        ).fetchone()
        if not grn:
            conn.close(); return jsonify({'status':'error','message':'GRN not found'}), 404

        # Correct per-box qty per line = qty_received / no_of_box.
        lines = conn.execute(
            """SELECT id AS grn_item_id, product_id, qty_received, no_of_box
                 FROM pm_grn_items WHERE grn_id=%s""",
            (grn_id,)
        ).fetchall()
        line_pbq = {}      # grn_item_id -> correct per_box_qty (or None if ambiguous)
        for l in lines:
            nob = int(l['no_of_box'] or 0)
            line_pbq[int(l['grn_item_id'])] = (round(float(l['qty_received'] or 0) / nob, 3)
                                               if nob > 0 else None)

        # All boxes for this GRN.
        status_clause = "" if include_moved else "AND b.current_status='in_stock'"
        boxes = conn.execute(
            f"""SELECT b.box_id, b.box_code, b.short_code, b.grn_item_id,
                       b.per_box_qty, b.current_status
                  FROM pm_boxes b
                 WHERE b.grn_id=%s {status_clause}
                 ORDER BY b.box_seq""",
            (grn_id,)
        ).fetchall()

        to_fix = []; skipped = []; ok = 0
        for b in boxes:
            gii = int(b['grn_item_id'] or 0)
            correct = line_pbq.get(gii)
            cur = float(b['per_box_qty'] or 0)
            if gii not in line_pbq:
                skipped.append({'box_code': b['box_code'], 'short_code': b['short_code'],
                                'reason': 'box.grn_item_id does not match any line on this GRN',
                                'current_per_box_qty': cur})
                continue
            if correct is None:
                skipped.append({'box_code': b['box_code'], 'short_code': b['short_code'],
                                'reason': 'line has no_of_box<=0; per-box qty cannot be derived',
                                'current_per_box_qty': cur})
                continue
            if abs(cur - correct) > 0.0005:
                to_fix.append({'box_id': b['box_id'], 'box_code': b['box_code'],
                               'short_code': b['short_code'], 'status': b['current_status'],
                               'from': cur, 'to': correct})
            else:
                ok += 1

        applied = 0
        if apply_changes and to_fix:
            for f in to_fix:
                conn.execute(
                    "UPDATE pm_boxes SET per_box_qty=%s WHERE box_id=%s",
                    (f['to'], f['box_id'])
                )
                applied += 1
            conn.commit()

        conn.close()
        return jsonify({
            'status':       'ok',
            'mode':         'applied' if apply_changes else 'dry_run',
            'grn_no':       grn['grn_no'],
            'boxes_checked': len(boxes),
            'already_correct': ok,
            'would_fix' if not apply_changes else 'fixed': to_fix,
            'fix_count':    len(to_fix),
            'applied_count': applied,
            'skipped':      skipped,          # boxes we refused to guess on
            'skipped_count': len(skipped),
            'note': ('Dry run — nothing changed. Re-POST with {"apply":true} to write.'
                     if not apply_changes else 'Changes written to pm_boxes.per_box_qty.'),
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/admin/resync_box_qty_all', methods=['POST'])
@_login_required
def api_admin_resync_box_qty_all():
    """ONE-TIME GLOBAL BACKFILL: resync per_box_qty across every GRN in the DB.

    Loops every GRN that has at least one pm_boxes row, runs the same dry-run /
    apply logic as the per-GRN endpoint, and rolls the counts up. Useful as a
    one-shot after upgrading to the version that enforces the
    `box_count = qty_received / no_of_box` invariant on save — boxes saved
    before that fix may still carry the wrong per_box_qty, and this endpoint
    sweeps them in one go without the operator having to open each GRN by hand.

    SAFETY:
      • Defaults to DRY-RUN. Pass {"apply": true} to actually write.
      • Only touches 'in_stock' boxes by default (history of consumed/dispatched
        boxes preserved). Pass {"include_moved": true} to widen the sweep.
      • Errors per GRN don't abort the sweep — partial progress sticks.
      • Returns a summary plus a per-GRN breakdown so the admin can see exactly
        which vouchers were affected.

    Admin only.
    """
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    body = request.get_json(silent=True) or {}
    apply_changes = bool(body.get('apply'))
    include_moved = bool(body.get('include_moved'))

    conn = sampling_portal.get_db_connection()
    try:
        # Pull every distinct grn_id that has at least one box. Cheap because
        # pm_boxes.grn_id has an index for this lookup.
        rows = conn.execute(
            "SELECT DISTINCT grn_id FROM pm_boxes WHERE grn_id IS NOT NULL AND grn_id > 0 "
            "ORDER BY grn_id"
        ).fetchall() or []
        grn_ids = [int(r['grn_id'] if hasattr(r, 'get') else r[0]) for r in rows]

        totals = {
            'grns_total':    len(grn_ids),
            'grns_clean':    0,         # GRNs with nothing to fix
            'grns_changed':  0,         # GRNs with at least one box to fix
            'grns_skipped':  0,         # GRNs where some boxes couldn't be classified
            'grns_failed':   0,         # GRNs that errored mid-resync
            'boxes_checked': 0,
            'boxes_already_correct': 0,
            'boxes_to_fix':  0,
            'boxes_applied': 0,
            'boxes_skipped': 0,
        }
        per_grn = []
        last_err = None

        for gid in grn_ids:
            try:
                # Re-fetch lines + boxes per GRN (single GRN at a time keeps
                # memory usage bounded even on a huge warehouse).
                grn = conn.execute(
                    "SELECT id, grn_no FROM pm_grn WHERE id=%s", (gid,)
                ).fetchone()
                if not grn:
                    continue
                grn_no = grn['grn_no'] if hasattr(grn, 'get') else grn[1]

                lines = conn.execute(
                    "SELECT id AS grn_item_id, product_id, qty_received, no_of_box "
                    "FROM pm_grn_items WHERE grn_id=%s",
                    (gid,)
                ).fetchall() or []
                line_pbq = {}
                for l in lines:
                    ld = dict(l) if hasattr(l, 'keys') else l
                    nob_l = int(ld.get('no_of_box') or 0)
                    line_pbq[int(ld['grn_item_id'])] = (
                        round(float(ld.get('qty_received') or 0) / nob_l, 3)
                        if nob_l > 0 else None
                    )

                status_clause = "" if include_moved else "AND current_status='in_stock'"
                boxes = conn.execute(
                    f"""SELECT box_id, box_code, short_code, grn_item_id, per_box_qty, current_status
                        FROM pm_boxes
                        WHERE grn_id=%s {status_clause}""",
                    (gid,)
                ).fetchall() or []

                grn_fix     = 0
                grn_ok      = 0
                grn_skip    = 0
                grn_applied = 0
                for b in boxes:
                    bd = dict(b) if hasattr(b, 'keys') else b
                    gii = int(bd.get('grn_item_id') or 0)
                    cur_pbq = float(bd.get('per_box_qty') or 0)
                    if gii not in line_pbq:
                        grn_skip += 1
                        continue
                    correct = line_pbq[gii]
                    if correct is None:
                        grn_skip += 1
                        continue
                    if abs(cur_pbq - correct) > 0.0005:
                        grn_fix += 1
                        if apply_changes:
                            conn.execute(
                                "UPDATE pm_boxes SET per_box_qty=%s WHERE box_id=%s",
                                (correct, int(bd['box_id']))
                            )
                            grn_applied += 1
                    else:
                        grn_ok += 1
                if apply_changes and grn_applied:
                    conn.commit()

                totals['boxes_checked']         += len(boxes)
                totals['boxes_already_correct'] += grn_ok
                totals['boxes_to_fix']          += grn_fix
                totals['boxes_applied']         += grn_applied
                totals['boxes_skipped']         += grn_skip
                if grn_fix > 0:
                    totals['grns_changed'] += 1
                else:
                    totals['grns_clean'] += 1
                if grn_skip > 0:
                    totals['grns_skipped'] += 1

                if grn_fix > 0 or grn_skip > 0:
                    per_grn.append({
                        'grn_id':       gid,
                        'grn_no':       grn_no,
                        'boxes':        len(boxes),
                        'to_fix':       grn_fix,
                        'applied':      grn_applied,
                        'already_ok':   grn_ok,
                        'skipped':      grn_skip,
                    })
            except Exception as _grn_err:
                totals['grns_failed'] += 1
                last_err = f'GRN {gid}: {_grn_err}'
                continue

        conn.close()
        return jsonify({
            'status':         'ok',
            'mode':           'applied' if apply_changes else 'dry_run',
            'totals':         totals,
            'per_grn':        per_grn[:200],   # cap for response size
            'per_grn_total':  len(per_grn),
            'last_error':     last_err,
            'note': ('Dry run — nothing changed. Re-POST with {"apply":true} to write.'
                     if not apply_changes else 'Changes written.'),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ══════════════════════════════════════════════════════════════════════
#  DUPLICATE-BOX DIAGNOSTIC + CLEANUP
#  Investigates GRN lines where pm_boxes has more rows than the GRN
#  line's no_of_box says it should. Built after the PM-GRN/0343/26-27
#  incident (Jun 2026) where two box batches existed for one grn_items
#  row — one for each grn_item_id that had been created during two
#  sequential save attempts. The orphan grn_item_id values no longer
#  resolve in pm_grn_items so the boxes are now floating against a
#  valid pm_grn but a missing item row.
#
#  This pair of endpoints is the SAFE way to clean those up:
#    • GET  /admin/duplicate_boxes              — list affected GRNs
#    • POST /admin/duplicate_boxes/cleanup      — soft-delete the extras
#                                                 (recoverable via Recycle Bin)
#
#  Detection rule: for every (grn_id, product_id), find every distinct
#  grn_item_id under pm_boxes. If more than one grn_item_id exists OR
#  the total box count doesn't match the surviving pm_grn_items row's
#  no_of_box, surface the GRN for review.
# ══════════════════════════════════════════════════════════════════════

@pm_stock_bp.route('/api/pm_stock/admin/duplicate_boxes', methods=['GET'])
@_login_required
def api_admin_duplicate_boxes():
    """List GRN lines whose pm_boxes rowcount doesn't match the line's
    no_of_box, grouped by (grn_id, product_id). Each entry returns a list
    of box "batches" (one per distinct grn_item_id present in pm_boxes)
    so the admin can decide which batch is canonical and which to discard.

    Returns:
      {
        status: 'ok',
        count: <N>,
        rows: [
          {
            grn_id, grn_no, grn_date,
            product_id, product_code, product_name,
            expected_no_of_box,           # from pm_grn_items.no_of_box (sum if >1 row)
            expected_qty,                 # qty_received
            current_grn_item_ids,         # ids that still exist in pm_grn_items
            actual_box_count,             # total rows in pm_boxes
            extra,                        # actual - expected
            batches: [
              {
                grn_item_id,              # may be orphan (not in pm_grn_items)
                is_orphan: bool,
                box_count,
                min_seq, max_seq,
                first_created, last_created,
                first_box_code, last_box_code,
                any_moved: bool,          # any box has movement beyond grn_create
              },
              ...
            ]
          },
          ...
        ]
      }
    Admin only.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        # Step 1: per (grn_id, product_id) find all batches in pm_boxes and
        # the expected count from pm_grn_items.
        cand = conn.execute("""
            SELECT b.grn_id,
                   b.product_id,
                   COUNT(*)                        AS actual_box_count,
                   COALESCE(SUM(gi.no_of_box), 0)  AS expected_no_of_box,
                   COALESCE(SUM(gi.qty_received), 0) AS expected_qty,
                   COUNT(DISTINCT b.grn_item_id)   AS batch_count
            FROM pm_boxes b
            LEFT JOIN pm_grn_items gi
              ON gi.grn_id = b.grn_id AND gi.product_id = b.product_id
            GROUP BY b.grn_id, b.product_id
            HAVING COUNT(*) <> COALESCE(SUM(DISTINCT gi.no_of_box), 0)
                OR COUNT(DISTINCT b.grn_item_id) > 1
        """).fetchall()
        # NOTE: the HAVING uses SUM(DISTINCT no_of_box) because the LEFT JOIN
        # multiplies pm_grn_items rows when the box-count outer set has
        # multiple boxes per item — we want the line's own no_of_box,
        # not its repeated value. For the common one-item-per-product case
        # this collapses to the correct expected number.

        rows = []
        for c in cand:
            grn_id  = int(c['grn_id'])
            prod_id = int(c['product_id'])
            head = conn.execute(
                """SELECT g.id, g.grn_no, g.grn_date,
                          COALESCE(p.product_code,'') AS product_code,
                          COALESCE(p.product_name,'') AS product_name
                     FROM pm_grn g
                     LEFT JOIN pm_products p ON p.id=%s
                    WHERE g.id=%s""",
                (prod_id, grn_id)
            ).fetchone()
            if not head:
                continue  # GRN was deleted but boxes orphaned — surface separately later if needed
            # Per-batch breakdown
            batches = conn.execute("""
                SELECT b.grn_item_id,
                       COUNT(*) AS box_count,
                       MIN(b.box_seq) AS min_seq,
                       MAX(b.box_seq) AS max_seq,
                       MIN(b.created_at) AS first_created,
                       MAX(b.created_at) AS last_created,
                       (SELECT box_code FROM pm_boxes
                        WHERE grn_id=b.grn_id AND product_id=b.product_id
                          AND grn_item_id <=> b.grn_item_id
                        ORDER BY box_seq ASC LIMIT 1) AS first_box_code,
                       (SELECT box_code FROM pm_boxes
                        WHERE grn_id=b.grn_id AND product_id=b.product_id
                          AND grn_item_id <=> b.grn_item_id
                        ORDER BY box_seq DESC LIMIT 1) AS last_box_code,
                       EXISTS (
                         SELECT 1 FROM pm_box_movements m
                          JOIN pm_boxes b2 ON b2.box_id = m.box_id
                         WHERE b2.grn_id=b.grn_id AND b2.product_id=b.product_id
                           AND b2.grn_item_id <=> b.grn_item_id
                           AND m.movement_type IN ('out','in','consume','adjust')
                       ) AS any_moved
                FROM pm_boxes b
                WHERE b.grn_id=%s AND b.product_id=%s
                GROUP BY b.grn_item_id
                ORDER BY MIN(b.box_seq)
            """, (grn_id, prod_id)).fetchall()

            # Mark which grn_item_ids still exist in pm_grn_items (current).
            current_ids = {r['id'] for r in conn.execute(
                "SELECT id FROM pm_grn_items WHERE grn_id=%s AND product_id=%s",
                (grn_id, prod_id)
            ).fetchall()}
            batch_list = []
            for b in batches:
                gii = b['grn_item_id']
                batch_list.append({
                    'grn_item_id':   int(gii) if gii is not None else None,
                    'is_orphan':     gii is None or int(gii) not in current_ids,
                    'box_count':     int(b['box_count'] or 0),
                    'min_seq':       int(b['min_seq'] or 0),
                    'max_seq':       int(b['max_seq'] or 0),
                    'first_created': str(b['first_created']) if b['first_created'] is not None else None,
                    'last_created':  str(b['last_created'])  if b['last_created']  is not None else None,
                    'first_box_code': b['first_box_code'] or '',
                    'last_box_code':  b['last_box_code']  or '',
                    'any_moved':     bool(b['any_moved']),
                })

            rows.append({
                'grn_id':                grn_id,
                'grn_no':                head['grn_no'],
                'grn_date':              str(head['grn_date']) if head['grn_date'] else None,
                'product_id':            prod_id,
                'product_code':          head['product_code'],
                'product_name':          head['product_name'],
                'expected_no_of_box':    int(c['expected_no_of_box'] or 0),
                'expected_qty':          float(c['expected_qty'] or 0),
                'current_grn_item_ids':  sorted(current_ids),
                'actual_box_count':      int(c['actual_box_count'] or 0),
                'extra':                 int(c['actual_box_count'] or 0) - int(c['expected_no_of_box'] or 0),
                'batches':               batch_list,
            })

        # Sort: biggest discrepancies first
        rows.sort(key=lambda r: (-r['extra'], r['grn_no']))
        conn.close()
        return jsonify({'status':'ok', 'count': len(rows), 'rows': rows})
    except Exception as e:
        import traceback; traceback.print_exc()
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/admin/duplicate_boxes/cleanup', methods=['POST'])
@_login_required
def api_admin_duplicate_boxes_cleanup():
    """Soft-delete the box rows the admin selected as duplicates. The rows
    AND their pm_box_movements children are snapshotted into pm_recycle_bin
    so a wrong call is recoverable via the Recycle Bin UI.

    Body:
      {
        grn_id:       <int, required>,
        product_id:   <int, required>,
        box_ids:      [<int>, ...] — must all belong to (grn_id, product_id),
        reason:       optional short text (≤ 500 chars)
      }

    Returns: { status, deleted, bin_id, message }

    Refuses to delete any box that has movements beyond 'grn_create' —
    that would orphan transfer/dispatch history. Admin only.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    d = request.get_json() or {}
    try:
        grn_id     = int(d.get('grn_id') or 0)
        product_id = int(d.get('product_id') or 0)
        box_ids    = [int(x) for x in (d.get('box_ids') or []) if x]
        reason     = (d.get('reason') or '').strip() or None
    except (TypeError, ValueError):
        return jsonify({'status':'error','message':'Invalid payload'}), 400
    if not grn_id or not product_id or not box_ids:
        return jsonify({'status':'error','message':'grn_id, product_id and non-empty box_ids required'}), 400
    if len(box_ids) > 5000:
        return jsonify({'status':'error','message':'Refusing to delete >5000 boxes in one call'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # Validate: every box_id must belong to the named (grn_id, product_id).
        # Anything else is a sign of a stale request — refuse the whole batch.
        placeholders = ','.join(['%s'] * len(box_ids))
        check = conn.execute(
            f"""SELECT box_id, grn_id, product_id, current_status
                  FROM pm_boxes
                 WHERE box_id IN ({placeholders})""",
            tuple(box_ids)
        ).fetchall()
        found_ids = {int(r['box_id']) for r in check}
        missing = [bid for bid in box_ids if bid not in found_ids]
        if missing:
            conn.close()
            return jsonify({'status':'error',
                'message': f'{len(missing)} box_ids not found: {missing[:10]}{"..." if len(missing)>10 else ""}'
            }), 400
        mismatched = [int(r['box_id']) for r in check
                      if int(r['grn_id']) != grn_id or int(r['product_id']) != product_id]
        if mismatched:
            conn.close()
            return jsonify({'status':'error',
                'message': f'{len(mismatched)} box_ids do not belong to grn_id={grn_id}, product_id={product_id}'
            }), 400

        # Refuse if any selected box has movement history beyond grn_create.
        # We don't want to silently delete boxes that may have been scanned
        # out / received / consumed — that would orphan ledger history.
        moved = conn.execute(
            f"""SELECT m.box_id, m.movement_type
                  FROM pm_box_movements m
                 WHERE m.box_id IN ({placeholders})
                   AND m.movement_type IN ('out','in','consume','adjust')
                 LIMIT 5""",
            tuple(box_ids)
        ).fetchall()
        if moved:
            mvd_ids = sorted({int(m['box_id']) for m in moved})
            conn.close()
            return jsonify({'status':'error',
                'message': f'Refusing to delete: {len(mvd_ids)} of the selected boxes have movement history (e.g. box_id={mvd_ids[0]}). Reverse those movements first.'
            }), 409

        # GRN label for the bin entry (for human-readable recycle bin display)
        head = conn.execute(
            """SELECT g.grn_no, COALESCE(p.product_name,'') AS pname
                 FROM pm_grn g
                 LEFT JOIN pm_products p ON p.id=%s
                WHERE g.id=%s""",
            (product_id, grn_id)
        ).fetchone()
        grn_no = head['grn_no'] if head else f'GRN#{grn_id}'
        pname  = head['pname']  if head else f'#{product_id}'

        # Snapshot the boxes + their movements directly to pm_recycle_bin so
        # this is reversible. We don't use _bin_soft_delete here because its
        # contract is "delete one parent row + its children" — and our parent
        # (the GRN header) must stay intact. So we hand-roll the bin entry.
        import json as _json
        # Pull full row data for both tables (preserve column order so
        # restore is just "INSERT with these keys").
        box_rows = conn.execute(
            f"SELECT * FROM pm_boxes WHERE box_id IN ({placeholders})",
            tuple(box_ids)
        ).fetchall()
        mvt_rows = conn.execute(
            f"SELECT * FROM pm_box_movements WHERE box_id IN ({placeholders})",
            tuple(box_ids)
        ).fetchall()
        payload = {
            'parent_table': 'pm_grn',     # informational only; not used on restore
            'parent':       {'id': grn_id, 'grn_no': grn_no},
            'children': {
                'pm_box_movements': {
                    'where': f'box_id IN ({placeholders})',
                    'rows':  [_row_to_dict(r) for r in mvt_rows],
                },
                'pm_boxes': {
                    'where': f'box_id IN ({placeholders})',
                    'rows':  [_row_to_dict(r) for r in box_rows],
                },
            }
        }
        payload_json = _json.dumps(payload, ensure_ascii=False, default=_json_safe)
        cur = conn.execute(
            """INSERT INTO pm_recycle_bin
                  (entity_type, entity_label, entity_id, payload, payload_summary,
                   deleted_by, reason)
                VALUES (%s, %s, %s, %s, %s, %s, %s)""",
            ('duplicate_box_cleanup',
             f"GRN {grn_no} · {pname} (duplicate-box cleanup × {len(box_ids)})"[:200],
             grn_id,
             payload_json,
             f'Removed {len(box_ids)} duplicate boxes from GRN {grn_no} / {pname}'[:500],
             _user(),
             reason)
        )
        bin_id = cur.lastrowid

        # Now the actual delete — movements first (FK-safe), then boxes.
        conn.execute(
            f"DELETE FROM pm_box_movements WHERE box_id IN ({placeholders})",
            tuple(box_ids)
        )
        conn.execute(
            f"DELETE FROM pm_boxes WHERE box_id IN ({placeholders})",
            tuple(box_ids)
        )

        conn.commit()
        conn.close()
        return jsonify({
            'status':  'ok',
            'deleted': len(box_ids),
            'bin_id':  bin_id,
            'message': f'Removed {len(box_ids)} duplicate boxes. Reversible from Recycle Bin (entry #{bin_id}).'
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        try: conn.rollback(); conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/admin/duplicate_boxes/batch_ids', methods=['GET'])
@_login_required
def api_admin_duplicate_boxes_batch_ids():
    """Return the box_ids belonging to a specific batch (grn_id, product_id,
    grn_item_id) — used by the cleanup UI right before POSTing the delete.

    Query params:
        grn_id          (required)
        product_id      (required)
        grn_item_id     (optional — pass empty string '' or omit for null/orphan
                                    batches, which means grn_item_id IS NULL)

    Returns: { status, box_ids:[..], count }
    Admin only.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    try:
        grn_id     = int(request.args.get('grn_id') or 0)
        product_id = int(request.args.get('product_id') or 0)
    except (TypeError, ValueError):
        return jsonify({'status':'error','message':'grn_id and product_id required'}), 400
    if not grn_id or not product_id:
        return jsonify({'status':'error','message':'grn_id and product_id required'}), 400

    gii_raw = request.args.get('grn_item_id', None)
    use_null = (gii_raw is None) or (gii_raw == '') or (gii_raw == 'null')
    try:
        gii_val = None if use_null else int(gii_raw)
    except (TypeError, ValueError):
        return jsonify({'status':'error','message':'grn_item_id must be an integer or empty for null'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        if use_null:
            rows = conn.execute(
                """SELECT box_id FROM pm_boxes
                    WHERE grn_id=%s AND product_id=%s AND grn_item_id IS NULL
                    ORDER BY box_seq""",
                (grn_id, product_id)
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT box_id FROM pm_boxes
                    WHERE grn_id=%s AND product_id=%s AND grn_item_id=%s
                    ORDER BY box_seq""",
                (grn_id, product_id, gii_val)
            ).fetchall()
        ids = [int(r['box_id']) for r in rows]
        conn.close()
        return jsonify({'status':'ok', 'box_ids': ids, 'count': len(ids)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/grn/list')
@_login_required
def api_grn_list():
    from_date = request.args.get('from_date') or ''
    to_date   = request.args.get('to_date')   or ''
    search    = (request.args.get('search') or '').strip()

    conn = sampling_portal.get_db_connection()
    params = []
    where  = []
    if from_date:
        where.append("g.grn_date >= %s"); params.append(from_date)
    if to_date:
        where.append("g.grn_date <= %s"); params.append(to_date)
    if search:
        where.append("(g.grn_no LIKE %s OR g.supplier LIKE %s OR g.po_number LIKE %s)")
        s = f'%{search}%'
        params += [s, s, s]

    where_clause = ("WHERE " + " AND ".join(where)) if where else ""
    rows = conn.execute(f"""
        SELECT g.id, g.grn_no, g.grn_date, g.po_number, g.po_date,
               g.supplier, g.godown_id, COALESCE(gd.name,'') AS godown_name,
               g.remarks, g.created_by, g.created_at,
               COALESCE(g.party_invoice_no,'') AS party_invoice_no,
               g.party_invoice_date,
               -- GRN verification fields (toggle-aware). 'verified' is the
               -- default for legacy rows and for the toggle-off path.
               COALESCE(g.verification_status, 'verified') AS verification_status,
               g.verified_at, COALESCE(g.verified_by,'') AS verified_by,
               COUNT(gi.id) AS item_count,
               COALESCE(SUM(gi.qty_received),0) AS total_qty,
               (SELECT COUNT(*) FROM pm_grn_files f
                  WHERE f.grn_id = g.id AND f.kind='invoice') AS invoice_file_count
        FROM pm_grn g
        LEFT JOIN procurement_godowns gd ON gd.id = g.godown_id
        LEFT JOIN pm_grn_items gi ON gi.grn_id = g.id
        {where_clause}
        GROUP BY g.id
        ORDER BY g.grn_date DESC, g.created_at DESC
    """, params).fetchall()
    conn.close()

    result = []
    for r in rows:
        d = dict(r)
        if hasattr(d.get('grn_date'), 'isoformat'):
            d['grn_date'] = d['grn_date'].isoformat()
        if hasattr(d.get('po_date'), 'isoformat') and d['po_date']:
            d['po_date'] = d['po_date'].isoformat()
        if hasattr(d.get('party_invoice_date'), 'isoformat') and d['party_invoice_date']:
            d['party_invoice_date'] = d['party_invoice_date'].isoformat()
        if hasattr(d.get('created_at'), 'isoformat'):
            d['created_at'] = str(d['created_at'])
        if d.get('verified_at') is not None:
            d['verified_at'] = str(d['verified_at'])
        result.append(d)
    return jsonify(result)

# ── PM Purchase Orders ────────────────────────────────────────────────────
# Phase A: read-only list endpoint. Returns one row per PO header with
# aggregated item counts. Filters mirror the GRN list (date range + search)
# plus PO-specific filters (status, approval_status, supplier, product).
#
# The two-axis state model is reflected in two filter params: 'status' (the
# receipt lifecycle) and 'approval_status' (admin approval state). When the
# UI wants to show "anything that's actionable right now" it filters
# approval_status='approved' + status IN ('open','partial').
@pm_stock_bp.route('/api/pm_stock/purchase_orders/list')
@_login_required
def api_pm_po_list():
    from_date       = request.args.get('from_date') or ''
    to_date         = request.args.get('to_date')   or ''
    search          = (request.args.get('search') or '').strip()
    status_filter   = (request.args.get('status') or '').strip()
    appr_filter     = (request.args.get('approval_status') or '').strip()
    supplier_filter = (request.args.get('supplier') or '').strip()
    product_filter  = (request.args.get('product_id') or '').strip()

    conn = sampling_portal.get_db_connection()
    # Honest pre-flight: do the tables exist? If not, return a clear
    # message instead of a confusing 500. This is the most likely failure
    # mode on first deploy (Flask reloaded the new route but the helpers.py
    # migrations didn't run).
    try:
        conn.execute("SELECT 1 FROM pm_purchase_orders LIMIT 1").fetchone()
    except Exception:
        conn.close()
        return jsonify({
            'status': 'error',
            'message': 'Migrations have not run yet. Restart Flask, or hit /api/pm_stock/_version to see which migrations are missing.',
            'missing_table': 'pm_purchase_orders',
        }), 503  # service-unavailable — semantically correct for "DB schema not ready"

    params = []
    where  = []
    if from_date:
        where.append("p.po_date >= %s"); params.append(from_date)
    if to_date:
        where.append("p.po_date <= %s"); params.append(to_date)
    if status_filter:
        # Allow comma-separated lists too: "open,partial"
        wanted = [s.strip() for s in status_filter.split(',') if s.strip()]
        if wanted:
            placeholders = ','.join(['%s'] * len(wanted))
            where.append(f"p.status IN ({placeholders})")
            params += wanted
    if appr_filter:
        wanted = [s.strip() for s in appr_filter.split(',') if s.strip()]
        if wanted:
            placeholders = ','.join(['%s'] * len(wanted))
            where.append(f"p.approval_status IN ({placeholders})")
            params += wanted
    if supplier_filter:
        # Search by supplier name (denormalized on the PO header)
        where.append("p.supplier_name LIKE %s")
        params.append(f'%{supplier_filter}%')
    if product_filter:
        # Sub-select to find POs that have this product on any line
        where.append("EXISTS (SELECT 1 FROM pm_po_items i WHERE i.po_id=p.id AND i.product_id=%s)")
        params.append(int(product_filter))
    if search:
        where.append("(p.po_num LIKE %s OR p.supplier_name LIKE %s OR p.remarks LIKE %s)")
        s = f'%{search}%'
        params += [s, s, s]

    where_clause = ("WHERE " + " AND ".join(where)) if where else ""
    rows = conn.execute(f"""
        SELECT p.id, p.po_num, p.po_date,
               p.supplier_id, COALESCE(p.supplier_name,'') AS supplier_name,
               p.godown_id, COALESCE(gd.name,'') AS godown_name,
               p.delivery_date, p.delivery_days,
               p.status, p.approval_status,
               COALESCE(p.approved_by,'') AS approved_by,
               p.approved_at,
               COALESCE(p.rejection_reason,'') AS rejection_reason,
               p.grand_total, p.freight_charge, p.packing_charge,
               COALESCE(p.remarks,'') AS remarks,
               COALESCE(p.created_by,'') AS created_by, p.created_at,
               COALESCE(p.updated_by,'') AS updated_by, p.updated_at,
               COUNT(i.id) AS item_count,
               COALESCE(SUM(i.qty_primary),0) AS total_qty_primary,
               -- How much has been received against this PO across all
               -- GRNs linked to it. pm_grn has no 'cancelled' status column
               -- (unlike procurement_grn) — deletions remove the row entirely,
               -- so every present GRN counts. If a soft-cancel flag is added
               -- to pm_grn later, fold it into this WHERE.
               (SELECT COALESCE(SUM(gi.qty_received),0)
                  FROM pm_grn g
                  JOIN pm_grn_items gi ON gi.grn_id = g.id
                 WHERE g.po_id = p.id) AS received_qty_total
        FROM pm_purchase_orders p
        LEFT JOIN procurement_godowns gd ON gd.id = p.godown_id
        LEFT JOIN pm_po_items i ON i.po_id = p.id
        {where_clause}
        GROUP BY p.id
        ORDER BY p.po_date DESC, p.created_at DESC
    """, params).fetchall()
    conn.close()

    result = []
    for r in rows:
        d = dict(r)
        for fld in ('po_date','delivery_date'):
            if hasattr(d.get(fld), 'isoformat') and d[fld] is not None:
                d[fld] = d[fld].isoformat()
        for fld in ('created_at','updated_at','approved_at'):
            if d.get(fld) is not None:
                d[fld] = str(d[fld])
        # Cast numerics that the DB driver may return as Decimal/None
        for fld in ('grand_total','freight_charge','packing_charge','total_qty_primary','received_qty_total'):
            d[fld] = float(d.get(fld) or 0)
        result.append(d)
    return jsonify(result)


# ── ═══════════════════════════════════════════════════════════════════
#    Purchase Orders — Phase B: detail, save, approve, reject, cancel
#    ═══════════════════════════════════════════════════════════════════
#
# All write endpoints share a few invariants:
#   • approval_status starts 'pending' on create.
#   • status starts at whatever the client sent ('draft' or 'open');
#     'closed' / 'partial' are computed on GRN linkage, not set by hand.
#   • Editing a PO with approval='approved' is restricted: only the lines'
#     ordered qty can NOT be lowered below what's already received; we
#     allow header tweaks (delivery date, remarks) but block destructive
#     line edits with a clear error.
#   • Cancelling a PO that already has linked GRNs is blocked.
#   • All four write endpoints honour the same access decorators that
#     other voucher endpoints use, so admin-defined access roles work.

def _po_recompute_status(conn, po_id):
    """Recompute status from received qty across linked GRNs.
       Called on PO save and from grn_save after a GRN is linked.
       Rule: open → partial (some received) → closed (all received).
       'draft' / 'cancelled' / 'rejected' are sticky and NEVER touched here."""
    head = conn.execute(
        "SELECT id, status, approval_status FROM pm_purchase_orders WHERE id=%s",
        (po_id,)
    ).fetchone()
    if not head: return
    # Don't auto-flip terminal/manual states
    if head['status'] in ('draft', 'cancelled'): return
    if head['approval_status'] == 'rejected': return
    totals = conn.execute("""
        SELECT
          COALESCE(SUM(i.qty_primary),0) AS ordered_total,
          (SELECT COALESCE(SUM(gi.qty_received),0)
             FROM pm_grn g JOIN pm_grn_items gi ON gi.grn_id=g.id
             WHERE g.po_id=%s) AS received_total
        FROM pm_po_items i WHERE i.po_id=%s
    """, (po_id, po_id)).fetchone()
    ordered = float(totals['ordered_total'] or 0)
    received = float(totals['received_total'] or 0)
    if ordered <= 0: return
    new_status = 'open'
    if received >= ordered:
        new_status = 'closed'
    elif received > 0:
        new_status = 'partial'
    if new_status != head['status']:
        conn.execute(
            "UPDATE pm_purchase_orders SET status=%s, updated_by=%s WHERE id=%s",
            (new_status, _user(), po_id)
        )


@pm_stock_bp.route('/api/pm_stock/purchase_orders/<int:po_id>')
@_login_required
def api_pm_po_detail(po_id):
    conn = sampling_portal.get_db_connection()
    try:
        head = conn.execute("""
            SELECT p.*, COALESCE(gd.name,'') AS godown_name
            FROM pm_purchase_orders p
            LEFT JOIN procurement_godowns gd ON gd.id=p.godown_id
            WHERE p.id=%s
        """, (po_id,)).fetchone()
        if not head:
            conn.close()
            return jsonify({'status':'error','message':'PO not found'}), 404
        items = conn.execute("""
            SELECT i.*, p.product_name, p.pm_type, p.product_code,
                   COALESCE(p.primary_uom,'Nos') AS primary_uom,
                   -- How much of THIS line has already arrived via linked GRNs.
                   -- We sum GRN-item qtys for the same product across all GRNs
                   -- linked to this PO. (Per-line link isn't modelled; product
                   -- match is the proxy. Good enough for status tracking.)
                   (SELECT COALESCE(SUM(gi.qty_received),0)
                      FROM pm_grn g
                      JOIN pm_grn_items gi ON gi.grn_id=g.id
                     WHERE g.po_id=%s AND gi.product_id=i.product_id) AS received_qty
            FROM pm_po_items i
            LEFT JOIN pm_products p ON p.id=i.product_id
            WHERE i.po_id=%s
            ORDER BY i.id
        """, (po_id, po_id)).fetchall()
        # GRNs linked to this PO (for the "Receipts" tab on PO detail)
        grns = conn.execute("""
            SELECT g.id, g.grn_no, g.grn_date, g.supplier,
                   COALESCE(gd.name,'') AS godown_name,
                   COALESCE(g.verification_status,'verified') AS vstatus,
                   (SELECT COALESCE(SUM(gi.qty_received),0)
                      FROM pm_grn_items gi WHERE gi.grn_id=g.id) AS total_qty
            FROM pm_grn g LEFT JOIN procurement_godowns gd ON gd.id=g.godown_id
            WHERE g.po_id=%s
            ORDER BY g.grn_date DESC, g.id DESC
        """, (po_id,)).fetchall()
        conn.close()

        h = dict(head)
        for fld in ('po_date','delivery_date'):
            if hasattr(h.get(fld), 'isoformat') and h[fld] is not None:
                h[fld] = h[fld].isoformat()
        for fld in ('created_at','updated_at','approved_at'):
            if h.get(fld) is not None: h[fld] = str(h[fld])
        for fld in ('grand_total','freight_charge','packing_charge'):
            h[fld] = float(h.get(fld) or 0)

        out_items = []
        for r in items:
            d = dict(r)
            for fld in ('qty','entered_qty','qty_primary','rate','amount',
                        'gst_rate','cgst_amount','sgst_amount','received_qty'):
                d[fld] = float(d.get(fld) or 0)
            out_items.append(d)

        out_grns = []
        for g in grns:
            d = dict(g)
            if hasattr(d.get('grn_date'), 'isoformat') and d['grn_date'] is not None:
                d['grn_date'] = d['grn_date'].isoformat()
            d['total_qty'] = float(d.get('total_qty') or 0)
            out_grns.append(d)

        return jsonify({'status':'ok', 'header':h, 'items':out_items, 'grns':out_grns})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/purchase_orders/save', methods=['POST'])
@_login_required
def api_pm_po_save():
    """Create or update a PO. Body shape:
       { id?: int,                  # present → update; absent → create
         po_date, supplier_id?, supplier_name, godown_id?,
         delivery_date?, delivery_days?,
         status: 'draft'|'open',    # 'partial'/'closed'/'cancelled' set elsewhere
         remarks?, freight_charge?, packing_charge?,
         items: [
           { id?: int, product_id, qty, entered_uom?, entered_qty?,
             qty_primary, rate?, amount?, gst_rate?, remarks? }
         ]
       }"""
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked

    d = request.get_json() or {}
    po_id    = d.get('id')
    po_date  = d.get('po_date') or str(date.today())
    sup_id   = d.get('supplier_id') or None
    sup_name = (d.get('supplier_name') or '').strip() or None
    godown_id = d.get('godown_id') or None
    delivery_date = d.get('delivery_date') or None
    delivery_days = d.get('delivery_days') or None
    status   = (d.get('status') or 'draft').strip().lower()
    if status not in ('draft','open'):
        return jsonify({'status':'error','message':"status must be 'draft' or 'open'"}), 400
    remarks  = (d.get('remarks') or '').strip() or None
    freight  = d.get('freight_charge') or 0
    packing  = d.get('packing_charge') or 0
    items    = d.get('items') or []

    if not sup_name:
        return jsonify({'status':'error','message':'Supplier is required'}), 400
    valid_items = [it for it in items if int(it.get('product_id') or 0) and float(it.get('qty') or 0) > 0]
    if not valid_items:
        return jsonify({'status':'error','message':'Add at least one item with qty > 0'}), 400

    # Compute grand_total from items + freight + packing for the header
    line_total = sum(float(it.get('amount') or 0) for it in valid_items)
    grand_total = line_total + float(freight or 0) + float(packing or 0)

    conn = sampling_portal.get_db_connection()
    try:
        if po_id:
            # ── UPDATE ─────────────────────────────────────────────────
            head = conn.execute(
                "SELECT id, status, approval_status FROM pm_purchase_orders WHERE id=%s FOR UPDATE",
                (po_id,)
            ).fetchone()
            if not head:
                conn.close(); return jsonify({'status':'error','message':'PO not found'}), 404
            # Block edits on cancelled / closed
            if head['status'] in ('cancelled','closed'):
                conn.close(); return jsonify({'status':'error','message':f"Cannot edit a {head['status']} PO"}), 400
            # If already approved, refuse line changes that would invalidate
            # received GRN quantities. Header tweaks are fine.
            if head['approval_status'] == 'approved':
                # Check each existing line: new qty_primary must be ≥ received_qty
                existing = conn.execute(
                    "SELECT id, product_id, qty_primary FROM pm_po_items WHERE po_id=%s",
                    (po_id,)
                ).fetchall()
                received_by_pid = {}
                rcv_rows = conn.execute("""
                    SELECT gi.product_id, COALESCE(SUM(gi.qty_received),0) AS r
                    FROM pm_grn g JOIN pm_grn_items gi ON gi.grn_id=g.id
                    WHERE g.po_id=%s GROUP BY gi.product_id
                """, (po_id,)).fetchall()
                for r in rcv_rows:
                    received_by_pid[int(r['product_id'])] = float(r['r'] or 0)
                # Sum new qty per product
                new_qty_by_pid = {}
                for it in valid_items:
                    pid = int(it.get('product_id') or 0)
                    new_qty_by_pid[pid] = new_qty_by_pid.get(pid, 0) + float(it.get('qty_primary') or 0)
                for pid, rcv in received_by_pid.items():
                    if new_qty_by_pid.get(pid, 0) < rcv:
                        conn.close()
                        return jsonify({'status':'error',
                            'message': f'Cannot reduce ordered qty for product #{pid} below already-received {rcv}. Use Cancel + new PO instead.'
                        }), 400

            conn.execute("""
                UPDATE pm_purchase_orders SET
                  po_date=%s, supplier_id=%s, supplier_name=%s, godown_id=%s,
                  delivery_date=%s, delivery_days=%s, status=%s,
                  remarks=%s, freight_charge=%s, packing_charge=%s, grand_total=%s,
                  updated_by=%s
                WHERE id=%s
            """, (po_date, sup_id, sup_name, godown_id, delivery_date, delivery_days,
                  status, remarks, freight, packing, grand_total, _user(), po_id))
            # Wipe + re-insert items. Simpler than diffing, and pm_po_items has
            # no downstream FK references.
            conn.execute("DELETE FROM pm_po_items WHERE po_id=%s", (po_id,))
            for it in valid_items:
                conn.execute("""
                    INSERT INTO pm_po_items
                      (po_id, product_id, qty, entered_uom, entered_qty, qty_primary,
                       rate, amount, gst_rate, cgst_amount, sgst_amount, remarks)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (po_id, int(it.get('product_id')),
                      float(it.get('qty') or 0),
                      (it.get('entered_uom') or '').strip() or None,
                      float(it.get('entered_qty') or 0),
                      float(it.get('qty_primary') or it.get('qty') or 0),
                      float(it.get('rate') or 0),
                      float(it.get('amount') or 0),
                      float(it.get('gst_rate') or 0),
                      float(it.get('cgst_amount') or 0),
                      float(it.get('sgst_amount') or 0),
                      (it.get('remarks') or '').strip() or None))
            _po_recompute_status(conn, po_id)
            conn.commit()
            conn.close()
            return jsonify({'status':'ok','id':po_id,'mode':'update'})
        else:
            # ── CREATE ─────────────────────────────────────────────────
            po_num = _next_voucher_no(conn, 'PM-PO', po_date)
            cur = conn.execute("""
                INSERT INTO pm_purchase_orders
                  (po_num, po_date, supplier_id, supplier_name, godown_id,
                   delivery_date, delivery_days, status, approval_status,
                   remarks, freight_charge, packing_charge, grand_total,
                   created_by, updated_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'pending',%s,%s,%s,%s,%s,%s)
            """, (po_num, po_date, sup_id, sup_name, godown_id,
                  delivery_date, delivery_days, status,
                  remarks, freight, packing, grand_total,
                  _user(), _user()))
            new_id = cur.lastrowid
            for it in valid_items:
                conn.execute("""
                    INSERT INTO pm_po_items
                      (po_id, product_id, qty, entered_uom, entered_qty, qty_primary,
                       rate, amount, gst_rate, cgst_amount, sgst_amount, remarks)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (new_id, int(it.get('product_id')),
                      float(it.get('qty') or 0),
                      (it.get('entered_uom') or '').strip() or None,
                      float(it.get('entered_qty') or 0),
                      float(it.get('qty_primary') or it.get('qty') or 0),
                      float(it.get('rate') or 0),
                      float(it.get('amount') or 0),
                      float(it.get('gst_rate') or 0),
                      float(it.get('cgst_amount') or 0),
                      float(it.get('sgst_amount') or 0),
                      (it.get('remarks') or '').strip() or None))
            conn.commit()
            conn.close()
            return jsonify({'status':'ok','id':new_id,'po_num':po_num,'mode':'create'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/purchase_orders/<int:po_id>/approve', methods=['POST'])
@_login_required
def api_pm_po_approve(po_id):
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    # Approving is gated by the same access flag for now. If a separate
    # "approver" role is added later, switch this to that key.
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked
    conn = sampling_portal.get_db_connection()
    try:
        head = conn.execute(
            "SELECT id, status, approval_status FROM pm_purchase_orders WHERE id=%s FOR UPDATE",
            (po_id,)
        ).fetchone()
        if not head:
            conn.close(); return jsonify({'status':'error','message':'PO not found'}), 404
        if head['status'] == 'cancelled':
            conn.close(); return jsonify({'status':'error','message':'PO is cancelled'}), 400
        if head['approval_status'] == 'approved':
            conn.close(); return jsonify({'status':'ok','message':'Already approved'})
        conn.execute("""
            UPDATE pm_purchase_orders SET
              approval_status='approved', approved_by=%s, approved_at=NOW(),
              rejection_reason=NULL,
              -- An approved PO that was still 'draft' is implicitly 'open' now.
              status = CASE WHEN status='draft' THEN 'open' ELSE status END,
              updated_by=%s
            WHERE id=%s
        """, (_user(), _user(), po_id))
        conn.commit()
        conn.close()
        return jsonify({'status':'ok'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/purchase_orders/<int:po_id>/reject', methods=['POST'])
@_login_required
def api_pm_po_reject(po_id):
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked
    d = request.get_json() or {}
    reason = (d.get('reason') or '').strip()
    if not reason:
        return jsonify({'status':'error','message':'Rejection reason is required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        head = conn.execute(
            "SELECT id, status, approval_status FROM pm_purchase_orders WHERE id=%s FOR UPDATE",
            (po_id,)
        ).fetchone()
        if not head:
            conn.close(); return jsonify({'status':'error','message':'PO not found'}), 404
        # Block rejection of POs that already have GRNs against them.
        rcv = conn.execute(
            "SELECT COUNT(*) AS n FROM pm_grn WHERE po_id=%s", (po_id,)
        ).fetchone()
        if int(rcv['n'] or 0) > 0:
            conn.close()
            return jsonify({'status':'error',
                'message':'PO has linked GRNs — cannot reject. Cancel instead.'}), 400
        conn.execute("""
            UPDATE pm_purchase_orders SET
              approval_status='rejected', rejection_reason=%s,
              approved_by=NULL, approved_at=NULL, updated_by=%s
            WHERE id=%s
        """, (reason, _user(), po_id))
        conn.commit()
        conn.close()
        return jsonify({'status':'ok'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/purchase_orders/<int:po_id>/cancel', methods=['POST'])
@_login_required
def api_pm_po_cancel(po_id):
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked
    conn = sampling_portal.get_db_connection()
    try:
        head = conn.execute(
            "SELECT id, status FROM pm_purchase_orders WHERE id=%s FOR UPDATE",
            (po_id,)
        ).fetchone()
        if not head:
            conn.close(); return jsonify({'status':'error','message':'PO not found'}), 404
        if head['status'] == 'cancelled':
            conn.close(); return jsonify({'status':'ok','message':'Already cancelled'})
        # Block cancellation if any GRN already linked. Tarak's flow is:
        # if you started receiving against it, you can't undo the PO; either
        # finish receiving or create a return DN.
        rcv = conn.execute(
            "SELECT COUNT(*) AS n FROM pm_grn WHERE po_id=%s", (po_id,)
        ).fetchone()
        if int(rcv['n'] or 0) > 0:
            conn.close()
            return jsonify({'status':'error',
                'message':'PO has linked GRNs — cannot cancel.'}), 400
        conn.execute(
            "UPDATE pm_purchase_orders SET status='cancelled', updated_by=%s WHERE id=%s",
            (_user(), po_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'status':'ok'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


# ── End of Purchase Orders Phase B endpoints ──────────────────────────


@pm_stock_bp.route('/api/pm_stock/grn/<int:grn_id>')
@_login_required
def api_grn_detail(grn_id):
    conn = sampling_portal.get_db_connection()
    g = conn.execute("""
        SELECT g.*, COALESCE(gd.name,'') AS godown_name
        FROM pm_grn g
        LEFT JOIN procurement_godowns gd ON gd.id = g.godown_id
        WHERE g.id = %s
    """, (grn_id,)).fetchone()
    if not g:
        conn.close()
        return jsonify({'status': 'error', 'message': 'GRN not found'}), 404

    items = conn.execute("""
        SELECT gi.id, gi.product_id, p.product_name, p.pm_type,
               gi.qty_received, COALESCE(gi.no_of_box,0) AS no_of_box,
               COALESCE(gi.box_count,0) AS box_count,
               gi.remarks,
               COALESCE(gi.product_version, '') AS product_version,
               COALESCE(gi.entered_uom, '')    AS entered_uom,
               COALESCE(p.primary_uom, 'Nos')  AS primary_uom,
               COALESCE(p.alt_uom, '')         AS alt_uom,
               COALESCE(fl.fifo_code,'') AS fifo_code
        FROM pm_grn_items gi
        JOIN pm_products p ON p.id = gi.product_id
        LEFT JOIN pm_fifo_lots fl
               ON fl.lot_kind='grn' AND fl.lot_ref=gi.grn_id
              AND fl.product_id=gi.product_id
        WHERE gi.grn_id = %s
    """, (grn_id,)).fetchall()
    # Include party_invoice_date + grn_date normalization
    if g.get('party_invoice_date') and hasattr(g['party_invoice_date'], 'isoformat'):
        g['party_invoice_date'] = g['party_invoice_date'].isoformat()
    conn.close()

    d = dict(g)
    for k in ('grn_date', 'po_date', 'created_at'):
        if d.get(k) and hasattr(d[k], 'isoformat'):
            d[k] = d[k].isoformat() if k != 'created_at' else str(d[k])
    d['items'] = [dict(i) for i in items]
    return jsonify(d)

@pm_stock_bp.route('/api/pm_stock/grn/save', methods=['POST'])
@_login_required
def api_grn_save():
    """
    Create a PM GRN. Each item automatically creates a godown 'inward' transaction.
    Body: {
        grn_date, po_number (opt), po_date (opt), supplier (opt),
        godown_id, remarks,
        items: [{product_id, qty_received, remarks}, ...]
    }
    """
    blocked = _block_if_disabled('grn')
    if blocked is not None: return blocked
    # Requesters (FACTORY users) can't create GRNs — they submit a
    # Material Request and a fulfiller creates the OUT instead.
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    # New-voucher-entries access check (per-user access control)
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked
    d         = request.get_json() or {}
    grn_date  = d.get('grn_date')  or str(date.today())
    po_number = (d.get('po_number') or '').strip() or None
    po_date   = d.get('po_date') or None
    # Optional FK to pm_purchase_orders.id when this GRN is being raised
    # against an existing PO (picked via the PO picker in the GRN modal).
    po_id_raw = d.get('po_id')
    try: po_id = int(po_id_raw) if po_id_raw else None
    except (TypeError, ValueError): po_id = None
    supplier  = (d.get('supplier') or '').strip()
    godown_id = d.get('godown_id') or None
    remarks   = (d.get('remarks') or '').strip()
    party_invoice_no   = (d.get('party_invoice_no') or '').strip() or None
    party_invoice_date = d.get('party_invoice_date') or None
    supervisor_name    = (d.get('supervisor_name') or '').strip() or None
    items     = d.get('items', [])

    if not items:
        return jsonify({'status': 'error', 'message': 'At least one item required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # Enforce per-user home godown (non-admins only)
        ok, msg = _enforce_home_godown(conn, ('godown_id', godown_id))
        if not ok:
            conn.close()
            return jsonify({'status':'error','message':msg}), 403

        grn_no = _next_voucher_no(conn, 'PM-GRN', grn_date)
        # ── Verification-required toggle ──────────────────────────────
        # When the admin has turned on `grn_verify_required`, GRN saves
        # are parked in 'pending' state — boxes are still created so the
        # operator can print labels — but inward stock is NOT posted
        # to pm_godown_txn until /verify succeeds. This gives a true
        # physical receipt-counted inventory rather than a paper one.
        # Legacy / toggle-off path: status='verified', stock posts now.
        _raw_vr = _setting_get(conn, 'grn_verify_required', '0')
        verify_required = str(_raw_vr).strip().lower() in ('1','true','t','yes','y','on')
        v_status        = 'pending' if verify_required else 'verified'
        verified_at_val = None      if verify_required else datetime.now()
        verified_by_val = None      if verify_required else _user()
        cur = conn.execute(
            """INSERT INTO pm_grn (grn_no, grn_date, po_number, po_date, po_id, supplier, godown_id, remarks, created_by, party_invoice_no, party_invoice_date, supervisor_name, verification_status, verified_at, verified_by)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (grn_no, grn_date, po_number, po_date or None, po_id, supplier, godown_id, remarks, _user(),
             party_invoice_no, party_invoice_date or None, supervisor_name,
             v_status, verified_at_val, verified_by_val)
        )
        grn_id = cur.lastrowid
        # If this GRN was raised against a PO, recompute that PO's status
        # (open → partial → closed) based on cumulative received vs ordered.
        # Best-effort; if it fails we still want the GRN save to succeed.
        if po_id:
            try: _po_recompute_status(conn, po_id)
            except Exception: pass

        # Track FIFO codes assigned per product so we can echo them back to
        # the client and stamp on labels without a second round-trip.
        item_fifo = {}  # product_id -> fifo_code

        # Parallel to input items[] — None for rows that got skipped
        # (no product or zero qty). The frontend uses this list to stamp
        # the new grn_item_id onto each row, so label printing can look
        # up the correct short_codes per grn_item_id (not per product —
        # multi-line same-product GRNs need item-level resolution).
        new_item_ids = []

        for item in items:
            pid = item.get('product_id')
            qty = float(item.get('qty_received') or 0)
            nob = int(item.get('no_of_box') or 0)
            bc  = int(item.get('box_count') or 0)
            irm = (item.get('remarks') or '').strip()
            pver = (item.get('product_version') or '').strip() or None
            # Phase 2 UOM label — pure metadata; no conversion applied to qty.
            # NULL means "fall back to the product's primary UOM at display".
            euom = (item.get('entered_uom') or '').strip() or None
            # Unit rate at receipt (optional). Used by ABC Analysis to compute
            # material value. Missing / null / negative → stored as 0, which
            # means "rate unknown" and contributes nothing to the value pool.
            try:
                rate = float(item.get('rate') or 0)
                if rate < 0: rate = 0.0
            except (TypeError, ValueError):
                rate = 0.0
            if not pid or qty <= 0:
                new_item_ids.append(None)
                continue
            # ── INVARIANT GUARD: box_count must equal qty_received / no_of_box ──
            # The "Fix Box Qty" repair tool exists because, historically, boxes
            # could end up with per_box_qty != qty/no_of_box (user typo, or
            # legacy code that took box_count from the form without checking).
            # Permanent fix: server-side, force the derived value. Any UI form
            # that has both `qty_received` and `no_of_box` fields IS authoritatively
            # describing the per-box qty as their ratio. If the front-end sends
            # a `box_count` that disagrees, we silently correct it here so the
            # data is always self-consistent. Tarak's "120-box scans as 170"
            # bug becomes impossible going forward — line A (qty 120, nob 1)
            # always yields bc=120; line B (qty 5950, nob 35) always yields
            # bc=170 — never the other way around.
            if nob > 0:
                derived_bc = round(qty / nob, 3)
                # Only override when the client value clearly disagrees.
                # Accept tiny floating-point drift (< 0.01) as "same".
                if abs(float(bc) - derived_bc) > 0.01:
                    bc = derived_bc
                else:
                    bc = derived_bc   # normalise to the precise derived value
            cur = conn.execute(
                "INSERT INTO pm_grn_items (grn_id, product_id, qty_received, no_of_box, box_count, remarks, product_version, entered_uom, rate) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (grn_id, pid, qty, nob, bc, irm, pver, euom, rate)
            )
            grn_item_id = cur.lastrowid
            new_item_ids.append(grn_item_id)
            # Assign FIFO code for this lot (one per (grn_id, product_id)).
            # Idempotent — if the GRN is re-saved or re-imported, the same
            # code is returned.
            try:
                fi = _get_or_assign_fifo(conn, 'grn', grn_id, pid)
                item_fifo[int(pid)] = fi.get('fifo_code', '') or ''
            except Exception as _fe:
                # Non-fatal: FIFO is a labelling aid, not stock-affecting.
                # Log and continue so a FIFO failure can never block a GRN.
                import sys as _sys
                print(f"[pm_stock_routes] FIFO assign failed for GRN {grn_id}/{pid}: {_fe}", file=_sys.stderr)
            # Auto create godown inward — UNLESS verify_required is ON,
            # in which case stock posts later via /api/pm_stock/grn/<id>/verify
            # after the operator scans every box.
            if not verify_required:
                txn_vno = _next_voucher_no(conn, 'PM-GTXN', grn_date)
                conn.execute(
                    """INSERT INTO pm_godown_txn (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
                       VALUES (%s, %s, 'inward', %s, %s, %s, %s, %s)""",
                    (pid, grn_date, qty,
                     f"[PM GRN: {grn_no}] {irm}".strip(), _user(), txn_vno, godown_id)
                )
            # Auto-create individual box records (one row per physical box)
            if nob > 0 and bc > 0:
                pcode_row = conn.execute(
                    "SELECT COALESCE(product_code,'') AS pc FROM pm_products WHERE id=%s", (pid,)
                ).fetchone()
                product_code = (pcode_row['pc'] if pcode_row else '') or ''
                _create_boxes_for_grn_item(
                    conn,
                    grn_id          = grn_id,
                    grn_no          = grn_no,
                    grn_item_id     = grn_item_id,
                    product_id      = pid,
                    product_code    = product_code,
                    no_of_box       = nob,
                    per_box_qty     = bc,
                    godown_id       = godown_id,
                    grn_date        = grn_date,
                    user            = _user(),
                    product_version = pver,
                )
                # Label-print log: each box printed at GRN time counts as a
                # 'grn_fresh' print. Box codes follow the standard pattern.
                try:
                    _label_print_record(
                        conn,
                        print_kind='grn_fresh',
                        voucher_kind='grn',
                        voucher_id=grn_id,
                        voucher_no=grn_no,
                        product_id=pid,
                        label_count=int(nob),
                    )
                except Exception: pass

        # Audit the GRN creation (single audit row for the whole voucher)
        try:
            _audit_record(
                conn,
                action='grn.create',
                entity='grn',
                entity_id=grn_id,
                summary=f"GRN {grn_no} · {len(items)} item(s) · supplier={supplier or '—'}",
                before=None,
                after={
                    'grn_no': grn_no, 'grn_date': str(grn_date),
                    'supplier': supplier, 'godown_id': godown_id,
                    'item_count': len(items),
                },
            )
        except Exception: pass

        # ── INVARIANT GUARD ─────────────────────────────────────────────────
        # Verify pm_boxes count matches pm_grn_items.no_of_box sum before
        # committing. Catches orphan/missing-box corruption (see history of
        # PM-GRN/0312 and PM-GRN/0342, 26-27). On mismatch, raises; the
        # outer except returns an error and the un-committed transaction is
        # auto-rolled-back when the connection closes. Healthy data: silent.
        _assert_grn_box_invariant(conn, grn_id)

        conn.commit()
        conn.close()
        return jsonify({
            'status':    'ok',
            'grn_no':    grn_no,
            'id':        grn_id,
            'fifo_codes': item_fifo,    # {product_id: fifo_code} for label stamping
            'item_ids':  new_item_ids,  # parallel to input items[]; None for skipped rows
            'verification_status': v_status,  # 'pending' triggers verify modal, 'verified' goes straight to list
        })
    except RuntimeError as _inv_err:
        # Invariant guard fired — log the technical detail to stderr and
        # return a friendly message to the user. Transaction auto-rolls-back
        # on conn.close() (we never called commit).
        import sys as _sys
        print(f"[pm_stock_routes] api_grn_save: {_inv_err}", file=_sys.stderr)
        conn.close()
        return jsonify({
            'status':  'error',
            'message': ('GRN data consistency check failed. The save was '
                        'cancelled to prevent corruption. Please retry; if '
                        'this persists, contact the developer with the GRN '
                        'number.')
        }), 500
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════
# GRN VERIFICATION — box-scan post-save flow
# ═══════════════════════════════════════════════════════════════════════════
#
# When the admin toggle `grn_verify_required` is ON, GRN save parks the
# voucher in 'pending' state with no inward stock posted. The operator
# then opens the Verify modal, scans every box, and on success the
# endpoint here posts the inward stock atomically and flips status to
# 'verified'. On mismatch (missing boxes, unknown codes, qty drift) a
# discrepancy report is recorded — stock is NEVER posted on mismatch.
#
# Endpoints:
#   GET  /api/pm_stock/grn/<id>/verify_status  — expected boxes + totals
#   POST /api/pm_stock/grn/<id>/verify         — commit verification
#   GET  /api/pm_stock/grn/discrepancy_report/<rid>/pdf  — download report
#
# Helpers (internal):
#   _next_disc_report_no   — per-GRN sequential report number
#   _record_grn_discrepancy — write report row (PDF best-effort)
# ═══════════════════════════════════════════════════════════════════════════

def _next_disc_report_no(conn, grn_no):
    """Generate a per-GRN sequential report number: DR-<grn_no>/<NN>.
    Unique within a GRN by counting prior reports. Format keeps the GRN
    visible in the report number so it's traceable in audit trails."""
    n = 1
    try:
        r = conn.execute(
            "SELECT COUNT(*) AS c FROM pm_grn_discrepancy_reports WHERE grn_no=%s",
            (grn_no,)
        ).fetchone()
        if r:
            try:
                n = int(r.get('c') or 0) + 1
            except Exception:
                n = int(r['c'] or 0) + 1
    except Exception:
        pass
    return f"DR-{grn_no}/{n:02d}"


def _record_grn_discrepancy(conn, *, grn_id, grn_no, mismatch_kind, payload, note=''):
    """Insert a discrepancy report row. PDF rendering is deferred / optional
    (reportlab may not be installed). Returns a dict the verify endpoint
    embeds in its response.

    The frontend's `lastReport` state uses:
        report_id, report_no, mismatch_kind, pdf_available, summary, kind_label

    payload schema (matches what verify endpoint builds):
        {
          expected_count, scanned_count,
          duplicates: [str], unknown: [str], missing: [str],
          scanned_codes: [str], expected_codes: [str],
          box_total, item_total
        }
    """
    report_no = _next_disc_report_no(conn, grn_no)
    try:
        cur = conn.execute(
            """INSERT INTO pm_grn_discrepancy_reports
                 (grn_id, grn_no, report_no, mismatch_kind, scanned_count,
                  expected_count, box_total, item_total, payload_json, note,
                  created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (grn_id, grn_no, report_no, mismatch_kind,
             int(payload.get('scanned_count') or 0),
             int(payload.get('expected_count') or 0),
             float(payload.get('box_total') or 0),
             float(payload.get('item_total') or 0),
             json.dumps(payload, default=str)[:16000000],
             (note or '')[:1000],
             _user())
        )
        report_id = cur.lastrowid
    except Exception as _e:
        import sys as _sys
        print(f"[pm_stock] discrepancy insert failed: {_e}", file=_sys.stderr)
        return {
            'report_id':     None,
            'report_no':     report_no,
            'mismatch_kind': mismatch_kind,
            'pdf_available': False,
        }

    # PDF rendering would happen here. Skipped for now since reportlab may
    # not be installed in this environment. The frontend already handles
    # `pdf_available=false` — it falls back to text-only WhatsApp sharing
    # and disables the Download button.
    kind_label = {
        'verification_mismatch': 'Box-set mismatch',
        'qty_mismatch':          'Quantity mismatch',
    }.get(mismatch_kind, mismatch_kind.replace('_', ' ').title())

    # Build a one-liner summary for the toast/share message.
    summary_bits = []
    if payload.get('missing'):    summary_bits.append(f"{len(payload['missing'])} missing")
    if payload.get('unknown'):    summary_bits.append(f"{len(payload['unknown'])} unknown")
    if payload.get('duplicates'): summary_bits.append(f"{len(payload['duplicates'])} duplicate")
    if mismatch_kind == 'qty_mismatch':
        summary_bits.append(f"qty {payload.get('box_total',0):.2f} vs {payload.get('item_total',0):.2f}")
    summary = ' · '.join(summary_bits) or 'see report'

    return {
        'report_id':     report_id,
        'report_no':     report_no,
        'mismatch_kind': mismatch_kind,
        'kind_label':    kind_label,
        'summary':       summary,
        'pdf_available': False,
    }


@pm_stock_bp.route('/api/pm_stock/grn/<int:grn_id>/verify_status', methods=['GET'])
@_login_required
def api_grn_verify_status(grn_id):
    """Return the data the Verify modal needs: expected box list + totals.

    Response shape:
        {
          status: 'ok',
          grn_no, verification_status, verified_at, verified_by,
          item_total_qty, total_qty,
          boxes: [ { box_code, product_id, product_name, per_box_qty, box_seq, total_boxes }, ... ]
        }
    """
    conn = sampling_portal.get_db_connection()
    try:
        g = conn.execute(
            """SELECT id, grn_no, grn_date,
                      COALESCE(verification_status,'verified') AS verification_status,
                      verified_at, COALESCE(verified_by,'') AS verified_by
                 FROM pm_grn WHERE id=%s""",
            (grn_id,)
        ).fetchone()
        if not g:
            conn.close()
            return jsonify({'status':'error','message':'GRN not found'}), 404

        boxes = conn.execute(
            """SELECT b.box_code, b.short_code, b.product_id, b.per_box_qty,
                      COALESCE(b.box_seq,0)     AS box_seq,
                      COALESCE(b.total_boxes,0) AS total_boxes,
                      COALESCE(p.product_name,'') AS product_name
                 FROM pm_boxes b
                 LEFT JOIN pm_products p ON p.id = b.product_id
                 WHERE b.grn_id=%s
                 ORDER BY b.product_id, b.box_seq""",
            (grn_id,)
        ).fetchall()
        items = conn.execute(
            "SELECT product_id, qty_received FROM pm_grn_items WHERE grn_id=%s",
            (grn_id,)
        ).fetchall()

        out_boxes = []
        sum_box_qty = 0.0
        for b in boxes:
            d = dict(b)
            d['per_box_qty'] = float(d.get('per_box_qty') or 0)
            sum_box_qty += d['per_box_qty']
            out_boxes.append(d)
        item_total = float(sum(float(i['qty_received'] or 0) for i in items))

        # Response shape: frontend (openGrnVerifyModal in pm_stock_grn_mtv.js)
        # destructures `d.grn` and reads `g.grn_no`, `g.boxes`, `g.item_total_qty`,
        # `g.total_qty`, `g.verification_status` from it. Wrap accordingly.
        # The duplicated top-level keys (grn_no, boxes, etc.) are kept for
        # any other caller that might read this endpoint without the wrapper.
        grn_payload = {
            'grn_no':              g['grn_no'],
            'verification_status': g['verification_status'],
            'verified_at':         str(g['verified_at']) if g.get('verified_at') else None,
            'verified_by':         g.get('verified_by') or '',
            'item_total_qty':      item_total,
            'total_qty':           sum_box_qty,
            'boxes':               out_boxes,
        }
        result = {
            'status': 'ok',
            'grn':    grn_payload,
            # Backwards-compat top-level copies (for any tooling that
            # introspected the previous shape).
            **grn_payload,
        }
        conn.close()
        return jsonify(result)
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/grn/<int:grn_id>/verify', methods=['POST'])
@_login_required
def api_grn_verify(grn_id):
    """Verify a pending GRN by matching scanned box codes to expected boxes,
    then post inward stock and flip status to 'verified'.

    Body:
        box_codes: [str, ...]
        note:      str (optional — added to discrepancy report on fail)

    On mismatch (missing/unknown/duplicate box codes OR qty drift) a
    discrepancy report row is created and the GRN stays 'pending' —
    stock is NEVER posted on mismatch.
    """
    d = request.get_json() or {}
    raw_codes = d.get('box_codes') or []
    note      = (d.get('note') or '').strip()
    if not isinstance(raw_codes, list):
        return jsonify({'status':'error','message':'box_codes must be a list'}), 400
    scanned = [str(c).strip().upper() for c in raw_codes if str(c).strip()]
    if not scanned:
        return jsonify({'status':'error','message':'No box codes provided. Scan all boxes for this GRN.'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        grn = conn.execute(
            """SELECT id, grn_no, grn_date, godown_id, supplier,
                      COALESCE(verification_status,'verified') AS verification_status
                 FROM pm_grn WHERE id=%s""",
            (grn_id,)
        ).fetchone()
        if not grn:
            conn.close()
            return jsonify({'status':'error','message':'GRN not found'}), 404
        if grn['verification_status'] == 'verified':
            conn.close()
            return jsonify({
                'status': 'error', 'code': 'already_verified',
                'message': f"GRN {grn['grn_no']} is already verified — stock has been posted."
            }), 409
        if grn['verification_status'] != 'pending':
            conn.close()
            return jsonify({
                'status': 'error',
                'message': f"GRN {grn['grn_no']} is in unexpected state '{grn['verification_status']}'."
            }), 409

        grn_no    = grn['grn_no']
        grn_date  = grn['grn_date']
        godown_id = grn['godown_id']

        expected_rows = conn.execute(
            "SELECT box_id, box_code, short_code, product_id, per_box_qty FROM pm_boxes WHERE grn_id=%s",
            (grn_id,)
        ).fetchall()
        # Build a code → box lookup that accepts BOTH the long box_code
        # (e.g. BEARTUBE34-G0161-B001) AND the compact short_code (e.g.
        # A0004187). Operators may scan either format depending on which
        # generation of QR label is stuck on the physical box. Both
        # codes resolve to the same expected box row.
        expected_by_code = {}
        for r in expected_rows:
            bc = (r['box_code'] or '').upper().strip()
            sc = (r['short_code'] or '').upper().strip()
            if bc: expected_by_code[bc] = r
            if sc: expected_by_code[sc] = r
        # expected_codes is the set of long box_codes — used downstream
        # for the "scanned == expected" check. We normalise every scanned
        # input back to its long box_code (via expected_by_code) before
        # adding to the scanned_set so short-code scans count as matches
        # for the corresponding box.
        expected_codes = { (r['box_code'] or '').upper().strip() for r in expected_rows if r['box_code'] }

        # Detect duplicates in scanned list AND normalise short-codes back
        # to long box_codes so the comparison below works correctly.
        seen, duplicates = set(), []
        normalised_scanned = []
        for c in scanned:
            cu = (c or '').upper().strip()
            # Resolve via short→long if present
            box = expected_by_code.get(cu)
            if box:
                cu = (box['box_code'] or '').upper().strip()
            normalised_scanned.append(cu)
            if cu in seen:
                duplicates.append(cu)
            else:
                seen.add(cu)
        scanned_set = seen
        unknown = sorted(scanned_set - expected_codes)
        missing = sorted(expected_codes - scanned_set)

        sum_box_qty = float(sum(float(r['per_box_qty'] or 0) for r in expected_rows))
        item_rows = conn.execute(
            "SELECT product_id, qty_received, COALESCE(remarks,'') AS remarks FROM pm_grn_items WHERE grn_id=%s",
            (grn_id,)
        ).fetchall()
        sum_item_qty = float(sum(float(r['qty_received'] or 0) for r in item_rows))

        # ── Check 1: box-set match ──
        if duplicates or unknown or missing:
            payload = {
                'expected_count':  len(expected_codes),
                'scanned_count':   len(scanned),
                'duplicates':      duplicates,
                'unknown':         unknown,
                'missing':         missing,
                'scanned_codes':   sorted(list(scanned_set)),
                'expected_codes':  sorted(list(expected_codes)),
                'box_total':       sum_box_qty,
                'item_total':      sum_item_qty,
            }
            report = _record_grn_discrepancy(
                conn, grn_id=grn_id, grn_no=grn_no,
                mismatch_kind='verification_mismatch', payload=payload, note=note,
            )
            conn.commit(); conn.close()
            return jsonify({
                'status': 'error', 'code': 'verification_mismatch',
                'message': 'Scanned boxes do not match GRN expected set. Discrepancy report generated.',
                'expected_count': len(expected_codes),
                'scanned_count':  len(scanned),
                'duplicates': duplicates, 'unknown': unknown, 'missing': missing,
                'report': report, 'grn_no': grn_no,
            }), 400

        # ── Check 2: quantity match (tolerance 0.001 for DECIMAL float surprises) ──
        if abs(sum_box_qty - sum_item_qty) > 0.001:
            payload = {
                'expected_count':  len(expected_codes),
                'scanned_count':   len(scanned),
                'duplicates':      [],
                'unknown':         [],
                'missing':         [],
                'scanned_codes':   sorted(list(scanned_set)),
                'expected_codes':  sorted(list(expected_codes)),
                'box_total':       sum_box_qty,
                'item_total':      sum_item_qty,
            }
            report = _record_grn_discrepancy(
                conn, grn_id=grn_id, grn_no=grn_no,
                mismatch_kind='qty_mismatch', payload=payload, note=note,
            )
            conn.commit(); conn.close()
            return jsonify({
                'status':  'error', 'code': 'qty_mismatch',
                'message': (f'Quantity mismatch: scanned boxes total {sum_box_qty:.2f} '
                            f'but GRN declared {sum_item_qty:.2f}. Discrepancy report generated.'),
                'box_total':  sum_box_qty, 'item_total': sum_item_qty,
                'report': report, 'grn_no': grn_no,
            }), 400

        # ── All checks passed — post inward stock atomically ──
        for item in item_rows:
            pid = item['product_id']
            qty = float(item['qty_received'] or 0)
            irm = (item['remarks'] or '').strip()
            txn_vno = _next_voucher_no(conn, 'PM-GTXN', grn_date)
            conn.execute(
                """INSERT INTO pm_godown_txn (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
                   VALUES (%s, %s, 'inward', %s, %s, %s, %s, %s)""",
                (pid, grn_date, qty,
                 f"[PM GRN: {grn_no}] {irm} (verified)".strip(), _user(), txn_vno, godown_id)
            )

        conn.execute(
            "UPDATE pm_grn SET verification_status='verified', verified_at=%s, verified_by=%s WHERE id=%s",
            (datetime.now(), _user(), grn_id)
        )

        try:
            _audit_record(
                conn,
                action='grn.verify',
                entity='grn',
                entity_id=grn_id,
                summary=f"GRN {grn_no} verified · {len(expected_codes)} boxes · {sum_item_qty:.2f} units posted",
                before={'verification_status': 'pending'},
                after={
                    'verification_status': 'verified',
                    'verified_by': _user(),
                    'box_count':   len(expected_codes),
                    'qty_total':   sum_item_qty,
                },
            )
        except Exception: pass

        conn.commit(); conn.close()
        return jsonify({
            'status':              'ok',
            'grn_no':              grn_no,
            'verification_status': 'verified',
            'box_count':           len(expected_codes),
            'qty_total':           sum_item_qty,
            'message':             f'✓ GRN {grn_no} verified — {sum_item_qty:.2f} units posted to inventory.'
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/grn/discrepancy_report/<int:report_id>/pdf', methods=['GET'])
@_login_required
def api_grn_discrepancy_pdf(report_id):
    """Download stub — PDF rendering is not yet wired in this build.
    The discrepancy data lives in pm_grn_discrepancy_reports.payload_json
    and the frontend already falls back to a text-only WhatsApp share
    when pdf_available is false. Returning 501 makes the failure explicit.
    """
    return jsonify({
        'status':'error',
        'message': 'PDF rendering not enabled in this build. The discrepancy data is stored in pm_grn_discrepancy_reports.payload_json; render via a reporting tool or copy the text from the modal.'
    }), 501


@pm_stock_bp.route('/api/pm_stock/fifo/lookup', methods=['POST'])
@_login_required
def api_fifo_lookup():
    """
    POST body:
      { "lots": [
          {"kind": "grn", "ref": <grn_id>,  "product_id": <pid>},
          {"kind": "op",  "ref": <op_seq>,  "product_id": <pid>},
          ...
        ] }

    Returns:
      { "status": "ok",
        "results": [
          {"kind": "grn", "ref": ..., "product_id": ..., "fifo_code": "A12"},
          ...
        ] }

    Idempotent — if a lot doesn't have a FIFO code yet (legacy data, e.g.
    GRN created before FIFO tracking existed), one is auto-assigned now.
    """
    d = request.get_json() or {}
    lots = d.get('lots') or []
    if not isinstance(lots, list):
        return jsonify({'status': 'error', 'message': 'lots must be a list'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        out = []
        for lot in lots:
            try:
                kind = (lot.get('kind') or '').lower()
                ref  = int(lot.get('ref') or 0)
                pid  = int(lot.get('product_id') or 0)
            except Exception:
                continue
            if kind not in ('grn', 'op') or ref <= 0 or pid <= 0:
                out.append({'kind': kind, 'ref': ref, 'product_id': pid, 'fifo_code': ''})
                continue
            fi = _get_or_assign_fifo(conn, kind, ref, pid)
            out.append({
                'kind':       kind,
                'ref':        ref,
                'product_id': pid,
                'fifo_code':  fi.get('fifo_code', '') or '',
                'fifo_seq':   fi.get('fifo_seq', 0)
            })
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'results': out})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/grn/<int:grn_id>/box_short_codes', methods=['GET'])
@_login_required
def api_grn_box_short_codes(grn_id):
    """Return a {product_id: {box_seq: short_code}} map for one GRN.

    Used by the first-print labels modal so it can encode the compact 8-char
    short_code into each QR (instead of the long box_code). The same map
    structure works for the multi-product GRN case — the JS picks the
    correct short_code per label by (productId, boxNum).

    Boxes without a short_code (legacy data created before this feature)
    show up as empty strings here; the JS falls back to the long box_code
    for those, which still scans correctly via the OR-clause in scan
    endpoints.

    Also returns `by_item`: { <grn_item_id>: [short_code_1, short_code_2,
    ...] } sorted by box_seq. This is the AUTHORITATIVE shape — the legacy
    `codes` map keyed by (product_id, box_seq) silently produced
    duplicate-QR labels when one product appeared on multiple GRN lines
    (e.g. one line of 37×220 plus one line of 1×50 — both lines start at
    a local 'box 1 of N' in the print loop, both look up codes[pid][1],
    and both get the SAME short_code). by_item is keyed by grn_item_id
    so each line gets its own array; the client just walks the array
    positionally.
    """
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(
            """SELECT product_id, grn_item_id, box_seq, short_code
               FROM pm_boxes
               WHERE grn_id=%s
               ORDER BY grn_item_id, box_seq""",
            (grn_id,)
        ).fetchall()
        codes   = {}
        by_item = {}
        for r in rows:
            pid = int(r['product_id'] or 0)
            iid = int(r['grn_item_id'] or 0)
            seq = int(r['box_seq'] or 0)
            sc  = (r.get('short_code') or '')
            if pid <= 0 or seq <= 0:
                continue
            codes.setdefault(pid, {})[seq] = sc
            if iid > 0:
                by_item.setdefault(iid, []).append(sc)
        conn.close()
        return jsonify({'status': 'ok', 'codes': codes, 'by_item': by_item})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/grn/delete', methods=['POST'])
@_login_required
def api_grn_delete():
    """Soft-delete a GRN. Snapshots the parent + line items + auto-posted
    inward txns + boxes into the recycle bin, then removes the originals.
    Admin-only. Restorable from the bin."""
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only — non-admins cannot delete GRNs.'}), 403
    d      = request.get_json() or {}
    grn_id = d.get('id')
    reason = (d.get('reason') or '').strip() or None
    if not grn_id:
        return jsonify({'status': 'error', 'message': 'id required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        g = conn.execute("SELECT grn_no, grn_date, supplier FROM pm_grn WHERE id=%s", (grn_id,)).fetchone()
        if not g:
            conn.close()
            return jsonify({'status': 'error', 'message': 'GRN not found'}), 404

        # Refuse to delete if any box from this GRN has been moved (out/in/consume).
        # Otherwise we'd orphan transfer history. (Same guard as before — the
        # bin entry would also be hard to restore if downstream movements
        # depend on it.)
        moved = conn.execute("""
            SELECT COUNT(*) AS c
            FROM pm_box_movements m
            JOIN pm_boxes b ON b.box_id = m.box_id
            WHERE b.grn_id = %s AND m.movement_type IN ('out','in','consume','adjust')
        """, (grn_id,)).fetchone()
        if int((moved or {}).get('c') or 0) > 0:
            conn.close()
            return jsonify({
                'status': 'error',
                'message': 'Cannot delete GRN — boxes from this GRN have already been moved or consumed. Reverse those movements first.'
            }), 409

        # Soft-delete: snapshot parent + all dependents into bin, then remove originals
        bin_id = _bin_soft_delete(
            conn,
            entity_type  = 'grn',
            entity_id    = grn_id,
            entity_label = f"GRN {g['grn_no']}",
            parent_table = 'pm_grn',
            parent_where = 'id=%s',
            parent_params= (grn_id,),
            children     = [
                {'table': 'pm_grn_items',
                 'where': 'grn_id=%s', 'params': (grn_id,)},
                {'table': 'pm_godown_txn',
                 'where': "remarks LIKE %s AND txn_date=%s AND txn_type='inward'",
                 'params': (f"[PM GRN: {g['grn_no']}]%", g['grn_date'])},
                {'table': 'pm_box_movements',
                 'where': 'box_id IN (SELECT box_id FROM pm_boxes WHERE grn_id=%s)',
                 'params': (grn_id,)},
                {'table': 'pm_boxes',
                 'where': 'grn_id=%s', 'params': (grn_id,)},
                # Attached invoice files (metadata rows). The actual files on
                # disk stay put — the relative_path is preserved in the bin
                # snapshot so a restore can re-link them seamlessly.
                {'table': 'pm_grn_files',
                 'where': 'grn_id=%s', 'params': (grn_id,)},
            ],
            summary = f"GRN {g['grn_no']} · supplier: {g.get('supplier') or '—'} · dated {g['grn_date']}",
            reason  = reason
        )

        try:
            _audit_record(
                conn,
                action='grn.delete',
                entity='grn',
                entity_id=grn_id,
                summary=f"Deleted GRN {g['grn_no']} (recycle bin id={bin_id})",
                before={'grn_no': g['grn_no'], 'grn_date': str(g['grn_date']),
                        'supplier': g.get('supplier'), 'bin_id': bin_id, 'reason': reason},
                after=None,
            )
        except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'bin_id': bin_id, 'message': 'GRN moved to recycle bin. Admin can restore from Recycle Bin.'})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/voucher_sequences')
@_login_required
def api_voucher_sequences():
    conn = sampling_portal.get_db_connection()
    rows = conn.execute("SELECT * FROM pm_voucher_sequences ORDER BY voucher_type").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@pm_stock_bp.route('/api/pm_stock/voucher_sequences/save', methods=['POST'])
@_login_required
def api_voucher_sequences_save():
    items = request.get_json() or []
    conn  = sampling_portal.get_db_connection()
    try:
        # Snapshot full table before changes for reversal
        try:
            before_rows = conn.execute(
                "SELECT voucher_type, prefix, last_num, pad_digits, reset_yearly "
                "FROM pm_voucher_sequences"
            ).fetchall()
            before_map = {r['voucher_type']: dict(r) for r in before_rows}
        except Exception:
            before_map = {}

        changed = []
        for item in items:
            vtype = item.get('voucher_type')
            if not vtype:
                continue
            conn.execute("""
                INSERT INTO pm_voucher_sequences (voucher_type, prefix, last_num, pad_digits, reset_yearly)
                VALUES (%s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    prefix=%s, pad_digits=%s, reset_yearly=%s
            """, (
                vtype,
                item.get('prefix', ''),
                item.get('last_num', 0),
                item.get('pad_digits', 4),
                int(item.get('reset_yearly', 1)),
                item.get('prefix', ''),
                item.get('pad_digits', 4),
                int(item.get('reset_yearly', 1)),
            ))
            new_state = {
                'voucher_type': vtype,
                'prefix': item.get('prefix',''),
                'pad_digits': int(item.get('pad_digits',4)),
                'reset_yearly': int(item.get('reset_yearly',1)),
            }
            old_state = before_map.get(vtype)
            # Only audit a real change (or a fresh insert)
            if not old_state or any(old_state.get(k) != new_state.get(k)
                                    for k in ('prefix','pad_digits','reset_yearly')):
                changed.append((vtype, old_state, new_state))

        try:
            for vtype, old_st, new_st in changed:
                _audit_record(
                    conn,
                    action='voucher_seq.update',
                    entity='voucher_sequence',
                    entity_id=vtype,
                    summary=f"Voucher numbering '{vtype}' updated",
                    before=old_st,
                    after=new_st,
                )
        except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/edit_txn', methods=['POST'])
@_login_required
def api_edit_txn():
    d      = request.get_json() or {}
    tid    = d.get('id')
    source = d.get('source')
    qty    = d.get('qty')
    txn_date = d.get('txn_date')
    remarks  = d.get('remarks', '')
    if not tid or source not in ('godown', 'floor'):
        return jsonify({'status': 'error', 'message': 'id and source required'}), 400
    table = 'pm_godown_txn' if source == 'godown' else 'pm_floor_txn'
    conn  = sampling_portal.get_db_connection()
    try:
        # Snapshot before
        before_row = conn.execute(
            f"SELECT id, product_id, txn_type, qty, txn_date, remarks FROM {table} WHERE id=%s", (tid,)
        ).fetchone()
        before_state = {k: (str(v) if hasattr(v,'isoformat') else v) for k,v in dict(before_row or {}).items()}

        updates, params = [], []
        if qty      is not None: updates.append("qty=%s");      params.append(float(qty))
        if txn_date is not None: updates.append("txn_date=%s"); params.append(txn_date)
        if remarks  is not None: updates.append("remarks=%s");  params.append(remarks)
        if not updates:
            conn.close(); return jsonify({'status': 'ok'})
        params.append(tid)
        conn.execute(f"UPDATE {table} SET {','.join(updates)} WHERE id=%s", params)

        try:
            after_state = {**before_state}
            if qty      is not None: after_state['qty']      = float(qty)
            if txn_date is not None: after_state['txn_date'] = str(txn_date)
            if remarks  is not None: after_state['remarks']  = remarks
            if before_state != after_state:
                _audit_record(
                    conn,
                    action='godown_txn.create' if source == 'godown' else 'floor_txn.create',
                    entity=table,
                    entity_id=tid,
                    summary=f"Edited {source} txn #{tid} ({before_state.get('txn_type','?')})",
                    before=before_state,
                    after=after_state,
                )
        except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/delete_txn', methods=['POST'])
@_login_required
def api_delete_txn():
    """Delete an individual stock txn row. Admin-only as of Phase 1.
    Phase 2 will convert this to soft-delete with restore-from-bin."""
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only — non-admins cannot delete stock transactions.'}), 403
    d      = request.get_json() or {}
    tid    = d.get('id')
    source = d.get('source')
    if not tid or source not in ('godown', 'floor'):
        return jsonify({'status': 'error', 'message': 'id and source required'}), 400
    table  = 'pm_godown_txn' if source == 'godown' else 'pm_floor_txn'
    conn   = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            f"SELECT product_id, txn_type, qty, txn_date FROM {table} WHERE id = %s", (tid,)
        ).fetchone()
        conn.execute(f"DELETE FROM {table} WHERE id = %s", (tid,))
        if row and source == 'godown' and row['txn_type'] == 'outward':
            conn.execute("""
                DELETE FROM pm_floor_txn
                WHERE product_id = %s AND txn_type = 'issue'
                  AND qty = %s AND txn_date = %s
                  AND remarks LIKE '[Auto: Issue from Godown]%%'
                ORDER BY id DESC LIMIT 1
            """, (row['product_id'], row['qty'], row['txn_date']))
        elif row and source == 'floor' and row['txn_type'] == 'pm_return':
            conn.execute("""
                DELETE FROM pm_godown_txn
                WHERE product_id = %s AND txn_type = 'inward'
                  AND qty = %s AND txn_date = %s
                  AND remarks LIKE '[Auto: Return from Floor]%%'
                ORDER BY id DESC LIMIT 1
            """, (row['product_id'], row['qty'], row['txn_date']))

        try:
            if row:
                snap = {k: (str(v) if hasattr(v,'isoformat') else v) for k,v in dict(row).items()}
                _audit_record(
                    conn,
                    action='godown_txn.delete' if source == 'godown' else 'floor_txn.delete',
                    entity=table,
                    entity_id=tid,
                    summary=f"Deleted {source} txn #{tid} ({snap.get('txn_type','?')}, qty={snap.get('qty')})",
                    before=snap,
                    after=None,
                )
        except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/delete_product', methods=['POST'])
@_login_required
def api_delete_product():
    """Soft-delete (deactivate) products. Open to all logged-in users.
    Also deletes the product's pm_godown_txn and pm_floor_txn ledger rows,
    so this is destructive. Every deactivation is recorded in the audit log
    with the user who performed it so the action can be traced after the
    fact (the deleted ledger rows themselves are NOT recoverable)."""
    d   = request.get_json() or {}
    ids = d.get('ids', [])
    if not ids:
        return jsonify({'status': 'error', 'message': 'No product ids'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        placeholders = ','.join(['%s'] * len(ids))

        # Snapshot for audit before deactivating
        try:
            rows_before = conn.execute(
                f"SELECT id, product_name, pm_type, brand_id, product_code FROM pm_products WHERE id IN ({placeholders})",
                ids
            ).fetchall()
            snapshot = [dict(r) for r in rows_before]
        except Exception:
            snapshot = []

        conn.execute(f"DELETE FROM pm_godown_txn WHERE product_id IN ({placeholders})", ids)
        conn.execute(f"DELETE FROM pm_floor_txn  WHERE product_id IN ({placeholders})", ids)
        conn.execute(f"UPDATE pm_products SET is_active=0 WHERE id IN ({placeholders})", ids)

        # Audit each deactivated product separately so reversal can target one row
        try:
            for snap in snapshot:
                _audit_record(
                    conn,
                    action='product.delete',
                    entity='product',
                    entity_id=snap.get('id'),
                    summary=f"Deactivated product '{snap.get('product_name','')}' [{snap.get('pm_type','')}]",
                    before=snap,
                    after={'is_active': 0},
                )
        except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/delete_product_permanent', methods=['POST'])
@_login_required
def api_delete_product_permanent():
    """PERMANENTLY delete products from the database — admin only.

    Safe by design: a product is hard-deleted ONLY if it has no dependent
    records (no boxes, no stock ledger rows, no material-request lines, no
    transfer lines, no FIFO lots, no material-lock rules). If anything
    references it, we refuse and report what's blocking — the caller should
    use the normal (soft) delete for products that have history.

    Body: { ids:[...], confirm:true }
    """
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only — non-admins cannot permanently delete products.'}), 403
    d = request.get_json() or {}
    ids = d.get('ids', [])
    if not ids:
        return jsonify({'status': 'error', 'message': 'No product ids'}), 400
    if not d.get('confirm'):
        return jsonify({'status': 'error', 'message': 'confirm:true required for permanent deletion.'}), 400

    # Tables that would orphan / break if the product vanished. (table, column)
    REF_CHECKS = [
        ('pm_boxes',                  'product_id', 'boxes'),
        ('pm_godown_txn',             'product_id', 'godown transactions'),
        ('pm_floor_txn',              'product_id', 'factory transactions'),
        ('pm_material_request_items', 'product_id', 'material-request lines'),
        ('pm_transfer_items',         'product_id', 'transfer lines'),
        ('pm_fifo_lots',              'product_id', 'FIFO lots'),
        ('pm_material_locks',         'product_id', 'material-lock rules'),
    ]

    conn = sampling_portal.get_db_connection()
    try:
        deleted, blocked = [], []
        for pid in ids:
            try:
                pid_int = int(pid)
            except (TypeError, ValueError):
                continue
            prow = conn.execute(
                "SELECT id, product_name, pm_type, brand_id, product_code FROM pm_products WHERE id=%s",
                (pid_int,)
            ).fetchone()
            if not prow:
                continue

            # Tally references across every dependent table.
            reasons = []
            for table, col, label in REF_CHECKS:
                try:
                    c = conn.execute(f"SELECT COUNT(*) AS n FROM {table} WHERE {col}=%s", (pid_int,)).fetchone()
                    n = int(c['n'] or 0) if c else 0
                    if n > 0:
                        reasons.append(f"{n} {label}")
                except Exception:
                    # If a table doesn't exist in this install, skip it safely.
                    pass

            if reasons:
                blocked.append({
                    'id': pid_int,
                    'product_name': prow['product_name'],
                    'reasons': reasons,
                })
                continue

            # Clean product → safe to hard-delete.
            conn.execute("DELETE FROM pm_products WHERE id=%s", (pid_int,))
            try:
                _audit_record(
                    conn, action='product.delete_permanent', entity='product',
                    entity_id=pid_int,
                    summary=f"PERMANENTLY deleted product '{prow['product_name']}' [{prow['pm_type']}]",
                    before=dict(prow), after=None,
                )
            except Exception: pass
            deleted.append({'id': pid_int, 'product_name': prow['product_name']})

        conn.commit()
        conn.close()
        msg_parts = []
        if deleted: msg_parts.append(f"{len(deleted)} permanently deleted")
        if blocked: msg_parts.append(f"{len(blocked)} kept (have history)")
        return jsonify({
            'status': 'ok',
            'deleted': deleted,
            'blocked': blocked,
            'message': ' · '.join(msg_parts) or 'Nothing to delete',
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/brands')
@_login_required
def api_pm_brands():
    conn = sampling_portal.get_db_connection()
    rows = conn.execute(
        "SELECT id, name, color FROM procurement_brands ORDER BY name"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@pm_stock_bp.route('/api/pm_stock/brands', methods=['POST'])
@_login_required
def api_pm_create_brand():
    d     = request.get_json() or {}
    name  = (d.get('name') or '').strip()
    color = (d.get('color') or '#6366f1').strip()
    if not name:
        return jsonify({'status': 'error', 'message': 'Brand name required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        cur = conn.execute(
            "INSERT INTO procurement_brands (name, color) VALUES (%s, %s)", (name, color)
        )
        conn.commit()
        brand_id = cur.lastrowid
        conn.close()
        return jsonify({'status': 'ok', 'id': brand_id, 'name': name, 'color': color})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/assign_brand', methods=['POST'])
@_login_required
def api_pm_assign_brand():
    # Open to all logged-in users. Bulk brand assignment was previously
    # admin-only but it's a routine catalog operation, not a privileged one
    # (it doesn't move stock or change ledgers). The BATCH_CAP below still
    # guards against accidental mass-overwrite from a stale "Select All".
    d        = request.get_json() or {}
    ids      = d.get('ids', [])
    brand_id = d.get('brand_id')
    confirm  = bool(d.get('confirm_large_batch'))   # explicit confirmation flag for big batches
    if not ids:
        return jsonify({'status': 'error', 'message': 'No product ids'}), 400

    # ── Safety cap: refuse very large batches unless explicitly confirmed.
    # This is a defence-in-depth against accidental mass-overwrite from a
    # stale "Select All" plus filter. The cap is intentionally generous so
    # legitimate bulk operations on a few hundred items go through.
    BATCH_CAP = 500
    try:
        ids = [int(x) for x in ids]
    except Exception:
        return jsonify({'status': 'error', 'message': 'ids must be integers'}), 400
    if len(ids) > BATCH_CAP and not confirm:
        return jsonify({
            'status':  'error',
            'message': f'Batch too large ({len(ids)} products). Maximum without explicit confirmation is {BATCH_CAP}. ' \
                       f'If this is intentional, narrow the filter or split into smaller batches.',
            'requested': len(ids), 'cap': BATCH_CAP
        }), 413

    conn = sampling_portal.get_db_connection()
    try:
        placeholders = ','.join(['%s'] * len(ids))

        # Snapshot before-state per product so reversal can put them back
        try:
            rows_before = conn.execute(
                f"SELECT id, product_name, brand_id FROM pm_products WHERE id IN ({placeholders})",
                ids
            ).fetchall()
            before_map = {r['id']: dict(r) for r in rows_before}
        except Exception:
            before_map = {}

        new_bid = int(brand_id) if brand_id else None
        if brand_id:
            conn.execute(
                f"UPDATE pm_products SET brand_id=%s WHERE id IN ({placeholders})",
                ([brand_id] + list(ids))
            )
        else:
            conn.execute(
                f"UPDATE pm_products SET brand_id=NULL WHERE id IN ({placeholders})", list(ids)
            )

        # Audit per product so reversal can target one row at a time
        try:
            for pid in ids:
                snap = before_map.get(pid)
                if not snap: continue
                if (snap.get('brand_id') or None) == new_bid: continue  # no change
                _audit_record(
                    conn,
                    action='product.brand_assign',
                    entity='product',
                    entity_id=pid,
                    summary=f"Brand assigned for '{snap.get('product_name','')}' "
                            f"(brand_id {snap.get('brand_id') or 'None'} → {new_bid or 'None'})",
                    before={'brand_id': snap.get('brand_id')},
                    after={'brand_id': new_bid},
                )
        except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'updated': len(ids)})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/products/assign_uom', methods=['POST'])
@_login_required
def api_pm_assign_uom():
    """Bulk-assign Primary UOM (and optionally Alternate UOM + conversion) to
    a selection of products. Admin only. Mirrors api_pm_assign_brand for the
    selection/large-batch/audit pattern.

    Body:
      ids: [int]                       — products to update (required)
      primary_uom: str                 — required, e.g. 'Nos', 'Kg' (free text,
                                         preserves casing). Defaults to 'Nos'.
      alt_uom: str | null              — optional. If provided, ratio is required.
      alt_to_primary_ratio: number     — required when alt_uom is set; means
                                         "1 primary = ratio × alternate"
                                         (Tally style). Must be > 0.
      clear_alt: bool                  — when true, sets alt_uom + ratio to NULL
                                         regardless of what's passed; Primary is
                                         left UNCHANGED on the row (so users can
                                         drop just the alternate without losing
                                         the primary they already set).
      confirm_large_batch: bool        — required when ids > 500.
    """
    if not _is_admin():
        return jsonify({'status':'error',
                        'message':'Only an admin can bulk-assign UOM.'}), 403
    d        = request.get_json() or {}
    ids      = d.get('ids', [])
    clear_alt = bool(d.get('clear_alt'))
    confirm  = bool(d.get('confirm_large_batch'))
    if not ids:
        return jsonify({'status':'error','message':'No product ids'}), 400
    try:
        ids = [int(x) for x in ids]
    except Exception:
        return jsonify({'status':'error','message':'ids must be integers'}), 400

    BATCH_CAP = 500
    if len(ids) > BATCH_CAP and not confirm:
        return jsonify({
            'status':'error',
            'message': f'Batch too large ({len(ids)} products). Maximum without explicit confirmation is {BATCH_CAP}.',
            'requested': len(ids), 'cap': BATCH_CAP
        }), 413

    # Parse + validate UOM fields.
    primary_uom = (d.get('primary_uom') or 'Nos').strip() or 'Nos'
    if clear_alt:
        alt_uom = None; alt_ratio = None
    else:
        alt_raw = (d.get('alt_uom') or '').strip()
        alt_uom = alt_raw or None
        alt_ratio = None
        if alt_uom:
            try:
                alt_ratio = float(d.get('alt_to_primary_ratio'))
            except (TypeError, ValueError):
                return jsonify({'status':'error',
                                'message':'alt_to_primary_ratio must be a number when alt_uom is set'}), 400
            if alt_ratio <= 0:
                return jsonify({'status':'error',
                                'message':'alt_to_primary_ratio must be greater than zero'}), 400
            if alt_uom.lower() == primary_uom.lower():
                return jsonify({'status':'error',
                                'message':'Alternate UOM must differ from primary UOM'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        placeholders = ','.join(['%s'] * len(ids))

        # Snapshot before-state per product (for audit + change detection).
        rows_before = conn.execute(
            f"""SELECT id, product_name,
                       COALESCE(primary_uom,'Nos') AS primary_uom,
                       COALESCE(alt_uom,'')        AS alt_uom,
                       alt_to_primary_ratio        AS alt_ratio
                  FROM pm_products WHERE id IN ({placeholders})""",
            ids
        ).fetchall()
        before_map = {r['id']: dict(r) for r in rows_before}

        # Apply.
        if clear_alt:
            # Leave primary_uom alone (user only asked to clear alt).
            conn.execute(
                f"""UPDATE pm_products
                       SET alt_uom=NULL, alt_to_primary_ratio=NULL
                     WHERE id IN ({placeholders})""",
                list(ids)
            )
        else:
            conn.execute(
                f"""UPDATE pm_products
                       SET primary_uom=%s, alt_uom=%s, alt_to_primary_ratio=%s
                     WHERE id IN ({placeholders})""",
                [primary_uom, alt_uom, alt_ratio] + list(ids)
            )

        # Per-product audit so a future reversal can revert one row at a time.
        updated_count = 0
        for pid in ids:
            snap = before_map.get(pid)
            if not snap: continue
            old = {
                'primary_uom': snap.get('primary_uom') or 'Nos',
                'alt_uom':     snap.get('alt_uom') or None,
                'alt_to_primary_ratio': float(snap['alt_ratio']) if snap.get('alt_ratio') is not None else None,
            }
            if clear_alt:
                new = {'primary_uom': old['primary_uom'], 'alt_uom': None, 'alt_to_primary_ratio': None}
            else:
                new = {'primary_uom': primary_uom, 'alt_uom': alt_uom, 'alt_to_primary_ratio': alt_ratio}
            if old == new:
                continue  # no change, skip audit noise
            updated_count += 1
            try:
                _audit_record(
                    conn,
                    action='product.uom_assign',
                    entity='product',
                    entity_id=pid,
                    summary=f"UOM updated for '{snap.get('product_name','')}'",
                    before=old, after=new,
                )
            except Exception:
                pass

        conn.commit()
        conn.close()
        return jsonify({'status':'ok',
                        'updated_count': updated_count,
                        'total_ids': len(ids)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/set_min_stock', methods=['POST'])
@_login_required
def api_set_min_stock():
    d   = request.get_json() or {}
    pid = d.get('product_id')
    qty = d.get('min_stock')
    if pid is None or qty is None:
        return jsonify({'status': 'error', 'message': 'product_id and min_stock required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        # Snapshot the previous min_stock for audit
        prev_row = conn.execute(
            "SELECT product_name, COALESCE(min_stock,0) AS min_stock FROM pm_products WHERE id=%s", (pid,)
        ).fetchone()
        prev_qty = int((prev_row or {}).get('min_stock', 0) or 0)
        prev_name = (prev_row or {}).get('product_name', '')

        conn.execute("UPDATE pm_products SET min_stock=%s WHERE id=%s", (int(qty), pid))

        if prev_qty != int(qty):
            try:
                _audit_record(
                    conn,
                    action='product.threshold',
                    entity='product',
                    entity_id=pid,
                    summary=f"Min stock for '{prev_name}' changed: {prev_qty} → {int(qty)}",
                    before={'min_stock': prev_qty},
                    after={'min_stock': int(qty)},
                )
            except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/update_product', methods=['POST'])
@_login_required
def api_update_product():
    d    = request.get_json() or {}
    pid  = d.get('id')
    name = (d.get('product_name') or '').strip()
    pm   = (d.get('pm_type')      or '').strip()
    bid  = d.get('brand_id')
    # UOM (Phase 1) — same validation as Add Product (preserves user casing).
    primary_uom = (d.get('primary_uom') or 'Nos').strip() or 'Nos'
    alt_uom_raw = (d.get('alt_uom') or '').strip()
    alt_uom     = alt_uom_raw or None
    alt_ratio   = None
    if alt_uom:
        try:
            alt_ratio = float(d.get('alt_to_primary_ratio'))
        except (TypeError, ValueError):
            return jsonify({'status':'error',
                            'message':'alt_to_primary_ratio must be a number when alt_uom is set'}), 400
        if alt_ratio <= 0:
            return jsonify({'status':'error',
                            'message':'alt_to_primary_ratio must be greater than zero'}), 400
        if alt_uom.lower() == primary_uom.lower():
            return jsonify({'status':'error',
                            'message':'Alternate UOM must differ from primary UOM'}), 400
    if not pid or not name or not pm:
        return jsonify({'status': 'error', 'message': 'id, product_name, pm_type required'}), 400
    if not bid:
        return jsonify({'status': 'error', 'message': 'Brand is required (used to generate product code)'}), 400

    is_admin = (session.get('User_Type', '').lower() == 'admin')

    conn = sampling_portal.get_db_connection()
    try:
        # Read current row to detect changes
        cur = conn.execute(
            "SELECT product_name, pm_type, brand_id, product_code FROM pm_products WHERE id=%s", (pid,)
        ).fetchone()
        if not cur:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Product not found'}), 404

        before_state = {
            'product_name': cur['product_name'],
            'pm_type':      cur['pm_type'],
            'brand_id':     cur['brand_id'],
            'product_code': cur['product_code'] or '',
        }

        brand_name = _brand_name_by_id(conn, bid)
        if not brand_name:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Selected brand not found'}), 400

        # Duplicate guard on rename: block renaming into an existing product
        # (whitespace/case-insensitive), excluding this row itself.
        dup = _find_duplicate_product(conn, name, pm, exclude_id=int(pid))
        if dup:
            conn.close()
            return jsonify({
                'status':  'error',
                'message': (f"Another product already has this name: "
                            f"\"{dup['product_name']}\" ({dup['product_code'] or 'no code'}). "
                            "Rename blocked to avoid a duplicate.")
            }), 409
        # Brand or PM Type, because either change forces a code regeneration.
        # Non-admins can still rename the product (name change keeps the code).
        existing_code  = cur['product_code'] or ''
        brand_changed  = str(cur['brand_id'] or '') != str(bid or '')
        pm_changed     = (cur['pm_type'] or '') != pm
        if existing_code and (brand_changed or pm_changed) and not is_admin:
            conn.close()
            return jsonify({
                'status':  'error',
                'message': 'Only admins can change Brand or PM Type for a product that already has a code (would force regenerate). Ask an administrator if needed.'
            }), 403

        new_code = existing_code
        # Regenerate when: no code yet, OR brand changed, OR pm_type changed
        if (not new_code) or brand_changed or pm_changed:
            new_code = _generate_product_code(conn, brand_name, pm, exclude_id=pid)

        # ── Safety: product_code change would invalidate printed box labels ──
        # If a product_code is changing AND there are existing pm_boxes whose
        # box_code starts with the old code, those labels (already stuck on
        # physical boxes) would silently stop matching the DB. Block the
        # change unless the caller explicitly opts in via `force_code_change`.
        # This is what caused the "Box ... not found" scan failures on GRN-0008
        # where M65/M95 labels were printed but the DB later said M64/M08.
        if existing_code and new_code != existing_code:
            affected = conn.execute(
                """SELECT COUNT(*) AS n FROM pm_boxes
                   WHERE product_id=%s
                     AND product_code=%s
                     AND current_status='in_stock'""",
                (pid, existing_code)
            ).fetchone()
            n_affected = int(affected['n'] or 0) if affected else 0
            force = bool(d.get('force_code_change'))
            if n_affected > 0 and not force:
                conn.close()
                return jsonify({
                    'status':  'error',
                    'code':    'product_code_change_blocked',
                    'message': (
                        f"Changing brand/PM type would regenerate product_code "
                        f"from '{existing_code}' to '{new_code}'. There are "
                        f"{n_affected} in-stock box(es) with labels printed using "
                        f"the old code — those physical labels would stop scanning. "
                        f"Either: (1) consume those boxes first, (2) keep brand/pm "
                        f"unchanged, or (3) re-submit with force_code_change=true "
                        f"and plan to reprint+restick those box labels."
                    ),
                    'old_code':       existing_code,
                    'new_code':       new_code,
                    'affected_boxes': n_affected,
                }), 409

        conn.execute(
            """UPDATE pm_products
                  SET product_name=%s, pm_type=%s, brand_id=%s, product_code=%s,
                      primary_uom=%s, alt_uom=%s, alt_to_primary_ratio=%s
                WHERE id=%s""",
            (name, pm, bid, new_code, primary_uom, alt_uom, alt_ratio, pid)
        )

        after_state = {
            'product_name': name, 'pm_type': pm,
            'brand_id': int(bid), 'product_code': new_code,
        }

        # Audit trail — only record if something actually changed
        if before_state != after_state:
            try:
                _audit_record(
                    conn,
                    action='product.update',
                    entity='product',
                    entity_id=pid,
                    summary=f"Edited product '{name}' [{pm}]",
                    before=before_state,
                    after=after_state,
                )
            except Exception: pass

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'product_code': new_code})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/products/regenerate_codes', methods=['POST'])
@_login_required
def api_pm_regenerate_codes():
    """
    Regenerate product codes for a specific list of product IDs passed in
    the request body. Touches ONLY pm_products.product_code — never
    brand_id, never any other column. Skips products without a brand_id.

    Available to all logged-in users.

    Request: {"product_ids": [1, 2, 3, ...]}
    """
    d   = request.get_json() or {}
    ids = d.get('product_ids') or []
    if not isinstance(ids, list) or not ids:
        return jsonify({'status': 'error', 'message': 'product_ids (non-empty list) required'}), 400

    # Sanitise IDs to integers
    try:
        ids = [int(x) for x in ids if str(x).strip().isdigit() or isinstance(x, int)]
    except Exception:
        return jsonify({'status': 'error', 'message': 'product_ids must be integers'}), 400
    if not ids:
        return jsonify({'status': 'error', 'message': 'No valid product_ids supplied'}), 400

    conn = sampling_portal.get_db_connection()
    generated = 0
    skipped_no_brand = 0
    failed    = 0
    errors    = []
    try:
        placeholders = ','.join(['%s'] * len(ids))
        rows = conn.execute(f"""
            SELECT p.id, p.product_name, p.pm_type, p.brand_id,
                   COALESCE(b.name,'') AS brand_name
            FROM pm_products p
            LEFT JOIN procurement_brands b ON b.id = p.brand_id
            WHERE p.id IN ({placeholders}) AND p.is_active=1
        """, ids).fetchall()

        for r in rows:
            pid = r['id']
            try:
                brand_name = (r['brand_name'] or '').strip()
                pm_type    = (r['pm_type'] or '').strip()
                if not r['brand_id'] or not brand_name:
                    skipped_no_brand += 1
                    errors.append({'id': pid, 'name': r['product_name'], 'reason': 'No brand assigned'})
                    continue
                if not pm_type:
                    failed += 1
                    errors.append({'id': pid, 'name': r['product_name'], 'reason': 'Missing PM type'})
                    continue
                code = _generate_product_code(conn, brand_name, pm_type, exclude_id=pid)
                # IMPORTANT: only product_code is touched. Nothing else.
                conn.execute("UPDATE pm_products SET product_code=%s WHERE id=%s", (code, pid))
                generated += 1
            except Exception as ie:
                failed += 1
                errors.append({'id': pid, 'name': r['product_name'], 'reason': str(ie)})
        conn.commit()
        conn.close()
        return jsonify({
            'status':           'ok',
            'requested':        len(ids),
            'generated':        generated,
            'skipped_no_brand': skipped_no_brand,
            'failed':           failed,
            'errors':           errors[:30],
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e), 'generated': generated}), 500

@pm_stock_bp.route('/api/pm_stock/products/by_code')
@_login_required
def api_pm_product_by_code():
    """Look up a single product by its 10-char product code (used by QR scanner)."""
    code = (request.args.get('code') or '').strip().upper()
    if not code:
        return jsonify({'status': 'error', 'message': 'code required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute("""
            SELECT p.id, p.product_name, p.pm_type,
                   COALESCE(p.product_code,'') AS product_code,
                   COALESCE(p.brand_id,0)      AS brand_id,
                   COALESCE(b.name,'')         AS brand_name,
                   COALESCE(b.color,'')        AS brand_color
            FROM pm_products p
            LEFT JOIN procurement_brands b ON b.id = p.brand_id
            WHERE p.product_code=%s AND p.is_active=1
            LIMIT 1
        """, (code,)).fetchone()
        conn.close()
        if not row:
            return jsonify({'status': 'not_found'})
        return jsonify({'status': 'ok', 'product': dict(row)})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/boxes/by_code')
@_login_required
def api_box_by_code():
    """
    Look up a box by its full box_code (e.g. 'BEARTUBE12-G0234-B003').
    Used by the QR scanner to resolve a scanned box to its product / qty / location.
    """
    code = (request.args.get('code') or '').strip().upper()
    if not code:
        return jsonify({'status': 'error', 'message': 'code required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute("""
            SELECT b.*,
                   p.product_name, p.pm_type,
                   COALESCE(br.name,'')          AS brand_name,
                   COALESCE(br.color,'')         AS brand_color,
                   COALESCE(g.name,'')           AS current_godown_name,
                   COALESCE(g.type,'')           AS current_godown_type
            FROM pm_boxes b
            JOIN pm_products p           ON p.id = b.product_id
            LEFT JOIN procurement_brands br ON br.id = p.brand_id
            LEFT JOIN procurement_godowns g ON g.id = b.current_godown_id
            WHERE b.short_code = %s OR b.box_code = %s
            LIMIT 1
        """, (code, code)).fetchone()
        conn.close()
        if not row:
            return jsonify({'status': 'not_found'})
        d = dict(row)
        # Normalise types for JSON
        for k in ('per_box_qty',):
            if d.get(k) is not None:
                d[k] = float(d[k])
        for k in ('created_at', 'updated_at'):
            if d.get(k) is not None:
                d[k] = str(d[k])
        return jsonify({'status': 'ok', 'box': d})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/label_reissue/request', methods=['POST'])
@_login_required
def api_label_reissue_request():
    """A user requests a REPLACEMENT label for a box whose QR is damaged /
    won't scan. Admin approval is required; on approval a brand-new short
    code is stamped on the box and the requester prints the new label.

    Body: { code (str) } OR { box_id (int) }, plus reason (str, required),
    plus optional new_per_box_qty (number — proposed corrected per-box qty).
    """
    d = request.get_json() or {}
    code   = (d.get('code') or '').strip().upper()
    box_id = d.get('box_id')
    reason = (d.get('reason') or '').strip()[:500]
    new_pbq = d.get('new_per_box_qty')
    new_gid = d.get('new_godown_id')
    if not reason:
        return jsonify({'status': 'error', 'message': 'A reason is required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        ok, msg, req_id = _label_reissue_insert_one(conn, code, box_id, reason, new_pbq, new_gid)
        if not ok:
            conn.close()
            return jsonify({'status': 'error', 'message': msg}), 409
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'req_id': req_id})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


def _label_reissue_insert_one(conn, code, box_id, reason, new_per_box_qty=None, new_godown_id=None):
    """Insert ONE pending label-reissue request. Shared by the single and
    bulk request endpoints. Caller manages the transaction + HTTP shaping.
    Returns (ok: bool, message: str, req_id: int|None).

    `new_per_box_qty` (optional): a corrected per-box quantity the requester
    proposes. Captured here as a pending change; applied to the box only when
    an admin approves. None / blank / equal to current = no change.

    `new_godown_id` (optional): a corrected location. On approval the box is
    moved there AND balancing ledger txns are posted at both locations (a
    real box move). None / blank / equal to current = no change.
    """
    code   = (code or '').strip().upper()
    reason = (reason or '').strip()[:500]
    if not reason:
        return False, 'A reason is required for every box', None

    brow = None
    if box_id:
        try:
            box_id = int(box_id)
        except (TypeError, ValueError):
            return False, 'box_id must be an integer', None
        brow = conn.execute(
            """SELECT b.box_id, b.box_code, b.short_code, b.grn_no, b.product_id,
                      b.per_box_qty, b.current_godown_id, p.product_name
               FROM pm_boxes b JOIN pm_products p ON p.id=b.product_id
               WHERE b.box_id=%s LIMIT 1""", (box_id,)
        ).fetchone()
    elif code:
        brow = conn.execute(
            """SELECT b.box_id, b.box_code, b.short_code, b.grn_no, b.product_id,
                      b.per_box_qty, b.current_godown_id, p.product_name
               FROM pm_boxes b JOIN pm_products p ON p.id=b.product_id
               WHERE b.short_code=%s OR b.box_code=%s LIMIT 1""", (code, code)
        ).fetchone()
    else:
        return False, 'code or box_id is required', None

    if not brow:
        return False, f'Box not found ({code or box_id})', None
    bid = int(brow['box_id'])
    old_qty = float(brow['per_box_qty'] or 0)
    old_gid = brow['current_godown_id']

    # Parse the optional proposed quantity. Only store it when it's a valid
    # positive number that differs from the current per-box qty.
    new_qty = None
    if new_per_box_qty is not None and str(new_per_box_qty).strip() != '':
        try:
            cand = float(new_per_box_qty)
        except (TypeError, ValueError):
            return False, 'new_per_box_qty must be a number', None
        if cand <= 0:
            return False, 'new_per_box_qty must be greater than 0', None
        if abs(cand - old_qty) > 1e-9:
            new_qty = cand

    # Parse the optional proposed location. Only store it when it's a valid,
    # active godown that differs from the current one.
    new_gid = None
    if new_godown_id is not None and str(new_godown_id).strip() != '':
        try:
            gcand = int(new_godown_id)
        except (TypeError, ValueError):
            return False, 'new_godown_id must be an integer', None
        if old_gid is None or gcand != int(old_gid):
            grow = conn.execute(
                "SELECT id FROM procurement_godowns WHERE id=%s LIMIT 1", (gcand,)
            ).fetchone()
            if not grow:
                return False, f'Location #{gcand} not found', None
            new_gid = gcand

    dupe = conn.execute(
        """SELECT req_id FROM pm_label_reissue_requests
           WHERE box_id=%s AND requested_by=%s AND status IN ('pending','approved')
           ORDER BY requested_at DESC LIMIT 1""",
        (bid, _user())
    ).fetchone()
    if dupe:
        return False, f'Box {brow["box_code"]} already has an open request (#{dupe["req_id"]})', None

    cur = conn.execute(
        """INSERT INTO pm_label_reissue_requests
             (box_id, box_code, old_short_code, product_id, product_name,
              grn_no, requested_by, reason, status, old_per_box_qty, new_per_box_qty,
              old_godown_id, new_godown_id)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'pending',%s,%s,%s,%s)""",
        (bid, brow['box_code'], (brow['short_code'] or None),
         brow['product_id'], brow['product_name'], brow['grn_no'],
         _user(), reason, old_qty, new_qty,
         (int(old_gid) if old_gid is not None else None), new_gid)
    )
    req_id = cur.lastrowid
    try:
        _qty_note = (f"; qty {old_qty:g} → {new_qty:g}" if new_qty is not None else "")
        _loc_note = (f"; location → #{new_gid}" if new_gid is not None else "")
        _audit_record(
            conn, action='box.reissue_label', entity='label_reissue_request',
            entity_id=req_id,
            summary=f"Label reissue requested for box {brow['box_code']}{_qty_note}{_loc_note}",
            before=None,
            after={'box_code': brow['box_code'], 'reason': reason,
                   'old_per_box_qty': old_qty, 'new_per_box_qty': new_qty,
                   'old_godown_id': (int(old_gid) if old_gid is not None else None),
                   'new_godown_id': new_gid},
        )
    except Exception:
        pass
    return True, 'ok', req_id


@pm_stock_bp.route('/api/pm_stock/label_reissue/request_bulk', methods=['POST'])
@_login_required
def api_label_reissue_request_bulk():
    """Submit reissue requests for MULTIPLE boxes at once. Each item carries
    its own reason and identifies the box by code or box_id.

    Body: { items: [ { box_id|code, reason }, ... ] }

    Each item is processed independently; the response carries a per-item
    result. All successful inserts commit together.
    """
    d = request.get_json() or {}
    items = d.get('items') or []
    if not isinstance(items, list) or not items:
        return jsonify({'status': 'error', 'message': 'items must be a non-empty list'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        results = []
        created_ids = []
        for it in items:
            if not isinstance(it, dict):
                results.append({'ok': False, 'message': 'Invalid item'}); continue
            ok, msg, rid = _label_reissue_insert_one(
                conn, it.get('code'), it.get('box_id'), it.get('reason'),
                it.get('new_per_box_qty'), it.get('new_godown_id')
            )
            results.append({'box_id': it.get('box_id'), 'code': it.get('code'),
                            'ok': ok, 'message': msg, 'req_id': rid})
            if ok:
                created_ids.append(rid)
        conn.commit()
        conn.close()
        return jsonify({
            'status':      'ok',
            'created':     len(created_ids),
            'skipped':     len(results) - len(created_ids),
            'total':       len(results),
            'created_ids': created_ids,
            'per_item':    results,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/label_reissue/requests', methods=['GET'])
@_login_required
def api_label_reissue_list():
    """List reissue requests. Admins see all; non-admins see their own."""
    status_f = (request.args.get('status') or '').strip().lower()
    where, params = [], []
    if status_f:
        where.append("r.status=%s"); params.append(status_f)
    is_admin = _is_admin()
    if not is_admin:
        where.append("r.requested_by=%s"); params.append(_user())
    wc = (' WHERE ' + ' AND '.join(where)) if where else ''
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(f"""
            SELECT r.req_id, r.box_id, r.box_code, r.old_short_code, r.new_short_code,
                   r.product_id, r.product_name, r.grn_no, r.requested_by, r.requested_at,
                   r.reason, r.status, r.decided_by, r.decided_at, r.decided_note,
                   r.printed_at, r.printed_by, r.old_per_box_qty, r.new_per_box_qty,
                   r.old_godown_id, r.new_godown_id,
                   COALESCE(og.name,'') AS old_godown_name,
                   COALESCE(ng.name,'') AS new_godown_name,
                   b.current_godown_id AS cur_godown_id,
                   COALESCE(cg.name,'') AS cur_godown_name,
                   b.current_status AS box_status
            FROM pm_label_reissue_requests r
            LEFT JOIN procurement_godowns og ON og.id = r.old_godown_id
            LEFT JOIN procurement_godowns ng ON ng.id = r.new_godown_id
            LEFT JOIN pm_boxes b ON b.box_id = r.box_id
            LEFT JOIN procurement_godowns cg ON cg.id = b.current_godown_id
            {wc}
            ORDER BY
              CASE r.status WHEN 'pending' THEN 0 WHEN 'approved' THEN 1 ELSE 2 END,
              r.requested_at DESC
            LIMIT 300
        """, params).fetchall()
        conn.close()
        out = []
        for r in rows:
            dd = dict(r)
            for k in ('requested_at', 'decided_at', 'printed_at'):
                if dd.get(k) is not None: dd[k] = str(dd[k])
            for k in ('old_per_box_qty', 'new_per_box_qty'):
                if dd.get(k) is not None:
                    try: dd[k] = float(dd[k])
                    except Exception: dd[k] = None
            out.append(dd)
        return jsonify({'status': 'ok', 'requests': out, 'count': len(out)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/label_reissue/pending_count', methods=['GET'])
@_login_required
def api_label_reissue_pending_count():
    """Badge poll. Admins: pending count. Non-admins: own pending+approved
    (i.e. still actionable by them — approved means 'ready to print')."""
    conn = sampling_portal.get_db_connection()
    try:
        if _is_admin():
            row = conn.execute(
                "SELECT COUNT(*) AS c FROM pm_label_reissue_requests WHERE status='pending'"
            ).fetchone()
        else:
            row = conn.execute(
                """SELECT COUNT(*) AS c FROM pm_label_reissue_requests
                   WHERE requested_by=%s AND status IN ('pending','approved')""",
                (_user(),)
            ).fetchone()
        conn.close()
        return jsonify({'status': 'ok', 'count': int((row or {}).get('c') or 0)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/label_reissue/<int:req_id>/approve', methods=['POST'])
@_login_required
def api_label_reissue_approve(req_id):
    """Admin approves a reissue request. Stamps a fresh short code on the box
    NOW and records it, so the requester can print the replacement label."""
    if not _user_has_access('label_reissue'):
        return jsonify({'status': 'error', 'message': 'You do not have access to Label Reissue approvals'}), 403
    d = request.get_json() or {}
    note = (d.get('note') or '').strip()[:500] or None
    conn = sampling_portal.get_db_connection()
    try:
        ok, msg, new_code = _label_reissue_approve_one(conn, req_id, note)
        if not ok:
            conn.close()
            return jsonify({'status': 'error', 'message': msg}), 409
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'req_id': req_id, 'new_code': new_code})
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


def _label_reissue_approve_one(conn, req_id, note):
    """Approve a single pending reissue request: stamp a fresh short code on
    the box and record it. If the request proposed a corrected per-box qty,
    update the box AND post a balancing stock movement so godown/factory
    stock stays correct. Shared by single + bulk approve endpoints. Caller
    manages the transaction. Returns (ok, message, new_code|None)."""
    row = conn.execute(
        """SELECT box_id, box_code, status, product_id, new_per_box_qty,
                  old_per_box_qty, new_godown_id, old_godown_id
           FROM pm_label_reissue_requests WHERE req_id=%s""",
        (req_id,)
    ).fetchone()
    if not row:
        return False, f'Request #{req_id} not found', None
    if row['status'] != 'pending':
        return False, f"Request #{req_id} is '{row['status']}', not pending", None

    new_code, old_code = _reissue_box_short_code(conn, int(row['box_id']))

    # ── Apply a proposed per-box quantity change (if any) ───────────────
    # IMPORTANT: a per-box-qty correction is a DATA correction of an
    # existing box's recorded count — the physical box does NOT move and no
    # goods enter or leave a location. We therefore update the box value
    # ONLY and deliberately do NOT post a godown/floor transaction.
    #
    # Why no ledger row: posting one through _post_stock_movement would have
    # to be expressed as inward/outward (godown) or issue/dispatch (floor) —
    # there is no neutral "adjustment" type in those enums — which mislabels
    # a count fix as goods moving and pollutes the dispatched/inward totals
    # (this caused a spurious "Factory Dispatch" row + downstream confusion).
    #
    # The new per_box_qty flows into every FUTURE transfer automatically,
    # because OUT/IN scans read the box's live per_box_qty at scan time.
    qty_applied = None
    if row['new_per_box_qty'] is not None:
        try:
            box = conn.execute(
                """SELECT box_id, product_id, per_box_qty, current_godown_id,
                          current_status
                   FROM pm_boxes WHERE box_id=%s LIMIT 1""",
                (int(row['box_id']),)
            ).fetchone()
            if box:
                cur_qty = float(box['per_box_qty'] or 0)
                new_qty = float(row['new_per_box_qty'])
                if abs(new_qty - cur_qty) > 1e-9:
                    conn.execute(
                        "UPDATE pm_boxes SET per_box_qty=%s WHERE box_id=%s",
                        (new_qty, int(row['box_id']))
                    )
                    qty_applied = (cur_qty, new_qty)
        except Exception as _qe:
            import sys as _sys
            print(f"[pm_stock] reissue qty change failed (req {req_id}): {_qe}", file=_sys.stderr)

    # ── Apply a proposed location change (if any) ───────────────────────
    # Unlike a qty correction, a location change IS a real box move: the box
    # leaves one location's books and enters another's. Location stock totals
    # come from the ledger, so we MUST post balancing movements — outward at
    # the old location, inward at the new — exactly as a transfer would, plus
    # a pm_box_movements row. We use the box's live per_box_qty for the move.
    loc_applied = None
    if row['new_godown_id'] is not None:
        try:
            box = conn.execute(
                """SELECT box_id, product_id, per_box_qty, current_godown_id,
                          current_status, box_code
                   FROM pm_boxes WHERE box_id=%s LIMIT 1""",
                (int(row['box_id']),)
            ).fetchone()
            if box:
                from_gid = box['current_godown_id']
                to_gid   = int(row['new_godown_id'])
                if from_gid is None or int(from_gid) != to_gid:
                    move_qty = float(box['per_box_qty'] or 0)
                    import datetime as _dt
                    _today = _dt.date.today().isoformat()
                    # Post OUT at the old location (only if it had one and the
                    # box was live stock there — otherwise there's nothing to
                    # remove from that location's books).
                    if from_gid is not None and box['current_status'] == 'in_stock':
                        try:
                            _post_stock_movement(
                                conn, product_id=int(box['product_id']),
                                godown_id=int(from_gid), qty=move_qty, direction='out',
                                transfer_no=f"PM-RELOC/{req_id}", transfer_id=None,
                                txn_date=_today, user=_user(),
                            )
                        except Exception as _pe:
                            import sys as _sys
                            print(f"[pm_stock] reloc OUT post failed (req {req_id}): {_pe}", file=_sys.stderr)
                    # Post IN at the new location.
                    try:
                        _post_stock_movement(
                            conn, product_id=int(box['product_id']),
                            godown_id=to_gid, qty=move_qty, direction='in',
                            transfer_no=f"PM-RELOC/{req_id}", transfer_id=None,
                            txn_date=_today, user=_user(),
                        )
                    except Exception as _pe:
                        import sys as _sys
                        print(f"[pm_stock] reloc IN post failed (req {req_id}): {_pe}", file=_sys.stderr)
                    # Move the box + log a box-movement row (the box-history trail).
                    conn.execute(
                        "UPDATE pm_boxes SET current_godown_id=%s, current_status='in_stock' WHERE box_id=%s",
                        (to_gid, int(row['box_id']))
                    )
                    try:
                        conn.execute(
                            """INSERT INTO pm_box_movements
                                 (box_id, transfer_id, movement_type, from_godown_id,
                                  to_godown_id, qty, moved_by, remarks)
                               VALUES (%s,NULL,'in',%s,%s,%s,%s,%s)""",
                            (int(row['box_id']), (int(from_gid) if from_gid is not None else None),
                             to_gid, move_qty, _user(),
                             f"Location corrected via reissue #{req_id}")
                        )
                    except Exception:
                        pass
                    loc_applied = (from_gid, to_gid)
        except Exception as _le:
            import sys as _sys
            print(f"[pm_stock] reissue location change failed (req {req_id}): {_le}", file=_sys.stderr)

    conn.execute(
        """UPDATE pm_label_reissue_requests
           SET status='approved', decided_by=%s, decided_at=NOW(),
               decided_note=%s, new_short_code=%s, old_short_code=%s
           WHERE req_id=%s""",
        (_user(), note, new_code, old_code, req_id)
    )
    try:
        _qn = (f"; qty {qty_applied[0]:g} → {qty_applied[1]:g}" if qty_applied else "")
        _ln = (f"; location #{loc_applied[0]} → #{loc_applied[1]}" if loc_applied else "")
        _audit_record(
            conn, action='box.reissue_label', entity='label_reissue_request',
            entity_id=req_id,
            summary=f"Label reissue #{req_id} approved for box {row['box_code']}: {old_code or '(none)'} → {new_code}{_qn}{_ln}",
            before={'status': 'pending', 'short_code': old_code,
                    'per_box_qty': (qty_applied[0] if qty_applied else None),
                    'godown_id': (loc_applied[0] if loc_applied else None)},
            after={'status': 'approved', 'short_code': new_code, 'note': note,
                   'per_box_qty': (qty_applied[1] if qty_applied else None),
                   'godown_id': (loc_applied[1] if loc_applied else None)},
        )
    except Exception:
        pass
    return True, 'ok', new_code


@pm_stock_bp.route('/api/pm_stock/label_reissue/approve_bulk', methods=['POST'])
@_login_required
def api_label_reissue_approve_bulk():
    """Admin approves multiple pending reissue requests at once. Each gets a
    fresh short code stamped on its box. Body: {req_ids:[...], note?}.
    All successful approvals commit together."""
    if not _user_has_access('label_reissue'):
        return jsonify({'status': 'error', 'message': 'You do not have access to Label Reissue approvals'}), 403
    d = request.get_json() or {}
    raw_ids = d.get('req_ids') or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return jsonify({'status': 'error', 'message': 'req_ids must be a non-empty list'}), 400
    try:
        req_ids = sorted(set(int(x) for x in raw_ids))
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'req_ids must contain integers'}), 400
    note = (d.get('note') or '').strip()[:500] or None

    conn = sampling_portal.get_db_connection()
    try:
        results, approved_ids = [], []
        for rid in req_ids:
            ok, msg, new_code = _label_reissue_approve_one(conn, rid, note)
            results.append({'req_id': rid, 'ok': ok, 'message': msg, 'new_code': new_code})
            if ok:
                approved_ids.append(rid)
        if approved_ids:
            try:
                _audit_record(
                    conn, action='box.reissue_label', entity='label_reissue_request',
                    entity_id=None,
                    summary=f"Bulk approved {len(approved_ids)} label reissue(s): {approved_ids}",
                    before=None, after={'approved_ids': approved_ids, 'note': note},
                )
            except Exception:
                pass
        conn.commit()
        conn.close()
        return jsonify({
            'status':      'ok',
            'approved':    len(approved_ids),
            'skipped':     len(req_ids) - len(approved_ids),
            'total':       len(req_ids),
            'per_request': results,
        })
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/label_reissue/<int:req_id>/reject', methods=['POST'])
@_login_required
def api_label_reissue_reject(req_id):
    """Admin rejects a pending reissue request."""
    if not _user_has_access('label_reissue'):
        return jsonify({'status': 'error', 'message': 'You do not have access to Label Reissue approvals'}), 403
    d = request.get_json() or {}
    note = (d.get('note') or '').strip()[:500] or None
    conn = sampling_portal.get_db_connection()
    try:
        ok, msg = _label_reissue_reject_one(conn, req_id, note)
        if not ok:
            conn.close()
            return jsonify({'status': 'error', 'message': msg}), 409
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


def _label_reissue_reject_one(conn, req_id, note):
    """Reject a single pending reissue request. Shared by single + bulk.
    Caller manages the transaction. Returns (ok, message)."""
    row = conn.execute(
        "SELECT status, box_code FROM pm_label_reissue_requests WHERE req_id=%s", (req_id,)
    ).fetchone()
    if not row:
        return False, f'Request #{req_id} not found'
    if row['status'] != 'pending':
        return False, f"Request #{req_id} is '{row['status']}', not pending"
    conn.execute(
        """UPDATE pm_label_reissue_requests
           SET status='rejected', decided_by=%s, decided_at=NOW(), decided_note=%s
           WHERE req_id=%s""",
        (_user(), note, req_id)
    )
    try:
        _audit_record(
            conn, action='box.reissue_label', entity='label_reissue_request',
            entity_id=req_id,
            summary=f"Label reissue #{req_id} rejected for box {row['box_code']}",
            before={'status': 'pending'}, after={'status': 'rejected', 'note': note},
        )
    except Exception:
        pass
    return True, 'ok'


@pm_stock_bp.route('/api/pm_stock/label_reissue/reject_bulk', methods=['POST'])
@_login_required
def api_label_reissue_reject_bulk():
    """Admin rejects multiple pending reissue requests at once.
    Body: {req_ids:[...], note?}."""
    if not _user_has_access('label_reissue'):
        return jsonify({'status': 'error', 'message': 'You do not have access to Label Reissue approvals'}), 403
    d = request.get_json() or {}
    raw_ids = d.get('req_ids') or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return jsonify({'status': 'error', 'message': 'req_ids must be a non-empty list'}), 400
    try:
        req_ids = sorted(set(int(x) for x in raw_ids))
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'req_ids must contain integers'}), 400
    note = (d.get('note') or '').strip()[:500] or None

    conn = sampling_portal.get_db_connection()
    try:
        results, rejected_ids = [], []
        for rid in req_ids:
            ok, msg = _label_reissue_reject_one(conn, rid, note)
            results.append({'req_id': rid, 'ok': ok, 'message': msg})
            if ok:
                rejected_ids.append(rid)
        if rejected_ids:
            try:
                _audit_record(
                    conn, action='box.reissue_label', entity='label_reissue_request',
                    entity_id=None,
                    summary=f"Bulk rejected {len(rejected_ids)} label reissue(s): {rejected_ids}",
                    before=None, after={'rejected_ids': rejected_ids, 'note': note},
                )
            except Exception:
                pass
        conn.commit()
        conn.close()
        return jsonify({
            'status':      'ok',
            'rejected':    len(rejected_ids),
            'skipped':     len(req_ids) - len(rejected_ids),
            'total':       len(req_ids),
            'per_request': results,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/label_reissue/<int:req_id>/print', methods=['POST'])
@_login_required
def api_label_reissue_print(req_id):
    """Requester (or admin) fetches the assembled replacement-label payload
    for an APPROVED request and marks it printed. The label encodes the new
    short code stamped at approval time."""
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            """SELECT req_id, box_id, box_code, new_short_code, status, requested_by
               FROM pm_label_reissue_requests WHERE req_id=%s""", (req_id,)
        ).fetchone()
        if not row:
            conn.close(); return jsonify({'status': 'error', 'message': 'Request not found'}), 404
        # Only the requester or an admin may print.
        if not _is_admin() and row['requested_by'] != _user():
            conn.close(); return jsonify({'status': 'error', 'message': 'Not your request'}), 403

        is_admin = _is_admin()
        # One-time print for the requester: once it's been printed, a
        # non-admin cannot print it again. Admins may still reprint (e.g. if
        # the first physical print jammed/misfed).
        if row['status'] == 'printed' and not is_admin:
            conn.close(); return jsonify({
                'status': 'error',
                'message': 'This replacement label has already been printed. Ask an admin if you need it reprinted.'
            }), 409
        if row['status'] not in ('approved', 'printed'):
            conn.close(); return jsonify({
                'status': 'error',
                'message': f"Request is '{row['status']}' — only approved requests can be printed."
            }), 409

        label = _assemble_box_label(conn, int(row['box_id']))
        if not label:
            conn.close(); return jsonify({'status': 'error', 'message': 'Box not found'}), 404
        # Force the QR payload to the freshly-assigned code recorded on the request.
        if row['new_short_code']:
            label['new_short_code'] = row['new_short_code']

        # Mark printed on the FIRST print (approved → printed). Admin reprints
        # of an already-printed request leave the original printed_at/by intact.
        if row['status'] == 'approved':
            conn.execute(
                """UPDATE pm_label_reissue_requests
                   SET status='printed', printed_at=NOW(), printed_by=%s
                   WHERE req_id=%s AND status='approved'""",
                (_user(), req_id)
            )
            try:
                _audit_record(
                    conn, action='box.reissue_label', entity='label_reissue_request',
                    entity_id=req_id,
                    summary=f"Replacement label #{req_id} printed for box {row['box_code']} ({row['new_short_code'] or ''})",
                    before={'status': 'approved'}, after={'status': 'printed', 'printed_by': _user()},
                )
            except Exception:
                pass
            conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'label': label,
                        'new_code': row['new_short_code'] or label['new_short_code']})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/label_reissue/cleanup_stray_txns', methods=['POST'])
@_login_required
def api_label_reissue_cleanup_stray_txns():
    """One-time repair (admin only): remove the spurious godown/floor
    transaction rows an earlier version of the qty-correction feature posted
    with voucher_no like 'PM-REISSUE/...'. Those rows mislabelled a per-box
    count correction as an inward/outward/dispatch movement and skewed
    location stock totals. A per-box-qty correction no longer posts any
    ledger row, so these stray rows can be safely deleted.

    GET-style preview: pass {"preview": true} to see counts without deleting.
    """
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    d = request.get_json() or {}
    preview = bool(d.get('preview'))
    conn = sampling_portal.get_db_connection()
    try:
        gd = conn.execute(
            "SELECT COUNT(*) AS c FROM pm_godown_txn WHERE voucher_no LIKE 'PM-REISSUE/%%'"
        ).fetchone()
        fl = conn.execute(
            "SELECT COUNT(*) AS c FROM pm_floor_txn WHERE voucher_no LIKE 'PM-REISSUE/%%'"
        ).fetchone()
        gd_n = int((gd or {}).get('c') or 0)
        fl_n = int((fl or {}).get('c') or 0)
        if preview:
            conn.close()
            return jsonify({'status': 'ok', 'preview': True,
                            'godown_rows': gd_n, 'floor_rows': fl_n, 'total': gd_n + fl_n})
        conn.execute("DELETE FROM pm_godown_txn WHERE voucher_no LIKE 'PM-REISSUE/%%'")
        conn.execute("DELETE FROM pm_floor_txn  WHERE voucher_no LIKE 'PM-REISSUE/%%'")
        try:
            _audit_record(
                conn, action='box.reissue_label', entity='label_reissue_request',
                entity_id=None,
                summary=f"Cleaned up {gd_n + fl_n} stray PM-REISSUE stock txn row(s) (qty-correction repair)",
                before={'godown_rows': gd_n, 'floor_rows': fl_n}, after={'deleted': gd_n + fl_n},
            )
        except Exception:
            pass
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'deleted': gd_n + fl_n,
                        'godown_rows': gd_n, 'floor_rows': fl_n})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/boxes/list')
@_login_required
def api_box_list():
    """
    List boxes with filters. Optional query params:
      grn_id, grn_no, product_id, godown_id, status, limit (default 200)
    """
    grn_id     = request.args.get('grn_id')
    grn_no     = (request.args.get('grn_no') or '').strip()
    product_id = request.args.get('product_id')
    godown_id  = request.args.get('godown_id')
    status     = (request.args.get('status') or '').strip()
    try:
        limit = max(1, min(int(request.args.get('limit') or 200), 1000))
    except Exception:
        limit = 200

    where, params = [], []
    if grn_id:     where.append("b.grn_id=%s");     params.append(int(grn_id))
    if grn_no:     where.append("b.grn_no=%s");     params.append(grn_no)
    if product_id: where.append("b.product_id=%s"); params.append(int(product_id))
    if godown_id:  where.append("b.current_godown_id=%s"); params.append(int(godown_id))
    if status:     where.append("b.current_status=%s");    params.append(status)
    sql_where = (' WHERE ' + ' AND '.join(where)) if where else ''

    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(f"""
            SELECT b.box_id, b.box_code, b.grn_id, b.grn_no, b.product_id, b.product_code,
                   b.box_seq, b.total_boxes, b.per_box_qty, b.current_godown_id,
                   b.current_status, b.created_at, b.updated_at,
                   p.product_name, p.pm_type,
                   COALESCE(g.name,'') AS current_godown_name
            FROM pm_boxes b
            JOIN pm_products p ON p.id = b.product_id
            LEFT JOIN procurement_godowns g ON g.id = b.current_godown_id
            {sql_where}
            ORDER BY b.box_id DESC
            LIMIT %s
        """, params + [limit]).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            d['per_box_qty'] = float(d.get('per_box_qty') or 0)
            for k in ('created_at', 'updated_at'):
                if d.get(k) is not None:
                    d[k] = str(d[k])
            out.append(d)
        return jsonify({'status': 'ok', 'boxes': out, 'count': len(out)})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/boxes/<int:box_id>/movements')
@_login_required
def api_box_movements(box_id):
    """Return the full movement history for a single box (newest first)."""
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute("""
            SELECT m.*,
                   COALESCE(fg.name,'') AS from_name,
                   COALESCE(tg.name,'') AS to_name,
                   COALESCE(t.transfer_no,'') AS transfer_no
            FROM pm_box_movements m
            LEFT JOIN procurement_godowns fg ON fg.id = m.from_godown_id
            LEFT JOIN procurement_godowns tg ON tg.id = m.to_godown_id
            LEFT JOIN pm_transfers t        ON t.transfer_id = m.transfer_id
            WHERE m.box_id = %s
            ORDER BY m.movement_at DESC, m.movement_id DESC
        """, (box_id,)).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            d['qty'] = float(d.get('qty') or 0)
            if d.get('movement_at') is not None:
                d['movement_at'] = str(d['movement_at'])
            out.append(d)
        return jsonify({'status': 'ok', 'movements': out})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/boxes/backfill', methods=['POST'])
@_login_required
def api_box_backfill():
    """
    ADMIN-ONLY. Generate pm_boxes records for all historical GRN items that
    don't already have boxes (no_of_box > 0, box_count > 0). Locations are
    set to the original GRN godown — this is a best-effort starting point;
    boxes that have already been moved historically via MTV won't reflect
    those movements (MTV history is not box-aware).

    Request body (optional): {"dry_run": true}
    Response: {generated_grns, generated_boxes, skipped, details: [...]}
    """
    if session.get('User_Type', '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403

    d        = request.get_json() or {}
    dry_run  = bool(d.get('dry_run'))

    conn = sampling_portal.get_db_connection()
    generated_grns  = 0
    generated_boxes = 0
    skipped         = 0
    details         = []
    try:
        # Find candidate GRN items: nob>0, bc>0, and no boxes already exist for that item
        rows = conn.execute("""
            SELECT gi.id          AS grn_item_id,
                   gi.grn_id      AS grn_id,
                   gi.product_id  AS product_id,
                   gi.no_of_box   AS no_of_box,
                   gi.box_count   AS per_box_qty,
                   g.grn_no       AS grn_no,
                   g.grn_date     AS grn_date,
                   g.godown_id    AS godown_id,
                   COALESCE(p.product_code, '') AS product_code
            FROM pm_grn_items gi
            JOIN pm_grn      g ON g.id = gi.grn_id
            JOIN pm_products p ON p.id = gi.product_id
            WHERE gi.no_of_box > 0
              AND gi.box_count > 0
              AND NOT EXISTS (
                  SELECT 1 FROM pm_boxes b WHERE b.grn_item_id = gi.id
              )
            ORDER BY g.id ASC, gi.id ASC
        """).fetchall()

        for r in rows:
            if not r['product_code']:
                skipped += 1
                details.append({
                    'grn_no':   r['grn_no'],
                    'grn_item_id': r['grn_item_id'],
                    'reason':   'product has no product_code; assign one first'
                })
                continue
            if dry_run:
                generated_boxes += int(r['no_of_box'] or 0)
                generated_grns  += 1
                continue
            created = _create_boxes_for_grn_item(
                conn,
                grn_id      = r['grn_id'],
                grn_no      = r['grn_no'],
                grn_item_id = r['grn_item_id'],
                product_id  = r['product_id'],
                product_code= r['product_code'],
                no_of_box   = r['no_of_box'],
                per_box_qty = r['per_box_qty'],
                godown_id   = r['godown_id'],
                grn_date    = r['grn_date'],
                user        = _user()
            )
            generated_boxes += created
            if created > 0:
                generated_grns += 1

        if not dry_run:
            conn.commit()
        conn.close()
        return jsonify({
            'status':          'ok',
            'dry_run':         dry_run,
            'generated_grns':  generated_grns,
            'generated_boxes': generated_boxes,
            'skipped':         skipped,
            'details':         details[:50]
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/op_batches/list', methods=['GET'])
@_login_required
def api_pm_op_batches_list():
    """List Opening Stock label batches that have been generated previously,
    so users can request a reprint for a specific batch.

    Query params:
      search    : optional substring matched against op_label, product code,
                  or product name
      from_date : optional YYYY-MM-DD; matches the box's pm_box_movements
                  earliest 'adjust' date (when the batch was created)
      to_date   : optional YYYY-MM-DD
      limit     : optional, default 200, max 500

    Response: { status:'ok', batches: [
      { op_seq, op_label, product_id, product_code, product_name, pm_type,
        brand_name, godown_id, godown_name, no_of_box, per_box_qty, total_qty,
        created_at, created_by }
    ] }
    """
    search   = (request.args.get('search')    or '').strip()
    fromd    = (request.args.get('from_date') or '').strip()
    tod      = (request.args.get('to_date')   or '').strip()
    try:
        lim = max(1, min(500, int(request.args.get('limit') or 200)))
    except Exception:
        lim = 200

    where, params = ["b.grn_no LIKE 'PM-OP/%%'"], []
    if search:
        where.append("(b.grn_no LIKE %s OR p.product_code LIKE %s OR p.product_name LIKE %s)")
        sp = f"%{search}%"; params.extend([sp, sp, sp])

    conn = sampling_portal.get_db_connection()
    try:
        # Aggregate by (grn_no = OP label, product_id, godown) — each batch
        # corresponds to one product at one location. no_of_box is COUNT of
        # boxes; per_box_qty is taken from MAX (they are uniform within a batch).
        rows = conn.execute(f"""
            SELECT b.grn_no                    AS op_label,
                   b.product_id,
                   COALESCE(b.product_code,'') AS product_code,
                   p.product_name,
                   COALESCE(p.pm_type,'')      AS pm_type,
                   COALESCE(br.name,'')        AS brand_name,
                   b.current_godown_id         AS godown_id,
                   COALESCE(g.name,'')         AS godown_name,
                   COUNT(*)                    AS no_of_box,
                   MAX(b.per_box_qty)          AS per_box_qty,
                   SUM(b.per_box_qty)          AS total_qty,
                   MIN(b.created_at)           AS created_at,
                   MIN(b.created_by)           AS created_by
            FROM pm_boxes b
            JOIN pm_products p ON p.id = b.product_id
            LEFT JOIN procurement_brands br ON br.id = p.brand_id
            LEFT JOIN procurement_godowns g  ON g.id  = b.current_godown_id
            WHERE {' AND '.join(where)}
            GROUP BY b.grn_no, b.product_id, b.current_godown_id,
                     b.product_code, p.product_name, p.pm_type, br.name, g.name
            ORDER BY MIN(b.created_at) DESC
            LIMIT %s
        """, (*params, lim)).fetchall()
        conn.close()

        out = []
        for r in rows:
            d = dict(r)
            # Normalise types
            d['no_of_box']   = int(d.get('no_of_box') or 0)
            d['per_box_qty'] = float(d.get('per_box_qty') or 0)
            d['total_qty']   = float(d.get('total_qty')   or 0)
            if d.get('created_at') is not None: d['created_at'] = str(d['created_at'])
            # Extract op_seq from "PM-OP/0042"
            try:
                d['op_seq'] = int(str(d['op_label']).split('/')[-1])
            except Exception:
                d['op_seq'] = 0
            # Optional date filter (applied client-side here against created_at)
            if fromd and d.get('created_at') and d['created_at'][:10] < fromd: continue
            if tod   and d.get('created_at') and d['created_at'][:10] > tod:   continue
            out.append(d)

        # Bulk-fetch FIFO codes for all (op_seq, product_id) pairs in one query.
        # Idempotent — does not assign new codes for legacy batches; assignment
        # happens lazily on first reprint. Empty string for unassigned.
        if out:
            keys = [(r['op_seq'], r['product_id']) for r in out
                    if r.get('op_seq') and r.get('product_id')]
            if keys:
                # Build IN clause efficiently
                placeholders = ','.join(['(%s,%s)'] * len(keys))
                flat = []
                for op_s, pid in keys: flat.extend([op_s, pid])
                conn2 = sampling_portal.get_db_connection()
                try:
                    fifo_rows = conn2.execute(
                        f"""SELECT lot_ref AS op_seq, product_id, fifo_code
                            FROM pm_fifo_lots
                            WHERE lot_kind='op' AND (lot_ref, product_id) IN ({placeholders})""",
                        tuple(flat)
                    ).fetchall()
                    fifo_map = {(int(r['op_seq']), int(r['product_id'])): r['fifo_code']
                                for r in fifo_rows}
                    for r in out:
                        r['fifo_code'] = fifo_map.get(
                            (int(r.get('op_seq') or 0), int(r.get('product_id') or 0)),
                            '')
                finally:
                    conn2.close()

        return jsonify({'status': 'ok', 'batches': out, 'count': len(out)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/op_batches/payload', methods=['GET'])
@_login_required
def api_pm_op_batch_payload():
    """Reconstruct the printOpeningLabels() payload for an existing OP batch.
    Used by the redeem-and-print flow so reprints fire the existing label
    rendering function with the correct data.

    Query params: op_seq=<int>, product_id=<int>, godown_id=<int>
       (all three needed: same op_seq can have multiple products if batches
        were appended; godown disambiguates when boxes have moved.)

    Response: shape compatible with printOpeningLabels — includes box_codes
    array, op_label, product details, godown_id.
    """
    try:
        op_seq    = int(request.args.get('op_seq'))
        product_id = int(request.args.get('product_id'))
        godown_id  = int(request.args.get('godown_id'))
    except Exception:
        return jsonify({'status':'error','message':'op_seq, product_id, godown_id (all integers) required'}), 400

    op_label = f'PM-OP/{op_seq:04d}'

    conn = sampling_portal.get_db_connection()
    try:
        prod = conn.execute("""
            SELECT p.product_name, COALESCE(p.product_code,'') AS code,
                   COALESCE(p.pm_type,'')   AS pm_type,
                   COALESCE(b.name,'')      AS brand_name
            FROM pm_products p
            LEFT JOIN procurement_brands b ON b.id = p.brand_id
            WHERE p.id=%s
        """, (product_id,)).fetchone()
        if not prod:
            conn.close(); return jsonify({'status':'error','message':'Product not found'}), 404

        boxes = conn.execute("""
            SELECT box_code, short_code, per_box_qty, current_godown_id
            FROM pm_boxes
            WHERE grn_no=%s AND product_id=%s
            ORDER BY box_seq
        """, (op_label, product_id)).fetchall()
        if not boxes:
            conn.close(); return jsonify({'status':'error','message':'No boxes found for this OP batch + product'}), 404

        # Pick the godown's boxes only (current_godown_id may differ from the
        # original if some boxes have moved). For reprint, we render labels
        # for boxes that still belong to the requested godown.
        selected = [b for b in boxes if int(b['current_godown_id']) == godown_id]
        # Fallback: if no boxes match (all moved), use all of them — better
        # to print labels than to fail.
        target = selected if selected else boxes

        codes        = [b['box_code']  for b in target]
        # Parallel short_codes aligned 1:1 with codes. Legacy boxes have NULL
        # short_code — the JS treats those as "use box_code in the QR" so
        # old reprints still scan via the OR-clause in scan endpoints.
        short_codes  = [(b.get('short_code') or '') for b in target]
        pbq   = float(target[0]['per_box_qty'] or 0)

        # Find the original creation date / remarks from the first movement
        first_mv = conn.execute("""
            SELECT m.movement_at, m.remarks
            FROM pm_box_movements m
            JOIN pm_boxes b ON b.box_id = m.box_id
            WHERE b.grn_no=%s AND b.product_id=%s
            ORDER BY m.movement_id ASC
            LIMIT 1
        """, (op_label, product_id)).fetchone()
        op_date = ''
        remarks = ''
        if first_mv:
            try: op_date = str(first_mv['movement_at'])[:10]
            except Exception: pass
            try: remarks = first_mv['remarks'] or ''
            except Exception: pass

        # FIFO code for this lot — assign on the fly if missing (legacy data)
        fifo_code = ''
        try:
            fi = _get_or_assign_fifo(conn, 'op', op_seq, product_id)
            fifo_code = fi.get('fifo_code', '') or ''
            conn.commit()
        except Exception:
            pass

        conn.close()
        return jsonify({
            'status':       'ok',
            'op_label':     op_label,
            'op_seq':       op_seq,
            'product_id':   product_id,
            'product_code': prod['code'],
            'product_name': prod['product_name'],
            'pm_type':      prod['pm_type']    or '',
            'brand_name':   prod['brand_name'] or '',
            'godown_id':    godown_id,
            'no_of_box':    len(codes),
            'per_box_qty':  pbq,
            'total_qty':    pbq * len(codes),
            'box_codes':    codes,
            'short_codes':  short_codes,
            'op_date':      op_date,
            'remarks':      remarks,
            'fifo_code':    fifo_code
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/op_batches/<int:op_seq>/reprint_with_edits', methods=['POST'])
@_login_required
def api_pm_op_batch_reprint_with_edits(op_seq):
    """Recreate an Opening Stock label batch with new no_of_box / per_box_qty.

    Body: {
      product_id, godown_id, new_no_of_box, new_per_box_qty,
      req_id (required for non-admin), token (required for non-admin)
    }

    Effects (all in one transaction):
      1. Snapshot existing pm_boxes + pm_box_movements + the matching opening
         ledger row into pm_recycle_bin (entity_type='op_batch_edit')
      2. Delete the old pm_boxes / pm_box_movements rows for (op_label, product_id)
      3. Create new pm_boxes (with NEW box_codes — BNNN sequence resets to 1) and
         matching pm_box_movements rows
      4. UPSERT the opening ledger row at new total qty (preserves the
         single-row-per-(product,godown) invariant established earlier)
      5. Return the print payload so frontend fires printOpeningLabels
    """
    d = request.get_json() or {}
    try:
        product_id  = int(d.get('product_id'))
        godown_id   = int(d.get('godown_id'))
        new_nob     = int(d.get('new_no_of_box') or 0)
        new_pbq     = float(d.get('new_per_box_qty') or 0)
    except Exception:
        return jsonify({'status':'error','message':'product_id, godown_id, new_no_of_box, new_per_box_qty (numeric) are required'}), 400
    if new_nob <= 0 or new_pbq <= 0:
        return jsonify({'status':'error','message':'new_no_of_box and new_per_box_qty must be > 0'}), 400
    req_id = d.get('req_id')
    token  = d.get('token')

    op_label = f'PM-OP/{op_seq:04d}'
    new_total = new_nob * new_pbq

    conn = sampling_portal.get_db_connection()
    try:
        # Verify caller's authority
        ok, err, _req_row = _verify_reprint_auth(
            conn,
            scope_type='voucher_op', voucher_kind='op', voucher_id=op_seq,
            req_id=req_id, token=token,
            expected_product_id=product_id, expected_godown_id=godown_id,
            expected_new_nob=new_nob, expected_new_pbq=new_pbq
        )
        if not ok:
            conn.close(); return err

        # Load product details for the print payload + box code generation
        prod = conn.execute(
            """SELECT product_name, COALESCE(product_code,'') AS code,
                      COALESCE(pm_type,'') AS pm_type
               FROM pm_products WHERE id=%s""", (product_id,)
        ).fetchone()
        if not prod or not prod['code']:
            conn.close(); return jsonify({'status':'error','message':'Product not found or has no product code'}), 404
        brand = conn.execute(
            """SELECT COALESCE(b.name,'') AS name FROM pm_products p
               LEFT JOIN procurement_brands b ON b.id = p.brand_id
               WHERE p.id=%s""", (product_id,)
        ).fetchone()
        brand_name = brand['name'] if brand else ''

        # Confirm batch exists
        existing_boxes = conn.execute(
            """SELECT * FROM pm_boxes
               WHERE grn_no=%s AND product_id=%s AND current_godown_id=%s
               ORDER BY box_seq""",
            (op_label, product_id, godown_id)
        ).fetchall()
        if not existing_boxes:
            conn.close(); return jsonify({'status':'error','message':'No matching OP batch found'}), 404

        # ── Snapshot for recycle bin BEFORE mutating ──────────────────────
        # We bundle: the boxes + their movements + the opening ledger row.
        # Restore from this entry would re-INSERT all those rows (admin can use
        # the recycle bin if they need to roll back).
        bin_id = _bin_soft_delete(
            conn,
            entity_type  = 'op_batch_edit',
            entity_id    = op_seq,
            entity_label = f"OP batch {op_label} · {prod['product_name']} (edited)",
            parent_table = 'pm_boxes',
            parent_where = 'grn_no=%s AND product_id=%s AND current_godown_id=%s',
            parent_params= (op_label, product_id, godown_id),
            children = [
                # Snapshot ALL boxes (the parent above is just one anchor row;
                # the children list captures the rest of the batch).
                {'table': 'pm_boxes',
                 'where': 'grn_no=%s AND product_id=%s AND current_godown_id=%s',
                 'params': (op_label, product_id, godown_id)},
                # Their movements
                {'table': 'pm_box_movements',
                 'where': 'box_id IN (SELECT box_id FROM pm_boxes WHERE grn_no=%s AND product_id=%s AND current_godown_id=%s)',
                 'params': (op_label, product_id, godown_id)},
                # Opening ledger row(s) — we capture both godown and floor
                {'table': 'pm_godown_txn',
                 'where': "product_id=%s AND godown_id=%s AND txn_type='opening'",
                 'params': (product_id, godown_id)},
                {'table': 'pm_floor_txn',
                 'where': "product_id=%s AND godown_id=%s AND txn_type='floor_opening'",
                 'params': (product_id, godown_id)},
            ],
            summary = f"Edit reprint of {op_label}: {len(existing_boxes)}×{float(existing_boxes[0]['per_box_qty'] or 0):g} → {new_nob}×{new_pbq:g}",
            reason  = (d.get('reason') or '')[:500] or None
        )

        # _bin_soft_delete already deleted parent + children (boxes,
        # movements, opening ledger rows). Recreate everything fresh.

        # 1. Recreate boxes with new dimensions, fresh BNNN sequence
        new_codes = []
        # Parallel array of 8-char sequential short_codes, one per box.
        # The label printer prefers _shortCode for the QR payload and keeps
        # box_code as the visible text.
        new_short_codes = []
        first_box_id_for_movements = None
        for seq in range(1, new_nob + 1):
            box_code = _make_op_box_code(prod['code'], op_seq, seq)
            cur = conn.execute(
                """INSERT INTO pm_boxes
                     (box_code, grn_id, grn_no, grn_item_id, product_id, product_code,
                      box_seq, total_boxes, per_box_qty, current_godown_id, current_status,
                      created_by)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'in_stock',%s)""",
                (box_code, 0, op_label, 0, product_id, prod['code'],
                 seq, new_nob, new_pbq, godown_id, _user())
            )
            box_id = cur.lastrowid
            sc = _assign_box_short_code(conn, box_id) or ''
            if first_box_id_for_movements is None:
                first_box_id_for_movements = box_id
            conn.execute(
                """INSERT INTO pm_box_movements
                     (box_id, movement_type, to_godown_id, qty, moved_by, remarks)
                   VALUES (%s, 'adjust', %s, %s, %s, %s)""",
                (box_id, godown_id, new_pbq, _user(),
                 f'Opening stock {op_label} box {seq}/{new_nob} (edited reprint)')
            )
            new_codes.append(box_code)
            new_short_codes.append(sc)

        # 2. Update the ledger opening to new_total (UPSERT scoped to this
        # product+godown). Determine if this godown is a floor.
        is_floor = _is_floor_godown(conn, godown_id)
        if is_floor:
            ex = conn.execute(
                """SELECT id FROM pm_floor_txn
                   WHERE product_id=%s AND godown_id=%s AND txn_type='floor_opening'
                   ORDER BY id LIMIT 1""",
                (product_id, godown_id)
            ).fetchone()
            if ex:
                conn.execute(
                    """UPDATE pm_floor_txn
                       SET qty=%s, remarks=%s, created_by=%s
                       WHERE id=%s""",
                    (new_total, f'[{op_label}] Opening (edited) · {new_nob}×{new_pbq:g}',
                     _user(), ex['id'])
                )
            else:
                conn.execute(
                    """INSERT INTO pm_floor_txn (product_id, txn_date, txn_type, qty, remarks, created_by, godown_id)
                       VALUES (%s, CURDATE(), 'floor_opening', %s, %s, %s, %s)""",
                    (product_id, new_total,
                     f'[{op_label}] Opening (edited) · {new_nob}×{new_pbq:g}',
                     _user(), godown_id)
                )
        else:
            ex = conn.execute(
                """SELECT id FROM pm_godown_txn
                   WHERE product_id=%s AND godown_id=%s AND txn_type='opening'
                   ORDER BY id LIMIT 1""",
                (product_id, godown_id)
            ).fetchone()
            if ex:
                conn.execute(
                    """UPDATE pm_godown_txn
                       SET qty=%s, remarks=%s, created_by=%s
                       WHERE id=%s""",
                    (new_total, f'[{op_label}] Opening (edited) · {new_nob}×{new_pbq:g}',
                     _user(), ex['id'])
                )
            else:
                conn.execute(
                    """INSERT INTO pm_godown_txn (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
                       VALUES (%s, CURDATE(), 'opening', %s, %s, %s, %s, %s)""",
                    (product_id, new_total,
                     f'[{op_label}] Opening (edited) · {new_nob}×{new_pbq:g}',
                     _user(), op_label, godown_id)
                )

        # Log the reprint event
        try:
            _label_print_record(
                conn,
                print_kind='op_reprint',
                voucher_kind='op',
                voucher_id=op_seq,
                voucher_no=op_label,
                product_id=product_id,
                label_count=int(new_nob),
                box_codes=new_codes,
            )
        except Exception: pass

        conn.commit()
        conn.close()

        # Return the print payload (mirror printOpeningLabels expected shape)
        return jsonify({
            'status':       'ok',
            'op_label':     op_label,
            'op_seq':       op_seq,
            'product_id':   product_id,
            'product_code': prod['code'],
            'product_name': prod['product_name'],
            'pm_type':      prod['pm_type'] or '',
            'brand_name':   brand_name,
            'godown_id':    godown_id,
            'no_of_box':    new_nob,
            'per_box_qty':  new_pbq,
            'total_qty':    new_total,
            'box_codes':    new_codes,
            'short_codes':  new_short_codes,
            'op_date':      str(date.today()),
            'remarks':      f'Edited reprint · was {len(existing_boxes)}×{float(existing_boxes[0]["per_box_qty"] or 0):g}',
            'bin_id':       bin_id,
            'edited':       True
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/grn/<int:grn_id>/reprint_labels_with_edits', methods=['POST'])
@_login_required
def api_grn_reprint_with_edits(grn_id):
    """Recreate the box labels for one GRN line item with new dimensions, and
    update the GRN line + auto-posted inward txn to match.

    Body: {
      product_id, new_no_of_box, new_per_box_qty,
      req_id (required for non-admin), token (required for non-admin)
    }

    Effects (one transaction):
      1. Snapshot the GRN line + its boxes + their movements + the matching
         inward txn into pm_recycle_bin (entity_type='grn_item_edit')
      2. Delete the old boxes/movements for (grn_id, product_id)
      3. Recreate boxes with new dimensions (codes: PRODCODE-G####-B###)
      4. UPDATE the pm_grn_items row for this product (qty, no_of_box, per_box_qty)
      5. UPDATE the auto-posted inward txn (pm_godown_txn) to match new qty
      6. Return print payload (data shape consumed by grnOpenLabels' printer)
    """
    d = request.get_json() or {}
    try:
        product_id = int(d.get('product_id'))
        new_nob    = int(d.get('new_no_of_box') or 0)
        new_pbq    = float(d.get('new_per_box_qty') or 0)
    except Exception:
        return jsonify({'status':'error','message':'product_id, new_no_of_box, new_per_box_qty (numeric) are required'}), 400
    if new_nob <= 0 or new_pbq <= 0:
        return jsonify({'status':'error','message':'new_no_of_box and new_per_box_qty must be > 0'}), 400
    req_id = d.get('req_id')
    token  = d.get('token')

    new_total = new_nob * new_pbq

    conn = sampling_portal.get_db_connection()
    try:
        # Authorize
        ok, err, _req_row = _verify_reprint_auth(
            conn,
            scope_type='voucher_grn', voucher_kind='grn', voucher_id=grn_id,
            req_id=req_id, token=token,
            expected_product_id=product_id,
            expected_new_nob=new_nob, expected_new_pbq=new_pbq
        )
        if not ok:
            conn.close(); return err

        # Load GRN header + line item
        grn = conn.execute(
            """SELECT id, grn_no, grn_date, supplier, godown_id
               FROM pm_grn WHERE id=%s""", (grn_id,)
        ).fetchone()
        if not grn:
            conn.close(); return jsonify({'status':'error','message':'GRN not found'}), 404
        grn_no   = grn['grn_no']
        grn_date = grn['grn_date']
        godown_id = grn['godown_id']

        line = conn.execute(
            """SELECT * FROM pm_grn_items WHERE grn_id=%s AND product_id=%s LIMIT 1""",
            (grn_id, product_id)
        ).fetchone()
        if not line:
            conn.close(); return jsonify({'status':'error','message':'GRN line item not found for this product'}), 404
        old_qty = float(line.get('qty') or 0)
        old_nob = int(line.get('no_of_box') or 0)
        old_pbq = float(line.get('per_box_qty') or 0)
        # Preserve the original line's rate across a split so ABC's value
        # totals don't shift just because we re-keyed the boxes.
        old_rate = float(line.get('rate') or 0)

        prod = conn.execute(
            """SELECT product_name, COALESCE(product_code,'') AS code,
                      COALESCE(pm_type,'') AS pm_type
               FROM pm_products WHERE id=%s""", (product_id,)
        ).fetchone()
        if not prod or not prod['code']:
            conn.close(); return jsonify({'status':'error','message':'Product not found or has no product code'}), 404
        brand = conn.execute(
            """SELECT COALESCE(b.name,'') AS name FROM pm_products p
               LEFT JOIN procurement_brands b ON b.id = p.brand_id
               WHERE p.id=%s""", (product_id,)
        ).fetchone()
        brand_name = brand['name'] if brand else ''

        # Snapshot before mutating
        bin_id = _bin_soft_delete(
            conn,
            entity_type  = 'grn_item_edit',
            entity_id    = int(line.get('id') or 0),
            entity_label = f"GRN {grn_no} · {prod['product_name']} (edited)",
            parent_table = 'pm_grn_items',
            parent_where = 'id=%s',
            parent_params= (line['id'],),
            children = [
                {'table': 'pm_boxes',
                 'where': 'grn_id=%s AND product_id=%s', 'params': (grn_id, product_id)},
                {'table': 'pm_box_movements',
                 'where': 'box_id IN (SELECT box_id FROM pm_boxes WHERE grn_id=%s AND product_id=%s)',
                 'params': (grn_id, product_id)},
                # The auto-posted inward — we need to capture it before removing.
                # The remarks are tagged "[PM GRN: <grn_no>]" by the GRN-save flow
                # so we filter on that + product_id + qty (best-effort).
                {'table': 'pm_godown_txn',
                 'where': "product_id=%s AND godown_id=%s AND txn_type='inward' AND remarks LIKE %s AND txn_date=%s",
                 'params': (product_id, godown_id, f'[PM GRN: {grn_no}]%', grn_date)},
            ],
            summary = f"Edit reprint of GRN {grn_no} line for {prod['product_name']}: {old_nob}×{old_pbq:g} ({old_qty:g}) → {new_nob}×{new_pbq:g} ({new_total:g})",
            reason  = (d.get('reason') or '')[:500] or None
        )
        # NOTE: _bin_soft_delete deleted parent (the GRN line) AND its children,
        # including the inward txn. We need to recreate the GRN line and
        # the inward txn fresh.

        # Recreate the GRN line item with new dimensions
        # NOTE: pm_grn_items uses (qty_received, no_of_box, box_count) — there
        # are no `qty` or `per_box_qty` columns. box_count holds the per-box
        # quantity. Earlier code referenced non-existent columns and crashed
        # on edit-reprint; corrected here.
        cur = conn.execute(
            """INSERT INTO pm_grn_items
                 (grn_id, product_id, qty_received, no_of_box, box_count, remarks, product_version, entered_uom, rate)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (grn_id, product_id, new_total, new_nob, int(new_pbq) if float(new_pbq).is_integer() else new_pbq,
             f'Edited reprint · was {old_nob}×{old_pbq:g}',
             (line.get('product_version') or None),
             (line.get('entered_uom') or None),
             old_rate)
        )
        new_grn_item_id = cur.lastrowid
        try:
            _get_or_assign_fifo(conn, 'grn', grn_id, product_id)
        except Exception as _fe:
            import sys as _sys
            print(f"[pm_stock_routes] FIFO assign failed (edit-reprint) for GRN {grn_id}/{product_id}: {_fe}", file=_sys.stderr)

        # Recreate boxes
        _create_boxes_for_grn_item(
            conn,
            grn_id=grn_id, grn_no=grn_no, grn_item_id=new_grn_item_id,
            product_id=product_id, product_code=prod['code'],
            no_of_box=new_nob, per_box_qty=new_pbq,
            godown_id=godown_id, grn_date=grn_date, user=_user(),
            product_version=(line.get('product_version') or None),
        )

        # Repost the inward txn for the new total
        conn.execute(
            """INSERT INTO pm_godown_txn
                 (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
               VALUES (%s, %s, 'inward', %s, %s, %s, %s, %s)""",
            (product_id, grn_date, new_total,
             f'[PM GRN: {grn_no}] Edited reprint · {new_nob}×{new_pbq:g}',
             _user(), grn_no, godown_id)
        )

        # Build print payload — match the shape grnLabelDoPrint expects
        new_codes = [_make_box_code(prod['code'], grn_no, s) for s in range(1, new_nob + 1)]

        # Parallel short_codes — fetch the newly-created boxes' short_codes
        # in box_seq order so the array aligns 1:1 with new_codes.
        sc_rows = conn.execute(
            """SELECT short_code
               FROM pm_boxes
               WHERE grn_id=%s AND product_id=%s
               ORDER BY box_seq""",
            (grn_id, product_id)
        ).fetchall()
        new_short_codes = [(r.get('short_code') or '') for r in sc_rows]

        # Pull current FIFO code (idempotent — assigned earlier)
        fifo_code = ''
        try:
            fi = _get_or_assign_fifo(conn, 'grn', grn_id, product_id)
            fifo_code = fi.get('fifo_code', '') or ''
        except Exception:
            pass

        # Log the reprint event for the audit/label report
        try:
            _label_print_record(
                conn,
                print_kind='grn_reprint',
                voucher_kind='grn',
                voucher_id=grn_id,
                voucher_no=grn_no,
                product_id=product_id,
                label_count=int(new_nob),
                box_codes=new_codes,
            )
        except Exception: pass

        # ── INVARIANT GUARD ─────────────────────────────────────────────────
        # See note in api_grn_save. The "edit reprint" path soft-deletes the
        # old line/boxes and re-creates them with new dimensions; if anything
        # goes wrong with that swap, we want a hard failure, not silent
        # corruption.
        _assert_grn_box_invariant(conn, grn_id)

        conn.commit()
        conn.close()
        return jsonify({
            'status':       'ok',
            'grn_id':       grn_id,
            'grn_no':       grn_no,
            'grn_date':     str(grn_date) if grn_date else '',
            'godown_id':    godown_id,
            'product_id':   product_id,
            'product_code': prod['code'],
            'product_name': prod['product_name'],
            'pm_type':      prod['pm_type'] or '',
            'brand_name':   brand_name,
            'no_of_box':    new_nob,
            'per_box_qty':  new_pbq,
            'total_qty':    new_total,
            'box_codes':    new_codes,
            'short_codes':  new_short_codes,
            'bin_id':       bin_id,
            'edited':       True,
            'fifo_code':    fifo_code
        })
    except RuntimeError as _inv_err:
        import sys as _sys
        print(f"[pm_stock_routes] api_grn_reprint_with_edits: {_inv_err}", file=_sys.stderr)
        try: conn.close()
        except Exception: pass
        return jsonify({
            'status':  'error',
            'message': ('GRN data consistency check failed. The reprint edit '
                        'was cancelled to prevent corruption. Please retry; '
                        'if this persists, contact the developer with the '
                        'GRN number.')
        }), 500
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/opening_stock/list', methods=['GET'])
@_login_required
def api_pm_opening_stock_list():
    """List products that have opening stock > 0 at a specific location.
    Used by the Generate Opening Stock Labels modal to populate its product
    picker — only products with existing opening stock are eligible for
    label generation.

    Query params:
      godown_id (required): the godown/floor location to look up

    Response: { status:'ok', is_floor: bool, products: [
        { id, product_name, product_code, pm_type, brand_name,
          opening_qty, labelled_qty, remaining_qty }
    ] }
    where:
      - opening_qty is the total qty in the opening ledger row(s)
      - labelled_qty is the total qty already covered by OP-labelled boxes
      - remaining_qty = opening_qty - labelled_qty (what's still un-labelled)
    """
    gid_raw = request.args.get('godown_id')
    if not gid_raw:
        return jsonify({'status': 'error', 'message': 'godown_id required'}), 400
    try:
        gid = int(gid_raw)
    except Exception:
        return jsonify({'status': 'error', 'message': 'godown_id must be an integer'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        is_floor = _is_floor_godown(conn, gid)

        # Pull (product_id, opening_qty) for this location from whichever ledger
        # table is appropriate. Group by product so multi-row openings sum.
        if is_floor:
            opening_rows = conn.execute(
                """SELECT product_id, COALESCE(SUM(qty),0) AS opening_qty
                   FROM pm_floor_txn
                   WHERE godown_id=%s AND txn_type='floor_opening'
                   GROUP BY product_id
                   HAVING SUM(qty) > 0""",
                (gid,)
            ).fetchall()
        else:
            opening_rows = conn.execute(
                """SELECT product_id, COALESCE(SUM(qty),0) AS opening_qty
                   FROM pm_godown_txn
                   WHERE godown_id=%s AND txn_type='opening'
                   GROUP BY product_id
                   HAVING SUM(qty) > 0""",
                (gid,)
            ).fetchall()

        if not opening_rows:
            conn.close()
            return jsonify({'status': 'ok', 'is_floor': is_floor, 'products': []})

        # Build product_id list for second query
        pid_list   = [int(r['product_id']) for r in opening_rows]
        opening_by = {int(r['product_id']): float(r['opening_qty']) for r in opening_rows}

        # Pull labelled-qty per product at this location: how much qty is
        # already covered by OP-labelled pm_boxes rows so the UI can warn the
        # user not to over-label.
        #
        # IMPORTANT: only count boxes that are STILL physically present here.
        # Filter to current_status IN ('in_stock','in_transit'):
        #   * 'consumed' / 'damaged' / 'lost' boxes have left the building —
        #     they still carry the old current_godown_id (the system doesn't
        #     null it on consume), so including them double-counts ancient
        #     OP batches that have long since been issued / discarded.
        #   * 'superseded' boxes are split parents whose qty is already
        #     represented by their child boxes — counting both would inflate.
        # Without this filter, opening that was previously high and has since
        # been consumed shows up as "already labelled" forever, blocking any
        # new opening-label printing even after the opening qty was edited
        # down.
        placeholders = ','.join(['%s'] * len(pid_list))
        # ── Labelled-qty per product at THIS location ────────────────────────
        # The original implementation joined pm_boxes by current_godown_id.
        # That's wrong for boxes whose current_godown_id has been flipped by
        # the admin /api/pm_stock/box_location_reconcile endpoint — those
        # OP labels still semantically belong to their ORIGINAL location
        # (which is where the opening-stock ledger entry lives).
        #
        # Fix: derive each box's original location from its earliest
        # pm_box_movements row (the 'adjust' creation row written by
        # _create_opening_boxes). Fall back to current_godown_id for boxes
        # without a movement row (split children — _create_opening_boxes
        # inserts a movement per direct OP box, but box-split children
        # are inserted into pm_boxes without their own creation row).
        #
        # Filter current_status IN ('in_stock','in_transit') so consumed
        # / damaged / lost / superseded boxes don't inflate the count.
        labelled = conn.execute(
            f"""SELECT b.product_id,
                       COALESCE(SUM(b.per_box_qty),0) AS labelled_qty
                FROM pm_boxes b
                LEFT JOIN (
                    SELECT m.box_id, m.to_godown_id
                    FROM pm_box_movements m
                    JOIN (
                        SELECT box_id, MIN(movement_id) AS first_mid
                        FROM pm_box_movements
                        GROUP BY box_id
                    ) firsts
                      ON firsts.box_id    = m.box_id
                     AND firsts.first_mid = m.movement_id
                ) creat ON creat.box_id = b.box_id
                WHERE COALESCE(creat.to_godown_id, b.current_godown_id) = %s
                  AND b.product_id IN ({placeholders})
                  AND b.box_code LIKE '%%-OP%%'
                  AND b.current_status IN ('in_stock','in_transit')
                GROUP BY b.product_id""",
            (gid, *pid_list)
        ).fetchall()
        labelled_by = {int(r['product_id']): float(r['labelled_qty'] or 0) for r in labelled}

        # Get product details
        prods = conn.execute(
            f"""SELECT p.id, p.product_name, COALESCE(p.product_code,'') AS product_code,
                       COALESCE(p.pm_type,'') AS pm_type,
                       COALESCE(b.name,'')    AS brand_name
                FROM pm_products p
                LEFT JOIN procurement_brands b ON b.id = p.brand_id
                WHERE p.id IN ({placeholders})
                  AND p.is_active=1
                ORDER BY p.product_name""",
            tuple(pid_list)
        ).fetchall()
        conn.close()

        out = []
        for p in prods:
            pid = int(p['id'])
            op_qty = opening_by.get(pid, 0.0)
            lab    = labelled_by.get(pid, 0.0)
            out.append({
                'id':            pid,
                'product_name':  p['product_name'],
                'product_code':  p['product_code'],
                'pm_type':       p['pm_type'],
                'brand_name':    p['brand_name'],
                'opening_qty':   op_qty,
                'labelled_qty':  lab,
                'remaining_qty': max(0.0, op_qty - lab),
            })
        return jsonify({'status': 'ok', 'is_floor': is_floor, 'products': out})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/opening_boxes/create', methods=['POST'])
@_login_required
def api_create_opening_boxes():
    """
    Create QR-coded label boxes for an EXISTING opening-stock entry.
    This endpoint is print-only — it does NOT add stock. Opening stock must
    already exist (entered via the Add Product modal or /opening/save).

    Writes:
      - N rows in pm_boxes (with codes PRODUCTCODE-OPNNNN-BNNN)
      - N rows in pm_box_movements (movement_type='adjust', remarks tag = OP label)

    Does NOT write:
      - pm_godown_txn / pm_floor_txn — stock is unchanged

    Validates:
      - Opening stock for (product, godown) must already exist (>0)
      - Total qty (no_of_box × per_box_qty) must not exceed existing opening qty

    Request: {
      product_id, godown_id,
      // Single-group (legacy):
      no_of_box, per_box_qty,
      // OR multi-group (one OP label, mixed per-box quantities):
      groups: [
        {no_of_box, per_box_qty},
        {no_of_box, per_box_qty},
        ...
      ],
      op_date  (YYYY-MM-DD, optional, defaults today),
      remarks  (optional)
    }
    Returns the OP label, all generated box codes, and a print URL hint.
    """
    blocked = _block_if_disabled('opening')
    if blocked is not None: return blocked
    # Requesters can't create vouchers — they only submit Material Requests.
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    # New-voucher-entries access check (per-user access control)
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked
    d = request.get_json() or {}
    pid = d.get('product_id')
    gid = d.get('godown_id')

    # Parse either single-group or multi-group input.
    groups_in = d.get('groups')
    norm_groups = []
    if isinstance(groups_in, list) and groups_in:
        for g in groups_in:
            try:
                gn = int(g.get('no_of_box') or 0)
                gp = float(g.get('per_box_qty') or 0)
            except Exception:
                return jsonify({'status':'error','message':'Each group needs numeric no_of_box and per_box_qty'}), 400
            if gn <= 0 or gp <= 0:
                return jsonify({'status':'error','message':'Each group must have positive no_of_box and per_box_qty'}), 400
            norm_groups.append({'no_of_box': gn, 'per_box_qty': gp})
        if not norm_groups:
            return jsonify({'status':'error','message':'At least one group is required'}), 400
    else:
        try:
            nob = int(d.get('no_of_box') or 0)
            pbq = float(d.get('per_box_qty') or 0)
        except Exception:
            return jsonify({'status': 'error', 'message': 'no_of_box and per_box_qty must be numbers'}), 400
        if nob <= 0 or pbq <= 0:
            return jsonify({'status': 'error', 'message': 'no_of_box and per_box_qty must be positive'}), 400
        norm_groups.append({'no_of_box': nob, 'per_box_qty': pbq})

    if not pid or not gid:
        return jsonify({'status': 'error', 'message': 'product_id and godown_id required'}), 400

    op_date = (d.get('op_date') or '').strip() or str(date.today())
    remarks = (d.get('remarks') or '').strip()

    conn = sampling_portal.get_db_connection()
    try:
        # Need product_code on the product — opening boxes can't be made without one
        prow = conn.execute(
            """SELECT p.product_name, p.pm_type,
                      COALESCE(p.product_code,'') AS code,
                      COALESCE(b.name,'')         AS brand_name
               FROM pm_products p
               LEFT JOIN procurement_brands b ON b.id = p.brand_id
               WHERE p.id=%s AND p.is_active=1""",
            (int(pid),)
        ).fetchone()
        if not prow:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Product not found'}), 404
        if not prow['code']:
            conn.close()
            return jsonify({
                'status': 'error',
                'message': f"Product '{prow['product_name']}' has no product code yet. Generate it from the Products tab first."
            }), 400

        # ── Validate existing opening stock at this location ─────────────────
        # Labels can only be printed against opening stock that ALREADY exists.
        # This prevents the old behaviour where label generation silently added
        # stock and could double-count on re-prints.
        is_floor = _is_floor_godown(conn, int(gid))
        if is_floor:
            existing = conn.execute(
                """SELECT COALESCE(SUM(qty),0) AS opening_qty
                   FROM pm_floor_txn
                   WHERE product_id=%s AND godown_id=%s AND txn_type='floor_opening'""",
                (int(pid), int(gid))
            ).fetchone()
        else:
            existing = conn.execute(
                """SELECT COALESCE(SUM(qty),0) AS opening_qty
                   FROM pm_godown_txn
                   WHERE product_id=%s AND godown_id=%s AND txn_type='opening'""",
                (int(pid), int(gid))
            ).fetchone()
        existing_qty = float((existing or {}).get('opening_qty') or 0)

        # Also subtract qty already covered by previously-printed OP labels at
        # this location, so re-printing additional labels doesn't go past the
        # opening stock total.
        #
        # Must match the picker's list endpoint exactly — only count boxes
        # that are STILL physically present here. Status enum semantics:
        #   * 'consumed' / 'damaged' / 'lost' — box has left the building.
        #     pm_boxes.current_godown_id is NOT nulled on these transitions
        #     (see MTV-receive at line ~11717, DN-consume at line ~13911),
        #     so without this filter the row keeps showing up at the source
        #     godown forever — blocking new label printing for products
        #     whose opening was later edited down.
        #   * 'superseded' — split parent; qty already counted in children.
        #   * 'in_stock' + 'in_transit' — legitimately occupy opening qty
        #     at this location and must be counted.
        # ── Labelled-qty validation: same logic as the picker endpoint ──
        # See /api/pm_stock/opening_stock/list for the full rationale. tl;dr:
        # join against the earliest pm_box_movements row to find the OP
        # creation godown; fall back to current_godown_id for split-children
        # that don't have their own creation movement row.
        already_labelled_row = conn.execute(
            """SELECT COALESCE(SUM(b.per_box_qty),0) AS lab
               FROM pm_boxes b
               LEFT JOIN (
                   SELECT m.box_id, m.to_godown_id
                   FROM pm_box_movements m
                   JOIN (
                       SELECT box_id, MIN(movement_id) AS first_mid
                       FROM pm_box_movements
                       GROUP BY box_id
                   ) firsts
                     ON firsts.box_id    = m.box_id
                    AND firsts.first_mid = m.movement_id
               ) creat ON creat.box_id = b.box_id
               WHERE b.product_id=%s
                 AND COALESCE(creat.to_godown_id, b.current_godown_id)=%s
                 AND b.box_code LIKE '%%-OP%%'
                 AND b.current_status IN ('in_stock','in_transit')""",
            (int(pid), int(gid))
        ).fetchone()
        already_labelled_qty = float((already_labelled_row or {}).get('lab') or 0)
        remaining_qty = max(0.0, existing_qty - already_labelled_qty)
        total_qty     = sum(g['no_of_box'] * g['per_box_qty'] for g in norm_groups)
        total_boxes   = sum(g['no_of_box'] for g in norm_groups)

        if existing_qty <= 0:
            conn.close()
            return jsonify({
                'status': 'error',
                'message': f"No opening stock found for '{prow['product_name']}' at this location. Add opening stock first via the Add Product modal, then come back to print labels."
            }), 400

        if total_qty > remaining_qty + 0.001:   # small float tolerance
            conn.close()
            return jsonify({
                'status': 'error',
                'message': (f"Total qty ({total_qty:g}) exceeds remaining un-labelled "
                            f"opening stock ({remaining_qty:g}) for this product at "
                            f"this location. Opening total: {existing_qty:g}, already "
                            f"labelled: {already_labelled_qty:g}.")
            }), 400

        # ── Create the box records (no stock-ledger write) ──────────────────
        result = _create_opening_boxes(
            conn,
            product_id   = int(pid),
            product_code = prow['code'],
            groups       = norm_groups,
            godown_id    = int(gid),
            op_date      = op_date,
            op_remarks   = remarks or None,
            user         = _user()
        )

        # NOTE: No INSERT into pm_godown_txn / pm_floor_txn. Stock is unchanged
        # by this operation — labels are a print-only artefact. Existing
        # opening stock entries (entered via the Add Product modal or
        # /opening/save) remain the source of truth for qty.

        # Audit + label-print log
        try:
            _audit_record(
                conn,
                action='op.create',
                entity='op_batch',
                entity_id=result['op_seq'],
                summary=f"OP batch {result['op_label']} · {total_boxes} box(es) · "
                        f"{prow['product_name']} [{prow['pm_type'] or '—'}]",
                before=None,
                after={
                    'op_seq': result['op_seq'],
                    'product_id': pid,
                    'godown_id': int(gid),
                    'no_of_box': total_boxes,
                    'total_qty': total_qty,
                    'op_date': op_date,
                    'groups': result.get('groups', norm_groups),
                },
            )
            _label_print_record(
                conn,
                print_kind='op_fresh',
                voucher_kind='op',
                voucher_id=result['op_seq'],
                voucher_no=result['op_label'],
                product_id=pid,
                label_count=total_boxes,
                box_codes=result.get('codes'),
            )
        except Exception: pass

        conn.commit()
        conn.close()
        # For back-compat: report no_of_box as the TOTAL across groups, and
        # per_box_qty as the FIRST group's value (older clients used it as
        # a single value). Authoritative info is in `per_box_qtys` + `groups`.
        first_pbq = norm_groups[0]['per_box_qty']
        return jsonify({
            'status':       'ok',
            'op_label':     result['op_label'],
            'op_seq':       result['op_seq'],
            'product_code': prow['code'],
            'product_name': prow['product_name'],
            'pm_type':      prow['pm_type'] or '',
            'brand_name':   prow['brand_name'] or '',
            'no_of_box':    total_boxes,         # total across all groups
            'per_box_qty':  first_pbq,           # first group's pbq (legacy)
            'per_box_qtys': result.get('per_box_qtys', []),  # one per box
            'groups':       result.get('groups', norm_groups),
            'total_qty':    total_qty,
            'box_codes':    result['codes'],
            'short_codes':  result.get('short_codes') or [],
            'godown_id':    int(gid),
            'op_date':      op_date,
            'remarks':      remarks or '',
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ── Voucher health check & heal (admin only) ─────────────────────────────────
# These endpoints catch the class of bug seen on PMT/26-27/0099 where
# pm_transfer_items had two OUT ledger rows (1 box × 1,140 and 1 box × 1,130)
# for the same product, but the IN modal's client-side aggregation was
# last-write-wins and showed only one of them as ✅ OK — masking a real
# 1,130-qty discrepancy. The frontend aggregation is now fixed, but the
# underlying DB drift can still exist on legacy vouchers where the
# uq_pm_xfer_item unique key was added AFTER the duplicate rows were
# created. These endpoints find and (optionally) consolidate such drift.

@pm_stock_bp.route('/api/pm_stock/transfers/sync_check', methods=['GET'])
@_login_required
def api_transfers_sync_check():
    """
    Audit every active voucher for line-ledger vs box-scan drift.
    Returns vouchers where pm_transfer_items totals don't match
    pm_box_movements scan counts for the same (transfer, side, product).
    Read-only — does not modify anything.
    Admin only.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        # For each voucher, compute per-product (side, no_of_box, total_qty)
        # from BOTH sources and surface rows where they disagree.
        rows = conn.execute("""
            SELECT t.transfer_id, t.transfer_no, t.status, t.has_discrepancy,
                   ti.side, ti.product_id, p.product_name,
                   COUNT(*)             AS ti_row_count,
                   SUM(ti.no_of_box)    AS ti_boxes,
                   SUM(ti.total_qty)    AS ti_qty,
                   bm_agg.bm_count      AS bm_boxes,
                   bm_agg.bm_qty        AS bm_qty
            FROM pm_transfer_items ti
            JOIN pm_transfers t  ON t.transfer_id = ti.transfer_id
            JOIN pm_products  p  ON p.id          = ti.product_id
            LEFT JOIN (
              SELECT bm.transfer_id, bm.movement_type AS side,
                     b.product_id,
                     COUNT(*) AS bm_count, SUM(b.per_box_qty) AS bm_qty
              FROM pm_box_movements bm
              JOIN pm_boxes b ON b.box_id = bm.box_id
              WHERE bm.movement_type IN ('out','in')
              GROUP BY bm.transfer_id, bm.movement_type, b.product_id
            ) bm_agg ON bm_agg.transfer_id = ti.transfer_id
                    AND bm_agg.side        = ti.side
                    AND bm_agg.product_id  = ti.product_id
            WHERE t.status <> 'cancelled'
            GROUP BY t.transfer_id, t.transfer_no, t.status, t.has_discrepancy,
                     ti.side, ti.product_id, p.product_name, bm_agg.bm_count, bm_agg.bm_qty
        """).fetchall()
        # Filter to rows where totals don't match (we do this in Python
        # rather than HAVING so we can also flag duplicate-row cases).
        bad = []
        for r in rows:
            ti_boxes = int(r['ti_boxes'] or 0)
            ti_qty   = float(r['ti_qty']   or 0)
            bm_boxes = int(r['bm_boxes'] or 0) if r['bm_boxes'] is not None else None
            bm_qty   = float(r['bm_qty']   or 0) if r['bm_qty']   is not None else None
            row_ct   = int(r['ti_row_count'] or 0)
            issues = []
            if row_ct > 1:
                issues.append(f'duplicate_rows({row_ct})')
            if bm_boxes is not None and ti_boxes != bm_boxes:
                issues.append(f'box_count_drift(ledger={ti_boxes}, scans={bm_boxes})')
            if bm_qty is not None and abs(ti_qty - bm_qty) > 0.001:
                issues.append(f'qty_drift(ledger={ti_qty}, scans={bm_qty})')
            if issues:
                bad.append({
                    'transfer_id':   r['transfer_id'],
                    'transfer_no':   r['transfer_no'],
                    'status':        r['status'],
                    'has_discrepancy': bool(r['has_discrepancy']),
                    'side':          r['side'],
                    'product_id':    r['product_id'],
                    'product_name':  r['product_name'],
                    'ledger_boxes':  ti_boxes,
                    'ledger_qty':    ti_qty,
                    'scanned_boxes': bm_boxes,
                    'scanned_qty':   bm_qty,
                    'ledger_row_count': row_ct,
                    'issues':        issues,
                })
        conn.close()
        return jsonify({
            'status':         'ok',
            'affected_count': len(bad),
            'affected':       bad,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/heal_line_items', methods=['POST'])
@_login_required
def api_transfer_heal_line_items(tid):
    """
    Consolidate duplicate pm_transfer_items rows for one voucher.
    For each (transfer_id, side, product_id) group with >1 row, sums
    no_of_box and total_qty into the MIN(item_id) row and deletes the rest.
    Idempotent — no-op if there are no duplicates.
    Recomputes _check_discrepancy after consolidation.
    Admin only.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute(
            "SELECT transfer_id, transfer_no FROM pm_transfers WHERE transfer_id=%s",
            (tid,)
        ).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Voucher not found'}), 404

        dup_groups = conn.execute(
            """SELECT side, product_id,
                      MIN(item_id)   AS keep_id,
                      COUNT(*)       AS n,
                      SUM(no_of_box) AS sum_box,
                      SUM(total_qty) AS sum_qty
               FROM pm_transfer_items
               WHERE transfer_id=%s
               GROUP BY side, product_id
               HAVING COUNT(*) > 1""",
            (tid,)
        ).fetchall()

        consolidated = []
        for g in dup_groups:
            conn.execute(
                """UPDATE pm_transfer_items
                   SET no_of_box=%s, total_qty=%s
                   WHERE item_id=%s""",
                (int(g['sum_box'] or 0), float(g['sum_qty'] or 0), g['keep_id'])
            )
            conn.execute(
                """DELETE FROM pm_transfer_items
                   WHERE transfer_id=%s AND side=%s AND product_id=%s
                     AND item_id <> %s""",
                (tid, g['side'], g['product_id'], g['keep_id'])
            )
            consolidated.append({
                'side':       g['side'],
                'product_id': g['product_id'],
                'rows_before': int(g['n'] or 0),
                'rows_after':  1,
                'kept_item_id': g['keep_id'],
                'merged_boxes': int(g['sum_box'] or 0),
                'merged_qty':   float(g['sum_qty'] or 0),
            })

        # Recompute discrepancy flag so the banner state reflects the
        # consolidated reality (some "discrepancies" might disappear if
        # they were just duplicate-row aggregation artefacts; others
        # like PMT/0099 remain real and stay flagged).
        try:
            has_d, mismatches = _check_discrepancy(conn, tid)
        except Exception:
            has_d, mismatches = (None, None)

        conn.commit()
        conn.close()
        return jsonify({
            'status':            'ok',
            'transfer_no':       t['transfer_no'],
            'groups_consolidated': len(consolidated),
            'consolidated':      consolidated,
            'has_discrepancy_after': has_d,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transfers/heal_line_items_bulk', methods=['POST'])
@_login_required
def api_transfers_heal_line_items_bulk():
    """
    Sweep every non-cancelled voucher for duplicate pm_transfer_items rows
    and consolidate them in one transaction. Same per-voucher logic as
    /transfers/<tid>/heal_line_items, applied across the whole system.

    Idempotent. Vouchers with no duplicates are skipped (zero cost). Returns
    a per-voucher summary so the admin can see what actually changed.

    Use this after deploying the IN-modal aggregation fix to clean up the
    underlying DB drift on legacy vouchers (those created before the
    uq_pm_xfer_item unique key was added, where two ledger rows for the
    same product on the same side coexist).

    Admin only.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        # Find every (transfer, side, product) group with >1 row, then
        # GROUP by transfer to determine which vouchers need work.
        affected_tids = conn.execute("""
            SELECT DISTINCT ti.transfer_id
            FROM pm_transfer_items ti
            JOIN pm_transfers t ON t.transfer_id = ti.transfer_id
            WHERE t.status <> 'cancelled'
            GROUP BY ti.transfer_id, ti.side, ti.product_id
            HAVING COUNT(*) > 1
        """).fetchall()
        tids = sorted({int(r['transfer_id']) for r in affected_tids})

        per_voucher = []
        total_groups   = 0
        total_rows_dropped = 0
        for tid in tids:
            t = conn.execute(
                "SELECT transfer_no FROM pm_transfers WHERE transfer_id=%s",
                (tid,)
            ).fetchone()
            transfer_no = t['transfer_no'] if t else f'<tid={tid}>'

            dup_groups = conn.execute(
                """SELECT side, product_id,
                          MIN(item_id)   AS keep_id,
                          COUNT(*)       AS n,
                          SUM(no_of_box) AS sum_box,
                          SUM(total_qty) AS sum_qty
                   FROM pm_transfer_items
                   WHERE transfer_id=%s
                   GROUP BY side, product_id
                   HAVING COUNT(*) > 1""",
                (tid,)
            ).fetchall()

            groups_here = 0
            rows_dropped_here = 0
            for g in dup_groups:
                conn.execute(
                    "UPDATE pm_transfer_items SET no_of_box=%s, total_qty=%s WHERE item_id=%s",
                    (int(g['sum_box'] or 0), float(g['sum_qty'] or 0), g['keep_id'])
                )
                conn.execute(
                    """DELETE FROM pm_transfer_items
                       WHERE transfer_id=%s AND side=%s AND product_id=%s
                         AND item_id <> %s""",
                    (tid, g['side'], g['product_id'], g['keep_id'])
                )
                groups_here       += 1
                rows_dropped_here += int(g['n'] or 0) - 1

            # Refresh discrepancy flag now that ledger is consolidated
            try:
                has_d, _ = _check_discrepancy(conn, tid)
            except Exception:
                has_d = None

            per_voucher.append({
                'transfer_id':         tid,
                'transfer_no':         transfer_no,
                'groups_consolidated': groups_here,
                'rows_dropped':        rows_dropped_here,
                'has_discrepancy_after': has_d,
            })
            total_groups       += groups_here
            total_rows_dropped += rows_dropped_here

        conn.commit()
        conn.close()
        return jsonify({
            'status':              'ok',
            'vouchers_touched':    len(per_voucher),
            'total_groups':        total_groups,
            'total_rows_dropped':  total_rows_dropped,
            'per_voucher':         per_voucher,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/recalc_out_from_boxes', methods=['POST'])
@_login_required
def api_transfer_recalc_out_from_boxes(tid):
    """
    'Recalculate from scanned boxes' repair — boxes are the source of truth.

    For an OUT voucher whose stored line totals (pm_transfer_items.no_of_box /
    total_qty) have drifted from the physical scans, this recomputes each OUT
    line directly from the actual pm_box_movements (movement_type='out') rows.

    The line uniqueness key is (transfer_id, side, product_id, per_box_qty), so
    we aggregate scanned boxes by (product_id, per_box_qty) — NOT just product —
    to map exactly one aggregate to one line row. We update matching lines, and
    report any movement groups that have NO matching line (orphans) and any
    lines that have NO scanned boxes (stale lines) so the admin can see them.

    Also refreshes pm_transfers.total_boxes / total_qty (header) and the
    discrepancy flag. Idempotent — running twice changes nothing the 2nd time.
    Admin only.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute(
            "SELECT transfer_id, transfer_no FROM pm_transfers WHERE transfer_id=%s",
            (tid,)
        ).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Voucher not found'}), 404

        # Actual scanned OUT boxes, grouped by (product, per_box_qty) — matches
        # the line uniqueness contract exactly.
        agg = conn.execute(
            """SELECT b.product_id,
                      b.per_box_qty            AS per_box_qty,
                      COUNT(*)                 AS box_count,
                      SUM(b.per_box_qty)       AS qty
                 FROM pm_box_movements m
                 JOIN pm_boxes b ON b.box_id = m.box_id
                WHERE m.transfer_id=%s AND m.movement_type='out'
                GROUP BY b.product_id, b.per_box_qty""",
            (tid,)
        ).fetchall()

        # Existing OUT lines for this voucher.
        lines = conn.execute(
            """SELECT item_id, product_id, per_box_qty, no_of_box, total_qty
                 FROM pm_transfer_items
                WHERE transfer_id=%s AND side='out'""",
            (tid,)
        ).fetchall()
        line_by_key = {(int(l['product_id']), float(l['per_box_qty'])): l for l in lines}

        updated = []
        orphans = []          # scanned boxes with no matching line
        hdr_boxes = 0
        hdr_qty   = 0.0
        for a in agg:
            key = (int(a['product_id']), float(a['per_box_qty']))
            new_boxes = int(a['box_count'] or 0)
            new_qty   = float(a['qty'] or 0)
            hdr_boxes += new_boxes
            hdr_qty   += new_qty
            ln = line_by_key.get(key)
            if ln is None:
                orphans.append({'product_id': key[0], 'per_box_qty': key[1],
                                'scanned_boxes': new_boxes, 'scanned_qty': new_qty})
                continue
            old_boxes = int(ln['no_of_box'] or 0); old_qty = float(ln['total_qty'] or 0)
            if old_boxes != new_boxes or abs(old_qty - new_qty) > 0.0005:
                conn.execute(
                    "UPDATE pm_transfer_items SET no_of_box=%s, total_qty=%s WHERE item_id=%s",
                    (new_boxes, new_qty, ln['item_id'])
                )
                updated.append({'item_id': ln['item_id'], 'product_id': key[0],
                                'per_box_qty': key[1],
                                'boxes': {'from': old_boxes, 'to': new_boxes},
                                'qty':   {'from': old_qty,   'to': new_qty}})

        # Lines with NO scanned boxes at all (stale) — report, don't auto-delete.
        scanned_keys = {(int(a['product_id']), float(a['per_box_qty'])) for a in agg}
        stale_lines = [
            {'item_id': l['item_id'], 'product_id': int(l['product_id']),
             'per_box_qty': float(l['per_box_qty']), 'line_boxes': int(l['no_of_box'] or 0)}
            for l in lines
            if (int(l['product_id']), float(l['per_box_qty'])) not in scanned_keys
        ]

        # Refresh header totals to match the scanned reality.
        conn.execute(
            "UPDATE pm_transfers SET total_boxes=%s, total_qty=%s WHERE transfer_id=%s",
            (hdr_boxes, hdr_qty, tid)
        )

        if updated:
            _log_transfer_edit(conn, tid, 'recalc_out_from_boxes',
                               f'{len(updated)} line(s) recomputed from scanned boxes')

        try:
            has_d, _ = _check_discrepancy(conn, tid)
        except Exception:
            has_d = None

        conn.commit()
        conn.close()
        return jsonify({
            'status':       'ok',
            'transfer_no':  t['transfer_no'],
            'lines_updated': len(updated),
            'updated':       updated,
            'orphan_box_groups': orphans,      # scanned but no line (needs attention)
            'stale_lines':       stale_lines,  # line but nothing scanned (needs attention)
            'header_boxes':  hdr_boxes,
            'header_qty':    hdr_qty,
            'has_discrepancy_after': has_d,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transfers/start', methods=['POST'])
@_login_required
def api_transfer_start():
    """Create a new transfer header. Returns transfer_id + transfer_no.

    Optional `request_id` on the payload ties this OUT to a Material
    Request — on save_out, the request's qty_fulfilled is updated and
    its status auto-progresses.
    """
    d         = request.get_json() or {}
    from_id   = d.get('from_godown_id')
    to_id     = d.get('to_godown_id')
    remarks   = (d.get('remarks') or '').strip()
    request_id= d.get('request_id')
    if not from_id or not to_id:
        return jsonify({'status': 'error', 'message': 'from_godown_id and to_godown_id required'}), 400
    if int(from_id) == int(to_id):
        return jsonify({'status': 'error', 'message': 'Source and destination must differ'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        transfer_no = _next_voucher_no(conn, 'PM-MT', date.today())
        cur = conn.execute(
            """INSERT INTO pm_transfers (transfer_no, from_godown_id, to_godown_id, status, out_by, remarks, request_id)
               VALUES (%s,%s,%s,'out_started',%s,%s,%s)""",
            (transfer_no, int(from_id), int(to_id), _user(), remarks or None,
             int(request_id) if request_id else None)
        )
        tid = cur.lastrowid
        _log_transfer_edit(conn, tid, 'start_out',
                           f'From={from_id} → To={to_id}'
                           + (f' · for Request #{request_id}' if request_id else ''))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'transfer_id': tid, 'transfer_no': transfer_no})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/scan_out', methods=['POST'])
@_login_required
def api_transfer_scan_out(tid):
    """
    Scan a single box for OUT. Validates that:
      - transfer is still in out_started state
      - box exists and is currently in_stock at the transfer's source godown
      - box hasn't already been scanned for this transfer
    Marks box as in_transit (pre-emptively, so a second scan can't double-count).
    """
    d   = request.get_json() or {}
    code = (d.get('box_code') or '').strip().upper()
    if not code:
        return jsonify({'status': 'error', 'message': 'box_code required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute(
            "SELECT transfer_id, transfer_no, from_godown_id, to_godown_id, status FROM pm_transfers WHERE transfer_id=%s",
            (tid,)
        ).fetchone()
        if not t:
            conn.close(); return jsonify({'status': 'error', 'message': 'Transfer not found'}), 404
        if t['status'] != 'out_started':
            conn.close(); return jsonify({'status': 'error', 'message': f"Transfer is in '{t['status']}' state — cannot scan OUT"}), 409

        b = conn.execute(
            """SELECT b.box_id, b.box_code, b.product_id, b.product_code, b.per_box_qty,
                      b.current_godown_id, b.current_status, p.product_name, p.pm_type
               FROM pm_boxes b
               JOIN pm_products p ON p.id = b.product_id
               WHERE b.short_code = %s OR b.box_code = %s LIMIT 1""",
            (code, code)
        ).fetchone()
        if not b:
            conn.close(); return jsonify({'status': 'error', 'message': f'Box {code} not found'}), 404
        # Location check WITH SELF-HEAL — see api_voucher_scan_box for full
        # rationale. Same logic mirrored here so both scan endpoints heal
        # consistently. Heal records itself as a pm_box_movements row.
        box_auto_corrected = False
        if b['current_godown_id'] != t['from_godown_id']:
            _verdict, _why = _heal_box_location(
                conn,
                box_id=b['box_id'],
                expected_godown_id=t['from_godown_id'],
                current_godown_id=b['current_godown_id'],
                user=_user(),
            )
            if _verdict == 'heal':
                b = dict(b)
                b['current_godown_id'] = t['from_godown_id']
                box_auto_corrected = True
            else:
                conn.close()
                return jsonify({
                    'status': 'error',
                    'message': f"Box is at a different location — cannot OUT from this source",
                    'box_current_godown_id': b['current_godown_id'],
                    'expected_godown_id':    t['from_godown_id'],
                    'heal_reason':           _why,
                }), 409
        if b['current_status'] != 'in_stock':
            conn.close()
            return jsonify({
                'status': 'error',
                'message': f"Box is currently '{b['current_status']}' — must be 'in_stock' to scan OUT"
            }), 409

        # Check we haven't already scanned this box for this transfer
        already = conn.execute(
            """SELECT 1 FROM pm_box_movements
               WHERE box_id=%s AND transfer_id=%s AND movement_type='out' LIMIT 1""",
            (b['box_id'], tid)
        ).fetchone()
        if already:
            conn.close()
            return jsonify({'status': 'error', 'message': 'This box was already scanned for this transfer'}), 409

        # Mark box as in_transit + log movement
        conn.execute(
            "UPDATE pm_boxes SET current_status='in_transit' WHERE box_id=%s",
            (b['box_id'],)
        )
        conn.execute(
            """INSERT INTO pm_box_movements
                 (box_id, transfer_id, movement_type, from_godown_id, to_godown_id, qty, moved_by, remarks)
               VALUES (%s,%s,'out',%s,%s,%s,%s,%s)""",
            (b['box_id'], tid, t['from_godown_id'], t['to_godown_id'],
             float(b['per_box_qty'] or 0), _user(), f"Transfer {t['transfer_no']} OUT")
        )
        _log_transfer_edit(conn, tid, 'scan_out', f"Box {b['box_code']}")
        conn.commit()
        conn.close()
        return jsonify({
            'status': 'ok',
            'box': {
                'box_id':       b['box_id'],
                'box_code':     b['box_code'],
                'product_id':   b['product_id'],
                'product_name': b['product_name'],
                'pm_type':      b['pm_type'],
                'product_code': b['product_code'],
                'per_box_qty':  float(b['per_box_qty'] or 0)
            },
            'auto_corrected': bool(locals().get('box_auto_corrected', False)),
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/scanned', methods=['GET'])
@_login_required
def api_transfer_scanned(tid):
    """Return the list of boxes scanned for this transfer (used to rehydrate UI)."""
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute(
            """SELECT t.*, fg.name AS from_name, tg.name AS to_name
               FROM pm_transfers t
               LEFT JOIN procurement_godowns fg ON fg.id=t.from_godown_id
               LEFT JOIN procurement_godowns tg ON tg.id=t.to_godown_id
               WHERE t.transfer_id=%s""",
            (tid,)
        ).fetchone()
        if not t:
            conn.close(); return jsonify({'status': 'error', 'message': 'Transfer not found'}), 404
        rows = conn.execute(
            """SELECT b.box_id, b.box_code, b.product_id, b.product_code, b.per_box_qty,
                      b.box_seq, b.total_boxes, p.product_name, p.pm_type,
                      m.movement_type, m.movement_at
               FROM pm_box_movements m
               JOIN pm_boxes b ON b.box_id = m.box_id
               JOIN pm_products p ON p.id = b.product_id
               WHERE m.transfer_id=%s
               ORDER BY m.movement_at ASC""",
            (tid,)
        ).fetchall()
        conn.close()

        boxes_out, boxes_in = [], []
        for r in rows:
            d = dict(r)
            d['per_box_qty'] = float(d.get('per_box_qty') or 0)
            d['movement_at'] = str(d.get('movement_at') or '')
            mt = d.pop('movement_type')
            if   mt == 'out': boxes_out.append(d)
            elif mt == 'in':  boxes_in.append(d)
        td = dict(t)
        for k in ('out_at','in_at'):
            if td.get(k) is not None: td[k] = str(td[k])
        td['total_qty'] = float(td.get('total_qty') or 0)
        return jsonify({'status':'ok', 'transfer': td, 'out_boxes': boxes_out, 'in_boxes': boxes_in})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/submit_out', methods=['POST'])
@_login_required
def api_transfer_submit_out(tid):
    """
    Finalise the OUT side: insert source-decrementing stock txns for every
    scanned box (grouped by product), set status='in_pending'.
    """
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute(
            "SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)
        ).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Transfer not found'}), 404
        if t['status'] != 'out_started':
            conn.close(); return jsonify({'status':'error','message':f"Transfer is in '{t['status']}' state"}), 409

        # Aggregate scanned boxes by product
        agg = conn.execute(
            """SELECT b.product_id, COUNT(*) AS box_count, SUM(b.per_box_qty) AS qty
               FROM pm_box_movements m
               JOIN pm_boxes b ON b.box_id = m.box_id
               WHERE m.transfer_id=%s AND m.movement_type='out'
               GROUP BY b.product_id""",
            (tid,)
        ).fetchall()
        if not agg:
            conn.close()
            return jsonify({'status':'error','message':'No boxes scanned — cannot submit empty transfer'}), 400

        total_boxes, total_qty = 0, 0.0
        today = str(date.today())
        for r in agg:
            qty = float(r['qty'] or 0)
            total_boxes += int(r['box_count'])
            total_qty   += qty
            _post_stock_movement(
                conn,
                product_id  = r['product_id'],
                godown_id   = t['from_godown_id'],
                qty         = qty,
                direction   = 'out',
                transfer_no = t['transfer_no'],
                transfer_id = tid,
                txn_date    = today,
                user        = _user()
            )

        conn.execute(
            """UPDATE pm_transfers
                  SET status='in_pending', total_boxes=%s, total_qty=%s
                WHERE transfer_id=%s""",
            (total_boxes, total_qty, tid)
        )

        # Reconcile the per-line totals (pm_transfer_items) to the actual
        # scanned boxes so the voucher display can never drift from the box
        # count. Group by (product_id, per_box_qty) to match the line unique
        # key. This is what prevents the "154 vs 155" mismatch at the source.
        line_agg = conn.execute(
            """SELECT b.product_id, b.per_box_qty AS per_box_qty,
                      COUNT(*) AS box_count, SUM(b.per_box_qty) AS qty
                 FROM pm_box_movements m
                 JOIN pm_boxes b ON b.box_id = m.box_id
                WHERE m.transfer_id=%s AND m.movement_type='out'
                GROUP BY b.product_id, b.per_box_qty""",
            (tid,)
        ).fetchall()
        for la in line_agg:
            conn.execute(
                """UPDATE pm_transfer_items
                      SET no_of_box=%s, total_qty=%s
                    WHERE transfer_id=%s AND side='out'
                      AND product_id=%s AND per_box_qty=%s""",
                (int(la['box_count'] or 0), float(la['qty'] or 0),
                 tid, la['product_id'], la['per_box_qty'])
            )

        _log_transfer_edit(conn, tid, 'submit_out', f'{total_boxes} box · qty {total_qty}')
        conn.commit()
        conn.close()
        return jsonify({
            'status':'ok','transfer_no':t['transfer_no'],
            'total_boxes':total_boxes,'total_qty':total_qty
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/cancel', methods=['POST'])
@_login_required
def api_transfer_cancel(tid):
    """
    Cancel a transfer that's still in out_started state. Reverts any
    in_transit boxes back to in_stock at the source.
    Cannot cancel after submit_out (use receipt-discrepancy flow instead).
    """
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute(
            "SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)
        ).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Transfer not found'}), 404
        if t['status'] != 'out_started':
            conn.close(); return jsonify({'status':'error','message':"Cannot cancel after OUT submitted"}), 409

        # Revert each in_transit box back to in_stock at source
        boxes = conn.execute(
            """SELECT b.box_id FROM pm_box_movements m
               JOIN pm_boxes b ON b.box_id = m.box_id
               WHERE m.transfer_id=%s AND m.movement_type='out'""",
            (tid,)
        ).fetchall()
        for r in boxes:
            conn.execute(
                "UPDATE pm_boxes SET current_status='in_stock' WHERE box_id=%s",
                (r['box_id'],)
            )
            conn.execute(
                """INSERT INTO pm_box_movements
                     (box_id, transfer_id, movement_type, to_godown_id, qty, moved_by, remarks)
                   VALUES (%s,%s,'cancel',%s,0,%s,%s)""",
                (r['box_id'], tid, t['from_godown_id'], _user(), 'Transfer cancelled — reverted to in_stock')
            )
        conn.execute(
            "UPDATE pm_transfers SET status='cancelled' WHERE transfer_id=%s", (tid,)
        )
        _log_transfer_edit(conn, tid, 'cancel', f'{len(boxes)} box reverted')
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','reverted_boxes': len(boxes)})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/pending', methods=['GET'])
@_login_required
def api_transfer_pending():
    """List transfers with status='in_pending' (waiting to be received).

    Separation of duties: non-admin users do NOT see vouchers they themselves
    created the OUT for. They cannot perform IN actions on those, so showing
    them in the IN queue would be misleading. Admins see everything.
    """
    godown_id = request.args.get('godown_id')
    conn = sampling_portal.get_db_connection()
    try:
        where, params = ["t.status='in_pending'"], []
        if godown_id:
            where.append("t.to_godown_id=%s"); params.append(int(godown_id))
        # Hide own OUT vouchers from the IN queue for non-admins
        if not _is_admin():
            me = (_user() or '').strip()
            if me and me != 'Unknown':
                where.append("(LOWER(COALESCE(t.out_by,'')) <> LOWER(%s))")
                params.append(me)
        rows = conn.execute(f"""
            SELECT t.transfer_id, t.transfer_no, t.from_godown_id, t.to_godown_id,
                   t.out_at, t.out_by, t.total_boxes, t.total_qty, t.remarks,
                   COALESCE(fg.name,'') AS from_name, COALESCE(tg.name,'') AS to_name
            FROM pm_transfers t
            LEFT JOIN procurement_godowns fg ON fg.id=t.from_godown_id
            LEFT JOIN procurement_godowns tg ON tg.id=t.to_godown_id
            WHERE {' AND '.join(where)}
            ORDER BY t.out_at DESC
        """, params).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            d['out_at']    = str(d.get('out_at') or '')
            d['total_qty']= float(d.get('total_qty') or 0)
            out.append(d)
        return jsonify({'status':'ok','transfers':out, 'count':len(out)})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/scan_in', methods=['POST'])
@_login_required
def api_transfer_scan_in(tid):
    """
    Mark a single box as received at destination. Box must currently be
    in_transit and belong to this transfer's OUT scan list.
    """
    d    = request.get_json() or {}
    code = (d.get('box_code') or '').strip().upper()
    if not code:
        return jsonify({'status': 'error', 'message': 'box_code required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute(
            "SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)
        ).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Transfer not found'}), 404
        if t['status'] != 'in_pending':
            conn.close(); return jsonify({'status':'error','message':f"Transfer is in '{t['status']}' state"}), 409

        # Separation of duties: OUT creator cannot perform IN actions
        if _is_out_creator(conn, tid):
            conn.close()
            return jsonify({
                'status':'error',
                'message':'Separation of duties: you created the Material OUT for this transfer, so a different user must perform the Material IN. You can still view or print the voucher from History after it is received.'
            }), 403

        # Confirm box is actually part of this transfer's OUT scans
        b = conn.execute(
            """SELECT b.box_id, b.box_code, b.product_id, b.product_code, b.per_box_qty,
                      b.current_status, p.product_name, p.pm_type
               FROM pm_boxes b
               JOIN pm_products p ON p.id = b.product_id
               WHERE b.short_code=%s OR b.box_code=%s LIMIT 1""",
            (code, code)
        ).fetchone()
        if not b:
            conn.close(); return jsonify({'status':'error','message':f'Box {code} not found'}), 404

        belongs = conn.execute(
            """SELECT 1 FROM pm_box_movements
               WHERE box_id=%s AND transfer_id=%s AND movement_type='out' LIMIT 1""",
            (b['box_id'], tid)
        ).fetchone()
        if not belongs:
            conn.close()
            return jsonify({'status':'error','message':'This box was not part of the OUT scan for this transfer'}), 409

        already_in = conn.execute(
            """SELECT 1 FROM pm_box_movements
               WHERE box_id=%s AND transfer_id=%s AND movement_type='in' LIMIT 1""",
            (b['box_id'], tid)
        ).fetchone()
        if already_in:
            conn.close(); return jsonify({'status':'error','message':'Box already scanned IN for this transfer'}), 409

        if b['current_status'] != 'in_transit':
            conn.close()
            return jsonify({'status':'error','message':f"Box status is '{b['current_status']}' — expected in_transit"}), 409

        # Mark box as in_stock at destination + log movement
        conn.execute(
            "UPDATE pm_boxes SET current_status='in_stock', current_godown_id=%s WHERE box_id=%s",
            (t['to_godown_id'], b['box_id'])
        )
        conn.execute(
            """INSERT INTO pm_box_movements
                 (box_id, transfer_id, movement_type, from_godown_id, to_godown_id, qty, moved_by, remarks)
               VALUES (%s,%s,'in',%s,%s,%s,%s,%s)""",
            (b['box_id'], tid, t['from_godown_id'], t['to_godown_id'],
             float(b['per_box_qty'] or 0), _user(), f"Transfer {t['transfer_no']} IN")
        )
        _log_transfer_edit(conn, tid, 'scan_in', f"Box {b['box_code']}")
        conn.commit()
        conn.close()
        return jsonify({
            'status':'ok',
            'box': {
                'box_id':       b['box_id'],
                'box_code':     b['box_code'],
                'product_id':   b['product_id'],
                'product_name': b['product_name'],
                'pm_type':      b['pm_type'],
                'product_code': b['product_code'],
                'per_box_qty':  float(b['per_box_qty'] or 0)
            }
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/confirm_receipt', methods=['POST'])
@_login_required
def api_transfer_confirm_receipt(tid):
    """
    Finalise IN side: requires ALL expected boxes to have been scanned
    (zero tolerance). Inserts destination-incrementing stock txns,
    sets status='received'.
    """
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute(
            "SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)
        ).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Transfer not found'}), 404
        if t['status'] != 'in_pending':
            conn.close(); return jsonify({'status':'error','message':f"Transfer is in '{t['status']}' state"}), 409

        # Separation of duties: OUT creator cannot confirm receipt of their own transfer
        if _is_out_creator(conn, tid):
            conn.close()
            return jsonify({
                'status':'error',
                'message':'Separation of duties: you created the Material OUT for this transfer, so a different user must confirm receipt. You can still view or print the voucher from History after it is received.'
            }), 403

        # Compare OUT count vs IN count
        counts = conn.execute(
            """SELECT
                  SUM(CASE WHEN movement_type='out' THEN 1 ELSE 0 END) AS out_n,
                  SUM(CASE WHEN movement_type='in'  THEN 1 ELSE 0 END) AS in_n
               FROM pm_box_movements WHERE transfer_id=%s""",
            (tid,)
        ).fetchone()
        out_n = int((counts or {}).get('out_n') or 0)
        in_n  = int((counts or {}).get('in_n')  or 0)
        if in_n < out_n:
            conn.close()
            return jsonify({
                'status':'error',
                'message': f'Cannot confirm — {out_n - in_n} box(es) still missing. Scan all {out_n} expected boxes first.',
                'expected': out_n, 'scanned_in': in_n
            }), 409

        # Aggregate scanned-IN boxes by product → post stock_in to destination
        agg = conn.execute(
            """SELECT b.product_id, COUNT(*) AS box_count, SUM(b.per_box_qty) AS qty
               FROM pm_box_movements m
               JOIN pm_boxes b ON b.box_id = m.box_id
               WHERE m.transfer_id=%s AND m.movement_type='in'
               GROUP BY b.product_id""",
            (tid,)
        ).fetchall()

        today = str(date.today())
        for r in agg:
            _post_stock_movement(
                conn,
                product_id  = r['product_id'],
                godown_id   = t['to_godown_id'],
                qty         = float(r['qty'] or 0),
                direction   = 'in',
                transfer_no = t['transfer_no'],
                transfer_id = tid,
                txn_date    = today,
                user        = _user()
            )

        conn.execute(
            """UPDATE pm_transfers
                  SET status='received', in_at=NOW(), in_by=%s
                WHERE transfer_id=%s""",
            (_user(), tid)
        )
        _log_transfer_edit(conn, tid, 'confirm_receipt', f"In={in_n}/Out={out_n}")
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','transfer_no':t['transfer_no']})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/force_delete', methods=['POST'])
@_login_required
def api_transfer_force_delete(tid):
    """
    Admin-only: soft-delete a transfer that's stuck in 'out_started' with no
    movements. Used to clean up orphans created by failed voucher-no allocation.
    Refuses to delete if any boxes have movements pointing at this transfer
    (those are real work-in-progress records that need real cancel + revert).
    Snapshots into recycle bin — restorable.
    """
    if not _is_admin():
        return jsonify({'status':'error', 'message':'Admin only'}), 403
    d = request.get_json(silent=True) or {}
    reason = (d.get('reason') or '').strip() or None
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute("SELECT transfer_no, status FROM pm_transfers WHERE transfer_id=%s", (tid,)).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Not found'}), 404
        if t['status'] != 'out_started':
            conn.close(); return jsonify({'status':'error','message':f"Refusing — status is '{t['status']}'. Use Cancel for in_pending transfers."}), 409
        # Reject if any boxes are tied to this transfer
        mv = conn.execute(
            "SELECT COUNT(*) AS c FROM pm_box_movements WHERE transfer_id=%s", (tid,)
        ).fetchone()
        if mv and (mv['c'] or 0) > 0:
            conn.close(); return jsonify({'status':'error','message':f"Refusing — {mv['c']} box movements exist. Cancel first to revert boxes."}), 409

        bin_id = _bin_soft_delete(
            conn,
            entity_type  = 'transfer',
            entity_id    = tid,
            entity_label = f"Empty Transfer {t['transfer_no']}",
            parent_table = 'pm_transfers',
            parent_where = 'transfer_id=%s',
            parent_params= (tid,),
            children = [
                {'table': 'pm_transfer_edits',
                 'where': 'transfer_id=%s', 'params': (tid,)},
            ],
            summary = f"Empty draft transfer {t['transfer_no']} · status={t['status']}",
            reason  = reason
        )
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','deleted_id':tid, 'bin_id': bin_id, 'message': 'Transfer moved to recycle bin.'})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/admin_delete', methods=['POST'])
@_login_required
def api_transfer_admin_delete(tid):
    """
    Admin-only soft-delete with full stock reversal. For any voucher in any
    state. The reversal logic mirrors the original posting:

    - If status was 'in_pending' or 'received' (i.e. save_out has run):
        → write 'inward' txns to SOURCE for each OUT line (refund source)
    - If status was 'received' (save_in has run):
        → write 'outward' txns to DESTINATION for each IN line (un-receive)
    - All boxes that touched this voucher get reverted to 'in_stock' at their
      original source location

    Phase 1 soft-delete: a complete snapshot of the voucher header, items,
    edit log, and box movements is captured into pm_recycle_bin BEFORE any
    reversal or deletion happens. Restore re-INSERTs those original rows but
    does NOT undo the reversal stock txns that were written today — those
    must be reviewed and undone manually if a full restore is desired. The
    bin entry's payload includes a 'reversal_summary' for traceability.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin access required'}), 403

    d = request.get_json(silent=True) or {}
    reason = (d.get('reason') or '').strip() or None

    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute("SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Transfer not found'}), 404

        old_status = t['status']
        from_gid   = t['from_godown_id']
        to_gid     = t['to_godown_id']
        xfer_no    = t['transfer_no']
        today      = str(date.today())

        # ── Snapshot EVERYTHING into the recycle bin BEFORE doing anything else.
        # This way, if reversal or delete fails, we still rolled back the
        # whole operation cleanly. We snapshot the header, items, edits, and
        # all box_movements — but we do NOT delete box_movements (we just
        # null out their transfer_id below), so the bin's box_movements rows
        # are a HISTORICAL record, not something to be re-INSERTed on
        # restore.
        bin_id = _bin_soft_delete(
            conn,
            entity_type  = 'transfer',
            entity_id    = tid,
            entity_label = f"Transfer {xfer_no}",
            parent_table = 'pm_transfers',
            parent_where = 'transfer_id=%s',
            parent_params= (tid,),
            children = [
                {'table': 'pm_transfer_items',
                 'where': 'transfer_id=%s', 'params': (tid,)},
                {'table': 'pm_transfer_edits',
                 'where': 'transfer_id=%s', 'params': (tid,)},
            ],
            summary = f"Transfer {xfer_no} · {old_status} · from godown #{from_gid} → #{to_gid}",
            reason  = reason
        )

        # ── Reversal stock-movement logic (unchanged from original hard-delete) ──
        reversal = {'source_refunded':0.0, 'dest_unreceived':0.0, 'boxes_reverted':0}

        # Step 1: refund source for any OUT lines that had decremented stock.
        # We need to read items from BIN (since we just deleted them).
        bin_row = conn.execute(
            "SELECT payload FROM pm_recycle_bin WHERE bin_id=%s", (bin_id,)
        ).fetchone()
        try:
            bin_payload = json.loads(bin_row['payload'] or '{}')
        except Exception:
            bin_payload = {}
        items_snapshot = (bin_payload.get('children') or {}).get('pm_transfer_items', {}).get('rows') or []

        if old_status in ('in_pending','received'):
            for r in items_snapshot:
                if (r.get('side') or '').lower() != 'out':
                    continue
                qty = float(r.get('total_qty') or 0)
                if qty <= 0: continue
                _post_stock_movement(
                    conn,
                    product_id=r['product_id'], godown_id=from_gid,
                    qty=qty, direction='in',
                    transfer_no=xfer_no, transfer_id=tid,
                    txn_date=today, user=_user()
                )
                reversal['source_refunded'] += qty

        if old_status == 'received':
            for r in items_snapshot:
                if (r.get('side') or '').lower() != 'in':
                    continue
                qty = float(r.get('total_qty') or 0)
                if qty <= 0: continue
                _post_stock_movement(
                    conn,
                    product_id=r['product_id'], godown_id=to_gid,
                    qty=qty, direction='out',
                    transfer_no=xfer_no, transfer_id=tid,
                    txn_date=today, user=_user()
                )
                reversal['dest_unreceived'] += qty

        # Step 3: revert each box that ever touched this voucher to in_stock at
        # source. Box movements stay in place but lose their transfer_id link.
        boxes = conn.execute(
            """SELECT DISTINCT b.box_id
               FROM pm_box_movements m
               JOIN pm_boxes b ON b.box_id = m.box_id
               WHERE m.transfer_id = %s""",
            (tid,)
        ).fetchall()
        for r in boxes:
            conn.execute(
                "UPDATE pm_boxes SET current_status='in_stock', current_godown_id=%s WHERE box_id=%s",
                (from_gid, r['box_id'])
            )
            reversal['boxes_reverted'] += 1
        # Detach historical box-movement rows from the (now-deleted) transfer
        conn.execute(
            "UPDATE pm_box_movements SET transfer_id=NULL, remarks=CONCAT(COALESCE(remarks,''),' [transfer ',%s,' soft-deleted by admin]') WHERE transfer_id=%s",
            (xfer_no, tid)
        )

        # If this transfer was fulfilling a Material Request, unlink the
        # request rows so qty_fulfilled is decremented and status auto-
        # rolls-back. Best-effort: never blocks the delete itself.
        try:
            _mr_unlink_transfer_from_request(conn, int(tid))
        except Exception as _unl_err:
            import sys as _sys
            print(f"[pm_stock] MR unlink failed for transfer {tid}: {_unl_err}", file=_sys.stderr)

        conn.commit()
        conn.close()
        return jsonify({
            'status':       'ok',
            'deleted_id':   tid,
            'transfer_no':  xfer_no,
            'old_status':   old_status,
            'reversal':     reversal,
            'bin_id':       bin_id,
            'message':      'Transfer moved to recycle bin. Stock reversal txns written; restore will re-create the voucher header but reversal txns must be reviewed manually.'
        })
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/voucher/create', methods=['POST'])
@_login_required
def api_voucher_create():
    """Create an empty draft transfer voucher. Returns transfer_id, transfer_no.

    Optional `request_id` ties the OUT to a Material Request — on save_out
    the request's qty_fulfilled is updated and its status auto-progresses.
    """
    blocked = _block_if_disabled('mtv')
    if blocked is not None: return blocked
    # Requesters cannot create Material OUT vouchers. They use the
    # Material Request feature to ask for material; a fulfiller then
    # creates the OUT to deliver it.
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    # New-voucher-entries access check (per-user access control)
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked
    d = request.get_json() or {}
    from_id = d.get('from_godown_id')
    to_id   = d.get('to_godown_id')
    remarks = (d.get('remarks') or '').strip()
    request_id_raw = d.get('request_id')
    if not from_id or not to_id:
        return jsonify({'status':'error','message':'from_godown_id and to_godown_id required'}), 400
    if int(from_id) == int(to_id):
        return jsonify({'status':'error','message':'Source and destination must differ'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        # Per-user home godown enforcement — non-admins must transfer
        # FROM their home godown (source-only). Receivers don't go through
        # Material OUT — they wait for an incoming transfer and complete
        # it via Material IN.
        if not _is_admin():
            home = _user_home_godown(conn)
            if home is not None and int(from_id) != home:
                gname = ''
                try:
                    r = conn.execute("SELECT name FROM procurement_godowns WHERE id=%s", (home,)).fetchone()
                    if r: gname = r['name']
                except Exception:
                    pass
                conn.close()
                return jsonify({
                    'status':'error',
                    'message': f"Locked to your home location ({gname or '#'+str(home)}). Source must be your home godown — you can only send transfers from there."
                }), 403

        transfer_no = _next_voucher_no(conn, 'PM-MT', date.today())
        cur = conn.execute(
            """INSERT INTO pm_transfers (transfer_no, from_godown_id, to_godown_id, status, out_by, remarks, request_id)
               VALUES (%s,%s,%s,'out_started',%s,%s,%s)""",
            (transfer_no, int(from_id), int(to_id), _user(), remarks or None,
             int(request_id_raw) if request_id_raw else None)
        )
        tid = cur.lastrowid
        _log_transfer_edit(conn, tid, 'voucher_create',
                           f'From={from_id} → To={to_id}'
                           + (f' · for Request #{request_id_raw}' if request_id_raw else ''))
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','transfer_id':tid,'transfer_no':transfer_no})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transfers/voucher/create_allotment', methods=['POST'])
@_login_required
def api_voucher_create_allotment():
    """
    Create an empty draft Material Allotment voucher (PM-AL/...).

    Allotment vouchers cover material moved within or between godowns for FG
    packing work. Unlike a regular Transfer (which forbids same-godown), an
    allotment may have from_godown_id == to_godown_id. When that's the case,
    the IN side does NOT post a stock-increment txn at save_in time — only
    the OUT decrement is recorded, treating the material as utilised in FG
    packing rather than relocated.

    Body: { from_godown_id, to_godown_id, remarks? }
    Returns: { status, transfer_id, transfer_no }
    """
    blocked = _block_if_disabled('mtv')
    if blocked is not None:
        return blocked
    # Requesters can't create vouchers — they only submit Material Requests.
    blocked = _block_if_requester()
    if blocked is not None:
        return blocked
    # New-voucher-entries access check (per-user access control)
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None:
        return blocked
    d = request.get_json() or {}
    from_id = d.get('from_godown_id')
    to_id   = d.get('to_godown_id')
    remarks = (d.get('remarks') or '').strip()
    if not from_id or not to_id:
        return jsonify({'status':'error','message':'from_godown_id and to_godown_id required'}), 400
    # NOTE: deliberately NO from==to rejection here. Allotments allow same-godown.

    conn = sampling_portal.get_db_connection()
    try:
        # Same home-godown enforcement as regular voucher_create — non-admins
        # can only allot from their home location.
        if not _is_admin():
            home = _user_home_godown(conn)
            if home is not None and int(from_id) != home:
                gname = ''
                try:
                    r = conn.execute("SELECT name FROM procurement_godowns WHERE id=%s", (home,)).fetchone()
                    if r: gname = r['name']
                except Exception:
                    pass
                conn.close()
                return jsonify({
                    'status':'error',
                    'message': f"Locked to your home location ({gname or '#'+str(home)}). Source must be your home godown."
                }), 403

        transfer_no = _next_voucher_no(conn, 'PM-AL', date.today())
        cur = conn.execute(
            """INSERT INTO pm_transfers (transfer_no, from_godown_id, to_godown_id, voucher_type, status, out_by, remarks)
               VALUES (%s,%s,%s,'allotment','out_started',%s,%s)""",
            (transfer_no, int(from_id), int(to_id), _user(), remarks or None)
        )
        tid = cur.lastrowid
        same = (int(from_id) == int(to_id))
        _log_transfer_edit(conn, tid, 'allotment_create',
                           f'From={from_id} → To={to_id}{" (same godown)" if same else ""}')
        conn.commit()
        conn.close()
        return jsonify({
            'status':'ok',
            'transfer_id': tid,
            'transfer_no': transfer_no,
            'voucher_type':'allotment',
            'same_godown': same,
        })
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transfers/voucher/create_allotment_with_box', methods=['POST'])
@_login_required
def api_voucher_create_allotment_with_box():
    """
    One-shot creation of a Material Allotment voucher seeded by a scanned box.

    The picker modal calls this when the operator scans a box code instead of
    picking godowns by hand. The box's current godown becomes the source AND
    destination by default (same-godown allotment for FG packing). The first
    OUT-side scan is recorded in the same transaction so the operator can
    keep scanning more boxes immediately in the OUT modal that opens next.

    Body: { box_code: str, to_godown_id?: int, remarks?: str }
      - to_godown_id is optional. If omitted, defaults to the box's current
        godown (same-godown allotment). If supplied, used as-is — allows the
        operator to scan a box at Godown A but allot it to Godown B if they
        explicitly chose that destination in the picker before scanning.
      - same_godown is computed from the resulting from/to pair.

    Returns: { status, transfer_id, transfer_no, voucher_type:'allotment',
               same_godown, scanned_box: {...} }
    Errors mirror api_voucher_scan_box for box-related rejections so the
    picker can show a single, consistent toast.
    """
    blocked = _block_if_disabled('mtv')
    if blocked is not None:
        return blocked
    # Requesters can't create vouchers — they only submit Material Requests.
    blocked = _block_if_requester()
    if blocked is not None:
        return blocked
    # New-voucher-entries access check (per-user access control)
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None:
        return blocked
    d = request.get_json() or {}
    code = (d.get('box_code') or '').strip().upper()
    to_id_in = d.get('to_godown_id')
    remarks  = (d.get('remarks') or '').strip()
    if not code:
        return jsonify({'status':'error','message':'box_code required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # ── Look up the box ───────────────────────────────────────────────
        b = conn.execute(
            """SELECT b.box_id, b.box_code, b.product_id, b.product_code, b.per_box_qty,
                      b.current_godown_id, b.current_status, p.product_name, p.pm_type
               FROM pm_boxes b
               JOIN pm_products p ON p.id = b.product_id
               WHERE b.short_code=%s OR b.box_code=%s LIMIT 1""",
            (code, code)
        ).fetchone()
        if not b:
            conn.close(); return jsonify({'status':'error','message':f'Box {code} not found'}), 404
        if b['current_status'] != 'in_stock':
            conn.close()
            return jsonify({'status':'error','message':f"Box status is '{b['current_status']}' — must be in_stock to allot"}), 409
        if not b['current_godown_id']:
            conn.close()
            return jsonify({'status':'error','message':'Box has no current godown — cannot determine source'}), 409

        from_id = int(b['current_godown_id'])
        to_id   = int(to_id_in) if to_id_in else from_id
        same    = (from_id == to_id)

        # Home-godown lock for non-admin: must allot from your own godown.
        if not _is_admin():
            home = _user_home_godown(conn)
            if home is not None and from_id != int(home):
                gname = ''
                try:
                    r = conn.execute("SELECT name FROM procurement_godowns WHERE id=%s", (home,)).fetchone()
                    if r: gname = r['name']
                except Exception:
                    pass
                conn.close()
                return jsonify({
                    'status':'error',
                    'message': f"Box is at a different godown than your home location ({gname or '#'+str(home)}) — cannot allot."
                }), 403

        # ── Create the voucher ────────────────────────────────────────────
        transfer_no = _next_voucher_no(conn, 'PM-AL', date.today())
        cur = conn.execute(
            """INSERT INTO pm_transfers (transfer_no, from_godown_id, to_godown_id, voucher_type, status, out_by, remarks)
               VALUES (%s,%s,%s,'allotment','out_started',%s,%s)""",
            (transfer_no, from_id, to_id, _user(), remarks or None)
        )
        tid = cur.lastrowid
        _log_transfer_edit(conn, tid, 'allotment_create',
                           f'From={from_id} → To={to_id}{" (same godown)" if same else ""} · seeded by scan {code}')

        # ── Record the first OUT scan ─────────────────────────────────────
        # Mirrors the OUT-side flow inside api_voucher_scan_box. We
        # deliberately skip the FIFO check here because allotment voucher
        # creation is the operator's first action — they're scanning what's
        # in front of them, not choosing a box from a list. FIFO will still
        # apply on subsequent scans through the regular OUT modal flow.
        conn.execute(
            """INSERT INTO pm_box_movements
                 (box_id, transfer_id, movement_type, from_godown_id, to_godown_id, qty, moved_by, remarks)
               VALUES (%s,%s,'out',%s,%s,%s,%s,%s)""",
            (b['box_id'], tid, from_id, to_id,
             float(b['per_box_qty'] or 0), _user(),
             f'Allotment first-scan: {code}')
        )
        conn.execute(
            "UPDATE pm_boxes SET current_status='in_transit' WHERE box_id=%s",
            (b['box_id'],)
        )

        # Add OUT line for this product. Under the v3 unique key
        #   (transfer_id, side, product_id, per_box_qty)
        # ON DUPLICATE KEY only fires when the EXACT same (product, pack-size)
        # is scanned again in the same voucher — i.e. genuinely another box of
        # the same flavour. In that case just bump no_of_box; do NOT touch
        # per_box_qty because (a) it's part of the key and (b) the values are
        # the same by definition. A second pack size of the same product gets
        # its own row because the wider unique key lets it.
        per_box = float(b['per_box_qty'] or 0)
        conn.execute(
            """INSERT INTO pm_transfer_items
                 (transfer_id, side, product_id, no_of_box, per_box_qty, total_qty)
               VALUES (%s, 'out', %s, 1, %s, %s)
               ON DUPLICATE KEY UPDATE
                 no_of_box = no_of_box + 1,
                 total_qty = total_qty + VALUES(per_box_qty)""",
            (tid, b['product_id'], per_box, per_box)
        )

        _refresh_transfer_totals(conn, tid)
        conn.commit()
        conn.close()
        return jsonify({
            'status':'ok',
            'transfer_id': tid,
            'transfer_no': transfer_no,
            'voucher_type':'allotment',
            'same_godown': same,
            'from_godown_id': from_id,
            'to_godown_id': to_id,
            'scanned_box': {
                'box_id'       : b['box_id'],
                'box_code'     : b['box_code'],
                'product_name' : b['product_name'],
                'pm_type'      : b['pm_type'],
                'per_box_qty'  : float(b['per_box_qty'] or 0),
            },
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transfers/voucher/<int:tid>/scan_box', methods=['POST'])
@_login_required
def api_voucher_scan_box(tid):
    """
    Scan a box into a voucher. side='out' or 'in' (default out).
    For OUT: validates box currently lives at source, marks it in_transit, adds
             to OUT lines.
    For IN:  destination scan; allows extras, just records what arrived.
    Behaviour: if line for this product exists in this voucher+side, increment
               its no_of_box by 1; else create new line with per_box_qty from
               the box.
    """
    d = request.get_json() or {}
    side = (d.get('side') or 'out').lower()
    if side not in ('out','in'):
        return jsonify({'status':'error','message':'side must be out or in'}), 400
    code = (d.get('box_code') or '').strip().upper()
    if not code:
        return jsonify({'status':'error','message':'box_code required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute(
            "SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)
        ).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Transfer not found'}), 404

        if side == 'out' and t['status'] != 'out_started':
            conn.close(); return jsonify({'status':'error','message':f"OUT scanning closed — status is '{t['status']}'"}), 409
        if side == 'in' and t['status'] != 'in_pending':
            conn.close(); return jsonify({'status':'error','message':f"IN scanning not open — status is '{t['status']}'"}), 409

        # Separation of duties: OUT creator cannot perform IN scans on their own transfer
        if side == 'in' and _is_out_creator(conn, tid):
            conn.close()
            return jsonify({
                'status':'error',
                'message':'Separation of duties: you created the Material OUT for this transfer, so a different user must scan the Material IN. You can still view or print the voucher from History.'
            }), 403

        # Non-admin home godown lock: OUT scans only allowed if you are at source;
        # IN scans only allowed if you are at destination.
        if not _is_admin():
            home = _user_home_godown(conn)
            if home is not None:
                expected = t['from_godown_id'] if side == 'out' else t['to_godown_id']
                if int(home) != int(expected):
                    conn.close()
                    return jsonify({
                        'status':'error',
                        'message': f"You're locked to a different location and cannot {side.upper()}-scan this voucher."
                    }), 403

        b = conn.execute(
            """SELECT b.box_id, b.box_code, b.product_id, b.product_code, b.per_box_qty,
                      b.current_godown_id, b.current_status, p.product_name, p.pm_type
               FROM pm_boxes b
               JOIN pm_products p ON p.id = b.product_id
               WHERE b.short_code=%s OR b.box_code=%s LIMIT 1""",
            (code, code)
        ).fetchone()
        if not b:
            conn.close(); return jsonify({'status':'error','message':f'Box {code} not found'}), 404

        # OUT-side validation: box must be at source and in_stock
        if side == 'out':
            # Location check WITH SELF-HEAL. If the cached current_godown_id
            # on pm_boxes disagrees with the scan source, ask the helper
            # to check the box's last-movement row. If the ledger says
            # the box IS where the operator is scanning, the row is just
            # stale — heal it silently and let the scan proceed (with
            # auto_corrected:true in the response so the frontend can
            # show a toast). If the ledger backs up the cached value,
            # genuine mismatch — error as before.
            box_auto_corrected = False
            if b['current_godown_id'] != t['from_godown_id']:
                _verdict, _why = _heal_box_location(
                    conn,
                    box_id=b['box_id'],
                    expected_godown_id=t['from_godown_id'],
                    current_godown_id=b['current_godown_id'],
                    user=_user(),
                )
                if _verdict == 'heal':
                    # Update the local row dict so downstream code sees
                    # the healed location (we already wrote pm_boxes).
                    b = dict(b)
                    b['current_godown_id'] = t['from_godown_id']
                    box_auto_corrected = True
                else:
                    conn.close()
                    return jsonify({
                        'status':  'error',
                        'message': 'Box is at a different location — cannot OUT from this source',
                        'box_current_godown_id': b['current_godown_id'],
                        'expected_godown_id':    t['from_godown_id'],
                        'heal_reason':           _why,
                    }), 409
            if b['current_status'] == 'superseded':
                # Box was split. Look up the children so the operator knows
                # which boxes to scan instead.
                kids = conn.execute(
                    "SELECT box_code FROM pm_boxes WHERE parent_box_id=%s ORDER BY box_seq",
                    (b['box_id'],)
                ).fetchall()
                kids_str = ', '.join(r['box_code'] for r in kids) if kids else '(unknown)'
                conn.close()
                return jsonify({
                    'status':  'error',
                    'code':    'box_superseded',
                    'message': f'Box was split into {len(kids)} children: {kids_str}. Scan one of those instead.'
                }), 409
            if b['current_status'] != 'in_stock':
                conn.close()
                return jsonify({'status':'error','message':f"Box status is '{b['current_status']}' — must be in_stock"}), 409
            # Already scanned for this voucher's OUT?
            already = conn.execute(
                """SELECT 1 FROM pm_box_movements
                   WHERE box_id=%s AND transfer_id=%s AND movement_type='out' LIMIT 1""",
                (b['box_id'], tid)
            ).fetchone()
            if already:
                conn.close()
                return jsonify({'status':'error','message':'This box was already scanned for OUT'}), 409

            # ── Material Lock enforcement (INDEPENDENT of FIFO) ──────────
            # A Manager/admin may lock material so it can't be OUT-scanned at
            # a location — either everything received before a cutoff date, or
            # a specific GRN. This gate runs BEFORE FIFO and is unrelated to
            # it: a locked box is refused outright, with no override path.
            try:
                _lk = conn.execute(
                    """SELECT b.grn_id, b.grn_no, g.grn_date
                       FROM pm_boxes b
                       LEFT JOIN pm_grn g ON g.id = b.grn_id
                       WHERE b.box_id=%s LIMIT 1""",
                    (b['box_id'],)
                ).fetchone()
                _is_op = bool(_lk) and (
                    (_lk['grn_id'] is None) or
                    str(_lk['grn_no'] or '').upper().startswith('PM-OP/')
                )
                _blocked, _why = _material_lock_check(
                    conn,
                    product_id=b['product_id'],
                    godown_id=b['current_godown_id'],
                    grn_id=(_lk['grn_id'] if _lk else None),
                    grn_date=(_lk['grn_date'] if _lk else None),
                    is_opening=_is_op,
                )
            except Exception:
                _blocked, _why = (False, None)
            if _blocked:
                conn.close()
                return jsonify({
                    'status':  'error',
                    'code':    'material_locked',
                    'message': f"🔒 {_why}. Contact a Manager to unlock."
                }), 423

            # ── FIFO enforcement ─────────────────────────────────────────
            # Block OUT scans that violate FIFO (box from a newer lot when
            # an older lot still has stock at this source). Admins can
            # bypass with `force_fifo_override: true` in the request body —
            # the override is logged on the transfer's edit history for
            # audit. Non-admins can never override.
            try:
                fifo_violation = _fifo_check_oldest(
                    conn,
                    product_id     = b['product_id'],
                    godown_id      = t['from_godown_id'],
                    scanned_box_id = b['box_id']
                )
            except Exception as _fe:
                # FIFO check is advisory — don't bomb the scan if it fails.
                import sys as _sys
                print(f"[pm_stock_routes] FIFO check failed for box {b['box_code']}: {_fe}", file=_sys.stderr)
                fifo_violation = None

            if fifo_violation:
                force = bool(d.get('force_fifo_override'))
                is_admin_user = _is_admin()

                # Non-admin path: they can't force directly. But if an admin
                # has already APPROVED an override request for this exact box
                # on this transfer, let it through (single-use) and consume
                # the approval. Otherwise, return a 'requestable' response so
                # the client shows the "Request override" UI instead of a
                # dead-end block.
                approved_req_id = None
                if not is_admin_user:
                    approved_req_id = _fifo_override_find_approved(
                        conn, tid, b['box_id'], _user()
                    )
                    if approved_req_id:
                        # Consume the pass and log it on the transfer trail.
                        _fifo_override_mark_used(conn, approved_req_id, _user())
                        try:
                            _log_transfer_edit(
                                conn, tid, 'fifo_override',
                                f"Approved override #{approved_req_id} used by {_user()}: "
                                f"scanned {fifo_violation['scanned_fifo_code']} "
                                f"despite older lot {fifo_violation['oldest_fifo_code']} "
                                f"({fifo_violation['oldest_box_count']} box(es) pending) "
                                f"from voucher {fifo_violation['oldest_voucher']}"
                            )
                        except Exception:
                            pass
                        try:
                            _audit_record(
                                conn,
                                action='fifo_override.used',
                                entity='fifo_override_request',
                                entity_id=approved_req_id,
                                summary=f"FIFO override #{approved_req_id} consumed on {t['transfer_no']} box {b['box_code']}",
                                before={'status': 'approved'},
                                after={'status': 'used', 'box_code': b['box_code']},
                            )
                        except Exception:
                            pass
                        # fall through — scan proceeds normally below
                    else:
                        # No approved pass. Block, but tell the client this
                        # can be requested. Surface any pending request id so
                        # the UI can say "already requested".
                        pending = None
                        try:
                            pr = conn.execute(
                                """SELECT req_id, status FROM pm_fifo_override_requests
                                   WHERE transfer_id=%s AND box_id=%s AND requested_by=%s
                                     AND status IN ('pending','rejected')
                                   ORDER BY requested_at DESC LIMIT 1""",
                                (tid, b['box_id'], _user())
                            ).fetchone()
                            if pr:
                                pending = {'req_id': int(pr['req_id']), 'status': pr['status']}
                        except Exception:
                            pending = None
                        conn.close()
                        sf = fifo_violation['scanned_fifo_code'] or '(no FIFO)'
                        of = fifo_violation['oldest_fifo_code']  or '(no FIFO)'
                        bc = fifo_violation['oldest_box_count']
                        qt = fifo_violation['oldest_total_qty']
                        msg_lines = [
                            f"FIFO violation — scanned {sf}, but lot {of} must go first.",
                            f"Lot {of}: {bc} box(es) · {qt:g} pcs still at this location.",
                            "Ask an admin to approve an override.",
                        ]
                        return jsonify({
                            'status':  'error',
                            'code':    'fifo_override_requestable',
                            'message': '\n'.join(msg_lines),
                            'fifo':    fifo_violation,
                            'box_id':  b['box_id'],
                            'pending': pending,
                        }), 409

                elif not force:
                    conn.close()
                    # Admin who hasn't forced — build the rich block payload.
                    sf = fifo_violation['scanned_fifo_code'] or '(no FIFO)'
                    of = fifo_violation['oldest_fifo_code']  or '(no FIFO)'
                    bc = fifo_violation['oldest_box_count']
                    qt = fifo_violation['oldest_total_qty']
                    vno = fifo_violation['oldest_voucher']
                    sup = fifo_violation['oldest_supplier']
                    dt  = fifo_violation['oldest_date']
                    msg_lines = [
                        f"FIFO violation — scanned {sf}, but lot {of} must go first.",
                        f"Lot {of}: {bc} box(es) · {qt:g} pcs still at this location.",
                    ]
                    if vno:
                        bits = [vno]
                        if dt: bits.append(f"dated {dt}")
                        if sup: bits.append(f"from {sup}")
                        msg_lines.append('Source voucher: ' + ' · '.join(bits))
                    return jsonify({
                        'status': 'error',
                        'code':   'fifo_violation',
                        'message': '\n'.join(msg_lines),
                        'fifo':   fifo_violation
                    }), 409
                else:
                    # Admin override — log it on the transfer audit trail.
                    try:
                        _log_transfer_edit(
                            conn, tid, 'fifo_override',
                            f"Admin override: scanned {fifo_violation['scanned_fifo_code']} "
                            f"despite older lot {fifo_violation['oldest_fifo_code']} "
                            f"({fifo_violation['oldest_box_count']} box(es) pending) "
                            f"from voucher {fifo_violation['oldest_voucher']}"
                        )
                    except Exception:
                        pass
            # ─────────────────────────────────────────────────────────────

            # Mark in_transit + log box movement
            conn.execute("UPDATE pm_boxes SET current_status='in_transit' WHERE box_id=%s", (b['box_id'],))
            conn.execute(
                """INSERT INTO pm_box_movements
                     (box_id, transfer_id, movement_type, from_godown_id, to_godown_id, qty, moved_by, remarks)
                   VALUES (%s,%s,'out',%s,%s,%s,%s,%s)""",
                (b['box_id'], tid, t['from_godown_id'], t['to_godown_id'],
                 float(b['per_box_qty'] or 0), _user(), f"Voucher {t['transfer_no']} OUT scan")
            )

        # IN-side: must have been part of the OUT (or extra — flagged later in discrepancy)
        else:  # side == 'in'
            already_in = conn.execute(
                """SELECT 1 FROM pm_box_movements
                   WHERE box_id=%s AND transfer_id=%s AND movement_type='in' LIMIT 1""",
                (b['box_id'], tid)
            ).fetchone()
            if already_in:
                conn.close()
                return jsonify({'status':'error','message':'Box already scanned for IN'}), 409

            # ── Strict membership check ──────────────────────────────────────
            # The box must have been part of THIS voucher's OUT scan. Without
            # this check, an operator could accidentally pick up a box from
            # somewhere else (a different in-transit voucher, or a box that
            # was supposed to stay at source) and "receive" it into this
            # voucher, silently mis-allocating stock.
            #
            # Match logic: there must exist a pm_box_movements row with
            #   transfer_id = this transfer
            #   movement_type = 'out'
            #   box_id = the scanned box
            # i.e. the operator who did the OUT scan included this exact box.
            in_out_batch = conn.execute(
                """SELECT 1 FROM pm_box_movements
                   WHERE box_id=%s AND transfer_id=%s AND movement_type='out'
                   LIMIT 1""",
                (b['box_id'], tid)
            ).fetchone()
            if not in_out_batch:
                conn.close()
                # Try to give a useful hint — is this box currently in another
                # in-transit voucher? Helps operator find where to scan it.
                hint = ''
                try:
                    other = conn.execute(
                        """SELECT t.transfer_no
                           FROM pm_box_movements m
                           JOIN pm_transfers t ON t.transfer_id = m.transfer_id
                           WHERE m.box_id=%s AND m.movement_type='out'
                             AND m.transfer_id <> %s
                             AND t.status IN ('in_pending','out_started')
                           ORDER BY m.movement_id DESC LIMIT 1""",
                        (b['box_id'], tid)
                    ).fetchone()
                    if other:
                        hint = f" (this box belongs to {other['transfer_no']})"
                except Exception:
                    pass
                return jsonify({
                    'status':  'error',
                    'code':    'box_not_in_voucher',
                    'message': f'Item not in this voucher{hint}',
                }), 409

            # Move box to destination + log
            conn.execute(
                "UPDATE pm_boxes SET current_status='in_stock', current_godown_id=%s WHERE box_id=%s",
                (t['to_godown_id'], b['box_id'])
            )
            conn.execute(
                """INSERT INTO pm_box_movements
                     (box_id, transfer_id, movement_type, from_godown_id, to_godown_id, qty, moved_by, remarks)
                   VALUES (%s,%s,'in',%s,%s,%s,%s,%s)""",
                (b['box_id'], tid, t['from_godown_id'], t['to_godown_id'],
                 float(b['per_box_qty'] or 0), _user(), f"Voucher {t['transfer_no']} IN scan")
            )

        # Increment-or-insert the line item for this product+per_box_qty bucket.
        #
        # IMPORTANT: We key the merge on (product_id, per_box_qty) — NOT just
        # product_id. A single product can have multiple GRN lines with
        # different per-box quantities (e.g. 9 boxes of 4000 PLUS 1 box of
        # 2400 — same Plix Tray product, different pack sizes). When boxes
        # from both packs scan into the same MTV, each per-box-qty bucket
        # must stay on its own line so the printed voucher reads:
        #     Plix Tray  9 × 4000 = 36000
        #     Plix Tray  1 × 2400 =  2400
        # Without the per_box_qty match, both scans would collapse into one
        # line showing "10 × 4000" with total_qty=38400, which is arithmetic
        # nonsense (10 × 4000 = 40000, not 38400) and silently misrepresents
        # what physically arrived.
        per_box = float(b['per_box_qty'] or 0)
        existing = conn.execute(
            """SELECT item_id, no_of_box, total_qty
               FROM pm_transfer_items
               WHERE transfer_id=%s AND side=%s AND product_id=%s
                 AND ABS(per_box_qty - %s) < 0.001
               LIMIT 1""",
            (tid, side, b['product_id'], per_box)
        ).fetchone()
        if existing:
            new_nob = int(existing['no_of_box']) + 1
            new_qty = float(existing['total_qty']) + per_box
            conn.execute(
                "UPDATE pm_transfer_items SET no_of_box=%s, total_qty=%s WHERE item_id=%s",
                (new_nob, new_qty, existing['item_id'])
            )
        else:
            conn.execute(
                """INSERT INTO pm_transfer_items
                     (transfer_id, side, product_id, no_of_box, per_box_qty, total_qty)
                   VALUES (%s,%s,%s,1,%s,%s)""",
                (tid, side, b['product_id'], per_box, per_box)
            )

        # Refresh aggregate totals on the header
        if side == 'out':
            _refresh_transfer_totals(conn, tid)
        _log_transfer_edit(conn, tid, f'scan_{side}', f"Box {b['box_code']}")

        # ── Request target info (overscan warning) ────────────────────
        # If this transfer was created to fulfil a Material Request, surface
        # the request line for THIS product so the frontend can pop a
        # "Scanned qty exceeds requested qty — continue?" confirmation modal.
        # Sent ONLY on OUT side; IN side doesn't affect request fulfilment
        # counters and doesn't need the warning.
        request_target = None
        try:
            if side == 'out':
                req_id_on_t = t.get('request_id') if hasattr(t, 'get') else (t['request_id'] if 'request_id' in t.keys() else None)
                if req_id_on_t:
                    # Fetch the request item for THIS product. There's only one
                    # line per product in a request (UI enforces it), so we
                    # take the first match.
                    ri = conn.execute(
                        """SELECT ri.id, ri.qty_requested, ri.qty_fulfilled
                           FROM pm_material_request_items ri
                           WHERE ri.request_id=%s AND ri.product_id=%s
                           LIMIT 1""",
                        (int(req_id_on_t), int(b['product_id']))
                    ).fetchone()
                    # Fetch the total qty already scanned for this product on
                    # THIS transfer (so far in the OUT side). This is the
                    # number we compare against qty_requested.
                    so_far_row = conn.execute(
                        """SELECT COALESCE(SUM(total_qty), 0) AS s
                           FROM pm_transfer_items
                           WHERE transfer_id=%s AND side='out' AND product_id=%s""",
                        (tid, int(b['product_id']))
                    ).fetchone()
                    so_far = float(so_far_row['s']) if so_far_row else 0.0
                    if ri:
                        req_qty = float(ri['qty_requested'] or 0)
                        already_ful_other = float(ri['qty_fulfilled'] or 0)  # from prior OUT vouchers
                        # Remaining qty BEFORE this OUT started
                        remaining_at_start = max(0.0, req_qty - already_ful_other)
                        # How much over the requested amount we are right now,
                        # counting this scan
                        scanned_total = so_far  # already includes this scan (we inserted/updated above)
                        request_target = {
                            'request_id':         int(req_id_on_t),
                            'qty_requested':      req_qty,
                            'qty_fulfilled_other':already_ful_other,  # from earlier OUT vouchers
                            'qty_remaining_at_start': remaining_at_start,
                            'qty_scanned_so_far': scanned_total,        # in THIS OUT, for THIS product
                            'over_by':            max(0.0, scanned_total - remaining_at_start),
                            'exceeded':           scanned_total > remaining_at_start + 0.001,
                        }
                    else:
                        # Box scanned isn't on the request — flag it so the
                        # frontend can warn (e.g. "this product wasn't requested,
                        # are you sure?"). Same modal pattern.
                        request_target = {
                            'request_id':   int(req_id_on_t),
                            'not_in_request': True,
                            'qty_scanned_so_far': so_far,
                        }
        except Exception as _rt_err:
            # Best-effort: a failure here must not block the scan itself
            import sys as _sys
            print(f"[pm_stock] scan_box request_target lookup failed: {_rt_err}", file=_sys.stderr)
            request_target = None

        conn.commit()
        conn.close()
        return jsonify({
            'status':'ok',
            'box': {
                'box_code':     b['box_code'],
                'product_id':   b['product_id'],
                'product_name': b['product_name'],
                'pm_type':      b['pm_type'],
                'product_code': b['product_code'],
                'per_box_qty':  per_box,
            },
            'request_target': request_target,
            # Auto-heal flag: when true, the box's stored current_godown_id
            # had drifted from where the ledger said it should be and we
            # silently corrected it. The frontend can show an info toast
            # so operators know the system self-corrected rather than
            # accepting the scan from the wrong location. Only present
            # on the OUT path — IN scans always set the location, so the
            # concept doesn't apply there.
            'auto_corrected': bool(locals().get('box_auto_corrected', False)),
        })
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/voucher/<int:tid>', methods=['GET'])
@_login_required
def api_voucher_get(tid):
    """Return full voucher: header + OUT lines + IN lines + discrepancy state."""
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute("""
            SELECT t.*,
                   COALESCE(fg.name,'') AS from_name, COALESCE(fg.type,'') AS from_type,
                   COALESCE(tg.name,'') AS to_name,   COALESCE(tg.type,'') AS to_type
            FROM pm_transfers t
            LEFT JOIN procurement_godowns fg ON fg.id=t.from_godown_id
            LEFT JOIN procurement_godowns tg ON tg.id=t.to_godown_id
            WHERE t.transfer_id=%s
        """, (tid,)).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Not found'}), 404

        def _fetch(side):
            return conn.execute("""
                SELECT i.item_id, i.product_id, i.no_of_box, i.per_box_qty, i.total_qty,
                       COALESCE(i.remarks,'') AS remarks,
                       p.product_name, p.pm_type, COALESCE(p.product_code,'') AS product_code,
                       COALESCE(b.name,'') AS brand_name,
                       -- UOM (Phase 3) — pull product master UOM so the print shows
                       -- the primary unit label, and (if this transfer line was
                       -- fulfilling a Material Request line that was entered in an
                       -- alternate UOM) pull the alt UOM + ratio so the print can
                       -- show "120 Nos = 0.008 Kg" style conversion matrix.
                       -- The LEFT JOIN through pm_material_request_links is loose:
                       -- if the same product appears on multiple linked MR lines,
                       -- MAX() picks any one entered_uom — they should all be the
                       -- same since the alt UOM is set per-product, not per-line.
                       COALESCE(p.primary_uom,'Nos')    AS primary_uom,
                       COALESCE(p.alt_uom,'')           AS alt_uom,
                       p.alt_to_primary_ratio           AS alt_to_primary_ratio,
                       (SELECT MAX(ri.entered_uom)
                          FROM pm_material_request_links lk
                          JOIN pm_material_request_items ri ON ri.id = lk.request_item_id
                         WHERE lk.transfer_id=i.transfer_id
                           AND ri.product_id=i.product_id
                           AND ri.entered_uom IS NOT NULL
                       ) AS linked_entered_uom,
                       (SELECT SUM(ri.entered_qty)
                          FROM pm_material_request_links lk
                          JOIN pm_material_request_items ri ON ri.id = lk.request_item_id
                         WHERE lk.transfer_id=i.transfer_id
                           AND ri.product_id=i.product_id
                           AND ri.entered_qty IS NOT NULL
                       ) AS linked_entered_qty_total
                FROM pm_transfer_items i
                JOIN pm_products p ON p.id = i.product_id
                LEFT JOIN procurement_brands b ON b.id = p.brand_id
                WHERE i.transfer_id=%s AND i.side=%s
                ORDER BY p.product_name
            """, (tid, side)).fetchall()
        out_items = [dict(r) for r in _fetch('out')]
        in_items  = [dict(r) for r in _fetch('in')]
        for arr in (out_items, in_items):
            for r in arr:
                r['per_box_qty'] = float(r.get('per_box_qty') or 0)
                r['total_qty']   = float(r.get('total_qty')   or 0)
                r['no_of_box']   = int(r.get('no_of_box')     or 0)
                # Normalize UOM fields for JSON: ratio→float, nulls→empty str
                r['primary_uom'] = r.get('primary_uom') or 'Nos'
                r['alt_uom']     = r.get('alt_uom') or ''
                _ratio = r.get('alt_to_primary_ratio')
                r['alt_to_primary_ratio'] = float(_ratio) if _ratio is not None else None
                r['linked_entered_uom'] = r.get('linked_entered_uom') or ''
                _leq = r.get('linked_entered_qty_total')
                r['linked_entered_qty_total'] = float(_leq) if _leq is not None else None

        # Compute discrepancy view fresh (don't trust stale flag if lines moved)
        has_d, mismatches = _check_discrepancy(conn, tid)
        # Separation of duties: is this user the OUT creator? If so, IN actions
        # are blocked for them (read-only / print only).
        in_locked = _is_out_creator(conn, tid)

        # ── Per-line scanned-box detail ──────────────────────────────
        # The OUT-side modal lets operators remove individual scanned
        # boxes (via api_voucher_unscan_box). To render the "click × on
        # this box code" UI we need to know which boxes belong to which
        # line. Group by (product_id, per_box_qty within tolerance) —
        # same grain as pm_transfer_items uniqueness — so each line
        # gets exactly its own scanned boxes.
        out_box_rows = conn.execute(
            """SELECT m.movement_id, m.box_id, m.movement_type,
                      b.box_code, b.short_code, b.product_id, b.per_box_qty,
                      b.current_status
               FROM pm_box_movements m
               JOIN pm_boxes b ON b.box_id = m.box_id
               WHERE m.transfer_id=%s AND m.movement_type IN ('out','in')
               ORDER BY m.movement_id ASC""",
            (tid,)
        ).fetchall()

        def _attach_boxes(items, side):
            for it in items:
                it['boxes'] = []
            for r in out_box_rows:
                if r['movement_type'] != side:
                    continue
                for it in items:
                    if int(it['product_id']) == int(r['product_id'] or 0) \
                       and abs(float(it['per_box_qty'] or 0) - float(r['per_box_qty'] or 0)) < 0.001:
                        it['boxes'].append({
                            'box_id':     int(r['box_id']),
                            'box_code':   r['box_code'] or '',
                            'short_code': r['short_code'] or '',
                        })
                        break
        _attach_boxes(out_items, 'out')
        _attach_boxes(in_items,  'in')

        conn.commit()  # save refreshed flag
        conn.close()

        hdr = dict(t)
        for k in ('out_at','in_at'):
            if hdr.get(k) is not None: hdr[k] = str(hdr[k])
        hdr['total_qty']       = float(hdr.get('total_qty') or 0)
        hdr['has_discrepancy'] = bool(hdr.get('has_discrepancy'))
        return jsonify({
            'status':'ok',
            'header': hdr,
            'out_items': out_items,
            'in_items':  in_items,
            'has_discrepancy': has_d,
            'mismatches': mismatches,
            'in_locked_for_user': bool(in_locked),
            'in_locked_reason': ('You created the Material OUT for this transfer; a different user must perform the Material IN. View / print only.' if in_locked else None)
        })
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/voucher/<int:tid>/lines/<int:item_id>', methods=['PATCH','DELETE'])
@_login_required
def api_voucher_edit_line(tid, item_id):
    """
    Edit or delete a single line item on a voucher.
    PATCH body: { no_of_box?, per_box_qty? }

    Permissions:
      * PATCH (quantity edits) — admin only, always.
      * DELETE (remove line)   — admin always; non-admins ONLY when the
                                  voucher is still a draft (out_started) and
                                  the line is on the OUT side. This lets the
                                  scanning operator clean up a mis-scanned
                                  product without needing an admin escalation.
                                  Posted / in-transit / IN-side lines stay
                                  admin-only so audited stock movements can't
                                  be silently rewritten.

    Stock txns are recalculated by the save_out / save_in flow next time, but
    if save_out has already happened we adjust the corresponding godown txn
    here directly so totals stay consistent.
    """
    is_admin = (session.get('User_Type','').lower() == 'admin')

    # A PATCH that ONLY sets remarks is allowed for everyone — a per-item note
    # isn't a stock change. Quantity edits (no_of_box / per_box_qty) stay
    # admin-only. Detect the remarks-only case up front.
    _patch_body = request.get_json() or {} if request.method == 'PATCH' else {}
    _remarks_only = (
        request.method == 'PATCH'
        and 'remarks' in _patch_body
        and _patch_body.get('no_of_box') is None
        and _patch_body.get('per_box_qty') is None
    )

    # Quantity edits stay admin-only (remarks-only edits are exempt).
    if request.method == 'PATCH' and not is_admin and not _remarks_only:
        return jsonify({'status':'error','message':'Admin only'}), 403

    conn = sampling_portal.get_db_connection()
    try:
        line = conn.execute("""
            SELECT i.*, t.status AS xfer_status, t.from_godown_id, t.to_godown_id, t.transfer_no
            FROM pm_transfer_items i
            JOIN pm_transfers t ON t.transfer_id=i.transfer_id
            WHERE i.item_id=%s AND i.transfer_id=%s
        """, (item_id, tid)).fetchone()
        if not line:
            conn.close(); return jsonify({'status':'error','message':'Line not found'}), 404

        # Remarks-only edit: just persist the note and return. No stock math.
        if _remarks_only:
            rmk = (_patch_body.get('remarks') or '').strip()[:255] or None
            conn.execute(
                "UPDATE pm_transfer_items SET remarks=%s WHERE item_id=%s",
                (rmk, item_id)
            )
            try:
                _log_transfer_edit(conn, tid, f"edit_{line['side']}_remark",
                                   f"product_id={line['product_id']} remark set")
            except Exception:
                pass
            conn.commit(); conn.close()
            return jsonify({'status':'ok'})

        old_qty  = float(line['total_qty'] or 0)
        old_side = line['side']
        prod_id  = line['product_id']

        # DELETE permission check — non-admins can only delete OUT-side lines
        # on a draft voucher. Anything else stays admin-only.
        if request.method == 'DELETE' and not is_admin:
            xfer_status = line.get('xfer_status') or ''
            if xfer_status != 'out_started':
                conn.close()
                return jsonify({
                    'status':'error',
                    'message':'Only admins can edit a voucher after stock has been posted.'
                }), 403
            if old_side != 'out':
                conn.close()
                return jsonify({
                    'status':'error',
                    'message':'Only admins can remove lines on the receiving side.'
                }), 403

        if request.method == 'DELETE':
            # ── Free the scanned boxes so they can be re-scanned ─────────
            # Symptom this fixes: after deleting a product line from a draft
            # Material Out, the boxes that had been scanned for that line
            # remain in pm_box_movements with movement_type='out' AND in
            # pm_boxes with current_status='in_transit'. The next scan
            # attempt rejects them with "Box status is 'in_transit' — must
            # be in_stock", forcing the user to abandon the voucher and
            # start over.
            #
            # Scope: only the boxes that match this line's
            # (product_id, per_box_qty) — multi-pack-size vouchers (e.g.
            # 368-pack and 416-pack of the same product, two lines) MUST
            # not free boxes from the OTHER line when only one is deleted.
            #
            # Only fires on draft (out_started) — for posted vouchers we
            # leave box state alone because stock txns are already in place
            # and reverting them would desync the ledgers.
            xfer_status_for_cleanup = line.get('xfer_status') or ''
            if xfer_status_for_cleanup == 'out_started' and old_side == 'out':
                line_pbq = float(line['per_box_qty'] or 0)
                affected = conn.execute(
                    """SELECT m.movement_id, m.box_id
                       FROM pm_box_movements m
                       JOIN pm_boxes b ON b.box_id = m.box_id
                       WHERE m.transfer_id=%s
                         AND m.movement_type='out'
                         AND b.product_id=%s
                         AND ABS(COALESCE(b.per_box_qty,0) - %s) < 0.001""",
                    (tid, prod_id, line_pbq)
                ).fetchall()
                if affected:
                    box_ids = [r['box_id'] for r in affected]
                    mv_ids  = [r['movement_id'] for r in affected]
                    # Wipe the OUT movements so the "already scanned" guard
                    # stops firing. Integers only, no user input — safe to
                    # interpolate placeholders inline.
                    ph_mv = ','.join(['%s'] * len(mv_ids))
                    conn.execute(
                        f"DELETE FROM pm_box_movements WHERE movement_id IN ({ph_mv})",
                        tuple(mv_ids)
                    )
                    # Revert in_transit → in_stock. Conditional on status
                    # so boxes already committed elsewhere stay untouched.
                    ph_b = ','.join(['%s'] * len(box_ids))
                    conn.execute(
                        f"UPDATE pm_boxes SET current_status='in_stock' "
                        f"WHERE box_id IN ({ph_b}) AND current_status='in_transit'",
                        tuple(box_ids)
                    )

            conn.execute("DELETE FROM pm_transfer_items WHERE item_id=%s", (item_id,))
            new_qty = 0.0
            _log_transfer_edit(conn, tid, f'edit_{old_side}_delete', f"product_id={prod_id}")
        else:
            d = request.get_json() or {}
            try:
                nob = int(d.get('no_of_box')   if d.get('no_of_box')   is not None else line['no_of_box'])
                pbq = float(d.get('per_box_qty') if d.get('per_box_qty') is not None else line['per_box_qty'])
            except Exception:
                conn.close(); return jsonify({'status':'error','message':'Invalid number'}), 400
            if nob < 0 or pbq < 0:
                conn.close(); return jsonify({'status':'error','message':'Values must be >= 0'}), 400
            new_qty = nob * pbq
            conn.execute(
                "UPDATE pm_transfer_items SET no_of_box=%s, per_box_qty=%s, total_qty=%s WHERE item_id=%s",
                (nob, pbq, new_qty, item_id)
            )
            _log_transfer_edit(conn, tid, f'edit_{old_side}', f"product_id={prod_id} → {nob}box × {pbq}")

        # If voucher already posted stock, write a delta correction txn
        already_posted_out = (line['xfer_status'] in ('in_pending','received','cancelled'))
        already_posted_in  = (line['xfer_status'] == 'received')
        delta_qty = new_qty - old_qty
        if abs(delta_qty) > 0.001:
            today = str(date.today())
            if old_side == 'out' and already_posted_out:
                # Source stock went down by old_qty; should have gone down by new_qty.
                # Delta correction: insert another OUT line for the difference.
                _post_stock_movement(
                    conn,
                    product_id=prod_id, godown_id=line['from_godown_id'],
                    qty=abs(delta_qty),
                    direction='out' if delta_qty > 0 else 'in',
                    transfer_no=line['transfer_no'], transfer_id=tid,
                    txn_date=today, user=_user()
                )
            elif old_side == 'in' and already_posted_in:
                _post_stock_movement(
                    conn,
                    product_id=prod_id, godown_id=line['to_godown_id'],
                    qty=abs(delta_qty),
                    direction='in' if delta_qty > 0 else 'out',
                    transfer_no=line['transfer_no'], transfer_id=tid,
                    txn_date=today, user=_user()
                )

        if old_side == 'out': _refresh_transfer_totals(conn, tid)
        _check_discrepancy(conn, tid)
        conn.commit()
        conn.close()
        return jsonify({'status':'ok'})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transfers/voucher/<int:tid>/unscan_box', methods=['POST'])
@_login_required
def api_voucher_unscan_box(tid):
    """Remove a single scanned box from a draft Material Out voucher.

    Why this endpoint exists:
      Operators sometimes scan the wrong box (similar codes, mis-pick from
      the shelf, etc.). Before this endpoint they had to delete the whole
      product line and re-scan everything — painful when 49 of 50 boxes
      were correct. This reverts just one box:

        1. DELETE the pm_box_movements 'out' row for (box, transfer)
        2. Flip pm_boxes.current_status 'in_transit' → 'in_stock'
        3. Decrement the matching pm_transfer_items line's no_of_box /
           total_qty (matching by product_id + per_box_qty within
           tolerance — same grain as the unique key on pm_transfer_items).
           If this was the last box on the line, DELETE the line.

    Body:    { box_id: int }
    Returns: { status, line_deleted: bool, remaining_boxes: int }

    Permissions: mirror api_voucher_edit_line DELETE — admins always;
    non-admins on draft (out_started) OUT-side only. Posted vouchers stay
    admin-locked because stock txns are already on the books.
    """
    d = request.get_json() or {}
    box_id = d.get('box_id')
    if not box_id:
        return jsonify({'status':'error','message':'box_id required'}), 400
    try:
        box_id = int(box_id)
    except Exception:
        return jsonify({'status':'error','message':'box_id must be an integer'}), 400

    is_admin = (session.get('User_Type','').lower() == 'admin')

    conn = sampling_portal.get_db_connection()
    try:
        # Look up the OUT movement + box state + transfer state in one query.
        # If there's no matching movement we treat it as a no-op (idempotent
        # — double-click safe).
        mv = conn.execute(
            """SELECT m.movement_id, m.box_id,
                      b.product_id, b.per_box_qty, b.current_status,
                      t.status AS xfer_status
               FROM pm_box_movements m
               JOIN pm_boxes b ON b.box_id = m.box_id
               JOIN pm_transfers t ON t.transfer_id = m.transfer_id
               WHERE m.transfer_id=%s AND m.box_id=%s AND m.movement_type='out'
               LIMIT 1""",
            (tid, box_id)
        ).fetchone()
        if not mv:
            conn.close()
            return jsonify({
                'status': 'ok',
                'line_deleted': False,
                'remaining_boxes': 0,
                'message': 'Box was not scanned for this voucher (no-op)',
            })

        xfer_status = mv.get('xfer_status') or ''
        if not is_admin and xfer_status != 'out_started':
            conn.close()
            return jsonify({
                'status':'error',
                'message':'Only admins can edit a voucher after stock has been posted.'
            }), 403

        prod_id = int(mv['product_id'])
        per_box = float(mv['per_box_qty'] or 0)

        # 1. Remove the movement
        conn.execute(
            "DELETE FROM pm_box_movements WHERE movement_id=%s",
            (mv['movement_id'],)
        )
        # 2. Revert box status (only if it's still in_transit — defensive
        # against weird interleavings where the box's already been moved on).
        conn.execute(
            "UPDATE pm_boxes SET current_status='in_stock' "
            "WHERE box_id=%s AND current_status='in_transit'",
            (box_id,)
        )
        # 3. Decrement the matching pm_transfer_items line. Match by
        # (product_id, per_box_qty within tolerance) — same grain as the
        # row uniqueness on the table.
        line = conn.execute(
            """SELECT item_id, no_of_box, total_qty, per_box_qty
               FROM pm_transfer_items
               WHERE transfer_id=%s AND side='out' AND product_id=%s
                 AND ABS(per_box_qty - %s) < 0.001
               LIMIT 1""",
            (tid, prod_id, per_box)
        ).fetchone()
        line_deleted = False
        remaining = 0
        if line:
            new_nob = max(0, int(line['no_of_box'] or 0) - 1)
            new_tot = max(0.0, float(line['total_qty'] or 0) - per_box)
            if new_nob <= 0:
                # Last box on this line — remove the ghost line entirely.
                conn.execute(
                    "DELETE FROM pm_transfer_items WHERE item_id=%s",
                    (line['item_id'],)
                )
                line_deleted = True
            else:
                conn.execute(
                    "UPDATE pm_transfer_items SET no_of_box=%s, total_qty=%s WHERE item_id=%s",
                    (new_nob, new_tot, line['item_id'])
                )
                remaining = new_nob

        _log_transfer_edit(
            conn, tid, 'unscan_out',
            f"box_id={box_id} product_id={prod_id} pbq={per_box}"
        )
        _refresh_transfer_totals(conn, tid)
        _check_discrepancy(conn, tid)
        conn.commit()
        conn.close()
        return jsonify({
            'status': 'ok',
            'line_deleted': line_deleted,
            'remaining_boxes': remaining,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transfers/voucher/<int:tid>/save_out', methods=['POST'])
@_login_required
def api_voucher_save_out(tid):
    """Finalise OUT: write source-decrement stock txns, status='in_pending'.

    Allotment auto-complete
    ───────────────────────
    For allotments — vouchers where source == destination (production
    consumption at the same location) or `voucher_type='allotment'` — the
    IN-side handshake doesn't make physical sense. The boxes are picked
    from the godown and then immediately consumed in production at the
    same site; there's nothing to "receive" later. So instead of leaving
    the voucher in 'in_pending' for a second user to scan-IN and risk
    a phantom discrepancy, we mirror the OUT lines into IN-side rows,
    post the IN movements (so the destination ledger balances), mark
    the boxes 'consumed', and jump straight to status='received'.

    Discrepancy is bypassed entirely: OUT and IN are constructed as
    exact copies, so they cannot mismatch.
    """
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute("SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Not found'}), 404
        if t['status'] != 'out_started':
            conn.close(); return jsonify({'status':'error','message':f"Already in '{t['status']}' state"}), 409
        items = conn.execute(
            "SELECT item_id, product_id, total_qty FROM pm_transfer_items WHERE transfer_id=%s AND side='out'",
            (tid,)
        ).fetchall()
        if not items:
            conn.close(); return jsonify({'status':'error','message':'No items in voucher — scan at least one box first'}), 400
        today = str(date.today())
        # 1. Post the OUT-side stock txns (source godown decrement). This
        # is the same for every transfer type — boxes leave the source.
        for r in items:
            _post_stock_movement(
                conn,
                product_id=r['product_id'], godown_id=t['from_godown_id'],
                qty=float(r['total_qty'] or 0), direction='out',
                transfer_no=t['transfer_no'], transfer_id=tid,
                txn_date=today, user=_user()
            )
        _refresh_transfer_totals(conn, tid)

        # 1b. If this transfer was created to fulfill a Material Request,
        # write link rows now so the request's qty_fulfilled is updated
        # and its status auto-progresses (pending → in_progress → fulfilled).
        # Best-effort: a failure here won't roll back the OUT itself.
        req_id_on_transfer = t.get('request_id') if hasattr(t, 'get') else (t['request_id'] if 'request_id' in t.keys() else None)
        if req_id_on_transfer:
            try:
                _mr_link_transfer_to_request(
                    conn,
                    request_id=int(req_id_on_transfer),
                    transfer_id=int(tid),
                    transfer_items=[
                        {'id': int(r['item_id']), 'product_id': int(r['product_id']), 'qty': float(r['total_qty'] or 0)}
                        for r in items
                    ],
                    fulfilled_by=_user(),
                )
            except Exception as _link_err:
                import sys as _sys
                print(f"[pm_stock] MR link failed for transfer {tid}: {_link_err}", file=_sys.stderr)

        # 2. Is this an allotment? Two ways to qualify:
        #   a) voucher_type explicitly tagged 'allotment'
        #   b) source godown == destination godown (production consumption
        #      at the same physical location — boxes go from godown to the
        #      factory floor that lives at the same site)
        # Either way → auto-complete with no IN-side handshake or
        # discrepancy check.
        #
        # Important: voucher_type may legitimately be MISSING from older
        # rows (column was added later via ALTER) — we treat missing as
        # blank, which falls through to the same-godown check.
        voucher_type_raw = t.get('voucher_type') if hasattr(t, 'get') else (t['voucher_type'] if 'voucher_type' in t.keys() else None)
        voucher_type_lc = (voucher_type_raw or '').strip().lower()
        same_godown = (int(t['from_godown_id']) == int(t['to_godown_id']))
        is_allotment = (voucher_type_lc == 'allotment' or same_godown)

        # Diagnostic log so we can debug "stuck in_pending" reports without
        # adding print statements after the fact. Visible in Flask stderr.
        import sys as _sys
        print(
            f"[pm_stock] save_out tid={tid} no={t['transfer_no']!r} "
            f"voucher_type={voucher_type_raw!r} (lc={voucher_type_lc!r}) "
            f"from={t['from_godown_id']} to={t['to_godown_id']} "
            f"same_godown={same_godown} → is_allotment={is_allotment}",
            file=_sys.stderr, flush=True
        )

        if is_allotment:
            # 2a. Mirror OUT lines into IN-side rows. Use the same product
            # + per_box_qty grain so the two sides match exactly — no way
            # for a future discrepancy check to trip on this voucher.
            out_lines = conn.execute(
                """SELECT product_id, no_of_box, per_box_qty, total_qty
                   FROM pm_transfer_items
                   WHERE transfer_id=%s AND side='out'""",
                (tid,)
            ).fetchall()
            for ol in out_lines:
                # The pm_transfer_items unique key includes per_box_qty, so
                # different pack sizes of the same product get separate
                # rows on the IN side just like on the OUT side. Use
                # INSERT ... ON DUPLICATE KEY UPDATE so re-runs (idempotency)
                # don't accumulate duplicate IN rows.
                conn.execute(
                    """INSERT INTO pm_transfer_items
                         (transfer_id, side, product_id, no_of_box, per_box_qty, total_qty)
                       VALUES (%s, 'in', %s, %s, %s, %s)
                       ON DUPLICATE KEY UPDATE
                         no_of_box = VALUES(no_of_box),
                         total_qty = VALUES(total_qty)""",
                    (tid, ol['product_id'], int(ol['no_of_box'] or 0),
                     float(ol['per_box_qty'] or 0), float(ol['total_qty'] or 0))
                )

            # 2b. Mirror OUT box-movement records into IN-side box-movement
            # records. Same set of physical boxes, just now also marked as
            # received at the destination. Use INSERT IGNORE so re-runs
            # don't duplicate.
            out_moves = conn.execute(
                """SELECT box_id FROM pm_box_movements
                   WHERE transfer_id=%s AND movement_type='out'""",
                (tid,)
            ).fetchall()
            user_now = _user()
            for mv in out_moves:
                conn.execute(
                    """INSERT INTO pm_box_movements
                         (box_id, movement_type, from_godown_id, to_godown_id,
                          qty, transfer_id, moved_by, remarks)
                       SELECT box_id, 'in', %s, %s, per_box_qty, %s, %s,
                              'allotment auto-complete'
                       FROM pm_boxes
                       WHERE box_id=%s
                       LIMIT 1""",
                    (t['from_godown_id'], t['to_godown_id'], tid, user_now, mv['box_id'])
                )

            # 2c. Post IN stock txns at destination (production consumption).
            # For allotments to the factory floor, this becomes the
            # consumption ledger entry.
            in_lines = conn.execute(
                "SELECT product_id, total_qty FROM pm_transfer_items WHERE transfer_id=%s AND side='in'",
                (tid,)
            ).fetchall()
            for r in in_lines:
                _post_stock_movement(
                    conn,
                    product_id=r['product_id'], godown_id=t['to_godown_id'],
                    qty=float(r['total_qty'] or 0), direction='in',
                    transfer_no=t['transfer_no'], transfer_id=tid,
                    txn_date=today, user=user_now
                )

            # 2d. Mark every box on this transfer as consumed. They've
            # left the source AND been received at destination AND
            # immediately consumed in production — they don't sit around
            # in_transit or in_stock anywhere. This also removes them
            # from any "available for scan" listings.
            box_id_rows = conn.execute(
                """SELECT DISTINCT box_id FROM pm_box_movements
                   WHERE transfer_id=%s AND movement_type IN ('out','in')""",
                (tid,)
            ).fetchall()
            if box_id_rows:
                ids = [int(r['box_id']) for r in box_id_rows]
                ph = ','.join(['%s'] * len(ids))
                conn.execute(
                    f"UPDATE pm_boxes SET current_status='consumed' "
                    f"WHERE box_id IN ({ph})",
                    tuple(ids)
                )

            # 2e. Settle the transfer outright. No has_discrepancy flag
            # is touched (defaults to 0) — by construction OUT and IN
            # are identical, so reconcile would always pass.
            conn.execute(
                """UPDATE pm_transfers
                   SET status='received',
                       in_at=NOW(),
                       in_by=%s,
                       has_discrepancy=0
                   WHERE transfer_id=%s""",
                (user_now, tid)
            )
            _refresh_transfer_totals(conn, tid)
            _log_transfer_edit(
                conn, tid, 'allotment_auto_complete',
                f'{len(items)} line(s) auto-settled '
                f'({"same-location" if int(t["from_godown_id"])==int(t["to_godown_id"]) else "voucher_type=allotment"})'
            )
            conn.commit()
            conn.close()
            return jsonify({
                'status': 'ok',
                'transfer_no': t['transfer_no'],
                'auto_completed': True,
                'final_status': 'received',
            })

        # Non-allotment path: standard OUT-only finalise, waits for IN.
        conn.execute("UPDATE pm_transfers SET status='in_pending' WHERE transfer_id=%s", (tid,))
        _log_transfer_edit(conn, tid, 'save_out', f'{len(items)} line(s)')
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','transfer_no':t['transfer_no']})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transfers/voucher/<int:tid>/retro_complete_allotment', methods=['POST'])
@_login_required
def api_voucher_retro_complete_allotment(tid):
    """Retro-complete a stuck allotment voucher.

    Use case: a voucher created BEFORE the auto-complete logic shipped (or
    that fell through it for any reason) is stuck in 'in_pending' with a
    permanent discrepancy banner because no IN-scan will ever arrive at a
    same-location allotment. This endpoint takes such a voucher and
    finishes the auto-complete steps that save_out would have done if the
    fix had been in place at save time:

      1. Mirror OUT lines into IN-side pm_transfer_items rows
      2. Mirror OUT box movements into IN-side movements
      3. Post the IN-side stock txns at destination
      4. Mark all the scanned boxes as 'consumed'
      5. Set transfer status='received', has_discrepancy=0

    Admin-only. Refuses to run on non-allotment vouchers (must be either
    voucher_type='allotment' or same source/dest) and on vouchers already
    in 'received' or 'cancelled' state.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403

    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute("SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Not found'}), 404

        # Permission gate: only allotment-like vouchers qualify.
        voucher_type_lc = (t.get('voucher_type') if hasattr(t,'get') else (t['voucher_type'] if 'voucher_type' in t.keys() else '')) or ''
        voucher_type_lc = voucher_type_lc.strip().lower()
        same_godown = (int(t['from_godown_id']) == int(t['to_godown_id']))
        is_allotment = (voucher_type_lc == 'allotment' or same_godown)
        if not is_allotment:
            conn.close()
            return jsonify({
                'status':'error',
                'message':f"Refusing — voucher_type={voucher_type_lc or 'NULL'} and from!=to. Not an allotment."
            }), 400

        if t['status'] not in ('in_pending', 'out_started'):
            conn.close()
            return jsonify({
                'status':'error',
                'message':f"Refusing — status is '{t['status']}'. Only in_pending or out_started can be retro-completed."
            }), 409

        today    = str(date.today())
        user_now = _user()

        # If status is out_started, also post the OUT-side stock txns first
        # (mirrors what save_out would have done).
        if t['status'] == 'out_started':
            out_items = conn.execute(
                "SELECT product_id, total_qty FROM pm_transfer_items WHERE transfer_id=%s AND side='out'",
                (tid,)
            ).fetchall()
            for r in out_items:
                _post_stock_movement(
                    conn,
                    product_id=r['product_id'], godown_id=t['from_godown_id'],
                    qty=float(r['total_qty'] or 0), direction='out',
                    transfer_no=t['transfer_no'], transfer_id=tid,
                    txn_date=today, user=user_now
                )

        # Mirror OUT lines into IN-side rows. ON DUPLICATE KEY UPDATE handles
        # the case where a partial IN had already been started.
        out_lines = conn.execute(
            """SELECT product_id, no_of_box, per_box_qty, total_qty
               FROM pm_transfer_items WHERE transfer_id=%s AND side='out'""",
            (tid,)
        ).fetchall()
        for ol in out_lines:
            conn.execute(
                """INSERT INTO pm_transfer_items
                     (transfer_id, side, product_id, no_of_box, per_box_qty, total_qty)
                   VALUES (%s, 'in', %s, %s, %s, %s)
                   ON DUPLICATE KEY UPDATE
                     no_of_box = VALUES(no_of_box),
                     total_qty = VALUES(total_qty)""",
                (tid, ol['product_id'], int(ol['no_of_box'] or 0),
                 float(ol['per_box_qty'] or 0), float(ol['total_qty'] or 0))
            )

        # Mirror OUT box movements into IN-side movements. INSERT IGNORE in
        # case some IN movements already exist (partial scan + abandon).
        out_moves = conn.execute(
            """SELECT box_id FROM pm_box_movements
               WHERE transfer_id=%s AND movement_type='out'""",
            (tid,)
        ).fetchall()
        for mv in out_moves:
            # Skip if an IN movement already exists for this (box, transfer).
            existing = conn.execute(
                """SELECT 1 FROM pm_box_movements
                   WHERE transfer_id=%s AND box_id=%s AND movement_type='in'
                   LIMIT 1""",
                (tid, mv['box_id'])
            ).fetchone()
            if existing:
                continue
            conn.execute(
                """INSERT INTO pm_box_movements
                     (box_id, movement_type, from_godown_id, to_godown_id,
                      qty, transfer_id, moved_by, remarks)
                   SELECT box_id, 'in', %s, %s, per_box_qty, %s, %s,
                          'allotment retro-complete'
                   FROM pm_boxes WHERE box_id=%s LIMIT 1""",
                (t['from_godown_id'], t['to_godown_id'], tid, user_now, mv['box_id'])
            )

        # Post IN stock txns at destination.
        in_lines = conn.execute(
            "SELECT product_id, total_qty FROM pm_transfer_items WHERE transfer_id=%s AND side='in'",
            (tid,)
        ).fetchall()
        for r in in_lines:
            _post_stock_movement(
                conn,
                product_id=r['product_id'], godown_id=t['to_godown_id'],
                qty=float(r['total_qty'] or 0), direction='in',
                transfer_no=t['transfer_no'], transfer_id=tid,
                txn_date=today, user=user_now
            )

        # Mark every box on this transfer as consumed.
        box_id_rows = conn.execute(
            """SELECT DISTINCT box_id FROM pm_box_movements
               WHERE transfer_id=%s AND movement_type IN ('out','in')""",
            (tid,)
        ).fetchall()
        if box_id_rows:
            ids = [int(r['box_id']) for r in box_id_rows]
            ph = ','.join(['%s'] * len(ids))
            conn.execute(
                f"UPDATE pm_boxes SET current_status='consumed' "
                f"WHERE box_id IN ({ph})",
                tuple(ids)
            )

        # Settle.
        conn.execute(
            """UPDATE pm_transfers
               SET status='received', in_at=NOW(), in_by=%s, has_discrepancy=0
               WHERE transfer_id=%s""",
            (user_now, tid)
        )
        _refresh_transfer_totals(conn, tid)
        _log_transfer_edit(
            conn, tid, 'allotment_retro_complete',
            f"by admin · was status='{t['status']}'"
        )
        conn.commit()
        conn.close()
        return jsonify({
            'status':'ok',
            'transfer_no': t['transfer_no'],
            'final_status': 'received',
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/voucher/<int:tid>/save_in', methods=['POST'])
@_login_required
def api_voucher_save_in(tid):
    """
    Finalise IN: write destination-increment txns. Compare OUT vs IN per product.
    If exact match → status='received', has_discrepancy=0.
    If mismatch    → status='in_pending' (stays), has_discrepancy=1.
    """
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute("SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Not found'}), 404
        if t['status'] not in ('in_pending','received'):
            conn.close(); return jsonify({'status':'error','message':f"Cannot save IN in '{t['status']}' state"}), 409

        # Separation of duties: OUT creator cannot save the IN side of their own transfer
        if _is_out_creator(conn, tid):
            conn.close()
            return jsonify({
                'status':'error',
                'message':'Separation of duties: you created the Material OUT for this transfer, so a different user must save the Material IN. You can still view or print the voucher from History.'
            }), 403

        items = conn.execute(
            "SELECT product_id, total_qty FROM pm_transfer_items WHERE transfer_id=%s AND side='in'",
            (tid,)
        ).fetchall()
        if not items:
            conn.close(); return jsonify({'status':'error','message':'No IN items — scan at least one box first'}), 400

        # Has IN already been posted? If so we need to skip re-posting.
        already_posted = (t['status'] == 'received')

        if not already_posted:
            today = str(date.today())
            for r in items:
                _post_stock_movement(
                    conn,
                    product_id=r['product_id'], godown_id=t['to_godown_id'],
                    qty=float(r['total_qty'] or 0), direction='in',
                    transfer_no=t['transfer_no'], transfer_id=tid,
                    txn_date=today, user=_user()
                )

        has_d, mismatches = _check_discrepancy(conn, tid)
        if has_d:
            # Stay in_pending so it remains in the in-transit grid
            conn.execute(
                "UPDATE pm_transfers SET in_at=NOW(), in_by=%s, status='in_pending' WHERE transfer_id=%s",
                (_user(), tid)
            )
            _log_transfer_edit(conn, tid, 'save_in_discrepancy', f'{len(mismatches)} mismatch(es)')
        else:
            conn.execute(
                "UPDATE pm_transfers SET status='received', in_at=NOW(), in_by=%s WHERE transfer_id=%s",
                (_user(), tid)
            )
            _log_transfer_edit(conn, tid, 'save_in_complete', f'{len(items)} line(s)')
        conn.commit()
        conn.close()
        return jsonify({
            'status':'ok',
            'has_discrepancy': has_d,
            'mismatches': mismatches,
            'transfer_no': t['transfer_no']
        })
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/debug_floor_dupes', methods=['GET','POST'])
@_login_required
def api_debug_floor_dupes():
    """
    DEBUG: Returns the raw rows in pm_floor_txn AND pm_godown_txn so we can
    see exactly what's stored. Optional ?vno=PMT/26-27/0025 filters to a
    specific voucher.
    """
    if (session.get('User_Type','').lower() != 'admin'):
        return jsonify({'status':'error','message':'Admin only'}), 403
    vno_filter = (request.args.get('vno') or '').strip()
    conn = sampling_portal.get_db_connection()
    try:
        # Build WHERE clauses based on filter
        where_floor  = ""
        where_godown = ""
        params_f = []
        params_g = []
        if vno_filter:
            where_floor  = " WHERE voucher_no = %s OR remarks LIKE %s"
            where_godown = " WHERE voucher_no = %s OR remarks LIKE %s"
            params_f = [vno_filter, f'%{vno_filter}%']
            params_g = [vno_filter, f'%{vno_filter}%']

        floor_rows = conn.execute(
            f"""SELECT id, product_id, godown_id, txn_date, txn_type,
                       qty, COALESCE(remarks,'') AS remarks,
                       COALESCE(voucher_no,'') AS voucher_no,
                       COALESCE(created_by,'') AS created_by
                  FROM pm_floor_txn{where_floor}
                 ORDER BY id DESC
                 LIMIT 50""",
            tuple(params_f)
        ).fetchall()

        godown_rows = conn.execute(
            f"""SELECT id, product_id, godown_id, txn_date, txn_type,
                       qty, COALESCE(remarks,'') AS remarks,
                       COALESCE(voucher_no,'') AS voucher_no,
                       COALESCE(created_by,'') AS created_by
                  FROM pm_godown_txn{where_godown}
                 ORDER BY id DESC
                 LIMIT 50""",
            tuple(params_g)
        ).fetchall()

        def cleanup(rows):
            out = []
            for r in rows:
                d = dict(r)
                if d.get('txn_date') is not None: d['txn_date'] = str(d['txn_date'])
                d['qty'] = float(d.get('qty') or 0)
                out.append(d)
            return out

        # Also list godowns so we know if godown_id 1 is floor or not
        godowns = conn.execute(
            "SELECT id, name, COALESCE(type,'') AS type FROM procurement_godowns ORDER BY id"
        ).fetchall()

        conn.close()
        return jsonify({
            'status':       'ok',
            'filter_vno':   vno_filter,
            'floor_rows':   cleanup(floor_rows),
            'godown_rows':  cleanup(godown_rows),
            'godowns':      [dict(r) for r in godowns],
            'floor_count':  len(floor_rows),
            'godown_count': len(godown_rows),
        })
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/repair_floor_in', methods=['POST'])
@_login_required
def api_repair_floor_transfer_in():
    """
    Admin-only repair for floor IN transfers.

    Two-phase clean-up:
      Phase A — DEDUPE: For every (product_id, godown_id, voucher_no, txn_type='issue')
        triple, if there are multiple rows, keep ONLY the most recent one (highest id)
        and delete the rest. This recovers from earlier double-inserts caused by
        running the old repair after a live save_in had already posted.
      Phase B — INSERT MISSING: For every received transfer to a floor destination,
        if no 'issue' row exists at all, insert the missing one (idempotent).
      Phase C — CLEAN STALE: Delete leftover bogus 'consumed' rows posted by the
        old buggy code (matched by remarks tag '[PM-MT:...] IN').

    Idempotent end-to-end — safe to run repeatedly.
    """
    if (session.get('User_Type','').lower() != 'admin'):
        return jsonify({'status':'error','message':'Admin only'}), 403

    conn = sampling_portal.get_db_connection()
    try:
        # ──────────────────────────────────────────────────────────────────
        # PHASE A — Dedupe duplicate 'issue' rows for the same transfer voucher.
        #
        # Identifies duplicates by a STABLE TAG embedded in the remarks column.
        # Both the live save_in path and the repair tool always write a remarks
        # value containing "[PM-MT:<voucher_no>]". So we extract that tag and
        # group by it — this works even if voucher_no column is NULL on older
        # rows (legacy bug) or if one row says "(repaired)" and another doesn't.
        # ──────────────────────────────────────────────────────────────────
        import re as _re
        dupes_deleted_per_vno = {}   # transfer_no → count of duplicates removed
        dupes_deleted = 0

        all_issue_rows = conn.execute("""
            SELECT id, product_id, godown_id, remarks, voucher_no
              FROM pm_floor_txn
             WHERE txn_type='issue'
               AND remarks LIKE %s
        """, ('%PM-MT:%',)).fetchall()

        TAG_RE = _re.compile(r'PM-MT:([^\]\s]+)')
        groups = {}
        for r in all_issue_rows:
            m = TAG_RE.search(r['remarks'] or '')
            if not m:
                continue
            vno_tag = m.group(1).strip().rstrip(']').rstrip()
            key = (r['product_id'], r['godown_id'], vno_tag)
            groups.setdefault(key, []).append(r['id'])

        for key, ids in groups.items():
            if len(ids) <= 1:
                continue
            ids_sorted = sorted(ids, reverse=True)   # highest id first
            for stale_id in ids_sorted[1:]:
                conn.execute("DELETE FROM pm_floor_txn WHERE id=%s", (stale_id,))
                dupes_deleted += 1
                vno = key[2]
                dupes_deleted_per_vno[vno] = dupes_deleted_per_vno.get(vno, 0) + 1

        # ──────────────────────────────────────────────────────────────────
        # PHASE A2 — Cross-table duplicates between pm_floor_txn and pm_godown_txn.
        #
        # Buggy historical case: same transfer voucher posted an 'inward' to
        # pm_godown_txn (because at the time the destination's type was wrongly
        # 'godown', not 'floor') AND an 'issue' to pm_floor_txn (from a later
        # repair run). Same voucher, same product, same godown → counted twice
        # in the unified ledger.
        #
        # Resolution: if the destination IS a floor location (per the fixed
        # _is_floor_godown helper), the floor 'issue' row is the correct one
        # to keep — delete the matching pm_godown_txn 'inward' row.
        # ──────────────────────────────────────────────────────────────────
        cross_deleted = 0
        cross_deleted_per_vno = {}
        godown_inward_rows = conn.execute("""
            SELECT id, product_id, godown_id, remarks
              FROM pm_godown_txn
             WHERE txn_type='inward'
               AND remarks LIKE %s
        """, ('%PM-MT:%',)).fetchall()
        for gr in godown_inward_rows:
            m = TAG_RE.search(gr['remarks'] or '')
            if not m:
                continue
            vno_tag = m.group(1).strip().rstrip(']').rstrip()
            # Is this godown actually a floor (per the corrected detection)?
            if not _is_floor_godown(conn, gr['godown_id']):
                continue   # genuinely a godown destination — keep the inward row
            # Floor side has the correct row already?
            floor_match = conn.execute(
                "SELECT id FROM pm_floor_txn "
                "WHERE product_id=%s AND godown_id=%s "
                "AND txn_type='issue' AND remarks LIKE %s "
                "LIMIT 1",
                (gr['product_id'], gr['godown_id'], f'%PM-MT:{vno_tag}%')
            ).fetchone()
            if not floor_match:
                # No matching floor row — MOVE this row over instead of deleting,
                # so we don't lose stock data
                conn.execute(
                    """INSERT INTO pm_floor_txn (product_id, txn_date, txn_type, qty, remarks, created_by, godown_id, voucher_no)
                       SELECT product_id, txn_date, 'issue', qty, remarks, created_by, godown_id, voucher_no
                         FROM pm_godown_txn WHERE id=%s""",
                    (gr['id'],)
                )
            # Now delete the wrong-table row
            conn.execute("DELETE FROM pm_godown_txn WHERE id=%s", (gr['id'],))
            cross_deleted += 1
            cross_deleted_per_vno[vno_tag] = cross_deleted_per_vno.get(vno_tag, 0) + 1
        # Roll cross-deletes into the same dupes_deleted counter for reporting
        dupes_deleted += cross_deleted
        for vno, n in cross_deleted_per_vno.items():
            dupes_deleted_per_vno[vno] = dupes_deleted_per_vno.get(vno, 0) + n

        # ──────────────────────────────────────────────────────────────────
        # PHASE A3 — Remove stray transfer rows that don't belong.
        #
        # For a correct transfer, exactly TWO stock rows should exist:
        #   - source_godown:  outward (or 'dispatch' if source is a floor)
        #   - destination:    inward  (or 'issue'    if destination is a floor)
        #
        # In some recorded data we find FOUR rows per transfer:
        #   - source.outward         ✅ correct
        #   - source.inward          🚨 stray (likely from a buggy mirror-write)
        #   - destination.outward    🚨 stray
        #   - destination.<inward/issue>  ✅ correct
        #
        # This phase walks every transfer and deletes any row that doesn't
        # match the canonical (source.outward / destination.inward) pattern.
        # ──────────────────────────────────────────────────────────────────
        misplaced_deleted = 0
        misplaced_per_vno = {}
        all_xfer_rows = conn.execute("""
            SELECT 'godown' AS src, id, product_id, godown_id, txn_type, remarks
              FROM pm_godown_txn
             WHERE remarks LIKE %s
            UNION ALL
            SELECT 'floor' AS src, id, product_id, godown_id, txn_type, remarks
              FROM pm_floor_txn
             WHERE remarks LIKE %s
        """, ('%PM-MT:%', '%PM-MT:%')).fetchall()

        # Cache transfer header lookups by voucher number
        xfer_cache = {}
        def _get_xfer(vno):
            if vno in xfer_cache: return xfer_cache[vno]
            row = conn.execute(
                "SELECT from_godown_id, to_godown_id FROM pm_transfers WHERE transfer_no=%s",
                (vno,)
            ).fetchone()
            xfer_cache[vno] = row
            return row

        for r in all_xfer_rows:
            m = TAG_RE.search(r['remarks'] or '')
            if not m:
                continue
            vno_tag = m.group(1).strip().rstrip(']').rstrip()
            t = _get_xfer(vno_tag)
            if not t:
                continue
            from_id = int(t['from_godown_id'] or 0)
            to_id   = int(t['to_godown_id']   or 0)
            this_id = int(r['godown_id'])
            txn_t   = (r['txn_type'] or '').lower()

            # Decide: is this row valid?
            valid = False
            if this_id == from_id:
                # Source side: only 'outward' or 'dispatch' is valid
                valid = txn_t in ('outward', 'dispatch')
            elif this_id == to_id:
                # Destination side: only 'inward' or 'issue' is valid
                valid = txn_t in ('inward', 'issue')
            else:
                # Row at a godown that isn't either side of the transfer
                # — orphaned, also stray
                valid = False

            if valid:
                continue

            # Stray — delete from the appropriate table
            tbl = 'pm_godown_txn' if r['src'] == 'godown' else 'pm_floor_txn'
            conn.execute(f"DELETE FROM {tbl} WHERE id=%s", (r['id'],))
            misplaced_deleted += 1
            misplaced_per_vno[vno_tag] = misplaced_per_vno.get(vno_tag, 0) + 1

        dupes_deleted += misplaced_deleted
        for vno, n in misplaced_per_vno.items():
            dupes_deleted_per_vno[vno] = dupes_deleted_per_vno.get(vno, 0) + n

        # ── Find all 'received' transfers whose destination is a floor godown ──
        rows = conn.execute("""
            SELECT t.transfer_id, t.transfer_no, t.to_godown_id, t.in_at
              FROM pm_transfers t
              JOIN procurement_godowns g ON g.id = t.to_godown_id
             WHERE t.status = 'received'
               AND (COALESCE(g.type,'') = 'floor'
                    OR LOWER(g.name) IN ('factory','factory floor','production floor'))
        """).fetchall()

        inserted_total = 0
        deleted_bogus  = 0
        fixed_vouchers = []
        for r in rows:
            tid     = r['transfer_id']
            vno     = r['transfer_no']
            gid     = r['to_godown_id']
            raw_in  = r['in_at']
            if raw_in is None:
                in_date = str(date.today())
            elif hasattr(raw_in, 'date'):
                in_date = str(raw_in.date())
            else:
                in_date = str(raw_in)[:10] or str(date.today())

            items = conn.execute(
                "SELECT product_id, total_qty FROM pm_transfer_items "
                "WHERE transfer_id=%s AND side='in'",
                (tid,)
            ).fetchall()
            if not items:
                continue

            v_inserted = 0
            v_deleted  = 0
            for it in items:
                pid = it['product_id']
                qty = float(it['total_qty'] or 0)
                if qty <= 0:
                    continue

                # ── PHASE B — Already a valid 'issue' row for this transfer? ──
                existing = conn.execute(
                    "SELECT id FROM pm_floor_txn "
                    "WHERE product_id=%s AND godown_id=%s "
                    "AND txn_type='issue' "
                    "AND voucher_no=%s "
                    "LIMIT 1",
                    (pid, gid, vno)
                ).fetchone()
                if not existing:
                    # No row at all — insert the missing inflow row
                    conn.execute(
                        """INSERT INTO pm_floor_txn (product_id, txn_date, txn_type, qty, remarks, created_by, godown_id, voucher_no)
                           VALUES (%s,%s,'issue',%s,%s,%s,%s,%s)""",
                        (pid, in_date, qty,
                         f"[PM-MT:{vno}] IN (repaired)", 'system_repair', gid, vno)
                    )
                    v_inserted += 1
                    inserted_total += 1

                # ── PHASE C — Clean up bogus 'consumed' rows from the old buggy code ──
                bogus = conn.execute(
                    "SELECT id FROM pm_floor_txn "
                    "WHERE product_id=%s AND godown_id=%s "
                    "AND txn_type='consumed' "
                    "AND remarks LIKE %s",
                    (pid, gid, f"[PM-MT:{vno}]%")
                ).fetchall()
                for b in bogus:
                    conn.execute("DELETE FROM pm_floor_txn WHERE id=%s", (b['id'],))
                    v_deleted += 1
                    deleted_bogus += 1

            # Roll Phase A dedupe results into the per-voucher detail
            phase_a_deleted = dupes_deleted_per_vno.get(vno, 0)

            if v_inserted or v_deleted or phase_a_deleted:
                fixed_vouchers.append({
                    'transfer_no':   vno,
                    'inserted':      v_inserted,
                    'deleted_bogus': v_deleted,
                    'duplicates':    phase_a_deleted
                })

        conn.commit()
        conn.close()
        return jsonify({
            'status':              'ok',
            'transfers_scanned':   len(rows),
            'transfers_fixed':     len(fixed_vouchers),
            'rows_inserted':       inserted_total,
            'rows_deleted_bogus':  deleted_bogus,
            'duplicates_deleted':  dupes_deleted,
            # Per-phase breakdown for debugging
            'phase_a_rows_seen':   len(all_issue_rows),
            'phase_a2_rows_seen':  len(godown_inward_rows),
            'phase_a2_deleted':    cross_deleted,
            'phase_a3_rows_seen':  len(all_xfer_rows),
            'phase_a3_deleted':    misplaced_deleted,
            'detail':              fixed_vouchers
        })
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/voucher/<int:tid>/reconcile/preview', methods=['GET'])
@_login_required
def api_voucher_reconcile_preview(tid):
    """
    Returns a per-product impact preview for both possible settlements.
    Used by the reconcile modal to show admin exactly what will happen
    BEFORE they commit. Read-only — no stock changes here.

    Response shape:
    {
      status: 'ok',
      transfer_no: 'PMT/...',
      from_godown: {id, name},
      to_godown:   {id, name},
      mismatches: [
        {
          product_id, product_name, pm_type,
          out_qty, in_qty, qty_delta,   // delta = in - out
          settle_out_action: {
            // What happens if admin picks OUT
            target_godown: 'to'|'from',
            target_godown_name: '...',
            direction: 'in'|'out',     // stock movement direction
            qty: <abs delta>,
            description: 'Credit destination 2400 — fills missing diff'
          },
          settle_in_action: {  /* same shape for IN side */ }
        }, ...
      ]
    }
    """
    if (session.get('User_Type','').lower() != 'admin'):
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute("SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Not found'}), 404
        if not t['has_discrepancy']:
            conn.close(); return jsonify({'status':'ok','mismatches':[]})

        from_g = conn.execute("SELECT id,name FROM procurement_godowns WHERE id=%s",
                              (t['from_godown_id'],)).fetchone()
        to_g   = conn.execute("SELECT id,name FROM procurement_godowns WHERE id=%s",
                              (t['to_godown_id'],)).fetchone()
        from_name = (from_g['name'] if from_g else '') or f"#{t['from_godown_id']}"
        to_name   = (to_g['name']   if to_g   else '') or f"#{t['to_godown_id']}"

        _, mismatches = _check_discrepancy(conn, tid)

        # Enrich each mismatch with what each settlement choice will do.
        out_mismatches = []
        for m in mismatches:
            qd = float(m['qty_delta'] or 0)   # in_qty - out_qty
            absqd = abs(qd)
            pinfo = conn.execute(
                "SELECT product_name, pm_type FROM pm_products WHERE id=%s",
                (m['product_id'],)
            ).fetchone()
            pname = pinfo['product_name'] if pinfo else f"#{m['product_id']}"
            ptype = pinfo['pm_type']      if pinfo else ''

            # Settle-to-OUT: destination gets adjusted to match OUT qty.
            if qd < 0:
                # Less arrived. To match OUT, credit destination with the missing.
                out_act = {
                    'target_godown':      'to',
                    'target_godown_name': to_name,
                    'direction':          'in',
                    'qty':                absqd,
                    'description':        f"Credit {to_name} +{absqd:g} (less arrived than was sent)",
                }
            elif qd > 0:
                # More arrived. To match OUT, debit dest of the extras.
                out_act = {
                    'target_godown':      'to',
                    'target_godown_name': to_name,
                    'direction':          'out',
                    'qty':                absqd,
                    'description':        f"Debit {to_name} -{absqd:g} (extra boxes arrived; remove from dest)",
                }
            else:
                out_act = {'qty': 0, 'description': 'No adjustment'}

            # Settle-to-IN: source gets adjusted to match IN qty.
            if qd < 0:
                # Less arrived. The missing portion never shipped — refund source.
                in_act = {
                    'target_godown':      'from',
                    'target_godown_name': from_name,
                    'direction':          'in',
                    'qty':                absqd,
                    'description':        f"Refund {from_name} +{absqd:g} (treat shortfall as never-shipped)",
                }
            elif qd > 0:
                # More arrived. The extras must have come from source — debit it.
                in_act = {
                    'target_godown':      'from',
                    'target_godown_name': from_name,
                    'direction':          'out',
                    'qty':                absqd,
                    'description':        f"Debit {from_name} -{absqd:g} (extras came from source)",
                }
            else:
                in_act = {'qty': 0, 'description': 'No adjustment'}

            out_mismatches.append({
                'product_id':         m['product_id'],
                'product_name':       pname,
                'pm_type':            ptype,
                'out_qty':            float(m.get('out_qty') or 0),
                'in_qty':             float(m.get('in_qty')  or 0),
                'qty_delta':          qd,
                'settle_out_action':  out_act,
                'settle_in_action':   in_act,
            })

        conn.close()
        return jsonify({
            'status':       'ok',
            'transfer_no':  t['transfer_no'],
            'from_godown':  {'id': t['from_godown_id'], 'name': from_name},
            'to_godown':    {'id': t['to_godown_id'],   'name': to_name},
            'mismatches':   out_mismatches,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/transfers/voucher/<int:tid>/reconcile', methods=['POST'])
@_login_required
def api_voucher_reconcile(tid):
    """
    Force-reconcile a discrepancy. Admin only. Writes any needed delta stock
    txns so the books balance, rewrites pm_transfer_items so both sides match,
    sets status='received', clears the red flag.

    Body:
      { side: 'out' | 'in', note: '...' }

    Settle-to-OUT: destination is adjusted to match OUT. Source unchanged.
      • qty_delta < 0 (less arrived): +delta to dest (credit, fills missing)
      • qty_delta > 0 (extras):       -delta from dest (debit, removes extras)
      • IN items rewritten to equal OUT items.

    Settle-to-IN: source is adjusted to match IN. Destination unchanged.
      • qty_delta < 0 (less arrived): +delta to source (refund never-shipped)
      • qty_delta > 0 (extras):       -delta from source (extras came from source)
      • OUT items rewritten to equal IN items.

    Both modes also flush any boxes still 'in_transit' to the destination
    godown (sequenced AFTER the ledger postings) so box-level state matches
    the chosen settlement direction.
    """
    if (session.get('User_Type','').lower() != 'admin'):
        return jsonify({'status':'error','message':'Admin only'}), 403
    d = request.get_json(silent=True) or {}
    side = (d.get('side') or '').strip().lower()
    note = (d.get('note') or '').strip()
    if side not in ('out', 'in'):
        return jsonify({'status':'error',
            'message':"side must be 'out' or 'in'"}), 400
    if not note:
        return jsonify({'status':'error',
            'message':'Reconciliation note required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        t = conn.execute("SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)).fetchone()
        if not t:
            conn.close(); return jsonify({'status':'error','message':'Not found'}), 404
        if not t['has_discrepancy']:
            conn.close(); return jsonify({'status':'error','message':'No discrepancy to reconcile'}), 409

        # Snapshot before-state for audit trail.
        _, mismatches = _check_discrepancy(conn, tid)
        items_before = conn.execute(
            """SELECT side, product_id, no_of_box, per_box_qty, total_qty
               FROM pm_transfer_items WHERE transfer_id=%s ORDER BY side, item_id""",
            (tid,)
        ).fetchall()
        before_snapshot = [dict(r) for r in items_before]

        today    = str(date.today())
        postings = []   # log of each stock movement, for audit

        for m in mismatches:
            qd = float(m['qty_delta'] or 0)   # in - out
            if abs(qd) <= 0.001: continue
            absqd = abs(qd)
            pid   = m['product_id']

            if side == 'out':
                # Adjust DESTINATION to match OUT.
                if qd < 0:
                    # Less arrived; credit dest with missing diff.
                    _post_stock_movement(
                        conn,
                        product_id=pid, godown_id=t['to_godown_id'],
                        qty=absqd, direction='in',
                        transfer_no=t['transfer_no'], transfer_id=tid,
                        txn_date=today, user=_user()
                    )
                    postings.append({'side':'out','product_id':pid,
                        'godown':'to','direction':'in','qty':absqd})
                else:
                    # Extras at dest; debit them off (they shouldn't have arrived).
                    _post_stock_movement(
                        conn,
                        product_id=pid, godown_id=t['to_godown_id'],
                        qty=absqd, direction='out',
                        transfer_no=t['transfer_no'], transfer_id=tid,
                        txn_date=today, user=_user()
                    )
                    postings.append({'side':'out','product_id':pid,
                        'godown':'to','direction':'out','qty':absqd})

            else:  # side == 'in'
                # Adjust SOURCE to match IN.
                if qd < 0:
                    # Less arrived; refund source for the never-shipped portion.
                    _post_stock_movement(
                        conn,
                        product_id=pid, godown_id=t['from_godown_id'],
                        qty=absqd, direction='in',
                        transfer_no=t['transfer_no'], transfer_id=tid,
                        txn_date=today, user=_user()
                    )
                    postings.append({'side':'in','product_id':pid,
                        'godown':'from','direction':'in','qty':absqd})
                else:
                    # Extras at dest came from source — debit source.
                    _post_stock_movement(
                        conn,
                        product_id=pid, godown_id=t['from_godown_id'],
                        qty=absqd, direction='out',
                        transfer_no=t['transfer_no'], transfer_id=tid,
                        txn_date=today, user=_user()
                    )
                    postings.append({'side':'in','product_id':pid,
                        'godown':'from','direction':'out','qty':absqd})

        # ── Rewrite pm_transfer_items so both sides match the chosen reference ─
        # After settle-to-OUT: IN items become identical to OUT items.
        # After settle-to-IN:  OUT items become identical to IN items.
        # This makes future voucher prints show a clean, consistent state.
        if side == 'out':
            ref_rows = conn.execute(
                """SELECT product_id, no_of_box, per_box_qty, total_qty
                   FROM pm_transfer_items
                   WHERE transfer_id=%s AND side='out'""",
                (tid,)
            ).fetchall()
            conn.execute(
                "DELETE FROM pm_transfer_items WHERE transfer_id=%s AND side='in'",
                (tid,)
            )
            for r in ref_rows:
                conn.execute(
                    """INSERT INTO pm_transfer_items
                         (transfer_id, side, product_id, no_of_box, per_box_qty, total_qty)
                       VALUES (%s,'in',%s,%s,%s,%s)""",
                    (tid, r['product_id'], r['no_of_box'],
                     r['per_box_qty'], r['total_qty'])
                )
        else:
            ref_rows = conn.execute(
                """SELECT product_id, no_of_box, per_box_qty, total_qty
                   FROM pm_transfer_items
                   WHERE transfer_id=%s AND side='in'""",
                (tid,)
            ).fetchall()
            conn.execute(
                "DELETE FROM pm_transfer_items WHERE transfer_id=%s AND side='out'",
                (tid,)
            )
            for r in ref_rows:
                conn.execute(
                    """INSERT INTO pm_transfer_items
                         (transfer_id, side, product_id, no_of_box, per_box_qty, total_qty)
                       VALUES (%s,'out',%s,%s,%s,%s)""",
                    (tid, r['product_id'], r['no_of_box'],
                     r['per_box_qty'], r['total_qty'])
                )

        # Push every still-in_transit box for this transfer over to
        # destination so the box-level location matches the ledger we
        # just balanced. Without this, unscanned boxes would stay in
        # 'in_transit' status forever, blocking any future move-out.
        # (Same behaviour for both sides — the boxes physically went
        # to dest in either interpretation.)
        in_transit_boxes = conn.execute(
            """SELECT b.box_id, b.per_box_qty
               FROM pm_box_movements m
               JOIN pm_boxes b ON b.box_id = m.box_id
               WHERE m.transfer_id=%s AND m.movement_type='out'
                 AND b.current_status='in_transit'
                 AND NOT EXISTS (
                   SELECT 1 FROM pm_box_movements mi
                   WHERE mi.box_id=m.box_id AND mi.transfer_id=%s
                         AND mi.movement_type='in'
                 )""",
            (tid, tid)
        ).fetchall()
        for ib in in_transit_boxes:
            conn.execute(
                "UPDATE pm_boxes SET current_status='in_stock', current_godown_id=%s WHERE box_id=%s",
                (t['to_godown_id'], ib['box_id'])
            )
            conn.execute(
                """INSERT INTO pm_box_movements
                     (box_id, transfer_id, movement_type, from_godown_id, to_godown_id, qty, moved_by, remarks)
                   VALUES (%s,%s,'in',%s,%s,%s,%s,%s)""",
                (ib['box_id'], tid, t['from_godown_id'], t['to_godown_id'],
                 float(ib['per_box_qty'] or 0), _user(),
                 f"Reconcile auto-IN: voucher {t['transfer_no']} (settled to {side.upper()})")
            )

        conn.execute(
            """UPDATE pm_transfers
               SET status='received', has_discrepancy=0,
                   discrepancy_note=CONCAT('RECONCILED to ', %s,
                       ': ', %s, ' [was: ', COALESCE(discrepancy_note,''), ']')
               WHERE transfer_id=%s""",
            (side.upper(), note, tid)
        )
        try:
            _refresh_transfer_totals(conn, tid)
        except Exception:
            pass
        _log_transfer_edit(conn, tid, f'reconcile_{side}', note[:200])

        # Audit log
        try:
            _audit_record(
                conn,
                action='transfer.reconcile',
                entity='pm_transfers',
                entity_id=str(tid),
                summary=f"Reconciled {t['transfer_no']} settled to {side.upper()} side",
                before={
                    'transfer_no': t['transfer_no'],
                    'mismatches': [
                        {'product_id': m['product_id'],
                         'out_qty':    float(m.get('out_qty') or 0),
                         'in_qty':     float(m.get('in_qty')  or 0),
                         'qty_delta':  float(m['qty_delta'] or 0)}
                        for m in mismatches
                    ],
                    'items': before_snapshot,
                    'has_discrepancy': True,
                },
                after={
                    'side':     side,
                    'note':     note,
                    'postings': postings,
                    'has_discrepancy': False,
                },
                reversal_class=AUDIT_REVERSAL_FINAL,    # not auto-reversible
            )
        except Exception:
            pass

        conn.commit()
        conn.close()
        return jsonify({'status':'ok',
            'side': side,
            'postings_count': len(postings)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/in_transit', methods=['GET'])
@_login_required
def api_voucher_in_transit():
    """Return all transfers in 'in_pending' state for the in-transit grid.

    Separation of duties: non-admin users do NOT see vouchers they themselves
    created the OUT for. They cannot perform IN actions on those, so showing
    them in the IN-transit list would be misleading. Admins see everything.
    """
    conn = sampling_portal.get_db_connection()
    try:
        where, params = ["t.status='in_pending'"], []
        if not _is_admin():
            me = (_user() or '').strip()
            if me and me != 'Unknown':
                where.append("(LOWER(COALESCE(t.out_by,'')) <> LOWER(%s))")
                params.append(me)
        rows = conn.execute(f"""
            SELECT t.transfer_id, t.transfer_no, t.from_godown_id, t.to_godown_id,
                   t.out_at, t.out_by, t.in_at, t.in_by,
                   t.total_boxes, t.total_qty, t.has_discrepancy, t.discrepancy_note,
                   COALESCE(fg.name,'') AS from_name,
                   COALESCE(tg.name,'') AS to_name
            FROM pm_transfers t
            LEFT JOIN procurement_godowns fg ON fg.id=t.from_godown_id
            LEFT JOIN procurement_godowns tg ON tg.id=t.to_godown_id
            WHERE {' AND '.join(where)}
            ORDER BY t.has_discrepancy DESC, t.out_at DESC
        """, params).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            for k in ('out_at','in_at'):
                if d.get(k) is not None: d[k] = str(d[k])
            d['total_qty']       = float(d.get('total_qty') or 0)
            d['has_discrepancy'] = bool(d.get('has_discrepancy'))
            out.append(d)
        return jsonify({'status':'ok','transfers':out, 'count':len(out)})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/voucher/items', methods=['GET'])
@_login_required
def api_voucher_items():
    """Return line items for any voucher type, formatted for inline display.

    Query params:
      type = 'grn' | 'mt' | 'dn'
      id   = numeric id of the voucher

    Response: {
      status: 'ok',
      type:   'grn'|'mt'|'dn',
      id:     int,
      voucher_no: str,
      from:   str (label),
      to:     str (label),
      items:  [
        { product_id, product_code, product_name, pm_type, brand_name,
          no_of_box, per_box_qty, total_qty,
          side (only for 'mt': 'out'|'in') }
      ]
    }
    """
    vtype = (request.args.get('type') or '').strip().lower()
    vid_raw = request.args.get('id')
    if vtype not in ('grn', 'mt', 'dn'):
        return jsonify({'status': 'error', 'message': "type must be grn|mt|dn"}), 400
    try:
        vid = int(vid_raw)
    except Exception:
        return jsonify({'status': 'error', 'message': 'id required (integer)'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        if vtype == 'grn':
            head = conn.execute("""
                SELECT g.grn_no, g.supplier, COALESCE(gd.name,'') AS godown_name
                FROM pm_grn g
                LEFT JOIN procurement_godowns gd ON gd.id = g.godown_id
                WHERE g.id=%s
            """, (vid,)).fetchone()
            if not head:
                conn.close(); return jsonify({'status':'error','message':'GRN not found'}), 404
            items = conn.execute("""
                SELECT i.product_id,
                       COALESCE(p.product_code,'') AS product_code,
                       p.product_name,
                       COALESCE(p.pm_type,'')      AS pm_type,
                       COALESCE(b.name,'')         AS brand_name,
                       COALESCE(i.no_of_box,1)     AS no_of_box,
                       CASE
                         WHEN COALESCE(i.no_of_box,0) > 0
                           THEN COALESCE(i.qty_received,0) / i.no_of_box
                         ELSE COALESCE(i.qty_received,0)
                       END AS per_box_qty,
                       COALESCE(i.qty_received,0)  AS total_qty
                FROM pm_grn_items i
                JOIN pm_products p ON p.id = i.product_id
                LEFT JOIN procurement_brands b ON b.id = p.brand_id
                WHERE i.grn_id=%s
                ORDER BY i.id
            """, (vid,)).fetchall()
            conn.close()
            return jsonify({
                'status': 'ok', 'type': 'grn', 'id': vid,
                'voucher_no': head['grn_no'],
                'from': head['supplier'] or '—',
                'to':   head['godown_name'] or '—',
                'items': [dict(i) for i in items]
            })

        if vtype == 'mt':
            head = conn.execute("""
                SELECT t.transfer_no,
                       COALESCE(fg.name,'') AS from_name,
                       COALESCE(tg.name,'') AS to_name
                FROM pm_transfers t
                LEFT JOIN procurement_godowns fg ON fg.id=t.from_godown_id
                LEFT JOIN procurement_godowns tg ON tg.id=t.to_godown_id
                WHERE t.transfer_id=%s
            """, (vid,)).fetchone()
            if not head:
                conn.close(); return jsonify({'status':'error','message':'Transfer not found'}), 404
            items = conn.execute("""
                SELECT i.product_id, i.side,
                       COALESCE(p.product_code,'') AS product_code,
                       p.product_name,
                       COALESCE(p.pm_type,'')   AS pm_type,
                       COALESCE(b.name,'')      AS brand_name,
                       COALESCE(i.no_of_box,1)  AS no_of_box,
                       COALESCE(i.per_box_qty,0) AS per_box_qty,
                       COALESCE(i.total_qty, 0)  AS total_qty
                FROM pm_transfer_items i
                JOIN pm_products p ON p.id = i.product_id
                LEFT JOIN procurement_brands b ON b.id = p.brand_id
                WHERE i.transfer_id=%s
                ORDER BY i.side, i.item_id
            """, (vid,)).fetchall()
            conn.close()
            return jsonify({
                'status': 'ok', 'type': 'mt', 'id': vid,
                'voucher_no': head['transfer_no'],
                'from': head['from_name'] or '—',
                'to':   head['to_name']   or '—',
                'items': [dict(i) for i in items]
            })

        # vtype == 'dn'
        head = conn.execute("""
            SELECT d.dn_no, d.supplier, COALESCE(gd.name,'') AS godown_name
            FROM pm_dn d
            LEFT JOIN procurement_godowns gd ON gd.id = d.godown_id
            WHERE d.id=%s
        """, (vid,)).fetchone()
        if not head:
            conn.close(); return jsonify({'status':'error','message':'DN not found'}), 404
        items = conn.execute("""
            SELECT i.product_id,
                   COALESCE(p.product_code,'') AS product_code,
                   p.product_name,
                   COALESCE(p.pm_type,'')      AS pm_type,
                   COALESCE(b.name,'')         AS brand_name,
                   COALESCE(i.no_of_box,1)     AS no_of_box,
                   CASE
                     WHEN COALESCE(i.no_of_box,0) > 0
                       THEN COALESCE(i.qty_delivered,0) / i.no_of_box
                     ELSE COALESCE(i.qty_delivered,0)
                   END AS per_box_qty,
                   COALESCE(i.qty_delivered,0) AS total_qty
            FROM pm_dn_items i
            JOIN pm_products p ON p.id = i.product_id
            LEFT JOIN procurement_brands b ON b.id = p.brand_id
            WHERE i.dn_id=%s
            ORDER BY i.id
        """, (vid,)).fetchall()
        conn.close()
        return jsonify({
            'status': 'ok', 'type': 'dn', 'id': vid,
            'voucher_no': head['dn_no'],
            'from': head['godown_name'] or '—',
            'to':   head['supplier']    or '—',
            'items': [dict(i) for i in items]
        })

    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/discrepancies', methods=['GET'])
@_login_required
def api_voucher_discrepancies():
    """Return count + list of currently-flagged discrepancies for the
    sticky red banner.

    Visibility rules (per user request):
      - Admins see ALL flagged discrepancies (regardless of location)
      - Non-admin users only see discrepancies on transfers whose
        DESTINATION godown matches their home godown — source-side users at
        a different godown should not be alarmed by an issue that exists
        downstream from them
      - Only transfers with status='received' are returned; a transfer
        in-transit (status='in_pending') has no IN scan yet, so the
        OUT-vs-IN comparison has not actually been made and no real
        discrepancy can exist
    """
    conn = sampling_portal.get_db_connection()
    try:
        is_admin = _is_admin()
        home_gid = None if is_admin else _user_home_godown(conn)

        where = ["t.has_discrepancy=1", "t.status='received'"]
        params = []

        # Non-admin filter: only show discrepancies on transfers TO the user's
        # home godown. If a non-admin user has no home godown set we fall back
        # to showing none — they have no business location of responsibility,
        # so the banner shouldn't fire for them.
        if not is_admin:
            if home_gid is None:
                conn.close()
                return jsonify({'status': 'ok', 'count': 0, 'discrepancies': []})
            where.append("t.to_godown_id = %s")
            params.append(home_gid)

        rows = conn.execute(f"""
            SELECT t.transfer_id, t.transfer_no, t.discrepancy_note,
                   t.from_godown_id, t.to_godown_id,
                   COALESCE(fg.name,'') AS from_name,
                   COALESCE(tg.name,'') AS to_name
            FROM pm_transfers t
            LEFT JOIN procurement_godowns fg ON fg.id=t.from_godown_id
            LEFT JOIN procurement_godowns tg ON tg.id=t.to_godown_id
            WHERE {' AND '.join(where)}
            ORDER BY t.transfer_id DESC
        """, params).fetchall()
        conn.close()
        return jsonify({
            'status':'ok',
            'count': len(rows),
            'discrepancies': [dict(r) for r in rows]
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/list', methods=['GET'])
@_login_required
def api_transfer_list():
    """All transfers (history). Optional filters: status, from_date, to_date, search.

    Search scope
    ────────────
    By default `search=foo` matches against transfer_no, from/to godown
    name, and remarks (cheap LIKE-on-the-header).

    When `deep=1` is set, the search ALSO matches against:
      - product names + product codes on any OUT line of the transfer
      - box_code / short_code of any box ever moved by this transfer
    This is heavier (joins to pm_transfer_items, pm_products, pm_box_movements,
    pm_boxes) so the frontend only triggers it on a debounced/Enter event,
    not per-keystroke.

    The match is a UNION of header-level and deep matches, deduped by
    transfer_id, ordered by recency. The LIMIT 500 still applies.
    """
    status   = (request.args.get('status')    or '').strip()
    q_from   = (request.args.get('from_date') or '').strip()
    q_to     = (request.args.get('to_date')   or '').strip()
    q_search = (request.args.get('search')    or '').strip()
    deep     = (request.args.get('deep') == '1')
    conn = sampling_portal.get_db_connection()
    try:
        where, params = [], []
        if status:
            where.append("t.status=%s"); params.append(status)
        if q_from:
            where.append("DATE(t.out_at) >= %s"); params.append(q_from)
        if q_to:
            where.append("DATE(t.out_at) <= %s"); params.append(q_to)

        # ── Search clause ────────────────────────────────────────────
        # Light search (default): header-only LIKE — fast, no joins.
        # Deep search: header LIKE OR a sub-EXISTS against transfer items
        # (matching product_name/product_code) OR a sub-EXISTS against box
        # movements (matching box_code/short_code). EXISTS short-circuits
        # so we don't blow up the result set with line-level rows.
        if q_search:
            sp = f"%{q_search}%"
            if deep:
                where.append("""(
                    t.transfer_no LIKE %s OR fg.name LIKE %s OR tg.name LIKE %s OR t.remarks LIKE %s
                    OR EXISTS (
                        SELECT 1 FROM pm_transfer_items ti
                        JOIN pm_products p ON p.id = ti.product_id
                        WHERE ti.transfer_id = t.transfer_id
                          AND (p.product_name LIKE %s OR p.product_code LIKE %s)
                    )
                    OR EXISTS (
                        SELECT 1 FROM pm_box_movements bm
                        JOIN pm_boxes b ON b.box_id = bm.box_id
                        WHERE bm.transfer_id = t.transfer_id
                          AND (b.box_code LIKE %s OR b.short_code LIKE %s)
                    )
                )""")
                params.extend([sp, sp, sp, sp, sp, sp, sp, sp])
            else:
                where.append("(t.transfer_no LIKE %s OR fg.name LIKE %s OR tg.name LIKE %s OR t.remarks LIKE %s)")
                params.extend([sp, sp, sp, sp])

        sw = (' WHERE ' + ' AND '.join(where)) if where else ''
        rows = conn.execute(f"""
            SELECT t.*,
                   COALESCE(fg.name,'') AS from_name,
                   COALESCE(tg.name,'') AS to_name,
                   (SELECT COUNT(DISTINCT i.product_id)
                      FROM pm_transfer_items i
                      WHERE i.transfer_id = t.transfer_id AND i.side='out') AS item_count
            FROM pm_transfers t
            LEFT JOIN procurement_godowns fg ON fg.id=t.from_godown_id
            LEFT JOIN procurement_godowns tg ON tg.id=t.to_godown_id
            {sw}
            ORDER BY COALESCE(t.out_at, t.in_at) DESC, t.transfer_id DESC
            LIMIT 500
        """, params).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = dict(r)
            for k in ('out_at','in_at'):
                if d.get(k) is not None: d[k] = str(d[k])
            d['total_qty'] = float(d.get('total_qty') or 0)
            d['item_count'] = int(d.get('item_count') or 0)
            d['has_discrepancy'] = bool(d.get('has_discrepancy'))
            out.append(d)
        return jsonify({'status':'ok','transfers':out,'deep':deep,'search':q_search})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/voucher_data')
@_login_required
def api_transfer_voucher_data(tid):
    """
    Return everything needed to print a transfer voucher:
      - header (transfer_no, dates, status, remarks, full from/to godown details)
      - items grouped by product (product_name, pm_type, brand, no_of_box,
        per_box_qty, total_qty)
      - edit history (every state-changing action with user + timestamp)

    Query parameter: ?type=out (default) | ?type=in
      - type=out → items aggregated from OUT scans (what was sent)
      - type=in  → items aggregated from IN scans (what was received)
    """
    voucher_type = (request.args.get('type') or 'out').lower()
    if voucher_type not in ('out', 'in'):
        voucher_type = 'out'
    conn = sampling_portal.get_db_connection()
    try:
        # Header with both godowns fully expanded (matches MTV print's "From / To" boxes)
        t = conn.execute("""
            SELECT t.*,
                   fg.name AS from_name, fg.address AS from_address, fg.contact AS from_contact,
                   fg.phone AS from_phone, fg.email AS from_email,
                   COALESCE(fg.type,'') AS from_type,
                   tg.name AS to_name,   tg.address AS to_address,   tg.contact AS to_contact,
                   tg.phone AS to_phone, tg.email AS to_email,
                   COALESCE(tg.type,'') AS to_type
            FROM pm_transfers t
            LEFT JOIN procurement_godowns fg ON fg.id = t.from_godown_id
            LEFT JOIN procurement_godowns tg ON tg.id = t.to_godown_id
            WHERE t.transfer_id = %s
        """, (tid,)).fetchone()
        if not t:
            conn.close(); return jsonify({'status': 'error', 'message': 'Transfer not found'}), 404

        # Aggregate scanned boxes by product (filter by movement_type per voucher type)
        items = conn.execute("""
            SELECT b.product_id,
                   p.product_name,
                   p.pm_type,
                   COALESCE(p.product_code,'') AS product_code,
                   COALESCE(br.name, '')      AS brand_name,
                   COUNT(*)                   AS no_of_box,
                   MAX(b.per_box_qty)         AS per_box_qty,
                   SUM(b.per_box_qty)         AS total_qty
            FROM pm_box_movements m
            JOIN pm_boxes b   ON b.box_id     = m.box_id
            JOIN pm_products p ON p.id        = b.product_id
            LEFT JOIN procurement_brands br ON br.id = p.brand_id
            WHERE m.transfer_id = %s
              AND m.movement_type = %s
            GROUP BY b.product_id, p.product_name, p.pm_type, p.product_code, br.name
            ORDER BY p.product_name
        """, (tid, voucher_type)).fetchall()

        # Per-product list of individual scanned boxes (for the dropdown detail
        # rows under each print item). One row per box scan, sorted oldest-first.
        # Includes box_id and short_code so the OUT-Scanned Boxes modal can
        # offer a per-box "remove" action (which calls unscan_box server-side
        # using box_id) and show the compact short_code instead of the long
        # box_code where useful.
        box_rows = conn.execute("""
            SELECT b.box_id,
                   b.product_id,
                   b.box_code,
                   COALESCE(b.short_code,'') AS short_code,
                   b.per_box_qty,
                   m.movement_at,
                   COALESCE(m.moved_by,'') AS moved_by
            FROM pm_box_movements m
            JOIN pm_boxes b ON b.box_id = m.box_id
            WHERE m.transfer_id = %s
              AND m.movement_type = %s
            ORDER BY b.product_id, m.movement_at ASC, m.movement_id ASC
        """, (tid, voucher_type)).fetchall()
        boxes_by_product = {}
        for br_ in box_rows:
            d = dict(br_)
            ts = d.get('movement_at')
            d['movement_at'] = str(ts) if ts is not None else ''
            d['per_box_qty'] = float(d.get('per_box_qty') or 0)
            boxes_by_product.setdefault(d['product_id'], []).append({
                'box_id':      int(d['box_id']),
                'box_code':    d['box_code'],
                'short_code':  d['short_code'] or '',
                'per_box_qty': d['per_box_qty'],
                'movement_at': d['movement_at'],
                'moved_by':    d['moved_by'],
            })

        # Edit history — every state-changing action, oldest first
        edits = conn.execute("""
            SELECT action, edited_by, edited_at, COALESCE(details,'') AS details
            FROM pm_transfer_edits
            WHERE transfer_id = %s
            ORDER BY edited_at ASC, edit_id ASC
        """, (tid,)).fetchall()

        conn.close()

        hdr = dict(t)
        for k in ('out_at', 'in_at'):
            if hdr.get(k) is not None: hdr[k] = str(hdr[k])
        hdr['total_qty']   = float(hdr.get('total_qty') or 0)

        # Distill edit history into a compact summary suitable for the voucher footer
        edit_rows = []
        for r in edits:
            d = dict(r)
            d['edited_at'] = str(d.get('edited_at') or '')
            edit_rows.append(d)

        # Distinct editor names in chronological order, deduped
        seen, distinct_editors = set(), []
        for r in edit_rows:
            u = (r.get('edited_by') or '').strip()
            if u and u not in seen:
                seen.add(u); distinct_editors.append(u)
        creator      = distinct_editors[0] if distinct_editors else (hdr.get('out_by') or '')
        last_editor  = distinct_editors[-1] if distinct_editors else creator
        edits_differ = (creator != last_editor) and (len(distinct_editors) > 1)

        item_rows = []
        for r in items:
            d = dict(r)
            d['no_of_box']   = int(d.get('no_of_box') or 0)
            d['per_box_qty'] = float(d.get('per_box_qty') or 0)
            d['total_qty']   = float(d.get('total_qty') or 0)
            d['boxes']       = boxes_by_product.get(d['product_id'], [])
            item_rows.append(d)

        return jsonify({
            'status':           'ok',
            'voucher_type':     voucher_type,
            'header':           hdr,
            'items':            item_rows,
            'edits':            edit_rows,
            'creator':          creator,
            'last_editor':      last_editor,
            'editors_differ':   edits_differ,
            'distinct_editors': distinct_editors
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/item_detail')
@_login_required
def api_item_detail():
    product_id = request.args.get('product_id')
    month      = request.args.get('month')  # YYYY-MM
    if not product_id:
        return jsonify({'status': 'error', 'message': 'product_id required'}), 400

    conn = sampling_portal.get_db_connection()
    g = _godown_summary(conn, product_id=int(product_id))
    f = _floor_summary(conn, product_id=int(product_id))
    g = g[0] if g else {}
    f = f[0] if f else {}

    rows = []
    if month:
        y, m = month.split('-')
        from_dt = f"{y}-{m}-01"
        import calendar
        last_day = calendar.monthrange(int(y), int(m))[1]
        to_dt    = f"{y}-{m}-{last_day:02d}"

        g_rows = conn.execute("""
            SELECT txn_date, 'godown' AS source, txn_type, qty, remarks, created_by,
                   COALESCE(voucher_no,'') AS voucher_no,
                   COALESCE(gd.name,'') AS godown_name
            FROM pm_godown_txn gt
            LEFT JOIN procurement_godowns gd ON gd.id = gt.godown_id
            WHERE product_id=%s AND txn_date BETWEEN %s AND %s
            ORDER BY txn_date DESC, gt.id DESC
        """, (product_id, from_dt, to_dt)).fetchall()
        f_rows = conn.execute("""
            SELECT txn_date, 'floor' AS source, txn_type, qty, remarks, created_by,
                   COALESCE(voucher_no,'') AS voucher_no,
                   COALESCE(gd.name,'Factory') AS godown_name
            FROM pm_floor_txn ft
            LEFT JOIN procurement_godowns gd ON gd.id = ft.godown_id
            WHERE product_id=%s AND txn_date BETWEEN %s AND %s
            ORDER BY txn_date DESC, ft.id DESC
        """, (product_id, from_dt, to_dt)).fetchall()
        rows = [dict(r) for r in g_rows] + [dict(r) for r in f_rows]
        for r in rows:
            if hasattr(r.get('txn_date'), 'isoformat'):
                r['txn_date'] = r['txn_date'].isoformat()
        rows.sort(key=lambda r: r.get('txn_date', ''), reverse=True)

    conn.close()
    return jsonify({
        'godown': g,
        'floor':  f,
        'transactions': rows,
    })

@pm_stock_bp.route('/api/pm_stock/import_products', methods=['POST'])
@_login_required
def api_pm_import_products():
    """
    Accepts multi-godown import payload.
    New format (preferred):
      [{product_name, pm_type, op_date, openings: [{godown_id, is_floor, qty}]}]
    Legacy format (still supported for backward compat):
      [{product_name, pm_type, op_date, godown_op, floor_op, godown_id}]
    """
    blocked = _block_if_disabled('opening')
    if blocked is not None: return blocked
    # Requesters can't create vouchers — they only submit Material Requests.
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    # New-voucher-entries access check (per-user access control)
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked
    items = request.get_json() or []
    if not isinstance(items, list):
        return jsonify({'status': 'error', 'message': 'Expected JSON array'}), 400

    from datetime import date as _date
    default_date = str(_date.today())
    user         = _user()

    conn    = sampling_portal.get_db_connection()
    added   = 0
    skipped = 0
    op_added = 0
    try:
        # Cache godown types for floor/godown detection on legacy payloads
        godown_rows = conn.execute(
            "SELECT id, COALESCE(type,'godown') AS godown_type FROM procurement_godowns"
        ).fetchall()
        godown_type_map = {r['id']: r['godown_type'] for r in godown_rows}

        for item in items:
            if not isinstance(item, dict):
                skipped += 1; continue
            name    = (item.get('product_name') or '').strip()
            pm      = (item.get('pm_type')      or '').strip()
            op_date = (item.get('op_date') or default_date).strip()

            if not name or not pm:
                skipped += 1; continue

            # Duplicate guard (whitespace/case-insensitive) — reuse the
            # existing product instead of creating a visual duplicate.
            existing = _find_duplicate_product(conn, name, pm)

            if existing:
                product_id = existing['id']; skipped += 1
            else:
                cur = conn.execute(
                    "INSERT INTO pm_products (product_name, pm_type) VALUES (%s, %s)", (name, pm)
                )
                product_id = cur.lastrowid; added += 1

            # ── Build list of (godown_id, is_floor, qty) openings ──
            openings = []
            if isinstance(item.get('openings'), list):
                # New format
                for o in item['openings']:
                    try:
                        qty = float(o.get('qty') or 0)
                    except (TypeError, ValueError):
                        qty = 0.0
                    if qty <= 0:
                        continue
                    gid = o.get('godown_id')
                    is_floor = bool(o.get('is_floor'))
                    # If is_floor not supplied, infer from godown_type_map
                    if 'is_floor' not in o and gid in godown_type_map:
                        is_floor = (godown_type_map[gid] == 'floor')
                    openings.append((gid, is_floor, qty))
            else:
                # Legacy format
                _gop = item.get('godown_op')
                _fop = item.get('floor_op')
                godown_op = float(_gop) if _gop not in (None, '', False) else 0.0
                floor_op  = float(_fop) if _fop not in (None, '', False) else 0.0
                gid = item.get('godown_id') or None
                if godown_op > 0:
                    openings.append((gid, False, godown_op))
                if floor_op > 0:
                    openings.append((gid, True, floor_op))

            # Insert each opening as its own txn
            for gid, is_floor, qty in openings:
                if is_floor:
                    # Factory opening — only once per (product, godown_id)
                    ex = conn.execute(
                        "SELECT id FROM pm_floor_txn WHERE product_id=%s AND txn_type='floor_opening' "
                        "AND (godown_id = %s OR (godown_id IS NULL AND %s IS NULL))",
                        (product_id, gid, gid)
                    ).fetchone()
                    if not ex:
                        conn.execute(
                            """INSERT INTO pm_floor_txn
                               (product_id, txn_date, txn_type, qty, remarks, created_by, godown_id)
                               VALUES (%s,%s,'floor_opening',%s,'Factory Opening Balance (Import)',%s,%s)""",
                            (product_id, op_date, qty, user, gid)
                        )
                        op_added += 1
                else:
                    # Godown opening — only once per (product, godown_id)
                    ex = conn.execute(
                        "SELECT id FROM pm_godown_txn WHERE product_id=%s AND txn_type='opening' "
                        "AND (godown_id = %s OR (godown_id IS NULL AND %s IS NULL))",
                        (product_id, gid, gid)
                    ).fetchone()
                    if not ex:
                        conn.execute(
                            """INSERT INTO pm_godown_txn
                               (product_id, txn_date, txn_type, qty, remarks, created_by, godown_id)
                               VALUES (%s,%s,'opening',%s,'Opening Balance (Import)',%s,%s)""",
                            (product_id, op_date, qty, user, gid)
                        )
                        op_added += 1

        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

    conn.close()
    return jsonify({'status': 'ok', 'added': added, 'skipped': skipped, 'op_added': op_added})

@pm_stock_bp.route('/api/pm_stock/reset/categories', methods=['GET'])
@_login_required
def api_pm_reset_list_categories():
    """Return the menu of categories the admin can clear, with row counts so the
    UI can show 'GRN Vouchers (45 rows)' etc."""
    if (session.get('User_Type','').lower() != 'admin'):
        return jsonify({'status':'error','message':'Admin access required'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        out = []
        for key, spec in _RESET_CATEGORIES.items():
            total = 0
            per_table = []
            for tbl in spec['tables']:
                try:
                    r = conn.execute(f"SELECT COUNT(*) AS c FROM {tbl}").fetchone()
                    rows = int(r['c']) if r else 0
                except Exception:
                    rows = 0
                total += rows
                per_table.append({'table': tbl, 'rows': rows})
            out.append({
                'key':   key,
                'label': spec['label'],
                'desc':  spec['desc'],
                'tables': per_table,
                'total_rows': total
            })
        conn.close()
        return jsonify({'status':'ok', 'categories': out})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/reset/categories', methods=['POST'])
@_login_required
def api_pm_reset_clear_categories():
    """Clear the requested categories. Body: {categories: ['transactions','grns',...]}"""
    if (session.get('User_Type','').lower() != 'admin'):
        return jsonify({'status':'error','message':'Admin access required'}), 403
    d = request.get_json() or {}
    cats = d.get('categories') or []
    if not isinstance(cats, list) or not cats:
        return jsonify({'status':'error','message':'categories required (non-empty list)'}), 400
    invalid = [c for c in cats if c not in _RESET_CATEGORIES]
    if invalid:
        return jsonify({'status':'error','message':f'Unknown categories: {invalid}'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        summary = _do_clear_categories(conn, cats)
        conn.commit()
        conn.close()
        total_deleted = sum(s.get('deleted', 0) for s in summary)
        any_errors = [s for s in summary if 'error' in s]
        return jsonify({
            'status':         'ok',
            'message':        f"Cleared {total_deleted} row(s) across {len(cats)} categor{'y' if len(cats)==1 else 'ies'}" + (f' · {len(any_errors)} error(s)' if any_errors else ''),
            'summary':        summary,
            'errors':         any_errors,
            'total_deleted':  total_deleted
        })
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/reset_all', methods=['POST'])
@_login_required
def api_pm_reset_all():
    """
    Legacy endpoint kept for backward compatibility. Internally delegates to
    the category clearer with all transactional categories selected; products
    only included if wipe_products=true.
    """
    role = session.get('User_Type', '')
    if role.lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin access required'}), 403
    d = request.get_json() or {}
    wipe_products = bool(d.get('wipe_products', False))
    cats = ['transactions', 'transfers', 'mtvs', 'grns', 'dns', 'boxes', 'audit']
    if wipe_products:
        cats.append('products')
    conn = sampling_portal.get_db_connection()
    try:
        summary = _do_clear_categories(conn, cats)
        conn.commit()
        conn.close()
        msg = "All PM transactions cleared" + (" and products deleted" if wipe_products else "")
        return jsonify({'status': 'ok', 'message': msg, 'summary': summary})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/voucher_numbering/list')
@_login_required
def api_pm_vn_list():
    # User-facing PM voucher types. The corresponding `_next_voucher_no`
    # entries in helpers.py route each of these to procurement_voucher_numbering
    # so any style configured here takes immediate effect; if no style is
    # configured, the helper falls back to pm_voucher_sequences (legacy default).
    # Internal-only ledger types (pm_gtxn, pm_ftxn) are intentionally omitted
    # — those are auto-created stock-movement vouchers, not user-created.
    PM_TYPES = ('pm_grn', 'pm_dn', 'pm_mt', 'pm_mtv', 'pm_al', 'pm_aud', 'pm_op', 'pm_mr')
    conn = sampling_portal.get_db_connection()
    rows = conn.execute(
        "SELECT * FROM procurement_voucher_numbering WHERE voucher_type IN %s ORDER BY voucher_type, valid_from DESC",
        (PM_TYPES,)
    ).fetchall()
    conn.close()
    return jsonify({'status': 'ok', 'styles': [dict(r) for r in rows]})

@pm_stock_bp.route('/api/pm_stock/voucher_numbering/save', methods=['POST'])
@_login_required
def api_pm_vn_save():
    # PM users cannot modify voucher numbering — admin / other roles only.
    if (session.get('User_Type','') or '').strip() == 'PM':
        return jsonify({'status':'error','message':'PM users cannot modify voucher numbering'}), 403
    d = request.get_json() or {}
    _ALLOWED_PM_VN = ('pm_grn', 'pm_dn', 'pm_mt', 'pm_mtv', 'pm_al', 'pm_aud', 'pm_op', 'pm_mr')
    if d.get('voucher_type') not in _ALLOWED_PM_VN:
        return jsonify({
            'status':  'error',
            'message': f"voucher_type must be one of: {', '.join(_ALLOWED_PM_VN)}"
        }), 400
    conn = sampling_portal.get_db_connection()
    try:
        sid = d.get('id')
        if sid:
            conn.execute(
                "UPDATE procurement_voucher_numbering SET prefix=%s,suffix=%s,digits=%s,start_num=%s,valid_from=%s,valid_to=%s WHERE id=%s",
                (d.get('prefix',''), d.get('suffix',''), int(d.get('digits',4)),
                 int(d.get('start_num',1)), d.get('valid_from'), d.get('valid_to'), sid)
            )
        else:
            conn.execute(
                "INSERT INTO procurement_voucher_numbering (voucher_type,prefix,suffix,digits,start_num,valid_from,valid_to) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (d.get('voucher_type'), d.get('prefix',''), d.get('suffix',''),
                 int(d.get('digits',4)), int(d.get('start_num',1)), d.get('valid_from'), d.get('valid_to'))
            )
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/voucher_numbering/delete', methods=['POST'])
@_login_required
def api_pm_vn_delete():
    """Delete a voucher-numbering rule. Admin-only as of Phase 1."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only — only admins can modify voucher numbering.'}), 403
    d = request.get_json() or {}
    sid = d.get('id')
    if not sid:
        return jsonify({'status': 'error', 'message': 'id required'}), 400
    conn = sampling_portal.get_db_connection()
    conn.execute("DELETE FROM procurement_voucher_numbering WHERE id=%s", (sid,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})

@pm_stock_bp.route('/api/pm_stock/voucher_numbering/next_preview')
@_login_required
def api_pm_vn_next_preview():
    import re as _re
    vtype = request.args.get('type', 'pm_grn')
    today = str(date.today())
    conn  = sampling_portal.get_db_connection()
    try:
        vn = conn.execute(
            "SELECT prefix,suffix,digits FROM procurement_voucher_numbering WHERE voucher_type=%s AND valid_from<=%s AND valid_to>=%s ORDER BY id DESC LIMIT 1",
            (vtype, today, today)
        ).fetchone()
        if not vn:
            conn.close()
            return jsonify({'status': 'ok', 'preview': None})
        prefix = (vn['prefix'] or '').strip()
        suffix = (vn['suffix'] or '').strip()
        digits = int(vn['digits'] or 4)
        tbl_map = {'pm_grn': ('pm_grn','grn_no'), 'pm_mtv': ('pm_mtv','mtv_no'), 'pm_mt': ('pm_transfers','transfer_no')}
        tbl, col = tbl_map.get(vtype, ('pm_grn','grn_no'))
        pattern  = (prefix + '/%') if prefix else '%'
        rows = conn.execute(f"SELECT {col} FROM {tbl} WHERE {col} LIKE %s", (pattern,)).fetchall()
        max_seq = 0
        for row in rows:
            nums = _re.findall(r'(\d{' + str(digits) + r',})', str(row[col] or ''))
            if nums: max_seq = max(max_seq, int(nums[-1]))
        preview = '/'.join(p for p in [prefix, str(max_seq+1).zfill(digits), suffix] if p)
        conn.close()
        return jsonify({'status': 'ok', 'preview': preview})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'ok', 'preview': None})

@pm_stock_bp.route('/api/pm_stock/grn/update', methods=['POST'])
@_login_required
def api_grn_update():
    d         = request.get_json() or {}
    grn_id    = d.get('id')
    grn_date  = d.get('grn_date') or str(date.today())
    godown_id = d.get('godown_id')
    supplier  = (d.get('supplier') or '').strip()
    po_num    = (d.get('po_number') or '').strip() or None
    po_date   = d.get('po_date') or None
    remarks   = (d.get('remarks') or '').strip()
    party_invoice_no   = (d.get('party_invoice_no') or '').strip() or None
    party_invoice_date = d.get('party_invoice_date') or None
    supervisor_name    = (d.get('supervisor_name') or '').strip() or None
    items     = d.get('items', [])
    if not grn_id or not godown_id or not items:
        return jsonify({'status':'error','message':'id, godown_id, items required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        grn = conn.execute("SELECT grn_no FROM pm_grn WHERE id=%s", (grn_id,)).fetchone()
        if not grn: conn.close(); return jsonify({'status':'error','message':'GRN not found'}), 404
        grn_no = grn['grn_no']

        # Refuse if any box from this GRN has already been moved beyond grn_create.
        # Editing the GRN would invalidate downstream movement history.
        moved = conn.execute("""
            SELECT COUNT(*) AS c
            FROM pm_box_movements m
            JOIN pm_boxes b ON b.box_id = m.box_id
            WHERE b.grn_id = %s AND m.movement_type IN ('out','in','consume','adjust')
        """, (grn_id,)).fetchone()
        if int((moved or {}).get('c') or 0) > 0:
            conn.close()
            return jsonify({
                'status': 'error',
                'message': 'Cannot edit GRN — boxes from this GRN have already been moved or consumed. Reverse those movements first.'
            }), 409

        conn.execute(
            "UPDATE pm_grn SET grn_date=%s,godown_id=%s,supplier=%s,po_number=%s,po_date=%s,remarks=%s,party_invoice_no=%s,party_invoice_date=%s,supervisor_name=%s WHERE id=%s",
            (grn_date,godown_id,supplier,po_num,po_date,remarks,party_invoice_no,party_invoice_date or None,supervisor_name,grn_id)
        )
        # Remove old inward txns
        old = conn.execute("SELECT product_id, qty_received FROM pm_grn_items WHERE grn_id=%s",(grn_id,)).fetchall()
        for o in old:
            conn.execute("DELETE FROM pm_godown_txn WHERE product_id=%s AND txn_type='inward' AND qty=%s AND godown_id=%s AND remarks LIKE %s ORDER BY id DESC LIMIT 1",
                         (o['product_id'],o['qty_received'],godown_id,f'%{grn_no}%'))
        # Wipe boxes for this GRN (safe — we already verified none have moved)
        _delete_boxes_for_grn(conn, grn_id)
        conn.execute("DELETE FROM pm_grn_items WHERE grn_id=%s",(grn_id,))
        for item in items:
            pid = item.get('product_id'); qty = float(item.get('qty_received') or 0); irm = (item.get('remarks') or '').strip()
            nob = int(item.get('no_of_box') or 0)
            bc  = int(item.get('box_count') or 0)
            pver = (item.get('product_version') or '').strip() or None
            # Phase 2: UOM label only (no conversion).
            euom = (item.get('entered_uom') or '').strip() or None
            # ABC: unit rate at receipt. Same defensive handling as api_grn_save.
            try:
                rate = float(item.get('rate') or 0)
                if rate < 0: rate = 0.0
            except (TypeError, ValueError):
                rate = 0.0
            if not pid or qty<=0: continue
            # Same invariant as api_grn_save: force bc = qty / nob so old
            # form-data inconsistencies can't survive a re-save of the GRN.
            # See api_grn_save for the rationale.
            if nob > 0:
                derived_bc = round(qty / nob, 3)
                if abs(float(bc) - derived_bc) > 0.01:
                    bc = derived_bc
                else:
                    bc = derived_bc
            cur = conn.execute(
                "INSERT INTO pm_grn_items (grn_id,product_id,qty_received,no_of_box,box_count,remarks,product_version,entered_uom,rate) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (grn_id,pid,qty,nob,bc,irm,pver,euom,rate)
            )
            grn_item_id = cur.lastrowid
            try:
                _get_or_assign_fifo(conn, 'grn', grn_id, pid)
            except Exception as _fe:
                import sys as _sys
                print(f"[pm_stock_routes] FIFO assign failed (edit) for GRN {grn_id}/{pid}: {_fe}", file=_sys.stderr)
            vno = _next_voucher_no(conn,'PM-GTXN',grn_date)
            conn.execute("INSERT INTO pm_godown_txn (product_id,txn_date,txn_type,qty,remarks,created_by,voucher_no,godown_id) VALUES (%s,%s,'inward',%s,%s,%s,%s,%s)",
                         (pid,grn_date,qty,f'[GRN:{grn_no}] {irm}'.strip(),_user(),vno,godown_id))
            # Recreate boxes
            if nob > 0 and bc > 0:
                pcode_row = conn.execute(
                    "SELECT COALESCE(product_code,'') AS pc FROM pm_products WHERE id=%s", (pid,)
                ).fetchone()
                product_code = (pcode_row['pc'] if pcode_row else '') or ''
                _create_boxes_for_grn_item(
                    conn, grn_id=grn_id, grn_no=grn_no, grn_item_id=grn_item_id,
                    product_id=pid, product_code=product_code,
                    no_of_box=nob, per_box_qty=bc, godown_id=godown_id,
                    grn_date=grn_date, user=_user(),
                    product_version=pver,
                )

        # ── INVARIANT GUARD ─────────────────────────────────────────────────
        # See note in api_grn_save. This is the endpoint where the historic
        # orphan-box corruption (PM-GRN/0312, PM-GRN/0342) most likely
        # originated. The guard ensures any future regression in the
        # rewrite logic (delete-old-then-insert-new) cannot silently leave
        # orphan or missing box rows.
        _assert_grn_box_invariant(conn, grn_id)

        conn.commit(); conn.close()
        return jsonify({'status':'ok','grn_no':grn_no})
    except RuntimeError as _inv_err:
        import sys as _sys
        print(f"[pm_stock_routes] api_grn_update: {_inv_err}", file=_sys.stderr)
        conn.close()
        return jsonify({
            'status':  'error',
            'message': ('GRN data consistency check failed. The edit was '
                        'cancelled to prevent corruption. Please retry; if '
                        'this persists, contact the developer with the GRN '
                        'number.')
        }), 500
    except Exception as e:
        conn.close(); return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/dn/list')
@_login_required
def api_dn_list():
    """List delivery notes with optional date + supplier + godown filters."""
    conn = sampling_portal.get_db_connection()
    q_from     = request.args.get('from') or None
    q_to       = request.args.get('to') or None
    q_supplier = (request.args.get('supplier') or '').strip()
    q_godown   = request.args.get('godown') or None

    where = []
    params = []
    if q_from: where.append("d.dn_date >= %s"); params.append(q_from)
    if q_to:   where.append("d.dn_date <= %s"); params.append(q_to)
    if q_supplier:
        where.append("d.supplier LIKE %s"); params.append(f"%{q_supplier}%")
    if q_godown:
        where.append("d.from_godown = %s"); params.append(int(q_godown))
    wc = ("WHERE " + " AND ".join(where)) if where else ""

    try:
        rows = conn.execute(f"""
            SELECT d.id, d.dn_no, d.dn_date, d.supplier, d.from_godown,
                   COALESCE(gd.name,'') AS godown_name,
                   d.reason, d.remarks, d.supervisor_name,
                   d.reference_no, d.reference_date, d.created_by, d.created_at,
                   COUNT(di.id) AS item_count,
                   COALESCE(SUM(di.qty_delivered),0) AS total_qty
            FROM pm_dn d
            LEFT JOIN pm_dn_items di ON di.dn_id = d.id
            LEFT JOIN procurement_godowns gd ON gd.id = d.from_godown
            {wc}
            GROUP BY d.id
            ORDER BY d.dn_date DESC, d.id DESC
        """, params).fetchall()
    except Exception as e:
        conn.close(); return jsonify({'status':'error','message':str(e)}), 500
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        for k in ('dn_date','reference_date','created_at'):
            if d.get(k) and hasattr(d[k],'isoformat'):
                d[k] = d[k].isoformat() if k!='created_at' else str(d[k])
        d['total_qty'] = float(d['total_qty'] or 0)
        out.append(d)
    return jsonify({'status':'ok','dns': out})

@pm_stock_bp.route('/api/pm_stock/dn/<int:dn_id>')
@_login_required
def api_dn_detail(dn_id):
    """
    Returns DN header + items + per-item box list.

    Each item gets a 'boxes' array of {box_id, box_code, short_code,
    per_box_qty, prior_godown_id} — the boxes the operator scanned into
    that line. Frontend renders these as chips under the row.
    """
    conn = sampling_portal.get_db_connection()
    d = conn.execute("""
        SELECT d.*, COALESCE(gd.name,'') AS godown_name
        FROM pm_dn d
        LEFT JOIN procurement_godowns gd ON gd.id = d.from_godown
        WHERE d.id = %s
    """, (dn_id,)).fetchone()
    if not d: conn.close(); return jsonify({'status':'error','message':'DN not found'}), 404
    items = conn.execute("""
        SELECT di.id, di.product_id, p.product_name, p.pm_type,
               di.qty_delivered, COALESCE(di.no_of_box,0) AS no_of_box,
               COALESCE(di.box_count,0) AS box_count,
               di.remarks
        FROM pm_dn_items di
        JOIN pm_products p ON p.id = di.product_id
        WHERE di.dn_id = %s
    """, (dn_id,)).fetchall()

    # Box scans for this DN — group by dn_item_id (and fall back to
    # (product_id, per_box_qty) for legacy rows where dn_item_id is NULL
    # because the linking happened before item re-IDs settled).
    box_rows = conn.execute("""
        SELECT s.id, s.dn_item_id, s.box_id, s.box_code, s.product_id,
               s.per_box_qty, s.prior_godown_id, s.created_at, s.created_by,
               COALESCE(b.short_code,'') AS short_code,
               b.current_status AS box_current_status
        FROM pm_dn_box_scans s
        LEFT JOIN pm_boxes b ON b.box_id = s.box_id
        WHERE s.dn_id = %s
        ORDER BY s.id ASC
    """, (dn_id,)).fetchall()
    conn.close()

    # Build a lookup: dn_item_id -> [box dicts]
    by_item = {}
    orphan_by_pid_qty = {}  # (product_id, per_box_qty) -> [box dicts]
    for r in box_rows:
        bd = dict(r)
        for k in ('per_box_qty',):
            if bd.get(k) is not None: bd[k] = float(bd[k])
        for k in ('created_at',):
            if bd.get(k) is not None: bd[k] = str(bd[k])
        if bd.get('dn_item_id'):
            by_item.setdefault(int(bd['dn_item_id']), []).append(bd)
        else:
            key = (int(bd['product_id']), round(float(bd['per_box_qty'] or 0), 3))
            orphan_by_pid_qty.setdefault(key, []).append(bd)

    out = dict(d)
    for k in ('dn_date','reference_date','created_at'):
        if out.get(k) and hasattr(out[k],'isoformat'):
            out[k] = out[k].isoformat() if k!='created_at' else str(out[k])

    items_out = []
    for it in items:
        idict = dict(it)
        # Direct match
        boxes = list(by_item.get(int(idict['id']), []))
        # Orphan adoption — re-attach legacy rows by product+per_box_qty match
        key = (int(idict['product_id']), round(float(idict.get('box_count') or 0), 3))
        adopted = orphan_by_pid_qty.pop(key, [])
        if adopted:
            boxes.extend(adopted)
        idict['boxes'] = boxes
        items_out.append(idict)

    out['items'] = items_out
    return jsonify(out)


# ── DN box scan helpers ─────────────────────────────────────────────────────
#
# The DN page lets the operator scan boxes directly into a DN row. Boxes
# are recorded immediately (per the user's choice in JS state) but stock
# status changes happen at SAVE time — so a half-typed DN can be closed
# without leaking inventory.
#
# Flow:
#   1. User scans → frontend calls /dn/box/check to validate the box can
#      go on this DN (lives at the right godown, not already on another
#      open DN, etc.) and gets back its product/qty for line-building.
#   2. User saves → frontend posts each item with its `boxes:[code,...]`.
#      Server inside the save transaction:
#        a. Walks each box → marks it 'consumed' in pm_boxes (snapshot
#           prior godown so unscan can restore).
#        b. Writes a pm_box_movements row of type 'out' referencing the
#           DN (transfer_id is NULL — DN doesn't live in pm_transfers,
#           but the remarks identify the DN_no).
#        c. Inserts a pm_dn_box_scans row linking dn_id → box_id.
#   3. User edits & removes a box → on update, server diffs old vs new
#      box lists, restoring removed boxes back to 'in_stock' at their
#      prior_godown_id.
#   4. DN deleted → all linked boxes get restored to 'in_stock'.

def _dn_validate_box_for_scan(conn, code, from_godown_id, dn_id=None):
    """
    Resolve a scan code (short_code or box_code) and validate it can be
    attached to a DN sourced at from_godown_id. dn_id (if given) is the
    DN currently being edited — boxes already on THIS DN are still OK.

    Returns (box_row_dict, error_string). Exactly one is None.
    """
    if not code:
        return None, 'No box code provided'
    b = conn.execute("""
        SELECT b.box_id, b.box_code, b.short_code, b.product_id, b.product_code,
               b.per_box_qty, b.current_godown_id, b.current_status,
               p.product_name, p.pm_type
        FROM pm_boxes b
        JOIN pm_products p ON p.id = b.product_id
        WHERE b.short_code = %s OR b.box_code = %s
        LIMIT 1
    """, (code, code)).fetchone()
    if not b:
        return None, f'Box {code} not found'
    bd = dict(b)
    # Allow re-scanning a box that's already on the SAME DN (edit reopens it)
    already_here = False
    if dn_id:
        chk = conn.execute(
            "SELECT 1 FROM pm_dn_box_scans WHERE dn_id=%s AND box_id=%s LIMIT 1",
            (dn_id, bd['box_id'])
        ).fetchone()
        already_here = bool(chk)
    if already_here:
        # Caller can decide whether to treat as "already on this DN" — return
        # the box with an annotation; not an error per se.
        bd['_already_on_this_dn'] = True
        return bd, None
    if bd['current_status'] == 'superseded':
        return None, f'Box {bd["box_code"]} was split — scan one of its children instead'
    if bd['current_status'] != 'in_stock':
        return None, (f"Box {bd['box_code']} is '{bd['current_status']}' — must be in_stock"
                      + (" (already attached to another DN?)" if bd['current_status']=='consumed' else ''))
    if int(bd['current_godown_id'] or 0) != int(from_godown_id):
        return None, (f"Box {bd['box_code']} is at a different godown — cannot DN from this source")
    return bd, None


def _dn_consume_box(conn, dn_id, dn_no, dn_item_id, box, from_godown_id, supplier):
    """Mark a box consumed-on-DN: status flip + movement row + junction row.
    Snapshots prior_godown_id so unscan can restore the box later.
    """
    prior_godown = int(box['current_godown_id'] or 0) or None
    conn.execute(
        "UPDATE pm_boxes SET current_status='consumed' WHERE box_id=%s "
        "AND current_status='in_stock'",
        (box['box_id'],)
    )
    # Log the movement so stock ledger / audit trail shows the consumption
    try:
        conn.execute(
            """INSERT INTO pm_box_movements
                 (box_id, transfer_id, movement_type, from_godown_id,
                  to_godown_id, qty, moved_by, remarks)
               VALUES (%s, NULL, 'consume', %s, NULL, %s, %s, %s)""",
            (box['box_id'], prior_godown, float(box['per_box_qty'] or 0),
             _user(),
             f"[PM DN: {dn_no}] To: {supplier} — box {box['box_code']}")
        )
    except Exception:
        # pm_box_movements may not exist on extremely old installs — don't
        # let movement-log failures block the save.
        pass
    conn.execute(
        """INSERT INTO pm_dn_box_scans
             (dn_id, dn_item_id, box_id, box_code, product_id, per_box_qty,
              prior_godown_id, created_by)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
        (dn_id, dn_item_id, box['box_id'], box['box_code'],
         box['product_id'], float(box['per_box_qty'] or 0),
         prior_godown, _user())
    )


def _dn_restore_box(conn, dn_id, scan_row):
    """Reverse of _dn_consume_box: status back to in_stock at prior godown,
    junction row deleted. scan_row is a dict from pm_dn_box_scans.
    """
    prior = scan_row.get('prior_godown_id')
    # Only restore boxes that we previously marked consumed — don't stomp
    # over a box that the operator manually reassigned somewhere else.
    if prior:
        conn.execute(
            "UPDATE pm_boxes SET current_status='in_stock', current_godown_id=%s "
            "WHERE box_id=%s AND current_status='consumed'",
            (int(prior), int(scan_row['box_id']))
        )
    else:
        conn.execute(
            "UPDATE pm_boxes SET current_status='in_stock' "
            "WHERE box_id=%s AND current_status='consumed'",
            (int(scan_row['box_id']),)
        )
    try:
        conn.execute(
            """INSERT INTO pm_box_movements
                 (box_id, transfer_id, movement_type, from_godown_id,
                  to_godown_id, qty, moved_by, remarks)
               VALUES (%s, NULL, 'cancel', NULL, %s, %s, %s, %s)""",
            (int(scan_row['box_id']), prior,
             float(scan_row.get('per_box_qty') or 0),
             _user(), f"DN-scan reversed (dn_id={dn_id}, box {scan_row.get('box_code') or scan_row['box_id']})")
        )
    except Exception:
        pass
    conn.execute(
        "DELETE FROM pm_dn_box_scans WHERE dn_id=%s AND box_id=%s",
        (dn_id, int(scan_row['box_id']))
    )


@pm_stock_bp.route('/api/pm_stock/dn/box/check', methods=['POST'])
@_login_required
def api_dn_box_check():
    """
    Pre-validate a box scan for a DN row. Called from the scanner input
    BEFORE save — gives the user immediate feedback ("not at this godown",
    "already on another DN", etc.) so they don't discover problems only
    when they hit Save.

    Body: {code: 'BOX-CODE', from_godown_id: <int>, dn_id?: <int>}
    Returns: {status:'ok', box: {...}} or {status:'error', message:'...'}
    """
    d = request.get_json() or {}
    code = (d.get('code') or '').strip().upper()
    from_godown = d.get('from_godown_id')
    dn_id = d.get('dn_id')
    if not code:
        return jsonify({'status':'error','message':'code required'}), 400
    if not from_godown:
        return jsonify({'status':'error','message':'from_godown_id required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        box, err = _dn_validate_box_for_scan(conn, code, int(from_godown),
                                             int(dn_id) if dn_id else None)
        conn.close()
        if err:
            return jsonify({'status':'error','message':err}), 409
        # Normalise types for JSON
        out = dict(box)
        if out.get('per_box_qty') is not None:
            out['per_box_qty'] = float(out['per_box_qty'])
        return jsonify({'status':'ok','box': out})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/dn/save', methods=['POST'])
@_login_required
def api_dn_save():
    """Create a Delivery Note (HCP → Supplier).
    Each item creates an 'outward' godown transaction — reduces stock.
    Items may include a `boxes: [code, ...]` list; each box gets marked
    'consumed' and linked via pm_dn_box_scans.
    """
    blocked = _block_if_disabled('dn')
    if blocked is not None: return blocked
    # Requesters can't create vouchers — they only submit Material Requests.
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    # New-voucher-entries access check (per-user access control)
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked
    d = request.get_json() or {}
    dn_date        = d.get('dn_date') or str(date.today())
    supplier       = (d.get('supplier') or '').strip()
    from_godown    = d.get('from_godown') or None
    reason         = (d.get('reason') or '').strip()
    remarks        = (d.get('remarks') or '').strip()
    supervisor     = (d.get('supervisor_name') or '').strip() or None
    reference_no   = (d.get('reference_no') or '').strip() or None
    reference_date = d.get('reference_date') or None
    items          = d.get('items', [])

    if not supplier:
        return jsonify({'status':'error','message':'Supplier is required'}), 400
    if not from_godown:
        return jsonify({'status':'error','message':'From godown is required'}), 400
    if not items:
        return jsonify({'status':'error','message':'At least one item required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # Enforce per-user home godown
        ok, msg = _enforce_home_godown(conn, ('from_godown', from_godown))
        if not ok:
            conn.close()
            return jsonify({'status':'error','message':msg}), 403

        # Pre-validate ALL scanned boxes BEFORE inserting anything. If one
        # box fails (e.g. someone else's MTV started in the meantime),
        # we'd rather reject the whole save than create a half-committed DN.
        validated = []  # list of (item_idx, box_row)
        for idx, item in enumerate(items):
            for code in (item.get('boxes') or []):
                code = (str(code) or '').strip().upper()
                if not code: continue
                box, err = _dn_validate_box_for_scan(conn, code, int(from_godown))
                if err:
                    conn.close()
                    return jsonify({'status':'error',
                                    'message': f'Cannot save: {err}'}), 409
                validated.append((idx, box))

        dn_no = _next_voucher_no(conn, 'PM-DN', dn_date)
        cur = conn.execute(
            """INSERT INTO pm_dn (dn_no, dn_date, supplier, from_godown, reason, remarks,
                                  supervisor_name, reference_no, reference_date, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (dn_no, dn_date, supplier, int(from_godown), reason, remarks,
             supervisor, reference_no, reference_date or None, _user())
        )
        dn_id = cur.lastrowid

        item_ids = []  # parallel to items[], so we can attach boxes by index
        # Aggregate outward stock txns by product. Multiple DN line items
        # for the same product (e.g. one full-box line of 750 and a
        # half-box line of 785) collapse into a single outward ledger
        # row per product so the stock view reads cleanly. The voucher_no
        # on each merged row is the DN number itself — readers don't
        # need to chase a PG-NNNN cross-reference back to its DN.
        # Across products, you'll still see one row per product (each
        # product has its own running balance), but the voucher_no will
        # be the same DN number on each.
        agg_by_pid = {}   # product_id -> {qty: float, remarks: [list of per-line remarks]}
        for item in items:
            pid = item.get('product_id')
            qty = float(item.get('qty_delivered') or 0)
            nob = int(item.get('no_of_box') or 0)
            bc  = int(item.get('box_count') or 0)
            irm = (item.get('remarks') or '').strip()
            if not pid or qty <= 0:
                item_ids.append(None)
                continue
            icur = conn.execute(
                """INSERT INTO pm_dn_items (dn_id, product_id, qty_delivered, no_of_box, box_count, remarks)
                   VALUES (%s,%s,%s,%s,%s,%s)""",
                (dn_id, pid, qty, nob, bc, irm)
            )
            item_ids.append(icur.lastrowid)
            slot = agg_by_pid.setdefault(int(pid), {'qty': 0.0, 'remarks': []})
            slot['qty'] += qty
            if irm:
                slot['remarks'].append(irm)

        # Now post the aggregated outward txns — one row per distinct product.
        for pid, slot in agg_by_pid.items():
            joined_remarks = '; '.join(slot['remarks']) if slot['remarks'] else ''
            conn.execute(
                """INSERT INTO pm_godown_txn (product_id, txn_date, txn_type, qty,
                                              remarks, created_by, voucher_no, godown_id)
                   VALUES (%s, %s, 'outward', %s, %s, %s, %s, %s)""",
                (pid, dn_date, slot['qty'],
                 (f"[PM DN: {dn_no}] To: {supplier}"
                  + (f" — {joined_remarks}" if joined_remarks else "")).strip(),
                 _user(), dn_no, int(from_godown))
            )

        # Now attach boxes (status flip + junction insert) — all pre-validated
        for idx, box in validated:
            dn_item_id = item_ids[idx] if idx < len(item_ids) else None
            _dn_consume_box(conn, dn_id, dn_no, dn_item_id, box,
                            int(from_godown), supplier)

        conn.commit(); conn.close()
        return jsonify({'status':'ok','dn_no':dn_no,'id':dn_id})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/dn/update', methods=['POST'])
@_login_required
def api_dn_update():
    """Update an existing DN. Diffs the box list: boxes added → consumed,
    boxes removed → restored to in_stock at their prior godown."""
    d = request.get_json() or {}
    dn_id          = d.get('id')
    dn_date        = d.get('dn_date') or str(date.today())
    supplier       = (d.get('supplier') or '').strip()
    from_godown    = d.get('from_godown') or None
    reason         = (d.get('reason') or '').strip()
    remarks        = (d.get('remarks') or '').strip()
    supervisor     = (d.get('supervisor_name') or '').strip() or None
    reference_no   = (d.get('reference_no') or '').strip() or None
    reference_date = d.get('reference_date') or None
    items          = d.get('items', [])

    if not dn_id or not supplier or not from_godown or not items:
        return jsonify({'status':'error','message':'id, supplier, from_godown and items required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # Enforce per-user home godown
        ok, msg = _enforce_home_godown(conn, ('from_godown', from_godown))
        if not ok:
            conn.close()
            return jsonify({'status':'error','message':msg}), 403

        dn = conn.execute("SELECT dn_no FROM pm_dn WHERE id=%s", (dn_id,)).fetchone()
        if not dn: conn.close(); return jsonify({'status':'error','message':'DN not found'}), 404
        dn_no = dn['dn_no']

        # Pre-validate all box scans against the (possibly new) source
        # godown. Boxes already on THIS DN are allowed even if their
        # 'current_status' is 'consumed' (set by the previous save).
        validated = []   # list of (item_idx, box_row, is_already_here)
        submitted_box_ids = set()
        for idx, item in enumerate(items):
            for code in (item.get('boxes') or []):
                code = (str(code) or '').strip().upper()
                if not code: continue
                box, err = _dn_validate_box_for_scan(conn, code, int(from_godown),
                                                     dn_id=int(dn_id))
                if err:
                    conn.close()
                    return jsonify({'status':'error',
                                    'message': f'Cannot save: {err}'}), 409
                already_here = bool(box.pop('_already_on_this_dn', False))
                validated.append((idx, box, already_here))
                submitted_box_ids.add(int(box['box_id']))

        # Restore boxes that were on the DN before but are NOT in the new list
        existing_scans = conn.execute(
            "SELECT id, box_id, box_code, per_box_qty, prior_godown_id "
            "FROM pm_dn_box_scans WHERE dn_id=%s", (int(dn_id),)
        ).fetchall()
        for s in existing_scans:
            if int(s['box_id']) not in submitted_box_ids:
                _dn_restore_box(conn, int(dn_id), dict(s))

        conn.execute(
            """UPDATE pm_dn SET dn_date=%s, supplier=%s, from_godown=%s, reason=%s,
                                remarks=%s, supervisor_name=%s, reference_no=%s, reference_date=%s
               WHERE id=%s""",
            (dn_date, supplier, int(from_godown), reason, remarks,
             supervisor, reference_no, reference_date or None, dn_id)
        )
        # Remove ALL old outward txns linked to this DN. Pre-merge those
        # rows were created one-per-line with PG-NNNN voucher numbers and
        # per-line qty; post-merge they're one-per-product with the DN
        # number itself as voucher_no. Either way, the remarks string
        # contains the DN number, so a LIKE clause catches both formats.
        conn.execute(
            """DELETE FROM pm_godown_txn
               WHERE txn_type='outward' AND remarks LIKE %s""",
            (f'%[PM DN: {dn_no}]%',)
        )
        conn.execute("DELETE FROM pm_dn_items WHERE dn_id=%s",(dn_id,))

        item_ids = []
        # Same aggregation pattern as api_dn_save — one outward row per
        # distinct product, voucher_no = the DN number.
        agg_by_pid = {}   # product_id -> {qty: float, remarks: [list of per-line remarks]}
        for item in items:
            pid = item.get('product_id'); qty = float(item.get('qty_delivered') or 0)
            nob = int(item.get('no_of_box') or 0); bc = int(item.get('box_count') or 0)
            irm = (item.get('remarks') or '').strip()
            if not pid or qty <= 0:
                item_ids.append(None)
                continue
            icur = conn.execute(
                """INSERT INTO pm_dn_items (dn_id, product_id, qty_delivered, no_of_box, box_count, remarks)
                   VALUES (%s,%s,%s,%s,%s,%s)""",
                (dn_id, pid, qty, nob, bc, irm)
            )
            item_ids.append(icur.lastrowid)
            slot = agg_by_pid.setdefault(int(pid), {'qty': 0.0, 'remarks': []})
            slot['qty'] += qty
            if irm:
                slot['remarks'].append(irm)

        for pid, slot in agg_by_pid.items():
            joined_remarks = '; '.join(slot['remarks']) if slot['remarks'] else ''
            conn.execute(
                """INSERT INTO pm_godown_txn (product_id, txn_date, txn_type, qty,
                                              remarks, created_by, voucher_no, godown_id)
                   VALUES (%s, %s, 'outward', %s, %s, %s, %s, %s)""",
                (pid, dn_date, slot['qty'],
                 (f"[PM DN: {dn_no}] To: {supplier}"
                  + (f" — {joined_remarks}" if joined_remarks else "")).strip(),
                 _user(), dn_no, int(from_godown))
            )

        # Re-attach box scans:
        #  - Boxes that were already on this DN → relink dn_item_id only
        #    (don't re-flip status, already consumed).
        #  - New boxes → full consume flow.
        for idx, box, already_here in validated:
            dn_item_id = item_ids[idx] if idx < len(item_ids) else None
            if already_here:
                conn.execute(
                    "UPDATE pm_dn_box_scans SET dn_item_id=%s "
                    "WHERE dn_id=%s AND box_id=%s",
                    (dn_item_id, int(dn_id), int(box['box_id']))
                )
            else:
                _dn_consume_box(conn, int(dn_id), dn_no, dn_item_id, box,
                                int(from_godown), supplier)

        conn.commit(); conn.close()
        return jsonify({'status':'ok','dn_no':dn_no})
    except Exception as e:
        conn.close()
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/dn/delete', methods=['POST'])
@_login_required
def api_dn_delete():
    """Soft-delete a Delivery Note. Admin-only. Restorable from recycle bin.
    All boxes attached to this DN are restored to 'in_stock' at their
    prior godown."""
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only — non-admins cannot delete Delivery Notes.'}), 403
    d = request.get_json() or {}
    dn_id = d.get('id')
    reason = (d.get('reason') or '').strip() or None
    if not dn_id: return jsonify({'status':'error','message':'id required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute("SELECT dn_no, dn_date, supplier FROM pm_dn WHERE id=%s",(dn_id,)).fetchone()
        if not row: conn.close(); return jsonify({'status':'error','message':'DN not found'}), 404
        dn_no = row['dn_no']

        # Restore every box attached to this DN before the soft-delete
        # tucks pm_dn_items / pm_dn rows away. Without this the boxes
        # would stay 'consumed' forever even though the DN no longer
        # exists from the user's POV.
        scans = conn.execute(
            "SELECT id, box_id, box_code, per_box_qty, prior_godown_id "
            "FROM pm_dn_box_scans WHERE dn_id=%s", (int(dn_id),)
        ).fetchall()
        for s in scans:
            _dn_restore_box(conn, int(dn_id), dict(s))

        bin_id = _bin_soft_delete(
            conn,
            entity_type  = 'dn',
            entity_id    = dn_id,
            entity_label = f"DN {dn_no}",
            parent_table = 'pm_dn',
            parent_where = 'id=%s',
            parent_params= (dn_id,),
            children = [
                {'table': 'pm_dn_items',
                 'where': 'dn_id=%s', 'params': (dn_id,)},
                {'table': 'pm_godown_txn',
                 'where': "remarks LIKE %s AND txn_type='outward'",
                 'params': (f'%{dn_no}%',)},
            ],
            summary = f"DN {dn_no} · supplier: {row.get('supplier') or '—'} · dated {row.get('dn_date')}",
            reason  = reason
        )
        conn.commit(); conn.close()
        return jsonify({'status':'ok','dn_no':dn_no, 'bin_id': bin_id, 'message': 'DN moved to recycle bin.'})
    except Exception as e:
        conn.close(); return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/supplier_type_assoc', methods=['GET'])
@_login_required
def api_pm_sup_assoc_get():
    """Return all supplier types with association flag for pm_stock page."""
    conn = sampling_portal.get_db_connection()
    try:
        types = conn.execute(
            "SELECT id, type_name FROM supplier_type ORDER BY sort_order, type_name"
        ).fetchall()
        assoc = {r['supplier_type_id'] for r in conn.execute(
            "SELECT supplier_type_id FROM pm_supplier_type_assoc WHERE page='pm_stock'"
        ).fetchall()}
        conn.close()
        return jsonify({'status': 'ok', 'types': [
            {'id': r['id'], 'type_name': r['type_name'], 'associated': r['id'] in assoc}
            for r in types
        ]})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/supplier_type_assoc', methods=['POST'])
@_login_required
def api_pm_sup_assoc_save():
    """Save supplier type associations for pm_stock page (admin only)."""
    if session.get('User_Type', '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    d = request.get_json() or {}
    type_ids = d.get('type_ids', [])   # list of ints to associate
    conn = sampling_portal.get_db_connection()
    try:
        conn.execute("DELETE FROM pm_supplier_type_assoc WHERE page='pm_stock'")
        for tid in type_ids:
            conn.execute(
                "INSERT IGNORE INTO pm_supplier_type_assoc (supplier_type_id, page) VALUES (%s,'pm_stock')",
                (int(tid),)
            )
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'saved': len(type_ids)})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/suppliers/debug')
@_login_required
def api_pm_suppliers_debug():
    """Dumps raw contents of the 3 supplier-related tables for diagnosis."""
    conn = sampling_portal.get_db_connection()
    try:
        types = [dict(r) for r in conn.execute(
            "SELECT id, type_name FROM supplier_type ORDER BY id"
        ).fetchall()]
        assoc = [dict(r) for r in conn.execute(
            "SELECT id, supplier_type_id, page FROM pm_supplier_type_assoc"
        ).fetchall()]
        suppliers = [dict(r) for r in conn.execute(
            """SELECT s.id, s.supplier_name, s.supplier_type_id, s.status,
                      COALESCE(st.type_name,'(no type)') AS supplier_type_name
               FROM procurement_suppliers s
               LEFT JOIN supplier_type st ON st.id = s.supplier_type_id
               ORDER BY s.id"""
        ).fetchall()]
        # Build lookup
        assoc_for_pm = [a['supplier_type_id'] for a in assoc if a.get('page') == 'pm_stock']
        conn.close()
        return jsonify({
            'status':       'ok',
            'supplier_types':              types,
            'pm_supplier_type_assoc':      assoc,
            'assoc_ids_for_pm_stock':      assoc_for_pm,
            'suppliers':                   suppliers,
            'totals': {
                'types':      len(types),
                'assoc':      len(assoc),
                'assoc_pm':   len(assoc_for_pm),
                'suppliers':  len(suppliers),
            }
        })
    except Exception as e:
        conn.close()
        import traceback
        return jsonify({'status': 'error', 'message': str(e), 'trace': traceback.format_exc()}), 500

@pm_stock_bp.route('/api/pm_stock/suppliers')
@_login_required
def api_pm_suppliers_list():
    """
    Return suppliers whose type is associated with pm_stock page.
    If no suppliers match the strict filters, progressively widens the query
    and returns a diagnostic field explaining what was tried.
    """
    search = (request.args.get('search') or '').strip().lower()
    conn = sampling_portal.get_db_connection()
    try:
        # Gather association info
        assoc = [r['supplier_type_id'] for r in conn.execute(
            "SELECT supplier_type_id FROM pm_supplier_type_assoc WHERE page='pm_stock'"
        ).fetchall()]

        # Count total suppliers for diagnostic
        total_all = conn.execute("SELECT COUNT(*) AS c FROM procurement_suppliers").fetchone()
        total_all = total_all['c'] if total_all else 0

        diagnostic = {
            'assoc_count':       len(assoc),
            'total_suppliers':   total_all,
            'filter_used':       'none',
        }

        rows = []

        # ── Path 1: strict — associated types AND "PM" in type name
        if assoc:
            ph = ','.join(['%s'] * len(assoc))
            rows = conn.execute(f"""
                SELECT s.*, st.type_name AS supplier_type_name
                FROM procurement_suppliers s
                LEFT JOIN supplier_type st ON st.id = s.supplier_type_id
                WHERE s.supplier_type_id IN ({ph})
                  AND UPPER(COALESCE(st.type_name,'')) LIKE '%%PM%%'
                ORDER BY s.supplier_name ASC
            """, assoc).fetchall()
            diagnostic['filter_used'] = 'assoc + PM-name'

        # ── Path 2: association by ID only (relax name filter)
        if not rows and assoc:
            ph = ','.join(['%s'] * len(assoc))
            rows = conn.execute(f"""
                SELECT s.*, st.type_name AS supplier_type_name
                FROM procurement_suppliers s
                LEFT JOIN supplier_type st ON st.id = s.supplier_type_id
                WHERE s.supplier_type_id IN ({ph})
                ORDER BY s.supplier_name ASC
            """, assoc).fetchall()
            diagnostic['filter_used'] = 'assoc-only (name filter relaxed)'

        # ── Path 3: no association set — fall back to PM-named types
        if not rows and not assoc:
            rows = conn.execute("""
                SELECT s.*, st.type_name AS supplier_type_name
                FROM procurement_suppliers s
                LEFT JOIN supplier_type st ON st.id = s.supplier_type_id
                WHERE UPPER(COALESCE(st.type_name,'')) LIKE '%%PM%%'
                ORDER BY s.supplier_name ASC
            """).fetchall()
            diagnostic['filter_used'] = 'PM-name (no association set)'

        # ── Path 4: last-resort — ALL suppliers
        # Only when no rows from any filtered path AND DB actually has suppliers.
        # This prevents the "empty directory" UX when associations/types are misconfigured.
        if not rows and total_all > 0:
            rows = conn.execute("""
                SELECT s.*, st.type_name AS supplier_type_name
                FROM procurement_suppliers s
                LEFT JOIN supplier_type st ON st.id = s.supplier_type_id
                ORDER BY s.supplier_name ASC
            """).fetchall()
            diagnostic['filter_used']  = 'all-suppliers (fallback — configure Associate Types)'
            diagnostic['needs_config'] = True

        # Serialise + apply search filter
        result = []
        for r in rows:
            d = dict(r)
            for f in ('created_at', 'updated_at'):
                if d.get(f): d[f] = str(d[f])
            if search:
                hay = ' '.join([
                    str(d.get('supplier_name')    or ''),
                    str(d.get('contact_person')   or ''),
                    str(d.get('supplier_type_name') or ''),
                    str(d.get('gst_number')       or ''),
                ]).lower()
                if search not in hay:
                    continue
            result.append(d)

        conn.close()
        return jsonify({
            'status':     'ok',
            'suppliers':  result,
            'diagnostic': diagnostic,
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/suppliers/save', methods=['POST'])
@_login_required
def api_pm_suppliers_save():
    """Insert or update a supplier in procurement_suppliers."""
    d = request.get_json() or {}
    sup_id   = d.get('id')
    name     = (d.get('supplier_name') or '').strip()
    if not name:
        return jsonify({'status': 'error', 'message': 'Supplier name required'}), 400

    fields = {
        'supplier_name':    name,
        'contact_person':   (d.get('contact_person') or '').strip() or None,
        'phone':            (d.get('phone') or '').strip() or None,
        'email':            (d.get('email') or '').strip() or None,
        'address':          (d.get('address') or '').strip() or None,
        'gst_number':       (d.get('gst_number') or '').strip().upper() or None,
        'pan_number':       (d.get('pan_number') or '').strip().upper() or None,
        'payment_terms':    (d.get('payment_terms') or '').strip() or None,
        'payment_type':     (d.get('payment_type') or '').strip() or None,
        'credit_days':      int(d['credit_days']) if d.get('credit_days') else None,
        'currency':         d.get('currency') or 'INR',
        'lead_time_days':   int(d['lead_time_days']) if d.get('lead_time_days') else None,
        'moq':              float(d['moq']) if d.get('moq') else None,
        'rating':           int(d['rating']) if d.get('rating') else None,
        'status':           d.get('status') or 'active',
        'supplier_type_id': int(d['supplier_type_id']) if d.get('supplier_type_id') else None,
        'updated_by':       _user(),
    }

    conn = sampling_portal.get_db_connection()
    try:
        if sup_id:
            set_clause = ', '.join(f"{k}=%s" for k in fields)
            conn.execute(
                f"UPDATE procurement_suppliers SET {set_clause} WHERE id=%s",
                list(fields.values()) + [int(sup_id)]
            )
        else:
            cols = ', '.join(fields.keys())
            ph   = ', '.join(['%s'] * len(fields))
            conn.execute(
                f"INSERT INTO procurement_suppliers ({cols}) VALUES ({ph})",
                list(fields.values())
            )
        conn.commit()
        row = conn.execute(
            "SELECT id FROM procurement_suppliers WHERE supplier_name=%s", (name,)
        ).fetchone()
        conn.close()
        return jsonify({'status': 'ok', 'id': row['id'] if row else None})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/suppliers/delete', methods=['POST'])
@_login_required
def api_pm_suppliers_delete():
    """Delete a supplier. Admin-only as of Phase 1. Phase 2 will route through
    the recycle bin for restorability."""
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only — non-admins cannot delete suppliers.'}), 403
    d = request.get_json() or {}
    sup_id = d.get('id')
    if not sup_id:
        return jsonify({'status': 'error', 'message': 'id required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        conn.execute("DELETE FROM procurement_suppliers WHERE id=%s", (int(sup_id),))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/suppliers/pm_grn_ledger')
@_login_required
def api_pm_supplier_ledger():
    """Return PM GRNs for a supplier (for PM Stock supplier ledger)."""
    supplier = (request.args.get('supplier') or '').strip()
    if not supplier:
        return jsonify({'status': 'error', 'message': 'supplier required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        grns = conn.execute("""
            SELECT g.id, g.grn_no, g.grn_date, g.po_number, g.remarks,
                   g.created_by, COALESCE(gd.name,'') AS location,
                   COUNT(gi.id) AS item_count,
                   COALESCE(SUM(gi.qty_received),0) AS total_qty
            FROM pm_grn g
            LEFT JOIN procurement_godowns gd ON gd.id = g.godown_id
            LEFT JOIN pm_grn_items gi ON gi.grn_id = g.id
            WHERE LOWER(TRIM(g.supplier)) = LOWER(TRIM(%s))
            GROUP BY g.id
            ORDER BY g.grn_date DESC, g.created_at DESC
        """, (supplier,)).fetchall()
        result = []
        for r in grns:
            d2 = dict(r)
            if hasattr(d2.get('grn_date'), 'isoformat'): d2['grn_date'] = d2['grn_date'].isoformat()
            result.append(d2)
        conn.close()
        return jsonify({'status': 'ok', 'grns': result})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/export_excel', methods=['POST'])
@_login_required
def api_pm_export_excel():
    """Export PM stock summary to Excel using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'status': 'error', 'message': 'openpyxl not installed on server'}), 500

    d            = request.get_json() or {}
    scope        = d.get('scope', 'combined')   # godown | floor | combined
    to_date      = d.get('to_date') or str(date.today())
    from_date    = d.get('from_date') or None
    search       = (d.get('search') or '').strip().lower()
    pm_type_f    = (d.get('pm_type') or '').strip()
    brand_id_f   = d.get('brand_id') or None
    stock_filter = (d.get('stock_filter') or '').strip()
    ids          = d.get('ids') or []

    conn = sampling_portal.get_db_connection()
    try:
        godown_data = _godown_summary(conn, from_date=from_date, to_date=to_date)
        floor_data  = _floor_summary(conn, from_date=from_date, to_date=to_date)

        # Build brand map
        brands = {r['id']: r['name'] for r in conn.execute(
            "SELECT id, name FROM procurement_brands"
        ).fetchall()} if _table_exists(conn, 'procurement_brands') else {}

        # Build per-godown stock map: product_id -> {godown_id: stock}
        # Reuses the same logic as api_pm_summary_per_godown
        storage_godowns = [g for g in _get_godowns(conn)
                          if g.get('godown_type','godown') not in ('floor','billing','shipping')
                          and not g.get('is_floor')]
        pg_rows = conn.execute("""
            SELECT
                t.product_id, t.godown_id,
                COALESCE(SUM(CASE WHEN t.txn_type='opening' THEN t.qty ELSE 0 END),0) AS op,
                COALESCE(SUM(CASE WHEN t.txn_type='inward'  THEN t.qty ELSE 0 END),0) AS inward,
                COALESCE(SUM(CASE WHEN t.txn_type='outward' THEN t.qty ELSE 0 END),0) AS outward
            FROM pm_godown_txn t
            WHERE (%s IS NULL OR t.txn_date <= %s)
            GROUP BY t.product_id, t.godown_id
        """, (to_date, to_date)).fetchall()
        per_godown = {}  # product_id -> {godown_id: stock}
        for r in pg_rows:
            pid = r['product_id']
            stock = float(r['op']) + float(r['inward']) - float(r['outward'])
            per_godown.setdefault(pid, {})[r['godown_id']] = round(stock, 3)
    finally:
        conn.close()

    # Merge godown + floor into combined
    floor_idx = {r['id']: r for r in floor_data}
    combined = []
    for g in godown_data:
        f = floor_idx.get(g['id'], {})
        row = {**g,
            'floor_op':  f.get('floor_op', 0),
            'issue':     f.get('issue', 0),
            'dispatch':  f.get('dispatch', 0),
            'rejection': f.get('rejection', 0),
            'pm_return': f.get('pm_return', 0),
            'remaining': f.get('remaining', 0),
            'floor_last_txn': f.get('last_txn_date', ''),
        }
        row['total_stock'] = row['godown_stock'] + row['remaining']
        row['brand_name']  = brands.get(g.get('brand_id'), '')
        # Per-godown breakdown — one key per godown.
        # Legacy NULL-godown opening balances are attributed to the first godown to avoid double-counting.
        pg = per_godown.get(g['id'], {})
        for idx, gd in enumerate(storage_godowns):
            base = pg.get(gd['id'], 0)
            if idx == 0:
                base += pg.get(None, 0)  # fold legacy NULL-godown stock into first godown
            row[f'gd_{gd["id"]}'] = round(base, 3)
        combined.append(row)

    # Apply filters
    if ids:           combined = [r for r in combined if r['id'] in ids]
    if search:
        # Search matches product name, PM type, OR product code (case-insensitive).
        combined = [r for r in combined
                    if search in r['product_name'].lower()
                    or search in r['pm_type'].lower()
                    or search in (r.get('product_code') or '').lower()]
    if pm_type_f:     combined = [r for r in combined if r['pm_type'].lower() == pm_type_f.lower()]
    if brand_id_f:
        try:
            bid = int(brand_id_f)
            combined = [r for r in combined if r.get('brand_id') == bid]
        except: pass
    if stock_filter == 'nonzero':      combined = [r for r in combined if r['godown_stock'] > 0]
    elif stock_filter == 'zero':       combined = [r for r in combined if r['godown_stock'] <= 0]
    elif stock_filter == 'has_inward': combined = [r for r in combined if r['inward'] > 0]
    elif stock_filter == 'has_outward':combined = [r for r in combined if r['outward'] > 0]
    elif stock_filter == 'has_movement':combined = [r for r in combined if r['inward'] > 0 or r['outward'] > 0]

    # Build workbook
    wb = openpyxl.Workbook()
    hdr_fill  = PatternFill('solid', fgColor='0D9488')
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    num_align = Alignment(horizontal='right', vertical='center')
    thin      = Side(style='thin', color='D1D5DB')
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def make_sheet(ws, rows, cols):
        ws.row_dimensions[1].height = 28
        for ci, (hdr, key, width, is_num) in enumerate(cols, 1):
            c = ws.cell(1, ci, hdr)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = hdr_align; c.border = bdr
            ws.column_dimensions[get_column_letter(ci)].width = width
        for ri, r in enumerate(rows, 2):
            for ci, (hdr, key, width, is_num) in enumerate(cols, 1):
                val = r.get(key, 0 if is_num else '')
                c = ws.cell(ri, ci, val)
                c.border = bdr
                if is_num:
                    c.alignment = num_align
                    c.number_format = '#,##0.###'
            ws.row_dimensions[ri].height = 16

    GODOWN_COLS = [
        ('#',           'row',          4,  True),
        ('Brand',       'brand_name',   14, False),
        ('Product',     'product_name', 40, False),
        ('PM Type',     'pm_type',      14, False),
        ('Opening',     'op',           12, True),
        ('Inward',      'inward',       12, True),
        ('Outward',     'outward',      12, True),
        ('Godown Stock','godown_stock', 14, True),
        ('Min Stock',   'min_stock',    10, True),
        ('Last Txn',    'last_txn_date',12, False),
    ]
    FLOOR_COLS = [
        ('#',           'row',          4,  True),
        ('Brand',       'brand_name',   14, False),
        ('Product',     'product_name', 40, False),
        ('PM Type',     'pm_type',      14, False),
        ('Factory OP',  'floor_op',     12, True),
        ('Issued',      'issue',        12, True),
        ('Dispatched',  'dispatch',     12, True),
        ('Rejection',   'rejection',    12, True),
        ('PM Return',   'pm_return',    12, True),
        ('Remaining',   'remaining',    12, True),
    ]
    # Build COMBINED_COLS dynamically — one col per storage godown
    gd_cols = [(f'🏢 {g["name"]}', f'gd_{g["id"]}', 14, True) for g in storage_godowns]
    COMBINED_COLS = [
        ('#',           'row',          4,  True),
        ('Brand',       'brand_name',   14, False),
        ('Product',     'product_name', 40, False),
        ('PM Type',     'pm_type',      14, False),
        *gd_cols,
        ('Total Godown','godown_stock', 14, True),
        ('Factory Rem.','remaining',    14, True),
        ('Total Stock', 'total_stock',  14, True),
        ('Last Txn',    'last_txn_date',12, False),
    ]

    # Add row numbers
    for i, r in enumerate(combined, 1): r['row'] = i

    if scope in ('godown', 'combined'):
        ws = wb.active; ws.title = 'Godown Stock'
        ws.freeze_panes = 'E2'
        make_sheet(ws, combined, GODOWN_COLS)

    if scope in ('floor', 'combined'):
        ws2 = wb.create_sheet('Factory Stock')
        ws2.freeze_panes = 'E2'
        make_sheet(ws2, combined, FLOOR_COLS)

    if scope == 'combined':
        ws3 = wb.create_sheet('Combined')
        ws3.freeze_panes = 'E2'
        make_sheet(ws3, combined, COMBINED_COLS)

    # Remove default empty sheet if other sheets were created
    if scope != 'godown' and 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'PM_Stock_{to_date}.xlsx'
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@pm_stock_bp.route('/api/pm_stock/mtv/list')
@_login_required
def api_mtv_list():
    from_date = request.args.get('from_date') or ''
    to_date   = request.args.get('to_date')   or ''
    search    = (request.args.get('search') or '').strip()

    conn = sampling_portal.get_db_connection()
    params = []
    where  = []
    if from_date:
        where.append("m.mtv_date >= %s"); params.append(from_date)
    if to_date:
        where.append("m.mtv_date <= %s"); params.append(to_date)
    if search:
        where.append("m.mtv_no LIKE %s")
        params.append(f'%{search}%')

    wc = ("WHERE " + " AND ".join(where)) if where else ""
    rows = conn.execute(f"""
        SELECT m.id, m.mtv_no, m.mtv_date, m.from_godown, m.to_godown,
               m.from_type, m.to_type, m.remarks, m.created_by, m.created_at,
               fg.name AS from_name, tg.name AS to_name,
               COUNT(mi.id) AS item_count,
               COALESCE(SUM(mi.qty), 0) AS total_qty
        FROM pm_mtv m
        LEFT JOIN procurement_godowns fg ON fg.id = m.from_godown
        LEFT JOIN procurement_godowns tg ON tg.id = m.to_godown
        LEFT JOIN pm_mtv_items mi ON mi.mtv_id = m.id
        {wc}
        GROUP BY m.id
        ORDER BY m.mtv_date DESC, m.created_at DESC
    """, params).fetchall()
    conn.close()

    result = []
    for r in rows:
        d = dict(r)
        if hasattr(d.get('mtv_date'), 'isoformat'):
            d['mtv_date'] = d['mtv_date'].isoformat()
        if hasattr(d.get('created_at'), 'isoformat'):
            d['created_at'] = str(d['created_at'])
        result.append(d)
    return jsonify(result)

@pm_stock_bp.route('/api/pm_stock/mtv/save', methods=['POST'])
@_login_required
def api_mtv_save():
    """
    Create a Material Transfer Voucher.
    Moves stock: debit from_godown outward, credit to_godown inward (or floor equivalent).
    Body: {
        mtv_date, from_godown, to_godown, from_type, to_type, remarks,
        items: [{product_id, qty, remarks}, ...]
    }
    """
    blocked = _block_if_disabled('mtv')
    if blocked is not None: return blocked
    # Requesters can't create vouchers — they only submit Material Requests.
    blocked = _block_if_requester()
    if blocked is not None: return blocked
    # New-voucher-entries access check (per-user access control)
    blocked = _block_if_no_access('new_voucher_entries')
    if blocked is not None: return blocked
    d          = request.get_json() or {}
    mtv_date   = d.get('mtv_date')    or str(date.today())
    from_gd    = d.get('from_godown')
    to_gd      = d.get('to_godown')
    from_type  = d.get('from_type', 'godown')   # godown | floor
    to_type    = d.get('to_type', 'godown')      # godown | floor
    remarks    = (d.get('remarks') or '').strip()
    items      = d.get('items', [])

    if not from_gd or not to_gd:
        return jsonify({'status': 'error', 'message': 'from_godown and to_godown required'}), 400
    if not items:
        return jsonify({'status': 'error', 'message': 'At least one item required'}), 400
    if str(from_gd) == str(to_gd) and from_type == to_type:
        return jsonify({'status': 'error', 'message': 'Source and destination are the same'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        mtv_no = _next_voucher_no(conn, 'PM-MTV', mtv_date)
        cur = conn.execute(
            """INSERT INTO pm_mtv (mtv_no, mtv_date, from_godown, to_godown, from_type, to_type, remarks, created_by)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
            (mtv_no, mtv_date, from_gd, to_gd, from_type, to_type, remarks, _user())
        )
        mtv_id = cur.lastrowid

        for item in items:
            pid = item.get('product_id')
            qty = float(item.get('qty') or 0)
            irm = (item.get('remarks') or '').strip()
            if not pid or qty <= 0:
                continue
            conn.execute(
                "INSERT INTO pm_mtv_items (mtv_id, product_id, qty, remarks) VALUES (%s,%s,%s,%s)",
                (mtv_id, pid, qty, irm)
            )
            mtv_rem = f"[MTV: {mtv_no}] {irm}".strip()
            # Debit: outward from source
            if from_type == 'godown':
                vno = _next_voucher_no(conn, 'PM-GTXN', mtv_date)
                conn.execute(
                    """INSERT INTO pm_godown_txn (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
                       VALUES (%s,%s,'outward',%s,%s,%s,%s,%s)""",
                    (pid, mtv_date, qty, mtv_rem, _user(), vno, from_gd)
                )
            else:  # floor outward = dispatch
                vno = _next_voucher_no(conn, 'PM-FTXN', mtv_date)
                conn.execute(
                    """INSERT INTO pm_floor_txn (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
                       VALUES (%s,%s,'dispatch',%s,%s,%s,%s,%s)""",
                    (pid, mtv_date, qty, mtv_rem, _user(), vno, from_gd)
                )
            # Credit: inward to destination
            if to_type == 'godown':
                vno2 = _next_voucher_no(conn, 'PM-GTXN', mtv_date)
                conn.execute(
                    """INSERT INTO pm_godown_txn (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
                       VALUES (%s,%s,'inward',%s,%s,%s,%s,%s)""",
                    (pid, mtv_date, qty, mtv_rem, _user(), vno2, to_gd)
                )
            else:  # floor inward = issue
                vno2 = _next_voucher_no(conn, 'PM-FTXN', mtv_date)
                conn.execute(
                    """INSERT INTO pm_floor_txn (product_id, txn_date, txn_type, qty, remarks, created_by, voucher_no, godown_id)
                       VALUES (%s,%s,'issue',%s,%s,%s,%s,%s)""",
                    (pid, mtv_date, qty, mtv_rem, _user(), vno2, to_gd)
                )

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'mtv_no': mtv_no, 'id': mtv_id})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/mtv/<int:mtv_id>')
@_login_required
def api_mtv_detail(mtv_id):
    conn = sampling_portal.get_db_connection()
    m = conn.execute("""
        SELECT m.*, fg.name AS from_name, tg.name AS to_name
        FROM pm_mtv m
        LEFT JOIN procurement_godowns fg ON fg.id=m.from_godown
        LEFT JOIN procurement_godowns tg ON tg.id=m.to_godown
        WHERE m.id=%s
    """, (mtv_id,)).fetchone()
    if not m: conn.close(); return jsonify({'status':'error','message':'MTV not found'}), 404
    items = conn.execute("""
        SELECT mi.product_id, p.product_name, p.pm_type, mi.qty, mi.remarks
        FROM pm_mtv_items mi JOIN pm_products p ON p.id=mi.product_id
        WHERE mi.mtv_id=%s
    """, (mtv_id,)).fetchall()
    conn.close()
    d = dict(m)
    for k in ('mtv_date','created_at'):
        if d.get(k) and hasattr(d[k],'isoformat'): d[k] = d[k].isoformat() if k!='created_at' else str(d[k])
    d['items'] = [dict(i) for i in items]
    return jsonify(d)

@pm_stock_bp.route('/api/pm_stock/mtv/update', methods=['POST'])
@_login_required
def api_mtv_update():
    d         = request.get_json() or {}
    mtv_id    = d.get('id')
    mtv_date  = d.get('mtv_date') or str(date.today())
    from_gd   = d.get('from_godown')
    to_gd     = d.get('to_godown')
    from_type = d.get('from_type', 'godown')   # godown | floor
    to_type   = d.get('to_type',   'godown')   # godown | floor
    remarks   = (d.get('remarks') or '').strip()
    items     = d.get('items', [])

    if not all([mtv_id, from_gd, to_gd, items]):
        return jsonify({'status': 'error', 'message': 'id, from_godown, to_godown, items required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        mtv = conn.execute("SELECT mtv_no, from_type, to_type FROM pm_mtv WHERE id=%s", (mtv_id,)).fetchone()
        if not mtv:
            conn.close()
            return jsonify({'status': 'error', 'message': 'MTV not found'}), 404
        mtv_no         = mtv['mtv_no']
        old_from_type  = mtv['from_type']  or 'godown'
        old_to_type    = mtv['to_type']    or 'godown'

        # ── Reverse old movements (delete from correct tables) ────────────────
        old_items = conn.execute(
            "SELECT product_id, qty FROM pm_mtv_items WHERE mtv_id=%s", (mtv_id,)
        ).fetchall()
        for o in old_items:
            pid = o['product_id']; qty = o['qty']
            rem_pat = f'%{mtv_no}%'
            # Reverse source debit
            if old_from_type == 'godown':
                conn.execute(
                    "DELETE FROM pm_godown_txn WHERE product_id=%s AND qty=%s AND txn_type='outward' AND remarks LIKE %s ORDER BY id DESC LIMIT 1",
                    (pid, qty, rem_pat))
            else:
                conn.execute(
                    "DELETE FROM pm_floor_txn WHERE product_id=%s AND qty=%s AND txn_type='dispatch' AND remarks LIKE %s ORDER BY id DESC LIMIT 1",
                    (pid, qty, rem_pat))
            # Reverse destination credit
            if old_to_type == 'godown':
                conn.execute(
                    "DELETE FROM pm_godown_txn WHERE product_id=%s AND qty=%s AND txn_type='inward' AND remarks LIKE %s ORDER BY id DESC LIMIT 1",
                    (pid, qty, rem_pat))
            else:
                conn.execute(
                    "DELETE FROM pm_floor_txn WHERE product_id=%s AND qty=%s AND txn_type='issue' AND remarks LIKE %s ORDER BY id DESC LIMIT 1",
                    (pid, qty, rem_pat))

        # ── Update MTV header ─────────────────────────────────────────────────
        conn.execute(
            "UPDATE pm_mtv SET mtv_date=%s, from_godown=%s, to_godown=%s, from_type=%s, to_type=%s, remarks=%s WHERE id=%s",
            (mtv_date, from_gd, to_gd, from_type, to_type, remarks, mtv_id)
        )
        conn.execute("DELETE FROM pm_mtv_items WHERE mtv_id=%s", (mtv_id,))

        # ── Re-insert corrected movements ─────────────────────────────────────
        for item in items:
            pid = item.get('product_id'); qty = float(item.get('qty') or 0)
            if not pid or qty <= 0:
                continue
            conn.execute(
                "INSERT INTO pm_mtv_items (mtv_id, product_id, qty, remarks) VALUES (%s,%s,%s,'')",
                (mtv_id, pid, qty)
            )
            rem = f'[MTV:{mtv_no}]'
            # Debit source
            if from_type == 'godown':
                vno1 = _next_voucher_no(conn, 'PM-GTXN', mtv_date)
                conn.execute(
                    "INSERT INTO pm_godown_txn (product_id,txn_date,txn_type,qty,remarks,created_by,voucher_no,godown_id) VALUES (%s,%s,'outward',%s,%s,%s,%s,%s)",
                    (pid, mtv_date, qty, rem, _user(), vno1, from_gd))
            else:
                vno1 = _next_voucher_no(conn, 'PM-FTXN', mtv_date)
                conn.execute(
                    "INSERT INTO pm_floor_txn (product_id,txn_date,txn_type,qty,remarks,created_by,voucher_no,godown_id) VALUES (%s,%s,'dispatch',%s,%s,%s,%s,%s)",
                    (pid, mtv_date, qty, rem, _user(), vno1, from_gd))
            # Credit destination
            if to_type == 'godown':
                vno2 = _next_voucher_no(conn, 'PM-GTXN', mtv_date)
                conn.execute(
                    "INSERT INTO pm_godown_txn (product_id,txn_date,txn_type,qty,remarks,created_by,voucher_no,godown_id) VALUES (%s,%s,'inward',%s,%s,%s,%s,%s)",
                    (pid, mtv_date, qty, rem, _user(), vno2, to_gd))
            else:
                vno2 = _next_voucher_no(conn, 'PM-FTXN', mtv_date)
                conn.execute(
                    "INSERT INTO pm_floor_txn (product_id,txn_date,txn_type,qty,remarks,created_by,voucher_no,godown_id) VALUES (%s,%s,'issue',%s,%s,%s,%s,%s)",
                    (pid, mtv_date, qty, rem, _user(), vno2, to_gd))

        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'mtv_no': mtv_no})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/mtv/delete', methods=['POST'])
@_login_required
def api_mtv_delete():
    """Soft-delete a legacy MTV. Snapshots the parent + line items + auto-posted
    ledger txns into the recycle bin. Admin-only. Restorable from the bin."""
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only — non-admins cannot delete MTV vouchers.'}), 403
    d = request.get_json() or {}; mtv_id = d.get('id')
    reason = (d.get('reason') or '').strip() or None
    if not mtv_id: return jsonify({'status':'error','message':'id required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        mtv = conn.execute("SELECT mtv_no, mtv_date, from_type, to_type FROM pm_mtv WHERE id=%s",(mtv_id,)).fetchone()
        if not mtv: conn.close(); return jsonify({'status':'error','message':'MTV not found'}), 404
        mtv_no    = mtv['mtv_no']
        from_type = mtv['from_type']  or 'godown'
        to_type   = mtv['to_type']    or 'godown'

        # Snapshot ALL ledger rows tagged with this voucher number, regardless
        # of where they landed. Bulk-pattern is safer than the old per-item
        # DELETE … LIMIT 1 logic which could leave duplicates behind.
        bin_id = _bin_soft_delete(
            conn,
            entity_type  = 'mtv',
            entity_id    = mtv_id,
            entity_label = f"MTV {mtv_no}",
            parent_table = 'pm_mtv',
            parent_where = 'id=%s',
            parent_params= (mtv_id,),
            children = [
                {'table': 'pm_mtv_items',
                 'where': 'mtv_id=%s', 'params': (mtv_id,)},
                {'table': 'pm_godown_txn',
                 'where': "remarks LIKE %s",
                 'params': (f'%{mtv_no}%',)},
                {'table': 'pm_floor_txn',
                 'where': "remarks LIKE %s",
                 'params': (f'%{mtv_no}%',)},
            ],
            summary = f"Legacy MTV {mtv_no} · {from_type} → {to_type} · dated {mtv.get('mtv_date')}",
            reason  = reason
        )
        conn.commit(); conn.close()
        return jsonify({'status':'ok','mtv_no':mtv_no, 'bin_id': bin_id, 'message': 'MTV moved to recycle bin.'})
    except Exception as e:
        conn.close(); return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/mtv/repair', methods=['GET', 'POST'])
@_login_required
def api_mtv_repair():
    """
    Full repair for MTVs going to a floor/factory destination.

    Handles TWO cases:
      Case A — wrong inward row exists in pm_godown_txn: delete it, insert as pm_floor_txn issue
      Case B — wrong inward row already deleted but pm_floor_txn issue is missing:
               rebuild it from pm_mtv_items

    Floor location detection: uses pm_mtv.to_type='floor' (already fixed by previous runs)
    OR any location that has issue/dispatch entries in pm_floor_txn.
    """
    conn = sampling_portal.get_db_connection()
    try:
        # Detect floor location IDs from actual data
        floor_ids = set()
        for r in conn.execute(
            "SELECT DISTINCT to_godown FROM pm_mtv WHERE to_type='floor'"
        ).fetchall():
            floor_ids.add(r['to_godown'])
        for r in conn.execute(
            "SELECT DISTINCT godown_id FROM pm_floor_txn WHERE txn_type IN ('issue','dispatch') AND godown_id IS NOT NULL"
        ).fetchall():
            floor_ids.add(r['godown_id'])

        # Get all MTVs going to a floor location
        all_mtvs = conn.execute("""
            SELECT m.id, m.mtv_no, m.to_godown, m.mtv_date, m.to_type
            FROM pm_mtv m
            WHERE m.to_type = 'floor'
            ORDER BY m.id
        """).fetchall()

        # Also include MTVs where to_godown is in floor_ids even if to_type not set
        if floor_ids:
            ph = ','.join(['%s'] * len(floor_ids))
            extra = conn.execute(f"""
                SELECT m.id, m.mtv_no, m.to_godown, m.mtv_date, m.to_type
                FROM pm_mtv m
                WHERE m.to_godown IN ({ph}) AND m.to_type != 'floor'
                ORDER BY m.id
            """, list(floor_ids)).fetchall()
            all_mtvs = list(all_mtvs) + list(extra)

        fixed_mtvs = 0
        rows_inserted = 0
        rows_deleted  = 0
        detail = []

        for mtv in all_mtvs:
            mtv_id   = mtv['id']
            mtv_no   = mtv['mtv_no']
            to_gd    = mtv['to_godown']
            mtv_date = mtv['mtv_date']

            # Case A: delete any wrong inward entries still in pm_godown_txn
            wrong = conn.execute("""
                SELECT id, product_id, qty, created_by, voucher_no, txn_date
                FROM pm_godown_txn
                WHERE txn_type='inward' AND godown_id=%s AND remarks LIKE %s
            """, (to_gd, f'%{mtv_no}%')).fetchall()

            for w in wrong:
                conn.execute("DELETE FROM pm_godown_txn WHERE id=%s", (w['id'],))
                rows_deleted += 1

            # Case B: ensure every item in pm_mtv_items has a floor issue entry
            items = conn.execute("""
                SELECT i.product_id, i.qty
                FROM pm_mtv_items i
                WHERE i.mtv_id = %s
            """, (mtv_id,)).fetchall()

            inserted_this = 0
            for item in items:
                pid = item['product_id']
                qty = item['qty']

                exists = conn.execute("""
                    SELECT id FROM pm_floor_txn
                    WHERE product_id=%s AND txn_type='issue'
                      AND godown_id=%s AND remarks LIKE %s
                """, (pid, to_gd, f'%{mtv_no}%')).fetchone()

                if not exists:
                    conn.execute("""
                        INSERT INTO pm_floor_txn
                          (product_id, txn_date, txn_type, qty, remarks,
                           created_by, voucher_no, godown_id)
                        VALUES (%s, %s, 'issue', %s, %s, %s, %s, %s)
                    """, (pid, mtv_date, qty,
                          f'[MTV:{mtv_no}]',
                          'System Repair', None, to_gd))
                    inserted_this += 1
                    rows_inserted += 1

            if wrong or inserted_this:
                fixed_mtvs += 1
                detail.append({
                    'mtv_no':   mtv_no,
                    'deleted':  len(wrong),
                    'inserted': inserted_this,
                })

            # Ensure to_type is 'floor' in pm_mtv header
            conn.execute(
                "UPDATE pm_mtv SET to_type='floor' WHERE id=%s", (mtv_id,))

        conn.commit()
        conn.close()
        return jsonify({
            'status':        'ok',
            'fixed_mtvs':    fixed_mtvs,
            'rows_deleted':  rows_deleted,
            'rows_inserted': rows_inserted,
            'detail':        detail,
        })
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/mtv/repair_preview', methods=['GET'])
@_login_required
def api_mtv_repair_preview():
    """
    Diagnostic: shows what the repair WOULD do without changing anything.
    Visit this URL in the browser to verify before running the actual repair.
    """
    conn = sampling_portal.get_db_connection()
    try:
        all_mtvs = conn.execute("""
            SELECT m.id, m.mtv_no, m.to_godown, m.mtv_date, m.to_type,
                   COALESCE(g.type,'godown') AS dest_type, g.name AS dest_name
            FROM pm_mtv m
            LEFT JOIN procurement_godowns g ON g.id = m.to_godown
            ORDER BY m.id
        """).fetchall()

        preview = []
        for mtv in all_mtvs:
            wrong_rows = conn.execute("""
                SELECT g.id, g.product_id, g.qty
                FROM pm_godown_txn g
                WHERE g.txn_type = 'inward'
                  AND g.godown_id = %s
                  AND g.remarks LIKE %s
            """, (mtv['to_godown'], f'%{mtv["mtv_no"]}%')).fetchall()

            preview.append({
                'mtv_no':     mtv['mtv_no'],
                'to_godown':  mtv['to_godown'],
                'dest_name':  mtv['dest_name'],
                'dest_type':  mtv['dest_type'],
                'to_type_db': mtv['to_type'],
                'wrong_inward_rows': len(wrong_rows),
                'rows': [{'product_id': r['product_id'], 'qty': float(r['qty'])} for r in wrong_rows],
            })

        conn.close()
        return jsonify({'status': 'ok', 'mtvs': preview})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/grn/<int:grn_id>/reprint_labels_selective',
                   methods=['POST'])
@_login_required
def api_grn_reprint_labels_selective(grn_id):
    """Selectively reprint a subset of a GRN line item's box labels.

    Body: {
      product_id: int,
      selections: [
        {box_id: int|null, box_seq: int, per_box_qty: float, is_new: bool}
      ],
      removed_box_ids: [int],   # existing boxes to delete (unchecked)
      reason: str (optional)
    }

    Constraint: sum(selections.per_box_qty) MUST equal the existing total
    qty for this GRN line (sum of pm_boxes.per_box_qty), so the godown
    inward stock and pm_grn_items.qty_received don't need any adjustment.

    Authorisation:
      - Admins: allowed unconditionally.
      - Non-admins: allowed when called with a valid `request_token`
        from an *approved* reprint request (i.e. the user has clicked
        "Print Now" on an approved request). The endpoint marks the
        request as `printed` and clears the token on success.
    """
    d = request.get_json(silent=True) or {}
    request_token = (d.get('request_token') or '').strip() or None
    request_req_id = d.get('request_req_id')

    # Authorise: admin OR valid approved-request token
    authorised_via_request = None     # set to (req_id, conn-less row) if matched
    if not _is_admin():
        if not request_token or not request_req_id:
            return jsonify({'status':'error',
                'message':'Selective reprint is admin-only. Use the "Request Reprint" flow and have an admin approve it.'}), 403
        # Validate token against the request — this happens inside the same
        # connection as the apply, just below, so the token check + the
        # delete + the rebuild are all in one transaction.

    try:
        product_id = int(d.get('product_id'))
    except Exception:
        return jsonify({'status':'error','message':'product_id is required'}), 400

    selections = d.get('selections') or []
    removed    = d.get('removed_box_ids') or []
    if not isinstance(selections, list) or not selections:
        return jsonify({'status':'error','message':'At least one selection is required'}), 400
    if not isinstance(removed, list):
        removed = []

    # Validate every selection has a positive per_box_qty
    norm_sel = []
    for s in selections:
        try:
            pbq = float(s.get('per_box_qty') or 0)
            seq = int(s.get('box_seq') or 0)
        except Exception:
            return jsonify({'status':'error','message':'Invalid per_box_qty/box_seq in selections'}), 400
        if pbq <= 0:
            return jsonify({'status':'error','message':f'Selection box_seq={seq} has non-positive per_box_qty'}), 400
        norm_sel.append({
            'box_id':      (int(s['box_id']) if s.get('box_id') is not None else None),
            'box_seq':     seq,
            'per_box_qty': pbq,
            'is_new':      bool(s.get('is_new')),
        })

    new_total_selected = sum(x['per_box_qty'] for x in norm_sel)

    conn = sampling_portal.get_db_connection()
    try:
        # If non-admin used a request_token, validate it before any work.
        # The token must reference an *approved* request whose voucher matches
        # this GRN. We DON'T clear the token yet — that happens after the
        # commit succeeds, so a partial/failed apply leaves the request in
        # a usable state for retry.
        if not _is_admin():
            req_row = conn.execute(
                """SELECT req_id, status, scope_type, voucher_kind, voucher_id,
                          requested_by, product_id
                   FROM pm_label_reprint_requests
                   WHERE req_id=%s AND print_token=%s""",
                (int(request_req_id), request_token)
            ).fetchone()
            if not req_row:
                conn.close()
                return jsonify({'status':'error','message':'Invalid request_token / req_id'}), 403
            if req_row['status'] != 'approved':
                conn.close()
                return jsonify({'status':'error',
                    'message': f"Request status is '{req_row['status']}'. Only approved requests can be applied."}), 409
            if req_row['requested_by'] != _user():
                conn.close()
                return jsonify({'status':'error',
                    'message':'Only the original requester can apply this request.'}), 403
            if req_row['scope_type'] != 'voucher_grn' or req_row['voucher_kind'] != 'grn' \
               or int(req_row['voucher_id'] or 0) != int(grn_id) \
               or int(req_row['product_id'] or 0) != int(product_id):
                conn.close()
                return jsonify({'status':'error',
                    'message':'Request token does not match this voucher/product.'}), 403
            authorised_via_request = int(req_row['req_id'])

        # Load GRN header
        grn = conn.execute(
            """SELECT id, grn_no, grn_date, supplier, godown_id
               FROM pm_grn WHERE id=%s""", (grn_id,)
        ).fetchone()
        if not grn:
            conn.close(); return jsonify({'status':'error','message':'GRN not found'}), 404
        grn_no    = grn['grn_no']
        grn_date  = grn['grn_date']
        godown_id = grn['godown_id']

        # Load product
        prod = conn.execute(
            """SELECT product_name, COALESCE(product_code,'') AS code,
                      COALESCE(pm_type,'') AS pm_type
               FROM pm_products WHERE id=%s""", (product_id,)
        ).fetchone()
        if not prod or not prod['code']:
            conn.close()
            return jsonify({'status':'error','message':'Product not found or has no product code'}), 404
        brand = conn.execute(
            """SELECT COALESCE(b.name,'') AS name FROM pm_products p
               LEFT JOIN procurement_brands b ON b.id = p.brand_id
               WHERE p.id=%s""", (product_id,)
        ).fetchone()
        brand_name = brand['name'] if brand else ''

        # Load existing boxes for this GRN+product
        existing = conn.execute(
            """SELECT box_id, box_seq, box_code, per_box_qty, current_status,
                      total_boxes, current_godown_id
               FROM pm_boxes
               WHERE grn_id=%s AND product_id=%s
               ORDER BY box_seq""",
            (grn_id, product_id)
        ).fetchall()
        if not existing:
            conn.close()
            return jsonify({'status':'error','message':'No boxes found for this GRN line'}), 404

        existing_total = sum(float(b['per_box_qty'] or 0) for b in existing)
        existing_by_id  = {int(b['box_id']): dict(b) for b in existing}
        existing_seqs   = {int(b['box_seq'] or 0) for b in existing}
        max_existing_seq = max(existing_seqs) if existing_seqs else 0

        # Validate constraint: total qty must match
        if abs(new_total_selected - existing_total) > 0.001:
            conn.close()
            return jsonify({
                'status':'error',
                'message': (f'Total qty must match: selected={new_total_selected:g}, '
                            f'original={existing_total:g}. Adjust per-box qty so the totals match.')
            }), 400

        # Validate that removed_box_ids belong to this GRN+product and aren't OUT
        bad_removed = []
        for bid in removed:
            try: bid = int(bid)
            except Exception: continue
            if bid not in existing_by_id:
                bad_removed.append(bid); continue
            if (existing_by_id[bid].get('current_status') or '') == 'out':
                conn.close()
                return jsonify({'status':'error',
                    'message': f'Cannot remove box_id={bid} - it has already moved out (status=out). Cancel/restore the transfer first.'
                }), 400
        if bad_removed:
            conn.close()
            return jsonify({'status':'error',
                'message': f'removed_box_ids contains boxes that do not belong to this GRN line: {bad_removed}'
            }), 400

        # Validate kept (selected, not new) box_ids exist and aren't OUT
        kept_ids = []
        new_to_create = []
        for s in norm_sel:
            if s['is_new']:
                new_to_create.append(s)
            else:
                if s['box_id'] is None or int(s['box_id']) not in existing_by_id:
                    conn.close()
                    return jsonify({'status':'error',
                        'message': f'Selection references unknown box_id={s.get("box_id")}'
                    }), 400
                if (existing_by_id[int(s['box_id'])].get('current_status') or '') == 'out':
                    conn.close()
                    return jsonify({'status':'error',
                        'message': f'Box id={s["box_id"]} is OUT - cannot reprint/edit. Cancel/restore the transfer first.'
                    }), 400
                kept_ids.append(int(s['box_id']))

        # Snapshot all affected rows for recycle bin BEFORE mutating
        bin_id = _bin_soft_delete(
            conn,
            entity_type  = 'grn_item_selective_edit',
            entity_id    = grn_id,
            entity_label = f"GRN {grn_no} | {prod['product_name']} (selective reprint)",
            parent_table = 'pm_boxes',
            parent_where = 'grn_id=%s AND product_id=%s',
            parent_params= (grn_id, product_id),
            children = [
                {'table': 'pm_boxes',
                 'where': 'grn_id=%s AND product_id=%s',
                 'params': (grn_id, product_id)},
                {'table': 'pm_box_movements',
                 'where': 'box_id IN (SELECT box_id FROM pm_boxes WHERE grn_id=%s AND product_id=%s)',
                 'params': (grn_id, product_id)},
            ],
            summary = (f'Selective reprint of GRN {grn_no} ({prod["product_name"]}): '
                       f'{len(existing)} existing, {len(removed)} removed, '
                       f'{len(new_to_create)} added, {len(kept_ids)} kept'),
            reason  = (d.get('reason') or '')[:500] or None
        )

        # _bin_soft_delete deletes parent + all children. Recreate the
        # final box population:
        #   - kept-or-untouched existing boxes (with edited pbq if selected)
        #   - new boxes from is_new=True selections
        #   - skip removed ones

        line = conn.execute(
            "SELECT id FROM pm_grn_items WHERE grn_id=%s AND product_id=%s LIMIT 1",
            (grn_id, product_id)
        ).fetchone()
        grn_item_id = int(line['id']) if line else 0

        kept_pbq_overrides = {
            int(s['box_id']): s['per_box_qty']
            for s in norm_sel if (not s['is_new']) and s.get('box_id') is not None
        }
        removed_set = {int(b) for b in removed}

        rebuild_plan = []
        for b in existing:
            if int(b['box_id']) in removed_set:
                continue
            seq  = int(b['box_seq'] or 0)
            code = b['box_code']
            pbq  = float(b['per_box_qty'] or 0)
            if int(b['box_id']) in kept_pbq_overrides:
                pbq = kept_pbq_overrides[int(b['box_id'])]
            rebuild_plan.append({
                'box_seq':     seq,
                'box_code':    code,
                'per_box_qty': pbq,
                'is_new':      False,
                'orig_status': b.get('current_status') or 'in_stock',
                'is_selected': int(b['box_id']) in kept_pbq_overrides,
            })

        next_seq = max_existing_seq
        for s in new_to_create:
            requested = int(s.get('box_seq') or 0)
            seqs_in_plan = {p['box_seq'] for p in rebuild_plan}
            if requested > max_existing_seq and requested not in seqs_in_plan:
                seq = requested
            else:
                next_seq += 1
                while next_seq in seqs_in_plan:
                    next_seq += 1
                seq = next_seq
            new_code = _make_box_code(prod['code'], grn_no, seq)
            rebuild_plan.append({
                'box_seq':     seq,
                'box_code':    new_code,
                'per_box_qty': float(s['per_box_qty']),
                'is_new':      True,
                'orig_status': 'in_stock',
                'is_selected': True,
            })

        rebuild_plan.sort(key=lambda x: x['box_seq'])
        final_total_boxes = len(rebuild_plan)

        # Recreate pm_boxes + 'grn_create' movements
        for p in rebuild_plan:
            cur = conn.execute(
                """INSERT INTO pm_boxes
                     (box_code, grn_id, grn_no, grn_item_id, product_id, product_code,
                      box_seq, total_boxes, per_box_qty, current_godown_id,
                      current_status, created_by)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (p['box_code'], grn_id, grn_no, grn_item_id, product_id, prod['code'],
                 p['box_seq'], final_total_boxes, p['per_box_qty'], godown_id,
                 p['orig_status'], _user())
            )
            box_id = cur.lastrowid
            _assign_box_short_code(conn, box_id)
            note = ('Selective reprint - new box ' if p['is_new']
                    else 'Selective reprint - ' +
                         ('edited ' if p['is_selected'] else 'preserved '))
            conn.execute(
                """INSERT INTO pm_box_movements
                     (box_id, movement_type, to_godown_id, qty, moved_by, remarks)
                   VALUES (%s, 'grn_create', %s, %s, %s, %s)""",
                (box_id, godown_id, p['per_box_qty'], _user(),
                 f'{note}{grn_no} box {p["box_seq"]}/{final_total_boxes}')
            )

        # Update pm_grn_items.no_of_box, box_count (per-box qty lives in box_count;
        # qty_received itself unchanged because the total physical qty doesn't shift
        # on a pure box-count edit).
        pbqs_in_plan = [p['per_box_qty'] for p in rebuild_plan]
        uniform_pbq = pbqs_in_plan[0] if pbqs_in_plan and all(
            abs(x - pbqs_in_plan[0]) < 0.001 for x in pbqs_in_plan
        ) else 0
        conn.execute(
            """UPDATE pm_grn_items
               SET no_of_box=%s, box_count=%s
               WHERE id=%s""",
            (final_total_boxes,
             int(uniform_pbq) if isinstance(uniform_pbq, (int, float)) and float(uniform_pbq).is_integer() else uniform_pbq,
             grn_item_id)
        )

        # Build print payload — only the selected rows
        printed = [p for p in rebuild_plan if p['is_selected']]
        printed.sort(key=lambda x: x['box_seq'])
        box_codes    = [p['box_code'] for p in printed]
        per_box_qtys = [p['per_box_qty'] for p in printed]
        box_seqs     = [p['box_seq']  for p in printed]   # for "N/total" labels

        # Parallel short_codes (8-char compact form) for the QR payload.
        # New boxes created above just got fresh short_codes; legacy rows
        # may have NULL — we fetch in one round trip and align by box_code.
        short_codes = []
        if box_codes:
            sc_rows = conn.execute(
                "SELECT box_code, short_code FROM pm_boxes "
                "WHERE grn_id=%s AND product_id=%s",
                (grn_id, product_id)
            ).fetchall()
            sc_map = {r['box_code']: (r.get('short_code') or '') for r in sc_rows}
            short_codes = [sc_map.get(bc, '') for bc in box_codes]

        # FIFO code — already assigned at GRN-create time. Look up.
        fifo_code = ''
        try:
            fi = _get_or_assign_fifo(conn, 'grn', grn_id, product_id)
            fifo_code = fi.get('fifo_code', '') or ''
        except Exception:
            pass

        # If this call was authorised by a request token, mark that
        # request as printed so the token can't be reused.
        if authorised_via_request is not None:
            conn.execute(
                """UPDATE pm_label_reprint_requests
                   SET status='printed', printed_at=NOW(), printed_by=%s, print_token=NULL
                   WHERE req_id=%s""",
                (_user(), authorised_via_request)
            )

        # Log the selective reprint event
        try:
            _label_print_record(
                conn,
                print_kind='grn_selective',
                voucher_kind='grn',
                voucher_id=grn_id,
                voucher_no=grn_no,
                product_id=product_id,
                label_count=len(box_codes),
                box_codes=box_codes,
                request_id=authorised_via_request,
            )
        except Exception: pass

        conn.commit()
        conn.close()

        return jsonify({
            'status':       'ok',
            'grn_id':       grn_id,
            'grn_no':       grn_no,
            'grn_date':     str(grn_date) if grn_date else '',
            'godown_id':    godown_id,
            'product_id':   product_id,
            'product_code': prod['code'],
            'product_name': prod['product_name'],
            'pm_type':      prod['pm_type'] or '',
            'brand_name':   brand_name,
            'no_of_box':    final_total_boxes,
            'box_codes':    box_codes,
            'short_codes':  short_codes,
            'per_box_qtys': per_box_qtys,
            'box_seqs':     box_seqs,
            'total_boxes':  final_total_boxes,
            'bin_id':       bin_id,
            'fifo_code':    fifo_code,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/op_batches/<int:op_seq>/reprint_selective',
                   methods=['POST'])
@_login_required
def api_op_batch_reprint_selective(op_seq):
    """Selectively reprint a subset of an Opening Stock batch's box labels.

    Body: {
      product_id: int,
      godown_id:  int,
      selections: [{box_id, box_seq, per_box_qty, is_new}],
      removed_box_ids: [int],
      reason: str (optional)
    }

    Authorisation:
      - Admins: allowed unconditionally.
      - Non-admins: allowed when called with a valid `request_token`
        from an *approved* reprint request (i.e. the user has clicked
        "Print Now" on an approved request). The endpoint marks the
        request as `printed` and clears the token on success.
    """
    d = request.get_json(silent=True) or {}
    request_token = (d.get('request_token') or '').strip() or None
    request_req_id = d.get('request_req_id')

    authorised_via_request = None
    if not _is_admin():
        if not request_token or not request_req_id:
            return jsonify({'status':'error',
                'message':'Selective reprint is admin-only. Use the "Request Reprint" flow and have an admin approve it.'}), 403
        # Token is validated inside the conn block below.

    try:
        product_id = int(d.get('product_id'))
        godown_id  = int(d.get('godown_id'))
    except Exception:
        return jsonify({'status':'error','message':'product_id and godown_id are required'}), 400

    selections = d.get('selections') or []
    removed    = d.get('removed_box_ids') or []
    if not isinstance(selections, list) or not selections:
        return jsonify({'status':'error','message':'At least one selection is required'}), 400
    if not isinstance(removed, list):
        removed = []

    norm_sel = []
    for s in selections:
        try:
            pbq = float(s.get('per_box_qty') or 0)
            seq = int(s.get('box_seq') or 0)
        except Exception:
            return jsonify({'status':'error','message':'Invalid per_box_qty/box_seq in selections'}), 400
        if pbq <= 0:
            return jsonify({'status':'error','message':f'Selection box_seq={seq} has non-positive per_box_qty'}), 400
        norm_sel.append({
            'box_id':      (int(s['box_id']) if s.get('box_id') is not None else None),
            'box_seq':     seq,
            'per_box_qty': pbq,
            'is_new':      bool(s.get('is_new')),
        })

    op_label = f'PM-OP/{op_seq:04d}'
    new_total_selected = sum(x['per_box_qty'] for x in norm_sel)

    conn = sampling_portal.get_db_connection()
    try:
        # Token validation for non-admin callers
        if not _is_admin():
            req_row = conn.execute(
                """SELECT req_id, status, scope_type, voucher_kind, voucher_id,
                          requested_by, product_id, godown_id
                   FROM pm_label_reprint_requests
                   WHERE req_id=%s AND print_token=%s""",
                (int(request_req_id), request_token)
            ).fetchone()
            if not req_row:
                conn.close()
                return jsonify({'status':'error','message':'Invalid request_token / req_id'}), 403
            if req_row['status'] != 'approved':
                conn.close()
                return jsonify({'status':'error',
                    'message': f"Request status is '{req_row['status']}'. Only approved requests can be applied."}), 409
            if req_row['requested_by'] != _user():
                conn.close()
                return jsonify({'status':'error',
                    'message':'Only the original requester can apply this request.'}), 403
            if req_row['scope_type'] != 'voucher_op' or req_row['voucher_kind'] != 'op' \
               or int(req_row['voucher_id'] or 0) != int(op_seq) \
               or int(req_row['product_id'] or 0) != int(product_id) \
               or int(req_row['godown_id'] or 0) != int(godown_id):
                conn.close()
                return jsonify({'status':'error',
                    'message':'Request token does not match this OP batch / product / godown.'}), 403
            authorised_via_request = int(req_row['req_id'])

        prod = conn.execute(
            """SELECT product_name, COALESCE(product_code,'') AS code,
                      COALESCE(pm_type,'') AS pm_type
               FROM pm_products WHERE id=%s""", (product_id,)
        ).fetchone()
        if not prod or not prod['code']:
            conn.close()
            return jsonify({'status':'error','message':'Product not found or has no product code'}), 404
        brand = conn.execute(
            """SELECT COALESCE(b.name,'') AS name FROM pm_products p
               LEFT JOIN procurement_brands b ON b.id = p.brand_id
               WHERE p.id=%s""", (product_id,)
        ).fetchone()
        brand_name = brand['name'] if brand else ''

        existing = conn.execute(
            """SELECT box_id, box_seq, box_code, per_box_qty, current_status,
                      total_boxes
               FROM pm_boxes
               WHERE grn_no=%s AND product_id=%s AND current_godown_id=%s
               ORDER BY box_seq""",
            (op_label, product_id, godown_id)
        ).fetchall()
        if not existing:
            conn.close()
            return jsonify({'status':'error','message':'No boxes found for this OP batch'}), 404

        existing_total   = sum(float(b['per_box_qty'] or 0) for b in existing)
        existing_by_id   = {int(b['box_id']): dict(b) for b in existing}
        max_existing_seq = max((int(b['box_seq'] or 0) for b in existing), default=0)

        if abs(new_total_selected - existing_total) > 0.001:
            conn.close()
            return jsonify({
                'status':'error',
                'message': (f'Total qty must match: selected={new_total_selected:g}, '
                            f'original={existing_total:g}. Adjust per-box qty so the totals match.')
            }), 400

        for bid in removed:
            try: bid = int(bid)
            except Exception: continue
            if bid not in existing_by_id:
                conn.close()
                return jsonify({'status':'error',
                    'message': f'removed_box_ids contains box {bid} that does not belong to this OP batch'
                }), 400
            if (existing_by_id[bid].get('current_status') or '') == 'out':
                conn.close()
                return jsonify({'status':'error',
                    'message': f'Cannot remove box_id={bid} - status=out. Cancel/restore the transfer first.'
                }), 400

        kept_pbq_overrides = {}
        new_to_create = []
        for s in norm_sel:
            if s['is_new']:
                new_to_create.append(s)
            else:
                if s['box_id'] is None or int(s['box_id']) not in existing_by_id:
                    conn.close()
                    return jsonify({'status':'error',
                        'message': f'Selection references unknown box_id={s.get("box_id")}'
                    }), 400
                if (existing_by_id[int(s['box_id'])].get('current_status') or '') == 'out':
                    conn.close()
                    return jsonify({'status':'error',
                        'message': f'Box id={s["box_id"]} is OUT - cannot reprint/edit.'
                    }), 400
                kept_pbq_overrides[int(s['box_id'])] = s['per_box_qty']

        bin_id = _bin_soft_delete(
            conn,
            entity_type  = 'op_batch_selective_edit',
            entity_id    = op_seq,
            entity_label = f"OP {op_label} | {prod['product_name']} (selective reprint)",
            parent_table = 'pm_boxes',
            parent_where = 'grn_no=%s AND product_id=%s AND current_godown_id=%s',
            parent_params= (op_label, product_id, godown_id),
            children = [
                {'table': 'pm_boxes',
                 'where': 'grn_no=%s AND product_id=%s AND current_godown_id=%s',
                 'params': (op_label, product_id, godown_id)},
                {'table': 'pm_box_movements',
                 'where': 'box_id IN (SELECT box_id FROM pm_boxes WHERE grn_no=%s AND product_id=%s AND current_godown_id=%s)',
                 'params': (op_label, product_id, godown_id)},
            ],
            summary = (f'Selective reprint of OP {op_label} ({prod["product_name"]}): '
                       f'{len(existing)} existing, {len(removed)} removed, '
                       f'{len(new_to_create)} added, {len(kept_pbq_overrides)} edited'),
            reason  = (d.get('reason') or '')[:500] or None
        )

        removed_set = {int(b) for b in removed}
        rebuild_plan = []
        for b in existing:
            if int(b['box_id']) in removed_set:
                continue
            seq = int(b['box_seq'] or 0)
            pbq = float(b['per_box_qty'] or 0)
            if int(b['box_id']) in kept_pbq_overrides:
                pbq = kept_pbq_overrides[int(b['box_id'])]
            rebuild_plan.append({
                'box_seq':     seq,
                'box_code':    b['box_code'],
                'per_box_qty': pbq,
                'is_new':      False,
                'orig_status': b.get('current_status') or 'in_stock',
                'is_selected': int(b['box_id']) in kept_pbq_overrides,
            })
        next_seq = max_existing_seq
        for s in new_to_create:
            requested = int(s.get('box_seq') or 0)
            seqs_in_plan = {p['box_seq'] for p in rebuild_plan}
            if requested > max_existing_seq and requested not in seqs_in_plan:
                seq = requested
            else:
                next_seq += 1
                while next_seq in seqs_in_plan:
                    next_seq += 1
                seq = next_seq
            new_code = _make_op_box_code(prod['code'], op_seq, seq)
            rebuild_plan.append({
                'box_seq':     seq,
                'box_code':    new_code,
                'per_box_qty': float(s['per_box_qty']),
                'is_new':      True,
                'orig_status': 'in_stock',
                'is_selected': True,
            })
        rebuild_plan.sort(key=lambda x: x['box_seq'])
        final_total_boxes = len(rebuild_plan)

        for p in rebuild_plan:
            cur = conn.execute(
                """INSERT INTO pm_boxes
                     (box_code, grn_id, grn_no, grn_item_id, product_id, product_code,
                      box_seq, total_boxes, per_box_qty, current_godown_id,
                      current_status, created_by)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (p['box_code'], 0, op_label, 0, product_id, prod['code'],
                 p['box_seq'], final_total_boxes, p['per_box_qty'], godown_id,
                 p['orig_status'], _user())
            )
            box_id = cur.lastrowid
            _assign_box_short_code(conn, box_id)
            note = ('Selective reprint - new box ' if p['is_new']
                    else 'Selective reprint - ' +
                         ('edited ' if p['is_selected'] else 'preserved '))
            conn.execute(
                """INSERT INTO pm_box_movements
                     (box_id, movement_type, to_godown_id, qty, moved_by, remarks)
                   VALUES (%s, 'adjust', %s, %s, %s, %s)""",
                (box_id, godown_id, p['per_box_qty'], _user(),
                 f'{note}{op_label} box {p["box_seq"]}/{final_total_boxes}')
            )

        # Opening ledger row (pm_godown_txn / pm_floor_txn) is NOT touched.
        # Total stays the same by constraint, so existing opening txn is correct.

        printed = [p for p in rebuild_plan if p['is_selected']]
        printed.sort(key=lambda x: x['box_seq'])
        box_codes    = [p['box_code'] for p in printed]
        per_box_qtys = [p['per_box_qty'] for p in printed]
        box_seqs     = [p['box_seq']  for p in printed]

        # Parallel short_codes (8-char compact form) aligned to box_codes.
        # Fetch in one round trip and index by box_code.
        short_codes = []
        if box_codes:
            sc_rows = conn.execute(
                "SELECT box_code, short_code FROM pm_boxes "
                "WHERE grn_no=%s AND product_id=%s",
                (op_label, product_id)
            ).fetchall()
            sc_map = {r['box_code']: (r.get('short_code') or '') for r in sc_rows}
            short_codes = [sc_map.get(bc, '') for bc in box_codes]

        # FIFO code — already assigned at OP-create time.
        fifo_code = ''
        try:
            fi = _get_or_assign_fifo(conn, 'op', op_seq, product_id)
            fifo_code = fi.get('fifo_code', '') or ''
        except Exception:
            pass

        # If this call was authorised by a request token, mark that
        # request as printed so the token can't be reused.
        if authorised_via_request is not None:
            conn.execute(
                """UPDATE pm_label_reprint_requests
                   SET status='printed', printed_at=NOW(), printed_by=%s, print_token=NULL
                   WHERE req_id=%s""",
                (_user(), authorised_via_request)
            )

        # Log the selective reprint event
        try:
            _label_print_record(
                conn,
                print_kind='op_selective',
                voucher_kind='op',
                voucher_id=op_seq,
                voucher_no=op_label,
                product_id=product_id,
                label_count=len(box_codes),
                box_codes=box_codes,
                request_id=authorised_via_request,
            )
        except Exception: pass

        conn.commit()
        conn.close()

        return jsonify({
            'status':       'ok',
            'op_seq':       op_seq,
            'op_label':     op_label,
            'godown_id':    godown_id,
            'product_id':   product_id,
            'product_code': prod['code'],
            'product_name': prod['product_name'],
            'pm_type':      prod['pm_type'] or '',
            'brand_name':   brand_name,
            'no_of_box':    final_total_boxes,
            'box_codes':    box_codes,
            'short_codes':  short_codes,
            'per_box_qtys': per_box_qtys,
            'box_seqs':     box_seqs,
            'total_boxes':  final_total_boxes,
            'bin_id':       bin_id,
            'fifo_code':    fifo_code,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/audit/summary', methods=['GET'])
@_login_required
def api_audit_summary():
    """Headline counts for the Audit Report dashboard."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    f, t = _parse_date_range(request.args)
    conn = sampling_portal.get_db_connection()
    try:
        # Products created (master data)
        prod_row = conn.execute(
            """SELECT COUNT(*) AS c FROM pm_products
               WHERE created_at >= %s AND created_at < (DATE_ADD(%s, INTERVAL 1 DAY))""",
            (f, t)
        ).fetchone()
        products_created = int((prod_row or {}).get('c') or 0)

        # Labels printed (by kind)
        label_rows = conn.execute(
            """SELECT print_kind, COUNT(*) AS runs, COALESCE(SUM(label_count),0) AS labels
               FROM pm_label_print_log
               WHERE ts >= %s AND ts < (DATE_ADD(%s, INTERVAL 1 DAY))
               GROUP BY print_kind""",
            (f, t)
        ).fetchall()
        labels_by_kind = {r['print_kind']: {'runs': int(r['runs']),
                                            'labels': int(r['labels'])} for r in label_rows}
        labels_total = sum(v['labels'] for v in labels_by_kind.values())

        # GRN count (created via grn.create audit OR direct count of pm_grn rows)
        grn_row = conn.execute(
            """SELECT COUNT(*) AS c FROM pm_grn
               WHERE created_at >= %s AND created_at < (DATE_ADD(%s, INTERVAL 1 DAY))""",
            (f, t)
        ).fetchone()
        grn_count = int((grn_row or {}).get('c') or 0)

        # Transfer counts — break out into started (Material OUT) and confirmed
        # (Material IN). pm_transfers status flow: out_started → in_pending → received.
        try:
            mt_started = conn.execute(
                """SELECT COUNT(*) AS c FROM pm_transfers
                   WHERE created_at >= %s AND created_at < (DATE_ADD(%s, INTERVAL 1 DAY))""",
                (f, t)
            ).fetchone()
            mt_received = conn.execute(
                """SELECT COUNT(*) AS c FROM pm_transfers
                   WHERE received_at >= %s AND received_at < (DATE_ADD(%s, INTERVAL 1 DAY))
                     AND status='received'""",
                (f, t)
            ).fetchone()
        except Exception:
            mt_started = mt_received = None
        material_out_count = int((mt_started or {}).get('c') or 0) if mt_started else 0
        material_in_count  = int((mt_received or {}).get('c') or 0) if mt_received else 0

        # OP batches created (from audit log)
        op_row = conn.execute(
            """SELECT COUNT(*) AS c FROM pm_audit_log
               WHERE action='op.create' AND ts >= %s AND ts < (DATE_ADD(%s, INTERVAL 1 DAY))""",
            (f, t)
        ).fetchone()
        op_batches_created = int((op_row or {}).get('c') or 0)

        # Edits volume by action
        edits_rows = conn.execute(
            """SELECT action, COUNT(*) AS c FROM pm_audit_log
               WHERE ts >= %s AND ts < (DATE_ADD(%s, INTERVAL 1 DAY))
               GROUP BY action ORDER BY c DESC""",
            (f, t)
        ).fetchall()
        edits_by_action = [{'action': r['action'], 'count': int(r['c'])} for r in edits_rows]
        edits_total = sum(e['count'] for e in edits_by_action)

        # Top editors
        editor_rows = conn.execute(
            """SELECT user_name, COUNT(*) AS c FROM pm_audit_log
               WHERE ts >= %s AND ts < (DATE_ADD(%s, INTERVAL 1 DAY))
                 AND user_name <> ''
               GROUP BY user_name ORDER BY c DESC LIMIT 10""",
            (f, t)
        ).fetchall()
        top_editors = [{'user': r['user_name'], 'count': int(r['c'])} for r in editor_rows]

        conn.close()
        return jsonify({
            'status': 'ok',
            'from': str(f), 'to': str(t),
            'products_created':   products_created,
            'labels_total':       labels_total,
            'labels_by_kind':     labels_by_kind,
            'grn_count':          grn_count,
            'material_out_count': material_out_count,
            'material_in_count':  material_in_count,
            'op_batches_created': op_batches_created,
            'edits_total':        edits_total,
            'edits_by_action':    edits_by_action,
            'top_editors':        top_editors,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/audit/feed', methods=['GET'])
@_login_required
def api_audit_feed():
    """Paginated audit log feed with optional filters."""
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    f, t = _parse_date_range(request.args)
    user_filter   = (request.args.get('user') or '').strip()
    action_filter = (request.args.get('action') or '').strip()
    entity_filter = (request.args.get('entity') or '').strip()
    try: page = max(1, int(request.args.get('page') or 1))
    except Exception: page = 1
    try: size = min(200, max(10, int(request.args.get('size') or 50)))
    except Exception: size = 50

    where = ['ts >= %s', 'ts < (DATE_ADD(%s, INTERVAL 1 DAY))']
    params = [f, t]
    if user_filter:
        where.append('user_name = %s'); params.append(user_filter)
    if action_filter:
        where.append('action = %s'); params.append(action_filter)
    if entity_filter:
        where.append('entity = %s'); params.append(entity_filter)
    where_sql = ' AND '.join(where)

    conn = sampling_portal.get_db_connection()
    try:
        total = int((conn.execute(
            f"SELECT COUNT(*) AS c FROM pm_audit_log WHERE {where_sql}", params
        ).fetchone() or {}).get('c') or 0)

        offset = (page - 1) * size
        rows = conn.execute(
            f"""SELECT id, ts, user_name, action, entity, entity_id, summary,
                       route_path, reversal_class, reversed_at, reversed_by,
                       reversal_audit_id
                FROM pm_audit_log
                WHERE {where_sql}
                ORDER BY id DESC
                LIMIT %s OFFSET %s""",
            params + [size, offset]
        ).fetchall()
        out = []
        for r in rows:
            d = dict(r)
            if d.get('ts')          and hasattr(d['ts'], 'isoformat'):          d['ts']          = d['ts'].isoformat()
            if d.get('reversed_at') and hasattr(d['reversed_at'], 'isoformat'): d['reversed_at'] = d['reversed_at'].isoformat()
            out.append(d)

        conn.close()
        return jsonify({
            'status': 'ok',
            'rows': out, 'total': total,
            'page': page, 'size': size,
            'pages': (total + size - 1) // size if size else 1,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/audit_log/<int:audit_id>', methods=['GET'])
@_login_required
def api_audit_detail(audit_id):
    """Full audit-log row (pm_audit_log) including before_json + after_json.

    URL changed from /api/pm_stock/audit/<id> to /api/pm_stock/audit_log/<id>
    in June 2026 because it was colliding with the audit-session detail
    endpoint /api/pm_stock/audit/<sid> (different feature: pm_audit_sessions,
    not pm_audit_log). Flask routes /api/pm_stock/audit/5 to whichever
    blueprint registered first — so the audit-session modal was getting
    back this endpoint's response shape (`{row:{...}}`) instead of the
    expected `{session:{...}}`, breaking the modal with
    "Cannot read properties of undefined (reading 'session_no')".

    No frontend code calls this route — pm_stock_admin.js only uses /feed,
    /summary, and /<id>/reverse — so renaming is safe.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM pm_audit_log WHERE id=%s", (audit_id,)
        ).fetchone()
        conn.close()
        if not row:
            return jsonify({'status':'error','message':'Not found'}), 404
        d = dict(row)
        if d.get('ts')          and hasattr(d['ts'], 'isoformat'):          d['ts']          = d['ts'].isoformat()
        if d.get('reversed_at') and hasattr(d['reversed_at'], 'isoformat'): d['reversed_at'] = d['reversed_at'].isoformat()
        return jsonify({'status':'ok','row': d})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500

@pm_stock_bp.route('/api/pm_stock/audit/<int:audit_id>/reverse', methods=['POST'])
@_login_required
def api_audit_reverse(audit_id):
    """
    Reverse one audit entry.

    Phase C — full implementation lands in turns 4–5. For now, this
    endpoint:
      • Always requires admin
      • Refuses to reverse rows already reversed
      • Refuses 'final' rows outright (those can't be reversed)
      • Performs a real revert for SAFE entries:
          product.update / product.threshold / product.brand_assign /
          product.code_regen / user_home.set / user_home.delete /
          voucher_seq.update
        — restores `before_json` to the source row.
      • For 'gated' entries, returns a 501 telling the caller to
        wait until turn 5 wires up data revert.
    """
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    import json as _json
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM pm_audit_log WHERE id=%s", (audit_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'status':'error','message':'Audit row not found'}), 404
        if row.get('reversed_at'):
            conn.close()
            return jsonify({'status':'error',
                'message':f"Already reversed at {row['reversed_at']} by {row.get('reversed_by') or '?'}"}), 409
        rclass = row.get('reversal_class') or 'final'
        if rclass == 'final':
            conn.close()
            return jsonify({'status':'error',
                'message':'This action is final and cannot be reversed (printed labels, confirmed receipts, etc.).'}), 400

        action  = row.get('action') or ''
        entity  = row.get('entity') or ''
        eid     = row.get('entity_id') or ''
        before  = _json.loads(row['before_json']) if row.get('before_json') else None
        after   = _json.loads(row['after_json'])  if row.get('after_json')  else None

        if rclass == 'gated':
            conn.close()
            return jsonify({'status':'error',
                'code':'reversal_not_yet_implemented',
                'message':'Stock-affecting reversals are not enabled yet — coming in a follow-up update.'}), 501

        # ── SAFE reversals ───────────────────────────────────────────────
        applied = False
        if action == 'product.update' and before:
            conn.execute(
                "UPDATE pm_products SET product_name=%s, pm_type=%s, brand_id=%s, product_code=%s WHERE id=%s",
                (before.get('product_name'), before.get('pm_type'),
                 before.get('brand_id'), before.get('product_code'), eid)
            )
            applied = True
        elif action == 'product.threshold' and before:
            conn.execute("UPDATE pm_products SET min_stock=%s WHERE id=%s",
                         (int(before.get('min_stock', 0) or 0), eid))
            applied = True
        elif action == 'product.brand_assign' and before:
            conn.execute("UPDATE pm_products SET brand_id=%s WHERE id=%s",
                         (before.get('brand_id'), eid))
            applied = True
        elif action == 'product.code_regen' and before:
            conn.execute("UPDATE pm_products SET product_code=%s WHERE id=%s",
                         (before.get('product_code'), eid))
            applied = True
        elif action == 'product.create':
            # Creating something is reversed by deactivating it (soft-delete).
            conn.execute("UPDATE pm_products SET is_active=0 WHERE id=%s", (eid,))
            applied = True
        elif action == 'user_home.set':
            if before:
                conn.execute("""
                    INSERT INTO pm_user_home_godown (user_name, godown_id, note, updated_by)
                    VALUES (%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE
                        godown_id=VALUES(godown_id),
                        note=VALUES(note),
                        updated_by=VALUES(updated_by)
                """, (eid, before.get('godown_id'), before.get('note'), _user()))
            else:
                # Was a fresh creation — reverse means delete
                conn.execute("DELETE FROM pm_user_home_godown WHERE user_name=%s", (eid,))
            applied = True
        elif action == 'user_home.delete' and before:
            conn.execute("""
                INSERT INTO pm_user_home_godown (user_name, godown_id, note, updated_by)
                VALUES (%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE
                    godown_id=VALUES(godown_id),
                    note=VALUES(note),
                    updated_by=VALUES(updated_by)
            """, (eid, before.get('godown_id'), before.get('note'), _user()))
            applied = True
        elif action == 'voucher_seq.update' and before:
            conn.execute("""
                UPDATE pm_voucher_sequences
                SET prefix=%s, pad_digits=%s, reset_yearly=%s
                WHERE voucher_type=%s
            """, (before.get('prefix'), int(before.get('pad_digits') or 4),
                  int(before.get('reset_yearly') or 1), eid))
            applied = True

        if not applied:
            conn.close()
            return jsonify({'status':'error',
                'code':'reversal_unsupported',
                'message':f'Reversal not implemented for action={action} (yet).'}), 501

        # Stamp original audit row as reversed
        rev_audit_id = _audit_record(
            conn,
            action='audit.reversed',
            entity=entity,
            entity_id=eid,
            summary=f"Reversed audit #{audit_id} (action={action})",
            before=after,    # the "current" state we're undoing
            after=before,    # what we restored to
        )
        conn.execute(
            """UPDATE pm_audit_log
               SET reversed_at=NOW(), reversed_by=%s, reversal_audit_id=%s
               WHERE id=%s""",
            (_user(), rev_audit_id, audit_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','reversed_audit_id': rev_audit_id})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500



# ═══════════════════════════════════════════════════════════════════════════
# FIFO SETTINGS — admin-only
# ═══════════════════════════════════════════════════════════════════════════
# GET  /api/pm_stock/settings/fifo
#      → { enabled: bool, start_date: 'YYYY-MM-DD' or null, updated_by, updated_at }
# POST /api/pm_stock/settings/fifo
#      Body: { enabled?: bool, start_date?: 'YYYY-MM-DD' or null,
#              reset_from_now?: bool }
#      • enabled       → toggles FIFO enforcement on/off
#      • start_date    → set lots assigned before this date as out-of-FIFO
#      • reset_from_now → shortcut: sets start_date to today
#      All admin only.
# ═══════════════════════════════════════════════════════════════════════════

# ── General PM settings (key/value bag) ──────────────────────────────────────
# Stored in pm_settings table (see ensure_pm_settings_table in helpers.py).
# Currently exposes: grn_verify_required (bool). New string/bool keys can be
# added without code changes — the frontend coerces booleans from the string
# representation.
#
# These two endpoints back the loadPmSettings() / togglePmSetting() functions
# in pm_stock.html which were 404'ing because only /settings/fifo existed.
#
# Read is open to all logged-in users (frontend branches on the toggle's value
# for non-admin saveGrn behaviour). Write is admin-only.
_PM_SETTINGS_KNOWN_BOOL_KEYS = ('grn_verify_required',)


def _coerce_setting_value(key, raw):
    """Convert stored string value to the proper Python type for known keys."""
    if raw is None:
        return False if key in _PM_SETTINGS_KNOWN_BOOL_KEYS else None
    if key in _PM_SETTINGS_KNOWN_BOOL_KEYS:
        return str(raw).strip().lower() in ('1', 'true', 't', 'yes', 'y', 'on')
    return raw


@pm_stock_bp.route('/api/pm_stock/settings', methods=['GET'])
@_login_required
def api_pm_settings_get():
    """
    Return all PM behaviour settings as a JSON object.
    Open to all logged-in users so saveGrn can branch on toggles without an
    admin-only round-trip.
    Shape: { status:'ok', settings: { grn_verify_required: bool, ... } }
    """
    conn = sampling_portal.get_db_connection()
    try:
        settings = {}
        # Known bool keys — read each and coerce. Unknown keys in the table are
        # also returned as raw strings so future additions don't need code.
        for k in _PM_SETTINGS_KNOWN_BOOL_KEYS:
            raw = _setting_get(conn, k, default=None)
            settings[k] = _coerce_setting_value(k, raw)

        # Also include any other rows in pm_settings (forward-compat). Skip
        # rows already covered above and skip FIFO settings (those have their
        # own dedicated endpoint to avoid double-exposing them).
        try:
            rows = conn.execute(
                "SELECT setting_key, setting_value FROM pm_settings"
            ).fetchall()
            for r in rows:
                k = r['setting_key']
                if k in settings:
                    continue
                if k.startswith('fifo_'):
                    continue
                settings[k] = r['setting_value']
        except Exception:
            pass  # pm_settings table may be unavailable; defaults already set
        conn.close()
        return jsonify({'status': 'ok', 'settings': settings})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/settings', methods=['POST'])
@_login_required
def api_pm_settings_set():
    """
    Update a single PM setting. Admin only.
    Body: { key: str, value: <bool|str> }

    Validates the key against a known set so a stray POST can't write garbage
    rows to the settings table. Booleans are normalised to '1'/'0' strings.
    """
    if not _is_admin():
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    d = request.get_json() or {}
    key = (d.get('key') or '').strip()
    if key not in _PM_SETTINGS_KNOWN_BOOL_KEYS:
        return jsonify({
            'status':  'error',
            'message': f"Unknown setting key: {key!r}. "
                       f"Known keys: {', '.join(_PM_SETTINGS_KNOWN_BOOL_KEYS)}"
        }), 400
    val = d.get('value')
    # Normalise to '1' / '0' for storage
    if isinstance(val, bool):
        stored = '1' if val else '0'
    elif isinstance(val, (int, float)):
        stored = '1' if val else '0'
    else:
        stored = '1' if str(val).strip().lower() in ('1', 'true', 'yes', 'on') else '0'

    conn = sampling_portal.get_db_connection()
    try:
        _setting_set(conn, key, stored, user=_user())
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'key': key, 'value': stored == '1'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/settings/fifo', methods=['GET'])
@_login_required
def api_fifo_settings_get():
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    conn = sampling_portal.get_db_connection()
    try:
        # Read both keys + their metadata in one query
        rows = conn.execute(
            """SELECT setting_key, setting_value, updated_by, updated_at
               FROM pm_settings
               WHERE setting_key IN ('fifo_enabled','fifo_start_date')"""
        ).fetchall()
        m = {r['setting_key']: r for r in rows}
        en = m.get('fifo_enabled')
        sd = m.get('fifo_start_date')
        enabled = (en['setting_value'] != '0') if en else True
        start_date = (sd['setting_value'] or None) if sd else None
        # Pick the most recent updated_at across both keys for "last changed"
        last_meta = None
        for r in (en, sd):
            if not r: continue
            if last_meta is None or (r['updated_at'] and r['updated_at'] > last_meta['updated_at']):
                last_meta = r
        updated_by = last_meta['updated_by'] if last_meta else ''
        updated_at = str(last_meta['updated_at']) if last_meta and last_meta['updated_at'] else ''
        conn.close()
        return jsonify({
            'status':     'ok',
            'enabled':    bool(enabled),
            'start_date': start_date,
            'updated_by': updated_by,
            'updated_at': updated_at,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/settings/fifo', methods=['POST'])
@_login_required
def api_fifo_settings_save():
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403
    d = request.get_json(silent=True) or {}
    has_enabled    = 'enabled' in d
    has_start_date = 'start_date' in d
    reset_from_now = bool(d.get('reset_from_now'))

    if not has_enabled and not has_start_date and not reset_from_now:
        return jsonify({'status':'error','message':'Nothing to update'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # Snapshot current values for audit trail
        before = {
            'enabled':    _fifo_is_enabled(conn),
            'start_date': _fifo_start_date(conn),
        }

        if has_enabled:
            new_val = '1' if bool(d.get('enabled')) else '0'
            _setting_set(conn, 'fifo_enabled', new_val)

        new_start_date = before['start_date']
        if reset_from_now:
            new_start_date = str(date.today())
            _setting_set(conn, 'fifo_start_date', new_start_date)
        elif has_start_date:
            sd = d.get('start_date')
            if sd in (None, '', 'null'):
                _setting_set(conn, 'fifo_start_date', None)
                new_start_date = None
            else:
                # Validate format
                try:
                    parsed = date.fromisoformat(str(sd)[:10])
                    new_start_date = str(parsed)
                    _setting_set(conn, 'fifo_start_date', new_start_date)
                except Exception:
                    conn.close()
                    return jsonify({'status':'error',
                        'message':'start_date must be YYYY-MM-DD or null'}), 400

        after = {
            'enabled': bool(d.get('enabled')) if has_enabled else before['enabled'],
            'start_date': new_start_date,
        }

        # Audit
        try:
            summary_bits = []
            if before['enabled'] != after['enabled']:
                summary_bits.append(f"enforcement {'ON' if after['enabled'] else 'OFF'}")
            if before['start_date'] != after['start_date']:
                summary_bits.append(f"start-date → {after['start_date'] or 'cleared'}")
            if summary_bits:
                _audit_record(
                    conn,
                    action='fifo.settings_update',
                    entity='pm_settings',
                    entity_id='fifo',
                    summary='FIFO settings: ' + ', '.join(summary_bits),
                    before=before,
                    after=after,
                    reversal_class=AUDIT_REVERSAL_SAFE,
                )
        except Exception:
            pass

        conn.commit()
        conn.close()
        return jsonify({'status':'ok',
                        'enabled':    after['enabled'],
                        'start_date': after['start_date']})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════
# TRANSFER ADMIN-EDIT — admin-only full edit with cascade
# ═══════════════════════════════════════════════════════════════════════════
# POST /api/pm_stock/transfers/<tid>/admin_edit
#
# Full-power edit of any transfer voucher in any state. The cascade rule:
#   1. Reverse all existing stock postings (refund source, un-receive dest)
#   2. Apply new header (date/godowns/remarks) + wipe & re-insert items
#   3. Re-post stock based on new state
#   4. If the transfer had scanned boxes whose totals no longer match the
#      new item totals, log a discrepancy warning to pm_transfer_edits and
#      to the audit log — but DO NOT touch the boxes themselves.
#      Tarak explicitly accepted this orphan risk.
#
# Body shape:
# {
#   "transfer_date": "YYYY-MM-DD",
#   "from_godown_id": int,
#   "to_godown_id":   int,
#   "remarks":        str,
#   "out_items":      [{product_id, no_of_box, per_box_qty, total_qty}, ...],
#   "in_items":       [{...same...}],
#   "reason":         str   (required, min 4 chars)
# }
# ═══════════════════════════════════════════════════════════════════════════

@pm_stock_bp.route('/api/pm_stock/transfers/<int:tid>/admin_edit', methods=['POST'])
@_login_required
def api_transfer_admin_edit(tid):
    if not _is_admin():
        return jsonify({'status':'error','message':'Admin only'}), 403

    d = request.get_json(silent=True) or {}
    reason = (d.get('reason') or '').strip()
    if len(reason) < 4:
        return jsonify({'status':'error',
            'message':'Reason is required (minimum 4 characters)'}), 400

    new_date    = d.get('transfer_date')
    new_from    = d.get('from_godown_id')
    new_to      = d.get('to_godown_id')
    new_remarks = (d.get('remarks') or '').strip()
    new_out     = d.get('out_items') or []
    new_in      = d.get('in_items')  or []

    # Basic field validation
    try:
        if new_date:
            date.fromisoformat(str(new_date)[:10])
    except Exception:
        return jsonify({'status':'error',
            'message':'transfer_date must be YYYY-MM-DD'}), 400
    if not new_from or not new_to:
        return jsonify({'status':'error',
            'message':'from_godown_id and to_godown_id are required'}), 400
    try:
        new_from = int(new_from); new_to = int(new_to)
    except Exception:
        return jsonify({'status':'error',
            'message':'godown ids must be integers'}), 400
    if new_from == new_to:
        return jsonify({'status':'error',
            'message':'Source and destination cannot be the same'}), 400

    def _normalize_items(arr, side_label):
        """Validate and coerce item rows."""
        out = []
        for i, r in enumerate(arr):
            try:
                pid = int(r.get('product_id'))
                nob = max(0, int(r.get('no_of_box') or 0))
                pbq = float(r.get('per_box_qty') or 0)
                tot = float(r.get('total_qty') or 0)
            except Exception as e:
                raise ValueError(f"{side_label} item #{i+1}: invalid number — {e}")
            if pid <= 0:
                raise ValueError(f"{side_label} item #{i+1}: product is required")
            if tot < 0 or pbq < 0:
                raise ValueError(f"{side_label} item #{i+1}: negative qty not allowed")
            out.append({
                'product_id':  pid,
                'no_of_box':   nob,
                'per_box_qty': pbq,
                'total_qty':   tot,
            })
        return out

    try:
        new_out_items = _normalize_items(new_out, 'OUT')
        new_in_items  = _normalize_items(new_in,  'IN')
    except ValueError as e:
        return jsonify({'status':'error','message':str(e)}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # ── Snapshot BEFORE state ────────────────────────────────────────────
        t = conn.execute(
            "SELECT * FROM pm_transfers WHERE transfer_id=%s", (tid,)
        ).fetchone()
        if not t:
            conn.close()
            return jsonify({'status':'error','message':'Transfer not found'}), 404

        old_status   = t['status']
        old_xfer_no  = t['transfer_no']
        old_from     = t['from_godown_id']
        old_to       = t['to_godown_id']
        # Schema column is `out_at` (DATETIME), not `transfer_date`.
        # We slice :10 to get the date portion (YYYY-MM-DD).
        old_date     = str(t['out_at'])[:10] if t['out_at'] else None
        old_remarks  = t['remarks'] or ''

        old_items = conn.execute(
            """SELECT side, product_id, no_of_box, per_box_qty, total_qty
               FROM pm_transfer_items WHERE transfer_id=%s
               ORDER BY side, item_id""", (tid,)
        ).fetchall()
        before_items = [dict(r) for r in old_items]

        # Count physically-scanned boxes that touched this transfer.
        # Used purely as a discrepancy signal — we don't change boxes here.
        scanned_count = conn.execute(
            "SELECT COUNT(*) AS c FROM pm_box_movements WHERE transfer_id=%s",
            (tid,)
        ).fetchone()
        scanned_count = int(scanned_count['c']) if scanned_count else 0

        # ── REVERSE existing stock postings ──────────────────────────────────
        # We mirror the logic in api_transfer_admin_delete: refund source for
        # any OUT lines, un-receive destination for any IN lines (only if
        # those side's posts had been written based on status).
        today      = str(date.today())
        reversal   = {'source_refunded': 0.0, 'dest_unreceived': 0.0}
        applied_dt = old_date or today

        if old_status in ('in_pending', 'received'):
            for r in before_items:
                if (r.get('side') or '').lower() != 'out': continue
                qty = float(r.get('total_qty') or 0)
                if qty <= 0: continue
                _post_stock_movement(
                    conn,
                    product_id=r['product_id'], godown_id=old_from,
                    qty=qty, direction='in',
                    transfer_no=old_xfer_no, transfer_id=tid,
                    txn_date=today, user=_user()
                )
                reversal['source_refunded'] += qty

        if old_status == 'received':
            for r in before_items:
                if (r.get('side') or '').lower() != 'in': continue
                qty = float(r.get('total_qty') or 0)
                if qty <= 0: continue
                _post_stock_movement(
                    conn,
                    product_id=r['product_id'], godown_id=old_to,
                    qty=qty, direction='out',
                    transfer_no=old_xfer_no, transfer_id=tid,
                    txn_date=today, user=_user()
                )
                reversal['dest_unreceived'] += qty

        # ── UPDATE header ────────────────────────────────────────────────────
        # transfer_no intentionally NOT editable per Tarak's lock-in.
        # Schema column is `out_at` (DATETIME). When the user changes the
        # date portion, MySQL accepts a date string and auto-fills 00:00:00
        # for the time. If the user kept the same date, we leave out_at
        # untouched to preserve the original timestamp.
        if new_date and new_date != old_date:
            conn.execute(
                """UPDATE pm_transfers
                   SET out_at=%s, from_godown_id=%s, to_godown_id=%s,
                       remarks=%s
                   WHERE transfer_id=%s""",
                (new_date, new_from, new_to, new_remarks, tid)
            )
        else:
            conn.execute(
                """UPDATE pm_transfers
                   SET from_godown_id=%s, to_godown_id=%s, remarks=%s
                   WHERE transfer_id=%s""",
                (new_from, new_to, new_remarks, tid)
            )

        # ── REPLACE items ────────────────────────────────────────────────────
        conn.execute(
            "DELETE FROM pm_transfer_items WHERE transfer_id=%s", (tid,)
        )
        for r in new_out_items:
            conn.execute(
                """INSERT INTO pm_transfer_items
                     (transfer_id, side, product_id, no_of_box, per_box_qty, total_qty)
                   VALUES (%s,'out',%s,%s,%s,%s)""",
                (tid, r['product_id'], r['no_of_box'], r['per_box_qty'], r['total_qty'])
            )
        for r in new_in_items:
            conn.execute(
                """INSERT INTO pm_transfer_items
                     (transfer_id, side, product_id, no_of_box, per_box_qty, total_qty)
                   VALUES (%s,'in',%s,%s,%s,%s)""",
                (tid, r['product_id'], r['no_of_box'], r['per_box_qty'], r['total_qty'])
            )

        # ── RE-POST stock with new state ────────────────────────────────────
        # Mirror the reversal logic but in the forward direction.
        # Use new godowns + new transfer_date.
        new_apply_dt = new_date or old_date or today
        new_posting  = {'source_decremented': 0.0, 'dest_credited': 0.0}

        if old_status in ('in_pending', 'received'):
            for r in new_out_items:
                qty = float(r['total_qty'] or 0)
                if qty <= 0: continue
                _post_stock_movement(
                    conn,
                    product_id=r['product_id'], godown_id=new_from,
                    qty=qty, direction='out',
                    transfer_no=old_xfer_no, transfer_id=tid,
                    txn_date=new_apply_dt, user=_user()
                )
                new_posting['source_decremented'] += qty

        if old_status == 'received':
            for r in new_in_items:
                qty = float(r['total_qty'] or 0)
                if qty <= 0: continue
                _post_stock_movement(
                    conn,
                    product_id=r['product_id'], godown_id=new_to,
                    qty=qty, direction='in',
                    transfer_no=old_xfer_no, transfer_id=tid,
                    txn_date=new_apply_dt, user=_user()
                )
                new_posting['dest_credited'] += qty

        # ── Refresh aggregate header totals from new items ─────────────────
        try:
            _refresh_transfer_totals(conn, tid)
        except Exception:
            pass

        # ── Discrepancy detection — boxes vs new totals ──────────────────────
        warnings = []
        if scanned_count > 0:
            new_out_box_total = sum(r['no_of_box'] for r in new_out_items)
            new_in_box_total  = sum(r['no_of_box'] for r in new_in_items)
            # Count actual scanned by side
            scn = conn.execute(
                """SELECT movement_type, COUNT(*) AS c
                   FROM pm_box_movements
                   WHERE transfer_id=%s
                   GROUP BY movement_type""", (tid,)
            ).fetchall()
            scn_map = {r['movement_type']: int(r['c']) for r in scn}
            out_scanned = scn_map.get('out', 0)
            in_scanned  = scn_map.get('in',  0)
            if old_status in ('in_pending', 'received') and new_out_box_total != out_scanned:
                warnings.append(
                    f"OUT total boxes changed to {new_out_box_total} but "
                    f"{out_scanned} boxes are physically scanned to OUT — "
                    f"orphan delta = {out_scanned - new_out_box_total}"
                )
            if old_status == 'received' and new_in_box_total != in_scanned:
                warnings.append(
                    f"IN total boxes changed to {new_in_box_total} but "
                    f"{in_scanned} boxes are physically scanned to IN — "
                    f"orphan delta = {in_scanned - new_in_box_total}"
                )

        # ── Append to per-transfer edit history ──────────────────────────────
        try:
            details_bits = [f"reason={reason}"]
            if old_date != (new_date or old_date):
                details_bits.append(f"date {old_date}→{new_date}")
            if old_from != new_from:
                details_bits.append(f"from #{old_from}→#{new_from}")
            if old_to != new_to:
                details_bits.append(f"to #{old_to}→#{new_to}")
            if old_remarks != new_remarks:
                details_bits.append("remarks changed")
            if warnings:
                details_bits.append('WARNINGS: ' + ' | '.join(warnings))
            _log_transfer_edit(conn, tid, 'admin_edit', '; '.join(details_bits))
        except Exception:
            pass

        # ── Audit log ────────────────────────────────────────────────────────
        try:
            summary = (
                f"Transfer {old_xfer_no} edited by admin · "
                f"status={old_status} · scanned={scanned_count} boxes"
            )
            if warnings:
                summary += ' · DISCREPANCY'
            _audit_record(
                conn,
                action='transfer.admin_edit',
                entity='pm_transfers',
                entity_id=str(tid),
                summary=summary,
                before={
                    'transfer_no': old_xfer_no,
                    'status':      old_status,
                    'date':        old_date,
                    'from':        old_from,
                    'to':          old_to,
                    'remarks':     old_remarks,
                    'items':       before_items,
                    'scanned_boxes': scanned_count,
                },
                after={
                    'transfer_no': old_xfer_no,
                    'status':      old_status,
                    'date':        new_date or old_date,
                    'from':        new_from,
                    'to':          new_to,
                    'remarks':     new_remarks,
                    'out_items':   new_out_items,
                    'in_items':    new_in_items,
                    'reason':      reason,
                    'warnings':    warnings,
                    'reversal':    reversal,
                    'reposting':   new_posting,
                },
                reversal_class=AUDIT_REVERSAL_GATED,
            )
        except Exception:
            pass

        conn.commit()
        conn.close()

        return jsonify({
            'status':       'ok',
            'transfer_no':  old_xfer_no,
            'old_status':   old_status,
            'reversal':     reversal,
            'reposting':    new_posting,
            'warnings':     warnings,
            'message':      'Transfer edited and stock recomputed.',
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':f'admin_edit failed: {e}'}), 500


# ═══════════════════════════════════════════════════════════════════════════
# BOX SPLIT — divide a parent box into N children of arbitrary qtys
# ═══════════════════════════════════════════════════════════════════════════
# POST /api/pm_stock/boxes/<box_id>/split
#
# Body:
# {
#   "splits": [{"qty": 50}, {"qty": 50}],
#   "reason": "Need to send only half to factory"
# }
#
# What happens:
#   1. Validate parent box exists and is in_stock
#   2. Validate sum(splits.qty) == parent.per_box_qty (no qty creation/loss)
#   3. For each child: create new pm_boxes row with same product/grn/godown,
#      new box_code (next free B-seq), parent_box_id pointing at parent
#   4. Mark parent: current_status='superseded', split_at=now
#   5. Insert pm_box_movements row of type='split' for audit
#   6. Audit log entry — class GATED (reversible only via merge, future)
#   7. Return list of child box_codes for label printing
#
# Stock postings: NONE. A split doesn't move stock — the same qty stays at
# the same godown, just packaged differently. pm_godown_txn unchanged.
#
# FIFO: children inherit parent's lot via grn_id/grn_no, so they get the
# same fifo_code (lookup is by lot+product, not by box).
# ═══════════════════════════════════════════════════════════════════════════

@pm_stock_bp.route('/api/pm_stock/boxes/<int:box_id>/split', methods=['POST'])
@_login_required
def api_box_split(box_id):
    d = request.get_json(silent=True) or {}
    splits = d.get('splits') or []
    reason = (d.get('reason') or '').strip()

    if not isinstance(splits, list) or len(splits) < 2:
        return jsonify({'status':'error',
            'message':'At least 2 child splits are required'}), 400
    if len(splits) > 50:
        # Sanity cap. Splitting one box into 50 children is already absurd.
        return jsonify({'status':'error',
            'message':'Maximum 50 children per split'}), 400

    # Coerce + validate qtys
    try:
        child_qtys = []
        for i, s in enumerate(splits):
            q = float(s.get('qty') or 0)
            if q <= 0:
                return jsonify({'status':'error',
                    'message':f'Child #{i+1}: qty must be > 0'}), 400
            child_qtys.append(q)
    except Exception as e:
        return jsonify({'status':'error',
            'message':f'Invalid split qty: {e}'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        parent = conn.execute(
            """SELECT b.*, p.product_name, p.product_code, p.pm_type,
                      COALESCE(g.name,'') AS godown_name
               FROM pm_boxes b
               LEFT JOIN pm_products p ON p.id = b.product_id
               LEFT JOIN procurement_godowns g ON g.id = b.current_godown_id
               WHERE b.box_id = %s""",
            (box_id,)
        ).fetchone()
        if not parent:
            conn.close()
            return jsonify({'status':'error','message':'Box not found'}), 404
        if parent['current_status'] != 'in_stock':
            conn.close()
            return jsonify({'status':'error',
                'message':f"Cannot split — box is currently '{parent['current_status']}'. "
                          f"Only in_stock boxes can be split."}), 400

        parent_qty = float(parent['per_box_qty'] or 0)
        total_split = sum(child_qtys)
        # Allow tiny float epsilon for decimal qty inputs
        if abs(total_split - parent_qty) > 0.001:
            conn.close()
            return jsonify({'status':'error',
                'message':f'Splits must total parent qty ({parent_qty:g}). '
                          f'You entered {total_split:g}.'}), 400

        # Find the next free box_seq for this GRN+product. We grab MAX(box_seq)
        # from pm_boxes filtered to same grn_id+product_id and increment from
        # there. This keeps the suffix readable (Bxxx) for new children.
        seq_row = conn.execute(
            """SELECT COALESCE(MAX(box_seq), 0) AS max_seq
               FROM pm_boxes
               WHERE grn_id = %s AND product_id = %s""",
            (parent['grn_id'], parent['product_id'])
        ).fetchone()
        next_seq = int(seq_row['max_seq'] or 0)

        # Create the children
        children = []
        is_op = (parent['grn_no'] or '').startswith('PM-OP/')
        for q in child_qtys:
            next_seq += 1
            if is_op:
                # OP-style box code: PRODCODE-OPnnnn-Bnnn
                op_seq_part = parent['grn_no'].split('/')[-1]
                child_code = (
                    f"{(parent['product_code'] or 'XXXXXXXXXX').upper()}"
                    f"-OP{op_seq_part}"
                    f"-B{str(next_seq).zfill(3)}"
                )
            else:
                child_code = _make_box_code(
                    parent['product_code'] or 'XXXXXXXXXX',
                    parent['grn_no'],
                    next_seq
                )

            cur = conn.execute(
                """INSERT INTO pm_boxes
                     (box_code, parent_box_id, grn_id, grn_no, grn_item_id,
                      product_id, product_code, box_seq, total_boxes,
                      per_box_qty, current_godown_id, current_status,
                      created_by)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'in_stock',%s)""",
                (child_code, parent['box_id'], parent['grn_id'], parent['grn_no'],
                 parent['grn_item_id'], parent['product_id'],
                 parent['product_code'], next_seq, parent['total_boxes'],
                 q, parent['current_godown_id'], _user())
            )
            try:
                child_id = cur.lastrowid
            except Exception:
                # Some DB drivers expose lastrowid via cursor differently;
                # fall back to a fresh SELECT.
                row = conn.execute(
                    "SELECT box_id FROM pm_boxes WHERE box_code = %s",
                    (child_code,)
                ).fetchone()
                child_id = row['box_id'] if row else None

            if child_id:
                _assign_box_short_code(conn, child_id)

            children.append({
                'box_id':   child_id,
                'box_code': child_code,
                'qty':      q,
            })

        # Mark parent superseded + record split timestamp
        conn.execute(
            """UPDATE pm_boxes
               SET current_status='superseded', split_at=NOW()
               WHERE box_id=%s""",
            (parent['box_id'],)
        )

        # Audit row in pm_box_movements
        child_codes_str = ', '.join(c['box_code'] for c in children)
        conn.execute(
            """INSERT INTO pm_box_movements
                 (box_id, movement_type, from_godown_id, to_godown_id,
                  qty, moved_by, remarks)
               VALUES (%s,'split',%s,%s,%s,%s,%s)""",
            (parent['box_id'], parent['current_godown_id'], parent['current_godown_id'],
             parent_qty, _user(),
             f"Split into {len(children)} children: {child_codes_str}"
             + (f" — Reason: {reason}" if reason else ""))
        )

        # Audit log
        try:
            _audit_record(
                conn,
                action='box.split',
                entity='pm_boxes',
                entity_id=str(parent['box_id']),
                summary=(
                    f"Split box {parent['box_code']} ({parent_qty:g}) → "
                    f"{len(children)} children at {parent['godown_name']}"
                ),
                before={
                    'box_code':    parent['box_code'],
                    'box_id':      parent['box_id'],
                    'product_id':  parent['product_id'],
                    'product':     parent['product_name'],
                    'per_box_qty': parent_qty,
                    'godown':      parent['godown_name'],
                    'status':      parent['current_status'],
                },
                after={
                    'parent_status': 'superseded',
                    'children': [
                        {'box_code': c['box_code'], 'qty': c['qty']}
                        for c in children
                    ],
                    'reason': reason,
                },
                reversal_class=AUDIT_REVERSAL_GATED,
            )
        except Exception:
            pass

        conn.commit()
        conn.close()

        return jsonify({
            'status':   'ok',
            'message':  f'Split {parent["box_code"]} into {len(children)} children',
            'parent':   {
                'box_id':   parent['box_id'],
                'box_code': parent['box_code'],
                'qty':      parent_qty,
            },
            'children': children,
            'product': {
                'id':           parent['product_id'],
                'name':         parent['product_name'],
                'product_code': parent['product_code'],
                'pm_type':      parent['pm_type'],
            },
            'grn_no':  parent['grn_no'],
            'godown':  parent['godown_name'],
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':f'Split failed: {e}'}), 500


# ═══════════════════════════════════════════════════════════════════════════
# BOX GROUPS — bag/lot scanning
# ═══════════════════════════════════════════════════════════════════════════
# A "group" or "bag" is a physical bundle of boxes shipped together.
# When operator scans a group code (BAG-PRODUCT-G0068-L001), the backend
# fans out and applies the same operation to every member box.
#
# Endpoints:
#   POST /api/pm_stock/groups/create_for_grn
#       Body: {grn_id, product_id, box_ids: [...], label?, remarks?}
#       Creates a group at GRN time. All boxes must be in_stock at the
#       same godown and share the product. Permissive on label.
#       Returns: {group_id, group_code, ...}
#
#   GET /api/pm_stock/groups/by_code?code=BAG-...
#       Lookup helper. Returns group + member list. 404 if not found.
#
#   POST /api/pm_stock/groups/<group_id>/scan
#       Body: {tid, side: 'out'|'in'}
#       Fans out: applies api_voucher_scan_box logic to each member.
#       Permissive: returns lists of succeeded vs failed boxes.
#       After fan-out, runs _refresh_group_status to update the group
#       status based on member outcomes.
#
#   POST /api/pm_stock/groups/<group_id>/break
#       Body: {reason?}
#       Dissolves the group: removes pm_boxes.group_id from members,
#       deletes member rows, marks group superseded. Boxes return to
#       individual handling.
# ═══════════════════════════════════════════════════════════════════════════

@pm_stock_bp.route('/api/pm_stock/groups/create_for_grn', methods=['POST'])
@_login_required
def api_group_create_for_grn():
    d = request.get_json(silent=True) or {}
    grn_id     = d.get('grn_id')
    product_id = d.get('product_id')
    box_ids    = d.get('box_ids') or []
    label      = (d.get('label') or '').strip() or None
    remarks    = (d.get('remarks') or '').strip() or None

    if not grn_id or not product_id:
        return jsonify({'status':'error',
            'message':'grn_id and product_id are required'}), 400
    if not isinstance(box_ids, list) or len(box_ids) < 2:
        return jsonify({'status':'error',
            'message':'At least 2 box_ids required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        # Look up GRN to get grn_no and the godown — needed for code generation
        grn = conn.execute(
            "SELECT grn_no, godown_id FROM pm_grn WHERE grn_id=%s",
            (grn_id,)
        ).fetchone()
        if not grn:
            conn.close()
            return jsonify({'status':'error','message':'GRN not found'}), 404

        try:
            grp = _create_group_for_boxes(
                conn,
                product_id=int(product_id),
                grn_id=int(grn_id),
                grn_no=grn['grn_no'],
                godown_id=grn['godown_id'],
                box_ids=[int(b) for b in box_ids],
                label=label,
                remarks=remarks,
                user=_user(),
            )
        except ValueError as e:
            conn.close()
            return jsonify({'status':'error','message':str(e)}), 400

        # Audit
        try:
            _audit_record(
                conn,
                action='box.group_create',
                entity='pm_box_groups',
                entity_id=str(grp['group_id']),
                summary=(
                    f"Created group {grp['group_code']} with "
                    f"{grp['member_count']} boxes (qty {grp['total_qty']:g})"
                ),
                before=None,
                after={
                    'group_code':   grp['group_code'],
                    'product_id':   grp['product_id'],
                    'grn_id':       grp['grn_id'],
                    'grn_no':       grp['grn_no'],
                    'box_ids':      grp['box_ids'],
                    'member_count': grp['member_count'],
                    'total_qty':    grp['total_qty'],
                    'label':        label,
                    'remarks':      remarks,
                },
                reversal_class=AUDIT_REVERSAL_GATED,
            )
        except Exception:
            pass

        conn.commit()
        conn.close()
        return jsonify({'status':'ok', 'group': grp})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':f'Group create failed: {e}'}), 500


@pm_stock_bp.route('/api/pm_stock/groups/by_code', methods=['GET'])
@_login_required
def api_group_by_code():
    code = (request.args.get('code') or '').strip().upper()
    if not code:
        return jsonify({'status':'error','message':'code parameter is required'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        g = conn.execute(
            """SELECT bg.*,
                      p.product_name, p.product_code, p.pm_type,
                      gd.name AS godown_name
               FROM pm_box_groups bg
               LEFT JOIN pm_products p ON p.id = bg.product_id
               LEFT JOIN procurement_godowns gd ON gd.id = bg.current_godown_id
               WHERE bg.group_code = %s""",
            (code,)
        ).fetchone()
        if not g:
            conn.close()
            return jsonify({'status':'error','message':f'Group {code} not found',
                            'code':'group_not_found'}), 404
        members = conn.execute(
            """SELECT b.box_id, b.box_code, b.per_box_qty, b.current_status,
                      b.current_godown_id
               FROM pm_box_group_members m
               JOIN pm_boxes b ON b.box_id = m.box_id
               WHERE m.group_id = %s
               ORDER BY b.box_seq""",
            (g['group_id'],)
        ).fetchall()
        conn.close()
        return jsonify({
            'status':  'ok',
            'group':   {k: _json_safe(v) for k, v in dict(g).items()},
            'members': [{k: _json_safe(v) for k, v in dict(m).items()} for m in members],
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/groups/<int:group_id>/scan', methods=['POST'])
@_login_required
def api_group_scan(group_id):
    """
    Returns the list of member boxes ready to be scanned by the frontend.

    DESIGN NOTE: Rather than fan out scans server-side (which would require
    duplicating ~300 lines of scan_box validation logic), this endpoint
    simply returns the member box codes. The frontend iterates and calls
    the existing /scan_box endpoint once per box. Each individual scan
    gets full validation, FIFO checks, audit logging — the same as a
    manual scan. The frontend collects per-box results and renders a
    summary, so the operator sees what worked and what didn't.

    Status precondition: group must be in_stock or partial. The frontend
    should still pass this through for permissive scanning.

    Body:
      { tid:  transfer id (int),
        side: 'out' | 'in' }

    Response:
      { status: 'ok',
        group_code:   '...',
        side:         'out' | 'in',
        boxes: [{box_id, box_code, per_box_qty, current_status, ...}, ...] }
    """
    d = request.get_json(silent=True) or {}
    tid  = d.get('tid')
    side = (d.get('side') or 'out').lower()
    if not tid:
        return jsonify({'status':'error','message':'tid is required'}), 400
    if side not in ('out', 'in'):
        return jsonify({'status':'error','message':"side must be 'out' or 'in'"}), 400

    conn = sampling_portal.get_db_connection()
    try:
        g = conn.execute(
            "SELECT * FROM pm_box_groups WHERE group_id=%s", (group_id,)
        ).fetchone()
        if not g:
            conn.close()
            return jsonify({'status':'error','message':'Group not found'}), 404
        if g['current_status'] in ('superseded', 'consumed'):
            conn.close()
            return jsonify({'status':'error',
                'message':f"Group is {g['current_status']} — cannot scan"}), 409

        members = conn.execute(
            """SELECT b.box_id, b.box_code, b.per_box_qty,
                      b.current_status, b.current_godown_id
               FROM pm_box_group_members m
               JOIN pm_boxes b ON b.box_id = m.box_id
               WHERE m.group_id = %s
               ORDER BY b.box_seq""",
            (group_id,)
        ).fetchall()
        if not members:
            conn.close()
            return jsonify({'status':'error','message':'Group has no members'}), 409

        boxes = [{k: _json_safe(v) for k, v in dict(m).items()} for m in members]
        conn.close()
        return jsonify({
            'status':     'ok',
            'group_id':   group_id,
            'group_code': g['group_code'],
            'side':       side,
            'tid':        tid,
            'boxes':      boxes,
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':f'Group scan prep failed: {e}'}), 500


@pm_stock_bp.route('/api/pm_stock/groups/<int:group_id>/scan_complete', methods=['POST'])
@_login_required
def api_group_scan_complete(group_id):
    """
    Called by the frontend AFTER it has finished iterating the box scans.
    Updates the group's current_status / godown / member_count to reflect
    the new state of its members, and writes ONE consolidated audit entry.

    Body:
      { tid:        transfer id,
        side:       'out' | 'in',
        succeeded:  [box_code, ...],
        failed:     [{box_code, reason}, ...] }
    """
    d = request.get_json(silent=True) or {}
    tid       = d.get('tid')
    side      = (d.get('side') or 'out').lower()
    succeeded = d.get('succeeded') or []
    failed    = d.get('failed') or []

    conn = sampling_portal.get_db_connection()
    try:
        g = conn.execute(
            "SELECT * FROM pm_box_groups WHERE group_id=%s", (group_id,)
        ).fetchone()
        if not g:
            conn.close()
            return jsonify({'status':'error','message':'Group not found'}), 404

        # Recompute group status from member current state
        try:
            _refresh_group_status(conn, group_id)
        except Exception:
            pass

        # Audit
        try:
            _audit_record(
                conn,
                action='box.group_scan',
                entity='pm_box_groups',
                entity_id=str(group_id),
                summary=(
                    f"Group {g['group_code']} scanned for {side.upper()} on "
                    f"transfer #{tid}: {len(succeeded)} ok, {len(failed)} failed"
                ),
                before=None,
                after={
                    'group_code':   g['group_code'],
                    'transfer_id':  tid,
                    'side':         side,
                    'succeeded':    succeeded,
                    'failed':       failed,
                },
                reversal_class=AUDIT_REVERSAL_SAFE,
            )
        except Exception:
            pass

        conn.commit()
        conn.close()
        return jsonify({'status':'ok'})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':f'scan_complete failed: {e}'}), 500


@pm_stock_bp.route('/api/pm_stock/groups/<int:group_id>/break', methods=['POST'])
@_login_required
def api_group_break(group_id):
    """Dissolve a group — members go back to individual handling."""
    d = request.get_json(silent=True) or {}
    reason = (d.get('reason') or '').strip()

    conn = sampling_portal.get_db_connection()
    try:
        g = conn.execute(
            "SELECT * FROM pm_box_groups WHERE group_id=%s", (group_id,)
        ).fetchone()
        if not g:
            conn.close()
            return jsonify({'status':'error','message':'Group not found'}), 404
        if g['current_status'] == 'superseded':
            conn.close()
            return jsonify({'status':'error','message':'Group is already superseded'}), 409

        # Snapshot members for audit
        members = conn.execute(
            """SELECT b.box_id, b.box_code
               FROM pm_box_group_members m
               JOIN pm_boxes b ON b.box_id = m.box_id
               WHERE m.group_id = %s""",
            (group_id,)
        ).fetchall()
        member_codes = [m['box_code'] for m in members]

        # Clear group_id on member boxes
        conn.execute(
            "UPDATE pm_boxes SET group_id=NULL WHERE group_id=%s",
            (group_id,)
        )
        # Delete member rows
        conn.execute(
            "DELETE FROM pm_box_group_members WHERE group_id=%s",
            (group_id,)
        )
        # Mark group superseded (audit trail kept)
        conn.execute(
            """UPDATE pm_box_groups
               SET current_status='superseded', member_count=0
               WHERE group_id=%s""",
            (group_id,)
        )

        # Audit
        try:
            _audit_record(
                conn,
                action='box.group_break',
                entity='pm_box_groups',
                entity_id=str(group_id),
                summary=f"Broke group {g['group_code']} ({len(members)} boxes returned to individual)",
                before={
                    'group_code':   g['group_code'],
                    'member_count': g['member_count'],
                    'total_qty':    float(g['total_qty'] or 0),
                    'members':      member_codes,
                },
                after={
                    'status': 'superseded',
                    'reason': reason or None,
                },
                reversal_class=AUDIT_REVERSAL_GATED,
            )
        except Exception:
            pass

        conn.commit()
        conn.close()
        return jsonify({'status':'ok',
                        'group_code': g['group_code'],
                        'released_count': len(members)})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':f'Group break failed: {e}'}), 500


# ── Helper endpoint for the GRN frontend's group-creation post-save flow ────
# After a GRN saves and the backend creates the boxes, the frontend calls
# this to fetch the list of box_ids per product_id so it can call
# /groups/create_for_grn for each line that the operator marked "bagged".
# Lightweight read-only — no auth changes, just a join.
@pm_stock_bp.route('/api/pm_stock/grn/<int:grn_id>/boxes_by_product', methods=['GET'])
@_login_required
def api_grn_boxes_by_product(grn_id):
    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(
            """SELECT product_id, box_id, box_code, box_seq, per_box_qty
               FROM pm_boxes
               WHERE grn_id = %s
               ORDER BY product_id, box_seq""",
            (grn_id,)
        ).fetchall()
        # Group by product_id
        by_pid = {}
        for r in rows:
            pid = int(r['product_id'])
            by_pid.setdefault(pid, []).append({
                'box_id':       int(r['box_id']),
                'box_code':     r['box_code'],
                'box_seq':      int(r['box_seq']),
                'per_box_qty':  float(r['per_box_qty'] or 0),
            })
        conn.close()
        return jsonify({'status':'ok', 'grn_id': grn_id, 'by_product': by_pid})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/grn/<int:grn_id>/groups', methods=['GET'])
@_login_required
def api_grn_groups(grn_id):
    """
    Returns the list of bag/groups created for this GRN, with their member
    box codes. Used by the GRN label print pipeline to append group sticker
    pages alongside the regular box labels.
    """
    conn = sampling_portal.get_db_connection()
    try:
        groups = conn.execute(
            """SELECT bg.group_id, bg.group_code, bg.group_label,
                      bg.product_id, bg.grn_no, bg.member_count, bg.total_qty,
                      bg.current_godown_id, bg.current_status,
                      p.product_name, p.product_code, p.pm_type, p.brand_id,
                      gd.name AS godown_name
               FROM pm_box_groups bg
               LEFT JOIN pm_products p ON p.id = bg.product_id
               LEFT JOIN procurement_godowns gd ON gd.id = bg.current_godown_id
               WHERE bg.grn_id = %s AND bg.current_status != 'superseded'
               ORDER BY bg.group_id""",
            (grn_id,)
        ).fetchall()
        out = []
        for g in groups:
            members = conn.execute(
                """SELECT b.box_id, b.box_code, b.per_box_qty
                   FROM pm_box_group_members m
                   JOIN pm_boxes b ON b.box_id = m.box_id
                   WHERE m.group_id = %s
                   ORDER BY b.box_seq""",
                (g['group_id'],)
            ).fetchall()
            row = {k: _json_safe(v) for k, v in dict(g).items()}
            # Resolve brand_name via the existing helper
            brand_name = ''
            if g['brand_id']:
                try: brand_name = _brand_name_by_id(conn, g['brand_id']) or ''
                except Exception: pass
            row['brand_name'] = brand_name
            row['members'] = [{k: _json_safe(v) for k, v in dict(m).items()} for m in members]
            out.append(row)
        conn.close()
        return jsonify({'status':'ok', 'grn_id': grn_id, 'groups': out})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


# ════════════════════════════════════════════════════════════════════════
# STOCK ADJUSTMENT VOUCHERS — pending/approved/rejected workflow
# ════════════════════════════════════════════════════════════════════════
#
# Non-admin creates a voucher with status='pending' — no ledger effect.
# Admin approves → ledger rows posted (pm_godown_txn inward/outward or
# pm_floor_txn pm_return/rejection) and item rows store godown_txn_id /
# floor_txn_id for future reversal traceability.
# Admin rejects → status='rejected' with reason. Requester can edit and
# resubmit (resets to pending, clears rejection metadata).
#
# Single-location per voucher: header carries godown_id + is_floor flag,
# applied to every line item.
#
# Access gating:
#   - List/view: any logged-in user with 'stock_adjustment' access. Non-
#     admins see ONLY their own vouchers; admins see all.
#   - Create/edit own pending: any logged-in user with 'stock_adjustment'.
#   - Delete: requester (own pending only) or admin (any).
#   - Approve / reject: admin only.
# ════════════════════════════════════════════════════════════════════════

def _adj_can_edit(row, current_user, is_admin):
    """Edit rights: own pending/rejected, or admin on pending/rejected.
    Approved vouchers are immutable (would require reversal flow).
    Returns (ok, reason).
    """
    if not row:
        return False, 'Voucher not found'
    status = (row.get('status') or '').lower()
    if status == 'approved':
        return False, 'Approved vouchers cannot be edited'
    if is_admin:
        return True, None
    if (row.get('requested_by') or '') != (current_user or ''):
        return False, 'You can only edit your own vouchers'
    return True, None


def _adj_resolve_godown(conn, godown_id):
    """Fetch godown row and figure out is_floor for the location. Returns
    (row, is_floor) or (None, False) if the godown id is invalid.
    """
    try:
        gid = int(godown_id)
    except Exception:
        return None, False
    row = conn.execute(
        "SELECT id, name, COALESCE(type,'') AS godown_type FROM procurement_godowns WHERE id=%s",
        (gid,)
    ).fetchone()
    if not row:
        return None, False
    is_floor = _is_floor_godown(conn, gid)
    return row, bool(is_floor)


@pm_stock_bp.route('/api/pm_stock/adjustments/list', methods=['GET'])
@_login_required
def api_pm_adj_list():
    """List stock adjustment vouchers.

    Query params:
      status    — 'pending'|'approved'|'rejected'|'all' (default 'all')
      mine_only — '1' to force only-own filter (non-admins always restricted)
      limit     — default 200

    Non-admins always see only their own vouchers regardless of mine_only.
    """
    blocked = _block_if_no_access('stock_adjustment')
    if blocked is not None: return blocked

    status = (request.args.get('status') or 'all').lower()
    mine_only = request.args.get('mine_only') == '1'
    try:
        limit = max(1, min(500, int(request.args.get('limit') or 200)))
    except Exception:
        limit = 200

    user = _user() or ''
    is_admin = _is_admin()
    where = []
    params = []
    if status in ('pending', 'approved', 'rejected'):
        where.append('a.status=%s')
        params.append(status)
    if mine_only or not is_admin:
        where.append('a.requested_by=%s')
        params.append(user)
    sql_where = (' WHERE ' + ' AND '.join(where)) if where else ''

    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(
            f"""SELECT a.id, a.adj_no, a.adj_date, a.godown_id, a.is_floor,
                       a.status, a.requested_by, a.requested_at,
                       a.approved_by, a.approved_at,
                       a.rejected_by, a.rejected_at, a.reject_reason,
                       a.voucher_remarks,
                       COALESCE(g.name,'') AS godown_name,
                       (SELECT COUNT(*) FROM pm_stock_adjustment_items i WHERE i.adj_id=a.id) AS line_count,
                       (SELECT COALESCE(SUM(CASE WHEN direction='increase' THEN qty ELSE 0 END),0)
                          FROM pm_stock_adjustment_items i WHERE i.adj_id=a.id) AS total_increase,
                       (SELECT COALESCE(SUM(CASE WHEN direction='decrease' THEN qty ELSE 0 END),0)
                          FROM pm_stock_adjustment_items i WHERE i.adj_id=a.id) AS total_decrease
                FROM pm_stock_adjustments a
                LEFT JOIN procurement_godowns g ON g.id = a.godown_id
                {sql_where}
                ORDER BY (a.status='pending') DESC, a.requested_at DESC, a.id DESC
                LIMIT {limit}""",
            tuple(params)
        ).fetchall()
        conn.close()
        out = []
        for r in rows:
            d = {k: _json_safe(v) for k, v in dict(r).items()}
            out.append(d)
        return jsonify({'status': 'ok', 'vouchers': out, 'is_admin': is_admin, 'me': user})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/adjustments/<int:adj_id>', methods=['GET'])
@_login_required
def api_pm_adj_detail(adj_id):
    """Detail view: header + all line items with product names."""
    blocked = _block_if_no_access('stock_adjustment')
    if blocked is not None: return blocked

    user = _user() or ''
    is_admin = _is_admin()

    conn = sampling_portal.get_db_connection()
    try:
        hdr = conn.execute(
            """SELECT a.*, COALESCE(g.name,'') AS godown_name
               FROM pm_stock_adjustments a
               LEFT JOIN procurement_godowns g ON g.id = a.godown_id
               WHERE a.id=%s""",
            (adj_id,)
        ).fetchone()
        if not hdr:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Voucher not found'}), 404
        # Non-admin can only see own
        if not is_admin and (hdr.get('requested_by') or '') != user:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Not allowed to view this voucher'}), 403

        items = conn.execute(
            """SELECT i.id, i.product_id, i.direction, i.qty, i.reason,
                      i.godown_txn_id, i.floor_txn_id,
                      COALESCE(p.product_name,'') AS product_name,
                      COALESCE(p.product_code,'') AS product_code,
                      COALESCE(p.pm_type,'')      AS pm_type,
                      COALESCE(p.primary_uom,'Nos') AS primary_uom
               FROM pm_stock_adjustment_items i
               LEFT JOIN pm_products p ON p.id = i.product_id
               WHERE i.adj_id=%s
               ORDER BY i.id""",
            (adj_id,)
        ).fetchall()
        conn.close()
        return jsonify({
            'status':  'ok',
            'voucher': {k: _json_safe(v) for k, v in dict(hdr).items()},
            'items':   [{k: _json_safe(v) for k, v in dict(r).items()} for r in items],
            'is_admin': is_admin,
            'me': user,
            'can_edit': _adj_can_edit(dict(hdr), user, is_admin)[0],
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status': 'error', 'message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/adjustments/save', methods=['POST'])
@_login_required
def api_pm_adj_save():
    """Create a new adjustment voucher OR update an existing pending/rejected
    one (idempotent re-save replaces all line items).

    Body:
      id          — optional; when present, update existing voucher
      adj_date    — YYYY-MM-DD; defaults to today
      godown_id   — required
      voucher_remarks — optional
      items       — required, list of:
        { product_id, direction:'increase'|'decrease', qty:>0, reason }

    If editing a 'rejected' voucher, this resaves it back to 'pending' and
    clears the rejection metadata (resubmit flow).
    """
    blocked = _block_if_no_access('stock_adjustment')
    if blocked is not None: return blocked

    d = request.get_json() or {}
    adj_id_in = d.get('id')
    godown_id = d.get('godown_id')
    adj_date  = (d.get('adj_date') or '').strip() or str(date.today())
    remarks   = (d.get('voucher_remarks') or '').strip()[:500]
    items_in  = d.get('items') or []

    if not godown_id:
        return jsonify({'status':'error','message':'godown_id required'}), 400
    if not isinstance(items_in, list) or not items_in:
        return jsonify({'status':'error','message':'At least one item is required'}), 400

    # Normalise items
    norm_items = []
    for ix, it in enumerate(items_in):
        try:
            pid = int(it.get('product_id') or 0)
            direction = (it.get('direction') or '').strip().lower()
            qty = float(it.get('qty') or 0)
            reason = (it.get('reason') or '').strip()[:500]
        except Exception:
            return jsonify({'status':'error','message':f'Row {ix+1}: bad numeric values'}), 400
        if pid <= 0:
            return jsonify({'status':'error','message':f'Row {ix+1}: pick a product'}), 400
        if direction not in ('increase','decrease'):
            return jsonify({'status':'error','message':f'Row {ix+1}: direction must be increase or decrease'}), 400
        if qty <= 0:
            return jsonify({'status':'error','message':f'Row {ix+1}: qty must be greater than zero'}), 400
        if not reason:
            return jsonify({'status':'error','message':f'Row {ix+1}: reason is required'}), 400
        norm_items.append({'product_id': pid, 'direction': direction,
                           'qty': qty, 'reason': reason})

    user = _user() or ''
    is_admin = _is_admin()

    conn = sampling_portal.get_db_connection()
    try:
        # Resolve location
        gdrow, is_floor = _adj_resolve_godown(conn, godown_id)
        if not gdrow:
            conn.close()
            return jsonify({'status':'error','message':'Invalid godown_id'}), 400

        # Validate products exist and are active
        pid_list = list({i['product_id'] for i in norm_items})
        placeholders = ','.join(['%s'] * len(pid_list))
        prod_rows = conn.execute(
            f"SELECT id FROM pm_products WHERE id IN ({placeholders}) AND is_active=1",
            tuple(pid_list)
        ).fetchall()
        valid_pids = {int(r['id']) for r in prod_rows}
        bad = [p for p in pid_list if p not in valid_pids]
        if bad:
            conn.close()
            return jsonify({'status':'error',
                            'message': f'Invalid or inactive product id(s): {bad}'}), 400

        # ─── UPDATE branch ─────────────────────────────────────────────
        if adj_id_in:
            try:
                adj_id = int(adj_id_in)
            except Exception:
                conn.close()
                return jsonify({'status':'error','message':'Bad id'}), 400
            existing = conn.execute(
                "SELECT * FROM pm_stock_adjustments WHERE id=%s", (adj_id,)
            ).fetchone()
            if not existing:
                conn.close()
                return jsonify({'status':'error','message':'Voucher not found'}), 404
            ok, reason = _adj_can_edit(dict(existing), user, is_admin)
            if not ok:
                conn.close()
                return jsonify({'status':'error','message': reason}), 403

            before_state = {
                'status': existing.get('status'),
                'godown_id': existing.get('godown_id'),
                'is_floor': existing.get('is_floor'),
                'adj_date': str(existing.get('adj_date') or ''),
                'voucher_remarks': existing.get('voucher_remarks'),
                'item_count': conn.execute(
                    "SELECT COUNT(*) AS c FROM pm_stock_adjustment_items WHERE adj_id=%s",
                    (adj_id,)
                ).fetchone()['c'],
            }
            was_rejected = (existing.get('status') or '').lower() == 'rejected'

            # Header update: edit always resets to 'pending' (clearing any
            # rejection state) so it goes back into the admin's queue.
            conn.execute(
                """UPDATE pm_stock_adjustments
                   SET adj_date=%s, godown_id=%s, is_floor=%s,
                       voucher_remarks=%s, status='pending',
                       rejected_by=NULL, rejected_at=NULL, reject_reason=NULL
                   WHERE id=%s""",
                (adj_date, gdrow['id'], 1 if is_floor else 0,
                 remarks, adj_id)
            )
            # Wipe + re-insert lines (simpler than per-row diff; safe because
            # no ledger rows exist yet on pending vouchers)
            conn.execute("DELETE FROM pm_stock_adjustment_items WHERE adj_id=%s", (adj_id,))
            for it in norm_items:
                conn.execute(
                    """INSERT INTO pm_stock_adjustment_items
                         (adj_id, product_id, direction, qty, reason)
                       VALUES (%s,%s,%s,%s,%s)""",
                    (adj_id, it['product_id'], it['direction'],
                     it['qty'], it['reason'])
                )
            action = 'stock_adj.resubmit' if was_rejected else 'stock_adj.update'
            try:
                _audit_record(
                    conn, action=action, entity='pm_stock_adjustment',
                    entity_id=adj_id,
                    summary=f"Adjustment {existing.get('adj_no')} "
                            f"{'resubmitted' if was_rejected else 'edited'} "
                            f"by {user} · {len(norm_items)} line(s)",
                    before=before_state,
                    after={'status':'pending', 'godown_id': gdrow['id'],
                           'is_floor': bool(is_floor), 'adj_date': adj_date,
                           'voucher_remarks': remarks,
                           'item_count': len(norm_items)},
                )
            except Exception: pass
            conn.commit()
            adj_no = existing.get('adj_no')
            conn.close()
            return jsonify({'status':'ok', 'mode':'updated',
                            'id': adj_id, 'adj_no': adj_no})

        # ─── CREATE branch ─────────────────────────────────────────────
        adj_no = _next_voucher_no(conn, 'PM-ADJ', adj_date)
        cur = conn.execute(
            """INSERT INTO pm_stock_adjustments
                 (adj_no, adj_date, godown_id, is_floor, status,
                  requested_by, voucher_remarks)
               VALUES (%s,%s,%s,%s,'pending',%s,%s)""",
            (adj_no, adj_date, gdrow['id'], 1 if is_floor else 0,
             user, remarks)
        )
        adj_id = cur.lastrowid
        for it in norm_items:
            conn.execute(
                """INSERT INTO pm_stock_adjustment_items
                     (adj_id, product_id, direction, qty, reason)
                   VALUES (%s,%s,%s,%s,%s)""",
                (adj_id, it['product_id'], it['direction'],
                 it['qty'], it['reason'])
            )
        try:
            _audit_record(
                conn, action='stock_adj.create', entity='pm_stock_adjustment',
                entity_id=adj_id,
                summary=f"Adjustment {adj_no} created by {user} "
                        f"at {gdrow['name']} · {len(norm_items)} line(s)",
                before=None,
                after={'adj_no': adj_no, 'godown_id': gdrow['id'],
                       'is_floor': bool(is_floor), 'adj_date': adj_date,
                       'item_count': len(norm_items)},
            )
        except Exception: pass
        conn.commit()
        conn.close()
        return jsonify({'status':'ok', 'mode':'created',
                        'id': adj_id, 'adj_no': adj_no})
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/adjustments/<int:adj_id>/delete', methods=['POST'])
@_login_required
def api_pm_adj_delete(adj_id):
    """Delete a pending/rejected adjustment voucher.

    Allowed if:
      - user is admin (any status except approved)
      - user is the requester AND voucher is still pending or rejected
    """
    blocked = _block_if_no_access('stock_adjustment')
    if blocked is not None: return blocked

    user = _user() or ''
    is_admin = _is_admin()

    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM pm_stock_adjustments WHERE id=%s", (adj_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'status':'error','message':'Voucher not found'}), 404
        if (row.get('status') or '').lower() == 'approved':
            conn.close()
            return jsonify({'status':'error',
                            'message':'Approved vouchers cannot be deleted '
                                      '(ledger entries exist).'}), 403
        if not is_admin and (row.get('requested_by') or '') != user:
            conn.close()
            return jsonify({'status':'error',
                            'message':'You can only delete your own vouchers'}), 403

        before = {k: _json_safe(v) for k, v in dict(row).items()}
        # CASCADE deletes item rows via FK
        conn.execute("DELETE FROM pm_stock_adjustments WHERE id=%s", (adj_id,))
        try:
            _audit_record(
                conn, action='stock_adj.delete', entity='pm_stock_adjustment',
                entity_id=adj_id,
                summary=f"Adjustment {row.get('adj_no')} deleted by {user}",
                before=before, after=None,
            )
        except Exception: pass
        conn.commit()
        conn.close()
        return jsonify({'status':'ok'})
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/adjustments/<int:adj_id>/approve', methods=['POST'])
@_login_required
def api_pm_adj_approve(adj_id):
    """Admin-only: approve a pending adjustment. Posts ledger rows for
    every line item and updates per-line godown_txn_id / floor_txn_id so
    each ledger entry stays traceable to its source voucher.

    Direction mapping:
      Godown:  increase → 'inward'    | decrease → 'outward'
      Floor:   increase → 'pm_return' | decrease → 'rejection'

    (Floor formula is: floor_opening + issue + pm_return - dispatch - rejection,
     so pm_return adds and rejection subtracts — the correct adjustment verbs.)
    """
    if not _is_admin():
        return jsonify({'status':'error',
                        'message':'Only an admin can approve stock adjustments'}), 403

    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM pm_stock_adjustments WHERE id=%s FOR UPDATE", (adj_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'status':'error','message':'Voucher not found'}), 404
        if (row.get('status') or '').lower() != 'pending':
            conn.close()
            return jsonify({'status':'error',
                            'message':f"Only pending vouchers can be approved "
                                      f"(this one is '{row.get('status')}')"}), 400

        items = conn.execute(
            """SELECT i.id, i.product_id, i.direction, i.qty, i.reason,
                      COALESCE(p.product_name,'') AS product_name
               FROM pm_stock_adjustment_items i
               LEFT JOIN pm_products p ON p.id=i.product_id
               WHERE i.adj_id=%s
               ORDER BY i.id""",
            (adj_id,)
        ).fetchall()
        if not items:
            conn.close()
            return jsonify({'status':'error',
                            'message':'Voucher has no line items'}), 400

        user = _user() or ''
        adj_date = str(row.get('adj_date') or date.today())
        adj_no   = row.get('adj_no') or ''
        gid      = int(row.get('godown_id') or 0)
        is_floor = bool(int(row.get('is_floor') or 0))

        posted = []
        for it in items:
            qty = float(it['qty'] or 0)
            direction = (it['direction'] or '').strip().lower()
            pid = int(it['product_id'])
            line_remark = f"[PM-ADJ:{adj_no}] {direction.upper()} · {it['reason']}"[:500]

            if is_floor:
                # increase → pm_return (adds to floor stock)
                # decrease → rejection (subtracts from floor stock)
                txn_type = 'pm_return' if direction == 'increase' else 'rejection'
                cur = conn.execute(
                    """INSERT INTO pm_floor_txn
                         (product_id, txn_date, txn_type, qty, remarks,
                          created_by, voucher_no, godown_id)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (pid, adj_date, txn_type, qty, line_remark,
                     user, adj_no, gid)
                )
                ftid = cur.lastrowid
                conn.execute(
                    "UPDATE pm_stock_adjustment_items SET floor_txn_id=%s WHERE id=%s",
                    (ftid, it['id'])
                )
                posted.append({'item_id': it['id'], 'floor_txn_id': ftid,
                               'txn_type': txn_type, 'qty': qty})
            else:
                # increase → inward, decrease → outward
                txn_type = 'inward' if direction == 'increase' else 'outward'
                cur = conn.execute(
                    """INSERT INTO pm_godown_txn
                         (product_id, txn_date, txn_type, qty, remarks,
                          created_by, voucher_no, godown_id)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (pid, adj_date, txn_type, qty, line_remark,
                     user, adj_no, gid)
                )
                gtid = cur.lastrowid
                conn.execute(
                    "UPDATE pm_stock_adjustment_items SET godown_txn_id=%s WHERE id=%s",
                    (gtid, it['id'])
                )
                posted.append({'item_id': it['id'], 'godown_txn_id': gtid,
                               'txn_type': txn_type, 'qty': qty})

        conn.execute(
            """UPDATE pm_stock_adjustments
               SET status='approved', approved_by=%s, approved_at=NOW()
               WHERE id=%s""",
            (user, adj_id)
        )
        try:
            _audit_record(
                conn, action='stock_adj.approve', entity='pm_stock_adjustment',
                entity_id=adj_id,
                summary=f"Adjustment {adj_no} APPROVED by {user} · "
                        f"{len(items)} ledger row(s) posted",
                before={'status':'pending'},
                after={'status':'approved', 'posted': posted},
            )
        except Exception: pass
        conn.commit()
        conn.close()
        return jsonify({'status':'ok', 'mode':'approved',
                        'id': adj_id, 'posted': posted})
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/adjustments/<int:adj_id>/reject', methods=['POST'])
@_login_required
def api_pm_adj_reject(adj_id):
    """Admin-only: reject a pending adjustment with a reason. Requester
    can subsequently edit + resave (which kicks status back to pending)."""
    if not _is_admin():
        return jsonify({'status':'error',
                        'message':'Only an admin can reject stock adjustments'}), 403

    d = request.get_json() or {}
    reject_reason = (d.get('reject_reason') or '').strip()[:500]
    if not reject_reason:
        return jsonify({'status':'error',
                        'message':'Reject reason is required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM pm_stock_adjustments WHERE id=%s", (adj_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'status':'error','message':'Voucher not found'}), 404
        if (row.get('status') or '').lower() != 'pending':
            conn.close()
            return jsonify({'status':'error',
                            'message':f"Only pending vouchers can be rejected "
                                      f"(this one is '{row.get('status')}')"}), 400

        user = _user() or ''
        conn.execute(
            """UPDATE pm_stock_adjustments
               SET status='rejected', rejected_by=%s, rejected_at=NOW(),
                   reject_reason=%s
               WHERE id=%s""",
            (user, reject_reason, adj_id)
        )
        try:
            _audit_record(
                conn, action='stock_adj.reject', entity='pm_stock_adjustment',
                entity_id=adj_id,
                summary=f"Adjustment {row.get('adj_no')} REJECTED by {user}: "
                        f"{reject_reason[:120]}",
                before={'status':'pending'},
                after={'status':'rejected', 'reject_reason': reject_reason},
            )
        except Exception: pass
        conn.commit()
        conn.close()
        return jsonify({'status':'ok', 'mode':'rejected', 'id': adj_id})
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message':str(e)}), 500


# ════════════════════════════════════════════════════════════════════════
# PM TESTING REQUISITION SLIPS (TRS)
# ────────────────────────────────────────────────────────────────────────
# Per-product certification slips raised against pm_grn lines.
#
# Flow:
#   1. User opens a GRN in the edit-GRN modal, checkboxes one or more
#      lines that share the same product, clicks "Generate TRS".
#   2. Frontend calls /trs/preflight to (a) confirm a TRS for those
#      lines doesn't already exist, and (b) get a list of fields that
#      need to be collected from the user (Physical State / Sample
#      Qty / Client Name — anything not in the GRN).
#   3. User fills the missing-data modal and the frontend calls
#      /trs/generate.  Server aggregates qty/box across the selected
#      lines, computes previous_supplier + new_or_old by looking at
#      prior GRNs for this product, and inserts one pm_grn_trs row.
#   4. The TRS is then editable (via /trs/<id>/update) and deletable
#      (via /trs/<id>/delete) until it leaves Pending — at which
#      point a 24h editability window starts running (enforced in
#      the QC-side approve endpoint, not here).  After lock, only an
#      admin can edit or delete.
#
# Approval / observation endpoints live in the QC blueprint (qc_bp).
# This file owns generation, preflight, editing, and the GRN-status
# lookup that decorates GRN line rows with "TRS generated" badges.
# ════════════════════════════════════════════════════════════════════════

# ── Internal helpers ──────────────────────────────────────────────────

def _pm_trs_lock_state(row, _now_dt=None):
    """Compute (is_locked, hours_remaining) for a pm_grn_trs row.

    A TRS row is locked when it has an approval_locked_at timestamp
    AND that timestamp is more than 24h old.  Before lock, the row is
    freely editable by its creator (and the QC user via the QC
    endpoints).  After lock, only an admin can change it.

    EXCEPTION: rows in 'Under Review' status are NEVER locked. Under
    Review is a non-terminal workflow state — the QC reviewer is still
    deciding the outcome, and the row will eventually flip to Approved
    or Rejected. The 24h lock countdown only applies to those two
    terminal states. (The upstream RM TRS portal historically set
    approval_locked_at on every non-Pending transition including Under
    Review; this guard makes our PM-side display correct regardless
    of what the upstream wrote.)
    """
    status = (row.get('approval_status') or '').strip().lower() if isinstance(row, dict) else ''
    if status in ('under review', 'under_review', 'underreview'):
        return False, None
    locked_at = row.get('approval_locked_at') if isinstance(row, dict) else None
    if not locked_at:
        return False, None
    try:
        from datetime import datetime as _dt
        # locked_at may already be a datetime, or a string — handle both.
        if isinstance(locked_at, str):
            lk = _dt.strptime(locked_at[:19], '%Y-%m-%d %H:%M:%S')
        else:
            lk = locked_at
        now = _now_dt or _dt.now()
        hrs = (now - lk).total_seconds() / 3600.0
        return (hrs >= 24.0), max(0.0, 24.0 - hrs)
    except Exception:
        return False, None


def _pm_trs_normalise_row(r, now_dt=None):
    """Coerce a pm_grn_trs row into a JSON-safe dict and add the
    is_locked / hours_remaining fields the frontend needs."""
    import json as _json
    d = {k: _json_safe(v) for k, v in dict(r).items()}
    # Parse checked_params JSON → list (frontend expects an array)
    cp = d.get('checked_params')
    if cp and isinstance(cp, str):
        try:
            parsed = _json.loads(cp)
            d['checked_params'] = parsed if isinstance(parsed, list) else [str(parsed)]
        except Exception:
            d['checked_params'] = []
    else:
        d['checked_params'] = []
    # Parse source_item_ids JSON → list of ints
    sids = d.get('source_item_ids')
    if sids and isinstance(sids, str):
        try:
            parsed = _json.loads(sids)
            d['source_item_ids'] = [int(x) for x in parsed] if isinstance(parsed, list) else []
        except Exception:
            d['source_item_ids'] = []
    else:
        d['source_item_ids'] = []
    # Locked flag
    locked, hrs = _pm_trs_lock_state(r if isinstance(r, dict) else dict(r), now_dt)
    d['is_locked']        = locked
    d['hours_remaining']  = hrs
    return d


def _pm_trs_resolve_previous_supplier(conn, product_id, current_grn_id):
    """For a given product, find the supplier from the most recent
    prior GRN (excluding the current GRN).  Returns (previous_supplier,
    new_or_old) where new_or_old is 'NEW' when there is no prior GRN
    or the prior supplier differs from the current one, else 'OLD'."""
    if not product_id:
        return '', 'NEW'
    try:
        # Pull current GRN's supplier first
        cur = conn.execute(
            "SELECT supplier FROM pm_grn WHERE id=%s",
            (int(current_grn_id),)
        ).fetchone()
        cur_sup = (cur or {}).get('supplier') or ''
        # Now look for the latest prior GRN with this product
        prev = conn.execute(
            """SELECT g.supplier, g.grn_date, g.id
               FROM pm_grn g
               JOIN pm_grn_items i ON i.grn_id = g.id
               WHERE i.product_id = %s AND g.id <> %s
               ORDER BY g.grn_date DESC, g.id DESC
               LIMIT 1""",
            (int(product_id), int(current_grn_id))
        ).fetchone()
        if not prev:
            return '', 'NEW'
        prev_sup = (prev.get('supplier') or '').strip()
        return prev_sup, ('OLD' if prev_sup and prev_sup == (cur_sup or '').strip() else 'NEW')
    except Exception:
        return '', 'NEW'


def _pm_trs_aggregate_lines(conn, grn_id, line_ids):
    """Given a GRN id and a list of pm_grn_items.id values, return
    (ok, error_or_aggregate).

    Validates that:
      - all line_ids belong to the given grn_id
      - all line_ids reference the SAME product (cross-product TRS is
        explicitly blocked per Tarak's spec)

    On success returns dict with:
      product_id, product_name, product_code, pm_type, primary_uom,
      no_of_box  (sum across lines),
      total_qty  (sum across lines),
      qty_per_pkg (total_qty / no_of_box, or first line's per-box
                   when no_of_box is 0 to avoid divide-by-zero),
      lines      (list of normalised line dicts for the missing-data
                  modal preview)
    """
    if not line_ids:
        return False, 'No line ids supplied'
    placeholders = ','.join(['%s'] * len(line_ids))
    rows = conn.execute(
        f"""SELECT i.id, i.grn_id, i.product_id, i.qty_received,
                   COALESCE(i.no_of_box, 0)   AS no_of_box,
                   COALESCE(i.box_count, 0)   AS box_count,
                   i.remarks,
                   p.product_name, COALESCE(p.product_code, '') AS product_code,
                   COALESCE(p.pm_type, '')      AS pm_type,
                   COALESCE(p.primary_uom, '')  AS primary_uom
            FROM pm_grn_items i
            JOIN pm_products p ON p.id = i.product_id
            WHERE i.id IN ({placeholders})""",
        tuple(line_ids)
    ).fetchall()
    if not rows:
        return False, 'No matching line items found'
    if len(rows) != len(line_ids):
        return False, 'Some line ids do not exist or were deleted'
    # Validate same GRN
    for r in rows:
        if int(r['grn_id']) != int(grn_id):
            return False, 'Line ids must all belong to the same GRN'
    # Validate same product
    pids = {int(r['product_id']) for r in rows}
    if len(pids) > 1:
        return False, ('Cannot generate one TRS across different products. '
                       'Select lines of the SAME item only.')
    first = rows[0]
    total_qty   = sum(float(r['qty_received'] or 0) for r in rows)
    # Prefer no_of_box, falling back to box_count for older rows
    total_boxes = sum(float(r['no_of_box'] or r['box_count'] or 0) for r in rows)
    if total_boxes > 0:
        qty_per_pkg = round(total_qty / total_boxes, 3)
    else:
        # Avoid divide-by-zero — fall back to the first row's snapshot
        qty_per_pkg = float(first['qty_received'] or 0)
    return True, {
        'product_id':    int(first['product_id']),
        'product_name':  first['product_name'] or '',
        'product_code':  first['product_code'] or '',
        'pm_type':       first['pm_type'] or '',
        'primary_uom':   first['primary_uom'] or '',
        'no_of_box':     total_boxes,
        'total_qty':     total_qty,
        'qty_per_pkg':   qty_per_pkg,
        'lines':         [{k: _json_safe(v) for k, v in dict(r).items()} for r in rows],
    }


def _pm_trs_existing_for_lines(conn, line_ids, grn_id=None):
    """Return the TRS id(s) that already cover any of the given line_ids
    (either as the canonical grn_item_id OR within source_item_ids JSON).

    When `grn_id` is provided (it should be, in practice) we scope the
    candidate set to that GRN — otherwise we'd scan the entire pm_grn_trs
    table on every preflight call. The function still works without it
    (legacy callers, tests).
    """
    import json as _json
    if not line_ids:
        return []
    placeholders = ','.join(['%s'] * len(line_ids))
    if grn_id:
        sql = (f"SELECT id, trs_num, grn_item_id, source_item_ids, approval_status "
               f"FROM pm_grn_trs WHERE grn_id=%s "
               f"AND (grn_item_id IN ({placeholders}) OR source_item_ids IS NOT NULL)")
        params = (int(grn_id), *line_ids)
    else:
        sql = (f"SELECT id, trs_num, grn_item_id, source_item_ids, approval_status "
               f"FROM pm_grn_trs "
               f"WHERE grn_item_id IN ({placeholders}) OR source_item_ids IS NOT NULL")
        params = tuple(line_ids)
    rows = conn.execute(sql, params).fetchall()
    line_set = {int(x) for x in line_ids}
    matches = []
    for r in rows:
        if r['grn_item_id'] and int(r['grn_item_id']) in line_set:
            matches.append({'id': int(r['id']), 'trs_num': r['trs_num'],
                            'status': r['approval_status'],
                            'matched_lines': [int(r['grn_item_id'])]})
            continue
        sids = r.get('source_item_ids')
        if sids:
            try:
                parsed = _json.loads(sids)
                shared = [x for x in parsed if int(x) in line_set]
                if shared:
                    matches.append({'id': int(r['id']), 'trs_num': r['trs_num'],
                                    'status': r['approval_status'],
                                    'matched_lines': [int(x) for x in shared]})
            except Exception:
                pass
    return matches


# ── Endpoints ────────────────────────────────────────────────────────

@pm_stock_bp.route('/api/pm_stock/trs/list', methods=['GET'])
@_login_required
def api_pm_trs_list():
    """List PM TRS rows. Supports optional filters:
      status   = 'Pending'|'Approved'|'Rejected'|'Under Review'|'All'
      grn_id   = N    (only TRS for one specific GRN)
      limit    = N    (default 500, max 2000)
    """
    blocked = _block_if_no_access('pm_trs')
    if blocked is not None: return blocked

    status = (request.args.get('status') or 'All').strip()
    try:
        grn_id = int(request.args.get('grn_id') or 0)
    except Exception:
        grn_id = 0
    try:
        limit = max(1, min(2000, int(request.args.get('limit') or 500)))
    except Exception:
        limit = 500

    where = []
    params = []
    if status and status != 'All':
        if status not in ('Pending', 'Approved', 'Rejected', 'Under Review'):
            return jsonify({'status':'error','message':'Bad status filter'}), 400
        where.append('approval_status=%s')
        params.append(status)
    if grn_id:
        where.append('grn_id=%s')
        params.append(grn_id)
    sql_where = (' WHERE ' + ' AND '.join(where)) if where else ''

    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(
            f"""SELECT id, trs_num, grn_id, grn_item_id, source_item_ids,
                       grn_num, grn_date,
                       product_id, material, product_code, pm_type,
                       no_of_box, qty_per_pkg, total_qty, uom,
                       supplier_name, previous_supplier, new_or_old,
                       physical_state, sample_qty, client_name,
                       generated_by, generated_at, verified_by,
                       approval_status, approved_by, approval_dt,
                       approval_remarks, approval_locked_at,
                       checked_params, rejection_reason
                FROM pm_grn_trs
                {sql_where}
                ORDER BY generated_at DESC, id DESC
                LIMIT {limit}""",
            tuple(params)
        ).fetchall() or []
        from datetime import datetime as _dt
        now_dt = _dt.now()
        out = [_pm_trs_normalise_row(r, now_dt) for r in rows]
        conn.close()
        return jsonify({'status':'ok', 'trs': out,
                        'is_admin': _is_admin(), 'me': _user() or ''})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/trs/<int:trs_id>', methods=['GET'])
@_login_required
def api_pm_trs_detail(trs_id):
    """Full TRS detail incl. parsed checked_params + lock state."""
    blocked = _block_if_no_access('pm_trs')
    if blocked is not None: return blocked

    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM pm_grn_trs WHERE id=%s",
            (trs_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'status':'error','message':'TRS not found'}), 404
        from datetime import datetime as _dt
        d = _pm_trs_normalise_row(row, _dt.now())
        conn.close()
        return jsonify({'status':'ok', 'trs': d,
                        'is_admin': _is_admin(), 'me': _user() or ''})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/trs/preflight', methods=['POST'])
@_login_required
def api_pm_trs_preflight():
    """Pre-generation validation.  Returns one of:

      { ok:true, mode:'ready', aggregate:{...}, missing_fields:[...],
        defaults:{...}, previous_supplier:'...', new_or_old:'NEW'|'OLD' }
        — proceed to show the missing-data modal.

      { ok:false, mode:'duplicate', existing:[{id, trs_num, status, ...}] }
        — one or more selected lines already have a TRS. Frontend
          shows a warning with link to open existing TRS instead.

      { ok:false, mode:'invalid', message:'...' }
        — line ids don't validate (cross-product, missing, etc.)

    Body: { grn_id:int, line_ids:[int,...] }
    """
    blocked = _block_if_no_access('pm_trs')
    if blocked is not None: return blocked

    d = request.get_json(silent=True) or {}
    try:
        grn_id = int(d.get('grn_id') or 0)
        line_ids = [int(x) for x in (d.get('line_ids') or [])]
    except Exception:
        return jsonify({'status':'error','message':'Bad payload'}), 400
    if not grn_id or not line_ids:
        return jsonify({'status':'error',
                        'message':'grn_id and line_ids required'}), 400

    # When the user has already acknowledged the duplicate (via the
    # "Force create another anyway" button), the frontend re-runs
    # preflight with skip_duplicate_check=true so it can get the
    # aggregate + missing-fields needed for the generate modal.
    skip_dup = bool(d.get('skip_duplicate_check'))

    conn = sampling_portal.get_db_connection()
    try:
        # Existing-TRS check first — most useful early signal
        if not skip_dup:
            existing = _pm_trs_existing_for_lines(conn, line_ids, grn_id=grn_id)
            if existing:
                conn.close()
                return jsonify({'status':'ok', 'ok': False,
                                'mode':'duplicate', 'existing': existing})

        # Validate + aggregate
        ok, agg_or_err = _pm_trs_aggregate_lines(conn, grn_id, line_ids)
        if not ok:
            conn.close()
            return jsonify({'status':'ok', 'ok': False,
                            'mode':'invalid', 'message': agg_or_err})

        agg = agg_or_err

        # GRN header
        grn = conn.execute(
            "SELECT grn_no, grn_date, supplier FROM pm_grn WHERE id=%s",
            (grn_id,)
        ).fetchone()
        if not grn:
            conn.close()
            return jsonify({'status':'ok', 'ok': False,
                            'mode':'invalid',
                            'message': 'GRN not found'}), 200

        # Previous supplier resolution
        prev_sup, new_or_old = _pm_trs_resolve_previous_supplier(
            conn, agg['product_id'], grn_id)

        # Always-prompted fields. Defaults are sensible starting points.
        missing_fields = [
            {'key':'physical_state',  'label':'Physical State',
             'type':'select',  'default':'OK',
             'options':['OK','Damaged','Wet','Contaminated','Other']},
            {'key':'sample_qty',      'label':'Sample Qty',
             'type':'number',  'default':1, 'min':0, 'step':0.01},
            {'key':'client_name',     'label':'Client Name (optional)',
             'type':'text',    'default':''},
        ]
        conn.close()

        # Build server-derived defaults that the modal pre-fills
        return jsonify({
            'status':            'ok',
            'ok':                True,
            'mode':              'ready',
            'grn':  {
                'id':         grn_id,
                'grn_no':     grn.get('grn_no') or '',
                'grn_date':   str(grn.get('grn_date') or '')[:10],
                'supplier':   grn.get('supplier') or '',
            },
            'aggregate':         agg,
            'previous_supplier': prev_sup,
            'new_or_old':        new_or_old,
            'missing_fields':    missing_fields,
            'defaults':          {
                'verified_by':    _user() or '',
            },
        })
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/trs/generate', methods=['POST'])
@_login_required
def api_pm_trs_generate():
    """Create a new pm_grn_trs row from the data the user filled in
    the missing-data modal + the auto-aggregated line data.

    Body:
      { grn_id, line_ids, physical_state, sample_qty, client_name,
        verified_by (optional), force_duplicate (bool, default false) }

    When force_duplicate is false and a duplicate exists, the request
    is rejected with 409.  (Frontend should call /preflight first to
    catch this naturally.)
    """
    blocked = _block_if_no_access('pm_trs')
    if blocked is not None: return blocked

    d = request.get_json(silent=True) or {}
    try:
        grn_id   = int(d.get('grn_id') or 0)
        line_ids = [int(x) for x in (d.get('line_ids') or [])]
    except Exception:
        return jsonify({'status':'error','message':'Bad payload'}), 400
    if not grn_id or not line_ids:
        return jsonify({'status':'error',
                        'message':'grn_id and line_ids required'}), 400

    phys_state   = (d.get('physical_state') or 'OK').strip()[:50]
    try:
        sample_qty = float(d.get('sample_qty') or 1)
    except Exception:
        sample_qty = 1.0
    if sample_qty < 0:
        sample_qty = 0.0
    client_name = (d.get('client_name') or '').strip()[:300]
    verified_by = (d.get('verified_by') or _user() or '').strip()[:100]
    force_duplicate = bool(d.get('force_duplicate'))

    conn = sampling_portal.get_db_connection()
    try:
        # Duplicate check unless explicitly forced
        if not force_duplicate:
            existing = _pm_trs_existing_for_lines(conn, line_ids, grn_id=grn_id)
            if existing:
                conn.close()
                return jsonify({
                    'status': 'error',
                    'code':   'duplicate',
                    'message': ('A TRS already exists for one or more of the '
                                'selected lines.'),
                    'existing': existing,
                }), 409

        ok, agg_or_err = _pm_trs_aggregate_lines(conn, grn_id, line_ids)
        if not ok:
            conn.close()
            return jsonify({'status':'error', 'message': agg_or_err}), 400
        agg = agg_or_err

        # GRN header
        grn = conn.execute(
            "SELECT grn_no, grn_date, supplier FROM pm_grn WHERE id=%s",
            (grn_id,)
        ).fetchone()
        if not grn:
            conn.close()
            return jsonify({'status':'error','message':'GRN not found'}), 404
        grn_num  = grn.get('grn_no')  or ''
        grn_date = grn.get('grn_date')
        supplier = grn.get('supplier') or ''

        prev_sup, new_or_old = _pm_trs_resolve_previous_supplier(
            conn, agg['product_id'], grn_id)

        # Mint TRS number
        trs_num = _next_voucher_no(conn, 'PM-TRS',
                                   ref_date=grn_date or None)

        import json as _json
        source_item_ids_json = _json.dumps(line_ids)
        # Pick the FIRST line as the canonical grn_item_id for indexing.
        # Aggregated set lives in source_item_ids.
        canonical_item_id = line_ids[0]

        cur = conn.execute(
            """INSERT INTO pm_grn_trs (
                 trs_num, grn_id, grn_item_id, source_item_ids,
                 grn_num, grn_date,
                 product_id, material, product_code, pm_type,
                 no_of_box, qty_per_pkg, total_qty, uom,
                 supplier_name, previous_supplier, new_or_old,
                 physical_state, sample_qty, client_name,
                 generated_by, verified_by, approval_status
               ) VALUES (
                 %s, %s, %s, %s,
                 %s, %s,
                 %s, %s, %s, %s,
                 %s, %s, %s, %s,
                 %s, %s, %s,
                 %s, %s, %s,
                 %s, %s, 'Pending'
               )""",
            (trs_num, grn_id, canonical_item_id, source_item_ids_json,
             grn_num, grn_date,
             agg['product_id'], agg['product_name'], agg['product_code'],
             agg['pm_type'],
             agg['no_of_box'], agg['qty_per_pkg'], agg['total_qty'],
             agg['primary_uom'],
             supplier, prev_sup, new_or_old,
             phys_state, sample_qty, client_name,
             _user() or '', verified_by)
        )
        trs_id = cur.lastrowid

        try:
            _audit_record(
                conn, action='pm_trs.generate', entity='pm_grn_trs',
                entity_id=trs_id,
                summary=(f"TRS {trs_num} generated for {agg['product_name']} "
                         f"on GRN {grn_num} ({len(line_ids)} line(s), "
                         f"total qty {agg['total_qty']:g})"),
                before=None,
                after={'trs_num': trs_num, 'grn_id': grn_id,
                       'line_ids': line_ids,
                       'product_id': agg['product_id'],
                       'product_name': agg['product_name'],
                       'total_qty': agg['total_qty'],
                       'no_of_box': agg['no_of_box']},
            )
        except Exception: pass
        conn.commit()
        conn.close()
        return jsonify({'status':'ok',
                        'mode':'created',
                        'id':       trs_id,
                        'trs_num':  trs_num})
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/trs/<int:trs_id>/update', methods=['POST'])
@_login_required
def api_pm_trs_update(trs_id):
    """Edit a TRS row.

    Permissions:
      - Before approval_locked_at is set: any user with pm_trs access
        may edit (this is the editable-while-Pending window).
      - After lock & status not Pending: admin only.
      - After lock+24h: admin only.

    Editable fields:
      physical_state, sample_qty, client_name, verified_by,
      supplier_name (admin only), previous_supplier (admin only),
      new_or_old (admin only).

    Quantity / box / line aggregation cannot be edited — to change
    those, delete & regenerate (per Tarak's confirmation that delete +
    recreate is acceptable; he kept the RM-style flow which only
    edits the manual fields, never the GRN-sourced ones).
    """
    blocked = _block_if_no_access('pm_trs')
    if blocked is not None: return blocked

    d = request.get_json(silent=True) or {}
    user     = _user() or ''
    is_admin = _is_admin()

    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM pm_grn_trs WHERE id=%s",
            (trs_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'status':'error','message':'TRS not found'}), 404

        rd = dict(row)
        # Compute lock state
        locked, _hrs = _pm_trs_lock_state(rd)
        if (rd.get('approval_status') or '').lower() != 'pending':
            if not is_admin:
                conn.close()
                return jsonify({
                    'status':'error',
                    'message': ('This TRS has already been decided. '
                                'Only an admin can edit it.'),
                }), 423
        if locked and not is_admin:
            conn.close()
            return jsonify({
                'status':'error',
                'message': ('This TRS is locked (more than 24 hours since '
                            'the QC decision). Admin override required.'),
            }), 423

        # Build update set
        sets = []
        params = []
        before = {}
        after  = {}
        def _capture(k, new):
            old = rd.get(k)
            if old != new:
                before[k] = old
                after[k]  = new

        if 'physical_state' in d:
            ns = (d.get('physical_state') or '').strip()[:50]
            _capture('physical_state', ns)
            sets.append('physical_state=%s'); params.append(ns)
        if 'sample_qty' in d:
            try: sq = float(d.get('sample_qty') or 0)
            except Exception: sq = 0.0
            if sq < 0: sq = 0.0
            _capture('sample_qty', sq)
            sets.append('sample_qty=%s'); params.append(sq)
        if 'client_name' in d:
            cn = (d.get('client_name') or '').strip()[:300]
            _capture('client_name', cn)
            sets.append('client_name=%s'); params.append(cn)
        if 'verified_by' in d:
            vb = (d.get('verified_by') or '').strip()[:100]
            _capture('verified_by', vb)
            sets.append('verified_by=%s'); params.append(vb)

        # Admin-only fields (supplier metadata)
        if is_admin:
            if 'supplier_name' in d:
                v = (d.get('supplier_name') or '').strip()[:300]
                _capture('supplier_name', v)
                sets.append('supplier_name=%s'); params.append(v)
            if 'previous_supplier' in d:
                v = (d.get('previous_supplier') or '').strip()[:300]
                _capture('previous_supplier', v)
                sets.append('previous_supplier=%s'); params.append(v)
            if 'new_or_old' in d:
                v = (d.get('new_or_old') or 'OLD').strip().upper()[:10]
                if v not in ('NEW','OLD'): v = 'OLD'
                _capture('new_or_old', v)
                sets.append('new_or_old=%s'); params.append(v)

        if not sets:
            conn.close()
            return jsonify({'status':'ok','message':'No changes'}), 200

        params.append(trs_id)
        conn.execute(
            f"UPDATE pm_grn_trs SET {', '.join(sets)} WHERE id=%s",
            tuple(params)
        )
        # Audit
        try:
            action = 'pm_trs.admin_override' if (locked and is_admin) else 'pm_trs.update'
            _audit_record(
                conn, action=action, entity='pm_grn_trs',
                entity_id=trs_id,
                summary=(f"TRS {rd.get('trs_num')} edited by {user}"
                        + (' (admin past lock)' if locked and is_admin else '')),
                before=before, after=after,
            )
        except Exception: pass
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','mode':'updated','id': trs_id})
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/trs/<int:trs_id>/delete', methods=['POST'])
@_login_required
def api_pm_trs_delete(trs_id):
    """Delete a TRS.  Same lock rules as update — admin only after
    the row has been decided + 24h.  Associated pm_trs_observations
    are CASCADE-deleted via FK."""
    blocked = _block_if_no_access('pm_trs')
    if blocked is not None: return blocked

    user     = _user() or ''
    is_admin = _is_admin()

    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM pm_grn_trs WHERE id=%s",
            (trs_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'status':'error','message':'TRS not found'}), 404
        rd = dict(row)
        locked, _hrs = _pm_trs_lock_state(rd)
        if (rd.get('approval_status') or '').lower() != 'pending' and not is_admin:
            conn.close()
            return jsonify({
                'status':'error',
                'message':('This TRS has already been decided by QC. '
                          'Only an admin can delete it.'),
            }), 423
        if locked and not is_admin:
            conn.close()
            return jsonify({
                'status':'error',
                'message':('TRS is locked (24h past QC decision). '
                          'Admin override required.'),
            }), 423

        before = {k: _json_safe(v) for k, v in rd.items()}
        conn.execute("DELETE FROM pm_grn_trs WHERE id=%s", (trs_id,))
        try:
            _audit_record(
                conn, action='pm_trs.delete', entity='pm_grn_trs',
                entity_id=trs_id,
                summary=f"TRS {rd.get('trs_num')} deleted by {user}",
                before=before, after=None,
            )
        except Exception: pass
        conn.commit()
        conn.close()
        return jsonify({'status':'ok','mode':'deleted','id': trs_id})
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message': str(e)}), 500


@pm_stock_bp.route('/api/pm_stock/trs/grn_status', methods=['GET'])
@_login_required
def api_pm_trs_grn_status():
    """For a given GRN, return a map of which lines have a TRS.

    Used by the edit-GRN modal to badge "TRS generated" pills on each
    line row.  Returns:

      { ok:true, by_line:{ "<grn_item_id>": { trs_id, trs_num, status,
                                              is_locked, hours_remaining } } }

    A single TRS can cover multiple lines (when same-product lines were
    aggregated at generation time); each covered line appears in the map.
    """
    blocked = _block_if_no_access('pm_trs')
    if blocked is not None: return blocked

    try:
        grn_id = int(request.args.get('grn_id') or 0)
    except Exception:
        grn_id = 0
    if not grn_id:
        return jsonify({'status':'error','message':'grn_id required'}), 400

    conn = sampling_portal.get_db_connection()
    try:
        rows = conn.execute(
            """SELECT id, trs_num, grn_item_id, source_item_ids,
                      approval_status, approval_locked_at, generated_at,
                      approved_by, approval_dt
               FROM pm_grn_trs
               WHERE grn_id=%s""",
            (grn_id,)
        ).fetchall() or []
        conn.close()

        from datetime import datetime as _dt
        now_dt = _dt.now()
        import json as _json
        by_line = {}
        for r in rows:
            rd = dict(r)
            locked, hrs = _pm_trs_lock_state(rd, now_dt)
            covered = set()
            if rd.get('grn_item_id'):
                covered.add(int(rd['grn_item_id']))
            sids = rd.get('source_item_ids')
            if sids:
                try:
                    parsed = _json.loads(sids)
                    for x in (parsed or []):
                        try: covered.add(int(x))
                        except Exception: pass
                except Exception:
                    pass
            entry = {
                'trs_id':           int(rd['id']),
                'trs_num':          rd.get('trs_num') or '',
                'status':           rd.get('approval_status') or 'Pending',
                'is_locked':        locked,
                'hours_remaining':  hrs,
                'approved_by':      rd.get('approved_by'),
                'approval_dt':      _json_safe(rd.get('approval_dt')),
                'generated_at':     _json_safe(rd.get('generated_at')),
            }
            for line_id in covered:
                # If a line somehow has multiple TRS rows, keep the most
                # recent one — generation order is roughly id-ascending.
                prev = by_line.get(str(line_id))
                if (not prev) or entry['trs_id'] > prev['trs_id']:
                    by_line[str(line_id)] = entry
        return jsonify({'status':'ok', 'by_line': by_line})
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════
# PM TRS — PDF rendering
# ────────────────────────────────────────────────────────────────────────
# Mirrors the TESTING REQUISITION SLIP (TRS) form provided by HCP
# (QR-854-07 R-00). One page A4, structured cells with header / item
# block / physical verification table / signature footer.
#
# Uses reportlab (already used by qc_routes for the in-process PDF), so
# no new dependency.
# ════════════════════════════════════════════════════════════════════════

def _draw_pm_trs_pdf(buf, trs_row):
    """Render one PM TRS as a single-page A4 PDF — same layout as before,
    just with more breathing room so labels and values sit comfortably
    within their cells.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    import re as _re

    W, H = A4
    c = canvas.Canvas(buf, pagesize=A4)
    c.setTitle(f"TRS {trs_row.get('trs_num') or ''}")

    # ── Date helper ─────────────────────────────────────────────────
    def dmy(v):
        if not v: return ''
        s = str(v)
        m = _re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
        if m:
            return f'{m.group(3)}-{m.group(2)}-{m.group(1)}'
        return s

    def fmt_num(v, max_dp=3):
        try:
            f = float(v)
            if f == int(f) and max_dp >= 0:
                return f'{f:.{max_dp}f}' if max_dp > 0 else f'{int(f)}'
            return f'{f:.{max_dp}f}'.rstrip('0').rstrip('.')
        except Exception:
            return str(v or '')

    margin = 12 * mm
    table_x = margin
    table_w = W - 2 * margin

    # ── Drawing primitives ──────────────────────────────────────────
    def hline(y_):
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.6)
        c.line(table_x, y_, table_x + table_w, y_)

    def vline(x_, y1, y2):
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.6)
        c.line(x_, y1, x_, y2)

    # Cell text helper. y_ is the TOP of the cell; we render the text
    # vertically centered using cap-height math (baseline ≈ bottom + h/2 −
    # cap_height/2). For Helvetica at 8.5pt cap height is ≈ 5.95pt ≈ 2.1mm.
    def cell(x_, y_, w, h, label='', value='', label_bold=True, value_bold=False,
             label_size=8, value_size=9, label_align='left', value_align='left',
             fill=None, value_color=colors.black):
        if fill:
            c.setFillColor(fill)
            c.rect(x_, y_ - h, w, h, stroke=0, fill=1)
            c.setFillColor(colors.black)
        # vertical baseline: bottom of cell + h/2 − cap_height/2
        # cap_height for both label and value is roughly equal in pt; use
        # the larger of the two so things stay aligned.
        bigger = max(label_size, value_size)
        cap_pt = bigger * 0.72            # rough cap height
        ty = y_ - h + (h - cap_pt) / 2
        if label:
            c.setFont('Helvetica-Bold' if label_bold else 'Helvetica', label_size)
            c.setFillColor(colors.black)
            if label_align == 'center':
                c.drawCentredString(x_ + w / 2, ty, label)
            elif label_align == 'right':
                c.drawRightString(x_ + w - 2 * mm, ty, label)
            else:
                c.drawString(x_ + 2 * mm, ty, label)
        if value:
            c.setFont('Helvetica-Bold' if value_bold else 'Helvetica', value_size)
            c.setFillColor(value_color)
            # Truncate to fit column. Helvetica averages ~1.4mm per char
            # at 9pt — was 1.7 (too conservative, caused truncation of
            # values that would have fit). Still leaves 2mm padding each side.
            max_chars = max(6, int((w - 4 * mm) / (1.4 * mm)))
            shown = value if len(value) <= max_chars else value[:max_chars - 1] + '…'
            if value_align == 'center':
                c.drawCentredString(x_ + w / 2, ty, shown)
            elif value_align == 'right':
                c.drawRightString(x_ + w - 2 * mm, ty, shown)
            else:
                c.drawString(x_ + 2 * mm, ty, shown)
            c.setFillColor(colors.black)

    # ══════════════════════════════════════════════════════════════
    # HEADER BAND — logo cell + title cell. Generous height so the
    # title and doc-code don't stack on top of each other.
    # ══════════════════════════════════════════════════════════════
    y = H - margin
    hdr_h = 22 * mm                       # was 16mm — gives both lines breathing space
    logo_w = 32 * mm

    # Logo cell
    c.setStrokeColor(colors.black)
    c.setLineWidth(0.6)
    c.rect(table_x, y - hdr_h, logo_w, hdr_h, stroke=1, fill=0)
    # Stack: HCP (bold) / company name underneath. Pulled apart so the
    # two lines breathe — HCP sits in the upper third of the cell, the
    # company name in the lower third, with a healthy gap between them.
    c.setFillColor(colors.HexColor('#1f2937'))
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(table_x + logo_w / 2, y - hdr_h / 2 + 4.5, 'HCP')
    c.setFont('Helvetica', 6.5)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawCentredString(table_x + logo_w / 2, y - hdr_h / 2 - 7.5, 'HCP WELLNESS PVT. LTD.')
    c.setFillColor(colors.black)

    # Title cell — split vertically into title (upper) and doc-code (lower).
    # Same breathing-room treatment as the logo cell: title sits in the
    # upper third, doc-code in the lower third.
    title_x = table_x + logo_w
    title_w = table_w - logo_w
    c.rect(title_x, y - hdr_h, title_w, hdr_h, stroke=1, fill=0)
    c.setFont('Helvetica-Bold', 14)
    c.drawCentredString(title_x + title_w / 2, y - hdr_h / 2 + 4.5, 'TESTING REQUISITION SLIP (TRS)')
    c.setFont('Helvetica-Bold', 9)
    c.setFillColor(colors.HexColor('#374151'))
    c.drawCentredString(title_x + title_w / 2, y - hdr_h / 2 - 8.5, 'QR-854-07    R-00')
    c.setFillColor(colors.black)

    y -= hdr_h

    # ══════════════════════════════════════════════════════════════
    # Standard row dimensions
    # ══════════════════════════════════════════════════════════════
    row_h   = 11 * mm                    # was 9mm — comfortably hosts 10pt text
    col_w   = table_w / 3.0
    label_w = 30 * mm                    # was 28mm

    # Helper to render a 3-column row: each column is label | value
    def row_3col(triples, y_, h=row_h):
        """triples is a list of 3 tuples (label, value, label_size, value_bold)."""
        c.rect(table_x, y_ - h, table_w, h, stroke=1, fill=0)
        for i in range(3):
            x0 = table_x + i * col_w
            # column separator
            if i > 0:
                vline(x0, y_, y_ - h)
            # label / value separator inside the column
            vline(x0 + label_w, y_, y_ - h)
        for i, t in enumerate(triples):
            label, value, lbl_sz, val_bold = t
            x0 = table_x + i * col_w
            cell(x0, y_, label_w, h, label=label, label_align='right', label_size=lbl_sz)
            cell(x0 + label_w, y_, col_w - label_w, h,
                 value=value, value_bold=val_bold)

    def row_full(label, value, y_, h=row_h, value_bold=False, lbl_sz=8):
        """A single label across the row with one wide value."""
        c.rect(table_x, y_ - h, table_w, h, stroke=1, fill=0)
        vline(table_x + label_w, y_, y_ - h)
        cell(table_x, y_, label_w, h, label=label, label_align='right', label_size=lbl_sz)
        cell(table_x + label_w, y_, table_w - label_w, h,
             value=value, value_bold=value_bold)

    # ── ROW: TRS No | TRS Date | Department ─────────────────────────
    row_3col([
        ('TRS No',     str(trs_row.get('trs_num') or ''),                          8, True),
        ('TRS Date',   dmy(trs_row.get('generated_at') or trs_row.get('grn_date')),8, False),
        ('Department', 'PM STORE',                                                 8, True),
    ], y)
    y -= row_h

    # ── ROW: Name of Item (full width) ──────────────────────────────
    row_full('Name of Item', str(trs_row.get('material') or ''), y, value_bold=True)
    y -= row_h

    # ── ROW: No. pkt./Box/Roll | GRN No | GRN Date ──────────────────
    row_3col([
        ('NO. pkt. / Box / Roll', fmt_num(trs_row.get('no_of_box')),       7, False),
        ('GRN No',                str(trs_row.get('grn_num') or ''),       8, False),
        ('GRN Date',              dmy(trs_row.get('grn_date')),            8, False),
    ], y)
    y -= row_h

    # ── ROW: Client Name (full width) ───────────────────────────────
    row_full('Client Name', str(trs_row.get('client_name') or ''), y)
    y -= row_h

    # ── ROW: Physical State | Sample Qty | (blank) ──────────────────
    # Third column intentionally blank — matches the original form
    # layout (the form has a blank pair on the right of this row).
    c.rect(table_x, y - row_h, table_w, row_h, stroke=1, fill=0)
    vline(table_x + col_w,     y, y - row_h)
    vline(table_x + 2 * col_w, y, y - row_h)
    vline(table_x + label_w,             y, y - row_h)
    vline(table_x + col_w + label_w,     y, y - row_h)
    cell(table_x,             y, label_w, row_h, label='Physical State', label_align='right')
    cell(table_x + col_w,     y, label_w, row_h, label='Sample Qty.',    label_align='right')
    cell(table_x + label_w,             y, col_w - label_w, row_h,
         value=str(trs_row.get('physical_state') or 'OK'))
    cell(table_x + col_w + label_w,     y, col_w - label_w, row_h,
         value=fmt_num(trs_row.get('sample_qty'), max_dp=0))
    y -= row_h

    # ── ROW: Qty. Per Box | Total Qty | Supplier Name ───────────────
    row_3col([
        ('Qty. Per Box',  fmt_num(trs_row.get('qty_per_pkg')),                       8, False),
        ('Total Qty.',    fmt_num(trs_row.get('total_qty'), max_dp=0),               8, True),
        ('Supplier Name', str(trs_row.get('supplier_name') or ''),                   7, False),
    ], y)
    y -= row_h

    # Small visual gap between data block and verification block
    y -= 1.5 * mm

    # ══════════════════════════════════════════════════════════════
    # SECTION HEADER — Physical Verification (centered, light fill)
    # ══════════════════════════════════════════════════════════════
    sec_h = 9 * mm                        # was 7mm
    c.setFillColor(colors.HexColor('#f3f4f6'))
    c.rect(table_x, y - sec_h, table_w, sec_h, stroke=1, fill=1)
    c.setFillColor(colors.black)
    c.setFont('Helvetica-Bold', 11)
    cap_pt = 11 * 0.72
    c.drawCentredString(table_x + table_w / 2, y - sec_h + (sec_h - cap_pt) / 2,
                        'Physical Verification')
    y -= sec_h

    # ══════════════════════════════════════════════════════════════
    # PHYSICAL VERIFICATION TABLE — Sr No | Parameter | Results
    # ══════════════════════════════════════════════════════════════
    pv_hdr_h  = 9 * mm                    # was 7mm
    srno_w    = 20 * mm
    param_w   = 70 * mm
    results_w = table_w - srno_w - param_w

    # Header strip
    c.setFillColor(colors.HexColor('#f9fafb'))
    c.rect(table_x, y - pv_hdr_h, table_w, pv_hdr_h, stroke=1, fill=1)
    c.setFillColor(colors.black)
    vline(table_x + srno_w,           y, y - pv_hdr_h)
    vline(table_x + srno_w + param_w, y, y - pv_hdr_h)
    c.setFont('Helvetica-Bold', 10)
    cap_pt = 10 * 0.72
    hdr_ty = y - pv_hdr_h + (pv_hdr_h - cap_pt) / 2
    c.drawCentredString(table_x + srno_w / 2,                 hdr_ty, 'Sr No')
    c.drawCentredString(table_x + srno_w + param_w / 2,       hdr_ty, 'Parameter')
    c.drawCentredString(table_x + srno_w + param_w + results_w / 2, hdr_ty, 'Results')
    y -= pv_hdr_h

    # Data rows — taller for readability
    pv_row_h = 10 * mm                    # was 7mm
    cp = trs_row.get('checked_params') or []
    if isinstance(cp, list) and cp:
        params_to_draw = cp[:12]
    else:
        params_to_draw = [{'name': 'COA Availability', 'observed': 'NO'}]

    for i, pr in enumerate(params_to_draw, start=1):
        c.rect(table_x, y - pv_row_h, table_w, pv_row_h, stroke=1, fill=0)
        vline(table_x + srno_w,           y, y - pv_row_h)
        vline(table_x + srno_w + param_w, y, y - pv_row_h)
        c.setFont('Helvetica', 10)
        cap_pt = 10 * 0.72
        ty = y - pv_row_h + (pv_row_h - cap_pt) / 2
        # Sr no
        c.drawCentredString(table_x + srno_w / 2, ty, str(i))
        # Param name
        param_name = ''
        observed = ''
        if isinstance(pr, dict):
            param_name = str(pr.get('name') or '')
            observed   = str(pr.get('observed') or '')
        else:
            param_name = str(pr)
        c.drawString(table_x + srno_w + 3 * mm, ty, param_name)
        # Result
        c.drawString(table_x + srno_w + param_w + 3 * mm, ty, observed)
        y -= pv_row_h

    # Visual gap before signature block
    y -= 1.5 * mm

    # ══════════════════════════════════════════════════════════════
    # VERIFIED BY HEADER + SIGNATURE BOX
    # ══════════════════════════════════════════════════════════════
    vb_h = 9 * mm                         # was 7mm
    c.setFillColor(colors.HexColor('#f3f4f6'))
    c.rect(table_x, y - vb_h, table_w, vb_h, stroke=1, fill=1)
    c.setFillColor(colors.black)
    c.setFont('Helvetica-Bold', 10.5)
    cap_pt = 10.5 * 0.72
    c.drawString(table_x + 3 * mm,
                 y - vb_h + (vb_h - cap_pt) / 2,
                 'Verified By')
    y -= vb_h

    # Subheader row: Name of Incharge | Sign | Date
    sub_h = 9 * mm                        # was 7mm
    name_w = table_w / 3.0
    sign_w = name_w
    date_w = table_w - name_w - sign_w
    c.rect(table_x, y - sub_h, table_w, sub_h, stroke=1, fill=0)
    vline(table_x + name_w,         y, y - sub_h)
    vline(table_x + name_w + sign_w, y, y - sub_h)
    c.setFont('Helvetica-Bold', 10)
    cap_pt = 10 * 0.72
    sub_ty = y - sub_h + (sub_h - cap_pt) / 2
    c.drawCentredString(table_x + name_w / 2,                       sub_ty, 'Name of Incharge')
    c.drawCentredString(table_x + name_w + sign_w / 2,              sub_ty, 'Sign')
    c.drawCentredString(table_x + name_w + sign_w + date_w / 2,     sub_ty, 'Date')
    y -= sub_h

    # Signature box — generous height so a real signature fits
    sig_h = 22 * mm                       # was 15mm
    c.rect(table_x, y - sig_h, table_w, sig_h, stroke=1, fill=0)
    vline(table_x + name_w,         y, y - sig_h)
    vline(table_x + name_w + sign_w, y, y - sig_h)
    # Name of Incharge — printed name near bottom-left
    c.setFont('Helvetica', 10)
    c.drawString(table_x + 3 * mm, y - sig_h + 3 * mm,
                 str(trs_row.get('verified_by') or trs_row.get('generated_by') or ''))
    # Date — bottom-left of date cell
    c.drawString(table_x + name_w + sign_w + 3 * mm, y - sig_h + 3 * mm,
                 dmy(trs_row.get('generated_at')))
    y -= sig_h

    # ══════════════════════════════════════════════════════════════
    # APPROVAL FOOTER (only when decided)
    # ══════════════════════════════════════════════════════════════
    status = trs_row.get('approval_status') or 'Pending'
    if status != 'Pending':
        y -= 3 * mm
        ap_h = 14 * mm
        if status == 'Approved':
            bg = colors.HexColor('#d1fae5'); fg = colors.HexColor('#065f46')
        elif status == 'Rejected':
            bg = colors.HexColor('#fee2e2'); fg = colors.HexColor('#991b1b')
        else:
            bg = colors.HexColor('#e0e7ff'); fg = colors.HexColor('#3730a3')
        c.setFillColor(bg)
        c.rect(table_x, y - ap_h, table_w, ap_h, stroke=1, fill=1)
        c.setFillColor(fg)
        c.setFont('Helvetica-Bold', 11)
        c.drawString(table_x + 4 * mm, y - 5 * mm, f"QC Decision: {status.upper()}")
        c.setFillColor(colors.black)
        c.setFont('Helvetica', 8.5)
        c.drawString(table_x + 4 * mm, y - 9 * mm,
                     f"By: {trs_row.get('approved_by') or '—'}    "
                     f"At: {trs_row.get('approval_dt') or '—'}")
        if trs_row.get('rejection_reason'):
            c.drawString(table_x + 4 * mm, y - 12 * mm,
                         f"Reason: {trs_row.get('rejection_reason')[:120]}")

    # Footer micro-line (page identifier)
    c.setFillColor(colors.HexColor('#9ca3af'))
    c.setFont('Helvetica', 7)
    c.drawRightString(W - margin, margin / 2,
                      f"Generated {dmy(trs_row.get('generated_at'))} · "
                      f"TRS {trs_row.get('trs_num') or ''}")

    c.showPage()
    c.save()

@pm_stock_bp.route('/api/pm_stock/trs/<int:trs_id>/pdf', methods=['GET'])
@_login_required
def api_pm_trs_pdf(trs_id):
    """Stream a PDF rendering of one PM TRS.

    Browser opens this via window.open() — Content-Disposition is
    'inline' so it shows in a new tab; users can print or download
    from there.
    """
    blocked = _block_if_no_access('pm_trs')
    if blocked is not None: return blocked

    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM pm_grn_trs WHERE id=%s",
            (trs_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'status':'error','message':'TRS not found'}), 404
        from datetime import datetime as _dt
        rec = _pm_trs_normalise_row(row, _dt.now())
        conn.close()

        import io as _io
        buf = _io.BytesIO()
        _draw_pm_trs_pdf(buf, rec)
        buf.seek(0)
        filename = f"TRS_{(rec.get('trs_num') or 'pm-trs').replace('/','_')}.pdf"
        return send_file(buf, mimetype='application/pdf',
                         as_attachment=False, download_name=filename)
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return jsonify({'status':'error','message': str(e)}), 500
