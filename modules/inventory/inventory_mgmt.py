r"""
inventory_mgmt.py  –  Unified Inventory Module (RM / PM / FG)
=============================================================
HCP Wellness Pvt Ltd

A single page that lists items across all departments (RM, PM, FG).
Does NOT replace procurement.py / pm_stock routes / fg_routes.py — it reuses
their existing tables so data stays single-source-of-truth.

Tables used (existing):
  • procurement_materials       – RM items
  • pm_products                 – PM items
      (cols: id, product_name, pm_type, brand_id, min_stock, material_type,
       fg_links, is_active, created_at, hsn_code, gst_rate, uom)
  • FG_Names                    – FG items
  • procurement_brands          – Brands (cols: id, name, color, text_color, created_at)
  • procurement_suppliers       – RM / FG suppliers  (rich schema)
  • purchase_suppliers          – PM suppliers       (simpler schema)
      (cols: id, supplier_name, contact_person, phone, email, address, gstin,
       payment_terms, rating, is_active, created_at, updated_at, created_by)
  • procurement_godowns         – Godowns (cols: id, name, address, contact, phone,
       email, is_default, type, gst_number, created_at, updated_at, state, city, pin)
  • pm_stock                    – PM stock per godown
  • FG_stock                    – FG stock (optional, if exists)
  • procurement_grn             – GRN headers           (shared with procurement.py)
  • procurement_grn_items       – GRN line items        (shared with procurement.py)
  • procurement_purchase_orders – POs (read-only here, used to link GRN to PO)
  • procurement_voucher_numbering – GRN voucher number prefix/suffix/digits style

Helper tables (created by this module, do not affect old code):
  • inventory_brand_dept        – brand_id ↔ department (many-to-many)
  • inventory_supplier_dept     – supplier_id + supplier_source ↔ department
      (supplier_source = 'procurement' | 'purchase', to disambiguate since
       id values overlap across the two tables)
  • inventory_last_purchase     – cached per-item last supplier / rate / date

Register in app.py:
    import inventory_mgmt
    inventory_mgmt.register_inventory_mgmt(app)

Page URL:     /inventory_mgmt
API prefix:   /api/inventory_mgmt/*
"""

from __future__ import annotations

import json
import os
import re
import shutil
import traceback
import uuid
from functools import wraps
from datetime import datetime, timedelta

from flask import (
    render_template, render_template_string, session, jsonify,
    redirect, url_for, request, send_file
)

import sampling_portal  # shared DB helper

# ── Quiet startup banner ─────────────────────────────────────────────────────
# The inventory sub-modules each print "✅ … registered/ready" lines at boot,
# which flood the Flask console. By default we suppress them for a clean screen;
# set env HCP_INVENTORY_VERBOSE=1 to see the full banner again. Errors/warnings
# (lines containing ⚠ or "error"/"failed") always pass through.
import builtins as _builtins

_INV_VERBOSE = (os.environ.get("HCP_INVENTORY_VERBOSE", "") or "").lower() in ("1", "true", "yes", "on")
_inv_real_print = _builtins.print


def _inv_quiet_print(*args, **kwargs):
    """Filtered print used during inventory startup: swallows the routine
    'ready/registered' banner lines but lets warnings/errors through."""
    try:
        msg = " ".join(str(a) for a in args)
    except Exception:
        msg = ""
    low = msg.lower()
    if ("⚠" in msg) or ("error" in low) or ("failed" in low) or ("could not" in low):
        _inv_real_print(*args, **kwargs)


try:
    from openpyxl import load_workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

DEFAULT_STKSUM_PATH = r"\\Tarakbhavsar\procurement new\CURRENT RM\StkSum.xlsx"

# Which supplier table belongs to which department
SUPPLIER_TABLE_BY_DEPT = {
    "RM": "procurement_suppliers",
    "FG": "procurement_suppliers",
    "PM": "purchase_suppliers",
}
SUPPLIER_SOURCE = {
    "procurement_suppliers": "procurement",
    "purchase_suppliers":    "purchase",
}

# ─────────────────────────────────────────────────────────────────────────────
# ACCESS CONTROL
# ─────────────────────────────────────────────────────────────────────────────

def _can_inventory() -> bool:
    return bool(session.get("logged_in"))


def _can_edit_inventory() -> bool:
    if not session.get("logged_in"):
        return False
    role = (session.get("User_Type") or "").strip().lower()
    uid  = (session.get("UID")       or "").strip().lower()
    return role in {"admin"} or uid in {"sonal", "tarak"}


def _login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


def _edit_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        if not _can_edit_inventory():
            return jsonify({"status": "error", "message": "Permission denied"}), 403
        return f(*args, **kwargs)
    return wrapper


# ── GRN-permission-aware guards ──
# These honour the fine-grained inventory_access switches (grn / grn_new /
# grn_edit / grn_view_print) so a non-admin user with the right toggles can
# actually create or edit GRNs. Admins always pass (legacy compat).
#
# Why three decorators: a user might be granted *only* "New GRN" rights,
# in which case they should be able to POST a new GRN but not edit or
# delete existing ones. The save route checks the request payload to
# decide which capability to require (presence of `id` field → edit,
# absent → create).
def _grn_caps_for_session():
    """Look up the current session user's GRN caps. Falls back to admin
    role if the access-rights helper isn't importable for any reason —
    that way a broken import can't lock admins out of their own system."""
    try:
        try:
            from .inventory_access import _inv_grn_caps
        except Exception:
            from inventory_access import _inv_grn_caps
        caps = _inv_grn_caps() or {}
    except Exception:
        caps = {}
    if _can_edit_inventory():
        # Admin gets everything regardless of caps row contents
        return {k: True for k in ("view", "print", "edit", "create", "delete")}
    return caps


def _grn_save_required(f):
    """Decorator for POST /grn/save. Reads the request body to decide
    whether this is a CREATE (no id) or EDIT (has id) and checks the
    matching cap. Falls back to admin-class edit rights for legacy paths
    that don't carry an id."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        try:
            payload = request.get_json(silent=True) or {}
        except Exception:
            payload = {}
        is_edit = bool(payload.get("id"))
        caps = _grn_caps_for_session()
        need = "edit" if is_edit else "create"
        if not caps.get(need):
            return jsonify({
                "status":  "error",
                "message": ("You don't have permission to edit existing GRNs"
                            if is_edit else
                            "You don't have permission to create new GRNs"),
            }), 403
        return f(*args, **kwargs)
    return wrapper


def _grn_delete_required(f):
    """Decorator for POST /grn/delete. Requires the delete cap (which
    follows edit rights — see _inv_grn_caps)."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        caps = _grn_caps_for_session()
        if not caps.get("delete"):
            return jsonify({
                "status":  "error",
                "message": "You don't have permission to delete GRNs",
            }), 403
        return f(*args, **kwargs)
    return wrapper


def _items_caps_for_session():
    """Look up the current session user's Items (RM master) caps.
    Falls back to admin role if the access helper isn't importable."""
    try:
        try:
            from .inventory_access import _inv_items_caps
        except Exception:
            from inventory_access import _inv_items_caps
        caps = _inv_items_caps() or {}
    except Exception:
        caps = {}
    if _can_edit_inventory():
        return {"view": True, "edit": True}
    return caps


def _items_edit_required(f):
    """Decorator for item create / edit / delete. Requires the Items cap at
    'edit' level (admins always pass). View-level users are blocked here."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        if not _items_caps_for_session().get("edit"):
            return jsonify({
                "status":  "error",
                "message": "You don't have permission to add, edit or delete items.",
            }), 403
        return f(*args, **kwargs)
    return wrapper


def _grn_attach_required(f):
    """Decorator for GRN attachment routes (file upload / delete). The
    user needs *any* edit-class right on GRNs — managing attachments is
    part of normal GRN work whether they're creating a new one or
    revising an existing one."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        caps = _grn_caps_for_session()
        if not (caps.get("create") or caps.get("edit")):
            return jsonify({
                "status":  "error",
                "message": "You don't have permission to manage GRN attachments",
            }), 403
        return f(*args, **kwargs)
    return wrapper


def _user():
    return session.get("User_Name") or session.get("UID") or "Unknown"


DEPT_RM = "RM"
DEPT_PM = "PM"
DEPT_FG = "FG"
VALID_DEPTS = {DEPT_RM, DEPT_PM, DEPT_FG}


def _norm_dept(d):
    if not d:
        return None
    d = str(d).strip().upper()
    return d if d in VALID_DEPTS else None


# ─────────────────────────────────────────────────────────────────────────────
# TABLE INITIALISATION  (creates helper tables only — does NOT alter existing ones)
# ─────────────────────────────────────────────────────────────────────────────

def _init_inventory_tables():
    conn = sampling_portal.get_db_connection()
    if not conn:
        print("[InventoryMgmt] ⚠️  DB connection failed — init skipped.")
        return

    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS inventory_brand_dept (
                id         INT AUTO_INCREMENT PRIMARY KEY,
                brand_id   INT          NOT NULL,
                department VARCHAR(10)  NOT NULL,
                created_at DATETIME     DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uq_brand_dept (brand_id, department)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()

        conn.execute("""
            CREATE TABLE IF NOT EXISTS inventory_supplier_dept (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                supplier_id     INT          NOT NULL,
                supplier_source VARCHAR(20)  NOT NULL DEFAULT 'procurement',
                department      VARCHAR(10)  NOT NULL,
                created_at      DATETIME     DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uq_sup_src_dept (supplier_id, supplier_source, department)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()

        # Migration: add supplier_source column on older installs
        try:
            conn.execute(
                "ALTER TABLE inventory_supplier_dept "
                "ADD COLUMN supplier_source VARCHAR(20) NOT NULL DEFAULT 'procurement'"
            )
            conn.commit()
        except Exception:
            pass

        conn.execute("""
            CREATE TABLE IF NOT EXISTS inventory_last_purchase (
                id             INT AUTO_INCREMENT PRIMARY KEY,
                department     VARCHAR(10)   NOT NULL,
                item_id        INT           NOT NULL,
                item_name      VARCHAR(500)  NOT NULL,
                last_supplier  VARCHAR(500)  DEFAULT NULL,
                last_rate      DECIMAL(15,4) DEFAULT NULL,
                last_date      DATE          DEFAULT NULL,
                gst_rate       DECIMAL(5,2)  DEFAULT NULL,
                updated_at     DATETIME      DEFAULT CURRENT_TIMESTAMP
                               ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_dept_item (department, item_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()

        # ── Seed all-dept mapping for every brand (idempotent) ──
        # Per user spec: brands apply to all departments by default.
        try:
            conn.execute("""
                INSERT IGNORE INTO inventory_brand_dept (brand_id, department)
                SELECT b.id, d.dept
                FROM procurement_brands b
                CROSS JOIN (
                    SELECT 'RM' AS dept UNION ALL SELECT 'PM' UNION ALL SELECT 'FG'
                ) d
            """)
            conn.commit()
        except Exception:
            traceback.print_exc()

        print("✅ [InventoryMgmt] helper tables ready")
    except Exception:
        traceback.print_exc()
    finally:
        try: conn.close()
        except Exception: pass


# ─────────────────────────────────────────────────────────────────────────────
# LIVE STOCK HELPERS
# ─────────────────────────────────────────────────────────────────────────────

# ── Real RM stock from rm_boxes (in_stock boxes only) ────────────────────────
# When USE_RM_BOX_STOCK is True, the items grid + godown breakdown reflect the
# actual package-level inventory (SUM of per_box_qty of in_stock boxes), totalled
# across all godowns — instead of the legacy Excel/Tally dump.
USE_RM_BOX_STOCK = True


def _read_rm_box_stock_by_material(conn) -> dict:
    """{material_id: total_in_stock_qty} summed across ALL godowns."""
    out = {}
    try:
        rows = conn.execute(
            "SELECT material_id, COALESCE(SUM(per_box_qty),0) AS q "
            "FROM rm_boxes WHERE current_status='in_stock' "
            "GROUP BY material_id"
        ).fetchall()
        for r in rows:
            mid = r["material_id"] if hasattr(r, "get") else r[0]
            q   = r["q"] if hasattr(r, "get") else r[1]
            if mid is not None:
                out[int(mid)] = float(q or 0)
    except Exception:
        pass
    return out


def _read_rm_box_stock_for_material(conn, material_id) -> dict:
    """{godown_id: qty} for one material, in_stock boxes only."""
    out = {}
    try:
        rows = conn.execute(
            "SELECT current_godown_id AS gid, COALESCE(SUM(per_box_qty),0) AS q "
            "FROM rm_boxes WHERE current_status='in_stock' AND material_id=%s "
            "GROUP BY current_godown_id",
            (int(material_id),),
        ).fetchall()
        for r in rows:
            gid = r["gid"] if hasattr(r, "get") else r[0]
            q   = r["q"] if hasattr(r, "get") else r[1]
            out[int(gid) if gid is not None else 0] = float(q or 0)
    except Exception:
        pass
    return out


def _read_rm_stock_from_xlsx() -> dict:
    if not OPENPYXL_OK:
        return {}
    path = DEFAULT_STKSUM_PATH
    try:
        conn = sampling_portal.get_db_connection()
        if conn:
            row = conn.execute(
                "SELECT setting_value FROM procurement_settings WHERE setting_key='stksum_path' LIMIT 1"
            ).fetchone()
            if row and row["setting_value"]:
                path = row["setting_value"]
            conn.close()
    except Exception:
        pass
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        out = {}
        for r in ws.iter_rows(min_row=16, values_only=True):
            if not r or len(r) < 3: continue
            name = r[0]; qty = r[2]
            if not name: continue
            try:
                out[str(name).strip().lower()] = float(qty or 0)
            except (TypeError, ValueError):
                out[str(name).strip().lower()] = 0.0
        wb.close()
        return out
    except Exception:
        return {}


def _read_pm_stock(conn) -> dict:
    """
    Aggregate PM stock per product across godowns AND factory floor.
    Uses the same transaction-log aggregation that /api/pm_stock/summary does,
    so the numbers match the PM Stock page exactly.

    Returns: { product_id: total_qty }  where total = godown_stock + floor_stock
    """
    out = {}
    try:
        # Try the exact same aggregation as pm_stock_routes. The canonical source
        # is the pm_stock_log transaction table (not a point-in-time snapshot).
        rows = conn.execute("""
            SELECT  product_id,
                    SUM(CASE
                        WHEN txn_type IN ('grn','opening','inward','return','mtv_in') THEN qty
                        WHEN txn_type IN ('issue','outward','mtv_out','adjustment_out','consumption') THEN -qty
                        WHEN txn_type IN ('adjustment_in') THEN qty
                        ELSE 0
                    END) AS total
            FROM pm_stock_log
            GROUP BY product_id
        """).fetchall()
        if rows:
            for r in rows:
                try:
                    out[int(r["product_id"])] = float(r["total"] or 0)
                except (TypeError, ValueError):
                    pass
            return out
    except Exception:
        pass

    # Fallback: if pm_stock_log doesn't exist (older install), try the
    # simpler pm_stock snapshot table if it exists.
    try:
        rows = conn.execute("""
            SELECT product_id, SUM(COALESCE(qty,0)) AS total
            FROM pm_stock
            GROUP BY product_id
        """).fetchall()
        for r in rows:
            try:
                out[int(r["product_id"])] = float(r["total"] or 0)
            except (TypeError, ValueError):
                pass
    except Exception:
        pass
    return out


def _read_fg_stock(conn) -> dict:
    out = {}
    try:
        rows = conn.execute("""
            SELECT fg_id, SUM(COALESCE(qty,0)) AS total
            FROM FG_stock
            GROUP BY fg_id
        """).fetchall()
        for r in rows:
            try:
                out[int(r["fg_id"])] = float(r["total"] or 0)
            except (TypeError, ValueError):
                pass
    except Exception:
        pass
    return out


# ─────────────────────────────────────────────────────────────────────────────
# REGISTRATION
# ─────────────────────────────────────────────────────────────────────────────

def register_inventory_mgmt(app):
    # Quiet the sub-module startup banner unless HCP_INVENTORY_VERBOSE=1.
    # We swap print() for a filtered version now and restore it at the end of
    # this function (see the restore just before the final summary line).
    if not _INV_VERBOSE:
        _builtins.print = _inv_quiet_print
    _init_inventory_tables()

    # ── Auto-register the User Access Control routes (Phase 1) ──
    # Attaching them here means app.py needs NO new lines: it already calls
    # register_inventory_mgmt(app), so the /api/inventory_mgmt/access/* routes
    # come along automatically. Guarded so a double-call (or a manual
    # register_inventory_access in app.py) can't register routes twice.
    if not getattr(app, "_inventory_access_registered", False):
        try:
            from inventory import inventory_access
        except Exception:
            try:
                import inventory_access  # fallback if not packaged
            except Exception:
                inventory_access = None
        if inventory_access is not None:
            try:
                inventory_access.register_inventory_access(app)
                app._inventory_access_registered = True
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register Material Request routes (Phase 2) ──
    if not getattr(app, "_inventory_mr_registered", False):
        try:
            from inventory import inventory_material_request as _imr
        except Exception:
            try:
                import inventory_material_request as _imr
            except Exception:
                _imr = None
        if _imr is not None:
            try:
                _imr.register_inventory_material_request(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register FEFO module (Phase 3) ──
    if not getattr(app, "_inventory_fefo_registered", False):
        try:
            from inventory import inventory_fefo as _ife
        except Exception:
            try:
                import inventory_fefo as _ife
            except Exception:
                _ife = None
        if _ife is not None:
            try:
                _ife.register_inventory_fefo(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register Material Lock module (Phase 4) ──
    if not getattr(app, "_inventory_matlock_registered", False):
        try:
            from inventory import inventory_material_lock as _iml
        except Exception:
            try:
                import inventory_material_lock as _iml
            except Exception:
                _iml = None
        if _iml is not None:
            try:
                _iml.register_inventory_material_lock(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register Label Reissue module (Phase 7) ──
    if not getattr(app, "_inventory_reissue_registered", False):
        try:
            from inventory import inventory_label_reissue as _ilr
        except Exception:
            try:
                import inventory_label_reissue as _ilr
            except Exception:
                _ilr = None
        if _ilr is not None:
            try:
                _ilr.register_inventory_label_reissue(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register Label Reprint module (Phase 7) ──
    if not getattr(app, "_inventory_reprint_registered", False):
        try:
            from inventory import inventory_label_reprint as _ilp
        except Exception:
            try:
                import inventory_label_reprint as _ilp
            except Exception:
                _ilp = None
        if _ilp is not None:
            try:
                _ilp.register_inventory_label_reprint(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register FEFO label-code module (RM labels) ──
    if not getattr(app, "_inventory_fefocode_registered", False):
        try:
            from inventory import inventory_fefo_code as _ifc
        except Exception:
            try:
                import inventory_fefo_code as _ifc
            except Exception:
                _ifc = None
        if _ifc is not None:
            try:
                _ifc.register_inventory_fefo_code(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register GRN Box Repair (admin tool) ──
    # Adds /inventory/grn_box_repair (admin HTML page) +
    # /api/inventory_mgmt/grn/{box_audit,repair_boxes} (admin API).
    # Fixes "ghost" GRN allocations where labels were printed but
    # rm_boxes rows were never created due to a silent material-lookup
    # failure during GRN save. Drop-in admin diagnostic + repair flow.
    if not getattr(app, "_inventory_grn_box_repair_registered", False):
        try:
            from inventory import inventory_grn_box_repair as _ibr
        except Exception:
            try:
                import inventory_grn_box_repair as _ibr
            except Exception:
                _ibr = None
        if _ibr is not None:
            try:
                _ibr.register_inventory_grn_box_repair(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register Voucher Numbering module (admin-only) ──
    # Provides /api/inventory_mgmt/voucher_numbering/{list,save,delete,preview}.
    # Self-contained — owns its own inventory_voucher_numbering table; no
    # PM Stock dependencies.
    if not getattr(app, "_inventory_vn_registered", False):
        try:
            from inventory import inventory_voucher_numbering as _ivn
        except Exception:
            try:
                import inventory_voucher_numbering as _ivn
            except Exception:
                _ivn = None
        if _ivn is not None:
            try:
                _ivn.register_inventory_voucher_numbering(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register Agent (pending-task reminder) ──
    # Read-only endpoint that returns the user's pending tasks across the
    # inventory module: expiring batches, below-MSL materials, pending
    # Material Requests, in-transit transfers, stale audits. Frontend
    # polls this every ~90 minutes and shows a 5-second top-right toast.
    if not getattr(app, "_inventory_agent_registered", False):
        try:
            from inventory import inventory_agent as _iag
        except Exception:
            try:
                import inventory_agent as _iag
            except Exception:
                _iag = None
        if _iag is not None:
            try:
                _iag.register_inventory_agent(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register Box Split module ──
    if not getattr(app, "_inventory_boxsplit_registered", False):
        try:
            from inventory import inventory_box_split as _ibs
        except Exception:
            try:
                import inventory_box_split as _ibs
            except Exception:
                _ibs = None
        if _ibs is not None:
            try:
                _ibs.register_inventory_box_split(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register Delivery Note module ──
    if not getattr(app, "_inventory_dn_registered", False):
        try:
            from inventory import inventory_delivery_note as _idn
        except Exception:
            try:
                import inventory_delivery_note as _idn
            except Exception:
                _idn = None
        if _idn is not None:
            try:
                _idn.register_inventory_delivery_note(app)
            except Exception:
                import traceback
                traceback.print_exc()

    if not getattr(app, "_inventory_audit_registered", False):
        try:
            from inventory import inventory_audit as _iad
        except Exception:
            try:
                import inventory_audit as _iad
            except Exception:
                _iad = None
        if _iad is not None:
            try:
                _iad.register_inventory_audit(app)
            except Exception:
                import traceback
                traceback.print_exc()

    if not getattr(app, "_inventory_reports_registered", False):
        try:
            from inventory import inventory_reports as _irep
        except Exception:
            try:
                import inventory_reports as _irep
            except Exception:
                _irep = None
        if _irep is not None:
            try:
                _irep.register_inventory_reports(app)
            except Exception:
                import traceback
                traceback.print_exc()

    if not getattr(app, "_inventory_reset_registered", False):
        try:
            from inventory import inventory_reset as _irst
        except Exception:
            try:
                import inventory_reset as _irst
            except Exception:
                _irst = None
        if _irst is not None:
            try:
                _irst.register_inventory_reset(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register Stock Transfers (QR scan OUT/IN voucher flow) ──
    # Routes live under /api/inventory_godown/transfers/* but the module is
    # its own file. inventory_transfers.js is loaded by the template, so the
    # Transfers UI is live and needs these endpoints. Guarded so a double
    # register_inventory_mgmt(app) can't register routes twice.
    if not getattr(app, "_inventory_transfers_registered", False):
        try:
            from inventory import inventory_transfers as _itr
        except Exception:
            try:
                import inventory_transfers as _itr
            except Exception:
                _itr = None
        if _itr is not None:
            try:
                _itr.register_inventory_transfers(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Auto-register Simple Transfer (manual voucher flow) ──
    # Routes under /api/inventory_simple_transfer/*. The module self-injects
    # its own sidebar panel at runtime once inventory_simple_transfer.js is
    # on the page, but the backend must be wired here regardless.
    if not getattr(app, "_inventory_simple_transfer_registered", False):
        try:
            from inventory import inventory_simple_transfer as _ist
        except Exception:
            try:
                import inventory_simple_transfer as _ist
            except Exception:
                _ist = None
        if _ist is not None:
            try:
                _ist.register_inventory_simple_transfer(app)
            except Exception:
                import traceback
                traceback.print_exc()

    # ── Detect whether procurement_materials has an `in_stock_qty` column ──
    # Older installs (where stock is tracked via an external StkSum.xlsx merge)
    # don't have this column. We probe once at startup so the GRN save/delete
    # logic can skip the UPDATE statements that would otherwise throw and
    # poison the transaction.
    _MAT_HAS_STOCK_COL = {"value": False}
    try:
        _probe_conn = sampling_portal.get_db_connection()
        if _probe_conn:
            try:
                _probe_conn.execute(
                    "SELECT in_stock_qty FROM procurement_materials LIMIT 1"
                ).fetchall()
                _MAT_HAS_STOCK_COL["value"] = True
                print("✅ [InventoryMgmt] procurement_materials.in_stock_qty present — stock-sync enabled")
            except Exception:
                _MAT_HAS_STOCK_COL["value"] = False
                print("ℹ️  [InventoryMgmt] procurement_materials.in_stock_qty missing — GRN stock-sync disabled (stock is tracked elsewhere)")
            try: _probe_conn.close()
            except Exception: pass
    except Exception:
        pass

    # ── Detect whether procurement_grn_items has a `manufacturer` column.
    # Added by the inventory_godown bootstrap (May 2026). If present we
    # save/load the per-item manufacturer field; otherwise the GRN save
    # paths skip it gracefully.
    #
    # NOTE: this is detected LAZILY (on first use) rather than once at boot.
    # A boot-time-only probe meant that if the column was added AFTER the
    # process started, the cached "missing" answer caused manufacturer to be
    # silently dropped on every save until a manual restart. value=None means
    # "not yet checked"; the first save/list re-probes the live connection and
    # caches the real answer, so adding the column takes effect immediately.
    _ITEMS_HAS_MFR = {"value": None}

    def _items_has_mfr(conn=None):
        """Return True if procurement_grn_items has a `manufacturer` column.
        Probes once on the given (or a fresh) connection, then caches."""
        if _ITEMS_HAS_MFR["value"] is not None:
            return _ITEMS_HAS_MFR["value"]
        _c = conn
        _own = False
        try:
            if _c is None:
                _c = sampling_portal.get_db_connection()
                _own = True
            if _c is None:
                return False  # don't cache a connection failure
            try:
                _c.execute(
                    "SELECT manufacturer FROM procurement_grn_items LIMIT 1"
                ).fetchall()
                _ITEMS_HAS_MFR["value"] = True
            except Exception:
                _ITEMS_HAS_MFR["value"] = False
        except Exception:
            return False
        finally:
            if _own and _c is not None:
                try: _c.close()
                except Exception: pass
        return _ITEMS_HAS_MFR["value"]

    # ── Bootstrap procurement_grn_files table (COA + Invoice attachments) ──
    # Files are now stored on the filesystem (not BLOBs), so the table only
    # holds metadata + relative_path. The previous BLOB column is dropped if
    # it exists — and any existing rows are wiped, since the truncated BLOBs
    # were broken anyway (decided May 2026, switching from BLOB to disk).
    #
    # Storage layout (relative to UPLOAD_ROOT):
    #   grn/<grn_id>/coa/<uuid>__<original-name>
    #   grn/<grn_id>/invoice/<uuid>__<original-name>
    try:
        _bs_conn = sampling_portal.get_db_connection()
        if _bs_conn:
            try:
                # 1. Create the table if it doesn't exist (fresh installs).
                _bs_conn.execute("""
                    CREATE TABLE IF NOT EXISTS procurement_grn_files (
                      id              INT AUTO_INCREMENT PRIMARY KEY,
                      grn_id          INT NOT NULL,
                      grn_item_id     INT NULL,
                      kind            ENUM('coa','invoice') NOT NULL,
                      original_name   VARCHAR(255) NOT NULL,
                      mime_type       VARCHAR(80)  NOT NULL,
                      size_bytes      INT          NOT NULL,
                      relative_path   VARCHAR(500) NULL,
                      uploaded_by     VARCHAR(80)  NULL,
                      uploaded_at     DATETIME DEFAULT CURRENT_TIMESTAMP,
                      INDEX idx_grn      (grn_id),
                      INDEX idx_grn_item (grn_item_id),
                      INDEX idx_kind     (kind)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                _bs_conn.commit()

                # 2. Detect existing columns — needed for migration from old BLOB schema.
                col_info = _bs_conn.execute("""
                    SELECT COLUMN_NAME
                    FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_SCHEMA = DATABASE()
                      AND TABLE_NAME   = 'procurement_grn_files'
                """).fetchall()
                cols = { (r["COLUMN_NAME"] if isinstance(r, dict) else r[0]).lower() for r in col_info }

                # 3. Add relative_path column if missing (older install).
                if "relative_path" not in cols:
                    _bs_conn.execute(
                        "ALTER TABLE procurement_grn_files "
                        "ADD COLUMN relative_path VARCHAR(500) NULL AFTER size_bytes"
                    )
                    _bs_conn.commit()
                    print("✅ [InventoryMgmt] Added relative_path column to procurement_grn_files")

                # 4. If the legacy file_data BLOB column exists, wipe all rows
                #    and drop the column.
                if "file_data" in cols:
                    # Best-effort: count rows before wipe so the log message is honest.
                    try:
                        cnt = _bs_conn.execute(
                            "SELECT COUNT(*) AS n FROM procurement_grn_files"
                        ).fetchone()
                        old_count = int(cnt["n"]) if cnt else 0
                    except Exception:
                        old_count = -1
                    _bs_conn.execute("DELETE FROM procurement_grn_files")
                    _bs_conn.execute("ALTER TABLE procurement_grn_files DROP COLUMN file_data")
                    _bs_conn.commit()
                    print(f"✅ [InventoryMgmt] Migrated procurement_grn_files to disk storage "
                          f"(wiped {old_count} legacy BLOB row(s), dropped file_data column)")

                print("✅ [InventoryMgmt] procurement_grn_files table ready (disk-storage mode)")
            except Exception as e:
                print(f"⚠️  [InventoryMgmt] procurement_grn_files setup error: {e}")
                traceback.print_exc()
            try: _bs_conn.close()
            except Exception: pass
    except Exception:
        pass

    # ── Bootstrap procurement_grn_trs table (Testing Requisition Slips) ──
    # One row per generated TRS, FK'd to the GRN line. Generating a TRS for
    # a line is idempotent — re-generating overwrites the existing row but
    # keeps trs_num stable (so re-printing always shows the same number).
    # Fields stored: every column shown on the printed slip, including the
    # ones we have to PROMPT for (physical_state, sample_qty, previous_supplier,
    # new_or_old). All others are denormalised from the GRN at gen-time so the
    # slip is reproducible even if the GRN is later edited.
    #
    # Also includes QC approval fields (approval_status / approved_by /
    # approval_dt / approval_remarks) — these are populated from the QC
    # Dashboard tab when a QC person reviews the slip. They default to
    # 'Pending' / NULL on TRS generation. The QC user can flip the row
    # to Approved / Rejected via /api/qc/trs/approve.
    try:
        import sampling_portal
        _bs_conn = sampling_portal.get_db_connection()
        if _bs_conn is not None:
            try:
                _bs_conn.execute("""
                    CREATE TABLE IF NOT EXISTS procurement_grn_trs (
                        id                INT AUTO_INCREMENT PRIMARY KEY,
                        trs_num           VARCHAR(50) NOT NULL UNIQUE,
                        grn_id            INT NOT NULL,
                        grn_item_id       INT NOT NULL UNIQUE,
                        -- Denormalised GRN snapshot at generation time
                        grn_num           VARCHAR(50)  DEFAULT NULL,
                        grn_date          DATE         DEFAULT NULL,
                        material          VARCHAR(255) DEFAULT NULL,
                        batch_num         VARCHAR(100) DEFAULT NULL,
                        packages          INT          DEFAULT NULL,
                        qty_per_pkg       DECIMAL(14,3) DEFAULT NULL,
                        total_qty         DECIMAL(14,3) DEFAULT NULL,
                        uom               VARCHAR(20)  DEFAULT 'KG',
                        manufacturer      VARCHAR(255) DEFAULT NULL,
                        mfg_date          DATE         DEFAULT NULL,
                        expiry_date       DATE         DEFAULT NULL,
                        supplier_name     VARCHAR(255) DEFAULT NULL,
                        -- Operator-supplied fields (prompted in the modal)
                        physical_state    VARCHAR(40)  DEFAULT NULL,
                        sample_qty        DECIMAL(14,3) DEFAULT NULL,
                        previous_supplier VARCHAR(255) DEFAULT NULL,
                        new_or_old        ENUM('NEW','OLD') DEFAULT NULL,
                        -- Audit
                        generated_by      VARCHAR(120) DEFAULT NULL,
                        generated_at      DATETIME     DEFAULT CURRENT_TIMESTAMP,
                        verified_by       VARCHAR(120) DEFAULT NULL,
                        -- QC approval (populated from QC Dashboard tab)
                        approval_status   ENUM('Pending','Approved','Rejected','Under Review')
                                              NOT NULL DEFAULT 'Pending',
                        approved_by       VARCHAR(120) DEFAULT NULL,
                        approval_dt       DATETIME     DEFAULT NULL,
                        approval_remarks  TEXT         DEFAULT NULL,
                        -- Lock: stamped on the FIRST transition out of Pending.
                        -- The row becomes uneditable 24 hours after this
                        -- timestamp — server enforces in /trs/approve.
                        approval_locked_at DATETIME    DEFAULT NULL,
                        -- For rejections: which parameters failed and why.
                        -- checked_params is a JSON array (stored as TEXT for
                        -- portability across MySQL versions that lack a JSON
                        -- type). rejection_reason is free-text.
                        checked_params    TEXT         DEFAULT NULL,
                        rejection_reason  TEXT         DEFAULT NULL,
                        INDEX idx_grn (grn_id),
                        INDEX idx_grn_item (grn_item_id),
                        INDEX idx_generated_at (generated_at),
                        INDEX idx_approval_status (approval_status)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                      COLLATE=utf8mb4_unicode_ci
                """)

                # Idempotent migration: add approval columns + index if the
                # table already existed from an earlier install without
                # them. Check information_schema first so reruns don't
                # bleat on duplicate-column errors.
                try:
                    have = set()
                    for c in _bs_conn.execute("SHOW COLUMNS FROM procurement_grn_trs").fetchall():
                        have.add(c["Field"] if hasattr(c, "get") else c[0])
                    add = []
                    if "approval_status" not in have:
                        add.append("ADD COLUMN approval_status ENUM('Pending','Approved','Rejected','Under Review') NOT NULL DEFAULT 'Pending'")
                    if "approved_by" not in have:
                        add.append("ADD COLUMN approved_by VARCHAR(120) DEFAULT NULL")
                    if "approval_dt" not in have:
                        add.append("ADD COLUMN approval_dt DATETIME DEFAULT NULL")
                    if "approval_remarks" not in have:
                        add.append("ADD COLUMN approval_remarks TEXT DEFAULT NULL")
                    if "approval_locked_at" not in have:
                        add.append("ADD COLUMN approval_locked_at DATETIME DEFAULT NULL")
                    if "checked_params" not in have:
                        add.append("ADD COLUMN checked_params TEXT DEFAULT NULL")
                    if "rejection_reason" not in have:
                        add.append("ADD COLUMN rejection_reason TEXT DEFAULT NULL")
                    if add:
                        _bs_conn.execute(
                            "ALTER TABLE procurement_grn_trs " + ", ".join(add)
                        )
                        print(f"✅ [InventoryMgmt] migrated procurement_grn_trs: {len(add)} column(s) added")
                    # Best-effort approval index (silent if already present)
                    try:
                        idx_row = _bs_conn.execute("""
                            SELECT COUNT(*) AS n FROM information_schema.STATISTICS
                            WHERE TABLE_SCHEMA = DATABASE()
                              AND TABLE_NAME   = 'procurement_grn_trs'
                              AND INDEX_NAME   = 'idx_approval_status'
                        """).fetchone()
                        has_idx = int((idx_row.get("n") if hasattr(idx_row, "get") else idx_row[0]) or 0) > 0
                    except Exception:
                        has_idx = False
                    if not has_idx:
                        try:
                            _bs_conn.execute(
                                "ALTER TABLE procurement_grn_trs "
                                "ADD INDEX idx_approval_status (approval_status)"
                            )
                        except Exception:
                            pass
                except Exception as me:
                    print(f"⚠️  [InventoryMgmt] procurement_grn_trs migration skipped: {me}")

                _bs_conn.commit()
                print("✅ [InventoryMgmt] procurement_grn_trs table ready")
            except Exception as e:
                print(f"⚠️  [InventoryMgmt] procurement_grn_trs setup error: {e}")
            try: _bs_conn.close()
            except Exception: pass
    except Exception:
        pass

    # ── Bootstrap qc_trs_parameter_library ──
    # A small library of QC parameters the user can pick from when
    # filling a TRS. Custom parameters added by QC users persist here
    # so they don't have to retype the same parameter for every slip.
    # Seeded with a starter set on first install; existing rows are
    # preserved on subsequent boots.
    try:
        import sampling_portal
        _bs_conn = sampling_portal.get_db_connection()
        if _bs_conn is not None:
            try:
                _bs_conn.execute("""
                    CREATE TABLE IF NOT EXISTS qc_trs_parameter_library (
                        id            INT AUTO_INCREMENT PRIMARY KEY,
                        name          VARCHAR(150) NOT NULL UNIQUE,
                        unit          VARCHAR(40)  DEFAULT NULL,
                        -- Default spec hint for the parameter — used to
                        -- pre-fill the spec column when the parameter is
                        -- ticked. spec_type='range' uses spec_from/spec_to;
                        -- spec_type='value' uses spec_target. Either may
                        -- be NULL (operator fills in per-slip).
                        spec_type     ENUM('range','value','text') DEFAULT 'range',
                        spec_from     VARCHAR(40)  DEFAULT NULL,
                        spec_to       VARCHAR(40)  DEFAULT NULL,
                        spec_target   VARCHAR(80)  DEFAULT NULL,
                        is_active     TINYINT(1)   NOT NULL DEFAULT 1,
                        sort_order    INT          NOT NULL DEFAULT 100,
                        created_by    VARCHAR(120) DEFAULT NULL,
                        created_at    DATETIME     DEFAULT CURRENT_TIMESTAMP,
                        INDEX idx_active_sort (is_active, sort_order)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                      COLLATE=utf8mb4_unicode_ci
                """)
                # Seed if the table is empty. Each row gives a reasonable
                # default — operators can override per-slip.
                row = _bs_conn.execute(
                    "SELECT COUNT(*) AS n FROM qc_trs_parameter_library"
                ).fetchone()
                cnt = int((row.get('n') if hasattr(row, 'get') else row[0]) or 0)
                if cnt == 0:
                    seed = [
                        # (name, unit, spec_type, from, to, target, sort)
                        ('Appearance',                None, 'text',  None, None, 'Matches reference',    10),
                        ('Colour',                    None, 'text',  None, None, 'As per standard',      20),
                        ('Odour',                     None, 'text',  None, None, 'Characteristic',       30),
                        ('Physical State',            None, 'text',  None, None, 'As declared',          40),
                        ('Particle Size',             'µm', 'range', None, None, None,                    50),
                        ('Moisture Content',          '%',  'range', '0', '5',  None,                     60),
                        ('pH Value',                  None, 'range', '6', '8',  None,                     70),
                        ('Density / Bulk Density',    'g/ml','range', None, None, None,                   80),
                        ('Viscosity',                 'cP', 'range', None, None, None,                    90),
                        ('Solubility',                None, 'text',  None, None, 'Soluble in water',     100),
                        ('Assay / Purity',            '%',  'range', '98', '102', None,                  110),
                        ('Identification',            None, 'text',  None, None, 'Positive',             120),
                        ('Microbial Load (Total)',    'cfu/g','range', '0', '1000', None,                130),
                        ('Foreign Matter',            '%',  'range', '0', '0.5', None,                   140),
                        ('Loss on Drying',            '%',  'range', '0', '5',  None,                    150),
                        ('Heavy Metals',              'ppm','range', '0', '10', None,                    160),
                        ('Ash Content',               '%',  'range', None, None, None,                   170),
                        ('COA Availability',          None, 'text',  None, None, 'YES',                  180),
                        ('Mfg / Expiry Date',         None, 'text',  None, None, 'Valid',                190),
                        ('Packaging Condition',       None, 'text',  None, None, 'Intact',               200),
                    ]
                    for s in seed:
                        try:
                            _bs_conn.execute("""
                                INSERT INTO qc_trs_parameter_library
                                  (name, unit, spec_type, spec_from, spec_to, spec_target, sort_order, created_by)
                                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                            """, (s[0], s[1], s[2], s[3], s[4], s[5], s[6], 'system'))
                        except Exception:
                            pass   # name already exists → skip
                    print(f"✅ [InventoryMgmt] qc_trs_parameter_library seeded with {len(seed)} entries")
                _bs_conn.commit()
                print("✅ [InventoryMgmt] qc_trs_parameter_library table ready")
            except Exception as e:
                print(f"⚠️  [InventoryMgmt] qc_trs_parameter_library setup error: {e}")
            try: _bs_conn.close()
            except Exception: pass
    except Exception:
        pass

    # ── Bootstrap qc_trs_item_params ──
    # Per-material parameter memory. When QC approves or rejects a TRS,
    # the parameter set used (specs + observations) is saved here keyed
    # to the material name. Next time the same material appears on a
    # TRS, the Checked Parameters modal pre-loads this list so the QC
    # operator doesn't have to re-pick the same parameters every time.
    #
    # Just the SPECS travel — observations are TRS-specific and start
    # blank each time. The 'params_json' column stores a JSON array of
    # { name, unit, spec_type, spec_from, spec_to, spec_target } objects.
    try:
        import sampling_portal
        _bs_conn = sampling_portal.get_db_connection()
        if _bs_conn is not None:
            try:
                _bs_conn.execute("""
                    CREATE TABLE IF NOT EXISTS qc_trs_item_params (
                        id            INT AUTO_INCREMENT PRIMARY KEY,
                        -- Stored case-folded so 'Behentronium chloride' and
                        -- 'BEHENTRONIUM CHLORIDE' resolve to the same row.
                        -- The original (display) form is kept in material_display.
                        material_key     VARCHAR(255) NOT NULL UNIQUE,
                        material_display VARCHAR(255) DEFAULT NULL,
                        params_json      TEXT NOT NULL,
                        last_used_by     VARCHAR(120) DEFAULT NULL,
                        last_used_at     DATETIME     DEFAULT CURRENT_TIMESTAMP
                                          ON UPDATE CURRENT_TIMESTAMP,
                        INDEX idx_last_used (last_used_at)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                      COLLATE=utf8mb4_unicode_ci
                """)
                _bs_conn.commit()
                print("✅ [InventoryMgmt] qc_trs_item_params table ready")
            except Exception as e:
                print(f"⚠️  [InventoryMgmt] qc_trs_item_params setup error: {e}")
            try: _bs_conn.close()
            except Exception: pass
    except Exception:
        pass

    # ── Boot-time repair: re-link orphaned TRS rows ──
    # When a GRN was edited before this fix landed, the line-item
    # DELETE+INSERT cycle in /grn/save would orphan every TRS row's
    # grn_item_id. The TRS data itself survives (so the QC dashboard
    # still works) but the GRN form's picker and line badges can't
    # find the link any more. This runs ONCE on Flask startup and
    # heals existing damage; the /grn/save patch prevents future
    # damage. Re-running this on every boot is safe — it only touches
    # rows where the link is genuinely broken.
    try:
        import sampling_portal
        _bs_conn = sampling_portal.get_db_connection()
        if _bs_conn is not None:
            try:
                # Repair pass: re-link orphaned TRS rows by matching on
                # (grn_id, material, batch_num). Using NOT EXISTS in place
                # of NOT IN (SELECT ... FROM derived) — the latter has
                # patchy support across MySQL versions when the inner
                # query correlates back to the outer table. NOT EXISTS
                # is rock-solid on every MySQL build we'd realistically
                # encounter.
                repair = _bs_conn.execute("""
                    UPDATE procurement_grn_trs t
                    JOIN procurement_grn_items i
                      ON i.grn_id = t.grn_id
                     AND LOWER(TRIM(i.material)) = LOWER(TRIM(t.material))
                     AND COALESCE(TRIM(i.batch_num),'') = COALESCE(TRIM(t.batch_num),'')
                    SET t.grn_item_id = i.id
                    WHERE NOT EXISTS (
                        SELECT 1 FROM procurement_grn_items i2
                        WHERE i2.id = t.grn_item_id
                    )
                """)
                _bs_conn.commit()
                # Most MySQL drivers expose rowcount on the cursor object.
                affected = getattr(repair, "rowcount", None)
                if affected and affected > 0:
                    print(f"✅ [InventoryMgmt] Repaired {affected} orphaned TRS link(s)")
                else:
                    print("✅ [InventoryMgmt] TRS links checked — no orphans found")
            except Exception as e:
                # Quiet failure: this is a maintenance task, not a critical
                # path. If the JOIN doesn't work on this MySQL version we'd
                # rather not block startup.
                import traceback; traceback.print_exc()
                print(f"ℹ️  [InventoryMgmt] TRS link repair skipped: {e}")
            try: _bs_conn.close()
            except Exception: pass
    except Exception:
        pass

    # ── Filesystem upload root ─────────────────────────────────────
    # Files live under <flask-root>/uploads/inventory_mgmt/.
    # flask-root = the parent folder of inventory_mgmt.py.
    _MODULE_DIR  = os.path.dirname(os.path.abspath(__file__))
    _APP_ROOT    = os.path.dirname(_MODULE_DIR)
    # If app.py imports from a flat tree, _APP_ROOT might be the module dir
    # itself — handle both by preferring whichever already has "uploads".
    if not os.path.isdir(os.path.join(_APP_ROOT, "uploads")) and \
       os.path.isdir(os.path.join(_MODULE_DIR, "uploads")):
        _APP_ROOT = _MODULE_DIR
    _GRN_UPLOAD_ROOT = os.path.join(_APP_ROOT, "uploads", "inventory_mgmt")
    try:
        os.makedirs(_GRN_UPLOAD_ROOT, exist_ok=True)
        print(f"✅ [InventoryMgmt] Upload root: {_GRN_UPLOAD_ROOT}")
    except Exception as e:
        print(f"⚠️  [InventoryMgmt] Could not create upload root {_GRN_UPLOAD_ROOT}: {e}")

    # Allowed upload MIME types — PDF + JPG + PNG per Tarak's spec.
    _GRN_FILE_ALLOWED_MIME = {"application/pdf", "image/jpeg", "image/jpg", "image/png"}
    _GRN_FILE_MAX_BYTES    = 10 * 1024 * 1024   # 10 MB

    def _safe_filename_segment(s):
        """Strip dangerous chars from a filename. Keep dots/letters/numbers/underscores/dashes."""
        s = str(s or "").strip()
        # Collapse whitespace to single underscore, drop everything else weird.
        s = re.sub(r"\s+", "_", s)
        s = re.sub(r"[^A-Za-z0-9._\-]", "", s)
        # Avoid empty / dot-only names.
        if not s or set(s) <= {"."}:
            s = "file"
        # Cap length (Windows path limits).
        return s[:120]

    def _grn_file_disk_path(relative_path):
        """Resolve a stored relative_path to an absolute filesystem path,
        with traversal guard."""
        # Normalize separators — relative_path uses forward slashes in DB.
        rp = (relative_path or "").lstrip("/\\").replace("\\", "/")
        abs_path = os.path.normpath(os.path.join(_GRN_UPLOAD_ROOT, *rp.split("/")))
        # Make sure the resolved path is still inside the upload root.
        root = os.path.normpath(_GRN_UPLOAD_ROOT)
        if not abs_path.startswith(root + os.sep) and abs_path != root:
            return None
        return abs_path

    @app.route("/inventory_mgmt")
    @_login_required
    def inventory_mgmt_page():
        """Render the inventory dashboard.

        Access-control model (redesigned May 2026):
        -------------------------------------------
        All gating decisions on this page derive from a single source of
        truth: the access dict computed by inventory_access._resolve_user_access().
        That resolver handles in order:
            admin role  →  full access (bypass)
            per-user row →  matched by candidate identifiers (case-insensitive)
            group row    →  matched the same way through group_members
            defaults     →  all 'off' (default-deny)

        The dict + diagnostic info ('source' = admin/row/group/defaults,
        'matched_name' = which session id matched, 'group_id' if from a
        group) is passed to the template. The template renders it into
        window.INV_CTX which IS the source of truth on the JS side too —
        we mirror it to window._invAccess synchronously so dependent
        modules (~30 of them) see the right values from page-load 0
        without waiting on the separate /access/me fetch.

        For features that have granular sub-permissions (GRN: view / new /
        edit / delete / view_print), we precompute the per-action flags
        here using _inv_grn_caps() / _inv_opening_caps() and pass them
        as separate Jinja vars (can_create_grn, etc.). Sections without
        a granular scheme (Items, Suppliers, Brands) fall back to the
        admin-only is_admin flag, which is unchanged from the previous
        design — those sections' UI gates haven't been re-modeled yet.
        """
        from flask import session
        # ── Resolve access via the unified resolver ──
        resolved = {"is_admin": False, "source": "defaults",
                    "matched_name": None, "group_id": None,
                    "access": {}}
        try:
            try:
                from .inventory_access import _resolve_user_access, _user_candidates
            except Exception:
                from inventory_access import _resolve_user_access, _user_candidates
            try:
                resolved = _resolve_user_access() or resolved
            except Exception:
                pass
        except Exception:
            _user_candidates = None

        access_caps    = resolved.get("access") or {}
        is_admin       = bool(resolved.get("is_admin"))
        access_source  = resolved.get("source") or "defaults"
        access_matched = resolved.get("matched_name")
        access_gid     = resolved.get("group_id")
        # access_has_row is now "did we resolve to anything beyond defaults?"
        # — kept as a Jinja var for backward-compat with any callers that
        # might still consult it, but the JS no longer uses it as a gating
        # decision (caps dict is authoritative).
        access_has_row = access_source in ("row", "group", "admin")

        # ── Per-feature edit caps (GRN granular permissions) ──
        try:
            try:
                from .inventory_access import _inv_grn_caps, _inv_opening_caps, _inv_items_caps
            except Exception:
                from inventory_access import _inv_grn_caps, _inv_opening_caps, _inv_items_caps
            grn_caps     = _inv_grn_caps() or {}
            opening_caps = _inv_opening_caps() or {}
            items_caps   = _inv_items_caps() or {}
        except Exception:
            grn_caps = {}
            opening_caps = {}
            items_caps = {}

        # Admin always passes; non-admins go through fine-grained caps.
        can_create_grn  = bool(is_admin or grn_caps.get("create"))
        can_edit_grn    = bool(is_admin or grn_caps.get("edit"))
        can_delete_grn  = bool(is_admin or grn_caps.get("delete"))
        can_edit_opening = bool(is_admin or opening_caps.get("edit"))
        # Items (RM master): view shows the tab; edit unlocks New/Edit/Delete.
        can_view_items   = bool(is_admin or items_caps.get("view"))
        can_edit_items   = bool(is_admin or items_caps.get("edit"))

        # ── Diagnostic log ──
        # Every page-load prints one line summarising what the backend
        # computed. Grep Flask logs for [InventoryMgmt] when a user
        # reports a permission issue — this line tells you exactly which
        # session keys the user has, which one matched the access row,
        # and the actual cap values resolved.
        try:
            import sys
            sess_keys = {k: session.get(k) for k in
                         ("UID","User_Name","username","user_name","User_Type")
                         if session.get(k)}
            sample = {k: access_caps.get(k) for k in
                      ("suppliers","grn","opening_stock","stock_transfer",
                       "manage_godown","label_reprint","material_request",
                       "simple_transfer")}
            print(f"[InventoryMgmt] page-load: session={sess_keys} "
                  f"is_admin={is_admin} source={access_source} "
                  f"matched={access_matched!r} gid={access_gid} "
                  f"grn={{create:{can_create_grn},edit:{can_edit_grn},"
                  f"delete:{can_delete_grn}}} sample={sample}",
                  file=sys.stderr)
        except Exception:
            pass

        # ── Location lock (May 2026) ──
        # Resolve the user's pinned godown so the template can bootstrap
        # window._invAccess.locked_godown_id synchronously. Admins are
        # never locked. NULL otherwise = no lock.
        locked_godown_id   = None
        locked_godown_name = None
        try:
            if not is_admin:
                try:
                    from .inventory_access import _locked_godown_id as _lg
                except Exception:
                    from inventory_access import _locked_godown_id as _lg
                locked_godown_id = _lg()
                if locked_godown_id:
                    try:
                        conn = sampling_portal.get_db_connection()
                        try:
                            row = conn.execute(
                                "SELECT name FROM procurement_godowns WHERE id=%s",
                                (locked_godown_id,)
                            ).fetchone()
                            if row:
                                locked_godown_name = row.get("name") if hasattr(row, "get") else row[0]
                        finally:
                            try: conn.close()
                            except Exception: pass
                    except Exception:
                        pass
        except Exception:
            pass

        # ── Pre-fetch voucher types server-side ──
        # The GRN Type dropdown is populated via /api/gop/voucher_types
        # (procurement endpoint), which 403s non-admins. The inventory
        # passthrough /api/inventory_mgmt/voucher_types also returns
        # nothing in some setups. To make sure non-admins always see
        # the full type list (per HCP's policy that any GRN-eligible user
        # should be able to pick from all configured GRN types), we read
        # gop_voucher_type_masters directly here and pass the rows through
        # to the template. The JS still tries the API for late-loaded types
        # but the dropdown is now usable even when the API returns empty.
        voucher_types_grn = []
        voucher_types_po  = []
        try:
            conn = sampling_portal.get_db_connection()
            if conn:
                try:
                    for parent_t, target in (("grn", voucher_types_grn),
                                             ("po",  voucher_types_po)):
                        try:
                            rs = conn.execute(
                                # COALESCE on sort_order so NULL values sort last
                                # instead of breaking the ORDER BY (some legacy
                                # rows have NULL sort_order).
                                "SELECT * FROM gop_voucher_type_masters "
                                "WHERE parent_type=%s "
                                "ORDER BY COALESCE(sort_order, 9999), name",
                                (parent_t,)
                            ).fetchall() or []
                            for r in rs:
                                d = dict(r) if hasattr(r, "keys") else r
                                # Normalise the active flag — different drivers
                                # return 0/1, True/False, or '0'/'1'.
                                d["is_active"] = bool(
                                    d.get("is_active") in (1, True, "1", "true", "True")
                                )
                                target.append(d)
                        except Exception:
                            # Table missing or unreadable — just leave the
                            # list empty and let the JS fallback run.
                            import traceback as _tb
                            _tb.print_exc()
                finally:
                    try: conn.close()
                    except Exception: pass
        except Exception:
            pass

        return render_template(
            "inventory/inventory_mgmt.html",
            user_name=session.get("User_Name"),
            role=session.get("User_Type"),
            # Admin-only flag — kept named `can_edit` for backward-compat
            # with existing {% if can_edit %} blocks in the template that
            # gate sections without granular permissions (Items, Suppliers,
            # Brands, admin tools). These still need is_admin because no
            # per-feature cap exists for them yet.
            can_edit=is_admin,
            is_admin=is_admin,
            can_create_grn=can_create_grn,
            can_edit_grn=can_edit_grn,
            can_delete_grn=can_delete_grn,
            can_edit_opening=can_edit_opening,
            can_view_items=can_view_items,
            can_edit_items=can_edit_items,
            access_caps=access_caps,
            access_has_row=access_has_row,
            access_source=access_source,
            access_matched=access_matched,
            access_group_id=access_gid,
            locked_godown_id=locked_godown_id,
            locked_godown_name=locked_godown_name,
            voucher_types_grn=voucher_types_grn,
            voucher_types_po=voucher_types_po,
        )

    # ── LOGOUT ────────────────────────────────────────────────────────────
    # Clears the user's session and redirects to the parent portal's login
    # screen. We try url_for('login') first to match the convention used
    # elsewhere in the app; if that endpoint isn't registered for any
    # reason we fall back to '/' (the portal redirects unauthenticated
    # users to login anyway). This mirrors pm_stock's logout pattern.
    @app.route("/inventory_mgmt/logout")
    def inventory_mgmt_logout():
        try:
            session.clear()
        except Exception:
            # Defensive: even if clear() fails, drop the auth keys explicitly.
            for k in ("logged_in", "User_Name", "UID", "User_Type"):
                session.pop(k, None)
        try:
            return redirect(url_for("login"))
        except Exception:
            return redirect("/")

    # ═══ LIST ITEMS ════════════════════════════════════════════════════════════
    @app.route("/api/inventory_mgmt/items")
    def api_inv_items():
        if not _can_inventory():
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        dept = _norm_dept(request.args.get("department"))
        if not dept:
            return jsonify({"status": "error", "message": "department is required (RM/PM/FG)"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500

        try:
            items = []

            if dept == DEPT_RM:
                # rm_stock (the Excel/Tally dump) is ONLY consulted in the
                # legacy fallback when USE_RM_BOX_STOCK is False. Reading and
                # parsing that workbook on every items load is expensive and
                # pointless when box-stock is the source, so skip it.
                rm_box_stock = _read_rm_box_stock_by_material(conn) if USE_RM_BOX_STOCK else {}
                rm_stock = {} if USE_RM_BOX_STOCK else _read_rm_stock_from_xlsx()
                rows = conn.execute("""
                    SELECT  m.id, m.material_name AS name, m.supplier_name,
                            m.last_purchase_rate, m.gst_rate, m.hsn_code,
                            m.uom, m.std_pack_size, m.msl, m.lead_time_days,
                            m.ordered_qty, m.buffer_qty,
                            m.group_id, m.material_type_id,
                            g.group_name,
                            t.type_name AS material_type,
                            m.updated_at, m.updated_by,
                            lp.last_supplier, lp.last_rate, lp.last_date
                    FROM procurement_materials m
                    LEFT JOIN procurement_material_groups g ON g.id = m.group_id
                    LEFT JOIN procurement_material_types  t ON t.id = m.material_type_id
                    LEFT JOIN inventory_last_purchase     lp
                           ON lp.department='RM' AND lp.item_id = m.id
                    ORDER BY m.material_name ASC
                """).fetchall()
                for r in rows:
                    key = (r["name"] or "").strip().lower()
                    items.append({
                        "id": int(r["id"]),
                        "department": "RM",
                        "name": r["name"] or "",
                        "last_supplier": r["last_supplier"] or r["supplier_name"] or "",
                        "last_rate": float(r["last_rate"] or r["last_purchase_rate"] or 0) or None,
                        "last_date": (str(r["last_date"]) if r["last_date"] else ""),
                        "gst_rate": float(r["gst_rate"]) if r["gst_rate"] is not None else None,
                        "hsn_code": r["hsn_code"] or "",
                        "uom": r["uom"] or "KG",
                        "std_pack_size": r["std_pack_size"] or "",
                        "msl": float(r["msl"]) if r["msl"] is not None else None,
                        "lead_time_days": int(r["lead_time_days"]) if r["lead_time_days"] else None,
                        "ordered_qty": float(r["ordered_qty"] or 0),
                        "buffer_qty": float(r["buffer_qty"] or 0),
                        "group_id": r["group_id"],
                        "group_name": r["group_name"] or "",
                        "material_type_id": r["material_type_id"],
                        "material_type": r["material_type"] or "",
                        "brand_id": None,
                        "brand_name": "",
                        "brand_color": "",
                        "in_stock": (rm_box_stock.get(int(r["id"]), 0.0)
                                     if USE_RM_BOX_STOCK
                                     else rm_stock.get(key, 0.0)),
                        "updated_at": (str(r["updated_at"])[:16] if r["updated_at"] else ""),
                        "updated_by": r["updated_by"] or "",
                    })

            elif dept == DEPT_PM:
                pm_stock = _read_pm_stock(conn)
                rows = conn.execute("""
                    SELECT  p.id, p.product_name AS name, p.pm_type,
                            p.brand_id, p.hsn_code, p.gst_rate, p.uom, p.min_stock,
                            p.material_type, p.is_active,
                            b.name AS brand_name, b.color AS brand_color,
                            lp.last_supplier, lp.last_rate, lp.last_date
                    FROM pm_products p
                    LEFT JOIN procurement_brands b ON b.id = p.brand_id
                    LEFT JOIN inventory_last_purchase lp
                           ON lp.department='PM' AND lp.item_id = p.id
                    ORDER BY p.product_name ASC
                """).fetchall()
                for r in rows:
                    items.append({
                        "id": int(r["id"]),
                        "department": "PM",
                        "name": r["name"] or "",
                        "pm_type": r["pm_type"] or "",
                        "material_type": r["material_type"] or "",
                        "last_supplier": r["last_supplier"] or "",
                        "last_rate": float(r["last_rate"]) if r["last_rate"] is not None else None,
                        "last_date": (str(r["last_date"]) if r["last_date"] else ""),
                        "gst_rate": float(r["gst_rate"]) if r["gst_rate"] is not None else None,
                        "hsn_code": r["hsn_code"] or "",
                        "uom": r["uom"] or "NOS",
                        "msl": float(r["min_stock"] or 0),
                        "brand_id": int(r["brand_id"]) if r["brand_id"] else None,
                        "brand_name": r["brand_name"] or "",
                        "brand_color": r["brand_color"] or "#6366f1",
                        "in_stock": float(pm_stock.get(int(r["id"]), 0)),
                        "is_active": int(r["is_active"] if r["is_active"] is not None else 1),
                    })

            elif dept == DEPT_FG:
                fg_stock = _read_fg_stock(conn)
                try:
                    rows = conn.execute("""
                        SELECT  f.id, f.fg_code, f.fg_name AS name, f.sku_size, f.uom,
                                f.brand_id, f.formulation_batch, f.pm_links, f.remarks, f.is_active,
                                f.hsn_code, f.gst_rate,
                                f.last_supplier, f.last_purchase_rate,
                                b.name AS brand_name, b.color AS brand_color
                        FROM FG_Names f
                        LEFT JOIN procurement_brands b ON b.id = f.brand_id
                        ORDER BY f.fg_name ASC
                    """).fetchall()
                except Exception:
                    rows = conn.execute("""
                        SELECT  f.id, f.fg_code, f.fg_name AS name, f.sku_size, f.uom,
                                f.brand_id, f.formulation_batch, f.remarks, f.is_active,
                                b.name AS brand_name, b.color AS brand_color
                        FROM FG_Names f
                        LEFT JOIN procurement_brands b ON b.id = f.brand_id
                        ORDER BY f.fg_name ASC
                    """).fetchall()

                for r in rows:
                    row_d = dict(r)
                    # Parse pm_links JSON → list of ints
                    pm_links_raw = row_d.get('pm_links')
                    pm_ids = []
                    if pm_links_raw:
                        try:
                            if isinstance(pm_links_raw, (list, tuple)):
                                pm_ids = [int(x) for x in pm_links_raw if x]
                            else:
                                parsed = json.loads(pm_links_raw) if isinstance(pm_links_raw, str) else pm_links_raw
                                if isinstance(parsed, list):
                                    pm_ids = [int(x) for x in parsed if x]
                        except Exception:
                            pm_ids = []

                    items.append({
                        "id": int(r["id"]),
                        "department": "FG",
                        "name": r["name"] or "",
                        "fg_code": r["fg_code"] or "",
                        "sku_size": r["sku_size"] or "",
                        "last_supplier": row_d.get("last_supplier") or "",
                        "last_rate": float(row_d["last_purchase_rate"]) if row_d.get("last_purchase_rate") is not None else None,
                        "last_date": "",
                        "gst_rate": float(row_d["gst_rate"]) if row_d.get("gst_rate") is not None else None,
                        "hsn_code": row_d.get("hsn_code") or "",
                        "uom": r["uom"] or "NOS",
                        "brand_id": int(r["brand_id"]) if r["brand_id"] else None,
                        "brand_name": r["brand_name"] or "",
                        "brand_color": r["brand_color"] or "#16a34a",
                        "formulation_batch": r["formulation_batch"] or "",
                        "pm_links": pm_ids,
                        "remarks": r["remarks"] or "",
                        "in_stock": float(fg_stock.get(int(r["id"]), 0)),
                        "is_active": int(r["is_active"] if r["is_active"] is not None else 1),
                    })

            return jsonify({"status": "ok", "department": dept, "items": items, "total": len(items)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ═══ STOCK BREAKDOWN (godown-wise) ════════════════════════════════════════
    @app.route("/api/inventory_mgmt/global_search")
    def api_inv_global_search():
        """Portal-wide fuzzy search for the floating dock. Returns hits across
        materials + vouchers (GRN / Delivery Note / Stock Transfer / Material
        Request) in one pass. Each hit carries a `kind` so the dock can route."""
        if not _can_inventory():
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        q = (request.args.get("q") or "").strip()
        if not q or len(q) < 2:
            return jsonify({"status": "ok", "materials": [], "vouchers": []})
        like = f"%{q}%"
        debug = request.args.get("debug")
        dbg = {}
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            # ── Materials ──
            materials = []
            try:
                mrows = conn.execute(
                    "SELECT id, material_name AS name, uom FROM procurement_materials "
                    "WHERE material_name LIKE %s ORDER BY material_name ASC LIMIT 10",
                    (like,),
                ).fetchall()
                materials = [{"id": int(r["id"]), "name": r["name"] or "", "uom": r["uom"] or ""}
                             for r in mrows]
                dbg["materials_count"] = len(materials)
            except Exception as me:
                dbg["materials_error"] = str(me)

            vouchers = []
            # ── GRN ── (only real GRNs, matching the GRN list filter; this
            #    excludes REJOUT / reissue vouchers which have no GRN-screen view)
            try:
                grn = conn.execute(
                    "SELECT 'grn' AS kind, id, grn_num AS voucher_no, supplier_name AS detail1, "
                    "grn_date AS detail2 FROM procurement_grn "
                    "WHERE (grn_type IS NULL OR grn_type = 'GRN' OR grn_type = '') "
                    "AND (grn_num LIKE %s OR supplier_name LIKE %s OR invoice_num LIKE %s) "
                    "ORDER BY id DESC LIMIT 5",
                    (like, like, like),
                ).fetchall()
                for r in grn: vouchers.append(dict(r))
                dbg["grn_count"] = len(grn)
            except Exception as e:
                dbg["grn_error"] = str(e)
            # ── Delivery Note ──
            try:
                dn = conn.execute(
                    "SELECT 'dn' AS kind, id, dn_no AS voucher_no, supplier_name AS detail1, "
                    "dn_date AS detail2 FROM inventory_dn "
                    "WHERE dn_no LIKE %s OR supplier_name LIKE %s ORDER BY id DESC LIMIT 5",
                    (like, like),
                ).fetchall()
                for r in dn: vouchers.append(dict(r))
                dbg["dn_count"] = len(dn)
            except Exception as e:
                dbg["dn_error"] = str(e)
            # ── Stock Transfer ──
            try:
                tr = conn.execute(
                    "SELECT 'transfer' AS kind, transfer_id AS id, transfer_no AS voucher_no, "
                    "status AS detail1, created_at AS detail2 FROM rm_stock_transfers "
                    "WHERE transfer_no LIKE %s ORDER BY transfer_id DESC LIMIT 5",
                    (like,),
                ).fetchall()
                for r in tr: vouchers.append(dict(r))
                dbg["transfer_count"] = len(tr)
            except Exception as e:
                dbg["transfer_error"] = str(e)
            # ── Material Request ──
            try:
                mr = conn.execute(
                    "SELECT 'mr' AS kind, id, request_no AS voucher_no, requested_by AS detail1, "
                    "request_date AS detail2 FROM inventory_material_requests "
                    "WHERE request_no LIKE %s OR requested_by LIKE %s ORDER BY id DESC LIMIT 5",
                    (like, like),
                ).fetchall()
                for r in mr: vouchers.append(dict(r))
                dbg["mr_count"] = len(mr)
            except Exception as e:
                dbg["mr_error"] = str(e)

            # rank: prefix match first, then newest
            ql = q.lower()
            def _score(v):
                vn = (v.get("voucher_no") or "").lower()
                return (0 if vn.startswith(ql) else (1 if ql in vn else 2), -(v.get("id") or 0))
            vouchers.sort(key=_score)
            vouchers = vouchers[:10]
            for v in vouchers:
                if v.get("detail2") is not None:
                    v["detail2"] = str(v["detail2"])
            conn.close()
            out = {"status": "ok", "materials": materials, "vouchers": vouchers}
            if debug:
                out["debug"] = dbg
            return jsonify(out)
        except Exception as e:
            try:
                conn.close()
            except Exception:
                pass
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/inventory_mgmt/box_track")
    def api_inv_box_track():
        """Box tracker: full movement timeline for one box (by box_id or code).
        Returns the box header + chronological movements (type, from→to, qty,
        when, by whom)."""
        if not session.get("logged_in") and not session.get("UID"):
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        code = (request.args.get("code") or "").strip()
        try:
            box_id = int(request.args.get("box_id") or 0)
        except (TypeError, ValueError):
            box_id = 0
        conn = sampling_portal.get_db_connection()
        if conn is None:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            if not box_id and code:
                br = conn.execute(
                    "SELECT box_id FROM rm_boxes WHERE box_code=%s LIMIT 1", (code,)
                ).fetchone()
                if br:
                    box_id = int(br["box_id"] if hasattr(br, "get") else br[0])
            if not box_id:
                conn.close(); return jsonify({"status": "error", "message": "Box not found"}), 404

            box = conn.execute(
                "SELECT b.box_id, b.box_code, b.material_id, b.per_box_qty, b.uom, "
                "b.current_status, b.source, b.grn_no, "
                "COALESCE(m.material_name,'') AS material_name, "
                "COALESCE(g.name,'') AS godown_name "
                "FROM rm_boxes b "
                "LEFT JOIN procurement_materials m ON m.id=b.material_id "
                "LEFT JOIN procurement_godowns g ON g.id=b.current_godown_id "
                "WHERE b.box_id=%s", (box_id,),
            ).fetchone()
            if not box:
                conn.close(); return jsonify({"status": "error", "message": "Box not found"}), 404

            gmap = {}
            for g in conn.execute("SELECT id, name FROM procurement_godowns").fetchall():
                gmap[int(g["id"])] = g["name"] or ""

            mv = conn.execute(
                "SELECT movement_type, from_godown_id, to_godown_id, qty, movement_at, "
                "moved_by, remarks FROM rm_box_movements "
                "WHERE box_id=%s ORDER BY movement_at ASC, movement_id ASC", (box_id,),
            ).fetchall()
            timeline = []
            for r in mv:
                timeline.append({
                    "type": r["movement_type"] or "",
                    "from": gmap.get(int(r["from_godown_id"] or 0), "") if r["from_godown_id"] else "",
                    "to": gmap.get(int(r["to_godown_id"] or 0), "") if r["to_godown_id"] else "",
                    "qty": round(float(r["qty"] or 0), 3),
                    "at": str(r["movement_at"])[:16] if r["movement_at"] else "",
                    "by": r["moved_by"] or "",
                    "remarks": r["remarks"] or "",
                })
            d = dict(box)
            conn.close()
            return jsonify({"status": "ok", "box": {
                "box_code": d.get("box_code", ""), "material_name": d.get("material_name", ""),
                "qty": round(float(d.get("per_box_qty") or 0), 3), "uom": d.get("uom", ""),
                "status": d.get("current_status", ""), "godown": d.get("godown_name", ""),
                "source": d.get("source", ""), "grn_no": d.get("grn_no", ""),
            }, "timeline": timeline})
        except Exception as e:
            try: conn.close()
            except Exception: pass
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/inventory_mgmt/item_godown_stock")
    def api_inv_item_godown_stock():
        """Stock check: a material's in-stock qty broken down by godown.
        Accepts material_id, or a box code/box_id to resolve the material."""
        if not session.get("logged_in") and not session.get("UID"):
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        code = (request.args.get("code") or "").strip()
        try:
            material_id = int(request.args.get("material_id") or 0)
        except (TypeError, ValueError):
            material_id = 0
        conn = sampling_portal.get_db_connection()
        if conn is None:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            if not material_id and code:
                br = conn.execute(
                    "SELECT material_id FROM rm_boxes WHERE box_code=%s LIMIT 1", (code,)
                ).fetchone()
                if br:
                    material_id = int(br["material_id"] if hasattr(br, "get") else br[0])
            if not material_id:
                conn.close(); return jsonify({"status": "error", "message": "Material not found"}), 404

            mat = conn.execute(
                "SELECT material_name AS name, uom, msl FROM procurement_materials WHERE id=%s",
                (material_id,),
            ).fetchone()
            # per-godown qty from in_stock boxes
            rows = conn.execute(
                "SELECT b.current_godown_id AS gid, COALESCE(g.name,'') AS godown, "
                "COALESCE(SUM(b.per_box_qty),0) AS qty, COUNT(*) AS boxes "
                "FROM rm_boxes b LEFT JOIN procurement_godowns g ON g.id=b.current_godown_id "
                "WHERE b.material_id=%s AND b.current_status='in_stock' "
                "GROUP BY b.current_godown_id HAVING qty <> 0 ORDER BY qty DESC",
                (material_id,),
            ).fetchall()
            breakdown = [{"godown": r["godown"] or "(unassigned)",
                          "qty": round(float(r["qty"] or 0), 3),
                          "boxes": int(r["boxes"] or 0)} for r in rows]
            total = round(sum(x["qty"] for x in breakdown), 3)
            conn.close()
            return jsonify({"status": "ok",
                "material": {"name": (mat["name"] if mat else "") or "",
                             "uom": (mat["uom"] if mat else "") or "",
                             "msl": float(mat["msl"] or 0) if mat else 0},
                "breakdown": breakdown, "total": total})
        except Exception as e:
            try: conn.close()
            except Exception: pass
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/inventory_mgmt/stock_breakdown")
    def api_inv_stock_breakdown():
        if not _can_inventory():
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        dept = _norm_dept(request.args.get("department"))
        try:
            item_id = int(request.args.get("item_id") or 0)
        except (TypeError, ValueError):
            item_id = 0
        if not dept or not item_id:
            return jsonify({"status": "error", "message": "department and item_id required"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            rows_out = []
            item_name = ""
            uom = ""

            if dept == DEPT_RM:
                it = conn.execute(
                    "SELECT material_name, uom FROM procurement_materials WHERE id=%s",
                    (item_id,)
                ).fetchone()
                if it:
                    item_name = it["material_name"] or ""
                    uom = it["uom"] or "KG"

                    # Per-godown stock: real package-level sums from rm_boxes
                    # (in_stock boxes), totalled across all godowns. Falls back
                    # to the legacy Excel total on the default godown if box
                    # tracking is disabled.
                    if USE_RM_BOX_STOCK:
                        per_g = _read_rm_box_stock_for_material(conn, item_id)
                    else:
                        per_g = None
                        stk = _read_rm_stock_from_xlsx()
                        total_qty = float(stk.get((item_name or "").strip().lower(), 0.0))

                    # Get all godowns so the breakdown shows the full list.
                    # Sort default-first then alphabetic.
                    gd_rows = conn.execute("""
                        SELECT id, name, type, city, state, is_default
                        FROM procurement_godowns
                        ORDER BY is_default DESC, name ASC
                    """).fetchall()

                    rows_out = []
                    default_assigned = False
                    for g in gd_rows:
                        is_default = bool(g["is_default"])
                        name = g["name"] or ""
                        if USE_RM_BOX_STOCK:
                            # Real per-godown quantity from rm_boxes.
                            display_name = name + (" (default)" if is_default else "")
                            qty = float((per_g or {}).get(int(g["id"]), 0.0))
                        else:
                            # Legacy placeholder: full total on the default godown.
                            if is_default and not default_assigned:
                                display_name = name + " (default)"
                                qty = total_qty
                                default_assigned = True
                            else:
                                display_name = name
                                qty = 0.0
                        rows_out.append({
                            "godown_id":   int(g["id"]),
                            "godown_name": display_name,
                            "godown_type": g["type"] or "",
                            "city":        g["city"] or "",
                            "state":       g["state"] or "",
                            "qty":         qty,
                        })

                    # Edge case: no godown flagged is_default (legacy mode only).
                    if not USE_RM_BOX_STOCK and not default_assigned and rows_out:
                        rows_out[0]["qty"] = total_qty
                        rows_out[0]["godown_name"] += " (no default — auto-assigned)"

                    # Edge case: no godowns at all. Fall back to a single
                    # synthetic row so the breakdown isn't empty.
                    if not rows_out:
                        fallback_qty = (
                            float(sum((_read_rm_box_stock_for_material(conn, item_id) or {}).values()))
                            if USE_RM_BOX_STOCK else total_qty
                        )
                        rows_out = [{
                            "godown_id":   None,
                            "godown_name": "All locations (no godowns configured)",
                            "godown_type": "system",
                            "city": "", "state": "",
                            "qty":  fallback_qty,
                        }]

            elif dept == DEPT_PM:
                it = conn.execute(
                    "SELECT product_name, uom FROM pm_products WHERE id=%s",
                    (item_id,)
                ).fetchone()
                if it:
                    item_name = it["product_name"] or ""
                    uom = it["uom"] or "NOS"

                # Try pm_stock_log first (canonical transaction log)
                rows_fetched = False
                try:
                    rows = conn.execute("""
                        SELECT  g.id   AS godown_id,
                                g.name AS godown_name,
                                g.type AS godown_type,
                                g.city, g.state,
                                SUM(CASE
                                    WHEN l.txn_type IN ('grn','opening','inward','return','mtv_in','adjustment_in') THEN l.qty
                                    WHEN l.txn_type IN ('issue','outward','mtv_out','adjustment_out','consumption') THEN -l.qty
                                    ELSE 0
                                END) AS qty
                        FROM pm_stock_log l
                        LEFT JOIN procurement_godowns g ON g.id = l.godown_id
                        WHERE l.product_id = %s
                        GROUP BY g.id, g.name, g.type, g.city, g.state
                        HAVING qty <> 0
                        ORDER BY qty DESC, g.name ASC
                    """, (item_id,)).fetchall()
                    rows_out = [{
                        "godown_name": r["godown_name"] or "Factory Floor",
                        "godown_type": r["godown_type"] or "floor",
                        "city":        r["city"] or "",
                        "state":       r["state"] or "",
                        "qty":         float(r["qty"] or 0),
                    } for r in rows]
                    rows_fetched = True
                except Exception:
                    pass

                if not rows_fetched:
                    # Fallback to pm_stock snapshot
                    try:
                        rows = conn.execute("""
                            SELECT  g.id         AS godown_id,
                                    g.name       AS godown_name,
                                    g.type       AS godown_type,
                                    g.city, g.state,
                                    COALESCE(SUM(s.qty),0) AS qty
                            FROM pm_stock s
                            LEFT JOIN procurement_godowns g ON g.id = s.godown_id
                            WHERE s.product_id = %s
                            GROUP BY g.id, g.name, g.type, g.city, g.state
                            HAVING qty <> 0
                            ORDER BY qty DESC, g.name ASC
                        """, (item_id,)).fetchall()
                        rows_out = [{
                            "godown_name": r["godown_name"] or "(unassigned)",
                            "godown_type": r["godown_type"] or "",
                            "city":        r["city"] or "",
                            "state":       r["state"] or "",
                            "qty":         float(r["qty"] or 0),
                        } for r in rows]
                    except Exception:
                        rows_out = []

            elif dept == DEPT_FG:
                it = conn.execute(
                    "SELECT fg_name, uom FROM FG_Names WHERE id=%s",
                    (item_id,)
                ).fetchone()
                if it:
                    item_name = it["fg_name"] or ""
                    uom = it["uom"] or "NOS"
                try:
                    rows = conn.execute("""
                        SELECT  g.id         AS godown_id,
                                g.name       AS godown_name,
                                g.type       AS godown_type,
                                g.city, g.state,
                                COALESCE(SUM(s.qty),0) AS qty
                        FROM FG_stock s
                        LEFT JOIN procurement_godowns g ON g.id = s.godown_id
                        WHERE s.fg_id = %s
                        GROUP BY g.id, g.name, g.type, g.city, g.state
                        HAVING qty <> 0
                        ORDER BY qty DESC, g.name ASC
                    """, (item_id,)).fetchall()
                    rows_out = [{
                        "godown_name": r["godown_name"] or "(unassigned)",
                        "godown_type": r["godown_type"] or "",
                        "city":        r["city"] or "",
                        "state":       r["state"] or "",
                        "qty":         float(r["qty"] or 0),
                    } for r in rows]
                except Exception:
                    rows_out = []

            total = sum(r["qty"] for r in rows_out)
            return jsonify({
                "status":    "ok",
                "department": dept,
                "item_id":   item_id,
                "item_name": item_name,
                "uom":       uom,
                "rows":      rows_out,
                "total":     total,
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ═══ ITEM SAVE ═════════════════════════════════════════════════════════════
    @app.route("/api/inventory_mgmt/items/save", methods=["POST"])
    @_items_edit_required
    def api_inv_item_save():
        d = request.get_json(silent=True) or {}
        dept = _norm_dept(d.get("department"))
        if not dept:
            return jsonify({"status": "error", "message": "department is required"}), 400
        name = (d.get("name") or "").strip()
        if not name:
            return jsonify({"status": "error", "message": "name is required"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500

        uid = _user()
        item_id = d.get("id")
        try:
            if dept == DEPT_RM:
                fields = {
                    "material_name":      name,
                    "supplier_name":      d.get("last_supplier") or d.get("supplier_name") or None,
                    "last_purchase_rate": d.get("last_rate") if d.get("last_rate") not in (None, "",) else None,
                    "gst_rate":           d.get("gst_rate") if d.get("gst_rate") not in (None, "",) else None,
                    "hsn_code":           d.get("hsn_code") or None,
                    "uom":                d.get("uom") or "KG",
                    "msl":                d.get("msl") if d.get("msl") not in (None, "",) else None,
                    "lead_time_days":     d.get("lead_time_days") if d.get("lead_time_days") not in (None, "",) else None,
                    "std_pack_size":      d.get("std_pack_size") or None,
                    "group_id":           d.get("group_id") or None,
                    "material_type_id":   d.get("material_type_id") or None,
                    "updated_by":         uid,
                }
                if item_id:
                    sets  = ", ".join(f"{k}=%s" for k in fields.keys())
                    vals  = list(fields.values()) + [int(item_id)]
                    conn.execute(f"UPDATE procurement_materials SET {sets} WHERE id=%s", tuple(vals))
                else:
                    # ── Block exact-name duplicates on CREATE (Inventory RM) ──
                    # Only on new items (no item_id); editing keeps its own name.
                    existing = conn.execute(
                        "SELECT id FROM procurement_materials WHERE material_name=%s LIMIT 1",
                        (name,),
                    ).fetchone()
                    if existing:
                        try:
                            conn.close()
                        except Exception:
                            pass
                        return jsonify({
                            "status": "error",
                            "code": "duplicate_name",
                            "message": f"A material named \u201c{name}\u201d already exists "
                                       f"(id #{int(existing['id'])}). Please use the "
                                       f"existing item or choose a different name.",
                            "existing_id": int(existing["id"]),
                        }), 409

                    cols  = ", ".join(fields.keys())
                    marks = ", ".join(["%s"]*len(fields))
                    conn.execute(
                        f"INSERT INTO procurement_materials ({cols}) VALUES ({marks}) "
                        "ON DUPLICATE KEY UPDATE updated_at=NOW()",
                        tuple(fields.values())
                    )
                    row = conn.execute(
                        "SELECT id FROM procurement_materials WHERE material_name=%s LIMIT 1", (name,)
                    ).fetchone()
                    if row: item_id = int(row["id"])
                conn.commit()

            elif dept == DEPT_PM:
                fields = {
                    "product_name":  name,
                    "pm_type":       d.get("pm_type") or None,
                    "brand_id":      d.get("brand_id") or None,
                    "gst_rate":      d.get("gst_rate") if d.get("gst_rate") not in (None, "",) else None,
                    "hsn_code":      d.get("hsn_code") or None,
                    "uom":           d.get("uom") or "NOS",
                    "min_stock":     d.get("msl") if d.get("msl") not in (None, "",) else 0,
                    "material_type": d.get("material_type") or None,
                    "is_active":     1 if d.get("is_active", 1) else 0,
                }
                if item_id:
                    sets  = ", ".join(f"{k}=%s" for k in fields.keys())
                    vals  = list(fields.values()) + [int(item_id)]
                    conn.execute(f"UPDATE pm_products SET {sets} WHERE id=%s", tuple(vals))
                else:
                    cols  = ", ".join(fields.keys())
                    marks = ", ".join(["%s"]*len(fields))
                    conn.execute(f"INSERT INTO pm_products ({cols}) VALUES ({marks})", tuple(fields.values()))
                    row = conn.execute(
                        "SELECT id FROM pm_products WHERE product_name=%s ORDER BY id DESC LIMIT 1", (name,)
                    ).fetchone()
                    if row: item_id = int(row["id"])
                conn.commit()

            elif dept == DEPT_FG:
                fg_fields = {
                    "fg_code":              d.get("fg_code") or None,
                    "fg_name":              name,
                    "sku_size":             d.get("sku_size") or None,
                    "uom":                  d.get("uom") or "NOS",
                    "brand_id":             d.get("brand_id") or None,
                    "formulation_batch":    d.get("formulation_batch") or None,
                    "remarks":              d.get("remarks") or None,
                    "is_active":            1 if d.get("is_active", 1) else 0,
                }
                optional_fields = {
                    "hsn_code":           d.get("hsn_code") or None,
                    "gst_rate":           d.get("gst_rate") if d.get("gst_rate") not in (None, "",) else None,
                    "last_supplier":      d.get("last_supplier") or None,
                    "last_purchase_rate": d.get("last_rate") if d.get("last_rate") not in (None, "",) else None,
                }

                def _do_write(fields):
                    nonlocal item_id
                    if item_id:
                        sets  = ", ".join(f"{k}=%s" for k in fields.keys())
                        vals  = list(fields.values()) + [int(item_id)]
                        conn.execute(f"UPDATE FG_Names SET {sets} WHERE id=%s", tuple(vals))
                    else:
                        cols  = ", ".join(fields.keys())
                        marks = ", ".join(["%s"]*len(fields))
                        conn.execute(f"INSERT INTO FG_Names ({cols}) VALUES ({marks})", tuple(fields.values()))
                        row = conn.execute(
                            "SELECT id FROM FG_Names WHERE fg_name=%s ORDER BY id DESC LIMIT 1", (name,)
                        ).fetchone()
                        if row: item_id = int(row["id"])

                try:
                    _do_write({**fg_fields, **optional_fields})
                except Exception:
                    conn.rollback()
                    _do_write(fg_fields)
                conn.commit()

            if d.get("last_supplier") or d.get("last_rate"):
                try:
                    conn.execute("""
                        INSERT INTO inventory_last_purchase
                            (department, item_id, item_name, last_supplier, last_rate, gst_rate, last_date)
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                        ON DUPLICATE KEY UPDATE
                            last_supplier=VALUES(last_supplier),
                            last_rate=VALUES(last_rate),
                            gst_rate=VALUES(gst_rate),
                            last_date=VALUES(last_date),
                            item_name=VALUES(item_name)
                    """, (
                        dept, int(item_id) if item_id else 0, name,
                        d.get("last_supplier") or None,
                        d.get("last_rate") or None,
                        d.get("gst_rate") or None,
                        d.get("last_date") or None
                    ))
                    conn.commit()
                except Exception:
                    pass

            return jsonify({"status": "ok", "id": item_id, "message": "Saved"})
        except Exception as e:
            conn.rollback(); traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ═══ ITEM DELETE ═══════════════════════════════════════════════════════════
    @app.route("/api/inventory_mgmt/items/delete", methods=["POST"])
    @_items_edit_required
    def api_inv_item_delete():
        d = request.get_json(silent=True) or {}
        dept = _norm_dept(d.get("department"))
        ids  = d.get("ids") or ([d.get("id")] if d.get("id") else [])
        ids  = [int(x) for x in ids if x]
        if not dept or not ids:
            return jsonify({"status": "error", "message": "department and ids required"}), 400
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        table = {"RM":"procurement_materials", "PM":"pm_products", "FG":"FG_Names"}[dept]
        placeholders = ",".join(["%s"]*len(ids))
        try:
            conn.execute(f"DELETE FROM {table} WHERE id IN ({placeholders})", tuple(ids))
            conn.execute(
                f"DELETE FROM inventory_last_purchase WHERE department=%s AND item_id IN ({placeholders})",
                (dept,) + tuple(ids)
            )
            conn.commit()
            return jsonify({"status": "ok", "deleted": len(ids)})
        except Exception as e:
            conn.rollback(); traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ═══ BRANDS ═══════════════════════════════════════════════════════════════
    @app.route("/api/inventory_mgmt/brands")
    def api_inv_brands():
        if not _can_inventory():
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        dept = _norm_dept(request.args.get("department"))
        # If "for_item=1" is passed, return ALL brands regardless of dept mapping —
        # this is used by the item-creation dropdown so newly added brands (without
        # a dept mapping yet) still appear and can be picked.
        for_item = request.args.get("for_item") in ("1", "true", "True", "yes")
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            if dept and not for_item:
                rows = conn.execute("""
                    SELECT b.id, b.name, b.color, b.text_color
                    FROM procurement_brands b
                    JOIN inventory_brand_dept bd ON bd.brand_id = b.id
                    WHERE bd.department = %s
                    ORDER BY b.name ASC
                """, (dept,)).fetchall()
                # If no mapping exists for this dept yet, fall back to ALL brands
                # so the UI isn't empty on a fresh install.
                if not rows:
                    rows = conn.execute(
                        "SELECT id, name, color, text_color FROM procurement_brands ORDER BY name ASC"
                    ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT id, name, color, text_color FROM procurement_brands ORDER BY name ASC"
                ).fetchall()

            all_mapping = conn.execute(
                "SELECT brand_id, department FROM inventory_brand_dept"
            ).fetchall()
            by_brand = {}
            for m in all_mapping:
                by_brand.setdefault(int(m["brand_id"]), []).append(m["department"])

            out = [{
                "id": int(r["id"]),
                "name": r["name"],
                "color": r["color"] or "#6366f1",
                "text_color": r["text_color"] or "#ffffff",
                "departments": by_brand.get(int(r["id"]), []),
            } for r in rows]
            return jsonify({"status": "ok", "brands": out})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    @app.route("/api/inventory_mgmt/brands/save", methods=["POST"])
    @_edit_required
    def api_inv_brand_save():
        d = request.get_json(silent=True) or {}
        brand_id = d.get("id")
        name     = (d.get("name") or "").strip()
        color    = d.get("color") or "#6366f1"
        text_color = d.get("text_color") or "#ffffff"
        # Per user spec: brands apply to all departments by default.
        # Accept an explicit list from the client, but if it's empty or missing,
        # default to all three departments.
        depts    = [x for x in (d.get("departments") or []) if _norm_dept(x)]
        if not depts:
            depts = ["RM", "PM", "FG"]
        if not name:
            return jsonify({"status": "error", "message": "name is required"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            if brand_id:
                conn.execute(
                    "UPDATE procurement_brands SET name=%s, color=%s, text_color=%s WHERE id=%s",
                    (name, color, text_color, int(brand_id))
                )
            else:
                conn.execute(
                    "INSERT INTO procurement_brands (name, color, text_color) VALUES (%s,%s,%s)",
                    (name, color, text_color)
                )
                row = conn.execute(
                    "SELECT id FROM procurement_brands WHERE name=%s LIMIT 1", (name,)
                ).fetchone()
                brand_id = int(row["id"]) if row else None

            if brand_id:
                conn.execute("DELETE FROM inventory_brand_dept WHERE brand_id=%s", (int(brand_id),))
                for dp in depts:
                    conn.execute(
                        "INSERT IGNORE INTO inventory_brand_dept (brand_id, department) VALUES (%s,%s)",
                        (int(brand_id), dp)
                    )
            conn.commit()
            return jsonify({"status": "ok", "id": brand_id})
        except Exception as e:
            conn.rollback(); traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    @app.route("/api/inventory_mgmt/brands/delete", methods=["POST"])
    @_edit_required
    def api_inv_brand_delete():
        d = request.get_json(silent=True) or {}
        ids = d.get("ids") or ([d.get("id")] if d.get("id") else [])
        ids = [int(x) for x in ids if x]
        if not ids:
            return jsonify({"status": "error", "message": "ids required"}), 400
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            placeholders = ",".join(["%s"]*len(ids))
            conn.execute(f"DELETE FROM inventory_brand_dept WHERE brand_id IN ({placeholders})", tuple(ids))
            conn.execute(f"DELETE FROM procurement_brands WHERE id IN ({placeholders})", tuple(ids))
            conn.commit()
            return jsonify({"status": "ok", "deleted": len(ids)})
        except Exception as e:
            conn.rollback(); traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ═══════════════════════════════════════════════════════════════════════
    # BRAND MANAGER endpoints — mirror /api/procurement/brands* exactly so
    # the Brand Manager modal (ported from procurement.html) works on this
    # page for users with inventory access (and not necessarily procurement).
    # Same DB table (`procurement_brands`); just inventory-scoped auth.
    # ═══════════════════════════════════════════════════════════════════════

    @app.route("/api/inventory_mgmt/brands", methods=["POST"])
    @_edit_required
    def api_inv_brands_create():
        d    = request.get_json() or {}
        name = (d.get("name") or "").strip()
        color= (d.get("color") or "#6366f1").strip()
        text_color = (d.get("text_color") or "#ffffff").strip()
        if not name:
            return jsonify({"status":"error","message":"name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            conn.execute(
                "INSERT INTO procurement_brands (name,color,text_color) VALUES (%s,%s,%s)",
                (name, color, text_color)
            )
            conn.commit()
            bid = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
            conn.close()
            return jsonify({"status":"ok","id":bid,"name":name,"color":color,"text_color":text_color})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/inventory_mgmt/brands/bulk", methods=["POST"])
    @_edit_required
    def api_inv_brands_bulk_create():
        """
        Create multiple brands at once. Pre-detects duplicates by name
        (case-insensitive). Body: {"brands":[{"name":..,"color":..,"text_color":..}, ...]}
        Returns: {"status":"ok","added":[...],"skipped":[{"name":..,"reason":..}]}
        """
        d = request.get_json() or {}
        items = d.get("brands") or []
        if not isinstance(items, list):
            return jsonify({"status":"error","message":"brands must be a list"}), 400

        cleaned, skipped, seen_lower = [], [], set()
        for it in items:
            if not isinstance(it, dict):
                continue
            name = (it.get("name") or "").strip()
            if not name:
                continue
            if len(name) > 200:
                skipped.append({"name": name[:60], "reason": "too_long"})
                continue
            color = (it.get("color") or "#6366f1").strip() or "#6366f1"
            text_color = (it.get("text_color") or "#ffffff").strip() or "#ffffff"
            low = name.lower()
            if low in seen_lower:
                skipped.append({"name": name, "reason": "duplicate"})
                continue
            seen_lower.add(low)
            cleaned.append((name, color, text_color))

        if not cleaned and not skipped:
            return jsonify({"status":"error","message":"no valid brands provided"}), 400

        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            existing = conn.execute("SELECT id, name FROM procurement_brands").fetchall() or []
            existing_lower = { (r["name"] or "").lower() for r in existing }

            to_insert = []
            for (nm, c, tc) in cleaned:
                if nm.lower() in existing_lower:
                    skipped.append({"name": nm, "reason": "duplicate"})
                else:
                    to_insert.append((nm, c, tc))

            added = []
            for (nm, c, tc) in to_insert:
                try:
                    conn.execute(
                        "INSERT INTO procurement_brands (name,color,text_color) VALUES (%s,%s,%s)",
                        (nm, c, tc)
                    )
                    bid = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
                    added.append({"id": bid, "name": nm, "color": c, "text_color": tc})
                except Exception:
                    skipped.append({"name": nm, "reason": "duplicate"})

            conn.commit(); conn.close()
            return jsonify({"status":"ok","added": added, "skipped": skipped})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/inventory_mgmt/brands/import_excel", methods=["POST"])
    @_edit_required
    def api_inv_brands_import_excel():
        """
        Import brands from an uploaded .xlsx file — column A is brand names.
        Auto-assigns a deterministic-per-name pleasant colour and contrasting
        text colour. Pre-detects duplicates (existing + within file).
        """
        import os, tempfile, uuid, hashlib, colorsys
        if not OPENPYXL_OK:
            return jsonify({"status":"error","message":"openpyxl not installed on server"}), 500
        if "file" not in request.files:
            return jsonify({"status":"error","message":"No file uploaded"}), 400
        f = request.files["file"]
        if not f.filename.lower().endswith(".xlsx"):
            return jsonify({"status":"error","message":"Only .xlsx files accepted"}), 400

        tmp_dir  = tempfile.gettempdir()
        tmp_path = os.path.join(tmp_dir, f"brand_imp_{uuid.uuid4().hex}.xlsx")
        f.save(tmp_path)

        def _color_for(name):
            h_int = int(hashlib.md5(name.lower().encode("utf-8")).hexdigest()[:8], 16)
            hue   = (h_int % 360) / 360.0
            r, g, b = colorsys.hls_to_rgb(hue, 0.50, 0.62)
            R, G, B = int(round(r*255)), int(round(g*255)), int(round(b*255))
            bg = "#{:02x}{:02x}{:02x}".format(R, G, B)
            lum = (0.2126*R + 0.7152*G + 0.0722*B) / 255.0
            return bg, ("#ffffff" if lum < 0.55 else "#111111")

        try:
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            raw_names = []
            for ri, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
                v = row[0] if row else None
                if v is None:
                    if raw_names: break
                    else: continue
                s = str(v).strip()
                if not s:
                    if raw_names: break
                    else: continue
                if ri == 1 and s.lower() in ("brand","brands","name","brand name","brand_name"):
                    continue
                raw_names.append(s)
            try: wb.close()
            except Exception: pass

            if not raw_names:
                return jsonify({"status":"error",
                                "message":"No brand names found in column A of the first sheet."}), 400

            cleaned, skipped, seen_lower = [], [], set()
            for nm in raw_names:
                if len(nm) > 200:
                    skipped.append({"name": nm[:60], "reason": "too_long"}); continue
                low = nm.lower()
                if low in seen_lower:
                    skipped.append({"name": nm, "reason": "duplicate"}); continue
                seen_lower.add(low)
                bg, tc = _color_for(nm)
                cleaned.append((nm, bg, tc))

            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            existing = conn.execute("SELECT name FROM procurement_brands").fetchall() or []
            existing_lower = { (r["name"] or "").lower() for r in existing }

            to_insert = []
            for (nm, c, tc) in cleaned:
                if nm.lower() in existing_lower:
                    skipped.append({"name": nm, "reason": "duplicate"})
                else:
                    to_insert.append((nm, c, tc))

            added = []
            for (nm, c, tc) in to_insert:
                try:
                    conn.execute(
                        "INSERT INTO procurement_brands (name,color,text_color) VALUES (%s,%s,%s)",
                        (nm, c, tc)
                    )
                    bid = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
                    added.append({"id": bid, "name": nm, "color": c, "text_color": tc})
                except Exception:
                    skipped.append({"name": nm, "reason": "duplicate"})

            conn.commit(); conn.close()
            return jsonify({"status":"ok","added": added, "skipped": skipped})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500
        finally:
            try: os.remove(tmp_path)
            except Exception: pass

    @app.route("/api/inventory_mgmt/brands/<int:brand_id>", methods=["PUT"])
    @_edit_required
    def api_inv_brands_update(brand_id):
        d    = request.get_json() or {}
        name = (d.get("name") or "").strip()
        color= (d.get("color") or "").strip()
        text_color = (d.get("text_color") or "#ffffff").strip()
        if not name:
            return jsonify({"status":"error","message":"name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            conn.execute(
                "UPDATE procurement_brands SET name=%s,color=%s,text_color=%s WHERE id=%s",
                (name, color or "#6366f1", text_color, brand_id)
            )
            conn.commit(); conn.close()
            return jsonify({"status":"ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/inventory_mgmt/brands/<int:brand_id>", methods=["DELETE"])
    @_edit_required
    def api_inv_brands_delete_one(brand_id):
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            # Unlink any references in procurement-side formulations
            try:
                conn.execute(
                    "UPDATE procurement_formulations SET brand_id=NULL WHERE brand_id=%s",
                    (brand_id,)
                )
            except Exception:
                pass  # table may not exist on installs without procurement
            # Clear inventory-side dept mapping
            conn.execute(
                "DELETE FROM inventory_brand_dept WHERE brand_id=%s",
                (brand_id,)
            )
            conn.execute("DELETE FROM procurement_brands WHERE id=%s", (brand_id,))
            conn.commit(); conn.close()
            return jsonify({"status":"ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    # ═══ SUPPLIERS (dept-branched) ═════════════════════════════════════════════
    def _supplier_source_for_dept(dept):
        table = SUPPLIER_TABLE_BY_DEPT.get(dept, "procurement_suppliers")
        return table, SUPPLIER_SOURCE[table]

    def _normalize_procurement_supplier(r):
        return {
            "id": int(r["id"]),
            "source": "procurement",
            "supplier_name":  r.get("supplier_name") or "",
            "supplier_code":  r.get("supplier_code") or "",
            "contact_person": r.get("contact_person") or "",
            "phone":          r.get("phone") or "",
            "email":          r.get("email") or "",
            "address":        r.get("address") or "",
            "gst_number":     r.get("gst_number") or "",
            "pan_number":     r.get("pan_number") or "",
            "payment_terms":  r.get("payment_terms") or "",
            "credit_days":    int(r["credit_days"]) if r.get("credit_days") else None,
            "payment_type":   r.get("payment_type") or "",
            "currency":       r.get("currency") or "INR",
            "lead_time_days": int(r["lead_time_days"]) if r.get("lead_time_days") else None,
            "rating":         int(r["rating"]) if r.get("rating") else None,
            "status":         r.get("status") or "active",
            # NEW: supplier_type from supplier_type table (joined in the query)
            "supplier_type_id": int(r["supplier_type_id"]) if r.get("supplier_type_id") else None,
            "type_name":        (r.get("type_name") or ""),
        }

    def _normalize_purchase_supplier(r):
        return {
            "id": int(r["id"]),
            "source": "purchase",
            "supplier_name":  r.get("supplier_name") or "",
            "supplier_code":  "",
            "contact_person": r.get("contact_person") or "",
            "phone":          r.get("phone") or "",
            "email":          r.get("email") or "",
            "address":        r.get("address") or "",
            "gst_number":     r.get("gstin") or "",
            "pan_number":     "",
            "payment_terms":  r.get("payment_terms") or "",
            "credit_days":    None,
            "payment_type":   "",
            "currency":       "INR",
            "lead_time_days": None,
            "rating":         int(r["rating"]) if r.get("rating") else None,
            "status":         ("active" if int(r.get("is_active") or 1) else "inactive"),
            # purchase_suppliers has no supplier_type column — we treat
            # every record as PM-typed implicitly (since this table is PM-only).
            "supplier_type_id": None,
            "type_name":        "PM SUPPLIER",
        }

    # ── Department → supplier_type_id mapping ───────────────────────
    # Maps the inventory dept (RM/PM/FG) to a list of supplier_type.id
    # values that count as belonging to that dept. Configure based on
    # your supplier_type table:
    #     id 1 = RM SUPPLIER
    #     id 2 = PM SUPPLIER
    #     id 3 = OTHER SUPPLIER  (currently shown only with "Show all")
    # If your supplier_type table grows, add the new ids here.
    SUPPLIER_TYPE_IDS_BY_DEPT = {
        "RM": [1],
        "PM": [2],
        "FG": [],   # no FG-only supplier_type today — relies on "Show all"
    }

    @app.route("/api/inventory_mgmt/suppliers")
    def api_inv_suppliers():
        if not _can_inventory():
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        dept = _norm_dept(request.args.get("department"))
        # show_all=1 disables the supplier_type filter — returns every
        # supplier from the relevant source table. Used by the toolbar
        # toggle so admins can spot/edit suppliers with no type set or
        # with an unmapped type (e.g. OTHER SUPPLIER).
        show_all = request.args.get("show_all") in ("1", "true", "yes")
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500

        try:
            out = []

            # Whether the supplier_type table exists (defensive). If not,
            # we degrade to returning all rows (same as show_all).
            _has_supplier_type = True
            try:
                conn.execute("SELECT 1 FROM supplier_type LIMIT 1").fetchall()
            except Exception:
                _has_supplier_type = False

            if dept:
                table, source = _supplier_source_for_dept(dept)

                if table == "procurement_suppliers":
                    # Build the query — LEFT JOIN supplier_type so we can
                    # show the type_name in the response. Filter only when
                    # show_all=false AND the dept has a configured id list.
                    type_ids = SUPPLIER_TYPE_IDS_BY_DEPT.get(dept, [])
                    if _has_supplier_type:
                        base_sql = (
                            "SELECT s.*, st.type_name "
                            "FROM procurement_suppliers s "
                            "LEFT JOIN supplier_type st ON st.id = s.supplier_type_id "
                        )
                    else:
                        base_sql = "SELECT s.* FROM procurement_suppliers s "

                    if show_all or not type_ids or not _has_supplier_type:
                        rows = conn.execute(
                            base_sql + "ORDER BY s.supplier_name ASC"
                        ).fetchall()
                    else:
                        # Filter by supplier_type_id IN (...)
                        placeholders = ",".join(["%s"] * len(type_ids))
                        rows = conn.execute(
                            base_sql
                            + f"WHERE s.supplier_type_id IN ({placeholders}) "
                            + "ORDER BY s.supplier_name ASC",
                            tuple(type_ids)
                        ).fetchall()
                    out = [_normalize_procurement_supplier(dict(r)) for r in rows]

                else:
                    # purchase_suppliers (PM only) — no supplier_type column,
                    # the table itself is the PM scope.
                    rows = conn.execute(
                        "SELECT * FROM purchase_suppliers ORDER BY supplier_name ASC"
                    ).fetchall()
                    out = [_normalize_purchase_supplier(dict(r)) for r in rows]
            else:
                # No specific department — return everything from both
                # source tables (procurement + purchase). JOIN supplier_type
                # so the response carries type_name for display.
                try:
                    if _has_supplier_type:
                        rows_p = conn.execute("""
                            SELECT s.*, st.type_name
                            FROM procurement_suppliers s
                            LEFT JOIN supplier_type st ON st.id = s.supplier_type_id
                            ORDER BY s.supplier_name ASC
                        """).fetchall()
                    else:
                        rows_p = conn.execute(
                            "SELECT * FROM procurement_suppliers ORDER BY supplier_name ASC"
                        ).fetchall()
                    out += [_normalize_procurement_supplier(dict(r)) for r in rows_p]
                except Exception: pass
                try:
                    rows_u = conn.execute(
                        "SELECT * FROM purchase_suppliers ORDER BY supplier_name ASC"
                    ).fetchall()
                    out += [_normalize_purchase_supplier(dict(r)) for r in rows_u]
                except Exception: pass

            # Backward-compatible `departments` field — derived from the
            # supplier_type now. The frontend still expects this array.
            for s in out:
                if s.get("source") == "purchase":
                    s["departments"] = ["PM"]
                else:
                    tid = s.get("supplier_type_id")
                    if tid == 1:   s["departments"] = ["RM"]
                    elif tid == 2: s["departments"] = ["PM"]
                    elif tid == 3: s["departments"] = []   # OTHER — uncategorized
                    else:          s["departments"] = []

            return jsonify({"status": "ok", "suppliers": out})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    @app.route("/api/inventory_mgmt/suppliers/save", methods=["POST"])
    @_edit_required
    def api_inv_supplier_save():
        d = request.get_json(silent=True) or {}
        sup_id = d.get("id")
        source = (d.get("source") or "").strip().lower()
        name   = (d.get("supplier_name") or "").strip()
        if not name:
            return jsonify({"status": "error", "message": "supplier_name is required"}), 400
        depts  = [x for x in (d.get("departments") or []) if _norm_dept(x)]
        if not depts:
            return jsonify({"status": "error", "message": "At least one department is required"}), 400

        if source not in ("procurement", "purchase"):
            first_dept = depts[0]
            table, source = _supplier_source_for_dept(first_dept)
        else:
            table = "purchase_suppliers" if source == "purchase" else "procurement_suppliers"

        # Dept list must not mix sources
        for dp in depts:
            expected_table, expected_src = _supplier_source_for_dept(dp)
            if expected_src != source:
                return jsonify({
                    "status": "error",
                    "message": f"Department {dp} does not belong to the '{source}' supplier table. "
                               "RM/FG → procurement_suppliers, PM → purchase_suppliers. "
                               "Create two separate supplier records instead."
                }), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            if table == "procurement_suppliers":
                fields = {
                    "supplier_name":   name,
                    "supplier_code":   d.get("supplier_code") or None,
                    "contact_person":  d.get("contact_person") or None,
                    "phone":           d.get("phone") or None,
                    "email":           d.get("email") or None,
                    "address":         d.get("address") or None,
                    "gst_number":      d.get("gst_number") or None,
                    "pan_number":      d.get("pan_number") or None,
                    "payment_terms":   d.get("payment_terms") or None,
                    "payment_type":    d.get("payment_type") or None,
                    "credit_days":     d.get("credit_days") if d.get("credit_days") not in (None, "",) else None,
                    "currency":        d.get("currency") or "INR",
                    "lead_time_days":  d.get("lead_time_days") if d.get("lead_time_days") not in (None, "",) else None,
                    "rating":          d.get("rating") if d.get("rating") not in (None, "",) else None,
                    "status":          d.get("status") or "active",
                    "updated_by":      _user(),
                }
            else:
                fields = {
                    "supplier_name":   name,
                    "contact_person":  d.get("contact_person") or None,
                    "phone":           d.get("phone") or None,
                    "email":           d.get("email") or None,
                    "address":         d.get("address") or None,
                    "gstin":           d.get("gst_number") or None,
                    "payment_terms":   d.get("payment_terms") or None,
                    "rating":          d.get("rating") if d.get("rating") not in (None, "",) else None,
                    "is_active":       1 if (d.get("status") or "active") == "active" else 0,
                }
                if not sup_id:
                    fields["created_by"] = _user()

            if sup_id:
                sets = ", ".join(f"{k}=%s" for k in fields.keys())
                vals = list(fields.values()) + [int(sup_id)]
                conn.execute(f"UPDATE {table} SET {sets} WHERE id=%s", tuple(vals))
            else:
                cols  = ", ".join(fields.keys())
                marks = ", ".join(["%s"]*len(fields))
                conn.execute(f"INSERT INTO {table} ({cols}) VALUES ({marks})", tuple(fields.values()))
                row = conn.execute(
                    f"SELECT id FROM {table} WHERE supplier_name=%s ORDER BY id DESC LIMIT 1", (name,)
                ).fetchone()
                sup_id = int(row["id"]) if row else None

            if sup_id:
                conn.execute(
                    "DELETE FROM inventory_supplier_dept "
                    "WHERE supplier_id=%s AND supplier_source=%s",
                    (int(sup_id), source)
                )
                for dp in depts:
                    conn.execute(
                        "INSERT IGNORE INTO inventory_supplier_dept "
                        "(supplier_id, supplier_source, department) VALUES (%s,%s,%s)",
                        (int(sup_id), source, dp)
                    )
            conn.commit()
            return jsonify({"status": "ok", "id": sup_id, "source": source})
        except Exception as e:
            conn.rollback(); traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    @app.route("/api/inventory_mgmt/suppliers/tag_dept", methods=["POST"])
    @_edit_required
    def api_inv_supplier_tag_dept():
        """
        Bulk-tag suppliers to a department.

        Body: {
            "items": [{"source":"procurement","id":123}, ...],
            "department": "RM" | "PM" | "FG",
            "remove": false  (optional — if true, removes the tag instead)
        }
        OR
        Body: {
            "tag_all_visible": true,   # tag everything in the source table
            "source": "procurement" | "purchase",
            "department": "RM" | "PM" | "FG"
        }

        Returns: { status:'ok', tagged: N, skipped: M }
        """
        d = request.get_json(silent=True) or {}
        dept = _norm_dept(d.get("department"))
        if not dept:
            return jsonify({"status": "error",
                            "message": "department is required (RM/PM/FG)"}), 400
        remove = bool(d.get("remove"))

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500

        # Cross-check: dept ↔ supplier_source must be consistent so we don't
        # tag a procurement_supplier as PM (which lives in purchase_suppliers).
        expected_table, expected_source = _supplier_source_for_dept(dept)

        try:
            # Resolve the list of (source, id) pairs to tag
            pairs = []
            if d.get("tag_all_visible"):
                src = (d.get("source") or expected_source).strip().lower()
                if src not in ("procurement", "purchase"):
                    return jsonify({"status": "error",
                                    "message": "Invalid source"}), 400
                if src != expected_source:
                    return jsonify({
                        "status": "error",
                        "message": f"Department {dept} belongs to '{expected_source}' "
                                   f"suppliers — cannot tag '{src}' rows with it."
                    }), 400
                table = "procurement_suppliers" if src == "procurement" else "purchase_suppliers"
                rows = conn.execute(f"SELECT id FROM {table}").fetchall()
                pairs = [(src, int(r["id"])) for r in rows]
            else:
                items = d.get("items") or []
                for it in items:
                    src = (it.get("source") or "").strip().lower()
                    if src not in ("procurement", "purchase"):
                        continue
                    try:
                        sid = int(it.get("id"))
                    except Exception:
                        continue
                    if src != expected_source:
                        # Skip mismatched rows silently — we don't want one bad
                        # row to abort the whole batch.
                        continue
                    pairs.append((src, sid))

            if not pairs:
                return jsonify({"status": "ok", "tagged": 0, "skipped": 0,
                                "message": "Nothing to tag"})

            # Apply the change in one transaction
            tagged = 0
            skipped = 0
            if remove:
                for (src, sid) in pairs:
                    cur = conn.execute(
                        "DELETE FROM inventory_supplier_dept "
                        "WHERE supplier_id=%s AND supplier_source=%s AND department=%s",
                        (sid, src, dept)
                    )
                    # rowcount may be unavailable on some drivers — just count attempts
                    tagged += 1
            else:
                for (src, sid) in pairs:
                    # Skip if already tagged
                    exists = conn.execute(
                        "SELECT 1 FROM inventory_supplier_dept "
                        "WHERE supplier_id=%s AND supplier_source=%s AND department=%s LIMIT 1",
                        (sid, src, dept)
                    ).fetchone()
                    if exists:
                        skipped += 1
                        continue
                    conn.execute(
                        "INSERT IGNORE INTO inventory_supplier_dept "
                        "(supplier_id, supplier_source, department) VALUES (%s,%s,%s)",
                        (sid, src, dept)
                    )
                    tagged += 1
            conn.commit()
            return jsonify({
                "status":      "ok",
                "tagged":      tagged,
                "skipped":     skipped,
                "department":  dept,
                "operation":   "remove" if remove else "add",
            })
        except Exception as e:
            try: conn.rollback()
            except Exception: pass
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    @app.route("/api/inventory_mgmt/suppliers/delete", methods=["POST"])
    @_edit_required
    def api_inv_supplier_delete():
        d = request.get_json(silent=True) or {}
        items = d.get("items") or []
        if not items and (d.get("ids") or d.get("id")):
            ids = d.get("ids") or [d.get("id")]
            source = (d.get("source") or "procurement").strip().lower()
            items = [{"id": int(x), "source": source} for x in ids if x]
        if not items:
            return jsonify({"status": "error", "message": "items required"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            deleted = 0
            for it in items:
                sup_id = int(it.get("id") or 0)
                source = (it.get("source") or "procurement").strip().lower()
                table  = "purchase_suppliers" if source == "purchase" else "procurement_suppliers"
                if not sup_id: continue
                conn.execute(
                    "DELETE FROM inventory_supplier_dept WHERE supplier_id=%s AND supplier_source=%s",
                    (sup_id, source)
                )
                conn.execute(f"DELETE FROM {table} WHERE id=%s", (sup_id,))
                deleted += 1
            conn.commit()
            return jsonify({"status": "ok", "deleted": deleted})
        except Exception as e:
            conn.rollback(); traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ═══ LOOKUPS ═══════════════════════════════════════════════════════════════
    @app.route("/api/inventory_mgmt/lookups")
    def api_inv_lookups():
        if not _can_inventory():
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            pm_types = []
            try:
                rows = conn.execute(
                    "SELECT DISTINCT pm_type FROM pm_products WHERE pm_type IS NOT NULL AND pm_type<>'' ORDER BY pm_type"
                ).fetchall()
                pm_types = [r["pm_type"] for r in rows]
            except Exception: pass

            pm_material_types = []
            try:
                rows = conn.execute(
                    "SELECT DISTINCT material_type FROM pm_products WHERE material_type IS NOT NULL AND material_type<>'' ORDER BY material_type"
                ).fetchall()
                pm_material_types = [r["material_type"] for r in rows]
            except Exception: pass

            groups = []
            try:
                rows = conn.execute(
                    "SELECT id, group_name, parent_id FROM procurement_material_groups ORDER BY group_name"
                ).fetchall()
                groups = [{"id": int(r["id"]), "group_name": r["group_name"], "parent_id": r["parent_id"]} for r in rows]
            except Exception: pass

            mtypes = []
            try:
                rows = conn.execute(
                    "SELECT id, type_name, abbreviation, color FROM procurement_material_types ORDER BY sort_order, type_name"
                ).fetchall()
                mtypes = [{"id": int(r["id"]), "type_name": r["type_name"], "abbreviation": r["abbreviation"], "color": r["color"]} for r in rows]
            except Exception: pass

            uoms = ["KG","G","MG","L","ML","NOS","PCS","BOX","PKT","ROLL","BOTTLE","JAR","DRUM"]

            return jsonify({
                "status": "ok",
                "pm_types":           pm_types,
                "pm_material_types":  pm_material_types,
                "material_groups":    groups,
                "material_types":     mtypes,
                "uoms":               uoms,
                "gst_rates":          [0, 0.1, 0.25, 1, 1.5, 3, 5, 12, 18, 28],
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ═══ LAST PURCHASE REFRESH ═════════════════════════════════════════════════
    @app.route("/api/inventory_mgmt/last_purchase/refresh", methods=["POST"])
    @_edit_required
    def api_inv_last_refresh():
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            try:
                conn.execute("""
                    INSERT INTO inventory_last_purchase
                        (department, item_id, item_name, last_supplier, last_rate, gst_rate, last_date)
                    SELECT 'RM', m.id, m.material_name,
                           g.supplier_name, gi.rate, gi.gst_rate, g.grn_date
                    FROM procurement_materials m
                    JOIN (
                        SELECT gi.material AS mat, MAX(g.grn_date) AS latest
                        FROM procurement_grn_items gi
                        JOIN procurement_grn g ON g.id = gi.grn_id
                        GROUP BY gi.material
                    ) lx ON lx.mat = m.material_name
                    JOIN procurement_grn g ON g.grn_date = lx.latest
                    JOIN procurement_grn_items gi ON gi.grn_id = g.id AND gi.material = m.material_name
                    ON DUPLICATE KEY UPDATE
                        last_supplier = VALUES(last_supplier),
                        last_rate     = VALUES(last_rate),
                        gst_rate      = VALUES(gst_rate),
                        last_date     = VALUES(last_date),
                        item_name     = VALUES(item_name)
                """)
                conn.commit()
            except Exception:
                traceback.print_exc()
            return jsonify({"status": "ok", "message": "Last purchase info refreshed"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ═══ SHARE CONTACTS ═══════════════════════════════════════════════════════
    @app.route("/api/inventory_mgmt/share/contacts")
    def api_inv_share_contacts():
        if not _can_inventory():
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        raw = (request.args.get("suppliers") or "").strip()
        if not raw:
            return jsonify({"status": "ok", "contacts": []})
        names = [n.strip() for n in raw.split("|") if n.strip()]
        if not names:
            return jsonify({"status": "ok", "contacts": []})

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            placeholders = ",".join(["%s"]*len(names))
            out = {}
            try:
                rows = conn.execute(
                    f"SELECT supplier_name, phone, email, contact_person "
                    f"FROM procurement_suppliers WHERE supplier_name IN ({placeholders})",
                    tuple(names)
                ).fetchall()
                for r in rows:
                    out[r["supplier_name"]] = {
                        "supplier_name": r["supplier_name"] or "",
                        "phone": (r["phone"] or "").strip(),
                        "email": (r["email"] or "").strip(),
                        "contact_person": r["contact_person"] or "",
                    }
            except Exception: pass
            try:
                rows = conn.execute(
                    f"SELECT supplier_name, phone, email, contact_person "
                    f"FROM purchase_suppliers WHERE supplier_name IN ({placeholders})",
                    tuple(names)
                ).fetchall()
                for r in rows:
                    out.setdefault(r["supplier_name"], {
                        "supplier_name": r["supplier_name"] or "",
                        "phone": (r["phone"] or "").strip(),
                        "email": (r["email"] or "").strip(),
                        "contact_person": r["contact_person"] or "",
                    })
            except Exception: pass
            return jsonify({"status": "ok", "contacts": list(out.values())})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    # ════════════════════════════════════════════════════════════════════════════
    # GRN MODULE  —  reads/writes the SAME `procurement_grn` and
    # `procurement_grn_items` tables used by procurement.py. Endpoints here are
    # gated by the inventory permissions (_login_required / _edit_required) so
    # users with inventory access can receive goods without needing the full
    # procurement role.
    # ════════════════════════════════════════════════════════════════════════════

    @app.route("/api/inventory_mgmt/grn/list", methods=["GET"])
    @_login_required
    def api_inv_grn_list():
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            # Aggregate distinct invoice numbers + dates from line items
            # so the list can show line-level invoices when no header invoice was set.
            #
            # NOTE: GROUP_CONCAT(DISTINCT ... ORDER BY ...) is rejected by MySQL
            # when the ORDER BY column isn't part of the DISTINCT set. We drop
            # ORDER BY and accept GROUP_CONCAT's natural order (which is fine
            # for display — typically there's only 1–2 distinct values anyway).
            #
            # We use CAST(... AS CHAR) for dates instead of DATE_FORMAT so the
            # SQL has zero percent signs — bypasses any %%/percent-escaping
            # quirks in the DB connection wrapper.
            # Aggregate the line-item MATERIAL NAMES and (if present) MANUFACTURERS
            # per GRN so the list search can match items inside a GRN and by
            # manufacturer. '|' separator (same reasoning as invoices below).
            _mfr_select = (
                "(SELECT GROUP_CONCAT(DISTINCT NULLIF(i.manufacturer,'') SEPARATOR '|') "
                " FROM procurement_grn_items i WHERE i.grn_id=g.id) AS line_manufacturers,"
                if _items_has_mfr(conn) else
                "'' AS line_manufacturers,"
            )
            rows = conn.execute("""
                SELECT g.*,
                    (SELECT COUNT(*) FROM procurement_grn_items i WHERE i.grn_id=g.id) AS item_count,
                    (SELECT COUNT(DISTINCT t.grn_item_id)
                       FROM procurement_grn_trs t
                       WHERE t.grn_id=g.id) AS trs_count,
                    -- TRS completion is judged per (material, batch) combo, NOT
                    -- per line: when a GRN has multiple lines with the SAME
                    -- material AND batch they share ONE TRS, so they count as a
                    -- single unit. combo_total = distinct material+batch combos
                    -- among line items; combo_done = those that have a TRS.
                    (SELECT COUNT(*) FROM (
                        SELECT 1 FROM procurement_grn_items i
                        WHERE i.grn_id=g.id
                        GROUP BY LOWER(TRIM(COALESCE(i.material,''))),
                                 TRIM(COALESCE(i.batch_num,''))
                     ) AS _combos) AS combo_total,
                    (SELECT COUNT(*) FROM (
                        SELECT 1 FROM procurement_grn_items i
                        WHERE i.grn_id=g.id
                          AND EXISTS (
                            SELECT 1 FROM procurement_grn_trs t
                            WHERE t.grn_id=g.id
                              AND LOWER(TRIM(COALESCE(t.material,''))) COLLATE utf8mb4_unicode_ci
                                  = LOWER(TRIM(COALESCE(i.material,''))) COLLATE utf8mb4_unicode_ci
                              AND TRIM(COALESCE(t.batch_num,'')) COLLATE utf8mb4_unicode_ci
                                  = TRIM(COALESCE(i.batch_num,'')) COLLATE utf8mb4_unicode_ci
                          )
                        GROUP BY LOWER(TRIM(COALESCE(i.material,''))),
                                 TRIM(COALESCE(i.batch_num,''))
                     ) AS _combos_done) AS combo_done,
                    (SELECT GROUP_CONCAT(DISTINCT NULLIF(i.material,'') SEPARATOR '|')
                       FROM procurement_grn_items i WHERE i.grn_id=g.id) AS line_materials,
                    """ + _mfr_select + """
                    (SELECT GROUP_CONCAT(DISTINCT NULLIF(i.invoice_num,'') SEPARATOR '|')
                       FROM procurement_grn_items i WHERE i.grn_id=g.id) AS line_invoice_nums,
                    (SELECT GROUP_CONCAT(DISTINCT NULLIF(CAST(i.invoice_date AS CHAR),'') SEPARATOR '|')
                       FROM procurement_grn_items i WHERE i.grn_id=g.id) AS line_invoice_dates,
                    (SELECT COUNT(*) FROM procurement_grn_files f
                       WHERE f.grn_id=g.id AND f.kind='invoice') AS invoice_file_count,
                    (SELECT COUNT(*) FROM procurement_grn_files f
                       WHERE f.grn_id=g.id AND f.kind='coa') AS coa_file_count
                FROM procurement_grn g
                WHERE (g.grn_type IS NULL OR g.grn_type = 'GRN' OR g.grn_type = '')
                ORDER BY g.created_at DESC
            """).fetchall()
            result = []
            for r in rows:
                d = dict(r)
                for f in ("grn_date", "invoice_date", "created_at", "updated_at"):
                    if d.get(f):
                        d[f] = str(d[f])
                if d.get("grand_total") is not None:
                    d["grand_total"] = float(d["grand_total"])
                if d.get("po_invoices"):
                    try:
                        d["po_invoices"] = json.loads(d["po_invoices"])
                    except Exception:
                        d["po_invoices"] = []
                else:
                    d["po_invoices"] = []
                # Parse the pipe-delimited line-item invoice aggregates into clean lists.
                # We use '|' as separator because invoice numbers commonly contain
                # commas (e.g. "INV/2026/12,304") which would be mis-split otherwise.
                line_inv_nums = (d.pop("line_invoice_nums", "") or "")
                line_inv_dates = (d.pop("line_invoice_dates", "") or "")
                # Line-item material names + manufacturers (for list search).
                line_mats = (d.pop("line_materials", "") or "")
                line_mfrs = (d.pop("line_manufacturers", "") or "")
                d["line_materials"] = [x for x in line_mats.split("|") if x]
                d["line_manufacturers"] = [x for x in line_mfrs.split("|") if x]
                d["line_invoices"] = []
                if line_inv_nums:
                    nums  = [x for x in line_inv_nums.split("|") if x]
                    dates = [x for x in line_inv_dates.split("|") if x]
                    # Pair them where possible. If counts differ, just use the nums
                    # with whatever dates were present (UI handles missing dates).
                    for i, num in enumerate(nums):
                        d["line_invoices"].append({
                            "invoice_num":  num,
                            "invoice_date": dates[i] if i < len(dates) else "",
                        })
                # ── TRS completion status for this GRN ──
                # Compare distinct grn_item_ids with TRS rows against the
                # GRN's total line-item count. Three states drive the GRN
                # list row tint, judged per (material, batch) combo so that
                # multiple identical lines sharing one TRS count as done:
                #   'all'     — every combo has a TRS
                #   'partial' — at least one combo has a TRS but not all
                #   'none'    — no TRS exists (or no line items yet)
                _combos = int(d.get("combo_total") or 0)
                _done   = int(d.get("combo_done")  or 0)
                if _combos > 0 and _done >= _combos:
                    d["trs_status"] = "all"
                elif _done > 0:
                    d["trs_status"] = "partial"
                else:
                    d["trs_status"] = "none"
                result.append(d)
            return jsonify({"status": "ok", "grns": result})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    @app.route("/api/inventory_mgmt/grn/get", methods=["GET"])
    @_login_required
    def api_inv_grn_get():
        grn_id = request.args.get("id")
        if not grn_id:
            return jsonify({"status": "error", "message": "id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            g = conn.execute(
                "SELECT * FROM procurement_grn WHERE id=%s", (grn_id,)
            ).fetchone()
            if not g:
                return jsonify({"status": "error", "message": "GRN not found"}), 404
            grn = dict(g)
            for f in ("grn_date", "invoice_date", "created_at", "updated_at"):
                if grn.get(f):
                    grn[f] = str(grn[f])
            if grn.get("po_invoices"):
                try:
                    grn["po_invoices"] = json.loads(grn["po_invoices"])
                except Exception:
                    grn["po_invoices"] = []
            else:
                grn["po_invoices"] = []

            # NEW (May 2026): parse the JSON columns added by the godown
            # bootstrap migration. If the columns don't exist on this install
            # the SELECT * above won't include them, so .get returns None.
            if grn.get("other_details"):
                try:
                    if isinstance(grn["other_details"], str):
                        grn["other_details"] = json.loads(grn["other_details"])
                except Exception:
                    grn["other_details"] = {}
            else:
                grn["other_details"] = {}
            if grn.get("unload_checklist"):
                try:
                    if isinstance(grn["unload_checklist"], str):
                        grn["unload_checklist"] = json.loads(grn["unload_checklist"])
                except Exception:
                    grn["unload_checklist"] = {}
            else:
                grn["unload_checklist"] = {}
            # supervisor_name comes through SELECT * automatically; ensure it's a string
            grn["supervisor_name"] = grn.get("supervisor_name") or ""

            items = conn.execute(
                "SELECT * FROM procurement_grn_items WHERE grn_id=%s ORDER BY id",
                (grn_id,)
            ).fetchall()

            # Build per-item PO reference by matching material to PO items
            po_ids_to_check = set()
            if grn.get("po_id"):
                po_ids_to_check.add(int(grn["po_id"]))
            for pi in grn.get("po_invoices", []):
                if pi.get("po_id"):
                    po_ids_to_check.add(int(pi["po_id"]))
            mat_po_map = {}
            for pid in po_ids_to_check:
                try:
                    po_info = conn.execute(
                        "SELECT po_num FROM procurement_purchase_orders WHERE id=%s",
                        (pid,)
                    ).fetchone()
                    po_items_list = conn.execute(
                        "SELECT material FROM procurement_po_items WHERE po_id=%s",
                        (pid,)
                    ).fetchall()
                    if po_info:
                        for pi_item in po_items_list:
                            mat_key = (pi_item["material"] or "").strip().lower()
                            if mat_key:
                                mat_po_map[mat_key] = {
                                    "po_id":  pid,
                                    "po_num": po_info["po_num"]
                                }
                except Exception:
                    pass

            grn["items"] = []
            for i in items:
                d2 = dict(i)
                for f in ("received_qty", "po_qty", "rate", "amount", "gst_rate"):
                    if d2.get(f) is not None:
                        d2[f] = float(d2[f])
                for f in ("invoice_date", "mfg_date", "expiry_date"):
                    if d2.get(f):
                        d2[f] = str(d2[f])
                # Ensure manufacturer is always a string (handles None and
                # also handles installs where the column doesn't exist yet).
                d2["manufacturer"] = (d2.get("manufacturer") or "")
                mat_key = (d2.get("material") or "").strip().lower()
                po_ref = mat_po_map.get(mat_key, {})
                d2["po_id"]  = po_ref.get("po_id")
                d2["po_num"] = po_ref.get("po_num", "")
                grn["items"].append(d2)

            # Attach file metadata (no blobs — those stream via the download endpoint).
            # COA files are paired to line items; invoice files belong to the GRN.
            grn["coa_by_item"] = {}    # { grn_item_id: [{id, original_name, mime_type, size_bytes}, ...] }
            grn["invoices"]    = []    # [ {id, original_name, mime_type, size_bytes, uploaded_at}, ... ]
            try:
                file_rows = conn.execute("""
                    SELECT id, grn_item_id, kind, original_name, mime_type, size_bytes, uploaded_at
                    FROM procurement_grn_files
                    WHERE grn_id=%s
                    ORDER BY id
                """, (grn_id,)).fetchall()
                for fr in file_rows:
                    fd = dict(fr)
                    if fd.get("uploaded_at"):
                        fd["uploaded_at"] = str(fd["uploaded_at"])
                    if fd["kind"] == "coa" and fd.get("grn_item_id"):
                        key = int(fd["grn_item_id"])
                        grn["coa_by_item"].setdefault(key, []).append(fd)
                    elif fd["kind"] == "invoice":
                        grn["invoices"].append(fd)
            except Exception as _e:
                # Table may not exist yet on a fresh install; silently skip.
                print(f"[InventoryMgmt] file list skipped: {_e}")

            return jsonify({"status": "ok", "grn": grn})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    # ════════════════════════════════════════════════════════════════════════════
    # GRN FILE ATTACHMENTS  —  COA (per line) + Invoice (per GRN, multiple)
    # ════════════════════════════════════════════════════════════════════════════
    #
    # Files are stored on the local filesystem under _GRN_UPLOAD_ROOT (default:
    # <flask-root>/uploads/inventory_mgmt/grn/<grn_id>/<kind>/<uuid>__<name>).
    # Only metadata + relative_path live in procurement_grn_files.
    # The same endpoint set handles both kinds; the `kind` form field discriminates.
    #
    # Endpoints:
    #   GET    /api/inventory_mgmt/grn/file/diag        — config + on-disk health check
    #   POST   /api/inventory_mgmt/grn/file/upload      — upload one file
    #   GET    /api/inventory_mgmt/grn/file/list?grn_id=…   — list files for a GRN
    #   GET    /api/inventory_mgmt/grn/file/<id>        — download / inline-view (streams from disk)
    #   DELETE /api/inventory_mgmt/grn/file/<id>        — remove (DB row + disk file)

    def _grn_file_kind_ok(kind):
        return kind in ("coa", "invoice")

    def _grn_file_mime_ok(mime):
        return (mime or "").lower() in _GRN_FILE_ALLOWED_MIME

    @app.route("/api/inventory_mgmt/grn/file/diag", methods=["GET"])
    @_login_required
    def api_inv_grn_file_diag():
        """
        Diagnostic for the file-upload subsystem.
        Reports the upload root, whether it's writable, disk-vs-DB consistency,
        and a list of recent files (with on-disk status).
        """
        out = {"status": "ok"}
        # Filesystem checks
        out["upload_root"] = _GRN_UPLOAD_ROOT
        out["upload_root_exists"]   = os.path.isdir(_GRN_UPLOAD_ROOT)
        out["upload_root_writable"] = os.access(_GRN_UPLOAD_ROOT, os.W_OK) if out["upload_root_exists"] else False
        try:
            # Total bytes on disk under the upload root.
            total = 0
            count = 0
            for root, _dirs, files in os.walk(_GRN_UPLOAD_ROOT):
                for fn in files:
                    p = os.path.join(root, fn)
                    try:
                        total += os.path.getsize(p)
                        count += 1
                    except Exception:
                        pass
            out["disk_files"]     = count
            out["disk_bytes"]     = total
        except Exception as e:
            out["disk_walk_error"] = str(e)

        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            try:
                vrow = conn.execute("SELECT VERSION() AS v").fetchone()
                out["mysql_version"] = vrow["v"] if vrow else None
            except Exception:
                pass
            # Per-file disk consistency check
            try:
                rows = conn.execute("""
                    SELECT id, grn_id, grn_item_id, kind, original_name,
                           mime_type, size_bytes, relative_path, uploaded_at
                    FROM procurement_grn_files
                    ORDER BY id DESC
                    LIMIT 50
                """).fetchall()
                missing = 0
                size_mismatch = 0
                recent = []
                for r in rows:
                    d = dict(r)
                    if d.get("uploaded_at"):
                        d["uploaded_at"] = str(d["uploaded_at"])
                    rp = d.get("relative_path")
                    if not rp:
                        d["disk_status"] = "no-path"
                        missing += 1
                    else:
                        abs_path = _grn_file_disk_path(rp)
                        if not abs_path or not os.path.isfile(abs_path):
                            d["disk_status"] = "missing-on-disk"
                            missing += 1
                        else:
                            actual = os.path.getsize(abs_path)
                            d["actual_bytes"] = actual
                            if actual != (d.get("size_bytes") or 0):
                                d["disk_status"] = "size-mismatch"
                                size_mismatch += 1
                            else:
                                d["disk_status"] = "ok"
                    recent.append(d)
                out["recent_files"]   = recent
                out["files_missing"]  = missing
                out["size_mismatch"]  = size_mismatch
            except Exception as e:
                out["recent_files_error"] = str(e)
            try: conn.close()
            except Exception: pass
            return jsonify(out)
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/inventory_mgmt/grn/file/upload", methods=["POST"])
    @_grn_attach_required
    def api_inv_grn_file_upload():
        kind        = (request.form.get("kind") or "").strip().lower()
        grn_id      = request.form.get("grn_id")
        grn_item_id = request.form.get("grn_item_id") or None

        if not _grn_file_kind_ok(kind):
            return jsonify({"status": "error", "message": "kind must be 'coa' or 'invoice'"}), 400
        if not grn_id:
            return jsonify({"status": "error", "message": "grn_id is required"}), 400
        if kind == "coa" and not grn_item_id:
            return jsonify({"status": "error", "message": "COA upload requires grn_item_id"}), 400
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "No file part in request"}), 400

        f = request.files["file"]
        if not f or not f.filename:
            return jsonify({"status": "error", "message": "Empty filename"}), 400

        mime = (f.mimetype or "").lower()
        if not _grn_file_mime_ok(mime):
            return jsonify({"status": "error",
                            "message": f"Unsupported file type: {mime or 'unknown'}. Allowed: PDF, JPG, PNG"}), 400

        # Stream-read to a temp location first so we can validate size before
        # committing to the final path. werkzeug already streams large uploads
        # to a SpooledTemporaryFile under the hood; .read() is safe here.
        blob = f.read()
        size = len(blob)
        if size == 0:
            return jsonify({"status": "error", "message": "File is empty"}), 400
        if size > _GRN_FILE_MAX_BYTES:
            return jsonify({"status": "error",
                            "message": f"File too large ({size//1024} KB). Max {_GRN_FILE_MAX_BYTES//(1024*1024)} MB"}), 400

        abs_disk_path = None  # for rollback on DB error
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500

            # Verify the GRN exists.
            grn_row = conn.execute(
                "SELECT id FROM procurement_grn WHERE id=%s", (grn_id,)
            ).fetchone()
            if not grn_row:
                return jsonify({"status": "error", "message": "GRN not found"}), 404

            # For COA: verify the line item belongs to this GRN.
            if grn_item_id:
                ok = conn.execute(
                    "SELECT id FROM procurement_grn_items WHERE id=%s AND grn_id=%s",
                    (grn_item_id, grn_id)
                ).fetchone()
                if not ok:
                    return jsonify({"status": "error",
                                    "message": "grn_item_id does not belong to this GRN"}), 400

            # Build a safe, collision-free filename:
            #   <uuid>__<sanitized-original-name>
            # Keeps the original name readable for anyone browsing the folder.
            safe_name = _safe_filename_segment(f.filename)
            disk_name = f"{uuid.uuid4().hex}__{safe_name}"

            # Relative path stored in DB (forward slashes, portable).
            rel_dir  = f"grn/{int(grn_id)}/{kind}"
            rel_path = f"{rel_dir}/{disk_name}"

            # Resolve to absolute, ensure directory exists, write the bytes.
            abs_dir = os.path.normpath(os.path.join(_GRN_UPLOAD_ROOT, *rel_dir.split("/")))
            os.makedirs(abs_dir, exist_ok=True)
            abs_disk_path = os.path.join(abs_dir, disk_name)

            # Atomic write: write to a temp file in the same dir, then rename.
            tmp_path = abs_disk_path + ".tmp"
            with open(tmp_path, "wb") as fh:
                fh.write(blob)
                fh.flush()
                try: os.fsync(fh.fileno())
                except Exception: pass
            os.replace(tmp_path, abs_disk_path)

            # Insert the metadata row.
            user = session.get("User_Name") or session.get("user_name")
            conn.execute("""
                INSERT INTO procurement_grn_files
                    (grn_id, grn_item_id, kind, original_name, mime_type, size_bytes,
                     relative_path, uploaded_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (grn_id, grn_item_id, kind, f.filename, mime, size, rel_path, user))
            file_id = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
            conn.commit()

            return jsonify({"status": "ok", "file": {
                "id":            file_id,
                "grn_id":        int(grn_id),
                "grn_item_id":   int(grn_item_id) if grn_item_id else None,
                "kind":          kind,
                "original_name": f.filename,
                "mime_type":     mime,
                "size_bytes":    size,
            }})
        except Exception as e:
            traceback.print_exc()
            # Roll back: if we wrote to disk but DB failed, delete the file.
            if abs_disk_path and os.path.exists(abs_disk_path):
                try: os.remove(abs_disk_path)
                except Exception: pass
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    @app.route("/api/inventory_mgmt/grn/file/list", methods=["GET"])
    @_login_required
    def api_inv_grn_file_list():
        grn_id = request.args.get("grn_id")
        if not grn_id:
            return jsonify({"status": "error", "message": "grn_id is required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            rows = conn.execute("""
                SELECT id, grn_id, grn_item_id, kind, original_name, mime_type,
                       size_bytes, uploaded_by, uploaded_at
                FROM procurement_grn_files
                WHERE grn_id=%s
                ORDER BY id
            """, (grn_id,)).fetchall()
            files = []
            for r in rows:
                d = dict(r)
                if d.get("uploaded_at"):
                    d["uploaded_at"] = str(d["uploaded_at"])
                files.append(d)
            return jsonify({"status": "ok", "files": files})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    @app.route("/api/inventory_mgmt/grn/file/<int:file_id>", methods=["GET"])
    @_login_required
    def api_inv_grn_file_download(file_id):
        # ?download=1 forces attachment disposition; default is inline view.
        force_download = request.args.get("download") in ("1", "true", "yes")
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            row = conn.execute("""
                SELECT original_name, mime_type, relative_path
                FROM procurement_grn_files
                WHERE id=%s
            """, (file_id,)).fetchone()
            if not row:
                return jsonify({"status": "error", "message": "File not found"}), 404
            d = dict(row)
            if not d.get("relative_path"):
                return jsonify({"status": "error",
                                "message": "Legacy file (no disk path) — please re-upload"}), 410

            abs_path = _grn_file_disk_path(d["relative_path"])
            if not abs_path or not os.path.isfile(abs_path):
                return jsonify({"status": "error",
                                "message": "File missing on disk — please re-upload"}), 404

            # send_file streams the file efficiently and sets correct headers.
            return send_file(
                abs_path,
                mimetype=d["mime_type"] or "application/octet-stream",
                as_attachment=force_download,
                download_name=d["original_name"],
                conditional=True,        # honour Range / If-Modified-Since
                max_age=300,
            )
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    @app.route("/api/inventory_mgmt/grn/file/<int:file_id>", methods=["DELETE"])
    @_grn_attach_required
    def api_inv_grn_file_delete(file_id):
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            row = conn.execute(
                "SELECT id, relative_path FROM procurement_grn_files WHERE id=%s",
                (file_id,)
            ).fetchone()
            if not row:
                return jsonify({"status": "error", "message": "File not found"}), 404
            d = dict(row)
            # Delete the DB row first — if disk delete fails after, we just
            # orphan a file on disk (cosmetic). The other way round risks
            # a phantom row pointing at nothing.
            conn.execute("DELETE FROM procurement_grn_files WHERE id=%s", (file_id,))
            conn.commit()
            # Best-effort disk cleanup.
            try:
                abs_path = _grn_file_disk_path(d.get("relative_path"))
                if abs_path and os.path.isfile(abs_path):
                    os.remove(abs_path)
            except Exception:
                pass
            return jsonify({"status": "ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    def _grn_create_boxes_for_line(conn, grn_id, grn_num, grn_item_id, material_id,
                                   location_name, packages, qty_per_pkg, received_qty,
                                   uom, user):
        """Create rm_boxes for one GRN line so received stock appears in the
        godown view / box system. Resolves the target godown from the line's
        location name; falls back to the default godown. Returns count created.
        Safe: on any problem it just creates nothing (GRN save still succeeds)."""
        try:
            # Lazy import to avoid circular import at module load.
            try:
                from inventory.inventory_godown import allocate_next_box_code
            except Exception:
                from inventory_godown import allocate_next_box_code
        except Exception:
            return 0
        if not material_id:
            return 0
        # Resolve godown id from the line's location name.
        gid = None
        if location_name:
            gr = conn.execute(
                "SELECT id FROM procurement_godowns WHERE LOWER(name)=LOWER(%s) LIMIT 1",
                (location_name.strip(),),
            ).fetchone()
            if gr:
                gid = int(gr["id"] if hasattr(gr, "get") else gr[0])
        if gid is None:
            dr = conn.execute(
                "SELECT id FROM procurement_godowns ORDER BY is_default DESC, id ASC LIMIT 1"
            ).fetchone()
            if dr:
                gid = int(dr["id"] if hasattr(dr, "get") else dr[0])
        if gid is None:
            return 0

        # Decide box count + per-box qty.
        nboxes = int(packages or 0)
        per = float(qty_per_pkg or 0)
        rq = float(received_qty or 0)
        if nboxes <= 0:
            # No package count → make a single box holding the whole received qty.
            nboxes = 1 if rq > 0 else 0
            per = rq
        elif per <= 0:
            # Packages given but no per-box qty → split received evenly.
            per = round(rq / nboxes, 3) if nboxes else 0
        if nboxes <= 0 or per <= 0:
            return 0
        if nboxes > 1000:
            nboxes = 1000  # safety cap

        # Reuse the codes already reserved for this line's labels, so the
        # printed label and the stored box always carry the SAME code.
        # Label printing reserves codes in rm_grn_box_codes keyed by
        # (grn_id, grn_item_id, box_seq); minting fresh codes here would
        # diverge from the labels and create un-scannable "ghost" boxes
        # (the bug behind GRN/RM/0008/26-27). We look up the reserved code
        # per box_seq first; only allocate fresh if none was reserved.
        reserved = {}
        try:
            rrows = conn.execute(
                "SELECT box_seq, box_code FROM rm_grn_box_codes "
                "WHERE grn_id=%s AND grn_item_id=%s",
                (grn_id, grn_item_id),
            ).fetchall()
            for rr in rrows:
                bs = rr["box_seq"] if hasattr(rr, "get") else rr[0]
                bc = rr["box_code"] if hasattr(rr, "get") else rr[1]
                if bs is not None and bc:
                    reserved[int(bs)] = bc
        except Exception:
            # rm_grn_box_codes may not exist on very old installs — fall back
            # to fresh allocation (legacy behaviour).
            reserved = {}

        created = 0
        for seq in range(1, nboxes + 1):
            _from_reservation = seq in reserved
            box_code = reserved.get(seq) or allocate_next_box_code(conn)
            # Guard: never reuse a code that's already a live box (e.g. a
            # reserved code that was somehow consumed elsewhere). If taken,
            # fall back to a fresh code rather than violating the unique key.
            try:
                dup = conn.execute(
                    "SELECT box_id FROM rm_boxes WHERE box_code=%s LIMIT 1",
                    (box_code,),
                ).fetchone()
                if dup:
                    box_code = allocate_next_box_code(conn)
                    _from_reservation = False
            except Exception:
                pass
            # If this code was freshly allocated (not from a reservation),
            # persist it into rm_grn_box_codes so the reservation table stays
            # the single source of truth — a later label reprint for this
            # (grn_id, grn_item_id, box_seq) returns the SAME code. Covers the
            # case where box count is increased on edit beyond what was
            # previously reserved.
            if not _from_reservation:
                try:
                    conn.execute(
                        "INSERT INTO rm_grn_box_codes "
                        "(grn_id, grn_item_id, box_seq, box_code, allocated_by) "
                        "VALUES (%s,%s,%s,%s,%s) "
                        "ON DUPLICATE KEY UPDATE box_code=VALUES(box_code)",
                        (grn_id, grn_item_id, seq, box_code, user),
                    )
                except Exception:
                    # table absent on old installs, or duplicate code key —
                    # safe to skip; the box itself is still created correctly.
                    pass
            cur = conn.execute(
                "INSERT INTO rm_boxes "
                "(box_code, grn_id, grn_no, grn_item_id, material_id, "
                " box_seq, total_boxes, per_box_qty, uom, "
                " current_godown_id, current_status, source, created_by) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'in_stock','grn',%s)",
                (box_code, grn_id, grn_num, grn_item_id, material_id,
                 seq, nboxes, per, uom, gid, user),
            )
            box_id = cur.lastrowid
            conn.execute(
                "INSERT INTO rm_box_movements "
                "(box_id, movement_type, from_godown_id, to_godown_id, qty, moved_by, remarks) "
                "VALUES (%s,'grn_create',NULL,%s,%s,%s,%s)",
                (box_id, gid, per, user, "Created from GRN " + (grn_num or "")),
            )
            created += 1
        return created

    @app.route("/api/inventory_mgmt/grn/save", methods=["POST"])
    @_grn_save_required
    def api_inv_grn_save():
        import re as _re
        d        = request.get_json() or {}
        grn_id   = d.get("id")
        supplier = (d.get("supplier_name") or "").strip()
        grn_date = d.get("grn_date") or None
        invoice_num  = (d.get("invoice_num") or "").strip() or None
        invoice_date = d.get("invoice_date") or None
        po_invoices_raw  = d.get("po_invoices") or []
        po_invoices_json = json.dumps(po_invoices_raw) if po_invoices_raw else None
        if po_invoices_raw and len(po_invoices_raw) == 1:
            invoice_num  = (po_invoices_raw[0].get("invoice_num") or "").strip() or invoice_num
            invoice_date = po_invoices_raw[0].get("invoice_date") or invoice_date
        po_id   = d.get("po_id") or None
        po_num  = (d.get("po_num") or "").strip() or None
        status  = (d.get("status")  or "open").strip()
        remarks = (d.get("remarks") or "").strip() or None
        items   = d.get("items") or []
        uid     = session.get("UID", "")
        try:
            freight = float(d.get("freight_charge") or 0) or None
        except Exception:
            freight = None
        try:
            packing = float(d.get("packing_charge") or 0) or None
        except Exception:
            packing = None
        try:
            other_charge = float(d.get("other_charge") or 0) or None
        except Exception:
            other_charge = None
        other_charge_label = (d.get("other_charge_label") or "").strip()[:60] or None
        supervisor_name = (d.get("supervisor_name") or "").strip()[:120] or None
        # other_details / unload_checklist are dicts; serialize to JSON
        import json as _json
        try:
            _od = d.get("other_details") or {}
            other_details_json = _json.dumps(_od) if _od else None
        except Exception:
            other_details_json = None
        try:
            _cl = d.get("unload_checklist") or {}
            unload_checklist_json = _json.dumps(_cl) if _cl else None
        except Exception:
            unload_checklist_json = None
        if not supplier:
            return jsonify({"status": "error", "message": "supplier_name required"}), 400

        unload_location   = (d.get("unload_location") or "").strip() or None
        voucher_type_name = (d.get("voucher_type_name") or "").strip() or None

        # Compute totals
        # Taxable = sum(received_qty × rate) for line items, plus all charges.
        # GST = each item's own gst_rate on its amount + 18% on the charges
        # (freight + packing + other). Stored grand_total = taxable + GST so it
        # matches the GRN form footer and printed voucher.
        line_taxable = 0.0
        gst_total = 0.0
        for item in items:
            try:
                rqty = float(item.get("received_qty") or 0)
                rate = float(item.get("rate") or 0)
                amt = rqty * rate
                line_taxable += amt
                try:
                    gpct = float(item.get("gst_rate") or 0)
                except Exception:
                    gpct = 0.0
                if gpct > 0 and amt > 0:
                    gst_total += round(amt * gpct / 100.0, 2)
            except Exception:
                pass
        charges_total = (freight or 0) + (packing or 0) + (other_charge or 0)
        if charges_total > 0:
            gst_total += round(charges_total * 18.0 / 100.0, 2)   # charges taxed at 18%
        taxable = line_taxable + charges_total
        grand_total = round(taxable + gst_total, 2)

        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500

            grn_num = None  # populated below for new GRNs

            # Check whether procurement_grn has the new columns (auto-added
            # by inventory_godown bootstrap). Build the SET / column-list
            # dynamically so older installs without these columns keep
            # working.
            _extra_cols = {}   # column_name → value
            try:
                _cols = conn.execute("SHOW COLUMNS FROM procurement_grn").fetchall()
                _col_names = {(c['Field'] if isinstance(c, dict) else c[0]).lower() for c in _cols}
                if 'other_charge'        in _col_names: _extra_cols['other_charge']        = other_charge
                if 'other_charge_label'  in _col_names: _extra_cols['other_charge_label']  = other_charge_label
                if 'supervisor_name'     in _col_names: _extra_cols['supervisor_name']     = supervisor_name
                if 'other_details'       in _col_names: _extra_cols['other_details']       = other_details_json
                if 'unload_checklist'    in _col_names: _extra_cols['unload_checklist']    = unload_checklist_json
            except Exception:
                pass

            _has_other_cols = bool(_extra_cols)   # legacy flag — true if any extra col present

            # Build the dynamic SQL pieces
            _extra_set    = ", ".join([f"{k}=%s" for k in _extra_cols.keys()]) if _extra_cols else ""
            _extra_cols_l = ", ".join(_extra_cols.keys()) if _extra_cols else ""
            _extra_phs    = ", ".join(["%s"] * len(_extra_cols)) if _extra_cols else ""
            _extra_vals   = list(_extra_cols.values())

            if grn_id:
                # Need grn_num for box labels/remarks on edit too.
                try:
                    _gr = conn.execute("SELECT grn_num FROM procurement_grn WHERE id=%s", (grn_id,)).fetchone()
                    if _gr: grn_num = _gr["grn_num"] if hasattr(_gr, "get") else _gr[0]
                except Exception:
                    pass
                # Reverse old stock before re-inserting
                old_items = conn.execute(
                    "SELECT id, material, batch_num, received_qty FROM procurement_grn_items WHERE grn_id=%s",
                    (grn_id,)
                ).fetchall()
                # Map (lc_material, batch) → old grn_item_id so we can RE-POINT
                # the reserved label codes (rm_grn_box_codes) from the old item
                # id to the new one after re-insert. This keeps previously
                # printed labels valid across edits (codes never change).
                _old_item_ids_by_mat_batch = {}
                try:
                    for _oi in old_items:
                        _om = (_oi["material"] if hasattr(_oi, "get") else _oi[1]) or ""
                        _ob = (_oi["batch_num"] if hasattr(_oi, "get") else _oi[2]) or ""
                        _oid = int(_oi["id"] if hasattr(_oi, "get") else _oi[0])
                        _old_item_ids_by_mat_batch[(_om.strip().lower(), _ob.strip())] = _oid
                except Exception:
                    _old_item_ids_by_mat_batch = {}
                if _has_other_cols:
                    _sql = (
                        "UPDATE procurement_grn SET "
                        "supplier_name=%s, grn_date=%s, invoice_num=%s, invoice_date=%s, "
                        "po_id=%s, po_num=%s, status=%s, remarks=%s, "
                        "freight_charge=%s, packing_charge=%s, "
                        "grand_total=%s, "
                        "po_invoices=%s, unload_location=%s, voucher_type_name=%s, updated_by=%s"
                        + (", " + _extra_set if _extra_set else "")
                        + " WHERE id=%s"
                    )
                    _params = ([supplier, grn_date, invoice_num, invoice_date,
                                po_id, po_num, status, remarks,
                                freight, packing,
                                grand_total,
                                po_invoices_json, unload_location, voucher_type_name, uid]
                               + _extra_vals
                               + [grn_id])
                    conn.execute(_sql, tuple(_params))
                else:
                    conn.execute("""
                        UPDATE procurement_grn
                        SET supplier_name=%s, grn_date=%s, invoice_num=%s, invoice_date=%s,
                            po_id=%s, po_num=%s, status=%s, remarks=%s,
                            freight_charge=%s, packing_charge=%s, grand_total=%s,
                            po_invoices=%s, unload_location=%s, voucher_type_name=%s, updated_by=%s
                        WHERE id=%s
                    """, (supplier, grn_date, invoice_num, invoice_date,
                          po_id, po_num, status, remarks,
                          freight, packing, grand_total,
                          po_invoices_json, unload_location, voucher_type_name, uid, grn_id))

                # Capture COA-file → material mapping BEFORE we wipe old items,
                # so we can re-link the files to the freshly-inserted item IDs
                # (otherwise editing a GRN would silently orphan all COA files).
                # We capture relative_path too so files for materials that get
                # removed can be cleaned up from disk.
                coa_by_material = {}   # { lowercase_material_name: [(file_id, relative_path), ...] }
                try:
                    coa_rows = conn.execute("""
                        SELECT f.id AS file_id, f.relative_path, i.material
                        FROM procurement_grn_files f
                        JOIN procurement_grn_items i ON i.id = f.grn_item_id
                        WHERE f.grn_id = %s AND f.kind = 'coa'
                    """, (grn_id,)).fetchall()
                    for cr in coa_rows:
                        mat_key = (cr["material"] or "").strip().lower()
                        if mat_key:
                            coa_by_material.setdefault(mat_key, []).append(
                                (cr["file_id"], cr.get("relative_path"))
                            )
                except Exception:
                    # Table may not exist yet — fine, nothing to preserve.
                    pass

                # Capture TRS rows BEFORE the line-item wipe so we can
                # re-link them to the freshly inserted grn_item_id values.
                # Otherwise editing a GRN orphans every TRS (the picker
                # then thinks no TRS exists and tries to re-generate, and
                # the GRN line shows no TRS badge). Keyed by (material,
                # batch_num) since two lines can share the same material
                # under different batches.
                trs_by_key = {}  # { (lc_mat, batch): trs_id }
                try:
                    trs_rows = conn.execute("""
                        SELECT t.id AS trs_id, t.material, t.batch_num
                        FROM procurement_grn_trs t
                        WHERE t.grn_id = %s
                    """, (grn_id,)).fetchall()
                    for tr in trs_rows:
                        td = dict(tr) if hasattr(tr, "keys") else tr
                        mat_key   = (td.get("material") or "").strip().lower()
                        batch_key = (td.get("batch_num") or "").strip()
                        if mat_key:
                            trs_by_key[(mat_key, batch_key)] = td["trs_id"]
                except Exception:
                    # Table may not exist on a very fresh install — fine.
                    pass

                conn.execute("DELETE FROM procurement_grn_items WHERE grn_id=%s", (grn_id,))
                # Re-saving a GRN must not double-count stock, AND must let the
                # SAME box codes be reused so previously-printed labels keep
                # scanning (codes are reserved in rm_grn_box_codes). Strategy:
                #   • If a previously-created 'grn' box is UNTOUCHED (only its
                #     grn_create/cancel movements, never transferred, never
                #     split) → HARD-DELETE it. This frees its box_code so the
                #     recreate step below can re-insert the identical code.
                #   • If it has any real history (a transfer, a split child,
                #     or any movement beyond grn_create/cancel) → keep it but
                #     mark 'cancelled' (preserves audit + referential rows).
                #     That line will get freshly-reserved codes instead.
                try:
                    old_boxes = conn.execute(
                        "SELECT box_id, box_code, current_godown_id, per_box_qty FROM rm_boxes "
                        "WHERE grn_id=%s AND source='grn' AND current_status='in_stock'",
                        (grn_id,),
                    ).fetchall()
                    for ob in old_boxes:
                        _bid  = int(ob["box_id"] if hasattr(ob, "get") else ob[0])
                        _bcode= ob["box_code"] if hasattr(ob, "get") else ob[1]
                        _ogid = ob["current_godown_id"] if hasattr(ob, "get") else ob[2]

                        # Is this box safe to hard-delete? It must have NO
                        # references other than its own creation movement.
                        # FAIL-SAFE: any check that cannot positively confirm
                        # "no references" leaves the box intact (cancelled).
                        safe_to_delete = True
                        try:
                            # any transfer-voucher line referencing it?
                            tref = conn.execute(
                                "SELECT 1 FROM rm_stock_transfer_boxes WHERE box_id=%s LIMIT 1",
                                (_bid,),
                            ).fetchone()
                            if tref:
                                safe_to_delete = False
                        except Exception:
                            # Table genuinely absent → no transfers possible →
                            # this check doesn't block deletion. (Other refs
                            # are still checked below.)
                            pass
                        if safe_to_delete:
                            try:
                                # any split children pointing at it?
                                cref = conn.execute(
                                    "SELECT 1 FROM rm_boxes WHERE parent_box_id=%s LIMIT 1",
                                    (_bid,),
                                ).fetchone()
                                if cref:
                                    safe_to_delete = False
                            except Exception:
                                # parent_box_id column may not exist (box-split
                                # never enabled) → no children possible → ok.
                                pass
                        if safe_to_delete:
                            try:
                                # Must have ONLY grn_create / cancel movements.
                                mref = conn.execute(
                                    "SELECT 1 FROM rm_box_movements "
                                    "WHERE box_id=%s AND movement_type NOT IN "
                                    "('grn_create','cancel') LIMIT 1",
                                    (_bid,),
                                ).fetchone()
                                if mref:
                                    safe_to_delete = False
                            except Exception:
                                # Can't verify movement history → DO NOT delete.
                                safe_to_delete = False

                        if safe_to_delete:
                            # Hard-delete: remove its movement rows first, then
                            # the box. Frees box_code for identical reuse.
                            try:
                                conn.execute(
                                    "DELETE FROM rm_box_movements WHERE box_id=%s", (_bid,)
                                )
                            except Exception:
                                pass
                            conn.execute(
                                "DELETE FROM rm_boxes WHERE box_id=%s", (_bid,)
                            )
                        else:
                            # Has history → keep as cancelled (audit-safe).
                            conn.execute(
                                "UPDATE rm_boxes SET current_status='cancelled' WHERE box_id=%s",
                                (_bid,),
                            )
                            conn.execute(
                                "INSERT INTO rm_box_movements "
                                "(box_id, movement_type, from_godown_id, to_godown_id, qty, moved_by, remarks) "
                                "VALUES (%s,'cancel',%s,NULL,0,%s,'GRN edited — boxes recreated')",
                                (_bid, _ogid, uid),
                            )
                except Exception:
                    traceback.print_exc()
                for oi in old_items:
                    omat = (oi["material"] or "").strip()
                    oqty = float(oi["received_qty"] or 0)
                    if omat and oqty > 0 and _MAT_HAS_STOCK_COL["value"]:
                        try:
                            conn.execute("""
                                UPDATE procurement_materials
                                SET in_stock_qty = COALESCE(in_stock_qty,0) - %s
                                WHERE material_name = %s
                            """, (oqty, omat))
                        except Exception:
                            pass
            else:
                # ── Unified voucher numbering for GRN ──
                prefix = "GRN"
                suffix = ""
                digits = 4
                today_str = datetime.now().date().isoformat()
                try:
                    vn_type_key = voucher_type_name or "grn"
                    vn_row = conn.execute(
                        "SELECT prefix, suffix, digits FROM procurement_voucher_numbering "
                        "WHERE voucher_type=%s AND valid_from <= %s AND valid_to >= %s "
                        "ORDER BY id DESC LIMIT 1",
                        (vn_type_key, today_str, today_str)
                    ).fetchone()
                    if not vn_row and voucher_type_name:
                        vn_row = conn.execute(
                            "SELECT prefix, suffix, digits FROM procurement_voucher_numbering "
                            "WHERE voucher_type='grn' AND valid_from <= %s AND valid_to >= %s "
                            "ORDER BY id DESC LIMIT 1",
                            (today_str, today_str)
                        ).fetchone()
                    if vn_row:
                        prefix = (vn_row["prefix"] or "GRN").strip()
                        suffix = (vn_row["suffix"] or "").strip()
                        digits = int(vn_row["digits"] or 4)
                except Exception:
                    pass

                conn.execute("SELECT GET_LOCK('grn_num_lock',10) AS lk")
                pattern = (prefix + "/%") if prefix else "GRN/%"
                rows = conn.execute(
                    "SELECT grn_num FROM procurement_grn WHERE grn_num LIKE %s",
                    (pattern,)
                ).fetchall()
                max_seq = 0
                for row in rows:
                    nums = _re.findall(r"(\d{" + str(digits) + r",})", row["grn_num"])
                    if nums:
                        max_seq = max(max_seq, int(nums[-1]))
                next_seq = max_seq + 1
                num_str = str(next_seq).zfill(digits)
                parts = []
                if prefix: parts.append(prefix)
                parts.append(num_str)
                if suffix: parts.append(suffix)
                grn_num = "/".join(parts)
                conn.execute("SELECT RELEASE_LOCK('grn_num_lock')")
                if _has_other_cols:
                    _sql = (
                        "INSERT INTO procurement_grn "
                        "(grn_num, supplier_name, grn_date, invoice_num, invoice_date, "
                        "po_id, po_num, status, remarks, "
                        "freight_charge, packing_charge, "
                        "grand_total, "
                        "po_invoices, unload_location, voucher_type_name, created_by, updated_by"
                        + (", " + _extra_cols_l if _extra_cols_l else "")
                        + ") VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s"
                        + ((", " + _extra_phs) if _extra_phs else "")
                        + ")"
                    )
                    _params = ([grn_num, supplier, grn_date, invoice_num, invoice_date,
                                po_id, po_num, status, remarks,
                                freight, packing,
                                grand_total,
                                po_invoices_json, unload_location, voucher_type_name, uid, uid]
                               + _extra_vals)
                    conn.execute(_sql, tuple(_params))
                else:
                    conn.execute("""
                        INSERT INTO procurement_grn
                            (grn_num, supplier_name, grn_date, invoice_num, invoice_date,
                             po_id, po_num, status, remarks,
                             freight_charge, packing_charge, grand_total,
                             po_invoices, unload_location, voucher_type_name, created_by, updated_by)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (grn_num, supplier, grn_date, invoice_num, invoice_date,
                          po_id, po_num, status, remarks,
                          freight, packing, grand_total,
                          po_invoices_json, unload_location, voucher_type_name, uid, uid))
                grn_id = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]

            # Insert items + update stock
            # `new_item_ids_by_material` lets us re-link preserved COA file rows
            # after the items get re-inserted (their primary keys change on edit).
            new_item_ids_by_material = {}   # { lowercase_material_name: new_item_id }
            # Same shape but keyed by (material, batch_num) — used to
            # re-link TRS rows after the DELETE+INSERT cycle. Two GRN
            # lines with the same material under different batches need
            # to resolve to distinct TRS rows.
            new_item_ids_by_mat_batch = {}  # { (lc_mat, batch): new_item_id }
            for item in items:
                mat = (item.get("material") or "").strip()
                if not mat:
                    continue
                try:
                    rqty   = float(item.get("received_qty") or 0) or None
                    po_qty = float(item.get("po_qty") or 0) or None
                    rate   = float(item.get("rate") or 0) or None
                    amount = (rqty or 0) * (rate or 0) or None
                except Exception:
                    rqty = po_qty = rate = amount = None
                hsn         = (item.get("hsn_code")    or "").strip() or None
                loc         = (item.get("location")    or "").strip() or None
                inv_num     = (item.get("invoice_num") or "").strip() or None
                inv_date    = item.get("invoice_date") or None
                batch_num   = (item.get("batch_num")   or "").strip() or None
                mfg_date    = item.get("mfg_date")     or None
                expiry_date = item.get("expiry_date")  or None
                manufacturer = (item.get("manufacturer") or "").strip()[:200] or None
                try:
                    gst = float(item.get("gst_rate") or 0) or None
                except Exception:
                    gst = None
                # Check whether procurement_grn_items has a manufacturer
                # column (detected lazily on first use). If so, include it;
                # otherwise fall back to the legacy INSERT shape.
                if _items_has_mfr(conn):
                    conn.execute("""
                        INSERT INTO procurement_grn_items
                            (grn_id, material, po_qty, received_qty, qty_per_pkg, rate, amount,
                             hsn_code, gst_rate, location, invoice_num, invoice_date,
                             batch_num, mfg_date, expiry_date, packages, uom, manufacturer)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (grn_id, mat, po_qty, rqty,
                          float(item.get("qty_per_pkg") or 0) or None,
                          rate, amount,
                          hsn, gst, loc, inv_num, inv_date,
                          batch_num, mfg_date, expiry_date,
                          int(item.get("packages")) if item.get("packages") else None,
                          (item.get("uom") or "KG").strip().upper() or None,
                          manufacturer))
                else:
                    conn.execute("""
                        INSERT INTO procurement_grn_items
                            (grn_id, material, po_qty, received_qty, qty_per_pkg, rate, amount,
                             hsn_code, gst_rate, location, invoice_num, invoice_date,
                             batch_num, mfg_date, expiry_date, packages, uom)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (grn_id, mat, po_qty, rqty,
                          float(item.get("qty_per_pkg") or 0) or None,
                          rate, amount,
                          hsn, gst, loc, inv_num, inv_date,
                          batch_num, mfg_date, expiry_date,
                          int(item.get("packages")) if item.get("packages") else None,
                          (item.get("uom") or "KG").strip().upper() or None))
                # Track the new item ID for COA re-linking.
                new_id = None
                try:
                    new_id = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
                    new_item_ids_by_material[mat.lower()] = new_id
                    # Also stash by (material, batch) so TRS re-linking can
                    # disambiguate between two lines of the same material
                    # received under different batches. batch_num may be
                    # NULL/empty — we still key with "" so single-line cases
                    # without batch still re-link cleanly.
                    new_item_ids_by_mat_batch[(mat.lower(), (batch_num or "").strip())] = new_id

                    # RE-POINT reserved label codes: move rm_grn_box_codes rows
                    # for this material/batch from the OLD grn_item_id to the
                    # NEW one, so _grn_create_boxes_for_line below finds them
                    # and recreates boxes with the SAME codes already printed
                    # on labels. Without this, an edit orphans the reserved
                    # codes (the root cause of un-scannable "ghost" labels).
                    try:
                        _old_iid = _old_item_ids_by_mat_batch.get(
                            (mat.lower(), (batch_num or "").strip())
                        )
                        # Fall back to material-only match if batch differs.
                        if _old_iid is None:
                            for (_m, _b), _i in _old_item_ids_by_mat_batch.items():
                                if _m == mat.lower():
                                    _old_iid = _i
                                    break
                        if _old_iid and new_id and _old_iid != new_id:
                            conn.execute(
                                "UPDATE rm_grn_box_codes SET grn_item_id=%s "
                                "WHERE grn_id=%s AND grn_item_id=%s",
                                (new_id, grn_id, _old_iid),
                            )
                    except Exception:
                        # rm_grn_box_codes may not exist on old installs — the
                        # box creator will just allocate fresh codes (safe).
                        pass
                except Exception:
                    pass
                # ── Option A: create godown boxes from this GRN line so the
                # received stock shows up in the godown view / box system. ──
                try:
                    _mrow = conn.execute(
                        "SELECT id FROM procurement_materials WHERE material_name=%s LIMIT 1",
                        (mat,),
                    ).fetchone()
                    _mid = int(_mrow["id"] if hasattr(_mrow, "get") else _mrow[0]) if _mrow else 0
                    if _mid:
                        _grn_create_boxes_for_line(
                            conn, grn_id, grn_num, new_id, _mid,
                            loc, item.get("packages"),
                            item.get("qty_per_pkg"), rqty,
                            (item.get("uom") or "KG").strip().upper(), uid,
                        )
                except Exception:
                    traceback.print_exc()
                # Update stock
                if rqty and rqty > 0 and _MAT_HAS_STOCK_COL["value"]:
                    try:
                        conn.execute("""
                            UPDATE procurement_materials
                            SET in_stock_qty = COALESCE(in_stock_qty,0) + %s
                            WHERE material_name = %s
                        """, (rqty, mat))
                    except Exception:
                        pass

            # Re-link preserved COA files (only relevant for edit-saves, since
            # `coa_by_material` is only populated in the edit branch above).
            try:
                if 'coa_by_material' in locals() and coa_by_material:
                    for mat_key, pairs in coa_by_material.items():
                        file_ids = [p[0] for p in pairs]
                        rel_paths = [p[1] for p in pairs]
                        new_id = new_item_ids_by_material.get(mat_key)
                        if not new_id:
                            # The material was removed from this GRN — drop its COA(s)
                            # from DB and clean the underlying disk files.
                            conn.execute(
                                "DELETE FROM procurement_grn_files "
                                "WHERE id IN (%s)" % ",".join(["%s"] * len(file_ids)),
                                tuple(file_ids)
                            )
                            for rp in rel_paths:
                                try:
                                    abs_p = _grn_file_disk_path(rp)
                                    if abs_p and os.path.isfile(abs_p):
                                        os.remove(abs_p)
                                except Exception:
                                    pass
                        else:
                            conn.execute(
                                "UPDATE procurement_grn_files SET grn_item_id=%%s "
                                "WHERE id IN (%s)" % ",".join(["%s"] * len(file_ids)),
                                (new_id, *file_ids)
                            )
            except Exception:
                # Non-fatal — files may briefly point to stale items, recoverable.
                traceback.print_exc()

            # Re-link TRS rows captured before the line-item wipe. Each
            # row points at the OLD grn_item_id; we update it to the
            # freshly inserted line that matches by (material, batch_num).
            # If the matching line was removed entirely on this save, we
            # leave the TRS row alone with its now-stale grn_item_id —
            # the operator may want to re-add the line later, and the
            # data is still valuable for audit. Worst case the badge
            # won't show; the row isn't lost.
            try:
                if 'trs_by_key' in locals() and trs_by_key:
                    for (mat_key, batch_key), trs_id in trs_by_key.items():
                        # Match strict (mat, batch) first; if a TRS row had
                        # no batch saved, fall back to mat-only lookup.
                        new_iid = new_item_ids_by_mat_batch.get((mat_key, batch_key))
                        if not new_iid and not batch_key:
                            new_iid = new_item_ids_by_material.get(mat_key)
                        if new_iid:
                            try:
                                conn.execute(
                                    "UPDATE procurement_grn_trs "
                                    "SET grn_item_id=%s WHERE id=%s",
                                    (new_iid, trs_id)
                                )
                            except Exception:
                                traceback.print_exc()
            except Exception:
                # Non-fatal — TRS data is still in the DB even if the
                # link isn't restored.
                traceback.print_exc()

            conn.commit()
            conn.close()
            conn = None

            # ── Auto-update linked PO status (mirrors procurement.py logic) ──
            linked_po_ids = set()
            updated_po_statuses = {}
            if po_id:
                linked_po_ids.add(int(po_id))
            for inv in (po_invoices_raw or []):
                if inv.get("po_id"):
                    linked_po_ids.add(int(inv["po_id"]))

            if linked_po_ids:
                try:
                    pc = sampling_portal.get_db_connection()
                    if pc:
                        for lpid in linked_po_ids:
                            po_items = pc.execute(
                                "SELECT material, qty FROM procurement_po_items WHERE po_id=%s",
                                (lpid,)
                            ).fetchall()
                            if not po_items:
                                continue
                            po_row = pc.execute(
                                "SELECT status FROM procurement_purchase_orders WHERE id=%s",
                                (lpid,)
                            ).fetchone()
                            if not po_row or po_row["status"] == "cancelled":
                                continue
                            total_received = {}
                            grn_rows = pc.execute(
                                "SELECT id FROM procurement_grn WHERE po_id=%s AND status <> 'cancelled'",
                                (lpid,)
                            ).fetchall()
                            json_grns = pc.execute(
                                "SELECT id FROM procurement_grn WHERE po_invoices LIKE %s AND status <> 'cancelled'",
                                ('%"po_id": ' + str(lpid) + '%',)
                            ).fetchall()
                            json_grns2 = pc.execute(
                                "SELECT id FROM procurement_grn WHERE po_invoices LIKE %s AND status <> 'cancelled'",
                                ('%"po_id":' + str(lpid) + '%',)
                            ).fetchall()
                            all_grn_ids = set(r["id"] for r in grn_rows)
                            all_grn_ids.update(r["id"] for r in json_grns)
                            all_grn_ids.update(r["id"] for r in json_grns2)
                            for grn_id_ref in all_grn_ids:
                                g_items = pc.execute(
                                    "SELECT material, received_qty FROM procurement_grn_items WHERE grn_id=%s",
                                    (grn_id_ref,)
                                ).fetchall()
                                for gi in g_items:
                                    mat = (gi["material"] or "").strip()
                                    rqty = float(gi["received_qty"] or 0)
                                    total_received[mat] = total_received.get(mat, 0) + rqty
                            # Determine new PO status
                            any_received = any(v > 0 for v in total_received.values())
                            all_full = all(
                                total_received.get((i["material"] or "").strip(), 0)
                                >= float(i["qty"] or 0) - 0.001
                                for i in po_items
                            )
                            if not any_received:
                                new_po_status = "open"
                            elif all_full:
                                new_po_status = "closed"
                            else:
                                new_po_status = "partial"
                            pc.execute(
                                "UPDATE procurement_purchase_orders SET status=%s WHERE id=%s",
                                (new_po_status, lpid)
                            )
                            updated_po_statuses[lpid] = new_po_status
                        pc.commit()
                        pc.close()
                except Exception:
                    traceback.print_exc()

            if not d.get("id"):
                return jsonify({
                    "status": "ok",
                    "id": grn_id,
                    "grn_num": grn_num,
                    "grand_total": grand_total,
                    "po_statuses": updated_po_statuses
                })
            return jsonify({
                "status": "ok",
                "id": grn_id,
                "grand_total": grand_total,
                "po_statuses": updated_po_statuses
            })
        except Exception as e:
            traceback.print_exc()
            try:
                if conn: conn.close()
            except Exception:
                pass
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.route("/api/inventory_mgmt/grn/backfill_boxes", methods=["POST"])
    @_edit_required
    def api_inv_grn_backfill_boxes():
        """One-time backfill: create godown boxes for EXISTING GRNs that don't
        have any yet. Admin-only. Defaults to DRY-RUN (preview only). Pass
        {"confirm":"BACKFILL"} to actually write.

        Safety:
          • Skips any GRN that already has source='grn' boxes (re-runnable).
          • Skips REJOUT / non-GRN vouchers.
          • Flags materials that already have opening-stock boxes (possible
            double-count) so the admin can decide.
        """
        if not _can_edit_inventory():
            return jsonify({"status": "error", "message": "Admin only"}), 403
        d = request.get_json(silent=True) or {}
        do_write = (d.get("confirm") == "BACKFILL")
        uid = session.get("UID", "") or "backfill"
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            # GRNs that already have boxes — skip these.
            done = set()
            for r in conn.execute(
                "SELECT DISTINCT grn_id FROM rm_boxes WHERE source='grn' AND grn_id IS NOT NULL"
            ).fetchall():
                done.add(int(r["grn_id"] if hasattr(r, "get") else r[0]))

            # Materials that already have opening-stock (for double-count warning).
            opening_mats = set()
            for r in conn.execute(
                "SELECT DISTINCT material_id FROM rm_boxes "
                "WHERE source='opening' AND current_status='in_stock' AND material_id IS NOT NULL"
            ).fetchall():
                opening_mats.add(int(r["material_id"] if hasattr(r, "get") else r[0]))

            # All real GRNs (exclude REJOUT/non-GRN).
            grns = conn.execute(
                "SELECT id, grn_num FROM procurement_grn "
                "WHERE (grn_type IS NULL OR grn_type='GRN' OR grn_type='') "
                "ORDER BY id ASC"
            ).fetchall()

            preview = []
            total_boxes = 0
            grns_done = 0
            warn_double = []
            for g in grns:
                gid = int(g["id"]); gnum = g["grn_num"] or ""
                if gid in done:
                    continue  # already has boxes
                items = conn.execute(
                    "SELECT id, material, received_qty, qty_per_pkg, packages, location, uom "
                    "FROM procurement_grn_items WHERE grn_id=%s ORDER BY id ASC",
                    (gid,),
                ).fetchall()
                gboxes = 0
                for it in items:
                    mat = (it["material"] or "").strip()
                    if not mat:
                        continue
                    mrow = conn.execute(
                        "SELECT id FROM procurement_materials WHERE material_name=%s LIMIT 1",
                        (mat,),
                    ).fetchone()
                    if not mrow:
                        continue
                    mid = int(mrow["id"] if hasattr(mrow, "get") else mrow[0])
                    if mid in opening_mats and mat not in warn_double:
                        warn_double.append(mat)
                    if do_write:
                        n = _grn_create_boxes_for_line(
                            conn, gid, gnum, int(it["id"]), mid,
                            it["location"], it["packages"], it["qty_per_pkg"],
                            it["received_qty"], (it["uom"] or "KG"), uid,
                        )
                        gboxes += n
                    else:
                        # dry-run: compute what WOULD be created
                        nb = int(it["packages"] or 0)
                        per = float(it["qty_per_pkg"] or 0)
                        rq = float(it["received_qty"] or 0)
                        if nb <= 0:
                            nb = 1 if rq > 0 else 0
                        gboxes += nb
                if gboxes > 0:
                    preview.append({"grn_num": gnum, "boxes": gboxes})
                    total_boxes += gboxes
                    grns_done += 1
            if do_write:
                conn.commit()
            conn.close()
            return jsonify({
                "status": "ok",
                "mode": "written" if do_write else "dry_run",
                "grns_processed": grns_done,
                "boxes_created" if do_write else "boxes_would_create": total_boxes,
                "double_count_warning_materials": warn_double,
                "preview": preview[:200],
            })
        except Exception as e:
            try: conn.close()
            except Exception: pass
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/inventory_mgmt/grn/delete", methods=["POST"])
    @_grn_delete_required
    def api_inv_grn_delete():
        d = request.get_json() or {}
        grn_id = d.get("id")
        if not grn_id:
            return jsonify({"status": "error", "message": "id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            # Reverse stock for items in this GRN before deleting
            items = conn.execute(
                "SELECT material, received_qty FROM procurement_grn_items WHERE grn_id=%s",
                (grn_id,)
            ).fetchall()
            for it in items:
                mat = (it["material"] or "").strip()
                rqty = float(it["received_qty"] or 0)
                if mat and rqty > 0 and _MAT_HAS_STOCK_COL["value"]:
                    try:
                        conn.execute("""
                            UPDATE procurement_materials
                            SET in_stock_qty = COALESCE(in_stock_qty,0) - %s
                            WHERE material_name = %s
                        """, (rqty, mat))
                    except Exception:
                        pass
            conn.execute("DELETE FROM procurement_grn WHERE id=%s", (grn_id,))
            # Also remove any attached COA / Invoice files for this GRN.
            try:
                conn.execute("DELETE FROM procurement_grn_files WHERE grn_id=%s", (grn_id,))
            except Exception:
                # Table may not exist on legacy installs; ignore.
                pass
            conn.commit()
            conn.close()
            # Best-effort: nuke the GRN's upload folder on disk.
            try:
                grn_dir = os.path.join(_GRN_UPLOAD_ROOT, "grn", str(grn_id))
                if os.path.isdir(grn_dir):
                    shutil.rmtree(grn_dir, ignore_errors=True)
            except Exception:
                pass
            return jsonify({"status": "ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    # ════════════════════════════════════════════════════════════════════════════
    # GRN SUPPORT LOOKUPS  —  thin passthroughs needed by the GRN form
    # ════════════════════════════════════════════════════════════════════════════

    @app.route("/api/inventory_mgmt/godowns", methods=["GET"])
    @_login_required
    def api_inv_godowns():
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            rows = conn.execute(
                "SELECT * FROM procurement_godowns ORDER BY is_default DESC, name ASC"
            ).fetchall()
            return jsonify({"status": "ok", "godowns": [dict(r) for r in rows]})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    @app.route("/api/inventory_mgmt/po/list", methods=["GET"])
    @_login_required
    def api_inv_po_list():
        """Light PO list — id, po_num, supplier, status, voucher_type_name only."""
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            try:
                rows = conn.execute("""
                    SELECT id, po_num, po_date, supplier_name, status,
                           voucher_type_name
                    FROM procurement_purchase_orders
                    ORDER BY created_at DESC
                """).fetchall()
            except Exception:
                # If voucher_type_name col doesn't exist on this install
                rows = conn.execute("""
                    SELECT id, po_num, po_date, supplier_name, status
                    FROM procurement_purchase_orders
                    ORDER BY created_at DESC
                """).fetchall()
            out = []
            for r in rows:
                d = dict(r)
                for f in ("po_date",):
                    if d.get(f):
                        d[f] = str(d[f])
                out.append(d)
            return jsonify({"status": "ok", "orders": out})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    @app.route("/api/inventory_mgmt/po/get", methods=["GET"])
    @_login_required
    def api_inv_po_get():
        """Full PO with items + per-item pending qty (already-received subtracted)."""
        po_id = request.args.get("id")
        if not po_id:
            return jsonify({"status": "error", "message": "id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            po = conn.execute(
                "SELECT * FROM procurement_purchase_orders WHERE id=%s", (po_id,)
            ).fetchone()
            if not po:
                return jsonify({"status": "error", "message": "PO not found"}), 404
            order = dict(po)
            for f in ("po_date", "created_at", "updated_at"):
                if order.get(f):
                    order[f] = str(order[f])

            items_raw = conn.execute(
                "SELECT * FROM procurement_po_items WHERE po_id=%s ORDER BY id",
                (po_id,)
            ).fetchall()

            # Already-received qty per material (across all GRNs linked to this PO)
            received = {}
            try:
                # Direct po_id link
                grn_items = conn.execute(
                    "SELECT gi.material, COALESCE(SUM(gi.received_qty),0) AS r "
                    "FROM procurement_grn_items gi "
                    "JOIN procurement_grn g ON gi.grn_id = g.id "
                    "WHERE g.po_id=%s AND g.status <> 'cancelled' "
                    "GROUP BY gi.material",
                    (po_id,)
                ).fetchall()
                for gi in grn_items:
                    received[(gi["material"] or "").strip().lower()] = float(gi["r"] or 0)
                # JSON-linked GRNs
                json_grns = conn.execute(
                    "SELECT id, po_invoices FROM procurement_grn "
                    "WHERE po_invoices LIKE %s AND status <> 'cancelled'",
                    ('%"po_id":%' + str(po_id) + '%',)
                ).fetchall()
                for jg in json_grns:
                    try:
                        invs = json.loads(jg["po_invoices"] or "[]")
                    except Exception:
                        invs = []
                    if not any(int(inv.get("po_id") or 0) == int(po_id) for inv in invs):
                        continue
                    g_items = conn.execute(
                        "SELECT material, received_qty FROM procurement_grn_items WHERE grn_id=%s",
                        (jg["id"],)
                    ).fetchall()
                    for gi in g_items:
                        k = (gi["material"] or "").strip().lower()
                        received[k] = received.get(k, 0) + float(gi["received_qty"] or 0)
            except Exception:
                pass

            items = []
            for i in items_raw:
                d2 = dict(i)
                for f in ("qty", "rate", "amount", "gst_rate"):
                    if d2.get(f) is not None:
                        d2[f] = float(d2[f])
                qty = float(d2.get("qty") or 0)
                rcvd = received.get((d2.get("material") or "").strip().lower(), 0)
                d2["received_qty"] = rcvd
                d2["pending_qty"]  = max(0, qty - rcvd)
                items.append(d2)
            order["items"] = items
            return jsonify({"status": "ok", "order": order})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    @app.route("/api/inventory_mgmt/grn_materials", methods=["GET"])
    @_login_required
    def api_inv_grn_materials():
        """
        Flat material list for the GRN form's autocomplete datalist.
        Mirrors procurement's join — group_name and mat_type_abbr come from
        joined tables, not procurement_materials itself. We don't expose
        in_stock_qty here (procurement merges that from StkSum.xlsx separately;
        not needed for the GRN line autocomplete).
        """
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            rows = conn.execute("""
                SELECT m.material_name,
                       m.uom,
                       m.hsn_code,
                       m.gst_rate,
                       g.group_name           AS group_name,
                       mt.abbreviation        AS mat_type_abbr
                FROM procurement_materials m
                LEFT JOIN procurement_material_groups g  ON m.group_id          = g.id
                LEFT JOIN procurement_material_types  mt ON m.material_type_id  = mt.id
                ORDER BY m.material_name ASC
            """).fetchall()
            out = []
            for r in rows:
                d = dict(r)
                if d.get("gst_rate") is not None:
                    try: d["gst_rate"] = float(d["gst_rate"])
                    except Exception: pass
                out.append(d)
            return jsonify({"status": "ok", "rows": out})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    @app.route("/api/inventory_mgmt/voucher_types", methods=["GET"])
    @_login_required
    def api_inv_voucher_types():
        """
        Reads gop_voucher_type_masters and returns rows matching parent_type.
        Errors are now logged loudly instead of being swallowed — if the
        dropdown is empty, the server log will tell us why.
        """
        parent_type = (request.args.get("parent_type") or "grn").strip().lower()
        conn = None
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                print(f"[inv_voucher_types] DB connection failed for parent_type={parent_type}")
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            try:
                rows = conn.execute(
                    "SELECT * FROM gop_voucher_type_masters "
                    "WHERE parent_type=%s "
                    "ORDER BY COALESCE(sort_order, 9999), name",
                    (parent_type,)
                ).fetchall()
                return jsonify({"status": "ok", "types": [dict(r) for r in rows]})
            except Exception as inner:
                # Most common: gop_voucher_type_masters table doesn't exist on
                # this install. Degrade gracefully — voucher types are optional.
                # Stay quiet for the expected "table doesn't exist" (1146) case
                # so it doesn't spam the console on every page load; only log
                # genuinely unexpected errors.
                msg = str(inner)
                is_missing_table = "1146" in msg or "doesn't exist" in msg or "does not exist" in msg
                if not is_missing_table:
                    print(f"[inv_voucher_types] SQL error for parent_type={parent_type}: {inner}")
                return jsonify({"status": "ok", "types": []})
        except Exception as e:
            import traceback as _tb
            _tb.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try:
                if conn: conn.close()
            except Exception: pass


    @app.route("/api/inventory_mgmt/voucher_numbering/next", methods=["GET"])
    @_login_required
    def api_inv_voucher_numbering_next():
        """Preview next GRN number based on procurement_voucher_numbering style."""
        import re as _re
        voucher_type = (request.args.get("voucher_type") or "grn").strip()
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            today_str = datetime.now().date().isoformat()
            prefix, suffix, digits = "GRN", "", 4
            try:
                vn_row = conn.execute(
                    "SELECT prefix, suffix, digits FROM procurement_voucher_numbering "
                    "WHERE voucher_type=%s AND valid_from <= %s AND valid_to >= %s "
                    "ORDER BY id DESC LIMIT 1",
                    (voucher_type, today_str, today_str)
                ).fetchone()
                if not vn_row:
                    vn_row = conn.execute(
                        "SELECT prefix, suffix, digits FROM procurement_voucher_numbering "
                        "WHERE voucher_type='grn' AND valid_from <= %s AND valid_to >= %s "
                        "ORDER BY id DESC LIMIT 1",
                        (today_str, today_str)
                    ).fetchone()
                if vn_row:
                    prefix = (vn_row["prefix"] or "GRN").strip()
                    suffix = (vn_row["suffix"] or "").strip()
                    digits = int(vn_row["digits"] or 4)
            except Exception:
                pass
            pattern = (prefix + "/%") if prefix else "GRN/%"
            rows = conn.execute(
                "SELECT grn_num FROM procurement_grn WHERE grn_num LIKE %s",
                (pattern,)
            ).fetchall()
            max_seq = 0
            for row in rows:
                nums = _re.findall(r"(\d{" + str(digits) + r",})", row["grn_num"])
                if nums:
                    max_seq = max(max_seq, int(nums[-1]))
            return jsonify({
                "status": "ok",
                "prefix": prefix, "suffix": suffix,
                "digits": digits, "next": max_seq + 1
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass


    # ═══ AUTH CHECK ════════════════════════════════════════════════════════════
    @app.route("/api/inventory_mgmt/can_access")
    def api_inv_can_access():
        return jsonify({
            "logged_in": bool(session.get("logged_in")),
            "can_view":  _can_inventory(),
            "can_edit":  _can_edit_inventory(),
            "uid":       session.get("UID"),
            "role":      session.get("User_Type"),
        })

    # ═══════════════════════════════════════════════════════════════════
    # TRS (Testing Requisition Slip) — per-GRN-line QC test request
    # ═══════════════════════════════════════════════════════════════════
    #
    # Flow on the page:
    #   1. User opens an existing (saved) GRN's edit view.
    #   2. Print menu → "Generate TRS" → modal lists line items with a
    #      ✅ badge on the ones that already have a TRS, and a radio (one
    #      at a time per spec) on the rest.
    #   3. Operator picks ONE line + clicks Generate → if the line already
    #      has manufacturer/batch/mfg/expiry from the GRN, the TRS is
    #      generated immediately. Otherwise a small modal prompts for the
    #      operator-supplied fields (Physical State / Sample Qty /
    #      Previous Supplier / NEW or OLD).
    #   4. POST /trs/generate creates the row and returns its trs_num +
    #      full snapshot. The JS opens a print-preview window.
    #
    # TRS numbering: {grn_num-after-last-slash}/{line_index_1based}
    # e.g. for GRN "RM/0558/26-27" line #1 → "RM/0558/26-27/1". This
    # mirrors the reference design (TRS No "RM/0558/26-27/1") so the
    # operator can see at a glance which GRN and which line a slip
    # belongs to without consulting the DB.

    def _trs_caps_for_session():
        """TRS generation is gated VERY broadly because generating a TRS
        is not a destructive operation on the GRN — it just creates a
        sibling row in procurement_grn_trs that flows to the QC
        dashboard. Locking it behind GRN edit-class caps mis-modelled
        the workflow: warehouse users who can see (but not edit) GRNs
        still need to send slips to QC.

        Gate: admin → pass; any logged-in user with ANY inventory
        access at all (any toggle on) → pass; QC-role users → pass.
        Anonymous / wholly-unprovisioned users → blocked.

        This intentionally decouples TRS generation from the
        `grn_edit` / `grn_new` toggles in User Access Control, so
        admins can grant TRS-only operators 'view'-level GRN access
        (or even no GRN toggle at all if they only need TRS via the
        QC dashboard) without inadvertently granting GRN editing."""
        if _can_edit_inventory():     # admin shortcut
            return {"generate": True, "view": True}
        try:
            try:
                from .inventory_access import _resolve_user_access
            except Exception:
                from inventory_access import _resolve_user_access
            resolved = _resolve_user_access() or {}
            access   = resolved.get("access") or {}
            # Any non-'off' toggle counts as "this user is a real
            # inventory user" — that's enough to originate a TRS.
            has_any_inv = any(
                str(v or 'off').lower() not in ('', 'off', '0', 'false', 'no')
                for v in access.values()
            )
        except Exception:
            has_any_inv = False
        # QC-role users may not have any inventory toggle but still
        # need to generate TRS from the GRN form when they're cross-
        # trained on receiving.
        role  = (session.get('User_Type') or '').lower()
        is_qc = role in ('qc', 'qc_common', 'purchase')
        ok = has_any_inv or is_qc
        return {"generate": ok, "view": ok}

    def _trs_generate_required(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not session.get("logged_in"):
                return jsonify({"status": "error", "message": "Not logged in"}), 401
            if not _trs_caps_for_session().get("generate"):
                return jsonify({"status": "error",
                                "message": "You don't have permission to generate TRS slips"}), 403
            return f(*args, **kwargs)
        return wrapper

    @app.route("/api/inventory_mgmt/trs/list", methods=["GET"])
    @_login_required
    def api_inv_trs_list():
        """List all TRS rows for a GRN. JS uses this to badge already-
        generated lines in the picker modal so operators don't accidentally
        re-generate the same slip.

        Lazy repair: TRS rows reference grn_item_id, but the GRN /save
        handler DELETEs and re-INSERTs line items on every save, which
        creates new ids and orphans the TRS references. The /grn/save
        path re-links TRS rows AFTER the wipe, but that only fixes
        future saves. For TRS rows that were already orphaned before
        the fix landed, we self-heal here: any row whose grn_item_id
        points to a non-existent line gets re-bound to the matching
        live line (by material + batch) before we return the list.
        Idempotent — does nothing on healthy rows."""
        grn_id = request.args.get("grn_id")
        if not grn_id:
            return jsonify({"status": "error", "message": "grn_id required"}), 400
        try:
            grn_id_i = int(grn_id)
        except Exception:
            return jsonify({"status": "error", "message": "grn_id must be int"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            try:
                # ── Lazy repair ──
                # Fetch TRS rows for this GRN along with their CURRENT
                # live line id (resolved by material+batch). If the
                # stored grn_item_id doesn't match the live one, fix it.
                # We use a Python loop rather than one UPDATE-JOIN so
                # any DDL/dialect quirk on a particular MySQL build
                # falls through safely without breaking the read path.
                try:
                    audit_rows = conn.execute("""
                        SELECT t.id            AS trs_row_id,
                               t.grn_item_id   AS stored_iid,
                               t.material      AS trs_mat,
                               t.batch_num     AS trs_batch,
                               (
                                 SELECT i.id FROM procurement_grn_items i
                                 WHERE i.grn_id = t.grn_id
                                   AND LOWER(TRIM(i.material)) = LOWER(TRIM(t.material))
                                   AND COALESCE(TRIM(i.batch_num),'') = COALESCE(TRIM(t.batch_num),'')
                                 LIMIT 1
                               ) AS live_iid
                        FROM procurement_grn_trs t
                        WHERE t.grn_id = %s
                    """, (grn_id_i,)).fetchall() or []
                    for ar in audit_rows:
                        ad = dict(ar) if hasattr(ar, "keys") else ar
                        stored = ad.get("stored_iid")
                        live   = ad.get("live_iid")
                        # Only repair when (a) a live match exists AND
                        # (b) the stored id differs from the live id.
                        # Don't touch rows where no live match exists —
                        # the matching line may have been removed, and
                        # we should preserve audit data rather than
                        # bind to an unrelated row.
                        if live and stored != live:
                            try:
                                conn.execute(
                                    "UPDATE procurement_grn_trs "
                                    "SET grn_item_id=%s WHERE id=%s",
                                    (live, ad["trs_row_id"])
                                )
                                print(f"[InventoryMgmt] TRS lazy-repair: "
                                      f"trs_id={ad['trs_row_id']} "
                                      f"grn_item_id {stored} → {live} "
                                      f"(mat={ad.get('trs_mat')!r} batch={ad.get('trs_batch')!r})")
                            except Exception:
                                traceback.print_exc()
                    conn.commit()
                except Exception:
                    # Repair is best-effort; never block the actual list.
                    traceback.print_exc()

                # ── Now do the actual read ──
                rows = conn.execute(
                    "SELECT id, trs_num, grn_item_id, material, batch_num, "
                    "       physical_state, generated_by, generated_at, "
                    "       approval_status, approved_by, approval_dt "
                    "FROM procurement_grn_trs WHERE grn_id=%s ORDER BY id",
                    (grn_id_i,)
                ).fetchall() or []
                out = []
                for r in rows:
                    d = dict(r) if hasattr(r, "keys") else r
                    for k in ("generated_at", "approval_dt"):
                        if d.get(k):
                            try:
                                d[k] = d[k].isoformat(sep=" ", timespec="seconds")
                            except Exception:
                                d[k] = str(d[k])
                    out.append(d)
                return jsonify({"status": "ok", "trs": out})
            finally:
                try: conn.close()
                except Exception: pass
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/inventory_mgmt/trs/get/<int:trs_id>", methods=["GET"])
    @_login_required
    def api_inv_trs_get(trs_id):
        """Fetch a single TRS by id — used by the print-again flow when
        an operator re-opens a previously-generated slip."""
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            try:
                row = conn.execute(
                    "SELECT * FROM procurement_grn_trs WHERE id=%s", (trs_id,)
                ).fetchone()
                if not row:
                    return jsonify({"status": "error", "message": "TRS not found"}), 404
                d = dict(row) if hasattr(row, "keys") else row
                for k in ("generated_at", "approval_dt"):
                    if d.get(k):
                        try:
                            d[k] = d[k].isoformat(sep=" ", timespec="seconds")
                        except Exception:
                            d[k] = str(d[k])
                for k in ("grn_date", "mfg_date", "expiry_date"):
                    if d.get(k):
                        try:
                            d[k] = d[k].isoformat()
                        except Exception:
                            d[k] = str(d[k])
                # Number coercion for JSON safety
                for k in ("qty_per_pkg", "total_qty", "sample_qty"):
                    if d.get(k) is not None:
                        try: d[k] = float(d[k])
                        except Exception: pass
                return jsonify({"status": "ok", "trs": d})
            finally:
                try: conn.close()
                except Exception: pass
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/inventory_mgmt/trs/coa/<int:trs_id>", methods=["GET"])
    @_login_required
    def api_inv_trs_coa(trs_id):
        """List COA attachments for the GRN line behind this TRS.

        QC dashboard's TRS-detail modal calls this when the modal opens.
        Returns the metadata (id, name, mime, size) — the file content
        itself is streamed by the existing /api/inventory_mgmt/grn/file/<id>
        endpoint, which is _login_required too, so QC users can view
        without needing GRN access caps.

        Matching strategy: we resolve COA files by joining
        procurement_grn_files → procurement_grn_items on the LIVE id,
        then filtering by (grn_id, material, batch_num) from the TRS
        row. This is intentionally id-agnostic: GRN /save does a
        DELETE+INSERT on procurement_grn_items, so TRS rows often hold
        a stale grn_item_id even when the file row itself was correctly
        re-linked. Joining by material+batch sidesteps both kinds of
        drift.

        We ALSO opportunistically heal the TRS row: if the TRS's
        grn_item_id doesn't match the live line we just resolved, update
        it so the next caller doesn't have to re-resolve.
        """
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            try:
                trs_row = conn.execute(
                    "SELECT id, grn_id, grn_item_id, material, batch_num "
                    "FROM procurement_grn_trs WHERE id=%s",
                    (trs_id,)
                ).fetchone()
                if not trs_row:
                    return jsonify({"status": "error", "message": "TRS not found"}), 404
                t = dict(trs_row) if hasattr(trs_row, "keys") else trs_row
                stored_iid = t.get("grn_item_id")
                grn_id_i   = t.get("grn_id")
                material   = (t.get("material") or "").strip()
                batch_num  = (t.get("batch_num") or "").strip()

                # Resolve the LIVE line id by material+batch. If the
                # GRN was edited after the TRS was generated, this is
                # the only reliable way to find the current row.
                live_iid = None
                try:
                    live_row = conn.execute("""
                        SELECT id FROM procurement_grn_items
                        WHERE grn_id = %s
                          AND LOWER(TRIM(material)) = LOWER(TRIM(%s))
                          AND COALESCE(TRIM(batch_num),'') = COALESCE(TRIM(%s),'')
                        LIMIT 1
                    """, (grn_id_i, material, batch_num)).fetchone()
                    if live_row:
                        live_iid = live_row.get("id") if hasattr(live_row, "get") else live_row[0]
                except Exception:
                    traceback.print_exc()

                # Opportunistic self-heal: if stored differs from live,
                # update the TRS row so this isn't recomputed every time.
                if live_iid and live_iid != stored_iid:
                    try:
                        conn.execute(
                            "UPDATE procurement_grn_trs "
                            "SET grn_item_id=%s WHERE id=%s",
                            (live_iid, trs_id)
                        )
                        conn.commit()
                        print(f"[InventoryMgmt] TRS COA self-heal: "
                              f"trs_id={trs_id} grn_item_id "
                              f"{stored_iid} → {live_iid}")
                    except Exception:
                        traceback.print_exc()

                # Look up COA files. We try two paths:
                #   (a) Files attached to the LIVE line id (the normal case
                #       after the COA re-link in /grn/save).
                #   (b) Files attached to the STORED (possibly stale) line
                #       id — covers the corner case where files were never
                #       re-linked because they pre-date the re-link patch.
                # We UNION the two and dedupe by file id, so we never
                # show duplicates and never miss attachments.
                files = []
                seen_ids = set()
                def _ingest(rows):
                    for fr in rows or []:
                        d = dict(fr) if hasattr(fr, "keys") else fr
                        fid = d.get("id")
                        if fid in seen_ids: continue
                        seen_ids.add(fid)
                        if d.get("uploaded_at"):
                            d["uploaded_at"] = str(d["uploaded_at"])
                        if d.get("size_bytes") is not None:
                            try: d["size_bytes"] = int(d["size_bytes"])
                            except Exception: pass
                        files.append(d)
                try:
                    if live_iid:
                        _ingest(conn.execute("""
                            SELECT id, original_name, mime_type, size_bytes, uploaded_at
                            FROM procurement_grn_files
                            WHERE kind='coa' AND grn_item_id=%s
                            ORDER BY uploaded_at DESC, id DESC
                        """, (live_iid,)).fetchall())
                    if stored_iid and stored_iid != live_iid:
                        _ingest(conn.execute("""
                            SELECT id, original_name, mime_type, size_bytes, uploaded_at
                            FROM procurement_grn_files
                            WHERE kind='coa' AND grn_item_id=%s
                            ORDER BY uploaded_at DESC, id DESC
                        """, (stored_iid,)).fetchall())
                    # Last-resort fallback: scan COA files across the
                    # whole GRN matched by material+batch via the JOIN.
                    # Catches the case where a COA file row was somehow
                    # orphaned to a totally dead grn_item_id that
                    # doesn't exist anywhere.
                    if not files and material:
                        _ingest(conn.execute("""
                            SELECT f.id, f.original_name, f.mime_type,
                                   f.size_bytes, f.uploaded_at
                            FROM procurement_grn_files f
                            JOIN procurement_grn_items i ON i.id = f.grn_item_id
                            WHERE f.kind='coa'
                              AND i.grn_id = %s
                              AND LOWER(TRIM(i.material)) = LOWER(TRIM(%s))
                              AND COALESCE(TRIM(i.batch_num),'') = COALESCE(TRIM(%s),'')
                            ORDER BY f.uploaded_at DESC, f.id DESC
                        """, (grn_id_i, material, batch_num)).fetchall())
                except Exception:
                    traceback.print_exc()

                return jsonify({
                    "status":      "ok",
                    "trs_id":      trs_id,
                    "grn_id":      grn_id_i,
                    "grn_item_id": live_iid or stored_iid,
                    "material":    material,
                    "batch_num":   batch_num,
                    "files":       files,
                })
            finally:
                try: conn.close()
                except Exception: pass
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/inventory_mgmt/trs/generate", methods=["POST"])
    @_trs_generate_required
    def api_inv_trs_generate():
        """Generate (or re-generate) a TRS for one GRN line.

        Body:
            {
              "grn_item_id":       int   (required),
              "physical_state":    str   (optional — required if missing on GRN line),
              "sample_qty":        float (optional — operator-supplied),
              "previous_supplier": str   (optional),
              "new_or_old":        "NEW"|"OLD" (optional)
            }
        """
        d = request.get_json(silent=True) or {}
        try:
            grn_item_id = int(d.get("grn_item_id") or 0)
        except Exception:
            grn_item_id = 0
        if not grn_item_id:
            return jsonify({"status": "error", "message": "grn_item_id required"}), 400

        physical_state    = (d.get("physical_state") or "").strip()
        sample_qty_raw    = d.get("sample_qty")
        previous_supplier = (d.get("previous_supplier") or "").strip() or None
        new_or_old        = (d.get("new_or_old") or "").strip().upper()
        if new_or_old and new_or_old not in ("NEW", "OLD"):
            new_or_old = None
        try:
            sample_qty = float(sample_qty_raw) if sample_qty_raw not in (None, "") else None
        except Exception:
            sample_qty = None

        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            try:
                # 1) Load the GRN line + parent GRN header in one go.
                line = conn.execute(
                    "SELECT * FROM procurement_grn_items WHERE id=%s",
                    (grn_item_id,)
                ).fetchone()
                if not line:
                    return jsonify({"status": "error",
                                    "message": "GRN line not found"}), 404
                line_d = dict(line) if hasattr(line, "keys") else line
                grn_id_i = int(line_d.get("grn_id") or 0)
                if not grn_id_i:
                    return jsonify({"status": "error",
                                    "message": "Line is not linked to a GRN"}), 400
                grn = conn.execute(
                    "SELECT id, grn_num, grn_date, supplier_name "
                    "FROM procurement_grn WHERE id=%s",
                    (grn_id_i,)
                ).fetchone()
                if not grn:
                    return jsonify({"status": "error",
                                    "message": "Parent GRN not found"}), 404
                grn_d = dict(grn) if hasattr(grn, "keys") else grn

                # 2) Compute the TRS number. Find this line's 1-based
                #    position among all lines on this GRN (stable: ordered
                #    by id). Format: {grn_num}/{position}
                pos_row = conn.execute(
                    "SELECT COUNT(*) AS n FROM procurement_grn_items "
                    "WHERE grn_id=%s AND id<=%s",
                    (grn_id_i, grn_item_id)
                ).fetchone()
                position = int((pos_row.get("n") if hasattr(pos_row, "get") else pos_row[0]) or 1)
                trs_num = f"{grn_d.get('grn_num') or 'GRN'}/{position}"

                # 3) Build the row. Denormalise so the slip stays correct
                #    even if the GRN is later edited.
                # Safe coercion: procurement_grn_items.packages and qty_per_pkg
                # come back as Decimal on most drivers, but legacy rows can
                # occasionally have empty strings or NULL. int('') and
                # float('') both throw, so we coerce via str()/strip() and
                # default to 0 on any parse failure.
                def _safe_int(v):
                    if v is None: return 0
                    s = str(v).strip()
                    if not s: return 0
                    try:    return int(float(s))   # 'float-first' tolerates "30.0"
                    except Exception: return 0
                def _safe_float(v):
                    if v is None: return 0.0
                    s = str(v).strip()
                    if not s: return 0.0
                    try:    return float(s)
                    except Exception: return 0.0
                packages = _safe_int(line_d.get("packages"))
                qpp      = _safe_float(line_d.get("qty_per_pkg"))
                total_q  = round(packages * qpp, 3) if (packages and qpp) else 0.0
                uom      = (line_d.get("uom") or "KG").strip() or "KG"

                # If the line has its own physical_state we don't have a
                # column for it on procurement_grn_items today, so the
                # operator-supplied value is the only source. Same for
                # the other prompted fields. Defaults are NULL.

                verified_by = _user()

                # 4) Upsert: if a TRS already exists for this line, update
                #    it (preserve the original trs_num & generated_at to
                #    keep the slip stable when re-printed). Otherwise
                #    insert a new row.
                existing = conn.execute(
                    "SELECT id, trs_num, generated_at FROM procurement_grn_trs "
                    "WHERE grn_item_id=%s", (grn_item_id,)
                ).fetchone()

                if existing:
                    ex_d = dict(existing) if hasattr(existing, "keys") else existing
                    trs_id  = int(ex_d.get("id"))
                    trs_num = ex_d.get("trs_num") or trs_num
                    conn.execute("""
                        UPDATE procurement_grn_trs SET
                            grn_id            = %s,
                            grn_num           = %s,
                            grn_date          = %s,
                            material          = %s,
                            batch_num         = %s,
                            packages          = %s,
                            qty_per_pkg       = %s,
                            total_qty         = %s,
                            uom               = %s,
                            manufacturer      = %s,
                            mfg_date          = %s,
                            expiry_date       = %s,
                            supplier_name     = %s,
                            physical_state    = %s,
                            sample_qty        = %s,
                            previous_supplier = %s,
                            new_or_old        = %s,
                            verified_by       = %s
                        WHERE id=%s
                    """, (
                        grn_id_i, grn_d.get("grn_num"), grn_d.get("grn_date"),
                        line_d.get("material"), line_d.get("batch_num"),
                        packages, qpp, total_q, uom,
                        line_d.get("manufacturer"),
                        line_d.get("mfg_date"), line_d.get("expiry_date"),
                        grn_d.get("supplier_name"),
                        physical_state or None, sample_qty,
                        previous_supplier, new_or_old, verified_by,
                        trs_id,
                    ))
                else:
                    cur = conn.execute("""
                        INSERT INTO procurement_grn_trs (
                            trs_num, grn_id, grn_item_id,
                            grn_num, grn_date,
                            material, batch_num,
                            packages, qty_per_pkg, total_qty, uom,
                            manufacturer, mfg_date, expiry_date,
                            supplier_name,
                            physical_state, sample_qty,
                            previous_supplier, new_or_old,
                            generated_by, verified_by
                        ) VALUES (
                            %s, %s, %s,
                            %s, %s,
                            %s, %s,
                            %s, %s, %s, %s,
                            %s, %s, %s,
                            %s,
                            %s, %s,
                            %s, %s,
                            %s, %s
                        )
                    """, (
                        trs_num, grn_id_i, grn_item_id,
                        grn_d.get("grn_num"), grn_d.get("grn_date"),
                        line_d.get("material"), line_d.get("batch_num"),
                        packages, qpp, total_q, uom,
                        line_d.get("manufacturer"),
                        line_d.get("mfg_date"), line_d.get("expiry_date"),
                        grn_d.get("supplier_name"),
                        physical_state or None, sample_qty,
                        previous_supplier, new_or_old,
                        verified_by, verified_by,
                    ))
                    # Get the new id back
                    nrow = conn.execute(
                        "SELECT id FROM procurement_grn_trs WHERE grn_item_id=%s",
                        (grn_item_id,)
                    ).fetchone()
                    trs_id = int((nrow.get("id") if hasattr(nrow, "get") else nrow[0]) or 0)

                conn.commit()

                # 5) Read back the canonical row so the client can render
                #    the print preview without a second round-trip.
                row = conn.execute(
                    "SELECT * FROM procurement_grn_trs WHERE id=%s", (trs_id,)
                ).fetchone()
                trs = dict(row) if hasattr(row, "keys") else row
                for k in ("generated_at", "approval_dt"):
                    if trs.get(k):
                        try:
                            trs[k] = trs[k].isoformat(sep=" ", timespec="seconds")
                        except Exception:
                            trs[k] = str(trs[k])
                for k in ("grn_date", "mfg_date", "expiry_date"):
                    if trs.get(k):
                        try:
                            trs[k] = trs[k].isoformat()
                        except Exception:
                            trs[k] = str(trs[k])
                for k in ("qty_per_pkg", "total_qty", "sample_qty"):
                    if trs.get(k) is not None:
                        try: trs[k] = float(trs[k])
                        except Exception: pass

                return jsonify({
                    "status":  "ok",
                    "trs":     trs,
                    "trs_num": trs_num,
                    "trs_id":  trs_id,
                    "was_update": bool(existing),
                })
            finally:
                try: conn.close()
                except Exception: pass
        except Exception as e:
            # Surface the real reason in the JSON response so the toast
            # the operator sees ("Network error: ...") is replaced by
            # the actual server-side cause ("TRS generate failed: …").
            # traceback.print_exc() still writes the full stack to Flask
            # logs for admins to grab.
            traceback.print_exc()
            tb_lines = traceback.format_exc().splitlines()
            last = tb_lines[-1] if tb_lines else str(e)
            return jsonify({
                "status":     "error",
                "message":    f"TRS generate failed: {last}",
                "error_type": type(e).__name__,
            }), 500

    # ═══════════════════════════════════════════════════════════════════
    # TRS Register — read-only audit grid for the inventory team
    # ═══════════════════════════════════════════════════════════════════
    #
    # Distinct from the QC Dashboard tab: the QC tab is *actionable*
    # (Approve / Reject), the register is *informational* (track what
    # we've generated and where QC is on each one). Inventory operators
    # who don't have QC role still need this view to follow up.
    #
    # Permission: any logged-in user with the 'grn' cap (view) can see
    # this page. Admins always pass.
    @app.route("/trs_register")
    @_login_required
    def trs_register_page():
        # Gating: open to anyone with ANY inventory presence (any cap
        # turned on) OR to QC-role users (they need to audit TRS state).
        # Admins always pass. The page is read-only — list view of all
        # generated TRS slips with status — so the bar is intentionally
        # low. The actionable bits (Approve/Reject, Generate) live on
        # the qc dashboard and the GRN form respectively.
        try:
            try:
                from .inventory_access import _resolve_user_access
            except Exception:
                from inventory_access import _resolve_user_access
            resolved = _resolve_user_access() or {}
            is_admin = bool(resolved.get('is_admin'))
            access   = resolved.get('access') or {}
            # Any access toggle != 'off' counts as "this user works in
            # inventory" and is allowed to see the register.
            has_any_inv = any(
                str(v or 'off').lower() not in ('', 'off', '0', 'false', 'no')
                for v in access.values()
            )
            # QC-role users can audit TRS state without needing a row
            # in inventory_user_access.
            role = (session.get('User_Type') or '').lower()
            is_qc = role in ('qc', 'qc_common', 'purchase')
            allowed = is_admin or has_any_inv or is_qc
        except Exception:
            is_admin = False
            allowed  = False
        if not allowed:
            return render_template_string(
                "<div style='padding:40px;font-family:Inter,sans-serif;"
                "color:#374151;text-align:center'>"
                "<h2 style='color:#dc2626'>Access Denied</h2>"
                "<p>You don't have permission to view the TRS Register.</p>"
                "<p>Ask an admin to grant you any inventory access "
                "(e.g. GRN view) or assign you to a QC role.</p>"
                "<p><a href='/inventory_mgmt' style='color:#4338ca'>"
                "← Back to Inventory</a></p></div>"
            ), 403
        return render_template(
            "inventory/trs_register.html",
            user_name=session.get("User_Name"),
            role=session.get("User_Type"),
            is_admin=is_admin,
        )

    @app.route("/api/inventory_mgmt/trs/register_list", methods=["GET"])
    @_login_required
    def api_inv_trs_register_list():
        """All-TRS list endpoint for the register page. Same data the QC
        dashboard /api/qc/trs/list returns, but here it's gated on broad
        inventory/QC access (read-only audit grid)."""
        try:
            try:
                from .inventory_access import _resolve_user_access
            except Exception:
                from inventory_access import _resolve_user_access
            resolved = _resolve_user_access() or {}
            is_admin = bool(resolved.get('is_admin'))
            access   = resolved.get('access') or {}
            has_any_inv = any(
                str(v or 'off').lower() not in ('', 'off', '0', 'false', 'no')
                for v in access.values()
            )
            role = (session.get('User_Type') or '').lower()
            is_qc = role in ('qc', 'qc_common', 'purchase')
            allowed = is_admin or has_any_inv or is_qc
        except Exception:
            allowed = False
        if not allowed:
            return jsonify({"status": "error", "message": "Access denied"}), 403
        try:
            conn = sampling_portal.get_db_connection()
            try:
                rows = conn.execute("""
                    SELECT id, trs_num, grn_id, grn_item_id,
                           grn_num, grn_date,
                           material, batch_num,
                           packages, qty_per_pkg, total_qty, uom,
                           manufacturer, mfg_date, expiry_date,
                           supplier_name, physical_state, sample_qty,
                           previous_supplier, new_or_old,
                           generated_by, generated_at, verified_by,
                           approval_status, approved_by, approval_dt,
                           approval_remarks, approval_locked_at,
                           checked_params, rejection_reason
                    FROM procurement_grn_trs
                    ORDER BY generated_at DESC, id DESC
                """).fetchall() or []
                # Compute "now" in IST once.
                try:
                    from datetime import datetime as _dt
                    # Match qc_routes._ist_now_str format
                    now_str = (datetime.utcnow() + timedelta(hours=5, minutes=30)).strftime('%Y-%m-%d %H:%M:%S')
                    now_ist = _dt.strptime(now_str, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    now_ist = None
                out = []
                import json as _json
                for r in rows:
                    d = dict(r) if hasattr(r, "keys") else r
                    # Date/datetime normalization
                    for k in ('generated_at', 'approval_dt', 'approval_locked_at'):
                        if d.get(k):
                            try:
                                d[k] = d[k].isoformat(sep=' ', timespec='seconds')
                            except Exception:
                                d[k] = str(d[k])
                    for k in ('grn_date', 'mfg_date', 'expiry_date'):
                        if d.get(k):
                            try:
                                d[k] = d[k].isoformat()
                            except Exception:
                                d[k] = str(d[k])
                    for k in ('qty_per_pkg', 'total_qty', 'sample_qty'):
                        if d.get(k) is not None:
                            try: d[k] = float(d[k])
                            except Exception: pass
                    # checked_params JSON → list
                    cp = d.get('checked_params')
                    if cp:
                        try:
                            parsed = _json.loads(cp)
                            d['checked_params'] = parsed if isinstance(parsed, list) else [str(parsed)]
                        except Exception:
                            d['checked_params'] = [s.strip() for s in
                                                   str(cp).replace(';', ',').split(',') if s.strip()]
                    else:
                        d['checked_params'] = []
                    d['is_locked'] = False
                    d['hours_remaining'] = None
                    locked_at_str = d.get('approval_locked_at')
                    if locked_at_str and now_ist:
                        try:
                            from datetime import datetime as _dt2
                            lk = _dt2.strptime(locked_at_str[:19], '%Y-%m-%d %H:%M:%S')
                            hrs = (now_ist - lk).total_seconds() / 3600.0
                            d['is_locked'] = hrs >= 24.0
                            d['hours_remaining'] = max(0.0, 24.0 - hrs)
                        except Exception:
                            pass
                    d['new_supplier'] = d.get('supplier_name') if d.get('new_or_old') == 'NEW' else None
                    if not d.get('approval_status'):
                        d['approval_status'] = 'Pending'
                    out.append(d)
                return jsonify({"status": "ok", "trs": out})
            finally:
                try: conn.close()
                except Exception: pass
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    # Restore the real print and show a single clean summary line.
    _builtins.print = _inv_real_print
    print("✅ Inventory Management module registered  →  /inventory_mgmt, /api/inventory_mgmt/*")
