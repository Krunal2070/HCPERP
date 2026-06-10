r"""
inventory_godown.py  —  RM Godown-Wise Stock Management (Phase 1)
==================================================================
HCP Wellness Pvt Ltd

Phase 1 scope (this file):
  • Auto-creates rm_boxes + rm_box_movements tables on first run
  • Godown master CRUD (procurement_godowns — shared with PM/FG)
  • Three-pane Godown View data feeds (godowns → items → boxes)
  • Box history lookup (for QR scan results)
  • Dev/seed endpoint for testing without Phase 2

Phase 2+ (not in this file yet):
  • RM GRN flow that auto-generates boxes
  • Label printing (PDF with QR codes)
  • Material Transfer Voucher (MTV) flow
  • Consumption flow
  • Physical audit

Box code format (NEW — short codes, May 2026):
  RM-{LETTER(S)}{7-DIGITS}
  e.g.  RM-A0000001 (1-letter prefix, 11 chars total)
        ...
        RM-Z9999999
        RM-AA0000001 (2-letter prefix, 12 chars total)
        ...
        RM-AAA0000001 (3-letter prefix, 13 chars total)

  Capacity:
    1-letter: 9,999,999 × 26 = ~260 million codes
    2-letter: 9,999,999 × 676 = ~6.76 billion codes
    3-letter: ~175 billion codes
  Stored in rm_boxes.box_code as VARCHAR(50) — plenty of headroom.

  Codes are assigned sequentially as boxes are created (one global
  counter per install). A box scanner can identify the box uniquely
  from its short code alone — no material/godown info embedded.

Legacy format (still accepted by the scanner regex for backwards
compatibility — boxes created via the previous scheme stay scannable):
  RM-MATCODE-G####-B###    (GRN-sourced)
  RM-MATCODE-OP####-B###   (Opening-stock-sourced)

Register in app.py AFTER inventory_mgmt:
    import inventory_godown
    inventory_godown.register_inventory_godown(app)

API surface: /api/inventory_godown/*
"""

from __future__ import annotations

import json as _json
import re
import traceback
from functools import wraps

from flask import jsonify, request, session

import sampling_portal


# ─────────────────────────────────────────────────────────────────────────────
# ACCESS CONTROL  (mirrors inventory_mgmt.py policy)
# ─────────────────────────────────────────────────────────────────────────────

def _can_view() -> bool:
    return bool(session.get("logged_in"))


def _can_edit() -> bool:
    if not session.get("logged_in"):
        return False
    role = (session.get("User_Type") or "").strip().lower()
    uid  = (session.get("UID")       or "").strip().lower()
    return role in {"admin"} or uid in {"sonal", "tarak"}


def _login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        return f(*args, **kwargs)
    return wrapper


def _edit_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        if not _can_edit():
            return jsonify({"status": "error", "message": "Permission denied"}), 403
        return f(*args, **kwargs)
    return wrapper


def _has_cap(cap_key: str) -> bool:
    """True if the current user is admin/edit-role, OR has the given access cap
    set to a non-off value. Used by per-feature route decorators so a user with
    'opening_stock' toggled ON can call opening-stock endpoints even though
    they're not an admin."""
    if _can_edit():
        return True
    try:
        try:
            from .inventory_access import _inv_user_has_access
        except Exception:
            from inventory_access import _inv_user_has_access
        return bool(_inv_user_has_access(cap_key))
    except Exception:
        return False


def _cap_required(cap_key: str):
    """Route decorator that allows either edit-role users OR users whose
    User Access Control toggle for `cap_key` is ON."""
    def deco(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not session.get("logged_in"):
                return jsonify({"status": "error", "message": "Not logged in"}), 401
            if not _has_cap(cap_key):
                return jsonify({"status": "error", "message": "Permission denied"}), 403
            return f(*args, **kwargs)
        return wrapper
    return deco


def _user() -> str:
    return session.get("User_Name") or session.get("UID") or "Unknown"


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

# Box code regex — accepts BOTH:
#   NEW short-code:  RM-{1-3 letters}{7 digits}      e.g. RM-A0000001, RM-AB0000045
#   LEGACY:          RM-MATCODE-G####-B###            e.g. RM-BLACKAMBE-G0234-B003
#                    RM-MATCODE-OP####-B###           e.g. RM-BLACKAMBE-OP0001-B001
# Both formats stay scannable in the field. New boxes use short codes.
_RM_BOX_CODE_RE = re.compile(
    r"^RM-("
    r"[A-Z]{1,3}\d{7}"                          # new short-code
    r"|"
    r"[A-Z0-9]{1,10}-(?:G|OP)\d{3,5}-B\d{2,4}"  # legacy
    r")$",
    re.IGNORECASE,
)


def derive_mat_code(material_name: str) -> str:
    """
    Legacy helper — derive a short alphanumeric code from a material name.
    Kept for backwards compatibility with the old box-code scheme. The new
    short-code scheme (allocate_next_box_code) doesn't use this.
    """
    if not material_name:
        return "MAT"
    cleaned = re.sub(r"[^A-Za-z0-9]", "", material_name).upper()
    return (cleaned[:10] or "MAT")


def _seq_to_short_code(seq: int) -> str:
    """
    Convert a sequential integer to a short-code suffix.

      seq 1         -> 'A0000001'
      seq 2         -> 'A0000002'
      seq 9,999,999 -> 'A9999999'
      seq 10,000,000-> 'B0000001'
      seq 9,999,999 * 26 -> 'Z9999999'
      next          -> 'AA0000001'
      ... grows to 'ZZ9999999' (2 letters, ~6.76 billion)
      then          -> 'AAA0000001' (3 letters, ~175 billion)

    The prefix letter(s) form a base-26 counter (A..Z, AA..ZZ, AAA..ZZZ),
    and the 7-digit numeric tail counts from 1 to 9,999,999 within each
    prefix slot.
    """
    if seq < 1:
        seq = 1
    PER_LETTER = 9_999_999
    # Determine how many letters we need and which letter-block we're in
    one_letter_cap  = PER_LETTER * 26              # ~260 million
    two_letter_cap  = PER_LETTER * 26 * 26         # ~6.76 billion
    three_letter_cap = PER_LETTER * 26 * 26 * 26   # ~175 billion

    if seq <= one_letter_cap:
        letter_idx = (seq - 1) // PER_LETTER       # 0..25
        num        = ((seq - 1) %  PER_LETTER) + 1 # 1..9_999_999
        prefix     = chr(ord('A') + letter_idx)
        return f"{prefix}{num:07d}"
    elif seq <= two_letter_cap:
        adj = seq - one_letter_cap                 # 1-based within 2-letter range
        letter_idx = (adj - 1) // PER_LETTER       # 0..675
        num        = ((adj - 1) %  PER_LETTER) + 1
        l1 = letter_idx // 26
        l2 = letter_idx %  26
        prefix = chr(ord('A') + l1) + chr(ord('A') + l2)
        return f"{prefix}{num:07d}"
    elif seq <= three_letter_cap:
        adj = seq - two_letter_cap
        letter_idx = (adj - 1) // PER_LETTER       # 0..17575
        num        = ((adj - 1) %  PER_LETTER) + 1
        l1 = letter_idx // (26 * 26)
        l2 = (letter_idx // 26) % 26
        l3 = letter_idx %  26
        prefix = chr(ord('A') + l1) + chr(ord('A') + l2) + chr(ord('A') + l3)
        return f"{prefix}{num:07d}"
    else:
        # Beyond ~175 billion — fall back to a numeric overflow code so
        # things don't crash. Realistically we'll never hit this.
        return f"OVERFLOW{seq}"


def allocate_next_box_code(conn) -> str:
    """
    Atomically allocate the next sequential box code.

    Strategy: keep a simple counter in a `rm_box_code_seq` table with one
    row {id:1, next_seq:N}. We UPDATE it with `next_seq = next_seq + 1`,
    capture the previous value, and use that for the new box.

    This is concurrency-safe: the UPDATE statement is atomic, and we read
    back the just-incremented value via LAST_INSERT_ID() trick so multiple
    Flask workers can't collide on the same seq.
    """
    # Ensure table exists (idempotent — no-op after first call)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS rm_box_code_seq (
            id        INT PRIMARY KEY,
            next_seq  BIGINT NOT NULL DEFAULT 1
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    conn.execute(
        "INSERT IGNORE INTO rm_box_code_seq (id, next_seq) VALUES (1, 1)"
    )
    # Atomic increment + read using MySQL's LAST_INSERT_ID() trick:
    #   UPDATE ... SET next_seq = LAST_INSERT_ID(next_seq + 1)
    # then SELECT LAST_INSERT_ID() returns the post-update value.
    conn.execute(
        "UPDATE rm_box_code_seq "
        "SET next_seq = LAST_INSERT_ID(next_seq + 1) "
        "WHERE id = 1"
    )
    row = conn.execute("SELECT LAST_INSERT_ID() AS v").fetchone()
    next_after = int(row["v"] if isinstance(row, dict) else row[0])
    # next_after is the post-increment value; the seq we allocated is one less
    allocated_seq = next_after - 1
    if allocated_seq < 1:
        allocated_seq = 1
    return "RM-" + _seq_to_short_code(allocated_seq)


# ─────────────────────────────────────────────────────────────────────────────
# SCHEMA BOOTSTRAP
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_schema():
    """Create rm_boxes + rm_box_movements tables if they don't exist."""
    conn = sampling_portal.get_db_connection()
    if not conn:
        print("⚠️  [InventoryGodown] DB unavailable — schema setup skipped")
        return
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS rm_boxes (
                box_id              INT AUTO_INCREMENT PRIMARY KEY,
                box_code            VARCHAR(50) NOT NULL UNIQUE,
                grn_id              INT DEFAULT NULL,
                grn_no              VARCHAR(50) DEFAULT NULL,
                grn_item_id         INT DEFAULT NULL,
                material_id         INT NOT NULL,
                material_code       VARCHAR(15) DEFAULT NULL,
                box_seq             INT NOT NULL,
                total_boxes         INT NOT NULL,
                per_box_qty         DECIMAL(14,3) NOT NULL DEFAULT 0,
                uom                 VARCHAR(20) DEFAULT NULL,
                current_godown_id   INT DEFAULT NULL,
                current_status      ENUM('in_stock','in_transit','consumed','damaged','lost','cancelled')
                                      NOT NULL DEFAULT 'in_stock',
                source              ENUM('grn','opening','adjustment') NOT NULL DEFAULT 'grn',
                created_at          DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at          DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                created_by          VARCHAR(50) DEFAULT NULL,
                INDEX ix_rm_boxes_grn      (grn_id),
                INDEX ix_rm_boxes_material (material_id),
                INDEX ix_rm_boxes_godown   (current_godown_id, current_status),
                INDEX ix_rm_boxes_status   (current_status)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS rm_box_movements (
                movement_id     INT AUTO_INCREMENT PRIMARY KEY,
                box_id          INT NOT NULL,
                transfer_id     INT DEFAULT NULL,
                movement_type   ENUM('grn_create','opening','out','in','consume','adjust','cancel')
                                  NOT NULL,
                from_godown_id  INT DEFAULT NULL,
                to_godown_id    INT DEFAULT NULL,
                qty             DECIMAL(14,3) NOT NULL DEFAULT 0,
                movement_at     DATETIME DEFAULT CURRENT_TIMESTAMP,
                moved_by        VARCHAR(50) DEFAULT NULL,
                remarks         VARCHAR(255) DEFAULT NULL,
                INDEX ix_rm_boxmv_box      (box_id),
                INDEX ix_rm_boxmv_transfer (transfer_id),
                INDEX ix_rm_boxmv_at       (movement_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # NOTE: The rm_stock_transfers and rm_stock_transfer_boxes tables
        # are created by inventory/inventory_transfers.py (the new two-stage
        # voucher module). Schema migration there handles upgrading legacy
        # single-stage tables in-place.
        # ── Auto-migration: ensure procurement_grn has new columns added by
        # the May 2026 GRN feature additions (other_charge, supervisor_name,
        # other_details JSON, unload_checklist JSON). Adds them only if
        # missing — safe to run on every startup.
        try:
            cols = conn.execute("SHOW COLUMNS FROM procurement_grn").fetchall()
            col_names = {(c['Field'] if isinstance(c, dict) else c[0]).lower() for c in cols}
            adds = []
            if 'other_charge' not in col_names:
                adds.append("ADD COLUMN other_charge DECIMAL(14,2) DEFAULT NULL AFTER packing_charge")
            if 'other_charge_label' not in col_names:
                adds.append("ADD COLUMN other_charge_label VARCHAR(60) DEFAULT NULL AFTER other_charge")
            if 'supervisor_name' not in col_names:
                adds.append("ADD COLUMN supervisor_name VARCHAR(120) DEFAULT NULL AFTER supplier_name")
            if 'other_details' not in col_names:
                adds.append("ADD COLUMN other_details JSON DEFAULT NULL")
            if 'unload_checklist' not in col_names:
                adds.append("ADD COLUMN unload_checklist JSON DEFAULT NULL")
            if adds:
                conn.execute("ALTER TABLE procurement_grn " + ", ".join(adds))
                print(f"✅ [InventoryGodown] added {len(adds)} column(s) to procurement_grn")
        except Exception as ex:
            # procurement_grn may not exist in some installs; or DB may not
            # support JSON (MySQL <5.7 / MariaDB <10.2). Skip silently.
            print(f"ℹ️  [InventoryGodown] skip procurement_grn migration: {ex}")

        # Add manufacturer column to procurement_grn_items
        try:
            icols = conn.execute("SHOW COLUMNS FROM procurement_grn_items").fetchall()
            iname_set = {(c['Field'] if isinstance(c, dict) else c[0]).lower() for c in icols}
            if 'manufacturer' not in iname_set:
                conn.execute("ALTER TABLE procurement_grn_items ADD COLUMN manufacturer VARCHAR(200) DEFAULT NULL")
                print("✅ [InventoryGodown] added procurement_grn_items.manufacturer column")
        except Exception as ex:
            print(f"ℹ️  [InventoryGodown] skip procurement_grn_items migration: {ex}")

        # ── Auto-migration: ensure rm_boxes.current_status ENUM includes
        # 'cancelled'. The opening-stock re-upload feature uses this value
        # to soft-delete the previous upload's rows. Pre-May-16-2026 DBs
        # were created without 'cancelled' in the enum and would throw
        # "Data truncated for column 'current_status'" on re-upload.
        # MODIFY COLUMN is idempotent — re-running has no effect.
        try:
            conn.execute("""
                ALTER TABLE rm_boxes
                MODIFY COLUMN current_status
                ENUM('in_stock','in_transit','consumed','damaged','lost','cancelled')
                NOT NULL DEFAULT 'in_stock'
            """)
            # No print on success — runs every startup, would be noisy
        except Exception as ex:
            print(f"ℹ️  [InventoryGodown] skip rm_boxes.current_status migration: {ex}")

        # Opening-stock boxes carry batch / expiry / manufacturer directly on
        # rm_boxes (GRN boxes get them via grn_item_id). These columns were
        # historically added lazily only when opening stock was first used, so
        # on an install that never touched opening stock they could be missing —
        # which made expiry reports/queries that read b.expiry_date crash with
        # "Unknown column". Ensure they exist at startup so every consumer is
        # safe. Each ADD is guarded individually (ADD COLUMN isn't idempotent).
        for _col, _ddl in (
            ("batch_num",    "ALTER TABLE rm_boxes ADD COLUMN batch_num VARCHAR(64) DEFAULT NULL"),
            ("expiry_date",  "ALTER TABLE rm_boxes ADD COLUMN expiry_date DATE DEFAULT NULL"),
            ("manufacturer", "ALTER TABLE rm_boxes ADD COLUMN manufacturer VARCHAR(128) DEFAULT NULL"),
        ):
            try:
                have = {(c["Field"] if hasattr(c, "get") else c[0])
                        for c in conn.execute("SHOW COLUMNS FROM rm_boxes").fetchall()}
                if _col not in have:
                    conn.execute(_ddl)
            except Exception as ex:
                print(f"ℹ️  [InventoryGodown] skip rm_boxes.{_col} migration: {ex}")

        conn.commit()
        print("✅ [InventoryGodown] rm_boxes & rm_box_movements ready")
    except Exception as e:
        print(f"⚠️  [InventoryGodown] schema setup error: {e}")
        traceback.print_exc()
    finally:
        try: conn.close()
        except Exception: pass


# ─────────────────────────────────────────────────────────────────────────────
# REGISTRATION
# ─────────────────────────────────────────────────────────────────────────────

def register_inventory_godown(app):
    """Register all routes onto the Flask app."""

    # Run schema setup at registration time
    _ensure_schema()

    # ════════════════════════════════════════════════════════════════════════
    # GODOWN MASTER CRUD
    # ════════════════════════════════════════════════════════════════════════

    @app.route("/api/inventory_godown/godowns/list", methods=["GET"])
    @_login_required
    def api_godown_list():
        """
        List all godowns + RM stock summary (boxes/qty per godown).
        """
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            godowns = conn.execute("""
                SELECT id, name, type, address, contact, phone, email,
                       is_default, gst_number, city, state, pin,
                       created_at, updated_at
                FROM procurement_godowns
                ORDER BY is_default DESC, name ASC
            """).fetchall()

            # Per-godown stock summary
            sums = {}
            try:
                rows = conn.execute("""
                    SELECT current_godown_id          AS gid,
                           COUNT(DISTINCT material_id) AS distinct_items,
                           COUNT(*)                    AS total_boxes,
                           COALESCE(SUM(per_box_qty),0) AS total_qty
                    FROM rm_boxes
                    WHERE current_status='in_stock'
                    GROUP BY current_godown_id
                """).fetchall()
                for r in rows:
                    gid = r["gid"]
                    if gid is None: continue
                    sums[int(gid)] = {
                        "distinct_items": int(r["distinct_items"] or 0),
                        "total_boxes":    int(r["total_boxes"] or 0),
                        "total_qty":      float(r["total_qty"] or 0),
                    }
            except Exception:
                pass  # tables fresh, no rows yet

            out = []
            for g in godowns:
                d = dict(g)
                s = sums.get(int(d["id"]), {"distinct_items": 0, "total_boxes": 0, "total_qty": 0.0})
                d["distinct_items"] = s["distinct_items"]
                d["total_boxes"]    = s["total_boxes"]
                d["total_qty"]      = s["total_qty"]
                for k in ("created_at", "updated_at"):
                    if d.get(k) is not None:
                        d[k] = str(d[k])
                out.append(d)

            return jsonify({"status": "ok", "godowns": out, "total": len(out)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    @app.route("/api/inventory_godown/godowns/save", methods=["POST"])
    @_cap_required("manage_godown")
    def api_godown_save():
        """Create or update a godown."""
        data = request.get_json(silent=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"status": "error", "message": "Name is required"}), 400

        gid = data.get("id")
        try:
            gid = int(gid) if gid not in (None, "", 0) else None
        except (TypeError, ValueError):
            gid = None

        gtype       = (data.get("type") or "godown").strip().lower()[:20]
        address     = (data.get("address") or "").strip()
        contact     = (data.get("contact") or "").strip()[:300]
        phone       = (data.get("phone") or "").strip()[:100]
        email       = (data.get("email") or "").strip()[:300]
        gst_number  = (data.get("gst_number") or "").strip()[:50]
        is_default  = 1 if int(data.get("is_default") or 0) else 0
        city        = (data.get("city") or "").strip()[:100]
        state       = (data.get("state") or "").strip()[:100]
        pin         = (data.get("pin") or "").strip()[:10]

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            if is_default:
                conn.execute("UPDATE procurement_godowns SET is_default=0")

            if gid:
                conn.execute("""
                    UPDATE procurement_godowns SET
                      name=%s, type=%s, address=%s, contact=%s, phone=%s, email=%s,
                      gst_number=%s, is_default=%s, city=%s, state=%s, pin=%s
                    WHERE id=%s
                """, (name, gtype, address, contact, phone, email,
                      gst_number, is_default, city, state, pin, gid))
                conn.commit()
                return jsonify({"status": "ok", "id": gid, "action": "updated"})
            else:
                dup = conn.execute(
                    "SELECT id FROM procurement_godowns WHERE LOWER(name)=LOWER(%s) LIMIT 1",
                    (name,)
                ).fetchone()
                if dup:
                    return jsonify({"status": "error",
                                    "message": f"A godown named '{name}' already exists"}), 400
                cur = conn.execute("""
                    INSERT INTO procurement_godowns
                      (name, type, address, contact, phone, email,
                       gst_number, is_default, city, state, pin)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (name, gtype, address, contact, phone, email,
                      gst_number, is_default, city, state, pin))
                conn.commit()
                return jsonify({"status": "ok", "id": cur.lastrowid, "action": "created"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    @app.route("/api/inventory_godown/godowns/delete", methods=["POST"])
    @_cap_required("manage_godown")
    def api_godown_delete():
        """Delete a godown — blocked if boxes still parked there, or if it's the default."""
        data = request.get_json(silent=True) or {}
        try:
            gid = int(data.get("id") or 0)
        except (TypeError, ValueError):
            gid = 0
        if not gid:
            return jsonify({"status": "error", "message": "id is required"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            row = conn.execute(
                "SELECT COUNT(*) AS n FROM rm_boxes WHERE current_godown_id=%s AND current_status='in_stock'",
                (gid,)
            ).fetchone()
            if row and int(row["n"] or 0) > 0:
                return jsonify({
                    "status": "error",
                    "message": f"Cannot delete — {int(row['n'])} RM package(s) are still in this godown. Transfer them out first."
                }), 400

            row = conn.execute(
                "SELECT is_default, name FROM procurement_godowns WHERE id=%s",
                (gid,)
            ).fetchone()
            if not row:
                return jsonify({"status": "error", "message": "Godown not found"}), 404
            if int(row["is_default"] or 0):
                return jsonify({
                    "status": "error",
                    "message": "Cannot delete the default godown. Mark another godown as default first."
                }), 400

            conn.execute("DELETE FROM procurement_godowns WHERE id=%s", (gid,))
            conn.commit()
            return jsonify({"status": "ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ════════════════════════════════════════════════════════════════════════
    # GODOWN VIEW DATA FEEDS  (three-pane drill-down)
    # ════════════════════════════════════════════════════════════════════════

    @app.route("/api/inventory_godown/items_at", methods=["GET"])
    @_login_required
    def api_items_at():
        """
        Pane 2 data: RM items at a godown.
        Query: ?godown_id=12  (0 = all godowns aggregated)
               ?status=in_stock|in_transit|all   (default in_stock)
        """
        try:
            gid = int(request.args.get("godown_id") or 0)
        except (TypeError, ValueError):
            gid = 0
        status = (request.args.get("status") or "in_stock").strip().lower()
        if status not in {"in_stock", "in_transit", "consumed", "damaged", "lost", "all"}:
            status = "in_stock"

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            where, params = [], []
            if gid:
                where.append("b.current_godown_id = %s")
                params.append(gid)
            if status != "all":
                where.append("b.current_status = %s")
                params.append(status)
            sql_where = (" WHERE " + " AND ".join(where)) if where else ""

            rows = conn.execute(f"""
                SELECT  b.material_id,
                        m.material_name,
                        m.uom,
                        COALESCE(g.group_name, '') AS group_name,
                        COUNT(b.box_id)            AS box_count,
                        COALESCE(SUM(b.per_box_qty),0) AS total_qty
                FROM rm_boxes b
                LEFT JOIN procurement_materials m  ON m.id = b.material_id
                LEFT JOIN procurement_material_groups g ON g.id = m.group_id
                {sql_where}
                GROUP BY b.material_id, m.material_name, m.uom, g.group_name
                ORDER BY m.material_name ASC
            """, params).fetchall()

            items = [{
                "material_id":   int(r["material_id"]),
                "material_name": r["material_name"] or "",
                "group_name":    r["group_name"] or "",
                "uom":           r["uom"] or "",
                "box_count":     int(r["box_count"] or 0),
                "total_qty":     float(r["total_qty"] or 0),
            } for r in rows]

            return jsonify({"status": "ok", "godown_id": gid,
                            "status_filter": status,
                            "items": items, "total": len(items)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    @app.route("/api/inventory_godown/boxes_at", methods=["GET"])
    @_login_required
    def api_boxes_at():
        """
        Pane 3 data: individual RM boxes for (godown, material).
        Query: ?godown_id=12&material_id=87&status=in_stock&limit=500
        """
        try:
            gid = int(request.args.get("godown_id") or 0)
        except (TypeError, ValueError):
            gid = 0
        try:
            mid = int(request.args.get("material_id") or 0)
        except (TypeError, ValueError):
            mid = 0
        try:
            limit = max(1, min(int(request.args.get("limit") or 500), 2000))
        except (TypeError, ValueError):
            limit = 500
        status = (request.args.get("status") or "in_stock").strip().lower()
        if status not in {"in_stock", "in_transit", "consumed", "damaged", "lost", "all"}:
            status = "in_stock"

        if not mid:
            return jsonify({"status": "error", "message": "material_id is required"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            where  = ["b.material_id = %s"]
            params = [mid]
            if gid:
                where.append("b.current_godown_id = %s")
                params.append(gid)
            if status != "all":
                where.append("b.current_status = %s")
                params.append(status)

            rows = conn.execute(f"""
                SELECT  b.box_id, b.box_code, b.grn_no, b.box_seq, b.total_boxes,
                        b.per_box_qty, b.uom, b.current_godown_id, b.current_status,
                        b.source, b.created_at, b.updated_at,
                        COALESCE(g.name,'') AS godown_name,
                        COALESCE(g.type,'') AS godown_type
                FROM rm_boxes b
                LEFT JOIN procurement_godowns g ON g.id = b.current_godown_id
                WHERE {' AND '.join(where)}
                ORDER BY b.box_id DESC
                LIMIT %s
            """, params + [limit]).fetchall()

            boxes = []
            for r in rows:
                d = dict(r)
                d["per_box_qty"] = float(d.get("per_box_qty") or 0)
                for k in ("created_at", "updated_at"):
                    if d.get(k) is not None:
                        d[k] = str(d[k])
                boxes.append(d)

            return jsonify({"status": "ok", "boxes": boxes, "total": len(boxes)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ════════════════════════════════════════════════════════════════════════
    # BOX HISTORY  (QR scan target)
    # ════════════════════════════════════════════════════════════════════════

    @app.route("/api/inventory_godown/box_history", methods=["GET"])
    @_login_required
    def api_box_history():
        """
        Full life history of one RM box.
        Query: ?box_id=N  OR  ?code=RM-MATCODE-G0234-B003
        """
        code = (request.args.get("code") or "").strip().upper()
        try:
            box_id = int(request.args.get("box_id") or 0)
        except (TypeError, ValueError):
            box_id = 0

        if not box_id and not code:
            return jsonify({"status": "error",
                            "message": "Provide box_id or code"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            if box_id:
                box_row = conn.execute("""
                    SELECT b.*,
                           m.material_name, m.uom AS m_uom,
                           COALESCE(g.name,'')    AS current_godown_name,
                           COALESCE(g.type,'')    AS current_godown_type
                    FROM rm_boxes b
                    LEFT JOIN procurement_materials m  ON m.id = b.material_id
                    LEFT JOIN procurement_godowns g    ON g.id = b.current_godown_id
                    WHERE b.box_id = %s
                    LIMIT 1
                """, (box_id,)).fetchone()
            else:
                box_row = conn.execute("""
                    SELECT b.*,
                           m.material_name, m.uom AS m_uom,
                           COALESCE(g.name,'')    AS current_godown_name,
                           COALESCE(g.type,'')    AS current_godown_type
                    FROM rm_boxes b
                    LEFT JOIN procurement_materials m  ON m.id = b.material_id
                    LEFT JOIN procurement_godowns g    ON g.id = b.current_godown_id
                    WHERE b.box_code = %s
                    LIMIT 1
                """, (code,)).fetchone()
            if not box_row:
                return jsonify({"status": "not_found",
                                "message": "Package not found"}), 404
            box = dict(box_row)
            box["per_box_qty"] = float(box.get("per_box_qty") or 0)
            for k in ("created_at", "updated_at"):
                if box.get(k) is not None:
                    box[k] = str(box[k])
            box_id = int(box["box_id"])

            # GRN origin (only for source='grn')
            grn = None
            if box.get("grn_id"):
                try:
                    grn_row = conn.execute("""
                        SELECT  h.id AS grn_id, h.grn_no, h.grn_date,
                                h.supplier_id, h.voucher_type_name,
                                COALESCE(s.supplier_name,'') AS supplier_name
                        FROM procurement_grn h
                        LEFT JOIN procurement_suppliers s ON s.id = h.supplier_id
                        WHERE h.id = %s
                        LIMIT 1
                    """, (box["grn_id"],)).fetchone()
                    if grn_row:
                        grn = dict(grn_row)
                        if grn.get("grn_date") is not None:
                            grn["grn_date"] = str(grn["grn_date"])
                except Exception:
                    grn = None

            movements = []
            try:
                mv_rows = conn.execute("""
                    SELECT m.movement_id, m.movement_type, m.qty,
                           m.movement_at, m.moved_by, m.remarks, m.transfer_id,
                           COALESCE(fg.name,'') AS from_name,
                           COALESCE(tg.name,'') AS to_name
                    FROM rm_box_movements m
                    LEFT JOIN procurement_godowns fg ON fg.id = m.from_godown_id
                    LEFT JOIN procurement_godowns tg ON tg.id = m.to_godown_id
                    WHERE m.box_id = %s
                    ORDER BY m.movement_at DESC, m.movement_id DESC
                """, (box_id,)).fetchall()
                for r in mv_rows:
                    d = dict(r)
                    d["qty"] = float(d.get("qty") or 0)
                    if d.get("movement_at") is not None:
                        d["movement_at"] = str(d["movement_at"])
                    movements.append(d)
            except Exception:
                pass

            return jsonify({
                "status":    "ok",
                "box":       box,
                "grn":       grn,
                "movements": movements,
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ════════════════════════════════════════════════════════════════════════
    # OPENING STOCK ENTRY  (single-box quick add — for Phase 1 testing)
    # ────────────────────────────────────────────────────────────────────────
    # Phase 2 will add bulk entry + label printing, but this minimal version
    # lets you start tracking real stock TODAY: select a material, godown,
    # number of boxes, qty per box → boxes are created with codes.
    # ════════════════════════════════════════════════════════════════════════

    @app.route("/api/inventory_godown/opening/log", methods=["GET"])
    def api_opening_log():
        """List opening-stock entries grouped by creation event
        (material + godown + created_at + batch + per_box_qty + created_by).
        Each group is one 'entry' with its box codes for reprint."""
        if not session.get("logged_in") and not session.get("UID"):
            return jsonify({"status": "error", "message": "Not logged in"}), 401
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            # only columns we know exist; batch/expiry/manufacturer may be absent
            have = set()
            for c in conn.execute("SHOW COLUMNS FROM rm_boxes").fetchall():
                have.add(c["Field"] if hasattr(c, "get") else c[0])
            sel_batch = "b.batch_num" if "batch_num" in have else "NULL"
            sel_exp   = "b.expiry_date" if "expiry_date" in have else "NULL"
            sel_mfr   = "b.manufacturer" if "manufacturer" in have else "NULL"

            rows = conn.execute(f"""
                SELECT b.box_id, b.box_code, b.material_id, b.current_godown_id AS gid,
                       b.per_box_qty, b.uom, b.current_status, b.source,
                       b.created_at, b.created_by, b.total_boxes,
                       {sel_batch} AS batch_num, {sel_exp} AS expiry_date, {sel_mfr} AS manufacturer,
                       COALESCE(m.material_name,'') AS material_name,
                       COALESCE(g.name,'') AS godown_name
                FROM rm_boxes b
                LEFT JOIN procurement_materials m ON m.id = b.material_id
                LEFT JOIN procurement_godowns  g ON g.id = b.current_godown_id
                WHERE b.source = 'opening'
                ORDER BY b.created_at DESC, b.box_id ASC
            """).fetchall()

            groups = {}
            order = []
            for r in rows:
                key = "|".join([
                    str(r["material_id"]), str(r["gid"]),
                    str(r["created_at"])[:19], str(r["batch_num"] or ""),
                    str(r["expiry_date"] or ""),
                    str(r["per_box_qty"]), str(r["created_by"] or ""),
                ])
                if key not in groups:
                    groups[key] = {
                        "material": r["material_name"], "godown": r["godown_name"],
                        "godown_id": r["gid"],
                        "created_at": str(r["created_at"])[:19],
                        "created_by": r["created_by"] or "",
                        "batch_num": r["batch_num"] or "",
                        "expiry_date": str(r["expiry_date"])[:10] if r["expiry_date"] else "",
                        "manufacturer": r["manufacturer"] or "",
                        "per_box_qty": float(r["per_box_qty"] or 0),
                        "uom": r["uom"] or "",
                        "boxes": [],
                        "box_ids": [],
                    }
                    order.append(key)
                groups[key]["boxes"].append({
                    "box_code": r["box_code"],
                    "status": r["current_status"],
                })
                groups[key]["box_ids"].append(r["box_id"])
            entries = []
            for k in order:
                gobj = groups[k]
                n = len(gobj["boxes"])
                active = sum(1 for b in gobj["boxes"] if b["status"] == "in_stock")
                gobj["no_of_box"] = n
                gobj["active_box"] = active
                gobj["total_qty"] = round(gobj["per_box_qty"] * n, 3)
                entries.append(gobj)
            conn.close()
            return jsonify({"status": "ok", "entries": entries, "count": len(entries)})
        except Exception as e:
            try: conn.close()
            except Exception: pass
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/inventory_godown/opening/create_boxes", methods=["POST"])
    @_cap_required("opening_stock")
    def api_opening_create_boxes():
        """
        Create N RM boxes for a material at a godown (opening-stock entry).

        Body:
        {
          "material_id":  123,
          "godown_id":    4,
          "no_of_box":    5,
          "per_box_qty":  50.0,
          "remarks":      "Initial physical count - Bhayla godown"
        }

        Response: { status: 'ok', op_seq: 1, boxes: [{box_id, box_code}, ...] }
        """
        # Additional check: opening_stock_view_print toggle means create/edit
        # is disallowed even though the user can see the page and print labels.
        try:
            try:
                from .inventory_access import _inv_block_opening_edit
            except Exception:
                from inventory_access import _inv_block_opening_edit
            _blk = _inv_block_opening_edit()
            if _blk is not None:
                return _blk
        except Exception:
            pass
        d = request.get_json(silent=True) or {}
        try:
            material_id = int(d.get("material_id") or 0)
            godown_id   = int(d.get("godown_id")   or 0)
            no_of_box   = int(d.get("no_of_box")   or 0)
            per_box_qty = float(d.get("per_box_qty") or 0)
        except (TypeError, ValueError):
            return jsonify({"status": "error", "message": "Invalid numeric input"}), 400

        remarks = (d.get("remarks") or "").strip()[:255]
        batch_num = (d.get("batch_num") or "").strip()[:64]
        expiry_date = (d.get("expiry_date") or "").strip() or None
        manufacturer = (d.get("manufacturer") or "").strip()[:128]

        if not material_id or not godown_id:
            return jsonify({"status": "error",
                            "message": "material_id and godown_id are required"}), 400
        if no_of_box <= 0 or no_of_box > 500:
            return jsonify({"status": "error",
                            "message": "Number of Packages must be between 1 and 500"}), 400
        if per_box_qty <= 0:
            return jsonify({"status": "error",
                            "message": "Quantity per Package must be greater than zero"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            mat = conn.execute(
                "SELECT id, material_name, uom FROM procurement_materials WHERE id=%s",
                (material_id,)
            ).fetchone()
            if not mat:
                return jsonify({"status": "error", "message": "Material not found"}), 404
            gd = conn.execute(
                "SELECT id, name FROM procurement_godowns WHERE id=%s",
                (godown_id,)
            ).fetchone()
            if not gd:
                return jsonify({"status": "error", "message": "Godown not found"}), 404

            mat_code = derive_mat_code(mat["material_name"] or "")  # kept in DB for legacy
            uom      = mat["uom"] or ""

            # Ensure rm_boxes has batch_num / expiry_date columns (opening stock
            # can carry these directly; GRN boxes get them via grn_item_id).
            try:
                cols = set()
                for c in conn.execute("SHOW COLUMNS FROM rm_boxes").fetchall():
                    cols.add(c["Field"] if hasattr(c, "get") else c[0])
                if "batch_num" not in cols:
                    conn.execute("ALTER TABLE rm_boxes ADD COLUMN batch_num VARCHAR(64) DEFAULT NULL")
                if "expiry_date" not in cols:
                    conn.execute("ALTER TABLE rm_boxes ADD COLUMN expiry_date DATE DEFAULT NULL")
                if "manufacturer" not in cols:
                    conn.execute("ALTER TABLE rm_boxes ADD COLUMN manufacturer VARCHAR(128) DEFAULT NULL")
            except Exception:
                pass

            # NEW: allocate sequential short codes via allocate_next_box_code.
            # Each box gets a globally-unique RM-A0000001 style code; no more
            # per-material OP-sequence tracking needed.
            created = []
            user = _user()
            for seq in range(1, no_of_box + 1):
                box_code = allocate_next_box_code(conn)
                cur = conn.execute("""
                    INSERT INTO rm_boxes
                      (box_code, grn_id, grn_no, grn_item_id,
                       material_id, material_code,
                       box_seq, total_boxes, per_box_qty, uom,
                       current_godown_id, current_status,
                       source, created_by, batch_num, expiry_date, manufacturer)
                    VALUES (%s, NULL, NULL, NULL,
                            %s, %s,
                            %s, %s, %s, %s,
                            %s, 'in_stock',
                            'opening', %s, %s, %s, %s)
                """, (box_code, material_id, mat_code,
                      seq, no_of_box, per_box_qty, uom,
                      godown_id, user, (batch_num or None), expiry_date,
                      (manufacturer or None)))
                box_id = cur.lastrowid
                conn.execute("""
                    INSERT INTO rm_box_movements
                      (box_id, movement_type, from_godown_id, to_godown_id,
                       qty, moved_by, remarks)
                    VALUES (%s, 'opening', NULL, %s, %s, %s, %s)
                """, (box_id, godown_id, per_box_qty, user,
                      remarks or "Opening stock entry"))
                created.append({"box_id": box_id, "box_code": box_code})

            conn.commit()
            return jsonify({
                "status":     "ok",
                "material":   mat["material_name"],
                "godown":     gd["name"],
                "no_of_box":  no_of_box,
                "per_box_qty": per_box_qty,
                "uom":        uom,
                "boxes":      created,
            })
        except Exception as e:
            try: conn.rollback()
            except Exception: pass
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ════════════════════════════════════════════════════════════════════════
    # OPENING-STOCK EDIT REQUESTS  (request → admin approve/reject → apply)
    # ────────────────────────────────────────────────────────────────────────
    # Any user with opening-stock access can request changes to an opening-
    # stock entry (all its boxes). Admins approve (changes apply to every box
    # in the entry) or reject. Mirrors the request/approval pattern used by the
    # FEFO override + GRN box repair flows.
    # ════════════════════════════════════════════════════════════════════════

    # Fields a requester is allowed to change, mapped to the rm_boxes column.
    _OP_EDIT_FIELDS = {
        "per_box_qty":  "per_box_qty",
        "uom":          "uom",
        "batch_num":    "batch_num",
        "expiry_date":  "expiry_date",
        "manufacturer": "manufacturer",
        "godown_id":    "current_godown_id",
    }

    def _op_edit_ensure_table(conn):
        conn.execute("""
            CREATE TABLE IF NOT EXISTS rm_opening_edit_requests (
                id            INT AUTO_INCREMENT PRIMARY KEY,
                box_ids       TEXT NOT NULL,
                material_id   INT DEFAULT NULL,
                material_name VARCHAR(200) DEFAULT NULL,
                box_count     INT NOT NULL DEFAULT 0,
                old_values    TEXT DEFAULT NULL,
                new_values    TEXT DEFAULT NULL,
                reason        VARCHAR(500) DEFAULT NULL,
                status        ENUM('pending','approved','rejected') NOT NULL DEFAULT 'pending',
                requested_by  VARCHAR(80) DEFAULT NULL,
                requested_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
                decided_by    VARCHAR(80) DEFAULT NULL,
                decided_at    DATETIME DEFAULT NULL,
                decide_note   VARCHAR(500) DEFAULT NULL,
                INDEX ix_oper_status (status),
                INDEX ix_oper_at     (requested_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

    def _op_is_admin() -> bool:
        return (session.get("User_Type") or "").strip().lower() == "admin"

    @app.route("/api/inventory_godown/opening/edit/request", methods=["POST"])
    @_cap_required("opening_stock")
    def api_opening_edit_request():
        """Submit a request to edit an opening-stock entry (all boxes in it)."""
        # Respect the view/print-only restriction.
        try:
            try:
                from .inventory_access import _inv_block_opening_edit
            except Exception:
                from inventory_access import _inv_block_opening_edit
            _blk = _inv_block_opening_edit()
            if _blk is not None:
                return _blk
        except Exception:
            pass

        d = request.get_json(silent=True) or {}
        box_ids = d.get("box_ids") or []
        reason  = (d.get("reason") or "").strip()
        changes = d.get("changes") or {}

        try:
            box_ids = [int(b) for b in box_ids if str(b).strip()]
        except (TypeError, ValueError):
            return jsonify({"status": "error", "message": "Invalid box_ids"}), 400
        if not box_ids:
            return jsonify({"status": "error", "message": "No boxes to edit"}), 400
        if not reason:
            return jsonify({"status": "error", "message": "A reason is required"}), 400

        # Keep only recognised, sanitised fields.
        clean = {}
        for k, v in (changes.items() if isinstance(changes, dict) else []):
            if k not in _OP_EDIT_FIELDS:
                continue
            if k == "per_box_qty":
                try:
                    fv = float(v)
                except (TypeError, ValueError):
                    return jsonify({"status": "error", "message": "Invalid qty"}), 400
                if fv <= 0:
                    return jsonify({"status": "error", "message": "Qty must be greater than zero"}), 400
                clean[k] = fv
            elif k == "godown_id":
                try:
                    iv = int(v)
                except (TypeError, ValueError):
                    return jsonify({"status": "error", "message": "Invalid godown"}), 400
                if iv > 0:
                    clean[k] = iv
            elif k == "expiry_date":
                clean[k] = (str(v).strip()[:10] or None)
            else:
                clean[k] = str(v).strip()[:128]
        if not clean:
            return jsonify({"status": "error", "message": "No valid changes supplied"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            _op_edit_ensure_table(conn)

            # Read current values from the first box (the entry shares values).
            ph = ",".join(["%s"] * len(box_ids))
            row = conn.execute(
                f"""SELECT b.material_id, b.per_box_qty, b.uom, b.batch_num,
                           b.expiry_date, b.manufacturer, b.current_godown_id,
                           COALESCE(m.material_name,'') AS material_name
                    FROM rm_boxes b
                    LEFT JOIN procurement_materials m ON m.id = b.material_id
                    WHERE b.box_id IN ({ph})
                    ORDER BY b.box_id ASC LIMIT 1""",
                tuple(box_ids)
            ).fetchone()
            if not row:
                return jsonify({"status": "error", "message": "Entry not found"}), 404

            # Resolve a friendly godown name for old/new display.
            def _gname(gid):
                if not gid:
                    return ""
                gr = conn.execute("SELECT name FROM procurement_godowns WHERE id=%s", (gid,)).fetchone()
                return (gr["name"] if gr else "") or ""

            old_vals, new_vals = {}, {}
            for k in clean:
                if k == "godown_id":
                    old_vals[k] = _gname(row["current_godown_id"])
                    new_vals[k] = _gname(clean[k])
                elif k == "per_box_qty":
                    old_vals[k] = float(row["per_box_qty"] or 0)
                    new_vals[k] = clean[k]
                elif k == "expiry_date":
                    old_vals[k] = str(row["expiry_date"])[:10] if row["expiry_date"] else ""
                    new_vals[k] = clean[k] or ""
                else:
                    old_vals[k] = row[k] or ""
                    new_vals[k] = clean[k]

            # Store the raw godown_id for apply-time (display map holds the name).
            apply_payload = dict(clean)

            conn.execute("""
                INSERT INTO rm_opening_edit_requests
                  (box_ids, material_id, material_name, box_count,
                   old_values, new_values, reason, status, requested_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,'pending',%s)
            """, (
                _json.dumps(box_ids), row["material_id"], row["material_name"],
                len(box_ids), _json.dumps(old_vals),
                _json.dumps({"_display": new_vals, "_apply": apply_payload}),
                reason[:500], _user(),
            ))
            conn.commit()
            return jsonify({"status": "ok", "message": "Edit request submitted for admin approval."})
        except Exception as e:
            try: conn.rollback()
            except Exception: pass
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    @app.route("/api/inventory_godown/opening/edit/requests", methods=["GET"])
    @_login_required
    def api_opening_edit_requests():
        """List edit requests. Admins see all + can act; others see their own."""
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            _op_edit_ensure_table(conn)
            is_admin = _op_is_admin()
            if is_admin:
                rows = conn.execute(
                    "SELECT * FROM rm_opening_edit_requests ORDER BY "
                    "(status='pending') DESC, requested_at DESC LIMIT 300"
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM rm_opening_edit_requests WHERE requested_by=%s "
                    "ORDER BY requested_at DESC LIMIT 300",
                    (_user(),)
                ).fetchall()

            items = []
            for r in rows:
                try:
                    nv_raw = _json.loads(r["new_values"] or "{}")
                except Exception:
                    nv_raw = {}
                nv_disp = nv_raw.get("_display", nv_raw) if isinstance(nv_raw, dict) else {}
                try:
                    ov = _json.loads(r["old_values"] or "{}")
                except Exception:
                    ov = {}
                items.append({
                    "id": r["id"],
                    "material_name": r["material_name"] or "",
                    "box_count": r["box_count"] or 0,
                    "old_values": ov,
                    "new_values": nv_disp,
                    "reason": r["reason"] or "",
                    "status": r["status"],
                    "requested_by": r["requested_by"] or "",
                    "requested_at": str(r["requested_at"])[:19] if r["requested_at"] else "",
                    "decided_by": r["decided_by"] or "",
                    "decided_at": str(r["decided_at"])[:19] if r["decided_at"] else "",
                    "decide_note": r["decide_note"] or "",
                })
            return jsonify({"status": "ok", "is_admin": is_admin, "items": items})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    @app.route("/api/inventory_godown/opening/edit/<int:rid>/<action>", methods=["POST"])
    @_login_required
    def api_opening_edit_decide(rid, action):
        """Admin approves (applies to all boxes) or rejects an edit request."""
        if action not in ("approve", "reject"):
            return jsonify({"status": "error", "message": "Unknown action"}), 400
        if not _op_is_admin():
            return jsonify({"status": "error", "message": "Only admins can decide requests"}), 403

        note = ((request.get_json(silent=True) or {}).get("note") or "").strip()[:500]

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            _op_edit_ensure_table(conn)
            req = conn.execute(
                "SELECT * FROM rm_opening_edit_requests WHERE id=%s FOR UPDATE", (rid,)
            ).fetchone()
            if not req:
                return jsonify({"status": "error", "message": "Request not found"}), 404
            if req["status"] != "pending":
                return jsonify({"status": "error",
                                "message": "This request was already " + req["status"] + "."}), 409

            if action == "reject":
                conn.execute(
                    "UPDATE rm_opening_edit_requests SET status='rejected', "
                    "decided_by=%s, decided_at=NOW(), decide_note=%s WHERE id=%s",
                    (_user(), note, rid)
                )
                conn.commit()
                return jsonify({"status": "ok", "message": "Request rejected."})

            # APPROVE → apply changes to all boxes in the entry.
            try:
                box_ids = _json.loads(req["box_ids"] or "[]")
                box_ids = [int(b) for b in box_ids]
            except Exception:
                box_ids = []
            if not box_ids:
                return jsonify({"status": "error", "message": "Request has no boxes"}), 400

            try:
                nv_raw = _json.loads(req["new_values"] or "{}")
            except Exception:
                nv_raw = {}
            apply_payload = nv_raw.get("_apply", {}) if isinstance(nv_raw, dict) else {}
            if not apply_payload:
                return jsonify({"status": "error", "message": "Nothing to apply"}), 400

            set_parts, vals = [], []
            for k, v in apply_payload.items():
                col = _OP_EDIT_FIELDS.get(k)
                if not col:
                    continue
                set_parts.append(col + "=%s")
                vals.append(v if v != "" else None)
            if not set_parts:
                return jsonify({"status": "error", "message": "No applicable columns"}), 400

            ph = ",".join(["%s"] * len(box_ids))
            conn.execute(
                "UPDATE rm_boxes SET " + ", ".join(set_parts) +
                f" WHERE box_id IN ({ph})",
                tuple(vals) + tuple(box_ids)
            )

            # Audit movement on each box.
            user = _user()
            for bid in box_ids:
                conn.execute(
                    "INSERT INTO rm_box_movements (box_id, movement_type, qty, moved_by, remarks) "
                    "VALUES (%s,'adjust',0,%s,%s)",
                    (bid, user, ("Opening-stock edit approved (req #%d)" % rid)[:255])
                )

            conn.execute(
                "UPDATE rm_opening_edit_requests SET status='approved', "
                "decided_by=%s, decided_at=NOW(), decide_note=%s WHERE id=%s",
                (user, note, rid)
            )
            conn.commit()
            return jsonify({"status": "ok",
                            "message": "Approved — changes applied to %d package(s)." % len(box_ids)})
        except Exception as e:
            try: conn.rollback()
            except Exception: pass
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    @app.route("/api/inventory_godown/opening/edit/pending_count", methods=["GET"])
    @_login_required
    def api_opening_edit_pending_count():
        """Badge count: admins see all pending; others see their own pending."""
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            _op_edit_ensure_table(conn)
            if _op_is_admin():
                row = conn.execute(
                    "SELECT COUNT(*) AS c FROM rm_opening_edit_requests WHERE status='pending'"
                ).fetchone()
            else:
                row = conn.execute(
                    "SELECT COUNT(*) AS c FROM rm_opening_edit_requests "
                    "WHERE status='pending' AND requested_by=%s",
                    (_user(),)
                ).fetchone()
            return jsonify({"status": "ok", "count": int(row["c"] if row else 0)})
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ════════════════════════════════════════════════════════════════════════
    # MATERIAL SEARCH  (for the opening-stock entry form's material picker)
    # ════════════════════════════════════════════════════════════════════════

    @app.route("/api/inventory_godown/materials/search", methods=["GET"])
    @_login_required
    def api_materials_search():
        """Quick search across RM materials for the opening-stock form."""
        q = (request.args.get("q") or "").strip()
        try:
            limit = max(1, min(int(request.args.get("limit") or 20), 100))
        except (TypeError, ValueError):
            limit = 20

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            if q:
                like = f"%{q}%"
                rows = conn.execute("""
                    SELECT id, material_name, uom
                    FROM procurement_materials
                    WHERE material_name LIKE %s
                    ORDER BY material_name ASC
                    LIMIT %s
                """, (like, limit)).fetchall()
            else:
                rows = conn.execute("""
                    SELECT id, material_name, uom
                    FROM procurement_materials
                    ORDER BY material_name ASC
                    LIMIT %s
                """, (limit,)).fetchall()
            return jsonify({"status": "ok",
                            "materials": [dict(r) for r in rows]})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ════════════════════════════════════════════════════════════════════════
    # ALLOCATE CODES  (for printing labels in batches, e.g. from GRN)
    # ════════════════════════════════════════════════════════════════════════

    @app.route("/api/inventory_godown/allocate_codes", methods=["POST"])
    @_login_required
    def api_allocate_codes():
        """
        Allocate sequential short codes for GRN label printing.

        Two payload shapes:

        ─── Persistent mode (preferred) ───────────────────────────────
        Body: { "grn_id": <int>, "lines": [
                  {"grn_item_id": <int>, "count": <int>}, ...
                ] }

          For each line, the server looks up codes already saved against
          (grn_id, grn_item_id) in rm_grn_box_codes. It returns those
          first, allocating fresh codes only for the shortfall, and
          persists new allocations so the SAME codes come back on every
          subsequent print. If count is LESS than what's saved, only the
          first `count` saved codes are returned (the rest stay reserved).

          This is what GRN label printing uses — codes for a saved GRN
          stay stable across reprints.

        ─── Legacy fragment mode (back-compat) ────────────────────────
        Body: { "count": N }
          Allocates N fresh codes with no persistence. Used by callers
          that don't (yet) have a saved GRN context, e.g. quick reprints
          on drafts. Note: GRN label print no longer takes this path —
          it blocks on the client side if grn_id is missing.

        Returns:
          Persistent: { "status":"ok", "lines":[
                          {"grn_item_id":<int>, "codes":[...]}, ...
                       ] }
          Legacy:     { "status":"ok", "codes":[...] }
        """
        d = request.get_json(silent=True) or {}
        grn_id = d.get("grn_id")
        lines  = d.get("lines")

        # Persistent mode
        if grn_id and isinstance(lines, list) and lines:
            try:
                grn_id = int(grn_id)
            except (TypeError, ValueError):
                return jsonify({"status":"error", "message":"Invalid grn_id"}), 400

            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            try:
                # Ensure the persistence table exists. Keyed by
                # (grn_id, grn_item_id, box_seq) — box_seq is just a within-line
                # ordering counter so we can return a deterministic list.
                # Same codes for the same (grn_id, grn_item_id) pair forever.
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS rm_grn_box_codes (
                        grn_id        INT NOT NULL,
                        grn_item_id   INT NOT NULL,
                        box_seq       INT NOT NULL,
                        box_code      VARCHAR(50) NOT NULL,
                        allocated_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
                        allocated_by  VARCHAR(64) DEFAULT NULL,
                        PRIMARY KEY (grn_id, grn_item_id, box_seq),
                        UNIQUE KEY ux_rm_grn_box_codes_code (box_code),
                        INDEX ix_rm_grn_box_codes_grn (grn_id)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                out_lines = []
                user = _user()

                for line in lines:
                    try:
                        item_id = int(line.get("grn_item_id") or 0)
                        cnt     = int(line.get("count") or 0)
                    except (TypeError, ValueError):
                        item_id, cnt = 0, 0
                    if item_id <= 0 or cnt <= 0 or cnt > 500:
                        # Skip invalid line — caller should fix and retry
                        out_lines.append({"grn_item_id": item_id, "codes": []})
                        continue

                    # ── Self-heal: align reservations to the ACTUAL boxes ──
                    # If rm_boxes already has real box rows for this line, those
                    # box_codes are the single source of truth (scanning matches
                    # rm_boxes). Historically a GRN re-save could recreate boxes
                    # with fresh codes while rm_grn_box_codes kept the old ones,
                    # so reprints printed codes that no longer scanned. Here we
                    # upsert rm_grn_box_codes to match rm_boxes by box_seq, so a
                    # reprint always carries the real, scannable code. Opening
                    # stock never reaches this path (no grn_id), so it's
                    # untouched. We only sync seqs that actually have a box;
                    # any shortfall is allocated below as before.
                    # Read the real boxes (guarded — a read failure is
                    # non-fatal and just skips the self-heal for this line).
                    real_pairs = []
                    try:
                        real_boxes = conn.execute("""
                            SELECT box_seq, box_code FROM rm_boxes
                            WHERE grn_id=%s AND grn_item_id=%s
                              AND box_seq IS NOT NULL AND box_code IS NOT NULL
                            ORDER BY box_seq ASC
                        """, (grn_id, item_id)).fetchall()
                        real_pairs = [
                            ((rb["box_seq"] if hasattr(rb, "get") else rb[0]),
                             (rb["box_code"] if hasattr(rb, "get") else rb[1]))
                            for rb in real_boxes
                        ]
                        real_pairs = [(s, c) for (s, c) in real_pairs if s is not None and c]
                    except Exception as _ex:
                        print(f"ℹ️  [InventoryGodown] allocate_codes box read skipped "
                              f"for grn_id={grn_id} item_id={item_id}: {_ex}")
                        real_pairs = []

                    if real_pairs:
                        # The real boxes are authoritative. Clear ALL of this
                        # line's reservations, then re-insert from rm_boxes.
                        # Wiping the whole line first avoids any clash with the
                        # UNIQUE(box_code) index when a real code is held under a
                        # different seq by a stale reservation. NOT wrapped in a
                        # local try/except: if the DELETE succeeds but an INSERT
                        # fails, we must let it propagate to the outer handler so
                        # the whole request rolls back — never commit a line with
                        # its reservations half-deleted.
                        conn.execute(
                            "DELETE FROM rm_grn_box_codes WHERE grn_id=%s AND grn_item_id=%s",
                            (grn_id, item_id),
                        )
                        for _seq, _code in real_pairs:
                            conn.execute("""
                                INSERT INTO rm_grn_box_codes
                                  (grn_id, grn_item_id, box_seq, box_code, allocated_by)
                                VALUES (%s, %s, %s, %s, %s)
                            """, (grn_id, item_id, _seq, _code, user))

                    # Get any existing codes for this (grn_id, grn_item_id),
                    # ordered by box_seq so the result is deterministic.
                    existing = conn.execute("""
                        SELECT box_seq, box_code FROM rm_grn_box_codes
                        WHERE grn_id=%s AND grn_item_id=%s
                        ORDER BY box_seq ASC
                    """, (grn_id, item_id)).fetchall()
                    existing_codes = [r["box_code"] for r in existing]
                    have = len(existing_codes)

                    if have >= cnt:
                        # Already have enough — return just the first `cnt`.
                        # The extras stay reserved (in case user re-edits to a
                        # higher number later).
                        codes_for_line = existing_codes[:cnt]
                    else:
                        # Allocate the shortfall and persist
                        codes_for_line = list(existing_codes)
                        for seq in range(have + 1, cnt + 1):
                            new_code = allocate_next_box_code(conn)
                            conn.execute("""
                                INSERT INTO rm_grn_box_codes
                                  (grn_id, grn_item_id, box_seq, box_code, allocated_by)
                                VALUES (%s, %s, %s, %s, %s)
                            """, (grn_id, item_id, seq, new_code, user))
                            codes_for_line.append(new_code)

                    out_lines.append({
                        "grn_item_id": item_id,
                        "codes":       codes_for_line,
                    })

                conn.commit()
                return jsonify({"status":"ok", "lines": out_lines})

            except Exception as e:
                try: conn.rollback()
                except Exception: pass
                traceback.print_exc()
                return jsonify({"status":"error","message":str(e)}), 500
            finally:
                try: conn.close()
                except Exception: pass

        # ─── Legacy fragment mode ───────────────────────────────────────
        try:
            n = int(d.get("count") or 0)
        except (TypeError, ValueError):
            n = 0
        if n < 1 or n > 500:
            return jsonify({"status":"error",
                            "message":"count must be between 1 and 500 "
                                      "(or use persistent mode with grn_id + lines)"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status":"error","message":"DB connection failed"}), 500
        try:
            codes = [allocate_next_box_code(conn) for _ in range(n)]
            conn.commit()
            return jsonify({"status":"ok", "codes": codes})
        except Exception as e:
            try: conn.rollback()
            except Exception: pass
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ════════════════════════════════════════════════════════════════════════
    # EXCEL BULK UPLOAD  (one-time opening stock import)
    # ════════════════════════════════════════════════════════════════════════

    @app.route("/api/inventory_godown/opening/upload_template", methods=["GET"])
    @_login_required
    def api_opening_upload_template():
        """Download a blank Excel template for the bulk opening-stock upload."""
        try:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "Opening Stock"
            headers = [
                "Material Name", "Godown", "No. of Packages",
                "Qty per Pkg.", "UOM", "Remarks (optional)",
                "Batch No. (optional)", "Expiry Date (optional, YYYY-MM-DD)",
                "Manufacturer (optional)",
                # ── Label-only columns (not stored in DB; printed on label only) ──
                "Supplier Name (label only)",
                "GRN No. (label only)",
                "GRN Date (label only, YYYY-MM-DD)",
                "Invoice No. (label only)",
                "Invoice Date (label only, YYYY-MM-DD)",
            ]
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF")
                # Highlight label-only columns in amber so users can see they
                # behave differently from the persisted columns.
                fill_color = "D97706" if col_idx >= 10 else "2563EB"
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Example rows so users see the expected format
            ws.append(["0032 Yum Pistachio", "BHAYLA OLD GODOWN", 5, 50.0, "KG", "Initial physical count", "B-1001", "2027-03-31", "Kothari Fragrances",
                       "Kothari Fragrances LLP", "GRN/2026/0123", "2026-05-15", "INV-9988", "2026-05-10"])
            ws.append(["00746 THE ONE",       "FACTORY",            3, 100.0,"KG", "", "", "", "",
                       "", "", "", "", ""])
            # Add a notes row at the bottom
            ws.append([])
            ws.append(["NOTES:"])
            ws.append(["- Material Name must EXACTLY match an existing item (case-insensitive)"])
            ws.append(["- Godown must EXACTLY match an existing godown name (case-insensitive)"])
            ws.append(["- No. of Packages: 1-500"])
            ws.append(["- Qty per Pkg.: any positive number"])
            ws.append(["- UOM is optional (defaults to material's UOM if blank)"])
            ws.append(["- Amber-coloured columns (Supplier, GRN, Invoice) are LABEL-ONLY — printed on the label but NOT stored in the database."])
            ws.append(["- Reprints from the Label Reprint module will NOT carry these values forward."])
            # Column widths
            widths = [38, 30, 14, 14, 8, 32, 22, 28, 28,
                      28, 22, 28, 22, 28]
            for i, w in enumerate(widths, 1):
                # Use openpyxl's get_column_letter for safety beyond Z
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(i)].width = w

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            from flask import send_file
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="opening_stock_template.xlsx",
            )
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/inventory_godown/opening/upload", methods=["POST"])
    @_cap_required("opening_stock")
    def api_opening_upload():
        """
        Bulk-create opening-stock boxes from an uploaded Excel file.

        Form fields:
          file:    the .xlsx file
          replace: '1' to delete existing opening-stock boxes for each
                   (material, godown) pair before re-creating them

        Returns a per-row results report:
          {
            "status": "ok",
            "summary": { rows_total, rows_ok, rows_skipped, rows_error,
                         boxes_created, boxes_replaced },
            "results": [ { row, material, godown, status, message, codes }, ... ]
          }
        """
        # Block create/edit when opening_stock_view_print is ON.
        try:
            try:
                from .inventory_access import _inv_block_opening_edit
            except Exception:
                from inventory_access import _inv_block_opening_edit
            _blk = _inv_block_opening_edit()
            if _blk is not None:
                return _blk
        except Exception:
            pass
        from io import BytesIO
        try:
            from openpyxl import load_workbook
        except ImportError:
            return jsonify({"status":"error",
                            "message":"openpyxl is not installed on the server"}), 500

        f = request.files.get("file")
        if not f:
            return jsonify({"status":"error","message":"No file uploaded"}), 400
        replace = request.form.get("replace") in ("1", "true", "yes")
        # Two-call confirmation protocol:
        #   • confirm != '1'  → validate only; if any hard conflict is found,
        #     return status='needs_confirmation' with a per-row analysis and
        #     insert NOTHING. The frontend shows a review screen.
        #   • confirm == '1'  → user has reviewed; proceed to insert, skipping
        #     any rows listed in exclude_rows (1-based Excel row numbers the
        #     user unticked on the review screen).
        confirmed = request.form.get("confirm") in ("1", "true", "yes")
        _excl_raw = (request.form.get("exclude_rows") or "").strip()
        excluded_rows = set()
        if _excl_raw:
            for _tok in _excl_raw.split(","):
                _tok = _tok.strip()
                if _tok.isdigit():
                    excluded_rows.add(int(_tok))

        try:
            data = f.read()
            wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return jsonify({"status":"error",
                            "message":f"Could not parse Excel file: {e}"}), 400

        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status":"error","message":"DB connection failed"}), 500

        try:
            # Build name → id lookups (case-insensitive)
            mat_rows = conn.execute(
                "SELECT id, material_name, uom FROM procurement_materials"
            ).fetchall()
            mat_lookup = {
                (r["material_name"] or "").strip().lower(): r
                for r in mat_rows
            }
            gd_rows = conn.execute(
                "SELECT id, name FROM procurement_godowns"
            ).fetchall()
            gd_lookup = {
                (r["name"] or "").strip().lower(): r
                for r in gd_rows
            }

            results = []
            summary = {
                "rows_total": 0, "rows_ok": 0, "rows_skipped": 0,
                "rows_error": 0, "boxes_created": 0, "boxes_replaced": 0,
            }
            user = _user()
            # Ensure rm_boxes carries batch/expiry (opening boxes store them directly)
            try:
                _cols = set()
                for _c in conn.execute("SHOW COLUMNS FROM rm_boxes").fetchall():
                    _cols.add(_c["Field"] if hasattr(_c, "get") else _c[0])
                if "batch_num" not in _cols:
                    conn.execute("ALTER TABLE rm_boxes ADD COLUMN batch_num VARCHAR(64) DEFAULT NULL")
                if "expiry_date" not in _cols:
                    conn.execute("ALTER TABLE rm_boxes ADD COLUMN expiry_date DATE DEFAULT NULL")
                if "manufacturer" not in _cols:
                    conn.execute("ALTER TABLE rm_boxes ADD COLUMN manufacturer VARCHAR(128) DEFAULT NULL")
            except Exception:
                pass

            # ════════════════════════════════════════════════════════════
            # VALIDATION PRE-PASS  (runs before any insert)
            # Collect every data row, group by (material, batch), and detect:
            #   • EXPIRY CONFLICT (hard): same material+batch with >1 expiry —
            #     a single batch physically cannot have two expiry dates. Also
            #     flagged if the file's expiry differs from EXISTING in-stock
            #     opening boxes of the same material+batch.
            #   • PACK-SIZE VARIANCE (soft): same material+batch+expiry with
            #     >1 per-box qty — allowed (e.g. 250kg & 300kg drums) but the
            #     user must confirm it's intentional.
            # Duplicates of the SAME material+batch+expiry are legitimate and
            # simply sum — never flagged.
            # ════════════════════════════════════════════════════════════
            def _parse_exp(v):
                if v is None or str(v).strip() == "":
                    return ""
                try:
                    if hasattr(v, "isoformat"):
                        return v.isoformat()[:10]
                    return str(v).strip()[:10]
                except Exception:
                    return ""

            pre_rows = []   # [{row, mat_name, batch, expiry, qty}]
            _pri = 0
            for _row in ws.iter_rows(values_only=True):
                _pri += 1
                if _pri == 1:
                    continue
                if not _row or all(c is None or str(c).strip() == "" for c in _row):
                    continue
                _first = str(_row[0] or "").strip()
                if _first.upper().startswith("NOTES") or _first.startswith("-"):
                    continue
                _mn = (str(_row[0]) if _row[0] is not None else "").strip()
                _bn = (str(_row[6]) if len(_row) > 6 and _row[6] is not None else "").strip()[:64]
                _ex = _parse_exp(_row[7] if len(_row) > 7 else None)
                try:    _q = float(_row[3]) if len(_row) > 3 and _row[3] is not None else 0.0
                except Exception: _q = 0.0
                if not _mn:
                    continue
                pre_rows.append({"row": _pri, "mat_name": _mn,
                                 "batch": _bn, "expiry": _ex, "qty": _q})

            # Group by (lc_material, batch) — only batches with a real batch
            # number participate in conflict checks (blank batch can't conflict
            # by definition). Note batches are surfaced separately as a warning.
            from collections import defaultdict
            _grp = defaultdict(list)
            for pr in pre_rows:
                _grp[(pr["mat_name"].lower(), pr["batch"])].append(pr)

            analysis = []   # one entry per (material,batch) group with a flag
            has_hard_conflict = False
            for (mlc, batch), items in _grp.items():
                disp_mat = items[0]["mat_name"]
                expiries = sorted({i["expiry"] for i in items if i["expiry"]})
                qtys_by_exp = defaultdict(set)
                for i in items:
                    qtys_by_exp[i["expiry"]].add(i["qty"])

                # Existing in-stock expiry for this material+batch (cross-upload
                # conflict detection — how the historical mess accumulated).
                existing_expiries = []
                if batch:
                    try:
                        _mrow = mat_lookup.get(mlc)
                        if _mrow:
                            _erows = conn.execute(
                                "SELECT DISTINCT expiry_date FROM rm_boxes "
                                "WHERE source='opening' AND current_status='in_stock' "
                                "AND material_id=%s AND batch_num=%s "
                                "AND expiry_date IS NOT NULL",
                                (int(_mrow["id"]), batch),
                            ).fetchall()
                            existing_expiries = sorted({
                                str(r["expiry_date"])[:10] for r in _erows
                            })
                    except Exception:
                        existing_expiries = []

                all_expiries = sorted(set(expiries) | set(existing_expiries))
                flag = "ok"; notes = []
                if batch and len(all_expiries) > 1:
                    flag = "conflict"; has_hard_conflict = True
                    note = "Multiple expiry dates for the same batch: " + ", ".join(all_expiries)
                    if existing_expiries and expiries and set(existing_expiries) - set(expiries):
                        note += " (some already in stock)"
                    notes.append(note)
                else:
                    # pack-size variance within a single expiry
                    for ex, qs in qtys_by_exp.items():
                        if len([q for q in qs if q > 0]) > 1:
                            flag = "warn" if flag == "ok" else flag
                            notes.append("Different pack sizes for "
                                         + (ex or "(no expiry)") + ": "
                                         + ", ".join(str(q) for q in sorted(qs)))
                if not batch:
                    flag = "warn" if flag == "ok" else flag
                    notes.append("No batch number — traceability limited")

                analysis.append({
                    "material": disp_mat, "batch": batch or "(blank)",
                    "rows": [i["row"] for i in items],
                    "file_expiries": expiries,
                    "existing_expiries": existing_expiries,
                    "flag": flag,
                    "note": "; ".join(notes),
                    "box_count": sum(1 for _ in items),  # group line count
                })

            # If there are hard conflicts and the user hasn't confirmed yet,
            # stop and return the review payload — insert nothing.
            if has_hard_conflict and not confirmed:
                conn.rollback()
                return jsonify({
                    "status": "needs_confirmation",
                    "message": "Some batches have conflicting expiry dates. "
                               "Review and choose which rows to import.",
                    "analysis": sorted(analysis,
                                       key=lambda a: {"conflict":0,"warn":1,"ok":2}[a["flag"]]),
                })

            row_idx = 0
            for row in ws.iter_rows(values_only=True):
                row_idx += 1
                if row_idx == 1:
                    continue  # skip header
                # Skip blank rows + notes rows (where first cell is "NOTES:" etc.)
                if not row or all(c is None or str(c).strip() == "" for c in row):
                    continue
                first = str(row[0] or "").strip()
                if first.upper().startswith("NOTES") or first.startswith("-"):
                    continue

                summary["rows_total"] += 1

                # User unticked this row on the review screen → skip it.
                if row_idx in excluded_rows:
                    summary["rows_skipped"] += 1
                    results.append({
                        "row": row_idx, "material": str(row[0] or "").strip(),
                        "godown": (str(row[1]) if len(row) > 1 and row[1] is not None else "").strip(),
                        "status": "skipped",
                        "message": "Excluded by user on review screen", "codes": [],
                    })
                    continue

                # Extract fields with defensive defaults
                mat_name = (str(row[0]) if row[0] is not None else "").strip()
                gd_name  = (str(row[1]) if len(row) > 1 and row[1] is not None else "").strip()
                try:    no_of_box = int(row[2]) if len(row) > 2 and row[2] is not None else 0
                except: no_of_box = 0
                try:    per_box_qty = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
                except: per_box_qty = 0.0
                uom_raw  = (str(row[4]) if len(row) > 4 and row[4] is not None else "").strip()
                remarks  = (str(row[5]) if len(row) > 5 and row[5] is not None else "").strip()[:255]
                batch_num = (str(row[6]) if len(row) > 6 and row[6] is not None else "").strip()[:64]
                # Expiry (col 8) — accept date or ISO string; blank → None.
                expiry_raw = row[7] if len(row) > 7 and row[7] is not None else None
                expiry_date = None
                if expiry_raw is not None and str(expiry_raw).strip():
                    try:
                        if hasattr(expiry_raw, "isoformat"):
                            expiry_date = expiry_raw.isoformat()[:10]
                        else:
                            expiry_date = str(expiry_raw).strip()[:10]
                    except Exception:
                        expiry_date = None
                manufacturer = (str(row[8]) if len(row) > 8 and row[8] is not None else "").strip()[:128]

                # ── Label-only fields (cols 10–14) ─────────────────────────
                # These are NOT persisted to rm_boxes — they only flow into
                # the response payload so the JS bulk-print can render them
                # on the printed label. A subsequent reprint from the Label
                # Reprint module will fall back to the OPENING defaults
                # because nothing is stored.
                def _cell_date(v):
                    """Excel date cell or string → 'YYYY-MM-DD' or ''."""
                    if v is None or str(v).strip() == "":
                        return ""
                    try:
                        if hasattr(v, "isoformat"):
                            return v.isoformat()[:10]
                        return str(v).strip()[:10]
                    except Exception:
                        return ""

                lbl_supplier   = (str(row[9])  if len(row) > 9  and row[9]  is not None else "").strip()[:200]
                lbl_grn_no     = (str(row[10]) if len(row) > 10 and row[10] is not None else "").strip()[:64]
                lbl_grn_date   = _cell_date(row[11] if len(row) > 11 else None)
                lbl_invoice_no = (str(row[12]) if len(row) > 12 and row[12] is not None else "").strip()[:64]
                lbl_invoice_dt = _cell_date(row[13] if len(row) > 13 else None)

                def fail(msg):
                    summary["rows_error"] += 1
                    results.append({
                        "row": row_idx, "material": mat_name, "godown": gd_name,
                        "status": "error", "message": msg, "codes": [],
                    })

                if not mat_name:
                    fail("Material Name is empty"); continue
                if not gd_name:
                    fail("Godown is empty"); continue
                if no_of_box <= 0 or no_of_box > 500:
                    fail(f"No. of Packages must be 1-500 (got {no_of_box})"); continue
                if per_box_qty <= 0:
                    fail(f"Qty per Pkg. must be > 0 (got {per_box_qty})"); continue

                mat = mat_lookup.get(mat_name.lower())
                if not mat:
                    fail(f"Material '{mat_name}' not found"); continue
                gd = gd_lookup.get(gd_name.lower())
                if not gd:
                    fail(f"Godown '{gd_name}' not found"); continue

                material_id = int(mat["id"])
                godown_id   = int(gd["id"])
                uom         = uom_raw or (mat["uom"] or "")

                # Optional: replace existing opening-stock boxes for this
                # (material, godown) pair.
                replaced_count = 0
                if replace:
                    old = conn.execute("""
                        SELECT box_id FROM rm_boxes
                        WHERE material_id=%s AND current_godown_id=%s
                          AND source='opening' AND current_status='in_stock'
                    """, (material_id, godown_id)).fetchall()
                    if old:
                        old_ids = [int(r["box_id"]) for r in old]
                        # Mark them as cancelled rather than DELETE — preserves
                        # the audit trail. Also remove from current stock.
                        placeholders = ",".join(["%s"] * len(old_ids))
                        conn.execute(
                            f"UPDATE rm_boxes SET current_status='cancelled' "
                            f"WHERE box_id IN ({placeholders})",
                            tuple(old_ids)
                        )
                        for bid in old_ids:
                            conn.execute(
                                "INSERT INTO rm_box_movements "
                                "(box_id, movement_type, from_godown_id, to_godown_id, "
                                " qty, moved_by, remarks) "
                                "VALUES (%s, 'cancel', %s, NULL, 0, %s, "
                                "        'Replaced by bulk Excel upload')",
                                (bid, godown_id, user)
                            )
                        replaced_count = len(old_ids)
                        summary["boxes_replaced"] += replaced_count

                # Create the boxes
                created_codes = []
                mat_code = derive_mat_code(mat["material_name"] or "")
                for seq in range(1, no_of_box + 1):
                    box_code = allocate_next_box_code(conn)
                    cur = conn.execute("""
                        INSERT INTO rm_boxes
                          (box_code, material_id, material_code, box_seq, total_boxes,
                           per_box_qty, uom, current_godown_id, current_status,
                           source, created_by, batch_num, expiry_date, manufacturer)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'in_stock',
                                'opening', %s, %s, %s, %s)
                    """, (box_code, material_id, mat_code, seq, no_of_box,
                          per_box_qty, uom, godown_id, user,
                          (batch_num or None), expiry_date, (manufacturer or None)))
                    box_id = cur.lastrowid
                    conn.execute("""
                        INSERT INTO rm_box_movements
                          (box_id, movement_type, from_godown_id, to_godown_id,
                           qty, moved_by, remarks)
                        VALUES (%s, 'opening', NULL, %s, %s, %s, %s)
                    """, (box_id, godown_id, per_box_qty, user,
                          remarks or "Opening stock — Excel bulk upload"))
                    created_codes.append(box_code)

                summary["rows_ok"] += 1
                summary["boxes_created"] += len(created_codes)
                msg = f"Created {len(created_codes)} package(s)"
                if replaced_count:
                    msg += f" (replaced {replaced_count} existing)"
                results.append({
                    "row": row_idx, "material": mat["material_name"],
                    "godown": gd["name"],
                    "status": "ok", "message": msg,
                    "codes": created_codes,
                    # label details so the frontend can print labels for this row
                    "uom": uom, "per_box_qty": per_box_qty,
                    "batch_num": batch_num, "expiry_date": expiry_date or "",
                    "manufacturer": manufacturer, "total_boxes": len(created_codes),
                    # Label-only (not stored in DB; the JS bulk-print uses
                    # these to render the printed label only).
                    "label_supplier":    lbl_supplier,
                    "label_grn_no":      lbl_grn_no,
                    "label_grn_date":    lbl_grn_date,
                    "label_invoice_no":  lbl_invoice_no,
                    "label_invoice_dt":  lbl_invoice_dt,
                })

            conn.commit()
            return jsonify({
                "status":  "ok",
                "summary": summary,
                "results": results,
            })
        except Exception as e:
            try: conn.rollback()
            except Exception: pass
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500
        finally:
            try: conn.close()
            except Exception: pass

    # ════════════════════════════════════════════════════════════════════════
    # STOCK TRANSFERS  →  see inventory/inventory_transfers.py
    # ────────────────────────────────────────────────────────────────────────
    # The transfer voucher flow (Out → In Transit → Received) moved to its
    # own module on 16 May 2026 for maintainability. The transfer endpoints
    # are still under /api/inventory_godown/transfers/* (URL unchanged) but
    # are registered by register_inventory_transfers(app), called from
    # app.py alongside this module.
    # ════════════════════════════════════════════════════════════════════════

    print("✅ [InventoryGodown] routes registered (RM phase 1)")
