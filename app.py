# ── Path bootstrap (reorganised HCPERP-style layout, June 2026) ──────────────
# The project was reorganised into folders: core/ models/ services/ modules/
# scripts/. These are added to sys.path so all existing flat imports
# (e.g. `import sampling_portal`, `from fg_routes import fg_bp`,
# `from inventory import inventory_mgmt`) keep resolving unchanged.
import os as _os, sys as _sys
_BASE_DIR = _os.path.dirname(_os.path.abspath(__file__))
for _pkg in ("core", "models", "services", "modules", "scripts"):
    _pkg_path = _os.path.join(_BASE_DIR, _pkg)
    if _pkg_path not in _sys.path:
        _sys.path.insert(0, _pkg_path)
# ─────────────────────────────────────────────────────────────────────────────

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, make_response
from functools import wraps 
from flask_cors import CORS


from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import io


from openpyxl import Workbook, load_workbook
from flask import send_file
import tempfile



# xlwings is optional — Windows + Excel par hi kaam karta hai.
# Agar installed nahi hai (ya Linux VPS hai) to app bina Petty-Cash
# Excel-sync ke chalega, crash nahi hoga.
try:
    import xlwings as xw
except ImportError:
    xw = None
from datetime import datetime, timedelta
import os
import secrets  # for session tokens used by login audit pairing
import re
import subprocess
import requests
import xmltodict
import sampling_portal  # Import the new SQLite bridge module
import cms_portal       # CMS v3 — Voucher-based Cash Management
from rd_sampling_routes import rd_sampling_bp          # R&D Sampling Blueprint
from rd_agent_dashboard import rd_agent_bp             # R&D Agent Dashboard (manual)
from planning_routes import planning_bp                # Planning Dashboard Blueprint
from pm_stock import pm_stock_bp                # PM Stock — Beardo
from pm_stock.pm_stock_audit_routes import pm_audit_bp   # PM Stock — Physical Stock Check (audit)
from fg_routes import fg_bp                            # Finished Goods Registry
from production_dept_routes import production_dept_bp  # Production Department Blueprint
from rm_store.production_initiater_routes import production_initiater_bp  # Production Initiater Blueprint (rm_store package)
from material_request_routes import material_request_bp          # Material Request (MRF) Blueprint
from hr_salary_routes import hr_salary_bp, ensure_hr_salary_tables  # HR Salary Calculation Blueprint
from hcp_stock_routes import hcp_stock_bp                # HCP Stock — Plix Stock Register Blueprint
import hcp_stock_db                                      # HCP Stock — DB schema bootstrap
from qc import qc_bp                                     # QC Dashboard + QC Sampling + Inprocess Approval Form
from procurement import register_procurement   # Procurement Dashboard module (now a package)
# Inventory modules — May 2026: relocated from project root into the
# `inventory/` package. The module-level functions register_inventory_mgmt
# and register_inventory_godown are unchanged, so the register lines below
# work identically. URL prefixes (/inventory_mgmt, /api/inventory_mgmt/*,
# /api/inventory_godown/*) are also unchanged.
from inventory import inventory_mgmt    # Inventory Management module (RM / PM / FG unified)
from inventory import inventory_godown  # Inventory · Godown View + Manage Godowns + Package History
from inventory import inventory_transfers  # Inventory · Stock Transfer Voucher (Out → In Transit → In)
from inventory import inventory_simple_transfer  # Inventory · Simple (manual / non-QR) Stock Transfer Voucher
import general_op        # General Operations module (Godowns, Voucher Types, Numbering)
import backup_system     # Enhanced Backup System (DB + App ZIP)
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import atexit


app = Flask(__name__)
CORS(app)
app.register_blueprint(rd_sampling_bp)       # R&D Sampling routes
app.register_blueprint(rd_agent_bp)          # R&D Agent Dashboard routes
app.register_blueprint(planning_bp)          # Planning Dashboard routes
app.register_blueprint(pm_stock_bp)          # PM Stock routes
app.register_blueprint(pm_audit_bp)          # PM Stock Physical Audit routes
app.register_blueprint(fg_bp)                # Finished Goods routes
app.register_blueprint(production_dept_bp)   # Production Department routes
app.register_blueprint(production_initiater_bp)   # Production Initiater routes
app.register_blueprint(material_request_bp)       # Material Request (MRF) routes
app.register_blueprint(hr_salary_bp)         # HR Salary Calculation routes
ensure_hr_salary_tables()                    # Create hr_employees / periods / days tables
app.register_blueprint(hcp_stock_bp)         # HCP Stock routes
hcp_stock_db.ensure_tables()                 # Create hcp_stock_* tables on first run
app.register_blueprint(qc_bp)                # QC Dashboard / QC Sampling / Inprocess Approval Form

# ===== CRM · Leads module (modules/crm/) =====
from crm import crm_bp, ensure_lead_tables   # CRM Leads blueprint + table bootstrap
from mail import mail_bp                      # Mail Master (CRM) Blueprint -> modules/mail/
app.register_blueprint(crm_bp)               # CRM Leads routes  ->  /crm/leads
app.register_blueprint(mail_bp)              # Mail Master routes -> /mail/master
ensure_lead_tables()                         # lead_* tables ko first-run pe auto-create

# ===== NPD / EPD Projects module (modules/npd/) =====
from npd import npd_bp, ensure_npd_tables    # NPD Projects blueprint + table bootstrap
app.register_blueprint(npd_bp)               # NPD routes  ->  /npd
ensure_npd_tables()                          # npd_* tables ko first-run pe auto-create

# User-administration lookup tables (Department / Designation / User Type / Access Level)
import sys as _ua_sys
from user_admin import user_admin_bp, ensure_user_admin_tables, _fetch_lookups
app.register_blueprint(user_admin_bp)        # /users/lookup/* CRUD endpoints

# Device Access Control — cookie-based device allowlist (desktop + mobile)
from device_access import (
    device_access_bp,
    device_check_at_login,
    get_or_issue_device_cookie,
    get_admin_dashboard_context,
    inactivity_check_at_login,
    log_login_event,
    log_logout_event,
    single_session_enabled,
    set_active_session_token,
    is_session_token_current,
    clear_active_session_token,
)
app.register_blueprint(device_access_bp)     # /api/device/*, /api/admin/devices*, /device_pending, /device_recover

# DB work runs lazily on first use — never at boot — so it can't crash the worker.
_ua_tables_ready = {'done': False}
def _ua_lookups():
    if not _ua_tables_ready['done']:
        try:
            ensure_user_admin_tables()
        except Exception as _e:
            print(f"[user_admin] ensure_user_admin_tables failed: {_e}", file=_ua_sys.stderr)
        finally:
            _ua_tables_ready['done'] = True
    try:
        data = _fetch_lookups()
    except Exception as _e:
        print(f"[user_admin] _fetch_lookups failed: {_e}", file=_ua_sys.stderr)
        return {'department': [], 'designation': [], 'user_type': [], 'access_level': []}
    # If every list is empty, the seed never ran (e.g. case-sensitivity issue on Linux MySQL
    # killed the first attempt silently). Re-try the seed once and re-fetch — uses INSERT IGNORE
    # so it's harmless if everything was actually fine.
    if all(len(v) == 0 for v in data.values()):
        try:
            ensure_user_admin_tables()
            data = _fetch_lookups()
        except Exception as _e:
            print(f"[user_admin] retry seed failed: {_e}", file=_ua_sys.stderr)
    return data

from flask import send_from_directory

# --- CONFIGURATION ---
# Original Petty Cash Excel path on server
SERVER_PATH = r"\\Hcp-server\d\DEPARTMENT COMMON\PURCHASE\PETTY CASH\PETTY CASH FROM 25-26 new.xlsx"
# Tally Configuration — fallback defaults (overridden per-user by saved credentials)
TALLY_URL    = "http://192.168.2.91:9000"
COMPANY_NAME = "HCP Wellness Pvt Ltd (from 1-Apr-25)"

# ── Tally credential helpers ───────────────────────────────────────────────────
import base64

_TALLY_CRED_KEY = b'hcp_tally_cred_key_2025'   # simple XOR key for obfuscation

def _obfuscate(text: str) -> str:
    """XOR-obfuscate + base64 encode — keeps passwords out of plain text in DB."""
    if not text:
        return ''
    key = _TALLY_CRED_KEY
    data = text.encode('utf-8')
    xored = bytes(b ^ key[i % len(key)] for i, b in enumerate(data))
    return base64.b64encode(xored).decode('ascii')

def _deobfuscate(token: str) -> str:
    """Reverse of _obfuscate."""
    if not token:
        return ''
    key = _TALLY_CRED_KEY
    try:
        xored = base64.b64decode(token.encode('ascii'))
        return bytes(b ^ key[i % len(key)] for i, b in enumerate(xored)).decode('utf-8')
    except Exception:
        return ''

def _ensure_tally_creds_table():
    """Create tally_credentials table if it doesn't exist."""
    conn = sampling_portal.get_db_connection()
    if not conn:
        return
    conn.execute("""
        CREATE TABLE IF NOT EXISTS tally_credentials (
            id           INT AUTO_INCREMENT PRIMARY KEY,
            user_id      INT NOT NULL UNIQUE,
            tally_url    VARCHAR(300) NOT NULL,
            company_name VARCHAR(500) NOT NULL,
            tally_user   VARCHAR(200) DEFAULT '',
            tally_pass   VARCHAR(500) DEFAULT '',
            saved_at     DATETIME     DEFAULT CURRENT_TIMESTAMP
                                      ON UPDATE CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

_ensure_tally_creds_table()

def _grant_lunch_coupons_dharmendra():
    """One-time: ensure user 'dharmendra' has lunch_coupons + cash management permissions."""
    conn = sampling_portal.get_db_connection()
    if not conn:
        return
    try:
        row = conn.execute(
            "SELECT id FROM `User_Tbl` WHERE LOWER(username) = %s",
            ('dharmendra',)
        ).fetchone()
        if row:
            uid = row['id']
            for perm in ('page:lunch_coupons', 'page:transaction', 'page:loan', 'page:scrap'):
                conn.execute("""
                    INSERT INTO user_permissions (user_id, perm_key, is_allowed, updated_by)
                    VALUES (%s, %s, 1, 'admin')
                    ON DUPLICATE KEY UPDATE is_allowed = 1, updated_by = 'admin', updated_at = NOW()
                """, (uid, perm))
            conn.commit()
            print("✅ lunch_coupons + cash management permissions granted to dharmendra")
        else:
            print("⚠️  User 'dharmendra' not found — permissions not set")
    except Exception as e:
        print(f"_grant_lunch_coupons_dharmendra error: {e}")
    finally:
        conn.close()

_grant_lunch_coupons_dharmendra()

def _get_tally_creds(user_id: int) -> dict | None:
    """Return saved Tally credentials for user_id, or None if not set."""
    conn = sampling_portal.get_db_connection()
    if not conn:
        return None
    row = conn.execute(
        "SELECT * FROM tally_credentials WHERE user_id=%s", (user_id,)
    ).fetchone()
    conn.close()
    if not row:
        return None
    row = dict(row)
    return {
        'tally_url':    row['tally_url'],
        'company_name': row['company_name'],
        'tally_user':   row.get('tally_user', ''),
        'tally_pass':   _deobfuscate(row.get('tally_pass', '')),
    }

# ── Auto-backup scheduler ──────────────────────────────────────────────────────
def _auto_backup_job():
    """Called by APScheduler — runs full backup (DB dump + App ZIP) to all destinations."""
    result = backup_system.run_full_backup(triggered_by="auto")
    status = result.get("status", "?")
    run_id = result.get("run_id", "")
    print(f"[AutoBackup] {status.upper()} — run_id={run_id}")

_scheduler = BackgroundScheduler(daemon=True)
# Every 2 hours from 07:00 to 19:00 inclusive → 07,09,11,13,15,17,19
_scheduler.add_job(_auto_backup_job, CronTrigger(hour="7-19/2", minute=0),
                   id="hcp_2hourly_backup", replace_existing=True, max_instances=1, coalesce=True)
# Final run at 20:00 (8 PM) — the /2 step lands on odd hours and skips 20:00, so add it explicitly
_scheduler.add_job(_auto_backup_job, CronTrigger(hour=20, minute=0),
                   id="hcp_evening_backup", replace_existing=True, max_instances=1, coalesce=True)
_scheduler.start()
atexit.register(lambda: _scheduler.shutdown(wait=False))
print("✅ Auto-backup scheduler started (every 2h · 07:00 → 20:00)")

app.secret_key = 'hcp_secret_key_123'  # Required for sessions

# ── Register Procurement Dashboard module ────────────────────
register_procurement(app)

# ── Register Inventory Management module (RM / PM / FG) ──────
inventory_mgmt.register_inventory_mgmt(app)
inventory_godown.register_inventory_godown(app)   # Godown View + Box History + Godown CRUD
inventory_transfers.register_inventory_transfers(app)  # Stock Transfer Voucher (Out → In Transit → In)
inventory_simple_transfer.register_inventory_simple_transfer(app)  # Simple Stock Transfer Voucher (manual / non-QR)

# ── Register General Operations module ───────────────────────
general_op.register_general_op(app)
from fg_routes import ensure_fg_tables; ensure_fg_tables()  # Create FG_Names table

# ── Service Worker must be served from root scope ──────────────────
@app.route('/sw.js')
def service_worker():
    return send_from_directory('static', 'sw.js',
                               mimetype='application/javascript')


# ---------------- SINGLE-SESSION ENFORCEMENT (before_request) ----------------
# If the global single-login toggle is on AND the current request belongs to
# a logged-in session, verify the session token still matches what's stored in
# the DB. If a newer login replaced it, drop this session and bounce to login
# with a friendly message.
#
# Why before_request: catches EVERY request the kicked device might make
# (pages, APIs, AJAX) without needing per-route changes.
#
# Skipped for the login/logout pages themselves and for static assets.
@app.before_request
def _enforce_single_session():
    try:
        # Cheap exit for unauthenticated requests
        if not session.get('logged_in'):
            return None
        # Skip paths where bouncing would loop or break assets
        p = (request.path or '')
        if p.startswith('/static/') or p in ('/login', '/logout', '/favicon.ico'):
            return None
        # Skip the polling and check endpoints used by the pending/recovery pages
        if p.startswith('/api/device/check-status') or p.startswith('/api/device/recover'):
            return None
        # Only enforce when the feature is ON
        if not single_session_enabled():
            return None
        uid = session.get('user_id')
        tok = session.get('single_sess_token')
        # If feature is on but this session predates the feature (no token in
        # session yet), let it through — it'll get a token on next login.
        if not uid or not tok:
            return None
        if not is_session_token_current(uid, tok):
            # Stale session — newer login replaced it. Kick this one out.
            session.clear()
            if p.startswith('/api/'):
                return jsonify({
                    'status': 'error',
                    'kicked': True,
                    'message': 'You were signed out because this account signed in from another device.'
                }), 401
            return render_template(
                'login.html',
                error='You were signed out because this account signed in from another device.'
            )
    except Exception:
        # On any error, fail open — never block traffic because of this check
        return None
    return None


# ---------------- LOGIN DECORATOR ----------------
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function



# ── Shared permission helpers (imported to avoid circular deps with blueprints) ──
from portal_helpers import (
    ROLE_DEFAULT_PAGES, _USER_PAGE_GRANTS,
    _get_all_permissions, _user_allowed_pages,
    can_access, can_do, _denied, _prod_role
)

def section_required(section_name):
    """Legacy decorator: maps old section names to page keys."""
    _map = {'transaction':'transaction','loan':'loan','scrap':'scrap',
            'sampling':'rd_sampling','qc_sampling':'qc_sampling'}
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not session.get('logged_in'):
                return redirect(url_for('login'))
            page_key = _map.get(section_name, section_name)
            if not can_access(page_key):
                return _denied(section_name)
            return f(*args, **kwargs)
        return wrapper
    return decorator





# Simple User Database for local portal


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        uid = request.form.get('UID')
        pwd = request.form.get('Pwd')

        user_record = sampling_portal.get_user_for_auth(uid)

        if user_record is None:
            log_login_event('login_fail', username=uid, fail_reason='user_not_found')
            return render_template('login.html', error="User not found")

        # Verify hashed password
        pwd_hash = sampling_portal.hash_password(pwd)
        if user_record['password_hash'] != pwd_hash:
            log_login_event('login_fail', user_record=user_record, fail_reason='wrong_password')
            return render_template('login.html', error="Wrong password")

        # ── Inactivity auto-lock check ─────────────────────────────────────
        # Runs AFTER password verification, BEFORE device check.
        # If user has been inactive > threshold days, mark locked and refuse.
        # Admin accounts (user_type='admin') are exempt.
        lock_verdict, lock_info = inactivity_check_at_login(user_record)
        if lock_verdict == 'locked':
            days = (lock_info or {}).get('days', 0)
            log_login_event('login_fail', user_record=user_record,
                            fail_reason=f'account_locked ({days} days inactive)')
            return render_template('login.html',
                error=f"Your account is locked due to {days} days of inactivity. "
                      "Please contact the system administrator to unlock it.")

        # ── Device Access Control ──────────────────────────────────────────
        # Run AFTER password is verified, BEFORE setting the session.
        # Returns one of: ('allow', did) | ('pending', did) | ('blocked', did)
        verdict, did = device_check_at_login(user_record)
        if verdict == 'pending':
            log_login_event('login_fail', user_record=user_record, fail_reason='device_pending')
            resp = make_response(redirect(url_for('device_access.device_pending_page')))
            if did:
                resp.set_cookie('hcp_device_id', did,
                                max_age=60*60*24*365*2, httponly=True,
                                samesite='Strict', secure=request.is_secure)
            return resp
        if verdict == 'blocked':
            log_login_event('login_fail', user_record=user_record, fail_reason='device_blocked')
            return render_template('login.html',
                error="This device has been blocked. Contact admin or use a recovery code.")
        # verdict == 'allow' → fall through to normal session setup below.

        # Check if user must reset password on first login
        if user_record['must_reset_password'] == 1:
            # Store minimal session for the reset page only
            session['reset_user_id'] = user_record['id']
            session['reset_username'] = user_record['username']
            return redirect(url_for('force_reset_password_page'))

        # Set session with enriched user data
        session['logged_in'] = True
        session['UID'] = user_record['username']
        session['User_Type'] = user_record['user_type']
        session['User_Name'] = user_record['full_name'] or user_record['username']
        session['user_id'] = user_record['id']
        session['role'] = user_record['role']
        session['department'] = user_record['department']
        session['email'] = user_record['email']
        session['profile_photo'] = user_record['profile_photo']
        session['is_dept_head'] = bool(user_record.get('is_dept_head', 0))
        # Mint a per-session token for audit-log pairing (login → logout).
        session['audit_token'] = secrets.token_hex(12)

        # Update last login timestamp
        sampling_portal.update_last_login(uid)
        # Audit: log the successful login (paired by audit_token with the eventual logout).
        log_login_event('login_success', user_record=user_record,
                        session_token=session['audit_token'])

        # ── Single-session-only enforcement ───────────────────────────────
        # If the global toggle is ON, stamp a fresh token onto the user row.
        # This invalidates any previous session for this same user account,
        # so the next request from the OLD device will fail the before_request
        # check and that user gets bounced back to /login.
        if single_session_enabled():
            sess_tok = secrets.token_hex(16)
            session['single_sess_token'] = sess_tok
            try:
                set_active_session_token(user_record['id'], sess_tok)
            except Exception as _e:
                # Never let this block a successful login
                pass

        role = user_record['user_type']

        # RD → R&D Sampling
        if role == "RD":
            return redirect(url_for('rd_sampling_page'))

        # QC → QC Dashboard
        if role in ('QC', 'qc_common', 'QC_Common'):
            return redirect(url_for('qc.qc_dashboard_page'))

        # Production → Production Department
        if (role or '').lower() == 'production':
            return redirect(url_for('production_dept.production_dept_page'))

        # Stores / RM_Store → Production Initiater
        prod = _prod_role()
        if prod == 'rm_store':
            return redirect(url_for('production_initiater.production_initiater_page'))

        # Planning → Planning Dashboard
        if (role or '').lower() == 'planning':
            return redirect('/planning_dashboard')

        # PM → PM Stock page
        if (role or '').lower() == 'pm':
            return redirect(url_for('pm_stock.pm_stock_page'))

        # All others (Purchase, admin, User, etc.) → Portal home
        return redirect(url_for('home'))

    return render_template('login.html')



@app.route('/force_reset_password', methods=['GET', 'POST'])
def force_reset_password_page():
    """First-login forced password reset page."""
    user_id = session.get('reset_user_id')
    username = session.get('reset_username')

    if not user_id:
        return redirect(url_for('login'))

    if request.method == 'POST':
        new_pwd = request.form.get('new_password', '')
        confirm_pwd = request.form.get('confirm_password', '')

        # Server-side validation: 4+ letters (1 uppercase) + 1 special + 3 numbers = min 8 chars
        import re
        letters = re.findall(r'[a-zA-Z]', new_pwd)
        uppercase = re.findall(r'[A-Z]', new_pwd)
        digits = re.findall(r'[0-9]', new_pwd)
        specials = re.findall(r'[^a-zA-Z0-9]', new_pwd)

        if new_pwd != confirm_pwd:
            return render_template('force_reset_password.html',
                                   username=username,
                                   error="Passwords do not match")

        if len(letters) < 4:
            return render_template('force_reset_password.html',
                                   username=username,
                                   error="Password must contain at least 4 letters")

        if len(uppercase) < 1:
            return render_template('force_reset_password.html',
                                   username=username,
                                   error="At least 1 letter must be uppercase")

        if len(specials) < 1:
            return render_template('force_reset_password.html',
                                   username=username,
                                   error="Password must contain at least 1 special character")

        if len(digits) < 3:
            return render_template('force_reset_password.html',
                                   username=username,
                                   error="Password must contain at least 3 numbers")

        # All validations passed — update password
        success, msg = sampling_portal.force_reset_password(user_id, new_pwd)

        if success:
            # Clear reset session and redirect to login
            session.pop('reset_user_id', None)
            session.pop('reset_username', None)
            return render_template('login.html',
                                   error=None,
                                   success="Password set successfully! Please login with your new password.")
        else:
            return render_template('force_reset_password.html',
                                   username=username,
                                   error=msg)

    return render_template('force_reset_password.html', username=username)



@app.route('/logout')
def logout():
    # Audit: log the logout BEFORE clearing the session so we still know who left.
    try:
        log_logout_event(
            username=session.get('UID') or '',
            session_token=session.get('audit_token') or ''
        )
    except Exception:
        pass
    # Clear the single-session token so a future login from any device starts fresh.
    try:
        uid_for_clear = session.get('user_id')
        if uid_for_clear:
            clear_active_session_token(uid_for_clear)
    except Exception:
        pass
    session.clear()
    return redirect(url_for('login'))





# =====================================================
# R&D SAMPLING PAGE
# =====================================================

@app.route('/rd_sampling')
@login_required
def rd_sampling_page():
    if not can_access('rd_sampling'): return _denied('R&D Sampling')

    page = request.args.get('page', 1, type=int)
    per_page = 50
    offset = (page - 1) * per_page

    conn = sampling_portal.get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM rd_sampling_requests
        ORDER BY id DESC
        LIMIT %s OFFSET %s
    """, (per_page, offset))
    rows = cursor.fetchall()

    records = [dict(row) for row in rows]

    cursor.execute("""
        SELECT COUNT(*) FROM rd_sampling_requests
    """)
    total = cursor.fetchone()[0]

    conn.close()

    total_pages = (total // per_page) + (1 if total % per_page else 0)

    # Resolve formulation permissions for this user
    user_role = session.get("User_Type", "")
    user_id   = session.get("user_id")
    is_admin  = user_role.lower() == "admin"

    if is_admin:
        can_fml_request = can_fml_print = can_fml_approve = True
    elif user_id:
        try:
            _fml_perms   = sampling_portal.get_user_permissions(user_id) or {}
            can_fml_request = bool(_fml_perms.get("rd_fml_request"))
            can_fml_print   = bool(_fml_perms.get("rd_fml_print"))
            can_fml_approve = bool(_fml_perms.get("rd_fml_approve"))
        except Exception:
            can_fml_request = can_fml_print = can_fml_approve = False
    else:
        can_fml_request = can_fml_print = can_fml_approve = False

    return render_template(
        "rd_sampling.html",
        records=records,
        page=page,
        total_pages=total_pages,
        role=user_role,
        can_fml_request=can_fml_request,
        can_fml_print=can_fml_print,
        can_fml_approve=can_fml_approve,
    )




# =====================================================
# QC SAMPLING PAGE
# =====================================================
# NOTE: The /qc_sampling page route now lives in qc/qc_routes.py (blueprint qc_bp).
# Template moved to templates/qc/qc_sampling.html.
# The /save_qc_sampling, /delete_qc_sampling, /import_qc_sampling, and /trs_view
# handlers below are kept here because they're consumed by the existing
# templates/qc/qc_sampling.html JS code unchanged.


@app.route('/save_qc_sampling', methods=['POST'])
@login_required
def save_qc_sampling_handler():

    role = session.get("User_Type")

    if role not in ('admin', 'QC', 'Purchase') and not can_do("qc_add"):
        return jsonify({"status": "error", "message": "Access Denied"})

    data = request.json

    try:
        sampling_portal.save_qc_sampling(
            data,
            role,
            session.get("UID")
        )
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})



@app.route('/delete_qc_sampling', methods=['POST'])
@login_required
def delete_qc_sampling_route():

    role = session.get("User_Type")

    if role not in ('admin', 'QC', 'Purchase') and not can_do("qc_edit"):
        return jsonify({"status": "error", "message": "Access Denied"})

    data = request.json

    try:
        sampling_portal.delete_qc_sampling(
            data.get("id")
        )
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})




@app.route('/import_qc_sampling', methods=['POST'])
@login_required
def import_qc_sampling():

    role = session.get("User_Type")

    if role not in ('admin', 'QC', 'Purchase') and not can_do("qc_add"):
        return jsonify({"status":"error","message":"Access Denied"})

    rows = request.json

    try:
        sampling_portal.import_qc_sampling_data(
            rows,
            session.get("UID")
        )

        return jsonify({
            "status":"success",
            "message":"QC Excel imported successfully"
        })

    except Exception as e:
        return jsonify({
            "status":"error",
            "message":str(e)
        })


@app.route('/trs_view/<int:record_id>')
@login_required
def trs_view(record_id):
    if not can_access('trs_view'): return _denied('TRS View')

    appearance = request.args.get("appearance", "")

    conn = sampling_portal.get_db_connection()
    row = conn.execute(
        "SELECT * FROM qc_sampling_records WHERE id=%s",
        (record_id,)
    ).fetchone()
    conn.close()

    if not row:
        return "Record not found"

    return render_template(
        "trs_view.html",
        r=row,
        appearance=appearance
    )
       

@app.route('/dashboard')
@login_required
def index():
    return ("Not Found", 404)  # canteen page removed

    role = session.get('User_Type')

    selected_month = request.args.get(
        "month",
        datetime.now().strftime("%Y-%m")
    )

    conn = sampling_portal.get_db_connection()

    # =========================
    # EMPLOYEES + LOG DATA
    # =========================
    employees = conn.execute("SELECT * FROM canteen_employees").fetchall()

    lunches = conn.execute("""
        SELECT * FROM canteen_lunch_entries
        WHERE LEFT(date,7) = %s
        ORDER BY date DESC
    """, (selected_month,)).fetchall()

    expenses_list = conn.execute("""
        SELECT * FROM canteen_expenses
        WHERE LEFT(date,7) = %s
        ORDER BY date DESC
    """, (selected_month,)).fetchall()

    payments_list = conn.execute("""
        SELECT * FROM canteen_payments
        WHERE LEFT(date,7) = %s
        ORDER BY date DESC
    """, (selected_month,)).fetchall()

    # =========================
    # CALCULATE INCOME
    # =========================
    income_row = conn.execute("""
        SELECT SUM(amount) FROM canteen_payments
        WHERE LEFT(date,7) = %s
    """, (selected_month,)).fetchone()

    income = income_row[0] if income_row[0] else 0

    # =========================
    # CALCULATE EXPENSES
    # =========================
    expense_row = conn.execute("""
        SELECT SUM(amount) FROM canteen_expenses
        WHERE LEFT(date,7) = %s
    """, (selected_month,)).fetchone()

    expenses = expense_row[0] if expense_row[0] else 0

    # =========================
    # LOAD / CREATE MONTH SUMMARY
    # =========================
    summary = conn.execute("""
    SELECT * FROM canteen_monthly_summary
        WHERE month = %s
           OR month = DATE_FORMAT(%s, '%%b-%%y')
    """, (selected_month, selected_month)).fetchone()

    if summary:
        opening_balance = summary["opening_balance"]
    else:
        opening_balance = 0
        conn.execute("""
            INSERT INTO canteen_monthly_summary (month, opening_balance)
            VALUES (%s, %s)
        """, (selected_month, 0))
        conn.commit()

    # =========================
    # CLOSING CALCULATION
    # =========================
    balance = opening_balance + income - expenses

    # =========================
    # UPDATE MONTHLY SUMMARY
    # =========================
    conn.execute("""
        UPDATE canteen_monthly_summary
        SET income = %s,
            expenses = %s,
            closing_balance = %s
        WHERE month = %s
    """, (income, expenses, balance, selected_month))

    conn.commit()

    conn.close()

    # =========================
    # OUTSTANDING
    # =========================
    total_outstanding = sampling_portal.get_total_outstanding(selected_month)

    if total_outstanding is None:
        total_outstanding = 0

    return render_template(
        "canteen.html",
        user_name=session.get("User_Name"),
        user_type=role,

        employees=employees,
        lunches=lunches,
        expenses_list=expenses_list,
        payments_list=payments_list,

        selected_month=selected_month,

        opening_balance=opening_balance,
        income=income,
        expenses=expenses,
        balance=balance,
        dashboard_outstanding=total_outstanding
    )

# ----------------------------------------------------------------
# EXCEL HELPER FUNCTIONS (RETAINED FROM ORIGINAL)
# ----------------------------------------------------------------

def get_excel_instance():
    """Connects to the Petty Cash Excel workbook using xlwings."""
    if xw is None:
        print("xlwings not installed — Petty Cash Excel sync disabled.")
        return None
    try:
        try:
            return xw.books['PETTY CASH FROM 25-26 new.xlsx']
        except:
            return xw.Book(SERVER_PATH)
    except Exception as e:
        print(f"Excel Connection Error: {e}")
        return None

def get_transaction_data():
    """Fetches petty cash transactions from Excel."""
    wb = get_excel_instance()
    if not wb: return [], 1
    sheet = wb.sheets["EXPENSE DETAILS"]
    last_row = sheet.range('A' + str(sheet.cells.last_cell.row)).end('up').row
    vouchers = sheet.range(f"B5:B{last_row}").value
    max_v = 0
    if vouchers:
        if not isinstance(vouchers, list): vouchers = [vouchers]
        for v in vouchers:
            try:
                val = int(float(v))
                if val > max_v: max_v = val
            except: continue
    
    data = sheet.range(f"A5:F{last_row}").value if last_row >= 5 else []
    if last_row == 5 and data: data = [data]
    
    rows = []
    for i, row in enumerate(data):
        if row and row[0]:
            d_obj = row[0] if isinstance(row[0], datetime) else datetime.now()
            rows.append({
                "row_idx": i + 5, 
                "date_display": d_obj.strftime("%d-%m-%Y"), 
                "voucher": row[1], 
                "particulars": str(row[2] or ""), 
                "credit": row[3] or 0, 
                "debit": row[4] or 0, 
                "balance": row[5] or 0
            })
    return rows[::-1], max_v + 1

def get_loan_data():
    """Fetches staff loan/advance data from Excel."""
    wb = get_excel_instance()
    if not wb: return [], 0
    sheet = wb.sheets["ADV. EXP CASH GIVEN"]
    total_loan = sheet.range('D2').value or 0
    last_row = sheet.range('C' + str(sheet.cells.last_cell.row)).end('up').row
    data = sheet.range(f"C5:D{last_row}").value if last_row >= 5 else []
    if last_row == 5 and data: data = [data]
    rows = [{"row_idx": i + 5, "name": str(row[0] or ""), "amount": row[1] or 0} 
            for i, row in enumerate(data or []) if row and row[0]]
    return rows[::-1], total_loan

def get_scrap_data():
    """Fetches scrap history data from Excel."""
    wb = get_excel_instance()
    if not wb: return []
    sheet = wb.sheets["SCRAP HISAB - JAGDAMBA"]
    last_row = sheet.range('A' + str(sheet.cells.last_cell.row)).end('up').row
    data = sheet.range(f"A6:E{last_row}").value if last_row >= 6 else []
    if last_row == 6 and data: data = [data]
    rows = []
    for i, row in enumerate(data or []):
        if row and row[0]:
            d_str = row[1].strftime("%Y-%m-%d") if isinstance(row[1], datetime) else str(row[1])
            rows.append({
                "row_idx": i + 6, 
                "receipt": str(row[0] or ""), 
                "date": d_str,
                "credit": row[2] or 0, 
                "debit": row[3] or 0, 
                "balance": row[4] or 0
            })
    return rows[::-1]
    
def add_new_user(user, pwd):
    conn = sampling_portal.get_db_connection()
    try:
        pwd_hash = sampling_portal.hash_password(pwd)
        with conn:
            conn.execute('INSERT INTO `User_Tbl` (username, password_hash, user_type) VALUES (%s, %s, %s)', 
                         (user, pwd_hash, 'user'))
        print(f"User {user} created successfully!")
    except:
        print("User already exists!")
    finally:
        conn.close()

# ----------------------------------------------------------------
# FLASK ROUTES
# ----------------------------------------------------------------

@app.route('/')
@login_required
def home():

    role = session.get('User_Type')

    # ── Role-based direct redirects ───────────────────────────────────────────
    if role == "RD":
        return redirect(url_for('rd_sampling_page'))
    if role in ('QC', 'qc_common', 'QC_Common'):
        return redirect(url_for('qc.qc_dashboard_page'))
    if (role or '').lower() == 'production':
        return redirect(url_for('production_dept.production_dept_page'))
    if (role or '').lower() == 'planning':
        return redirect('/planning_dashboard')
    if (role or '').lower() == 'pm':
        return redirect(url_for('pm_stock.pm_stock_page'))
    if _prod_role() == 'rm_store':
        return redirect(url_for('production_initiater.production_initiater_page'))

    # ── Everyone else sees the portal dashboard ───────────────────────────────
    all_pages = _user_allowed_pages()
    # `allowed_sections` is consumed by index.html to gate every KPI card.
    # Previously this was filtered to just 4 cash keys, which silently hid every
    # other module card even when Access Control granted it. Pass the full set.
    allowed_sections = list(all_pages)

    return render_template(
        'index.html',
        role=role,
        user_name=session.get('User_Name'),
        allowed_sections=allowed_sections,
        # ── Index Access Control modal (admin only) ───────────────────
        is_admin=((role or '').lower() == 'admin'),
        index_perm_keys=INDEX_PERM_KEYS,
        # ── Device Access Control KPI card (admin only) ───────────────
        **get_admin_dashboard_context(),
    )


# ══════════════════════════════════════════════════════════════════════════════
# CASH MANAGEMENT PAGE  (/cash_management?section=transaction|loan|scrap)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/cash_management')
@login_required
def cash_management_page():

    role = session.get('User_Type')

    _INDEX_SECTION_ORDER = ['transaction', 'loan', 'scrap']
    all_pages = _user_allowed_pages()
    allowed_sections = [s for s in _INDEX_SECTION_ORDER if s in all_pages]

    if not allowed_sections:
        return _denied('Cash Management')

    section = request.args.get('section')
    if not section or section not in allowed_sections:
        section = allowed_sections[0]

    trans_list     = []
    loan_list      = []
    scrap_list     = []
    next_v         = None
    total_loan_amt = None
    cash_in_hand   = None
    safe_cash      = None

    wb = get_excel_instance()
    if not wb:
        return render_template(
            'cash_management.html',
            role=role,
            allowed_sections=allowed_sections,
            entries=[], next_voucher=None,
            loan_entries=[], total_loan=None,
            scrap_entries=[],
            cash_in_hand=None, safe_cash=None,
            active_section=section,
            excel_error="Excel file unavailable — make sure the Petty Cash file is open on the server.",
        )

    try:
        sh_exp = wb.sheets["EXPENSE DETAILS"]

        if 'transaction' in allowed_sections:
            trans_list, next_v = get_transaction_data()
            cash_in_hand = sh_exp.range('F2').value or 0
            safe_cash    = sh_exp.range('D3').value or 0

        if 'loan' in allowed_sections:
            loan_list, total_loan_amt = get_loan_data()

        if 'scrap' in allowed_sections:
            scrap_list = get_scrap_data()

    except Exception as e:
        import traceback; traceback.print_exc()
        return render_template(
            'cash_management.html',
            role=role,
            allowed_sections=allowed_sections,
            entries=[], next_voucher=None,
            loan_entries=[], total_loan=None,
            scrap_entries=[],
            cash_in_hand=None, safe_cash=None,
            active_section=section,
            excel_error=f"Excel error — make sure the Petty Cash file is open on the server. ({e})",
        )

    return render_template(
        'cash_management.html',
        role=role,
        allowed_sections=allowed_sections,
        entries=trans_list,
        next_voucher=next_v,
        loan_entries=loan_list,
        total_loan=total_loan_amt,
        scrap_entries=scrap_list,
        cash_in_hand=cash_in_hand,
        safe_cash=safe_cash,
        active_section=section,
    )

# ── END CASH MANAGEMENT PAGE ──────────────────────────────────────────────────


@app.route('/save', methods=['POST'])
def save_entry():
    """
    Smart Switch: Directs data to either Excel or SQLite based on type.
    """
    data = request.json
    t = data.get('type')
    


    # Handle Original Excel Logic (Transaction, Loan, Scrap)
    wb = get_excel_instance()
    if not wb: return jsonify({"status": "error", "message": "Excel Unavailable"})
    
    sheet = wb.sheets["SCRAP HISAB - JAGDAMBA"] if t == 'scrap' else \
           (wb.sheets["ADV. EXP CASH GIVEN"] if t == 'loan' else wb.sheets["EXPENSE DETAILS"])
    
    nr = int(data.get('row_idx')) if data.get('row_idx') else \
         sheet.range(('C' if t == 'loan' else 'A') + str(sheet.cells.last_cell.row)).end('up').row + 1
    
    if t == 'loan':
        sheet.range(f'C{nr}').value = data['name']
        sheet.range(f'D{nr}').value = float(data.get('added_loan', 0)) + float(data.get('prev_loan', 0))
    elif t == 'scrap':
        sheet.range(f'A{nr}').value = data['receipt']
        sheet.range(f'B{nr}').value = datetime.strptime(data['date'], '%Y-%m-%d')
        sheet.range(f'C{nr}').value = float(data['credit'] or 0)
        sheet.range(f'D{nr}').value = float(data['debit'] or 0)
    else:
        # Transaction Logic
        if not data.get('row_idx'): sheet.range(f'A{nr}').value = datetime.now()
        sheet.range(f'B{nr}').value = data['voucher']
        sheet.range(f'C{nr}').value = data['particulars']
        sheet.range(f'D{nr}').value = float(data['credit'] or 0)
        sheet.range(f'E{nr}').value = float(data['debit'] or 0)
    
    
    
    wb.save()
    return jsonify({"status": "success"})
    

@app.route('/delete_general', methods=['POST'])
def delete_general():
    """Handles deletion for Transactions, Loans, and Scrap from Excel."""
    data = request.json
    t = data.get('type')
    row_idx = data.get('row_idx')

    if not row_idx:
        return jsonify({"status": "error", "message": "No row selected"})

    # 1. Connect to the Excel workbook
    wb = get_excel_instance()
    if not wb: 
        return jsonify({"status": "error", "message": "Excel Unavailable"})
    
    # 2. Select the correct sheet based on the 'type'
    if t == 'scrap':
        sheet = wb.sheets["SCRAP HISAB - JAGDAMBA"]
    elif t == 'loan':
        sheet = wb.sheets["ADV. EXP CASH GIVEN"]
    else:
        sheet = wb.sheets["EXPENSE DETAILS"]

    try:
        # 3. Delete the specific row
        # row_idx comes from the frontend as the actual Excel row number
        target_row = int(row_idx)
        sheet.range(f"{target_row}:{target_row}").delete()
        
        # 4. Save the changes to the file
        wb.save()
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})



@app.route('/create_user', methods=['GET', 'POST'])
def create_user():

    # Must be logged in
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    # Must be admin
    if (session.get('User_Type') or '').lower() != 'admin':
        return "<h2>Access Denied</h2>"

    if request.method == 'POST':

        uid = request.form.get('username', '').strip()
        pwd = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '').strip()
        mobile = request.form.get('mobile', '').strip()
        employee_id = request.form.get('employee_id', '').strip()
        department = request.form.get('department', '').strip()
        designation = request.form.get('designation', '').strip()
        role = request.form.get('role', '').strip()
        user_type = request.form.get('user_type', 'user').strip()
        access_level = request.form.get('access_level', '1').strip()

        # Validation
        if not uid or not pwd:
            return render_template('create_user.html', lookups=_ua_lookups(),
                                   error="Username and Password are required")

        if pwd != confirm:
            return render_template('create_user.html', lookups=_ua_lookups(),
                                   error="Passwords do not match")

        if len(pwd) < 4:
            return render_template('create_user.html', lookups=_ua_lookups(),
                                   error="Password must be at least 4 characters")

        # Handle profile photo upload
        profile_photo_path = ''
        if 'profile_photo' in request.files:
            photo = request.files['profile_photo']
            if photo.filename:
                import os
                upload_dir = os.path.join(app.root_path, 'static', 'profile_photos')
                os.makedirs(upload_dir, exist_ok=True)
                ext = os.path.splitext(photo.filename)[1]
                filename = f"{uid}{ext}"
                photo.save(os.path.join(upload_dir, filename))
                profile_photo_path = f"/static/profile_photos/{filename}"

        user_data = {
            'username': uid,
            'password': pwd,
            'full_name': full_name,
            'email': email,
            'mobile': mobile,
            'employee_id': employee_id,
            'department': department,
            'designation': designation,
            'role': role,
            'user_type': user_type,
            'access_level': int(access_level) if access_level.isdigit() else 1,
            'profile_photo': profile_photo_path
        }

        success, message = sampling_portal.create_new_user(
            user_data, created_by=session.get('UID', 'admin')
        )

        if success:
            return render_template('create_user.html', lookups=_ua_lookups(), success=message)
        else:
            return render_template('create_user.html', lookups=_ua_lookups(), error=message)

    return render_template('create_user.html', lookups=_ua_lookups())


# ── USER MANAGEMENT API ROUTES ──

@app.route('/manage_users')
@login_required
def manage_users():
    if not can_access('manage_users'): return _denied('Manage Users')
    if (session.get('User_Type') or '').lower() != 'admin':
        return "<h2>Access Denied</h2>"
    users = sampling_portal.get_all_users()
    return render_template('manage_users.html',
                           users=users,
                           role=session.get('User_Type'),
                           lookups=_ua_lookups(),
                           single_session_enabled=single_session_enabled())


@app.route('/api/users', methods=['GET'])
@login_required
def api_get_users():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    users = sampling_portal.get_all_users()
    return jsonify(users)


@app.route('/api/users/<int:user_id>', methods=['GET'])
@login_required
def api_get_user(user_id):
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    user = sampling_portal.get_user_by_id(user_id)
    if user:
        return jsonify(user)
    return jsonify({'error': 'User not found'}), 404


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
def api_update_user(user_id):
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    data = request.get_json()
    success, msg = sampling_portal.update_user(user_id, data)
    return jsonify({'success': success, 'message': msg})


@app.route('/api/users/<int:user_id>/toggle', methods=['POST'])
@login_required
def api_toggle_user(user_id):
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    data = request.get_json()
    is_active = data.get('is_active', True)
    success, msg = sampling_portal.toggle_user_active(user_id, is_active)
    return jsonify({'success': success, 'message': msg})


@app.route('/api/users/<int:user_id>/reset_password', methods=['POST'])
@login_required
def api_reset_password(user_id):
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    data = request.get_json()
    new_pwd = data.get('new_password', '')
    if len(new_pwd) < 4:
        return jsonify({'success': False, 'message': 'Password must be at least 4 characters'})
    success, msg = sampling_portal.reset_user_password(user_id, new_pwd)
    return jsonify({'success': success, 'message': msg})

@app.route('/api/access_control/save', methods=['POST'])
@login_required
def api_access_control_save():

    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({"status":"error","message":"Access denied"}),403

    data = request.json

    user_id = data.get("user_id")
    permissions = data.get("permissions",{})

    sampling_portal.save_user_permissions(user_id, permissions)

    return jsonify({"status":"success"})



@app.route('/tally_daybook', methods=['POST'])
def get_tally_daybook():
    """Original Tally Integration Logic."""
    auth = request.json
    xml_req = f"""<ENVELOPE>
        <HEADER><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>
        <BODY>
            <EXPORTDATA>
                <REQUESTDESC>
                    <REPORTNAME>Day Book</REPORTNAME>
                    <STATICVARIABLES>
                        <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
                        <SVCURRENTCOMPANY>{COMPANY_NAME}</SVCURRENTCOMPANY>
                    </STATICVARIABLES>
                </REQUESTDESC>
            </EXPORTDATA>
        </BODY>
    </ENVELOPE>"""
    try:
        if auth.get('user') and auth.get('pass'):
            resp = requests.post(TALLY_URL, data=xml_req, auth=(auth['user'], auth['pass']), timeout=5)
        else:
            resp = requests.post(TALLY_URL, data=xml_req, timeout=5)
        return jsonify({"status": "success", "data": xmltodict.parse(resp.text)})
    except:
        return jsonify({"status": "error", "message": "Tally Connection Error."})

@app.route('/open_calc')
def open_calc():
    """Launches the Windows Calculator."""
    os.startfile('calc.exe')
    return jsonify({"status": "success"})
    
from waitress import serve


# ============================================================
# CANTEEN API ROUTES
# ============================================================

# ============================================================
# MONTH FORMAT + ENSURE RECORD
# ============================================================

from datetime import datetime

def format_month_label(month_yyyy_mm):
    """
    Convert '2026-01' → 'Jan-26'
    """
    dt = datetime.strptime(month_yyyy_mm + "-01", "%Y-%m-%d")
    return dt.strftime("%b-%y")


def ensure_month_record(conn, month_yyyy_mm):

    month_label = format_month_label(month_yyyy_mm)

    row = conn.execute("""
        SELECT id FROM canteen_monthly_summary
        WHERE month=%s
    """, (month_label,)).fetchone()

    if not row:
        conn.execute("""
            INSERT INTO canteen_monthly_summary (month)
            VALUES (%s)
        """, (month_label,))
        conn.commit()

    return month_label


# ============================================================
# ADMIN ONLY - EXPORT EMPLOYEE TABLE TO EXCEL
# ============================================================

import pandas as pd
from flask import send_file
from io import BytesIO





@app.route('/admin/export-canteen-employees')
@login_required
def admin_export_canteen_employees():

    # 🔒 ADMIN SECURITY CHECK
    if session.get("User_Type") != "admin":
        return "<h3>Access Denied</h3>"

    conn = sampling_portal.get_db_connection()

    df = pd.read_sql_query("""
        SELECT 
            emp_id AS "Employee ID",
            emp_name AS "Employee Name",
            department AS "Department",
            category AS "Category",
            contact_number AS "Contact Number",
            opening_balance AS "Opening Balance"
        FROM canteen_employees
        ORDER BY emp_id
    """, conn)

    conn.close()

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Canteen Employees')

    output.seek(0)

    return send_file(
        output,
        download_name="canteen_employees.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/api/canteen/delete-employee', methods=['POST'])
@login_required
def delete_canteen_employee():

    data = request.json
    emp_id = data.get("id")

    if not emp_id:
        return jsonify({"status": "error", "message": "No employee selected."})

    conn = sampling_portal.get_db_connection()
    conn.execute(
        "DELETE FROM canteen_employees WHERE emp_id=%s",
        (emp_id,)
    )
    conn.commit()
    conn.close()

    return jsonify({"status": "success"})

@app.route('/canteen')
@login_required
def canteen_page():
    return ("Not Found", 404)  # canteen feature removed
    if not can_access('canteen'): return _denied('Canteen')

    role = session.get('User_Type')

    # -------------------------
    # GET FILTERS
    # -------------------------
    selected_month = request.args.get("month")
    if not selected_month:
        selected_month = datetime.now().strftime("%Y-%m")

    selected_date = request.args.get("date")
    if not selected_date:
        selected_date = None

    month_start = selected_month + "-01"
    month_end = selected_month + "-31"

    conn = sampling_portal.get_db_connection()

    
    # ---------------------------------
    # FORMAT CURRENT MONTH LABEL
    # ---------------------------------
    month_label = format_month_label(selected_month)

    # ---------------------------------
    # CHECK IF MONTH RECORD EXISTS
    # ---------------------------------
    row = conn.execute("""
        SELECT opening_balance
        FROM canteen_monthly_summary
        WHERE month=%s
    """, (selected_month,)).fetchone()

    # If record exists → opening already set
    if row:
        opening_balance = row["opening_balance"] or 0
        opening_set = True
    else:
        # No record → fetch last month closing
        last_month_row = conn.execute("""
            SELECT closing_balance
            FROM canteen_monthly_summary
            ORDER BY id DESC
            LIMIT 1
        """).fetchone()

        if last_month_row:
            opening_balance = last_month_row["closing_balance"] or 0
        else:
            opening_balance = 0

        opening_set = False
    
    

    # -------------------------
    # EMPLOYEES
    # -------------------------
    employees = conn.execute(
        "SELECT * FROM canteen_employees ORDER BY emp_id"
    ).fetchall()

    # -------------------------
    # LUNCHES
    # -------------------------
    if selected_date:
        lunches = conn.execute(
            """
            SELECT * FROM canteen_lunch_entries
            WHERE date = %s
            ORDER BY id DESC
            """,
            (selected_date,)
        ).fetchall()
    else:
        lunches = conn.execute(
            """
            SELECT * FROM canteen_lunch_entries
            WHERE date BETWEEN %s AND %s
            ORDER BY id DESC
            """,
            (month_start, month_end)
        ).fetchall()

    # -------------------------
    # EXPENSES
    # -------------------------
    expenses_list = conn.execute(
        """
        SELECT * FROM canteen_expenses
        WHERE date BETWEEN %s AND %s
        ORDER BY id DESC
        """,
        (month_start, month_end)
    ).fetchall()

    # -------------------------
    # PAYMENTS
    # -------------------------
    payments_list = conn.execute(
        """
        SELECT * FROM canteen_payments
        WHERE date BETWEEN %s AND %s
        ORDER BY id DESC
        """,
        (month_start, month_end)
    ).fetchall()

    # -------------------------
    # CALCULATE TOTALS
    # -------------------------
    income = conn.execute(
        """
        SELECT SUM(amount)
        FROM canteen_payments
        WHERE date BETWEEN %s AND %s
        """,
        (month_start, month_end)
    ).fetchone()[0] or 0

    expenses = conn.execute(
        """
        SELECT SUM(amount)
        FROM canteen_expenses
        WHERE date BETWEEN %s AND %s
        """,
        (month_start, month_end)
    ).fetchone()[0] or 0

    # 🔥 Make expenses negative
    expenses = -abs(float(expenses))
    
    # -------------------------
    # CALCULATE CLOSING
    # -------------------------
    closing_balance = opening_balance + income + expenses

    # ---------------------------------
    # UPDATE MONTH SUMMARY (ONLY IF EXISTS)
    # ---------------------------------
    if opening_set:
        conn.execute("""
            UPDATE canteen_monthly_summary
            SET income=%s,
                expenses=%s,
                closing_balance=%s
            WHERE month=%s
        """, (income, expenses, closing_balance, selected_month))

    conn.commit()
    conn.close()
    
    # =========================
    # TOTAL OUTSTANDING
    # =========================
    selected_month = request.args.get(
        "month",
        datetime.now().strftime("%Y-%m")
    )

    total_outstanding = sampling_portal.get_total_outstanding(selected_month)

    if total_outstanding is None:
        total_outstanding = 0
    
   # =========================
    # FINANCIAL CALCULATIONS
    # =========================

    conn = sampling_portal.get_db_connection()

    # Total Payments (Income)
    income_row = conn.execute("""
        SELECT SUM(amount) FROM canteen_payments
        WHERE LEFT(date,7) = %s
    """, (selected_month,)).fetchone()

    income = income_row[0] if income_row[0] else 0

    # Total Expenses
    expense_row = conn.execute("""
        SELECT SUM(amount) FROM canteen_expenses
        WHERE LEFT(date,7) = %s
    """, (selected_month,)).fetchone()

    expenses = expense_row[0] if expense_row[0] else 0

    conn.close()

    

    # =========================
    # RENDER
    # =========================

    return render_template(
        "canteen.html",
        user_name=session.get("User_Name"),
        user_type=role,

        employees=employees,
        lunches=lunches,
        expenses_list=expenses_list,
        payments_list=payments_list,

        selected_month=selected_month,
        selected_date=selected_date or '',

        opening_balance=opening_balance,   # ✅ FROM DATABASE
        income=income,
        expenses=expenses,
        balance=closing_balance,

        dashboard_outstanding=total_outstanding
    )


# =====================================================
# SAVE OPENING BALANCE
# =====================================================

@app.route('/api/canteen/set-opening', methods=['POST'])
@login_required
def set_opening_balance():

    role = session.get("User_Type")

    if not can_do("ctn_pay"):
        return jsonify({"status": "error", "message": "Access Denied"}), 403

    data = request.json
    month = data.get("month")
    amount = float(data.get("amount") or 0)

    conn = sampling_portal.get_db_connection()

    try:
        conn.execute("""
            INSERT INTO canteen_opening_balance (month, opening_amount)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE opening_amount = VALUES(opening_amount)
        """, (month, amount))

        conn.commit()

        return jsonify({"status": "success"})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

    finally:
        conn.close()


@app.route('/api/canteen/employees')
@login_required
def api_canteen_employees():

    conn = sampling_portal.get_db_connection()
    rows = conn.execute("SELECT * FROM canteen_employees ORDER BY emp_id").fetchall()
    conn.close()

    return jsonify({
        "status": "success",
        "data": [dict(r) for r in rows]
    })

@app.route('/api/canteen/lunch/add', methods=['POST'])
@login_required
def api_add_lunch():

    data = request.json
    date = data.get("date")
    entries = data.get("entries", [])

    if not date:
        return jsonify({"status": "error", "message": "Date required"})

    conn = sampling_portal.get_db_connection()

    success_count = 0
    duplicate_count = 0

    for e in entries:

        # 🔒 Duplicate Protection
        existing = conn.execute("""
            SELECT 1 FROM canteen_lunch_entries
            WHERE date=%s AND emp_id=%s
        """, (
            date,
            e["emp_id"]
        )).fetchone()

        if existing:
            duplicate_count += 1
            continue

        amount = -abs(float(e["amount"]))

        conn.execute("""
            INSERT INTO canteen_lunch_entries
            (date, emp_id, emp_name, category, amount)
            VALUES (%s,%s,%s,%s,%s)
        """, (
            date,
            e["emp_id"],
            e["emp_name"],
            e["category"],
            amount
        ))

        success_count += 1

    conn.commit()
    conn.close()

    if duplicate_count > 0 and success_count > 0:
        return jsonify({
            "status": "partial",
            "message": f"{duplicate_count} duplicate entries skipped."
        })

    elif duplicate_count > 0:
        return jsonify({
            "status": "error",
            "message": "All selected entries already exist."
        })

    else:
        return jsonify({"status": "success"})
        
        
        
@app.route('/api/canteen/lunch/delete', methods=['POST'])
@login_required
def api_delete_lunch():

    ids = request.json.get("ids", [])

    conn = sampling_portal.get_db_connection()

    for i in ids:
        conn.execute("DELETE FROM canteen_lunch_entries WHERE id=%s", (i,))

    conn.commit()
    conn.close()

    return jsonify({"status": "success"})



@app.route('/api/canteen/lunch/list')
@login_required
def api_list_lunch():

    conn = sampling_portal.get_db_connection()
    rows = conn.execute(
        "SELECT * FROM canteen_lunch_entries ORDER BY id DESC"
    ).fetchall()
    conn.close()

    return jsonify({
        "status": "success",
        "data": [dict(r) for r in rows]
    })


@app.route('/api/canteen/payments/list')
@login_required
def api_list_payments():

    conn = sampling_portal.get_db_connection()
    rows = conn.execute(
        "SELECT * FROM canteen_payments ORDER BY id DESC"
    ).fetchall()
    conn.close()

    return jsonify({
        "status": "success",
        "data": [dict(r) for r in rows]
    })


@app.route('/api/canteen/export-employees')
@login_required
def export_canteen_employees():

    if session.get("User_Type") != "admin":
        return jsonify({"status": "error", "message": "Admin only access"})

    conn = sampling_portal.get_db_connection()

    rows = conn.execute(
        "SELECT * FROM canteen_employees ORDER BY emp_id"
    ).fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "canteen_employees"

    if rows:
        headers = rows[0].keys()
        ws.append(list(headers))

        for r in rows:
            ws.append([r[h] for h in headers])

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)

    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name="canteen_employees.xlsx"
    )
    

@app.route('/api/canteen/save-expense', methods=['POST'])
@login_required
def save_canteen_expense():

    if not can_do("ctn_pay"):
        return jsonify({"status": "error", "message": "Access Denied"})

    data = request.json

    record_id = data.get("id")  # 🔥 detect edit mode
    expense_date = data.get("date") or None  # NEW: custom date
    invoice_no = data.get("invoice_no")
    particulars = data.get("particulars")
    category = data.get("category")
    amount = data.get("amount")

    if not amount:
        return jsonify({"status": "error", "message": "Amount required"})

    conn = sampling_portal.get_db_connection()

    try:

        # ==============================
        # UPDATE MODE
        # ==============================
        if record_id:

            conn.execute("""
                UPDATE canteen_expenses
                SET invoice_no=%s,
                    particulars=%s,
                    category=%s,
                    amount=%s
                WHERE id=%s
            """, (
                invoice_no,
                particulars,
                category,
                float(amount),
                record_id
            ))

        # ==============================
        # INSERT MODE
        # ==============================
        else:

            conn.execute("""
                INSERT INTO canteen_expenses
                (date, invoice_no, particulars, category, amount)
                VALUES (COALESCE(%s, CURDATE()), %s, %s, %s, %s)
            """, (
                expense_date,
                invoice_no,
                particulars,
                category,
                float(amount)
            ))

        conn.commit()
        conn.close()

        return jsonify({"status": "success"})

    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)})   

@app.route('/api/canteen/save-employee', methods=['POST'])
@login_required
def save_employee():

    data = request.json
    conn = sampling_portal.get_db_connection()

    try:

        # ==========================
        # UPDATE MODE
        # ==========================
        if data.get("id"):

            conn.execute("""
                UPDATE canteen_employees
                SET emp_name=%s,
                    department=%s,
                    category=%s,
                    contact_number=%s,
                    opening_balance=%s
                WHERE id=%s
            """, (
                data.get("emp_name"),
                data.get("department"),
                data.get("category"),
                data.get("contact_number"),
                float(data.get("opening_balance") or 0),
                data.get("id")
            ))

        # ==========================
        # INSERT MODE
        # ==========================
        else:

            conn.execute("""
                INSERT INTO canteen_employees
                (emp_id, emp_name, department, category, contact_number, opening_balance)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                data.get("emp_id"),
                data.get("emp_name"),
                data.get("department"),
                data.get("category"),
                data.get("contact_number"),
                float(data.get("opening_balance") or 0)
            ))

        conn.commit()
        conn.close()

        return jsonify({"status": "success"})

    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)})
        


@app.route('/api/canteen/employee/next-id/<category>')
@login_required
def get_next_employee_id(category):

    conn = sampling_portal.get_db_connection()

    try:

        if category.lower() == "monthly":
            prefix = "M"
        elif category.lower() == "daily":
            prefix = "D"
        else:
            return jsonify({"status": "error", "message": "Invalid category"})

        # Get highest existing ID for that prefix
        result = conn.execute("""
            SELECT emp_id FROM canteen_employees
            WHERE emp_id LIKE %s
            ORDER BY emp_id DESC
            LIMIT 1
        """, (prefix + "%",)).fetchone()

        if result:
            last_id = result["emp_id"]  # Example: M007
            last_number = int(last_id[1:])  # remove prefix
            next_number = last_number + 1
        else:
            next_number = 1

        new_id = f"{prefix}{next_number:03d}"

        conn.close()

        return jsonify({
            "status": "success",
            "emp_id": new_id
        })

    except Exception as e:
        conn.close()
        return jsonify({
            "status": "error",
            "message": str(e)
        })




@app.route('/api/canteen/save-payment', methods=['POST'])
@login_required
def save_canteen_payment():

    if not can_do("ctn_pay"):
        return jsonify({"status": "error", "message": "Access Denied"})

    data = request.json
    conn = sampling_portal.get_db_connection()
    payment_date = data.get("date") or None  # NEW: custom date

    try:

        # ==========================
        # UPDATE MODE
        # ==========================
        if data.get("id"):

            conn.execute("""
                UPDATE canteen_payments
                SET vch_no=%s,
                    emp_name=%s,
                    amount=%s
                WHERE id=%s
            """, (
                data.get("vch_no"),
                data.get("emp_name"),
                float(data.get("amount") or 0),
                data.get("id")
            ))

        # ==========================
        # INSERT MODE
        # ==========================
        else:

            conn.execute("""
                INSERT INTO canteen_payments
                (date, vch_no, emp_name, amount)
                VALUES (COALESCE(%s, CURDATE()), %s, %s, %s)
            """, (
                payment_date,
                data.get("vch_no"),
                data.get("emp_name"),
                float(data.get("amount") or 0)
            ))

        conn.commit()
        conn.close()

        return jsonify({"status": "success"})

    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)})


@app.route('/api/canteen/import-employees', methods=['POST'])
@login_required
def import_canteen_employees():

    if session.get("User_Type") != "admin":
        return jsonify({"status": "error", "message": "Admin only access"})

    file = request.files.get("file")

    if not file:
        return jsonify({"status": "error", "message": "No file uploaded"})

    wb = load_workbook(file)

    if "canteen_employees" not in wb.sheetnames:
        return jsonify({"status": "error", "message": "Invalid Excel format"})

    ws = wb["canteen_employees"]
    rows = list(ws.values)

    if not rows:
        return jsonify({"status": "error", "message": "Empty sheet"})

    headers = rows[0]
    data_rows = rows[1:]

    conn = sampling_portal.get_db_connection()

    try:
        # Clear existing employee table
        conn.execute("DELETE FROM canteen_employees")

        for row in data_rows:
            placeholders = ",".join(["?"] * len(headers))
            conn.execute(
                f"INSERT INTO canteen_employees ({','.join(headers)}) VALUES ({placeholders})",
                row
            )

        conn.commit()
        conn.close()

        return jsonify({"status": "success", "message": "Employees imported successfully"})

    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)})



@app.route('/api/canteen/summary')
@login_required
def api_canteen_summary():

    conn = sampling_portal.get_db_connection()

    income = conn.execute(
        "SELECT SUM(amount) FROM canteen_payments"
    ).fetchone()[0] or 0

    expense = conn.execute(
        "SELECT SUM(amount) FROM canteen_lunch_entries"
    ).fetchone()[0] or 0

    conn.close()

    return jsonify({
        "status": "success",
        "income": income,
        "expense": expense,
        "balance": income - expense
    })




@app.route('/api/canteen/report/user-summary/<month>')
@login_required
def api_user_summary(month):

    conn = sampling_portal.get_db_connection()

    if month == "all":
        month_filter = ""
        params = ()
        payment_filter = ""
        payment_params = ()
    else:
        month_start = month + "-01"
        month_end = month + "-31"

        month_filter = "AND l.date BETWEEN %s AND ?"
        payment_filter = "AND p.date BETWEEN %s AND ?"

        params = (month_start, month_end)
        payment_params = (month_start, month_end)

    query = f"""
        SELECT 
            e.emp_id,
            e.emp_name,
            e.category,
            IFNULL(SUM(l.amount),0) as total_lunch,
            IFNULL((
                SELECT SUM(p.amount)
                FROM canteen_payments p
                WHERE p.emp_name = e.emp_name
                {payment_filter}
            ),0) as total_payment
        FROM canteen_employees e
        LEFT JOIN canteen_lunch_entries l
            ON l.emp_id = e.emp_id
            {month_filter}
        GROUP BY e.emp_id, e.emp_name, e.category
        ORDER BY e.emp_name
    """

    if month == "all":
        rows = conn.execute(query).fetchall()
    else:
        rows = conn.execute(query, params + payment_params).fetchall()

    data = []

    for r in rows:
        pending = (r["total_lunch"] or 0) + (r["total_payment"] or 0)

        data.append({
            "emp_id": r["emp_id"],
            "emp_name": r["emp_name"],
            "category": r["category"],
            "total_lunch": r["total_lunch"] or 0,
            "total_payment": r["total_payment"] or 0,
            "pending": pending
        })

    conn.close()

    return jsonify({
        "status": "success",
        "data": data
    })




import os
import time
import base64
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


@app.route('/api/canteen/send-lunch-report', methods=['POST'])
@login_required
def send_lunch_report():

    data = request.get_json()
    image_base64 = data.get("image")

    if not image_base64:
        return jsonify({"status": "error", "message": "No image received"})

    # Save file
    file_path = os.path.join(os.getcwd(), "canteen_lunch_report.png")

    image_data = base64.b64decode(image_base64.split(",")[1])

    with open(file_path, "wb") as f:
        f.write(image_data)

    try:
        # Launch Chrome
        options = webdriver.ChromeOptions()
        options.add_argument("--user-data-dir=./whatsapp-session")  # keep login session

        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=options
        )

        driver.get("https://web.whatsapp.com")

        wait = WebDriverWait(driver, 60)

        # Wait for search box
        search_box = wait.until(
            EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="3"]'))
        )

        time.sleep(3)

        # Search group
        search_box.click()
        search_box.send_keys("Canteen Group")
        time.sleep(2)
        search_box.send_keys(Keys.ENTER)

        time.sleep(3)

        # Click attachment button
        attach_btn = driver.find_element(By.XPATH, '//span[@data-icon="clip"]')
        attach_btn.click()

        time.sleep(2)

        # Upload file
        file_input = driver.find_element(By.XPATH, '//input[@type="file"]')
        file_input.send_keys(file_path)

        time.sleep(3)

        # Add caption
        caption_box = driver.find_element(By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]')
        caption_box.send_keys("CLEAR THE PAYMENTS ON AN URGENT BASIS")

        time.sleep(2)

        caption_box.send_keys(Keys.ENTER)

        time.sleep(5)

        driver.quit()

        return jsonify({"status": "success"})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})



@app.route('/api/canteen/pending/<emp_name>')
@login_required
def get_employee_pending(emp_name):

    conn = sampling_portal.get_db_connection()

    total_lunch = conn.execute("""
        SELECT SUM(amount)
        FROM canteen_lunch_entries
        WHERE emp_name=%s
    """, (emp_name,)).fetchone()[0] or 0

    total_payment = conn.execute("""
        SELECT SUM(amount)
        FROM canteen_payments
        WHERE emp_name=%s
    """, (emp_name,)).fetchone()[0] or 0

    conn.close()

    return jsonify({
        "status": "success",
        "pending": total_lunch + total_payment
    })
    
@app.route('/api/canteen/payment/<int:record_id>')
@login_required
def get_payment_by_id(record_id):

    conn = sampling_portal.get_db_connection()

    row = conn.execute("""
        SELECT * FROM canteen_payments
        WHERE id=%s
    """, (record_id,)).fetchone()

    conn.close()

    if not row:
        return jsonify({"status": "error", "message": "Not found"})

    return jsonify({
        "status": "success",
        "data": dict(row)
    })


@app.route('/api/canteen/delete', methods=['POST'])
@login_required
def delete_canteen_record():

    data = request.json
    table = data.get("table")
    record_id = data.get("id")

    if table not in [
        "canteen_employees",
        "canteen_payments",
        "canteen_expenses"
    ]:
        return jsonify({"status": "error", "message": "Invalid table"})

    conn = sampling_portal.get_db_connection()

    conn.execute(f"DELETE FROM {table} WHERE id=%s", (record_id,))
    conn.commit()
    conn.close()

    return jsonify({"status": "success"})
    

@app.route('/api/canteen/get-opening/<month>')
@login_required
def get_opening_balance(month):

    conn = sampling_portal.get_db_connection()

    row = conn.execute("""
        SELECT opening_balance
        FROM canteen_monthly_summary
        WHERE month=%s
    """, (month,)).fetchone()

    conn.close()

    if row:
        return jsonify({
            "status": "exists",
            "opening_balance": row["opening_balance"]
        })
    else:
        return jsonify({
            "status": "new",
            "opening_balance": 0
        })


@app.route('/api/canteen/opening-balance', methods=['POST'])
@login_required
def save_opening_balance():

    role = session.get("User_Type")

    if not can_do("ctn_pay"):
        return jsonify({"status": "error", "message": "Access Denied"}), 403

    data = request.json
    month = data.get("month")
    amount = float(data.get("amount") or 0)

    conn = sampling_portal.get_db_connection()

    existing = conn.execute("""
        SELECT id FROM canteen_monthly_summary
        WHERE month=%s
    """, (month,)).fetchone()

    if existing:
        conn.execute("""
            UPDATE canteen_monthly_summary
            SET opening_balance=%s
            WHERE month=%s
        """, (amount, month))
        action = "updated"
    else:
        conn.execute("""
            INSERT INTO canteen_monthly_summary (month, opening_balance)
            VALUES (%s, %s)
        """, (month, amount))
        action = "created"

    conn.commit()
    conn.close()

    return jsonify({
        "status": "success",
        "action": action
    })
    

@app.route('/api/canteen/report/employee-ledger')
@login_required
def employee_ledger_report():

    emp_name = request.args.get("emp_name")
    month = request.args.get("month")

    if not emp_name or not month:
        return jsonify({"status": "error", "message": "Missing data"})

    month_start = month + "-01"
    month_end = month + "-31"

    conn = sampling_portal.get_db_connection()

    # -------------------------
    # GET EMPLOYEE
    # -------------------------
    employee = conn.execute("""
        SELECT * FROM canteen_employees
        WHERE emp_name=%s
    """, (emp_name,)).fetchone()

    if not employee:
        conn.close()
        return jsonify({"status": "error", "message": "Employee not found"})

    opening_balance = employee["opening_balance"] or 0

    # -------------------------
    # FETCH MONTH DATA
    # -------------------------
    lunches = conn.execute("""
        SELECT date, lunch_taken, amount
        FROM canteen_lunch_entries
        WHERE emp_name=%s AND date BETWEEN %s AND %s
    """, (emp_name, month_start, month_end)).fetchall()

    payments = conn.execute("""
        SELECT date, amount
        FROM canteen_payments
        WHERE emp_name=%s AND date BETWEEN %s AND %s
    """, (emp_name, month_start, month_end)).fetchall()

    total_lunch = sum([r["amount"] for r in lunches])
    total_payment = sum([r["amount"] for r in payments])

    # -------------------------
    # CALCULATE BALANCE
    # -------------------------
    current_balance = opening_balance + total_lunch + total_payment

    carry_forward = 0

    if employee["category"] == "Monthly" and current_balance < 0:
        carry_forward = current_balance

    conn.close()

    return jsonify({
        "status": "success",
        "employee": dict(employee),
        "lunches": [dict(r) for r in lunches],
        "payments": [dict(r) for r in payments],
        "opening_balance": opening_balance,
        "pending": current_balance,
        "carry_forward": carry_forward
    })



# ============================================================
# MONTH CLOSE & CARRY FORWARD SYSTEM
# ============================================================

@app.route('/api/canteen/close-month', methods=['POST'])
@login_required
def close_month():

    role = session.get("User_Type")

    if not can_do("ctn_pay"):
        return jsonify({"status": "error", "message": "Access Denied"})

    data = request.json
    month = data.get("month")

    if not month:
        return jsonify({"status": "error", "message": "Month required"})

    month_start = month + "-01"
    month_end = month + "-31"

    conn = sampling_portal.get_db_connection()

    employees = conn.execute("""
        SELECT * FROM canteen_employees
    """).fetchall()

    for emp in employees:

        emp_name = emp["emp_name"]
        opening_balance = emp["opening_balance"] or 0

        lunches = conn.execute("""
            SELECT SUM(amount)
            FROM canteen_lunch_entries
            WHERE emp_name=%s AND date BETWEEN %s AND %s
        """, (emp_name, month_start, month_end)).fetchone()[0] or 0

        payments = conn.execute("""
            SELECT SUM(amount)
            FROM canteen_payments
            WHERE emp_name=%s AND date BETWEEN %s AND %s
        """, (emp_name, month_start, month_end)).fetchone()[0] or 0

        closing_balance = opening_balance + lunches + payments

        # Carry forward rule
        if emp["category"] == "Monthly" and closing_balance < 0:
            new_opening = closing_balance
        else:
            new_opening = 0

        # Update employee opening balance
        conn.execute("""
            UPDATE canteen_employees
            SET opening_balance=%s
            WHERE emp_id=%s
        """, (new_opening, emp["emp_id"]))

    conn.commit()
    conn.close()

    return jsonify({"status": "success", "message": "Month closed successfully"})


@app.route("/api/canteen/lunch/update-single", methods=["POST"])
@login_required
def update_single_lunch():

    data = request.json
    record_id = data.get("id")
    date = data.get("date")
    amount = data.get("amount")

    if not record_id:
        return jsonify({"status": "error", "message": "Missing ID"})

    conn = sampling_portal.get_db_connection()

    try:
        conn.execute("""
            UPDATE canteen_lunch
            SET date = %s, amount = %s
            WHERE id = %s
        """, (date, amount, record_id))

        conn.commit()
        conn.close()

        return jsonify({"status": "success"})

    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)})


@app.route("/api/canteen/payment/next-voucher")
@login_required
def get_next_payment_voucher():

    from datetime import datetime

    today = datetime.now()

    # DB format
    today_db = today.strftime("%Y-%m-%d")

    # Display format → 02Mar26
    today_display = today.strftime("%d%b%y")

    conn = sampling_portal.get_db_connection()

    # Count today's payments
    count = conn.execute("""
        SELECT COUNT(*) 
        FROM canteen_payments
        WHERE date = %s
    """, (today_db,)).fetchone()[0]

    conn.close()

    next_number = count + 1

    voucher = f"{today_display}/{str(next_number).zfill(4)}"

    return jsonify({
        "status": "success",
        "voucher": voucher
    })



@app.route('/api/canteen/lunch/by-date/<date>')
@login_required
def api_lunch_by_date(date):
    conn = sampling_portal.get_db_connection()
    rows = conn.execute("""
        SELECT * FROM canteen_lunch_entries WHERE date=%s ORDER BY id DESC
    """, (date,)).fetchall()
    conn.close()
    return jsonify({"status": "success", "data": [dict(r) for r in rows]})


@app.route('/api/canteen/monthly-summaries')
@login_required
def api_monthly_summaries():
    conn = sampling_portal.get_db_connection()
    rows = conn.execute("""
        SELECT * FROM canteen_monthly_summary ORDER BY id DESC
    """).fetchall()
    conn.close()
    return jsonify({"status": "success", "data": [dict(r) for r in rows]})



# ─────────────────────────────────────────────────────────────────
# HOLIDAY LIST ROUTES
# ─────────────────────────────────────────────────────────────────

def _ensure_holiday_table():
    conn = sampling_portal.get_db_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS canteen_holidays (
            id   INT AUTO_INCREMENT PRIMARY KEY,
            date DATE NOT NULL,
            name VARCHAR(200) NOT NULL,
            type VARCHAR(50) NOT NULL DEFAULT 'National'
        )
    """)
    conn.commit()
    conn.close()

_ensure_holiday_table()


@app.route('/api/canteen/holidays')
@login_required
def api_get_holidays():
    month = request.args.get('month', '')
    conn  = sampling_portal.get_db_connection()
    if month:
        rows = conn.execute(
            "SELECT * FROM canteen_holidays WHERE date LIKE %s ORDER BY date",
            (month + '%',)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM canteen_holidays ORDER BY date").fetchall()
    conn.close()
    return jsonify({"status": "success", "holidays": [dict(r) for r in rows]})


@app.route('/api/canteen/save-holiday', methods=['POST'])
@login_required
def api_save_holiday():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({"status": "error", "message": "Admin only"})
    data = request.json
    conn = sampling_portal.get_db_connection()
    try:
        if data.get('id'):
            conn.execute(
                "UPDATE canteen_holidays SET date=%s, name=%s, type=%s WHERE id=%s",
                (data['date'], data['name'], data.get('type','National'), data['id'])
            )
        else:
            conn.execute(
                "INSERT INTO canteen_holidays (date, name, type) VALUES (%s,%s,%s)",
                (data['date'], data['name'], data.get('type','National'))
            )
        conn.commit()
        conn.close()
        return jsonify({"status": "success"})
    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)})


@app.route('/api/canteen/delete-holiday', methods=['POST'])
@login_required
def api_delete_holiday():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({"status": "error", "message": "Admin only"})
    hid = request.json.get('id')
    conn = sampling_portal.get_db_connection()
    conn.execute("DELETE FROM canteen_holidays WHERE id=%s", (hid,))
    conn.commit()
    conn.close()
    return jsonify({"status": "success"})


# ─────────────────────────────────────────────────────────────────
# MONTHLY ADVANCE PAYMENT TRACKING ROUTES
# ─────────────────────────────────────────────────────────────────

def _ensure_advance_table():
    conn = sampling_portal.get_db_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS canteen_advance_payments (
            id       INT AUTO_INCREMENT PRIMARY KEY,
            month    VARCHAR(20) NOT NULL,
            emp_id   VARCHAR(20) NOT NULL,
            emp_name VARCHAR(200),
            paid     TINYINT(1) DEFAULT 0,
            paid_on  DATE,
            UNIQUE KEY uq_month_emp (month, emp_id)
        )
    """)
    conn.commit()
    conn.close()

_ensure_advance_table()


@app.route('/api/canteen/advance-payments')
@login_required
def api_get_advance_payments():
    month = request.args.get('month', '')
    conn  = sampling_portal.get_db_connection()

    # Get all Monthly-category employees
    employees = conn.execute(
        "SELECT emp_id, emp_name, department, contact_number FROM canteen_employees WHERE category='Monthly' ORDER BY emp_name"
    ).fetchall()

    # Ensure a record exists for each employee this month
    for emp in employees:
        conn.execute("""
            INSERT IGNORE INTO canteen_advance_payments (month, emp_id, emp_name)
            VALUES (%s, %s, %s)
        """, (month, emp['emp_id'], emp['emp_name']))
    conn.commit()

    # Fetch merged data
    rows = conn.execute("""
        SELECT e.emp_id, e.emp_name, e.department, e.contact_number,
               COALESCE(a.paid,0) as paid, a.paid_on
        FROM canteen_employees e
        LEFT JOIN canteen_advance_payments a
            ON a.emp_id = e.emp_id AND a.month = %s
        WHERE e.category = 'Monthly'
        ORDER BY e.emp_name
    """, (month,)).fetchall()
    conn.close()
    return jsonify({"status": "success", "records": [dict(r) for r in rows]})


@app.route('/api/canteen/mark-advance-paid', methods=['POST'])
@login_required
def api_mark_advance_paid():
    data    = request.json
    month   = data.get('month')
    emp_id  = data.get('emp_id')
    emp_name= data.get('emp_name','')
    paid    = 1 if data.get('paid') else 0
    paid_on = __import__('datetime').date.today().isoformat() if paid else None

    conn = sampling_portal.get_db_connection()
    try:
        conn.execute("""
            INSERT INTO canteen_advance_payments (month, emp_id, emp_name, paid, paid_on)
            VALUES (%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE paid=VALUES(paid), paid_on=VALUES(paid_on)
        """, (month, emp_id, emp_name, paid, paid_on))
        conn.commit()
        conn.close()
        return jsonify({"status": "success"})
    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)})


# ═══════════════════════════════════════════════════════════════
# TASK REMINDERS — admin & Purchase only
# ═══════════════════════════════════════════════════════════════

def _ensure_task_reminder_table():
    conn = sampling_portal.get_db_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS task_reminders (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            title       TEXT NOT NULL,
            description TEXT,
            due_date    DATE,
            priority    VARCHAR(20) DEFAULT 'Medium',
            status      VARCHAR(30) DEFAULT 'Pending',
            assigned_to VARCHAR(100),
            created_by  VARCHAR(100),
            created_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at  DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

_ensure_task_reminder_table()


@app.route('/task_reminders')
@login_required
def task_reminders_page():
    if not can_access('task_reminders'):
        return _denied('Task Reminders')

    role = session.get('User_Type')
    uid  = session.get('UID', '')
    conn = sampling_portal.get_db_connection()

    # Admin and Purchase see ALL tasks
    if role in ('admin', 'Purchase'):
        tasks = conn.execute("""
            SELECT * FROM task_reminders ORDER BY
                CASE priority WHEN 'Critical' THEN 1 WHEN 'High' THEN 2 WHEN 'Medium' THEN 3 ELSE 4 END,
                due_date ASC, id DESC
        """).fetchall()
    else:
        # Other roles see only:
        # 1. Tasks assigned specifically to them (by UID)
        # 2. Tasks assigned to their role/department
        # 3. Tasks with no assignment (blank/null = visible to all)
        tasks = conn.execute("""
            SELECT * FROM task_reminders
            WHERE (
                assigned_to = %s
                OR assigned_to = %s
                OR assigned_to = 'All'
                OR assigned_to = ''
                OR assigned_to IS NULL
            )
            ORDER BY
                CASE priority WHEN 'Critical' THEN 1 WHEN 'High' THEN 2 WHEN 'Medium' THEN 3 ELSE 4 END,
                due_date ASC, id DESC
        """, (uid, role)).fetchall()
    conn.close()

    tasks = [dict(t) for t in tasks]
    return render_template('task_reminders.html',
        tasks=tasks,
        role=role,
        user=uid
    )


@app.route('/api/task_reminders/save', methods=['POST'])
@login_required
def api_save_task():
    role = session.get('User_Type')
    if not can_access('task_reminders'):
        return jsonify({'status': 'error', 'message': 'Access Denied'})

    data    = request.json
    task_id = data.get('id')
    title   = data.get('title', '').strip()
    if not title:
        return jsonify({'status': 'error', 'message': 'Title is required'})

    conn = sampling_portal.get_db_connection()
    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        if task_id:
            conn.execute("""
                UPDATE task_reminders
                SET title=%s, description=%s, due_date=%s, priority=%s,
                    status=%s, assigned_to=%s, updated_at=%s
                WHERE id=%s
            """, (
                title,
                data.get('description', ''),
                data.get('due_date') or None,
                data.get('priority', 'Medium'),
                data.get('status', 'Pending'),
                data.get('assigned_to', ''),
                now, task_id
            ))
        else:
            conn.execute("""
                INSERT INTO task_reminders
                    (title, description, due_date, priority, status, assigned_to, created_by, created_at, updated_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                title,
                data.get('description', ''),
                data.get('due_date') or None,
                data.get('priority', 'Medium'),
                data.get('status', 'Pending'),
                data.get('assigned_to', ''),
                session.get('UID'),
                now, now
            ))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/api/task_reminders/delete', methods=['POST'])
@login_required
def api_delete_task():
    role = session.get('User_Type')
    if not can_access('task_reminders'):
        return jsonify({'status': 'error', 'message': 'Access Denied'})

    task_id = request.json.get('id')
    if not task_id:
        return jsonify({'status': 'error', 'message': 'Missing ID'})

    conn = sampling_portal.get_db_connection()
    conn.execute("DELETE FROM task_reminders WHERE id=%s", (task_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})


@app.route('/api/task_reminders/send_reminder', methods=['POST'])
@login_required
def api_send_reminder():
    """Admin-only: instantly push reminder to assigned user via SSE queue."""
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'})

    task_id = request.json.get('task_id')
    if not task_id:
        return jsonify({'status': 'error', 'message': 'Missing task_id'})

    conn = sampling_portal.get_db_connection()
    task = conn.execute("SELECT * FROM task_reminders WHERE id=%s", (task_id,)).fetchone()

    if not task:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Task not found'})

    task = dict(task)
    assigned_to = task.get('assigned_to', '')

    if not assigned_to:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Task has no assigned user'})

    if task['status'] == 'Done':
        conn.close()
        return jsonify({'status': 'error', 'message': 'Task is already done'})

    # Store in DB (for users not currently connected via SSE)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS task_push_reminders (
            id         INT AUTO_INCREMENT PRIMARY KEY,
            task_id    INT NOT NULL,
            target_uid VARCHAR(100) NOT NULL,
            sent_by    VARCHAR(100),
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            delivered  TINYINT(1) DEFAULT 0
        )
    """)
    conn.execute("""
        INSERT INTO task_push_reminders (task_id, target_uid, sent_by)
        VALUES (%s, %s, %s)
    """, (task_id, assigned_to, session.get('UID')))
    conn.commit()
    conn.close()

    # Push instantly to SSE queue if user is connected
    import json as _json
    payload = _json.dumps({
        'task_id':     task['id'],
        'title':       task['title'],
        'description': task['description'] or '',
        'priority':    task['priority'],
        'due_date':    task['due_date'] or '',
        'status':      task['status'],
        'assigned_to': task['assigned_to'],
        'sent_by':     session.get('UID'),
        'is_overdue':  bool(task['due_date'] and task['due_date'] < datetime.now().strftime('%Y-%m-%d'))
    })

    # Push to all SSE listeners for this user
    queues = _sse_clients.get(assigned_to, [])
    for q in list(queues):
        try:
            q.put_nowait(payload)
        except Exception:
            pass

    return jsonify({
        'status':      'success',
        'assigned_to': assigned_to,
        'task_title':  task['title'],
        'sse_clients': len(queues)
    })


# ── SSE client registry (in-memory, per-user queues) ──────────────
import queue as _queue_module
_sse_clients = {}   # { uid: [Queue, Queue, ...] }


@app.route('/api/task_reminders/stream')
@login_required
def api_task_reminder_stream():
    """
    SSE endpoint — each logged-in browser tab connects here.
    Keeps connection open; server pushes reminder events instantly.
    Also delivers any unread DB reminders on connect.
    """
    uid  = session.get('UID')
    if not uid or not can_access('task_reminders'):
        return '', 403

    q = _queue_module.Queue()

    # Register this client
    if uid not in _sse_clients:
        _sse_clients[uid] = []
    _sse_clients[uid].append(q)

    # Deliver any pending DB reminders immediately on connect
    def get_pending():
        try:
            conn = sampling_portal.get_db_connection()
            conn.execute("""
                CREATE TABLE IF NOT EXISTS task_push_reminders (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    task_id INT NOT NULL,
                    target_uid VARCHAR(100) NOT NULL,
                    sent_by VARCHAR(100),
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    delivered TINYINT(1) DEFAULT 0
                )
            """)
            rows = conn.execute("""
                SELECT p.id, p.task_id, p.sent_by,
                       t.title, t.description, t.priority,
                       t.due_date, t.status, t.assigned_to
                FROM task_push_reminders p
                JOIN task_reminders t ON t.id = p.task_id
                WHERE p.target_uid=%s AND p.delivered=0 AND t.status!='Done'
                ORDER BY p.created_at DESC
            """, (uid,)).fetchall()
            if rows:
                ids = [r['id'] for r in rows]
                conn.execute(
                    "UPDATE task_push_reminders SET delivered=1 WHERE id IN ({})".format(
                        ','.join(['?']*len(ids))), ids)
                conn.commit()
            conn.close()
            return rows
        except Exception:
            return []

    import json as _json

    def event_stream():
        # Send pending DB messages first
        for r in get_pending():
            t = dict(r)
            t['is_overdue'] = bool(t['due_date'] and t['due_date'] < datetime.now().strftime('%Y-%m-%d'))
            yield f"data: {_json.dumps(t)}\n\n"

        # Send heartbeat every 25s to keep connection alive through proxies
        import time
        while True:
            try:
                # Block up to 25 seconds waiting for a new message
                msg = q.get(timeout=25)
                yield f"data: {msg}\n\n"
            except _queue_module.Empty:
                # Heartbeat — keeps connection alive
                yield ": heartbeat\n\n"
            except GeneratorExit:
                break

    def cleanup():
        try:
            _sse_clients[uid].remove(q)
            if not _sse_clients[uid]:
                del _sse_clients[uid]
        except (KeyError, ValueError):
            pass

    from flask import stream_with_context, Response

    def guarded_stream():
        try:
            yield from event_stream()
        finally:
            cleanup()

    return Response(
        stream_with_context(guarded_stream()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control':     'no-cache',
            'X-Accel-Buffering': 'no',
        }
    )


@app.route('/api/task_reminders/check_push')
@login_required
def api_check_push_reminders():
    """Fallback poll for browsers that don't support SSE (rare)."""
    uid = session.get('UID')
    if not uid:
        return jsonify({'reminders': []})

    conn = sampling_portal.get_db_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS task_push_reminders (
            id INTEGER PRIMARY KEY AUTO_INCREMENT,
            task_id INTEGER NOT NULL, target_uid TEXT NOT NULL,
            sent_by TEXT, created_at TEXT DEFAULT (NOW()),
            delivered INTEGER DEFAULT 0
        )
    """)
    rows = conn.execute("""
        SELECT p.id, p.task_id, p.sent_by,
               t.title, t.description, t.priority, t.due_date,
               t.status, t.assigned_to
        FROM task_push_reminders p
        JOIN task_reminders t ON t.id = p.task_id
        WHERE p.target_uid=%s AND p.delivered=0 AND t.status!='Done'
        ORDER BY p.created_at DESC
    """, (uid,)).fetchall()
    if rows:
        ids = [r['id'] for r in rows]
        conn.execute(
            "UPDATE task_push_reminders SET delivered=1 WHERE id IN ({})".format(
                ','.join(['?']*len(ids))), ids)
        conn.commit()
    conn.close()

    today = datetime.now().strftime('%Y-%m-%d')
    reminders = [{**dict(r), 'is_overdue': bool(r['due_date'] and r['due_date'] < today)} for r in rows]
    return jsonify({'reminders': reminders})




@app.route('/api/task_reminders/toggle_status', methods=['POST'])
@login_required
def api_toggle_task_status():
    role = session.get('User_Type')
    if not can_access('task_reminders'):
        return jsonify({'status': 'error', 'message': 'Access Denied'})

    data       = request.json
    task_id    = data.get('id')
    new_status = data.get('status')
    if not task_id or not new_status:
        return jsonify({'status': 'error', 'message': 'Missing data'})

    conn = sampling_portal.get_db_connection()
    conn.execute("""
        UPDATE task_reminders SET status=%s, updated_at=NOW()
        WHERE id=%s
    """, (new_status, task_id))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})


@app.route('/api/task_reminders/overdue_count')
@login_required
def api_task_overdue_count():
    role = session.get('User_Type')
    if not can_access('task_reminders'):
        return jsonify({'count': 0})

    conn = sampling_portal.get_db_connection()
    count = conn.execute("""
        SELECT COUNT(*) FROM task_reminders
        WHERE status != 'Done'
          AND due_date IS NOT NULL
          AND due_date < CURDATE()
    """).fetchone()[0]
    conn.close()
    return jsonify({'count': count})


@app.route('/api/task_reminders/assignable_users')
@login_required
def api_task_assignable_users():
    """Return only admin & Purchase users for task assignment dropdown."""
    role = session.get('User_Type')
    if not can_access('task_reminders'):
        return jsonify({'status': 'error', 'message': 'Access Denied'})

    conn = sampling_portal.get_db_connection()
    rows = conn.execute("""
        SELECT username FROM `User_Tbl`
        WHERE user_type IN ('admin', 'Purchase')
        AND is_active = 1
        ORDER BY username
    """).fetchall()
    conn.close()
    return jsonify({'status': 'success', 'users': [r['username'] for r in rows]})


@app.route('/api/task_reminders/my_pending')
@login_required
def api_my_pending_tasks():
    """
    Polled every 2 hours by each logged-in user's browser.
    Returns pending/in-progress tasks assigned to that user.
    """
    role = session.get('User_Type')
    if not can_access('task_reminders'):
        return jsonify({'tasks': []})

    uid = session.get('UID')
    role = session.get('User_Type', '')

    conn = sampling_portal.get_db_connection()
    rows = conn.execute("""
        SELECT id, title, description, due_date, priority, status, assigned_to
        FROM task_reminders
        WHERE status != 'Done'
          AND (
            assigned_to = %s
            OR assigned_to = %s
            OR assigned_to = 'All'
            OR assigned_to = ''
            OR assigned_to IS NULL
          )
        ORDER BY
            CASE priority WHEN 'Critical' THEN 1 WHEN 'High' THEN 2 WHEN 'Medium' THEN 3 ELSE 4 END,
            due_date ASC
        LIMIT 20
    """, (uid, role)).fetchall()
    conn.close()

    today = datetime.now().strftime('%Y-%m-%d')
    tasks = []
    for r in rows:
        t = dict(r)
        t['is_overdue'] = bool(t['due_date'] and t['due_date'] < today)
        tasks.append(t)

    return jsonify({'tasks': tasks})


@app.route('/api/task_reminders/performance')
@login_required
def api_task_performance():
    """Performance chart data — admin & Purchase only."""
    role = session.get('User_Type')
    if not can_access('task_reminders'):
        return jsonify({'status': 'error', 'message': 'Access Denied'})

    conn = sampling_portal.get_db_connection()

    # ── Per-user summary ──────────────────────────────────────────
    user_rows = conn.execute("""
        SELECT
            COALESCE(NULLIF(assigned_to,''), 'Unassigned') AS user,
            COUNT(*)                                        AS total,
            SUM(CASE WHEN status='Done'        THEN 1 ELSE 0 END) AS done,
            SUM(CASE WHEN status='In-Progress' THEN 1 ELSE 0 END) AS in_progress,
            SUM(CASE WHEN status='Pending'     THEN 1 ELSE 0 END) AS pending,
            SUM(CASE WHEN status!='Done'
                      AND due_date IS NOT NULL
                      AND due_date < CURDATE() THEN 1 ELSE 0 END) AS overdue
        FROM task_reminders
        GROUP BY COALESCE(NULLIF(assigned_to,''), 'Unassigned')
        ORDER BY total DESC
    """).fetchall()

    # ── Priority breakdown ────────────────────────────────────────
    priority_rows = conn.execute("""
        SELECT priority, COUNT(*) AS cnt
        FROM task_reminders
        GROUP BY priority
    """).fetchall()

    # ── Status breakdown ─────────────────────────────────────────
    status_rows = conn.execute("""
        SELECT status, COUNT(*) AS cnt
        FROM task_reminders
        GROUP BY status
    """).fetchall()

    # ── Daily completions last 14 days ────────────────────────────
    daily_rows = conn.execute("""
        SELECT DATE(updated_at) AS day, COUNT(*) AS cnt
        FROM task_reminders
        WHERE status = 'Done'
          AND updated_at >= date('now', '-14 days')
        GROUP BY DATE(updated_at)
        ORDER BY day
    """).fetchall()

    conn.close()

    return jsonify({
        'status':    'success',
        'users':     [dict(r) for r in user_rows],
        'priority':  [dict(r) for r in priority_rows],
        'status_breakdown': [dict(r) for r in status_rows],
        'daily':     [dict(r) for r in daily_rows]
    })



# ═══════════════════════════════════════════════════════════════
# ACCESS CONTROL PANEL  ─  Super Admin only
# ═══════════════════════════════════════════════════════════════

@app.route('/access_control')
@login_required
def access_control_page():

    if not can_access('access_control'):
        return _denied('Access Control')

    if (session.get('User_Type') or '').lower() != 'admin':
        return _denied('Access Control')

    users = sampling_portal.get_all_users()

    return render_template(
        "access_control.html",
        users=users,
        role=session.get("User_Type")
    )


@app.route('/api/access_control/get/<int:user_id>')
@login_required
def api_get_user_permissions(user_id):
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'error': 'Access Denied'}), 403
    perms = sampling_portal.get_user_permissions(user_id)
    return jsonify({'status': 'success', 'permissions': perms})


@app.route('/api/access_control/save', methods=['POST'])
@login_required
def api_save_user_permissions():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Access Denied'}), 403
    data = request.get_json()
    user_id = data.get('user_id')
    permissions = data.get('permissions', {})
    if not user_id:
        return jsonify({'status': 'error', 'message': 'Missing user_id'})
    success, msg = sampling_portal.save_user_permissions(
        user_id, permissions, updated_by=session.get('UID', 'admin')
    )
    return jsonify({'status': 'success' if success else 'error', 'message': msg})


# ═══════════════════════════════════════════════════════════════════════════════
# INDEX-PAGE ACCESS CONTROL MODAL  ─  Admin only
# ─────────────────────────────────────────────────────────────────────────────
# Modal opened from the "Access Control" KPI card on /index.
# Mirrors the RM Store / Production Initiater pattern: one endpoint returns
# all users with their current permissions for the 18 KPI cards inline.
# Save uses the existing /api/access_control/save endpoint.
# ═══════════════════════════════════════════════════════════════════════════════

# (key, label) for each of the 18 KPI cards shown on the index page.
# Keys match the strings already passed to can_access() throughout app.py
# so existing routes pick up the toggles automatically.
INDEX_PERM_KEYS = [
    # Operations
    ('crm',                     'CRM · Leads'),
    ('production_initiater',    'Production Initiater'),
    ('production_dept',         'Production Department'),
    ('planning',                'Planning Dashboard'),
    ('qc_dashboard',            'QC Dashboard'),
    ('qc_sampling',             'QC Sampling'),
    ('rd_sampling',             'R&D Sampling'),
    ('procurement',             'Procurement Dashboard'),
    ('inventory_mgmt',          'Inventory Management'),
    ('packing',                 'Packing Entry'),
    ('invoice_checklist',       'Invoice Document Checklist'),
    ('pm_stock',                'PM Stock'),
    ('hcp_stock',               'Stock Register'),
    # Management & Admin
    ('task_reminders',          'Task Reminders'),
    ('task_scheduler',          'Task Scheduler'),
    ('manage_users',            'Manage Users'),
    ('access_control',          'Access Control'),
    ('backup_manager',          'Backup Manager'),
]
INDEX_PERM_KEY_SET = {k for k, _ in INDEX_PERM_KEYS}


@app.route('/api/index_access_control/users', methods=['GET'])
@login_required
def api_index_access_control_users():
    """
    Return all users with their index-page perms inline.
    For admin users, all 19 keys are reported as True (UI shows 'Full Access'
    banner instead of toggles).
    """
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403

    try:
        users = sampling_portal.get_all_users() or []
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

    out = []
    for u in users:
        u = dict(u)
        # Serialise datetimes
        for k in ('created_at', 'last_login'):
            v = u.get(k)
            if hasattr(v, 'isoformat'):
                u[k] = v.isoformat()
        uid = u.get('id')
        is_admin_user = (u.get('user_type') or '').lower() == 'admin'
        if is_admin_user:
            page_perms = {k: True for k in INDEX_PERM_KEY_SET}
        else:
            try:
                all_perms = sampling_portal.get_user_permissions(uid) or {}
            except Exception:
                all_perms = {}
            page_perms = {k: bool(all_perms.get(k, False)) for k in INDEX_PERM_KEY_SET}
        u['is_admin']   = is_admin_user
        u['page_perms'] = page_perms
        out.append(u)

    return jsonify({
        'status': 'ok',
        'users':  out,
        'perm_keys': INDEX_PERM_KEYS,
    })


# ═══════════════════════════════════════════════════════════════
# TASK SCHEDULER  ─  dept-aware kanban board
# ═══════════════════════════════════════════════════════════════

@app.route('/task_scheduler')
@login_required
def task_scheduler_page():
    if not can_access('task_scheduler'): return _denied('Task Scheduler')
    role    = session.get('User_Type')
    dept    = session.get('department', '')
    is_dh   = session.get('is_dept_head', False)
    uid     = session.get('UID')
    u_name  = session.get('User_Name')

    # Map to scheduler role
    if role == 'admin':
        sched_role = 'admin'
    elif is_dh:
        sched_role = 'dept_head'
    else:
        sched_role = 'user'

    conn = sampling_portal.get_db_connection()
    if sched_role == 'admin':
        tasks = conn.execute("""
            SELECT * FROM task_reminders
            ORDER BY CASE priority WHEN 'Critical' THEN 1 WHEN 'High' THEN 2
                                   WHEN 'Medium' THEN 3 ELSE 4 END,
                     due_date ASC, id DESC
        """).fetchall()
    else:
        tasks = conn.execute("""
            SELECT * FROM task_reminders WHERE department=%s
            ORDER BY CASE priority WHEN 'Critical' THEN 1 WHEN 'High' THEN 2
                                   WHEN 'Medium' THEN 3 ELSE 4 END,
                     due_date ASC, id DESC
        """, (dept,)).fetchall()
    conn.close()

    return render_template('task_scheduler.html',
        tasks=[dict(t) for t in tasks],
        sched_role=sched_role,
        user_dept=dept or '',
        user_name=u_name,
        uid=uid
    )


@app.route('/api/task_scheduler/save', methods=['POST'])
@login_required
def api_task_scheduler_save():
    role   = session.get('User_Type')
    dept   = session.get('department', '')
    is_dh  = session.get('is_dept_head', False)
    uid    = session.get('UID')
    sched_role = 'admin' if role == 'admin' else ('dept_head' if is_dh else 'user')

    if sched_role == 'user':
        return jsonify({'status': 'error', 'message': 'Permission denied'})

    data      = request.get_json()
    task_dept = data.get('department', dept)

    if sched_role == 'dept_head' and task_dept != dept:
        return jsonify({'status': 'error', 'message': 'You can only manage your own department'})

    conn    = sampling_portal.get_db_connection()
    task_id = data.get('id')
    try:
        if task_id:
            existing = conn.execute("SELECT department FROM task_reminders WHERE id=%s", (task_id,)).fetchone()
            if existing and sched_role == 'dept_head' and existing['department'] != dept:
                conn.close()
                return jsonify({'status': 'error', 'message': 'Permission denied'})
            conn.execute("""
                UPDATE task_reminders SET
                    title=%s, description=%s, due_date=%s, priority=%s,
                    status=%s, assigned_to=%s, department=%s, recurrence=%s, updated_at=NOW()
                WHERE id=%s
            """, (data.get('title'), data.get('description'), data.get('due_date'),
                  data.get('priority','Medium'), data.get('status','Pending'),
                  data.get('assigned_to'), task_dept, data.get('recurrence','None'), task_id))
            msg = "Task updated"
        else:
            conn.execute("""
                INSERT INTO task_reminders
                    (title, description, due_date, priority, status, assigned_to,
                     created_by, department, recurrence)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (data.get('title'), data.get('description'), data.get('due_date'),
                  data.get('priority','Medium'), data.get('status','Pending'),
                  data.get('assigned_to'), uid, task_dept, data.get('recurrence','None')))
            msg = "Task created"
        conn.commit()
        new_id = conn.lastrowid if not task_id else task_id
        conn.close()
        return jsonify({'status': 'success', 'message': msg, 'id': new_id})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/api/task_scheduler/delete', methods=['POST'])
@login_required
def api_task_scheduler_delete():
    role   = session.get('User_Type')
    dept   = session.get('department', '')
    is_dh  = session.get('is_dept_head', False)
    sched_role = 'admin' if role == 'admin' else ('dept_head' if is_dh else 'user')

    if sched_role == 'user':
        return jsonify({'status': 'error', 'message': 'Permission denied'})

    task_id = request.get_json().get('id')
    conn = sampling_portal.get_db_connection()
    existing = conn.execute("SELECT department FROM task_reminders WHERE id=%s", (task_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Task not found'})
    if sched_role == 'dept_head' and existing['department'] != dept:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Permission denied'})
    conn.execute("DELETE FROM task_reminders WHERE id=%s", (task_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})


@app.route('/api/task_scheduler/status', methods=['POST'])
@login_required
def api_task_scheduler_status():
    role   = session.get('User_Type')
    dept   = session.get('department', '')
    is_dh  = session.get('is_dept_head', False)
    sched_role = 'admin' if role == 'admin' else ('dept_head' if is_dh else 'user')

    if sched_role == 'user':
        return jsonify({'status': 'error', 'message': 'Permission denied'})

    data    = request.get_json()
    task_id = data.get('id')
    conn = sampling_portal.get_db_connection()
    existing = conn.execute("SELECT department FROM task_reminders WHERE id=%s", (task_id,)).fetchone()
    if existing and sched_role == 'dept_head' and existing['department'] != dept:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Permission denied'})
    conn.execute("UPDATE task_reminders SET status=%s, updated_at=NOW() WHERE id=%s",
                 (data.get('status'), task_id))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})


# ═══════════════════════════════════════════════════════════════
# LUNCH COUPONS  ─  print page, canteen admin only
# ═══════════════════════════════════════════════════════════════

@app.route('/lunch_coupons')
@login_required
def lunch_coupons_page():
    return ("Not Found", 404)  # lunch coupons feature removed
    if not can_access('lunch_coupons'): return _denied('Lunch Coupons')
    return render_template('lunch_coupons.html',
        user_name=session.get('User_Name'),
        role=session.get('User_Type')
    )


# CASH MANAGEMENT SYSTEM (CMS) — Routes
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/cms')
@login_required
def cms_page():
    if not can_access('cms'):
        return _denied('Cash Management System')
    sampling_portal.cms_init_db()
    sampling_portal.cms_init_scrap_vendor_tables()
    cms_portal.cms_v3_init_db()
    return render_template(
        'cms.html',
        role=(session.get('User_Type') or '').lower(),
        user=session.get('User_Name', session.get('UID', ''))
    )

@app.route('/api/cms/dashboard')
@login_required
def cms_dashboard():
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    data = sampling_portal.cms_get_dashboard(from_date, to_date)
    return jsonify(data)

@app.route('/api/cms/income')
@login_required
def cms_income_list():
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    rec_id    = request.args.get('id', None)
    source    = request.args.get('source', '')
    data = sampling_portal.cms_get_income(from_date, to_date, rec_id)
    if source and not rec_id:
        data['records'] = [r for r in data['records'] if r.get('source','') == source]
        data['total'] = sum(float(r['amount']) for r in data['records'])
    return jsonify(data)

@app.route('/api/cms/income/save', methods=['POST'])
@login_required
def cms_income_save():
    payload = request.get_json()
    payload['created_by'] = session.get('UID', '')
    result = sampling_portal.cms_save_income(payload)
    return jsonify(result)

@app.route('/api/cms/income/delete', methods=['POST'])
@login_required
def cms_income_delete():
    data = request.get_json()
    result = sampling_portal.cms_delete_income(data['id'])
    return jsonify(result)

@app.route('/api/cms/expense')
@login_required
def cms_expense_list():
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    category  = request.args.get('category', '')
    rec_id    = request.args.get('id', None)
    data = sampling_portal.cms_get_expenses(from_date, to_date, category, rec_id)
    return jsonify(data)

@app.route('/api/cms/expense/save', methods=['POST'])
@login_required
def cms_expense_save():
    payload = request.get_json()
    payload['created_by'] = session.get('UID', '')
    result = sampling_portal.cms_save_expense(payload)
    return jsonify(result)

@app.route('/api/cms/expense/delete', methods=['POST'])
@login_required
def cms_expense_delete():
    data = request.get_json()
    result = sampling_portal.cms_delete_expense(data['id'])
    return jsonify(result)

@app.route('/api/cms/loans')
@login_required
def cms_loans_list():
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    emp_id    = request.args.get('employee_id', None)
    rec_id    = request.args.get('id', None)
    data = sampling_portal.cms_get_loans(from_date, to_date, emp_id, rec_id)
    return jsonify(data)

@app.route('/api/cms/loans/save', methods=['POST'])
@login_required
def cms_loans_save():
    payload = request.get_json()
    payload['created_by'] = session.get('UID', '')
    result = sampling_portal.cms_save_loan(payload)
    return jsonify(result)

@app.route('/api/cms/loans/delete', methods=['POST'])
@login_required
def cms_loans_delete():
    data = request.get_json()
    result = sampling_portal.cms_delete_loan(data['id'])
    return jsonify(result)


@app.route('/api/cms/employees')
@login_required
def cms_employees():
    data = sampling_portal.cms_get_employees()
    return jsonify(data)

@app.route('/api/cms/employees/save', methods=['POST'])
@login_required
def cms_employees_save():
    payload = request.get_json()
    result = sampling_portal.cms_save_employee(payload)
    return jsonify(result)

@app.route('/api/cms/employees/delete', methods=['POST'])
@login_required
def cms_employees_delete():
    data = request.get_json()
    result = sampling_portal.cms_delete_employee(data['id'])
    return jsonify(result)

@app.route('/api/cms/employee/<int:emp_id>/ledger')
@login_required
def cms_employee_ledger(emp_id):
    data = sampling_portal.cms_get_employee_ledger(emp_id)
    return jsonify(data)

@app.route('/api/cms/opening_balance')
@login_required
def cms_opening_balance_get():
    data = sampling_portal.cms_get_opening_balance()
    return jsonify(data)

@app.route('/api/cms/opening_balance/save', methods=['POST'])
@login_required
def cms_opening_balance_save():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Only admins can modify opening balance'})
    payload = request.get_json()
    payload['changed_by'] = session.get('UID', '')
    result = sampling_portal.cms_save_opening_balance(payload)
    return jsonify(result)

@app.route('/api/cms/opening_balance/history')
@login_required
def cms_opening_balance_history():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'history': []})
    data = sampling_portal.cms_get_ob_history()
    return jsonify(data)

@app.route('/api/cms/report/daily')
@login_required
def cms_report_daily():
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    data = sampling_portal.cms_report_daily(from_date, to_date)
    return jsonify(data)

@app.route('/api/cms/report/category')
@login_required
def cms_report_category():
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    data = sampling_portal.cms_report_category(from_date, to_date)
    return jsonify(data)

@app.route('/api/cms/report/ledger')
@login_required
def cms_report_ledger():
    emp_id    = request.args.get('employee_id', '')
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    data = sampling_portal.cms_report_ledger(emp_id, from_date, to_date)
    return jsonify(data)

@app.route('/api/cms/report/expense_pdf')
@login_required
def cms_expense_pdf():
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    category  = request.args.get('category', '')
    data      = sampling_portal.cms_get_expenses(from_date, to_date, category)
    records   = data.get('records', [])
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=18*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story  = []
    story.append(Paragraph('HCP Wellness Pvt Ltd', styles['Title']))
    story.append(Paragraph('Cash Expense Report', styles['Heading2']))
    story.append(Paragraph(f"Period: {from_date or 'All'} to {to_date or 'All'}" +
                           (f"  |  Category: {category}" if category else ''), styles['Normal']))
    story.append(Spacer(1, 8*mm))
    table_data = [['Date', 'Category', 'Description', 'Voucher', 'Amount']]
    total = 0
    for r in records:
        table_data.append([r['date'], r['category'],
                           (r.get('description') or '')[:40],
                           r.get('voucher_no') or '',
                           f"Rs. {float(r['amount']):,.2f}"])
        total += float(r.get('amount', 0))
    table_data.append(['', '', '', 'TOTAL', f"Rs. {total:,.2f}"])
    t = Table(table_data, colWidths=[25*mm, 38*mm, 60*mm, 28*mm, 32*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d9488')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#f3f4f6')]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#fee2e2')),
        ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#d1d5db')),
        ('ALIGN',      (4,0), (4,-1), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f'HCP_Expenses_{from_date}_{to_date}.pdf')

@app.route('/api/cms/report/daily_pdf')
@login_required
def cms_daily_pdf():
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    data      = sampling_portal.cms_report_daily(from_date, to_date)
    rows      = data.get('rows', [])
    summary   = data.get('summary', {})
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=18*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story  = []
    story.append(Paragraph('HCP Wellness Pvt Ltd', styles['Title']))
    story.append(Paragraph('Daily Cash Report', styles['Heading2']))
    story.append(Paragraph(f"Period: {from_date} to {to_date}", styles['Normal']))
    story.append(Spacer(1, 6*mm))
    summ_data = [
        ['Total Income',  f"Rs. {float(summary.get('total_income',0)):,.2f}"],
        ['Total Expense', f"Rs. {float(summary.get('total_expense',0)):,.2f}"],
        ['Net Balance',   f"Rs. {float(summary.get('net',0)):,.2f}"],
    ]
    st = Table(summ_data, colWidths=[60*mm, 60*mm])
    st.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#f3f4f6')),
        ('FONTNAME',   (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#d1d5db')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(st)
    story.append(Spacer(1, 6*mm))
    table_data = [['Date', 'Opening Balance', 'Total Income', 'Total Expense', 'Closing Balance']]
    for r in rows:
        table_data.append([r['date'],
                           f"Rs. {float(r['opening']):,.2f}",
                           f"Rs. {float(r['income']):,.2f}",
                           f"Rs. {float(r['expense']):,.2f}",
                           f"Rs. {float(r['closing']):,.2f}"])
    t = Table(table_data, colWidths=[30*mm, 37*mm, 35*mm, 35*mm, 38*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d9488')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#d1d5db')),
        ('ALIGN',      (1,0), (-1,-1), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f'HCP_Daily_Report_{from_date}_{to_date}.pdf')

@app.route('/api/cms/report/category_pdf')
@login_required
def cms_category_pdf():
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    data      = sampling_portal.cms_report_category(from_date, to_date)
    rows      = data.get('rows', [])
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=18*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story  = []
    story.append(Paragraph('HCP Wellness Pvt Ltd', styles['Title']))
    story.append(Paragraph('Category-wise Expense Report', styles['Heading2']))
    story.append(Paragraph(f"Period: {from_date} to {to_date}", styles['Normal']))
    story.append(Spacer(1, 8*mm))
    total = sum(float(r['total']) for r in rows)
    table_data = [['Category', 'No. of Entries', 'Total Amount', '% of Total']]
    for r in rows:
        pct = f"{(float(r['total'])/total*100):.1f}%" if total > 0 else '0%'
        table_data.append([r['category'], str(r['count']),
                           f"Rs. {float(r['total']):,.2f}", pct])
    table_data.append(['TOTAL', '', f"Rs. {total:,.2f}", '100%'])
    t = Table(table_data, colWidths=[65*mm, 35*mm, 50*mm, 30*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d9488')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#f3f4f6')]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#fee2e2')),
        ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#d1d5db')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f'HCP_Category_Report_{from_date}_{to_date}.pdf')

@app.route('/api/cms/report/ledger_pdf')
@login_required
def cms_ledger_pdf():
    emp_id    = request.args.get('employee_id', '')
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    data      = sampling_portal.cms_report_ledger(emp_id, from_date, to_date)
    rows      = data.get('rows', [])
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=18*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story  = []
    story.append(Paragraph('HCP Wellness Pvt Ltd', styles['Title']))
    story.append(Paragraph('Employee Ledger Report', styles['Heading2']))
    story.append(Paragraph(f"Period: {from_date or 'All'} to {to_date or 'All'}", styles['Normal']))
    story.append(Spacer(1, 6*mm))
    summ_data = [
        ['Total Loan Advanced', f"Rs. {float(data.get('total_advanced',0)):,.2f}"],
        ['Total Recovered',     f"Rs. {float(data.get('total_recovered',0)):,.2f}"],
        ['Outstanding Balance', f"Rs. {float(data.get('outstanding',0)):,.2f}"],
    ]
    st = Table(summ_data, colWidths=[60*mm, 60*mm])
    st.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#f3f4f6')),
        ('FONTNAME',   (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#d1d5db')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(st)
    story.append(Spacer(1, 6*mm))
    table_data = [['Employee', 'Date', 'Type', 'Description', 'Loan Given', 'Recovered', 'Balance']]
    for r in rows:
        loan = f"Rs. {float(r['amount']):,.2f}" if r['txn_type'] == 'Employee Loan' else '—'
        rcvd = f"Rs. {float(r['amount']):,.2f}" if r['txn_type'] == 'Employee Repayment' else '—'
        table_data.append([r['employee_name'], r['date'],
                           'LOAN' if r['txn_type'] == 'Employee Loan' else 'REPAY',
                           (r.get('description') or '')[:30],
                           loan, rcvd, f"Rs. {float(r['running_balance']):,.2f}"])
    t = Table(table_data, colWidths=[30*mm, 22*mm, 18*mm, 40*mm, 24*mm, 24*mm, 24*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d9488')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 7),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#d1d5db')),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     download_name='HCP_Ledger_Report.pdf')


@app.route('/api/cms/loan_accounts/summary')
@login_required
def cms_loan_accounts_summary():
    return jsonify(sampling_portal.cms_get_loan_account_summary())


@app.route('/api/cms/loan_accounts/detail')
@login_required
def cms_loan_accounts_detail():
    account   = request.args.get('account', 'General Advances')
    emp_id    = request.args.get('employee_id', '')
    from_date = request.args.get('from', '')
    to_date   = request.args.get('to', '')
    return jsonify(sampling_portal.cms_get_loan_account_detail(account, emp_id or None, from_date, to_date))


# ── RECEIVABLES / PAYABLES ───────────────────────────────────────────────────

@app.route('/api/cms/receivables')
@login_required
def cms_receivables():
    cat = request.args.get('category', '')
    return jsonify(sampling_portal.cms_get_receivables(cat))


@app.route('/api/cms/payables')
@login_required
def cms_payables():
    cat = request.args.get('category', '')
    return jsonify(sampling_portal.cms_get_payables(cat))


@app.route('/api/cms/loan_accounts')
@login_required
def cms_loan_accounts():
    return jsonify(sampling_portal.cms_get_loan_accounts())


# ── SCRAP VENDORS ────────────────────────────────────────────────────────────

@app.route('/api/cms/scrap_vendors')
@login_required
def cms_scrap_vendors_list():
    return jsonify(sampling_portal.cms_get_scrap_vendors())


@app.route('/api/cms/scrap_vendors/save', methods=['POST'])
@login_required
def cms_scrap_vendors_save():
    data = request.get_json()
    data['created_by'] = session.get('User_Name', session.get('UID', ''))
    return jsonify(sampling_portal.cms_save_scrap_vendor(data))


@app.route('/api/cms/scrap_vendors/delete', methods=['POST'])
@login_required
def cms_scrap_vendors_delete():
    data = request.get_json()
    return jsonify(sampling_portal.cms_delete_scrap_vendor(data['id']))


@app.route('/api/cms/scrap_vendor/<int:vendor_id>/ledger')
@login_required
def cms_scrap_vendor_ledger(vendor_id):
    return jsonify(sampling_portal.cms_get_scrap_vendor_ledger(vendor_id))


@app.route('/api/cms/scrap_vendor_txn/save', methods=['POST'])
@login_required
def cms_scrap_vendor_txn_save():
    data = request.get_json()
    data['created_by'] = session.get('User_Name', session.get('UID', ''))
    return jsonify(sampling_portal.cms_save_scrap_vendor_txn(data))


@app.route('/api/cms/scrap_vendor_txn/<int:txn_id>')
@login_required
def cms_scrap_vendor_txn_get(txn_id):
    return jsonify(sampling_portal.cms_get_scrap_vendor_txn(txn_id))


@app.route('/api/cms/scrap_vendor_txn/delete', methods=['POST'])
@login_required
def cms_scrap_vendor_txn_delete():
    data = request.get_json()
    return jsonify(sampling_portal.cms_delete_scrap_vendor_txn(data['id']))


# ── BACKUP ROUTES (admin only) ───────────────────────────────────────────────

@app.route('/api/cms/backup/trigger', methods=['POST'])
@login_required
def cms_backup_trigger():
    """Manual backup — admin only."""
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'})
    user = session.get('UID', 'admin')
    result = sampling_portal.cms_create_backup(triggered_by=f"manual:{user}")
    return jsonify(result)

@app.route('/api/cms/backup/list')
@login_required
def cms_backup_list():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'files': []})
    return jsonify(sampling_portal.cms_list_backups())

@app.route('/api/cms/backup/log')
@login_required
def cms_backup_log():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'logs': []})
    return jsonify(sampling_portal.cms_get_backup_log())

@app.route('/api/cms/backup/download/<path:filename>')
@login_required
def cms_backup_download(filename):
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    import re
    if not re.match(r'^cms_backup_[\w_]+\.sql\.gz$', filename):
        return jsonify({'status': 'error', 'message': 'Invalid filename'}), 400
    backup_dir = sampling_portal.get_active_backup_dir()
    return send_from_directory(backup_dir, filename, as_attachment=True)

@app.route('/api/cms/backup/delete', methods=['POST'])
@login_required
def cms_backup_delete():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'})
    filename = (request.get_json() or {}).get('filename', '')
    import re
    if not re.match(r'^cms_backup_[\w_]+\.sql\.gz$', filename):
        return jsonify({'status': 'error', 'message': 'Invalid filename'})
    fpath = os.path.join(sampling_portal.get_active_backup_dir(), filename)
    try:
        os.remove(fpath)
        return jsonify({'status': 'ok'})
    except FileNotFoundError:
        return jsonify({'status': 'error', 'message': 'File not found'})


# ══════════════════════════════════════════════════════════════════════════════
# CMS v3 ROUTES  ── Voucher-based Cash Management
# ══════════════════════════════════════════════════════════════════════════════

# ── LEDGER GROUPS ─────────────────────────────────────────────────────────────
@app.route('/api/cms/ledger_groups')
@login_required
def api_cms_ledger_groups():
    return jsonify(cms_portal.cms_get_ledger_groups())

@app.route('/api/cms/ledger_groups/save', methods=['POST'])
@login_required
def api_cms_ledger_group_save():
    return jsonify(cms_portal.cms_save_ledger_group(request.get_json()))

@app.route('/api/cms/ledger_groups/delete', methods=['POST'])
@login_required
def api_cms_ledger_group_delete():
    d = request.get_json()
    return jsonify(cms_portal.cms_delete_ledger_group(d['id']))


# ── LEDGERS ───────────────────────────────────────────────────────────────────
@app.route('/api/cms/ledgers')
@login_required
def api_cms_ledgers():
    return jsonify(cms_portal.cms_get_ledgers(
        ledger_type=request.args.get('type', ''),
        group_id=request.args.get('group_id'),
        search=request.args.get('q', '')
    ))

@app.route('/api/cms/ledgers/save', methods=['POST'])
@login_required
def api_cms_ledger_save():
    return jsonify(cms_portal.cms_save_ledger(request.get_json()))

@app.route('/api/cms/ledgers/delete', methods=['POST'])
@login_required
def api_cms_ledger_delete():
    d = request.get_json()
    return jsonify(cms_portal.cms_delete_ledger(d['id']))

@app.route('/api/cms/ledgers/<int:lid>/report')
@login_required
def api_cms_ledger_report(lid):
    return jsonify(cms_portal.cms_ledger_report(
        lid,
        from_date=request.args.get('from', ''),
        to_date=request.args.get('to', '')
    ))

@app.route('/api/cms/ledgers/<int:lid>/advances')
@login_required
def api_cms_ledger_advances(lid):
    return jsonify(cms_portal.cms_get_pending_advances(lid))


# ── VOUCHERS ──────────────────────────────────────────────────────────────────
@app.route('/api/cms/vouchers')
@login_required
def api_cms_vouchers():
    return jsonify(cms_portal.cms_get_vouchers(
        from_date=request.args.get('from', ''),
        to_date=request.args.get('to', ''),
        vtype=request.args.get('type', ''),
        ledger_id=request.args.get('ledger_id'),
        page=int(request.args.get('page', 1)),
        per_page=int(request.args.get('per_page', 25))
    ))

@app.route('/api/cms/vouchers/<int:vid>')
@login_required
def api_cms_voucher_get(vid):
    v = cms_portal.cms_get_voucher(vid)
    if not v:
        return jsonify({'error': 'not found'}), 404
    return jsonify(v)

@app.route('/api/cms/vouchers/save', methods=['POST'])
@login_required
def api_cms_voucher_save():
    data = request.get_json()
    data['created_by'] = session.get('UID', '')
    return jsonify(cms_portal.cms_save_voucher(data))

@app.route('/api/cms/vouchers/delete', methods=['POST'])
@login_required
def api_cms_voucher_delete():
    d = request.get_json()
    return jsonify(cms_portal.cms_delete_voucher(d['id']))


# ── DASHBOARD  (v3 uses /v3/ prefix to avoid clash with old route) ────────────
@app.route('/api/cms/v3/dashboard')
@login_required
def api_cms_v3_dashboard():
    return jsonify(cms_portal.cms_dashboard(
        from_date=request.args.get('from', ''),
        to_date=request.args.get('to', '')
    ))


# ── REPORTS ───────────────────────────────────────────────────────────────────
@app.route('/api/cms/report/advances')
@login_required
def api_cms_advances():
    return jsonify(cms_portal.cms_advance_report())

@app.route('/api/cms/report/receivables')
@login_required
def api_cms_receivables_v3():
    return jsonify(cms_portal.cms_receivables())

@app.route('/api/cms/report/payables')
@login_required
def api_cms_payables_v3():
    return jsonify(cms_portal.cms_payables())

@app.route('/api/cms/v3/report/daily')
@login_required
def api_cms_v3_daily():
    from datetime import date as _date
    dt = request.args.get('date', _date.today().isoformat())
    return jsonify(cms_portal.cms_daily_report(dt))

@app.route('/api/cms/v3/report/category')
@login_required
def api_cms_v3_category():
    return jsonify(cms_portal.cms_category_report(
        from_date=request.args.get('from', ''),
        to_date=request.args.get('to', ''),
        nature=request.args.get('nature', 'expense')
    ))


# ── REMINDERS ─────────────────────────────────────────────────────────────────
@app.route('/api/cms/reminders')
@login_required
def api_cms_reminders():
    done = request.args.get('done', '0') == '1'
    return jsonify(cms_portal.cms_get_reminders(done))

@app.route('/api/cms/reminders/save', methods=['POST'])
@login_required
def api_cms_reminder_save():
    d = request.get_json()
    d['created_by'] = session.get('UID', '')
    return jsonify(cms_portal.cms_save_reminder(d))


# ── CMS ADMIN RESET ──────────────────────────────────────────────────────────
@app.route('/api/cms/admin/reset', methods=['POST'])
@login_required
def api_cms_admin_reset():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    data = request.get_json() or {}
    if data.get('confirm') != 'RESET ALL CONFIRMED':
        return jsonify({'status': 'error', 'message': 'Invalid confirmation token'}), 400
    conn = cms_portal.get_db()
    cur  = conn.cursor()
    try:
        cur.execute("SET FOREIGN_KEY_CHECKS = 0")
        wiped = []
        for t in ['cms_voucher_lines','cms_vouchers','cms_ledger_balances','cms_ob_history','cms_reminders','cms_ledgers']:
            try:
                cur.execute("TRUNCATE TABLE `%s`" % t)
                wiped.append(t)
            except Exception:
                pass
        cur.execute("SET FOREIGN_KEY_CHECKS = 1")
        for name, nature in [
            ('Cash & Bank','asset'),('Employee Advances','asset'),
            ('Accounts Receivable','asset'),('Sales / Scrap','income'),
            ('Other Income','income'),('Operating Expense','expense'),
            ('Staff Expense','expense'),('Admin Expense','expense'),
            ('Accounts Payable','liability'),
        ]:
            cur.execute("INSERT IGNORE INTO cms_ledger_groups (name,nature,is_system) VALUES (%s,%s,1)", (name, nature))
        cur.execute("SELECT id FROM cms_ledger_groups WHERE name='Cash & Bank' LIMIT 1")
        grp = cur.fetchone()
        if grp:
            cur.execute("INSERT INTO cms_ledgers (name,ledger_group_id,ledger_type,opening_balance,ob_type) VALUES ('Cash in Hand',%s,'cash',0,'dr')", (grp['id'],))
        conn.commit()
        import datetime as _dt
        uid = session.get('UID','admin')
        print("[CMS RESET] by='%s' at=%s wiped=%s" % (uid, _dt.datetime.now().isoformat(), wiped))
        return jsonify({'status':'ok','message':'Reset complete','wiped':wiped,'by':uid})
    except Exception as e:
        conn.rollback()
        try: cur.execute("SET FOREIGN_KEY_CHECKS = 1")
        except: pass
        return jsonify({'status':'error','message':str(e)}), 500
    finally:
        conn.close()


# ── CMS: Cashbook Excel Export ────────────────────────────────────────────────
@app.route('/api/cms/export/cashbook')
@login_required
def api_cms_cashbook_export():
    from_date = request.args.get('from','')
    to_date   = request.args.get('to','')
    vtype     = request.args.get('type','')
    data      = cms_portal.cms_get_cashbook_export(from_date, to_date, vtype)
    return jsonify(data)

# ── END CMS ROUTES ─────────────────────────────────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════════════
# PRODUCTION DEPARTMENT ROUTES — moved to production_dept_routes.py (Blueprint)
# Registered above via: app.register_blueprint(production_dept_bp)
# ══════════════════════════════════════════════════════════════════════════════

# _can_prod_dept() is kept here so other modules in app.py (e.g. QC status_map)
# can still call it without importing from the blueprint.
def _can_prod_dept():
    role = session.get('User_Type', '')
    return 'production_dept' in _user_allowed_pages() or role.lower() == 'production'

# ══════════════════════════════════════════════════════════════════════════════
# BACKUP MANAGER ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/backup')
@app.route('/backup_dashboard')
@login_required
def backup_page():
    if (session.get('User_Type') or '').lower() != 'admin':
        return _denied('Backup Manager')
    return render_template('backup_dashboard.html')


@app.route('/api/backup/config', methods=['GET'])
@login_required
def api_backup_get_config():
    """Return current backup path configuration — admin only."""
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    # Read raw DB row directly so we can return unconfigured state without raising
    sampling_portal._ensure_backup_config_table()
    conn = sampling_portal.get_db_connection()
    try:
        row = conn.execute(
            'SELECT primary_path, drive_d_path, network_path, updated_at, updated_by '
            'FROM backup_config WHERE id = 1'
        ).fetchone()
    finally:
        conn.close()
    primary = (row['primary_path'] if row else '').strip()
    drive_d = (row['drive_d_path'] if row else '').strip()
    network = (row['network_path'] if row else '').strip()
    missing = [n for n, v in [('Primary (Server Local)', primary),
                               ('Local Drive D', drive_d),
                               ('Network Server', network)] if not v]
    return jsonify({
        'status': 'ok',
        'config': {
            'primary_path': primary,
            'drive_d_path': drive_d,
            'network_path': network,
            'updated_at':   str(row['updated_at']) if row and row.get('updated_at') else None,
            'updated_by':   row['updated_by'] if row else None,
        },
        'paths_configured': len(missing) == 0,
        'missing_paths': missing,
    })


@app.route('/api/backup/config', methods=['POST'])
@login_required
def api_backup_set_config():
    """Save admin-chosen backup paths to the DB — admin only.
    ALL THREE paths are required; saving with any blank path is rejected.
    """
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    data = request.get_json() or {}
    primary = data.get('primary_path', '').strip()
    drive_d = data.get('drive_d_path', '').strip()
    network = data.get('network_path', '').strip()

    # Enforce: all three paths must be provided — no silent fallback to defaults
    missing = [n for n, v in [('Primary (Server Local)', primary),
                               ('Local Drive D', drive_d),
                               ('Network Server', network)] if not v]
    if missing:
        return jsonify({
            'status': 'error',
            'message': 'All three backup paths are required. Missing: ' + ', '.join(missing)
        }), 400

    # Validate: primary directory must be reachable/creatable on this server
    try:
        os.makedirs(primary, exist_ok=True)
    except Exception as e:
        return jsonify({'status': 'error',
                        'message': f'Cannot create/access Primary path: {e}'}), 400

    try:
        updated = sampling_portal.set_backup_config(
            primary_path=primary,
            drive_d_path=drive_d,
            network_path=network,
            updated_by=session.get('Username', 'admin'),
        )
        return jsonify({'status': 'ok', 'config': updated})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/backup/run', methods=['POST'])
@login_required
def api_backup_run():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    try:
        result = backup_system.run_full_backup(triggered_by="manual")
        return jsonify(result)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/backup/log')
@login_required
def api_backup_log():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    limit = int(request.args.get('limit', 100))
    return jsonify({'status': 'ok', 'logs': backup_system.get_backup_log(limit)})


@app.route('/api/backup/stats')
@login_required
def api_backup_stats():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    return jsonify({'status': 'ok', 'stats': backup_system.get_backup_stats()})


@app.route('/api/backup/files')
@login_required
def api_backup_files():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    return jsonify({'status': 'ok', 'files': backup_system.list_backup_files()})


@app.route('/api/backup/download/<path:filename>')
@login_required
def api_backup_download(filename):
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    if not re.match(r'^hcp_(db|appfiles)_(auto|manual)_\d{8}_\d{6}\.(sql\.gz|zip)$', filename):
        return jsonify({'status': 'error', 'message': 'Invalid filename'}), 400
    return send_from_directory(sampling_portal.get_active_backup_dir(), filename, as_attachment=True)


@app.route('/api/backup/prune', methods=['POST'])
@login_required
def api_backup_prune():
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    try:
        deleted = backup_system.prune_old_backups()
        return jsonify({'status': 'ok', 'deleted': deleted})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/backup/restore', methods=['POST'])
@login_required
def api_backup_restore():
    """Restore MySQL database from a selected .sql.gz backup file. Admin only."""
    if (session.get('User_Type') or '').lower() != 'admin':
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    data     = request.get_json() or {}
    filename = data.get('filename', '').strip()
    confirm  = data.get('confirm', False)
    if not filename:
        return jsonify({'status': 'error', 'message': 'No filename provided'})
    if not confirm:
        return jsonify({'status': 'error', 'message': 'Restore not confirmed — set confirm:true'})
    try:
        result = backup_system.restore_database(filename)
        return jsonify(result)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ── END BACKUP MANAGER ROUTES ──────────────────────────────────────────────────


# ══════════════════════════════════════════════════════════════════════════════
# QC DASHBOARD ROUTES
# ══════════════════════════════════════════════════════════════════════════════
# All routes for /qc_dashboard, /qc_sampling, /api/qc/* and /api/qc_sampling/*
# have been migrated to the qc blueprint (see qc/qc_routes.py). The blueprint
# is registered at the top of this file via:  app.register_blueprint(qc_bp)
#
# Only the thin _can_qc_dashboard() helper below is kept here, since legacy
# routes in app.py (e.g. the Packing routes) still reference it for shared
# access checks. For all new logic, use the helper inside the blueprint.

def _can_qc_dashboard():
    role = session.get('User_Type', '')
    if role == 'admin' or can_access('qc_dashboard'):
        return True
    if (role or '').lower() in ('qc', 'qc_common', 'purchase'):
        return True
    return False


# ── QC Label PDF — pure Python via reportlab (Windows compatible, no wkhtmltopdf) ──
@app.route('/api/qc/label_pdf', methods=['POST'])
@login_required
def api_qc_label_pdf():
    if not _can_qc_dashboard():
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    try:
        from reportlab.lib.pagesizes import mm
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        import io

        d    = request.get_json() or {}
        rows = d.get('rows', [])
        if not rows:
            return jsonify({'status': 'error', 'message': 'No label data'}), 400

        W, H   = 100*mm, 50*mm
        lbl_w  = 22*mm
        buf    = io.BytesIO()
        pdf    = canvas.Canvas(buf, pagesize=(W, H))

        def draw_label(row, drum_num, total_drums):
            pdf.setPageSize((W, H))

            # Outer border
            pdf.setLineWidth(2)
            pdf.setStrokeColor(colors.black)
            pdf.rect(0, 0, W, H, stroke=1, fill=0)

            y = H

            # STATUS BAR
            bar_h = 8*mm
            y -= bar_h
            pdf.setFillColor(colors.white)
            pdf.rect(0, y, W, bar_h, stroke=0, fill=1)
            pdf.setLineWidth(2)
            pdf.line(0, y, W, y)
            pdf.setFillColor(colors.black)
            pdf.setFont('Helvetica-Bold', 11)
            status_label = row.get('status_label', 'QC - PENDING')
            pdf.drawCentredString(W/2, y + 2.5*mm, status_label)

            # BATCH NAME
            name_h = 7*mm
            y -= name_h
            pdf.setFillColor(colors.white)
            pdf.rect(0, y, W, name_h, stroke=0, fill=1)
            pdf.setLineWidth(2)
            pdf.line(0, y, W, y)
            pdf.setFillColor(colors.black)
            pdf.setFont('Helvetica-Bold', 9)
            batch = str(row.get('batch_name', ''))
            while pdf.stringWidth(batch, 'Helvetica-Bold', 9) > W - 4*mm and len(batch) > 5:
                batch = batch[:-1]
            pdf.drawCentredString(W/2, y + 2*mm, batch)

            # GRID — 5 rows x 2 columns + 1 drums row
            grid_top = y
            fields = [
                ('TRS No :',       row.get('trs_no', '—')),
                ('TRS Date :',     row.get('trs_date', '—')),
                ('Batch No :',     row.get('product_code', '—')),
                ('Batch Size :',   str(row.get('batch_size', '—')) + ' Kg'),
                ('Batch Date :',   row.get('batch_date', '—')),
                ('Incharge :',     row.get('operator_name', '—')),
                ('Sampling Date :', row.get('sampling_date', '—')),
                ('Approval Date :', row.get('approval_dt', '—')),
                ('Sample Size :',  row.get('sample_qty', '—')),
                ('Approved By :',  row.get('approved_by', '—')),
            ]
            n_rows  = len(fields) // 2   # 5
            drums_h = 5*mm
            row_h   = (grid_top - drums_h) / n_rows
            col_w   = W / 2

            pdf.setLineWidth(1.2)
            for i, (lbl, val) in enumerate(fields):
                col   = i % 2
                r_idx = i // 2
                rx    = col * col_w
                ry    = grid_top - (r_idx + 1) * row_h

                # Cell fill
                pdf.setFillColor(colors.white)
                pdf.rect(rx, ry, col_w, row_h, stroke=0, fill=1)
                pdf.setStrokeColor(colors.black)
                pdf.line(rx, ry, rx + col_w, ry)           # bottom
                if col == 0:
                    pdf.line(rx + col_w, ry, rx + col_w, ry + row_h)  # right

                # Label bg
                pdf.setFillColor(colors.HexColor('#f0f0f0'))
                pdf.rect(rx, ry, lbl_w, row_h, stroke=0, fill=1)
                pdf.setFillColor(colors.black)
                pdf.line(rx + lbl_w, ry, rx + lbl_w, ry + row_h)

                # Label text
                pdf.setFont('Helvetica-Bold', 5.5)
                pdf.drawString(rx + 1.5*mm, ry + 1.5*mm, str(lbl))

                # Value text
                pdf.setFont('Helvetica-Bold', 6.5)
                val_str = str(val) if val else '—'
                max_w   = col_w - lbl_w - 2*mm
                while pdf.stringWidth(val_str, 'Helvetica-Bold', 6.5) > max_w and len(val_str) > 3:
                    val_str = val_str[:-1]
                pdf.drawString(rx + lbl_w + 1.5*mm, ry + 1.5*mm, val_str)

            # DRUMS ROW (full width)
            dr_y = grid_top - (n_rows + 1) * row_h
            pdf.setFillColor(colors.white)
            pdf.rect(0, dr_y, W, drums_h, stroke=0, fill=1)
            pdf.setLineWidth(1.2)
            pdf.setStrokeColor(colors.black)
            pdf.line(0, dr_y, W, dr_y)
            # Label bg
            pdf.setFillColor(colors.HexColor('#f0f0f0'))
            pdf.rect(0, dr_y, lbl_w, drums_h, stroke=0, fill=1)
            pdf.setFillColor(colors.black)
            pdf.line(lbl_w, dr_y, lbl_w, dr_y + drums_h)
            pdf.setFont('Helvetica-Bold', 5.5)
            pdf.drawString(1.5*mm, dr_y + 1.5*mm, 'No of Drums :')
            pdf.setFont('Helvetica-Bold', 8)
            drum_str = 'Drums {} of {}'.format(drum_num, total_drums)
            pdf.drawCentredString(lbl_w + (W - lbl_w)/2, dr_y + 1.5*mm, drum_str)

        for row_data in rows:
            total = int(row_data.get('total_drums', 1))
            for drum_num in range(1, total + 1):
                draw_label(row_data, drum_num, total)
                pdf.showPage()

        pdf.save()
        buf.seek(0)
        pdf_bytes = buf.read()

        batch_name = rows[0].get('batch_name', 'label') if rows else 'label'
        safe_name  = ''.join(ch for ch in batch_name if ch.isalnum() or ch in ' _-')[:40].strip().replace(' ', '_')
        filename   = 'QC_Label_{}.pdf'.format(safe_name)

        from flask import Response
        return Response(
            pdf_bytes,
            mimetype='application/pdf',
            headers={
                'Content-Disposition': 'attachment; filename="{}"'.format(filename),
                'Content-Type': 'application/pdf',
            }
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ── END QC DASHBOARD ROUTES ───────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# PACKING DEPARTMENT — Sample Receipt / Entry Log
# Table creation is handled by sampling_portal._ensure_packing_table()
# ══════════════════════════════════════════════════════════════════════════════

# NOTE: The /packing page route was removed (template packing.html no longer
# exists). The API endpoints (/api/packing/list, /save, /delete, /brands)
# remain functional for any external integrations or future re-introduction
# of the page. Users with 'packing' role now fall through to the default
# portal home (see login + home dispatch above) instead of being redirected
# to a missing page.

@app.route('/packing')
@login_required
def packing_page():
    # Re-introduced June 2026: packing.html restored from backup template.
    # Packing role users + QC dashboard users may view this page.
    if not can_access('packing') and not _can_qc_dashboard():
        return _denied('Packing')
    return render_template(
        'packing.html',
        user_name=session.get('User_Name'),
        role=session.get('User_Type'),
    )


@app.route('/api/packing/list')
@login_required
def api_packing_list():
    if not can_access('packing') and not _can_qc_dashboard():
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    try:
        from_date = request.args.get('from', '')
        to_date   = request.args.get('to', '')
        rows = sampling_portal.packing_list(from_date, to_date)
        # Serialise date/datetime objects
        for r in rows:
            for k, v in r.items():
                if hasattr(v, 'isoformat'):
                    r[k] = v.isoformat()
        return jsonify({'status': 'ok', 'rows': rows})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/packing/save', methods=['POST'])
@login_required
def api_packing_save():
    if not can_access('packing') and not _can_qc_dashboard():
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    d = request.get_json() or {}
    if not (d.get('product_name') or '').strip():
        return jsonify({'status': 'error', 'message': 'Product Name is required'})
    # QC-only save: only status, received_date, remark fields updated
    role = session.get('User_Type', '')
    if (role or '').lower() in ('qc', 'qc_common', 'qc_common') and d.get('id'):
        d['_qc_only'] = True
    try:
        new_id = sampling_portal.packing_save(d, created_by=session.get('UID', ''))
        return jsonify({'status': 'ok', 'id': new_id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/packing/delete', methods=['POST'])
@login_required
def api_packing_delete():
    role = session.get('User_Type', '')
    if (role or '').lower() in ('qc', 'qc_common', 'qc_common'):
        return jsonify({'status': 'error', 'message': 'QC users cannot delete packing entries'}), 403
    if not can_access('packing'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    rid = (request.get_json() or {}).get('id')
    if not rid:
        return jsonify({'status': 'error', 'message': 'Missing id'}), 400
    try:
        sampling_portal.packing_delete(rid)
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ── END PACKING ROUTES ──────────────────────────────────────────────────────


# ── Procurement Brands — for packing entry brand autocomplete ────────────────
@app.route('/api/packing/brands')
@login_required
def api_packing_brands():
    """Return all brands from procurement_brands table, ordered by name."""
    try:
        conn = sampling_portal.get_db_connection()
        rows = conn.execute(
            "SELECT id, name, color FROM procurement_brands ORDER BY name ASC"
        ).fetchall()
        conn.close()
        return jsonify({'status': 'ok', 'brands': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ── END BRANDS ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("🚀 Starting HCP Portal on http://0.0.0.0:80")
    try:
        serve(
        app,
        host='127.0.0.1',
        port=8000,
        threads=32,
        connection_limit=500,
        channel_timeout=120,
    )
    except KeyboardInterrupt:
        # Python 3.14 + Waitress compatibility fix:
        # Waitress shutdown raises KeyboardInterrupt internally on Ctrl+C.
        # Catching it here gives a clean exit instead of a traceback.
        print("\n⛔ Server stopped.")
