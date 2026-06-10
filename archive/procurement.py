r"""
procurement.py  –  Procurement Dashboard Module

Access : admin (role) | sonal (username)
Template: templates/procurement.html
DB tables: procurement_materials  (editable columns, no in_stock_qty)
             procurement_settings    (key-value config: stksum_path, etc.)

In Stock Qty is ALWAYS read live from StkSum.xlsx on every API call:
  Path  : \\Tarakbhavsar\procurement new\CURRENT RM\StkSum.xlsx
  Col A : Material Name  (matched case-insensitively)
  Col C : Quantity
  Rows  : 16 onwards
  If a material is not found in the file → qty = 0

Import / Export
  Export : client-side via SheetJS (all columns; In Stock Qty is read-only info)
  Import : POST /api/procurement/import  { rows:[...] }
           Upserts by material_name. in_stock_qty in the import file is IGNORED
           (qty always comes from StkSum.xlsx, never from user input).
"""

from __future__ import annotations

import os
import traceback
from datetime import datetime
from functools import wraps

from flask import render_template, session, jsonify, redirect, url_for, request

import sampling_portal

try:
    from openpyxl import load_workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_STKSUM_PATH = r"\\Tarakbhavsar\procurement new\CURRENT RM\StkSum.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# HARDCODED MATERIAL RATES
# Fallback rates used in ALL cost calculations when a material has no
# last_purchase_rate set in procurement_materials.
# Keys are lower-cased. DB value always takes priority over these.
# Add more entries here as needed — no other code changes required.
# ─────────────────────────────────────────────────────────────────────────────
HARDCODED_MATERIAL_RATES = {
    "demineralized water":  0.43,
    "demineralised water":  0.43,
    "dm water":             0.43,
    "d m water":            0.43,
    "d.m. water":           0.43,
    "demi water":           0.43,
}

def _hardcoded_rate(material_name):
    """
    Return a hardcoded fallback rate (float) for a material name, or None.
    Case-insensitive, strips leading/trailing whitespace.
    DB-sourced rates always take priority — only call this when rate is None.
    """
    return HARDCODED_MATERIAL_RATES.get((material_name or "").strip().lower())


PROCUREMENT_ALLOWED_ROLES = {"admin"}
PROCUREMENT_ALLOWED_UIDS  = {"sonal"}


# ─────────────────────────────────────────────────────────────────────────────
# ACCESS
# ─────────────────────────────────────────────────────────────────────────────

def _can_procurement() -> bool:
    if not session.get("logged_in"):
        return False
    role = (session.get("User_Type") or "").strip().lower()
    uid  = (session.get("UID")       or "").strip().lower()
    return role in PROCUREMENT_ALLOWED_ROLES or uid in PROCUREMENT_ALLOWED_UIDS


def procurement_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        if not _can_procurement():
            return (
                """<!DOCTYPE html><html><head><title>Access Denied</title>
<style>body{font-family:'Segoe UI',sans-serif;background:#f8fafc;display:flex;
align-items:center;justify-content:center;height:100vh;margin:0}
.box{background:#fff;border-radius:16px;padding:48px 56px;text-align:center;
box-shadow:0 8px 32px rgba(0,0,0,.1);max-width:420px}
h1{color:#e11d48;font-size:1.8rem;margin-bottom:12px}p{color:#64748b;line-height:1.6}
a{color:#0d9488;text-decoration:none;font-weight:700}</style></head>
<body><div class="box"><h1>🔒 Access Denied</h1>
<p>You do not have permission to view the <strong>Procurement</strong> page.</p>
<p style="margin-top:24px"><a href="/">&#8592; Back to Portal</a></p>
</div></body></html>""",
                403,
            )
        return f(*args, **kwargs)
    return wrapper


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL READER  —  StkSum.xlsx
# ─────────────────────────────────────────────────────────────────────────────

def _get_stksum_path() -> str:
    """Returns the configured StkSum.xlsx path from DB, or the default."""
    try:
        conn = sampling_portal.get_db_connection()
        if conn:
            row = conn.execute(
                "SELECT setting_value FROM procurement_settings WHERE setting_key='stksum_path'"
            ).fetchone()
            conn.close()
            if row and row['setting_value']:
                return row['setting_value'].strip()
    except Exception:
        pass
    return DEFAULT_STKSUM_PATH


def _read_stksum() -> dict[str, float]:
    """
    Reads StkSum.xlsx from the configured path (or default).
    Returns { material_name_lowercase: qty } for fast lookup.

    StkSum.xlsx structure:
      - Row 14: Group header rows (e.g. 'Acids', 'Wax & Butter') — skipped
      - Row 15 onwards: Material rows — Col A = name, Col C = closing stock qty
      - The FIRST section (rows 15 up to the first 'Grand Total') is the
        combined total across all godowns — this is what we want.
      - After the first 'Grand Total' the file repeats per-godown — we stop
        there to avoid overwriting correct totals with partial godown values.
    """
    stock: dict[str, float] = {}
    STKSUM_PATH = _get_stksum_path()

    if not OPENPYXL_OK:
        print("[Procurement] ⚠️  openpyxl not installed — In Stock Qty will be 0 for all.")
        return stock

    if not os.path.exists(STKSUM_PATH):
        print(f"[Procurement] ⚠️  StkSum.xlsx not found at {STKSUM_PATH!r}")
        return stock

    try:
        wb = load_workbook(STKSUM_PATH, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=15, values_only=True):
            if not row:
                continue
            mat_name = row[0]   # Column A = Material Name
            qty_val  = row[2]   # Column C = Current Stock
            if not mat_name:
                continue
            name_str = str(mat_name).strip()
            if not name_str:
                continue
            # Stop at Grand Total — everything after is per-godown detail
            if name_str == 'Grand Total':
                break
            key = name_str.lower()
            try:
                stock[key] = float(qty_val) if qty_val is not None else 0.0
            except (TypeError, ValueError):
                stock[key] = 0.0
        wb.close()
    except Exception:
        traceback.print_exc()

    return stock


# ─────────────────────────────────────────────────────────────────────────────
# DB INIT
# ─────────────────────────────────────────────────────────────────────────────

def _init_procurement_table():
    """
    Creates procurement_materials table.
    NOTE: in_stock_qty is NOT stored — it is always sourced live from StkSum.xlsx.
    """
    conn = sampling_portal.get_db_connection()
    if not conn:
        print("[Procurement] ⚠️  DB connection failed — table init skipped.")
        return
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_materials (
                id                  INT AUTO_INCREMENT PRIMARY KEY,
                material_name       VARCHAR(500)   NOT NULL UNIQUE,
                ordered_qty         DECIMAL(15,3)  DEFAULT NULL,
                buffer_qty          DECIMAL(15,3)  DEFAULT NULL,
                supplier_name       VARCHAR(500)   DEFAULT NULL,
                last_purchase_rate  DECIMAL(15,4)  DEFAULT NULL,
                std_pack_size       VARCHAR(100)   DEFAULT NULL,
                msl                 DECIMAL(15,3)  DEFAULT NULL,
                lead_time_days      INT            DEFAULT NULL,
                hsn_code            VARCHAR(50)    DEFAULT NULL,
                gst_rate            DECIMAL(5,2)   DEFAULT NULL,
                taxability          VARCHAR(50)    DEFAULT NULL,
                type_of_supply      VARCHAR(50)    DEFAULT NULL,
                updated_at          DATETIME       DEFAULT CURRENT_TIMESTAMP
                                    ON UPDATE CURRENT_TIMESTAMP,
                updated_by          VARCHAR(200)   DEFAULT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()

        # Migration: drop in_stock_qty if it exists from an older version
        try:
            conn.execute("ALTER TABLE procurement_materials DROP COLUMN in_stock_qty")
            conn.commit()
            print("[Procurement] ℹ️  Dropped in_stock_qty column (now sourced from StkSum.xlsx).")
        except Exception:
            pass   # Column didn't exist — fine

        # Migration: add GST columns if not present
        for _col, _def in [
            ("hsn_code",       "VARCHAR(50) DEFAULT NULL"),
            ("gst_rate",       "DECIMAL(5,2) DEFAULT NULL"),
            ("taxability",     "VARCHAR(50) DEFAULT NULL"),
            ("type_of_supply", "VARCHAR(50) DEFAULT NULL"),
            ("aliases",        "TEXT DEFAULT NULL"),
            ("description",    "TEXT DEFAULT NULL"),
            ("supplier_code",  "VARCHAR(50) DEFAULT NULL AFTER supplier_name"),
            ("group_id",       "INT DEFAULT NULL"),
            ("uom",            "VARCHAR(20) DEFAULT 'KG'"),
            ("material_type_id", "INT DEFAULT NULL"),
        ]:
            try:
                conn.execute(f"ALTER TABLE procurement_materials ADD COLUMN {_col} {_def}")
                conn.commit()
            except Exception:
                pass

        # ── Material Groups table ────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_material_groups (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                group_name  VARCHAR(300) NOT NULL UNIQUE,
                parent_id   INT          DEFAULT NULL,   -- NULL = top-level (Father Group)
                description TEXT         DEFAULT NULL,
                created_at  DATETIME     DEFAULT CURRENT_TIMESTAMP,
                updated_at  DATETIME     DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                created_by  VARCHAR(200) DEFAULT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        # Ensure Father Group (root) exists with id=1
        fg = conn.execute(
            "SELECT id FROM procurement_material_groups WHERE group_name='Father Group' LIMIT 1"
        ).fetchone()
        if not fg:
            conn.execute(
                "INSERT INTO procurement_material_groups (id, group_name, parent_id, description) "
                "VALUES (1, 'Father Group', NULL, 'Root group — materials must be assigned to a sub-group')"
            )
            conn.commit()

        # ── Material Types table ────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_material_types (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                type_name   VARCHAR(100) NOT NULL UNIQUE,
                abbreviation VARCHAR(10) DEFAULT NULL,
                description  TEXT        DEFAULT NULL,
                color        VARCHAR(20) DEFAULT NULL,
                sort_order   INT         DEFAULT 0,
                created_at   DATETIME    DEFAULT CURRENT_TIMESTAMP,
                created_by   VARCHAR(200) DEFAULT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        # Seed default types if empty
        existing_types = conn.execute("SELECT COUNT(*) AS n FROM procurement_material_types").fetchone()
        if existing_types and existing_types["n"] == 0:
            for _name, _abbr, _color in [
                ("Raw Material",      "RM",  "#2563eb"),
                ("Packing Material",  "PM",  "#7c3aed"),
                ("Finished Goods",    "FG",  "#16a34a"),
                ("Other",             "OT",  "#6b7280"),
            ]:
                conn.execute(
                    "INSERT INTO procurement_material_types (type_name, abbreviation, color, sort_order) VALUES (%s,%s,%s,%s)",
                    (_name, _abbr, _color, {"RM":1,"PM":2,"FG":3,"OT":4}[_abbr])
                )
            conn.commit()

        # Settings table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_settings (
                setting_key   VARCHAR(100) NOT NULL PRIMARY KEY,
                setting_value TEXT         DEFAULT NULL,
                updated_at    DATETIME     DEFAULT CURRENT_TIMESTAMP
                              ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        # Formulations vs Qty table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_formulations (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                batch_name      VARCHAR(700)   NOT NULL,
                product_code    VARCHAR(200)   DEFAULT NULL,
                material_name   VARCHAR(500)   NOT NULL,
                supplier_name   VARCHAR(500)   DEFAULT NULL,
                concentration   VARCHAR(200)   DEFAULT NULL,
                qty_kg          VARCHAR(100)   DEFAULT NULL,
                batch_size      VARCHAR(100)   DEFAULT NULL,
                batch_date      VARCHAR(100)   DEFAULT NULL,
                num_batches     VARCHAR(50)    DEFAULT NULL,
                imported_at     DATETIME       DEFAULT CURRENT_TIMESTAMP,
                imported_by     VARCHAR(200)   DEFAULT NULL,
                source_batch_name VARCHAR(700)   DEFAULT NULL,
                UNIQUE KEY uq_batch_mat (batch_name(350), material_name(300))
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        # Migration: add new columns if upgrading older schema
        for _col, _def in [
            ("product_code",       "VARCHAR(200) DEFAULT NULL AFTER batch_name"),
            ("supplier_name",      "VARCHAR(500) DEFAULT NULL AFTER material_name"),
            ("qty_kg",             "VARCHAR(100) DEFAULT NULL AFTER concentration"),
            ("batch_size",         "VARCHAR(100) DEFAULT NULL AFTER qty_kg"),
            ("batch_date",         "VARCHAR(100) DEFAULT NULL AFTER batch_size"),
            ("num_batches",        "VARCHAR(50)  DEFAULT NULL AFTER batch_date"),
            ("source_batch_name",  "VARCHAR(700) DEFAULT NULL"),
            ("manuf_process",       "TEXT DEFAULT NULL"),
            ("is_active",           "TINYINT(1) NOT NULL DEFAULT 1"),
        ]:
            try:
                conn.execute(f"ALTER TABLE procurement_formulations ADD COLUMN {_col} {_def}")
                conn.commit()
            except Exception:
                pass
        # Update log table — records every Add/Deduct procurement action
        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_update_log (
                id            INT AUTO_INCREMENT PRIMARY KEY,
                batch_name    VARCHAR(700) NOT NULL,
                action_type   VARCHAR(20)  NOT NULL,   -- 'add' or 'deduct'
                qty_changed   DECIMAL(15,3) NOT NULL,
                size_before   DECIMAL(15,3) DEFAULT NULL,
                size_after    DECIMAL(15,3) DEFAULT NULL,
                action_by     VARCHAR(200)  DEFAULT NULL,
                action_at     DATETIME      DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        # Formulation change log
        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_formulation_changelog (
                id           INT AUTO_INCREMENT PRIMARY KEY,
                batch_name   VARCHAR(700)  NOT NULL,
                changed_by   VARCHAR(200)  DEFAULT NULL,
                changed_at   DATETIME      DEFAULT CURRENT_TIMESTAMP,
                change_type  VARCHAR(50)   NOT NULL,  -- 'update','import','link'
                ingredients_before  TEXT  DEFAULT NULL,  -- JSON snapshot
                ingredients_after   TEXT  DEFAULT NULL   -- JSON snapshot
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()

        # Supplier master table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_suppliers (
                id               INT AUTO_INCREMENT PRIMARY KEY,

                -- Basic Info
                supplier_name    VARCHAR(500)  NOT NULL UNIQUE,
                supplier_code    VARCHAR(50)   DEFAULT NULL,
                contact_person   VARCHAR(300)  DEFAULT NULL,
                phone            VARCHAR(100)  DEFAULT NULL,
                email            VARCHAR(300)  DEFAULT NULL,
                address          TEXT          DEFAULT NULL,

                -- Business Info
                gst_number       VARCHAR(50)   DEFAULT NULL,
                pan_number       VARCHAR(50)   DEFAULT NULL,
                payment_terms    VARCHAR(200)  DEFAULT NULL,
                currency         VARCHAR(10)   DEFAULT 'INR',

                -- Operational Info
                lead_time_days   INT           DEFAULT NULL,
                moq              DECIMAL(15,3) DEFAULT NULL,
                rating           TINYINT       DEFAULT NULL,   -- 1-5
                status           VARCHAR(20)   DEFAULT 'active',  -- active / inactive

                -- Meta
                created_at       DATETIME      DEFAULT CURRENT_TIMESTAMP,
                updated_at       DATETIME      DEFAULT CURRENT_TIMESTAMP
                                 ON UPDATE CURRENT_TIMESTAMP,
                updated_by       VARCHAR(200)  DEFAULT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()

        # Migration: add columns if upgrading older schema
        for _col, _def in [
            ("supplier_code",  "VARCHAR(50)  DEFAULT NULL AFTER supplier_name"),
            ("contact_person", "VARCHAR(300) DEFAULT NULL"),
            ("phone",          "VARCHAR(100) DEFAULT NULL"),
            ("email",          "VARCHAR(300) DEFAULT NULL"),
            ("address",        "TEXT         DEFAULT NULL"),
            ("gst_number",     "VARCHAR(50)  DEFAULT NULL"),
            ("pan_number",     "VARCHAR(50)  DEFAULT NULL"),
            ("payment_terms",  "VARCHAR(200) DEFAULT NULL"),
            ("currency",       "VARCHAR(10)  DEFAULT 'INR'"),
            ("lead_time_days", "INT          DEFAULT NULL"),
            ("moq",            "DECIMAL(15,3) DEFAULT NULL"),
            ("rating",         "TINYINT      DEFAULT NULL"),
            ("status",         "VARCHAR(20)  DEFAULT 'active'"),
            ("updated_by",     "VARCHAR(200) DEFAULT NULL"),
            ("payment_type",    "VARCHAR(100) DEFAULT NULL"),
            ("credit_days",     "INT          DEFAULT NULL"),
            ("declaration_id",  "INT          DEFAULT NULL"),
            ("supplier_type_id","INT          DEFAULT NULL"),
        ]:
            try:
                conn.execute(f"ALTER TABLE procurement_suppliers ADD COLUMN {_col} {_def}")
                conn.commit()
            except Exception:
                pass

        # ── Supplier Types (shared, no prefix) ───────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS supplier_type (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                type_name   VARCHAR(200) NOT NULL UNIQUE,
                sort_order  INT          DEFAULT 0,
                created_at  DATETIME     DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()

        # ── Purchase Orders ──────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_purchase_orders (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                po_num          VARCHAR(100) NOT NULL,
                po_date         DATE         DEFAULT NULL,
                supplier_name   VARCHAR(500) DEFAULT NULL,
                status          VARCHAR(30)  DEFAULT 'open',
                delivery_date   DATE         DEFAULT NULL,
                delivery_days   INT          DEFAULT NULL,
                remarks         TEXT         DEFAULT NULL,
                grand_total     DECIMAL(18,4) DEFAULT NULL,
                tc_list_id      INT          DEFAULT NULL,
                declaration_id  INT          DEFAULT NULL,
                created_by      VARCHAR(200) DEFAULT NULL,
                updated_by      VARCHAR(200) DEFAULT NULL,
                created_at      DATETIME     DEFAULT CURRENT_TIMESTAMP,
                updated_at      DATETIME     DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_po_num (po_num)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_po_items (
                id           INT AUTO_INCREMENT PRIMARY KEY,
                po_id        INT           NOT NULL,
                material     VARCHAR(500)  NOT NULL,
                qty          DECIMAL(15,3) DEFAULT NULL,
                rate         DECIMAL(15,4) DEFAULT NULL,
                amount       DECIMAL(18,4) DEFAULT NULL,
                hsn_code     VARCHAR(50)   DEFAULT NULL,
                gst_rate     DECIMAL(5,2)  DEFAULT NULL,
                cgst_amount  DECIMAL(18,4) DEFAULT NULL,
                sgst_amount  DECIMAL(18,4) DEFAULT NULL,
                packages     INT           DEFAULT NULL,
                qty_per_pkg  DECIMAL(15,3) DEFAULT NULL,
                uom          VARCHAR(50)   DEFAULT NULL,
                FOREIGN KEY (po_id) REFERENCES procurement_purchase_orders(id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        # Migration: add new columns to purchase_orders if missing
        for _col, _def in [
            ("delivery_days",   "INT DEFAULT NULL"),
            ("tc_list_id",      "INT DEFAULT NULL"),
            ("declaration_id",    "INT DEFAULT NULL"),
            ("freight_charge",    "DECIMAL(15,4) DEFAULT NULL"),
            ("packing_charge",    "DECIMAL(15,4) DEFAULT NULL"),
            ("voucher_type_name", "VARCHAR(200) DEFAULT NULL"),  # links to gop_voucher_type_masters.name
        ]:
            try:
                conn.execute(f"ALTER TABLE procurement_purchase_orders ADD COLUMN {_col} {_def}")
                conn.commit()
            except Exception:
                pass

        # Migration: add GST columns to po_items if missing
        for _col, _def in [
            ("hsn_code",    "VARCHAR(50) DEFAULT NULL"),
            ("gst_rate",    "DECIMAL(5,2) DEFAULT NULL"),
            ("cgst_amount", "DECIMAL(18,4) DEFAULT NULL"),
            ("sgst_amount", "DECIMAL(18,4) DEFAULT NULL"),
            ("packages",    "INT DEFAULT NULL"),
            ("qty_per_pkg", "DECIMAL(15,3) DEFAULT NULL"),
            ("uom",         "VARCHAR(50) DEFAULT NULL"),
        ]:
            try:
                conn.execute(f"ALTER TABLE procurement_po_items ADD COLUMN {_col} {_def}")
                conn.commit()
            except Exception:
                pass

        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_declarations (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                name        VARCHAR(200) NOT NULL,
                text        TEXT         DEFAULT NULL,
                created_at  DATETIME     DEFAULT CURRENT_TIMESTAMP,
                updated_at  DATETIME     DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_po_settings (
                id       INT AUTO_INCREMENT PRIMARY KEY,
                key_name VARCHAR(100) NOT NULL UNIQUE,
                val      TEXT         DEFAULT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()

        # ── Godowns & Addresses ───────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_godowns (
                id           INT AUTO_INCREMENT PRIMARY KEY,
                name         VARCHAR(300) NOT NULL,
                address      TEXT         DEFAULT NULL,
                contact      VARCHAR(300) DEFAULT NULL,
                phone        VARCHAR(100) DEFAULT NULL,
                email        VARCHAR(300) DEFAULT NULL,
                is_default   TINYINT(1)   DEFAULT 0,
                type         VARCHAR(20)  DEFAULT 'godown',
                gst_number   VARCHAR(50)  DEFAULT NULL,
                created_at   DATETIME     DEFAULT CURRENT_TIMESTAMP,
                updated_at   DATETIME     DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # Migration: add state, city, pin columns to godowns
        for _gcol, _gdef in [
            ("state", "VARCHAR(100) DEFAULT NULL"),
            ("city",  "VARCHAR(100) DEFAULT NULL"),
            ("pin",   "VARCHAR(10)  DEFAULT NULL"),
        ]:
            try:
                conn.execute(f"ALTER TABLE procurement_godowns ADD COLUMN {_gcol} {_gdef}")
                conn.commit()
            except Exception:
                pass
        # Company billing address — stored as a single row in procurement_settings
        conn.commit()

        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_tc_lists (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                name            VARCHAR(300) NOT NULL,
                delivery_days   INT          DEFAULT NULL,
                delivery_mode   VARCHAR(100) DEFAULT NULL,
                delivery_notes  TEXT         DEFAULT NULL,
                payment_type    VARCHAR(100) DEFAULT NULL,
                credit_days     INT          DEFAULT NULL,
                payment_notes   TEXT         DEFAULT NULL,
                other_terms     TEXT         DEFAULT NULL,
                created_by      VARCHAR(200) DEFAULT NULL,
                created_at      DATETIME     DEFAULT CURRENT_TIMESTAMP,
                updated_at      DATETIME     DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()

        # ── Brands table ──────────────────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS procurement_brands (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                name        VARCHAR(300) NOT NULL UNIQUE,
                color       VARCHAR(20)  DEFAULT '#6366f1',
                text_color  VARCHAR(20)  DEFAULT '#ffffff',
                created_at  DATETIME     DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()

        # ── text_color column on brands table (migration for existing DBs) ──────
        try:
            conn.execute("ALTER TABLE procurement_brands ADD COLUMN text_color VARCHAR(20) DEFAULT '#ffffff'")
            conn.commit()
        except Exception:
            pass

        # ── brand_id column on formulations (one brand per batch) ─────────────
        for _col, _def in [
            ("brand_id", "INT DEFAULT NULL"),
        ]:
            try:
                conn.execute(f"ALTER TABLE procurement_formulations ADD COLUMN {_col} {_def}")
                conn.commit()
            except Exception:
                pass

        # Migrate existing 'draft' POs → 'open' (draft status removed)
        try:
            conn.execute("UPDATE procurement_purchase_orders SET status='open' WHERE status='draft'")
            conn.commit()
        except Exception:
            pass
        print("✅ all procurement tables ready")

        # ── GRN Tables — use fresh connection to avoid transaction contamination ──
        try:
            conn.close()
        except Exception:
            pass

        grn_conn = sampling_portal.get_db_connection()
        if grn_conn:
            try:
                grn_conn.execute("""
                    CREATE TABLE IF NOT EXISTS procurement_grn (
                        id              INT AUTO_INCREMENT PRIMARY KEY,
                        grn_num         VARCHAR(100) NOT NULL UNIQUE,
                        grn_date        DATE         DEFAULT NULL,
                        po_id           INT          DEFAULT NULL,
                        po_num          VARCHAR(100) DEFAULT NULL,
                        supplier_name   VARCHAR(500) DEFAULT NULL,
                        invoice_num     VARCHAR(200) DEFAULT NULL,
                        invoice_date    DATE         DEFAULT NULL,
                        status          VARCHAR(30)  DEFAULT 'open',
                        freight_charge  DECIMAL(15,4) DEFAULT NULL,
                        packing_charge  DECIMAL(15,4) DEFAULT NULL,
                        grand_total     DECIMAL(18,4) DEFAULT NULL,
                        remarks         TEXT         DEFAULT NULL,
                        created_by      VARCHAR(200) DEFAULT NULL,
                        updated_by      VARCHAR(200) DEFAULT NULL,
                        created_at      DATETIME     DEFAULT CURRENT_TIMESTAMP,
                        updated_at      DATETIME     DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                grn_conn.execute("""
                    CREATE TABLE IF NOT EXISTS procurement_grn_items (
                        id           INT AUTO_INCREMENT PRIMARY KEY,
                        grn_id       INT           NOT NULL,
                        material     VARCHAR(500)  NOT NULL,
                        po_qty       DECIMAL(15,3) DEFAULT NULL,
                        received_qty DECIMAL(15,3) DEFAULT NULL,
                        rate         DECIMAL(15,4) DEFAULT NULL,
                        amount       DECIMAL(18,4) DEFAULT NULL,
                        hsn_code     VARCHAR(50)   DEFAULT NULL,
                        gst_rate     DECIMAL(5,2)  DEFAULT NULL,
                        packages     INT           DEFAULT NULL,
                        qty_per_pkg  DECIMAL(15,3) DEFAULT NULL,
                        uom          VARCHAR(50)   DEFAULT NULL,
                        FOREIGN KEY (grn_id) REFERENCES procurement_grn(id) ON DELETE CASCADE
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                grn_conn.commit()
                # Add po_id FK separately — ignore if already exists
                try:
                    grn_conn.execute("""
                        ALTER TABLE procurement_grn
                        ADD CONSTRAINT fk_grn_po FOREIGN KEY (po_id)
                        REFERENCES procurement_purchase_orders(id) ON DELETE SET NULL
                    """)
                    grn_conn.commit()
                except Exception:
                    pass
                # Migration: add packages + uom columns to grn_items
                for _col, _def in [("packages","INT DEFAULT NULL"),("qty_per_pkg","DECIMAL(15,3) DEFAULT NULL"),("uom","VARCHAR(50) DEFAULT NULL")]:
                    try:
                        grn_conn.execute(f"ALTER TABLE procurement_grn_items ADD COLUMN {_col} {_def}")
                        grn_conn.commit()
                    except Exception:
                        pass
                # Migration: add po_invoices column for multi-PO invoice tracking
                try:
                    grn_conn.execute("ALTER TABLE procurement_grn ADD COLUMN po_invoices TEXT DEFAULT NULL")
                    grn_conn.commit()
                except Exception:
                    pass  # column already exists
                # Migration: add unload_location column
                try:
                    grn_conn.execute("ALTER TABLE procurement_grn ADD COLUMN unload_location VARCHAR(300) DEFAULT NULL")
                    grn_conn.commit()
                except Exception:
                    pass
                # Migration: add voucher_type_name for GOP voucher type support
                try:
                    grn_conn.execute("ALTER TABLE procurement_grn ADD COLUMN voucher_type_name VARCHAR(200) DEFAULT NULL")
                    grn_conn.commit()
                except Exception:
                    pass
                # Migration: add per-item invoice, batch, mfg/expiry date, location columns
                for _col, _def in [
                    ("location",     "VARCHAR(300) DEFAULT NULL"),
                    ("invoice_num",  "VARCHAR(200) DEFAULT NULL"),
                    ("invoice_date", "DATE DEFAULT NULL"),
                    ("batch_num",    "VARCHAR(100) DEFAULT NULL"),
                    ("mfg_date",     "DATE DEFAULT NULL"),
                    ("expiry_date",  "DATE DEFAULT NULL"),
                ]:
                    try:
                        grn_conn.execute(f"ALTER TABLE procurement_grn_items ADD COLUMN {_col} {_def}")
                        grn_conn.commit()
                    except Exception:
                        pass
                print("✅ GRN tables ready")

                # Cleanup: drop orphan duplicate tables if they exist
                for _orphan in ("procurement_grn_lines",):
                    try:
                        _exists = grn_conn.execute(
                            "SELECT COUNT(*) AS n FROM information_schema.tables WHERE table_schema=DATABASE() AND table_name=%s",
                            (_orphan,)
                        ).fetchone()
                        if _exists and _exists["n"] > 0:
                            grn_conn.execute(f"DROP TABLE {_orphan}")
                            grn_conn.commit()
                            print(f"  🗑 Dropped orphan table: {_orphan}")
                    except Exception:
                        pass
            except Exception as e:
                print(f"[Procurement] ⚠ GRN table init (non-fatal): {e}")
                try: grn_conn.rollback()
                except Exception: pass
            finally:
                try: grn_conn.close()
                except Exception: pass

        # ── MTV Tables ────────────────────────────────────────────────
        try:
            mtv_conn = sampling_portal.get_db_connection()
            mtv_conn.execute("""
                CREATE TABLE IF NOT EXISTS procurement_mtv (
                    id          INT AUTO_INCREMENT PRIMARY KEY,
                    mtv_num     VARCHAR(100) NOT NULL,
                    mtv_date    DATE         NOT NULL,
                    from_loc    VARCHAR(300) DEFAULT NULL,
                    to_loc      VARCHAR(300) DEFAULT NULL,
                    status      VARCHAR(30)  DEFAULT 'open',
                    remarks     TEXT         DEFAULT NULL,
                    voucher_type_name VARCHAR(200) DEFAULT NULL,
                    created_by  VARCHAR(200) DEFAULT NULL,
                    updated_by  VARCHAR(200) DEFAULT NULL,
                    created_at  DATETIME     DEFAULT CURRENT_TIMESTAMP,
                    updated_at  DATETIME     DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            mtv_conn.execute("""
                CREATE TABLE IF NOT EXISTS procurement_mtv_items (
                    id            INT AUTO_INCREMENT PRIMARY KEY,
                    mtv_id        INT           NOT NULL,
                    material_id   INT           DEFAULT NULL,
                    material_name VARCHAR(500)  NOT NULL,
                    qty           DECIMAL(15,3) DEFAULT NULL,
                    qty_per_pkg   DECIMAL(15,3) DEFAULT NULL,
                    uom           VARCHAR(50)   DEFAULT 'kg',
                    packages      INT           DEFAULT NULL,
                    remarks       TEXT          DEFAULT NULL,
                    FOREIGN KEY (mtv_id) REFERENCES procurement_mtv(id) ON DELETE CASCADE
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            mtv_conn.commit()
            # Migration: add voucher_type_name if missing
            try:
                mtv_conn.execute("ALTER TABLE procurement_mtv ADD COLUMN voucher_type_name VARCHAR(200) DEFAULT NULL")
                mtv_conn.commit()
            except Exception:
                pass
            try:
                mtv_conn.execute("ALTER TABLE procurement_mtv_items ADD COLUMN packages INT DEFAULT NULL")
                mtv_conn.commit()
            except Exception:
                pass
            try:
                mtv_conn.execute("ALTER TABLE procurement_mtv_items ADD COLUMN qty_per_pkg DECIMAL(15,3) DEFAULT NULL")
                mtv_conn.commit()
            except Exception:
                pass
            print("✅ MTV tables ready")
        except Exception as e:
            print(f"[Procurement] ⚠ MTV table init (non-fatal): {e}")
        finally:
            try: mtv_conn.close()
            except Exception: pass

        # ── Voucher Numbering Styles ──────────────────────────────────
        try:
            vn_conn = sampling_portal.get_db_connection()
            vn_conn.execute("""
                CREATE TABLE IF NOT EXISTS procurement_voucher_numbering (
                    id            INT AUTO_INCREMENT PRIMARY KEY,
                    voucher_type  VARCHAR(50)  NOT NULL DEFAULT 'po',
                    prefix        VARCHAR(100) DEFAULT '',
                    suffix        VARCHAR(100) DEFAULT '',
                    digits        INT          DEFAULT 4,
                    start_num     INT          DEFAULT 1,
                    valid_from    DATE         NOT NULL,
                    valid_to      DATE         NOT NULL,
                    created_at    DATETIME     DEFAULT CURRENT_TIMESTAMP,
                    updated_at    DATETIME     DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            vn_conn.commit()
            vn_conn.close()
            print("✅ Voucher numbering table ready")
        except Exception as e:
            print(f"[Procurement] ⚠ Voucher numbering table init (non-fatal): {e}")

    except Exception as e:
        print(f"[Procurement] ❌ Table init error: {e}")
        traceback.print_exc()
    finally:
        try:
            conn.close()
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# DB HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _get_all_materials() -> list[dict]:
    """
    Fetches all rows from DB, then merges:
      - In Stock Qty from StkSum.xlsx
      - Required Qty computed from procurement_formulations
          required_qty = SUM( concentration × batch_size )
          across all formulation rows for this material where
          both concentration and batch_size are valid numbers.
    """
    conn = sampling_portal.get_db_connection()
    if not conn:
        return []
    try:
        # 1. Fetch procurement materials
        rows = conn.execute("""
            SELECT m.*, g.group_name, mt.type_name AS material_type, mt.abbreviation AS mat_type_abbr, mt.color AS mat_type_color
            FROM procurement_materials m
            LEFT JOIN procurement_material_groups g  ON m.group_id = g.id
            LEFT JOIN procurement_material_types  mt ON m.material_type_id = mt.id
            ORDER BY m.material_name ASC
        """).fetchall()
        db_rows = [dict(r) for r in rows]

        # 2. Compute required_qty per material from formulations
        req_rows = conn.execute("""
            SELECT
                LOWER(TRIM(material_name)) AS mat_key,
                concentration,
                batch_size
            FROM procurement_formulations
            WHERE material_name  IS NOT NULL
              AND concentration   IS NOT NULL
              AND batch_size      IS NOT NULL
        """).fetchall()

        required_map = {}
        import re as _re
        for rr in req_rows:
            mk = (rr["mat_key"] or "").strip().lower()
            if not mk:
                continue
            try:
                conc_f = float(rr["concentration"])
            except (TypeError, ValueError):
                continue
            bs_str   = str(rr["batch_size"]).strip()
            bs_match = _re.search(r"[\d.]+", bs_str)
            if not bs_match:
                continue
            try:
                bs_f = float(bs_match.group())
            except (TypeError, ValueError):
                continue
            required_map[mk] = required_map.get(mk, 0.0) + conc_f * bs_f

        # 2b. Fetch last used rate from po_items for materials with no last_purchase_rate
        try:
            po_rate_rows = conn.execute("""
                SELECT LOWER(TRIM(pi.material)) AS mat_key, pi.rate
                FROM procurement_po_items pi
                INNER JOIN (
                    SELECT material, MAX(id) AS max_id
                    FROM procurement_po_items
                    WHERE rate IS NOT NULL AND rate > 0
                    GROUP BY material
                ) latest ON pi.material = latest.material AND pi.id = latest.max_id
            """).fetchall()
            po_rate_map = {r["mat_key"]: float(r["rate"]) for r in po_rate_rows if r["rate"]}
        except Exception:
            po_rate_map = {}

    except Exception:
        traceback.print_exc()
        return []
    finally:
        try:
            conn.close()
        except Exception:
            pass

    # 3. Read live stock quantities from StkSum.xlsx
    stock_map = _read_stksum()   # { lowercase_name: qty }

    # 4. Merge and assign sr_no
    result = []
    for i, r in enumerate(db_rows, start=1):
        key = r["material_name"].strip().lower()
        r["in_stock_qty"]  = stock_map.get(key, 0.0)
        raw_req            = required_map.get(key, 0.0)
        r["required_qty"]  = round(raw_req, 3) if raw_req > 0 else None
        r["sr_no"]         = i
        # Fill last_purchase_rate from PO history if not set in material master
        if not r.get("last_purchase_rate") and key in po_rate_map:
            r["last_purchase_rate"] = po_rate_map[key]
        result.append(r)

    return result


def _upsert_material(mat: str, fields: dict, uid: str) -> None:
    """
    INSERT ... ON DUPLICATE KEY UPDATE for a single material.
    in_stock_qty is NEVER written to the DB — it comes from StkSum.xlsx.
    """
    conn = sampling_portal.get_db_connection()
    if not conn:
        raise RuntimeError("DB connection failed")
    # Read old rate before overwriting (for cost-impact detection)
    old_rate_row = conn.execute(
        "SELECT last_purchase_rate FROM procurement_materials WHERE material_name=%s",
        (mat,)
    ).fetchone()
    old_rate = float(old_rate_row["last_purchase_rate"]) if old_rate_row and old_rate_row["last_purchase_rate"] is not None else None

    def _v(key):
        val = fields.get(key)
        if val is None or str(val).strip() in ("", "None"):
            return None
        try:
            if key in ("ordered_qty", "buffer_qty", "last_purchase_rate", "msl"):
                return float(str(val).replace(",", ""))
            if key in ("lead_time_days",):
                return int(float(str(val).replace(",", "")))
            if key == "gst_rate":
                return float(str(val).replace(",", ""))
            return str(val).strip()
        except (ValueError, TypeError):
            return None

    try:
        conn.execute("""
            INSERT INTO procurement_materials
                (material_name, ordered_qty, buffer_qty,
                 supplier_name, supplier_code, last_purchase_rate, std_pack_size,
                 msl, lead_time_days, hsn_code, gst_rate, taxability, type_of_supply,
                 aliases, description, group_id, uom, material_type_id, updated_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
                ordered_qty         = VALUES(ordered_qty),
                buffer_qty          = VALUES(buffer_qty),
                supplier_name       = VALUES(supplier_name),
                supplier_code       = VALUES(supplier_code),
                last_purchase_rate  = VALUES(last_purchase_rate),
                std_pack_size       = VALUES(std_pack_size),
                msl                 = VALUES(msl),
                lead_time_days      = VALUES(lead_time_days),
                hsn_code            = VALUES(hsn_code),
                gst_rate            = VALUES(gst_rate),
                taxability          = VALUES(taxability),
                type_of_supply      = VALUES(type_of_supply),
                aliases             = VALUES(aliases),
                description         = VALUES(description),
                group_id            = VALUES(group_id),
                uom                 = VALUES(uom),
                material_type_id    = VALUES(material_type_id),
                updated_by          = VALUES(updated_by)
        """, (
            mat,
            _v("ordered_qty"), _v("buffer_qty"),
            _v("supplier_name"), _v("supplier_code"), _v("last_purchase_rate"), _v("std_pack_size"),
            _v("msl"), _v("lead_time_days"),
            _v("hsn_code"), _v("gst_rate"), _v("taxability"), _v("type_of_supply"),
            _v("aliases"), _v("description"),
            int(fields["group_id"]) if fields.get("group_id") else None,
            str(fields.get("uom") or "KG").strip().upper() or "KG",
            int(fields["material_type_id"]) if fields.get("material_type_id") else None,
            uid,
        ))
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            pass
    # Return old_rate so caller can check cost impact
    old_rate_out = old_rate if 'old_rate' in dir() else None
    return old_rate_out


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────

def register_procurement(app):
    """Call once from app.py:  procurement.register_procurement(app)"""

    _init_procurement_table()

    # ── Main page ─────────────────────────────────────────────────────────
    @app.route("/procurement")
    @procurement_required
    def procurement_main():
        return render_template("procurement.html")

    # ── API: all materials (DB + live StkSum.xlsx merge) ──────────────────
    @app.route("/api/procurement/stock_summary")
    @procurement_required
    def api_procurement_stock_summary():
        try:
            rows = _get_all_materials()
            return jsonify({
                "status":       "ok",
                "rows":         rows,
                "count":        len(rows),
                "refreshed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    # ── API: save / add material ──────────────────────────────────────────
    @app.route("/api/procurement/save_material", methods=["POST"])
    @procurement_required
    def api_procurement_save_material():
        d   = request.get_json() or {}
        mat = (d.get("material_name") or "").strip()
        if not mat:
            return jsonify({"status": "error", "message": "material_name required"}), 400
        try:
            # in_stock_qty is intentionally excluded — comes from StkSum.xlsx
            old_rate = _upsert_material(mat, d, session.get("UID", ""))
            new_rate_val = d.get("last_purchase_rate")
            new_rate = float(str(new_rate_val).replace(",","")) if new_rate_val not in (None,"") else None
            rate_changed = (old_rate is not None and new_rate is not None and
                            abs(old_rate - new_rate) > 0.0001)
            return jsonify({
                "status":       "ok",
                "old_rate":     old_rate,
                "new_rate":     new_rate,
                "rate_changed": rate_changed
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    # ── API: import ───────────────────────────────────────────────────────
    @app.route("/api/procurement/import", methods=["POST"])
    @procurement_required
    def api_procurement_import():
        """
        Accepts { rows: [ { material_name, ordered_qty, ... } ] }
        in_stock_qty is silently ignored even if present in the payload —
        it is always sourced from StkSum.xlsx.
        """
        d    = request.get_json() or {}
        rows = d.get("rows", [])
        if not rows:
            return jsonify({"status": "error", "message": "No rows provided"}), 400

        uid      = session.get("UID", "")
        imported = 0
        skipped  = 0
        errors   = []

        for row in rows:
            mat = str(row.get("material_name") or "").strip()
            if not mat:
                skipped += 1
                continue
            try:
                _upsert_material(mat, row, uid)
                imported += 1
            except Exception as e:
                errors.append(f"{mat}: {e}")
                skipped += 1

        return jsonify({
            "status":   "ok",
            "imported": imported,
            "skipped":  skipped,
            "errors":   errors[:10],
        })

    # ── API: get settings ────────────────────────────────────────────────
    @app.route("/api/procurement/settings", methods=["GET"])
    @procurement_required
    def api_procurement_get_settings():
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            rows = conn.execute("SELECT setting_key, setting_value FROM procurement_settings").fetchall()
            conn.close()
            settings = {r["setting_key"]: r["setting_value"] for r in rows}
            # Always return the effective path (DB or default)
            settings.setdefault("stksum_path", DEFAULT_STKSUM_PATH)
            return jsonify({"status": "ok", "settings": settings})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    # ── API: save settings ────────────────────────────────────────────────
    @app.route("/api/procurement/settings", methods=["POST"])
    @procurement_required
    def api_procurement_save_settings():
        d = request.get_json() or {}
        allowed_keys = {"stksum_path"}
        conn = sampling_portal.get_db_connection()
        if not conn:
            return jsonify({"status": "error", "message": "DB connection failed"}), 500
        try:
            for key, value in d.items():
                if key not in allowed_keys:
                    continue
                value = (value or "").strip()
                conn.execute("""
                    INSERT INTO procurement_settings (setting_key, setting_value)
                    VALUES (%s, %s)
                    ON DUPLICATE KEY UPDATE setting_value = VALUES(setting_value)
                """, (key, value or None))
            conn.commit()
            return jsonify({"status": "ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try:
                conn.close()
            except Exception:
                pass

    # ── API: test StkSum path ─────────────────────────────────────────────
    @app.route("/api/procurement/test_stksum_path", methods=["POST"])
    @procurement_required
    def api_test_stksum_path():
        d    = request.get_json() or {}
        path = (d.get("path") or "").strip()
        if not path:
            path = DEFAULT_STKSUM_PATH
        exists = os.path.exists(path)
        row_count = 0
        if exists and OPENPYXL_OK:
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=15, values_only=True):
                    if row and row[0] and str(row[0]).strip():
                        if str(row[0]).strip() == 'Grand Total':
                            break
                        row_count += 1
                wb.close()
            except Exception:
                pass
        return jsonify({
            "status":    "ok",
            "path":      path,
            "exists":    exists,
            "row_count": row_count,
        })


    # ── API: update batch_size (Add / Deduct Procurement) ────────
    @app.route("/api/procurement/formulations/update_batch_size", methods=["POST"])
    @procurement_required
    def api_formulations_update_batch_size():
        """
        Body: {
          batch_name:   str,
          batch_size:   str,    e.g. "62.5 KG"
          action_type:  str,    "add" | "deduct"
          qty_changed:  float,
          size_before:  float
        }
        Updates batch_size and writes to procurement_update_log.
        """
        d           = request.get_json() or {}
        batch_name  = (d.get("batch_name")  or "").strip()
        batch_size  = (d.get("batch_size")  or "").strip() or None
        action_type = (d.get("action_type") or "").strip() or "update"
        qty_changed = d.get("qty_changed")
        size_before = d.get("size_before")
        uid         = session.get("UID", "")

        if not batch_name:
            return jsonify({"status":"error","message":"batch_name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500

            # Update all formulation rows for this batch
            conn.execute(
                "UPDATE procurement_formulations SET batch_size=%s WHERE batch_name=%s",
                (batch_size, batch_name)
            )

            # Parse numeric size_after from batch_size string
            try:
                import re as _re
                size_after_num  = float(_re.sub(r'[^\d.]', '', batch_size or '0')) if batch_size else None
                size_before_num = float(size_before) if size_before is not None else None
                qty_changed_num = float(qty_changed) if qty_changed is not None else None
            except (ValueError, TypeError):
                size_after_num = size_before_num = qty_changed_num = None

            # Write log entry
            conn.execute("""
                INSERT INTO procurement_update_log
                    (batch_name, action_type, qty_changed, size_before, size_after, action_by)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (batch_name, action_type, qty_changed_num, size_before_num, size_after_num, uid))

            conn.commit()
            conn.close()
            return jsonify({"status":"ok","batch_name":batch_name,"batch_size":batch_size})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    # ── API: get update log ───────────────────────────────────────
    @app.route("/api/procurement/formulations/update_log")
    @procurement_required
    def api_formulations_update_log():
        """Returns update log entries, optionally filtered by batch_name."""
        batch_name = request.args.get("batch_name","").strip()
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            if batch_name:
                rows = conn.execute("""
                    SELECT * FROM procurement_update_log
                    WHERE batch_name=%s ORDER BY action_at DESC
                """, (batch_name,)).fetchall()
            else:
                rows = conn.execute("""
                    SELECT * FROM procurement_update_log
                    ORDER BY action_at DESC LIMIT 500
                """).fetchall()
            conn.close()
            result = []
            for i, r in enumerate(rows, start=1):
                d2 = dict(r)
                d2["sr_no"] = i
                result.append(d2)
            return jsonify({"status":"ok","rows":result,"count":len(result)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    # ── API: bulk delete batches by name list ────────────────────
    @app.route("/api/procurement/formulations/delete_batches", methods=["POST"])
    @procurement_required
    def api_formulations_delete_batches():
        d = request.get_json() or {}
        batch_names = d.get("batch_names") or []
        if not batch_names:
            return jsonify({"status":"error","message":"batch_names required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            fmt = ','.join(['%s']*len(batch_names))
            conn.execute(f"DELETE FROM procurement_formulations WHERE batch_name IN ({fmt})", batch_names)
            conn.commit()
            conn.close()
            return jsonify({"status":"ok","deleted":len(batch_names)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    # ── API: update supplier_name for material in formulations ────
    @app.route("/api/procurement/formulations/update_supplier", methods=["POST"])
    @procurement_required
    def api_formulations_update_supplier():
        """
        Called when a material's supplier_name is updated in Tab 1.
        Updates ALL procurement_formulations rows where material_name matches.
        """
        d             = request.get_json() or {}
        material_name = (d.get("material_name") or "").strip()
        supplier_name = (d.get("supplier_name") or "").strip() or None
        if not material_name:
            return jsonify({"status":"error","message":"material_name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            result = conn.execute(
                "UPDATE procurement_formulations SET supplier_name=%s WHERE LOWER(material_name)=LOWER(%s)",
                (supplier_name, material_name)
            )
            conn.commit()
            affected = result.rowcount if hasattr(result,'rowcount') else 0
            conn.close()
            return jsonify({"status":"ok","updated_rows":affected})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    # ── API: update formulation rows (inline edit or re-import) ──
    @app.route("/api/procurement/formulations/update_rows", methods=["POST"])
    @procurement_required
    def api_formulations_update_rows():
        """
        Body: {
          batch_name: str,
          rows: [{ material_name, supplier_name, concentration, qty_kg }, ...]
        }
        Replaces ALL ingredient rows for the given batch_name with new values.
        Preserves batch_size, batch_date, num_batches, product_code from existing rows.
        """
        d          = request.get_json() or {}
        batch_name = (d.get("batch_name") or "").strip()
        new_rows   = d.get("rows") or []
        if not batch_name:
            return jsonify({"status":"error","message":"batch_name required"}), 400
        if not new_rows:
            return jsonify({"status":"error","message":"rows required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            uid = session.get("UID","")
            # Get existing meta (batch_size, batch_date, num_batches, product_code)
            existing = conn.execute(
                """SELECT batch_size, batch_date, num_batches, product_code
                   FROM procurement_formulations WHERE batch_name=%s LIMIT 1""",
                (batch_name,)
            ).fetchone()
            meta = dict(existing) if existing else {}
            # Delete all existing rows for this batch
            conn.execute(
                "DELETE FROM procurement_formulations WHERE batch_name=%s",
                (batch_name,)
            )
            # Re-insert with new values
            imported = 0
            for r in new_rows:
                mat = (r.get("material_name") or "").strip()
                if not mat:
                    continue
                sup  = (r.get("supplier_name") or "").strip() or None
                conc = r.get("concentration")
                qty  = r.get("qty_kg")
                # Normalize numeric strings
                def _nf(v):
                    if v is None or str(v).strip() == "": return None
                    try: return f"{float(str(v).replace(',','')):.6f}".rstrip('0').rstrip('.')
                    except: return str(v).strip()
                conn.execute("""
                    INSERT INTO procurement_formulations
                        (batch_name, product_code, material_name, supplier_name,
                         concentration, qty_kg,
                         batch_size, batch_date, num_batches, imported_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    batch_name,
                    meta.get("product_code"),
                    mat, sup,
                    _nf(conc), _nf(qty),
                    meta.get("batch_size"),
                    meta.get("batch_date"),
                    meta.get("num_batches"),
                    uid
                ))
                imported += 1
            conn.commit()
            # Check how many batches are linked to this one
            linked_rows = conn.execute("""
                SELECT DISTINCT batch_name FROM procurement_formulations
                WHERE source_batch_name=%s
            """, (batch_name,)).fetchall()
            linked_names = [r["batch_name"] for r in linked_rows]
            conn.close()
            return jsonify({
                "status":       "ok",
                "imported":     imported,
                "batch_name":   batch_name,
                "linked_batches": linked_names   # frontend shows propagation dialog if non-empty
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    # ── API: link a batch to a source formulation ────────────────
    @app.route("/api/procurement/formulations/link_batch", methods=["POST"])
    @procurement_required
    def api_formulations_link_batch():
        """
        Body: { new_batch_name, source_batch_name, batch_size (optional) }
        Creates a new batch by copying all ingredient rows from source,
        setting source_batch_name so we know they are linked.
        """
        d              = request.get_json() or {}
        new_name       = (d.get("new_batch_name")    or "").strip()
        source_name    = (d.get("source_batch_name") or "").strip()
        new_batch_size = (d.get("batch_size")        or "").strip() or None
        uid            = session.get("UID","")
        if not new_name:
            return jsonify({"status":"error","message":"new_batch_name required"}), 400
        if not source_name:
            return jsonify({"status":"error","message":"source_batch_name required"}), 400
        if new_name == source_name:
            return jsonify({"status":"error","message":"New name must differ from source"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            # Check source exists
            src_rows = conn.execute(
                "SELECT * FROM procurement_formulations WHERE batch_name=%s ORDER BY id ASC",
                (source_name,)
            ).fetchall()
            if not src_rows:
                conn.close()
                return jsonify({"status":"error","message":f"Source batch '{source_name}' not found"}), 404
            # If batch already exists, delete it first (user chose Link explicitly)
            conn.execute(
                "DELETE FROM procurement_formulations WHERE batch_name=%s",
                (new_name,)
            )
            conn.commit()
            # Copy rows, tagging source_batch_name
            count = 0
            for r in src_rows:
                conn.execute("""
                    INSERT INTO procurement_formulations
                        (batch_name, product_code, material_name, supplier_name,
                         concentration, qty_kg,
                         batch_size, batch_date, num_batches,
                         imported_by, source_batch_name)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    new_name,
                    r["product_code"],
                    r["material_name"], r["supplier_name"],
                    r["concentration"], r["qty_kg"],
                    new_batch_size or r["batch_size"],
                    r["batch_date"], r["num_batches"],
                    uid, source_name
                ))
                count += 1
            conn.commit()
            conn.close()
            return jsonify({"status":"ok","linked":count,"new_batch_name":new_name,"source_batch_name":source_name})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    # ── API: get linked batches for a given batch ─────────────────
    @app.route("/api/procurement/formulations/linked_batches")
    @procurement_required
    def api_formulations_linked_batches():
        """
        ?batch_name=X  → returns all batches where source_batch_name=X
        """
        batch_name = request.args.get("batch_name","").strip()
        if not batch_name:
            return jsonify({"status":"error","message":"batch_name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            rows = conn.execute("""
                SELECT DISTINCT batch_name FROM procurement_formulations
                WHERE source_batch_name=%s
                ORDER BY batch_name ASC
            """, (batch_name,)).fetchall()
            conn.close()
            return jsonify({"status":"ok","linked_batches":[r["batch_name"] for r in rows]})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    # ── API: propagate ingredient changes to linked batches ────────
    @app.route("/api/procurement/formulations/propagate_to_linked", methods=["POST"])
    @procurement_required
    def api_formulations_propagate_to_linked():
        """
        Body: { source_batch_name, rows: [...] }
        Updates ingredient rows of ALL batches linked to source_batch_name
        with the new ingredient list, preserving each batch's own batch_size.
        """
        d           = request.get_json() or {}
        source_name = (d.get("source_batch_name") or "").strip()
        new_rows    = d.get("rows") or []
        uid         = session.get("UID","")
        if not source_name or not new_rows:
            return jsonify({"status":"error","message":"source_batch_name and rows required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            # Get all linked batches
            linked = conn.execute("""
                SELECT DISTINCT batch_name FROM procurement_formulations
                WHERE source_batch_name=%s
            """, (source_name,)).fetchall()
            linked_names = [r["batch_name"] for r in linked]
            updated_batches = []
            for lbn in linked_names:
                # Get this linked batch's own meta
                meta_row = conn.execute("""
                    SELECT batch_size,batch_date,num_batches,product_code
                    FROM procurement_formulations WHERE batch_name=%s LIMIT 1
                """, (lbn,)).fetchone()
                meta = dict(meta_row) if meta_row else {}
                # Delete old rows
                conn.execute("DELETE FROM procurement_formulations WHERE batch_name=%s", (lbn,))
                # Re-insert with new ingredients, keeping this batch's meta + source link
                def _nf(v):
                    if v is None or str(v).strip()=="": return None
                    try: return f"{float(str(v).replace(',','')):.6f}".rstrip('0').rstrip('.')
                    except: return str(v).strip()
                for r in new_rows:
                    mat = (r.get("material_name") or "").strip()
                    if not mat: continue
                    sup  = (r.get("supplier_name") or "").strip() or None
                    conn.execute("""
                        INSERT INTO procurement_formulations
                            (batch_name, product_code, material_name, supplier_name,
                             concentration, qty_kg,
                             batch_size, batch_date, num_batches,
                             imported_by, source_batch_name)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        lbn, meta.get("product_code"),
                        mat, sup,
                        _nf(r.get("concentration")), _nf(r.get("qty_kg")),
                        meta.get("batch_size"), meta.get("batch_date"), meta.get("num_batches"),
                        uid, source_name
                    ))
                updated_batches.append(lbn)
            conn.commit()
            conn.close()
            return jsonify({"status":"ok","updated_batches":updated_batches,"count":len(updated_batches)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    # ── API: unlink a batch (make independent copy) ────────────────
    @app.route("/api/procurement/formulations/unlink_batch", methods=["POST"])
    @procurement_required
    def api_formulations_unlink_batch():
        """
        Body: { batch_name }
        Sets source_batch_name=NULL for all rows of this batch,
        making it an independent formulation.
        """
        d          = request.get_json() or {}
        batch_name = (d.get("batch_name") or "").strip()
        if not batch_name:
            return jsonify({"status":"error","message":"batch_name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            conn.execute(
                "UPDATE procurement_formulations SET source_batch_name=NULL WHERE batch_name=%s",
                (batch_name,)
            )
            conn.commit()
            conn.close()
            return jsonify({"status":"ok","batch_name":batch_name})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    # ── API: linked formulations report ─────────────────────────
    @app.route("/api/procurement/formulations/linked_report")
    @procurement_required
    def api_formulations_linked_report():
        """Returns all source→linked relationships."""
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            rows = conn.execute("""
                SELECT DISTINCT batch_name, source_batch_name
                FROM procurement_formulations
                WHERE source_batch_name IS NOT NULL
                ORDER BY source_batch_name, batch_name
            """).fetchall()
            conn.close()
            return jsonify({"status":"ok","rows":[dict(r) for r in rows]})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}),500

    # ── API: batch cost per KG report ────────────────────────────
    @app.route("/api/procurement/formulations/cost_per_kg")
    @procurement_required
    def api_formulations_cost_per_kg():
        """
        For each batch, computes cost per KG using 1 KG as the standard batch size.
        This means cost_per_kg = SUM(concentration × 1 × last_purchase_rate)
        = SUM(concentration × rate)
        Works for ALL batches regardless of whether a procurement size is set.
        """
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            rows = conn.execute("""
                SELECT f.batch_name,
                       f.material_name,
                       f.concentration,
                       m.last_purchase_rate,
                       (SELECT MAX(f2.product_code)
                        FROM procurement_formulations f2
                        WHERE f2.batch_name = f.batch_name
                          AND f2.product_code IS NOT NULL
                          AND f2.product_code != '') AS product_code
                FROM   procurement_formulations f
                LEFT   JOIN procurement_materials m
                       ON LOWER(TRIM(f.material_name))=LOWER(TRIM(m.material_name))
                ORDER  BY f.batch_name, f.id
            """).fetchall()
            conn.close()
            from collections import defaultdict
            # Always use 1 KG batch size — cost_per_kg = SUM(conc × rate)
            batches = defaultdict(lambda: {"cost_per_kg":0.0,"has_rate":False,"missing_rate":[],"product_code":None})
            for r in rows:
                bn   = r["batch_name"]
                conc = r["concentration"]
                rate = r["last_purchase_rate"]
                # Capture product_code (first non-null wins)
                if r["product_code"] and not batches[bn]["product_code"]:
                    batches[bn]["product_code"] = r["product_code"]
                try: conc_f = float(conc) if conc else None
                except: conc_f = None
                try: rate_f = float(rate) if rate else None
                except: rate_f = None
                # --- Hardcoded fallback: applied BEFORE missing_rate check ---
                if rate_f is None:
                    rate_f = _hardcoded_rate(r["material_name"])
                # ------------------------------------------------------------
                if conc_f is not None and rate_f is not None:
                    batches[bn]["cost_per_kg"] += conc_f * rate_f   # × 1 KG
                    batches[bn]["has_rate"] = True
                elif conc_f is not None and rate_f is None:
                    batches[bn]["missing_rate"].append(r["material_name"])
            result = []
            for bn, d in batches.items():
                cpk = round(d["cost_per_kg"], 4) if d["cost_per_kg"] > 0 else None
                result.append({
                    "batch_name":    bn,
                    "product_code":  d["product_code"],
                    "cost_per_kg":   cpk,
                    "has_rate":      d["has_rate"],
                    "missing_rate":  d["missing_rate"]
                })
            result.sort(key=lambda x: x["batch_name"])
            return jsonify({"status":"ok","batches":result})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}),500

    # ── API: formulation change log ──────────────────────────────
    @app.route("/api/procurement/formulations/changelog")
    @procurement_required
    def api_formulations_changelog():
        batch_name = request.args.get("batch_name","").strip()
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            if batch_name:
                rows = conn.execute("""
                    SELECT * FROM procurement_formulation_changelog
                    WHERE batch_name=%s ORDER BY changed_at DESC LIMIT 200
                """, (batch_name,)).fetchall()
            else:
                rows = conn.execute("""
                    SELECT * FROM procurement_formulation_changelog
                    ORDER BY changed_at DESC LIMIT 500
                """).fetchall()
            conn.close()
            result = []
            for i,r in enumerate(rows,1):
                d = dict(r); d["sr_no"]=i; result.append(d)
            return jsonify({"status":"ok","rows":result})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}),500

    # ── API: log a formulation change ─────────────────────────────
    @app.route("/api/procurement/formulations/log_change", methods=["POST"])
    @procurement_required
    def api_formulations_log_change():
        d = request.get_json() or {}
        batch_name = (d.get("batch_name") or "").strip()
        change_type = (d.get("change_type") or "update").strip()
        before = d.get("ingredients_before")
        after  = d.get("ingredients_after")
        uid    = session.get("UID","")
        if not batch_name:
            return jsonify({"status":"error","message":"batch_name required"}),400
        import json as _json
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            conn.execute("""
                INSERT INTO procurement_formulation_changelog
                    (batch_name, changed_by, change_type, ingredients_before, ingredients_after)
                VALUES (%s,%s,%s,%s,%s)
            """, (batch_name, uid, change_type,
                  _json.dumps(before) if before else None,
                  _json.dumps(after)  if after  else None))
            conn.commit(); conn.close()
            return jsonify({"status":"ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}),500

    # ── API: cost impact of rate change ──────────────────────────
    @app.route("/api/procurement/formulations/cost_impact", methods=["POST"])
    @procurement_required
    def api_formulations_cost_impact():
        """
        Body: { material_name, old_rate, new_rate }
        Returns all batches using this material with old/new costs.
        """
        d             = request.get_json() or {}
        material_name = (d.get("material_name") or "").strip()
        old_rate      = d.get("old_rate")
        new_rate      = d.get("new_rate")
        if not material_name:
            return jsonify({"status":"error","message":"material_name required"}),400
        try:
            old_rate_f = float(old_rate) if old_rate is not None else None
            new_rate_f = float(new_rate) if new_rate is not None else None
        except (ValueError,TypeError):
            return jsonify({"status":"error","message":"invalid rate values"}),400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            # Get all formulation rows for this material
            rows = conn.execute("""
                SELECT f.batch_name, f.concentration, f.batch_size,
                       f.material_name
                FROM   procurement_formulations f
                WHERE  LOWER(TRIM(f.material_name))=LOWER(TRIM(%s))
                ORDER  BY f.batch_name
            """, (material_name,)).fetchall()
            # For each batch, also get ALL ingredients to compute total costs
            batch_names = list({r["batch_name"] for r in rows})
            all_ings = conn.execute("""
                SELECT f.batch_name, f.material_name, f.concentration, f.batch_size,
                       m.last_purchase_rate
                FROM   procurement_formulations f
                LEFT   JOIN procurement_materials m
                       ON LOWER(TRIM(f.material_name))=LOWER(TRIM(m.material_name))
                WHERE  f.batch_name IN ({})
                ORDER  BY f.batch_name, f.id
            """.format(','.join(['%s']*len(batch_names))), batch_names).fetchall() if batch_names else []
            conn.close()
            import re as _re
            from collections import defaultdict
            def _bs(bs_str):
                m2 = _re.search(r"[\d.]+", str(bs_str or ""))
                return float(m2.group()) if m2 else None
            def _f(v):
                try: return float(v) if v is not None else None
                except: return None

            # Group all ingredients by batch
            batch_ings = defaultdict(list)
            for r in all_ings:
                batch_ings[r["batch_name"]].append(r)

            results = []
            for bn in batch_names:
                ings = batch_ings[bn]
                bs   = _bs(ings[0]["batch_size"]) if ings else None
                # Include batches even when bs=None or bs=0:
                # 1KG cost always computable; procurement cost only when bs>0
                bs_for_proc = bs if (bs and bs > 0) else None
                # old_total/new_total = procurement size cost (c × bs × rate), 0 if no bs
                # old_1kg/new_1kg     = standard 1 KG cost (c × rate, bs=1)
                old_total=0.0; new_total=0.0
                old_1kg  =0.0; new_1kg  =0.0
                affected_ing=None
                ing_rows=[]
                for ing in ings:
                    c = _f(ing["concentration"])
                    if c is None: continue
                    qty = round(c * bs_for_proc, 3) if bs_for_proc else None
                    mat = ing["material_name"]
                    if mat.strip().lower()==material_name.lower():
                        r_old = old_rate_f or 0
                        r_new = new_rate_f or 0
                        affected_ing = {"material_name":mat,"qty_kg":qty,
                                        "old_rate":r_old,"new_rate":r_new,
                                        "old_cost":round(qty*r_old,2) if qty is not None else None,
                                        "new_cost":round(qty*r_new,2) if qty is not None else None}
                    else:
                        _db_rate = _f(ing["last_purchase_rate"])
                        if _db_rate is None:
                            _db_rate = _hardcoded_rate(ing["material_name"])
                        r_old = r_new = _db_rate or 0
                    if bs_for_proc:
                        old_total += c * bs_for_proc * r_old   # procurement cost
                        new_total += c * bs_for_proc * r_new
                    old_1kg   += c * r_old        # 1 KG standard cost
                    new_1kg   += c * r_new
                    ing_rows.append({"material_name":mat,"qty_kg":qty,
                                     "rate":r_new,"cost":round(c*bs_for_proc*r_new,2) if bs_for_proc else None})
                if affected_ing:
                    # cost_per_kg = 1 KG standard cost (always computed)
                    cpk_old = round(old_1kg, 4)
                    cpk_new = round(new_1kg, 4)
                    # Show batch_size as None if not set, so frontend shows "—"
                    results.append({
                        "batch_name":        bn,
                        "batch_size":        bs,
                        "old_total_cost":    round(old_total,2),
                        "new_total_cost":    round(new_total,2),
                        "cost_diff":         round(new_total-old_total,2),
                        "old_cost_per_kg":   cpk_old,
                        "new_cost_per_kg":   cpk_new,
                        "affected_ingredient": affected_ing,
                        "ingredients":       ing_rows
                    })
            return jsonify({"status":"ok","material_name":material_name,
                            "old_rate":old_rate_f,"new_rate":new_rate_f,
                            "affected_batches":results})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}),500


    # ── API: delete material ─────────────────────────────────
    @app.route("/api/procurement/delete_material", methods=["POST"])
    @procurement_required
    def api_procurement_delete_material():
        """
        Body: { material_name: str }
        Refuses deletion if the material appears in any formulation row.
        """
        d   = request.get_json() or {}
        mat = (d.get("material_name") or "").strip()
        if not mat:
            return jsonify({"status":"error","message":"material_name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            # Check formulation usage
            used = conn.execute(
                """SELECT DISTINCT batch_name FROM procurement_formulations
                   WHERE LOWER(TRIM(material_name))=LOWER(TRIM(%s)) LIMIT 10""",
                (mat,)
            ).fetchall()
            if used:
                batch_list = ", ".join(r["batch_name"] for r in used[:5])
                extra = f" (+{len(used)-5} more)" if len(used) > 5 else ""
                conn.close()
                return jsonify({
                    "status":  "error",
                    "code":    "in_use",
                    "message": f"Cannot delete — used in {len(used)} formulation batch{'es' if len(used)>1 else ''}: {batch_list}{extra}"
                }), 409
            conn.execute(
                "DELETE FROM procurement_materials WHERE material_name=%s", (mat,)
            )
            conn.commit()
            conn.close()
            return jsonify({"status":"ok","deleted":mat})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    # ── API: bulk delete materials ───────────────────────────
    @app.route("/api/procurement/delete_materials", methods=["POST"])
    @procurement_required
    def api_procurement_delete_materials():
        """
        Body: { material_names: [str, ...] }
        Skips any material that is used in a formulation, reports results.
        """
        d     = request.get_json() or {}
        names = [n.strip() for n in (d.get("material_names") or []) if str(n).strip()]
        if not names:
            return jsonify({"status":"error","message":"material_names required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            deleted=[];  blocked=[]
            for mat in names:
                used = conn.execute(
                    """SELECT COUNT(DISTINCT batch_name) AS cnt
                       FROM procurement_formulations
                       WHERE LOWER(TRIM(material_name))=LOWER(TRIM(%s))""",
                    (mat,)
                ).fetchone()
                if used and used["cnt"] > 0:
                    blocked.append({"material_name":mat,"batch_count":used["cnt"]})
                else:
                    conn.execute("DELETE FROM procurement_materials WHERE material_name=%s",(mat,))
                    deleted.append(mat)
            conn.commit()
            conn.close()
            return jsonify({"status":"ok","deleted":deleted,"blocked":blocked})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    # ── API: check for duplicate formulations before import ──────
    @app.route("/api/procurement/formulations/check_duplicates", methods=["POST"])
    @procurement_required
    def api_formulations_check_duplicates():
        d = request.get_json() or {}
        sheets = d.get("sheets") or []
        if not sheets:
            return jsonify({"status":"ok","results":[]}), 200
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500

            def _fp_conc(v):
                try: return round(float(v), 6)
                except: return None

            rows = conn.execute("""
                SELECT batch_name, material_name, concentration
                FROM procurement_formulations ORDER BY batch_name
            """).fetchall()
            from collections import defaultdict
            existing = defaultdict(list)
            for row in rows:
                existing[row["batch_name"]].append(row)
            existing_fps = {
                bn: frozenset(
                    (str(r["material_name"]).strip().lower(), _fp_conc(r["concentration"]))
                    for r in brows if r["material_name"]
                )
                for bn, brows in existing.items()
            }

            results = []
            for sheet in sheets:
                batch_name  = (sheet.get("batch_name") or "").strip()
                ingredients = sheet.get("ingredients") or []
                proposed_fp = frozenset(
                    (str(r.get("material_name","")).strip().lower(), _fp_conc(r.get("concentration")))
                    for r in ingredients if r.get("material_name","").strip()
                )
                matching_batch = next(
                    (bn for bn, fp in existing_fps.items()
                     if proposed_fp == fp and len(proposed_fp) > 0),
                    None
                )
                results.append({
                    "batch_name":       batch_name,
                    "is_duplicate":     matching_batch is not None,
                    "matching_batch":   matching_batch,
                    "exact_exists":     batch_name in existing,
                    "ingredient_count": len(proposed_fp),
                })
            conn.close()
            return jsonify({"status":"ok","results":results})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    # ── API: correct formulations — rename materials, report missing suppliers ──
    @app.route("/api/procurement/formulations/correct", methods=["POST"])
    @procurement_required
    def api_formulations_correct():
        """
        Body: {
            action: "preview" | "apply",
            tasks:  [{ type: "rename", from: str, to: str }]   # for action=apply
        }
        Returns preview of what will change, or applies the corrections.
        Also always returns:
          - missing_supplier: [ {material_name, batch_count} ]
            materials in formulations with no supplier in procurement_materials
        """
        import json as _json
        d      = request.get_json() or {}
        action = d.get("action", "preview")
        tasks  = d.get("tasks", [])
        uid    = session.get("UID", "")
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500

            # ── Collect rename previews ──────────────────────────────────────
            rename_preview = []
            for t in tasks:
                if t.get("type") != "rename": continue
                frm = (t.get("from") or "").strip()
                to  = (t.get("to")   or "").strip()
                if not frm or not to or frm == to: continue
                rows = conn.execute("""
                    SELECT DISTINCT batch_name
                    FROM procurement_formulations
                    WHERE LOWER(TRIM(material_name)) = LOWER(TRIM(%s))
                    ORDER BY batch_name
                """, (frm,)).fetchall()
                count = len(rows)
                batches = [r["batch_name"] for r in rows[:5]]
                rename_preview.append({
                    "from":    frm,
                    "to":      to,
                    "count":   count,
                    "batches": batches,
                    "extra":   max(0, count - 5)
                })

            # ── Apply renames ────────────────────────────────────────────────
            applied = []
            if action == "apply":
                for t in tasks:
                    if t.get("type") != "rename": continue
                    frm = (t.get("from") or "").strip()
                    to  = (t.get("to")   or "").strip()
                    if not frm or not to or frm == to: continue
                    result = conn.execute("""
                        UPDATE procurement_formulations
                        SET    material_name = %s
                        WHERE  LOWER(TRIM(material_name)) = LOWER(TRIM(%s))
                    """, (to, frm))
                    rows_updated = result.rowcount
                    if rows_updated:
                        applied.append({"from": frm, "to": to, "rows_updated": rows_updated})
                conn.commit()

            # ── Missing supplier report ──────────────────────────────────────
            # Materials in formulations that have no entry in procurement_materials
            missing_rows = conn.execute("""
                SELECT   f.material_name,
                         COUNT(DISTINCT f.batch_name) AS batch_count
                FROM     procurement_formulations f
                LEFT JOIN procurement_materials m
                         ON LOWER(TRIM(f.material_name)) = LOWER(TRIM(m.material_name))
                WHERE    m.material_name IS NULL
                   OR    (m.supplier_name IS NULL OR TRIM(m.supplier_name) = '')
                GROUP BY f.material_name
                ORDER BY batch_count DESC, f.material_name ASC
            """).fetchall()

            missing_supplier = [
                {"material_name": r["material_name"], "batch_count": r["batch_count"]}
                for r in missing_rows
            ]

            conn.close()
            return jsonify({
                "status":           "ok",
                "rename_preview":   rename_preview,
                "applied":          applied,
                "missing_supplier": missing_supplier
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    # ── Rename a batch (all rows in procurement_formulations + update_log) ──
    @app.route("/api/procurement/formulations/rename_batch", methods=["POST"])
    @procurement_required
    def api_formulations_rename_batch():
        d        = request.get_json() or {}
        old_name = (d.get("old_name") or "").strip()
        new_name = (d.get("new_name") or "").strip()
        uid      = session.get("UID", "")
        if not old_name:
            return jsonify({"status":"error","message":"old_name required"}), 400
        if not new_name:
            return jsonify({"status":"error","message":"new_name required"}), 400
        if old_name == new_name:
            return jsonify({"status":"error","message":"New name is the same"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            # Check old batch exists
            exists = conn.execute(
                "SELECT 1 FROM procurement_formulations WHERE batch_name=%s LIMIT 1",
                (old_name,)
            ).fetchone()
            if not exists:
                conn.close()
                return jsonify({"status":"error","message":f"Batch '{old_name}' not found"}), 404
            # Check new name not already taken
            conflict = conn.execute(
                "SELECT 1 FROM procurement_formulations WHERE batch_name=%s LIMIT 1",
                (new_name,)
            ).fetchone()
            if conflict:
                conn.close()
                return jsonify({"status":"error","message":f"Batch '{new_name}' already exists"}), 409
            # Rename in formulations
            r1 = conn.execute(
                "UPDATE procurement_formulations SET batch_name=%s WHERE batch_name=%s",
                (new_name, old_name)
            )
            # Rename in update_log
            conn.execute(
                "UPDATE procurement_update_log SET batch_name=%s WHERE batch_name=%s",
                (new_name, old_name)
            )
            # Rename source_batch_name references (linked children)
            conn.execute(
                "UPDATE procurement_formulations SET source_batch_name=%s WHERE source_batch_name=%s",
                (new_name, old_name)
            )
            # Rename in changelog
            conn.execute(
                "UPDATE procurement_formulation_changelog SET batch_name=%s WHERE batch_name=%s",
                (new_name, old_name)
            )
            conn.commit()
            conn.close()
            return jsonify({"status":"ok","old_name":old_name,"new_name":new_name,"rows_updated":r1.rowcount})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    @app.route("/api/procurement/formulations/cost_impact_multi", methods=["POST"])
    @procurement_required
    def api_formulations_cost_impact_multi():
        """Combined cost impact for multiple material rate changes."""
        import re as _re
        from collections import defaultdict
        d       = request.get_json() or {}
        changes = d.get("changes", [])
        if not changes:
            return jsonify({"status":"error","message":"No changes provided"}), 400
        rate_map = {}
        for ch in changes:
            name = (ch.get("material_name") or "").strip()
            if not name: continue
            try:
                old_r = float(ch["old_rate"]) if ch.get("old_rate") is not None else None
                new_r = float(ch["new_rate"]) if ch.get("new_rate") is not None else None
            except (ValueError, TypeError):
                continue
            rate_map[name.lower()] = {"name": name, "old": old_r, "new": new_r}
        if not rate_map:
            return jsonify({"status":"error","message":"No valid changes"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            names = [v["name"] for v in rate_map.values()]
            ph    = ','.join(['%s'] * len(names))
            rows  = conn.execute(
                "SELECT DISTINCT batch_name FROM procurement_formulations "
                "WHERE LOWER(TRIM(material_name)) IN ({}) ORDER BY batch_name".format(ph),
                names).fetchall()
            batch_names = [r["batch_name"] for r in rows]
            if not batch_names:
                conn.close()
                return jsonify({"status":"ok","affected_batches":[],"changes":changes})
            all_ings = conn.execute(
                "SELECT f.batch_name,f.material_name,f.concentration,f.batch_size,"
                "m.last_purchase_rate FROM procurement_formulations f "
                "LEFT JOIN procurement_materials m "
                "ON LOWER(TRIM(f.material_name))=LOWER(TRIM(m.material_name)) "
                "WHERE f.batch_name IN ({}) ORDER BY f.batch_name,f.id".format(
                    ','.join(['%s']*len(batch_names))),
                batch_names).fetchall()
            conn.close()
            def _bs(s):
                m2 = _re.search("[0-9.]+", str(s or ""))
                return float(m2.group()) if m2 else None
            def _f(v):
                try: return float(v) if v is not None else None
                except: return None
            bi = defaultdict(list)
            for r in all_ings:
                bi[r["batch_name"]].append(r)
            results = []
            for bn in batch_names:
                ings = bi[bn]
                bs   = _bs(ings[0]["batch_size"]) if ings else None
                bfp  = bs if (bs and bs > 0) else None
                o1 = n1 = ot = nt = 0.0
                for ing in ings:
                    c = _f(ing["concentration"])
                    if c is None: continue
                    ml = (ing["material_name"] or "").strip().lower()
                    if ml in rate_map:
                        ro = rate_map[ml]["old"] or 0
                        rn = rate_map[ml]["new"] or 0
                    else:
                        db = _f(ing["last_purchase_rate"])
                        if db is None:
                            db = _hardcoded_rate(ing["material_name"])
                        ro = rn = db or 0
                    o1 += c * ro;  n1 += c * rn
                    if bfp: ot += c * bfp * ro;  nt += c * bfp * rn
                results.append({
                    "batch_name":      bn,
                    "batch_size":      bfp,
                    "old_cost_per_kg": round(o1, 4),
                    "new_cost_per_kg": round(n1, 4),
                    "old_total_cost":  round(ot, 2),
                    "new_total_cost":  round(nt, 2),
                    "cost_diff":       round(nt - ot, 2),
                })
            results.sort(key=lambda x: abs(x["new_cost_per_kg"]-x["old_cost_per_kg"]), reverse=True)
            return jsonify({"status":"ok","affected_batches":results,"changes":changes})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    # ── Admin Data Reset ─────────────────────────────────────────────────────
    @app.route("/api/procurement/admin/reset", methods=["POST"])
    @procurement_required
    def api_procurement_admin_reset():
        role = (session.get("User_Type") or "").strip().lower()
        if role != "admin":
            return jsonify({"status":"error","message":"Admin only"}), 403
        d             = request.get_json() or {}
        scope         = d.get("scope", [])
        confirm_token = (d.get("confirm_token") or "").strip()
        if confirm_token != "CONFIRM-DELETE":
            return jsonify({"status":"error","message":"Invalid confirmation token"}), 400
        if not scope:
            return jsonify({"status":"error","message":"Nothing selected"}), 400
        # Human-readable dependency descriptions for FK violations
        DEPENDENCY_HINTS = {
            "procurement_materials":   "Materials are referenced by Purchase Orders or GRNs. Clear Purchase Orders and GRNs first.",
            "procurement_suppliers":   "Suppliers are referenced by Purchase Orders. Clear Purchase Orders first.",
            "procurement_godowns":     "Godowns are referenced by Purchase Orders (billing/shipping). Clear Purchase Orders first.",
            "procurement_tc_lists":    "T&C lists are linked to Purchase Orders. Clear Purchase Orders first.",
            "procurement_purchase_orders": "Purchase Orders are referenced by GRNs. Clear GRNs first.",
            "procurement_formulations": "Formulations may be referenced by other records.",
        }
        TABLE_MAP = {
            "formulations": "procurement_formulations",
            "materials":    "procurement_materials",
            "update_log":   "procurement_update_log",
            "changelog":    "procurement_formulation_changelog",
            "settings":     "procurement_settings",
            "tc_lists":     "procurement_tc_lists",
            "suppliers":    "procurement_suppliers",
            "godowns":      "procurement_godowns",
        }

        def _is_fk_error(e):
            msg = str(e).lower()
            return "foreign key" in msg or "cannot delete" in msg or "a foreign" in msg or "1451" in msg

        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500

            deleted  = {}
            blocked  = {}  # key → reason string shown to user

            # ── GRN cascade (items first, then header) ──
            if "grn" in scope:
                try:
                    conn.execute("DELETE FROM procurement_grn_items")
                    conn.execute("DELETE FROM procurement_grn")
                    deleted["procurement_grn"] = "all"
                except Exception as e:
                    conn.execute("ROLLBACK") if hasattr(conn, "execute") else None
                    if _is_fk_error(e):
                        blocked["grn"] = "GRNs are referenced by other records and could not be deleted."
                    else:
                        blocked["grn"] = f"GRN delete failed: {e}"

            # ── Purchase Orders cascade (items first, then header) ──
            if "purchase_orders" in scope:
                try:
                    conn.execute("DELETE FROM procurement_po_items")
                    conn.execute("DELETE FROM procurement_purchase_orders")
                    deleted["procurement_purchase_orders"] = "all"
                except Exception as e:
                    conn.execute("ROLLBACK") if hasattr(conn, "execute") else None
                    if _is_fk_error(e):
                        blocked["purchase_orders"] = DEPENDENCY_HINTS.get("procurement_purchase_orders",
                            "Purchase Orders are referenced by GRNs. Clear GRNs first.")
                    else:
                        blocked["purchase_orders"] = f"PO delete failed: {e}"

            # ── All other tables ──
            for key in scope:
                if key in ("purchase_orders", "grn"):
                    continue
                tbl = TABLE_MAP.get(key)
                if not tbl:
                    continue
                try:
                    result = conn.execute(f"DELETE FROM {tbl}")
                    deleted[tbl] = result.rowcount
                except Exception as e:
                    try: conn.execute("ROLLBACK")
                    except Exception: pass
                    if _is_fk_error(e):
                        blocked[key] = DEPENDENCY_HINTS.get(tbl,
                            f"'{tbl}' has linked records in another table. Remove those dependencies first.")
                    else:
                        blocked[key] = f"Delete failed: {e}"

            conn.commit()
            conn.close()

            if blocked:
                # Some tables were blocked — report what succeeded and what was blocked
                return jsonify({
                    "status":  "partial" if deleted else "blocked",
                    "deleted": deleted,
                    "blocked": blocked,
                    "message": "Some tables could not be cleared due to linked records."
                })
            return jsonify({"status":"ok","deleted":deleted})

        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    # ── Manufacturing process per batch ──────────────────────────────────────
    @app.route("/api/procurement/formulations/manuf_process", methods=["POST"])
    @procurement_required
    def api_procurement_manuf_process():
        d          = request.get_json() or {}
        batch_name = (d.get("batch_name") or "").strip()
        action     = (d.get("action")     or "get").strip()
        if not batch_name:
            return jsonify({"status":"error","message":"batch_name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            if action == "save":
                text = (d.get("text") or "").strip()
                conn.execute(
                    "UPDATE procurement_formulations SET manuf_process=%s WHERE batch_name=%s",
                    (text or None, batch_name)
                )
                conn.commit()
                conn.close()
                return jsonify({"status":"ok","batch_name":batch_name,"text":text})
            else:
                row = conn.execute(
                    "SELECT manuf_process FROM procurement_formulations WHERE batch_name=%s LIMIT 1",
                    (batch_name,)
                ).fetchone()
                conn.close()
                return jsonify({"status":"ok","text":row["manuf_process"] if row else ""})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    # ── RM Requirement from DB formulations ──────────────────────────────────
    @app.route("/api/procurement/formulations/rm_requirement", methods=["POST"])
    @procurement_required
    def api_procurement_rm_requirement():
        import re as _re
        from collections import defaultdict
        d          = request.get_json() or {}
        cart_items = d.get("cart_items", [])
        if not cart_items:
            return jsonify({"status":"error","message":"No batches in cart"}), 400
        size_override = {}
        batch_names   = []
        for item in cart_items:
            bn = (item.get("batch_name") or "").strip()
            bs = item.get("batch_size")
            if bn:
                batch_names.append(bn)
                if bs is not None:
                    try: size_override[bn] = float(bs)
                    except: pass
        batch_names = list(dict.fromkeys(batch_names))
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            ph  = ','.join(['%s'] * len(batch_names))
            sql = ("SELECT f.batch_name, f.material_name, f.concentration, f.batch_size "
                   "FROM procurement_formulations f "
                   "WHERE f.batch_name IN ({}) ORDER BY f.batch_name, f.id").format(ph)
            rows = conn.execute(sql, batch_names).fetchall()
            mat_rows = conn.execute(
                "SELECT material_name, last_purchase_rate, supplier_name FROM procurement_materials"
            ).fetchall()
            conn.close()
            rate_map     = {r["material_name"].strip().lower(): r["last_purchase_rate"] for r in mat_rows}
            supplier_map = {r["material_name"].strip().lower(): r["supplier_name"]      for r in mat_rows}
            stock_map  = _read_stksum()
            stk_error  = None if stock_map else "StkSum.xlsx could not be read"
            batch_info = {}
            for r in rows:
                bn = r["batch_name"]
                if bn not in batch_info:
                    if bn in size_override:
                        batch_info[bn] = size_override[bn]
                    else:
                        bs_m = _re.search("[0-9.]+", str(r["batch_size"] or "0"))
                        batch_info[bn] = float(bs_m.group()) if bs_m else 0.0
            mat_totals  = defaultdict(float)
            mat_batches = defaultdict(set)
            for r in rows:
                bn   = r["batch_name"]
                bs   = batch_info.get(bn, 0)
                conc = float(r["concentration"] or 0)
                if bs > 0 and conc > 0:
                    mat_totals[r["material_name"]]  += round(conc * bs, 6)
                    mat_batches[r["material_name"]].add(bn)
            _rmIgnore = ['demineralized water','d.m. water','dm water','demineralised water']
            materials = []
            for mat_name, total_qty in mat_totals.items():
                if mat_name.strip().lower() in _rmIgnore:
                    continue
                k             = mat_name.strip().lower()
                current_stock = stock_map.get(k)
                stock_diff    = round(current_stock - total_qty, 4) if current_stock is not None else None
                materials.append({
                    "name":          mat_name,
                    "total_qty":     round(total_qty, 4),
                    "current_stock": round(current_stock, 4) if current_stock is not None else None,
                    "stock_diff":    stock_diff,
                    "supplier":      supplier_map.get(k),
                    "rate":          float(rate_map[k]) if k in rate_map and rate_map[k] is not None else None,
                    "batches":       sorted(mat_batches[mat_name]),
                })
            materials.sort(key=lambda m: (0 if (m["stock_diff"] is not None and m["stock_diff"]<0) else 1, m["name"].lower()))
            batches_out = [{"batch_name":bn,"batch_size":batch_info.get(bn,0)} for bn in batch_names if bn in batch_info]
            return jsonify({"status":"ok","materials":materials,"batches":batches_out,"stk_error":stk_error})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    # ── Access check ─────────────────────────────────────────────────────
    @app.route("/api/procurement/can_access")
    def api_procurement_can_access():
        return jsonify({"can_access": _can_procurement()})


    # ═══════════════════════════════════════════════════════════
    # FORMULATIONS VS QTY — routes
    # ═══════════════════════════════════════════════════════════

    # ── Step 1: Upload file, return filename + worksheet names ──
    @app.route("/api/procurement/formulations/inspect", methods=["POST"])
    @procurement_required
    def api_formulations_inspect():
        import tempfile, uuid
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "No file uploaded"}), 400
        f = request.files["file"]
        if not f.filename.lower().endswith(".xlsx"):
            return jsonify({"status": "error", "message": "Only .xlsx files accepted"}), 400
        tmp_dir  = tempfile.gettempdir()
        tmp_name = f"procform_{uuid.uuid4().hex}.xlsx"
        tmp_path = os.path.join(tmp_dir, tmp_name)
        f.save(tmp_path)
        try:
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            # Strip whitespace from sheet names; build map stripped->original
            _sheet_map = {}
            for _raw in wb.sheetnames:
                _stripped = _raw.strip()
                if _stripped.lower() == "index":
                    continue
                _sheet_map[_stripped] = _raw
            all_sheets = list(_sheet_map.keys())

            sheets            = []   # valid (A13 == 1)
            invalid_sheets    = {}   # { name: reason }
            sheet_ingredients = {}   # { name: [{material_name, concentration}] }

            for sname in all_sheets:
                ws = wb[_sheet_map.get(sname, sname)]
                trigger_cell = ws.cell(row=13, column=1).value
                try:
                    trigger_val = int(float(str(trigger_cell).strip()))                         if trigger_cell not in (None, "") else None
                except (ValueError, TypeError):
                    trigger_val = None

                if trigger_val != 1:
                    reason = ("Cell A13 is empty — expected the number 1 in A13 "
                              "as the first row marker." if (trigger_cell is None or
                              str(trigger_cell).strip() == "")
                              else f"Cell A13 contains '{trigger_cell}' instead of 1. "
                                   "Expected the number 1 in A13 as the first row marker.")
                    invalid_sheets[sname] = reason
                    continue

                sheets.append(sname)
                ings = []
                for ri in range(13, (ws.max_row or 13) + 1):
                    ca = ws.cell(row=ri, column=1).value
                    if ca is None or str(ca).strip() == "":
                        break
                    mat  = str(ws.cell(row=ri, column=2).value or "").strip()
                    conc = ws.cell(row=ri, column=4).value
                    if mat:
                        try:
                            cf = f"{float(conc):.6f}".rstrip('0').rstrip('.')                                 if conc is not None else None
                        except (TypeError, ValueError):
                            cf = None
                        ings.append({"material_name": mat, "concentration": cf})
                sheet_ingredients[sname] = ings

            wb.close()
        except Exception as e:
            os.remove(tmp_path)
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        session["_procform_tmp"] = tmp_path
        basename = os.path.splitext(os.path.basename(f.filename))[0]
        return jsonify({
            "status":            "ok",
            "basename":          basename,
            "sheets":            sheets,
            "invalid_sheets":    invalid_sheets,
            "sheet_ingredients": sheet_ingredients,
            "tmp_key":           tmp_name
        })

    # ── Step 2: Parse selected sheets and store into DB ─────────
    @app.route("/api/procurement/formulations/import", methods=["POST"])
    @procurement_required
    def api_formulations_import():
        d        = request.get_json() or {}
        basename = (d.get("basename") or "").strip()
        # sheets: list of strings OR list of {sheet, batch_name} dicts
        raw_sheets = d.get("sheets") or []
        sheets_with_names = []
        for _s in raw_sheets:
            if isinstance(_s, dict):
                _sn = (_s.get("sheet") or "").strip()
                _bn = (_s.get("batch_name") or "").strip() or None
                sheets_with_names.append((_sn, _bn))
            else:
                sheets_with_names.append((str(_s).strip(), None))
        tmp_path = session.get("_procform_tmp", "")
        if not sheets_with_names:
            return jsonify({"status":"error","message":"No sheets selected"}), 400
        if not tmp_path or not os.path.exists(tmp_path):
            return jsonify({"status":"error","message":"Upload session expired — please re-upload the file"}), 400
        uid = session.get("UID","")
        total_imported=0; total_skipped=0; sheet_results=[]; errors=[]
        try:
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":f"Cannot open file: {e}"}), 500
        conn = sampling_portal.get_db_connection()
        if not conn:
            wb.close()
            return jsonify({"status":"error","message":"DB connection failed"}), 500

        def _cs(ws,row,col):
            v=ws.cell(row=row,column=col).value
            if v is None: return None
            s=str(v).strip(); return s if s else None

        def _ff(v):
            if v is None: return None
            try:
                f=float(v); return f"{f:.6f}".rstrip('0').rstrip('.')
            except: return str(v).strip()

        try:
            for sheet_name, custom_batch in sheets_with_names:
                if sheet_name not in wb.sheetnames:
                    sheet_results.append({"sheet":sheet_name,"imported":0,"skipped":0,"note":"not found"})
                    continue
                ws=wb[sheet_name]; # Avoid duplicating sheet name when filename already contains it
                _bn_check = sheet_name.strip().lower()
                _base_check = basename.strip().lower()
                if _bn_check in _base_check:
                    batch_name = basename.strip()
                else:
                    batch_name = f"{basename} – {sheet_name}"
                imported=0; skipped=0
                # Batch meta not imported — user enters manually in popup
                batch_size  = None
                batch_date  = None
                num_batches = None
                # Scan rows 1-20 for a cell whose text starts with "Product Code"
                # The product code value is in the NEXT cell to the right,
                # or after the label text in the same cell.
                product_code = None
                for _pr in range(1, min(20, ws.max_row + 1)):
                    for _pc in range(1, min(15, (ws.max_column or 15) + 1)):
                        _cv = ws.cell(row=_pr, column=_pc).value
                        if _cv is None: continue
                        _cvs = str(_cv).strip()
                        if _cvs.lower().startswith('product code'):
                            # Value in next cell to the right
                            _nxt = ws.cell(row=_pr, column=_pc + 1).value
                            if _nxt and str(_nxt).strip():
                                product_code = str(_nxt).strip()
                            # Same cell: text after "Product Code" / "Product Code:"
                            elif len(_cvs) > len('product code'):
                                _rest = _cvs[len('product code'):].lstrip(' :\t\u2013-').strip()
                                if _rest:
                                    product_code = _rest
                            if product_code:
                                break
                    if product_code:
                        break
                # Ingredient rows
                for ri in range(13, ws.max_row+1):
                    ca=ws.cell(row=ri,column=1).value
                    if ca is None or str(ca).strip()=="": break
                    cb=_cs(ws,ri,2); cc=_cs(ws,ri,3)
                    cd=ws.cell(row=ri,column=4).value; ce=ws.cell(row=ri,column=5).value
                    mat=cb or ""
                    if not mat: skipped+=1; continue
                    try:
                        conn.execute("""
                            INSERT INTO procurement_formulations
                                (batch_name,product_code,material_name,supplier_name,
                                 concentration,qty_kg,
                                 batch_size,batch_date,num_batches,imported_by)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            ON DUPLICATE KEY UPDATE
                                product_code=VALUES(product_code),
                                supplier_name=VALUES(supplier_name),
                                concentration=VALUES(concentration),
                                qty_kg=VALUES(qty_kg),
                                batch_size=VALUES(batch_size),
                                batch_date=VALUES(batch_date),
                                num_batches=VALUES(num_batches),
                                imported_by=VALUES(imported_by),
                                imported_at=CURRENT_TIMESTAMP
                        """,(batch_name,product_code,mat,cc,_ff(cd),_ff(ce),batch_size,batch_date,num_batches,uid))
                        imported+=1
                    except Exception as re:
                        errors.append(f"{batch_name}/{mat}: {re}"); skipped+=1
                conn.commit()
                total_imported+=imported; total_skipped+=skipped
                sheet_results.append({"sheet":sheet_name,"batch":batch_name,
                    "batch_size":batch_size,"batch_date":batch_date,
                    "imported":imported,"skipped":skipped})
        except Exception as e:
            traceback.print_exc(); errors.append(str(e))
        finally:
            wb.close()
            try: conn.close()
            except: pass
            try:
                if tmp_path and os.path.exists(tmp_path): os.remove(tmp_path)
            except: pass
            session.pop("_procform_tmp",None)
        return jsonify({"status":"ok","total_imported":total_imported,
            "total_skipped":total_skipped,"sheet_results":sheet_results,"errors":errors[:10]})

    # ── API: import manufacturing process from Excel ─────────────
    @app.route("/api/procurement/formulations/import_manuf_process", methods=["POST"])
    @procurement_required
    def api_import_manuf_process():
        """
        Upload same Excel file used for formulation import.
        For each sheet:
          1. Derive batch_name using identical logic as formulation import.
          2. Find where ingredient rows end (Col A becomes empty from row 13).
          3. Scan the next 10 rows for a cell containing "MANUFACTURING" or
             "PRODUCT SPECIFICATION" (case-insensitive).
          4. From that anchor row, read all remaining rows to end of sheet.
          5. Build an HTML table from those rows and save as manuf_process.
        """
        import tempfile, uuid, html as _html
        if "file" not in request.files:
            return jsonify({"status": "error", "message": "No file uploaded"}), 400
        f = request.files["file"]
        if not f.filename.lower().endswith(".xlsx"):
            return jsonify({"status": "error", "message": "Only .xlsx files accepted"}), 400

        tmp_dir  = tempfile.gettempdir()
        tmp_name = f"procmp_{uuid.uuid4().hex}.xlsx"
        tmp_path = os.path.join(tmp_dir, tmp_name)
        f.save(tmp_path)

        basename = os.path.splitext(os.path.basename(f.filename))[0]
        results  = []   # {sheet, batch_name, status, message}
        updated  = 0
        skipped  = 0
        not_found = 0

        try:
            wb = load_workbook(tmp_path, read_only=True, data_only=True)

            # Build sheet map (strip whitespace, skip Index)
            sheet_map = {}
            for raw in wb.sheetnames:
                stripped = raw.strip()
                if stripped.lower() == "index":
                    continue
                sheet_map[stripped] = raw

            conn = sampling_portal.get_db_connection()
            if not conn:
                wb.close()
                return jsonify({"status": "error", "message": "DB connection failed"}), 500

            # Row layout constants
            # Page 1: header rows 1-12, ingredients rows 13 to ~(13+n)
            # Gap:    ~10 rows between ingredient end and process start
            # Page 2: manufacturing process, ~40 rows
            MANUF_GAP       = 10   # rows to skip after ingredient loop breaks
            MANUF_PAGE_ROWS = 40   # rows to read for manufacturing process

            for sname, raw_name in sheet_map.items():
                ws = wb[raw_name]

                # ── Derive batch name (identical to formulation import) ──
                _bn_check   = sname.strip().lower()
                _base_check = basename.strip().lower()
                if _bn_check in _base_check:
                    batch_name = basename.strip()
                else:
                    batch_name = f"{basename} – {sname}"

                # ── Find where ingredient rows end (Col A empty from row 13) ──
                ingredient_end_row = 13
                for ri in range(13, (ws.max_row or 13) + 1):
                    ca = ws.cell(row=ri, column=1).value
                    if ca is None or str(ca).strip() == "":
                        ingredient_end_row = ri
                        break

                # ── Manufacturing process starts 10 rows after ingredient end ──
                manuf_start = ingredient_end_row + MANUF_GAP
                manuf_end   = manuf_start + MANUF_PAGE_ROWS

                # ── Read rows; preserve newlines inside cells ──
                html_rows = []
                max_col   = ws.max_column or 1

                for ri in range(manuf_start, manuf_end):
                    row_cells = []
                    has_content = False
                    for col in range(1, max_col + 1):
                        cv = ws.cell(row=ri, column=col).value
                        # Preserve embedded newlines; only trim leading/trailing whitespace
                        cell_text = str(cv).strip() if cv is not None else ""
                        if cell_text:
                            has_content = True
                        row_cells.append(cell_text)

                    if not has_content:
                        html_rows.append(None)   # blank spacer
                        continue

                    # Strip trailing empty cells
                    while row_cells and not row_cells[-1]:
                        row_cells.pop()

                    html_rows.append(row_cells)

                # Remove leading and trailing blank spacers
                while html_rows and html_rows[0]  is None: html_rows.pop(0)
                while html_rows and html_rows[-1] is None: html_rows.pop()

                if not html_rows:
                    skipped += 1
                    results.append({
                        "sheet": sname, "batch_name": batch_name,
                        "status": "skipped",
                        "message": f"No content found in rows {manuf_start}–{manuf_end} (ingredient ended at row {ingredient_end_row})"
                    })
                    continue

                # ══════════════════════════════════════════════════════════════
                # Build CANONICAL clean HTML (matches Excel layout exactly)
                # Structure:
                #   <table class="mp-spec"> ... </table>    — spec block
                #   <table class="mp-steps"> ... </table>   — process block
                # Both paths (this importer + the JS cleaner) produce identical
                # structures so print rendering is consistent.
                # ══════════════════════════════════════════════════════════════
                import re as _re_mp

                _SPEC_KW  = ('product specification',)
                _PROC_KW  = ('manufacturing process',)

                def _row_text(row):
                    if row is None: return ""
                    return " ".join(c.strip() for c in row if c and c.strip())

                def _is_spec_header(row):
                    t = _row_text(row).lower()
                    return any(kw in t for kw in _SPEC_KW)

                def _is_proc_header(row):
                    t = _row_text(row).lower()
                    return any(kw in t for kw in _PROC_KW)

                def _strip_left_empty_cols(rows):
                    """Remove columns from the left that are empty in every data row."""
                    if not rows: return rows
                    while True:
                        any_left = False
                        for r in rows:
                            if r and len(r) > 0 and r[0].strip():
                                any_left = True; break
                        if any_left: break
                        stripped_any = False
                        for r in rows:
                            if r and len(r) > 0:
                                r.pop(0); stripped_any = True
                        if not stripped_any: break
                    return rows

                def _cell_to_html(text):
                    """Escape text and convert embedded newlines.
                    If the cell has multiple lines and the first line looks like a
                    subheading (short, Title Case, optionally with parenthetical),
                    wrap the first line in <strong> and separate the rest with <br>.
                    """
                    if not text: return ""
                    # Normalise newlines
                    lines = [ln.strip() for ln in text.replace('\r\n','\n').replace('\r','\n').split('\n')]
                    lines = [ln for ln in lines if ln]
                    if not lines:
                        return ""
                    if len(lines) == 1:
                        return _html.escape(lines[0])
                    # Multi-line: check if first line is a subheading
                    first = lines[0]
                    rest  = lines[1:]
                    # Subheading heuristic: short, starts with uppercase, doesn't end with sentence punctuation
                    looks_like_heading = (
                        len(first) <= 60
                        and first[0:1].isupper()
                        and not first.rstrip().endswith(('.', '?', '!', ';'))
                    )
                    if looks_like_heading:
                        return (
                            "<strong>" + _html.escape(first) + "</strong><br>"
                            + "<br>".join(_html.escape(ln) for ln in rest)
                        )
                    else:
                        return "<br>".join(_html.escape(ln) for ln in lines)

                # ── Split row stream into: spec rows, manufFor row, manufSub row, step rows ──
                spec_rows    = []     # [(param, observation), ...]
                spec_header  = ""     # "PRODUCT SPECIFICATION" if present
                manuf_for    = ""     # "MANUFACTURING PROCESS FOR <product>"
                manuf_sub    = ""     # "Manufacturing Process"
                step_rows    = []     # [(serial, content), ...]
                extra_blocks = []     # anything else → appended as raw <p>

                # First pass: partition
                mode = None  # None → waiting, 'spec' → in spec block, 'steps' → in steps block
                for row in html_rows:
                    if row is None:
                        continue
                    rt = _row_text(row).strip()
                    if not rt:
                        continue

                    if _is_spec_header(row):
                        spec_header = rt
                        mode = 'spec'
                        continue

                    if _is_proc_header(row):
                        # Distinguish "MANUFACTURING PROCESS FOR ..." vs plain "Manufacturing Process"
                        rt_lower = rt.lower()
                        if 'for' in rt_lower and rt_lower.index('for') < 30:
                            manuf_for = rt
                        else:
                            manuf_sub = rt
                        mode = 'steps'
                        continue

                    if mode == 'spec':
                        # Collect spec rows (drop empty leading cols)
                        trimmed = [c for c in row]  # copy
                        # Remove leading empty cells
                        while trimmed and not trimmed[0].strip():
                            trimmed.pop(0)
                        # Keep only first two non-empty cells (Parameter / Observation)
                        non_empty = [c for c in trimmed if c.strip()]
                        if len(non_empty) >= 2:
                            spec_rows.append((non_empty[0], non_empty[1]))
                        elif len(non_empty) == 1:
                            spec_rows.append((non_empty[0], ""))
                        continue

                    if mode == 'steps':
                        # A step row = col containing a serial number (e.g. "1", "2.", "10")
                        # followed by the step content
                        trimmed = list(row)
                        while trimmed and not trimmed[0].strip():
                            trimmed.pop(0)
                        non_empty = [c for c in trimmed if c.strip()]
                        if not non_empty:
                            continue
                        first = non_empty[0].strip()
                        m = _re_mp.match(r'^(\d+)\.?$', first)
                        if m and len(non_empty) >= 2:
                            # Serial + content
                            serial  = int(m.group(1))
                            # Content = longest non-serial cell
                            rest_cells = non_empty[1:]
                            content = max(rest_cells, key=len) if rest_cells else ""
                            step_rows.append((serial, content))
                        elif m and len(non_empty) == 1:
                            # Serial only, no content — skip
                            continue
                        else:
                            # Non-serial row in step area: probably a continuation
                            # or a subheading. Attach to previous step if exists.
                            longest = max(non_empty, key=len)
                            if step_rows:
                                prev_s, prev_c = step_rows[-1]
                                step_rows[-1] = (prev_s, (prev_c + "\n" + longest).strip() if prev_c else longest)
                            else:
                                extra_blocks.append(longest)
                        continue

                    # Pre-spec / pre-proc content
                    extra_blocks.append(rt)

                # ── Emit canonical HTML ──
                html_pieces = []

                # Any pre-content blocks (rare)
                for blk in extra_blocks:
                    html_pieces.append(f'<p style="font-size:12px;margin:2px 0">{_html.escape(blk)}</p>')

                # Spec table
                if spec_header or spec_rows:
                    parts = ['<table class="mp-spec" style="border-collapse:collapse;width:100%;font-size:11px;table-layout:fixed;margin:0 0 8px 0">']
                    parts.append('<colgroup><col style="width:35%"><col style="width:65%"></colgroup>')
                    if spec_header:
                        parts.append(
                            f'<tr><th colspan="2" style="border:1px solid #cbd5e1;padding:4px 6px;'
                            f'font-weight:700;background:#f1f5f9;text-align:center;font-size:11px">'
                            f'{_html.escape(spec_header)}</th></tr>'
                        )
                    # Column header row — first spec row is Parameters/Observation header if it
                    # matches, otherwise synthesise the header
                    data_rows = list(spec_rows)
                    first_is_header = False
                    if data_rows:
                        p0, o0 = data_rows[0]
                        if p0.lower().startswith(('parameter','paramerter')) or o0.lower().startswith('observation'):
                            first_is_header = True
                    if first_is_header:
                        p0, o0 = data_rows.pop(0)
                        parts.append(
                            f'<tr><th style="border:1px solid #cbd5e1;padding:4px 6px;font-weight:700;'
                            f'background:#f8fafc;text-align:center;font-size:10px">{_html.escape(p0)}</th>'
                            f'<th style="border:1px solid #cbd5e1;padding:4px 6px;font-weight:700;'
                            f'background:#f8fafc;text-align:center;font-size:10px">{_html.escape(o0)}</th></tr>'
                        )
                    else:
                        parts.append(
                            '<tr><th style="border:1px solid #cbd5e1;padding:4px 6px;font-weight:700;'
                            'background:#f8fafc;text-align:center;font-size:10px">Parameters</th>'
                            '<th style="border:1px solid #cbd5e1;padding:4px 6px;font-weight:700;'
                            'background:#f8fafc;text-align:center;font-size:10px">Observation (Result)</th></tr>'
                        )
                    for p, o in data_rows:
                        parts.append(
                            f'<tr><td style="border:1px solid #cbd5e1;padding:3px 6px;vertical-align:top;font-size:11px">'
                            f'{_html.escape(p)}</td>'
                            f'<td style="border:1px solid #cbd5e1;padding:3px 6px;vertical-align:top;font-size:11px">'
                            f'{_html.escape(o)}</td></tr>'
                        )
                    parts.append('</table>')
                    html_pieces.append(''.join(parts))

                # Steps table
                if manuf_for or manuf_sub or step_rows:
                    parts = ['<table class="mp-steps" style="border-collapse:collapse;width:100%;font-size:11px;table-layout:fixed;margin:0 0 6px 0">']
                    parts.append('<colgroup><col style="width:28px"><col></colgroup>')
                    if manuf_for:
                        parts.append(
                            f'<tr><th colspan="2" style="border:1px solid #cbd5e1;padding:4px 6px;'
                            f'font-weight:700;background:#f1f5f9;text-align:center;font-size:11px">'
                            f'{_html.escape(manuf_for)}</th></tr>'
                        )
                    if manuf_sub:
                        parts.append(
                            f'<tr><th colspan="2" style="border:1px solid #cbd5e1;padding:3px 6px;'
                            f'font-weight:700;background:#f8fafc;text-align:center;font-size:10px">'
                            f'{_html.escape(manuf_sub)}</th></tr>'
                        )
                    for serial, content in step_rows:
                        cell_html = _cell_to_html(content)
                        parts.append(
                            f'<tr>'
                            f'<td class="sr" style="border:1px solid #cbd5e1;padding:3px 4px;text-align:center;'
                            f'color:#64748b;font-size:10px;vertical-align:top">{serial}</td>'
                            f'<td style="border:1px solid #cbd5e1;padding:3px 6px;vertical-align:top;'
                            f'font-size:11px;line-height:1.5">{cell_html}</td>'
                            f'</tr>'
                        )
                    parts.append('</table>')
                    html_pieces.append(''.join(parts))

                manuf_html = ''.join(html_pieces)

                # ── Check batch exists AND whether manuf_process is already set ──
                existing = conn.execute(
                    "SELECT batch_name, manuf_process FROM procurement_formulations WHERE batch_name=%s LIMIT 1",
                    (batch_name,)
                ).fetchone()

                if not existing:
                    not_found += 1
                    results.append({
                        "sheet": sname, "batch_name": batch_name,
                        "status": "not_found",
                        "message": f"Batch '{batch_name}' not found in DB"
                    })
                    continue

                # ── Skip if manufacturing process already exists ──
                if existing["manuf_process"] and str(existing["manuf_process"]).strip():
                    skipped += 1
                    results.append({
                        "sheet": sname, "batch_name": batch_name,
                        "status": "skipped",
                        "message": "Already has manufacturing process — skipped"
                    })
                    continue

                # ── Save to DB ──
                conn.execute(
                    "UPDATE procurement_formulations SET manuf_process=%s WHERE batch_name=%s",
                    (manuf_html, batch_name)
                )
                conn.commit()
                updated += 1
                results.append({
                    "sheet": sname, "batch_name": batch_name,
                    "status": "updated", "message": "OK"
                })

            conn.close()
            wb.close()

        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500
        finally:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except:
                pass

        return jsonify({
            "status":    "ok",
            "updated":   updated,
            "skipped":   skipped,
            "not_found": not_found,
            "total":     updated + skipped + not_found,
            "results":   results
        })


    # ══════════════════════════════════════════════════════════════
    #  BRAND MANAGEMENT APIs
    # ══════════════════════════════════════════════════════════════

    @app.route("/api/procurement/brands", methods=["GET"])
    @procurement_required
    def api_brands_list():
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            brands = conn.execute(
                "SELECT id,name,color,text_color,created_at FROM procurement_brands ORDER BY name ASC"
            ).fetchall()
            conn.close()
            return jsonify({"status":"ok","brands":[dict(b) for b in brands]})
        except Exception as e:
            traceback.print_exc(); return jsonify({"status":"error","message":str(e)}),500

    @app.route("/api/procurement/brands", methods=["POST"])
    @procurement_required
    def api_brands_create():
        d    = request.get_json() or {}
        name = (d.get("name") or "").strip()
        color= (d.get("color") or "#6366f1").strip()
        text_color = (d.get("text_color") or "#ffffff").strip()
        if not name: return jsonify({"status":"error","message":"name required"}),400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            conn.execute(
                "INSERT INTO procurement_brands (name,color,text_color) VALUES (%s,%s,%s)",
                (name, color, text_color)
            )
            conn.commit()
            bid = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
            conn.close()
            return jsonify({"status":"ok","id":bid,"name":name,"color":color,"text_color":text_color})
        except Exception as e:
            traceback.print_exc(); return jsonify({"status":"error","message":str(e)}),500

    @app.route("/api/procurement/brands/bulk", methods=["POST"])
    @procurement_required
    def api_brands_bulk_create():
        """
        Create multiple brands at once. Pre-detects duplicates by name (case-insensitive,
        matching the existing UNIQUE constraint behaviour) so we never raise on insert.
        Body: {"brands":[{"name":..,"color":..,"text_color":..}, ...]}
        Returns: {"status":"ok","added":[{id,name,color,text_color},...],
                                 "skipped":[{"name":..,"reason":"duplicate"|"empty"|"too_long"}]}
        """
        d = request.get_json() or {}
        items = d.get("brands") or []
        if not isinstance(items, list):
            return jsonify({"status":"error","message":"brands must be a list"}), 400

        # Normalise + classify input rows BEFORE touching DB
        cleaned = []   # [(name, color, text_color)]
        skipped = []   # [{"name":.., "reason":..}]
        seen_lower = set()
        for it in items:
            if not isinstance(it, dict):
                continue
            name = (it.get("name") or "").strip()
            if not name:
                continue
            if len(name) > 200:
                skipped.append({"name": name[:60], "reason": "too_long"})
                continue
            color = (it.get("color") or "#6366f1").strip() or "#6366f1"
            text_color = (it.get("text_color") or "#ffffff").strip() or "#ffffff"
            low = name.lower()
            if low in seen_lower:
                skipped.append({"name": name, "reason": "duplicate"})
                continue
            seen_lower.add(low)
            cleaned.append((name, color, text_color))

        if not cleaned and not skipped:
            return jsonify({"status":"error","message":"no valid brands provided"}), 400

        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500

            # Pre-detect collisions against existing rows (case-insensitive)
            existing = conn.execute(
                "SELECT id, name FROM procurement_brands"
            ).fetchall() or []
            existing_lower = { (r["name"] or "").lower() for r in existing }

            to_insert = []
            for (nm, c, tc) in cleaned:
                if nm.lower() in existing_lower:
                    skipped.append({"name": nm, "reason": "duplicate"})
                else:
                    to_insert.append((nm, c, tc))

            added = []
            for (nm, c, tc) in to_insert:
                try:
                    conn.execute(
                        "INSERT INTO procurement_brands (name,color,text_color) VALUES (%s,%s,%s)",
                        (nm, c, tc)
                    )
                    bid = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
                    added.append({"id": bid, "name": nm, "color": c, "text_color": tc})
                except Exception as ie:
                    # Defensive: if a race or unforeseen constraint hits, treat as skipped
                    skipped.append({"name": nm, "reason": "duplicate"})

            conn.commit()
            conn.close()
            return jsonify({"status":"ok","added": added, "skipped": skipped})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/brands/import_excel", methods=["POST"])
    @procurement_required
    def api_brands_import_excel():
        """
        Import brands from an uploaded .xlsx file. Reads brand names from the
        first column (column A); empty rows end a contiguous block. Detects an
        optional header in row 1 if it looks like a label ('brand', 'name',
        'brand name'). Generates a random but deterministic-per-name pleasant
        colour and an auto-contrasting text colour. Pre-detects duplicates
        against existing rows and within the uploaded set, then inserts only
        the unique survivors.
        Returns: {"status":"ok","added":[{id,name,color,text_color},...],
                                 "skipped":[{"name":..,"reason":..}]}
        """
        import tempfile, uuid, hashlib, colorsys

        if "file" not in request.files:
            return jsonify({"status":"error","message":"No file uploaded"}), 400
        f = request.files["file"]
        if not f.filename.lower().endswith(".xlsx"):
            return jsonify({"status":"error","message":"Only .xlsx files accepted"}), 400

        tmp_dir  = tempfile.gettempdir()
        tmp_name = f"brand_imp_{uuid.uuid4().hex}.xlsx"
        tmp_path = os.path.join(tmp_dir, tmp_name)
        f.save(tmp_path)

        # Helpers --------------------------------------------------------------
        def _color_for(name):
            """Deterministic-per-name pleasant HSL colour + contrasting text."""
            # Seed hue from md5 of name → hue in [0, 360)
            h_int = int(hashlib.md5(name.lower().encode("utf-8")).hexdigest()[:8], 16)
            hue   = (h_int % 360) / 360.0
            sat   = 0.62                    # 62% saturation — vivid but not neon
            light = 0.50                    # 50% lightness  — mid-tone, readable
            r, g, b = colorsys.hls_to_rgb(hue, light, sat)
            R, G, B = int(round(r * 255)), int(round(g * 255)), int(round(b * 255))
            bg = "#{:02x}{:02x}{:02x}".format(R, G, B)
            # Perceptual luminance → choose white or black for text
            lum = (0.2126 * R + 0.7152 * G + 0.0722 * B) / 255.0
            tc  = "#ffffff" if lum < 0.55 else "#111111"
            return bg, tc

        try:
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]                      # first sheet only

            # Collect column A names. Skip first row only if it looks like a header.
            raw_names = []
            for ri, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
                v = row[0] if row else None
                if v is None:
                    # treat as terminator only if we've already collected something;
                    # keep scanning if file has a blank row above the data
                    if raw_names:
                        break
                    else:
                        continue
                s = str(v).strip()
                if not s:
                    if raw_names:
                        break
                    else:
                        continue
                if ri == 1 and s.lower() in ("brand", "brands", "name", "brand name", "brand_name"):
                    continue   # skip header
                raw_names.append(s)

            if not raw_names:
                try: wb.close()
                except Exception: pass
                return jsonify({"status":"error",
                                "message":"No brand names found in column A of the first sheet."}), 400

            # Normalise + classify -------------------------------------------
            cleaned = []     # list[(name, color, text_color)]
            skipped = []     # list[{"name":.., "reason":..}]
            seen_lower = set()
            for nm in raw_names:
                if len(nm) > 200:
                    skipped.append({"name": nm[:60], "reason": "too_long"})
                    continue
                low = nm.lower()
                if low in seen_lower:
                    skipped.append({"name": nm, "reason": "duplicate"})
                    continue
                seen_lower.add(low)
                bg, tc = _color_for(nm)
                cleaned.append((nm, bg, tc))

            try: wb.close()
            except Exception: pass

            # Insert with pre-detected DB collisions -------------------------
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500

            existing = conn.execute("SELECT name FROM procurement_brands").fetchall() or []
            existing_lower = { (r["name"] or "").lower() for r in existing }

            to_insert = []
            for (nm, c, tc) in cleaned:
                if nm.lower() in existing_lower:
                    skipped.append({"name": nm, "reason": "duplicate"})
                else:
                    to_insert.append((nm, c, tc))

            added = []
            for (nm, c, tc) in to_insert:
                try:
                    conn.execute(
                        "INSERT INTO procurement_brands (name,color,text_color) VALUES (%s,%s,%s)",
                        (nm, c, tc)
                    )
                    bid = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
                    added.append({"id": bid, "name": nm, "color": c, "text_color": tc})
                except Exception:
                    skipped.append({"name": nm, "reason": "duplicate"})

            conn.commit()
            conn.close()
            return jsonify({"status":"ok","added": added, "skipped": skipped})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500
        finally:
            try: os.remove(tmp_path)
            except Exception: pass

    @app.route("/api/procurement/brands/<int:brand_id>", methods=["PUT"])
    @procurement_required
    def api_brands_update(brand_id):
        d    = request.get_json() or {}
        name = (d.get("name") or "").strip()
        color= (d.get("color") or "").strip()
        text_color = (d.get("text_color") or "#ffffff").strip()
        if not name: return jsonify({"status":"error","message":"name required"}),400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            conn.execute(
                "UPDATE procurement_brands SET name=%s,color=%s,text_color=%s WHERE id=%s",
                (name, color or "#6366f1", text_color, brand_id)
            )
            conn.commit(); conn.close()
            return jsonify({"status":"ok"})
        except Exception as e:
            traceback.print_exc(); return jsonify({"status":"error","message":str(e)}),500

    @app.route("/api/procurement/brands/<int:brand_id>", methods=["DELETE"])
    @procurement_required
    def api_brands_delete(brand_id):
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            # Unlink formulations first
            conn.execute(
                "UPDATE procurement_formulations SET brand_id=NULL WHERE brand_id=%s",
                (brand_id,)
            )
            conn.execute("DELETE FROM procurement_brands WHERE id=%s",(brand_id,))
            conn.commit(); conn.close()
            return jsonify({"status":"ok"})
        except Exception as e:
            traceback.print_exc(); return jsonify({"status":"error","message":str(e)}),500

    @app.route("/api/procurement/formulations/set_brand", methods=["POST"])
    @procurement_required
    def api_formulations_set_brand():
        """Assign brand_id to one or more batch_names."""
        d          = request.get_json() or {}
        batch_names= d.get("batch_names") or []
        brand_id   = d.get("brand_id")   # None to clear
        if not batch_names: return jsonify({"status":"error","message":"batch_names required"}),400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            for bn in batch_names:
                conn.execute(
                    "UPDATE procurement_formulations SET brand_id=%s WHERE batch_name=%s",
                    (brand_id, bn)
                )
            conn.commit(); conn.close()
            return jsonify({"status":"ok","updated":len(batch_names)})
        except Exception as e:
            traceback.print_exc(); return jsonify({"status":"error","message":str(e)}),500

    @app.route("/api/procurement/brands/report")
    @procurement_required
    def api_brands_report():
        """Returns each brand with count of distinct batch_names and the batch list."""
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            brands = conn.execute(
                "SELECT id,name,color,text_color FROM procurement_brands ORDER BY name ASC"
            ).fetchall()
            rows = conn.execute("""
                SELECT DISTINCT batch_name, brand_id, source_batch_name
                FROM   procurement_formulations
                ORDER  BY batch_name ASC
            """).fetchall()
            conn.close()
            brand_map = {b["id"]: {"id":b["id"],"name":b["name"],"color":b["color"],"text_color":b.get("text_color","#ffffff"),"batches":[]} for b in brands}
            unbranded = []
            for r in rows:
                if r["brand_id"] and r["brand_id"] in brand_map:
                    if r["batch_name"] not in brand_map[r["brand_id"]]["batches"]:
                        brand_map[r["brand_id"]]["batches"].append(r["batch_name"])
                else:
                    if r["batch_name"] not in unbranded:
                        unbranded.append(r["batch_name"])
            result = list(brand_map.values())
            result.append({"id":None,"name":"Unbranded","color":"#94a3b8","batches":unbranded})
            return jsonify({"status":"ok","report":result})
        except Exception as e:
            traceback.print_exc(); return jsonify({"status":"error","message":str(e)}),500

    @app.route("/api/procurement/formulations/create_manual", methods=["POST"])
    @procurement_required
    def api_formulations_create_manual():
        """
        Create a formulation manually from a list of ingredients.
        Body: { batch_name, product_code, batch_size, batch_date, brand_id,
                ingredients: [{material_name, supplier_name, concentration, qty_kg}] }
        """
        d          = request.get_json() or {}
        batch_name = (d.get("batch_name") or "").strip()
        if not batch_name:
            return jsonify({"status":"error","message":"batch_name required"}),400
        ingredients = d.get("ingredients") or []
        if not ingredients:
            return jsonify({"status":"error","message":"At least one ingredient required"}),400
        product_code = (d.get("product_code") or "").strip() or None
        batch_size   = (d.get("batch_size")   or "").strip() or None
        batch_date   = (d.get("batch_date")   or "").strip() or None
        brand_id     = d.get("brand_id") or None
        uid          = session.get("UID","")
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            # Check name not already taken
            exists = conn.execute(
                "SELECT 1 FROM procurement_formulations WHERE batch_name=%s LIMIT 1",
                (batch_name,)
            ).fetchone()
            if exists:
                conn.close()
                return jsonify({"status":"error","message":f"Batch '{batch_name}' already exists"}),409
            count = 0
            for ing in ingredients:
                mat  = (str(ing.get("material_name",""))).strip()
                if not mat: continue
                conn.execute("""
                    INSERT INTO procurement_formulations
                        (batch_name,product_code,material_name,supplier_name,
                         concentration,qty_kg,batch_size,batch_date,imported_by,brand_id)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    batch_name, product_code, mat,
                    (ing.get("supplier_name") or "").strip() or None,
                    (str(ing.get("concentration",""))).strip() or None,
                    (str(ing.get("qty_kg",""))).strip() or None,
                    batch_size, batch_date, uid, brand_id
                ))
                count += 1
            conn.commit(); conn.close()
            return jsonify({"status":"ok","batch_name":batch_name,"inserted":count})
        except Exception as e:
            traceback.print_exc(); return jsonify({"status":"error","message":str(e)}),500

    # ── API: list batches (summary) + full detail ───────────────
    @app.route("/api/procurement/formulations/list")
    @procurement_required
    def api_formulations_list():
        try:
            conn=sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            rows=conn.execute("""
                SELECT id,batch_name,product_code,material_name,supplier_name,
                       concentration,qty_kg,
                       batch_size,batch_date,num_batches,
                       manuf_process,imported_at,imported_by,source_batch_name,brand_id,
                       COALESCE(is_active,1) as is_active
                FROM   procurement_formulations
                ORDER  BY batch_name ASC,id ASC
            """).fetchall()
            conn.close()
            all_rows=[dict(r) for r in rows]
            seen={}; batch_list=[]
            for r in all_rows:
                bn=r["batch_name"]
                if bn not in seen:
                    seen[bn]={"batch_name":bn,"product_code":r.get("product_code"),
                              "batch_size":r.get("batch_size"),
                              "batch_date":r.get("batch_date"),"num_batches":r.get("num_batches"),
                              "imported_at":r.get("imported_at"),"source_batch_name":r.get("source_batch_name"),
                              "manuf_process":r.get("manuf_process"),
                              "brand_id":r.get("brand_id"),
                              "is_active":int(r.get("is_active",1)),"item_count":0}
                    batch_list.append(seen[bn])
                seen[bn]["item_count"]+=1
            # ── Fill manuf_process from source for linked batches that have none ──
            mp_by_name = {b["batch_name"]: b.get("manuf_process") for b in batch_list}
            for b in batch_list:
                if not b.get("manuf_process") and b.get("source_batch_name"):
                    src_mp = mp_by_name.get(b["source_batch_name"])
                    if src_mp:
                        import re as _re_list, html as _html_list
                        escaped = _html_list.escape(b["batch_name"])
                        b["manuf_process"] = _re_list.sub(
                            r'<!-- BATCH_TITLE -->.*?<!-- /BATCH_TITLE -->',
                            f'<!-- BATCH_TITLE -->{escaped}<!-- /BATCH_TITLE -->',
                            src_mp, flags=_re_list.DOTALL
                        )

            for i,b in enumerate(batch_list,start=1): b["sr_no"]=i
            return jsonify({"status":"ok","batches":batch_list,"detail":all_rows,
                "batch_names":[b["batch_name"] for b in batch_list],"count":len(batch_list)})
        except Exception as e:
            traceback.print_exc(); return jsonify({"status":"error","message":str(e)}),500

    # ── API: delete a batch ───────────────────────────────────────
    @app.route("/api/procurement/formulations/delete_batch", methods=["POST"])
    @procurement_required
    def api_formulations_delete_batch():
        d=request.get_json() or {}
        batch=(d.get("batch_name") or "").strip()
        if not batch: return jsonify({"status":"error","message":"batch_name required"}),400
        try:
            conn=sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            conn.execute("DELETE FROM procurement_formulations WHERE batch_name=%s",(batch,))
            conn.commit(); conn.close()
            return jsonify({"status":"ok"})
        except Exception as e:
            traceback.print_exc(); return jsonify({"status":"error","message":str(e)}),500

    # ── API: toggle active / inactive for a batch ─────────────────
    @app.route("/api/procurement/formulations/toggle_active", methods=["POST"])
    @procurement_required
    def api_formulations_toggle_active():
        d = request.get_json() or {}
        batch = (d.get("batch_name") or "").strip()
        if not batch:
            return jsonify({"status":"error","message":"batch_name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500
            row = conn.execute(
                "SELECT is_active FROM procurement_formulations WHERE batch_name=%s LIMIT 1",
                (batch,)
            ).fetchone()
            if not row:
                conn.close()
                return jsonify({"status":"error","message":"Batch not found"}), 404
            new_state = 0 if row["is_active"] else 1
            conn.execute(
                "UPDATE procurement_formulations SET is_active=%s WHERE batch_name=%s",
                (new_state, batch)
            )
            conn.commit(); conn.close()
            return jsonify({"status":"ok","batch_name":batch,"is_active":new_state})
        except Exception as e:
            traceback.print_exc(); return jsonify({"status":"error","message":str(e)}), 500

    # ─────────────────────────────────────────────────────────────────────────
    # SUPPLIER MASTER  —  CRUD
    # ─────────────────────────────────────────────────────────────────────────

    # ── Supplier Types ────────────────────────────────────────────

    @app.route("/api/supplier_types", methods=["GET"])
    @procurement_required
    def api_supplier_types_list():
        """Return all supplier types."""
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            rows = conn.execute(
                "SELECT id, type_name, sort_order FROM supplier_type ORDER BY sort_order ASC, type_name ASC"
            ).fetchall()
            conn.close()
            return jsonify({"status": "ok", "types": [dict(r) for r in rows]})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/supplier_types/save", methods=["POST"])
    @procurement_required
    def api_supplier_types_save():
        """Insert or update a supplier type."""
        d = request.get_json() or {}
        type_name = (d.get("type_name") or "").strip()
        if not type_name:
            return jsonify({"status": "error", "message": "type_name is required"}), 400
        type_id   = d.get("id")
        sort_order = int(d.get("sort_order") or 0)
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            if type_id:
                conn.execute(
                    "UPDATE supplier_type SET type_name=%s, sort_order=%s WHERE id=%s",
                    (type_name, sort_order, type_id)
                )
                conn.commit(); conn.close()
                return jsonify({"status": "ok", "action": "updated", "id": type_id})
            else:
                conn.execute(
                    "INSERT INTO supplier_type (type_name, sort_order) VALUES (%s, %s)",
                    (type_name, sort_order)
                )
                conn.commit()
                new_id = conn.execute(
                    "SELECT id FROM supplier_type WHERE type_name=%s", (type_name,)
                ).fetchone()["id"]
                conn.close()
                return jsonify({"status": "ok", "action": "created", "id": new_id})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/supplier_types/delete", methods=["POST"])
    @procurement_required
    def api_supplier_types_delete():
        """Delete a supplier type by id. Clears the FK on suppliers first."""
        d = request.get_json() or {}
        type_id = d.get("id")
        if not type_id:
            return jsonify({"status": "error", "message": "id is required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            conn.execute(
                "UPDATE procurement_suppliers SET supplier_type_id=NULL WHERE supplier_type_id=%s",
                (type_id,)
            )
            conn.execute("DELETE FROM supplier_type WHERE id=%s", (type_id,))
            conn.commit(); conn.close()
            return jsonify({"status": "ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    # ── /Supplier Types ───────────────────────────────────────────

    @app.route("/api/procurement/suppliers/bulk_assign_type", methods=["POST"])
    @procurement_required
    def api_suppliers_bulk_assign_type():
        """Bulk-assign a supplier_type_id to multiple suppliers."""
        d = request.get_json() or {}
        supplier_ids = d.get("supplier_ids") or []
        type_id = d.get("supplier_type_id")
        if not supplier_ids:
            return jsonify({"status": "error", "message": "No suppliers selected"}), 400
        if not type_id:
            return jsonify({"status": "error", "message": "No type selected"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            ph = ",".join(["%s"] * len(supplier_ids))
            conn.execute(
                f"UPDATE procurement_suppliers SET supplier_type_id=%s WHERE id IN ({ph})",
                [int(type_id)] + [int(i) for i in supplier_ids]
            )
            conn.commit()
            updated = conn.execute(
                f"SELECT COUNT(*) AS c FROM procurement_suppliers WHERE id IN ({ph})",
                [int(i) for i in supplier_ids]
            ).fetchone()["c"]
            conn.close()
            return jsonify({"status": "ok", "updated": updated})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/suppliers", methods=["GET"])
    @procurement_required
    def api_suppliers_list():
        """Return all suppliers with material counts joined from procurement_materials."""
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            suppliers = conn.execute("""
                SELECT s.*, st.type_name AS supplier_type_name
                FROM procurement_suppliers s
                LEFT JOIN supplier_type st ON st.id = s.supplier_type_id
                ORDER BY s.supplier_name ASC
            """).fetchall()
            rows = [dict(r) for r in suppliers]
            mat_counts = conn.execute("""
                SELECT LOWER(TRIM(supplier_name)) AS sup_key, COUNT(*) AS mat_count
                FROM procurement_materials
                WHERE supplier_name IS NOT NULL AND supplier_name != ''
                GROUP BY LOWER(TRIM(supplier_name))
            """).fetchall()
            count_map = {r["sup_key"]: r["mat_count"] for r in mat_counts}
            for r in rows:
                r["mat_count"] = count_map.get((r["supplier_name"] or "").strip().lower(), 0)
                for f in ("created_at", "updated_at"):
                    if r.get(f): r[f] = str(r[f])
            conn.close()
            return jsonify({"status": "ok", "suppliers": rows, "count": len(rows)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.route("/api/procurement/suppliers/save", methods=["POST"])
    @procurement_required
    def api_suppliers_save():
        """Insert or update a supplier record."""
        d = request.get_json() or {}
        name = (d.get("supplier_name") or "").strip()
        if not name:
            return jsonify({"status": "error", "message": "supplier_name is required"}), 400
        supplier_id = d.get("id")
        uid = session.get("UID", "")
        supplier_code = (d.get("supplier_code") or "").strip() or None
        def _i(k): return int(d[k]) if d.get(k) not in (None, "", "null") else None
        def _f(k): return float(d[k]) if d.get(k) not in (None, "", "null") else None
        def _s(k, default=None): return (d.get(k) or "").strip() or default
        fields = {
            "supplier_name":  name,
            "supplier_code":  supplier_code,
            "contact_person": _s("contact_person"),
            "phone":          _s("phone"),
            "email":          _s("email"),
            "address":        _s("address"),
            "gst_number":     (_s("gst_number") or "").upper() or None,
            "pan_number":     (_s("pan_number") or "").upper() or None,
            "payment_terms":  _s("payment_terms"),
            "payment_type":   _s("payment_type"),
            "credit_days":    _i("credit_days"),
            "currency":       _s("currency", "INR"),
            "lead_time_days": _i("lead_time_days"),
            "moq":            _f("moq"),
            "rating":         _i("rating"),
            "status":          _s("status", "active"),
            "updated_by":      uid,
            "declaration_id":  _i("declaration_id"),
            "supplier_type_id":_i("supplier_type_id"),
        }
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            if supplier_id:
                # For existing suppliers: if code is blank or old SUP- format, re-assign HCPRMS
                # but only if it wasn't manually set to a valid HCPRMS code
                existing = conn.execute(
                    "SELECT supplier_code FROM procurement_suppliers WHERE id=%s", (supplier_id,)
                ).fetchone()
                old_code = (existing["supplier_code"] or "") if existing else ""
                import re as _re_sup
                if fields["supplier_code"] and _re_sup.match(r"^HCPRMS-\d+$", fields["supplier_code"]):
                    pass  # keep what was sent — valid HCPRMS code
                elif not fields["supplier_code"] or not _re_sup.match(r"^HCPRMS-\d+$", old_code):
                    # Need to assign a new HCPRMS code
                    conn.execute("SELECT GET_LOCK('hcprms_code_lock',10) AS lk")
                    try:
                        rows = conn.execute(
                            "SELECT supplier_code FROM procurement_suppliers WHERE supplier_code LIKE 'HCPRMS-%%'"
                        ).fetchall()
                        max_n = 0
                        for r in rows:
                            m = _re_sup.search(r"HCPRMS-(\d+)", r["supplier_code"] or "")
                            if m: max_n = max(max_n, int(m.group(1)))
                        fields["supplier_code"] = f"HCPRMS-{max_n+1:04d}"
                    finally:
                        conn.execute("SELECT RELEASE_LOCK('hcprms_code_lock')")
                else:
                    fields["supplier_code"] = old_code  # keep existing valid HCPRMS code
                set_clause = ", ".join(f"{k}=%s" for k in fields)
                conn.execute(
                    f"UPDATE procurement_suppliers SET {set_clause} WHERE id=%s",
                    list(fields.values()) + [supplier_id]
                )
                conn.commit(); conn.close()
                return jsonify({"status": "ok", "action": "updated", "id": supplier_id,
                                "supplier_code": fields["supplier_code"]})
            else:
                # New supplier — always assign fresh HCPRMS code atomically
                import re as _re_sup2
                conn.execute("SELECT GET_LOCK('hcprms_code_lock',10) AS lk")
                try:
                    rows = conn.execute(
                        "SELECT supplier_code FROM procurement_suppliers WHERE supplier_code LIKE 'HCPRMS-%%'"
                    ).fetchall()
                    max_n = 0
                    for r in rows:
                        m = _re_sup2.search(r"HCPRMS-(\d+)", r["supplier_code"] or "")
                        if m: max_n = max(max_n, int(m.group(1)))
                    fields["supplier_code"] = f"HCPRMS-{max_n+1:04d}"
                    cols = ", ".join(fields.keys())
                    ph   = ", ".join(["%s"] * len(fields))
                    conn.execute(f"INSERT INTO procurement_suppliers ({cols}) VALUES ({ph})", list(fields.values()))
                    conn.commit()
                finally:
                    conn.execute("SELECT RELEASE_LOCK('hcprms_code_lock')")
                new_id = conn.execute(
                    "SELECT id FROM procurement_suppliers WHERE supplier_name=%s", (name,)
                ).fetchone()["id"]
                conn.close()
                return jsonify({"status": "ok", "action": "created", "id": new_id,
                                "supplier_code": fields["supplier_code"]})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.route("/api/procurement/suppliers/delete", methods=["POST"])
    @procurement_required
    def api_suppliers_delete():
        """Delete a supplier by id."""
        d = request.get_json() or {}
        supplier_id = d.get("id")
        if not supplier_id:
            return jsonify({"status": "error", "message": "id is required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            conn.execute("DELETE FROM procurement_suppliers WHERE id=%s", (supplier_id,))
            conn.commit(); conn.close()
            return jsonify({"status": "ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.route("/api/procurement/suppliers/get", methods=["GET"])
    @procurement_required
    def api_suppliers_get():
        """Get a single supplier by id."""
        supplier_id = request.args.get("id")
        if not supplier_id:
            return jsonify({"status": "error", "message": "id is required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            row = conn.execute(
                "SELECT * FROM procurement_suppliers WHERE id=%s", (supplier_id,)
            ).fetchone()
            conn.close()
            if not row:
                return jsonify({"status": "error", "message": "Supplier not found"}), 404
            r = dict(row)
            for f in ("created_at", "updated_at"):
                if r.get(f): r[f] = str(r[f])
            return jsonify({"status": "ok", "supplier": r})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.route("/api/procurement/suppliers/sync_from_materials", methods=["POST"])
    @procurement_required
    def api_suppliers_sync():
        """
        Seeds procurement_suppliers from unique supplier_name values in
        procurement_materials. Only inserts rows that don't already exist.
        Call once after deployment to migrate existing supplier names.
        """
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            mat_sups = conn.execute("""
                SELECT DISTINCT TRIM(supplier_name) AS sname,
                       MIN(lead_time_days) AS ltd
                FROM procurement_materials
                WHERE supplier_name IS NOT NULL AND TRIM(supplier_name) != ''
                GROUP BY TRIM(supplier_name)
            """).fetchall()
            existing = {r["supplier_name"].strip().lower()
                        for r in conn.execute("SELECT supplier_name FROM procurement_suppliers").fetchall()}
            counter = conn.execute("SELECT COUNT(*) AS c FROM procurement_suppliers").fetchone()["c"]
            inserted = 0
            for r in mat_sups:
                sname = (r["sname"] or "").strip()
                if not sname or sname.lower() in existing:
                    continue
                counter += 1
                conn.execute("""
                    INSERT INTO procurement_suppliers
                        (supplier_name, supplier_code, lead_time_days, status, currency)
                    VALUES (%s, %s, %s, 'active', 'INR')
                """, (sname, f"SUP-{counter:04d}", r["ltd"]))
                existing.add(sname.lower())
                inserted += 1
            conn.commit(); conn.close()
            return jsonify({"status": "ok", "inserted": inserted})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    # ══════════════════════════════════════════════════════════════════════════
    # MATERIAL GROUPS
    # ══════════════════════════════════════════════════════════════════════════
    # MATERIAL TYPES
    # ══════════════════════════════════════════════════════════════════════════

    @app.route("/api/procurement/material_types", methods=["GET"])
    @procurement_required
    def api_material_types_list():
        try:
            conn = sampling_portal.get_db_connection()
            rows = conn.execute(
                "SELECT * FROM procurement_material_types ORDER BY sort_order, type_name"
            ).fetchall()
            conn.close()
            return jsonify({"status": "ok", "types": [dict(r) for r in rows]})
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/material_types/save", methods=["POST"])
    @procurement_required
    def api_material_types_save():
        d    = request.get_json() or {}
        tid  = d.get("id")
        name = (d.get("type_name") or "").strip()
        abbr = (d.get("abbreviation") or "").strip().upper() or None
        color= (d.get("color") or "").strip() or None
        desc = (d.get("description") or "").strip() or None
        if not name:
            return jsonify({"status": "error", "message": "type_name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if tid:
                conn.execute(
                    "UPDATE procurement_material_types SET type_name=%s, abbreviation=%s, color=%s, description=%s WHERE id=%s",
                    (name, abbr, color, desc, tid)
                )
            else:
                conn.execute(
                    "INSERT INTO procurement_material_types (type_name, abbreviation, color, description, created_by) VALUES (%s,%s,%s,%s,%s)",
                    (name, abbr, color, desc, session.get("UID", ""))
                )
                tid = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
            conn.commit(); conn.close()
            return jsonify({"status": "ok", "id": tid, "type_name": name, "abbreviation": abbr})
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

    # ══════════════════════════════════════════════════════════════════════════

    @app.route("/api/procurement/material_groups", methods=["GET"])
    @procurement_required
    def api_material_groups_list():
        """Return all material groups with material counts. Father Group always first."""
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            groups = conn.execute(
                "SELECT * FROM procurement_material_groups ORDER BY id ASC"
            ).fetchall()
            counts = conn.execute("""
                SELECT group_id, COUNT(*) AS cnt
                FROM procurement_materials
                WHERE group_id IS NOT NULL
                GROUP BY group_id
            """).fetchall()
            count_map = {r["group_id"]: r["cnt"] for r in counts}
            result = []
            for g in groups:
                d = dict(g)
                d["mat_count"] = count_map.get(d["id"], 0)
                for f in ("created_at","updated_at"):
                    if d.get(f): d[f] = str(d[f])
                result.append(d)
            conn.close()
            return jsonify({"status":"ok","groups":result})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}),500

    @app.route("/api/procurement/material_groups/save", methods=["POST"])
    @procurement_required
    def api_material_groups_save():
        """Create or update a material group. Cannot rename/modify Father Group (id=1)."""
        d = request.get_json() or {}
        gid  = d.get("id")
        name = (d.get("group_name") or "").strip()
        desc = (d.get("description") or "").strip() or None
        uid  = session.get("UID","")
        if not name:
            return jsonify({"status":"error","message":"group_name required"}),400
        if name.lower() == "father group" and not gid:
            return jsonify({"status":"error","message":"Cannot create a group named 'Father Group'"}),400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            if gid:
                if int(gid) == 1:
                    return jsonify({"status":"error","message":"Cannot modify Father Group"}),400
                conn.execute(
                    "UPDATE procurement_material_groups SET group_name=%s, description=%s WHERE id=%s",
                    (name, desc, gid)
                )
            else:
                conn.execute(
                    "INSERT INTO procurement_material_groups (group_name, description, parent_id, created_by) VALUES (%s,%s,1,%s)",
                    (name, desc, uid)
                )
                gid = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
            conn.commit(); conn.close()
            return jsonify({"status":"ok","id":gid,"group_name":name})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}),500

    @app.route("/api/procurement/material_groups/delete", methods=["POST"])
    @procurement_required
    def api_material_groups_delete():
        """Delete a group. Blocked if materials are assigned to it. Father Group cannot be deleted."""
        d   = request.get_json() or {}
        gid = d.get("id")
        if not gid:
            return jsonify({"status":"error","message":"id required"}),400
        if int(gid) == 1:
            return jsonify({"status":"error","message":"Cannot delete Father Group"}),400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            count = conn.execute(
                "SELECT COUNT(*) AS n FROM procurement_materials WHERE group_id=%s", (gid,)
            ).fetchone()["n"]
            if count:
                conn.close()
                return jsonify({"status":"error","message":f"{count} material(s) assigned to this group — reassign them first"}),400
            conn.execute("DELETE FROM procurement_material_groups WHERE id=%s AND id!=1", (gid,))
            conn.commit(); conn.close()
            return jsonify({"status":"ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}),500


    @app.route("/api/procurement/material_groups/bulk_assign", methods=["POST"])
    @procurement_required
    def api_material_groups_bulk_assign():
        """Assign (or clear) a group_id for a list of material_names in one shot."""
        d     = request.get_json() or {}
        gid   = d.get("group_id")        # None / "" = clear group
        names = d.get("material_names") or []
        if not names:
            return jsonify({"status": "error", "message": "material_names required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            if gid:
                exists = conn.execute(
                    "SELECT id FROM procurement_material_groups WHERE id=%s AND id!=1", (gid,)
                ).fetchone()
                if not exists:
                    conn.close()
                    return jsonify({"status": "error", "message": "Invalid group id"}), 400
                gid_val = int(gid)
            else:
                gid_val = None
            fmt = ",".join(["%s"] * len(names))
            conn.execute(
                f"UPDATE procurement_materials SET group_id=%s WHERE material_name IN ({fmt})",
                [gid_val] + list(names)
            )
            conn.commit()
            conn.close()
            return jsonify({"status": "ok", "updated": len(names)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/bulk_assign_gst", methods=["POST"])
    @procurement_required
    def api_bulk_assign_gst():
        """Assign a gst_rate to a list of material_names in one shot."""
        d     = request.get_json() or {}
        rate  = d.get("gst_rate")
        names = d.get("material_names") or []
        if not names:
            return jsonify({"status": "error", "message": "material_names required"}), 400
        if rate is None or str(rate).strip() == "":
            return jsonify({"status": "error", "message": "gst_rate required"}), 400
        try:
            rate_val = float(str(rate).strip())
        except (ValueError, TypeError):
            return jsonify({"status": "error", "message": "Invalid gst_rate value"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            fmt = ",".join(["%s"] * len(names))
            conn.execute(
                f"UPDATE procurement_materials SET gst_rate=%s WHERE material_name IN ({fmt})",
                [rate_val] + list(names)
            )
            conn.commit()
            conn.close()
            return jsonify({"status": "ok", "updated": len(names)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/bulk_assign_uom", methods=["POST"])
    @procurement_required
    def api_bulk_assign_uom():
        """Assign a UOM to a list of material_names in one shot."""
        d     = request.get_json() or {}
        uom   = (d.get("uom") or "").strip().upper()
        names = d.get("material_names") or []
        if not names:
            return jsonify({"status": "error", "message": "material_names required"}), 400
        if not uom:
            return jsonify({"status": "error", "message": "uom required"}), 400
        valid_uoms = {"KG", "G", "L", "ML", "PCS", "MT", "TON", "BOX", "PKT"}
        if uom not in valid_uoms:
            return jsonify({"status": "error", "message": f"Invalid UOM. Must be one of: {', '.join(sorted(valid_uoms))}"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            fmt = ",".join(["%s"] * len(names))
            conn.execute(
                f"UPDATE procurement_materials SET uom=%s WHERE material_name IN ({fmt})",
                [uom] + list(names)
            )
            conn.commit()
            conn.close()
            return jsonify({"status": "ok", "updated": len(names)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/bulk_assign_material_type", methods=["POST"])
    @procurement_required
    def api_bulk_assign_material_type():
        """Assign a material_type_id to a list of material_names in one shot."""
        d       = request.get_json() or {}
        type_id = d.get("material_type_id")
        names   = d.get("material_names") or []
        if not names:
            return jsonify({"status": "error", "message": "material_names required"}), 400
        if not type_id:
            return jsonify({"status": "error", "message": "material_type_id required"}), 400
        try:
            type_id = int(type_id)
        except (ValueError, TypeError):
            return jsonify({"status": "error", "message": "Invalid material_type_id"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            fmt = ",".join(["%s"] * len(names))
            conn.execute(
                f"UPDATE procurement_materials SET material_type_id=%s WHERE material_name IN ({fmt})",
                [type_id] + list(names)
            )
            conn.commit()
            conn.close()
            return jsonify({"status": "ok", "updated": len(names)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    # ══════════════════════════════════════════════════════════════════════════
    # SUPPLIER CODE — HCPRMS format migrate + next-code API
    # ══════════════════════════════════════════════════════════════════════════

    @app.route("/api/procurement/suppliers/next_code", methods=["GET"])
    @procurement_required
    def api_suppliers_next_code():
        """
        Returns the next available HCPRMS-NNNN code using GET_LOCK for concurrency safety.
        Does NOT assign it — assignment happens on save.
        """
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            conn.execute("SELECT GET_LOCK('hcprms_code_lock',10) AS lk")
            try:
                rows = conn.execute(
                    "SELECT supplier_code FROM procurement_suppliers WHERE supplier_code LIKE 'HCPRMS-%%'"
                ).fetchall()
                max_n = 0
                import re as _re2
                for r in rows:
                    m = _re2.search(r"HCPRMS-(\d+)", r["supplier_code"] or "")
                    if m: max_n = max(max_n, int(m.group(1)))
                next_code = f"HCPRMS-{max_n+1:04d}"
            finally:
                conn.execute("SELECT RELEASE_LOCK('hcprms_code_lock')")
            conn.close()
            return jsonify({"status":"ok","next_code":next_code})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}),500

    @app.route("/api/procurement/suppliers/migrate_codes", methods=["POST"])
    @procurement_required
    def api_suppliers_migrate_codes():
        """
        Admin-only: Renumber ALL supplier codes to HCPRMS-NNNN format (sorted by existing code/name).
        Old non-HCPRMS codes are replaced. Suppliers without a code get assigned one.
        Uses GET_LOCK to prevent concurrent runs.
        """
        if (session.get("User_Type","") or "").lower() != "admin":
            return jsonify({"status":"error","message":"Admin only"}),403
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            conn.execute("SELECT GET_LOCK('hcprms_migrate_lock',30) AS lk")
            try:
                sups = conn.execute(
                    "SELECT id, supplier_name, supplier_code FROM procurement_suppliers ORDER BY id ASC"
                ).fetchall()
                updated = 0
                import re as _re3
                counter = 0
                for sup in sups:
                    counter += 1
                    new_code = f"HCPRMS-{counter:04d}"
                    old_code = sup["supplier_code"] or ""
                    if old_code != new_code:
                        conn.execute(
                            "UPDATE procurement_suppliers SET supplier_code=%s WHERE id=%s",
                            (new_code, sup["id"])
                        )
                        updated += 1
                conn.commit()
            finally:
                conn.execute("SELECT RELEASE_LOCK('hcprms_migrate_lock')")
            conn.close()
            return jsonify({"status":"ok","updated":updated,"total":len(sups)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}),500

    # ══════════════════════════════════════════════════════════════════════════
    # WHOLE SUPPLIER LIST — share (admin only)
    # ══════════════════════════════════════════════════════════════════════════

    @app.route("/api/procurement/suppliers/list_for_share", methods=["GET"])
    @procurement_required
    def api_suppliers_list_for_share():
        """
        Returns all active suppliers with full details for WhatsApp/Email sharing.
        Admin-only.
        """
        if (session.get("User_Type","") or "").lower() != "admin":
            return jsonify({"status":"error","message":"Admin only"}),403
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}),500
            rows = conn.execute("""
                SELECT s.*, COUNT(m.id) AS mat_count
                FROM procurement_suppliers s
                LEFT JOIN procurement_materials m
                       ON LOWER(TRIM(m.supplier_name))=LOWER(TRIM(s.supplier_name))
                GROUP BY s.id
                ORDER BY s.supplier_name ASC
            """).fetchall()
            result = []
            for r in rows:
                d = dict(r)
                for f in ("created_at","updated_at"):
                    if d.get(f): d[f] = str(d[f])
                result.append(d)
            conn.close()
            return jsonify({"status":"ok","suppliers":result,"count":len(result)})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}),500

    # ══════════════════════════════════════════════════════════════════════════
    # PURCHASE ORDERS
    # ══════════════════════════════════════════════════════════════════════════

    # ─── Declaration CRUD ───────────────────────────────────────────────────────
    @app.route("/api/procurement/declarations", methods=["GET"])
    @procurement_required
    def api_decl_list():
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            rows = conn.execute("SELECT * FROM procurement_declarations ORDER BY name").fetchall()
            conn.close()
            return jsonify({"status":"ok","declarations":[dict(r) for r in rows]})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/declarations/save", methods=["POST"])
    @procurement_required
    def api_decl_save():
        d      = request.get_json() or {}
        did    = d.get("id")
        name   = (d.get("name") or "").strip()
        text   = (d.get("text") or "").strip()
        if not name:
            return jsonify({"status":"error","message":"name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            if did:
                conn.execute("UPDATE procurement_declarations SET name=%s,text=%s WHERE id=%s",
                             (name, text, did))
            else:
                conn.execute("INSERT INTO procurement_declarations (name,text) VALUES (%s,%s)",
                             (name, text))
                did = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
            conn.commit()
            conn.close()
            return jsonify({"status":"ok","id":did})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/declarations/delete", methods=["POST"])
    @procurement_required
    def api_decl_delete():
        d   = request.get_json() or {}
        did = d.get("id")
        if not did:
            return jsonify({"status":"error","message":"id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            conn.execute("DELETE FROM procurement_declarations WHERE id=%s", (did,))
            conn.commit()
            conn.close()
            return jsonify({"status":"ok"})
        except Exception as e:
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/po/list", methods=["GET"])
    @procurement_required
    def api_po_list():
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            orders = conn.execute("""
                SELECT o.*,
                       (SELECT COUNT(*) FROM procurement_po_items i WHERE i.po_id = o.id) AS item_count
                FROM procurement_purchase_orders o
                ORDER BY o.created_at DESC
            """).fetchall()

            # Build GRN receipt totals per PO — check BOTH po_id column AND po_invoices JSON
            grn_totals = {}  # po_id → {material → total_received}
            try:
                # Method 1: GRNs with direct po_id
                grn_rows = conn.execute("""
                    SELECT g.po_id, gi.material, SUM(gi.received_qty) AS total_received
                    FROM procurement_grn g
                    JOIN procurement_grn_items gi ON gi.grn_id = g.id
                    WHERE g.po_id IS NOT NULL AND g.status <> 'cancelled'
                    GROUP BY g.po_id, gi.material
                """).fetchall()
                for gr in grn_rows:
                    pid = gr["po_id"]
                    if pid not in grn_totals: grn_totals[pid] = {}
                    mat = (gr["material"] or "").strip()
                    grn_totals[pid][mat] = grn_totals[pid].get(mat, 0) + float(gr["total_received"] or 0)

                # Method 2: GRNs linked via po_invoices JSON (multi-PO GRNs)
                import json as _json2
                json_grns = conn.execute("""
                    SELECT g.id, g.po_invoices, gi.material, gi.received_qty
                    FROM procurement_grn g
                    JOIN procurement_grn_items gi ON gi.grn_id = g.id
                    WHERE g.po_id IS NULL AND g.po_invoices IS NOT NULL AND g.status <> 'cancelled'
                """).fetchall()
                for jg in json_grns:
                    try:
                        inv_list = _json2.loads(jg["po_invoices"] or "[]")
                        for inv in inv_list:
                            pid = inv.get("po_id")
                            if not pid: continue
                            pid = int(pid)
                            if pid not in grn_totals: grn_totals[pid] = {}
                            mat = (jg["material"] or "").strip()
                            grn_totals[pid][mat] = grn_totals[pid].get(mat, 0) + float(jg["received_qty"] or 0)
                    except Exception:
                        pass
            except Exception:
                pass  # GRN table may not exist yet

            # Get PO items for status computation
            po_items_map = {}  # po_id → [{material, qty}]
            try:
                po_items_rows = conn.execute(
                    "SELECT po_id, material, qty FROM procurement_po_items"
                ).fetchall()
                for pi in po_items_rows:
                    pid = pi["po_id"]
                    if pid not in po_items_map:
                        po_items_map[pid] = []
                    po_items_map[pid].append({
                        "material": (pi["material"] or "").strip(),
                        "qty": float(pi["qty"] or 0)
                    })
            except Exception:
                pass

            rows = []
            for o in orders:
                r = dict(o)
                for f in ("po_date", "delivery_date", "created_at", "updated_at"):
                    if r.get(f): r[f] = str(r[f])
                if r.get("grand_total") is not None:
                    r["grand_total"] = float(r["grand_total"])

                # Compute status from GRN data (don't override cancelled)
                stored_status = r.get("status") or "open"
                if stored_status not in ("cancelled",):
                    pid = r["id"]
                    po_items = po_items_map.get(pid, [])
                    received = grn_totals.get(pid, {})
                    if po_items and received:
                        any_received = any(received.get(i["material"], 0) > 0 for i in po_items)
                        all_full     = all(received.get(i["material"], 0) >= i["qty"] - 0.001 for i in po_items)
                        if all_full:
                            computed = "closed"
                        elif any_received:
                            computed = "partial"
                        else:
                            computed = "open"
                    else:
                        computed = stored_status if stored_status in ("approved", "not_approved") else "open"
                    r["status"] = computed
                    # Also persist if changed
                    if computed != stored_status:
                        try:
                            conn.execute(
                                "UPDATE procurement_purchase_orders SET status=%s WHERE id=%s",
                                (computed, pid)
                            )
                        except Exception:
                            pass

                # Add per-item pending summary for display in PO list
                po_items  = po_items_map.get(pid, [])
                received  = grn_totals.get(pid, {})
                # Build a lowercased lookup for case-insensitive matching
                received_lower = {k.lower(): v for k, v in received.items()}
                pending_items = []
                for pi in po_items:
                    mat     = pi["material"]
                    po_qty  = float(pi["qty"] or 0)
                    rcvd    = float(received_lower.get(mat.strip().lower(), 0))
                    pending = max(0.0, round(po_qty - rcvd, 3))
                    if pending > 0:
                        pending_items.append({"material": mat, "pending_qty": pending, "po_qty": po_qty})
                r["pending_items"] = pending_items

                rows.append(r)

            try: conn.commit()
            except Exception: pass
            conn.close()
            return jsonify({"status": "ok", "orders": rows})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.route("/api/procurement/po/get", methods=["GET"])
    @procurement_required
    def api_po_get():
        po_id = request.args.get("id")
        if not po_id:
            return jsonify({"status": "error", "message": "id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            o = conn.execute(
                "SELECT * FROM procurement_purchase_orders WHERE id=%s", (po_id,)
            ).fetchone()
            if not o:
                return jsonify({"status": "error", "message": "PO not found"}), 404
            order = dict(o)
            for f in ("po_date", "delivery_date", "created_at", "updated_at"):
                if order.get(f): order[f] = str(order[f])
            items = conn.execute(
                "SELECT * FROM procurement_po_items WHERE po_id=%s ORDER BY id", (po_id,)
            ).fetchall()

            # Calculate already-received qty per material for this PO
            # Covers: (1) GRNs with po_id = this PO, (2) GRNs with po_invoices referencing this PO
            received_map = {}  # material (lower) → total received qty
            try:
                # Method 1: GRNs directly linked by po_id
                grn_items = conn.execute(
                    """SELECT gi.material, gi.received_qty
                       FROM procurement_grn_items gi
                       JOIN procurement_grn g ON gi.grn_id = g.id
                       WHERE g.po_id = %s
                         AND (g.status IS NULL OR g.status != 'cancelled')""",
                    (po_id,)
                ).fetchall()
                for gi in grn_items:
                    mat = (gi["material"] or "").strip().lower()
                    received_map[mat] = received_map.get(mat, 0) + float(gi["received_qty"] or 0)

                # Method 2: GRNs linked via po_invoices JSON (multi-PO or manual GRNs)
                # Includes GRNs where po_id IS NULL or a different PO
                import json as _json
                grn_multi = conn.execute(
                    """SELECT gi.material, gi.received_qty, g.po_invoices
                       FROM procurement_grn_items gi
                       JOIN procurement_grn g ON gi.grn_id = g.id
                       WHERE g.po_invoices IS NOT NULL
                         AND (g.po_id IS NULL OR g.po_id != %s)
                         AND (g.status IS NULL OR g.status != 'cancelled')""",
                    (po_id,)
                ).fetchall()
                for gi in grn_multi:
                    try:
                        invs = _json.loads(gi["po_invoices"] or "[]")
                        if any(str(inv.get("po_id")) == str(po_id) for inv in invs):
                            mat = (gi["material"] or "").strip().lower()
                            received_map[mat] = received_map.get(mat, 0) + float(gi["received_qty"] or 0)
                    except Exception:
                        pass

                # Method 3: GRNs linked by po_num match (legacy fallback for older GRNs)
                po_num_val = (o.get("po_num") or "").strip()
                if po_num_val:
                    grn_by_num = conn.execute(
                        """SELECT gi.material, gi.received_qty
                           FROM procurement_grn_items gi
                           JOIN procurement_grn g ON gi.grn_id = g.id
                           WHERE g.po_num = %s
                             AND (g.po_id IS NULL OR g.po_id != %s)
                             AND (g.status IS NULL OR g.status != 'cancelled')""",
                        (po_num_val, po_id)
                    ).fetchall()
                    for gi in grn_by_num:
                        mat = (gi["material"] or "").strip().lower()
                        received_map[mat] = received_map.get(mat, 0) + float(gi["received_qty"] or 0)

            except Exception:
                pass  # GRN table may not exist yet — safe fallback

            order["items"] = []
            for i in items:
                it = {**dict(i),
                      "qty":    float(i["qty"])    if i["qty"]    is not None else None,
                      "rate":   float(i["rate"])   if i["rate"]   is not None else None,
                      "amount": float(i["amount"]) if i["amount"] is not None else None}
                # Add received and pending quantities
                mat_key = (i["material"] or "").strip().lower()
                received = received_map.get(mat_key, 0)
                po_qty   = float(i["qty"] or 0)
                pending  = max(0.0, round(po_qty - received, 6))
                it["received_qty"] = round(received, 6)
                it["pending_qty"]  = pending
                order["items"].append(it)

            conn.close()
            return jsonify({"status": "ok", "order": order})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.route("/api/procurement/po/pdf", methods=["GET"])
    @procurement_required
    def api_po_pdf():
        """Generate PO as PDF using pdfkit and return as downloadable file."""
        import pdfkit, tempfile, os as _os
        po_id = request.args.get("id")
        if not po_id:
            return jsonify({"status": "error", "message": "id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            o = conn.execute("SELECT * FROM procurement_purchase_orders WHERE id=%s", (po_id,)).fetchone()
            if not o:
                return jsonify({"status": "error", "message": "PO not found"}), 404
            order = dict(o)
            for f in ("po_date", "delivery_date"):
                if order.get(f): order[f] = str(order[f])
            items = conn.execute("SELECT * FROM procurement_po_items WHERE po_id=%s ORDER BY id", (po_id,)).fetchall()
            order["items"] = [{**dict(i), "qty": float(i["qty"] or 0), "rate": float(i["rate"] or 0),
                               "amount": float(i["amount"] or 0)} for i in items]
            # Supplier details
            sup = conn.execute("SELECT * FROM procurement_suppliers WHERE LOWER(TRIM(supplier_name))=LOWER(TRIM(%s))",
                               (order.get("supplier_name",""),)).fetchone()
            sup = dict(sup) if sup else {}
            # Godown / billing
            bill_godown = conn.execute("SELECT * FROM procurement_godowns WHERE type='billing' OR is_default=1 LIMIT 1").fetchone()
            ship_godown_id = order.get("shipping_godown_id")
            ship_godown = conn.execute("SELECT * FROM procurement_godowns WHERE id=%s", (ship_godown_id,)).fetchone() if ship_godown_id else None
            conn.close()

            # Build HTML
            MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
            def fd(d):
                if not d: return '—'
                p = str(d).split('-')
                return f"{p[2]}/{MONTHS[int(p[1])-1]}/{p[0]}" if len(p)==3 else str(d)

            po_num  = order.get("po_num","—")
            status  = (order.get("status","open")).upper()
            po_date = fd(order.get("po_date",""))
            supplier_name = order.get("supplier_name","—")

            def fi(n):
                try: return f"₹{float(n):,.2f}"
                except: return "—"

            total = sum(float(i.get("amount",0) or 0) for i in order["items"])
            freight = float(order.get("freight_charge") or 0)
            packing = float(order.get("packing_charge") or 0)
            cgst = sgst = 0
            for it in order["items"]:
                gp = float(it.get("gst_rate") or 0)
                amt = float(it.get("amount") or 0)
                if gp > 0 and amt > 0:
                    c = round(amt * (gp/2) / 100, 2)
                    cgst += c; sgst += c
            grand = total + freight + packing + cgst + sgst

            item_rows = ""
            for i, it in enumerate(order["items"], 1):
                qty  = float(it.get("qty") or 0)
                rate = float(it.get("rate") or 0)
                amt  = qty * rate
                gst  = it.get("gst_rate")
                item_rows += f"""<tr>
                    <td style="padding:7px 9px;text-align:center;color:#888">{i}</td>
                    <td style="padding:7px 9px"><strong>{it.get('material','')}</strong>
                    {'<br><span style="font-size:9px;color:#888">HSN: '+str(it.get('hsn_code',''))+'</span>' if it.get('hsn_code') else ''}</td>
                    <td style="padding:7px 9px;text-align:right;font-family:monospace">{f"{qty:,.3f} Kgs" if qty else "—"}</td>
                    <td style="padding:7px 9px;text-align:right;font-family:monospace">{fi(rate) if rate else "—"}</td>
                    <td style="padding:7px 9px;text-align:center;font-family:monospace">{str(gst)+"%" if gst else "—"}</td>
                    <td style="padding:7px 9px;text-align:right;font-family:monospace;font-weight:700">{fi(amt) if amt else "—"}</td>
                </tr>"""

            sup_lines = f"<strong>{supplier_name}</strong>"
            if sup.get("address"): sup_lines += f"<br>{sup['address']}"
            if sup.get("gst_number"): sup_lines += f"<br>GSTIN: <strong>{sup['gst_number']}</strong>"
            if sup.get("contact_person"): sup_lines += f"<br>Contact: {sup['contact_person']}" + (f" | {sup['phone']}" if sup.get("phone") else "")
            if sup.get("email"): sup_lines += f"<br>E-Mail: {sup['email']}"

            bill_lines = ""
            if bill_godown:
                bg = dict(bill_godown)
                if bg.get("name"): bill_lines += f"<strong>{bg['name']}</strong><br>"
                if bg.get("address"): bill_lines += f"{bg['address']}<br>"
                if bg.get("gst"): bill_lines += f"GSTIN: <strong>{bg['gst']}</strong><br>"

            html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
            <title>{po_num}</title>
            <style>
            *{{box-sizing:border-box;margin:0;padding:0;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
            body{{font-family:Arial,Helvetica,sans-serif;font-size:11px;color:#111;background:#fff;padding:20px 28px}}
            .hdr{{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:2px solid #1e3a8a;padding-bottom:8px;margin-bottom:0}}
            .co{{font-size:22px;font-weight:900;color:#1e3a8a;text-transform:uppercase}}
            .cosub{{font-size:9px;color:#666;margin-top:2px}}
            .pstat{{font-size:11px;font-weight:800;color:#1e40af;padding:3px 10px;background:#eff6ff;border-radius:4px}}
            .bar{{display:grid;border:1px solid #ccc;border-top:none}}
            .bar4{{grid-template-columns:1fr 1fr 1fr 1fr}}
            .bar3{{grid-template-columns:1fr 1fr 1fr}}
            .bc{{padding:5px 9px;border-right:1px solid #ccc}}.bc:last-child{{border-right:none}}
            .bl{{font-size:7px;font-weight:800;color:#888;text-transform:uppercase;letter-spacing:.6px;margin-bottom:1px}}
            .bv{{font-size:10.5px;font-weight:600}}
            .adg{{display:grid;grid-template-columns:1fr 1fr 1fr;border:1px solid #ccc;border-top:none}}
            .ab{{padding:8px 10px;border-right:1px solid #ccc;font-size:10px;line-height:1.65}}.ab:last-child{{border-right:none}}
            .al{{font-size:7px;font-weight:800;color:#888;text-transform:uppercase;letter-spacing:.6px;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid #eee}}
            table{{width:100%;border-collapse:collapse}}
            thead tr{{background:#1e3a8a}}
            th{{color:#fff;padding:7px 9px;font-size:8.5px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;border-right:1px solid rgba(255,255,255,.2);text-align:right}}
            th:first-child{{text-align:center}}th:nth-child(2){{text-align:left}}th:last-child{{border-right:none}}
            tbody tr{{border-bottom:1px solid #ddd}}
            tbody tr:nth-child(odd){{background:#f9fafb}}
            td{{padding:7px 9px;font-size:11px;vertical-align:middle;border-right:1px solid #eee}}
            td:last-child{{border-right:none}}
            .ftrow td{{padding:5px 9px;border-right:1px solid #eee;font-size:10px}}
            .ftrow td:last-child{{border-right:none}}
            .ftrow-total td{{font-weight:800;font-size:12px;background:#eff6ff;border-top:2px solid #1e3a8a}}
            .amt-words{{border:1px solid #ccc;border-top:none;padding:6px 10px;font-size:10px}}
            .sig{{display:grid;grid-template-columns:1fr 1fr;border:1px solid #ccc;border-top:none}}
            .sb{{padding:9px 10px;border-right:1px solid #ccc;min-height:48px}}.sb:last-child{{border-right:none;text-align:right}}
            .footer{{text-align:center;font-size:8.5px;color:#94a3b8;margin-top:6px;border-top:1px solid #eee;padding-top:5px}}
            </style></head><body>
            <div class="hdr">
              <div><div class="co">Purchase Order</div><div class="cosub">HCP Wellness Pvt Ltd</div></div>
              <span class="pstat">{status}</span>
            </div>
            <div class="bar bar4">
              <div class="bc"><div class="bl">Voucher No.</div><div class="bv">{po_num}</div></div>
              <div class="bc"><div class="bl">Reference No.</div><div class="bv">{po_num}</div></div>
              <div class="bc"><div class="bl">Dated</div><div class="bv">{po_date}</div></div>
              <div class="bc"><div class="bl">Mode / Terms of Payment</div><div class="bv">—</div></div>
            </div>
            <div class="adg">
              <div class="ab"><div class="al">Supplier (Bill From)</div>{sup_lines}</div>
              <div class="ab"><div class="al">Invoice To</div>{bill_lines}</div>
              <div class="ab"><div class="al">Remarks</div>{order.get('remarks','') or '—'}</div>
            </div>
            <table><thead><tr>
              <th style="width:26px;text-align:center">Sl<br>No.</th>
              <th style="text-align:left">Description of Goods</th>
              <th style="width:100px">Quantity</th>
              <th style="width:100px">Rate (&#8377;)</th>
              <th style="width:60px;text-align:center">GST %</th>
              <th style="width:110px">Amount (₹)</th>
            </tr></thead>
            <tbody>{item_rows}</tbody>
            <tfoot>
              {'<tr class="ftrow"><td colspan="5" style="text-align:right;color:#555">Freight</td><td style="text-align:right;font-family:monospace">'+fi(freight)+'</td></tr>' if freight else ''}
              {'<tr class="ftrow"><td colspan="5" style="text-align:right;color:#555">Packing</td><td style="text-align:right;font-family:monospace">'+fi(packing)+'</td></tr>' if packing else ''}
              <tr class="ftrow"><td colspan="5" style="text-align:right;color:#555">Taxable Amount</td><td style="text-align:right;font-family:monospace">{fi(total)}</td></tr>
              {'<tr class="ftrow"><td colspan="5" style="text-align:right;color:#555">CGST</td><td style="text-align:right;font-family:monospace">'+fi(cgst)+'</td></tr>' if cgst else ''}
              {'<tr class="ftrow"><td colspan="5" style="text-align:right;color:#555">SGST</td><td style="text-align:right;font-family:monospace">'+fi(sgst)+'</td></tr>' if sgst else ''}
              <tr class="ftrow-total"><td colspan="5" style="text-align:right">Total</td><td style="text-align:right;font-family:monospace;color:#1e3a8a;font-size:13px">{fi(grand)}</td></tr>
            </tfoot></table>
            <div class="sig">
              <div class="sb"></div>
              <div class="sb" style="text-align:right"><div style="font-size:7px;font-weight:800;color:#888;text-transform:uppercase;letter-spacing:.6px;margin-bottom:4px">for HCP Wellness Pvt Ltd</div>
              <br><div style="font-size:10px;color:#666;margin-top:16px">Authorised Signatory</div></div>
            </div>
            <div class="footer">SUBJECT TO AHMEDABAD JURISDICTION | This is a Computer Generated Document</div>
            </body></html>"""

            options = {
                'page-size': 'A4',
                'margin-top':    '10mm',
                'margin-right':  '10mm',
                'margin-bottom': '10mm',
                'margin-left':   '10mm',
                'encoding': 'UTF-8',
                'no-outline': None,
                'disable-smart-shrinking': None,
                'enable-local-file-access': None,
                'quiet': None,
            }
            tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            tmp.close()
            try:
                pdfkit.from_string(html, tmp.name, options=options)
            except OSError as pdf_err:
                if not _os.path.exists(tmp.name) or _os.path.getsize(tmp.name) == 0:
                    raise pdf_err
            with open(tmp.name, 'rb') as f:
                pdf_bytes = f.read()
            _os.unlink(tmp.name)

            from flask import Response
            safe_num = po_num.replace('/', '_')
            return Response(
                pdf_bytes,
                mimetype='application/pdf',
                headers={'Content-Disposition': f'attachment; filename="PO_{safe_num}.pdf"'}
            )
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500



    @app.route("/api/procurement/po/save", methods=["POST"])
    @procurement_required
    def api_po_save():
        d        = request.get_json() or {}
        po_id    = d.get("id")          # None = create, int = update
        po_num   = (d.get("po_num") or "").strip()
        supplier = (d.get("supplier_name") or "").strip()
        status   = (d.get("status") or "open").strip()
        po_date  = d.get("po_date")     or None
        delivery = d.get("delivery_date") or None
        remarks  = (d.get("remarks") or "").strip() or None
        voucher_type_name = (d.get("voucher_type_name") or "").strip() or None
        try:
            delivery_days = int(d.get("delivery_days") or 0) or None
        except (TypeError, ValueError):
            delivery_days = None
        try:
            tc_list_id = int(d.get("tc_list_id") or 0) or None
        except (TypeError, ValueError):
            tc_list_id = None
        try:
            declaration_id = int(d.get("declaration_id") or 0) or None
        except (TypeError, ValueError):
            declaration_id = None
        try:
            freight_charge = float(d.get("freight_charge") or 0) or None
        except (TypeError, ValueError):
            freight_charge = None
        try:
            packing_charge = float(d.get("packing_charge") or 0) or None
        except (TypeError, ValueError):
            packing_charge = None
        items    = d.get("items") or []   # [{material, qty, rate}]
        uid      = session.get("UID", "")

        if not supplier:
            return jsonify({"status": "error", "message": "supplier_name required"}), 400
        if not items:
            return jsonify({"status": "error", "message": "at least one line item required"}), 400
        # For updates, po_num is required (user can't change it but it must be sent)
        if po_id and not po_num:
            return jsonify({"status": "error", "message": "po_num required for update"}), 400

        # Compute grand total (items + freight + packing)
        grand_total = 0.0
        for item in items:
            try:
                qty = float(item.get("qty") or 0)
                rate = float(item.get("rate") or 0)
                grand_total += qty * rate
            except (TypeError, ValueError):
                pass
        grand_total += (freight_charge or 0) + (packing_charge or 0)

        # ── Helper: atomically generate next PO number ──────────────────────
        def _assign_po_number(conn, cfg):
            """
            Uses GET_LOCK + MAX query to atomically assign the next PO number.
            cfg = {prefix, suffix, digits} from the frontend settings.
            If cfg is empty, looks up active style from procurement_voucher_numbering.
            Returns the assigned po_num string.
            """
            import re as _re
            prefix  = (cfg.get("prefix") or "").strip()
            suffix  = (cfg.get("suffix") or "").strip()
            digits  = int(cfg.get("digits") or 4)

            # Server-side fallback: if frontend sent no prefix/suffix,
            # look up the active voucher numbering style from DB
            if not prefix and not suffix:
                today_str = __import__('datetime').date.today().isoformat()
                try:
                    # If a specific named voucher type was chosen (e.g. "RM Purchase Order"),
                    # look up its numbering style by name first, then fall back to parent 'po'
                    vn_type_key = voucher_type_name or "po"
                    vn_row = conn.execute(
                        "SELECT prefix, suffix, digits, start_num "
                        "FROM procurement_voucher_numbering "
                        "WHERE voucher_type=%s AND valid_from <= %s AND valid_to >= %s "
                        "ORDER BY id DESC LIMIT 1",
                        (vn_type_key, today_str, today_str)
                    ).fetchone()
                    # Fallback to parent 'po' type if no style found for the named type
                    if not vn_row and voucher_type_name:
                        vn_row = conn.execute(
                            "SELECT prefix, suffix, digits, start_num "
                            "FROM procurement_voucher_numbering "
                            "WHERE voucher_type='po' AND valid_from <= %s AND valid_to >= %s "
                            "ORDER BY id DESC LIMIT 1",
                            (today_str, today_str)
                        ).fetchone()
                    if vn_row:
                        prefix = (vn_row["prefix"] or "").strip()
                        suffix = (vn_row["suffix"] or "").strip()
                        digits = int(vn_row["digits"] or 4)
                except Exception:
                    pass  # non-fatal — fall through to plain numbering

            # Build search pattern for existing POs with this prefix/suffix
            pattern = (prefix + "/%") if prefix else "%"
            lock_name = f"po_num_lock_{prefix}_{suffix}"

            # Acquire named lock (timeout 10s) — serialises concurrent saves
            conn.execute("SELECT GET_LOCK(%s, 10) AS locked", (lock_name,))
            try:
                rows = conn.execute(
                    "SELECT po_num FROM procurement_purchase_orders WHERE po_num LIKE %s",
                    (pattern,)
                ).fetchall()
                max_seq = 0
                for row in rows:
                    nums = _re.findall(r"(\d{" + str(digits) + r",})", row["po_num"])
                    if nums:
                        max_seq = max(max_seq, int(nums[-1]))
                next_seq = max_seq + 1
                num_str  = str(next_seq).zfill(digits)
                parts = []
                if prefix: parts.append(prefix)
                parts.append(num_str)
                if suffix: parts.append(suffix)
                return "/".join(parts)
            finally:
                conn.execute("SELECT RELEASE_LOCK(%s)", (lock_name,))

        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500

            if po_id:
                # ── UPDATE existing PO ──────────────────────────────────────
                conn.execute("""
                    UPDATE procurement_purchase_orders
                    SET po_num=%s, supplier_name=%s, status=%s, po_date=%s,
                        delivery_date=%s, delivery_days=%s, remarks=%s,
                        grand_total=%s, tc_list_id=%s, declaration_id=%s,
                        freight_charge=%s, packing_charge=%s,
                        voucher_type_name=%s, updated_by=%s
                    WHERE id=%s
                """, (po_num, supplier, status, po_date, delivery, delivery_days,
                      remarks, grand_total, tc_list_id, declaration_id,
                      freight_charge, packing_charge, voucher_type_name, uid, po_id))
                conn.execute("DELETE FROM procurement_po_items WHERE po_id=%s", (po_id,))

            else:
                # ── CREATE new PO — server assigns number atomically ────────
                num_cfg = d.get("num_cfg") or {}   # {prefix, suffix, digits} from frontend
                po_num  = _assign_po_number(conn, num_cfg)
                conn.execute("""
                    INSERT INTO procurement_purchase_orders
                        (po_num, supplier_name, status, po_date, delivery_date,
                         delivery_days, remarks, grand_total,
                         tc_list_id, declaration_id, freight_charge, packing_charge,
                         voucher_type_name, created_by, updated_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (po_num, supplier, status, po_date, delivery, delivery_days,
                      remarks, grand_total, tc_list_id, declaration_id,
                      freight_charge, packing_charge, voucher_type_name, uid, uid))
                po_id = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]

            # ── Insert line items ───────────────────────────────────────────
            for item in items:
                mat = (item.get("material") or "").strip()
                if not mat:
                    continue
                try:
                    qty    = float(item.get("qty")  or 0) or None
                    rate   = float(item.get("rate") or 0) or None
                    amount = (qty or 0) * (rate or 0) or None
                except (TypeError, ValueError):
                    qty = rate = amount = None
                hsn_code = (item.get("hsn_code") or "").strip() or None
                try:
                    gst_rate = float(item.get("gst_rate") or 0) or None
                except (TypeError, ValueError):
                    gst_rate = None
                try:
                    cgst_amount = float(item.get("cgst_amount") or 0) or None
                    sgst_amount = float(item.get("sgst_amount") or 0) or None
                except (TypeError, ValueError):
                    cgst_amount = sgst_amount = None
                conn.execute("""
                    INSERT INTO procurement_po_items
                        (po_id, material, qty, qty_per_pkg, rate, amount, hsn_code, gst_rate, cgst_amount, sgst_amount, packages, uom)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (po_id, mat, qty,
                         float(item.get("qty_per_pkg") or 0) or None,
                         rate, amount, hsn_code, gst_rate, cgst_amount, sgst_amount,
                         int(item.get("packages")) if item.get("packages") else None,
                         (item.get("uom") or "KG").strip().upper() or None))
                # Write back rate to material master so future POs auto-fill it
                if rate:
                    try:
                        conn.execute("""
                            UPDATE procurement_materials
                            SET last_purchase_rate=%s
                            WHERE material_name=%s AND (last_purchase_rate IS NULL OR last_purchase_rate=0)
                        """, (rate, mat))
                    except Exception:
                        pass

            conn.commit()
            conn.close()
            return jsonify({"status": "ok", "id": po_id, "po_num": po_num,
                            "grand_total": grand_total})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.route("/api/procurement/po/delete", methods=["POST"])
    @procurement_required
    def api_po_delete():
        d     = request.get_json() or {}
        po_id = d.get("id")
        if not po_id:
            return jsonify({"status": "error", "message": "id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            conn.execute("DELETE FROM procurement_purchase_orders WHERE id=%s", (po_id,))
            conn.commit()
            conn.close()
            return jsonify({"status": "ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.route("/api/procurement/po/cancel", methods=["POST"])
    @procurement_required
    def api_po_cancel():
        """Cancel a PO: clear all line items and field values, keep PO number, mark cancelled.
        The PO number is preserved so the blank PO can be reused/edited again."""
        d     = request.get_json() or {}
        po_id = d.get("id")
        if not po_id:
            return jsonify({"status": "error", "message": "id required"}), 400
        uid = session.get("UID", "")
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            # Delete all line items
            conn.execute("DELETE FROM procurement_po_items WHERE po_id=%s", (po_id,))
            # Clear PO header fields — keep po_num, reset everything else
            conn.execute("""
                UPDATE procurement_purchase_orders
                SET status='cancelled',
                    supplier_name=NULL,
                    po_date=NULL,
                    delivery_date=NULL,
                    delivery_days=NULL,
                    remarks=NULL,
                    grand_total=0,
                    tc_list_id=NULL,
                    declaration_id=NULL,
                    updated_by=%s
                WHERE id=%s
            """, (uid, po_id))
            conn.commit()
            conn.close()
            return jsonify({"status": "ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.route("/api/procurement/po/last_rate", methods=["GET"])
    @procurement_required
    def api_po_last_rate():
        """Returns the last used rate for a material from PO history."""
        mat = (request.args.get("material") or "").strip()
        if not mat:
            return jsonify({"status": "error", "message": "material required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            # Check material master first
            row = conn.execute(
                "SELECT last_purchase_rate FROM procurement_materials WHERE material_name=%s",
                (mat,)
            ).fetchone()
            rate = float(row["last_purchase_rate"]) if row and row["last_purchase_rate"] else None
            # If not in material master, check po_items history
            if not rate:
                po_row = conn.execute("""
                    SELECT rate FROM procurement_po_items
                    WHERE material=%s AND rate IS NOT NULL AND rate > 0
                    ORDER BY id DESC LIMIT 1
                """, (mat,)).fetchone()
                rate = float(po_row["rate"]) if po_row and po_row["rate"] else None
            conn.close()
            return jsonify({"status": "ok", "rate": rate})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.route("/api/procurement/po/next_num", methods=["GET"])
    @procurement_required
    def api_po_next_num():
        """Returns the next available PO sequence number for a given prefix+period."""
        prefix  = request.args.get("prefix", "HCP")
        period  = request.args.get("period", "")
        digits  = int(request.args.get("digits", 4))
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            # Find highest existing number with this prefix/period
            pattern = f"{prefix}/{period}/%" if period else f"{prefix}/%"
            rows = conn.execute(
                "SELECT po_num FROM procurement_purchase_orders WHERE po_num LIKE %s",
                (pattern,)
            ).fetchall()
            conn.close()
            max_seq = 0
            import re as _re
            for row in rows:
                nums = _re.findall(r"(\d{" + str(digits) + r",})", row["po_num"])
                if nums:
                    max_seq = max(max_seq, int(nums[-1]))
            return jsonify({"status": "ok", "next": max_seq + 1})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500



    # ══════════════════════════════════════════════════════════════════════════
    # VOUCHER NUMBERING — CRUD + next-number preview
    # ══════════════════════════════════════════════════════════════════════════

    @app.route("/api/procurement/voucher_numbering/list", methods=["GET"])
    @procurement_required
    def api_vn_list():
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            rows = conn.execute(
                "SELECT * FROM procurement_voucher_numbering ORDER BY voucher_type, valid_from DESC"
            ).fetchall()
            conn.close()
            styles = []
            for r in rows:
                styles.append({
                    "id":           r["id"],
                    "voucher_type": r["voucher_type"],
                    "prefix":       r["prefix"] or "",
                    "suffix":       r["suffix"] or "",
                    "digits":       r["digits"] or 4,
                    "start_num":    r["start_num"] or 1,
                    "valid_from":   str(r["valid_from"]) if r["valid_from"] else "",
                    "valid_to":     str(r["valid_to"]) if r["valid_to"] else "",
                })
            return jsonify({"status": "ok", "styles": styles})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/voucher_numbering/save", methods=["POST"])
    @procurement_required
    def api_vn_save():
        d = request.get_json() or {}
        vtype  = (d.get("voucher_type") or "po").strip().lower()
        prefix = (d.get("prefix") or "").strip()
        suffix = (d.get("suffix") or "").strip()
        digits = int(d.get("digits") or 4)
        start  = int(d.get("start_num") or 1)
        vfrom  = (d.get("valid_from") or "").strip()
        vto    = (d.get("valid_to") or "").strip()
        sid    = d.get("id")
        if not vfrom or not vto:
            return jsonify({"status": "error", "message": "valid_from and valid_to required"}), 400
        if vfrom > vto:
            return jsonify({"status": "error", "message": "Valid From must be before Valid To"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500

            # ── Duplicate / overlap check ──────────────────────────────
            # Block if another style for the same voucher_type has ANY
            # overlapping date range — prevents ambiguity in number assignment.
            overlap_check = conn.execute("""
                SELECT id, prefix, suffix, valid_from, valid_to
                FROM procurement_voucher_numbering
                WHERE voucher_type = %s
                  AND id != %s
                  AND valid_from <= %s
                  AND valid_to   >= %s
            """, (vtype, sid or 0, vto, vfrom)).fetchone()
            if overlap_check:
                def _fd(d):
                    if not d: return '—'
                    try:
                        parts = str(d).split('-')
                        months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
                        return parts[2]+'/'+months[int(parts[1])-1]+'/'+parts[0]
                    except: return str(d)
                existing_pfx = overlap_check['prefix'] or '—'
                existing_sfx = overlap_check['suffix'] or '—'
                same = (existing_pfx.lower().strip() == prefix.lower().strip() and
                        existing_sfx.lower().strip() == suffix.lower().strip())
                msg = (
                    f"{'Duplicate' if same else 'Period clash'}: style '{existing_pfx}/{existing_sfx}' "
                    f"already covers {_fd(overlap_check['valid_from'])} → {_fd(overlap_check['valid_to'])}. "
                    f"Adjust the date range so periods don't overlap."
                )
                return jsonify({"status": "error", "message": msg}), 400
            if sid:
                conn.execute("""
                    UPDATE procurement_voucher_numbering
                    SET voucher_type=%s, prefix=%s, suffix=%s, digits=%s,
                        start_num=%s, valid_from=%s, valid_to=%s
                    WHERE id=%s
                """, (vtype, prefix, suffix, digits, start, vfrom, vto, sid))
            else:
                conn.execute("""
                    INSERT INTO procurement_voucher_numbering
                        (voucher_type, prefix, suffix, digits, start_num, valid_from, valid_to)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (vtype, prefix, suffix, digits, start, vfrom, vto))
                sid = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
            conn.commit()
            conn.close()
            return jsonify({"status": "ok", "id": sid})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/voucher_numbering/delete", methods=["POST"])
    @procurement_required
    def api_vn_delete():
        d = request.get_json() or {}
        sid = d.get("id")
        if not sid:
            return jsonify({"status": "error", "message": "id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            conn.execute("DELETE FROM procurement_voucher_numbering WHERE id=%s", (sid,))
            conn.commit()
            conn.close()
            return jsonify({"status": "ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/voucher_numbering/next", methods=["GET"])
    @procurement_required
    def api_vn_next():
        """Preview the next sequence number for a given voucher type + prefix/suffix.
        If no prefix/suffix passed, looks up the active style from DB."""
        vtype  = request.args.get("voucher_type", "po")
        prefix = request.args.get("prefix", "")
        suffix = request.args.get("suffix", "")
        digits = int(request.args.get("digits", 4))
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500

            # Auto-detect from DB if not provided
            if not prefix and not suffix:
                from datetime import date as _date
                today_str = _date.today().isoformat()
                vn_row = conn.execute(
                    "SELECT prefix, suffix, digits FROM procurement_voucher_numbering "
                    "WHERE voucher_type=%s AND valid_from <= %s AND valid_to >= %s "
                    "ORDER BY id DESC LIMIT 1",
                    (vtype, today_str, today_str)
                ).fetchone()
                if vn_row:
                    prefix = (vn_row["prefix"] or "").strip()
                    suffix = (vn_row["suffix"] or "").strip()
                    digits = int(vn_row["digits"] or 4)

            # Determine the table based on voucher type
            if vtype in ("grn", "rov"):
                table = "procurement_grn"
                col   = "grn_num"
            else:
                table = "procurement_purchase_orders"
                col   = "po_num"

            # Build search pattern
            pattern = (prefix + "/%") if prefix else "%"
            rows = conn.execute(
                f"SELECT {col} AS vnum FROM {table} WHERE {col} LIKE %s",
                (pattern,)
            ).fetchall()
            conn.close()

            import re as _re
            max_seq = 0
            for row in rows:
                nums = _re.findall(r"(\d{" + str(digits) + r",})", row["vnum"] or "")
                if nums:
                    max_seq = max(max_seq, int(nums[-1]))
            return jsonify({"status": "ok", "next": max_seq + 1,
                           "prefix": prefix, "suffix": suffix, "digits": digits})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    # ══════════════════════════════════════════════════════════════════════════
    # GODOWNS & BILLING ADDRESS
    # ══════════════════════════════════════════════════════════════════════════

    @app.route("/api/procurement/godowns", methods=["GET"])
    @procurement_required
    def api_godowns_list():
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            rows = conn.execute("SELECT * FROM procurement_godowns ORDER BY is_default DESC, name ASC").fetchall()
            conn.close()
            return jsonify({"status":"ok","godowns":[dict(r) for r in rows]})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/godowns/save", methods=["POST"])
    @procurement_required
    def api_godowns_save():
        d = request.get_json() or {}
        gid = d.get("id"); name = (d.get("name") or "").strip()
        if not name: return jsonify({"status":"error","message":"name required"}), 400
        state = (d.get("state") or "").strip() or None
        city  = (d.get("city") or "").strip() or None
        pin   = (d.get("pin") or "").strip() or None
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            if d.get("is_default"):
                conn.execute("UPDATE procurement_godowns SET is_default=0")
            if gid:
                conn.execute("UPDATE procurement_godowns SET name=%s,address=%s,contact=%s,phone=%s,email=%s,state=%s,city=%s,pin=%s,is_default=%s WHERE id=%s",
                    (name,d.get("address"),d.get("contact"),d.get("phone"),d.get("email"),state,city,pin,1 if d.get("is_default") else 0,gid))
            else:
                conn.execute("INSERT INTO procurement_godowns (name,address,contact,phone,email,state,city,pin,is_default) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (name,d.get("address"),d.get("contact"),d.get("phone"),d.get("email"),state,city,pin,1 if d.get("is_default") else 0))
                gid = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
            conn.commit(); conn.close()
            return jsonify({"status":"ok","id":gid})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/godowns/delete", methods=["POST"])
    @procurement_required
    def api_godowns_delete():
        d = request.get_json() or {}; gid = d.get("id")
        if not gid: return jsonify({"status":"error","message":"id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            conn.execute("DELETE FROM procurement_godowns WHERE id=%s",(gid,))
            conn.commit(); conn.close()
            return jsonify({"status":"ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/billing", methods=["GET","POST"])
    @procurement_required
    def api_billing_address():
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            if request.method == "POST":
                d = request.get_json() or {}
                for k, v in d.items():
                    if isinstance(v, (dict, list)):
                        continue
                    conn.execute(
                        "INSERT INTO procurement_settings (setting_key, setting_value) VALUES (%s,%s) ON DUPLICATE KEY UPDATE setting_value=VALUES(setting_value)",
                        ("billing_" + k, str(v) if v is not None else None)
                    )
                conn.commit(); conn.close()
                return jsonify({"status": "ok"})
            else:
                rows = conn.execute(
                    "SELECT setting_key, setting_value FROM procurement_settings WHERE setting_key LIKE %s",
                    ("billing_%%",)
                ).fetchall()
                conn.close()
                data = {r["setting_key"].replace("billing_", ""): r["setting_value"] for r in rows}
                return jsonify({"status": "ok", "billing": data})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    @app.route("/api/procurement/godowns/seed", methods=["POST"])
    @procurement_required
    def api_godowns_seed():
        d = request.get_json() or {}
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB failed"}), 500
            existing = conn.execute("SELECT COUNT(*) AS n FROM procurement_godowns").fetchone()
            seeded_godowns = 0
            seeded_billing = False
            if (existing["n"] if existing else 0) == 0:
                for g in (d.get("godowns") or []):
                    name = (g.get("name") or "").strip()
                    if not name: continue
                    conn.execute(
                        "INSERT INTO procurement_godowns (name,address,contact,phone,email,is_default) VALUES (%s,%s,%s,%s,%s,%s)",
                        (name, g.get("address",""), g.get("contact",""), g.get("phone",""), g.get("email",""), 1 if g.get("is_default") else 0)
                    )
                    seeded_godowns += 1
            billing = d.get("billing") or {}
            if billing and billing.get("addr1"):
                for k, v in billing.items():
                    if isinstance(v, (dict, list)):
                        continue
                    conn.execute(
                        "INSERT INTO procurement_settings (setting_key, setting_value) VALUES (%s,%s) ON DUPLICATE KEY UPDATE setting_value=VALUES(setting_value)",
                        ("billing_" + k, str(v) if v is not None else None)
                    )
                seeded_billing = True
            conn.commit(); conn.close()
            return jsonify({"status":"ok","seeded_godowns":seeded_godowns,"seeded_billing":seeded_billing})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/tc/list", methods=["GET"])
    @procurement_required
    def api_tc_list():
        try:
            import json as _j
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            rows = conn.execute("SELECT * FROM procurement_tc_lists ORDER BY name ASC").fetchall()
            conn.close()
            result = []
            for r in rows:
                d2 = dict(r)
                d2["other_terms"] = _j.loads(d2["other_terms"]) if d2.get("other_terms") else []
                for f in ("created_at","updated_at"):
                    if d2.get(f): d2[f] = str(d2[f])
                result.append(d2)
            return jsonify({"status":"ok","tc_lists":result})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/tc/save", methods=["POST"])
    @procurement_required
    def api_tc_save():
        import json as _j
        d = request.get_json() or {}
        tc_id = d.get("id")
        name  = (d.get("name") or "").strip()
        if not name: return jsonify({"status":"error","message":"name required"}), 400
        other = _j.dumps(d.get("other_terms") or [])
        uid   = session.get("UID","")
        fields = (name, d.get("delivery_days") or None, d.get("delivery_mode") or None,
                  d.get("delivery_notes") or None, d.get("payment_type") or None,
                  d.get("credit_days") or None, d.get("payment_notes") or None, other)
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            if tc_id:
                conn.execute(
                    "UPDATE procurement_tc_lists SET name=%s,delivery_days=%s,delivery_mode=%s,"
                    "delivery_notes=%s,payment_type=%s,credit_days=%s,payment_notes=%s,other_terms=%s WHERE id=%s",
                    fields + (tc_id,))
            else:
                conn.execute(
                    "INSERT INTO procurement_tc_lists (name,delivery_days,delivery_mode,delivery_notes,"
                    "payment_type,credit_days,payment_notes,other_terms,created_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    fields + (uid,))
                tc_id = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]
            conn.commit(); conn.close()
            return jsonify({"status":"ok","id":tc_id})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/tc/delete", methods=["POST"])
    @procurement_required
    def api_tc_delete():
        d = request.get_json() or {}
        tc_id = d.get("id")
        if not tc_id: return jsonify({"status":"error","message":"id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            conn.execute("DELETE FROM procurement_tc_lists WHERE id=%s", (tc_id,))
            conn.commit(); conn.close()
            return jsonify({"status":"ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    # ══════════════════════════════════════════════════════════════════
    # GRN — Goods Receipt Note
    # ══════════════════════════════════════════════════════════════════

    @app.route("/api/procurement/grn/list", methods=["GET"])
    @procurement_required
    def api_grn_list():
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            rows = conn.execute("""
                SELECT g.*,
                    (SELECT COUNT(*) FROM procurement_grn_items i WHERE i.grn_id=g.id) AS item_count
                FROM procurement_grn g
                WHERE (g.grn_type IS NULL OR g.grn_type = 'GRN' OR g.grn_type = '')
                ORDER BY g.created_at DESC
            """).fetchall()
            result = []
            for r in rows:
                d = dict(r)
                for f in ("grn_date","invoice_date","created_at","updated_at"):
                    if d.get(f): d[f] = str(d[f])
                if d.get("grand_total") is not None: d["grand_total"] = float(d["grand_total"])
                # Parse po_invoices JSON
                import json as _json
                if d.get("po_invoices"):
                    try: d["po_invoices"] = _json.loads(d["po_invoices"])
                    except Exception: d["po_invoices"] = []
                else:
                    d["po_invoices"] = []
                result.append(d)
            conn.close()
            return jsonify({"status":"ok","grns":result})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/grn/get", methods=["GET"])
    @procurement_required
    def api_grn_get():
        grn_id = request.args.get("id")
        if not grn_id: return jsonify({"status":"error","message":"id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            g = conn.execute("SELECT * FROM procurement_grn WHERE id=%s", (grn_id,)).fetchone()
            if not g: return jsonify({"status":"error","message":"GRN not found"}), 404
            grn = dict(g)
            for f in ("grn_date","invoice_date","created_at","updated_at"):
                if grn.get(f): grn[f] = str(grn[f])
            # Parse po_invoices JSON
            import json as _json
            if grn.get("po_invoices"):
                try: grn["po_invoices"] = _json.loads(grn["po_invoices"])
                except Exception: grn["po_invoices"] = []
            else:
                grn["po_invoices"] = []
            items = conn.execute(
                "SELECT * FROM procurement_grn_items WHERE grn_id=%s ORDER BY id", (grn_id,)
            ).fetchall()

            # Build per-item PO reference by matching material to PO items
            po_ids_to_check = set()
            if grn.get("po_id"): po_ids_to_check.add(int(grn["po_id"]))
            for pi in grn.get("po_invoices", []):
                if pi.get("po_id"): po_ids_to_check.add(int(pi["po_id"]))
            # Build material→po_num mapping
            mat_po_map = {}
            for pid in po_ids_to_check:
                try:
                    po_info = conn.execute("SELECT po_num FROM procurement_purchase_orders WHERE id=%s", (pid,)).fetchone()
                    po_items_list = conn.execute("SELECT material FROM procurement_po_items WHERE po_id=%s", (pid,)).fetchall()
                    if po_info:
                        for pi_item in po_items_list:
                            mat_key = (pi_item["material"] or "").strip().lower()
                            if mat_key:
                                mat_po_map[mat_key] = {"po_id": pid, "po_num": po_info["po_num"]}
                except Exception:
                    pass

            grn["items"] = []
            for i in items:
                d2 = dict(i)
                for f in ("received_qty","po_qty","rate","amount","gst_rate"):
                    if d2.get(f) is not None: d2[f] = float(d2[f])
                for f in ("invoice_date","mfg_date","expiry_date"):
                    if d2.get(f): d2[f] = str(d2[f])
                # Add per-item PO reference
                mat_key = (d2.get("material") or "").strip().lower()
                po_ref = mat_po_map.get(mat_key, {})
                d2["po_id"] = po_ref.get("po_id")
                d2["po_num"] = po_ref.get("po_num", "")
                grn["items"].append(d2)
            conn.close()
            return jsonify({"status":"ok","grn":grn})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/grn/pdf", methods=["GET"])
    @procurement_required
    def api_grn_pdf():
        """Generate GRN as PDF using pdfkit."""
        import pdfkit, tempfile, os as _os, json as _json
        grn_id = request.args.get("id")
        if not grn_id:
            return jsonify({"status": "error", "message": "id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            g = conn.execute("SELECT * FROM procurement_grn WHERE id=%s", (grn_id,)).fetchone()
            if not g:
                return jsonify({"status": "error", "message": "GRN not found"}), 404
            grn = dict(g)
            for f in ("grn_date","invoice_date","created_at","updated_at"):
                if grn.get(f): grn[f] = str(grn[f])
            if grn.get("po_invoices"):
                try: grn["po_invoices"] = _json.loads(grn["po_invoices"])
                except: grn["po_invoices"] = []
            items_raw = conn.execute(
                "SELECT * FROM procurement_grn_items WHERE grn_id=%s ORDER BY id", (grn_id,)
            ).fetchall()
            items = []
            for i in items_raw:
                d2 = dict(i)
                for f in ("received_qty","po_qty","rate","amount","gst_rate"):
                    if d2.get(f) is not None: d2[f] = float(d2[f])
                for f in ("invoice_date","mfg_date","expiry_date"):
                    if d2.get(f): d2[f] = str(d2[f])
                items.append(d2)
            sup = conn.execute(
                "SELECT * FROM procurement_suppliers WHERE LOWER(TRIM(supplier_name))=LOWER(TRIM(%s))",
                (grn.get("supplier_name",""),)
            ).fetchone()
            sup = dict(sup) if sup else {}
            conn.close()

            MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
            def fd(d):
                if not d: return '—'
                p = str(d).split('-')
                return f"{p[2]}/{MONTHS[int(p[1])-1]}/{p[0]}" if len(p)==3 else str(d)
            def fi(n):
                try: return f"₹{float(n):,.2f}"
                except: return "—"

            grn_num = grn.get("grn_num","—")
            grn_date = fd(grn.get("grn_date",""))
            supplier = grn.get("supplier_name","—")

            sup_lines = f"<strong>{supplier}</strong>"
            if sup.get("address"): sup_lines += f"<br>{sup['address']}"
            if sup.get("gst_number"): sup_lines += f"<br>GSTIN: <strong>{sup['gst_number']}</strong>"
            if sup.get("contact_person"): sup_lines += f"<br>Contact: {sup['contact_person']}" + (f" | {sup['phone']}" if sup.get("phone") else "")
            if sup.get("email"): sup_lines += f"<br>E-Mail: {sup['email']}"

            po_invs = grn.get("po_invoices") or []
            inv_rows = ""
            for inv in po_invs:
                if inv.get("po_num"):
                    inv_rows += f'<tr><td style="padding:3px 8px;font-size:10px;font-weight:700">{inv.get("po_num","")}</td><td style="padding:3px 8px;font-size:10px">—</td></tr>'

            total = 0; cgst = 0; sgst = 0
            item_rows = ""
            for i, it in enumerate(items, 1):
                pkgs  = int(it.get("packages") or 0)
                qpp   = float(it.get("qty_per_pkg") or 0)
                rqty  = float(it.get("received_qty") or 0)
                if pkgs > 0 and qpp > 0:
                    total_qty = pkgs * qpp
                elif pkgs > 0 and rqty > 0:
                    qpp = round(rqty / pkgs, 3); total_qty = rqty
                else:
                    total_qty = rqty
                rate = float(it.get("rate") or 0)
                amt  = total_qty * rate
                total += amt
                gp = float(it.get("gst_rate") or 0)
                if gp > 0 and amt > 0:
                    cv = round(amt*(gp/2)/100, 2); cgst += cv; sgst += cv
                sub = []
                if it.get("invoice_num"): sub.append("Inv: <strong>" + str(it["invoice_num"]) + "</strong>" + (" (" + fd(it["invoice_date"]) + ")" if it.get("invoice_date") else ""))
                if it.get("batch_num"):   sub.append("Batch: " + str(it["batch_num"]))
                if it.get("mfg_date"):    sub.append("Mfg: " + fd(it["mfg_date"]))
                if it.get("expiry_date"): sub.append("Exp: " + fd(it["expiry_date"]))
                if it.get("location"):    sub.append("\U0001f4cd " + str(it["location"]))
                sub_line = ('<br><span style="font-size:8.5px;color:#555">' + " &nbsp;|&nbsp; ".join(sub) + "</span>") if sub else ""
                pqty     = float(it.get("po_qty") or 0)
                uom      = (it.get("uom") or "KG").upper()
                if pkgs > 0 and qpp > 0:
                    pkgs_cell = (str(pkgs) + '&nbsp;<span style="font-size:9px;color:#888">pkgs</span>'
                                 + '&nbsp;<span style="color:#aaa">&times;</span>&nbsp;'
                                 + f"{qpp:,.3f}")
                else:
                    pkgs_cell = "&mdash;"
                item_rows += (
                    "<tr>"
                    + '<td style="padding:7px 9px;text-align:center;color:#888">' + str(i) + "</td>"
                    + '<td style="padding:7px 9px"><strong>' + str(it.get("material","")) + "</strong>" + sub_line + "</td>"
                    + '<td style="padding:7px 9px;text-align:center;font-family:monospace;font-size:9.5px;color:#7c3aed">' + pkgs_cell + "</td>"
                    + '<td style="padding:7px 9px;text-align:center;font-size:9px;color:#64748b">' + uom + "</td>"
                    + '<td style="padding:7px 9px;text-align:right;font-family:monospace;font-weight:800;color:#0d9488">' + (f"{total_qty:,.3f}" if total_qty else "&mdash;") + "</td>"
                    + '<td style="padding:7px 9px;text-align:right;font-family:monospace;color:#888">' + (f"{pqty:,.3f}" if pqty else "&mdash;") + "</td>"
                    + '<td style="padding:7px 9px;text-align:right;font-family:monospace">' + (fi(rate) if rate else "&mdash;") + "</td>"
                    + '<td style="padding:7px 9px;text-align:right;font-family:monospace;font-weight:700">' + (fi(amt) if amt else "&mdash;") + "</td>"
                    + "</tr>"
                )

            freight = float(grn.get("freight_charge") or 0)
            packing = float(grn.get("packing_charge") or 0)
            taxable = total + freight + packing
            grand   = taxable + cgst + sgst

            html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
            <title>{grn_num}</title>
            <style>
            *{{box-sizing:border-box;margin:0;padding:0;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
            body{{font-family:Arial,Helvetica,sans-serif;font-size:11px;color:#111;background:#fff;padding:20px 28px}}
            .hdr{{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:2px solid #166534;padding-bottom:8px}}
            .co{{font-size:22px;font-weight:900;color:#166534;text-transform:uppercase}}
            .cosub{{font-size:9px;color:#666;margin-top:2px}}
            .pnum{{font-size:13px;font-weight:800;font-family:monospace;color:#166534;text-align:right}}
            .bar{{display:grid;border:1px solid #ccc;border-top:none}}
            .bar3{{grid-template-columns:1fr 1fr 1fr}}
            .bc{{padding:5px 9px;border-right:1px solid #ccc}}.bc:last-child{{border-right:none}}
            .bl{{font-size:7px;font-weight:800;color:#888;text-transform:uppercase;letter-spacing:.6px;margin-bottom:1px}}
            .bv{{font-size:10.5px;font-weight:600}}
            .adg{{display:grid;grid-template-columns:1fr 1fr;border:1px solid #ccc;border-top:none}}
            .ab{{padding:8px 10px;border-right:1px solid #ccc;font-size:10px;line-height:1.65}}.ab:last-child{{border-right:none}}
            .al{{font-size:7px;font-weight:800;color:#888;text-transform:uppercase;letter-spacing:.6px;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid #eee}}
            table{{width:100%;border-collapse:collapse}}
            thead tr{{background:#166534}}
            th{{color:#fff;padding:6px 8px;font-size:8px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;border-right:1px solid rgba(255,255,255,.2);text-align:right}}
            th:first-child{{text-align:center}}th:nth-child(2){{text-align:left}}th:last-child{{border-right:none}}
            tbody tr{{border-bottom:1px solid #ddd}}
            tbody tr:nth-child(odd){{background:#f9fafb}}
            td{{padding:6px 8px;font-size:10.5px;vertical-align:top;border-right:1px solid #eee}}td:last-child{{border-right:none}}
            .ftrow td{{padding:4px 8px;font-size:10px}}
            .ftrow-total td{{font-weight:800;font-size:12px;background:#f0fdf4;border-top:2px solid #166534}}
            .amt-words{{border:1px solid #ccc;border-top:none;padding:6px 10px;font-size:10px}}
            .sig{{display:grid;grid-template-columns:1fr 1fr;border:1px solid #ccc;border-top:none}}
            .sb{{padding:9px 10px;border-right:1px solid #ccc;min-height:48px}}.sb:last-child{{border-right:none;text-align:right}}
            .footer{{text-align:center;font-size:8.5px;color:#94a3b8;margin-top:6px;border-top:1px solid #eee;padding-top:5px}}
            </style></head><body>
            <div class="hdr">
              <div><div class="co">Goods Receipt Note</div><div class="cosub">HCP Wellness Pvt Ltd</div></div>
              <div class="pnum">{grn_num}</div>
            </div>
            <div class="bar bar3">
              <div class="bc"><div class="bl">GRN Number</div><div class="bv">{grn_num}</div></div>
              <div class="bc"><div class="bl">GRN Date</div><div class="bv">{grn_date}</div></div>
              <div class="bc"><div class="bl">Supplier</div><div class="bv">{supplier}</div></div>
            </div>
            <div class="adg">
              <div class="ab"><div class="al">Supplier Details</div>{sup_lines}</div>
              <div class="ab"><div class="al">Linked PO Details</div>
                {'<table style="width:100%;border-collapse:collapse;margin-top:2px"><thead><tr style="background:#f1f5f9"><th style="padding:3px 8px;text-align:left;font-size:8px;color:#666;font-weight:700">PO NUMBER</th><th style="padding:3px 8px;text-align:left;font-size:8px;color:#666;font-weight:700">PO DATE</th></tr></thead><tbody>'+inv_rows+'</tbody></table>' if inv_rows else '—'}
              </div>
            </div>
            <table><thead><tr>
              <th style="width:22px;text-align:center">Sl</th>
              <th style="text-align:left">Material Description &amp; Details</th>
              <th style="width:130px;text-align:center">Pkgs &times; Qty/Pkg</th>
              <th style="width:40px;text-align:center">UOM</th>
              <th style="width:85px">Recd Qty</th>
              <th style="width:75px">PO Qty</th>
              <th style="width:80px">Rate (₹)</th>
              <th style="width:100px">Amount (₹)</th>
            </tr></thead>
            <tbody>{item_rows}</tbody>
            <tfoot>
              {'<tr class="ftrow"><td colspan="5" style="text-align:right;color:#555">Freight</td><td style="text-align:right;font-family:monospace">'+fi(freight)+'</td></tr>' if freight else ''}
              {'<tr class="ftrow"><td colspan="5" style="text-align:right;color:#555">Packing</td><td style="text-align:right;font-family:monospace">'+fi(packing)+'</td></tr>' if packing else ''}
              <tr class="ftrow"><td colspan="5" style="text-align:right;color:#555">Taxable Amount</td><td style="text-align:right;font-family:monospace">{fi(taxable)}</td></tr>
              {'<tr class="ftrow"><td colspan="5" style="text-align:right;color:#555">CGST</td><td style="text-align:right;font-family:monospace">'+fi(cgst)+'</td></tr>' if cgst else ''}
              {'<tr class="ftrow"><td colspan="5" style="text-align:right;color:#555">SGST</td><td style="text-align:right;font-family:monospace">'+fi(sgst)+'</td></tr>' if sgst else ''}
              <tr class="ftrow-total"><td colspan="5" style="text-align:right;text-transform:uppercase;letter-spacing:.5px">Grand Total</td><td style="text-align:right;font-family:monospace;font-size:13px">{fi(grand)}</td></tr>
            </tfoot></table>
            <div class="amt-words"><strong>Amount in Words:</strong>&nbsp; INR {grand:,.2f} Only</div>
            <div class="sig">
              <div class="sb"><div style="font-size:7px;font-weight:800;color:#888;text-transform:uppercase;letter-spacing:.6px;margin-bottom:4px">Received By (Store)</div><div style="margin-top:28px;font-size:9px;color:#888">Name &amp; Signature</div></div>
              <div class="sb"><div style="font-size:7px;font-weight:800;color:#888;text-transform:uppercase;letter-spacing:.6px;margin-bottom:4px">Authorised By</div><div style="margin-top:28px;font-size:9px;color:#888">for HCP Wellness Pvt Ltd</div></div>
            </div>
            <div class="footer">SUBJECT TO AHMEDABAD JURISDICTION | This is a Computer Generated Document</div>
            </body></html>"""

            options = {'page-size':'A4','margin-top':'10mm','margin-right':'10mm',
                       'margin-bottom':'10mm','margin-left':'10mm','encoding':'UTF-8',
                       'no-outline':None,'disable-smart-shrinking':None,'enable-local-file-access':None,
                       'quiet':None}
            tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            tmp.close()
            try:
                pdfkit.from_string(html, tmp.name, options=options)
            except OSError as pdf_err:
                # wkhtmltopdf exits non-zero on warnings — check if PDF was still created
                if not _os.path.exists(tmp.name) or _os.path.getsize(tmp.name) == 0:
                    raise pdf_err
            with open(tmp.name,'rb') as f: pdf_bytes = f.read()
            _os.unlink(tmp.name)
            from flask import Response
            safe = grn_num.replace('/','_')
            return Response(pdf_bytes, mimetype='application/pdf',
                headers={'Content-Disposition': f'attachment; filename="GRN_{safe}.pdf"'})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    @app.route("/api/procurement/suppliers/ledger_pdf", methods=["POST"])
    @procurement_required
    def api_supplier_ledger_pdf():
        """Generate full Supplier PO & GRN Ledger as PDF."""
        import pdfkit, tempfile, os as _os, json as _json
        d = request.get_json() or {}
        supplier_name = (d.get("supplier_name") or "").strip()
        po_ids  = d.get("po_ids", [])
        grn_ids = d.get("grn_ids", [])
        if not supplier_name:
            return jsonify({"status":"error","message":"supplier_name required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status":"error","message":"DB connection failed"}), 500

            MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
            def fd(d):
                if not d: return '—'
                p = str(d).split('-')
                return f"{p[2]}/{MONTHS[int(p[1])-1]}/{p[0]}" if len(p)==3 else str(d)
            def fi(n):
                try: return f"₹{float(n):,.2f}"
                except: return "—"

            # Supplier details
            sup = conn.execute(
                "SELECT * FROM procurement_suppliers WHERE LOWER(TRIM(supplier_name))=LOWER(TRIM(%s))",
                (supplier_name,)
            ).fetchone()
            sup = dict(sup) if sup else {}

            # POs with items
            pos = []
            for pid in po_ids:
                po = conn.execute("SELECT * FROM procurement_purchase_orders WHERE id=%s",(pid,)).fetchone()
                if not po: continue
                po = dict(po)
                for f in ("po_date","delivery_date"): 
                    if po.get(f): po[f] = str(po[f])
                items = conn.execute("SELECT * FROM procurement_po_items WHERE po_id=%s ORDER BY id",(pid,)).fetchall()
                po["items"] = [{**dict(i),"qty":float(i["qty"] or 0),"rate":float(i["rate"] or 0),"amount":float(i["amount"] or 0)} for i in items]
                pos.append(po)

            # GRNs with items
            grns = []
            for gid in grn_ids:
                grn = conn.execute("SELECT * FROM procurement_grn WHERE id=%s",(gid,)).fetchone()
                if not grn: continue
                grn = dict(grn)
                for f in ("grn_date","invoice_date","created_at","updated_at"):
                    if grn.get(f): grn[f] = str(grn[f])
                items = conn.execute("SELECT * FROM procurement_grn_items WHERE grn_id=%s ORDER BY id",(gid,)).fetchall()
                grn["items"] = [{**dict(i),"received_qty":float(i["received_qty"] or 0),"rate":float(i["rate"] or 0),"amount":float(i["amount"] or 0)} for i in items]
                grns.append(grn)
            conn.close()

            # Summary stats
            total_po_val  = sum(float(p.get("grand_total") or 0) for p in pos)
            total_grn_val = sum(float(g.get("grand_total") or 0) for g in grns)
            open_pos      = [p for p in pos if p.get("status") in ("open","partial","approved")]
            closed_pos    = [p for p in pos if p.get("status") == "closed"]
            pending_val   = sum(float(p.get("grand_total") or 0) for p in open_pos)

            sup_lines = f"<strong>{supplier_name}</strong>"
            if sup.get("address"):        sup_lines += f"<br>{sup['address']}"
            if sup.get("gst_number"):     sup_lines += f"<br>GSTIN: <strong>{sup['gst_number']}</strong>"
            if sup.get("contact_person"): sup_lines += f"<br>Contact: {sup['contact_person']}" + (f" | {sup['phone']}" if sup.get("phone") else "")
            if sup.get("email"):          sup_lines += f"<br>E-Mail: {sup['email']}"

            # Build PO tables
            po_html = ""
            for po in pos:
                status_color = {"closed":"#16a34a","partial":"#d97706","open":"#2563eb","cancelled":"#dc2626"}.get(po.get("status","open"),"#2563eb")
                po_html += f"""
                <div style="margin-bottom:16px;border:1px solid #ddd;border-radius:6px;overflow:hidden">
                  <div style="display:flex;justify-content:space-between;align-items:center;padding:8px 12px;background:#f8fafc;border-bottom:1px solid #e2e8f0">
                    <div style="display:flex;align-items:center;gap:12px">
                      <strong style="font-family:monospace;color:#0d9488">{po.get('po_num','—')}</strong>
                      <span style="font-size:9px;font-weight:800;padding:2px 8px;border-radius:20px;background:{status_color}22;color:{status_color}">{(po.get('status','open')).upper()}</span>
                      <span style="font-size:10px;color:#666">Date: {fd(po.get('po_date'))}</span>
                      {'<span style="font-size:10px;color:#666">Expected: '+fd(po.get("delivery_date"))+'</span>' if po.get("delivery_date") else ''}
                    </div>
                    <strong style="font-family:monospace">{fi(po.get('grand_total'))}</strong>
                  </div>
                  <table style="width:100%;border-collapse:collapse;font-size:10px">
                    <thead><tr style="background:#f1f5f9">
                      <th style="padding:5px 10px;text-align:center;color:#666;font-weight:700;width:24px">#</th>
                      <th style="padding:5px 10px;text-align:left;color:#666;font-weight:700">Material</th>
                      <th style="padding:5px 10px;text-align:right;color:#666;font-weight:700;width:90px">Qty (kg)</th>
                      <th style="padding:5px 10px;text-align:right;color:#666;font-weight:700;width:90px">Rate (₹)</th>
                      <th style="padding:5px 10px;text-align:right;color:#666;font-weight:700;width:100px">Amount (₹)</th>
                    </tr></thead><tbody>"""
                for ii, it in enumerate(po.get("items",[]), 1):
                    qty = float(it.get("qty") or 0); rate = float(it.get("rate") or 0); amt = qty*rate
                    po_html += f'<tr style="border-bottom:1px solid #f0f0f0"><td style="padding:5px 10px;text-align:center;color:#888">{ii}</td><td style="padding:5px 10px"><strong>{it.get("material","")}</strong></td><td style="padding:5px 10px;text-align:right;font-family:monospace">{f"{qty:,.3f}" if qty else "—"}</td><td style="padding:5px 10px;text-align:right;font-family:monospace">{fi(rate) if rate else "—"}</td><td style="padding:5px 10px;text-align:right;font-family:monospace;font-weight:700">{fi(amt) if amt else "—"}</td></tr>'
                po_html += "</tbody></table></div>"

            # Build GRN tables
            grn_html = ""
            for grn in grns:
                po_invs = []
                if grn.get("po_invoices"):
                    try:
                        pis = _json.loads(grn["po_invoices"]) if isinstance(grn["po_invoices"],str) else grn["po_invoices"]
                        po_invs = [i.get("po_num","") for i in pis if i.get("po_num")]
                    except: pass
                if grn.get("po_num") and grn["po_num"] not in po_invs:
                    po_invs.insert(0, grn["po_num"])
                grn_html += f"""
                <div style="margin-bottom:16px;border:1px solid #ddd;border-left:3px solid #16a34a;border-radius:6px;overflow:hidden">
                  <div style="display:flex;justify-content:space-between;align-items:center;padding:8px 12px;background:#f0fdf4;border-bottom:1px solid #dcfce7">
                    <div style="display:flex;align-items:center;gap:12px">
                      <strong style="font-family:monospace;color:#16a34a">{grn.get('grn_num','—')}</strong>
                      <span style="font-size:10px;color:#666">Date: {fd(grn.get('grn_date'))}</span>
                      {'<span style="font-size:10px;color:#0d9488">POs: '+', '.join(po_invs)+'</span>' if po_invs else ''}
                    </div>
                    <strong style="font-family:monospace">{fi(grn.get('grand_total'))}</strong>
                  </div>
                  <table style="width:100%;border-collapse:collapse;font-size:10px">
                    <thead><tr style="background:#f1f5f9">
                      <th style="padding:5px 10px;text-align:center;color:#666;font-weight:700;width:24px">#</th>
                      <th style="padding:5px 10px;text-align:left;color:#666;font-weight:700">Material</th>
                      <th style="padding:5px 10px;text-align:right;color:#666;font-weight:700;width:80px">PO Qty</th>
                      <th style="padding:5px 10px;text-align:right;color:#666;font-weight:700;width:80px">Recd Qty</th>
                      <th style="padding:5px 10px;text-align:right;color:#666;font-weight:700;width:90px">Rate (₹)</th>
                      <th style="padding:5px 10px;text-align:right;color:#666;font-weight:700;width:100px">Amount (₹)</th>
                    </tr></thead><tbody>"""
                for ii, it in enumerate(grn.get("items",[]), 1):
                    pqty=float(it.get("po_qty") or 0); rqty=float(it.get("received_qty") or 0); rate=float(it.get("rate") or 0); amt=rqty*rate
                    grn_html += f'<tr style="border-bottom:1px solid #f0f0f0"><td style="padding:5px 10px;text-align:center;color:#888">{ii}</td><td style="padding:5px 10px"><strong>{it.get("material","")}</strong></td><td style="padding:5px 10px;text-align:right;font-family:monospace;color:#888">{f"{pqty:,.3f}" if pqty else "—"}</td><td style="padding:5px 10px;text-align:right;font-family:monospace;font-weight:700;color:#16a34a">{f"{rqty:,.3f}" if rqty else "—"}</td><td style="padding:5px 10px;text-align:right;font-family:monospace">{fi(rate) if rate else "—"}</td><td style="padding:5px 10px;text-align:right;font-family:monospace;font-weight:700">{fi(amt) if amt else "—"}</td></tr>'
                grn_html += "</tbody></table></div>"

            from datetime import date as _date
            today_str = fd(str(_date.today()))

            html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
            <title>Ledger — {supplier_name}</title>
            <style>
            *{{box-sizing:border-box;margin:0;padding:0;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
            body{{font-family:Arial,Helvetica,sans-serif;font-size:11px;color:#111;background:#fff;padding:20px 28px}}
            h2{{font-size:15px;font-weight:900;color:#1e3a8a;margin-bottom:2px}}
            .sub{{font-size:9px;color:#666;margin-bottom:0}}
            .hdr{{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:2px solid #1e3a8a;padding-bottom:10px;margin-bottom:12px}}
            .summary{{display:grid;grid-template-columns:repeat(5,1fr);gap:8px;margin-bottom:16px}}
            .sc{{padding:10px 12px;border:1px solid #e2e8f0;border-radius:6px;background:#f8fafc}}
            .sl{{font-size:8px;font-weight:700;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-bottom:3px}}
            .sv{{font-size:16px;font-weight:900}}
            .ss{{font-size:9px;color:#888;margin-top:2px}}
            .sec{{font-size:11px;font-weight:800;color:#1e3a8a;text-transform:uppercase;letter-spacing:.6px;border-bottom:2px solid #e2e8f0;padding-bottom:5px;margin-bottom:10px;margin-top:16px}}
            .footer{{text-align:center;font-size:8px;color:#94a3b8;margin-top:10px;border-top:1px solid #eee;padding-top:5px}}
            </style></head><body>
            <div class="hdr">
              <div>
                <div style="font-size:8px;color:#888;text-transform:uppercase;letter-spacing:.8px;margin-bottom:2px">HCP WELLNESS PVT LTD</div>
                <h2>Supplier PO &amp; GRN Ledger</h2>
                <div class="sub">{sup_lines}</div>
              </div>
              <div style="text-align:right">
                <div style="font-size:8px;color:#888">Generated: {today_str}</div>
              </div>
            </div>
            <div class="summary">
              <div class="sc"><div class="sl">Total POs</div><div class="sv" style="color:#1d4ed8">{len(pos)}</div><div class="ss">{fi(total_po_val)}</div></div>
              <div class="sc"><div class="sl">Pending/Open</div><div class="sv" style="color:#d97706">{len(open_pos)}</div><div class="ss">{fi(pending_val)} pending</div></div>
              <div class="sc"><div class="sl">Closed</div><div class="sv" style="color:#16a34a">{len(closed_pos)}</div><div class="ss">fully received</div></div>
              <div class="sc"><div class="sl">GRNs Received</div><div class="sv" style="color:#0d9488">{len(grns)}</div><div class="ss">{fi(total_grn_val)}</div></div>
              <div class="sc"><div class="sl">Net Payable</div><div class="sv" style="color:#7c3aed">{fi(total_grn_val)}</div><div class="ss">GRN value</div></div>
            </div>
            <div class="sec">Purchase Orders</div>
            {po_html if po_html else '<div style="color:#888;padding:10px">No Purchase Orders</div>'}
            <div class="sec">Goods Receipt Notes</div>
            {grn_html if grn_html else '<div style="color:#888;padding:10px">No GRNs</div>'}
            <div class="footer">SUBJECT TO AHMEDABAD JURISDICTION | HCP Wellness Pvt Ltd | This is a Computer Generated Document</div>
            </body></html>"""

            options = {'page-size':'A4','margin-top':'10mm','margin-right':'8mm',
                       'margin-bottom':'10mm','margin-left':'8mm','encoding':'UTF-8',
                       'no-outline':None,'disable-smart-shrinking':None,'enable-local-file-access':None}
            tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            tmp.close()
            try:
                pdfkit.from_string(html, tmp.name, options=options)
            except OSError as pdf_err:
                if not _os.path.exists(tmp.name) or _os.path.getsize(tmp.name) == 0:
                    raise pdf_err
            with open(tmp.name,'rb') as f: pdf_bytes = f.read()
            _os.unlink(tmp.name)
            from flask import Response
            safe = supplier_name.replace(' ','_')[:30]
            return Response(pdf_bytes, mimetype='application/pdf',
                headers={'Content-Disposition': f'attachment; filename="Ledger_{safe}.pdf"'})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500


    @app.route("/api/procurement/grn/save", methods=["POST"])
    @procurement_required
    def api_grn_save():
        import re as _re, json as _json
        d          = request.get_json() or {}
        grn_id     = d.get("id")
        supplier   = (d.get("supplier_name") or "").strip()
        grn_date   = d.get("grn_date") or None
        # Legacy single invoice (for backwards compat / single-PO GRN)
        invoice_num  = (d.get("invoice_num") or "").strip() or None
        invoice_date = d.get("invoice_date") or None
        # Multi-PO invoices: [{po_id, po_num, invoice_num, invoice_date}]
        po_invoices_raw = d.get("po_invoices") or []
        po_invoices_json = _json.dumps(po_invoices_raw) if po_invoices_raw else None
        # If only one PO-invoice entry, also populate the legacy fields
        if po_invoices_raw and len(po_invoices_raw) == 1:
            invoice_num  = (po_invoices_raw[0].get("invoice_num") or "").strip() or invoice_num
            invoice_date = po_invoices_raw[0].get("invoice_date") or invoice_date
        po_id      = d.get("po_id") or None
        po_num     = (d.get("po_num") or "").strip() or None
        status     = (d.get("status") or "open").strip()
        remarks    = (d.get("remarks") or "").strip() or None
        items      = d.get("items") or []
        uid        = session.get("UID", "")
        try:
            freight = float(d.get("freight_charge") or 0) or None
        except: freight = None
        try:
            packing = float(d.get("packing_charge") or 0) or None
        except: packing = None
        if not supplier:
            return jsonify({"status":"error","message":"supplier_name required"}), 400

        # Compute totals
        unload_location = (d.get("unload_location") or "").strip() or None
        voucher_type_name = (d.get("voucher_type_name") or "").strip() or None

        grand_total = 0.0
        for item in items:
            try:
                rqty = float(item.get("received_qty") or 0)
                rate = float(item.get("rate") or 0)
                grand_total += rqty * rate
            except: pass
        grand_total += (freight or 0) + (packing or 0)

        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500

            if grn_id:
                # Fetch old items to reverse stock before re-inserting
                old_items = conn.execute(
                    "SELECT material, received_qty FROM procurement_grn_items WHERE grn_id=%s", (grn_id,)
                ).fetchall()
                conn.execute("""
                    UPDATE procurement_grn
                    SET supplier_name=%s, grn_date=%s, invoice_num=%s, invoice_date=%s,
                        po_id=%s, po_num=%s, status=%s, remarks=%s,
                        freight_charge=%s, packing_charge=%s, grand_total=%s,
                        po_invoices=%s, unload_location=%s, voucher_type_name=%s, updated_by=%s
                    WHERE id=%s
                """, (supplier, grn_date, invoice_num, invoice_date,
                      po_id, po_num, status, remarks,
                      freight, packing, grand_total,
                      po_invoices_json, unload_location, voucher_type_name, uid, grn_id))
                conn.execute("DELETE FROM procurement_grn_items WHERE grn_id=%s", (grn_id,))
                # Reverse old stock
                for oi in old_items:
                    omat = (oi["material"] or "").strip()
                    oqty = float(oi["received_qty"] or 0)
                    if omat and oqty > 0:
                        try:
                            conn.execute("""
                                UPDATE procurement_materials
                                SET in_stock_qty = COALESCE(in_stock_qty,0) - %s
                                WHERE material_name = %s
                            """, (oqty, omat))
                        except Exception: pass
            else:
                # ── Unified voucher numbering for GRN ──
                import re as _re2
                prefix = "GRN"
                suffix = ""
                digits = 4
                today_str = __import__('datetime').date.today().isoformat()
                try:
                    # Use voucher_type_name first, fall back to parent 'grn' key
                    vn_type_key = voucher_type_name or "grn"
                    vn_row = conn.execute(
                        "SELECT prefix, suffix, digits FROM procurement_voucher_numbering "
                        "WHERE voucher_type=%s AND valid_from <= %s AND valid_to >= %s "
                        "ORDER BY id DESC LIMIT 1",
                        (vn_type_key, today_str, today_str)
                    ).fetchone()
                    # Fall back to parent 'grn' if custom type has no style
                    if not vn_row and voucher_type_name:
                        vn_row = conn.execute(
                            "SELECT prefix, suffix, digits FROM procurement_voucher_numbering "
                            "WHERE voucher_type='grn' AND valid_from <= %s AND valid_to >= %s "
                            "ORDER BY id DESC LIMIT 1",
                            (today_str, today_str)
                        ).fetchone()
                    if vn_row:
                        prefix = (vn_row["prefix"] or "GRN").strip()
                        suffix = (vn_row["suffix"] or "").strip()
                        digits = int(vn_row["digits"] or 4)
                except Exception:
                    pass  # fall back to GRN/0001

                conn.execute("SELECT GET_LOCK('grn_num_lock',10) AS lk")
                pattern = (prefix + "/%") if prefix else "GRN/%"
                rows = conn.execute(
                    "SELECT grn_num FROM procurement_grn WHERE grn_num LIKE %s",
                    (pattern,)
                ).fetchall()
                max_seq = 0
                for row in rows:
                    nums = _re2.findall(r"(\d{" + str(digits) + r",})", row["grn_num"])
                    if nums: max_seq = max(max_seq, int(nums[-1]))
                next_seq = max_seq + 1
                num_str = str(next_seq).zfill(digits)
                parts = []
                if prefix: parts.append(prefix)
                parts.append(num_str)
                if suffix: parts.append(suffix)
                grn_num = "/".join(parts)
                conn.execute("SELECT RELEASE_LOCK('grn_num_lock')")
                conn.execute("""
                    INSERT INTO procurement_grn
                        (grn_num, supplier_name, grn_date, invoice_num, invoice_date,
                         po_id, po_num, status, remarks,
                         freight_charge, packing_charge, grand_total,
                         po_invoices, unload_location, voucher_type_name, created_by, updated_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (grn_num, supplier, grn_date, invoice_num, invoice_date,
                      po_id, po_num, status, remarks,
                      freight, packing, grand_total,
                      po_invoices_json, unload_location, voucher_type_name, uid, uid))
                grn_id = conn.execute("SELECT LAST_INSERT_ID() AS id").fetchone()["id"]

            # Insert items + update stock
            for item in items:
                mat = (item.get("material") or "").strip()
                if not mat: continue
                try:
                    rqty   = float(item.get("received_qty") or 0) or None
                    po_qty = float(item.get("po_qty") or 0) or None
                    rate   = float(item.get("rate") or 0) or None
                    amount = (rqty or 0) * (rate or 0) or None
                except: rqty = po_qty = rate = amount = None
                hsn        = (item.get("hsn_code")    or "").strip() or None
                loc        = (item.get("location")    or "").strip() or None
                inv_num    = (item.get("invoice_num") or "").strip() or None
                inv_date   = item.get("invoice_date") or None
                batch_num  = (item.get("batch_num")   or "").strip() or None
                mfg_date   = item.get("mfg_date")     or None
                expiry_date= item.get("expiry_date")  or None
                try: gst = float(item.get("gst_rate") or 0) or None
                except: gst = None
                conn.execute("""
                    INSERT INTO procurement_grn_items
                        (grn_id, material, po_qty, received_qty, qty_per_pkg, rate, amount,
                         hsn_code, gst_rate, location, invoice_num, invoice_date,
                         batch_num, mfg_date, expiry_date, packages, uom)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (grn_id, mat, po_qty, rqty,
                      float(item.get("qty_per_pkg") or 0) or None,
                      rate, amount,
                      hsn, gst, loc, inv_num, inv_date,
                      batch_num, mfg_date, expiry_date,
                      int(item.get("packages")) if item.get("packages") else None,
                      (item.get("uom") or "KG").strip().upper() or None))
                # Update stock
                if rqty and rqty > 0:
                    try:
                        conn.execute("""
                            UPDATE procurement_materials
                            SET in_stock_qty = COALESCE(in_stock_qty,0) + %s
                            WHERE material_name = %s
                        """, (rqty, mat))
                    except Exception: pass

            conn.commit(); conn.close()

            # ── Auto-update linked PO status based on GRN receipts ──────────────
            # Collect all PO IDs referenced by this GRN
            linked_po_ids = set()
            updated_po_statuses = {}  # {po_id: new_status}
            if po_id:
                linked_po_ids.add(int(po_id))
            for inv in (po_invoices_raw or []):
                if inv.get("po_id"):
                    linked_po_ids.add(int(inv["po_id"]))

            if linked_po_ids:
                try:
                    pc = sampling_portal.get_db_connection()
                    if pc:
                        for lpid in linked_po_ids:
                            # Get PO ordered qty per material
                            po_items = pc.execute(
                                "SELECT material, qty FROM procurement_po_items WHERE po_id=%s",
                                (lpid,)
                            ).fetchall()
                            if not po_items:
                                continue
                            # Don't touch cancelled POs
                            po_row = pc.execute(
                                "SELECT status FROM procurement_purchase_orders WHERE id=%s",
                                (lpid,)
                            ).fetchone()
                            if not po_row or po_row["status"] == "cancelled":
                                continue
                            # Sum received qty across ALL non-cancelled GRNs for this PO
                            total_received = {}
                            # Find all GRNs linked to this PO — via po_id column OR po_invoices JSON
                            grn_rows = pc.execute(
                                "SELECT id FROM procurement_grn WHERE po_id=%s AND status <> 'cancelled'",
                                (lpid,)
                            ).fetchall()
                            # Also find GRNs where PO is referenced inside po_invoices JSON
                            json_grns = pc.execute(
                                "SELECT id FROM procurement_grn WHERE po_invoices LIKE %s AND status <> 'cancelled'",
                                ('%"po_id": '+str(lpid)+'%',)
                            ).fetchall()
                            # Also try without space after colon
                            json_grns2 = pc.execute(
                                "SELECT id FROM procurement_grn WHERE po_invoices LIKE %s AND status <> 'cancelled'",
                                ('%"po_id":'+str(lpid)+'%',)
                            ).fetchall()
                            all_grn_ids = set(r["id"] for r in grn_rows)
                            all_grn_ids.update(r["id"] for r in json_grns)
                            all_grn_ids.update(r["id"] for r in json_grns2)
                            for grn_id_ref in all_grn_ids:
                                g_items = pc.execute(
                                    "SELECT material, received_qty FROM procurement_grn_items WHERE grn_id=%s",
                                    (grn_id_ref,)
                                ).fetchall()
                                for gi in g_items:
                                    mat = (gi["material"] or "").strip()
                                    rqty = float(gi["received_qty"] or 0)
                                    total_received[mat] = total_received.get(mat, 0) + rqty

                            # Subtract ROV (Rejection Out) quantities for this PO
                            # ROVs linked via tracking_grn_id to any of this PO's GRNs, or via po_id
                            rov_rows = pc.execute(
                                "SELECT id FROM procurement_grn WHERE grn_type='ROV' AND (po_id=%s OR tracking_grn_id IN (SELECT id FROM procurement_grn WHERE po_id=%s AND grn_type!='ROV')) AND status <> 'cancelled'",
                                (lpid, lpid)
                            ).fetchall()
                            # Also ROVs linked via po_invoices JSON
                            rov_json = pc.execute(
                                "SELECT id FROM procurement_grn WHERE grn_type='ROV' AND po_invoices LIKE %s AND status <> 'cancelled'",
                                ('%"po_id": '+str(lpid)+'%',)
                            ).fetchall()
                            all_rov_ids = set(r["id"] for r in rov_rows)
                            all_rov_ids.update(r["id"] for r in rov_json)
                            for rov_id_ref in all_rov_ids:
                                rov_items = pc.execute(
                                    "SELECT material, received_qty FROM procurement_grn_items WHERE grn_id=%s",
                                    (rov_id_ref,)
                                ).fetchall()
                                for ri in rov_items:
                                    mat = (ri["material"] or "").strip()
                                    rqty = float(ri["received_qty"] or 0)
                                    if mat and rqty > 0:
                                        total_received[mat] = total_received.get(mat, 0) - rqty

                            # Determine new PO status
                            any_received = any(v > 0 for v in total_received.values())
                            all_full = all(
                                total_received.get((i["material"] or "").strip(), 0)
                                >= float(i["qty"] or 0) - 0.001
                                for i in po_items
                            )
                            if not any_received:
                                new_po_status = "open"
                            elif all_full:
                                new_po_status = "closed"
                            else:
                                new_po_status = "partial"
                            pc.execute(
                                "UPDATE procurement_purchase_orders SET status=%s WHERE id=%s",
                                (new_po_status, lpid)
                            )
                            updated_po_statuses[lpid] = new_po_status
                        pc.commit()
                        pc.close()
                except Exception as _pe:
                    traceback.print_exc()  # non-fatal — log but don't fail the GRN save

            # Fetch grn_num if new
            if not d.get("id"):
                return jsonify({"status":"ok","id":grn_id,"grn_num":grn_num,"grand_total":grand_total,"po_statuses":updated_po_statuses})
            return jsonify({"status":"ok","id":grn_id,"grand_total":grand_total,"po_statuses":updated_po_statuses})

        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    @app.route("/api/procurement/grn/delete", methods=["POST"])
    @procurement_required
    def api_grn_delete():
        d = request.get_json() or {}
        grn_id = d.get("id")
        if not grn_id: return jsonify({"status":"error","message":"id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn: return jsonify({"status":"error","message":"DB connection failed"}), 500
            conn.execute("DELETE FROM procurement_grn WHERE id=%s", (grn_id,))
            conn.commit(); conn.close()
            return jsonify({"status":"ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status":"error","message":str(e)}), 500

    # ══════════════════════════════════════════════════════════════
    # MTV ENDPOINTS
    # ══════════════════════════════════════════════════════════════
    # ══════════════════════════════════════════════════════════════════════════

    @app.route("/api/procurement/mtv/list", methods=["GET"])
    @procurement_required
    def api_mtv_list():
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            rows = conn.execute("""
                SELECT m.*,
                       (SELECT COUNT(*) FROM procurement_mtv_items i WHERE i.mtv_id = m.id) AS item_count
                FROM procurement_mtv m
                ORDER BY m.mtv_date DESC, m.id DESC
            """).fetchall()
            conn.close()
            return jsonify({"status": "ok", "rows": [dict(r) for r in rows]})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/mtv/get", methods=["GET"])
    @procurement_required
    def api_mtv_get():
        mtv_id = request.args.get("id")
        if not mtv_id:
            return jsonify({"status": "error", "message": "id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            mtv = conn.execute(
                "SELECT * FROM procurement_mtv WHERE id=%s", (mtv_id,)
            ).fetchone()
            if not mtv:
                return jsonify({"status": "error", "message": "MTV not found"}), 404
            items = conn.execute(
                "SELECT * FROM procurement_mtv_items WHERE mtv_id=%s ORDER BY id", (mtv_id,)
            ).fetchall()
            conn.close()
            return jsonify({
                "status": "ok",
                "mtv":   dict(mtv),
                "items": [dict(i) for i in items]
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/mtv/save", methods=["POST"])
    @procurement_required
    def api_mtv_save():
        import re as _re
        d        = request.get_json() or {}
        mtv_id   = d.get("id")
        mtv_num  = (d.get("mtv_num")  or "").strip()
        mtv_date = d.get("mtv_date")  or None
        from_loc = (d.get("from_loc") or "").strip() or None
        to_loc   = (d.get("to_loc")   or "").strip() or None
        remarks  = (d.get("remarks")  or "").strip() or None
        status   = (d.get("status")   or "open").strip()
        items    = d.get("items")     or []
        voucher_type_name = (d.get("voucher_type_name") or "").strip() or None
        uid      = session.get("UID", "")

        if not mtv_num:
            return jsonify({"status": "error", "message": "mtv_num required"}), 400
        if not mtv_date:
            return jsonify({"status": "error", "message": "mtv_date required"}), 400

        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500

            # Auto-generate voucher number if new and no number supplied
            if not mtv_id:
                # Use unified voucher numbering
                today_str = __import__('datetime').date.today().isoformat()
                prefix = "MTV"; suffix = ""; digits = 4
                try:
                    vn_type_key = voucher_type_name or "mtv"
                    vn_row = conn.execute(
                        "SELECT prefix, suffix, digits FROM procurement_voucher_numbering "
                        "WHERE voucher_type=%s AND valid_from <= %s AND valid_to >= %s "
                        "ORDER BY id DESC LIMIT 1",
                        (vn_type_key, today_str, today_str)
                    ).fetchone()
                    # Fall back to parent 'mtv' if custom type has no style
                    if not vn_row and voucher_type_name:
                        vn_row = conn.execute(
                            "SELECT prefix, suffix, digits FROM procurement_voucher_numbering "
                            "WHERE voucher_type='mtv' AND valid_from <= %s AND valid_to >= %s "
                            "ORDER BY id DESC LIMIT 1",
                            (today_str, today_str)
                        ).fetchone()
                    if vn_row:
                        prefix = (vn_row["prefix"] or "MTV").strip()
                        suffix = (vn_row["suffix"] or "").strip()
                        digits = int(vn_row["digits"] or 4)
                except Exception:
                    pass
                # Find max seq from existing MTVs
                pattern = (prefix + "/%") if prefix else "%"
                existing = conn.execute(
                    "SELECT mtv_num FROM procurement_mtv WHERE mtv_num LIKE %s",
                    (pattern,)
                ).fetchall()
                max_seq = 0
                for row in existing:
                    nums = _re.findall(r"(\d{" + str(digits) + r",})", row["mtv_num"] or "")
                    if nums:
                        max_seq = max(max_seq, int(nums[-1]))
                seq     = max_seq + 1
                mtv_num = prefix + "/" + str(seq).zfill(digits) + ("/" + suffix if suffix else "")

            was_completed = False
            if mtv_id:
                old = conn.execute("SELECT status FROM procurement_mtv WHERE id=%s", (mtv_id,)).fetchone()
                was_completed = old and old["status"] == "completed"

                # If editing a completed MTV, reverse stock for old items first
                if was_completed and status != "completed":
                    old_items = conn.execute(
                        "SELECT material_name, qty FROM procurement_mtv_items WHERE mtv_id=%s", (mtv_id,)
                    ).fetchall()
                    for oi in old_items:
                        omat = (oi["material_name"] or "").strip()
                        oqty = float(oi["qty"] or 0)
                        if omat and oqty > 0:
                            try:
                                conn.execute("""
                                    UPDATE procurement_materials
                                    SET ordered_qty = GREATEST(0, COALESCE(ordered_qty,0) + %s)
                                    WHERE material_name = %s
                                """, (oqty, omat))
                            except Exception:
                                pass

                conn.execute("""
                    UPDATE procurement_mtv
                    SET mtv_num=%s, mtv_date=%s, from_loc=%s, to_loc=%s,
                        status=%s, remarks=%s, voucher_type_name=%s, updated_by=%s
                    WHERE id=%s
                """, (mtv_num, mtv_date, from_loc, to_loc, status, remarks, voucher_type_name, uid, mtv_id))
                conn.execute("DELETE FROM procurement_mtv_items WHERE mtv_id=%s", (mtv_id,))
            else:
                cur = conn.execute("""
                    INSERT INTO procurement_mtv
                    (mtv_num, mtv_date, from_loc, to_loc, status, remarks, voucher_type_name, created_by, updated_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (mtv_num, mtv_date, from_loc, to_loc, status, remarks, voucher_type_name, uid, uid))
                mtv_id = cur.lastrowid

            # ── Stock availability check (only when completing) ──────────────
            if status == "completed" and not was_completed:
                stock_map = _read_stksum()
                stock_errors = []
                for item in items:
                    mat_name = (item.get("material_name") or "").strip()
                    try:
                        qty = float(item.get("qty") or 0)
                    except Exception:
                        qty = 0
                    if not mat_name or qty <= 0:
                        continue
                    available = stock_map.get(mat_name.lower(), 0.0)
                    if qty > available:
                        stock_errors.append({
                            "material": mat_name,
                            "requested": qty,
                            "available": round(available, 3)
                        })
                if stock_errors:
                    conn.close()
                    return jsonify({
                        "status":      "error",
                        "message":     "Insufficient stock for " + str(len(stock_errors)) + " item(s)",
                        "stock_errors": stock_errors
                    }), 400

            # Insert items
            for item in items:
                mat_name = (item.get("material_name") or "").strip()
                if not mat_name:
                    continue
                try:
                    qty = float(item.get("qty") or 0)
                except Exception:
                    qty = 0
                if qty <= 0:
                    continue
                mat_id  = item.get("material_id") or None
                uom     = (item.get("uom") or "kg").strip()
                it_rmk  = (item.get("remarks") or "").strip() or None
                conn.execute("""
                    INSERT INTO procurement_mtv_items
                    (mtv_id, material_id, material_name, qty, qty_per_pkg, uom, packages, remarks)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """, (mtv_id, mat_id, mat_name, qty,
                         float(item.get("qty_per_pkg") or qty or 0) or None,
                         uom,
                         int(item.get("packages")) if item.get("packages") else None,
                         it_rmk))

            # If completing: deduct ordered_qty (proxy for in-transit/used stock)
            # Note: this adjusts ordered_qty to reflect material moved out.
            # You can swap this for in_stock_qty if your stock field differs.
            if status == "completed" and not was_completed:
                for item in items:
                    mat_name = (item.get("material_name") or "").strip()
                    if not mat_name:
                        continue
                    try:
                        qty = float(item.get("qty") or 0)
                    except Exception:
                        qty = 0
                    if qty <= 0:
                        continue
                    try:
                        conn.execute("""
                            UPDATE procurement_materials
                            SET ordered_qty = GREATEST(0, COALESCE(ordered_qty,0) - %s)
                            WHERE material_name = %s
                        """, (qty, mat_name))
                    except Exception:
                        pass

            conn.commit()
            conn.close()
            return jsonify({"status": "ok", "id": mtv_id, "mtv_num": mtv_num})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/mtv/cancel", methods=["POST"])
    @procurement_required
    def api_mtv_cancel():
        d = request.get_json() or {}
        mtv_id = d.get("id")
        if not mtv_id:
            return jsonify({"status": "error", "message": "id required"}), 400
        uid = session.get("UID", "")
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            conn.execute(
                "UPDATE procurement_mtv SET status='cancelled', updated_by=%s WHERE id=%s",
                (uid, mtv_id)
            )
            conn.commit()
            conn.close()
            return jsonify({"status": "ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/mtv/delete", methods=["POST"])
    @procurement_required
    def api_mtv_delete():
        d = request.get_json() or {}
        mtv_id = d.get("id")
        if not mtv_id:
            return jsonify({"status": "error", "message": "id required"}), 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return jsonify({"status": "error", "message": "DB connection failed"}), 500
            conn.execute("DELETE FROM procurement_mtv WHERE id=%s", (mtv_id,))
            conn.commit()
            conn.close()
            return jsonify({"status": "ok"})
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route("/api/procurement/mtv/print", methods=["GET"])
    @procurement_required
    def api_mtv_print():
        """Returns a simple printable HTML page for an MTV."""
        mtv_id = request.args.get("id")
        if not mtv_id:
            return "id required", 400
        try:
            conn = sampling_portal.get_db_connection()
            if not conn:
                return "DB connection failed", 500
            mtv = conn.execute("SELECT * FROM procurement_mtv WHERE id=%s", (mtv_id,)).fetchone()
            if not mtv:
                return "MTV not found", 404
            items = conn.execute(
                "SELECT * FROM procurement_mtv_items WHERE mtv_id=%s ORDER BY id", (mtv_id,)
            ).fetchall()
            conn.close()

            from datetime import date as _date

            def fd(d):
                if not d:
                    return "—"
                import datetime as _dt
                if isinstance(d, (_dt.date, _dt.datetime)):
                    dt = d
                else:
                    try:
                        dt = _dt.datetime.strptime(str(d)[:10], "%Y-%m-%d")
                    except Exception:
                        return str(d)
                months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
                return f"{str(dt.day).zfill(2)}/{months[dt.month-1]}/{dt.year}"

            def fmt_qty(v):
                try: return f"{float(v):,.3f}" if v else ""
                except: return str(v or "")
            def calc_row(item):
                pkgs     = float(item['packages'] or 0)
                qty_pp   = float(item['qty_per_pkg'] or 0)
                total    = float(item['qty'] or 0)
                # Derive qty_per_pkg if not stored (old records)
                if qty_pp == 0 and pkgs > 0 and total > 0:
                    qty_pp = total / pkgs
                elif qty_pp == 0:
                    qty_pp = total
                # Recalculate total from pkgs × qty_per_pkg when both present
                if pkgs > 0 and qty_pp > 0:
                    total = pkgs * qty_pp
                return pkgs, qty_pp, total
            rows_html = "".join((
                lambda idx, item: (
                    lambda pkgs, qpp, total: f"""
                <tr>
                  <td class='center muted'>{idx+1}</td>
                  <td class='mat-name'>{item['material_name'] or '—'}</td>
                  <td class='num' style='color:#7c3aed;font-weight:700'>{int(pkgs) if pkgs else '—'}</td>
                  <td class='center' style='color:#94a3b8;font-size:11px'>&times;</td>
                  <td class='num'>{fmt_qty(qpp) if qpp else ''}</td>
                  <td class='center' style='color:#0891b2;font-weight:600'>{item['uom'] or ''}</td>
                  <td class='num' style='color:#0d9488;font-weight:800;font-size:13px'>{fmt_qty(total)}</td>
                  <td class='muted'>{item['remarks'] or ''}</td>
                </tr>""")(* calc_row(item))
            )(idx, item) for idx, item in enumerate(items))

            status_color = {"open":"#3b82f6","completed":"#16a34a","cancelled":"#ef4444"}.get(mtv.get("status",""), "#64748b")

            vtname = (mtv.get('voucher_type_name') or 'Material Transfer').strip()

            import datetime as _dt2
            today_str = fd(_dt2.date.today())
            def _item_total(it):
                pkgs = float(it['packages'] or 0)
                qpp  = float(it['qty_per_pkg'] or 0)
                tot  = float(it['qty'] or 0)
                if qpp == 0 and pkgs > 0 and tot > 0: qpp = tot / pkgs
                return pkgs * qpp if pkgs > 0 and qpp > 0 else tot
            total_qty  = sum(_item_total(it) for it in items)
            total_pkgs = sum(int(it['packages'] or 0) for it in items)
            total_pkgs_str = str(total_pkgs) if total_pkgs else '\u2014'

            html = (
                "<!DOCTYPE html><html><head><meta charset='UTF-8'>"
                "<title>MTV " + str(mtv['mtv_num']) + "</title>"
                "<style>"
                "*{box-sizing:border-box;margin:0;padding:0}"
                "body{font-family:'Segoe UI',Arial,sans-serif;font-size:12.5px;color:#1e293b;background:#fff}"
                "@page{size:A4;margin:16mm 18mm}"
                "@media print{body{padding:0}}"
                ".band{background:#1d4ed8;padding:14px 24px;display:flex;align-items:center;justify-content:space-between}"
                ".band-left .co{font-size:17px;font-weight:800;color:#fff;letter-spacing:.2px}"
                ".band-left .co-sub{font-size:10px;color:rgba(255,255,255,.7);margin-top:1px;letter-spacing:.5px;text-transform:uppercase}"
                ".band-right .doc-title{font-size:15px;font-weight:700;color:#fff;text-align:right}"
                ".band-right .doc-num{font-family:monospace;font-size:12px;color:rgba(255,255,255,.85);text-align:right;margin-top:2px}"
                ".strip{background:#f0f9ff;border-bottom:1px solid #bae6fd;padding:6px 24px;display:flex;align-items:center;gap:10px}"
                ".sbadge{font-size:10px;font-weight:800;padding:2px 10px;border-radius:20px;text-transform:uppercase;letter-spacing:.6px;color:#fff;background:" + status_color + "}"
                ".vtlabel{font-size:10.5px;font-weight:600;color:#1d4ed8;background:#dbeafe;padding:2px 9px;border-radius:10px}"
                ".sdate{font-size:11px;color:#475569}"
                ".body{padding:20px 24px 16px}"
                ".route-card{display:flex;align-items:stretch;margin-bottom:18px;border:1px solid #e2e8f0;border-radius:10px;overflow:hidden}"
                ".route-loc{flex:1;padding:14px 18px;background:#f8fafc}"
                ".route-loc .loc-label{font-size:10px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.6px;margin-bottom:4px}"
                ".route-loc .loc-name{font-size:14px;font-weight:800;color:#0f172a}"
                ".route-arrow{width:52px;text-align:center;flex-shrink:0;background:#fff;border-left:1px solid #e2e8f0;border-right:1px solid #e2e8f0;color:#94a3b8;font-size:20px;display:flex;align-items:center;justify-content:center}"
                ".route-loc.to{background:#f0fdf4}"
                ".route-loc.to .loc-label{color:#15803d}"
                ".route-loc.to .loc-name{color:#14532d}"
                ".meta-row{display:flex;margin-bottom:18px;border:1px solid #e2e8f0;border-radius:10px;overflow:hidden}"
                ".mc{flex:1;padding:10px 16px;border-right:1px solid #e2e8f0;background:#fff}"
                ".mc:last-child{border-right:none}"
                ".mc .mcl{font-size:9.5px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.6px;margin-bottom:3px}"
                ".mc .mcv{font-size:12.5px;font-weight:700;color:#1e293b}"
                ".rmk-box{background:#fffbeb;border:1px solid #fde68a;border-radius:8px;padding:8px 14px;margin-bottom:16px;font-size:11.5px;color:#78350f}"
                ".rmk-box strong{font-size:9.5px;text-transform:uppercase;letter-spacing:.5px;color:#92400e;display:block;margin-bottom:2px}"
                ".tbl-wrap{border:1px solid #e2e8f0;border-radius:10px;overflow:hidden;margin-bottom:20px}"
                "table{width:100%;border-collapse:collapse}"
                "thead tr{background:#1d4ed8}"
                "th{padding:9px 10px;font-size:9.5px;font-weight:700;color:rgba(255,255,255,.9);text-transform:uppercase;letter-spacing:.5px}"
                "tbody tr{border-bottom:1px solid #f1f5f9}"
                "tbody tr:last-child{border-bottom:none}"
                "tbody tr:nth-child(even){background:#f8fafc}"
                "td{padding:8px 10px;font-size:12px;color:#1e293b;vertical-align:middle}"
                "td.num{font-family:monospace;text-align:right;font-weight:600}"
                "td.ctr{text-align:center}"
                "td.muted{color:#94a3b8;font-size:11px}"
                ".mat-name{font-weight:600}"
                ".sum-row td{background:#eff6ff;padding:8px 10px;font-size:11.5px;font-weight:700;color:#1d4ed8;border-top:2px solid #bfdbfe}"
                ".sig-section{display:grid;grid-template-columns:1fr 1fr 1fr;gap:20px;margin-top:24px}"
                ".sig-box{text-align:center}"
                ".sig-space{height:36px}"
                ".sig-line{border-top:1.5px solid #cbd5e1;margin-bottom:6px}"
                ".sig-label{font-size:10px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.5px}"
                ".sig-sub{font-size:9.5px;color:#94a3b8;margin-top:1px}"
                ".footer{margin-top:16px;padding-top:10px;border-top:1px solid #e2e8f0;display:flex;justify-content:space-between}"
                ".footer-l{font-size:9.5px;color:#94a3b8}"
                ".footer-r{font-size:9.5px;color:#94a3b8;font-family:monospace}"
                "</style></head><body>"

                "<div class='band'>"
                  "<div class='band-left'><div class='co'>HCP Wellness Pvt Ltd</div>"
                  "<div class='co-sub'>Material Transfer Voucher</div></div>"
                  "<div class='band-right'><div class='doc-title'>" + vtname + "</div>"
                  "<div class='doc-num'>" + str(mtv['mtv_num']) + "</div></div>"
                "</div>"

                "<div class='strip'>"
                  "<span class='sbadge'>" + (mtv.get('status') or 'open').upper() + "</span>"
                  "<span class='vtlabel'>" + vtname + "</span>"
                  "<span class='sdate'>Date: <strong>" + fd(mtv.get('mtv_date')) + "</strong></span>"
                  "<span class='sdate' style='margin-left:auto'>Printed: <strong>" + today_str + "</strong></span>"
                "</div>"

                "<div class='body'>"

                "<div class='route-card'>"
                  "<div class='route-loc'>"
                    "<div class='loc-label'>&#x2B06; From Location</div>"
                    "<div class='loc-name'>" + (mtv.get('from_loc') or '\u2014') + "</div>"
                  "</div>"
                  "<div class='route-arrow'>&#x2192;</div>"
                  "<div class='route-loc to'>"
                    "<div class='loc-label'>&#x2B07; To Location</div>"
                    "<div class='loc-name'>" + (mtv.get('to_loc') or '\u2014') + "</div>"
                  "</div>"
                "</div>"

                "<div class='meta-row'>"
                  "<div class='mc'><div class='mcl'>Transfer Date</div><div class='mcv'>" + fd(mtv.get('mtv_date')) + "</div></div>"
                  "<div class='mc'><div class='mcl'>MTV Number</div><div class='mcv' style='font-family:monospace;color:#1d4ed8'>" + str(mtv['mtv_num']) + "</div></div>"
                  "<div class='mc'><div class='mcl'>Total Items</div><div class='mcv'>" + str(len(items)) + "</div></div>"
                  "<div class='mc'><div class='mcl'>Prepared By</div><div class='mcv'>" + (mtv.get('created_by') or '\u2014') + "</div></div>"
                "</div>"

                + (("<div class='rmk-box'><strong>Remarks</strong>" + str(mtv.get('remarks','')) + "</div>") if mtv.get('remarks') else '')

                + "<div class='tbl-wrap'><table>"
                  "<thead><tr>"
                    "<th style='width:36px;text-align:center'>#</th>"
                    "<th>Material Name</th>"
                    "<th style='text-align:right;width:65px'>Pkgs</th>"
                    "<th style='text-align:center;width:20px'>&times;</th>"
                    "<th style='text-align:right;width:90px'>Qty/Pkg</th>"
                    "<th style='text-align:center;width:55px'>UOM</th>"
                    "<th style='text-align:right;width:95px'>Total Qty</th>"
                    "<th>Remarks</th>"
                  "</tr></thead>"
                  "<tbody>" + rows_html + "</tbody>"
                  "<tr class='sum-row'>"
                    "<td colspan='2'>Total</td>"
                    "<td class='num' style='color:#7c3aed'>" + total_pkgs_str + "</td>"
                    "<td></td>"
                    "<td></td>"
                    "<td></td>"
                    "<td class='num' style='color:#0d9488'>" + f"{total_qty:,.3f}" + "</td>"
                    "<td></td>"
                  "</tr>"
                "</table></div>"

                "<div class='sig-section'>"
                  "<div class='sig-box'><div class='sig-space'></div><div class='sig-line'></div>"
                    "<div class='sig-label'>Prepared By</div>"
                    "<div class='sig-sub'>" + (mtv.get('created_by') or '') + "</div></div>"
                  "<div class='sig-box'><div class='sig-space'></div><div class='sig-line'></div>"
                    "<div class='sig-label'>Checked By</div><div class='sig-sub'></div></div>"
                  "<div class='sig-box'><div class='sig-space'></div><div class='sig-line'></div>"
                    "<div class='sig-label'>Received By</div><div class='sig-sub'></div></div>"
                "</div>"

                "<div class='footer'>"
                  "<div class='footer-l'>HCP Wellness Pvt Ltd &nbsp;&middot;&nbsp; Material Transfer Document &nbsp;&middot;&nbsp; System Generated</div>"
                  "<div class='footer-r'>" + str(mtv['mtv_num']) + "</div>"
                "</div>"

                "</div>"
                "<script>window.onload=()=>window.print()</script>"
                "</body></html>"
            )
            return html, 200, {"Content-Type": "text/html; charset=utf-8"}
        except Exception as e:
            traceback.print_exc()
            return f"Error: {e}", 500


