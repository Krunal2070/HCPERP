# ══════════════════════════════════════════════════════════════════════════════
# production_initiater_routes.py
# Blueprint for all RM Store / Production Initiater routes.
# Registered in app.py via:
#   from production_initiater_routes import production_initiater_bp
#   app.register_blueprint(production_initiater_bp)
# ══════════════════════════════════════════════════════════════════════════════

from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, Response, send_file
from functools import wraps
from datetime import datetime
import os
import re
import subprocess
import requests
# LINUX_PATCH_APPLIED
try:
    import xlwings as xw  # Windows-only; safe to skip on Linux
except ImportError:
    xw = None  # Excel/petty-cash routes will return error if called
import tempfile
import io

from openpyxl import Workbook, load_workbook
import xmltodict

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

import sampling_portal

production_initiater_bp = Blueprint('production_initiater', __name__)


# ── Import shared helpers from portal_helpers (no circular imports) ───────────
from portal_helpers import can_access, _denied, _prod_role

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ═══════════════════════════════════════════════════════════════════════════════
# RM STORE DASHBOARD
# Access:  admin    → full edit + add batches
#          RM_Store → view + print only
# ═══════════════════════════════════════════════════════════════════════════════

FORMULATIONS_DIR = r'\\Hcp-server\secure data\FORMULATIONS'

# Worksheet names to exclude from the dropdown
_WS_EXCLUDE = {'INDEX', 'MAIN SHEET'}

# ── procurement_formulations (FVQ) helpers ───────────────────────────────────
# Table: procurement_formulations  (flat — one row per ingredient per batch)
# Columns: id, batch_name, product_code, material_name, supplier_name,
#          concentration, qty_kg, batch_size, batch_date, num_batches,
#          imported_at, imported_by, source_batch_name, manuf_process
_FVQ_TABLE = 'procurement_formulations'

def _fvq_get_ingredients(batch_name_key):
    """Return list of {material_name, supplier_name, concentration, qty_kg, batch_size} rows.
    qty_kg and batch_size are the stored reference values from procurement_formulations.
    The caller should scale: production_qty = stored_qty_kg * (production_batch_size / stored_batch_size)
    """
    conn = sampling_portal.get_db_connection()
    rows = conn.execute(
        f"SELECT material_name, supplier_name, concentration, qty_kg, batch_size "
        f"FROM `{_FVQ_TABLE}` "
        f"WHERE batch_name = %s ORDER BY id",
        (batch_name_key,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def _prod_role():
    """Return 'admin', 'rm_store', or None.
    Checks (in order):
      1. user_type == 'admin'
      2. user_type matches rm_store (any capitalisation)
      3. department field contains 'rm' or 'store' or 'production'
         (skipped if user_type is 'Production' — that role has its own page)
      4. user_permissions table has prod_view
    """
    if not session.get('logged_in'):
        return None
    r = session.get('User_Type', '') or ''
    if r.lower() == 'admin':
        return 'admin'
    if r.lower() in ('rm_store', 'rm store', 'rmstore', 'rm-store', 'stores', 'store'):
        return 'rm_store'

    # Production role has its own separate page — don't treat as rm_store
    if r.lower() == 'production':
        return None

    # Also check department field — catches users whose type wasn't set exactly
    dept = (session.get('department') or '').lower()
    if any(kw in dept for kw in ('rm_store', 'rm store', 'rmstore', 'production')):
        return 'rm_store'

    # Fallback: check permissions table
    user_id = session.get('user_id')
    if user_id:
        try:
            perms = sampling_portal.get_user_permissions(user_id) or {}
            if perms.get('prod_view'):
                if perms.get('prod_add') or perms.get('prod_print_sheets') or perms.get('prod_delete'):
                    return 'admin'
                return 'rm_store'
        except Exception:
            pass
    return None


def _ensure_processing_batches():
    """Create Processing_batches table if it doesn't exist."""
    conn = sampling_portal.get_db_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS Processing_batches (
            id                INT AUTO_INCREMENT PRIMARY KEY,
            file_name         VARCHAR(255)   NOT NULL,
            file_path         VARCHAR(500)   NOT NULL,
            worksheet         VARCHAR(255)   NOT NULL,
            batch_name        VARCHAR(500)   NOT NULL,
            batch_size        DECIMAL(10,3)  NOT NULL DEFAULT 0,
            no_of_batch       INT            NOT NULL DEFAULT 1,
            batch_type        VARCHAR(20)    NOT NULL DEFAULT 'Regular',
            sku_size          VARCHAR(100)   DEFAULT NULL,
            quantity          VARCHAR(100)   DEFAULT NULL,
            dispensed_batches INT            NOT NULL DEFAULT 0,
            added_by          VARCHAR(100)   NOT NULL,
            added_on          DATETIME       NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    try:
        conn.execute("ALTER TABLE Processing_batches ADD COLUMN dispensed_batches INT NOT NULL DEFAULT 0")
        conn.commit()
    except Exception:
        pass
    # Soft-delete support
    try:
        conn.execute("ALTER TABLE Processing_batches ADD COLUMN deleted_at DATETIME DEFAULT NULL")
        conn.commit()
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE Processing_batches ADD COLUMN deleted_by VARCHAR(100) DEFAULT NULL")
        conn.commit()
    except Exception:
        pass
    # Migrate quantity column from INT to VARCHAR(100) to support formats like "1000 / 1000 / 1000"
    try:
        conn.execute("ALTER TABLE Processing_batches MODIFY COLUMN quantity VARCHAR(100) DEFAULT NULL")
        conn.commit()
    except Exception:
        pass
    conn.close()

_ensure_processing_batches()


@production_initiater_bp.route('/production_initiater')
@login_required
def production_initiater_page():
    pr = _prod_role()
    if not pr:
        return _denied('RM Store Dashboard')
    # Load all batches from DB for the table
    today_str = datetime.now().strftime('%Y-%m-%d')
    conn = sampling_portal.get_db_connection()
    rows = conn.execute(
        "SELECT * FROM Processing_batches WHERE deleted_at IS NULL ORDER BY added_on DESC"
    ).fetchall()
    # Fetch today's dispensed counts from daily_dsp_summary (keyed by batch_id)
    dsp_today_rows = conn.execute(
        "SELECT batch_id, dispensed FROM daily_dsp_summary WHERE batch_date = %s",
        (today_str,)
    ).fetchall()
    conn.close()
    dsp_today_map = {int(r['batch_id']): int(r['dispensed']) for r in dsp_today_rows if r['batch_id']}
    batches = []
    for r in rows:
        b = dict(r)
        b['dispensed_today'] = dsp_today_map.get(int(b['id']), 0)
        batches.append(b)
    user_id   = session.get('user_id')
    user_type = session.get('User_Type', '')
    uid_lower = (session.get('UID') or '').lower()

    if user_type == 'admin':
        can_print_sheets    = True; can_add_batch = True
        can_delete          = True; can_print_labels = True
        can_edit            = True; can_dispensing_ready = True
    elif user_type in ('RM_Store', 'rm_store', 'Rm_Store') or pr == 'rm_store':
        # RM_Store: Print Labels + Add Batch + Dispensing Ready
        can_print_sheets    = False
        can_print_labels    = True   # ✅ Print Batch Labels
        can_edit            = True
        can_add_batch       = True   # ✅ Add Batch
        can_delete          = False
        can_dispensing_ready= True   # ✅ Dispensing Ready
    elif user_type == 'Stores':
        can_print_sheets    = False
        can_print_labels    = True
        can_edit            = True
        can_add_batch       = True
        can_delete          = True
        can_dispensing_ready= True
    else:
        try:    perms = sampling_portal.get_user_permissions(user_id) or {}
        except: perms = {}
        can_print_sheets    = bool(perms.get('prod_print_sheets'))
        can_add_batch       = bool(perms.get('prod_add'))
        can_delete          = bool(perms.get('prod_delete'))
        can_print_labels    = bool(perms.get('prod_print_labels', True))
        can_edit            = bool(perms.get('prod_edit')) or uid_lower == 'sonal' or uid_lower == 'rm_store'
        can_dispensing_ready= bool(perms.get('prod_dispensing_ready', True))

    return render_template('production_initiater.html',
        role=session.get('User_Type'),
        prod_role=pr,
        user_name=session.get('User_Name'),
        uid=session.get('UID', ''),
        batches=batches,
        can_print_sheets=can_print_sheets,
        can_add_batch=can_add_batch,
        can_delete=can_delete,
        can_print_labels=can_print_labels,
        can_edit=can_edit,
        can_dispensing_ready=can_dispensing_ready,
    )


# ==========================================
# OPEN PRODUCTION SHEET
# ==========================================

@production_initiater_bp.route("/open_production_sheet/<int:record_id>")
@login_required
def open_production_sheet(record_id):

    conn = sampling_portal.get_db_connection()

    row = conn.execute("""
        SELECT workbook_path
        FROM production_initiater
        WHERE id=%s
    """,(record_id,)).fetchone()

    conn.close()

    if not row:
        return jsonify({"status":"error","message":"File not found"})

    path = row["workbook_path"]

    try:
        os.startfile(path)
        return jsonify({"status":"success"})
    except Exception as e:
        return jsonify({"status":"error","message":str(e)})



# ── API: browse folders + files (supports subfolder navigation) ───────────────
@production_initiater_bp.route('/api/production/browse', methods=['GET'])
@login_required
def api_prod_browse():
    """Return folders and Excel files at the given path.
    Default path is FORMULATIONS_DIR. Subfolder param is relative to FORMULATIONS_DIR.
    Never allows navigating above FORMULATIONS_DIR (security guard)."""
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'error': 'Admin only'}), 403
    sub = request.args.get('sub', '').strip().strip('/\\')
    # Build real path, prevent traversal above root
    if sub:
        real_path = os.path.normpath(os.path.join(FORMULATIONS_DIR, sub))
        # Security: must still be inside FORMULATIONS_DIR
        if not real_path.startswith(os.path.normpath(FORMULATIONS_DIR)):
            return jsonify({'error': 'Access denied'}), 403
    else:
        real_path = FORMULATIONS_DIR
    try:
        entries = os.listdir(real_path)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    folders = sorted([
        e for e in entries
        if os.path.isdir(os.path.join(real_path, e)) and not e.startswith('~$')
    ], key=str.lower)

    files = sorted([
        e for e in entries
        if e.lower().endswith(('.xlsx', '.xlsm'))
        and not e.startswith('~$')
        and os.path.isfile(os.path.join(real_path, e))
    ], key=str.lower)

    # Compute relative sub-path for breadcrumb
    rel = os.path.relpath(real_path, FORMULATIONS_DIR)
    rel = '' if rel == '.' else rel.replace('\\', '/')
    # Parent sub-path (one level up)
    parent = '/'.join(rel.split('/')[:-1]) if rel else None

    return jsonify({
        'folders': folders,
        'files':   files,
        'sub':     rel,
        'parent':  parent,
        'root':    FORMULATIONS_DIR,
    })


# ── API: list worksheets of a chosen file ─────────────────────────────────────
@production_initiater_bp.route('/api/production/worksheets', methods=['GET'])
@login_required
def api_prod_worksheets():
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'error': 'Admin only'}), 403
    fname = request.args.get('file', '').strip()
    sub   = request.args.get('sub', '').strip().strip('/\\')   # relative subfolder only
    if not fname:
        return jsonify({'error': 'file param required'}), 400
    if not fname.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Not an Excel file'}), 400
    # Build path entirely server-side — never trust a full UNC path from client
    if sub:
        folder = os.path.normpath(os.path.join(FORMULATIONS_DIR, sub))
    else:
        folder = FORMULATIONS_DIR
    # Security: folder must still be inside FORMULATIONS_DIR
    norm_root = os.path.normpath(FORMULATIONS_DIR)
    norm_folder = os.path.normpath(folder)
    if not (norm_folder == norm_root or norm_folder.startswith(norm_root + os.sep)):
        return jsonify({'error': 'Access denied'}), 403
    fp = os.path.join(folder, fname)
    try:
        wb_tmp = load_workbook(fp, read_only=True, data_only=True)
        sheets = [
            s for s in wb_tmp.sheetnames
            if s.upper() not in _WS_EXCLUDE and not s.lower().startswith('sheet')
        ]
        wb_tmp.close()
        return jsonify({'sheets': sheets, 'path': folder})
    except FileNotFoundError:
        return jsonify({'error': f'File not found: {fp}'}), 404
    except Exception as e:
        import traceback
        print(f'[worksheets] ERROR for {fp}: {e}\n{traceback.format_exc()}')
        return jsonify({'error': str(e), 'file': fp}), 500


# ── API: list distinct FVQ batch names from procurement DB ────────────────────
@production_initiater_bp.route('/api/production/fvq_batches', methods=['GET'])
@login_required
def api_prod_fvq_batches():
    """Return all distinct batch names from the procurement_formulations table."""
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'error': 'Access denied'}), 403
    try:
        conn = sampling_portal.get_db_connection()
        rows = conn.execute(
            f"SELECT DISTINCT batch_name FROM `{_FVQ_TABLE}` ORDER BY batch_name"
        ).fetchall()
        conn.close()
        batches = [r['batch_name'] for r in rows]
        return jsonify({'status': 'ok', 'batches': batches})
    except Exception as e:
        import traceback
        print(f'[fvq_batches] ERROR: {e}\n{traceback.format_exc()}')
        return jsonify({'error': str(e)}), 500


# ── API: debug — return raw FVQ ingredient rows (admin only) ──────────────────
@production_initiater_bp.route('/api/production/debug_fvq', methods=['POST'])
@login_required
def api_debug_fvq():
    """Debug: return raw concentration and qty_kg values for a batch's ingredients."""
    if _prod_role() != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    d = request.json or {}
    fvq_name = d.get('fvq_name', '')
    conn = sampling_portal.get_db_connection()
    try:
        if not fvq_name:
            rows = conn.execute(
                f"SELECT batch_name, material_name, concentration, qty_kg, batch_size "
                f"FROM `{_FVQ_TABLE}` LIMIT 10"
            ).fetchall()
        else:
            rows = conn.execute(
                f"SELECT batch_name, material_name, concentration, qty_kg, batch_size "
                f"FROM `{_FVQ_TABLE}` WHERE batch_name = %s LIMIT 30",
                (fvq_name,)
            ).fetchall()
    finally:
        conn.close()
    return jsonify({'rows': [dict(r) for r in rows]})


# ── API: add batch to DB ───────────────────────────────────────────────────────
@production_initiater_bp.route('/api/production/add_batch', methods=['POST'])
@login_required
def api_prod_add_batch():
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    d = request.json or {}
    batch_size = d.get('batch_size')
    no_of_batch= d.get('no_of_batch')
    batch_type = d.get('batch_type', 'Regular')
    sku_size   = (d.get('sku_size') or '').strip() or None
    quantity   = d.get('quantity')

    # ── NEW: FVQ (procurement DB) source ──────────────────────────────────────
    fvq_batch_name = (d.get('fvq_batch_name') or '').strip()
    if fvq_batch_name:
        # Store sentinel values so label_data / rm_requirement know to use DB
        file_name  = '__FVQ__'
        file_path  = '__FVQ__'
        worksheet  = fvq_batch_name          # batch name is the lookup key
        batch_name = fvq_batch_name
    else:
        # ── LEGACY: Excel file source ─────────────────────────────────────────
        file_name = d.get('file_name', '').strip()
        sub       = d.get('sub', '').strip().strip('/\\')   # relative subfolder
        worksheet = d.get('worksheet', '').strip()
        if not file_name or not worksheet:
            return jsonify({'status': 'error', 'message': 'file_name and worksheet required'}), 400
        # Reconstruct full path server-side
        if sub:
            file_path = os.path.normpath(os.path.join(FORMULATIONS_DIR, sub))
        else:
            file_path = FORMULATIONS_DIR
        # Security check
        norm_root = os.path.normpath(FORMULATIONS_DIR)
        if not (os.path.normpath(file_path) == norm_root or
                os.path.normpath(file_path).startswith(norm_root + os.sep)):
            return jsonify({'status': 'error', 'message': 'Invalid path'}), 403
        batch_name = os.path.splitext(file_name)[0] + ' - ' + worksheet

    try:
        batch_size  = float(batch_size)
        no_of_batch = int(no_of_batch)
        quantity    = str(quantity).strip() if quantity not in (None, '', 'null') else None
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'message': 'Invalid numeric values'}), 400
    conn = sampling_portal.get_db_connection()
    try:
        conn.execute("""
            INSERT INTO Processing_batches
                (file_name, file_path, worksheet, batch_name,
                 batch_size, no_of_batch, batch_type, sku_size, quantity, added_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (file_name, file_path, worksheet, batch_name,
              batch_size, no_of_batch, batch_type, sku_size, quantity,
              session.get('User_Name', session.get('UID', 'admin'))))
        conn.commit()
        new_id = conn.execute("SELECT LAST_INSERT_ID()").fetchone()[0]
        conn.close()
        return jsonify({'status': 'ok', 'id': new_id, 'batch_name': batch_name})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ── API: delete batch from DB ──────────────────────────────────────────────────
# ── API: WhatsApp Batch Planning message ─────────────────────────────────────
@production_initiater_bp.route('/api/production/update_batch', methods=['POST'])
@login_required
def api_prod_update_batch():
    uid_lower = (session.get('UID') or '').lower()
    user_type = session.get('User_Type', '')
    if user_type != 'admin' and user_type != 'RM_Store' and uid_lower != 'sonal':
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    d   = request.get_json() or {}
    bid = d.get('id')
    if not bid:
        return jsonify({'status': 'error', 'message': 'Missing id'})
    allowed = ['batch_name','batch_size','no_of_batch','batch_type',
               'sku_size','quantity','dispensed_batches']
    fields = {k: d[k] for k in allowed if k in d}
    if not fields:
        return jsonify({'status': 'error', 'message': 'Nothing to update'})
    if 'batch_size'        in fields: fields['batch_size']        = float(fields['batch_size'])
    if 'no_of_batch'       in fields: fields['no_of_batch']       = int(fields['no_of_batch'])
    if 'dispensed_batches' in fields: fields['dispensed_batches'] = int(fields['dispensed_batches'])
    if 'quantity'          in fields:
        fields['quantity'] = str(fields['quantity']).strip() if fields['quantity'] not in (None,'') else None
    set_clause = ', '.join(f"`{k}`=%s" for k in fields)
    vals = list(fields.values()) + [bid]
    conn = sampling_portal.get_db_connection()
    conn.execute(f"UPDATE Processing_batches SET {set_clause} WHERE id=%s", vals)
    conn.commit()
    row = conn.execute("SELECT * FROM Processing_batches WHERE id=%s", (bid,)).fetchone()
    conn.close()
    return jsonify({'status': 'ok', 'batch': dict(row) if row else {}})


@production_initiater_bp.route('/api/production/whatsapp_batch', methods=['POST'])
@login_required
def api_prod_whatsapp_batch():
    """
    Builds a batch planning message and returns it as text.
    The frontend opens https://web.whatsapp.com/send?text=... directly —
    no Selenium, no Chrome automation required.
    """
    if not _prod_role():
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    d   = request.get_json() or {}
    ids = d.get('ids', [])
    if not ids:
        return jsonify({'status': 'error', 'message': 'No batches selected'}), 400

    placeholders = ','.join(['%s'] * len(ids))
    conn = sampling_portal.get_db_connection()
    rows = conn.execute(
        f"SELECT batch_name, batch_size FROM Processing_batches "
        f"WHERE id IN ({placeholders}) ORDER BY id",
        ids
    ).fetchall()
    conn.close()

    if not rows:
        return jsonify({'status': 'error', 'message': 'Batches not found'}), 404

    now      = datetime.now()
    time_str = now.strftime('%I:%M %p')
    date_str = now.strftime('%d %b %Y')

    lines = []
    for r in rows:
        lines.append(
            f"{r['batch_name']} @ {r['batch_size']} KG\n"
            f"Dispensing Time : {time_str}\n"
            f"Dispensing Date : {date_str}"
        )
    message = '\n\n'.join(lines)

    return jsonify({'status': 'ok', 'message': message})



@production_initiater_bp.route('/api/production/delete_batch', methods=['POST'])
@login_required
def api_prod_delete_batch():
    """Soft-delete: marks deleted_at/deleted_by. Permanent deletion happens via confirm_delete."""
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    bid = (request.json or {}).get('id')
    if not bid:
        return jsonify({'status': 'error', 'message': 'id required'}), 400
    deleted_by = session.get('User_Name') or session.get('UID') or 'unknown'
    conn = sampling_portal.get_db_connection()
    row = conn.execute(
        "SELECT * FROM Processing_batches WHERE id=%s AND deleted_at IS NULL", (bid,)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Batch not found'}), 404
    conn.execute(
        "UPDATE Processing_batches SET deleted_at=NOW(), deleted_by=%s WHERE id=%s",
        (deleted_by, bid)
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'batch': dict(row)})


@production_initiater_bp.route('/api/production/restore_batch', methods=['POST'])
@login_required
def api_prod_restore_batch():
    """Undo soft-delete: clears deleted_at/deleted_by within the 120s window."""
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    bid = (request.json or {}).get('id')
    if not bid:
        return jsonify({'status': 'error', 'message': 'id required'}), 400
    conn = sampling_portal.get_db_connection()
    # Only restore if still within 120s window
    row = conn.execute(
        "SELECT * FROM Processing_batches WHERE id=%s AND deleted_at IS NOT NULL "
        "AND deleted_at >= NOW() - INTERVAL 120 SECOND",
        (bid,)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Undo window has expired or batch not found'}), 404
    conn.execute(
        "UPDATE Processing_batches SET deleted_at=NULL, deleted_by=NULL WHERE id=%s", (bid,)
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok', 'batch': dict(row)})


@production_initiater_bp.route('/api/production/confirm_delete', methods=['POST'])
@login_required
def api_prod_confirm_delete():
    """Permanently delete a soft-deleted batch after the undo window has passed."""
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    bid = (request.json or {}).get('id')
    if not bid:
        return jsonify({'status': 'error', 'message': 'id required'}), 400
    conn = sampling_portal.get_db_connection()
    conn.execute(
        "DELETE FROM Processing_batches WHERE id=%s AND deleted_at IS NOT NULL", (bid,)
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


# ── API: list installed Windows printers ──────────────────────────────────────
@production_initiater_bp.route('/api/production/list_printers', methods=['GET'])
@login_required
def api_prod_list_printers():
    """Return list of installed Windows printers and the current default printer."""
    if not _prod_role():
        return jsonify({'error': 'Access denied'}), 403
    try:
        import win32print
        printers = [p[2] for p in win32print.EnumPrinters(
            win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS, None, 1
        )]
        default  = win32print.GetDefaultPrinter()
        return jsonify({'status': 'ok', 'printers': printers, 'default': default})
    except Exception as e:
        return jsonify({'status': 'ok', 'printers': [], 'default': '', 'error': str(e)})


# ── API: print batch sheets ───────────────────────────────────────────────────
@production_initiater_bp.route('/api/production/print_sheets', methods=['POST'])
@login_required
def api_prod_print_sheets():
    """
    Print batch front pages or process sheets with duplex-safe ordering.

    FRONT  page: print batches HIGH-id → LOW-id  (highest ID first)
                 → stack comes out with lowest ID on top face-up.
    PROCESS page: print batches LOW-id → HIGH-id (lowest ID first)
                 → after flipping the front stack, each process sheet
                   lands exactly behind its corresponding front page.

    Speed: Excel files are printed SEQUENTIALLY (one COM instance at a time)
    to avoid RPC/COM race conditions. Transient errors are retried up to 3×.
    """
    if not _prod_role():
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    d            = request.json or {}
    batch_ids    = d.get('batch_ids', [])
    print_type   = d.get('print_type', 'front')   # 'front' | 'process'
    printer_name = (d.get('printer_name') or '').strip()   # '' = use system default

    if not batch_ids:
        return jsonify({'status': 'error', 'message': 'No batches selected'}), 400

    # Fetch batch records from DB
    conn = sampling_portal.get_db_connection()
    placeholders = ','.join(['%s'] * len(batch_ids))
    rows = conn.execute(
        f"SELECT * FROM Processing_batches WHERE id IN ({placeholders})",
        tuple(batch_ids)
    ).fetchall()
    conn.close()

    if not rows:
        return jsonify({'status': 'error', 'message': 'Batches not found'}), 404

    # ── Sort order ────────────────────────────────────────────────────────────
    # Front  page: HIGH id first (83 → 68) so stack has 68 on top after printing
    # Process page: LOW  id first (68 → 83) so after flipping front stack,
    #               each process sheet is directly behind its front page
    rows_sorted = sorted(
        [dict(r) for r in rows],
        key=lambda r: r['id'],
        reverse=(print_type == 'front')   # True = descending for front
    )

    import xlwings as xw
    import time as _time
    from datetime import date as _date

    # COM error codes that are transient and worth retrying
    _COM_RETRYABLE = {
        -2147023170,   # The remote procedure call failed
        -2147023179,   # The RPC server is unavailable
        -2147024891,   # Access is denied (COM not ready)
        -2146959355,   # Server execution failed
        -2147417848,   # The object invoked has disconnected from its clients
    }
    _MAX_RETRIES   = 3
    _RETRY_DELAYS  = [2, 4, 8]   # seconds between attempts

    def _is_retryable(exc):
        """Return True if the exception is a transient COM/RPC error worth retrying."""
        s = str(exc)
        codes = [f'{c}' for c in _COM_RETRYABLE] + [hex(c & 0xFFFFFFFF) for c in _COM_RETRYABLE]
        keywords = ['remote procedure call', 'server execution failed',
                    'rpc server', 'disconnected from its clients']
        s_low = s.lower()
        return any(c in s for c in codes) or any(k in s_low for k in keywords)

    def _kill_xl_app(app_xl):
        """Safely quit an xlwings App instance, suppressing all errors."""
        try:
            if app_xl:
                app_xl.screen_updating = True
                app_xl.quit()
        except Exception:
            pass

    def _build_fvq_production_pdf(row):
        """Build a landscape A4 production sheet PDF for an FVQ batch.
        Layout: header | two info rows | ingredient table | footer.
        Returns PDF bytes suitable for browser display or printing."""
        import io as _io
        from reportlab.pdfgen import canvas as _rl_canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors as _rl_colors
        from reportlab.lib.units import mm
        from reportlab.lib.utils import simpleSplit
        from datetime import date as _dt

        batch_name  = row['batch_name']
        batch_size  = float(row['batch_size'])
        no_of_batch = int(row['no_of_batch'])
        batch_type  = row.get('batch_type', 'Regular')
        sku_size    = row.get('sku_size') or ''
        fvq_name    = row['worksheet']

        ingr_rows = _fvq_get_ingredients(fvq_name)

        # Fetch product_code + manuf_process
        try:
            _conn = sampling_portal.get_db_connection()
            meta_row = _conn.execute(
                f"SELECT product_code, manuf_process FROM `{_FVQ_TABLE}` "
                f"WHERE batch_name = %s LIMIT 1", (fvq_name,)
            ).fetchone()
            _conn.close()
            product_code = (meta_row['product_code'] or '') if meta_row else ''
            manuf        = (meta_row['manuf_process'] or '') if meta_row else ''
        except Exception:
            product_code = ''
            manuf        = ''

        # True landscape A4 (width=842pt, height=595pt)
        A4_W, A4_H = A4
        PW, PH     = A4_H, A4_W

        buf = _io.BytesIO()
        c   = _rl_canvas.Canvas(buf, pagesize=(PW, PH))

        _TEAL  = _rl_colors.HexColor('#0d9488')
        _GREY  = _rl_colors.HexColor('#475569')
        _LGREY = _rl_colors.HexColor('#f1f5f9')
        _BLACK = _rl_colors.black
        _WHITE = _rl_colors.white
        _BRDR  = _rl_colors.HexColor('#e2e8f0')

        margin = 12 * mm
        cw     = PW - 2 * margin

        HDR_H     = 10 * mm
        INFO1_H   =  9 * mm
        INFO2_H   =  8 * mm
        TBL_HDR_H =  7 * mm
        FOOTER_H  =  7 * mm

        fixed_h    = HDR_H + INFO1_H + INFO2_H + TBL_HDR_H + FOOTER_H + 2*margin + 8*mm
        available  = PH - fixed_h
        n_rows     = max(len(ingr_rows), 1)
        ROW_H      = max(5 * mm, min(8 * mm, available / n_rows))
        font_size  = max(6.5, min(8.5, (ROW_H / mm) * 0.85))

        y = PH - margin

        # Header bar
        c.setFillColor(_TEAL)
        c.rect(margin, y - HDR_H, cw, HDR_H, fill=1, stroke=0)
        c.setFont('Helvetica-Bold', 13)
        c.setFillColor(_WHITE)
        c.drawString(margin + 4*mm, y - 7*mm, 'PRODUCTION SHEET')
        c.setFont('Helvetica', 9)
        c.drawRightString(margin + cw - 3*mm, y - 7*mm,
                          f'Date: {_dt.today().strftime("%d-%m-%Y")}')
        y -= HDR_H

        # Info row 1: Batch Name | Product Code
        c.setFillColor(_LGREY)
        c.rect(margin, y - INFO1_H, cw, INFO1_H, fill=1, stroke=0)
        c.setFillColor(_BLACK)
        name_w = cw * 0.60
        code_w = cw * 0.40

        c.setFont('Helvetica-Bold', 8.5)
        c.drawString(margin + 2*mm, y - 6*mm, 'Batch:')
        lbl_px = c.stringWidth('Batch:', 'Helvetica-Bold', 8.5) + 3*mm
        c.setFont('Helvetica', 8.5)
        bname = batch_name
        while bname and c.stringWidth(bname, 'Helvetica', 8.5) > name_w - lbl_px - 4*mm:
            bname = bname[:-1]
        if bname != batch_name:
            bname = bname[:-1] + '…'
        c.drawString(margin + lbl_px + 2*mm, y - 6*mm, bname)

        xpc = margin + name_w
        c.setFont('Helvetica-Bold', 8.5)
        c.drawString(xpc + 2*mm, y - 6*mm, 'Product Code:')
        pc_lbl_px = c.stringWidth('Product Code:', 'Helvetica-Bold', 8.5) + 3*mm
        c.setFont('Helvetica', 8.5)
        pc = product_code or '—'
        while pc and c.stringWidth(pc, 'Helvetica', 8.5) > code_w - pc_lbl_px - 4*mm:
            pc = pc[:-1]
        if pc != (product_code or '—'):
            pc = pc[:-1] + '…'
        c.drawString(xpc + pc_lbl_px + 2*mm, y - 6*mm, pc)
        y -= INFO1_H

        # Info row 2: Batch Size | No. of Batches | SKU
        c.setFillColor(_rl_colors.HexColor('#e8f5f3'))
        c.rect(margin, y - INFO2_H, cw, INFO2_H, fill=1, stroke=0)
        c.setFillColor(_BLACK)
        seg3 = cw / 3
        nb_display = '1 (Trial)' if batch_type == 'Trial' else str(no_of_batch)
        row2_items = [
            ('Batch Size:',      f'{batch_size:g} KG'),
            ('No. of Batches:',  nb_display),
            ('SKU Size:',        sku_size or '—'),
        ]
        for ci, (lbl, val) in enumerate(row2_items):
            x0    = margin + ci * seg3
            lbl_w = c.stringWidth(lbl, 'Helvetica-Bold', 8) + 2*mm
            c.setFont('Helvetica-Bold', 8)
            c.drawString(x0 + 2*mm, y - 5.5*mm, lbl)
            c.setFont('Helvetica', 8)
            c.drawString(x0 + lbl_w + 2*mm, y - 5.5*mm, val)
        y -= INFO2_H

        y -= 2*mm

        # Ingredient table
        COL_NO   =  8 * mm
        COL_QTY  = 26 * mm
        COL_DISP = 26 * mm
        COL_SUP  = 65 * mm
        COL_NAME = cw - COL_NO - COL_QTY - COL_DISP - COL_SUP

        cx = [
            margin,
            margin + COL_NO,
            margin + COL_NO + COL_NAME,
            margin + COL_NO + COL_NAME + COL_SUP,
            margin + COL_NO + COL_NAME + COL_SUP + COL_QTY,
        ]

        c.setFillColor(_TEAL)
        c.rect(margin, y - TBL_HDR_H, cw, TBL_HDR_H, fill=1, stroke=0)
        c.setFillColor(_WHITE)
        c.setFont('Helvetica-Bold', 8)
        for i, hdr in enumerate(['#', 'Ingredient / Material', 'Supplier', 'Qty (KG)', 'Dispensed']):
            c.drawString(cx[i] + 1.5*mm, y - TBL_HDR_H + 2.2*mm, hdr)
        y -= TBL_HDR_H

        alt = False
        for idx, ingr in enumerate(ingr_rows, 1):
            mat  = str(ingr.get('material_name') or '').strip()
            sup  = str(ingr.get('supplier_name') or '').strip()
            stored_qty  = float(ingr.get('qty_kg') or 0)
            stored_size = float(ingr.get('batch_size') or 1) or 1
            qty  = round(stored_qty * (batch_size / stored_size), 3)

            if alt:
                c.setFillColor(_LGREY)
                c.rect(margin, y - ROW_H, cw, ROW_H, fill=1, stroke=0)
            alt = not alt

            c.setFillColor(_BLACK)
            c.setFont('Helvetica', font_size)

            def _fit(text, col_w, pad=3*mm):
                s = text
                while s and c.stringWidth(s, 'Helvetica', font_size) > col_w - pad:
                    s = s[:-1]
                return s + ('…' if s != text else '')

            txt_y = y - ROW_H + (ROW_H - font_size * 0.352778) / 2
            for rx, rv in [
                (cx[0], str(idx)),
                (cx[1], _fit(mat, COL_NAME)),
                (cx[2], _fit(sup, COL_SUP)),
                (cx[3], f'{qty:.3f}'),
                (cx[4], ''),
            ]:
                c.drawString(rx + 1.5*mm, txt_y, rv)

            c.setStrokeColor(_BRDR)
            c.line(margin, y - ROW_H, margin + cw, y - ROW_H)
            y -= ROW_H

        total_qty = sum(
            round(float(r.get('qty_kg') or 0) * (batch_size / (float(r.get('batch_size') or 1) or 1)), 3)
            for r in ingr_rows)

        y -= 2*mm
        c.setFont('Helvetica-Bold', 8.5)
        c.setFillColor(_TEAL)
        c.drawRightString(cx[3] - 2*mm, y, 'Total:')
        c.setFillColor(_BLACK)
        c.drawString(cx[3] + 1.5*mm, y, f'{total_qty:.3f} KG')
        y -= 7*mm

        if manuf and y > margin + 10*mm:
            c.setFont('Helvetica-Bold', 8)
            c.setFillColor(_TEAL)
            c.drawString(margin, y, 'Manufacturing Process')
            c.setFillColor(_BLACK)
            y -= 5*mm
            c.setFont('Helvetica', 7.5)
            for line in simpleSplit(manuf, 'Helvetica', 7.5, cw)[:5]:
                if y < margin + 5*mm:
                    break
                c.drawString(margin + 2*mm, y, line)
                y -= 4.5*mm

        # Footer
        c.setFont('Helvetica', 7)
        c.setFillColor(_GREY)
        c.drawString(margin, margin - 2*mm,
                     f'HCP Factory  ·  {batch_name}  ·  {_dt.today().strftime("%d-%m-%Y")}')
        c.drawRightString(margin + cw, margin - 2*mm, 'Production Sheet')
        c.save()
        return buf.getvalue()

    def _print_fvq_pdf(pdf_bytes, copies):
        """Send PDF bytes to the chosen printer (or system default if none).
        Prefers SumatraPDF CLI (reliable landscape); falls back to win32api ShellExecute."""
        import tempfile, subprocess, os as _os
        try:
            import win32api, win32print
        except ImportError:
            win32api = None
            win32print = None

        with tempfile.NamedTemporaryFile(
            suffix='.pdf', prefix='hcp_prod_', delete=False
        ) as tf:
            tf.write(pdf_bytes)
            tmp_path = tf.name

        # Resolve target printer: user selection wins; else Windows default
        target_printer = printer_name
        if not target_printer and win32print:
            try:
                target_printer = win32print.GetDefaultPrinter()
            except Exception:
                target_printer = None

        # Detect virtual/PDF printers that need 'printto' (Save-As dialog)
        _is_pdf_printer = False
        if target_printer:
            _tp_low = target_printer.lower()
            _is_pdf_printer = any(
                kw in _tp_low for kw in
                ('pdf', 'xps', 'onenote', 'fax', 'image writer', 'docucom')
            )

        # Prefer SumatraPDF for reliable landscape on physical printers
        _sumatra = None
        for _sp in [
            r'C:\Program Files\SumatraPDF\SumatraPDF.exe',
            r'C:\Program Files (x86)\SumatraPDF\SumatraPDF.exe',
        ]:
            if _os.path.exists(_sp):
                _sumatra = _sp
                break

        if _is_pdf_printer and win32api and target_printer:
            # Virtual printer — ShellExecute 'printto' triggers Save-As dialog
            for _ in range(copies):
                win32api.ShellExecute(
                    0, 'printto', tmp_path,
                    f'"{target_printer}"',
                    '.', 0
                )
        elif _sumatra and target_printer:
            # Physical printer — SumatraPDF with landscape forced
            for _ in range(copies):
                subprocess.run([
                    _sumatra,
                    '-print-to', target_printer,
                    '-print-settings', 'landscape',
                    '-silent',
                    tmp_path,
                ], timeout=60)
        elif win32api:
            # Fallback: temporarily set the chosen printer as default, then print
            _prev = None
            try:
                if target_printer and win32print:
                    try:
                        _prev = win32print.GetDefaultPrinter()
                        if target_printer != _prev:
                            win32print.SetDefaultPrinter(target_printer)
                    except Exception:
                        _prev = None
                for _ in range(copies):
                    win32api.ShellExecute(0, 'print', tmp_path, None, '.', 0)
            finally:
                if _prev and win32print:
                    try:
                        win32print.SetDefaultPrinter(_prev)
                    except Exception:
                        pass
        else:
            raise RuntimeError('No printing backend available (win32api not installed)')

    def _print_one_attempt(row):
        """
        One attempt: open one Excel COM instance, fill cells, print, save, close.
        For FVQ batches (file_name == '__FVQ__'), generates a PDF via ReportLab
        and sends it to the default printer.
        Raises on any failure so the caller can retry.
        Returns (result_dict, copies).
        """
        bid        = row['id']
        file_name  = row['file_name']
        file_path  = row['file_path']
        worksheet  = row['worksheet']
        batch_size = float(row['batch_size'])
        no_of_batch= int(row['no_of_batch'])
        batch_type = row['batch_type']

        # ── FVQ batch: generate PDF → send to printer ────────────────────────
        if file_name == '__FVQ__':
            copies    = no_of_batch if batch_type != 'Trial' else 1
            # Process sheet for FVQ = same PDF (FVQ has no separate process page)
            # If caller specifically asked for process and you add a second page
            # layout later, branch on print_type here.
            pdf_bytes = _build_fvq_production_pdf(row)
            _print_fvq_pdf(pdf_bytes, copies)
            return ({'id': bid, 'name': row['batch_name'],
                     'status': 'ok', 'copies': copies, 'source': 'db_pdf'}, copies)

        # ── Legacy Excel batch ────────────────────────────────────────────────
        fp         = os.path.join(file_path, file_name)
        app_xl     = None
        wb         = None
        try:
            app_xl = xw.App(visible=False)
            app_xl.display_alerts  = False
            app_xl.screen_updating = False

            # update_links=False avoids link-update dialogs that can hang COM
            wb = app_xl.books.open(fp, update_links=False, read_only=False)

            try:
                ws = wb.sheets[worksheet]
            except Exception:
                raise ValueError(f"Worksheet '{worksheet}' not found in {file_name}")

            # Set target printer if specified (else Excel uses Windows default)
            if printer_name:
                try:
                    app_xl.api.ActivePrinter = printer_name
                except Exception:
                    # If the printer name isn't resolvable by Excel, silently fall
                    # back to system default so the job still prints.
                    pass

            # Write batch info into cells
            i5_val = '1 Batches - Trial' if batch_type == 'Trial' else f'{no_of_batch} Batches'
            ws.range('I5').value = i5_val
            ws.range('I6').value = batch_size
            ws.range('I7').value = _date.today()

            copies = no_of_batch if batch_type != 'Trial' else 1

            if print_type == 'process':
                ws.api.PrintOut(
                    From=2, To=2, Copies=copies,
                    Collate=False, PrintToFile=False, IgnorePrintAreas=False,
                )
            else:
                ws.api.PrintOut(
                    From=1, To=1, Copies=copies,
                    Collate=False, PrintToFile=False, IgnorePrintAreas=False,
                )

            wb.save()
            wb.close()
            wb = None
            return ({'id': bid, 'name': row['batch_name'], 'status': 'ok', 'copies': copies}, copies)

        except Exception:
            # Close workbook before quitting app to avoid orphaned processes
            try:
                if wb:
                    wb.close()
            except Exception:
                pass
            raise
        finally:
            _kill_xl_app(app_xl)

    def _print_one(row):
        """
        Print one batch with retry on transient COM errors.
        Sequential — caller must NOT run multiple instances simultaneously.
        Returns (result_dict, error_dict).
        """
        bid  = row['id']
        name = row.get('batch_name', '?')
        last_err = None

        for attempt in range(_MAX_RETRIES):
            try:
                result, copies = _print_one_attempt(row)
                return (result, None)
            except Exception as e:
                last_err = e
                if _is_retryable(e) and attempt < _MAX_RETRIES - 1:
                    delay = _RETRY_DELAYS[attempt]
                    print(f'[print_sheets] Retry {attempt+1}/{_MAX_RETRIES-1} for ID {bid} '
                          f'after {delay}s — {e}')
                    _time.sleep(delay)
                else:
                    # Non-retryable or final attempt — give up
                    break

        return (None, {'id': bid, 'name': name, 'error': str(last_err)})

    # ── Sequential execution ──────────────────────────────────────────────────
    # Excel COM automation is NOT thread-safe. Running multiple simultaneous
    # xw.App() instances causes RPC failures ("remote procedure call failed",
    # "server execution failed"). Process one file at a time — each Excel
    # instance fully quits before the next opens. A short pause between files
    # lets the COM server settle and prevents spooler race conditions.
    results = []
    errors  = []

    for i, row in enumerate(rows_sorted):
        res, err = _print_one(row)
        if res:
            results.append(res)
        elif err:
            errors.append(err)
        # Brief settle time between files (skip after the last one)
        if i < len(rows_sorted) - 1:
            _time.sleep(1)

    return jsonify({
        'status':     'ok' if not errors else 'partial',
        'results':    results,
        'errors':     errors,
        'print_type': print_type,
        'order':      'high_to_low' if print_type == 'front' else 'low_to_high',
    })


# ── API: get all batches ───────────────────────────────────────────────────────
@production_initiater_bp.route('/api/production/batches', methods=['GET'])
@login_required
def api_prod_batches():
    if not _prod_role():
        return jsonify({'error': 'Access denied'}), 403
    conn = sampling_portal.get_db_connection()
    rows = conn.execute(
        "SELECT * FROM Processing_batches ORDER BY added_on DESC"
    ).fetchall()
    conn.close()
    return jsonify({'batches': [dict(r) for r in rows]})




# ── API: read ingredient data for batch labels ─────────────────────────────────
@production_initiater_bp.route('/api/production/label_data', methods=['POST'])
@login_required
def api_prod_label_data():
    """
    For each selected batch:
      - Opens the formulation Excel with openpyxl (read-only, no xlwings needed)
      - Reads from row 13 downward in the worksheet
      - Col A: ingredient present if not empty
      - Col B: material name
      - Col D: concentration % → multiplied by batch_size to get qty_kg
      - Returns JSON list of label objects
    """
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'error': 'Access denied'}), 403

    d = request.json or {}
    batch_ids = [int(x) for x in (d.get('batch_ids') or []) if str(x).isdigit()]
    if not batch_ids:
        return jsonify({'error': 'No batch IDs provided'}), 400

    conn = sampling_portal.get_db_connection()
    placeholders = ','.join(['%s'] * len(batch_ids))
    rows = conn.execute(
        f"SELECT * FROM Processing_batches WHERE id IN ({placeholders})",
        tuple(batch_ids)
    ).fetchall()
    conn.close()

    if not rows:
        return jsonify({'error': 'Batches not found'}), 404

    results = []

    for row in rows:
        row = dict(row)
        bid        = row['id']
        batch_name = row['batch_name']
        file_name  = row['file_name']
        file_path  = row['file_path']
        worksheet  = row['worksheet']
        batch_size = float(row['batch_size'])
        no_of_batch= int(row['no_of_batch'])

        # Product name = part before " - " in batch_name, or full name
        product_name = batch_name.split(' - ')[0].strip() if ' - ' in batch_name else batch_name

        try:
            labels = []

            # ── SOURCE: Procurement DB (FVQ) ─────────────────────────────────
            if file_name == '__FVQ__':
                fvq_name  = worksheet   # batch lookup key stored in worksheet column
                ingr_rows = _fvq_get_ingredients(fvq_name)
                if not ingr_rows:
                    raise ValueError(f"No ingredients found in DB for '{fvq_name}'")
                for ingr in ingr_rows:
                    material_name = str(ingr.get('material_name') or '').strip()
                    if not material_name:
                        continue
                    # qty_kg and batch_size stored — scale proportionally
                    stored_qty  = float(ingr.get('qty_kg') or 0)
                    stored_size = float(ingr.get('batch_size') or 1) or 1
                    qty_kg = round(stored_qty * (batch_size / stored_size), 6)
                    labels.append({
                        'material_name': material_name,
                        'qty_kg':        qty_kg,
                        'product_name':  product_name,
                        'batch_name':    batch_name,
                        'batch_no':      no_of_batch,
                        'batch_size':    batch_size,
                        'source':        'db',
                    })

            # ── SOURCE: Legacy Excel file ─────────────────────────────────────
            else:
                fp = os.path.join(file_path, file_name)
                wb_tmp = load_workbook(fp, read_only=True, data_only=True)
                try:
                    ws_obj = wb_tmp.worksheets[0]
                    # Find the correct worksheet by name
                    for s in wb_tmp.worksheets:
                        if s.title == worksheet:
                            ws_obj = s
                            break
                except Exception:
                    wb_tmp.close()
                    raise ValueError(f"Worksheet '{worksheet}' not found")

                # Read from row 13 downward
                # Col A=1, B=2, C=3, D=4
                for r in range(13, 500):
                    cell_a = ws_obj.cell(row=r, column=1).value
                    if cell_a is None or str(cell_a).strip() == '':
                        break  # stop at first empty row in col A
                    material_name = ws_obj.cell(row=r, column=2).value or ''
                    conc_raw = ws_obj.cell(row=r, column=4).value

                    try:
                        conc_percent = float(conc_raw) if conc_raw is not None else 0.0
                    except (TypeError, ValueError):
                        conc_percent = 0.0

                    # REQUIRED FORMULA
                    # Quantity = Batch Size × Concentration %
                    qty_kg = round((batch_size * conc_percent) / 100.0, 6)

                    labels.append({
                        'material_name': str(material_name).strip(),
                        'qty_kg': qty_kg,
                        'product_name': product_name,
                        'batch_name': batch_name,
                        'batch_no': no_of_batch,
                        'batch_size': batch_size,
                        'source':     'excel',
                    })

                wb_tmp.close()

            results.append({
                'id':               bid,
                'name':             batch_name,
                'status':           'ok',
                'ingredient_count': len(labels),
                'labels':           labels,
            })

        except FileNotFoundError:
            results.append({'id': bid, 'name': batch_name, 'status': 'error',
                            'error': f'Excel file not found: {os.path.join(file_path, file_name)}'})
        except Exception as e:
            import traceback
            print(f'[label_data] ERROR for batch {bid}: {e}\n{traceback.format_exc()}')
            results.append({'id': bid, 'name': batch_name, 'status': 'error', 'error': str(e)})

    return jsonify({'results': results})


# ── API: RM Requirement Calculator ───────────────────────────────────────────
@production_initiater_bp.route('/api/production/rm_requirement', methods=['POST'])
@login_required
def api_rm_requirement():
    """
    For each selected batch, opens its formulation Excel (read-only, openpyxl),
    reads ingredients from row 13 downward (Col A = present, Col B = material name,
    Col D = concentration %), aggregates required quantities across all batches,
    applies same exclusions as VBA macro, and returns a structured report.

    Response:
    {
      "status": "ok",
      "batches": [ {id, batch_name, batch_size, no_of_batch, total_size, status} ],
      "materials": [
        { "name": "...", "total_qty": 12.345,
          "batches": [ {"batch_name": "...", "qty": 3.45} ] }
      ],
      "errors": [ {"batch_name": "...", "error": "..."} ]
    }
    """
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403

    d = request.json or {}
    batch_ids = [int(x) for x in (d.get('batch_ids') or []) if str(x).isdigit()]
    if not batch_ids:
        return jsonify({'status': 'error', 'message': 'No batch IDs provided'}), 400

    conn = sampling_portal.get_db_connection()
    placeholders = ','.join(['%s'] * len(batch_ids))
    rows = conn.execute(
        f"SELECT * FROM Processing_batches WHERE id IN ({placeholders})",
        tuple(batch_ids)
    ).fetchall()
    conn.close()

    if not rows:
        return jsonify({'status': 'error', 'message': 'Batches not found'}), 404

    # Exclusion list — mirrors VBA macro
    EXCLUDED_SUBSTRINGS = ['water', 'aqua']
    EXCLUDED_EXACT      = {'PURE CASTOR OIL', 'SORBITOL', 'MFIL SILICA (PPT SILICA)'}

    def is_excluded(name):
        nu = name.upper().strip()
        if nu in EXCLUDED_EXACT:
            return True
        nl = name.lower().strip()
        return any(ex in nl for ex in EXCLUDED_SUBSTRINGS)

    # material_name → { total_qty, batches: [{batch_name, qty, no_of_batch}] }
    agg = {}
    batch_summaries = []
    errors = []

    for row in rows:
        row = dict(row)
        bid         = row['id']
        batch_name  = row['batch_name']
        file_name   = row['file_name']
        file_path   = row['file_path']
        worksheet   = row['worksheet']
        batch_size  = float(row['batch_size'])
        no_of_batch = int(row['no_of_batch'])
        # total material needed = batch_size × no_of_batch
        total_size  = round(batch_size * no_of_batch, 3)

        batch_summaries.append({
            'id':         bid,
            'batch_name': batch_name,
            'batch_size': batch_size,
            'no_of_batch': no_of_batch,
            'total_size': total_size,
        })

        try:
            # ── SOURCE: Procurement DB (FVQ) ─────────────────────────────────
            if file_name == '__FVQ__':
                fvq_name  = worksheet   # batch lookup key stored in worksheet column
                ingr_rows = _fvq_get_ingredients(fvq_name)
                if not ingr_rows:
                    raise ValueError(f"No ingredients found for '{fvq_name}' in procurement DB")

                for ingr in ingr_rows:
                    material_name = str(ingr.get('material_name') or '').strip()
                    if not material_name or is_excluded(material_name):
                        continue
                    # Scale stored qty_kg from reference batch_size to production batch_size
                    stored_qty  = float(ingr.get('qty_kg') or 0)
                    stored_size = float(ingr.get('batch_size') or 1) or 1
                    if stored_qty <= 0:
                        continue
                    qty_per_batch = round(stored_qty * (batch_size / stored_size), 6)
                    total_qty     = round(qty_per_batch * no_of_batch, 6)
                    if material_name not in agg:
                        agg[material_name] = {'total_qty': 0.0, 'batches': []}
                    agg[material_name]['total_qty'] = round(
                        agg[material_name]['total_qty'] + total_qty, 6)
                    agg[material_name]['batches'].append({
                        'batch_name':    batch_name,
                        'batch_size':    batch_size,
                        'no_of_batch':   no_of_batch,
                        'qty':           total_qty,
                        'qty_per_batch': qty_per_batch,
                    })

            # ── SOURCE: Legacy Excel file ─────────────────────────────────────
            else:
                fp = os.path.join(file_path, file_name)
                wb = load_workbook(fp, read_only=True, data_only=True)
                ws_obj = None
                for s in wb.worksheets:
                    if s.title == worksheet:
                        ws_obj = s
                        break
                if ws_obj is None:
                    wb.close()
                    raise ValueError(f"Worksheet '{worksheet}' not found in {file_name}")

                for r in range(13, 1000):
                    cell_a = ws_obj.cell(row=r, column=1).value
                    if cell_a is None or str(cell_a).strip() == '':
                        break

                    material_name = str(ws_obj.cell(row=r, column=2).value or '').strip()
                    if not material_name:
                        continue

                    # Skip excluded materials
                    if is_excluded(material_name):
                        continue

                    conc_raw = ws_obj.cell(row=r, column=4).value

                    # Mirror VBA exactly:
                    #   "%" in cell string  → strip % and divide by 100
                    #   plain numeric       → use as-is (already a fraction e.g. 0.125 = 12.5%)
                    conc = 0.0
                    if conc_raw is not None:
                        raw_str = str(conc_raw).strip()
                        if '%' in raw_str:
                            try:
                                conc = float(raw_str.replace('%', '').strip()) / 100.0
                            except (ValueError, TypeError):
                                conc = 0.0
                        else:
                            try:
                                conc = float(conc_raw)
                            except (ValueError, TypeError):
                                conc = 0.0

                    if conc <= 0:
                        continue

                    # VBA: requiredQty = batchSize * concentration
                    # where VBA batchSize = batch_size_col_E * no_of_batches_col_F
                    effective_size = batch_size * no_of_batch          # = VBA batchSize
                    total_qty      = round(effective_size * conc, 6)   # = VBA requiredQty
                    qty_per_batch  = round(batch_size * conc, 6)       # per single batch

                    if material_name not in agg:
                        agg[material_name] = {'total_qty': 0.0, 'batches': []}

                    agg[material_name]['total_qty'] = round(
                        agg[material_name]['total_qty'] + total_qty, 6)
                    agg[material_name]['batches'].append({
                        'batch_name':  batch_name,
                        'batch_size':  batch_size,
                        'no_of_batch': no_of_batch,
                        'qty':         total_qty,
                        'qty_per_batch': qty_per_batch,
                    })

                wb.close()

        except FileNotFoundError:
            errors.append({'batch_name': batch_name,
                           'error': f'Excel file not found: {os.path.join(file_path, file_name)}'})
        except Exception as e:
            import traceback
            print(f'[rm_requirement] ERROR {bid}: {e}\n{traceback.format_exc()}')
            errors.append({'batch_name': batch_name, 'error': str(e)})

    # ── Read StkSum.xlsx for current stock ──────────────────────────────────────
    # Layout: data starts row 12, Col A = ingredient name, Col C = current stock
    STK_SUM_PATH = r'\\Tarakbhavsar\procurement new\CURRENT RM\StkSum.xlsx'
    stock_map = {}   # ingredient_name_lower → {'name': original, 'qty': float}
    stk_error = None
    try:
        wb_stk = load_workbook(STK_SUM_PATH, read_only=True, data_only=True)
        ws_stk = wb_stk.active
        for stk_row in ws_stk.iter_rows(min_row=12, values_only=True):
            # Col A (index 0) = ingredient name, Col C (index 2) = current stock
            if not stk_row or stk_row[0] is None:
                continue
            stk_name = str(stk_row[0]).strip()
            stk_qty  = 0.0
            if len(stk_row) > 2 and stk_row[2] is not None:
                try:
                    stk_qty = float(stk_row[2])
                except (TypeError, ValueError):
                    stk_qty = 0.0
            if stk_name:
                stock_map[stk_name.lower()] = {'name': stk_name, 'qty': stk_qty}
        wb_stk.close()
    except Exception as stk_ex:
        import traceback as _tb
        stk_error = str(stk_ex)
        print(f'[rm_requirement] StkSum read error: {stk_ex}\n{_tb.format_exc()}')

    # Sort materials alphabetically and enrich with stock data
    def _build_material(name, data):
        total_qty = round(data['total_qty'], 4)

        # Build product names list (unique, joined with " / ")
        seen_products = []
        for b in data['batches']:
            bn = b['batch_name']
            if bn not in seen_products:
                seen_products.append(bn)
        products = ' / '.join(seen_products)

        # Stock lookup — exact match first, then case-insensitive
        stk_entry = stock_map.get(name.lower())
        current_stock = round(stk_entry['qty'], 4) if stk_entry else None
        stock_diff    = round(current_stock - total_qty, 4) if current_stock is not None else None

        return {
            'name':          name,
            'total_qty':     total_qty,
            'current_stock': current_stock,
            'stock_diff':    stock_diff,
            'products':      products,
            'batches':       data['batches'],
        }

    materials = sorted([
        _build_material(name, data)
        for name, data in agg.items()
    ], key=lambda x: x['name'].lower())

    return jsonify({
        'status':          'ok',
        'batches':         batch_summaries,
        'materials':       materials,
        'errors':          errors,
        'total_materials': len(materials),
        'stk_error':       stk_error,
    })


# ── Tally Stock Integration ───────────────────────────────────────────────────
# Fetches closing stock for all items from TallyPrime via XML port
# and fuzzy-matches against the RM names from the formulation files.

def _tally_fetch_all_stock(creds: dict = None):
    """
    Posts a TDL XML request to TallyPrime 5.0 asking for all stock items
    with closing balance (qty + rate + value).
    Returns a dict: { item_name_lower: { 'name': original_name, 'qty': float, 'unit': str } }
    or raises an exception on failure.

    creds: dict with keys tally_url, company_name, tally_user, tally_pass
           Falls back to global TALLY_URL / COMPANY_NAME if not provided.

    Uses the StockSummary report — works with TallyPrime 5.0 EditLog (port 9000).
    """
    url      = (creds or {}).get('tally_url', TALLY_URL).rstrip('/')
    company  = (creds or {}).get('company_name', COMPANY_NAME)
    t_user   = (creds or {}).get('tally_user', '')
    t_pass   = (creds or {}).get('tally_pass', '')

    xml_req = f"""<ENVELOPE>
  <HEADER>
    <TALLYREQUEST>Export Data</TALLYREQUEST>
  </HEADER>
  <BODY>
    <EXPORTDATA>
      <REQUESTDESC>
        <REPORTNAME>Stock Summary</REPORTNAME>
        <STATICVARIABLES>
          <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
          <SVCURRENTCOMPANY>{company}</SVCURRENTCOMPANY>
          <SVFROMDATE>$$MonthStart:$$SystemDate</SVFROMDATE>
          <SVTODATE>$$SystemDate</SVTODATE>
        </STATICVARIABLES>
      </REQUESTDESC>
    </EXPORTDATA>
  </BODY>
</ENVELOPE>"""

    req_kwargs = dict(
        data=xml_req.encode('utf-8'),
        headers={'Content-Type': 'text/xml; charset=utf-8'},
        timeout=10
    )
    if t_user and t_pass:
        req_kwargs['auth'] = (t_user, t_pass)

    resp = requests.post(url, **req_kwargs)
    resp.raise_for_status()

    parsed = xmltodict.parse(resp.text, force_list=('STOCKITEM',))
    stock_map = {}

    # Navigate: ENVELOPE → BODY → DATA → TALLYMESSAGE → STOCKITEM  (varies by Tally version)
    def _walk(node, target):
        if isinstance(node, dict):
            if target in node:
                return node[target]
            for v in node.values():
                result = _walk(v, target)
                if result is not None:
                    return result
        return None

    items = _walk(parsed, 'STOCKITEM') or []
    if isinstance(items, dict):
        items = [items]

    for item in items:
        if not isinstance(item, dict):
            continue

        # Item name — Tally uses @NAME or NAME attribute
        raw_name = (item.get('@NAME') or item.get('NAME') or
                    item.get('STOCKITEMNAME') or '').strip()
        if not raw_name:
            continue

        # Closing balance quantity — Tally wraps it in CLOSINGBALANCE or CLOSINGSTOCK
        closing = (item.get('CLOSINGBALANCE') or
                   item.get('CLOSINGSTOCK') or
                   item.get('CLOSINGQTY') or '0')

        if isinstance(closing, dict):
            closing = closing.get('#text', '0')

        closing_str = str(closing).strip()

        # Parse qty and unit — format is often "12.500 Kg" or "1500.000 Nos"
        parts = closing_str.split()
        try:
            qty = float(parts[0].replace(',', '')) if parts else 0.0
        except (ValueError, IndexError):
            qty = 0.0
        unit = parts[1].strip() if len(parts) > 1 else ''

        stock_map[raw_name.lower()] = {
            'name': raw_name,
            'qty':  round(qty, 4),
            'unit': unit,
        }


    return stock_map


def _exact_match(query, stock_map):
    """
    Case-insensitive exact match against Tally stock map.
    Names in Excel and Tally are kept identical, so no fuzzy logic needed.
    stock_map keys are already lowercase.
    Returns the stock_map entry dict or None.
    """
    return stock_map.get(query.lower().strip())


@production_initiater_bp.route('/api/production/tally_stock_check', methods=['POST'])
@login_required
def api_tally_stock_check():
    """
    Receives RM names + required quantities from the RM Requirement report.
    Fetches current closing stock from TallyPrime and does a case-insensitive
    exact match (names are identical between Excel formulations and Tally master).

    Request:  { "materials": [ {"name": "...", "total_qty": 24.0}, ... ] }
    Response: {
      "status": "ok",
      "items": [
        {
          "rm_name":      "Sodium Lauryl Sulfate",
          "tally_name":   "Sodium Lauryl Sulfate",   ← exact match
          "required_qty": 24.0,
          "tally_qty":    18.5,
          "unit":         "Kg",
          "shortage":     5.5,      ← positive = short, negative = surplus
          "status":       "SHORT"   ← "OK" / "SHORT" / "NOT IN TALLY"
        }
      ],
      "tally_error": null
    }
    """
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403

    d         = request.get_json() or {}
    materials = d.get('materials') or []
    if not materials:
        return jsonify({'status': 'error', 'message': 'No materials provided'}), 400

    # ── Load this user's saved Tally credentials ─────────────────────────────
    user_id = session.get('user_id')
    creds   = _get_tally_creds(user_id) if user_id else None
    if not creds:
        return jsonify({
            'status': 'error',
            'message': 'tally_not_configured',
            'detail': 'Please connect your Tally account first.'
        }), 400

    # ── Step 1: fetch all stock from Tally ──────────────────────────────────
    try:
        stock_map   = _tally_fetch_all_stock(creds)
        tally_error = None
    except requests.exceptions.ConnectionError:
        stock_map   = {}
        tally_error = f'Cannot connect to TallyPrime at {creds["tally_url"]} — make sure Tally is open and port 9000 is enabled.'
    except requests.exceptions.Timeout:
        stock_map   = {}
        tally_error = 'TallyPrime did not respond in time (timeout 10s).'
    except Exception as e:
        stock_map   = {}
        tally_error = f'Tally error: {str(e)}'

    # ── Step 2: exact match each RM name ────────────────────────────────────
    items = []
    for m in materials:
        rm_name      = str(m.get('name', '')).strip()
        required_qty = float(m.get('total_qty', 0))

        if not stock_map:
            items.append({
                'rm_name':      rm_name,
                'tally_name':   None,
                'required_qty': required_qty,
                'tally_qty':    None,
                'unit':         '',
                'shortage':     None,
                'status':       'NOT IN TALLY',
            })
            continue

        hit = _exact_match(rm_name, stock_map)

        if hit:
            tally_qty = hit['qty']
            shortage  = round(required_qty - tally_qty, 4)
            items.append({
                'rm_name':      rm_name,
                'tally_name':   hit['name'],
                'required_qty': required_qty,
                'tally_qty':    tally_qty,
                'unit':         hit['unit'],
                'shortage':     shortage,
                'status':       'OK' if shortage <= 0 else 'SHORT',
            })
        else:
            items.append({
                'rm_name':      rm_name,
                'tally_name':   None,
                'required_qty': required_qty,
                'tally_qty':    None,
                'unit':         '',
                'shortage':     None,
                'status':       'NOT IN TALLY',
            })

    ok_count      = sum(1 for i in items if i['status'] == 'OK')
    short_count   = sum(1 for i in items if i['status'] == 'SHORT')
    nomatch_count = sum(1 for i in items if i['status'] == 'NOT IN TALLY')

    return jsonify({
        'status':      'ok',
        'items':       items,
        'tally_error': tally_error,
        'summary': {
            'ok':       ok_count,
            'short':    short_count,
            'no_match': nomatch_count,
            'total':    len(items),
        }
    })


@production_initiater_bp.route('/api/production/tally_creds', methods=['GET'])
@login_required
def api_tally_creds_get():
    """Return saved Tally credentials for the current user (password masked)."""
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    user_id = session.get('user_id')
    creds   = _get_tally_creds(user_id) if user_id else None
    if not creds:
        return jsonify({'status': 'ok', 'configured': False})
    return jsonify({
        'status':       'ok',
        'configured':   True,
        'tally_url':    creds['tally_url'],
        'company_name': creds['company_name'],
        'tally_user':   creds['tally_user'],
        'has_password': bool(creds['tally_pass']),
    })


@production_initiater_bp.route('/api/production/tally_creds_save', methods=['POST'])
@login_required
def api_tally_creds_save():
    """Save (or update) Tally credentials for the current user."""
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403
    d            = request.get_json() or {}
    tally_url    = (d.get('tally_url') or '').strip().rstrip('/')
    company_name = (d.get('company_name') or '').strip()
    tally_user   = (d.get('tally_user') or '').strip()
    tally_pass   = (d.get('tally_pass') or '').strip()

    if not tally_url or not company_name:
        return jsonify({'status': 'error', 'message': 'Tally URL and Company Name are required'}), 400

    # Ensure URL has scheme
    if not tally_url.startswith('http'):
        tally_url = 'http://' + tally_url

    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'status': 'error', 'message': 'Session expired'}), 401

    obf_pass = _obfuscate(tally_pass) if tally_pass else ''

    conn = sampling_portal.get_db_connection()
    try:
        # If password field is blank on update, keep existing stored password
        existing = conn.execute(
            "SELECT tally_pass FROM tally_credentials WHERE user_id=%s", (user_id,)
        ).fetchone()

        if existing and not tally_pass:
            # Blank password = keep existing
            obf_pass = existing['tally_pass'] if existing else ''

        conn.execute("""
            INSERT INTO tally_credentials (user_id, tally_url, company_name, tally_user, tally_pass)
            VALUES (%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
                tally_url    = VALUES(tally_url),
                company_name = VALUES(company_name),
                tally_user   = VALUES(tally_user),
                tally_pass   = VALUES(tally_pass),
                saved_at     = CURRENT_TIMESTAMP
        """, (user_id, tally_url, company_name, tally_user, obf_pass))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok', 'message': 'Tally credentials saved.'})
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@production_initiater_bp.route('/api/production/tally_test', methods=['POST'])
@login_required
def api_tally_test():
    """
    Test a Tally connection with provided (or saved) credentials.
    Sends a lightweight 'List of Companies' request and verifies the target
    company exists in the response.
    """
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'status': 'error', 'message': 'Admin only'}), 403

    d = request.get_json() or {}
    # Allow testing with form values before saving
    tally_url    = (d.get('tally_url') or '').strip().rstrip('/')
    company_name = (d.get('company_name') or '').strip()
    tally_user   = (d.get('tally_user') or '').strip()
    tally_pass   = (d.get('tally_pass') or '').strip()

    # If any field missing, fall back to saved creds
    if not tally_url or not company_name:
        user_id = session.get('user_id')
        saved   = _get_tally_creds(user_id) if user_id else None
        if not saved:
            return jsonify({'status': 'error',
                            'message': 'Provide Tally URL and Company Name'}), 400
        tally_url    = tally_url    or saved['tally_url']
        company_name = company_name or saved['company_name']
        tally_user   = tally_user   or saved['tally_user']
        tally_pass   = tally_pass   or saved['tally_pass']

    if not tally_url.startswith('http'):
        tally_url = 'http://' + tally_url

    # Step 1 — basic connectivity
    xml_ping = """<ENVELOPE>
  <HEADER><TALLYREQUEST>Export Data</TALLYREQUEST></HEADER>
  <BODY><EXPORTDATA><REQUESTDESC>
    <REPORTNAME>List of Companies</REPORTNAME>
    <STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>
  </REQUESTDESC></EXPORTDATA></BODY>
</ENVELOPE>"""
    try:
        req_kwargs = dict(data=xml_ping.encode('utf-8'),
                          headers={'Content-Type': 'text/xml; charset=utf-8'},
                          timeout=8)
        if tally_user and tally_pass:
            req_kwargs['auth'] = (tally_user, tally_pass)

        resp = requests.post(tally_url, **req_kwargs)
        resp.raise_for_status()

        # Step 2 — check company exists in response
        text_lower    = resp.text.lower()
        company_found = company_name.lower() in text_lower

        # Step 3 — quick stock fetch to confirm company is open and accessible
        stock_ok = False
        stock_count = 0
        try:
            test_creds = {'tally_url': tally_url, 'company_name': company_name,
                          'tally_user': tally_user, 'tally_pass': tally_pass}
            smap = _tally_fetch_all_stock(test_creds)
            stock_ok    = True
            stock_count = len(smap)
        except Exception:
            pass

        return jsonify({
            'status':         'ok',
            'reachable':      True,
            'company_found':  company_found,
            'stock_ok':       stock_ok,
            'stock_count':    stock_count,
            'tally_url':      tally_url,
            'message':        (
                f'Connected ✓ · {stock_count} stock items found'
                if stock_ok else
                ('Company not found in Tally — check company name spelling'
                 if not company_found else
                 'Tally reachable but could not fetch stock — ensure company is open')
            )
        })

    except requests.exceptions.ConnectionError:
        return jsonify({'status': 'ok', 'reachable': False,
                        'message': f'Cannot connect to {tally_url} — check IP/port and that Tally is open'})
    except requests.exceptions.Timeout:
        return jsonify({'status': 'ok', 'reachable': False,
                        'message': 'Connection timed out (8s) — Tally may be busy or unreachable'})
    except Exception as e:
        return jsonify({'status': 'ok', 'reachable': False, 'message': str(e)})


# ── PDF helpers for Batch Confirmation + BST ─────────────────────────────────
import io
from datetime import datetime
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm

# Exact measurements from reference PDF (in points)
# Page: 595.3 x 841.9
# Content: left=27, right=572, width=545
# All y coords in reportlab (0=bottom)

TEAL   = colors.HexColor('#0d9488')
BLACK  = colors.black
LGREY  = colors.HexColor('#f5f5f5')   # very light — alternate rows only
MGREY  = colors.HexColor('#666666')

CHECK_ROWS = [
    ('RM Check Physical',      'Punam Singh'),
    ('RM Check Tally',         'Tarak Bhavsar'),
    ('PM Check Physical',      'Joitaji Thakor'),
    ('Sleeve /Corrugation',    'Joitaji Thakor'),
    ('Machines Check',         'Paresh Meraiya'),
    ('DM Water Check',         'Ashfaq Badhra'),
    ('Planning Head Approval', 'Arvind Purohit'),
    ('Production Dept',        'Chirag Rathod'),
]

def build_confirmation_pdf(batches):
    buf = io.BytesIO()
    PH = 841.9   # page height
    PW_PAGE = 595.3
    c = rl_canvas.Canvas(buf, pagesize=(PW_PAGE, PH))

    # Exact coords from reference
    L  = 27.0    # left border x
    R  = 572.4   # right border x
    W  = R - L   # 545.4

    # Vertical col dividers
    V1_TOP = 295.2   # top-section: Product|SKU divider
    V2_TOP = 384.6   # top-section: SKU|Date divider  (also qty + checklist col2)
    V1_QTY = 205.8   # qty section col divider 1
    V2_QTY = 384.6   # qty section col divider 2
    V1_CHK = 205.8   # checklist: Activity|CheckedBy
    V2_CHK = 384.6   # checklist: CheckedBy|Sign
    V3_CHK = 474.0   # checklist: Sign split (not needed for text)

    # Horizontal lines (reportlab y = 841.9 - pdfplumber_top)
    Y_SUBTITLE_UL  = 701.6   # underline below subtitle
    Y_INFO_TOP     = 682.3   # top of info section box
    Y_INFO_DIV1    = 661.9   # below label row (Product Name / SKU Size / Date labels)
    Y_INFO_DIV2    = 603.8   # below value row
    Y_QTY_LABEL_B  = 583.4   # bottom of QUANTITY label bar
    Y_QTY_HDR_TOP  = 542.6   # top of Batch Size / No of Batch / Total headers row
    Y_QTY_HDR_DIV  = 519.6   # between header and value
    Y_QTY_BOT      = 496.5   # bottom of qty section
    Y_CHK_TOP      = 473.5   # top of checklist (header row top)
    CHK_ROW_H      = 23.0    # each checklist row height
    Y_CHK_BOT      = 58.8    # bottom border

    def hl(y, x0=L, x1=R, lw=0.8):
        c.setStrokeColor(TEAL); c.setLineWidth(lw)
        c.line(x0, y, x1, y)

    def vl(x, y0, y1, lw=0.8):
        c.setStrokeColor(TEAL); c.setLineWidth(lw)
        c.line(x, y0, x, y1)

    for batch in batches:
        bname    = batch['batch_name']
        bsize    = float(batch['batch_size'])
        nb       = int(batch['no_of_batch'])
        sku      = batch['sku_size'] or ''
        total_bs = round(bsize * nb, 3)
        try:
            dt = datetime.strptime(str(batch['added_on'])[:10], '%Y-%m-%d')
            date_display = dt.strftime('%d-%m-%Y')
        except Exception:
            date_display = str(batch['added_on'])[:10]

        # ── Border lines (just lines, no filled rects) ──────────────────
        # Top horizontal border
        hl(Y_INFO_TOP, lw=0.8)
        # Bottom horizontal border  
        hl(Y_CHK_BOT, lw=0.8)
        # Left vertical border (full height of form)
        vl(L, Y_CHK_BOT, Y_INFO_TOP, lw=0.8)
        # Right vertical border
        vl(R, Y_CHK_BOT, Y_INFO_TOP, lw=0.8)

        # ── Company heading (text only, no background) ──────────────────
        c.setFont('Helvetica-Bold', 16)
        c.setFillColor(BLACK)
        c.drawCentredString(PW_PAGE/2, PH - 74, 'HCP WELLNESS PVT LTD')

        # Thin underline
        hl(Y_SUBTITLE_UL, x0=L+10, x1=R-10, lw=0.5)

        # Subtitle
        c.setFont('Helvetica-Bold', 9)
        c.setFillColor(MGREY)
        c.drawCentredString(PW_PAGE/2, PH - 120, 'PRODUCT MANUFACTURING PRE CHECK FORM')

        # ── Info section: horizontal lines ──────────────────────────────
        hl(Y_INFO_DIV1)    # below labels row
        hl(Y_INFO_DIV2)    # below values row

        # Info section: vertical dividers
        # Between Product Name and SKU: V1_TOP (but only for label row)
        vl(V1_TOP, Y_INFO_DIV2, Y_INFO_TOP)
        # Between SKU and Date: V2_TOP
        vl(V2_TOP, Y_INFO_DIV2, Y_INFO_TOP)

        # Label row text (top portion of info section)
        label_y = Y_INFO_DIV1 + 6   # text baseline in label row
        c.setFont('Helvetica-Bold', 8)
        c.setFillColor(BLACK)
        c.drawString(L + 86, label_y, 'Product Name')
        c.drawString(V1_TOP + 8, label_y, 'SKU Size')
        c.drawString(V2_TOP + 39, label_y, 'Date')

        # Value row text
        val_y = Y_INFO_DIV2 + 18   # text baseline in value row
        c.setFont('Helvetica', 9)

        # Batch name — truncate to fit in Product Name col (27 to 295)
        avail_w = V1_TOP - L - 10
        disp = bname
        while c.stringWidth(disp, 'Helvetica', 9) > avail_w and len(disp) > 4:
            disp = disp[:-2] + '\u2026'
        c.drawString(L + 5, val_y, disp)

        # SKU (295 to 384)
        c.drawCentredString((V1_TOP + V2_TOP)/2, val_y, str(sku))

        # Date (384 to 572)
        c.drawCentredString((V2_TOP + R)/2, val_y, date_display)

        # ── QUANTITY section ────────────────────────────────────────────
        # "QUANTITY" label bar  (Y_INFO_DIV2 down to Y_QTY_LABEL_B)
        hl(Y_QTY_LABEL_B)
        qty_label_mid = (Y_INFO_DIV2 + Y_QTY_LABEL_B) / 2
        c.setFont('Helvetica-Bold', 9)
        c.setFillColor(BLACK)
        c.drawCentredString(PW_PAGE/2, qty_label_mid - 4, 'QUANTITY')

        # Gap between QUANTITY bar and Batch Size section (Y_QTY_LABEL_B to Y_QTY_HDR_TOP)
        # just empty space — no lines needed

        # Batch size section top/bottom + inner dividers
        hl(Y_QTY_HDR_TOP)   # top of qty headers
        hl(Y_QTY_HDR_DIV)   # between header labels and values
        hl(Y_QTY_BOT)       # bottom of qty section

        # Vertical dividers for 3 qty columns
        vl(V1_QTY, Y_QTY_BOT, Y_QTY_HDR_TOP)
        vl(V2_QTY, Y_QTY_BOT, Y_QTY_HDR_TOP)

        # Column headers for qty
        col1_cx = (L + V1_QTY) / 2
        col2_cx = (V1_QTY + V2_QTY) / 2
        col3_cx = (V2_QTY + R) / 2

        c.setFont('Helvetica-Bold', 8)
        c.setFillColor(MGREY)
        hdr_label_y = Y_QTY_HDR_DIV + 6
        c.drawCentredString(col1_cx, hdr_label_y, 'Batch Size')
        c.drawCentredString(col2_cx, hdr_label_y, 'No of Batch')
        c.drawCentredString(col3_cx, hdr_label_y, 'Total Batch Size')

        # Qty values
        c.setFont('Helvetica-Bold', 11)
        c.setFillColor(BLACK)
        val_y2 = Y_QTY_BOT + 6
        c.drawCentredString(col1_cx, val_y2, f'{bsize:.3f} Kg')
        c.drawCentredString(col2_cx, val_y2, str(nb))
        c.drawCentredString(col3_cx, val_y2, f'{total_bs:.3f} Kg')

        # ── Checklist ───────────────────────────────────────────────────
        # Header row top line
        hl(Y_CHK_TOP)
        # Col dividers for header (full height)
        vl(V1_CHK, Y_CHK_BOT, Y_CHK_TOP)
        vl(V2_CHK, Y_CHK_BOT, Y_CHK_TOP)
        vl(V3_CHK, Y_CHK_BOT, Y_CHK_TOP)

        # Header text
        chk_hdr_y = Y_CHK_TOP - CHK_ROW_H + 7
        c.setFont('Helvetica-Bold', 8)
        c.setFillColor(BLACK)
        c.drawCentredString((L + V1_CHK)/2,        chk_hdr_y, 'Checking Activity')
        c.drawCentredString((V1_CHK + V2_CHK)/2,   chk_hdr_y, 'Checked By')
        c.drawCentredString((V2_CHK + R)/2,        chk_hdr_y, 'Checked Sign')

        # Header row bottom line
        y_after_hdr = Y_CHK_TOP - CHK_ROW_H
        hl(y_after_hdr)

        # Data rows
        for ri, (act, who) in enumerate(CHECK_ROWS):
            row_top = y_after_hdr - ri * CHK_ROW_H
            row_bot = row_top - CHK_ROW_H
            hl(row_bot, lw=0.5)

            # Very light alternate fill (minimal ink)
            if ri % 2 == 0:
                c.setFillColor(LGREY)
                c.rect(L, row_bot, W, CHK_ROW_H, fill=1, stroke=0)

            text_y = row_bot + 7
            c.setFillColor(BLACK)
            c.setFont('Helvetica', 8.5)
            c.drawString(L + 5, text_y, act)
            c.setFont('Helvetica-Bold', 8.5)
            c.drawString(V1_CHK + 5, text_y, who)

        c.showPage()

    c.save()
    buf.seek(0)
    return buf.read()


# ── BST (landscape) ──────────────────────────────────────────────────────────
def build_bst_pdf(batches, date_str):
    buf = io.BytesIO()
    W_PAGE, H_PAGE = landscape(A4)   # 841.9 x 595.3
    c = rl_canvas.Canvas(buf, pagesize=landscape(A4))

    TEAL2  = colors.HexColor('#0d9488')
    BLACK2 = colors.black
    LGREY2 = colors.HexColor('#f5f5f5')
    MGREY2 = colors.HexColor('#666666')

    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        date_display = dt.strftime('%d %B %Y')
    except Exception:
        date_display = date_str

    ML = 15*mm; MR = 15*mm; MT = 8*mm; MB = 8*mm
    PW = W_PAGE - ML - MR

    # Outer border
    c.setStrokeColor(TEAL2); c.setLineWidth(0.8)
    c.rect(ML, MB, PW, H_PAGE - MT - MB, fill=0, stroke=1)

    y = H_PAGE - MT

    # Title
    c.setFont('Helvetica-Bold', 14); c.setFillColor(BLACK2)
    c.drawCentredString(W_PAGE/2, y - 8*mm, 'BATCH SHEET TRACKER')
    c.setFont('Helvetica', 9); c.setFillColor(MGREY2)
    c.drawCentredString(W_PAGE/2, y - 14*mm, date_display)
    c.setStrokeColor(TEAL2); c.setLineWidth(0.6)
    c.line(ML, y - 16*mm, ML + PW, y - 16*mm)

    y -= 19*mm

    # Columns — match reference exactly
    C = {}
    C['no']   = 12*mm
    C['name'] = PW * 0.27
    C['size'] = 20*mm
    C['uom']  = 13*mm
    C['cop']  = 18*mm
    C['ri']   = 19*mm
    C['rs']   = 19*mm
    C['pi']   = 19*mm
    C['ps']   = 19*mm
    used = sum(v for k,v in C.items())
    C['rem']  = PW - used

    keys = ['no','name','size','uom','cop','ri','rs','pi','ps','rem']
    X = {}
    cx = ML
    for k in keys:
        X[k] = cx; cx += C[k]

    H1 = 9*mm   # header row 1 height
    H2 = 7*mm   # header row 2 height
    RH = 9*mm   # data row height
    MAX = 23

    def hl2(y2, lw=0.5):
        c.setStrokeColor(TEAL2); c.setLineWidth(lw)
        c.line(ML, y2, ML+PW, y2)

    def vl2(x2, y0, y1, lw=0.5):
        c.setStrokeColor(TEAL2); c.setLineWidth(lw)
        c.line(x2, y0, x2, y1)

    # Header row 1
    c.setFillColor(colors.HexColor('#f0faf9'))   # barely-there teal tint
    c.rect(ML, y - H1, PW, H1, fill=1, stroke=0)
    c.setStrokeColor(TEAL2); c.setLineWidth(0.5)
    c.rect(ML, y - H1, PW, H1, fill=0, stroke=1)

    c.setFont('Helvetica-Bold', 7.5); c.setFillColor(BLACK2)
    for k, lbl in [('no','#'),('name','Project Name'),('size','BATCH\nSIZE'),
                   ('uom','UOM'),('cop','No of\nCopies')]:
        lines = lbl.split('\n')
        cx2 = X[k] + C[k]/2
        if len(lines) == 2:
            c.drawCentredString(cx2, y - 3.5*mm, lines[0])
            c.drawCentredString(cx2, y - 7*mm,   lines[1])
        else:
            c.drawCentredString(cx2, y - H1/2 - 1.5, lbl)

    rm_cx = X['ri'] + (C['ri']+C['rs'])/2
    pr_cx = X['pi'] + (C['pi']+C['ps'])/2
    c.drawCentredString(rm_cx, y - H1/2 - 1.5, 'RM Store')
    c.drawCentredString(pr_cx, y - H1/2 - 1.5, 'Production')
    c.drawCentredString(X['rem'] + C['rem']/2, y - H1/2 - 1.5, 'Remarks')

    # vertical dividers hdr1
    for k in ['name','size','uom','cop','ri','pi','rem']:
        vl2(X[k], y-H1, y)
    vl2(X['rem']+C['rem'], y-H1, y)
    y -= H1

    # Header row 2
    c.setFillColor(colors.HexColor('#f0faf9'))
    c.rect(X['ri'], y-H2, C['ri']+C['rs']+C['pi']+C['ps'], H2, fill=1, stroke=0)
    c.setStrokeColor(TEAL2); c.setLineWidth(0.5)
    c.rect(X['ri'], y-H2, C['ri']+C['rs']+C['pi']+C['ps'], H2, fill=0, stroke=1)
    c.setFont('Helvetica-Bold', 7); c.setFillColor(BLACK2)
    for k, lbl in [('ri','ISSUE'),('rs','SUBMIT'),('pi','ISSUE'),('ps','SUBMIT')]:
        c.drawCentredString(X[k]+C[k]/2, y-H2/2-2, lbl)
    for k in ['rs','pi','ps']:
        vl2(X[k], y-H2, y)
    y -= H2

    # Data rows
    for ri in range(MAX):
        yr = y - (ri+1)*RH
        if ri % 2 == 0:
            c.setFillColor(LGREY2)
            c.rect(ML, yr, PW, RH, fill=1, stroke=0)
        c.setStrokeColor(colors.HexColor('#cccccc')); c.setLineWidth(0.25)
        c.rect(ML, yr, PW, RH, fill=0, stroke=1)
        for k in ['name','size','uom','cop','ri','rs','pi','ps','rem']:
            vl2(X[k], yr, yr+RH, 0.35)
        vl2(X['rem']+C['rem'], yr, yr+RH, 0.35)

        c.setFillColor(MGREY2); c.setFont('Helvetica', 7.5)
        c.drawCentredString(X['no']+C['no']/2, yr+2.5*mm, str(ri+1))

        if ri < len(batches):
            b = batches[ri]
            bname2 = b['batch_name']
            bsize2 = float(b['batch_size'])
            nb2    = int(b['no_of_batch'])
            avail  = C['name'] - 3*mm
            disp = bname2
            while c.stringWidth(disp,'Helvetica',7.5) > avail and len(disp) > 4:
                disp = disp[:-2]+'\u2026'
            c.setFillColor(BLACK2); c.setFont('Helvetica',7.5)
            c.drawString(X['name']+1.5*mm, yr+2.5*mm, disp)
            c.setFont('Helvetica',8)
            c.drawCentredString(X['size']+C['size']/2, yr+2.5*mm, f'{bsize2:.0f}')
            c.drawCentredString(X['uom']+C['uom']/2,   yr+2.5*mm, 'Kg')
            c.setFont('Helvetica-Bold',8)
            c.drawCentredString(X['cop']+C['cop']/2,   yr+2.5*mm, str(nb2))

    tbl_bot = y - MAX*RH
    c.setStrokeColor(TEAL2); c.setLineWidth(0.8)
    c.rect(ML, tbl_bot, PW, y - tbl_bot + H1 + H2, fill=0, stroke=1)

    c.save()
    buf.seek(0)
    return buf.read()



# ── API: Open Sheet — opens the Excel file on the server machine (admin only) ─
@production_initiater_bp.route('/api/production/open_sheet', methods=['POST'])
@login_required
def api_open_sheet():
    if _prod_role() != 'admin' or session.get('UID', '').lower() == 'rm_store':
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    d        = request.get_json() or {}
    batch_id = d.get('batch_id')

    # Always fetch from DB to get both folder path and file name
    if not batch_id:
        return jsonify({'status': 'error', 'message': 'batch_id required'}), 400

    conn = sampling_portal.get_db_connection()
    row  = conn.execute(
        "SELECT file_name, file_path FROM Processing_batches WHERE id=%s",
        (batch_id,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({'status': 'error', 'message': 'Batch not found'}), 404

    # file_path = folder, file_name = filename — join to get full path
    file_name = row['file_name']

    # FVQ batches have no Excel file to open — data is in procurement_formulations DB
    if file_name == '__FVQ__':
        return jsonify({
            'status':  'error',
            'message': 'This is a Formulation-DB batch — no Excel file to open. '
                       'Use the procurement module to view/edit the formulation.'
        }), 400

    full_path = os.path.join(row['file_path'], file_name)

    if not os.path.exists(full_path):
        return jsonify({'status': 'error', 'message': f'File not found: {full_path}'}), 404

    try:
        import platform
        system = platform.system()
        if system == 'Windows':
            os.startfile(full_path)
        elif system == 'Darwin':
            subprocess.Popen(['open', full_path])
        else:
            subprocess.Popen(['xdg-open', full_path])
        return jsonify({'status': 'ok', 'file_name': file_name, 'full_path': full_path})
    except Exception as ex:
        return jsonify({'status': 'error', 'message': str(ex)}), 500



# ── API: Batch Confirmation — page count (for duplex logic) ──────────────────
@production_initiater_bp.route('/api/production/batch_confirmation_count', methods=['GET'])
@login_required
def api_batch_confirmation_count():
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'error': 'Admin only'}), 403
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    conn = sampling_portal.get_db_connection()
    count = conn.execute(
        """SELECT COUNT(*) as c FROM Processing_batches
            WHERE DATE(added_on) = %s AND batch_type = 'Regular'""",
        (date_str,)
    ).fetchone()['c']
    conn.close()
    return jsonify({'count': count})


# ── API: Batch Confirmation Form PDF ─────────────────────────────────────────
@production_initiater_bp.route('/api/production/batch_confirmation_pdf', methods=['GET'])
@login_required
def api_batch_confirmation_pdf():
    """One page per Regular batch. Supports ?half=1 or ?half=2 for duplex printing."""
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'error': 'Admin only'}), 403
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    half     = request.args.get('half', None)   # '1' or '2' or None = all

    conn = sampling_portal.get_db_connection()
    rows = conn.execute(
        """SELECT batch_name, batch_size, no_of_batch, sku_size, added_on
             FROM Processing_batches
            WHERE DATE(added_on) = %s AND batch_type = 'Regular'
            ORDER BY added_on ASC""", (date_str,)
    ).fetchall()
    conn.close()
    if not rows:
        return jsonify({'error': 'No regular batches found for today'}), 404

    all_batches = [dict(r) for r in rows]
    total = len(all_batches)

    if half == '1':
        # First half: ceil(total/2) pages
        first_count = (total + 1) // 2
        batches = all_batches[:first_count]
    elif half == '2':
        # Second half: remaining pages
        first_count = (total + 1) // 2
        batches = all_batches[first_count:]
        if not batches:
            return jsonify({'error': 'No second-half pages'}), 404
    else:
        batches = all_batches

    pdf_bytes = build_confirmation_pdf(batches)
    buf = io.BytesIO(pdf_bytes)
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        suffix = f"_part{half}" if half else ""
        fname = f"Batch_Confirmation_{dt.strftime('%d%b%Y')}{suffix}.pdf"
    except Exception:
        fname = 'Batch_Confirmation.pdf'
    return send_file(buf, mimetype='application/pdf', download_name=fname)


# ── API: Batch Sheet Tracker PDF ──────────────────────────────────────────────
@production_initiater_bp.route('/api/production/bst_pdf', methods=['GET'])
@login_required
def api_bst_pdf():
    """Single landscape A4. Light-ink design."""
    if _prod_role() not in ('admin', 'rm_store'):
        return jsonify({'error': 'Admin only'}), 403
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    conn = sampling_portal.get_db_connection()
    rows = conn.execute(
        """SELECT batch_name, batch_size, no_of_batch, sku_size, added_on
             FROM Processing_batches
            WHERE DATE(added_on) = %s AND batch_type = 'Regular'
            ORDER BY added_on ASC""", (date_str,)
    ).fetchall()
    conn.close()
    if not rows:
        return jsonify({'error': 'No regular batches found for today'}), 404
    pdf_bytes = build_bst_pdf([dict(r) for r in rows], date_str)
    buf = io.BytesIO(pdf_bytes)
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        fname = f"BST_{dt.strftime('%d%b%Y')}.pdf"
    except Exception:
        fname = 'BatchSheetTracker.pdf'
    return send_file(buf, mimetype='application/pdf', download_name=fname)



# ══════════════════════════════════════════════════════════════════════════════

# ── API: Dispensing Ready — save batch to daily_dsp_summary ──────────────────
@production_initiater_bp.route('/api/production/dispensing_ready', methods=['POST'])
@login_required
def api_dispensing_ready():
    if not _prod_role():
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    d          = request.get_json() or {}
    batch_id   = d.get('batch_id')
    dispensed  = d.get('dispensed', 0)
    remarks    = d.get('remarks', '')
    if not batch_id:
        return jsonify({'status': 'error', 'message': 'batch_id required'}), 400
    conn = sampling_portal.get_db_connection()
    row  = conn.execute(
        "SELECT batch_name, batch_size, no_of_batch, dispensed_batches FROM Processing_batches WHERE id=%s",
        (batch_id,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({'status': 'error', 'message': 'Batch not found'}), 404
    batch_date = datetime.now().strftime('%Y-%m-%d')
    try:
        dispensed = int(dispensed)
    except (TypeError, ValueError):
        dispensed = 0

    no_of_batch    = int(row['no_of_batch'])
    prev_dispensed = int(row.get('dispensed_batches') or 0)

    # ── Cumulative cap: total dispensed across ALL dates must not exceed no_of_batch ──
    # Get today's existing entry so we can replace it (not add on top of it)
    conn_chk = sampling_portal.get_db_connection()
    today_row = conn_chk.execute(
        "SELECT COALESCE(dispensed, 0) AS d FROM daily_dsp_summary WHERE batch_id=%s AND batch_date=%s",
        (batch_id, batch_date)
    ).fetchone()
    conn_chk.close()
    today_existing  = int(today_row['d']) if today_row else 0
    prev_days_disp  = max(0, prev_dispensed - today_existing)   # what was done before today
    new_cumulative  = prev_days_disp + dispensed

    if new_cumulative > no_of_batch:
        max_allowed = max(0, no_of_batch - prev_days_disp)
        return jsonify({
            'status': 'error',
            'message': (
                f"Cannot dispense {dispensed} today — cumulative total would reach "
                f"{new_cumulative}, exceeding No. of Batches ({no_of_batch}). "
                f"Maximum allowed today: {max_allowed}."
            )
        }), 400

    initial_remaining = max(0, no_of_batch - prev_dispensed)
    row_id = sampling_portal.dsp_upsert_entry(
        batch_id          = batch_id,
        batch_name        = row['batch_name'],
        batch_date        = batch_date,
        batch_size        = float(row['batch_size']),
        no_of_batches     = int(row['no_of_batch']),
        dispensed         = dispensed,
        remarks           = remarks,
        initial_remaining = initial_remaining,
    )
    # ── dispensed_batches on Processing_batches = cumulative total (sum across all dates) ──
    conn2 = sampling_portal.get_db_connection()
    total_dispensed_all = conn2.execute(
        "SELECT COALESCE(SUM(dispensed),0) AS s FROM daily_dsp_summary WHERE batch_id=%s",
        (batch_id,)
    ).fetchone()['s']
    conn2.execute(
        "UPDATE Processing_batches SET dispensed_batches = %s WHERE id = %s",
        (int(total_dispensed_all), batch_id)
    )
    conn2.commit()
    conn2.close()
    return jsonify({'status': 'ok', 'id': row_id,
                    'batch_name': row['batch_name'],
                    'no_of_batches': int(row['no_of_batch']),
                    'dispensed_batches': int(total_dispensed_all),   # all-time total
                    'dispensed_today': dispensed})                    # today only


# ── API: Get daily dispensing summary ────────────────────────────────────────
@production_initiater_bp.route('/api/production/dsp_summary', methods=['GET'])
@login_required
def api_dsp_summary():
    if not _prod_role():
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    rows = sampling_portal.dsp_get_summary(date_str)
    # Convert date/datetime objects to strings for JSON serialisation
    for r in rows:
        for k, v in r.items():
            if hasattr(v, 'isoformat'):
                r[k] = v.isoformat()
    return jsonify({'status': 'ok', 'rows': rows, 'date': date_str})


# ── API: Update a dsp_summary entry (dispensed + remarks) ────────────────────
@production_initiater_bp.route('/api/production/dsp_update', methods=['POST'])
@login_required
def api_dsp_update():
    if not _prod_role():
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    d        = request.get_json() or {}
    row_id   = d.get('id')
    dispensed= d.get('dispensed', 0)
    remarks  = d.get('remarks', '')
    if not row_id:
        return jsonify({'status': 'error', 'message': 'id required'}), 400
    ok = sampling_portal.dsp_update_entry(row_id, int(dispensed), remarks)
    if not ok:
        return jsonify({'status': 'error', 'message': 'Entry not found'}), 404
    conn = sampling_portal.get_db_connection()
    row  = conn.execute(
        "SELECT * FROM daily_dsp_summary WHERE id=%s", (row_id,)
    ).fetchone()
    row_dict = dict(row) if row else {}
    # ── Recalculate cumulative dispensed_batches from ALL dates ─────────────
    batch_id = row_dict.get('batch_id')
    if batch_id:
        total_dispensed_all = conn.execute(
            "SELECT COALESCE(SUM(dispensed),0) AS s FROM daily_dsp_summary WHERE batch_id=%s",
            (batch_id,)
        ).fetchone()['s']
        conn.execute(
            "UPDATE Processing_batches SET dispensed_batches=%s WHERE id=%s",
            (int(total_dispensed_all), batch_id)
        )
        conn.commit()
    conn.close()
    for k, v in row_dict.items():
        if hasattr(v, 'isoformat'):
            row_dict[k] = v.isoformat()
    return jsonify({'status': 'ok', 'row': row_dict,
                    'batch_id': batch_id, 'dispensed_batches': int(total_dispensed_all if batch_id else dispensed)})


# ── API: Delete dsp_summary entries ──────────────────────────────────────────
@production_initiater_bp.route('/api/production/dsp_delete', methods=['POST'])
@login_required
def api_dsp_delete():
    if not _prod_role():
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    ids = (request.get_json() or {}).get('ids', [])
    if not ids:
        return jsonify({'status': 'error', 'message': 'ids required'}), 400
    sampling_portal.dsp_delete_entries([int(i) for i in ids])
    return jsonify({'status': 'ok'})


# ── API: WhatsApp message for daily dispensing summary ───────────────────────
@production_initiater_bp.route('/api/production/dsp_whatsapp', methods=['POST'])
@login_required
def api_dsp_whatsapp():
    if not _prod_role():
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    d        = request.get_json() or {}
    date_str = d.get('date', datetime.now().strftime('%Y-%m-%d'))
    rows     = sampling_portal.dsp_get_summary(date_str)
    if not rows:
        return jsonify({'status': 'error', 'message': 'No data for this date'}), 404
    try:
        dt         = datetime.strptime(date_str, '%Y-%m-%d')
        date_label = dt.strftime('%d %b %Y')
    except Exception:
        date_label = date_str
    lines = ['📋 *Daily Batch Dispensing Summary*', '📅 Date: ' + date_label, '']
    for i, r in enumerate(rows, 1):
        no_b       = r.get('no_of_batches', 0)
        total_disp = r.get('total_dispensed_all', r.get('dispensed', 0))
        disp_today = r.get('dispensed_today', r.get('dispensed', 0))
        pending    = max(0, int(no_b) - int(total_disp))
        pend_str   = '✅ Complete' if pending == 0 else '⏳ ' + str(pending) + ' pending'
        entry   = str(i) + '. *' + r['batch_name'] + '*\n'
        entry  += '   Batch Size: ' + str(r['batch_size']) + ' kg  |  No. of Batches: ' + str(no_b) + '\n'
        entry  += '   Total Dispensed: ' + str(total_disp) + '  |  Today: ' + str(disp_today) + '  |  ' + pend_str
        if r.get('remarks'):
            entry += '\n   Remarks: ' + r['remarks']
        lines.append(entry)
    return jsonify({'status': 'ok', 'message': '\n'.join(lines)})


